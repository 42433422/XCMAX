"""Legacy side-database importer retained for rollback-compatible upgrades.

New ERP/ETL writes must use ``erp_attendance_app_service``.  This module only
reads or repairs pre-migration customer data and must not be used as the
canonical attendance store.
"""

from __future__ import annotations

import json
import sqlite3
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
from typing import Any

from app.utils.operational_errors import RECOVERABLE_ERRORS


def _ensure_schema(conn: sqlite3.Connection) -> None:
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance_import_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            month_label TEXT NOT NULL,
            rows_in INTEGER NOT NULL,
            rows_written INTEGER NOT NULL,
            imported_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance_daily_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            month_label TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            employee_name TEXT NOT NULL,
            attendance_group TEXT NOT NULL,
            department TEXT NOT NULL,
            employee_no TEXT NOT NULL,
            position TEXT NOT NULL,
            user_id TEXT NOT NULL,
            work_date TEXT NOT NULL,
            shift_name TEXT NOT NULL,
            daily_times_json TEXT NOT NULL,
            raw_times_json TEXT NOT NULL,
            all_times_json TEXT NOT NULL,
            leave_hours REAL NOT NULL,
            absent_days REAL NOT NULL,
            late_count_hint REAL NOT NULL,
            early_count_hint REAL NOT NULL,
            missing_card_count REAL NOT NULL,
            notes_json TEXT NOT NULL,
            imported_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE UNIQUE INDEX IF NOT EXISTS ux_attendance_source_row
        ON attendance_daily_records (source_file, source_row)
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS ix_attendance_employee_date
        ON attendance_daily_records (employee_name, work_date)
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance_departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            department TEXT NOT NULL,
            main_department TEXT NOT NULL,
            attendance_group TEXT NOT NULL,
            UNIQUE(source_file, department, attendance_group)
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS attendance_employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            employee_name TEXT NOT NULL,
            department TEXT NOT NULL,
            main_department TEXT NOT NULL,
            attendance_group TEXT NOT NULL,
            employee_no TEXT NOT NULL,
            position TEXT NOT NULL,
            user_id TEXT NOT NULL,
            UNIQUE(source_file, employee_name, department)
        )
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS ix_employees_name
        ON attendance_employees (employee_name)
        """
    )
    conn.execute(
        """
        CREATE INDEX IF NOT EXISTS ix_departments_dept
        ON attendance_departments (department)
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL DEFAULT '',
            model_number TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL DEFAULT '',
            specification TEXT NOT NULL DEFAULT '',
            price REAL NOT NULL DEFAULT 0,
            unit TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL DEFAULT '',
            customer_name TEXT NOT NULL DEFAULT '',
            contact_person TEXT NOT NULL DEFAULT '',
            contact_phone TEXT NOT NULL DEFAULT '',
            address TEXT NOT NULL DEFAULT '',
            purchase_unit TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )


def _cell_first_line(value: Any) -> str:
    text = str(value or "").strip()
    return text.split("\n", 1)[0].strip() if text else ""


def _parse_dingtalk_daily_sheet(ws) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    seen_employees: dict[tuple[str, str], dict[str, str]] = OrderedDict()
    seen_departments: dict[tuple[str, str, str], dict[str, str]] = OrderedDict()

    for row in ws.iter_rows(min_row=4, values_only=True):
        name = str(row[0] or "").strip() if row else ""
        if not name:
            continue
        group = str(row[1] or "").strip()
        dept = str(row[2] or "").strip()
        main_dept = str(row[3] or "").strip()
        emp_no = str(row[4] or "").strip()
        position = str(row[5] or "").strip()
        uid = str(row[6] or "").strip()

        emp_key = (name, dept)
        if emp_key not in seen_employees:
            seen_employees[emp_key] = {
                "name": name,
                "group": group,
                "dept": dept,
                "main_dept": main_dept,
                "emp_no": emp_no,
                "position": position,
                "uid": uid,
            }

        dept_key = (dept, main_dept, group)
        if dept_key not in seen_departments:
            seen_departments[dept_key] = {
                "department": dept,
                "main_department": main_dept,
                "attendance_group": group,
            }

    return list(seen_departments.values()), list(seen_employees.values())


def _parse_attendance_detail_sheet(ws) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    seen_employees: dict[tuple[str, str], dict[str, str]] = OrderedDict()
    seen_departments: dict[tuple[str, str, str], dict[str, str]] = OrderedDict()

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        dept = _cell_first_line(row[0] if len(row) > 0 else "")
        nature = _cell_first_line(row[1] if len(row) > 1 else "")
        name = _cell_first_line(row[2] if len(row) > 2 else "")
        if not name or name == "姓名" or dept == "部门":
            continue

        emp_key = (name, dept)
        if emp_key not in seen_employees:
            seen_employees[emp_key] = {
                "name": name,
                "group": nature,
                "dept": dept,
                "main_dept": dept,
                "emp_no": "",
                "position": nature,
                "uid": "",
            }

        dept_key = (dept, dept, nature)
        if dept_key not in seen_departments:
            seen_departments[dept_key] = {
                "department": dept,
                "main_department": dept,
                "attendance_group": nature,
            }

    return list(seen_departments.values()), list(seen_employees.values())


def _infer_month_from_filename(excel_path: Path) -> str:
    import re

    text = excel_path.name
    match = re.search(r"(20\d{2})[-_年.]?\s*(0?[1-9]|1[0-2])", text)
    if not match:
        return ""
    return f"{match.group(1)}-{int(match.group(2)):02d}"


def _parse_workbook(excel_path: Path) -> tuple[list[dict[str, str]], list[dict[str, str]], str]:
    import openpyxl

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        if "每日统计" in names:
            return (*_parse_dingtalk_daily_sheet(wb["每日统计"]), "dingtalk")
        if "明细" in names:
            return (*_parse_attendance_detail_sheet(wb["明细"]), "mingxi")
        raise ValueError(f"未找到「每日统计」或「明细」工作表。实际工作表: {names!r}")
    finally:
        wb.close()


def _sync_products_customers(
    conn: sqlite3.Connection,
    source_file: str,
    employees: list[dict[str, str]],
) -> tuple[int, int]:
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute("DELETE FROM products WHERE source_file = ?", (source_file,))
    conn.execute("DELETE FROM customers WHERE source_file = ?", (source_file,))

    product_rows = 0
    for employee in employees:
        dept = (employee.get("dept") or "").strip()
        name = (employee.get("name") or "").strip()
        group = (employee.get("group") or "").strip()
        model_number = f"{dept}::{name}" if dept else name
        conn.execute(
            """
            INSERT INTO products (source_file, model_number, name, specification, price, unit, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (source_file, model_number, name, group, 0.0, dept, now, now),
        )
        product_rows += 1

    customer_rows = 0
    seen_depts: set[str] = set()
    for employee in employees:
        dept = (employee.get("dept") or "").strip()
        if not dept or dept in seen_depts:
            continue
        seen_depts.add(dept)
        conn.execute(
            """
            INSERT INTO customers (source_file, customer_name, contact_person, contact_phone, address, purchase_unit, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (source_file, dept, "", "", "", "", now, now),
        )
        customer_rows += 1

    return product_rows, customer_rows


def _to_datetime_text(values: list[Any]) -> list[str]:
    out: list[str] = []
    for value in values:
        if hasattr(value, "strftime"):
            out.append(value.strftime("%Y-%m-%d %H:%M:%S"))
        else:
            out.append(str(value))
    return out


def _import_daily_records_if_possible(
    conn: sqlite3.Connection,
    excel_path: Path,
    source_file: str,
    month: str,
) -> tuple[int, int, str]:
    try:
        from app.shell.taiyangniao_attendance.parser import parse_attendance_workbook
    except ModuleNotFoundError:
        return 0, 0, month

    parsed = parse_attendance_workbook(excel_path, month=month or None)
    month_label = parsed.month or month
    imported_at = datetime.now().isoformat(timespec="seconds")

    conn.execute("DELETE FROM attendance_daily_records WHERE source_file = ?", (source_file,))
    rows_written = 0
    for record in parsed.records:
        conn.execute(
            """
            INSERT INTO attendance_daily_records (
                source_file,
                month_label,
                source_row,
                employee_name,
                attendance_group,
                department,
                employee_no,
                position,
                user_id,
                work_date,
                shift_name,
                daily_times_json,
                raw_times_json,
                all_times_json,
                leave_hours,
                absent_days,
                late_count_hint,
                early_count_hint,
                missing_card_count,
                notes_json,
                imported_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                source_file,
                month_label,
                int(record.source_row),
                record.employee_name,
                record.attendance_group,
                record.department,
                record.employee_no,
                record.position,
                record.user_id,
                record.work_date.isoformat(),
                record.shift_name,
                json.dumps(_to_datetime_text(record.daily_times), ensure_ascii=False),
                json.dumps(_to_datetime_text(record.raw_times), ensure_ascii=False),
                json.dumps(_to_datetime_text(record.all_punch_times()), ensure_ascii=False),
                float(record.leave_hours),
                float(record.absent_days),
                float(record.late_count_hint),
                float(record.early_count_hint),
                float(record.missing_card_count),
                json.dumps(record.notes, ensure_ascii=False),
                imported_at,
            ),
        )
        rows_written += 1

    return int(parsed.rows_in), rows_written, month_label


def import_attendance_workbook(
    excel_path: Path,
    db_path: Path,
    *,
    source_file_key: str | None = None,
    sync_ui_tables: bool = False,
) -> dict[str, Any]:
    """Import into a legacy Sunbird private database during controlled migration.

    ``每日统计`` workbooks write daily records plus roster tables. Fixed-template
    ``明细`` workbooks write roster tables only, because converted result files no
    longer contain DingTalk raw punch records.
    """

    if not excel_path.is_file():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    departments, employees, workbook_kind = _parse_workbook(excel_path)
    source_file = (source_file_key or "").strip() or str(excel_path.resolve())
    month_label = _infer_month_from_filename(excel_path)
    imported_at = datetime.now().isoformat(timespec="seconds")

    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    try:
        _ensure_schema(conn)
        conn.execute("BEGIN")
        conn.execute("DELETE FROM attendance_departments WHERE source_file = ?", (source_file,))
        conn.execute("DELETE FROM attendance_employees WHERE source_file = ?", (source_file,))

        department_rows = 0
        for department in departments:
            conn.execute(
                """
                INSERT OR IGNORE INTO attendance_departments
                    (source_file, department, main_department, attendance_group)
                VALUES (?, ?, ?, ?)
                """,
                (
                    source_file,
                    department["department"],
                    department["main_department"],
                    department["attendance_group"],
                ),
            )
            department_rows += 1

        employee_rows = 0
        for employee in employees:
            conn.execute(
                """
                INSERT OR IGNORE INTO attendance_employees
                    (source_file, employee_name, department, main_department,
                     attendance_group, employee_no, position, user_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    source_file,
                    employee["name"],
                    employee["dept"],
                    employee["main_dept"],
                    employee["group"],
                    employee["emp_no"],
                    employee["position"],
                    employee["uid"],
                ),
            )
            employee_rows += 1

        product_rows = customer_rows = 0
        if sync_ui_tables:
            product_rows, customer_rows = _sync_products_customers(conn, source_file, employees)

        daily_rows_in = daily_rows_written = 0
        if workbook_kind == "dingtalk":
            daily_rows_in, daily_rows_written, month_label = _import_daily_records_if_possible(
                conn, excel_path, source_file, month_label
            )

        conn.execute(
            """
            INSERT INTO attendance_import_batches (
                source_file, month_label, rows_in, rows_written, imported_at
            ) VALUES (?, ?, ?, ?, ?)
            """,
            (
                source_file,
                month_label,
                daily_rows_in or employee_rows,
                daily_rows_written or employee_rows,
                imported_at,
            ),
        )
        batch_id = int(conn.execute("SELECT last_insert_rowid()").fetchone()[0])
        conn.commit()
        return {
            "source_file": source_file,
            "db_path": str(db_path),
            "workbook_kind": workbook_kind,
            "month_label": month_label,
            "batch_id": batch_id,
            "department_rows": department_rows,
            "employee_rows": employee_rows,
            "product_rows": product_rows,
            "customer_rows": customer_rows,
            "daily_rows_in": daily_rows_in,
            "daily_rows_written": daily_rows_written,
            "sync_ui_tables": sync_ui_tables,
        }
    except RECOVERABLE_ERRORS:
        conn.rollback()
        raise
    finally:
        conn.close()


__all__ = ["import_attendance_workbook"]
