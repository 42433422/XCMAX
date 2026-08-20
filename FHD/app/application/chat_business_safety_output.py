"""Verified attendance export and print actions for protected chat."""

from __future__ import annotations

from pathlib import Path
from typing import Any
from urllib.parse import quote

from app.application.chat_business_safety_attendance import (
    _attendance_error_payload,
    _attendance_rows,
    _json_list,
)
from app.application.chat_business_safety_core import (
    _DB_NAME,
    BusinessActorIdentity,
    BusinessChatIntent,
    _new_receipt,
    _not_executed,
    _payload,
)
from app.utils.operational_errors import RECOVERABLE_ERRORS

_EXPORT_COLUMNS = (
    ("work_date", "日期"),
    ("employee_name", "姓名"),
    ("employee_no", "工号"),
    ("department", "部门"),
    ("position", "岗位"),
    ("shift_name", "班次"),
    ("all_times_json", "打卡时间"),
    ("leave_hours", "请假小时"),
    ("absent_days", "缺勤天数"),
    ("late_count_hint", "迟到标记"),
    ("early_count_hint", "早退标记"),
    ("missing_card_count", "缺卡次数"),
)


def _create_attendance_export(rows: list[dict[str, Any]], meta: dict[str, Any]) -> tuple[Path, str]:
    import openpyxl

    _ = meta
    from app.infrastructure.workspace import (
        allocate_generated_workspace_file,
        workspace_root,
    )

    output = allocate_generated_workspace_file("attendance-export")
    relpath = output.relative_to(workspace_root()).as_posix()

    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "考勤明细"
    ws.append([label for _, label in _EXPORT_COLUMNS])
    for row in rows:
        values: list[Any] = []
        for key, _ in _EXPORT_COLUMNS:
            value = row.get(key)
            if key == "all_times_json":
                value = "、".join(_json_list(value))
            values.append(value)
        ws.append(values)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    from openpyxl.utils import get_column_letter

    for column_index, col in enumerate(ws.columns, start=1):
        width = min(36, max(10, max(len(str(cell.value or "")) for cell in col) + 2))
        ws.column_dimensions[get_column_letter(column_index)].width = width
    wb.save(output)
    wb.close()
    return output, relpath


def _handle_attendance_export(
    message: str, intent: BusinessChatIntent, *, actor: BusinessActorIdentity
) -> dict[str, Any]:
    rows, meta, error = _attendance_rows(message, actor=actor)
    if error:
        return _attendance_error_payload(intent, error, meta)
    if not rows:
        source = f"{_DB_NAME}:attendance_daily_records"
        return _not_executed(
            intent,
            "考勤表未生成：已查询真实考勤库，但指定范围没有记录。",
            reason="no_attendance_rows",
            source=source,
            details=meta,
        )
    try:
        output, relpath = _create_attendance_export(rows, meta)
    except RECOVERABLE_ERRORS as exc:  # noqa: BLE001 - converted to a truthful receipt
        return _not_executed(
            intent,
            "考勤表未生成：文件写出失败。",
            reason=f"attendance_export_failed:{exc}",
            source=f"{_DB_NAME}:attendance_daily_records",
            status="failed",
            details=meta,
        )
    url = f"/api/mod/taiyangniao-pro/attendance/download?relpath={quote(relpath)}"
    artifact = {
        "kind": "file",
        "filename": output.name,
        "path": str(output),
        "download_url": url,
        "size_bytes": output.stat().st_size,
    }
    receipt = _new_receipt(
        intent,
        status="completed",
        executed=True,
        verified=output.is_file() and output.stat().st_size > 0,
        source=f"{_DB_NAME}:attendance_daily_records",
        affected_rows=len(rows),
        artifacts=[artifact],
        details=meta,
    )
    text = (
        f"考勤表已根据真实记录生成，共 {len(rows)} 行。\n"
        f"[下载 {output.name}]({url})\n"
        f"导出回执：{receipt['receipt_id']}。"
    )
    return _payload(text, intent, receipt)


def _default_get_printer_service():
    from app.services import get_printer_service

    return get_printer_service()


def _handle_attendance_print(
    message: str, intent: BusinessChatIntent, *, actor: BusinessActorIdentity
) -> dict[str, Any]:
    rows, meta, error = _attendance_rows(message, actor=actor)
    if error:
        return _attendance_error_payload(intent, error, meta)
    if not rows:
        return _not_executed(
            intent,
            "打印未执行：指定范围没有可核验的考勤记录。",
            reason="no_attendance_rows",
            source=f"{_DB_NAME}:attendance_daily_records",
            details=meta,
        )
    try:
        from app.application import chat_business_safety as facade

        service = facade._get_printer_service()
        printers = service.get_printers()
    except RECOVERABLE_ERRORS as exc:  # noqa: BLE001
        return _not_executed(
            intent,
            "打印未执行：无法读取系统打印机状态。",
            reason=f"printer_status_failed:{exc}",
            source="printer_service",
            status="failed",
        )
    count = (
        int(printers.get("count") or len(printers.get("printers") or []))
        if isinstance(printers, dict)
        else 0
    )
    if not isinstance(printers, dict) or not printers.get("success") or count <= 0:
        return _not_executed(
            intent,
            "打印未执行：当前没有可用打印机。请先在“考勤打印机”中连接并验证打印机。",
            reason="no_available_printer",
            source="printer_service",
            details={"printer_count": count},
        )
    try:
        output, relpath = _create_attendance_export(rows, meta)
        result = service.print_document(str(output))
    except RECOVERABLE_ERRORS as exc:  # noqa: BLE001
        return _not_executed(
            intent,
            "打印未执行：考勤文件生成或提交打印失败。",
            reason=f"print_submit_failed:{exc}",
            source="printer_service",
            status="failed",
        )
    if not isinstance(result, dict) or not result.get("success"):
        reason = (
            str((result or {}).get("message") or "printer_rejected")
            if isinstance(result, dict)
            else "printer_rejected"
        )
        return _not_executed(
            intent,
            f"打印未执行：打印服务未接受任务（{reason}）。",
            reason=reason,
            source="printer_service",
            status="failed",
        )
    url = f"/api/mod/taiyangniao-pro/attendance/download?relpath={quote(relpath)}"
    printer_name = str(result.get("printer") or result.get("printer_name") or "系统默认打印机")
    receipt = _new_receipt(
        intent,
        status="submitted",
        executed=True,
        verified=True,
        source="printer_service",
        affected_rows=len(rows),
        artifacts=[
            {
                "kind": "file",
                "filename": output.name,
                "path": str(output),
                "download_url": url,
                "size_bytes": output.stat().st_size,
            }
        ],
        details={**meta, "printer": printer_name, "backend_result": result},
    )
    text = (
        f"考勤表已提交到 {printer_name}，共 {len(rows)} 行。打印机是否实际出纸仍以设备状态为准。\n"
        f"打印回执：{receipt['receipt_id']}；[下载打印文件]({url})。"
    )
    return _payload(text, intent, receipt)
