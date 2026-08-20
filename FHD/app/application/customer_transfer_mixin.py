import logging
from datetime import datetime
from typing import TYPE_CHECKING, Any

from app.utils.operational_errors import RECOVERABLE_ERRORS

logger = logging.getLogger(__name__)


class CustomerTransferMixin:
    if TYPE_CHECKING:

        def _get_session(self) -> Any: ...

    def import_data(
        self,
        data: list[dict[str, Any]],
        skip_duplicates: bool = True,
        validate_before_import: bool = True,
        clean_data: bool = True,
    ) -> dict[str, Any]:
        """导入客户数据（从解析后的数据列表）

        Args:
            data: 客户数据列表，每个元素包含 customer_name 等字段
            skip_duplicates: 是否跳过重复数据
            validate_before_import: 是否导入前验证
            clean_data: 是否清理数据

        Returns:
            导入结果统计
        """
        try:
            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                imported = 0
                skipped = 0
                failed = 0
                failed_items = []
                skipped_items = []

                for item in data:
                    try:
                        customer_name = (
                            item.get("customer_name")
                            or item.get("unit_name")
                            or item.get("name", "").strip()
                        )

                        if not customer_name:
                            skipped += 1
                            skipped_items.append({"reason": "客户名称为空", "item": item})
                            continue

                        if clean_data:
                            customer_name = customer_name.strip()

                        existing = (
                            session.query(PurchaseUnitModel)
                            .filter(PurchaseUnitModel.unit_name == customer_name)
                            .first()
                        )

                        if existing:
                            if skip_duplicates:
                                skipped += 1
                                skipped_items.append(
                                    {"reason": "客户已存在", "customer_name": customer_name}
                                )
                                continue
                            else:
                                existing.contact_person = (
                                    item.get("contact_person") or existing.contact_person
                                )
                                existing.contact_phone = (
                                    item.get("contact_phone") or existing.contact_phone
                                )
                                existing.address = (
                                    item.get("address")
                                    or item.get("contact_address")
                                    or existing.address
                                )
                        else:
                            unit = PurchaseUnitModel(
                                unit_name=customer_name,
                                contact_person=item.get("contact_person") or "",
                                contact_phone=item.get("contact_phone") or "",
                                address=item.get("address") or item.get("contact_address") or "",
                            )
                            session.add(unit)

                        imported += 1

                    except RECOVERABLE_ERRORS:
                        logger.exception("导入单条客户数据失败")
                        failed += 1
                        failed_items.append({"reason": "customer_import_failed", "item": item})

                session.commit()

                return {
                    "success": True,
                    "imported": imported,
                    "skipped": skipped,
                    "failed": failed,
                    "details": {"failed_items": failed_items, "skipped_items": skipped_items},
                }
            finally:
                session.close()

        except RECOVERABLE_ERRORS as e:
            logger.exception("导入客户数据失败: %s", e)
            return {
                "success": False,
                "imported": 0,
                "skipped": 0,
                "failed": 0,
                "details": {"failed_items": [], "skipped_items": []},
            }

    def import_from_excel(self, file) -> dict[str, Any]:
        """从 Excel 导入购买单位"""
        from app.db.sqlite_write_guard import sqlite_write_guard

        try:
            with sqlite_write_guard():
                return self._import_from_excel_locked(file)
        except RECOVERABLE_ERRORS as e:
            logger.exception("导入客户数据失败: %s", e)
            return {
                "success": False,
                "imported": 0,
                "skipped": 0,
                "failed": 0,
                "details": {"failed_items": [], "skipped_items": []},
            }

    def _import_from_excel_locked(self, file) -> dict[str, Any]:
        try:
            from openpyxl import load_workbook

            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                wb = load_workbook(file)
                ws = wb.active

                updated = 0
                inserted = 0
                skipped = 0

                assert ws is not None
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if not row or not row[0]:
                        continue

                    unit_name = str(row[0]).strip()
                    if not unit_name:
                        skipped += 1
                        continue

                    existing = (
                        session.query(PurchaseUnitModel)
                        .filter(PurchaseUnitModel.unit_name == unit_name)
                        .first()
                    )

                    if existing:
                        if row[1]:
                            existing.contact_person = str(row[1])
                        if row[2]:
                            existing.contact_phone = str(row[2])
                        if row[3]:
                            existing.address = str(row[3])
                        updated += 1
                    else:
                        unit = PurchaseUnitModel(
                            unit_name=unit_name,
                            contact_person=str(row[1]) if row[1] else "",
                            contact_phone=str(row[2]) if row[2] else "",
                            address=str(row[3]) if row[3] else "",
                        )
                        session.add(unit)
                        inserted += 1

                session.commit()

                return {
                    "success": True,
                    "message": "导入完成",
                    "updated": updated,
                    "inserted": inserted,
                    "skipped": skipped,
                }
            finally:
                session.close()

        except RECOVERABLE_ERRORS:
            logger.exception("导入失败")
            return {
                "success": False,
                "message": "客户导入失败",
                "updated": 0,
                "inserted": 0,
                "skipped": 0,
            }

    def export_to_excel(
        self, keyword: str | None = None, template_id: str | None = None
    ) -> dict[str, Any]:
        """导出购买单位到 Excel"""
        try:
            import os

            from openpyxl import Workbook

            from app.utils.excel.template_export_utils import fill_workbook_from_template
            from app.utils.path_io.path_utils import get_data_dir

            session = self._get_session()
            try:
                from app.db.models.purchase_unit import PurchaseUnit as PurchaseUnitModel

                query = session.query(PurchaseUnitModel).filter(PurchaseUnitModel.is_active == True)

                if keyword:
                    pattern = f"%{keyword}%"
                    query = query.filter(PurchaseUnitModel.unit_name.like(pattern))

                units = query.order_by(PurchaseUnitModel.unit_name).all()

                records = [
                    {
                        "id": unit.id,
                        "customer_name": unit.unit_name or "",
                        "contact_person": unit.contact_person or "",
                        "contact_phone": unit.contact_phone or "",
                        "address": unit.address or "",
                    }
                    for unit in units
                ]

                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"customers_{timestamp}.xlsx"

                export_dir = os.path.join(get_data_dir(), "exports")
                os.makedirs(export_dir, exist_ok=True)
                file_path = os.path.join(export_dir, filename)

                template_path = None
                if template_id:
                    try:
                        from app.application import get_template_app_service

                        templates = (get_template_app_service().get_templates() or {}).get(
                            "templates"
                        ) or []
                        target = next(
                            (t for t in templates if str(t.get("id")) == str(template_id)), None
                        )
                        if target:
                            candidate_path = str(
                                target.get("path") or target.get("file_path") or ""
                            ).strip()
                            if candidate_path and os.path.exists(candidate_path):
                                template_path = candidate_path
                    except RECOVERABLE_ERRORS:
                        template_path = None

                if template_path:
                    header_alias = {
                        "id": ["ID", "编号"],
                        "customer_name": ["客户名称", "购买单位", "单位名称"],
                        "contact_person": ["联系人", "联系人姓名"],
                        "contact_phone": ["电话", "联系电话", "手机号"],
                        "address": ["地址", "联系地址"],
                    }
                    wb = fill_workbook_from_template(
                        template_path=template_path,
                        records=records,
                        field_alias_map=header_alias,
                        sheet_name="客户列表",
                    )
                else:
                    wb = Workbook()
                    ws = wb.active
                    assert ws is not None
                    ws.title = "客户列表"
                    ws.append(["ID", "客户名称", "联系人", "电话", "地址"])
                    for row in records:
                        ws.append(
                            [
                                row["id"],
                                row["customer_name"],
                                row["contact_person"],
                                row["contact_phone"],
                                row["address"],
                            ]
                        )

                wb.save(file_path)

                return {
                    "success": True,
                    "message": f"成功导出 {len(units)} 条记录",
                    "file_path": str(file_path),
                    "filename": filename,
                }
            finally:
                session.close()

        except RECOVERABLE_ERRORS:
            logger.exception("导出失败")
            return {"success": False, "message": "客户导出失败"}
