"""Deterministic workbook inventory and per-document preview routing."""

from __future__ import annotations

import hashlib
import json
from collections import defaultdict
from typing import Any

from openpyxl.utils import get_column_letter

DOCUMENT_TARGETS = {
    "purchase_order": "purchase_orders",
    "delivery_note": "shipment_records",
    "attendance": "attendance",
    "customer_directory": "customers",
    "product_catalog": "products",
    "quotation": "export_xlsx",
    "invoice": "export_xlsx",
    "packing_list": "export_xlsx",
    # A ledger is historical/reference evidence. Importing it as a shipment
    # write without an explicit user choice is unsafe.
    "shipment_ledger": "export_xlsx",
    "generic_table": "export_xlsx",
}


def document_target(document_type: Any, *, hinted_target_type: str = "") -> str:
    target = DOCUMENT_TARGETS.get(str(document_type or "").strip())
    return target or str(hinted_target_type or "").strip() or "export_xlsx"


def _observed_range(sheet: dict[str, Any]) -> tuple[str, int]:
    cells = [
        cell
        for row in sheet.get("rows") or []
        if isinstance(row, dict)
        for cell in row.get("cells") or []
        if isinstance(cell, dict)
    ]
    if not cells:
        return "", 0
    min_row = min(int(cell.get("row") or 1) for cell in cells)
    max_row = max(int(cell.get("row") or 1) for cell in cells)
    min_column = min(int(cell.get("column") or 1) for cell in cells)
    max_column = max(int(cell.get("column") or 1) for cell in cells)
    return (
        f"{get_column_letter(min_column)}{min_row}:"
        f"{get_column_letter(max_column)}{max_row}",
        len(cells),
    )


def build_sheet_inventory(
    evidence: dict[str, Any],
    documents: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Return the first, target-independent stage of workbook understanding."""

    documents_by_sheet: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for document in documents or []:
        if not isinstance(document, dict):
            continue
        documents_by_sheet[str(document.get("sheet") or "")].append(document)

    evidence_by_sheet = {
        str(item.get("name") or ""): item
        for item in evidence.get("sheets") or []
        if isinstance(item, dict)
    }
    manifest = list(evidence.get("sheet_manifest") or evidence.get("sheets") or [])
    inventory: list[dict[str, Any]] = []
    for position, raw_sheet in enumerate(manifest, start=1):
        if not isinstance(raw_sheet, dict):
            continue
        name = str(raw_sheet.get("name") or f"Sheet{position}")
        sheet_evidence = evidence_by_sheet.get(name, {})
        observed_range, observed_cell_count = _observed_range(sheet_evidence)
        sheet_documents = documents_by_sheet.get(name, [])
        document_types = [
            str(document.get("document_type") or "generic_table")
            for document in sheet_documents
        ]
        max_row = int(raw_sheet.get("max_row") or 0)
        max_column = int(raw_sheet.get("max_column") or 0)
        is_empty = (
            observed_cell_count == 0
            and max_row <= 1
            and max_column <= 1
        )
        inspected = name in evidence_by_sheet
        sheet_requires_review = any(
            bool(document.get("requires_review"))
            for document in sheet_documents
        )
        if is_empty:
            structure = "empty"
            status = "empty"
        elif len(sheet_documents) > 1:
            structure = "multi_document"
            status = "review_required" if sheet_requires_review else "classified"
        elif len(sheet_documents) == 1:
            structure = "single_document"
            status = "review_required" if sheet_requires_review else "classified"
        elif inspected:
            structure = "unclassified"
            status = "review_required"
        else:
            structure = "not_inspected"
            status = "review_required"
        physical_range = ""
        if max_row > 0 and max_column > 0:
            physical_range = f"A1:{get_column_letter(max_column)}{max_row}"
        inventory.append(
            {
                "sheet_index": int(raw_sheet.get("index") or position),
                "sheet": name,
                "max_row": max_row,
                "max_column": max_column,
                "physical_range": physical_range,
                "observed_effective_range": observed_range,
                "observed_cell_count": observed_cell_count,
                "evidence_complete": bool(
                    inspected
                    and not sheet_evidence.get("truncated")
                    and max_row <= int((evidence.get("limits") or {}).get("rows_per_sheet") or 0)
                    and max_column <= int((evidence.get("limits") or {}).get("columns") or 0)
                ),
                "is_empty": is_empty,
                "status": status,
                "structure": structure,
                "document_count": len(sheet_documents),
                "document_ids": [
                    str(document.get("document_id") or "")
                    for document in sheet_documents
                ],
                "business_objects": list(dict.fromkeys(document_types)),
            }
        )
    return inventory


def build_document_routes(
    understanding: dict[str, Any],
    *,
    hinted_target_type: str,
) -> list[dict[str, Any]]:
    """Create stable one-document routes after the sheet inventory is complete."""

    sheet_index_by_name = {
        str(item.get("sheet") or ""): int(item.get("sheet_index") or 0)
        for item in understanding.get("sheet_inventory") or []
        if isinstance(item, dict)
    }
    routes: list[dict[str, Any]] = []
    for position, document in enumerate(understanding.get("documents") or [], start=1):
        if not isinstance(document, dict):
            continue
        document_type = str(document.get("document_type") or "generic_table")
        if document_type == "ignore":
            continue
        sheet = str(document.get("sheet") or "")
        document_id = str(document.get("document_id") or f"document-{position}")
        recommended_target_type = document_target(
            document_type,
            hinted_target_type=hinted_target_type,
        )
        target_type = recommended_target_type
        route_reason = "document_type_default"
        safety_review = False
        tables = [item for item in document.get("tables") or [] if isinstance(item, dict)]
        header_roles = {
            str(item.get("role") or "")
            for item in document.get("header_fields") or []
            if isinstance(item, dict)
        }
        planned_rows = sum(
            max(
                0,
                int(table.get("data_end_row") or 0)
                - int(table.get("data_start_row") or 0)
                + 1,
            )
            for table in tables
        )
        if (
            document_type == "delivery_note"
            and recommended_target_type == "shipment_records"
            and (planned_rows >= 80 or "customer" not in header_roles)
        ):
            target_type = "export_xlsx"
            route_reason = (
                "large_table_without_delivery_header"
                if planned_rows >= 80 and "customer" not in header_roles
                else (
                    "large_table_requires_document_segmentation"
                    if planned_rows >= 80
                    else "delivery_customer_missing"
                )
            )
            safety_review = True
        elif (
            document_type == "product_catalog"
            and recommended_target_type == "products"
            and "customer" not in header_roles
        ):
            target_type = "export_xlsx"
            route_reason = "product_owner_context_missing"
            safety_review = True
        route_seed = json.dumps(
            {
                "sheet_index": sheet_index_by_name.get(sheet, 0),
                "sheet": sheet,
                "document_id": document_id,
                "position": position,
                "target_type": target_type,
            },
            ensure_ascii=False,
            sort_keys=True,
        )
        route_id = hashlib.sha256(route_seed.encode("utf-8")).hexdigest()[:20]
        routes.append(
            {
                "route_id": route_id,
                "document_id": document_id,
                "document_index": position,
                "sheet_index": sheet_index_by_name.get(sheet, 0),
                "sheet": sheet,
                "document_type": document_type,
                "target_type": target_type,
                "recommended_target_type": recommended_target_type,
                "route_reason": route_reason,
                "planned_row_count": planned_rows,
                "table_count": len(tables),
                "data_ranges": [
                    {
                        "header_start_row": int(table.get("header_start_row") or 0),
                        "header_end_row": int(table.get("header_end_row") or 0),
                        "data_start_row": int(table.get("data_start_row") or 0),
                        "data_end_row": int(table.get("data_end_row") or 0),
                    }
                    for table in tables
                ],
                "confidence": float(document.get("confidence") or 0.0),
                "requires_review": bool(document.get("requires_review")) or safety_review,
                "status": "planned",
            }
        )
    return routes


def scoped_document_plan(
    understanding: dict[str, Any],
    route: dict[str, Any],
) -> dict[str, Any]:
    document_id = str(route.get("document_id") or "")
    position = int(route.get("document_index") or 0)
    documents = list(understanding.get("documents") or [])
    document = next(
        (
            item
            for index, item in enumerate(documents, start=1)
            if isinstance(item, dict)
            and str(item.get("document_id") or "") == document_id
            and (not position or index == position)
        ),
        None,
    )
    if document is None and 0 < position <= len(documents):
        candidate = documents[position - 1]
        document = candidate if isinstance(candidate, dict) else None
    return {
        **understanding,
        "file_structure": "single_document",
        "summary": (
            f"{route.get('sheet') or '工作表'} 中的"
            f"{route.get('document_type') or '业务单据'}独立预演"
        ),
        "recommended_target_type": str(route.get("target_type") or "export_xlsx"),
        "document_count": 1 if document is not None else 0,
        "documents": [document] if document is not None else [],
        "routing_scope": {
            "route_id": route.get("route_id"),
            "sheet_index": route.get("sheet_index"),
            "sheet": route.get("sheet"),
            "document_id": route.get("document_id"),
            "document_index": route.get("document_index"),
            "target_type": route.get("target_type"),
            "recommended_target_type": route.get("recommended_target_type"),
            "route_reason": route.get("route_reason"),
        },
        "workbook_document_count": int(
            understanding.get("document_count") or len(documents)
        ),
    }


__all__ = [
    "DOCUMENT_TARGETS",
    "build_document_routes",
    "build_sheet_inventory",
    "document_target",
    "scoped_document_plan",
]
