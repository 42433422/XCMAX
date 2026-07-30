"""Parse workbook rows according to an evidence-bound document plan."""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.application.etl.document_field_values import (
    HEADER_ROLE_LABELS,
    normalize_header_role_value,
)
from app.application.etl.errors import EtlError
from app.application.etl.parser_structure import (
    clean_cell_text,
    is_footer_or_note_row,
    is_repeated_header,
    semantic_key,
)
from app.application.etl.parser_types import ParsedDataset, ParsedRow

_DOCUMENT_TARGETS = {
    "purchase_order": "purchase_orders",
    "delivery_note": "shipment_records",
    "attendance": "attendance",
    "customer_directory": "customers",
    "product_catalog": "products",
    "shipment_ledger": "export_xlsx",
    "quotation": "export_xlsx",
    "invoice": "export_xlsx",
    "packing_list": "export_xlsx",
    "generic_table": "export_xlsx",
}


def _composed_headers(worksheet: Any, table: dict[str, Any]) -> list[str]:
    first_column = max(1, int(table.get("first_column") or 1))
    last_column = max(first_column, int(table.get("last_column") or first_column))
    start = max(1, int(table.get("header_start_row") or 1))
    end = max(start, int(table.get("header_end_row") or start))
    header_rows = list(
        worksheet.iter_rows(
            min_row=start,
            max_row=end,
            min_col=first_column,
            max_col=last_column,
            values_only=True,
        )
    )
    headers: list[str] = []
    used: dict[str, int] = {}
    for offset in range(last_column - first_column + 1):
        parts = []
        for row in header_rows:
            text = clean_cell_text(row[offset] if offset < len(row) else None)
            if text and text not in parts:
                parts.append(text)
        column = first_column + offset
        base = "/".join(parts)[-160:] if parts else f"未命名列{column}"
        used[base] = used.get(base, 0) + 1
        headers.append(base if used[base] == 1 else f"{base}_{used[base]}")
    return headers


def _header_values(document: dict[str, Any]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    header_fields = [
        item
        for item in document.get("header_fields") or []
        if isinstance(item, dict)
    ]
    for item in header_fields:
        role = str(item.get("role") or "")
        if role == "other":
            continue
        label = HEADER_ROLE_LABELS.get(role) or clean_cell_text(item.get("label")) or role
        value = normalize_header_role_value(
            role,
            item.get("value"),
            label=item.get("label"),
        )
        if label and value not in (None, ""):
            result[label] = value
    for role, canonical_label in HEADER_ROLE_LABELS.items():
        if canonical_label in result or role in {"total_amount", "remark"}:
            continue
        for item in header_fields:
            inferred = normalize_header_role_value(
                role,
                item.get("value"),
                label=item.get("label"),
            )
            raw = str(item.get("value") or "").strip()
            if inferred in (None, "") or str(inferred).strip() == raw:
                continue
            result[canonical_label] = inferred
            break
    return result


def _header_role_value(document: dict[str, Any], role: str) -> Any:
    for item in document.get("header_fields") or []:
        if isinstance(item, dict) and str(item.get("role") or "") == role:
            return normalize_header_role_value(
                role,
                item.get("value"),
                label=item.get("label"),
            )
    return None


def _party_name(value: Any) -> str:
    text = str(value or "").strip()
    return re.split(
        r"\s+(?=(?:联系人|经办人|日期|订单(?:号|编号)|单号|contact|date|no)\s*[:：])",
        text,
        maxsplit=1,
        flags=re.I,
    )[0].strip()


def _column_roles(
    table: dict[str, Any],
    headers: list[str],
) -> dict[str, str]:
    first_column = max(1, int(table.get("first_column") or 1))
    result: dict[str, str] = {}
    for item in table.get("columns") or []:
        if not isinstance(item, dict):
            continue
        column = int(item.get("column") or 0)
        role = str(item.get("role") or "")
        offset = column - first_column
        if 0 <= offset < len(headers) and role and role != "other":
            result[role] = headers[offset]
    return result


def _decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    text = str(value).strip().replace(",", "").replace("，", "").replace("￥", "").replace("¥", "")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError):
        return None


def _row_arithmetic_issues(
    values: dict[str, Any],
    roles: dict[str, str],
    *,
    target_type: str,
) -> list[dict[str, Any]]:
    unit_price = _decimal(values.get(roles.get("unit_price", "")))
    amount = _decimal(values.get(roles.get("amount", "")))
    quantity_headers = list(
        dict.fromkeys(
            [
                roles.get("quantity", ""),
                *[
                    header
                    for header in values
                    if any(
                        marker in semantic_key(header)
                        for marker in ("数量kg", "重量", "公斤数", "weight", "qtykg")
                    )
                ],
            ]
        )
    )
    quantity_headers = [header for header in quantity_headers if header]
    quantities = [
        (header, quantity)
        for header in quantity_headers
        if (quantity := _decimal(values.get(header))) is not None
    ]
    if not quantities or unit_price is None or amount is None:
        return []
    tolerance = max(Decimal("0.01"), abs(amount) * Decimal("0.0001"))
    candidates = [
        (header, quantity * unit_price)
        for header, quantity in quantities
    ]
    if any(abs(expected - amount) <= tolerance for _header, expected in candidates):
        return []
    quantity_header, expected = candidates[0]
    return [
        {
            "code": "ETL_LINE_AMOUNT_MISMATCH",
            "severity": "warning" if target_type in {"export_xlsx", "export_csv"} else "error",
            "field": roles.get("amount", ""),
            "message": (
                f"{quantity_header} × 单价为 {expected}，"
                f"与明细金额 {amount} 不一致"
            ),
        }
    ]


def _target_mismatch_issue(
    document: dict[str, Any],
    target_type: str,
    *,
    understanding_source: str,
    safety_routed: bool,
) -> list[dict[str, Any]]:
    if understanding_source != "llm" or safety_routed:
        return []
    expected = _DOCUMENT_TARGETS.get(str(document.get("document_type") or ""))
    if not expected or expected == target_type:
        return []
    if expected == "products" and target_type == "customer_products":
        return []
    return [
        {
            "code": "ETL_DOCUMENT_TARGET_MISMATCH",
            "severity": "error",
            "field": "",
            "message": f"识别为 {document.get('document_type')}，与当前导入目标 {target_type} 不一致",
        }
    ]


def parse_workbook_with_document_plan(
    path: str | Path,
    *,
    target_type: str,
    document_plan: dict[str, Any],
    max_rows: int,
) -> ParsedDataset | None:
    documents = document_plan.get("documents") or []
    if not documents:
        return None
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    routing_scope = document_plan.get("routing_scope") or {}
    safety_routed = bool(
        routing_scope.get("route_reason")
        and routing_scope.get("route_reason") != "document_type_default"
        and str(routing_scope.get("target_type") or "") == target_type
    )
    rows: list[ParsedRow] = []
    headers_all: list[str] = []
    warnings: list[dict[str, Any]] = []
    parsed_documents: list[dict[str, Any]] = []
    regions: list[dict[str, Any]] = []
    try:
        for document in documents:
            sheet_name = str(document.get("sheet") or "")
            if sheet_name not in workbook.sheetnames:
                warnings.append(
                    {
                        "code": "ETL_DOCUMENT_PLAN_SHEET_MISSING",
                        "message": f"结构计划引用了不存在的工作表：{sheet_name}",
                    }
                )
                continue
            worksheet = workbook[sheet_name]
            document_rows: list[ParsedRow] = []
            seen_source_rows: set[int] = set()
            duplicate_source_rows: set[int] = set()
            amount_values: list[Decimal] = []
            metadata = _header_values(document)
            for source_header in metadata:
                if source_header not in headers_all:
                    headers_all.append(source_header)
            for table_index, table in enumerate(document.get("tables") or [], start=1):
                headers = _composed_headers(worksheet, table)
                for header in headers:
                    if header not in headers_all:
                        headers_all.append(header)
                roles = _column_roles(table, headers)
                first_column = max(1, int(table.get("first_column") or 1))
                last_column = max(first_column, int(table.get("last_column") or first_column))
                data_start = max(1, int(table.get("data_start_row") or 1))
                data_end = min(
                    int(worksheet.max_row or data_start),
                    max(
                        data_start,
                        int(table.get("data_end_row") or worksheet.max_row or data_start),
                    ),
                )
                for row_number, row_values in enumerate(
                    worksheet.iter_rows(
                        min_row=data_start,
                        max_row=data_end,
                        min_col=first_column,
                        max_col=last_column,
                        values_only=True,
                    ),
                    start=data_start,
                ):
                    raw_values = list(row_values)
                    if not any(value not in (None, "") for value in raw_values):
                        continue
                    if is_repeated_header(raw_values, headers) or is_footer_or_note_row(raw_values):
                        continue
                    if row_number in seen_source_rows:
                        duplicate_source_rows.add(row_number)
                        continue
                    detail = {
                        headers[index]: value
                        for index, value in enumerate(raw_values)
                        if value not in (None, "")
                    }
                    if not detail:
                        continue
                    seen_source_rows.add(row_number)
                    values = {**metadata, **detail}
                    issues = [
                        *_target_mismatch_issue(
                            document,
                            target_type,
                            understanding_source=str(document_plan.get("source") or ""),
                            safety_routed=safety_routed,
                        ),
                        *_row_arithmetic_issues(
                            values,
                            roles,
                            target_type=target_type,
                        ),
                    ]
                    amount = _decimal(values.get(roles.get("amount", "")))
                    if amount is not None:
                        amount_values.append(amount)
                    if len(rows) >= max_rows:
                        raise EtlError(
                            "ETL_ROW_LIMIT_EXCEEDED",
                            f"文件超过 {max_rows} 行限制",
                            status_code=413,
                        )
                    parsed = ParsedRow(
                        sheet=sheet_name,
                        row_number=row_number,
                        values=values,
                        provenance={
                            "sheet": sheet_name,
                            "row": row_number,
                            "document_id": document.get("document_id"),
                            "document_type": document.get("document_type"),
                            "table_index": table_index,
                            "original_fragment": dict(zip(headers, raw_values, strict=False)),
                            "document_header": metadata,
                            "header_fields": document.get("header_fields") or [],
                            "columns": {
                                headers[index]: first_column + index
                                for index in range(len(headers))
                            },
                            "header_rows": {
                                "start": int(table.get("header_start_row") or 0),
                                "end": int(table.get("header_end_row") or 0),
                            },
                            "validation_issues": issues,
                        },
                    )
                    rows.append(parsed)
                    document_rows.append(parsed)
            if duplicate_source_rows:
                first_row = min(duplicate_source_rows)
                last_row = max(duplicate_source_rows)
                warnings.append(
                    {
                        "code": "ETL_OVERLAPPING_DOCUMENT_TABLE_ROWS_SKIPPED",
                        "severity": "warning",
                        "message": (
                            f"{sheet_name} 的明细区间存在重叠，已按物理行去重"
                            f"（{first_row}-{last_row}）"
                        ),
                    }
                )
            expected_total = _decimal(document.get("total_amount"))
            if expected_total is not None and amount_values:
                actual_total = sum(amount_values, Decimal("0"))
                tolerance = max(Decimal("0.01"), abs(expected_total) * Decimal("0.0001"))
                if abs(actual_total - expected_total) > tolerance:
                    issue = {
                        "code": "ETL_DOCUMENT_TOTAL_MISMATCH",
                        "severity": (
                            "warning"
                            if target_type in {"export_xlsx", "export_csv"}
                            else "error"
                        ),
                        "field": "",
                        "message": f"明细合计 {actual_total} 与单据总额 {expected_total} 不一致",
                    }
                    for parsed in document_rows:
                        parsed.provenance.setdefault("validation_issues", []).append(issue)
            parsed_documents.append(
                {
                    "document_id": document.get("document_id"),
                    "document_type": document.get("document_type"),
                    "sheet": sheet_name,
                    "row_count": len(document_rows),
                }
            )
            if str(document.get("document_type") or "") == "delivery_note":
                tables = [
                    table
                    for table in document.get("tables") or []
                    if isinstance(table, dict)
                ]
                if tables:
                    first_table = tables[0]
                    regions.append(
                        {
                            "id": str(
                                (document_plan.get("routing_scope") or {}).get("route_id")
                                or document.get("document_id")
                                or f"delivery:{sheet_name}"
                            ),
                            "status": "selected",
                            "role": "delivery_note_template_and_records",
                            "sheet": sheet_name,
                            "header_row": int(first_table.get("header_end_row") or 0),
                            "last_column": int(first_table.get("last_column") or 0),
                            "headers": _composed_headers(worksheet, first_table),
                            "customer_name": _party_name(
                                _header_role_value(document, "customer")
                            ),
                            "order_number": str(
                                _header_role_value(document, "document_number") or ""
                            ).strip(),
                            "row_count": len(document_rows),
                        }
                    )
    finally:
        workbook.close()
    return ParsedDataset(
        headers=headers_all,
        rows=rows,
        source_features={
            "kind": "workbook",
            "headers": headers_all,
            "structure_detection": "evidence_document_plan_v1",
            "documents": parsed_documents,
            "regions": regions,
        },
        warnings=warnings,
    )


__all__ = ["parse_workbook_with_document_plan"]
