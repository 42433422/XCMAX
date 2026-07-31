"""Workbook execution for multi-region customer/product parsing."""

from __future__ import annotations

from app.application.etl.parser_region_finalize import finalize_region_dataset
from app.utils.mixin_module_sync import sync_module_functions


def parse_customer_product_regions(
    path: Path,
    *,
    max_rows: int,
    target_type: str = "customer_products",
) -> ParsedDataset | None:
    """Parse explicit buyer + product-table regions without treating the whole sheet as one table."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    workbook_sheet_names = [str(name) for name in workbook.sheetnames]
    parsed_rows: list[ParsedRow] = []
    all_headers: list[str] = []
    regions: list[dict[str, Any]] = []
    probes: list[dict[str, Any]] = []
    skipped_sheets: list[dict[str, Any]] = []
    excluded_charge_rows: list[str] = []
    imported_by_sheet: dict[str, int] = {}
    history_product_count = 0
    companion_sheet_counts: dict[str, int] = {}
    companion_candidate_count = 0
    companion_stale_records_skipped = 0
    future_dated_source_rows: list[dict[str, Any]] = []
    same_date_source_conflicts = 0
    model_identity_ambiguity_count = 0
    sheet_domain_hints: dict[str, str] = {}
    try:
        for worksheet in workbook.worksheets:
            recent: deque[tuple[int, tuple[Any, ...]]] = deque(maxlen=5)
            active: dict[str, Any] | None = None
            sheet_candidates = 0
            for row_number, raw_values in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                values = tuple(raw_values)
                candidate = _header_candidate(values)
                if candidate is not None:
                    active = None
                    sheet_candidates += 1
                    context_rows = list(recent)
                    meta = _extract_meta(
                        context_rows,
                        max_col=int(candidate["last_col"]),
                    )
                    role = _region_role(
                        sheet_name=worksheet.title,
                        context_rows=context_rows,
                        meta=meta,
                        header=candidate,
                    )
                    region_id = (
                        f"{worksheet.title}!R{row_number}"
                        f"C{candidate['first_col']}:{candidate['last_col']}"
                    )
                    probe = {
                        "region_id": region_id,
                        "sheet": worksheet.title,
                        "header_row": row_number,
                        "headers": list(candidate["headers"]),
                        "context_rows": [
                            {
                                "row": number,
                                "text": _joined_row(
                                    row,
                                    max_col=int(candidate["last_col"]),
                                ),
                            }
                            for number, row in context_rows[-5:]
                            if _joined_row(
                                row,
                                max_col=int(candidate["last_col"]),
                            )
                        ],
                        "deterministic_role": role,
                        "explicit_customer": str(meta.get("customer_name") or ""),
                    }
                    probes.append(probe)
                    region = {
                        "id": region_id,
                        "sheet": worksheet.title,
                        "role": role,
                        "header_row": row_number,
                        "first_column": candidate["first_col"],
                        "last_column": candidate["last_col"],
                        "headers": list(candidate["headers"]),
                        "customer_name": str(meta.get("customer_name") or ""),
                        "contact_person": str(meta.get("contact_person") or ""),
                        "order_number": str(meta.get("order_number") or ""),
                        "order_date": str(meta.get("order_date") or ""),
                        "evidence_rows": list(meta.get("evidence_rows") or []),
                        "row_count": 0,
                        "status": "selected" if role == "delivery_note" else "excluded",
                    }
                    regions.append(region)
                    if role == "delivery_note" and meta.get("customer_name"):
                        active = {
                            "region": region,
                            "meta": meta,
                            "mapping": dict(candidate["by_field"]),
                            "source_by_col": _unique_source_headers(
                                dict(candidate["source_by_col"])
                            ),
                            "max_col": candidate["last_col"],
                        }
                    recent.append((row_number, values))
                    continue

                if active is not None:
                    if _is_total_row(values, max_col=int(active["max_col"])):
                        active = None
                        recent.append((row_number, values))
                        continue
                    mapping = active["mapping"]
                    name = clean_cell_text(_value_at(values, mapping.get("name")))
                    model = clean_cell_text(_value_at(values, mapping.get("model_number")))
                    if (name or model) and _has_measure(values, mapping):
                        business_name = name or model
                        if _NON_PRODUCT_RE.fullmatch(business_name):
                            excluded_charge_rows.append(f"{worksheet.title}:{row_number}")
                        else:
                            raw_values: dict[str, Any] = {
                                "customer_name": active["meta"]["customer_name"],
                            }
                            if active["meta"].get("contact_person"):
                                raw_values["contact_person"] = active["meta"]["contact_person"]
                            if active["meta"].get("contact_phone"):
                                raw_values["contact_phone"] = active["meta"]["contact_phone"]
                            original_fragment: dict[str, Any] = {}
                            columns: dict[str, int] = {}
                            for field, column in mapping.items():
                                value = _value_at(values, column)
                                if value not in (None, ""):
                                    raw_values[field] = value
                                    original_fragment[
                                        active["source_by_col"].get(column, field)
                                    ] = value
                                    columns[field] = column
                            source_values = project_delivery_region(
                                raw_values,
                                target_type=target_type,
                                meta=active["meta"],
                            )
                            if len(parsed_rows) >= max_rows:
                                from app.application.etl.errors import EtlError

                                raise EtlError(
                                    "ETL_ROW_LIMIT_EXCEEDED",
                                    f"文件超过 {max_rows} 行限制",
                                    status_code=413,
                                )
                            for header in source_values:
                                if header not in all_headers:
                                    all_headers.append(header)
                            region = active["region"]
                            parsed_rows.append(
                                ParsedRow(
                                    sheet=worksheet.title,
                                    row_number=row_number,
                                    values=source_values,
                                    provenance={
                                        "sheet": worksheet.title,
                                        "row": row_number,
                                        "source_kind": "delivery_note_region",
                                        "region_id": region["id"],
                                        "header_rows": {
                                            "start": region["header_row"],
                                            "end": region["header_row"],
                                        },
                                        "table_position": {
                                            "row": row_number,
                                            "first_column": region["first_column"],
                                            "last_column": region["last_column"],
                                        },
                                        "meta_evidence": list(
                                            active["meta"].get("evidence_rows") or []
                                        ),
                                        "external_order_no": active["meta"].get("order_number"),
                                        "order_date": active["meta"].get("order_date"),
                                        # Make selected delivery-note lines
                                        # comparable with appendix ledgers.
                                        # ``order_date`` remains untouched for
                                        # audit/display, while ``source_date``
                                        # is the normalized latest-record key.
                                        "source_date": _normalized_order_date(
                                            active["meta"].get("order_date")
                                        ),
                                        "original_fragment": original_fragment,
                                        "columns": columns,
                                    },
                                )
                            )
                            region["row_count"] += 1
                            imported_by_sheet[worksheet.title] = (
                                imported_by_sheet.get(worksheet.title, 0) + 1
                            )
                recent.append((row_number, values))
            if not imported_by_sheet.get(worksheet.title):
                skipped_sheets.append(
                    {
                        "name": worksheet.title,
                        "reason": (
                            "non_target_regions"
                            if sheet_candidates
                            else "no_explicit_customer_product_region"
                        ),
                    }
                )
        if target_type == "shipment_records":
            _attach_delivery_fingerprints(parsed_rows)
        # Every multi-sheet delivery workbook receives a deterministic appendix
        # plan.  A shipment-record preview deliberately leaves those supporting
        # rows out of the write target, while the linked customer-product
        # preview can reuse the exact same discovery result after the user asks
        # for it from the UI.
        sheet_domain_hints = {
            worksheet.title: _sheet_domain_hint(worksheet) for worksheet in workbook.worksheets
        }
        canonical_candidates: dict[str, set[str]] = {}
        for region in regions:
            if region.get("status") != "selected":
                continue
            customer_name = str(region.get("customer_name") or "").strip()
            alias = customer_alias_key(customer_name)
            if alias and customer_name:
                canonical_candidates.setdefault(alias, set()).add(customer_name)
        canonical_by_alias = {
            alias: next(iter(names))
            for alias, names in canonical_candidates.items()
            if len(names) == 1
        }
        # Customer/product previews are a current-fact view, not a replay of
        # every delivery line.  Start with selected delivery-note rows and let
        # newer supporting ledgers/quotes replace them by *business date*.
        # Shipment-record previews retain their actual line list unchanged.
        customer_product_latest: dict[tuple[str, str, str], ParsedRow] = {}
        same_date_conflict_rows: list[ParsedRow] = []
        if target_type == "customer_products":
            for row in parsed_rows:
                key = product_match_key(row.values)
                current = customer_product_latest.get(key)
                if current is None:
                    customer_product_latest[key] = row
                elif _same_date_conflict(row, current):
                    _mark_same_date_conflict(row, current)
                    same_date_conflict_rows.extend((current, row))
                    same_date_source_conflicts += 1
                elif _prefer_newer_companion(row, current):
                    customer_product_latest[key] = row
                    companion_stale_records_skipped += 1
                else:
                    companion_stale_records_skipped += 1

        existing_keys = {product_match_key(row.values) for row in parsed_rows}
        history_latest: dict[tuple[str, str, str], ParsedRow] = {}
        delivery_sheets = {
            str(region.get("sheet") or "")
            for region in regions
            if region.get("status") == "selected"
        }
        for worksheet in workbook.worksheets:
            if worksheet.title in delivery_sheets:
                continue
            if sheet_domain_hints.get(worksheet.title) == "finance_or_reconciliation":
                continue
            remaining = (
                max_rows - len(customer_product_latest)
                if target_type == "customer_products"
                else max_rows - len(history_latest)
            )
            if remaining <= 0:
                break
            companion_rows = parse_shipment_history_rows(
                worksheet,
                canonical_by_alias=canonical_by_alias,
                max_rows=remaining,
            )
            if not companion_rows:
                companion_rows = parse_structured_shipment_history_rows(
                    worksheet,
                    canonical_by_alias=canonical_by_alias,
                    max_rows=remaining,
                )
            if not companion_rows:
                companion_rows = parse_quote_rows(
                    worksheet,
                    canonical_by_alias=canonical_by_alias,
                    max_rows=remaining,
                )
            if companion_rows:
                companion_sheet_counts[worksheet.title] = len(companion_rows)
            for row in companion_rows:
                if _is_future_companion(row):
                    future_dated_source_rows.append(
                        {
                            "sheet": row.sheet,
                            "row": row.row_number,
                            "source_date": _companion_source_date(row),
                        }
                    )
                    continue
                key = product_match_key(row.values)
                if target_type == "customer_products":
                    existing = customer_product_latest.get(key)
                    if existing is None:
                        customer_product_latest[key] = row
                    elif _same_date_conflict(row, existing):
                        _mark_same_date_conflict(row, existing)
                        same_date_conflict_rows.extend((existing, row))
                        same_date_source_conflicts += 1
                    elif _prefer_newer_companion(row, existing):
                        customer_product_latest[key] = row
                        companion_stale_records_skipped += 1
                    else:
                        companion_stale_records_skipped += 1
                else:
                    if key in existing_keys:
                        continue
                    existing = history_latest.get(key)
                    if existing is None:
                        history_latest[key] = row
                    elif _prefer_newer_companion(row, existing):
                        history_latest[key] = row
                        companion_stale_records_skipped += 1
                    else:
                        companion_stale_records_skipped += 1

        companion_candidate_count = len(history_latest)
        if target_type == "customer_products":
            selected_rows: list[ParsedRow] = []
            seen_rows: set[tuple[str, int]] = set()
            for row in [*customer_product_latest.values(), *same_date_conflict_rows]:
                row_key = (str(row.sheet), int(row.row_number))
                if row_key in seen_rows:
                    continue
                seen_rows.add(row_key)
                selected_rows.append(row)
            parsed_rows = selected_rows
            history_rows = [
                row
                for row in parsed_rows
                if str(row.provenance.get("source_kind") or "")
                in {"shipment_history_ledger", "structured_shipment_history", "customer_quote"}
            ]
            history_product_count = len(history_rows)
            companion_candidate_count = history_product_count
            for row in parsed_rows:
                for header in row.values:
                    if header not in all_headers:
                        all_headers.append(header)
            if history_rows:
                for sheet in {row.sheet for row in history_rows}:
                    imported_by_sheet[sheet] = imported_by_sheet.get(sheet, 0) + sum(
                        row.sheet == sheet for row in history_rows
                    )

    finally:
        workbook.close()

    return finalize_region_dataset(
        parsed_rows=parsed_rows,
        target_type=target_type,
        probes=probes,
        regions=regions,
        excluded_charge_rows=excluded_charge_rows,
        future_dated_source_rows=future_dated_source_rows,
        same_date_source_conflicts=same_date_source_conflicts,
        history_product_count=history_product_count,
        model_identity_ambiguity_count=model_identity_ambiguity_count,
        companion_stale_records_skipped=companion_stale_records_skipped,
        companion_candidate_count=companion_candidate_count,
        workbook_sheet_names=workbook_sheet_names,
        companion_sheet_counts=companion_sheet_counts,
        sheet_domain_hints=sheet_domain_hints,
        skipped_sheets=skipped_sheets,
        all_headers=all_headers,
    )


sync_module_functions(
    target=globals(),
    source_module="app.application.etl.parser_regions",
    function_names=("parse_customer_product_regions",),
)
