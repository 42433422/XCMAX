# ruff: noqa
# mypy: ignore-errors
"""Customer/product multi-region parsing orchestration."""
from __future__ import annotations
import importlib

def _facade():
    return importlib.import_module('app.application.etl.parser_regions')

def parse_customer_product_regions(path: _facade().Path, *, max_rows: int, target_type: str='customer_products') -> _facade().ParsedDataset | None:
    """Parse explicit buyer + product-table regions without treating the whole sheet as one table."""
    from openpyxl import load_workbook
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    workbook_sheet_names = [str(name) for name in workbook.sheetnames]
    parsed_rows: list[_facade().ParsedRow] = []
    all_headers: list[str] = []
    regions: list[dict[str, _facade().Any]] = []
    probes: list[dict[str, _facade().Any]] = []
    skipped_sheets: list[dict[str, _facade().Any]] = []
    excluded_charge_rows: list[str] = []
    imported_by_sheet: dict[str, int] = {}
    history_product_count = 0
    companion_sheet_counts: dict[str, int] = {}
    companion_candidate_count = 0
    companion_stale_records_skipped = 0
    future_dated_source_rows: list[dict[str, _facade().Any]] = []
    same_date_source_conflicts = 0
    model_identity_ambiguity_count = 0
    sheet_domain_hints: dict[str, str] = {}
    try:
        for worksheet in workbook.worksheets:
            recent: _facade().deque[tuple[int, tuple[_facade().Any, ...]]] = _facade().deque(maxlen=5)
            active: dict[str, _facade().Any] | None = None
            sheet_candidates = 0
            for (row_number, raw_values) in enumerate(worksheet.iter_rows(values_only=True), start=1):
                values = tuple(raw_values)
                candidate = _facade()._header_candidate(values)
                if candidate is not None:
                    active = None
                    sheet_candidates += 1
                    context_rows = list(recent)
                    meta = _facade()._extract_meta(context_rows, max_col=int(candidate['last_col']))
                    role = _facade()._region_role(sheet_name=worksheet.title, context_rows=context_rows, meta=meta, header=candidate)
                    region_id = f"{worksheet.title}!R{row_number}C{candidate['first_col']}:{candidate['last_col']}"
                    probe = {'region_id': region_id, 'sheet': worksheet.title, 'header_row': row_number, 'headers': list(candidate['headers']), 'context_rows': [{'row': number, 'text': _facade()._joined_row(row, max_col=int(candidate['last_col']))} for (number, row) in context_rows[-5:] if _facade()._joined_row(row, max_col=int(candidate['last_col']))], 'deterministic_role': role, 'explicit_customer': str(meta.get('customer_name') or '')}
                    probes.append(probe)
                    region = {'id': region_id, 'sheet': worksheet.title, 'role': role, 'header_row': row_number, 'first_column': candidate['first_col'], 'last_column': candidate['last_col'], 'headers': list(candidate['headers']), 'customer_name': str(meta.get('customer_name') or ''), 'contact_person': str(meta.get('contact_person') or ''), 'order_number': str(meta.get('order_number') or ''), 'order_date': str(meta.get('order_date') or ''), 'evidence_rows': list(meta.get('evidence_rows') or []), 'row_count': 0, 'status': 'selected' if role == 'delivery_note' else 'excluded'}
                    regions.append(region)
                    if role == 'delivery_note' and meta.get('customer_name'):
                        active = {'region': region, 'meta': meta, 'mapping': dict(candidate['by_field']), 'source_by_col': _facade()._unique_source_headers(dict(candidate['source_by_col'])), 'max_col': candidate['last_col']}
                    recent.append((row_number, values))
                    continue
                if active is not None:
                    if _facade()._is_total_row(values, max_col=int(active['max_col'])):
                        active = None
                        recent.append((row_number, values))
                        continue
                    mapping = active['mapping']
                    name = _facade().clean_cell_text(_facade()._value_at(values, mapping.get('name')))
                    model = _facade().clean_cell_text(_facade()._value_at(values, mapping.get('model_number')))
                    if (name or model) and _facade()._has_measure(values, mapping):
                        business_name = name or model
                        if _facade()._NON_PRODUCT_RE.fullmatch(business_name):
                            excluded_charge_rows.append(f'{worksheet.title}:{row_number}')
                        else:
                            projected_input: dict[str, _facade().Any] = {'customer_name': active['meta']['customer_name']}
                            if active['meta'].get('contact_person'):
                                projected_input['contact_person'] = active['meta']['contact_person']
                            if active['meta'].get('contact_phone'):
                                projected_input['contact_phone'] = active['meta']['contact_phone']
                            original_fragment: dict[str, _facade().Any] = {}
                            columns: dict[str, int] = {}
                            for (field, column) in mapping.items():
                                value = _facade()._value_at(values, column)
                                if value not in (None, ''):
                                    projected_input[field] = value
                                    original_fragment[active['source_by_col'].get(column, field)] = value
                                    columns[field] = column
                            source_values = _facade().project_delivery_region(projected_input, target_type=target_type, meta=active['meta'])
                            if len(parsed_rows) >= max_rows:
                                from app.application.etl.errors import EtlError
                                raise EtlError('ETL_ROW_LIMIT_EXCEEDED', f'文件超过 {max_rows} 行限制', status_code=413)
                            for header in source_values:
                                if header not in all_headers:
                                    all_headers.append(header)
                            region = active['region']
                            parsed_rows.append(_facade().ParsedRow(sheet=worksheet.title, row_number=row_number, values=source_values, provenance={'sheet': worksheet.title, 'row': row_number, 'source_kind': 'delivery_note_region', 'region_id': region['id'], 'header_rows': {'start': region['header_row'], 'end': region['header_row']}, 'table_position': {'row': row_number, 'first_column': region['first_column'], 'last_column': region['last_column']}, 'meta_evidence': list(active['meta'].get('evidence_rows') or []), 'external_order_no': active['meta'].get('order_number'), 'order_date': active['meta'].get('order_date'), 'source_date': _facade()._normalized_order_date(active['meta'].get('order_date')), 'original_fragment': original_fragment, 'columns': columns}))
                            region['row_count'] += 1
                            imported_by_sheet[worksheet.title] = imported_by_sheet.get(worksheet.title, 0) + 1
                recent.append((row_number, values))
            if not imported_by_sheet.get(worksheet.title):
                skipped_sheets.append({'name': worksheet.title, 'reason': 'non_target_regions' if sheet_candidates else 'no_explicit_customer_product_region'})
        if target_type == 'shipment_records':
            _facade()._attach_delivery_fingerprints(parsed_rows)
        sheet_domain_hints = {worksheet.title: _facade()._sheet_domain_hint(worksheet) for worksheet in workbook.worksheets}
        canonical_candidates: dict[str, set[str]] = {}
        for region in regions:
            if region.get('status') != 'selected':
                continue
            customer_name = str(region.get('customer_name') or '').strip()
            alias = _facade().customer_alias_key(customer_name)
            if alias and customer_name:
                canonical_candidates.setdefault(alias, set()).add(customer_name)
        canonical_by_alias = {alias: next(iter(names)) for (alias, names) in canonical_candidates.items() if len(names) == 1}
        customer_product_latest: dict[tuple[str, str, str], _facade().ParsedRow] = {}
        same_date_conflict_rows: list[_facade().ParsedRow] = []
        if target_type == 'customer_products':
            for row in parsed_rows:
                key = _facade().product_match_key(row.values)
                current = customer_product_latest.get(key)
                if current is None:
                    customer_product_latest[key] = row
                elif _facade()._same_date_conflict(row, current):
                    _facade()._mark_same_date_conflict(row, current)
                    same_date_conflict_rows.extend((current, row))
                    same_date_source_conflicts += 1
                elif _facade()._prefer_newer_companion(row, current):
                    customer_product_latest[key] = row
                    companion_stale_records_skipped += 1
                else:
                    companion_stale_records_skipped += 1
        existing_keys = {_facade().product_match_key(row.values) for row in parsed_rows}
        history_latest: dict[tuple[str, str, str], _facade().ParsedRow] = {}
        delivery_sheets = {str(region.get('sheet') or '') for region in regions if region.get('status') == 'selected'}
        for worksheet in workbook.worksheets:
            if worksheet.title in delivery_sheets:
                continue
            if sheet_domain_hints.get(worksheet.title) == 'finance_or_reconciliation':
                continue
            remaining = max_rows - len(customer_product_latest) if target_type == 'customer_products' else max_rows - len(history_latest)
            if remaining <= 0:
                break
            companion_rows = _facade().parse_shipment_history_rows(worksheet, canonical_by_alias=canonical_by_alias, max_rows=remaining)
            if not companion_rows:
                companion_rows = _facade().parse_structured_shipment_history_rows(worksheet, canonical_by_alias=canonical_by_alias, max_rows=remaining)
            if not companion_rows:
                companion_rows = _facade().parse_quote_rows(worksheet, canonical_by_alias=canonical_by_alias, max_rows=remaining)
            if companion_rows:
                companion_sheet_counts[worksheet.title] = len(companion_rows)
            for row in companion_rows:
                if _facade()._is_future_companion(row):
                    future_dated_source_rows.append({'sheet': row.sheet, 'row': row.row_number, 'source_date': _facade()._companion_source_date(row)})
                    continue
                key = _facade().product_match_key(row.values)
                if target_type == 'customer_products':
                    existing = customer_product_latest.get(key)
                    if existing is None:
                        customer_product_latest[key] = row
                    elif _facade()._same_date_conflict(row, existing):
                        _facade()._mark_same_date_conflict(row, existing)
                        same_date_conflict_rows.extend((existing, row))
                        same_date_source_conflicts += 1
                    elif _facade()._prefer_newer_companion(row, existing):
                        customer_product_latest[key] = row
                        companion_stale_records_skipped += 1
                    else:
                        companion_stale_records_skipped += 1
                else:
                    if key in existing_keys:
                        continue
                    existing = history_latest.get(key)
                    if existing is None:
                        history_latest[key] = row
                    elif _facade()._prefer_newer_companion(row, existing):
                        history_latest[key] = row
                        companion_stale_records_skipped += 1
                    else:
                        companion_stale_records_skipped += 1
        companion_candidate_count = len(history_latest)
        if target_type == 'customer_products':
            selected_rows: list[_facade().ParsedRow] = []
            seen_rows: set[tuple[str, int]] = set()
            for row in [*customer_product_latest.values(), *same_date_conflict_rows]:
                row_key = (str(row.sheet), int(row.row_number))
                if row_key in seen_rows:
                    continue
                seen_rows.add(row_key)
                selected_rows.append(row)
            parsed_rows = selected_rows
            history_rows = [row for row in parsed_rows if str(row.provenance.get('source_kind') or '') in {'shipment_history_ledger', 'structured_shipment_history', 'customer_quote'}]
            history_product_count = len(history_rows)
            companion_candidate_count = history_product_count
            for row in parsed_rows:
                for header in row.values:
                    if header not in all_headers:
                        all_headers.append(header)
            if history_rows:
                for sheet in {row.sheet for row in history_rows}:
                    imported_by_sheet[sheet] = imported_by_sheet.get(sheet, 0) + sum((row.sheet == sheet for row in history_rows))
    finally:
        workbook.close()
    if not parsed_rows:
        return None
    if target_type == 'customer_products':
        identity_issues = _facade().source_model_ambiguity_issues([row.values for row in parsed_rows], unit_field='customer_name')
        if identity_issues:
            ambiguous_keys: set[tuple[str, str]] = set()
            for (index, issues) in identity_issues.items():
                row = parsed_rows[index]
                row.provenance.setdefault('validation_issues', []).extend(issues)
                ambiguous_keys.add(_facade().product_name_key(row.values, unit_field='customer_name'))
            model_identity_ambiguity_count = len(ambiguous_keys)
    llm = _facade().advise_workbook_regions(probes)
    llm_by_region = {str(item.get('region_id') or ''): item for item in list(llm.data.get('regions') or []) if isinstance(item, dict)}
    for region in regions:
        suggestion = llm_by_region.get(region['id'])
        if suggestion:
            region['llm_suggestion'] = suggestion
    selected = [region for region in regions if region['status'] == 'selected']
    excluded = [region for region in regions if region['status'] == 'excluded']
    target_label = '发货单' if target_type == 'shipment_records' else '客户产品'
    warnings: list[dict[str, _facade().Any]] = [{'code': 'ETL_MULTI_REGION_WORKBOOK_PLANNED', 'message': f'已从混合工作簿识别 {len(selected)} 个{target_label}业务区块，排除 {len(excluded)} 个其他业务区块。', 'selected_regions': len(selected), 'excluded_regions': len(excluded)}]
    if excluded_charge_rows:
        warnings.append({'code': 'ETL_NON_PRODUCT_CHARGES_SKIPPED', 'message': f'已跳过 {len(excluded_charge_rows)} 行运费等非产品费用。', 'count': len(excluded_charge_rows), 'rows': excluded_charge_rows[:50]})
    if future_dated_source_rows:
        warnings.append({'code': 'ETL_FUTURE_DATED_SOURCE_ROW', 'message': f'已隔离 {len(future_dated_source_rows)} 条日期晚于当前日期的历史/报价记录，它们不会被当作最新客户产品事实。', 'count': len(future_dated_source_rows), 'rows': future_dated_source_rows[:50]})
    if same_date_source_conflicts:
        warnings.append({'code': 'ETL_LATEST_SOURCE_CONFLICT', 'message': f'发现 {same_date_source_conflicts} 组同日跨表产品事实冲突，已保留为错误行，需人工确认。', 'count': same_date_source_conflicts})
    if history_product_count:
        warnings.append({'code': 'ETL_SHIPMENT_HISTORY_PRODUCTS_INCLUDED', 'message': f'已从出货历史增补 {history_product_count} 个客户产品候选，仅用于客户产品预演，不会作为新的发货记录写入。', 'count': history_product_count})
    if model_identity_ambiguity_count:
        warnings.append({'code': 'ETL_PRODUCT_MODEL_AMBIGUITY', 'message': f'发现 {model_identity_ambiguity_count} 组同客户同产品同时有型号和无型号的数据，已标为错误，需补全型号或拆分后重新预演。', 'count': model_identity_ambiguity_count})
    if companion_stale_records_skipped:
        warnings.append({'code': 'ETL_LATEST_PRODUCT_DATA_SELECTED', 'message': f'同一客户同一产品存在 {companion_stale_records_skipped} 条较早或同日旧记录，已按来源日期优先保留最新有效数据。', 'count': companion_stale_records_skipped})
    if companion_candidate_count and target_type == 'shipment_records':
        warnings.append({'code': 'ETL_COMPANION_CUSTOMER_PRODUCT_DATA_FOUND', 'message': f'已在附表发现 {companion_candidate_count} 个客户产品候选。它们不会写入当前发货记录预演；可新建客户及产品预演后确认。', 'count': companion_candidate_count})
    sheet_plan = _facade()._build_sheet_plan(workbook_sheet_names=workbook_sheet_names, regions=regions, companion_sheet_counts=companion_sheet_counts, sheet_domain_hints=sheet_domain_hints)
    return _facade().ParsedDataset(headers=all_headers, rows=parsed_rows, source_features={'structure_detection': 'deterministic_regions_v1', **_facade().region_source_features(target_type=target_type, regions=regions, rows=len(parsed_rows)), 'regions': regions, 'shipment_history_product_candidates': companion_candidate_count, 'latest_record_selection': {'basis': 'source_date_then_same_sheet_row', 'unique_candidates': companion_candidate_count, 'stale_records_skipped': companion_stale_records_skipped, 'future_dated_records_skipped': len(future_dated_source_rows), 'same_date_conflicts': same_date_source_conflicts, 'model_identity_ambiguity_groups': model_identity_ambiguity_count}, 'sheet_plan': sheet_plan, 'skipped_sheets': skipped_sheets, 'headers': all_headers, 'llm_structure': {**llm.public_metadata(), 'suggestion_count': len(llm_by_region)}}, warnings=warnings)
