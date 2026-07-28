"""Deterministic multi-region workbook parser for linked customer/product data."""

from __future__ import annotations

import hashlib
import json
import re
from collections import deque
from datetime import date
from pathlib import Path
from typing import Any

from app.application.etl.llm_assist import advise_workbook_regions
from app.application.etl.parser_region_projection import (
    project_delivery_region,
    region_source_features,
)
from app.application.etl.parser_shipment_history import (
    customer_alias_key,
    parse_quote_rows,
    parse_shipment_history_rows,
    parse_structured_shipment_history_rows,
    product_match_key,
)
from app.application.etl.parser_structure import clean_cell_text, semantic_key
from app.application.etl.parser_types import ParsedDataset, ParsedRow
from app.application.etl.product_identity import (
    product_name_key,
    source_model_ambiguity_issues,
)

_FINANCE_RE = re.compile(r"(回款|付款|收款|对账|欠款|余额|账龄|明细账)", re.I)
_FINANCE_CONTENT_RE = re.compile(
    r"(对账|回款|收款金额|付款明细|期末欠款|应收款|账龄|明细账|余额)",
    re.I,
)
_CATALOG_RE = re.compile(r"(报价|价格|价目|色漆编号|色号|样品|目录)", re.I)
_DELIVERY_RE = re.compile(r"(送货|发货|出货|delivery)", re.I)
_TOTAL_RE = re.compile(r"^(合计|总计|小计|汇总)(?:\s|[:：]|$)", re.I)
_NON_PRODUCT_RE = re.compile(
    r"^(运费|物流费|配送费|快递费|装卸费|服务费|折扣|优惠|税费)$",
    re.I,
)
_BUYER_LABEL_RE = re.compile(
    r"(?:(?:购货单位|购买单位|采购单位)(?:（[^）]*）|\([^)]*\))?"
    r"|(?:客户(?:名称)?|买方|buyer|bill\s*to|sold\s*to))"
    r"\s*[:：]\s*(.+?)(?=\s+(?:联系人|电话|手机|日期|订单|单号|编号|20\d{2})|$)",
    re.I,
)
_CONTACT_RE = re.compile(
    r"(?:联系人|经手人)\s*[:：]\s*(.+?)(?=\s+(?:电话|手机|日期|订单|单号|编号|20\d{2})|$)",
    re.I,
)
_PHONE_RE = re.compile(r"(?:电话|手机|联系方式)\s*[:：]\s*([+\d][\d\s-]{5,})", re.I)
_ORDER_RE = re.compile(
    r"(?:订单编号|订单号|单号|order\s*(?:no|number)?)\s*[:：#]?\s*([A-Za-z0-9][\w./-]*)",
    re.I,
)
_DATE_RE = re.compile(r"((?:19|20)\d{2}[年./-]\d{1,2}[月./-]\d{1,2}日?)")


def _field_for_header(value: Any) -> str:
    key = semantic_key(value)
    if not key:
        return ""
    if any(token in key for token in ("产品型号", "货品型号", "商品型号", "model", "sku")):
        return "model_number"
    if key in {"型号", "编号", "产品编号", "货号", "物料号", "partcode"}:
        return "model_number"
    if any(token in key for token in ("产品名称", "货品名称", "商品名称", "goodstitle")):
        return "name"
    if key in {"品名", "名称", "产品", "货品"}:
        return "name"
    if "数量kg" in key or "总重量" in key or "重量kg" in key:
        return "quantity_kg"
    if any(token in key for token in ("数量件", "数量桶", "件数", "桶数", "pcscount")):
        return "quantity_tins"
    if key == "数量":
        return "quantity_tins"
    if key.startswith("规格") or key in {"净重", "每桶kg", "netkgeach"}:
        return "specification"
    if any(token in key for token in ("单价", "价格", "现金价", "售价", "unitfee")):
        return "price"
    if any(token in key for token in ("金额", "价税合计", "linesum")):
        return "amount"
    if key in {"备注", "说明", "remark", "description"}:
        return "description"
    return ""


def _header_candidate(values: tuple[Any, ...] | list[Any]) -> dict[str, Any] | None:
    by_field: dict[str, int] = {}
    source_by_col: dict[int, str] = {}
    for index, value in enumerate(values, start=1):
        text = clean_cell_text(value)
        if not text:
            continue
        field = _field_for_header(text)
        if field and field not in by_field:
            by_field[field] = index
            source_by_col[index] = text[:160]
    identities = sum(field in by_field for field in ("model_number", "name"))
    commerce = sum(
        field in by_field
        for field in ("quantity_tins", "quantity_kg", "specification", "price", "amount")
    )
    if not (identities >= 1 and commerce >= 1) and not identities >= 2:
        return None
    if "name" not in by_field and "model_number" not in by_field:
        return None
    first_col = min(source_by_col)
    last_col = max(source_by_col)
    return {
        "by_field": by_field,
        "source_by_col": source_by_col,
        "first_col": first_col,
        "last_col": last_col,
        "headers": [source_by_col[col] for col in sorted(source_by_col)],
        "identity_count": identities,
        "commerce_count": commerce,
    }


def _joined_row(values: tuple[Any, ...] | list[Any], *, max_col: int = 30) -> str:
    return " ".join(text for value in list(values)[:max_col] if (text := clean_cell_text(value)))[
        :2000
    ]


def _extract_meta(
    context_rows: list[tuple[int, tuple[Any, ...]]],
    *,
    max_col: int,
) -> dict[str, Any]:
    result: dict[str, Any] = {
        "customer_name": "",
        "contact_person": "",
        "contact_phone": "",
        "order_number": "",
        "order_date": "",
        "evidence_rows": [],
    }
    for row_number, values in reversed(context_rows[-5:]):
        text = _joined_row(values, max_col=max_col)
        if not text:
            continue
        matched = False
        if not result["customer_name"] and (match := _BUYER_LABEL_RE.search(text)):
            result["customer_name"] = clean_cell_text(match.group(1))[:160]
            matched = True
        if not result["contact_person"] and (match := _CONTACT_RE.search(text)):
            result["contact_person"] = clean_cell_text(match.group(1))[:160]
            matched = True
        if not result["contact_phone"] and (match := _PHONE_RE.search(text)):
            result["contact_phone"] = re.sub(r"\s+", "", match.group(1))[:80]
            matched = True
        if not result["order_number"] and (match := _ORDER_RE.search(text)):
            result["order_number"] = clean_cell_text(match.group(1))[:160]
            matched = True
        if not result["order_date"] and (match := _DATE_RE.search(text)):
            result["order_date"] = clean_cell_text(match.group(1))[:80]
            matched = True
        if matched or _DELIVERY_RE.search(text):
            result["evidence_rows"].append({"row": row_number, "text": text[:800]})
    result["evidence_rows"].reverse()
    return result


def _region_role(
    *,
    sheet_name: str,
    context_rows: list[tuple[int, tuple[Any, ...]]],
    meta: dict[str, Any],
    header: dict[str, Any],
) -> str:
    context = " ".join(
        _joined_row(values, max_col=int(header["last_col"])) for _row, values in context_rows[-4:]
    )
    combined = f"{sheet_name} {context}"
    if _FINANCE_RE.search(combined):
        return "finance"
    if meta.get("customer_name") and (
        _DELIVERY_RE.search(combined) or header["commerce_count"] >= 2
    ):
        return "delivery_note"
    if _CATALOG_RE.search(combined):
        return "product_catalog"
    if _DELIVERY_RE.search(combined):
        return "shipment_ledger"
    return "ignore"


def _is_total_row(values: tuple[Any, ...] | list[Any], *, max_col: int) -> bool:
    cells = [clean_cell_text(value) for value in list(values)[:max_col] if clean_cell_text(value)]
    if not cells:
        return False
    first = semantic_key(cells[0])
    return first in {"合计", "总计", "小计", "汇总"} or bool(_TOTAL_RE.match(" ".join(cells[:2])))


def _build_sheet_plan(
    *,
    workbook_sheet_names: list[str],
    regions: list[dict[str, Any]],
    companion_sheet_counts: dict[str, int],
    sheet_domain_hints: dict[str, str],
) -> list[dict[str, Any]]:
    selected_by_sheet: dict[str, int] = {}
    for region in regions:
        if region.get("status") == "selected":
            sheet = str(region.get("sheet") or "")
            selected_by_sheet[sheet] = selected_by_sheet.get(sheet, 0) + int(
                region.get("row_count") or 0
            )
    plan: list[dict[str, Any]] = []
    for sheet in workbook_sheet_names:
        if sheet in selected_by_sheet:
            plan.append(
                {
                    "sheet": sheet,
                    "role": "delivery_note_template_and_records",
                    "status": "included",
                    "rows": selected_by_sheet[sheet],
                    "reason": "识别到购货单位、产品表头与合计行",
                }
            )
        elif sheet in companion_sheet_counts:
            plan.append(
                {
                    "sheet": sheet,
                    "role": "supporting_customer_product_data",
                    "status": "included",
                    "rows": companion_sheet_counts[sheet],
                    "reason": "识别到高置信出货历史或客户报价",
                }
            )
        elif sheet_domain_hints.get(sheet) == "finance_or_reconciliation" or _FINANCE_RE.search(
            sheet
        ):
            plan.append(
                {
                    "sheet": sheet,
                    "role": "finance_or_reconciliation",
                    "status": "excluded",
                    "rows": 0,
                    "reason": "财务/对账附表不写入客户产品或发货记录",
                }
            )
        elif sheet_domain_hints.get(sheet) == "reference_catalog" or _CATALOG_RE.search(sheet):
            plan.append(
                {
                    "sheet": sheet,
                    "role": "reference_catalog",
                    "status": "reviewed",
                    "rows": 0,
                    "reason": "已读取为参考目录，缺少可安全写入的客户或价格关系",
                }
            )
        else:
            plan.append(
                {
                    "sheet": sheet,
                    "role": "non_target_appendix",
                    "status": "excluded",
                    "rows": 0,
                    "reason": "未满足可安全写入的业务字段条件",
                }
            )
    return plan


def _sheet_domain_hint(worksheet: Any) -> str:
    """Classify an appendix by content when its tab name is generic (for example ``Sheet1``)."""
    text = " ".join(
        clean_cell_text(value)
        for row in worksheet.iter_rows(max_row=20, max_col=18, values_only=True)
        for value in row
        if clean_cell_text(value)
    )[:8000]
    # A delivery ledger often has a one-off “欠款” note.  Treat an unnamed
    # sheet as finance only when the content carries a stronger finance
    # signature (for example “对账单” or “收款金额”), rather than excluding
    # otherwise valid product lines because of that incidental note.
    if _FINANCE_RE.search(worksheet.title) or _FINANCE_CONTENT_RE.search(text):
        return "finance_or_reconciliation"
    if _CATALOG_RE.search(f"{worksheet.title} {text}"):
        return "reference_catalog"
    return ""


def _companion_source_date(row: ParsedRow) -> str:
    provenance = row.provenance if isinstance(row.provenance, dict) else {}
    value = str(provenance.get("source_date") or "").strip()
    # Parsers emit ISO dates. Keep an unparseable value out of ordering so a
    # malformed cell cannot masquerade as a newer business record.
    return value if re.fullmatch(r"(?:19|20)\d{2}-\d{2}-\d{2}", value) else ""


def _normalized_order_date(value: Any) -> str:
    """Normalize a delivery-note date into the same ordering evidence as ledgers."""

    text = clean_cell_text(value)
    match = _DATE_RE.search(text)
    if not match:
        return ""
    normalized = match.group(1).replace("年", "-").replace("月", "-").replace("日", "")
    try:
        return date.fromisoformat(normalized).isoformat()
    except ValueError:
        return ""


def _is_future_companion(row: ParsedRow) -> bool:
    source_date = _companion_source_date(row)
    if not source_date:
        return False
    try:
        return date.fromisoformat(source_date) > date.today()
    except ValueError:
        return False


def _same_date_conflict(candidate: ParsedRow, current: ParsedRow) -> bool:
    """Detect contradictory same-day facts from different sheets.

    A workbook tab order is not a business rule.  Preserve both candidates as
    blocking rows when the same customer/product/date disagrees on a value
    that would affect the product default or a printed document.
    """

    candidate_date = _companion_source_date(candidate)
    current_date = _companion_source_date(current)
    if not candidate_date or candidate_date != current_date or candidate.sheet == current.sheet:
        return False
    keys = ("model_number", "specification", "price")
    return any(
        str(candidate.values.get(key) or "").strip() != str(current.values.get(key) or "").strip()
        for key in keys
    )


def _mark_same_date_conflict(*rows: ParsedRow) -> None:
    issue = {
        "code": "ETL_LATEST_SOURCE_CONFLICT",
        "field": "source_date",
        "severity": "error",
        "message": "同一客户产品在不同工作表同日出现冲突价格、规格或型号，需人工确认后再导入",
    }
    for row in rows:
        issues = row.provenance.setdefault("validation_issues", [])
        if not any(isinstance(item, dict) and item.get("code") == issue["code"] for item in issues):
            issues.append(dict(issue))


def _prefer_newer_companion(candidate: ParsedRow, current: ParsedRow) -> bool:
    """Choose only evidenced newer data, never workbook-tab order by accident."""
    candidate_date = _companion_source_date(candidate)
    current_date = _companion_source_date(current)
    if candidate_date and not current_date:
        return True
    if not candidate_date or candidate_date != current_date:
        return bool(candidate_date and candidate_date > current_date)
    # Same dated ledger: a later source row is a deterministic correction only
    # within that same sheet. Between sheets we keep the first record rather
    # than silently inventing an ordering the workbook never supplied.
    return candidate.sheet == current.sheet and candidate.row_number > current.row_number


def _attach_delivery_fingerprints(rows: list[ParsedRow]) -> None:
    """Add old shipment-ETL compatible note and line fingerprints.

    The generic region parser is now preferred for delivery notes, but existing
    imports may have been recorded by the former shipment-profile parser.  Use
    the same content shape for its note fingerprint so a re-import stays
    idempotent across the upgrade; line fingerprints remain unique per source
    row within that note.
    """

    def number(value: Any) -> float:
        try:
            return float(value or 0)
        except (TypeError, ValueError):
            return 0.0

    groups: dict[str, list[ParsedRow]] = {}
    for row in rows:
        region_id = str(row.provenance.get("region_id") or "").strip()
        if region_id:
            groups.setdefault(region_id, []).append(row)
    for group_rows in groups.values():
        ordered = sorted(group_rows, key=lambda item: item.row_number)
        first = ordered[0]
        items = [
            {
                "m": str(row.values.get("model_number") or "").strip().upper(),
                "n": str(row.values.get("product_name") or "").strip(),
                "q": number(row.values.get("quantity_tins")),
                "k": number(row.values.get("quantity_kg")),
                "p": number(row.values.get("unit_price")),
            }
            for row in ordered
        ]
        note_payload = {
            "unit": str(first.values.get("purchase_unit") or "").strip(),
            "order": str(first.values.get("external_order_no") or "").strip(),
            "date": str(first.provenance.get("order_date") or "").strip(),
            "items": sorted(
                items, key=lambda item: (item["m"], item["n"], item["q"], item["k"], item["p"])
            ),
        }
        note_raw = json.dumps(
            note_payload, ensure_ascii=False, sort_keys=True, separators=(",", ":")
        )
        note_fingerprint = hashlib.sha256(note_raw.encode("utf-8")).hexdigest()[:28]
        for item_index, row in enumerate(ordered, start=1):
            line_payload = {
                "note": note_fingerprint,
                "index": item_index,
                "model": row.values.get("model_number"),
                "name": row.values.get("product_name"),
                "kg": row.values.get("quantity_kg"),
                "tins": row.values.get("quantity_tins"),
                "price": row.values.get("unit_price"),
            }
            row.values["legacy_note_fingerprint"] = note_fingerprint
            row.values["source_fingerprint"] = hashlib.sha256(
                json.dumps(
                    line_payload,
                    ensure_ascii=False,
                    sort_keys=True,
                    default=str,
                ).encode("utf-8")
            ).hexdigest()


def _value_at(values: tuple[Any, ...] | list[Any], column: int | None) -> Any:
    if not column or column < 1 or column > len(values):
        return None
    return values[column - 1]


def _has_measure(values: tuple[Any, ...] | list[Any], mapping: dict[str, int]) -> bool:
    for field in ("quantity_tins", "quantity_kg", "specification", "price", "amount"):
        value = _value_at(values, mapping.get(field))
        if value not in (None, ""):
            return True
    return False


def _unique_source_headers(source_by_col: dict[int, str]) -> dict[int, str]:
    seen: dict[str, int] = {}
    result: dict[int, str] = {}
    for column in sorted(source_by_col):
        base = source_by_col[column]
        seen[base] = seen.get(base, 0) + 1
        result[column] = base if seen[base] == 1 else f"{base}_{seen[base]}"
    return result


def parse_customer_product_regions(
    path: Path,
    *,
    max_rows: int,
    target_type: str = "customer_products",
) -> ParsedDataset | None:
    """Parse explicit buyer + product-table regions without treating the whole sheet as one table."""
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    workbook_sheet_names = [str(name) for name in workbook.sheetnames]
    parsed_rows: list[ParsedRow] = []
    all_headers: list[str] = []
    regions: list[dict[str, Any]] = []
    probes: list[dict[str, Any]] = []
    skipped_sheets: list[dict[str, Any]] = []
    excluded_charge_rows: list[str] = []
    imported_by_sheet: dict[str, int] = {}
    history_product_count = 0
    companion_sheet_counts: dict[str, int] = {}
    companion_candidate_count = 0
    companion_stale_records_skipped = 0
    future_dated_source_rows: list[dict[str, Any]] = []
    same_date_source_conflicts = 0
    model_identity_ambiguity_count = 0
    sheet_domain_hints: dict[str, str] = {}
    try:
        for worksheet in workbook.worksheets:
            recent: deque[tuple[int, tuple[Any, ...]]] = deque(maxlen=5)
            active: dict[str, Any] | None = None
            sheet_candidates = 0
            for row_number, raw_values in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                values = tuple(raw_values)
                candidate = _header_candidate(values)
                if candidate is not None:
                    active = None
                    sheet_candidates += 1
                    context_rows = list(recent)
                    meta = _extract_meta(
                        context_rows,
                        max_col=int(candidate["last_col"]),
                    )
                    role = _region_role(
                        sheet_name=worksheet.title,
                        context_rows=context_rows,
                        meta=meta,
                        header=candidate,
                    )
                    region_id = (
                        f"{worksheet.title}!R{row_number}"
                        f"C{candidate['first_col']}:{candidate['last_col']}"
                    )
                    probe = {
                        "region_id": region_id,
                        "sheet": worksheet.title,
                        "header_row": row_number,
                        "headers": list(candidate["headers"]),
                        "context_rows": [
                            {
                                "row": number,
                                "text": _joined_row(
                                    row,
                                    max_col=int(candidate["last_col"]),
                                ),
                            }
                            for number, row in context_rows[-5:]
                            if _joined_row(
                                row,
                                max_col=int(candidate["last_col"]),
                            )
                        ],
                        "deterministic_role": role,
                        "explicit_customer": str(meta.get("customer_name") or ""),
                    }
                    probes.append(probe)
                    region = {
                        "id": region_id,
                        "sheet": worksheet.title,
                        "role": role,
                        "header_row": row_number,
                        "first_column": candidate["first_col"],
                        "last_column": candidate["last_col"],
                        "headers": list(candidate["headers"]),
                        "customer_name": str(meta.get("customer_name") or ""),
                        "contact_person": str(meta.get("contact_person") or ""),
                        "order_number": str(meta.get("order_number") or ""),
                        "order_date": str(meta.get("order_date") or ""),
                        "evidence_rows": list(meta.get("evidence_rows") or []),
                        "row_count": 0,
                        "status": "selected" if role == "delivery_note" else "excluded",
                    }
                    regions.append(region)
                    if role == "delivery_note" and meta.get("customer_name"):
                        active = {
                            "region": region,
                            "meta": meta,
                            "mapping": dict(candidate["by_field"]),
                            "source_by_col": _unique_source_headers(
                                dict(candidate["source_by_col"])
                            ),
                            "max_col": candidate["last_col"],
                        }
                    recent.append((row_number, values))
                    continue

                if active is not None:
                    if _is_total_row(values, max_col=int(active["max_col"])):
                        active = None
                        recent.append((row_number, values))
                        continue
                    mapping = active["mapping"]
                    name = clean_cell_text(_value_at(values, mapping.get("name")))
                    model = clean_cell_text(_value_at(values, mapping.get("model_number")))
                    if (name or model) and _has_measure(values, mapping):
                        business_name = name or model
                        if _NON_PRODUCT_RE.fullmatch(business_name):
                            excluded_charge_rows.append(f"{worksheet.title}:{row_number}")
                        else:
                            raw_values: dict[str, Any] = {
                                "customer_name": active["meta"]["customer_name"],
                            }
                            if active["meta"].get("contact_person"):
                                raw_values["contact_person"] = active["meta"]["contact_person"]
                            if active["meta"].get("contact_phone"):
                                raw_values["contact_phone"] = active["meta"]["contact_phone"]
                            original_fragment: dict[str, Any] = {}
                            columns: dict[str, int] = {}
                            for field, column in mapping.items():
                                value = _value_at(values, column)
                                if value not in (None, ""):
                                    raw_values[field] = value
                                    original_fragment[
                                        active["source_by_col"].get(column, field)
                                    ] = value
                                    columns[field] = column
                            source_values = project_delivery_region(
                                raw_values,
                                target_type=target_type,
                                meta=active["meta"],
                            )
                            if len(parsed_rows) >= max_rows:
                                from app.application.etl.errors import EtlError

                                raise EtlError(
                                    "ETL_ROW_LIMIT_EXCEEDED",
                                    f"文件超过 {max_rows} 行限制",
                                    status_code=413,
                                )
                            for header in source_values:
                                if header not in all_headers:
                                    all_headers.append(header)
                            region = active["region"]
                            parsed_rows.append(
                                ParsedRow(
                                    sheet=worksheet.title,
                                    row_number=row_number,
                                    values=source_values,
                                    provenance={
                                        "sheet": worksheet.title,
                                        "row": row_number,
                                        "source_kind": "delivery_note_region",
                                        "region_id": region["id"],
                                        "header_rows": {
                                            "start": region["header_row"],
                                            "end": region["header_row"],
                                        },
                                        "table_position": {
                                            "row": row_number,
                                            "first_column": region["first_column"],
                                            "last_column": region["last_column"],
                                        },
                                        "meta_evidence": list(
                                            active["meta"].get("evidence_rows") or []
                                        ),
                                        "external_order_no": active["meta"].get("order_number"),
                                        "order_date": active["meta"].get("order_date"),
                                        # Make selected delivery-note lines
                                        # comparable with appendix ledgers.
                                        # ``order_date`` remains untouched for
                                        # audit/display, while ``source_date``
                                        # is the normalized latest-record key.
                                        "source_date": _normalized_order_date(
                                            active["meta"].get("order_date")
                                        ),
                                        "original_fragment": original_fragment,
                                        "columns": columns,
                                    },
                                )
                            )
                            region["row_count"] += 1
                            imported_by_sheet[worksheet.title] = (
                                imported_by_sheet.get(worksheet.title, 0) + 1
                            )
                recent.append((row_number, values))
            if not imported_by_sheet.get(worksheet.title):
                skipped_sheets.append(
                    {
                        "name": worksheet.title,
                        "reason": (
                            "non_target_regions"
                            if sheet_candidates
                            else "no_explicit_customer_product_region"
                        ),
                    }
                )
        if target_type == "shipment_records":
            _attach_delivery_fingerprints(parsed_rows)
        # Every multi-sheet delivery workbook receives a deterministic appendix
        # plan.  A shipment-record preview deliberately leaves those supporting
        # rows out of the write target, while the linked customer-product
        # preview can reuse the exact same discovery result after the user asks
        # for it from the UI.
        sheet_domain_hints = {
            worksheet.title: _sheet_domain_hint(worksheet) for worksheet in workbook.worksheets
        }
        canonical_candidates: dict[str, set[str]] = {}
        for region in regions:
            if region.get("status") != "selected":
                continue
            customer_name = str(region.get("customer_name") or "").strip()
            alias = customer_alias_key(customer_name)
            if alias and customer_name:
                canonical_candidates.setdefault(alias, set()).add(customer_name)
        canonical_by_alias = {
            alias: next(iter(names))
            for alias, names in canonical_candidates.items()
            if len(names) == 1
        }
        # Customer/product previews are a current-fact view, not a replay of
        # every delivery line.  Start with selected delivery-note rows and let
        # newer supporting ledgers/quotes replace them by *business date*.
        # Shipment-record previews retain their actual line list unchanged.
        customer_product_latest: dict[tuple[str, str, str], ParsedRow] = {}
        same_date_conflict_rows: list[ParsedRow] = []
        if target_type == "customer_products":
            for row in parsed_rows:
                key = product_match_key(row.values)
                current = customer_product_latest.get(key)
                if current is None:
                    customer_product_latest[key] = row
                elif _same_date_conflict(row, current):
                    _mark_same_date_conflict(row, current)
                    same_date_conflict_rows.extend((current, row))
                    same_date_source_conflicts += 1
                elif _prefer_newer_companion(row, current):
                    customer_product_latest[key] = row
                    companion_stale_records_skipped += 1
                else:
                    companion_stale_records_skipped += 1

        existing_keys = {product_match_key(row.values) for row in parsed_rows}
        history_latest: dict[tuple[str, str, str], ParsedRow] = {}
        delivery_sheets = {
            str(region.get("sheet") or "")
            for region in regions
            if region.get("status") == "selected"
        }
        for worksheet in workbook.worksheets:
            if worksheet.title in delivery_sheets:
                continue
            if sheet_domain_hints.get(worksheet.title) == "finance_or_reconciliation":
                continue
            remaining = (
                max_rows - len(customer_product_latest)
                if target_type == "customer_products"
                else max_rows - len(history_latest)
            )
            if remaining <= 0:
                break
            companion_rows = parse_shipment_history_rows(
                worksheet,
                canonical_by_alias=canonical_by_alias,
                max_rows=remaining,
            )
            if not companion_rows:
                companion_rows = parse_structured_shipment_history_rows(
                    worksheet,
                    canonical_by_alias=canonical_by_alias,
                    max_rows=remaining,
                )
            if not companion_rows:
                companion_rows = parse_quote_rows(
                    worksheet,
                    canonical_by_alias=canonical_by_alias,
                    max_rows=remaining,
                )
            if companion_rows:
                companion_sheet_counts[worksheet.title] = len(companion_rows)
            for row in companion_rows:
                if _is_future_companion(row):
                    future_dated_source_rows.append(
                        {
                            "sheet": row.sheet,
                            "row": row.row_number,
                            "source_date": _companion_source_date(row),
                        }
                    )
                    continue
                key = product_match_key(row.values)
                if target_type == "customer_products":
                    existing = customer_product_latest.get(key)
                    if existing is None:
                        customer_product_latest[key] = row
                    elif _same_date_conflict(row, existing):
                        _mark_same_date_conflict(row, existing)
                        same_date_conflict_rows.extend((existing, row))
                        same_date_source_conflicts += 1
                    elif _prefer_newer_companion(row, existing):
                        customer_product_latest[key] = row
                        companion_stale_records_skipped += 1
                    else:
                        companion_stale_records_skipped += 1
                else:
                    if key in existing_keys:
                        continue
                    existing = history_latest.get(key)
                    if existing is None:
                        history_latest[key] = row
                    elif _prefer_newer_companion(row, existing):
                        history_latest[key] = row
                        companion_stale_records_skipped += 1
                    else:
                        companion_stale_records_skipped += 1

        companion_candidate_count = len(history_latest)
        if target_type == "customer_products":
            selected_rows: list[ParsedRow] = []
            seen_rows: set[tuple[str, int]] = set()
            for row in [*customer_product_latest.values(), *same_date_conflict_rows]:
                row_key = (str(row.sheet), int(row.row_number))
                if row_key in seen_rows:
                    continue
                seen_rows.add(row_key)
                selected_rows.append(row)
            parsed_rows = selected_rows
            history_rows = [
                row
                for row in parsed_rows
                if str(row.provenance.get("source_kind") or "")
                in {"shipment_history_ledger", "structured_shipment_history", "customer_quote"}
            ]
            history_product_count = len(history_rows)
            companion_candidate_count = history_product_count
            for row in parsed_rows:
                for header in row.values:
                    if header not in all_headers:
                        all_headers.append(header)
            if history_rows:
                for sheet in {row.sheet for row in history_rows}:
                    imported_by_sheet[sheet] = imported_by_sheet.get(sheet, 0) + sum(
                        row.sheet == sheet for row in history_rows
                    )

    finally:
        workbook.close()

    if not parsed_rows:
        return None

    # A model-less price/history row is not safely interchangeable with an
    # otherwise identical modeled row.  Keep both source rows for review and
    # attach a blocking issue to each, rather than silently dropping newer data
    # or creating two products during a later preview.
    if target_type == "customer_products":
        identity_issues = source_model_ambiguity_issues(
            [row.values for row in parsed_rows],
            unit_field="customer_name",
        )
        if identity_issues:
            ambiguous_keys: set[tuple[str, str]] = set()
            for index, issues in identity_issues.items():
                row = parsed_rows[index]
                row.provenance.setdefault("validation_issues", []).extend(issues)
                ambiguous_keys.add(product_name_key(row.values, unit_field="customer_name"))
            model_identity_ambiguity_count = len(ambiguous_keys)

    llm = advise_workbook_regions(probes)
    llm_by_region = {
        str(item.get("region_id") or ""): item
        for item in list(llm.data.get("regions") or [])
        if isinstance(item, dict)
    }
    for region in regions:
        suggestion = llm_by_region.get(region["id"])
        if suggestion:
            region["llm_suggestion"] = suggestion

    selected = [region for region in regions if region["status"] == "selected"]
    excluded = [region for region in regions if region["status"] == "excluded"]
    target_label = "发货单" if target_type == "shipment_records" else "客户产品"
    warnings: list[dict[str, Any]] = [
        {
            "code": "ETL_MULTI_REGION_WORKBOOK_PLANNED",
            "message": (
                f"已从混合工作簿识别 {len(selected)} 个{target_label}业务区块，"
                f"排除 {len(excluded)} 个其他业务区块。"
            ),
            "selected_regions": len(selected),
            "excluded_regions": len(excluded),
        }
    ]
    if excluded_charge_rows:
        warnings.append(
            {
                "code": "ETL_NON_PRODUCT_CHARGES_SKIPPED",
                "message": f"已跳过 {len(excluded_charge_rows)} 行运费等非产品费用。",
                "count": len(excluded_charge_rows),
                "rows": excluded_charge_rows[:50],
            }
        )
    if future_dated_source_rows:
        warnings.append(
            {
                "code": "ETL_FUTURE_DATED_SOURCE_ROW",
                "message": (
                    f"已隔离 {len(future_dated_source_rows)} 条日期晚于当前日期的历史/报价记录，"
                    "它们不会被当作最新客户产品事实。"
                ),
                "count": len(future_dated_source_rows),
                "rows": future_dated_source_rows[:50],
            }
        )
    if same_date_source_conflicts:
        warnings.append(
            {
                "code": "ETL_LATEST_SOURCE_CONFLICT",
                "message": (
                    f"发现 {same_date_source_conflicts} 组同日跨表产品事实冲突，"
                    "已保留为错误行，需人工确认。"
                ),
                "count": same_date_source_conflicts,
            }
        )
    if history_product_count:
        warnings.append(
            {
                "code": "ETL_SHIPMENT_HISTORY_PRODUCTS_INCLUDED",
                "message": (
                    f"已从出货历史增补 {history_product_count} 个客户产品候选，"
                    "仅用于客户产品预演，不会作为新的发货记录写入。"
                ),
                "count": history_product_count,
            }
        )
    if model_identity_ambiguity_count:
        warnings.append(
            {
                "code": "ETL_PRODUCT_MODEL_AMBIGUITY",
                "message": (
                    f"发现 {model_identity_ambiguity_count} 组同客户同产品同时有型号和无型号的数据，"
                    "已标为错误，需补全型号或拆分后重新预演。"
                ),
                "count": model_identity_ambiguity_count,
            }
        )
    if companion_stale_records_skipped:
        warnings.append(
            {
                "code": "ETL_LATEST_PRODUCT_DATA_SELECTED",
                "message": (
                    f"同一客户同一产品存在 {companion_stale_records_skipped} 条较早或同日旧记录，"
                    "已按来源日期优先保留最新有效数据。"
                ),
                "count": companion_stale_records_skipped,
            }
        )
    elif companion_candidate_count and target_type == "shipment_records":
        warnings.append(
            {
                "code": "ETL_COMPANION_CUSTOMER_PRODUCT_DATA_FOUND",
                "message": (
                    f"已在附表发现 {companion_candidate_count} 个客户产品候选。"
                    "它们不会写入当前发货记录预演；可新建客户及产品预演后确认。"
                ),
                "count": companion_candidate_count,
            }
        )
    sheet_plan = _build_sheet_plan(
        workbook_sheet_names=workbook_sheet_names,
        regions=regions,
        companion_sheet_counts=companion_sheet_counts,
        sheet_domain_hints=sheet_domain_hints,
    )
    return ParsedDataset(
        headers=all_headers,
        rows=parsed_rows,
        source_features={
            "structure_detection": "deterministic_regions_v1",
            **region_source_features(
                target_type=target_type,
                regions=regions,
                rows=len(parsed_rows),
            ),
            "regions": regions,
            "shipment_history_product_candidates": companion_candidate_count,
            "latest_record_selection": {
                "basis": "source_date_then_same_sheet_row",
                "unique_candidates": companion_candidate_count,
                "stale_records_skipped": companion_stale_records_skipped,
                "future_dated_records_skipped": len(future_dated_source_rows),
                "same_date_conflicts": same_date_source_conflicts,
                "model_identity_ambiguity_groups": model_identity_ambiguity_count,
            },
            "sheet_plan": sheet_plan,
            "skipped_sheets": skipped_sheets,
            "headers": all_headers,
            "llm_structure": {
                **llm.public_metadata(),
                "suggestion_count": len(llm_by_region),
            },
        },
        warnings=warnings,
    )


__all__ = ["parse_customer_product_regions"]
