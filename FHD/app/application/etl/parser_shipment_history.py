"""Recognize headerless shipment-history rows as customer-product candidates.

Many customer workbooks keep a delivery-note layout beside a long running ledger
without column headers.  The ledger is useful for product master data, but it
must never be mistaken for a new delivery-note execution target.
"""

from __future__ import annotations

import re
from datetime import date, datetime
from typing import Any

from app.application.etl.parser_structure import clean_cell_text, semantic_key
from app.application.etl.parser_types import ParsedRow

_SHIPMENT_SHEET_RE = re.compile(r"(?:出货|发货|送货)", re.I)
_FINANCE_SHEET_RE = re.compile(r"(?:回款|付款|收款|对账|欠款|余额|账龄|明细账)", re.I)
_CUSTOMER_SUFFIX_RE = re.compile(
    r"(?:有限责任公司|有限公司|公司|家私|家具|商贸|贸易|建材|装饰)", re.I
)
_PARENTHETICAL_RE = re.compile(r"[（(][^）)]*[）)]")
_DATE_TEXT_RE = re.compile(
    r"(?P<year>(?:19|20)\d{2})[年./-](?P<month>\d{1,2})(?:[月./-](?P<day>\d{1,2}))?"
)
# Notes such as "未签单" are often written in the model-number column of a
# running shipment ledger.  They describe the order, not a sellable SKU.  This
# list is deliberately narrow and only matches complete, well-known status
# annotations, so a real Chinese model number is still preserved for review.
_NON_MODEL_ANNOTATION_RE = re.compile(
    r"^(?:未(?:签单|下单|发货)|(?:单)?下错(?:了)?|(?:已)?作废|取消|无(?:型号|编号|产品型号))$",
    re.I,
)


def customer_alias_key(value: Any) -> str:
    """Return a conservative key for matching a ledger alias to one note customer."""
    text = clean_cell_text(value)
    text = _PARENTHETICAL_RE.sub("", text)
    text = _CUSTOMER_SUFFIX_RE.sub("", text)
    return re.sub(r"[\s\-_/\\·,.，。:：]+", "", text).lower()


def _number(value: Any) -> float | None:
    if isinstance(value, bool) or value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _source_date(value: Any) -> str:
    """Return an ISO business date from a cell when it is unambiguous.

    Historical shipment ledgers commonly use Excel serial dates, while price
    sheets put an effective date into a title.  We preserve that evidence so
    deduplication can choose the newest record instead of whichever tab happens
    to be parsed last.
    """

    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean_cell_text(value)
    if match := _DATE_TEXT_RE.search(text):
        try:
            return date(
                int(match.group("year")),
                int(match.group("month")),
                int(match.group("day") or 1),
            ).isoformat()
        except ValueError:
            return ""
    serial = _number(value)
    # Excel serial dates for modern business ledgers. Values such as product
    # model numbers (9803) and prices must not be treated as dates.
    if serial is None or not 30_000 <= serial <= 70_000:
        return ""
    try:
        from openpyxl.utils.datetime import from_excel

        converted = from_excel(serial)
        if isinstance(converted, datetime):
            return converted.date().isoformat()
        if isinstance(converted, date):
            return converted.isoformat()
    except (TypeError, ValueError, OverflowError):
        return ""
    return ""


def _source_date_from_values(values: tuple[Any, ...], *, max_columns: int = 4) -> tuple[str, int]:
    for index, value in enumerate(values[:max_columns], start=1):
        if source_date := _source_date(value):
            return source_date, index
    return "", 0


def _looks_like_product(value: Any) -> bool:
    text = clean_cell_text(value)
    if len(text) < 2 or len(text) > 160:
        return False
    return bool(re.search(r"[\u4e00-\u9fffA-Za-z]", text))


def _looks_like_model(value: Any) -> bool:
    text = clean_cell_text(value)
    if (
        not text
        or re.fullmatch(r"\d+号?", text)
        or _NON_MODEL_ANNOTATION_RE.fullmatch(text)
    ):
        return False
    return bool(re.search(r"[\u4e00-\u9fffA-Za-z0-9]", text))


def _nearby_model_number(values: tuple[Any, ...], product_index: int) -> str:
    """Return a defensible model immediately before a headerless product row.

    A plain number normally remains unsafe: historical ledgers often put a
    delivery sequence (for example ``1号``) next to the product.  Some real
    ledgers place a *numeric* product code one cell closer to the product,
    though (``1号 | 9804 | … | PE白底漆``).  Preserve that code only when the
    surrounding row independently proves the preceding field is an order
    marker.  This avoids collapsing separate numeric models into a single
    name-only product during latest-record selection.
    """

    nearby = [
        clean_cell_text(candidate)
        for candidate in reversed(values[max(1, product_index - 4) : product_index])
        if clean_cell_text(candidate)
    ]
    for offset, text in enumerate(nearby):
        if _looks_like_model(text):
            return text
        if (
            offset == 0
            and re.fullmatch(r"\d{3,12}", text)
            and any(re.fullmatch(r"\d{1,4}号", earlier) for earlier in nearby[1:])
        ):
            return text
    return ""


def _line_candidate(values: tuple[Any, ...]) -> dict[str, Any] | None:
    """Find ``name, tins, spec, kg, price, amount`` in a headerless row."""
    best: dict[str, Any] | None = None
    for index in range(len(values) - 5):
        product_name = clean_cell_text(values[index])
        if not _looks_like_product(product_name):
            continue
        quantity_tins = _number(values[index + 1])
        specification = _number(values[index + 2])
        quantity_kg = _number(values[index + 3])
        price = _number(values[index + 4])
        amount = _number(values[index + 5])
        if None in {quantity_tins, specification, quantity_kg, price, amount}:
            continue
        if quantity_tins <= 0 or specification <= 0 or quantity_kg <= 0 or price < 0:
            continue
        kg_error = abs(quantity_kg - quantity_tins * specification)
        amount_error = abs(amount - quantity_kg * price)
        if kg_error > max(2.0, abs(quantity_kg) * 0.03):
            continue
        if amount_error > max(5.0, abs(amount) * 0.03):
            continue
        model_number = _nearby_model_number(values, index)
        score = 10 + (2 if model_number else 0)
        candidate = {
            "product_name": product_name,
            "model_number": model_number,
            "quantity_tins": quantity_tins,
            "specification": specification,
            "quantity_kg": quantity_kg,
            "price": price,
            "amount": amount,
            "first_column": index + 1,
            "last_column": index + 6,
            "score": score,
        }
        if best is None or candidate["score"] > best["score"]:
            best = candidate
    return best


def _canonical_customer(raw_name: str, canonical_by_alias: dict[str, str]) -> str:
    alias = customer_alias_key(raw_name)
    return canonical_by_alias.get(alias, raw_name)


def parse_shipment_history_rows(
    worksheet: Any,
    *,
    canonical_by_alias: dict[str, str],
    max_rows: int,
) -> list[ParsedRow]:
    """Parse a high-confidence historical shipment sheet for product-master preview only."""
    if _FINANCE_SHEET_RE.search(worksheet.title):
        return []

    rows: list[ParsedRow] = []
    last_customer_raw = ""
    for row_number, raw_values in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = tuple(raw_values)
        explicit_customer = clean_cell_text(values[0] if values else "")
        if explicit_customer and _FINANCE_SHEET_RE.search(explicit_customer):
            continue
        candidate = _line_candidate(values)
        if candidate is None:
            continue
        if explicit_customer:
            last_customer_raw = explicit_customer
        customer_raw = explicit_customer or last_customer_raw
        if not customer_raw:
            continue
        source_date, source_date_column = _source_date_from_values(values)
        customer_name = _canonical_customer(customer_raw, canonical_by_alias)
        values_out = {
            "customer_name": customer_name,
            "name": candidate["product_name"],
            "specification": candidate["specification"],
            "price": candidate["price"],
        }
        if candidate["model_number"]:
            values_out["model_number"] = candidate["model_number"]
        rows.append(
            ParsedRow(
                sheet=worksheet.title,
                row_number=row_number,
                values=values_out,
                provenance={
                    "sheet": worksheet.title,
                    "row": row_number,
                    "source_kind": "shipment_history_ledger",
                    "original_fragment": {
                        "customer_name": customer_raw,
                        "model_number": candidate["model_number"],
                        "name": candidate["product_name"],
                        "quantity_tins": candidate["quantity_tins"],
                        "specification": candidate["specification"],
                        "quantity_kg": candidate["quantity_kg"],
                        "price": candidate["price"],
                        "amount": candidate["amount"],
                    },
                    "table_position": {
                        "row": row_number,
                        "first_column": candidate["first_column"],
                        "last_column": candidate["last_column"],
                    },
                    "customer_alias_source": customer_raw,
                    "source_date": source_date,
                    "source_date_column": source_date_column or None,
                },
            )
        )
        if len(rows) >= max_rows:
            break
    # A named delivery/history sheet may contain a small valid table.  An
    # unnamed companion sheet has to prove itself with several matching lines
    # so that a one-off note or reconciliation page cannot turn into product
    # master data by accident.
    if not _SHIPMENT_SHEET_RE.search(worksheet.title) and len(rows) < 3:
        return []
    return rows


def product_match_key(values: dict[str, Any]) -> tuple[str, str, str]:
    customer = str(values.get("customer_name") or "").strip()
    model = str(values.get("model_number") or "").strip()
    return (customer, "model" if model else "name", model or str(values.get("name") or "").strip())


def _header_field(value: Any) -> str:
    key = semantic_key(value)
    if key in {"编号", "型号", "产品编号", "产品型号", "货号"}:
        return "model_number"
    if key in {"品名", "名称", "产品", "产品名称", "货品名称", "商品名称"}:
        return "name"
    if key.startswith("规格") or key in {"净重", "每桶kg"}:
        return "specification"
    if key in {"数量", "数量件", "数量桶", "件数", "桶数"}:
        return "quantity_tins"
    if "单价" in key or "现金价" in key or key == "价格":
        return "price"
    if "金额" in key:
        return "amount"
    return ""


def _customer_column(values: tuple[Any, ...]) -> int | None:
    for index, value in enumerate(values, start=1):
        key = semantic_key(value)
        if key in {"客户", "客户名称", "购买单位", "购货单位", "家具厂名称", "家具厂"}:
            return index
    return None


def _field_mapping(values: tuple[Any, ...]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        field = _header_field(value)
        if field and field not in mapping:
            mapping[field] = index
    return mapping


def _date_column(values: tuple[Any, ...]) -> int | None:
    for index, value in enumerate(values, start=1):
        key = semantic_key(value)
        if key in {"日期", "出货日期", "发货日期", "送货日期", "订单日期"}:
            return index
    return None


def parse_structured_shipment_history_rows(
    worksheet: Any,
    *,
    canonical_by_alias: dict[str, str],
    max_rows: int,
) -> list[ParsedRow]:
    """Read a headed shipment-history appendix with an explicit customer column."""
    if _FINANCE_SHEET_RE.search(worksheet.title):
        return []
    header_row = 0
    mapping: dict[str, int] = {}
    customer_column: int | None = None
    date_column: int | None = None
    for row_number, raw_values in enumerate(worksheet.iter_rows(max_row=60, values_only=True), start=1):
        values = tuple(raw_values)
        fields = _field_mapping(values)
        customer = _customer_column(values)
        if customer and {"name", "specification", "quantity_tins", "price"} <= set(fields):
            header_row = row_number
            mapping = fields
            customer_column = customer
            date_column = _date_column(values)
            break
    if not header_row or customer_column is None:
        return []

    rows: list[ParsedRow] = []
    for row_number, raw_values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        values = tuple(raw_values)
        customer_raw = clean_cell_text(values[customer_column - 1] if len(values) >= customer_column else "")
        product_name = clean_cell_text(values[mapping["name"] - 1] if len(values) >= mapping["name"] else "")
        quantity = _number(values[mapping["quantity_tins"] - 1] if len(values) >= mapping["quantity_tins"] else "")
        specification = _number(values[mapping["specification"] - 1] if len(values) >= mapping["specification"] else "")
        price = _number(values[mapping["price"] - 1] if len(values) >= mapping["price"] else "")
        if not customer_raw or not _looks_like_product(product_name):
            continue
        if quantity is None or quantity <= 0 or specification is None or specification <= 0 or price is None:
            continue
        model_number = clean_cell_text(
            values[mapping["model_number"] - 1]
            if mapping.get("model_number") and len(values) >= mapping["model_number"]
            else ""
        )
        values_out: dict[str, Any] = {
            "customer_name": _canonical_customer(customer_raw, canonical_by_alias),
            "name": product_name,
            "specification": specification,
            "price": price,
        }
        if _looks_like_model(model_number):
            values_out["model_number"] = model_number
        source_date = _source_date(
            values[date_column - 1] if date_column and len(values) >= date_column else ""
        )
        rows.append(
            ParsedRow(
                sheet=worksheet.title,
                row_number=row_number,
                values=values_out,
                provenance={
                    "sheet": worksheet.title,
                    "row": row_number,
                    "source_kind": "structured_shipment_history",
                    "header_rows": {"start": header_row, "end": header_row},
                    "customer_alias_source": customer_raw,
                    "source_date": source_date,
                    "source_date_column": date_column,
                },
            )
        )
        if len(rows) >= max_rows:
            break
    return rows


def _title_customer(worksheet: Any, canonical_by_alias: dict[str, str]) -> str:
    title = " ".join(
        clean_cell_text(value)
        for row in worksheet.iter_rows(max_row=3, values_only=True)
        for value in row
        if clean_cell_text(value)
    )
    normalized_title = customer_alias_key(title)
    matches = [
        customer
        for alias, customer in canonical_by_alias.items()
        if alias and alias in normalized_title
    ]
    return matches[0] if len(set(matches)) == 1 else ""


def parse_quote_rows(
    worksheet: Any,
    *,
    canonical_by_alias: dict[str, str],
    max_rows: int,
) -> list[ParsedRow]:
    """Read a customer-specific price appendix when its fields are explicit."""
    if "报价" not in worksheet.title and "报价" not in " ".join(
        clean_cell_text(value)
        for row in worksheet.iter_rows(max_row=3, values_only=True)
        for value in row
    ):
        return []
    customer_name = _title_customer(worksheet, canonical_by_alias)
    if not customer_name:
        return []
    header_row = 0
    mapping: dict[str, int] = {}
    for row_number, raw_values in enumerate(worksheet.iter_rows(max_row=30, values_only=True), start=1):
        candidate = _field_mapping(tuple(raw_values))
        if {"name", "specification", "price"} <= set(candidate):
            header_row, mapping = row_number, candidate
            break
    if not header_row:
        return []
    source_date, source_date_column = _source_date_from_values(
        tuple(
            value
            for row in worksheet.iter_rows(max_row=3, values_only=True)
            for value in row
        ),
        max_columns=80,
    )
    rows: list[ParsedRow] = []
    for row_number, raw_values in enumerate(
        worksheet.iter_rows(min_row=header_row + 1, values_only=True),
        start=header_row + 1,
    ):
        values = tuple(raw_values)
        name = clean_cell_text(values[mapping["name"] - 1] if len(values) >= mapping["name"] else "")
        specification = _number(values[mapping["specification"] - 1] if len(values) >= mapping["specification"] else "")
        price = _number(values[mapping["price"] - 1] if len(values) >= mapping["price"] else "")
        if not _looks_like_product(name) or specification is None or price is None:
            continue
        rows.append(
            ParsedRow(
                sheet=worksheet.title,
                row_number=row_number,
                values={
                    "customer_name": customer_name,
                    "name": name,
                    "specification": specification,
                    "price": price,
                },
                provenance={
                    "sheet": worksheet.title,
                    "row": row_number,
                    "source_kind": "customer_quote",
                    "header_rows": {"start": header_row, "end": header_row},
                    "source_date": source_date,
                    "source_date_column": source_date_column or None,
                },
            )
        )
        if len(rows) >= max_rows:
            break
    return rows


__all__ = [
    "customer_alias_key",
    "parse_quote_rows",
    "parse_shipment_history_rows",
    "parse_structured_shipment_history_rows",
    "product_match_key",
]
