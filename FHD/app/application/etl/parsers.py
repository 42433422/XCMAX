"""通用表格与 OCR 输入解析。"""

from __future__ import annotations

import csv
import itertools
from pathlib import Path
from typing import Any

from app.application.etl.errors import EtlError
from app.application.etl.parser_ocr_provenance import enrich_ocr_provenance
from app.application.etl.parser_structure import (
    detect_table_layout,
    header_match_score,
    is_auxiliary_sheet_name,
    is_footer_or_note_row,
    is_repeated_header,
)
from app.application.etl.parser_target_match import (
    covers_required_target_fields,
    target_header_hints,
)
from app.application.etl.parser_types import ParsedDataset, ParsedRow

MAX_ROWS = 100_000
STRUCTURED_SUFFIXES = {".xlsx", ".xlsm", ".csv"}
OCR_SUFFIXES = {".pdf", ".jpg", ".jpeg", ".png"}
KNOWLEDGE_ONLY_SUFFIXES = {".doc", ".docx", ".ppt", ".pptx"}
SUPPORTED_SUFFIXES = STRUCTURED_SUFFIXES | OCR_SUFFIXES | KNOWLEDGE_ONLY_SUFFIXES


def _aligned_headers_by_sheet(
    candidates: list[tuple[Any, Any]],
    target_type: str,
) -> dict[str, list[str]]:
    """Align equivalent aliases across sheets to one source key per target field."""
    from app.application.etl.targets import get_adapter

    adapter = get_adapter(target_type)
    if adapter.allow_dynamic_fields:
        return {
            worksheet.title: list(layout.headers)
            for worksheet, layout in candidates
            if layout is not None
        }
    canonical_by_field: dict[int, str] = {}
    result: dict[str, list[str]] = {}
    for worksheet, layout in candidates:
        if layout is None:
            continue
        headers = list(layout.headers)
        pairs = sorted(
            (
                (
                    header_match_score(
                        header,
                        (field.key, field.label, *field.aliases),
                    ),
                    field_index,
                    header_index,
                )
                for field_index, field in enumerate(adapter.fields)
                for header_index, header in enumerate(headers)
            ),
            reverse=True,
        )
        used_fields: set[int] = set()
        used_headers: set[int] = set()
        for score, field_index, header_index in pairs:
            # Only exact or contextual matches are safe to align automatically.
            if score < 0.9:
                break
            if field_index in used_fields or header_index in used_headers:
                continue
            canonical = canonical_by_field.setdefault(field_index, headers[header_index])
            headers[header_index] = canonical
            used_fields.add(field_index)
            used_headers.add(header_index)
        result[worksheet.title] = headers
    return result


def _parse_workbook(path: Path, max_rows: int, target_type: str) -> ParsedDataset:
    from openpyxl import load_workbook

    workbook = load_workbook(
        path,
        read_only=True,
        data_only=True,
        keep_links=False,
    )
    result: list[ParsedRow] = []
    all_headers: list[str] = []
    sheets: list[dict[str, Any]] = []
    skipped_sheets: list[dict[str, Any]] = []
    warnings: list[dict[str, Any]] = []
    repeated_header_count = 0
    footer_count = 0
    header_hints = target_header_hints(target_type)
    try:
        candidates: list[tuple[Any, Any]] = []
        for worksheet in workbook.worksheets:
            probe = list(worksheet.iter_rows(min_row=1, max_row=60, values_only=True))
            layout = detect_table_layout(probe, header_hints=header_hints)
            candidates.append((worksheet, layout))
        any_target_match = any(
            layout is not None and layout.matched_hint_count > 0 for _, layout in candidates
        )
        aligned_headers = _aligned_headers_by_sheet(candidates, target_type)

        for worksheet, layout in candidates:
            if layout is None:
                skipped_sheets.append({"name": worksheet.title, "reason": "no_tabular_header"})
                continue
            if (
                any_target_match
                and layout.matched_hint_count == 0
                and is_auxiliary_sheet_name(worksheet.title)
            ):
                skipped_sheets.append(
                    {
                        "name": worksheet.title,
                        "reason": "auxiliary_sheet_without_target_fields",
                    }
                )
                continue
            headers = aligned_headers.get(worksheet.title, layout.headers)
            header_aliases = {
                original: aligned
                for original, aligned in zip(layout.headers, headers, strict=False)
                if original != aligned
            }
            all_headers.extend(header for header in headers if header not in all_headers)
            sheet_count = 0
            for row_number, values in enumerate(
                worksheet.iter_rows(min_row=layout.header_end + 2, values_only=True),
                start=layout.header_end + 2,
            ):
                if is_repeated_header(values, layout.headers):
                    repeated_header_count += 1
                    continue
                if is_footer_or_note_row(values):
                    footer_count += 1
                    continue
                data = {
                    headers[index]: value
                    for index, value in enumerate(values[: len(headers)])
                    if value not in (None, "")
                }
                if not data:
                    continue
                if len(result) >= max_rows:
                    raise EtlError(
                        "ETL_ROW_LIMIT_EXCEEDED",
                        f"文件超过 {max_rows} 行限制",
                        status_code=413,
                    )
                result.append(
                    ParsedRow(
                        sheet=worksheet.title,
                        row_number=row_number,
                        values=data,
                        provenance={
                            "sheet": worksheet.title,
                            "row": row_number,
                            "original_fragment": {
                                layout.headers[index]: value
                                for index, value in enumerate(values[: len(headers)])
                            },
                            "header_aliases": header_aliases,
                            "columns": {headers[index]: index + 1 for index in range(len(headers))},
                            "header_rows": {
                                "start": layout.header_start + 1,
                                "end": layout.header_end + 1,
                            },
                        },
                    )
                )
                sheet_count += 1
            sheets.append(
                {
                    "name": worksheet.title,
                    "header_row": layout.header_end + 1,
                    "header_start_row": layout.header_start + 1,
                    "header_depth": layout.header_end - layout.header_start + 1,
                    "header_confidence": layout.confidence,
                    "header_reasons": list(layout.reasons),
                    "matched_target_headers": layout.matched_hint_count,
                    "source_headers": list(layout.headers),
                    "aligned_headers": list(headers),
                    "header_aliases": header_aliases,
                    "row_count": sheet_count,
                }
            )
    finally:
        workbook.close()
    if skipped_sheets:
        warnings.append(
            {
                "code": "ETL_AUXILIARY_SHEETS_SKIPPED",
                "message": f"已跳过 {len(skipped_sheets)} 个非业务或无法识别的工作表。",
                "sheets": skipped_sheets,
            }
        )
    if repeated_header_count:
        warnings.append(
            {
                "code": "ETL_REPEATED_HEADERS_SKIPPED",
                "message": f"已跳过 {repeated_header_count} 行重复表头。",
                "count": repeated_header_count,
            }
        )
    if footer_count:
        warnings.append(
            {
                "code": "ETL_FOOTER_ROWS_SKIPPED",
                "message": f"已跳过 {footer_count} 行合计、备注或签字尾行。",
                "count": footer_count,
            }
        )
    return ParsedDataset(
        headers=all_headers,
        rows=result,
        source_features={
            "kind": "workbook",
            "sheets": sheets,
            "skipped_sheets": skipped_sheets,
            "headers": all_headers,
            "structure_detection": "deterministic_v2",
        },
        warnings=warnings,
    )


def _csv_reader(path: Path):
    for encoding in ("utf-8-sig", "gb18030"):
        handle = None
        try:
            handle = path.open("r", encoding=encoding, newline="")
            sample = handle.read(8192)
            handle.seek(0)
            try:
                dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
            except csv.Error:
                dialect = csv.excel
            return handle, csv.reader(handle, dialect, skipinitialspace=True)
        except UnicodeDecodeError:
            if handle is not None:
                handle.close()
            continue
    raise EtlError("ETL_CSV_ENCODING_UNSUPPORTED", "CSV 编码无法识别")


def _parse_csv(path: Path, max_rows: int, target_type: str) -> ParsedDataset:
    handle, reader = _csv_reader(path)
    try:
        probe = list(itertools.islice(reader, 60))
        layout = detect_table_layout(
            probe,
            header_hints=target_header_hints(target_type),
        )
        if layout is None:
            return ParsedDataset(headers=[], rows=[], source_features={"kind": "csv"})
        headers = layout.headers
        rows: list[ParsedRow] = []
        repeated_header_count = 0
        footer_count = 0
        remaining_probe = probe[layout.header_end + 1 :]
        for row_number, values in enumerate(
            itertools.chain(remaining_probe, reader),
            start=layout.header_end + 2,
        ):
            if is_repeated_header(values, headers):
                repeated_header_count += 1
                continue
            if is_footer_or_note_row(values):
                footer_count += 1
                continue
            data = {
                headers[index]: value
                for index, value in enumerate(values[: len(headers)])
                if value != ""
            }
            if not data:
                continue
            if len(rows) >= max_rows:
                raise EtlError(
                    "ETL_ROW_LIMIT_EXCEEDED", f"文件超过 {max_rows} 行限制", status_code=413
                )
            rows.append(
                ParsedRow(
                    sheet="CSV",
                    row_number=row_number,
                    values=data,
                    provenance={
                        "sheet": "CSV",
                        "row": row_number,
                        "original_fragment": dict(zip(headers, values)),
                        "columns": {headers[index]: index + 1 for index in range(len(headers))},
                        "header_rows": {
                            "start": layout.header_start + 1,
                            "end": layout.header_end + 1,
                        },
                    },
                )
            )
        warnings: list[dict[str, Any]] = []
        if repeated_header_count:
            warnings.append(
                {
                    "code": "ETL_REPEATED_HEADERS_SKIPPED",
                    "message": f"已跳过 {repeated_header_count} 行重复表头。",
                    "count": repeated_header_count,
                }
            )
        if footer_count:
            warnings.append(
                {
                    "code": "ETL_FOOTER_ROWS_SKIPPED",
                    "message": f"已跳过 {footer_count} 行合计、备注或签字尾行。",
                    "count": footer_count,
                }
            )
        return ParsedDataset(
            headers=headers,
            rows=rows,
            source_features={
                "kind": "csv",
                "headers": headers,
                "row_count": len(rows),
                "header_row": layout.header_end + 1,
                "header_start_row": layout.header_start + 1,
                "header_depth": layout.header_end - layout.header_start + 1,
                "header_confidence": layout.confidence,
                "header_reasons": list(layout.reasons),
                "structure_detection": "deterministic_v2",
            },
            warnings=warnings,
        )
    finally:
        handle.close()


def parse_file(
    path: str | Path,
    *,
    target_type: str,
    max_rows: int = MAX_ROWS,
    compatibility_preset_id: str | None = None,
    document_plan: dict[str, Any] | None = None,
) -> ParsedDataset:
    source = Path(path).expanduser().resolve()
    suffix = source.suffix.lower()
    if not source.is_file():
        raise EtlError("ETL_UPLOAD_MISSING", "上传文件不存在", status_code=404)
    if suffix not in SUPPORTED_SUFFIXES:
        raise EtlError("ETL_FILE_TYPE_UNSUPPORTED", f"不支持的文件类型: {suffix}")
    if target_type == "knowledge":
        return ParsedDataset(
            headers=["document_path"],
            rows=[
                ParsedRow(
                    sheet="文档",
                    row_number=1,
                    values={"document_path": str(source)},
                    provenance={"original_file": source.name},
                )
            ],
            source_features={
                "kind": "document",
                "knowledge_only": suffix in KNOWLEDGE_ONLY_SUFFIXES,
            },
        )
    if suffix in KNOWLEDGE_ONLY_SUFFIXES:
        raise EtlError(
            "ETL_KNOWLEDGE_ONLY_FILE",
            "Word/PPT 仅可导入知识库，不支持结构化业务写入",
        )
    if suffix == ".csv":
        if compatibility_preset_id:
            raise EtlError(
                "ETL_COMPATIBILITY_PRESET_FILE_UNSUPPORTED",
                "兼容预设仅适用于 XLSX/XLSM 文件；CSV 请使用自动识别",
            )
        return _parse_csv(source, max_rows, target_type)
    if suffix in STRUCTURED_SUFFIXES:
        if (
            document_plan
            and suffix in {".xlsx", ".xlsm"}
            and not compatibility_preset_id
            and (
                target_type not in {"customer_products", "shipment_records"}
                or bool(document_plan.get("routing_scope"))
            )
        ):
            from app.application.etl.parser_document_plan import (
                parse_workbook_with_document_plan,
            )

            planned = parse_workbook_with_document_plan(
                source,
                target_type=target_type,
                document_plan=document_plan,
                max_rows=max_rows,
            )
            if planned is not None and (
                planned.rows or bool(document_plan.get("routing_scope"))
            ):
                return planned
        if compatibility_preset_id and target_type not in {
            "customer_products",
            "customers",
            "products",
            "shipment_records",
        }:
            raise EtlError(
                "ETL_COMPATIBILITY_PRESET_TARGET_MISMATCH",
                "兼容预设不适用于当前目标",
            )
        if target_type in {"customer_products", "shipment_records"} and not compatibility_preset_id:
            try:
                from app.application.etl.parser_regions import (
                    parse_customer_product_regions,
                )

                regional = parse_customer_product_regions(
                    source,
                    max_rows=max_rows,
                    target_type=target_type,
                )
            except EtlError:
                raise
            except Exception:  # noqa: BLE001 - legacy presets remain a safe fallback
                regional = None
            if regional is not None and covers_required_target_fields(regional, target_type):
                return regional
        from app.application.etl.shipment_compat_parser import (
            parse_delivery_note_with_compat_profile,
        )

        compatibility = parse_delivery_note_with_compat_profile(
            source,
            target_type=target_type,
            max_rows=max_rows,
            compatibility_preset_id=compatibility_preset_id,
        )
        if compatibility_preset_id:
            if compatibility is None:
                raise EtlError(
                    "ETL_COMPATIBILITY_PRESET_NO_MATCH",
                    "所选兼容预设未识别到可靠业务数据，请改用自动识别或其他预设",
                )
            return compatibility
        if compatibility is not None:
            if target_type == "shipment_records":
                return compatibility
            try:
                generic = _parse_workbook(source, max_rows, target_type)
            except EtlError:
                raise
            except Exception:  # noqa: BLE001 - a proven legacy preset remains a safe fallback
                return compatibility
            if generic.rows and covers_required_target_fields(generic, target_type):
                generic.warnings.insert(
                    0,
                    {
                        "code": "ETL_GENERIC_STRUCTURE_PREFERRED",
                        "message": "文件已包含完整业务字段，已优先使用通用表格结构解析。",
                    },
                )
                return generic
            return compatibility
        return _parse_workbook(source, max_rows, target_type)

    if compatibility_preset_id:
        raise EtlError(
            "ETL_COMPATIBILITY_PRESET_FILE_UNSUPPORTED",
            "兼容预设仅适用于 XLSX/XLSM 文件；OCR 文件请使用自动识别",
        )

    from app.application.shipment_excel_etl_ocr import ocr_source_to_workbook

    result = ocr_source_to_workbook(source)
    if not result.get("success"):
        raise EtlError(
            str(result.get("error_code") or "ETL_OCR_FAILED").upper(),
            "OCR 无法可靠还原表格，请更换清晰文件后重试",
        )
    dataset = _parse_workbook(Path(str(result["file_path"])), max_rows, target_type)
    return enrich_ocr_provenance(dataset, result, source_suffix=suffix)
