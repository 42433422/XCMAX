"""Extract one clean, reusable delivery-note layout from a mixed workbook."""

from __future__ import annotations

from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from app.application.etl.errors import EtlError
from app.application.etl.parser_structure import clean_cell_text, semantic_key


def _last_nonempty_row(worksheet: Any, start: int, end: int, last_col: int) -> int:
    for row in range(end, start - 1, -1):
        if any(
            clean_cell_text(worksheet.cell(row=row, column=column).value)
            for column in range(1, last_col + 1)
        ):
            return row
    return start


def _positive_int(value: Any) -> int:
    try:
        parsed = int(value or 0)
    except (TypeError, ValueError):
        return 0
    return parsed if parsed > 0 else 0


def _document_start_rows(source_features: dict[str, Any], *, sheet: str) -> list[int]:
    understanding = source_features.get("document_understanding")
    routes = understanding.get("document_routes") if isinstance(understanding, dict) else []
    starts: list[int] = []
    for route in routes or []:
        if not isinstance(route, dict) or str(route.get("sheet") or "") != sheet:
            continue
        header_rows = [
            _positive_int(item.get("header_start_row"))
            for item in route.get("data_ranges") or []
            if isinstance(item, dict)
        ]
        header_rows = [row for row in header_rows if row]
        if header_rows:
            starts.append(max(1, min(header_rows) - 2))
    return starts


def _template_bounds(
    source_features: dict[str, Any],
    regions: list[dict[str, Any]],
    *,
    source_region_id: str | None = None,
) -> tuple[dict[str, Any], int, int]:
    """Choose one explicitly selected delivery-note region.

    A workbook can carry several historical delivery-note layouts.  The
    caller may supply the audited region id selected in the ETL preview; when
    it does, never silently substitute the first layout on a different
    customer/document.  Legacy callers omit it and retain the deterministic
    first-region behaviour.
    """

    selected = [row for row in regions if row.get("status") == "selected"]
    if not selected:
        raise EtlError("ETL_SHIPMENT_TEMPLATE_REGION_MISSING", "未识别到可提取的送货单版式")
    requested = str(source_region_id or "").strip()
    if requested:
        region = next(
            (row for row in selected if str(row.get("id") or "") == requested),
            None,
        )
        if region is None:
            raise EtlError(
                "ETL_SHIPMENT_TEMPLATE_REGION_NOT_FOUND",
                "所选发货单版式不在当前预演中，请重新选择",
                status_code=409,
            )
    else:
        region = sorted(
            selected,
            key=lambda row: (str(row.get("sheet")), int(row.get("header_row") or 0)),
        )[0]
    header_row = int(region.get("header_row") or 0)
    evidence_rows = [
        int(item.get("row") or 0)
        for item in region.get("evidence_rows") or []
        if isinstance(item, dict) and int(item.get("row") or 0) > 0
    ]
    start = min(evidence_rows) if evidence_rows else max(1, header_row - 2)
    sheet_name = str(region.get("sheet") or "")
    next_starts = [
        max(1, int(item.get("header_row") or 0) - 2)
        for item in regions
        if str(item.get("sheet") or "") == sheet_name
        and int(item.get("header_row") or 0) > header_row
    ]
    next_starts.extend(
        row
        for row in _document_start_rows(source_features, sheet=sheet_name)
        if row > header_row
    )
    end_limit = min(next_starts) - 1 if next_starts else 0
    return region, start, end_limit


def _copy_sheet_region(
    source: Any,
    *,
    start_row: int,
    end_row: int,
    last_col: int,
) -> Workbook:
    output = Workbook()
    target = output.active
    target.title = "送货单"
    for source_row in range(start_row, end_row + 1):
        target_row = source_row - start_row + 1
        target.row_dimensions[target_row].height = source.row_dimensions[source_row].height
        target.row_dimensions[target_row].hidden = source.row_dimensions[source_row].hidden
        for column in range(1, last_col + 1):
            src = source.cell(row=source_row, column=column)
            dst = target.cell(row=target_row, column=column, value=src.value)
            if src.has_style:
                dst.font = copy(src.font)
                dst.fill = copy(src.fill)
                dst.border = copy(src.border)
                dst.number_format = src.number_format
                dst.alignment = copy(src.alignment)
                dst.protection = copy(src.protection)
    for column in range(1, last_col + 1):
        letter = get_column_letter(column)
        target.column_dimensions[letter].width = source.column_dimensions[letter].width
        target.column_dimensions[letter].hidden = source.column_dimensions[letter].hidden
    for merged in source.merged_cells.ranges:
        if merged.min_row >= start_row and merged.max_row <= end_row and merged.max_col <= last_col:
            target.merge_cells(
                start_row=merged.min_row - start_row + 1,
                end_row=merged.max_row - start_row + 1,
                start_column=merged.min_col,
                end_column=merged.max_col,
            )
    target.page_margins = copy(source.page_margins)
    target.page_setup = copy(source.page_setup)
    target.print_options = copy(source.print_options)
    target.sheet_properties = copy(source.sheet_properties)
    target.freeze_panes = source.freeze_panes
    target.print_area = f"A1:{get_column_letter(last_col)}{end_row - start_row + 1}"
    return output


def _clear_example_lines(worksheet: Any, *, header_row: int, last_col: int) -> None:
    total_row = 0
    for row in range(header_row + 1, worksheet.max_row + 1):
        values = [
            clean_cell_text(worksheet.cell(row=row, column=column).value)
            for column in range(1, min(last_col, 4) + 1)
        ]
        if semantic_key(" ".join(values)) in {"合计", "总计", "小计", "汇总"}:
            total_row = row
            break
        if any(semantic_key(value) in {"合计", "总计", "小计", "汇总"} for value in values):
            total_row = row
            break
    if not total_row:
        raise EtlError("ETL_SHIPMENT_TEMPLATE_TOTAL_MISSING", "送货单版式缺少合计行")
    for row in range(header_row + 1, total_row):
        for column in range(1, last_col + 1):
            cell = worksheet.cell(row=row, column=column)
            if not isinstance(cell, MergedCell):
                cell.value = None


def extract_shipment_template(
    source_path: str | Path,
    *,
    source_features: dict[str, Any],
    destination: str | Path,
    source_region_id: str | None = None,
) -> dict[str, Any]:
    regions = [item for item in source_features.get("regions") or [] if isinstance(item, dict)]
    region, start_row, end_limit = _template_bounds(
        source_features,
        regions,
        source_region_id=source_region_id,
    )
    workbook = load_workbook(source_path, data_only=False, keep_links=False)
    try:
        sheet_name = str(region.get("sheet") or "")
        if sheet_name not in workbook.sheetnames:
            raise EtlError("ETL_SHIPMENT_TEMPLATE_SHEET_MISSING", "送货单版式所在工作表不存在")
        source = workbook[sheet_name]
        last_col = int(region.get("last_column") or source.max_column)
        end_row = end_limit or source.max_row
        end_row = _last_nonempty_row(source, start_row, end_row, last_col)
        output = _copy_sheet_region(
            source,
            start_row=start_row,
            end_row=end_row,
            last_col=last_col,
        )
    finally:
        workbook.close()
    header_row = int(region.get("header_row") or 0) - start_row + 1
    _clear_example_lines(output.active, header_row=header_row, last_col=last_col)
    destination_path = Path(destination)
    destination_path.parent.mkdir(parents=True, exist_ok=True)
    output.save(destination_path)
    output.close()
    return {
        "file_path": str(destination_path),
        "sheet": sheet_name,
        "source_region_id": region.get("id"),
        "header_row": header_row,
        "rows": end_row - start_row + 1,
        "columns": last_col,
    }


__all__ = ["extract_shipment_template"]
