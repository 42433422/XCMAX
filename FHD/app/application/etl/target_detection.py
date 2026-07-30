"""Conservative input-target detection for the ETL Center's automatic mode."""

from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any, Iterable

from app.application.etl.parsers import KNOWLEDGE_ONLY_SUFFIXES

_DELIVERY_RE = re.compile(r"(送货单|发货单|出货单|delivery)", re.I)
_PRODUCT_RE = re.compile(r"(产品名称|商品名称|货品名称|品名)")
_QUANTITY_RE = re.compile(r"(数量|桶数|件数|重量|规格)")
_CUSTOMER_RE = re.compile(r"(购货单位|购买单位|客户名称|客户)")
_ATTENDANCE_RE = re.compile(r"(考勤|上班时间|下班时间|打卡)")
_PURCHASE_RE = re.compile(r"(采购订单|供应商|供货商)")
_EXPORT_DOCUMENT_RE = re.compile(r"(发票|报价单|装箱单|invoice|quotation|packing\s*list)", re.I)


def _joined(values: Iterable[Any]) -> str:
    return " ".join(str(value).strip() for value in values if value not in (None, ""))[:4000]


def _classify(lines: list[str], *, suffix: str) -> dict[str, Any]:
    text = "\n".join(lines)
    if _DELIVERY_RE.search(text) and _PRODUCT_RE.search(text) and _QUANTITY_RE.search(text):
        return {
            "target_type": "shipment_records",
            "document_type": "delivery_note_workbook",
            "confidence": 0.98,
            "reason": "delivery_title_and_product_table",
        }
    if _ATTENDANCE_RE.search(text):
        return {
            "target_type": "attendance",
            "document_type": "attendance_table",
            "confidence": 0.9,
            "reason": "attendance_headers",
        }
    if _EXPORT_DOCUMENT_RE.search(text):
        return {
            "target_type": "export_xlsx",
            "document_type": "reviewable_business_document",
            "confidence": 0.75,
            "reason": "document_requires_reviewable_export",
        }
    if _PURCHASE_RE.search(text) and _PRODUCT_RE.search(text):
        return {
            "target_type": "purchase_orders",
            "document_type": "purchase_order_table",
            "confidence": 0.86,
            "reason": "purchase_and_product_headers",
        }
    if _CUSTOMER_RE.search(text) and _PRODUCT_RE.search(text):
        return {
            "target_type": "customer_products",
            "document_type": "linked_customer_products",
            "confidence": 0.82,
            "reason": "customer_and_product_headers",
        }
    if _CUSTOMER_RE.search(text):
        return {
            "target_type": "customers",
            "document_type": "customer_table",
            "confidence": 0.74,
            "reason": "customer_headers",
        }
    if _PRODUCT_RE.search(text):
        return {
            "target_type": "products",
            "document_type": "product_table",
            "confidence": 0.72,
            "reason": "product_headers",
        }
    return {
        "target_type": "export_xlsx",
        "document_type": "generic_structured_table",
        "confidence": 0.35,
        "reason": f"safe_export_fallback:{suffix}",
    }


def detect_etl_target(path: str | Path, *, suffix: str = "") -> dict[str, Any]:
    source = Path(path)
    ext = str(suffix or source.suffix).lower()
    if ext in KNOWLEDGE_ONLY_SUFFIXES:
        return {
            "target_type": "knowledge",
            "document_type": "knowledge_document",
            "confidence": 1.0,
            "reason": "knowledge_only_suffix",
        }
    if ext == ".csv":
        lines: list[str] = []
        with source.open("r", encoding="utf-8-sig", errors="replace", newline="") as handle:
            for index, row in enumerate(csv.reader(handle)):
                lines.append(_joined(row))
                if index >= 30:
                    break
        return _classify(lines, suffix=ext)
    if ext in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook

        workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
        lines = []
        try:
            for worksheet in workbook.worksheets[:30]:
                lines.append(worksheet.title)
                for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                    lines.append(_joined(row))
                    if index >= 79:
                        break
        finally:
            workbook.close()
        return _classify(lines, suffix=ext)
    filename = source.name
    if _DELIVERY_RE.search(filename):
        return {
            "target_type": "shipment_records",
            "document_type": "delivery_note_ocr",
            "confidence": 0.68,
            "reason": "delivery_filename",
        }
    return {
        "target_type": "export_xlsx",
        "document_type": "unknown_business_document",
        "confidence": 0.25,
        "reason": "manual_review_required",
    }


__all__ = ["detect_etl_target"]
