"""Bounded, auditable workbook evidence for document understanding."""

from __future__ import annotations

import hashlib
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.application.etl.parser_structure import clean_cell_text, detect_table_layout

MAX_EVIDENCE_SHEETS = 30
MAX_EVIDENCE_ROWS_PER_SHEET = 160
MAX_EVIDENCE_COLUMNS = 40
MAX_EVIDENCE_CELLS = 2400

_UNIVERSAL_TABLE_HINTS = (
    "序号",
    "型号",
    "品名",
    "产品名称",
    "商品名称",
    "数量",
    "单价",
    "金额",
    "姓名",
    "部门",
    "日期",
    "no",
    "item",
    "sku",
    "part no",
    "description",
    "contents",
    "qty",
    "quantity",
    "unit price",
    "amount",
    "department",
    "employee",
)


def cell_evidence_id(sheet_index: int, row: int, column: int) -> str:
    return f"s{sheet_index}:r{row}:c{column}"


def _json_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _value_type(value: Any) -> str:
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (datetime, date)):
        return "date"
    if isinstance(value, (int, float)):
        return "number"
    return "text"


def _table_candidate(
    sheet_index: int,
    sheet: str,
    rows: list[list[Any]],
    *,
    max_row: int,
) -> dict[str, Any] | None:
    layout = detect_table_layout(rows, header_hints=_UNIVERSAL_TABLE_HINTS)
    if layout is None:
        return None
    header_row = layout.header_end + 1
    first_column = 1
    last_column = len(layout.headers)
    return {
        "candidate_id": f"table:s{sheet_index}:r{header_row}",
        "sheet": sheet,
        "header_start_row": layout.header_start + 1,
        "header_end_row": header_row,
        "data_start_row": header_row + 1,
        "data_end_row": max(max_row, len(rows)),
        "first_column": first_column,
        "last_column": last_column,
        "headers": list(layout.headers),
        "confidence": layout.confidence,
        "reasons": list(layout.reasons),
    }


def _key_value_candidates(
    *,
    sheet_index: int,
    sheet: str,
    rows: list[list[Any]],
    header_start_row: int,
) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row_index, values in enumerate(rows[: max(0, header_start_row - 1)], start=1):
        for column_index, value in enumerate(values, start=1):
            label = clean_cell_text(value)
            if not label:
                continue
            if "：" in label or ":" in label:
                separator = "：" if "：" in label else ":"
                left, right = (part.strip() for part in label.split(separator, 1))
                if left and right:
                    label_id = cell_evidence_id(sheet_index, row_index, column_index)
                    key = (label_id, label_id)
                    if key not in seen:
                        seen.add(key)
                        candidates.append(
                            {
                                "sheet": sheet,
                                "label": left[:160],
                                "value": right[:300],
                                "label_cell_id": label_id,
                                "value_cell_id": label_id,
                            }
                        )
            # Business forms often use merged cells, so the visible value can
            # be several physical columns away from its label.
            for value_column in range(column_index + 1, min(len(values), column_index + 12) + 1):
                candidate_value = values[value_column - 1]
                text = clean_cell_text(candidate_value)
                if not text:
                    continue
                label_id = cell_evidence_id(sheet_index, row_index, column_index)
                value_id = cell_evidence_id(sheet_index, row_index, value_column)
                key = (label_id, value_id)
                if key in seen:
                    continue
                seen.add(key)
                candidates.append(
                    {
                        "sheet": sheet,
                        "label": label[:160],
                        "value": _json_value(candidate_value),
                        "label_cell_id": label_id,
                        "value_cell_id": value_id,
                    }
                )
                break
    return candidates[:120]


def build_workbook_evidence(path: str | Path) -> dict[str, Any]:
    """Read a bounded workbook snapshot with stable cell references.

    The manifest is intentionally independent from any target schema. It gives
    document understanding access to title, preamble, tables, totals and notes
    while keeping every conclusion traceable to a supplied cell ID.
    """

    source = Path(path).expanduser().resolve()
    workbook = load_workbook(source, read_only=True, data_only=True, keep_links=False)
    sheets: list[dict[str, Any]] = []
    sheet_manifest = [
        {
            "index": index,
            "name": worksheet.title,
            "max_row": int(worksheet.max_row or 0),
            "max_column": int(worksheet.max_column or 0),
        }
        for index, worksheet in enumerate(workbook.worksheets, start=1)
    ]
    table_candidates: list[dict[str, Any]] = []
    key_value_candidates: list[dict[str, Any]] = []
    cell_index: dict[str, dict[str, Any]] = {}
    total_cells = 0
    truncated = False
    try:
        visible_sheet_count = min(len(workbook.worksheets), MAX_EVIDENCE_SHEETS)
        per_sheet_cell_budget = max(
            80,
            MAX_EVIDENCE_CELLS // max(1, visible_sheet_count),
        )
        for sheet_index, worksheet in enumerate(
            workbook.worksheets[:MAX_EVIDENCE_SHEETS],
            start=1,
        ):
            row_values: list[list[Any]] = []
            evidence_rows: list[dict[str, Any]] = []
            sheet_cells = 0
            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=1,
                    max_row=min(MAX_EVIDENCE_ROWS_PER_SHEET, worksheet.max_row or 1),
                    max_col=min(MAX_EVIDENCE_COLUMNS, worksheet.max_column or 1),
                    values_only=True,
                ),
                start=1,
            ):
                values = list(row)
                row_values.append(values)
                row_cells: list[dict[str, Any]] = []
                for column, value in enumerate(values, start=1):
                    if value in (None, ""):
                        continue
                    if sheet_cells >= per_sheet_cell_budget:
                        truncated = True
                        break
                    evidence_id = cell_evidence_id(sheet_index, row_number, column)
                    item = {
                        "id": evidence_id,
                        "sheet": worksheet.title,
                        "coordinate": f"{get_column_letter(column)}{row_number}",
                        "row": row_number,
                        "column": column,
                        "value": _json_value(value),
                        "text": clean_cell_text(value)[:500],
                        "value_type": _value_type(value),
                    }
                    row_cells.append(item)
                    cell_index[evidence_id] = item
                    total_cells += 1
                    sheet_cells += 1
                if row_cells:
                    evidence_rows.append({"row": row_number, "cells": row_cells})
                if sheet_cells >= per_sheet_cell_budget:
                    break
            candidate = _table_candidate(
                sheet_index,
                worksheet.title,
                row_values,
                max_row=int(worksheet.max_row or len(row_values)),
            )
            if candidate is not None:
                table_candidates.append(candidate)
                header_start_row = int(candidate["header_start_row"])
            else:
                header_start_row = min(len(row_values) + 1, 30)
            key_value_candidates.extend(
                _key_value_candidates(
                    sheet_index=sheet_index,
                    sheet=worksheet.title,
                    rows=row_values,
                    header_start_row=header_start_row,
                )
            )
            sheet_truncated = (
                sheet_cells >= per_sheet_cell_budget
                or int(worksheet.max_row or 0) > MAX_EVIDENCE_ROWS_PER_SHEET
                or int(worksheet.max_column or 0) > MAX_EVIDENCE_COLUMNS
            )
            sheets.append(
                {
                    "index": sheet_index,
                    "name": worksheet.title,
                    "max_row": int(worksheet.max_row or 0),
                    "max_column": int(worksheet.max_column or 0),
                    "truncated": sheet_truncated,
                    "rows": evidence_rows,
                }
            )
        if len(workbook.worksheets) > len(sheets):
            truncated = True
    finally:
        workbook.close()

    digest = hashlib.sha256()
    for evidence_id in sorted(cell_index):
        item = cell_index[evidence_id]
        digest.update(evidence_id.encode())
        digest.update(b"\0")
        digest.update(str(item["text"]).encode("utf-8", errors="replace"))
        digest.update(b"\0")
    return {
        "version": 1,
        "file_name": source.name,
        "evidence_hash": digest.hexdigest(),
        "truncated": truncated,
        "limits": {
            "sheets": MAX_EVIDENCE_SHEETS,
            "rows_per_sheet": MAX_EVIDENCE_ROWS_PER_SHEET,
            "columns": MAX_EVIDENCE_COLUMNS,
            "cells": MAX_EVIDENCE_CELLS,
        },
        "workbook_sheet_count": len(sheet_manifest),
        "sheet_manifest": sheet_manifest,
        "sheets": sheets,
        "table_candidates": table_candidates,
        "key_value_candidates": key_value_candidates[:240],
        "cell_index": cell_index,
    }


def public_evidence_summary(evidence: dict[str, Any]) -> dict[str, Any]:
    manifest = list(evidence.get("sheet_manifest") or evidence.get("sheets") or [])
    return {
        "version": evidence.get("version"),
        "file_name": evidence.get("file_name"),
        "evidence_hash": evidence.get("evidence_hash"),
        "truncated": bool(evidence.get("truncated")),
        "sheet_count": int(evidence.get("workbook_sheet_count") or len(manifest)),
        "cell_count": len(evidence.get("cell_index") or {}),
        "sheets": [
            {
                "name": item.get("name"),
                "max_row": item.get("max_row"),
                "max_column": item.get("max_column"),
            }
            for item in manifest
        ],
    }


__all__ = [
    "build_workbook_evidence",
    "cell_evidence_id",
    "public_evidence_summary",
]
