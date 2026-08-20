"""Workbook decomposition primitives shared by Excel-template entrypoints."""

from __future__ import annotations

import logging
import os
from datetime import date, datetime, time
from decimal import Decimal
from typing import Any, Callable

from app.utils.operational_errors import RECOVERABLE_ERRORS

logger = logging.getLogger(__name__)


def json_safe_cell_value(value: Any) -> Any:
    """Ensure values returned by spreadsheet engines are JSON serializable."""
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, (int, str)):
        return value
    if isinstance(value, float):
        try:
            import math

            if not math.isfinite(value):
                return None
        except RECOVERABLE_ERRORS:
            pass
        return value
    if isinstance(value, Decimal):
        return format(value, "f")
    return str(value)


def is_unreadable_workbook_error(error_message: str) -> bool:
    lowered = str(error_message or "").lower()
    markers = (
        "unable to read workbook",
        "could not read worksheets",
        "invalid xml",
        "badzipfile",
        "bad zip file",
        "bad magic number",
        "central directory",
        "file is not a zip file",
        "not a zip file",
        "does not support the old .xls",
    )
    return any(marker in lowered for marker in markers)


def pick_sheet_name(sheet_names: list[str], sheet_name: str | None) -> str:
    names = list(sheet_names or [])
    if sheet_name and sheet_name in names:
        return sheet_name
    for name in names:
        if "出货" in name:
            return name
    return names[0] if names else ""


def decompose_from_grid(
    file_path: str,
    sheet_title: str,
    nrows: int,
    ncols: int,
    get_cell_value: Callable[[int, int], Any],
    merged_cells_count: int,
    dimension: str,
    sheet_name: str | None,
    sample_rows: int,
) -> tuple[dict, int]:
    from openpyxl.utils import get_column_letter

    max_r = min(max(nrows, 1), 30)
    max_c = min(max(ncols, 1), 25)
    header_row_idx = None
    header_cells: list[dict[str, Any]] = []
    for row_index in range(1, max_r + 1):
        row_cells = []
        for column_index in range(1, max_c + 1):
            value = get_cell_value(row_index, column_index)
            if isinstance(value, str) and value.strip():
                row_cells.append(
                    {
                        "name": value.strip(),
                        "column": get_column_letter(column_index),
                        "column_index": column_index,
                    }
                )
        if len(row_cells) >= 4:
            header_row_idx = row_index
            header_cells = row_cells
            break
    if header_row_idx is None:
        header_row_idx = 1

    samples = []
    data_start = header_row_idx + 1
    data_end = min(max(nrows, 1), data_start + max(int(sample_rows), 1) - 1)
    for row_index in range(data_start, data_end + 1):
        row_data: dict[str, Any] = {}
        non_empty = False
        for header in header_cells:
            safe = json_safe_cell_value(get_cell_value(row_index, header["column_index"]))
            if safe is not None and safe != "":
                non_empty = True
            row_data[header["name"]] = safe
        if non_empty and row_data:
            samples.append(row_data)

    amount_related = [
        header
        for header in header_cells
        if any(keyword in header["name"] for keyword in ["金额", "单价", "价格", "数量"])
    ]
    return {
        "success": True,
        "template": {
            "name": os.path.basename(file_path),
            "path": file_path,
            "sheet": sheet_title,
            "dimension": dimension,
            "max_row": max(nrows, 1),
            "max_column": max(ncols, 1),
        },
        "decomposition": {
            "header_row": header_row_idx,
            "editable_entries": header_cells,
            "amount_related_entries": amount_related,
            "sample_rows": samples,
            "merged_cells_count": merged_cells_count,
        },
    }, 200


def decompose_template_xls_pandas(
    file_path: str, sheet_name: str | None = None, sample_rows: int = 5
) -> tuple[dict, int]:
    try:
        import pandas as pd
    except ImportError as exc:
        return {"success": False, "message": f"读取 .xls 需要 pandas：{exc}"}, 500
    try:
        workbook = pd.ExcelFile(file_path, engine="xlrd")
    except RECOVERABLE_ERRORS as exc:
        logger.error("pandas/xlrd 打开 .xls 失败: %s", exc)
        if is_unreadable_workbook_error(str(exc)):
            return {
                "success": False,
                "message": "无法读取该 .xls 文件（可能损坏）。请另存为 .xlsx 后重试。",
                "error_code": "UNREADABLE_WORKBOOK",
            }, 200
        hint = "请确认已安装 xlrd：pip install xlrd"
        return {"success": False, "message": f"{hint}。原始错误：{exc}"}, 500

    names = list(workbook.sheet_names)
    if not names:
        return {"success": False, "message": "工作簿中没有工作表"}, 200
    selected = pick_sheet_name(names, sheet_name)
    frame = pd.read_excel(file_path, sheet_name=selected, header=None, dtype=object, engine="xlrd")
    nrows, ncols = int(frame.shape[0]), int(frame.shape[1])

    def get_cell_value(row_index: int, column_index: int) -> Any:
        if row_index < 1 or column_index < 1 or row_index > nrows or column_index > ncols:
            return None
        value = frame.iat[row_index - 1, column_index - 1]
        if value is None or (isinstance(value, float) and pd.isna(value)):
            return None
        return value

    from openpyxl.utils import get_column_letter

    dimension = f"A1:{get_column_letter(max(ncols, 1))}{max(nrows, 1)}"
    return decompose_from_grid(
        file_path,
        str(selected),
        nrows,
        ncols,
        get_cell_value,
        0,
        dimension,
        sheet_name,
        sample_rows,
    )


def decompose_template_openpyxl(
    file_path: str, sheet_name: str | None = None, sample_rows: int = 5
) -> tuple[dict, int]:
    from openpyxl import load_workbook

    workbook = load_workbook(file_path, data_only=True)
    if sheet_name and sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    elif "出货" in workbook.sheetnames:
        sheet = workbook["出货"]
    else:
        sheet = workbook[workbook.sheetnames[0]]

    def get_cell_value(row_index: int, column_index: int) -> Any:
        return sheet.cell(row_index, column_index).value

    return decompose_from_grid(
        file_path,
        sheet.title,
        sheet.max_row,
        sheet.max_column,
        get_cell_value,
        len(sheet.merged_cells.ranges),
        sheet.calculate_dimension() or "A1:A1",
        sheet_name,
        sample_rows,
    )


def decompose_template(
    file_path: str, sheet_name: str | None = None, sample_rows: int = 5
) -> tuple[dict, int]:
    try:
        if not os.path.exists(file_path):
            return {"success": False, "message": f"模板文件不存在: {file_path}"}, 404
        if os.path.splitext(file_path)[1].lower() == ".xls":
            return decompose_template_xls_pandas(file_path, sheet_name, sample_rows)
        try:
            return decompose_template_openpyxl(file_path, sheet_name, sample_rows)
        except RECOVERABLE_ERRORS as exc:
            if is_unreadable_workbook_error(str(exc)):
                return {
                    "success": False,
                    "message": "模板文件损坏或格式异常，无法读取。请重新导出或另存为 .xlsx 后重试。",
                    "error_code": "UNREADABLE_WORKBOOK",
                }, 200
            raise
    except RECOVERABLE_ERRORS as exc:
        logger.error("分解 Excel 模板失败: %s", exc)
        if is_unreadable_workbook_error(str(exc)):
            return {
                "success": False,
                "message": "模板文件损坏或格式异常，无法读取。请重新导出或另存为 .xlsx 后重试。",
                "error_code": "UNREADABLE_WORKBOOK",
            }, 200
        return {"success": False, "message": str(exc)}, 500
