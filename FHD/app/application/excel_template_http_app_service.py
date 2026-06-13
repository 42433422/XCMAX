"""Excel 模板 HTTP 应用服务（自 fastapi_routes/excel_templates 下沉）。"""

from __future__ import annotations

import json
import logging
import os
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from fastapi import Body, File, Query, UploadFile
from fastapi.responses import FileResponse, JSONResponse
from sqlalchemy import text

from app.utils.operational_errors import RECOVERABLE_ERRORS

logger = logging.getLogger(__name__)

_REPO_ROOT = Path(__file__).resolve().parents[2]
TEMPLATE_DIR = str(_REPO_ROOT / "templates")
TEMP_EXCEL_DIR = str(_REPO_ROOT / "temp_excel")
os.makedirs(TEMPLATE_DIR, exist_ok=True)
os.makedirs(TEMP_EXCEL_DIR, exist_ok=True)


def get_base_dir() -> str:
    return str(_REPO_ROOT)


def _map_template_category(template_type: str) -> str:
    t = (template_type or "").strip().lower()
    if any(k in t for k in ["标签", "label", "print", "打印"]):
        return "label_print"
    return "excel"


def _normalize_template_dto(template: dict) -> dict:
    tpl = dict(template or {})
    template_type = tpl.get("template_type", "")
    category = tpl.get("category") or _map_template_category(str(template_type))
    file_path = tpl.get("file_path") or tpl.get("path")
    lower_fp = str(file_path or "").lower()
    if lower_fp.endswith((".docx", ".doc")):
        category = "word"
    normalized = {
        **tpl,
        "category": category,
        "file_path": file_path,
        "is_active": bool(tpl.get("is_active", True)),
        "preview_capable": bool(file_path and tpl.get("exists", False)),
    }
    return normalized


def _resolve_template_path(filename: str) -> str | None:
    base_dir = get_base_dir()
    path = os.path.join(base_dir, filename)
    if os.path.exists(path):
        return path
    alt_path = os.path.join(TEMPLATE_DIR, filename)
    if os.path.exists(alt_path):
        return alt_path
    return path if os.path.exists(path) else None


def _get_template_list():
    from app.application import get_template_app_service

    return get_template_app_service().get_templates().get("templates", [])


def _json_safe_cell_value(v: Any) -> Any:
    """Ensure cell values are JSON-serializable (openpyxl may return datetime/Decimal)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, datetime):
        return v.isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, time):
        return v.isoformat()
    if isinstance(v, (int, str)):
        return v
    if isinstance(v, float):
        try:
            import math

            if not math.isfinite(v):
                return None
        except RECOVERABLE_ERRORS:
            pass
        return v
    if isinstance(v, Decimal):
        return format(v, "f")
    return str(v)


def _is_unreadable_workbook_error(error_message: str) -> bool:
    text = str(error_message or "").lower()
    markers = [
        "unable to read workbook",
        "could not read worksheets",
        "invalid xml",
        "badzipfile",
        "bad zip file",
        "bad magic number",
        "central directory",
        "file is not a zip file",
        "not a zip file",
        "does not support the old .xls",
    ]
    return any(m in text for m in markers)


def _pick_sheet_name(sheet_names: list[str], sheet_name: str | None) -> str:
    names = list(sheet_names or [])
    if sheet_name and sheet_name in names:
        return sheet_name
    for n in names:
        if "出货" in n:
            return n
    return names[0] if names else ""


def _decompose_from_grid(
    file_path: str,
    sheet_title: str,
    nrows: int,
    ncols: int,
    get_cell_value,
    merged_cells_count: int,
    dimension: str,
    sheet_name: str | None,
    sample_rows: int,
) -> tuple[dict, int]:
    from openpyxl.utils import get_column_letter

    max_r = min(max(nrows, 1), 30)
    max_c = min(max(ncols, 1), 25)

    header_row_idx = None
    header_cells: list[dict[str, Any]] = []

    for r in range(1, max_r + 1):
        row_cells = []
        for c in range(1, max_c + 1):
            v = get_cell_value(r, c)
            if isinstance(v, str) and v.strip():
                row_cells.append(
                    {
                        "name": v.strip(),
                        "column": get_column_letter(c),
                        "column_index": c,
                    }
                )
        if len(row_cells) >= 4:
            header_row_idx = r
            header_cells = row_cells
            break

    if header_row_idx is None:
        header_row_idx = 1

    samples = []
    data_start = header_row_idx + 1
    data_end = min(max(nrows, 1), data_start + max(int(sample_rows), 1) - 1)

    for r in range(data_start, data_end + 1):
        row_data: dict[str, Any] = {}
        non_empty = False
        for h in header_cells:
            v = get_cell_value(r, h["column_index"])
            safe = _json_safe_cell_value(v)
            if safe is not None and safe != "":
                non_empty = True
            row_data[h["name"]] = safe
        if non_empty and row_data:
            samples.append(row_data)

    amount_related = [
        h for h in header_cells if any(k in h["name"] for k in ["金额", "单价", "价格", "数量"])
    ]

    result = {
        "success": True,
        "template": {
            "name": os.path.basename(file_path),
            "path": file_path,
            "sheet": sheet_title,
            "dimension": dimension,
            "max_row": max(nrows, 1),
            "max_column": max(ncols, 1),
        },
        "decomposition": {
            "header_row": header_row_idx,
            "editable_entries": header_cells,
            "amount_related_entries": amount_related,
            "sample_rows": samples,
            "merged_cells_count": merged_cells_count,
        },
    }
    return result, 200


def _decompose_template_xls_pandas(
    file_path: str, sheet_name=None, sample_rows=5
) -> tuple[dict, int]:
    try:
        import pandas as pd
    except ImportError as e:
        return {"success": False, "message": f"读取 .xls 需要 pandas：{e}"}, 500

    try:
        xl = pd.ExcelFile(file_path, engine="xlrd")
    except RECOVERABLE_ERRORS as e:
        logger.error("pandas/xlrd 打开 .xls 失败: %s", e)
        if _is_unreadable_workbook_error(str(e)):
            return {
                "success": False,
                "message": "无法读取该 .xls 文件（可能损坏）。请另存为 .xlsx 后重试。",
                "error_code": "UNREADABLE_WORKBOOK",
            }, 200
        hint = "请确认已安装 xlrd：pip install xlrd"
        if "xlrd" in str(e).lower() or "no module" in str(e).lower():
            return {"success": False, "message": f"{hint}。原始错误：{e}"}, 500
        return {"success": False, "message": f"{hint}。原始错误：{e}"}, 500

    names = list(xl.sheet_names)
    if not names:
        return {"success": False, "message": "工作簿中没有工作表"}, 200

    sn = _pick_sheet_name(names, sheet_name)
    df = pd.read_excel(file_path, sheet_name=sn, header=None, dtype=object, engine="xlrd")
    nrows, ncols = int(df.shape[0]), int(df.shape[1])

    def get_cell_value(r: int, c: int) -> Any:
        if r < 1 or c < 1 or r > nrows or c > ncols:
            return None
        v = df.iat[r - 1, c - 1]
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return None
        return v

    from openpyxl.utils import get_column_letter

    last_c = get_column_letter(max(ncols, 1))
    dimension = f"A1:{last_c}{max(nrows, 1)}"
    return _decompose_from_grid(
        file_path,
        str(sn),
        nrows,
        ncols,
        get_cell_value,
        0,
        dimension,
        sheet_name,
        sample_rows,
    )


def _decompose_template_openpyxl(file_path, sheet_name=None, sample_rows=5) -> tuple[dict, int]:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    elif "出货" in wb.sheetnames:
        ws = wb["出货"]
    else:
        ws = wb[wb.sheetnames[0]]

    def get_cell_value(r: int, c: int) -> Any:
        return ws.cell(r, c).value

    dim = ws.calculate_dimension() or "A1:A1"
    merged = len(ws.merged_cells.ranges)
    return _decompose_from_grid(
        file_path,
        ws.title,
        ws.max_row,
        ws.max_column,
        get_cell_value,
        merged,
        dim,
        sheet_name,
        sample_rows,
    )


def _decompose_template(file_path, sheet_name=None, sample_rows=5) -> tuple[dict, int]:
    try:
        if not os.path.exists(file_path):
            return {"success": False, "message": f"模板文件不存在: {file_path}"}, 404

        ext = os.path.splitext(file_path)[1].lower()
        if ext == ".xls":
            return _decompose_template_xls_pandas(file_path, sheet_name, sample_rows)

        try:
            return _decompose_template_openpyxl(file_path, sheet_name, sample_rows)
        except RECOVERABLE_ERRORS as oe:
            msg = str(oe)
            if _is_unreadable_workbook_error(msg):
                return {
                    "success": False,
                    "message": "模板文件损坏或格式异常，无法读取。请重新导出或另存为 .xlsx 后重试。",
                    "error_code": "UNREADABLE_WORKBOOK",
                }, 200
            raise

    except RECOVERABLE_ERRORS as e:
        logger.error("分解 Excel 模板失败: %s", e)
        if _is_unreadable_workbook_error(str(e)):
            return {
                "success": False,
                "message": "模板文件损坏或格式异常，无法读取。请重新导出或另存为 .xlsx 后重试。",
                "error_code": "UNREADABLE_WORKBOOK",
            }, 200
        return {"success": False, "message": str(e)}, 500


def list_templates_get():
    """与归档中后注册的 list_templates 一致（规范化 DTO）。"""
    try:
        templates = [_normalize_template_dto(t) for t in _get_template_list()]
        return JSONResponse({"success": True, "templates": templates}, status_code=200)
    except RECOVERABLE_ERRORS as e:
        logger.error("获取模板列表失败: %s", e)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


def get_templates_list():
    try:
        templates = _get_template_list()
        return JSONResponse({"success": True, "templates": templates})
    except RECOVERABLE_ERRORS as e:
        logger.error("获取模板列表失败：%s", e)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


def list_templates_by_type(
    type: str = Query(default="发货单"),
    active_only: str = Query(default="true"),
):
    try:
        from app.application import get_template_app_service

        active = active_only.lower() == "true"
        svc = get_template_app_service()
        templates = [_normalize_template_dto(t) for t in svc.list_by_type(type, active_only=active)]
        return JSONResponse(
            {"success": True, "templates": templates, "count": len(templates)}, status_code=200
        )
    except RECOVERABLE_ERRORS as e:
        logger.error("按类型获取模板列表失败：%s", e)
        return JSONResponse({"success": False, "message": f"获取失败：{str(e)}"}, status_code=500)


def get_default_template(type: str = Query(default="发货单")):
    try:
        from app.application import get_template_app_service

        svc = get_template_app_service()
        tpl = svc.get_default_for_type(type)
        if not tpl:
            return JSONResponse({"success": False, "message": "暂无可用模板"}, status_code=404)
        return JSONResponse(
            {"success": True, "template": _normalize_template_dto(tpl)}, status_code=200
        )
    except RECOVERABLE_ERRORS as e:
        logger.error("获取默认模板失败：%s", e)
        return JSONResponse({"success": False, "message": f"获取失败：{str(e)}"}, status_code=500)


def get_template_file(template_id: str):
    try:
        templates = _get_template_list()
        template = next((t for t in templates if t["id"] == template_id), None)
        if not template:
            return JSONResponse({"success": False, "message": "模板不存在"}, status_code=404)
        if not template.get("exists") or not template.get("path"):
            return JSONResponse({"success": False, "message": "模板文件不存在"}, status_code=404)
        return FileResponse(
            template["path"],
            filename=template["filename"],
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except RECOVERABLE_ERRORS as e:
        logger.error("获取模板文件失败: %s", e)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


def save_template(data: dict[str, Any] = Body(default_factory=dict)):
    try:
        from app.application import get_template_app_service

        source_name = data.get("source_name", "尹玉华132.xlsx")
        target_name = data.get("target_name", "发货单模板.xlsx")
        overwrite = bool(data.get("overwrite", False))
        result = get_template_app_service().save_template_file(source_name, target_name, overwrite)
        status = 200 if result.get("success") else 404
        return JSONResponse(result, status_code=status)
    except RECOVERABLE_ERRORS as e:
        logger.error("保存模板失败: %s", e)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


def decompose_template(data: dict[str, Any] = Body(default_factory=dict)):
    try:
        filename = data.get("filename")
        file_path = data.get("file_path")
        sheet_name = data.get("sheet_name")
        sample_rows = data.get("sample_rows", 5)
        if file_path:
            target_path = file_path
        elif filename:
            target_path = _resolve_template_path(filename)
        else:
            return JSONResponse(
                {"success": False, "message": "请提供 filename 或 file_path"}, status_code=400
            )
        if not target_path or not os.path.exists(target_path):
            return JSONResponse({"success": False, "message": "模板文件不存在"}, status_code=404)
        result, status = _decompose_template(target_path, sheet_name, sample_rows)
        return JSONResponse(result, status_code=status)
    except RECOVERABLE_ERRORS as e:
        logger.error("分解模板失败: %s", e)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


async def upload_excel(excel_file: UploadFile | None = File(default=None)):
    try:
        if excel_file is None or not excel_file.filename:
            return JSONResponse({"success": False, "message": "请上传 Excel 文件"}, status_code=400)
        if not excel_file.filename.lower().endswith((".xlsx", ".xls")):
            return JSONResponse(
                {"success": False, "message": "只支持 .xlsx 和 .xls 格式"}, status_code=400
            )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"excel_{timestamp}_{excel_file.filename}"
        file_path = os.path.join(TEMP_EXCEL_DIR, filename)
        body = await excel_file.read()
        with open(file_path, "wb") as f:
            f.write(body)
        logger.info("Excel 文件已上传: %s", file_path)
        return JSONResponse(
            {
                "success": True,
                "file_path": file_path,
                "filename": excel_file.filename,
                "message": "文件上传成功",
            }
        )
    except RECOVERABLE_ERRORS as e:
        logger.error("上传 Excel 文件失败: %s", e)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


def excel_templates_test():
    return JSONResponse(
        {
            "success": True,
            "message": "Excel 模板服务运行正常",
            "timestamp": datetime.now().isoformat(),
        }
    )


def get_template(template_id: int):
    try:
        from app.db.session import get_db

        with get_db() as db:
            result = db.execute(
                text("SELECT * FROM templates WHERE id = :id AND is_active = 1"),
                {"id": template_id},
            )
            row = result.fetchone()
            if not row:
                return JSONResponse({"success": False, "message": "模板不存在"}, status_code=404)
            template = {
                "id": row.id,
                "template_key": row.template_key,
                "template_name": row.template_name,
                "template_type": row.template_type,
                "original_file_path": row.original_file_path,
                "analyzed_data": json.loads(row.analyzed_data) if row.analyzed_data else None,
                "editable_config": json.loads(row.editable_config) if row.editable_config else None,
                "zone_config": json.loads(row.zone_config) if row.zone_config else None,
                "merged_cells_config": (
                    json.loads(row.merged_cells_config) if row.merged_cells_config else None
                ),
                "style_config": json.loads(row.style_config) if row.style_config else None,
                "business_rules": json.loads(row.business_rules) if row.business_rules else None,
                "created_at": row.created_at,
                "updated_at": row.updated_at,
            }
            return JSONResponse({"success": True, "template": template})
    except RECOVERABLE_ERRORS as e:
        logger.error("获取模板详情失败：%s", e)
        return JSONResponse({"success": False, "message": f"获取失败：{str(e)}"}, status_code=500)


def update_template(template_id: int, data: dict[str, Any] = Body(default_factory=dict)):
    try:
        from app.db.session import get_db

        with get_db() as db:
            result = db.execute(
                text("SELECT id FROM templates WHERE id = :id"), {"id": template_id}
            )
            if not result.fetchone():
                return JSONResponse({"success": False, "message": "模板不存在"}, status_code=404)
            updates = []
            params: dict[str, Any] = {"id": template_id}
            if "template_name" in data:
                updates.append("template_name = :template_name")
                params["template_name"] = data["template_name"]
            if "template_type" in data:
                updates.append("template_type = :template_type")
                params["template_type"] = data["template_type"]
            if "editable_config" in data:
                updates.append("editable_config = :editable_config")
                params["editable_config"] = json.dumps(data["editable_config"], ensure_ascii=False)
            if "zone_config" in data:
                updates.append("zone_config = :zone_config")
                params["zone_config"] = json.dumps(data["zone_config"], ensure_ascii=False)
            if "business_rules" in data:
                updates.append("business_rules = :business_rules")
                params["business_rules"] = json.dumps(data["business_rules"], ensure_ascii=False)
            updates.append("updated_at = :updated_at")
            params["updated_at"] = datetime.now()
            sql = "UPDATE templates SET " + ", ".join(updates) + " WHERE id = :id"
            db.execute(text(sql), params)
            db.commit()
            db.execute(
                text(
                    """
                    INSERT INTO template_usage_log (template_id, action, result)
                    VALUES (:template_id, 'update', :result)
                """
                ),
                {"template_id": template_id, "result": "更新模板配置"},
            )
            db.commit()
        return JSONResponse({"success": True, "message": "模板更新成功"})
    except RECOVERABLE_ERRORS as e:
        logger.error("更新模板失败：%s", e)
        return JSONResponse({"success": False, "message": f"更新失败：{str(e)}"}, status_code=500)


def delete_template(template_id: int):
    try:
        from app.db.session import get_db

        with get_db() as db:
            result = db.execute(
                text("SELECT id FROM templates WHERE id = :id"), {"id": template_id}
            )
            if not result.fetchone():
                return JSONResponse({"success": False, "message": "模板不存在"}, status_code=404)
            db.execute(
                text("UPDATE templates SET is_active = 0, updated_at = :updated_at WHERE id = :id"),
                {"id": template_id, "updated_at": datetime.now()},
            )
            db.execute(
                text(
                    """
                    INSERT INTO template_usage_log (template_id, action, result)
                    VALUES (:template_id, 'delete', :result)
                """
                ),
                {"template_id": template_id, "result": "删除模板"},
            )
            db.commit()
        return JSONResponse({"success": True, "message": "模板删除成功"})
    except RECOVERABLE_ERRORS as e:
        logger.error("删除模板失败：%s", e)
        return JSONResponse({"success": False, "message": f"删除失败：{str(e)}"}, status_code=500)
