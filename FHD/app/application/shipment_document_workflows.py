"""Reporting and document-generation workflows for shipment application service."""

from __future__ import annotations

import logging
import os
from datetime import datetime
from typing import Any, cast

from app.application.ports import ShipmentDocumentGeneratorPort, ShipmentRecordStorePort
from app.utils.operational_errors import RECOVERABLE_ERRORS

logger = logging.getLogger(__name__)


class ShipmentDocumentWorkflowMixin:
    _document_generator: ShipmentDocumentGeneratorPort | None
    _record_store: ShipmentRecordStorePort | None

    def get_shipment_records(self, unit_name: str | None = None) -> list[dict[str, Any]]:
        raise NotImplementedError

    def export_shipment_records(
        self,
        unit_name: str | None = None,
        template_id: str | None = None,
        status_filter: str | None = None,
    ) -> dict[str, Any]:
        """
        导出出货记录为 Excel 文件。
        说明：该功能偏 I/O/报表层，仍放在 application 层做编排。
        """
        try:
            from openpyxl import Workbook

            from app.utils.excel.template_export_utils import fill_workbook_from_template
            from app.utils.path_io.path_utils import get_data_dir

            records = self.get_shipment_records(unit_name)
            normalized_status = str(status_filter or "").strip().lower()
            if normalized_status:
                if normalized_status in ("printed", "已打印"):
                    records = [
                        r
                        for r in records
                        if str(r.get("status") or "").strip().lower() == "printed"
                    ]
                elif normalized_status in ("pending", "未打印"):
                    records = [
                        r
                        for r in records
                        if str(r.get("status") or "").strip().lower() in ("pending", "")
                    ]

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            unit_prefix = unit_name if unit_name else "all"
            filename = f"shipment_records_{unit_prefix}_{timestamp}.xlsx"

            export_dir = os.path.join(get_data_dir(), "exports")
            os.makedirs(export_dir, exist_ok=True)
            file_path = os.path.join(export_dir, filename)
            template_path = None
            if template_id:
                try:
                    from app.application import get_template_app_service

                    templates = (get_template_app_service().get_templates() or {}).get(
                        "templates"
                    ) or []
                    target = next(
                        (t for t in templates if str(t.get("id")) == str(template_id)), None
                    )
                    if not target:
                        return {
                            "success": False,
                            "message": "导出失败：所选模板不存在，请重新选择模板",
                            "file_path": None,
                            "filename": None,
                            "count": 0,
                        }
                    target_scope = str(target.get("business_scope") or "").strip()
                    target_type = str(target.get("template_type") or "").strip()
                    if (
                        target_scope
                        and target_scope != "shipmentRecords"
                        and target_type != "出货记录"
                    ):
                        return {
                            "success": False,
                            "message": "导出失败：所选模板不属于出货记录范围，请重新选择",
                            "file_path": None,
                            "filename": None,
                            "count": 0,
                        }
                    candidate_path = str(
                        target.get("path") or target.get("file_path") or ""
                    ).strip()
                    if not candidate_path:
                        return {
                            "success": False,
                            "message": "导出失败：所选模板未绑定 Excel 文件，请先在业务对接中上传并替换",
                            "file_path": None,
                            "filename": None,
                            "count": 0,
                        }
                    if not os.path.exists(candidate_path):
                        return {
                            "success": False,
                            "message": "导出失败：所选模板文件不存在，请重新上传模板后重试",
                            "file_path": None,
                            "filename": None,
                            "count": 0,
                        }
                    template_path = candidate_path
                    preview_data = (
                        target.get("preview_data")
                        if isinstance(target.get("preview_data"), dict)
                        else {}
                    )
                    business_rules = (
                        target.get("business_rules")
                        if isinstance(target.get("business_rules"), dict)
                        else {}
                    )
                    template_sheet_name = str(
                        preview_data.get("selected_sheet_name")
                        or business_rules.get("selected_sheet_name")
                        or preview_data.get("sheet_name")
                        or ""
                    ).strip()
                except RECOVERABLE_ERRORS:
                    logger.exception("读取出货记录模板信息失败")
                    return {
                        "success": False,
                        "message": "导出失败：读取模板信息异常",
                        "file_path": None,
                        "filename": None,
                        "count": 0,
                    }

            if template_path:
                header_alias = {
                    "purchase_unit": ["购买单位", "单位"],
                    "product_name": ["产品名称", "品名"],
                    "model_number": ["型号", "产品型号"],
                    "quantity_kg": ["数量", "数量/KG", "数量(kg)"],
                    "quantity_tins": ["数量/件", "数量/桶"],
                    "tin_spec": ["规格"],
                    "unit_price": ["单价", "单价/元"],
                    "amount": ["金额", "金额/元"],
                    "status": ["状态"],
                    "created_at": ["创建时间"],
                    "printed_at": ["打印时间"],
                    "printer_name": ["打印机"],
                }
                wb = fill_workbook_from_template(
                    template_path=template_path,
                    records=records,
                    field_alias_map=header_alias,
                    sheet_name=template_sheet_name or "出货记录",
                    clear_existing_data_rows_all_columns=True,
                    truncate_rows_after_data_area=True,
                    clear_rows_above_header=True,
                )
                # Export is delivered as shipment records only. Remove other worksheets
                # to avoid carrying historical/template auxiliary data into result file.
                used_sheet = (template_sheet_name or "出货记录").strip()
                if used_sheet not in wb.sheetnames and wb.sheetnames:
                    used_sheet = wb.sheetnames[0]
                for sheet in list(wb.sheetnames):
                    if sheet != used_sheet and len(wb.sheetnames) > 1:
                        wb.remove(wb[sheet])
                wb.save(file_path)
            else:
                wb = Workbook()
                ws = wb.active
                if ws is None:
                    ws = wb.create_sheet("出货记录")
                else:
                    ws.title = "出货记录"

                headers = [
                    "ID",
                    "购买单位",
                    "产品名称",
                    "型号",
                    "数量 (KG)",
                    "数量 (桶)",
                    "规格",
                    "单价",
                    "金额",
                    "状态",
                    "创建时间",
                    "打印时间",
                    "打印机",
                ]
                ws.append(headers)

                for r in records:
                    ws.append(
                        [
                            r.get("id"),
                            r.get("purchase_unit") or "",
                            r.get("product_name") or "",
                            r.get("model_number") or "",
                            r.get("quantity_kg") or 0,
                            r.get("quantity_tins") or 0,
                            r.get("tin_spec") or "",
                            r.get("unit_price") or 0,
                            r.get("amount") or 0,
                            r.get("status") or "",
                            (
                                created_at.strftime("%Y-%m-%d %H:%M:%S")
                                if (created_at := r.get("created_at")) is not None
                                and hasattr(created_at, "strftime")
                                else ""
                            ),
                            (
                                printed_at.strftime("%Y-%m-%d %H:%M:%S")
                                if (printed_at := r.get("printed_at")) is not None
                                and hasattr(printed_at, "strftime")
                                else ""
                            ),
                            r.get("printer_name") or "",
                        ]
                    )
                wb.save(file_path)

            return {
                "success": True,
                "file_path": str(file_path),
                "filename": filename,
                "count": len(records),
                "template_used": template_path or "",
            }
        except RECOVERABLE_ERRORS:
            logger.exception("导出出货记录失败")
            return {
                "success": False,
                "message": "导出失败，请稍后重试",
                "file_path": None,
                "filename": None,
                "count": 0,
            }

    def generate_shipment_document(
        self,
        *,
        unit_name: str,
        products: list[dict[str, Any]],
        date: str | None = None,
        template_name: str | None = None,
        order_number: str | None = None,
        template_id: str | None = None,
        preferred_template: str | None = None,
        intent: str = "shipment_generate",
        allow_products_from_db: bool = False,
        strict_template: bool | None = None,
        raw_text: str = "",
    ) -> dict[str, Any]:
        """生成发货单文档（用例编排）。

        未指定模版时从模版库按意图 / 客户 / 偏好解析默认模版，闭合
        ingest → templates → 打单 断点。
        """
        if not self._document_generator:
            return {
                "success": False,
                "message": "document_generator 未配置",
                "error_code": "DOCUMENT_GENERATOR_MISSING",
                "doc_name": None,
                "file_path": None,
            }

        products_source = "request"
        product_rows = list(products or [])
        if not product_rows and allow_products_from_db:
            from app.application.shipment_template_resolve import resolve_products_for_unit

            product_rows = resolve_products_for_unit(unit_name)
            if product_rows:
                products_source = "db_latest_shipment"
        if not product_rows:
            return {
                "success": False,
                "message": "产品列表不能为空",
                "error_code": "PRODUCTS_REQUIRED",
                "doc_name": None,
                "file_path": None,
            }

        from app.application.shipment_template_resolve import (
            resolve_shipment_template,
            shipment_template_strict_enabled,
        )

        strict = shipment_template_strict_enabled(strict_template)
        resolved_template = template_name
        template_meta: dict[str, Any] = {}
        try:
            resolved = resolve_shipment_template(
                template_id=template_id,
                template_name=template_name,
                preferred=preferred_template,
                unit_name=unit_name,
                intent=intent,
                strict=strict,
                log_usage=False,
            )
            template_meta = resolved
            if resolved.get("path"):
                resolved_template = str(resolved["path"])
            elif strict and not str(template_name or "").strip():
                return {
                    "success": False,
                    "message": "未找到可用发货单模版，请先入库或指定 template_id",
                    "error_code": resolved.get("error_code") or "TEMPLATE_NOT_FOUND",
                    "template_resolution": {
                        "template_id": resolved.get("template_id"),
                        "template_name": resolved.get("template_name"),
                        "reason": resolved.get("reason"),
                        "error_code": resolved.get("error_code"),
                    },
                    "doc_name": None,
                    "file_path": None,
                }
        except RECOVERABLE_ERRORS as exc:
            logger.warning("模版库解析失败，回退 legacy: %s", exc)
            if strict:
                return {
                    "success": False,
                    "message": f"模版解析失败：{exc}",
                    "error_code": "TEMPLATE_RESOLVE_FAILED",
                    "doc_name": None,
                    "file_path": None,
                }

        result = self._document_generator.generate(
            unit_name=unit_name,
            products=product_rows,
            date=date,
            template_name=resolved_template,
            order_number=order_number,
        )
        if isinstance(result, dict):
            result["products_source"] = products_source
            if template_meta:
                result["template_resolution"] = {
                    "template_id": template_meta.get("template_id"),
                    "template_name": template_meta.get("template_name"),
                    "template_type": template_meta.get("template_type"),
                    "source": template_meta.get("source"),
                    "reason": template_meta.get("reason"),
                    "path": template_meta.get("path"),
                    "score": template_meta.get("score"),
                    "error_code": template_meta.get("error_code"),
                    "ok": template_meta.get("ok"),
                }
                if result.get("success") and template_meta.get("template_id"):
                    try:
                        from app.application.shipment_template_resolve import log_template_usage

                        log_template_usage(
                            str(template_meta.get("template_id")),
                            action="generate",
                            result_text=(
                                f"{template_meta.get('reason')}|unit={unit_name}|"
                                f"ok={bool(result.get('success'))}"
                            ),
                        )
                    except RECOVERABLE_ERRORS:
                        pass
        if result.get("success") and self._record_store:
            try:
                record_products = result.get("parsed_products") or product_rows
                record_result = self._record_store.record_document_generation(
                    unit_name=result.get("purchase_unit") or unit_name,
                    unit_id=result.get("unit_id"),
                    products=record_products,
                    document_result=result,
                    raw_text=str(raw_text or ""),
                )
                record_id = (record_result or {}).get("record_id")
                if record_id:
                    # 向前兼容：历史前端把 order_id 当作 shipment_records 主键使用。
                    result["record_id"] = record_id
                    result["order_id"] = record_id
            except RECOVERABLE_ERRORS:
                # 记录写入失败不影响文档生成返回
                pass
        return cast("dict[str, Any]", result)
