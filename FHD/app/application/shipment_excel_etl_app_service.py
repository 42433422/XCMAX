"""Excel 单据 ETL：通用识别 → 知识库记忆 → 预览/入库 → 模板回写。

默认不依赖仓库内置送货单 YAML；版式来自：
- 知识库 ``excel_etl_kb``（同义词 + 可学习表头指纹）
- 可选 ``FHD_EXCEL_ETL_PROFILE_DIR`` 用户 YAML
- 仅当 ``FHD_EXCEL_ETL_ALLOW_BUILTIN=1`` 时加载 examples/

闭环能力：
- preview / execute（指纹幂等）
- batch 目录扫描
- 通用表写出 / 流水写出，并从 notes 反推再出单
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Any

from app.application.excel_etl_kb import (
    TemplateMemory,
    get_excel_etl_kb,
    sheet_layout_fingerprint,
)
from app.application.shipment_etl_profile import (
    ShipmentEtlProfile,
    column_rule_matches,
    get_shipment_etl_profile,
    header_groups_match,
)
from app.utils.operational_errors import RECOVERABLE_ERRORS

logger = logging.getLogger(__name__)


def _resolve_profile(
    profile: ShipmentEtlProfile | None = None, profile_id: str | None = None
) -> ShipmentEtlProfile:
    if profile is not None:
        return profile
    return get_shipment_etl_profile(profile_id)


def _profiles_for_parse(
    profile: ShipmentEtlProfile | None = None,
    profile_id: str | None = None,
) -> list[ShipmentEtlProfile]:
    """解析用 profile 列表：显式指定则单 profile；否则加载全部做竞分。"""
    if profile is not None:
        return [profile]
    raw = str(profile_id or "").strip()
    if not raw:
        import os

        raw = (
            str(os.environ.get("FHD_EXCEL_ETL_PROFILE") or "").strip()
            or str(os.environ.get("FHD_SHIPMENT_ETL_PROFILE") or "").strip()
        )
    if raw and raw.lower() not in {"auto", "*"}:
        return [get_shipment_etl_profile(raw)]
    from app.application.shipment_etl_profile import load_all_profiles

    profiles = load_all_profiles()
    return profiles or [get_shipment_etl_profile("universal")]


def _pick_best_profile_for_sheet(
    ws, profiles: list[ShipmentEtlProfile]
) -> tuple[ShipmentEtlProfile, int, int, str]:
    """返回 (profile, delivery_score, ledger_score, prefer_kind)."""
    best: tuple[ShipmentEtlProfile, int, int, str] | None = None
    best_score = -1
    for prof in profiles:
        d = _score_delivery_sheet(ws, prof)
        l = _score_ledger_sheet(ws, prof) if prof.has_ledger else 0
        if d >= l and d > best_score:
            best = (prof, d, l, "delivery_note")
            best_score = d
        elif l > best_score:
            best = (prof, d, l, "shipment_ledger")
            best_score = l
    if best is None:
        fallback = profiles[0]
        return fallback, 0, 0, "delivery_note"
    return best


def _norm_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ").strip()
    return re.sub(r"\s+", "", text)


def _norm_header(value: Any) -> str:
    return _norm_cell(value).lower()


def _to_float(value: Any, default: float = 0.0) -> float:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    try:
        return float(text)
    except ValueError:
        return default


def _to_int(value: Any, default: int = 0) -> int:
    try:
        return int(round(_to_float(value, float(default))))
    except (TypeError, ValueError):
        return default


def _row_texts(ws, row: int, max_col: int = 16) -> list[str]:
    out: list[str] = []
    for col in range(1, max_col + 1):
        raw = ws.cell(row, col).value
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            out.append(text)
    return out


def _joined_row(ws, row: int, max_col: int = 16) -> str:
    return " ".join(_row_texts(ws, row, max_col))


def _token_in_compact(token: str, compact: str) -> bool:
    """忽略斜杠差异的包含匹配。"""
    t = str(token or "")
    if not t:
        return False
    if t in compact:
        return True
    return t.replace("/", "") in compact.replace("/", "").lower() or t.lower() in compact.lower()


def _header_cell_texts(ws, header_row: int, max_col: int = 16) -> list[str]:
    out: list[str] = []
    for col in range(1, min(max_col, int(ws.max_column or 0) or max_col) + 1):
        raw = ws.cell(header_row, col).value
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            out.append(text)
    return out


def _kb_resolve_layout(ws) -> tuple[int | None, dict[str, int], str]:
    """按表头指纹查知识库；命中则返回 (header_row, columns, fingerprint)。"""
    kb = get_excel_etl_kb()
    max_row = int(ws.max_row or 0)
    max_col = min(16, int(ws.max_column or 0) or 16)
    for row in range(1, min(20, max_row) + 1):
        headers = _header_cell_texts(ws, row, max_col=max_col)
        if len(headers) < 2:
            continue
        fp = sheet_layout_fingerprint(
            sheet_title=str(ws.title or ""),
            header_cells=headers,
        )
        mem = kb.get_template(fp)
        if mem is None or not mem.columns:
            continue
        if mem.header_row is not None and int(mem.header_row) != row:
            # 指纹按当前行 headers 算；若记忆行号不一致仍以当前行为准
            pass
        if "product_name" not in mem.columns and "model_number" not in mem.columns:
            continue
        kb.touch(fp)
        return row, {str(k): int(v) for k, v in mem.columns.items()}, fp
    return None, {}, ""


def _remember_sheet_layout(
    ws,
    *,
    header_row: int,
    mapping: dict[str, int],
    profile: ShipmentEtlProfile,
    source: str = "learned",
) -> str:
    """解析成功后把表头映射写入知识库。"""
    if not mapping or header_row is None:
        return ""
    headers = _header_cell_texts(ws, header_row)
    if len(headers) < 2:
        return ""
    fp = sheet_layout_fingerprint(
        sheet_title=str(ws.title or ""),
        header_cells=headers,
    )
    try:
        get_excel_etl_kb().remember(
            TemplateMemory(
                fingerprint=fp,
                label=str(profile.label or profile.id),
                target=str(profile.target or "preview_only"),
                header_row=int(header_row),
                columns={str(k): int(v) for k, v in mapping.items()},
                meta={},
                write=dict(profile.write or {}),
                source=source,
            )
        )
    except RECOVERABLE_ERRORS:
        logger.debug("excel etl kb remember skipped", exc_info=True)
        return ""
    return fp


def _score_delivery_sheet(ws, profile: ShipmentEtlProfile) -> int:
    """内容指纹打分：规则来自 profile.detect.delivery。"""
    cfg = profile.detect.get("delivery") or {}
    probe_n = int(cfg.get("probe_rows") or 8)
    probe_rows = min(probe_n, int(ws.max_row or 0))
    blob = " ".join(_joined_row(ws, r) for r in range(1, probe_rows + 1))
    compact = _norm_cell(blob)
    score = 0
    if profile.meta_patterns.title.search(blob):
        score += int(cfg.get("title_weight") or 50)
    buyer_token = str(cfg.get("buyer_token") or "")
    if buyer_token and buyer_token in compact:
        score += int(cfg.get("buyer_weight") or 25)
    header_hits = 0
    for token in cfg.get("header_hit_tokens") or []:
        if _token_in_compact(str(token), compact):
            header_hits += 1
    score += min(header_hits, int(cfg.get("header_hit_cap") or 5)) * int(
        cfg.get("header_hit_weight") or 6
    )
    for bonus in cfg.get("bonus_tokens") or []:
        if not isinstance(bonus, dict):
            continue
        tok = str(bonus.get("token") or "")
        if tok and tok in compact:
            score += int(bonus.get("weight") or 0)
    return score


def _score_ledger_sheet(ws, profile: ShipmentEtlProfile) -> int:
    """出货流水打分：规则来自 profile.detect.ledger。"""
    cfg = profile.detect.get("ledger") or {}
    suppress_at = int(cfg.get("suppress_if_delivery_score_gte") or 60)
    if _score_delivery_sheet(ws, profile) >= suppress_at:
        return 0
    probe_n = int(cfg.get("probe_rows") or 10)
    probe_rows = min(probe_n, int(ws.max_row or 0))
    blob = " ".join(_joined_row(ws, r) for r in range(1, probe_rows + 1))
    compact = _norm_cell(blob)
    compact_l = compact.lower()
    score = 0
    sheet_hit = bool(profile.meta_patterns.ledger_sheet.search(str(ws.title or "")))
    content_tokens = [str(t) for t in (cfg.get("content_tokens") or [])]
    if sheet_hit or any(t.lower() in compact_l for t in content_tokens if t):
        score += int(cfg.get("sheet_weight") or 20)
    hits = 0
    for token in cfg.get("hit_tokens") or []:
        tok = str(token)
        if tok and (tok in compact or tok.lower() in compact_l):
            hits += 1
    score += min(hits, int(cfg.get("hit_cap") or 6)) * int(cfg.get("hit_weight") or 10)
    bonus_req = str(cfg.get("bonus_require_token") or "")
    bonus_exc = str(cfg.get("bonus_exclude_token") or "")
    if bonus_req and bonus_req in compact and (not bonus_exc or bonus_exc not in compact):
        score += int(cfg.get("bonus_weight") or 0)
    # 表头含「单号」列且无客户抬头 → 更像流水
    header_row = _find_ledger_header_row(ws, profile)
    if header_row is not None:
        mapping = _map_headers(ws, header_row, profile)
        if "order_number" in mapping and "客户" not in compact and "购货单位" not in compact:
            score += 25
    return score


def _find_header_row(ws, profile: ShipmentEtlProfile) -> int | None:
    cfg = profile.header_detect.get("delivery") or {}
    max_scan = int(cfg.get("max_scan_rows") or 12)
    groups = cfg.get("require_groups") or []
    for row in range(1, min(max_scan, int(ws.max_row or 0) + 1)):
        compact = _norm_header(_joined_row(ws, row))
        if header_groups_match(compact, groups):
            return row
    # 陌生表头：选「非空单元格最多」的候选行（至少 3 列）
    best_row = None
    best_count = 0
    for row in range(1, min(max_scan, int(ws.max_row or 0) + 1)):
        count = 0
        for col in range(1, min(16, int(ws.max_column or 0) or 16) + 1):
            raw = ws.cell(row, col).value
            if raw is not None and str(raw).strip():
                count += 1
        if count >= 3 and count > best_count:
            # 下一行最好有数据，避免把纯标题当表头
            has_body = False
            for r in range(row + 1, min(row + 4, int(ws.max_row or 0) + 1)):
                if any(
                    ws.cell(r, c).value not in (None, "")
                    for c in range(1, min(8, int(ws.max_column or 0) or 8) + 1)
                ):
                    has_body = True
                    break
            if has_body:
                best_row = row
                best_count = count
    return best_row


def _find_ledger_header_row(ws, profile: ShipmentEtlProfile) -> int | None:
    cfg = profile.header_detect.get("ledger") or {}
    max_scan = int(cfg.get("max_scan_rows") or 16)
    groups = cfg.get("require_groups") or []
    and_any = cfg.get("and_any_groups") or []
    for row in range(1, min(max_scan, int(ws.max_row or 0) + 1)):
        compact = _norm_header(_joined_row(ws, row))
        if not header_groups_match(compact, groups):
            continue
        if and_any:
            if not any(header_groups_match(compact, [g]) for g in and_any):
                continue
        return row
    return None


def _map_headers(ws, header_row: int, profile: ShipmentEtlProfile) -> dict[str, int]:
    mapping: dict[str, int] = {}
    # Preserve original field priority order from profile columns declaration.
    field_order = list(profile.columns.keys())
    for col in range(1, min(16, int(ws.max_column or 0) + 1)):
        key = _norm_header(ws.cell(header_row, col).value)
        if not key:
            continue
        for field_name in field_order:
            if field_name in mapping:
                continue
            for rule in profile.columns.get(field_name) or []:
                only_if_missing = [str(x) for x in (rule.get("only_if_missing") or [])]
                if only_if_missing and any(f in mapping for f in only_if_missing):
                    continue
                if column_rule_matches(key, rule):
                    mapping[field_name] = col
                    break
    return mapping


def _sample_values(ws, header_row: int, col: int, *, limit: int = 5) -> list[str]:
    out: list[str] = []
    for row in range(header_row + 1, min(header_row + 8, int(ws.max_row or 0) + 1)):
        raw = ws.cell(row, col).value
        if raw is None or str(raw).strip() == "":
            continue
        out.append(str(raw).strip())
        if len(out) >= limit:
            break
    return out


def _infer_columns_from_samples(
    ws,
    header_row: int,
    mapping: dict[str, int],
) -> dict[str, int]:
    """陌生表头：用样例值类型补列（不编造数值，只猜列位）。"""
    import os

    flag = str(os.environ.get("FHD_EXCEL_ETL_HEURISTIC") or "1").strip().lower()
    if flag in {"0", "false", "no", "off"}:
        return dict(mapping)
    out = dict(mapping)
    max_col = min(16, int(ws.max_column or 0) or 16)
    candidates: list[tuple[int, list[str], str]] = []
    for col in range(1, max_col + 1):
        if col in out.values():
            continue
        samples = _sample_values(ws, header_row, col)
        if not samples:
            continue
        header = _norm_header(ws.cell(header_row, col).value)
        joined = " ".join(samples)
        kind = "text"
        nums = 0
        for s in samples:
            try:
                float(str(s).replace(",", ""))
                nums += 1
            except ValueError:
                pass
        if nums >= max(1, len(samples) // 2 + 1):
            kind = "number"
        elif re.search(r"[A-Za-z0-9\-_/]{2,}", joined) and not re.search(
            r"[\u4e00-\u9fff]{2,}", joined
        ):
            kind = "code"
        elif re.search(r"[\u4e00-\u9fff]", joined):
            kind = "name"
        candidates.append((col, samples, kind if not header else f"{kind}:{header}"))

    def _take(field: str, predicate) -> None:
        if field in out:
            return
        for col, samples, kind in candidates:
            if col in out.values():
                continue
            if predicate(col, samples, kind):
                out[field] = col
                return

    _take(
        "model_number",
        lambda c, s, k: (
            k.startswith("code") or (k.startswith("text") and all(len(x) <= 24 for x in s))
        ),
    )
    _take("product_name", lambda c, s, k: k.startswith("name") or k.startswith("text"))
    # 数值列：按从左到右依次填 数量/规格/公斤/单价/金额
    num_cols = [c for c, s, k in candidates if k.startswith("number") and c not in out.values()]
    for field in ("quantity_tins", "tin_spec", "quantity_kg", "unit_price", "amount"):
        if field in out or not num_cols:
            continue
        out[field] = num_cols.pop(0)
    _take(
        "order_number",
        lambda c, s, k: any(re.search(r"[A-Za-z].*\d|\d.*[A-Za-z]", x) for x in s),
    )
    return out


def _classify_sheet_role(
    ws,
    profile: ShipmentEtlProfile,
    *,
    d_score: int,
    l_score: int,
) -> str:
    """多表混排：给工作表打角色 delivery / ledger / ignore / unknown。"""
    title = str(ws.title or "")
    if re.search(r"报价|价目|cover|目录|说明|readme", title, re.I):
        # 仍可能是单据；低分才忽略
        if d_score < 24 and l_score < 24:
            return "ignore"
    if profile.has_ledger and l_score >= 40 and l_score > d_score:
        return "ledger"
    if d_score >= 32:
        return "delivery"
    header = _find_header_row(ws, profile)
    ledger_header = _find_ledger_header_row(ws, profile) if profile.has_ledger else None
    if ledger_header and (header is None or l_score >= d_score):
        mapping = _map_headers(ws, ledger_header, profile)
        if "order_number" in mapping:
            return "ledger"
    if header is not None:
        mapping = _map_headers(ws, header, profile)
        if "product_name" in mapping or "model_number" in mapping:
            return "delivery"
    if d_score < 16 and l_score < 16:
        return "ignore"
    return "unknown"


def _parse_buyer_meta(ws, header_row: int, profile: ShipmentEtlProfile) -> dict[str, str]:
    meta = {
        "unit_name": "",
        "contact_person": "",
        "order_date": "",
        "order_number": "",
        "title": "",
    }
    mp = profile.meta_patterns
    for row in range(1, header_row):
        text = _joined_row(ws, row)
        if not text:
            continue
        if not meta["title"] and mp.title.search(text):
            meta["title"] = text.strip()
        buyer = mp.buyer.search(text.replace("　", " "))
        if buyer and not meta["unit_name"]:
            candidate = buyer.group(1).strip(" ：:　")
            candidate = re.sub(r"\s*\([^)]*\)\s*$", "", candidate).strip()
            if candidate and not _unit_name_looks_truncated(candidate):
                meta["unit_name"] = candidate
        contact = mp.contact.search(text)
        if contact and not meta["contact_person"]:
            meta["contact_person"] = contact.group(1).strip(" ：:　")
        date_m = mp.date.search(text)
        if date_m and not meta["order_date"]:
            meta["order_date"] = date_m.group(1).replace(" ", "")
        order_m = mp.order_no.search(text)
        if order_m and not meta["order_number"]:
            meta["order_number"] = order_m.group(1).strip()
    if not meta["unit_name"]:
        label = mp.buyer_label
        for row in range(1, header_row):
            text = _joined_row(ws, row)
            if label not in text:
                continue
            after = mp.buyer_split.split(text, maxsplit=1)
            if len(after) > 1:
                chunk = mp.buyer_stop.split(after[1], maxsplit=1)[0]
                meta["unit_name"] = chunk.strip(" ：:　")
                break
    adjacent = _extract_adjacent_buyer_meta(ws, header_row)
    from app.application.shipment_excel_etl_llm import unit_name_is_weak

    # 相邻格更准（Bill To / To: 分列）；整行拼接易吞到 Incoterms 等
    if adjacent.get("unit_name"):
        meta["unit_name"] = adjacent["unit_name"]
    if not meta["contact_person"] and adjacent.get("contact_person"):
        meta["contact_person"] = adjacent["contact_person"]
    if not meta["order_number"] and adjacent.get("order_number"):
        meta["order_number"] = adjacent["order_number"]
    if meta["unit_name"]:
        meta["unit_name"] = re.sub(r"\s*\([^)]*\)\s*$", "", meta["unit_name"]).strip()
        meta["unit_name"] = re.split(
            r"(?i)\s{2,}|\s+(?:Incoterms|Payment|Tel|Phone|地址|电话)\b",
            meta["unit_name"],
            maxsplit=1,
        )[0].strip()
    if meta["unit_name"] and (
        _unit_name_looks_truncated(meta["unit_name"]) or unit_name_is_weak(meta["unit_name"])
    ):
        meta["unit_name"] = ""
    return meta


_CORP_SUFFIX_ONLY = frozenset(
    {
        "ltd",
        "ltd.",
        "limited",
        "inc",
        "inc.",
        "llc",
        "pte",
        "pte.",
        "co",
        "co.",
        "corp",
        "corp.",
        "gmbh",
        "公司",
        "有限公司",
        "股份有限公司",
    }
)


def _unit_name_looks_truncated(unit: str) -> bool:
    text = str(unit or "").strip()
    if not text:
        return True
    if text.lower() in _CORP_SUFFIX_ONLY:
        return True
    # 单英词且过短，多为误切（如 Technologies→Ltd）
    if " " not in text and re.fullmatch(r"[A-Za-z.]{1,6}", text):
        return True
    return False


_BUYER_CELL_LABEL = re.compile(
    r"^(?:to|bill\s*to|sold\s*to|ship\s*to|consignee|customer|buyer|"
    r"购货单位|客户名称|客户|采购单位|收货单位|收货方|买方)\s*[:：]?$",
    re.IGNORECASE,
)
_BUYER_INLINE = re.compile(
    r"(?is)(?:bill\s*to|sold\s*to|ship\s*to|(?<![a-z])to|(?<![a-z])buyer|(?<![a-z])customer|"
    r"购货单位|客户名称|客户|采购单位)"
    r"\s*[:：]\s*([^\n·|]+?)(?=\s*(?:·|\||Incoterms|Payment|Tel|Phone|地址|电话)|$)"
)
_ATTN_CELL_LABEL = re.compile(r"^(?:attn|attention|联系人)\s*[:：]?$", re.IGNORECASE)
_ORDER_INLINE = re.compile(
    r"(?is)^(?:do\s*no|invoice\s*no|buyer\s*po|po\s*ref|订单号|订单编号|单号)\s*[:：]?\s*([A-Za-z0-9\-_/]+)"
)


def _extract_adjacent_buyer_meta(ws, header_row: int) -> dict[str, str]:
    """英文 DO/PI 常见：标签在 A 列、公司名在同行右侧单元格。"""
    from app.application.shipment_excel_etl_llm import unit_name_is_weak

    out = {"unit_name": "", "contact_person": "", "order_number": ""}
    max_col = min(int(getattr(ws, "max_column", 1) or 1), 16)
    scan_to = max(2, min(int(header_row or 2), 30))
    for row in range(1, scan_to):
        for col in range(1, max_col + 1):
            raw = str(ws.cell(row, col).value or "").strip()
            if not raw:
                continue
            inline = _BUYER_INLINE.search(raw)
            if inline and not out["unit_name"]:
                candidate = inline.group(1).strip().split("\n")[0].strip(" ：:　")
                # Bill To: Dukjil Trading Pte Ltd (BYR-001)
                candidate = re.sub(r"\s*\([^)]*\)\s*$", "", candidate).strip()
                if len(candidate) >= 2 and not unit_name_is_weak(candidate):
                    out["unit_name"] = candidate
                    continue
            if _BUYER_CELL_LABEL.match(raw) and not out["unit_name"]:
                for c2 in range(col + 1, min(col + 5, max_col + 1)):
                    val = str(ws.cell(row, c2).value or "").strip()
                    if not val or _BUYER_CELL_LABEL.match(val):
                        continue
                    candidate = val.split("\n")[0].strip()
                    candidate = re.sub(r"\s*\([^)]*\)\s*$", "", candidate).strip()
                    if candidate and not unit_name_is_weak(candidate):
                        out["unit_name"] = candidate
                    break
                continue
            if _ATTN_CELL_LABEL.match(raw) and not out["contact_person"]:
                for c2 in range(col + 1, min(col + 4, max_col + 1)):
                    val = str(ws.cell(row, c2).value or "").strip()
                    if val:
                        out["contact_person"] = val.split("\n")[0].strip()
                        break
                continue
            order_m = _ORDER_INLINE.match(raw)
            if order_m and not out["order_number"]:
                out["order_number"] = order_m.group(1).strip()
    return out


_NON_PRODUCT_TOKENS = frozenset(
    {
        "january",
        "february",
        "march",
        "april",
        "may",
        "june",
        "july",
        "august",
        "september",
        "october",
        "november",
        "december",
        "title",
        "identifier",
        "subject",
        "description",
        "notes",
        "creator",
        "accession",
        "my title",
        "another title",
        "the best image ever",
    }
)


def _looks_like_non_product_token(value: str) -> bool:
    text = str(value or "").strip().lower()
    return bool(text) and text in _NON_PRODUCT_TOKENS


def _looks_like_titleish(value: str) -> bool:
    text = str(value or "").strip().lower()
    if not text:
        return False
    if text in _NON_PRODUCT_TOKENS:
        return True
    return bool(re.fullmatch(r"[a-z][a-z\s\-']{2,40}", text)) and not re.search(r"\d", text)


def _build_item_from_row(ws, row: int, mapping: dict[str, int]) -> dict[str, Any] | None:
    model = ""
    name = ""
    if "model_number" in mapping:
        model = str(ws.cell(row, mapping["model_number"]).value or "").strip()
    if "product_name" in mapping:
        name = str(ws.cell(row, mapping["product_name"]).value or "").strip()
    if not name and model and not re.search(r"[A-Za-z0-9]", model):
        name, model = model, ""
    if not name and not model:
        return None
    # 拦截明显非商品语义（月份/档案标题当品名）
    if _looks_like_non_product_token(name) or _looks_like_non_product_token(model):
        return None
    tins = (
        _to_int(ws.cell(row, mapping["quantity_tins"]).value) if "quantity_tins" in mapping else 0
    )
    tin_spec = _to_float(ws.cell(row, mapping["tin_spec"]).value) if "tin_spec" in mapping else 0.0
    qty_kg = (
        _to_float(ws.cell(row, mapping["quantity_kg"]).value) if "quantity_kg" in mapping else 0.0
    )
    unit_price = (
        _to_float(ws.cell(row, mapping["unit_price"]).value) if "unit_price" in mapping else 0.0
    )
    amount = _to_float(ws.cell(row, mapping["amount"]).value) if "amount" in mapping else 0.0
    if tins <= 0 and qty_kg <= 0 and unit_price <= 0 and amount <= 0:
        return None
    # 年份当数量 + 标题当品名：档案表误入
    if 1990 <= tins <= 2035 and not model and _looks_like_titleish(name):
        return None
    if tin_spec <= 0 and tins > 0 and qty_kg > 0:
        tin_spec = qty_kg / tins
    if qty_kg <= 0 and tins > 0 and tin_spec > 0:
        qty_kg = tins * tin_spec
    if amount <= 0 and unit_price > 0 and qty_kg > 0:
        amount = unit_price * qty_kg
    if tins <= 0 and qty_kg > 0:
        tins = 1
        if tin_spec <= 0:
            tin_spec = qty_kg
    return {
        "product_name": name or model,
        "model_number": model,
        "quantity_tins": max(0, tins),
        "tin_spec": tin_spec or 0.0,
        "spec_per_tin": tin_spec or 0.0,
        "quantity_kg": qty_kg,
        "unit_price": unit_price,
        "amount": amount,
        "quantity": max(1, tins) if tins else 1,
    }


def _parse_items(
    ws, header_row: int, mapping: dict[str, int], profile: ShipmentEtlProfile
) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    max_row = int(ws.max_row or 0)
    for row in range(header_row + 1, max_row + 1):
        joined = _joined_row(ws, row)
        if not joined:
            continue
        if profile.meta_patterns.stop_row.search(joined):
            break
        item = _build_item_from_row(ws, row, mapping)
        if item:
            items.append(item)
    return items


def note_fingerprint(note: dict[str, Any]) -> str:
    """内容指纹：同客户+单号+明细再导入可幂等跳过。"""
    payload = {
        "unit": str(note.get("unit_name") or "").strip(),
        "order": str(note.get("order_number") or "").strip(),
        "date": str(note.get("order_date") or "").strip(),
        "items": sorted(
            [
                {
                    "m": str(i.get("model_number") or "").strip().upper(),
                    "n": str(i.get("product_name") or "").strip(),
                    "q": float(i.get("quantity_tins") or i.get("quantity") or 0),
                    "k": float(i.get("quantity_kg") or 0),
                    "p": float(i.get("unit_price") or 0),
                }
                for i in (note.get("items") or [])
            ],
            key=lambda x: (x["m"], x["n"], x["q"], x["k"], x["p"]),
        ),
    }
    raw = json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:28]


def _fingerprint_store_path() -> Path:
    """兼容旧测试 monkeypatch；真实幂等改走 SQLite。"""
    try:
        from app.utils.path_utils import get_data_dir

        root = Path(get_data_dir())
    except RECOVERABLE_ERRORS:
        root = Path.cwd() / "data"
    root.mkdir(parents=True, exist_ok=True)
    return root / "shipment_etl_fingerprints.json"


def _legacy_json_has_fingerprint(fingerprint: str) -> bool:
    path = _fingerprint_store_path()
    if not path.is_file():
        return False
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        entries = data.get("entries") if isinstance(data, dict) else None
        return bool(isinstance(entries, dict) and fingerprint in entries)
    except RECOVERABLE_ERRORS:
        return False


def _is_fingerprint_imported(tenant_key: str, fingerprint: str) -> bool:
    from app.application.shipment_excel_etl_fingerprint_store import has_fingerprint

    if has_fingerprint(tenant_key, fingerprint):
        return True
    # 兼容历史 JSON 指纹（无租户）
    return _legacy_json_has_fingerprint(fingerprint)


def _record_fingerprint_now(
    tenant_key: str,
    fingerprint: str,
    *,
    shipment_id: Any = None,
    unit_name: str = "",
    order_number: str = "",
    file_name: str = "",
) -> None:
    from app.application.shipment_excel_etl_fingerprint_store import record_fingerprint

    record_fingerprint(
        tenant_key,
        fingerprint,
        shipment_id=shipment_id,
        unit_name=unit_name,
        order_number=order_number,
        file_name=file_name,
    )


def _load_fingerprints() -> dict[str, Any]:
    path = _fingerprint_store_path()
    if not path.is_file():
        return {"entries": {}}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        if isinstance(data, dict) and isinstance(data.get("entries"), dict):
            return data
    except RECOVERABLE_ERRORS:
        logger.warning("failed to load shipment etl fingerprints", exc_info=True)
    return {"entries": {}}


def _save_fingerprints(data: dict[str, Any]) -> None:
    path = _fingerprint_store_path()
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def _enrich_note(note: dict[str, Any]) -> dict[str, Any]:
    out = dict(note)
    out["sheet_name"] = str(out.get("sheet_name") or out.get("sheet") or "")
    out["sheet"] = out["sheet_name"] or str(out.get("sheet") or "")
    out["fingerprint"] = note_fingerprint(out)
    out["item_count"] = len(out.get("items") or [])
    out["total_amount"] = round(
        sum(float(i.get("amount") or 0) for i in (out.get("items") or [])), 2
    )
    return out


def _build_sheet_probe(
    ws,
    profile: ShipmentEtlProfile,
    *,
    rule_hint: dict[str, Any] | None = None,
) -> Any:
    from app.application.shipment_excel_etl_llm import SheetProbe

    max_row = int(ws.max_row or 0)
    max_col = min(16, int(ws.max_column or 0) or 16)
    probe_n = min(10, max_row)
    probe_rows: list[dict[str, Any]] = []
    for row in range(1, probe_n + 1):
        cells = []
        for col in range(1, max_col + 1):
            raw = ws.cell(row, col).value
            if raw is None or str(raw).strip() == "":
                continue
            cells.append({"col": col, "text": str(raw).strip()[:80]})
        if cells:
            probe_rows.append({"row": row, "cells": cells})

    candidate_headers: list[dict[str, Any]] = []
    for row in range(1, min(16, max_row) + 1):
        cells = []
        for col in range(1, max_col + 1):
            raw = ws.cell(row, col).value
            text = str(raw).strip() if raw is not None else ""
            if not text:
                continue
            samples: list[str] = []
            for r in range(row + 1, min(row + 4, max_row + 1)):
                sv = ws.cell(r, col).value
                if sv is None or str(sv).strip() == "":
                    continue
                samples.append(str(sv).strip()[:40])
                if len(samples) >= 3:
                    break
            cells.append({"col": col, "header": text[:80], "samples": samples})
        if len(cells) >= 2:
            candidate_headers.append({"row": row, "cells": cells})

    return SheetProbe(
        profile_id=profile.id,
        sheet_title=str(ws.title or ""),
        probe_rows=probe_rows,
        candidate_headers=candidate_headers[:8],
        max_row=max_row,
        max_col=max_col,
        rule_hint=dict(rule_hint or {}),
    )


def _merge_meta(
    base: dict[str, str],
    overlay: dict[str, str],
    *,
    prefer_overlay: bool = False,
) -> dict[str, str]:
    out = dict(base)
    for key in ("unit_name", "contact_person", "order_date", "order_number", "title"):
        val = str((overlay or {}).get(key) or "").strip()
        if not val:
            continue
        if prefer_overlay or not str(out.get(key) or "").strip():
            out[key] = val
    return out


def _apply_llm_assist_to_layout(
    ws,
    profile: ShipmentEtlProfile,
    *,
    delivery_score: int,
    ledger_score: int,
    min_score: int,
    header_row: int | None,
    mapping: dict[str, int],
    meta: dict[str, str] | None,
    prefer_kind: str | None,
    fallback_unit: str = "",
) -> tuple[int | None, dict[str, int], dict[str, str], str | None, dict[str, Any]]:
    """低置信时请求 LLM；返回 (header_row, mapping, meta, source_kind, assist_public)."""
    from app.application.shipment_excel_etl_llm import (
        assist_sheet_layout,
        needs_llm_assist,
    )

    need, reason = needs_llm_assist(
        delivery_score=delivery_score,
        ledger_score=ledger_score,
        min_score=min_score,
        header_row=header_row,
        mapping=mapping,
        meta=meta,
        prefer_kind=prefer_kind,
        fallback_unit=fallback_unit,
    )
    assist_public: dict[str, Any] = {
        "used_llm": False,
        "cache_hit": False,
        "ok": False,
        "confidence": 1.0 if not need else 0.0,
        "reason": reason,
    }
    if not need:
        assist_public["ok"] = True
        return header_row, mapping, dict(meta or {}), prefer_kind, assist_public

    probe = _build_sheet_probe(
        ws,
        profile,
        rule_hint={
            "delivery_score": delivery_score,
            "ledger_score": ledger_score,
            "min_score": min_score,
            "prefer_kind": prefer_kind,
            "rule_header_row": header_row,
            "rule_mapping": mapping,
            "rule_meta": meta or {},
            "assist_reason": reason,
        },
    )
    assist = assist_sheet_layout(probe)
    assist_public = assist.as_public_dict()
    if not assist.ok:
        return header_row, mapping, dict(meta or {}), prefer_kind, assist_public

    new_header = assist.header_row if assist.header_row is not None else header_row
    new_mapping = dict(mapping)
    for field_name, col in (assist.columns or {}).items():
        if field_name not in new_mapping and isinstance(col, int) and col > 0:
            new_mapping[field_name] = col
    # Prefer LLM columns when rules were incomplete for that field
    for field_name, col in (assist.columns or {}).items():
        if field_name in {
            "product_name",
            "model_number",
            "order_number",
            "quantity_tins",
            "quantity_kg",
        }:
            if field_name not in mapping and isinstance(col, int) and col > 0:
                new_mapping[field_name] = col
    new_meta = _merge_meta(dict(meta or {}), assist.meta or {}, prefer_overlay=True)
    kind = (
        assist.source_kind
        if assist.source_kind in {"delivery_note", "shipment_ledger", "ignore"}
        else prefer_kind
    )
    return new_header, new_mapping, new_meta, kind, assist_public


def _parse_delivery_sheet(
    ws,
    *,
    fallback_unit: str,
    profile: ShipmentEtlProfile,
    allow_llm: bool = True,
) -> dict[str, Any] | None:
    d_score = _score_delivery_sheet(ws, profile)
    l_score = _score_ledger_sheet(ws, profile)
    kb_header, kb_mapping, kb_fp = _kb_resolve_layout(ws)
    header_row = kb_header if kb_header is not None else _find_header_row(ws, profile)
    mapping = (
        dict(kb_mapping)
        if kb_mapping
        else (_map_headers(ws, header_row, profile) if header_row is not None else {})
    )
    meta = (
        _parse_buyer_meta(ws, header_row, profile)
        if header_row is not None
        else {
            "unit_name": "",
            "contact_person": "",
            "order_date": "",
            "order_number": "",
            "title": "",
        }
    )
    assist_public: dict[str, Any] = {
        "used_llm": False,
        "cache_hit": bool(kb_fp),
        "ok": bool(kb_fp and mapping),
        "confidence": 1.0 if kb_fp else 1.0,
        "reason": "knowledge_base_hit" if kb_fp else "rules_only",
        "layout_fingerprint": kb_fp or "",
    }
    from app.application.shipment_excel_etl_llm import unit_name_is_weak

    unit_weak = unit_name_is_weak(str((meta or {}).get("unit_name") or ""), fallback=fallback_unit)
    # KB 命中但客户名仍弱/空时，仍允许 LLM 补 meta（不阻断规则列）
    if allow_llm and (not kb_fp or unit_weak):
        header_row, mapping, meta, kind, assist_public = _apply_llm_assist_to_layout(
            ws,
            profile,
            delivery_score=d_score,
            ledger_score=l_score,
            min_score=profile.delivery_min_score,
            header_row=header_row,
            mapping=mapping,
            meta=meta,
            prefer_kind="delivery_note",
            fallback_unit=fallback_unit,
        )
        if kind == "ignore":
            return None
        if kind == "shipment_ledger":
            # LLM reclassified as ledger — let caller handle via ledger path
            return None
        if kb_fp and not assist_public.get("layout_fingerprint"):
            assist_public["layout_fingerprint"] = kb_fp

    # 规则/LLM 仍缺关键列 → 样例启发式（陌生表头兜底；不编造数值）
    heuristic_on = str(os.environ.get("FHD_EXCEL_ETL_HEURISTIC") or "1").strip().lower() not in {
        "0",
        "false",
        "no",
        "off",
    }
    if (
        heuristic_on
        and header_row is not None
        and ("product_name" not in mapping and "model_number" not in mapping)
    ):
        inferred = _infer_columns_from_samples(ws, header_row, mapping)
        if "product_name" in inferred or "model_number" in inferred:
            mapping = inferred
            if not assist_public.get("used_llm"):
                assist_public = {
                    **assist_public,
                    "ok": True,
                    "confidence": 0.65,
                    "reason": "heuristic_samples",
                }

    if header_row is None:
        return None
    if "product_name" not in mapping and "model_number" not in mapping:
        return None
    items = _parse_items(ws, header_row, mapping, profile)
    if not items:
        return None
    remembered_fp = _remember_sheet_layout(
        ws,
        header_row=header_row,
        mapping=mapping,
        profile=profile,
        source="knowledge_base" if kb_fp else ("llm" if assist_public.get("used_llm") else "rules"),
    )
    if remembered_fp and not assist_public.get("layout_fingerprint"):
        assist_public["layout_fingerprint"] = remembered_fp
    unit = meta.get("unit_name") or fallback_unit
    note = _enrich_note(
        {
            "sheet": ws.title,
            "source_kind": "delivery_note",
            "score": d_score,
            "unit_name": unit,
            "contact_person": meta.get("contact_person") or "",
            "order_date": meta.get("order_date") or "",
            "order_number": meta.get("order_number") or "",
            "title": meta.get("title") or "",
            "items": items,
            "assist": assist_public,
        }
    )
    return note


def _excel_date_to_str(value: Any, profile: ShipmentEtlProfile) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return from_excel(value).strftime("%Y-%m-%d")
        except RECOVERABLE_ERRORS:
            return str(value)
    text = str(value).strip()
    date_m = profile.meta_patterns.date.search(text)
    return date_m.group(1).replace(" ", "") if date_m else text


def _parse_ledger_sheet(
    ws,
    *,
    fallback_unit: str,
    profile: ShipmentEtlProfile,
    allow_llm: bool = True,
) -> list[dict[str, Any]]:
    d_score = _score_delivery_sheet(ws, profile)
    l_score = _score_ledger_sheet(ws, profile)
    kb_header, kb_mapping, kb_fp = _kb_resolve_layout(ws)
    header_row = kb_header if kb_header is not None else _find_ledger_header_row(ws, profile)
    mapping = (
        dict(kb_mapping)
        if kb_mapping
        else (_map_headers(ws, header_row, profile) if header_row is not None else {})
    )
    meta: dict[str, str] = {
        "unit_name": "",
        "contact_person": "",
        "order_date": "",
        "order_number": "",
        "title": "",
    }
    assist_public: dict[str, Any] = {
        "used_llm": False,
        "cache_hit": bool(kb_fp),
        "ok": bool(kb_fp and mapping),
        "confidence": 1.0,
        "reason": "knowledge_base_hit" if kb_fp else "rules_only",
        "layout_fingerprint": kb_fp or "",
    }
    if allow_llm and not kb_fp:
        header_row, mapping, meta, kind, assist_public = _apply_llm_assist_to_layout(
            ws,
            profile,
            delivery_score=d_score,
            ledger_score=l_score,
            min_score=profile.delivery_min_score,
            header_row=header_row,
            mapping=mapping,
            meta=meta,
            prefer_kind="shipment_ledger",
            fallback_unit=fallback_unit,
        )
        if kind == "ignore":
            return []
        if kind == "delivery_note":
            return []

    if header_row is None:
        return []
    if "order_number" not in mapping:
        return []
    if "product_name" not in mapping and "model_number" not in mapping:
        return []

    _remember_sheet_layout(
        ws,
        header_row=header_row,
        mapping=mapping,
        profile=profile,
        source="knowledge_base" if kb_fp else ("llm" if assist_public.get("used_llm") else "rules"),
    )

    title_tpl = str((profile.ledger or {}).get("title_template") or "{unit}/{order_no}")
    unit_fallback = str(meta.get("unit_name") or fallback_unit).strip() or fallback_unit
    groups: dict[str, dict[str, Any]] = {}
    max_row = int(ws.max_row or 0)
    for row in range(header_row + 1, max_row + 1):
        joined = _joined_row(ws, row)
        if not joined:
            continue
        order_no = str(ws.cell(row, mapping["order_number"]).value or "").strip()
        if not order_no:
            continue
        item = _build_item_from_row(ws, row, mapping)
        if not item:
            continue
        order_date = ""
        if "order_date" in mapping:
            order_date = _excel_date_to_str(ws.cell(row, mapping["order_date"]).value, profile)
        bucket = groups.setdefault(
            order_no,
            {
                "sheet": ws.title,
                "source_kind": "shipment_ledger",
                "score": l_score,
                "unit_name": unit_fallback,
                "contact_person": meta.get("contact_person") or "",
                "order_date": order_date,
                "order_number": order_no,
                "title": title_tpl.format(unit=unit_fallback, order_no=order_no),
                "items": [],
                "assist": assist_public,
            },
        )
        if order_date and not bucket.get("order_date"):
            bucket["order_date"] = order_date
        bucket["items"].append(item)

    return [_enrich_note(g) for g in groups.values() if g.get("items")]


def parse_delivery_notes(
    file_path: str | Path,
    *,
    min_score: int | None = None,
    include_ledger: bool | str = "auto",
    unit_name_hint: str | None = None,
    profile_id: str | None = None,
    profile: ShipmentEtlProfile | None = None,
    allow_ocr: bool = True,
) -> dict[str, Any]:
    """解析工作簿：多 profile 竞分识别（通用表/流水/自定义 YAML）。

    include_ledger:
    - True: 主表 + 流水都收
    - False: 只收主表
    - "auto": 有主表时忽略同簿流水；无主表时再解析流水

    若路径是图片/PDF 且 allow_ocr=True，先走 OCR 桥接再解析。
    """
    from app.application.shipment_excel_etl_security import (
        ShipmentEtlPathError,
        resolve_etl_path,
    )

    try:
        path = resolve_etl_path(file_path, must_exist=False)
    except ShipmentEtlPathError:
        return {
            "success": False,
            "message": "非法文件路径",
            "notes": [],
            "error_code": "unsafe_path",
        }
    if allow_ocr:
        try:
            from app.application.shipment_excel_etl_ocr import is_ocr_source, parse_ocr_document

            if path.is_file() and is_ocr_source(path):
                return parse_ocr_document(
                    path,
                    include_ledger=include_ledger,
                    unit_name_hint=unit_name_hint,
                    profile_id=profile_id,
                )
        except RECOVERABLE_ERRORS:
            logger.debug("ocr auto-route skipped", exc_info=True)

    profiles = _profiles_for_parse(profile, profile_id)
    if not path.is_file():
        return {"success": False, "message": f"文件不存在: {path}", "notes": []}

    try:
        from openpyxl import load_workbook
    except ImportError:
        return {"success": False, "message": "缺少 openpyxl，无法解析 Excel", "notes": []}

    try:
        wb = load_workbook(str(path), data_only=True)
    except RECOVERABLE_ERRORS:
        return {"success": False, "message": "无法读取 Excel 文件", "notes": []}

    fallback_unit = (unit_name_hint or path.stem).strip() or path.stem
    delivery_notes: list[dict[str, Any]] = []
    ledger_notes: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    assist_summaries: list[dict[str, Any]] = []
    profile_hits: list[dict[str, Any]] = []
    sheet_roles: list[dict[str, Any]] = []
    try:
        for ws in wb.worksheets:
            prof, d_score, l_score, prefer = _pick_best_profile_for_sheet(ws, profiles)
            role = _classify_sheet_role(ws, prof, d_score=d_score, l_score=l_score)
            if role == "ledger":
                prefer = "shipment_ledger"
            elif role == "delivery":
                prefer = "delivery_note"
            score_floor = int(min_score if min_score is not None else prof.delivery_min_score)
            hit = {
                "sheet": ws.title,
                "profile_id": prof.id,
                "kind": prof.kind,
                "label": prof.label,
                "delivery_score": d_score,
                "ledger_score": l_score,
                "prefer": prefer,
                "role": role,
            }
            profile_hits.append(hit)
            sheet_roles.append({"sheet": ws.title, "role": role, "prefer": prefer})
            if role == "ignore":
                skipped.append(
                    {
                        "sheet": ws.title,
                        "score": max(d_score, l_score),
                        "reason": "sheet_ignored_mixed_workbook",
                        "profile_id": prof.id,
                        "role": role,
                    }
                )
                continue
            # 多表混排：角色已判定为单据时，允许较低分也尝试解析（陌生表头）
            delivery_gate = 24 if role in {"delivery", "unknown"} else 40
            if prefer == "delivery_note" and d_score >= delivery_gate:
                note = _parse_delivery_sheet(
                    ws, fallback_unit=fallback_unit, profile=prof, allow_llm=True
                )
                if note:
                    note["profile_id"] = prof.id
                    note["profile_kind"] = prof.kind
                    note["profile_label"] = prof.label
                    note["profile_target"] = prof.target
                    note["sheet_role"] = role
                    delivery_notes.append(note)
                    if isinstance(note.get("assist"), dict):
                        assist_summaries.append(
                            {
                                "sheet": ws.title,
                                "profile_id": prof.id,
                                **dict(note.get("assist") or {}),
                            }
                        )
                    continue
                if d_score >= score_floor:
                    skipped.append(
                        {
                            "sheet": ws.title,
                            "score": d_score,
                            "reason": "delivery_parse_failed",
                            "profile_id": prof.id,
                            "role": role,
                        }
                    )
                    continue

            if prefer == "shipment_ledger" and prof.has_ledger and l_score >= 40:
                parsed_ledger = _parse_ledger_sheet(
                    ws, fallback_unit=fallback_unit, profile=prof, allow_llm=True
                )
                if parsed_ledger:
                    for n in parsed_ledger:
                        n["profile_id"] = prof.id
                        n["profile_kind"] = prof.kind
                        n["profile_label"] = prof.label
                        n["profile_target"] = prof.target
                    ledger_notes.extend(parsed_ledger)
                    assist = (parsed_ledger[0] or {}).get("assist")
                    if isinstance(assist, dict):
                        assist_summaries.append(
                            {"sheet": ws.title, "profile_id": prof.id, **assist}
                        )
                    continue
                # 流水竞分失败 → 回退通用表解析（避免误伤单据表）
                note = _parse_delivery_sheet(
                    ws, fallback_unit=fallback_unit, profile=prof, allow_llm=True
                )
                if note:
                    note["profile_id"] = prof.id
                    note["profile_kind"] = prof.kind
                    note["profile_label"] = prof.label
                    note["profile_target"] = prof.target
                    delivery_notes.append(note)
                    if isinstance(note.get("assist"), dict):
                        assist_summaries.append(
                            {
                                "sheet": ws.title,
                                "profile_id": prof.id,
                                **dict(note.get("assist") or {}),
                            }
                        )
                    continue
                if l_score >= 50:
                    skipped.append(
                        {
                            "sheet": ws.title,
                            "score": l_score,
                            "reason": "ledger_empty",
                            "profile_id": prof.id,
                        }
                    )
                    continue

            if prof.has_ledger and l_score >= 40 and prefer != "shipment_ledger":
                parsed_ledger = _parse_ledger_sheet(
                    ws, fallback_unit=fallback_unit, profile=prof, allow_llm=True
                )
                if parsed_ledger:
                    for n in parsed_ledger:
                        n["profile_id"] = prof.id
                        n["profile_kind"] = prof.kind
                        n["profile_label"] = prof.label
                        n["profile_target"] = prof.target
                    ledger_notes.extend(parsed_ledger)
                    assist = (parsed_ledger[0] or {}).get("assist")
                    if isinstance(assist, dict):
                        assist_summaries.append(
                            {"sheet": ws.title, "profile_id": prof.id, **assist}
                        )
                elif l_score >= 50:
                    skipped.append(
                        {
                            "sheet": ws.title,
                            "score": l_score,
                            "reason": "ledger_empty",
                            "profile_id": prof.id,
                        }
                    )
                else:
                    skipped.append(
                        {
                            "sheet": ws.title,
                            "score": max(d_score, l_score),
                            "reason": "not_matched",
                            "profile_id": prof.id,
                        }
                    )
            else:
                note = _parse_delivery_sheet(
                    ws, fallback_unit=fallback_unit, profile=prof, allow_llm=True
                )
                if note:
                    note["profile_id"] = prof.id
                    note["profile_kind"] = prof.kind
                    note["profile_label"] = prof.label
                    note["profile_target"] = prof.target
                    delivery_notes.append(note)
                    if isinstance(note.get("assist"), dict):
                        assist_summaries.append(
                            {
                                "sheet": ws.title,
                                "profile_id": prof.id,
                                **dict(note.get("assist") or {}),
                            }
                        )
                else:
                    skipped.append(
                        {
                            "sheet": ws.title,
                            "score": max(d_score, l_score),
                            "reason": "not_matched",
                            "profile_id": prof.id,
                        }
                    )
    finally:
        wb.close()

    mode = include_ledger
    if isinstance(mode, str):
        mode_l = mode.strip().lower()
        if mode_l in {"1", "true", "yes", "on"}:
            mode = True
        elif mode_l in {"0", "false", "no", "off"}:
            mode = False
        else:
            mode = "auto"

    if mode is True:
        notes = delivery_notes + ledger_notes
    elif mode is False:
        notes = delivery_notes
        for n in ledger_notes:
            skipped.append(
                {"sheet": n.get("sheet"), "score": n.get("score"), "reason": "ledger_disabled"}
            )
    else:
        if delivery_notes:
            notes = delivery_notes
            for n in ledger_notes:
                skipped.append(
                    {
                        "sheet": n.get("sheet"),
                        "score": n.get("score"),
                        "reason": "ledger_skipped_auto_has_delivery",
                        "ledger_groups": 1,
                    }
                )
        else:
            notes = ledger_notes

    delivery_count = sum(1 for n in notes if n.get("source_kind") == "delivery_note")
    ledger_count = sum(1 for n in notes if n.get("source_kind") == "shipment_ledger")
    used_llm = any(bool(a.get("used_llm") and a.get("ok")) for a in assist_summaries)
    used_profile_ids = sorted(
        {str(n.get("profile_id") or "") for n in notes if n.get("profile_id")}
    )
    if len(used_profile_ids) == 1:
        result_profile_id = used_profile_ids[0]
    elif len(profiles) == 1:
        result_profile_id = profiles[0].id
    else:
        result_profile_id = "auto"
    return {
        "success": True,
        "file_path": str(path),
        "file_name": path.name,
        "profile_id": result_profile_id,
        "profile_ids": used_profile_ids,
        "profiles_available": [p.id for p in profiles],
        "profile_hits": profile_hits,
        "sheet_roles": sheet_roles,
        "mixed_workbook": len({r.get("role") for r in sheet_roles}) > 1,
        "note_count": len(notes),
        "delivery_note_count": delivery_count,
        "ledger_note_count": ledger_count,
        "ledger_available_count": len(ledger_notes),
        "include_ledger_mode": mode if mode in (True, False) else "auto",
        "notes": notes,
        "skipped_sheets": skipped,
        "assist": {
            "used_llm": used_llm,
            "sheets": assist_summaries,
        },
        "message": (
            (
                f"识别到 {len(notes)} 张单据（主表 {delivery_count} / 流水分组 {ledger_count}）"
                + (f"；profile={','.join(used_profile_ids)}" if used_profile_ids else "")
            )
            if notes
            else "未识别到可匹配的单据模板（可自定义 YAML profile）"
        ),
    }


def preview_shipment_excel_etl(
    file_path: str | Path,
    *,
    include_ledger: bool | str = "auto",
    unit_name_hint: str | None = None,
    workspace_root: str | Path | None = None,
    profile_id: str | None = None,
    profile: ShipmentEtlProfile | None = None,
) -> dict[str, Any]:
    from app.application.shipment_excel_etl_security import (
        ShipmentEtlPathError,
        resolve_etl_path,
        tenant_key_for_etl,
    )

    try:
        path = resolve_etl_path(file_path, workspace_root=workspace_root, must_exist=True)
    except ShipmentEtlPathError:
        return {
            "success": False,
            "message": "非法文件路径",
            "error_code": "unsafe_path",
            "notes": [],
        }

    parsed = parse_delivery_notes(
        path,
        include_ledger=include_ledger,
        unit_name_hint=unit_name_hint,
        profile_id=profile_id,
        profile=profile,
    )
    if not parsed.get("success"):
        return parsed
    notes = parsed.get("notes") or []
    tenant_key = tenant_key_for_etl()
    for note in notes:
        fp = str(note.get("fingerprint") or "")
        note["already_imported"] = bool(fp and _is_fingerprint_imported(tenant_key, fp))
    ledger_available = int(parsed.get("ledger_available_count") or 0)
    return {
        **parsed,
        "preview": True,
        "product_records": _notes_to_product_records(notes),
        "confirm_required": True,
        "duplicate_note_count": sum(1 for n in notes if n.get("already_imported")),
        "ledger_risk": ledger_available > 0 and int(parsed.get("ledger_note_count") or 0) == 0,
        "ledger_available_count": ledger_available,
        "message": parsed.get("message") + ("。确认后将写入客户、产品与发货单。" if notes else ""),
    }


def _notes_to_product_records(notes: list[dict[str, Any]]) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    seen: set[tuple[str, str, str]] = set()
    for note in notes:
        unit = str(note.get("unit_name") or "").strip()
        for item in note.get("items") or []:
            model = str(item.get("model_number") or "").strip().upper()
            name = str(item.get("product_name") or "").strip()
            key = (unit, model, name)
            if key in seen:
                continue
            seen.add(key)
            records.append(
                {
                    "unit_name": unit,
                    "product_name": name,
                    "model_number": model,
                    "unit_price": float(item.get("unit_price") or 0),
                }
            )
    return records


def execute_shipment_excel_etl(
    file_path: str | Path,
    *,
    import_products: bool = True,
    import_shipments: bool = True,
    notes: list[dict[str, Any]] | None = None,
    idempotent: bool = True,
    include_ledger: bool | str = False,
    confirm_ledger: bool = False,
    dry_run: bool = False,
    compensate_on_failure: bool = True,
    unit_name_hint: str | None = None,
    workspace_root: str | Path | None = None,
    profile_id: str | None = None,
    profile: ShipmentEtlProfile | None = None,
    direct: bool = False,
    force_shipment_target: bool = False,
) -> dict[str, Any]:
    """执行闭环：客户+产品+发货单（可幂等 / dry-run / 失败补偿）。

    生产默认 include_ledger=False；若要导入流水须 confirm_ledger=True。
    任一建单失败且 compensate_on_failure=True 时，取消本批已新建发货单并删除指纹。

    direct=True：无预览直写（需 FHD_EXCEL_ETL_ALLOW_DIRECT=1）。
    force_shipment_target=True：直写时把 preview_only notes 提升为 shipment。
    """
    from app.application.shipment_excel_etl_security import (
        ShipmentEtlPathError,
        direct_execute_allowed,
        resolve_etl_path,
        tenant_key_for_etl,
    )

    if direct and not dry_run and not direct_execute_allowed():
        return {
            "success": False,
            "message": (
                "无预览直写未开启。请设置 FHD_EXCEL_ETL_ALLOW_DIRECT=1 "
                "（或 FHD_SHIPMENT_ETL_ALLOW_DIRECT=1）并确认权限后再执行。"
            ),
            "error_code": "direct_execute_denied",
        }

    prof = _resolve_profile(profile, profile_id)
    path: Path | None = None
    file_name = "shipment.xlsx"
    if file_path:
        try:
            path = resolve_etl_path(
                file_path, workspace_root=workspace_root, must_exist=notes is None
            )
            file_name = path.name
        except ShipmentEtlPathError:
            return {
                "success": False,
                "message": "非法文件路径",
                "error_code": "unsafe_path",
            }

    if notes is None:
        if path is None:
            return {"success": False, "message": "缺少 file_path", "error_code": "missing_path"}
        parsed = parse_delivery_notes(
            path,
            include_ledger=include_ledger,
            unit_name_hint=unit_name_hint,
            profile=prof,
        )
        if not parsed.get("success"):
            return parsed
        notes = [_enrich_note(n) for n in (parsed.get("notes") or [])]
        file_name = str(parsed.get("file_name") or file_name)
        ledger_available = int(parsed.get("ledger_available_count") or 0)
    else:
        notes = [_enrich_note(n) for n in notes]
        ledger_available = sum(1 for n in notes if n.get("source_kind") == "shipment_ledger")

    if direct and force_shipment_target:
        for n in notes:
            if str(n.get("profile_target") or "").strip() in {"", "preview_only"}:
                n["profile_target"] = "shipment"
                n["direct_target_promoted"] = True

    ledger_notes = [n for n in notes if n.get("source_kind") == "shipment_ledger"]
    # 直写时可凭 confirm_ledger 放行流水；未确认仍拦截
    if ledger_notes and not confirm_ledger:
        return {
            "success": False,
            "message": (
                f"检测到 {len(ledger_notes)} 张出货流水分组，生产默认禁止直接入库。"
                "请传 confirm_ledger=1 并确认客户归属后再执行。"
            ),
            "error_code": "ledger_confirm_required",
            "ledger_note_count": len(ledger_notes),
            "note_count": len(notes),
            "direct": bool(direct),
        }

    if not notes:
        return {
            "success": False,
            "message": "没有可导入的单据",
            "error_code": "no_delivery_notes",
        }

    # 非 shipment target 的自定义模板：只允许 preview / 产品导入，禁止误建发货单
    non_shipment = [
        n
        for n in notes
        if str(n.get("profile_target") or "shipment").strip() not in {"", "shipment"}
    ]
    if non_shipment and import_shipments:
        targets = sorted({str(n.get("profile_target") or "preview_only") for n in non_shipment})
        return {
            "success": False,
            "message": (
                "识别到非发货单模板（target="
                + ",".join(targets)
                + "）。请改用 preview，或为该 YAML 设置 target: shipment。"
            ),
            "error_code": "unsupported_profile_target",
            "profile_ids": sorted(
                {str(n.get("profile_id") or "") for n in non_shipment if n.get("profile_id")}
            ),
            "note_count": len(notes),
        }

    tenant_key = tenant_key_for_etl()
    to_import: list[dict[str, Any]] = []
    skipped_duplicates: list[dict[str, Any]] = []
    for note in notes:
        fp = str(note.get("fingerprint") or note_fingerprint(note))
        note["fingerprint"] = fp
        if idempotent and _is_fingerprint_imported(tenant_key, fp):
            skipped_duplicates.append(
                {
                    "fingerprint": fp,
                    "unit_name": note.get("unit_name"),
                    "order_number": note.get("order_number"),
                }
            )
            continue
        to_import.append(note)

    if dry_run:
        return {
            "success": True,
            "dry_run": True,
            "direct": bool(direct),
            "message": (
                f"预演：将新建 {len(to_import)} 张，跳过重复 {len(skipped_duplicates)} 张；不会写库"
            ),
            "file_name": file_name,
            "note_count": len(notes),
            "would_create": len(to_import),
            "would_skip": len(skipped_duplicates),
            "notes": to_import,
            "skipped_duplicates": skipped_duplicates,
            "ledger_available_count": ledger_available,
            "closed_loop": False,
            "kind": "shipment_delivery_etl",
        }

    product_result: dict[str, Any] = {"success": True, "skipped": True}
    if import_products and to_import:
        from app.services.tools_workflow_registered import _execute_excel_import_records

        product_result = _execute_excel_import_records(_notes_to_product_records(to_import))
        if not bool(product_result.get("success", True)):
            return {
                "success": False,
                "message": f"客户/产品导入失败，已中止发货单写入：{product_result.get('message') or product_result}",
                "error_code": "product_import_failed",
                "product_result": product_result,
                "note_count": len(notes),
                "closed_loop": False,
            }

    shipment_created = 0
    shipment_failed = 0
    shipment_skipped = len(skipped_duplicates)
    shipment_ids: list[Any] = []
    created_pairs: list[tuple[Any, str]] = []  # (shipment_id, fingerprint)
    errors: list[str] = []
    compensated: list[Any] = []
    compensate_errors: list[str] = []

    if import_shipments and to_import:
        try:
            from app.bootstrap import get_shipment_app_service

            svc = get_shipment_app_service()
        except RECOVERABLE_ERRORS:
            return {
                "success": False,
                "message": "发货单服务不可用",
                "product_result": product_result,
            }

        for note in to_import:
            unit = str(note.get("unit_name") or "").strip()
            items = list(note.get("items") or [])
            if not unit or not items:
                shipment_failed += 1
                errors.append(f"缺少客户或明细: {note.get('order_number') or note.get('sheet')}")
                break
            result = svc.create_shipment(
                unit_name=unit,
                items_data=items,
                contact_person=str(note.get("contact_person") or ""),
                external_order_number=str(note.get("order_number") or ""),
                order_date=str(note.get("order_date") or ""),
                source_fingerprint=str(note.get("fingerprint") or ""),
                source_kind=str(note.get("source_kind") or ""),
            )
            if result.get("success"):
                shipment_created += 1
                shipment = result.get("shipment") or {}
                sid = shipment.get("id") if isinstance(shipment, dict) else None
                fp = str(note.get("fingerprint") or "")
                if sid is not None:
                    shipment_ids.append(sid)
                    created_pairs.append((sid, fp))
                if idempotent and fp:
                    try:
                        _record_fingerprint_now(
                            tenant_key,
                            fp,
                            shipment_id=sid,
                            unit_name=unit,
                            order_number=str(note.get("order_number") or ""),
                            file_name=file_name,
                        )
                    except RECOVERABLE_ERRORS:
                        logger.warning(
                            "failed to persist etl fingerprint immediately", exc_info=True
                        )
            else:
                shipment_failed += 1
                errors.append(str(result.get("message") or "create_shipment failed"))
                break

        # 未处理完的剩余 notes 计为未执行失败（避免静默漏导）
        processed = shipment_created + shipment_failed
        if processed < len(to_import) and shipment_failed:
            remaining = len(to_import) - processed
            errors.append(f"因失败中止，另有 {remaining} 张未执行")
            shipment_failed += remaining

        if shipment_failed and created_pairs and compensate_on_failure:
            from app.application.shipment_excel_etl_fingerprint_store import delete_fingerprint

            for sid, fp in created_pairs:
                try:
                    cancel = svc.cancel_shipment(int(sid))
                    if cancel.get("success"):
                        compensated.append(sid)
                    else:
                        # 取消失败则尝试删除
                        deleted = svc.delete_shipment(int(sid))
                        if deleted.get("success"):
                            compensated.append(sid)
                        else:
                            compensate_errors.append(
                                f"补偿失败 shipment_id={sid}: {cancel.get('message') or deleted.get('message')}"
                            )
                except RECOVERABLE_ERRORS as exc:
                    compensate_errors.append(f"补偿异常 shipment_id={sid}: {exc}")
                if fp:
                    try:
                        delete_fingerprint(tenant_key, fp)
                    except RECOVERABLE_ERRORS:
                        logger.warning(
                            "failed to delete etl fingerprint on compensate", exc_info=True
                        )
            shipment_created = max(0, shipment_created - len(compensated))
            shipment_ids = [sid for sid in shipment_ids if sid not in set(compensated)]

    ok = shipment_failed == 0 and bool(product_result.get("success", True))
    if not to_import and skipped_duplicates:
        ok = True
    compensated_ok = bool(
        shipment_failed and compensate_on_failure and created_pairs and not compensate_errors
    )
    if shipment_failed and compensate_on_failure:
        # 补偿成功后视为「未留下脏发货单」，success=False 但仍可安全重试
        ok = False
    return {
        "success": ok,
        "partial_success": bool(shipment_failed and shipment_ids and not compensate_on_failure),
        "compensated": compensated,
        "compensate_on_failure": compensate_on_failure,
        "compensate_errors": compensate_errors[:8],
        "safe_to_retry": (not shipment_ids) or compensated_ok or ok,
        "message": (
            f"送货单闭环完成：新建 {shipment_created}，跳过重复 {shipment_skipped}"
            + (f"，失败 {shipment_failed}" if shipment_failed else "")
            + (f"，已补偿撤销 {len(compensated)}" if compensated else "")
            + ("；客户/产品已同步" if import_products and to_import else "")
            + (
                "（部分成功，未启用补偿）"
                if (shipment_ids and shipment_failed and not compensate_on_failure)
                else ""
            )
        ),
        "file_name": file_name,
        "note_count": len(notes),
        "shipment_created": shipment_created,
        "shipment_failed": shipment_failed,
        "shipment_skipped": shipment_skipped,
        "shipment_ids": shipment_ids,
        "skipped_duplicates": skipped_duplicates,
        "product_result": product_result,
        "errors": errors[:20],
        "closed_loop": True,
        "idempotent": idempotent,
        "dry_run": False,
        "direct": bool(direct),
        "kind": "shipment_delivery_etl",
        "audit": {
            "tenant_key": tenant_key,
            "file_name": file_name,
            "created": shipment_created,
            "failed": shipment_failed,
            "skipped": shipment_skipped,
            "compensated": len(compensated),
            "direct": bool(direct),
            "force_shipment_target": bool(force_shipment_target),
        },
    }


def write_delivery_note_workbook(
    notes: list[dict[str, Any]],
    output_path: str | Path,
    *,
    seller_title: str | None = None,
    profile_id: str | None = None,
    profile: ShipmentEtlProfile | None = None,
) -> dict[str, Any]:
    """按 profile.write 版式写出送货单模板（可用于回环验证 / 测试数据）。"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return {"success": False, "message": "缺少 openpyxl，无法解析 Excel"}

    prof = _resolve_profile(profile, profile_id)
    write_cfg = prof.write or {}
    title = str(seller_title if seller_title is not None else write_cfg.get("seller_title") or "")
    headers = list(write_cfg.get("header_row") or [])
    item_cols = dict(write_cfg.get("item_columns") or {})
    date_fmt = str(write_cfg.get("date_format") or "%Y-%m-%d")
    meta_tpl = str(
        write_cfg.get("meta_line_template") or "{unit} {contact} {order_date} {order_no}"
    )
    footer = str(write_cfg.get("footer_label") or "")
    default_sheet = str(write_cfg.get("default_sheet_name") or "Sheet1")
    sheet_prefix = str(write_cfg.get("sheet_name_prefix") or "S")
    demo_meta = str(write_cfg.get("demo_meta_line") or meta_tpl)
    demo_item = dict(write_cfg.get("demo_item") or {})

    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    created = 0
    for idx, note in enumerate(notes or [], start=1):
        unit = str(note.get("unit_name") or f"客户{idx}").strip()
        sheet_name = (
            str(note.get("sheet_name") or note.get("sheet") or unit)[:28] or f"{sheet_prefix}{idx}"
        )
        sheet_name = re.sub(r"[\\/*?:\[\]]", "_", sheet_name)[:31]
        if idx == 1:
            ws = default
            ws.title = sheet_name
        else:
            base = sheet_name
            n = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base[:28]}_{n}"
                n += 1
            ws = wb.create_sheet(sheet_name)

        contact = str(note.get("contact_person") or "").strip()
        order_date = str(note.get("order_date") or datetime.now().strftime(date_fmt)).strip()
        order_no = str(note.get("order_number") or f"TEST-{idx:04d}").strip()
        ws["A1"] = title
        ws["A2"] = meta_tpl.format(
            unit=unit, contact=contact, order_date=order_date, order_no=order_no
        )
        for col, h in enumerate(headers, start=1):
            ws.cell(3, col, h)
        last_row = 3
        for r, item in enumerate(note.get("items") or [], start=4):
            if "model_number" in item_cols:
                ws.cell(r, int(item_cols["model_number"]), item.get("model_number") or "")
            if "product_name" in item_cols:
                ws.cell(r, int(item_cols["product_name"]), item.get("product_name") or "")
            if "quantity_tins" in item_cols:
                ws.cell(
                    r,
                    int(item_cols["quantity_tins"]),
                    item.get("quantity_tins") or item.get("quantity") or 0,
                )
            if "tin_spec" in item_cols:
                ws.cell(
                    r,
                    int(item_cols["tin_spec"]),
                    item.get("tin_spec") or item.get("spec_per_tin") or 0,
                )
            if "quantity_kg" in item_cols:
                ws.cell(r, int(item_cols["quantity_kg"]), item.get("quantity_kg") or 0)
            if "unit_price" in item_cols:
                ws.cell(r, int(item_cols["unit_price"]), item.get("unit_price") or 0)
            if "amount" in item_cols:
                ws.cell(r, int(item_cols["amount"]), item.get("amount") or 0)
            last_row = r
        if footer:
            ws.cell(last_row + 2, 1, footer)
        created += 1

    if created == 0:
        ws = default
        ws.title = default_sheet[:31]
        ws["A1"] = title
        ws["A2"] = demo_meta
        for col, h in enumerate(headers, start=1):
            ws.cell(3, col, h)
        r = 4
        if "model_number" in item_cols:
            ws.cell(r, int(item_cols["model_number"]), demo_item.get("model_number") or "")
        if "product_name" in item_cols:
            ws.cell(r, int(item_cols["product_name"]), demo_item.get("product_name") or "")
        if "quantity_tins" in item_cols:
            ws.cell(r, int(item_cols["quantity_tins"]), demo_item.get("quantity_tins") or 0)
        if "tin_spec" in item_cols:
            ws.cell(r, int(item_cols["tin_spec"]), demo_item.get("tin_spec") or 0)
        if "quantity_kg" in item_cols:
            ws.cell(r, int(item_cols["quantity_kg"]), demo_item.get("quantity_kg") or 0)
        if "unit_price" in item_cols:
            ws.cell(r, int(item_cols["unit_price"]), demo_item.get("unit_price") or 0)
        if "amount" in item_cols:
            ws.cell(r, int(item_cols["amount"]), demo_item.get("amount") or 0)
        created = 1

    wb.save(path)
    wb.close()
    return {
        "success": True,
        "file_path": str(path),
        "sheet_count": created,
        "profile_id": prof.id,
        "message": f"已生成送货单模板 {path.name}（{created} 张表）",
    }


def write_ledger_workbook(
    rows: list[dict[str, Any]],
    output_path: str | Path,
    *,
    sheet_name: str | None = None,
    unit_name: str | None = None,
    profile_id: str | None = None,
    profile: ShipmentEtlProfile | None = None,
) -> dict[str, Any]:
    """写出出货流水模板（表头/列位来自 profile.write）。"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return {"success": False, "message": "缺少 openpyxl，无法解析 Excel"}

    prof = _resolve_profile(profile, profile_id)
    write_cfg = prof.write or {}
    default_sheet = str(sheet_name or write_cfg.get("ledger_sheet_name") or "ledger")
    resolved_unit = str(unit_name or write_cfg.get("ledger_default_unit") or "unit")
    headers = list(write_cfg.get("ledger_header_row") or [])
    item_cols = dict(write_cfg.get("ledger_item_columns") or {})
    sample_rows = rows or list(write_cfg.get("ledger_sample_rows") or [])
    extra_sheet = str(write_cfg.get("ledger_extra_sheet") or "").strip()

    path = Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = default_sheet[:31] or "ledger"
    for col, h in enumerate(headers, start=1):
        ws.cell(1, col, h)
    for r, row in enumerate(sample_rows, start=2):
        for field_name, col_idx in item_cols.items():
            raw = row.get(field_name)
            if raw is None or raw == "":
                if field_name in {
                    "quantity_tins",
                    "tin_spec",
                    "quantity_kg",
                    "unit_price",
                    "amount",
                }:
                    raw = 0
                else:
                    raw = ""
            ws.cell(r, int(col_idx), raw)
    if extra_sheet and extra_sheet not in wb.sheetnames:
        wb.create_sheet(extra_sheet[:31])
    wb.save(path)
    wb.close()
    return {
        "success": True,
        "file_path": str(path),
        "unit_name": resolved_unit,
        "row_count": len(sample_rows),
        "profile_id": prof.id,
        "message": f"已生成出货流水模板 {path.name}",
    }


def regenerate_delivery_notes_from_file(
    file_path: str | Path,
    output_path: str | Path,
    *,
    include_ledger: bool | str = "auto",
    profile_id: str | None = None,
    profile: ShipmentEtlProfile | None = None,
) -> dict[str, Any]:
    """解析 → 按 profile 送货单版式再出单（模板反推闭环）。"""
    prof = _resolve_profile(profile, profile_id)
    parsed = parse_delivery_notes(file_path, include_ledger=include_ledger, profile=prof)
    if not parsed.get("success"):
        return parsed
    notes = parsed.get("notes") or []
    if not notes:
        return {"success": False, "message": "无可反推的单据", "error_code": "no_delivery_notes"}
    written = write_delivery_note_workbook(notes, output_path, profile=prof)
    if not written.get("success"):
        return written
    reparsed = parse_delivery_notes(output_path, include_ledger=False, profile=prof)
    return {
        "success": True,
        "source": parsed,
        "generated": written,
        "reparsed": reparsed,
        "fingerprint_match": (
            {n.get("fingerprint") for n in notes}
            == {n.get("fingerprint") for n in (reparsed.get("notes") or [])}
            if reparsed.get("success")
            else False
        ),
        "profile_id": prof.id,
        "message": "模板反推完成",
    }


def batch_preview_shipment_excel_etl(
    directory: str | Path,
    *,
    include_ledger: bool | str = "auto",
    pattern: str = "*.xlsx",
    workspace_root: str | Path | None = None,
    profile_id: str | None = None,
    profile: ShipmentEtlProfile | None = None,
) -> dict[str, Any]:
    from app.application.shipment_excel_etl_security import (
        ShipmentEtlPathError,
        resolve_etl_path,
    )

    prof = _resolve_profile(profile, profile_id)
    try:
        root = resolve_etl_path(directory, workspace_root=workspace_root, must_exist=True)
    except ShipmentEtlPathError:
        return {
            "success": False,
            "message": "非法目录",
            "error_code": "unsafe_path",
            "files": [],
        }
    if not root.is_dir():
        return {"success": False, "message": f"目录不存在: {root}", "files": []}
    files = sorted(root.glob(pattern))
    results = []
    total_notes = 0
    for path in files:
        if path.name.startswith("~$"):
            continue
        preview = preview_shipment_excel_etl(
            path,
            include_ledger=include_ledger,
            unit_name_hint=path.stem,
            workspace_root=workspace_root or root,
            profile=prof,
        )
        note_count = int(preview.get("note_count") or 0)
        total_notes += note_count
        results.append(
            {
                "file_path": str(path),
                "file_name": path.name,
                "success": bool(preview.get("success")),
                "note_count": note_count,
                "duplicate_note_count": preview.get("duplicate_note_count", 0),
                "message": preview.get("message"),
                "notes": preview.get("notes") or [],
            }
        )
    return {
        "success": True,
        "directory": str(root),
        "file_count": len(results),
        "note_count": total_notes,
        "files": results,
        "profile_id": prof.id,
        "message": f"批量预览完成：{len(results)} 个文件，共 {total_notes} 张单据",
    }


def batch_execute_shipment_excel_etl(
    directory: str | Path,
    *,
    include_ledger: bool | str = False,
    pattern: str = "*.xlsx",
    idempotent: bool = True,
    import_products: bool = True,
    import_shipments: bool = True,
    confirm_ledger: bool = False,
    dry_run: bool = False,
    workspace_root: str | Path | None = None,
    profile_id: str | None = None,
    profile: ShipmentEtlProfile | None = None,
) -> dict[str, Any]:
    from app.application.shipment_excel_etl_security import (
        ShipmentEtlPathError,
        batch_execute_allowed,
        resolve_etl_path,
    )

    prof = _resolve_profile(profile, profile_id)
    if not dry_run and not batch_execute_allowed():
        return {
            "success": False,
            "message": "批量入库默认关闭。需设置环境变量 FHD_SHIPMENT_ETL_ALLOW_BATCH=1",
            "error_code": "batch_disabled",
            "files": [],
        }
    try:
        root = resolve_etl_path(directory, workspace_root=workspace_root, must_exist=True)
    except ShipmentEtlPathError:
        return {
            "success": False,
            "message": "非法目录",
            "error_code": "unsafe_path",
            "files": [],
        }
    if not root.is_dir():
        return {"success": False, "message": f"目录不存在: {root}", "files": []}
    files = sorted(root.glob(pattern))
    if len(files) > 50:
        return {
            "success": False,
            "message": f"批量文件过多（{len(files)}），上限 50，请缩小范围",
            "error_code": "batch_too_large",
            "files": [],
        }
    results = []
    created = skipped = failed = 0
    for path in files:
        if path.name.startswith("~$"):
            continue
        result = execute_shipment_excel_etl(
            path,
            include_ledger=include_ledger,
            unit_name_hint=path.stem,
            idempotent=idempotent,
            import_products=import_products,
            import_shipments=import_shipments,
            confirm_ledger=confirm_ledger,
            dry_run=dry_run,
            workspace_root=workspace_root or root,
            profile=prof,
        )
        created += int(result.get("shipment_created") or result.get("would_create") or 0)
        skipped += int(result.get("shipment_skipped") or result.get("would_skip") or 0)
        failed += int(result.get("shipment_failed") or 0)
        results.append(
            {
                "file_path": str(path),
                "file_name": path.name,
                "success": bool(result.get("success")),
                "shipment_created": result.get("shipment_created", result.get("would_create", 0)),
                "shipment_skipped": result.get("shipment_skipped", result.get("would_skip", 0)),
                "shipment_failed": result.get("shipment_failed", 0),
                "message": result.get("message"),
                "error_code": result.get("error_code"),
            }
        )
    return {
        "success": failed == 0,
        "directory": str(root),
        "file_count": len(results),
        "shipment_created": created,
        "shipment_skipped": skipped,
        "shipment_failed": failed,
        "files": results,
        "closed_loop": not dry_run,
        "dry_run": dry_run,
        "profile_id": prof.id,
        "message": f"{'批量预演' if dry_run else '批量入库'}完成：新建/将建 {created}，跳过 {skipped}，失败 {failed}",
    }


class ShipmentExcelEtlApplicationService:
    def __init__(self, profile_id: str | None = None) -> None:
        self._profile_id = profile_id

    def _profile_kwargs(self, kwargs: dict[str, Any]) -> dict[str, Any]:
        if "profile" in kwargs or "profile_id" in kwargs:
            return kwargs
        if self._profile_id:
            return {**kwargs, "profile_id": self._profile_id}
        return kwargs

    def preview(self, file_path: str | Path, **kwargs: Any) -> dict[str, Any]:
        return preview_shipment_excel_etl(file_path, **self._profile_kwargs(kwargs))

    def execute(self, file_path: str | Path, **kwargs: Any) -> dict[str, Any]:
        return execute_shipment_excel_etl(file_path, **self._profile_kwargs(kwargs))

    def batch_preview(self, directory: str | Path, **kwargs: Any) -> dict[str, Any]:
        return batch_preview_shipment_excel_etl(directory, **self._profile_kwargs(kwargs))

    def batch_execute(self, directory: str | Path, **kwargs: Any) -> dict[str, Any]:
        return batch_execute_shipment_excel_etl(directory, **self._profile_kwargs(kwargs))

    def write_delivery_template(
        self, notes: list[dict[str, Any]], output_path: str | Path, **kwargs: Any
    ) -> dict[str, Any]:
        return write_delivery_note_workbook(notes, output_path, **self._profile_kwargs(kwargs))

    def write_ledger_template(
        self, rows: list[dict[str, Any]], output_path: str | Path, **kwargs: Any
    ) -> dict[str, Any]:
        return write_ledger_workbook(rows, output_path, **self._profile_kwargs(kwargs))

    def regenerate(
        self, file_path: str | Path, output_path: str | Path, **kwargs: Any
    ) -> dict[str, Any]:
        return regenerate_delivery_notes_from_file(
            file_path, output_path, **self._profile_kwargs(kwargs)
        )

    def ocr_preview(self, file_path: str | Path, **kwargs: Any) -> dict[str, Any]:
        from app.application.shipment_excel_etl_ocr import parse_ocr_document

        return parse_ocr_document(file_path, **self._profile_kwargs(kwargs))


_svc: ShipmentExcelEtlApplicationService | None = None


def get_shipment_excel_etl_app_service() -> ShipmentExcelEtlApplicationService:
    global _svc
    if _svc is None:
        _svc = ShipmentExcelEtlApplicationService()
    return _svc
