# ruff: noqa
# mypy: ignore-errors
"""Implementation extracted from the public facade module."""
from __future__ import annotations
import importlib

def _facade():
    return importlib.import_module('app.application.shipment_excel_etl_app_service')

def _unit_name_looks_truncated(unit: str) -> bool:
    text = str(unit or '').strip()
    if not text:
        return True
    if text.lower() in _facade()._CORP_SUFFIX_ONLY:
        return True
    if ' ' not in text and _facade().re.fullmatch('[A-Za-z.]{1,6}', text):
        return True
    return False

def _extract_adjacent_buyer_meta(ws, header_row: int) -> dict[str, str]:
    """英文 DO/PI 常见：标签在 A 列、公司名在同行右侧单元格。"""
    from app.application.shipment_excel_etl_llm import unit_name_is_weak
    out = {'unit_name': '', 'contact_person': '', 'order_number': ''}
    max_col = min(int(getattr(ws, 'max_column', 1) or 1), 16)
    scan_to = max(2, min(int(header_row or 2), 30))
    for row in range(1, scan_to):
        for col in range(1, max_col + 1):
            raw = str(ws.cell(row, col).value or '').strip()
            if not raw:
                continue
            inline = _facade()._BUYER_INLINE.search(raw)
            if inline and (not out['unit_name']):
                candidate = inline.group(1).strip().split('\n')[0].strip(' ：:\u3000')
                candidate = _facade().re.sub('\\s*\\([^)]*\\)\\s*$', '', candidate).strip()
                if len(candidate) >= 2 and (not unit_name_is_weak(candidate)):
                    out['unit_name'] = candidate
                    continue
            if _facade()._BUYER_CELL_LABEL.match(raw) and (not out['unit_name']):
                for c2 in range(col + 1, min(col + 5, max_col + 1)):
                    val = str(ws.cell(row, c2).value or '').strip()
                    if not val or _facade()._BUYER_CELL_LABEL.match(val):
                        continue
                    candidate = val.split('\n')[0].strip()
                    candidate = _facade().re.sub('\\s*\\([^)]*\\)\\s*$', '', candidate).strip()
                    if candidate and (not unit_name_is_weak(candidate)):
                        out['unit_name'] = candidate
                    break
                continue
            if _facade()._ATTN_CELL_LABEL.match(raw) and (not out['contact_person']):
                for c2 in range(col + 1, min(col + 4, max_col + 1)):
                    val = str(ws.cell(row, c2).value or '').strip()
                    if val:
                        out['contact_person'] = val.split('\n')[0].strip()
                        break
                continue
            order_m = _facade()._ORDER_INLINE.match(raw)
            if order_m and (not out['order_number']):
                out['order_number'] = order_m.group(1).strip()
    return out

def _looks_like_non_product_token(value: str) -> bool:
    text = str(value or '').strip().lower()
    return bool(text) and text in _facade()._NON_PRODUCT_TOKENS

def _looks_like_titleish(value: str) -> bool:
    text = str(value or '').strip().lower()
    if not text:
        return False
    if text in _facade()._NON_PRODUCT_TOKENS:
        return True
    return bool(_facade().re.fullmatch("[a-z][a-z\\s\\-']{2,40}", text)) and (not _facade().re.search('\\d', text))

def _build_item_from_row(ws, row: int, mapping: dict[str, int]) -> dict[str, _facade().Any] | None:
    model = ''
    name = ''
    if 'model_number' in mapping:
        model = str(ws.cell(row, mapping['model_number']).value or '').strip()
    if 'product_name' in mapping:
        name = str(ws.cell(row, mapping['product_name']).value or '').strip()
    if not name and model and (not _facade().re.search('[A-Za-z0-9]', model)):
        (name, model) = (model, '')
    if not name and (not model):
        return None
    if _facade()._looks_like_non_product_token(name) or _facade()._looks_like_non_product_token(model):
        return None
    tins = _facade()._to_int(ws.cell(row, mapping['quantity_tins']).value) if 'quantity_tins' in mapping else 0
    tin_spec = _facade()._to_float(ws.cell(row, mapping['tin_spec']).value) if 'tin_spec' in mapping else 0.0
    qty_kg = _facade()._to_float(ws.cell(row, mapping['quantity_kg']).value) if 'quantity_kg' in mapping else 0.0
    unit_price = _facade()._to_float(ws.cell(row, mapping['unit_price']).value) if 'unit_price' in mapping else 0.0
    amount = _facade()._to_float(ws.cell(row, mapping['amount']).value) if 'amount' in mapping else 0.0
    if tins <= 0 and qty_kg <= 0 and (unit_price <= 0) and (amount <= 0):
        return None
    if 1990 <= tins <= 2035 and (not model) and _facade()._looks_like_titleish(name):
        return None
    if tin_spec <= 0 and tins > 0 and (qty_kg > 0):
        tin_spec = qty_kg / tins
    if qty_kg <= 0 and tins > 0 and (tin_spec > 0):
        qty_kg = tins * tin_spec
    if amount <= 0 and unit_price > 0 and (qty_kg > 0):
        amount = unit_price * qty_kg
    if tins <= 0 and qty_kg > 0:
        tins = 1
        if tin_spec <= 0:
            tin_spec = qty_kg
    return {'product_name': name or model, 'model_number': model, 'quantity_tins': max(0, tins), 'tin_spec': tin_spec or 0.0, 'spec_per_tin': tin_spec or 0.0, 'quantity_kg': qty_kg, 'unit_price': unit_price, 'amount': amount, 'quantity': max(1, tins) if tins else 1}

def _parse_items(ws, header_row: int, mapping: dict[str, int], profile: _facade().ShipmentEtlProfile) -> list[dict[str, _facade().Any]]:
    items: list[dict[str, _facade().Any]] = []
    max_row = int(ws.max_row or 0)
    for row in range(header_row + 1, max_row + 1):
        joined = _facade()._joined_row(ws, row)
        if not joined:
            continue
        if profile.meta_patterns.stop_row.search(joined):
            break
        item = _facade()._build_item_from_row(ws, row, mapping)
        if item:
            items.append(item)
    return items

def note_fingerprint(note: dict[str, _facade().Any]) -> str:
    """内容指纹：同客户+单号+明细再导入可幂等跳过。"""
    payload = {'unit': str(note.get('unit_name') or '').strip(), 'order': str(note.get('order_number') or '').strip(), 'date': str(note.get('order_date') or '').strip(), 'items': sorted([{'m': str(i.get('model_number') or '').strip().upper(), 'n': str(i.get('product_name') or '').strip(), 'q': float(i.get('quantity_tins') or i.get('quantity') or 0), 'k': float(i.get('quantity_kg') or 0), 'p': float(i.get('unit_price') or 0)} for i in note.get('items') or []], key=lambda x: (x['m'], x['n'], x['q'], x['k'], x['p']))}
    raw = _facade().json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(',', ':'))
    return _facade().hashlib.sha256(raw.encode('utf-8')).hexdigest()[:28]

def _fingerprint_store_path() -> _facade().Path:
    """兼容旧测试 monkeypatch；真实幂等改走 SQLite。"""
    try:
        from app.utils.path_io.path_utils import get_data_dir
        root = _facade().Path(get_data_dir())
    except _facade().RECOVERABLE_ERRORS:
        root = _facade().Path.cwd() / 'data'
    root.mkdir(parents=True, exist_ok=True)
    return root / 'shipment_etl_fingerprints.json'

def _legacy_json_has_fingerprint(fingerprint: str) -> bool:
    path = _facade()._fingerprint_store_path()
    if not path.is_file():
        return False
    try:
        data = _facade().json.loads(path.read_text(encoding='utf-8'))
        entries = data.get('entries') if isinstance(data, dict) else None
        return bool(isinstance(entries, dict) and fingerprint in entries)
    except _facade().RECOVERABLE_ERRORS:
        return False

def _is_fingerprint_imported(tenant_key: str, fingerprint: str) -> bool:
    from app.application.shipment_excel_etl_fingerprint_store import has_fingerprint
    if has_fingerprint(tenant_key, fingerprint):
        return True
    return _facade()._legacy_json_has_fingerprint(fingerprint)

def _record_fingerprint_now(tenant_key: str, fingerprint: str, *, shipment_id: _facade().Any=None, unit_name: str='', order_number: str='', file_name: str='') -> None:
    from app.application.shipment_excel_etl_fingerprint_store import record_fingerprint
    record_fingerprint(tenant_key, fingerprint, shipment_id=shipment_id, unit_name=unit_name, order_number=order_number, file_name=file_name)

def _load_fingerprints() -> dict[str, _facade().Any]:
    path = _facade()._fingerprint_store_path()
    if not path.is_file():
        return {'entries': {}}
    try:
        data = _facade().json.loads(path.read_text(encoding='utf-8'))
        if isinstance(data, dict) and isinstance(data.get('entries'), dict):
            return data
    except _facade().RECOVERABLE_ERRORS:
        _facade().logger.warning('failed to load shipment etl fingerprints', exc_info=True)
    return {'entries': {}}

def _save_fingerprints(data: dict[str, _facade().Any]) -> None:
    path = _facade()._fingerprint_store_path()
    path.write_text(_facade().json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')

def _enrich_note(note: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    out = dict(note)
    out['sheet_name'] = str(out.get('sheet_name') or out.get('sheet') or '')
    out['sheet'] = out['sheet_name'] or str(out.get('sheet') or '')
    out['fingerprint'] = _facade().note_fingerprint(out)
    out['item_count'] = len(out.get('items') or [])
    out['total_amount'] = round(sum((float(i.get('amount') or 0) for i in out.get('items') or [])), 2)
    return out

def _build_sheet_probe(ws, profile: _facade().ShipmentEtlProfile, *, rule_hint: dict[str, _facade().Any] | None=None) -> _facade().Any:
    from app.application.shipment_excel_etl_llm import SheetProbe
    max_row = int(ws.max_row or 0)
    max_col = min(16, int(ws.max_column or 0) or 16)
    probe_n = min(10, max_row)
    probe_rows: list[dict[str, _facade().Any]] = []
    for row in range(1, probe_n + 1):
        cells = []
        for col in range(1, max_col + 1):
            raw = ws.cell(row, col).value
            if raw is None or str(raw).strip() == '':
                continue
            cells.append({'col': col, 'text': str(raw).strip()[:80]})
        if cells:
            probe_rows.append({'row': row, 'cells': cells})
    candidate_headers: list[dict[str, _facade().Any]] = []
    for row in range(1, min(16, max_row) + 1):
        cells = []
        for col in range(1, max_col + 1):
            raw = ws.cell(row, col).value
            text = str(raw).strip() if raw is not None else ''
            if not text:
                continue
            samples: list[str] = []
            for r in range(row + 1, min(row + 4, max_row + 1)):
                sv = ws.cell(r, col).value
                if sv is None or str(sv).strip() == '':
                    continue
                samples.append(str(sv).strip()[:40])
                if len(samples) >= 3:
                    break
            cells.append({'col': col, 'header': text[:80], 'samples': samples})
        if len(cells) >= 2:
            candidate_headers.append({'row': row, 'cells': cells})
    return SheetProbe(profile_id=profile.id, sheet_title=str(ws.title or ''), probe_rows=probe_rows, candidate_headers=candidate_headers[:8], max_row=max_row, max_col=max_col, rule_hint=dict(rule_hint or {}))

def _merge_meta(base: dict[str, str], overlay: dict[str, str], *, prefer_overlay: bool=False) -> dict[str, str]:
    out = dict(base)
    for key in ('unit_name', 'contact_person', 'order_date', 'order_number', 'title'):
        val = str((overlay or {}).get(key) or '').strip()
        if not val:
            continue
        if prefer_overlay or not str(out.get(key) or '').strip():
            out[key] = val
    return out

def _apply_llm_assist_to_layout(ws, profile: _facade().ShipmentEtlProfile, *, delivery_score: int, ledger_score: int, min_score: int, header_row: int | None, mapping: dict[str, int], meta: dict[str, str] | None, prefer_kind: str | None, fallback_unit: str='') -> tuple[int | None, dict[str, int], dict[str, str], str | None, dict[str, _facade().Any]]:
    """低置信时请求 LLM；返回 (header_row, mapping, meta, source_kind, assist_public)."""
    from app.application.shipment_excel_etl_llm import assist_sheet_layout, needs_llm_assist
    (need, reason) = needs_llm_assist(delivery_score=delivery_score, ledger_score=ledger_score, min_score=min_score, header_row=header_row, mapping=mapping, meta=meta, prefer_kind=prefer_kind, fallback_unit=fallback_unit)
    assist_public: dict[str, _facade().Any] = {'used_llm': False, 'cache_hit': False, 'ok': False, 'confidence': 1.0 if not need else 0.0, 'reason': reason}
    if not need:
        assist_public['ok'] = True
        return (header_row, mapping, dict(meta or {}), prefer_kind, assist_public)
    probe = _facade()._build_sheet_probe(ws, profile, rule_hint={'delivery_score': delivery_score, 'ledger_score': ledger_score, 'min_score': min_score, 'prefer_kind': prefer_kind, 'rule_header_row': header_row, 'rule_mapping': mapping, 'rule_meta': meta or {}, 'assist_reason': reason})
    assist = assist_sheet_layout(probe)
    assist_public = assist.as_public_dict()
    if not assist.ok:
        return (header_row, mapping, dict(meta or {}), prefer_kind, assist_public)
    new_header = assist.header_row if assist.header_row is not None else header_row
    new_mapping = dict(mapping)
    for (field_name, col) in (assist.columns or {}).items():
        if field_name not in new_mapping and isinstance(col, int) and (col > 0):
            new_mapping[field_name] = col
    for (field_name, col) in (assist.columns or {}).items():
        if field_name in {'product_name', 'model_number', 'order_number', 'quantity_tins', 'quantity_kg'}:
            if field_name not in mapping and isinstance(col, int) and (col > 0):
                new_mapping[field_name] = col
    new_meta = _facade()._merge_meta(dict(meta or {}), assist.meta or {}, prefer_overlay=True)
    kind = assist.source_kind if assist.source_kind in {'delivery_note', 'shipment_ledger', 'ignore'} else prefer_kind
    return (new_header, new_mapping, new_meta, kind, assist_public)

def _parse_delivery_sheet(ws, *, fallback_unit: str, profile: _facade().ShipmentEtlProfile, allow_llm: bool=True) -> dict[str, _facade().Any] | None:
    d_score = _facade()._score_delivery_sheet(ws, profile)
    l_score = _facade()._score_ledger_sheet(ws, profile)
    (kb_header, kb_mapping, kb_fp) = _facade()._kb_resolve_layout(ws)
    header_row = kb_header if kb_header is not None else _facade()._find_header_row(ws, profile)
    mapping = dict(kb_mapping) if kb_mapping else _facade()._map_headers(ws, header_row, profile) if header_row is not None else {}
    meta = _facade()._parse_buyer_meta(ws, header_row, profile) if header_row is not None else {'unit_name': '', 'contact_person': '', 'order_date': '', 'order_number': '', 'title': ''}
    assist_public: dict[str, _facade().Any] = {'used_llm': False, 'cache_hit': bool(kb_fp), 'ok': bool(kb_fp and mapping), 'confidence': 1.0 if kb_fp else 1.0, 'reason': 'knowledge_base_hit' if kb_fp else 'rules_only', 'layout_fingerprint': kb_fp or ''}
    from app.application.shipment_excel_etl_llm import unit_name_is_weak
    unit_weak = unit_name_is_weak(str((meta or {}).get('unit_name') or ''), fallback=fallback_unit)
    if allow_llm and (not kb_fp or unit_weak):
        (header_row, mapping, meta, kind, assist_public) = _facade()._apply_llm_assist_to_layout(ws, profile, delivery_score=d_score, ledger_score=l_score, min_score=profile.delivery_min_score, header_row=header_row, mapping=mapping, meta=meta, prefer_kind='delivery_note', fallback_unit=fallback_unit)
        if kind == 'ignore':
            return None
        if kind == 'shipment_ledger':
            return None
        if kb_fp and (not assist_public.get('layout_fingerprint')):
            assist_public['layout_fingerprint'] = kb_fp
    heuristic_on = str(_facade().os.environ.get('FHD_EXCEL_ETL_HEURISTIC') or '1').strip().lower() not in {'0', 'false', 'no', 'off'}
    if heuristic_on and header_row is not None and ('product_name' not in mapping and 'model_number' not in mapping):
        inferred = _facade()._infer_columns_from_samples(ws, header_row, mapping)
        if 'product_name' in inferred or 'model_number' in inferred:
            mapping = inferred
            if not assist_public.get('used_llm'):
                assist_public = {**assist_public, 'ok': True, 'confidence': 0.65, 'reason': 'heuristic_samples'}
    if header_row is None:
        return None
    if 'product_name' not in mapping and 'model_number' not in mapping:
        return None
    items = _facade()._parse_items(ws, header_row, mapping, profile)
    if not items:
        return None
    remembered_fp = _facade()._remember_sheet_layout(ws, header_row=header_row, mapping=mapping, profile=profile, source='knowledge_base' if kb_fp else 'llm' if assist_public.get('used_llm') else 'rules')
    if remembered_fp and (not assist_public.get('layout_fingerprint')):
        assist_public['layout_fingerprint'] = remembered_fp
    unit = meta.get('unit_name') or fallback_unit
    note = _facade()._enrich_note({'sheet': ws.title, 'source_kind': 'delivery_note', 'score': d_score, 'unit_name': unit, 'contact_person': meta.get('contact_person') or '', 'order_date': meta.get('order_date') or '', 'order_number': meta.get('order_number') or '', 'title': meta.get('title') or '', 'items': items, 'assist': assist_public})
    return note

def _excel_date_to_str(value: _facade().Any, profile: _facade().ShipmentEtlProfile) -> str:
    if value is None or value == '':
        return ''
    if isinstance(value, _facade().datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel
            return _facade().cast('str', from_excel(value).strftime('%Y-%m-%d'))
        except _facade().RECOVERABLE_ERRORS:
            return str(value)
    text = str(value).strip()
    date_m = profile.meta_patterns.date.search(text)
    return date_m.group(1).replace(' ', '') if date_m else text

def _parse_ledger_sheet(ws, *, fallback_unit: str, profile: _facade().ShipmentEtlProfile, allow_llm: bool=True) -> list[dict[str, _facade().Any]]:
    d_score = _facade()._score_delivery_sheet(ws, profile)
    l_score = _facade()._score_ledger_sheet(ws, profile)
    (kb_header, kb_mapping, kb_fp) = _facade()._kb_resolve_layout(ws)
    header_row = kb_header if kb_header is not None else _facade()._find_ledger_header_row(ws, profile)
    mapping = dict(kb_mapping) if kb_mapping else _facade()._map_headers(ws, header_row, profile) if header_row is not None else {}
    meta: dict[str, str] = {'unit_name': '', 'contact_person': '', 'order_date': '', 'order_number': '', 'title': ''}
    assist_public: dict[str, _facade().Any] = {'used_llm': False, 'cache_hit': bool(kb_fp), 'ok': bool(kb_fp and mapping), 'confidence': 1.0, 'reason': 'knowledge_base_hit' if kb_fp else 'rules_only', 'layout_fingerprint': kb_fp or ''}
    if allow_llm and (not kb_fp):
        (header_row, mapping, meta, kind, assist_public) = _facade()._apply_llm_assist_to_layout(ws, profile, delivery_score=d_score, ledger_score=l_score, min_score=profile.delivery_min_score, header_row=header_row, mapping=mapping, meta=meta, prefer_kind='shipment_ledger', fallback_unit=fallback_unit)
        if kind == 'ignore':
            return []
        if kind == 'delivery_note':
            return []
    if header_row is None:
        return []
    if 'order_number' not in mapping:
        return []
    if 'product_name' not in mapping and 'model_number' not in mapping:
        return []
    _facade()._remember_sheet_layout(ws, header_row=header_row, mapping=mapping, profile=profile, source='knowledge_base' if kb_fp else 'llm' if assist_public.get('used_llm') else 'rules')
    title_tpl = str((profile.ledger or {}).get('title_template') or '{unit}/{order_no}')
    unit_fallback = str(meta.get('unit_name') or fallback_unit).strip() or fallback_unit
    groups: dict[str, dict[str, _facade().Any]] = {}
    max_row = int(ws.max_row or 0)
    for row in range(header_row + 1, max_row + 1):
        joined = _facade()._joined_row(ws, row)
        if not joined:
            continue
        order_no = str(ws.cell(row, mapping['order_number']).value or '').strip()
        if not order_no:
            continue
        item = _facade()._build_item_from_row(ws, row, mapping)
        if not item:
            continue
        order_date = ''
        if 'order_date' in mapping:
            order_date = _facade()._excel_date_to_str(ws.cell(row, mapping['order_date']).value, profile)
        bucket = groups.setdefault(order_no, {'sheet': ws.title, 'source_kind': 'shipment_ledger', 'score': l_score, 'unit_name': unit_fallback, 'contact_person': meta.get('contact_person') or '', 'order_date': order_date, 'order_number': order_no, 'title': title_tpl.format(unit=unit_fallback, order_no=order_no), 'items': [], 'assist': assist_public})
        if order_date and (not bucket.get('order_date')):
            bucket['order_date'] = order_date
        bucket['items'].append(item)
    return [_facade()._enrich_note(g) for g in groups.values() if g.get('items')]
