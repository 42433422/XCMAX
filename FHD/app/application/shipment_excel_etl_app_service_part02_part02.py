# mypy: disable-error-code="no-any-return, valid-type"
"""Implementation extracted from the public facade module."""

from __future__ import annotations

import importlib


def _facade():
    return importlib.import_module("app.application.shipment_excel_etl_app_service")


def _apply_llm_assist_to_layout(
    ws,
    profile: _facade().ShipmentEtlProfile,
    *,
    delivery_score: int,
    ledger_score: int,
    min_score: int,
    header_row: int | None,
    mapping: dict[str, int],
    meta: dict[str, str] | None,
    prefer_kind: str | None,
    fallback_unit: str = "",
) -> tuple[int | None, dict[str, int], dict[str, str], str | None, dict[str, _facade().Any]]:
    """低置信时请求 LLM；返回 (header_row, mapping, meta, source_kind, assist_public)."""
    from app.application.shipment_excel_etl_llm import assist_sheet_layout, needs_llm_assist

    need, reason = needs_llm_assist(
        delivery_score=delivery_score,
        ledger_score=ledger_score,
        min_score=min_score,
        header_row=header_row,
        mapping=mapping,
        meta=meta,
        prefer_kind=prefer_kind,
        fallback_unit=fallback_unit,
    )
    assist_public: dict[str, _facade().Any] = {
        "used_llm": False,
        "cache_hit": False,
        "ok": False,
        "confidence": 1.0 if not need else 0.0,
        "reason": reason,
    }
    if not need:
        assist_public["ok"] = True
        return (header_row, mapping, dict(meta or {}), prefer_kind, assist_public)
    probe = _facade()._build_sheet_probe(
        ws,
        profile,
        rule_hint={
            "delivery_score": delivery_score,
            "ledger_score": ledger_score,
            "min_score": min_score,
            "prefer_kind": prefer_kind,
            "rule_header_row": header_row,
            "rule_mapping": mapping,
            "rule_meta": meta or {},
            "assist_reason": reason,
        },
    )
    assist = assist_sheet_layout(probe)
    assist_public = assist.as_public_dict()
    if not assist.ok:
        return (header_row, mapping, dict(meta or {}), prefer_kind, assist_public)
    new_header = assist.header_row if assist.header_row is not None else header_row
    new_mapping = dict(mapping)
    for field_name, col in (assist.columns or {}).items():
        if field_name not in new_mapping and isinstance(col, int) and (col > 0):
            new_mapping[field_name] = col
    for field_name, col in (assist.columns or {}).items():
        if field_name in {
            "product_name",
            "model_number",
            "order_number",
            "quantity_tins",
            "quantity_kg",
        }:
            if field_name not in mapping and isinstance(col, int) and (col > 0):
                new_mapping[field_name] = col
    new_meta = _facade()._merge_meta(dict(meta or {}), assist.meta or {}, prefer_overlay=True)
    kind = (
        assist.source_kind
        if assist.source_kind in {"delivery_note", "shipment_ledger", "ignore"}
        else prefer_kind
    )
    return (new_header, new_mapping, new_meta, kind, assist_public)


def _parse_delivery_sheet(
    ws, *, fallback_unit: str, profile: _facade().ShipmentEtlProfile, allow_llm: bool = True
) -> dict[str, _facade().Any] | None:
    d_score = _facade()._score_delivery_sheet(ws, profile)
    l_score = _facade()._score_ledger_sheet(ws, profile)
    kb_header, kb_mapping, kb_fp = _facade()._kb_resolve_layout(ws)
    header_row = kb_header if kb_header is not None else _facade()._find_header_row(ws, profile)
    mapping = (
        dict(kb_mapping)
        if kb_mapping
        else _facade()._map_headers(ws, header_row, profile)
        if header_row is not None
        else {}
    )
    meta = (
        _facade()._parse_buyer_meta(ws, header_row, profile)
        if header_row is not None
        else {
            "unit_name": "",
            "contact_person": "",
            "order_date": "",
            "order_number": "",
            "title": "",
        }
    )
    assist_public: dict[str, _facade().Any] = {
        "used_llm": False,
        "cache_hit": bool(kb_fp),
        "ok": bool(kb_fp and mapping),
        "confidence": 1.0 if kb_fp else 1.0,
        "reason": "knowledge_base_hit" if kb_fp else "rules_only",
        "layout_fingerprint": kb_fp or "",
    }
    from app.application.shipment_excel_etl_llm import unit_name_is_weak

    unit_weak = unit_name_is_weak(str((meta or {}).get("unit_name") or ""), fallback=fallback_unit)
    if allow_llm and (not kb_fp or unit_weak):
        header_row, mapping, meta, kind, assist_public = _facade()._apply_llm_assist_to_layout(
            ws,
            profile,
            delivery_score=d_score,
            ledger_score=l_score,
            min_score=profile.delivery_min_score,
            header_row=header_row,
            mapping=mapping,
            meta=meta,
            prefer_kind="delivery_note",
            fallback_unit=fallback_unit,
        )
        if kind == "ignore":
            return None
        if kind == "shipment_ledger":
            return None
        if kb_fp and (not assist_public.get("layout_fingerprint")):
            assist_public["layout_fingerprint"] = kb_fp
    heuristic_on = str(
        _facade().os.environ.get("FHD_EXCEL_ETL_HEURISTIC") or "1"
    ).strip().lower() not in {"0", "false", "no", "off"}
    if (
        heuristic_on
        and header_row is not None
        and ("product_name" not in mapping and "model_number" not in mapping)
    ):
        inferred = _facade()._infer_columns_from_samples(ws, header_row, mapping)
        if "product_name" in inferred or "model_number" in inferred:
            mapping = inferred
            if not assist_public.get("used_llm"):
                assist_public = {
                    **assist_public,
                    "ok": True,
                    "confidence": 0.65,
                    "reason": "heuristic_samples",
                }
    if header_row is None:
        return None
    if "product_name" not in mapping and "model_number" not in mapping:
        return None
    items = _facade()._parse_items(ws, header_row, mapping, profile)
    if not items:
        return None
    remembered_fp = _facade()._remember_sheet_layout(
        ws,
        header_row=header_row,
        mapping=mapping,
        profile=profile,
        source="knowledge_base" if kb_fp else "llm" if assist_public.get("used_llm") else "rules",
    )
    if remembered_fp and (not assist_public.get("layout_fingerprint")):
        assist_public["layout_fingerprint"] = remembered_fp
    unit = meta.get("unit_name") or fallback_unit
    note = _facade()._enrich_note(
        {
            "sheet": ws.title,
            "source_kind": "delivery_note",
            "score": d_score,
            "unit_name": unit,
            "contact_person": meta.get("contact_person") or "",
            "order_date": meta.get("order_date") or "",
            "order_number": meta.get("order_number") or "",
            "title": meta.get("title") or "",
            "items": items,
            "assist": assist_public,
        }
    )
    return note


def _excel_date_to_str(value: _facade().Any, profile: _facade().ShipmentEtlProfile) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, _facade().datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        try:
            from openpyxl.utils.datetime import from_excel

            return _facade().cast("str", from_excel(value).strftime("%Y-%m-%d"))
        except _facade().RECOVERABLE_ERRORS:
            return str(value)
    text = str(value).strip()
    date_m = profile.meta_patterns.date.search(text)
    return date_m.group(1).replace(" ", "") if date_m else text


def _parse_ledger_sheet(
    ws, *, fallback_unit: str, profile: _facade().ShipmentEtlProfile, allow_llm: bool = True
) -> list[dict[str, _facade().Any]]:
    d_score = _facade()._score_delivery_sheet(ws, profile)
    l_score = _facade()._score_ledger_sheet(ws, profile)
    kb_header, kb_mapping, kb_fp = _facade()._kb_resolve_layout(ws)
    header_row = (
        kb_header if kb_header is not None else _facade()._find_ledger_header_row(ws, profile)
    )
    mapping = (
        dict(kb_mapping)
        if kb_mapping
        else _facade()._map_headers(ws, header_row, profile)
        if header_row is not None
        else {}
    )
    meta: dict[str, str] = {
        "unit_name": "",
        "contact_person": "",
        "order_date": "",
        "order_number": "",
        "title": "",
    }
    assist_public: dict[str, _facade().Any] = {
        "used_llm": False,
        "cache_hit": bool(kb_fp),
        "ok": bool(kb_fp and mapping),
        "confidence": 1.0,
        "reason": "knowledge_base_hit" if kb_fp else "rules_only",
        "layout_fingerprint": kb_fp or "",
    }
    if allow_llm and (not kb_fp):
        header_row, mapping, meta, kind, assist_public = _facade()._apply_llm_assist_to_layout(
            ws,
            profile,
            delivery_score=d_score,
            ledger_score=l_score,
            min_score=profile.delivery_min_score,
            header_row=header_row,
            mapping=mapping,
            meta=meta,
            prefer_kind="shipment_ledger",
            fallback_unit=fallback_unit,
        )
        if kind == "ignore":
            return []
        if kind == "delivery_note":
            return []
    if header_row is None:
        return []
    if "order_number" not in mapping:
        return []
    if "product_name" not in mapping and "model_number" not in mapping:
        return []
    _facade()._remember_sheet_layout(
        ws,
        header_row=header_row,
        mapping=mapping,
        profile=profile,
        source="knowledge_base" if kb_fp else "llm" if assist_public.get("used_llm") else "rules",
    )
    title_tpl = str((profile.ledger or {}).get("title_template") or "{unit}/{order_no}")
    unit_fallback = str(meta.get("unit_name") or fallback_unit).strip() or fallback_unit
    groups: dict[str, dict[str, _facade().Any]] = {}
    max_row = int(ws.max_row or 0)
    for row in range(header_row + 1, max_row + 1):
        joined = _facade()._joined_row(ws, row)
        if not joined:
            continue
        order_no = str(ws.cell(row, mapping["order_number"]).value or "").strip()
        if not order_no:
            continue
        item = _facade()._build_item_from_row(ws, row, mapping)
        if not item:
            continue
        order_date = ""
        if "order_date" in mapping:
            order_date = _facade()._excel_date_to_str(
                ws.cell(row, mapping["order_date"]).value, profile
            )
        bucket = groups.setdefault(
            order_no,
            {
                "sheet": ws.title,
                "source_kind": "shipment_ledger",
                "score": l_score,
                "unit_name": unit_fallback,
                "contact_person": meta.get("contact_person") or "",
                "order_date": order_date,
                "order_number": order_no,
                "title": title_tpl.format(unit=unit_fallback, order_no=order_no),
                "items": [],
                "assist": assist_public,
            },
        )
        if order_date and (not bucket.get("order_date")):
            bucket["order_date"] = order_date
        bucket["items"].append(item)
    return [_facade()._enrich_note(g) for g in groups.values() if g.get("items")]
