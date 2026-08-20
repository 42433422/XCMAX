# mypy: disable-error-code="valid-type, attr-defined, no-any-return"
"""Implementation extracted from the public facade module."""

from __future__ import annotations

import importlib


def _facade():
    return importlib.import_module("app.application.shipment_excel_etl_app_service")


def write_delivery_note_workbook(
    notes: list[dict[str, _facade().Any]],
    output_path: str | _facade().Path,
    *,
    seller_title: str | None = None,
    profile_id: str | None = None,
    profile: _facade().ShipmentEtlProfile | None = None,
) -> dict[str, _facade().Any]:
    """按 profile.write 版式写出送货单模板（可用于回环验证 / 测试数据）。"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return {"success": False, "message": "缺少 openpyxl，无法解析 Excel"}
    prof = _facade()._resolve_profile(profile, profile_id)
    write_cfg = prof.write or {}
    title = str(seller_title if seller_title is not None else write_cfg.get("seller_title") or "")
    headers = list(write_cfg.get("header_row") or [])
    item_cols = dict(write_cfg.get("item_columns") or {})
    date_fmt = str(write_cfg.get("date_format") or "%Y-%m-%d")
    meta_tpl = str(
        write_cfg.get("meta_line_template") or "{unit} {contact} {order_date} {order_no}"
    )
    footer = str(write_cfg.get("footer_label") or "")
    default_sheet = str(write_cfg.get("default_sheet_name") or "Sheet1")
    sheet_prefix = str(write_cfg.get("sheet_name_prefix") or "S")
    demo_meta = str(write_cfg.get("demo_meta_line") or meta_tpl)
    demo_item = dict(write_cfg.get("demo_item") or {})
    path = _facade().Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    if default is None:
        default = wb.create_sheet(default_sheet[:31])
    created = 0
    for idx, note in enumerate(notes or [], start=1):
        unit = str(note.get("unit_name") or f"客户{idx}").strip()
        sheet_name = (
            str(note.get("sheet_name") or note.get("sheet") or unit)[:28] or f"{sheet_prefix}{idx}"
        )
        sheet_name = _facade().re.sub("[\\\\/*?:\\[\\]]", "_", sheet_name)[:31]
        if idx == 1:
            ws = default
            assert ws is not None
            ws.title = sheet_name
        else:
            base = sheet_name
            n = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{base[:28]}_{n}"
                n += 1
            ws = wb.create_sheet(sheet_name)
        contact = str(note.get("contact_person") or "").strip()
        order_date = str(
            note.get("order_date") or _facade().datetime.now().strftime(date_fmt)
        ).strip()
        order_no = str(note.get("order_number") or f"TEST-{idx:04d}").strip()
        ws["A1"] = title
        ws["A2"] = meta_tpl.format(
            unit=unit, contact=contact, order_date=order_date, order_no=order_no
        )
        for col, h in enumerate(headers, start=1):
            ws.cell(3, col, h)
        last_row = 3
        for r, item in enumerate(note.get("items") or [], start=4):
            if "model_number" in item_cols:
                ws.cell(r, int(item_cols["model_number"]), item.get("model_number") or "")
            if "product_name" in item_cols:
                ws.cell(r, int(item_cols["product_name"]), item.get("product_name") or "")
            if "quantity_tins" in item_cols:
                ws.cell(
                    r,
                    int(item_cols["quantity_tins"]),
                    item.get("quantity_tins") or item.get("quantity") or 0,
                )
            if "tin_spec" in item_cols:
                ws.cell(
                    r,
                    int(item_cols["tin_spec"]),
                    item.get("tin_spec") or item.get("spec_per_tin") or 0,
                )
            if "quantity_kg" in item_cols:
                ws.cell(r, int(item_cols["quantity_kg"]), item.get("quantity_kg") or 0)
            if "unit_price" in item_cols:
                ws.cell(r, int(item_cols["unit_price"]), item.get("unit_price") or 0)
            if "amount" in item_cols:
                ws.cell(r, int(item_cols["amount"]), item.get("amount") or 0)
            last_row = r
        if footer:
            ws.cell(last_row + 2, 1, footer)
        created += 1
    if created == 0:
        ws = default
        ws.title = default_sheet[:31]
        ws["A1"] = title
        ws["A2"] = demo_meta
        for col, h in enumerate(headers, start=1):
            ws.cell(3, col, h)
        r = 4
        if "model_number" in item_cols:
            ws.cell(r, int(item_cols["model_number"]), demo_item.get("model_number") or "")
        if "product_name" in item_cols:
            ws.cell(r, int(item_cols["product_name"]), demo_item.get("product_name") or "")
        if "quantity_tins" in item_cols:
            ws.cell(r, int(item_cols["quantity_tins"]), demo_item.get("quantity_tins") or 0)
        if "tin_spec" in item_cols:
            ws.cell(r, int(item_cols["tin_spec"]), demo_item.get("tin_spec") or 0)
        if "quantity_kg" in item_cols:
            ws.cell(r, int(item_cols["quantity_kg"]), demo_item.get("quantity_kg") or 0)
        if "unit_price" in item_cols:
            ws.cell(r, int(item_cols["unit_price"]), demo_item.get("unit_price") or 0)
        if "amount" in item_cols:
            ws.cell(r, int(item_cols["amount"]), demo_item.get("amount") or 0)
        created = 1
    wb.save(path)
    wb.close()
    return {
        "success": True,
        "file_path": str(path),
        "sheet_count": created,
        "profile_id": prof.id,
        "message": f"已生成送货单模板 {path.name}（{created} 张表）",
    }


def write_ledger_workbook(
    rows: list[dict[str, _facade().Any]],
    output_path: str | _facade().Path,
    *,
    sheet_name: str | None = None,
    unit_name: str | None = None,
    profile_id: str | None = None,
    profile: _facade().ShipmentEtlProfile | None = None,
) -> dict[str, _facade().Any]:
    """写出出货流水模板（表头/列位来自 profile.write）。"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return {"success": False, "message": "缺少 openpyxl，无法解析 Excel"}
    prof = _facade()._resolve_profile(profile, profile_id)
    write_cfg = prof.write or {}
    default_sheet = str(sheet_name or write_cfg.get("ledger_sheet_name") or "ledger")
    resolved_unit = str(unit_name or write_cfg.get("ledger_default_unit") or "unit")
    headers = list(write_cfg.get("ledger_header_row") or [])
    item_cols = dict(write_cfg.get("ledger_item_columns") or {})
    sample_rows = rows or list(write_cfg.get("ledger_sample_rows") or [])
    extra_sheet = str(write_cfg.get("ledger_extra_sheet") or "").strip()
    path = _facade().Path(output_path).expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = default_sheet[:31] or "ledger"
    for col, h in enumerate(headers, start=1):
        ws.cell(1, col, h)
    for r, row in enumerate(sample_rows, start=2):
        for field_name, col_idx in item_cols.items():
            raw = row.get(field_name)
            if raw is None or raw == "":
                if field_name in {
                    "quantity_tins",
                    "tin_spec",
                    "quantity_kg",
                    "unit_price",
                    "amount",
                }:
                    raw = 0
                else:
                    raw = ""
            ws.cell(r, int(col_idx), raw)
    if extra_sheet and extra_sheet not in wb.sheetnames:
        wb.create_sheet(extra_sheet[:31])
    wb.save(path)
    wb.close()
    return {
        "success": True,
        "file_path": str(path),
        "unit_name": resolved_unit,
        "row_count": len(sample_rows),
        "profile_id": prof.id,
        "message": f"已生成出货流水模板 {path.name}",
    }


def regenerate_delivery_notes_from_file(
    file_path: str | _facade().Path,
    output_path: str | _facade().Path,
    *,
    include_ledger: bool | str = "auto",
    profile_id: str | None = None,
    profile: _facade().ShipmentEtlProfile | None = None,
) -> dict[str, _facade().Any]:
    """解析 → 按 profile 送货单版式再出单（模板反推闭环）。"""
    prof = _facade()._resolve_profile(profile, profile_id)
    parsed = _facade().parse_delivery_notes(file_path, include_ledger=include_ledger, profile=prof)
    if not parsed.get("success"):
        return parsed
    notes = parsed.get("notes") or []
    if not notes:
        return {"success": False, "message": "无可反推的单据", "error_code": "no_delivery_notes"}
    written = _facade().write_delivery_note_workbook(notes, output_path, profile=prof)
    if not written.get("success"):
        return written
    reparsed = _facade().parse_delivery_notes(output_path, include_ledger=False, profile=prof)
    return {
        "success": True,
        "source": parsed,
        "generated": written,
        "reparsed": reparsed,
        "fingerprint_match": {n.get("fingerprint") for n in notes}
        == {n.get("fingerprint") for n in reparsed.get("notes") or []}
        if reparsed.get("success")
        else False,
        "profile_id": prof.id,
        "message": "模板反推完成",
    }


def batch_preview_shipment_excel_etl(
    directory: str | _facade().Path,
    *,
    include_ledger: bool | str = "auto",
    pattern: str = "*.xlsx",
    workspace_root: str | _facade().Path | None = None,
    profile_id: str | None = None,
    profile: _facade().ShipmentEtlProfile | None = None,
) -> dict[str, _facade().Any]:
    from app.application.shipment_excel_etl_security import ShipmentEtlPathError, resolve_etl_path

    prof = _facade()._resolve_profile(profile, profile_id)
    try:
        root = resolve_etl_path(directory, workspace_root=workspace_root, must_exist=True)
    except ShipmentEtlPathError:
        return {"success": False, "message": "非法目录", "error_code": "unsafe_path", "files": []}
    if not root.is_dir():
        return {"success": False, "message": f"目录不存在: {root}", "files": []}
    files = sorted(root.glob(pattern))
    results = []
    total_notes = 0
    for path in files:
        if path.name.startswith("~$"):
            continue
        preview = _facade().preview_shipment_excel_etl(
            path,
            include_ledger=include_ledger,
            unit_name_hint=path.stem,
            workspace_root=workspace_root or root,
            profile=prof,
        )
        note_count = int(preview.get("note_count") or 0)
        total_notes += note_count
        results.append(
            {
                "file_path": str(path),
                "file_name": path.name,
                "success": bool(preview.get("success")),
                "note_count": note_count,
                "duplicate_note_count": preview.get("duplicate_note_count", 0),
                "message": preview.get("message"),
                "notes": preview.get("notes") or [],
            }
        )
    return {
        "success": True,
        "directory": str(root),
        "file_count": len(results),
        "note_count": total_notes,
        "files": results,
        "profile_id": prof.id,
        "message": f"批量预览完成：{len(results)} 个文件，共 {total_notes} 张单据",
    }


def batch_execute_shipment_excel_etl(
    directory: str | _facade().Path,
    *,
    include_ledger: bool | str = False,
    pattern: str = "*.xlsx",
    idempotent: bool = True,
    import_products: bool = True,
    import_shipments: bool = True,
    confirm_ledger: bool = False,
    dry_run: bool = False,
    workspace_root: str | _facade().Path | None = None,
    profile_id: str | None = None,
    profile: _facade().ShipmentEtlProfile | None = None,
) -> dict[str, _facade().Any]:
    from app.application.shipment_excel_etl_security import (
        ShipmentEtlPathError,
        batch_execute_allowed,
        resolve_etl_path,
    )

    prof = _facade()._resolve_profile(profile, profile_id)
    if not dry_run and (not batch_execute_allowed()):
        return {
            "success": False,
            "message": "批量入库默认关闭。需设置环境变量 FHD_SHIPMENT_ETL_ALLOW_BATCH=1",
            "error_code": "batch_disabled",
            "files": [],
        }
    try:
        root = resolve_etl_path(directory, workspace_root=workspace_root, must_exist=True)
    except ShipmentEtlPathError:
        return {"success": False, "message": "非法目录", "error_code": "unsafe_path", "files": []}
    if not root.is_dir():
        return {"success": False, "message": f"目录不存在: {root}", "files": []}
    files = sorted(root.glob(pattern))
    if len(files) > 50:
        return {
            "success": False,
            "message": f"批量文件过多（{len(files)}），上限 50，请缩小范围",
            "error_code": "batch_too_large",
            "files": [],
        }
    results = []
    created = skipped = failed = 0
    for path in files:
        if path.name.startswith("~$"):
            continue
        result = _facade().execute_shipment_excel_etl(
            path,
            include_ledger=include_ledger,
            unit_name_hint=path.stem,
            idempotent=idempotent,
            import_products=import_products,
            import_shipments=import_shipments,
            confirm_ledger=confirm_ledger,
            dry_run=dry_run,
            workspace_root=workspace_root or root,
            profile=prof,
        )
        created += int(result.get("shipment_created") or result.get("would_create") or 0)
        skipped += int(result.get("shipment_skipped") or result.get("would_skip") or 0)
        failed += int(result.get("shipment_failed") or 0)
        results.append(
            {
                "file_path": str(path),
                "file_name": path.name,
                "success": bool(result.get("success")),
                "shipment_created": result.get("shipment_created", result.get("would_create", 0)),
                "shipment_skipped": result.get("shipment_skipped", result.get("would_skip", 0)),
                "shipment_failed": result.get("shipment_failed", 0),
                "message": result.get("message"),
                "error_code": result.get("error_code"),
            }
        )
    return {
        "success": failed == 0,
        "directory": str(root),
        "file_count": len(results),
        "shipment_created": created,
        "shipment_skipped": skipped,
        "shipment_failed": failed,
        "files": results,
        "closed_loop": not dry_run,
        "dry_run": dry_run,
        "profile_id": prof.id,
        "message": f"{('批量预演' if dry_run else '批量入库')}完成：新建/将建 {created}，跳过 {skipped}，失败 {failed}",
    }


class ShipmentExcelEtlApplicationService:
    def __init__(self, profile_id: str | None = None) -> None:
        self._profile_id = profile_id

    def _profile_kwargs(self, kwargs: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
        if "profile" in kwargs or "profile_id" in kwargs:
            return kwargs
        if self._profile_id:
            return {**kwargs, "profile_id": self._profile_id}
        return kwargs

    def preview(
        self, file_path: str | _facade().Path, **kwargs: _facade().Any
    ) -> dict[str, _facade().Any]:
        return _facade().preview_shipment_excel_etl(file_path, **self._profile_kwargs(kwargs))

    def execute(
        self, file_path: str | _facade().Path, **kwargs: _facade().Any
    ) -> dict[str, _facade().Any]:
        return _facade().execute_shipment_excel_etl(file_path, **self._profile_kwargs(kwargs))

    def batch_preview(
        self, directory: str | _facade().Path, **kwargs: _facade().Any
    ) -> dict[str, _facade().Any]:
        return _facade().batch_preview_shipment_excel_etl(directory, **self._profile_kwargs(kwargs))

    def batch_execute(
        self, directory: str | _facade().Path, **kwargs: _facade().Any
    ) -> dict[str, _facade().Any]:
        return _facade().batch_execute_shipment_excel_etl(directory, **self._profile_kwargs(kwargs))

    def write_delivery_template(
        self,
        notes: list[dict[str, _facade().Any]],
        output_path: str | _facade().Path,
        **kwargs: _facade().Any,
    ) -> dict[str, _facade().Any]:
        return _facade().write_delivery_note_workbook(
            notes, output_path, **self._profile_kwargs(kwargs)
        )

    def write_ledger_template(
        self,
        rows: list[dict[str, _facade().Any]],
        output_path: str | _facade().Path,
        **kwargs: _facade().Any,
    ) -> dict[str, _facade().Any]:
        return _facade().write_ledger_workbook(rows, output_path, **self._profile_kwargs(kwargs))

    def regenerate(
        self,
        file_path: str | _facade().Path,
        output_path: str | _facade().Path,
        **kwargs: _facade().Any,
    ) -> dict[str, _facade().Any]:
        return _facade().regenerate_delivery_notes_from_file(
            file_path, output_path, **self._profile_kwargs(kwargs)
        )

    def ocr_preview(
        self, file_path: str | _facade().Path, **kwargs: _facade().Any
    ) -> dict[str, _facade().Any]:
        from app.application.shipment_excel_etl_ocr import parse_ocr_document

        return parse_ocr_document(file_path, **self._profile_kwargs(kwargs))


def get_shipment_excel_etl_app_service() -> ShipmentExcelEtlApplicationService:
    global _svc
    if _facade()._svc is None:
        _facade()._svc = _facade().ShipmentExcelEtlApplicationService()
    return _facade()._svc
