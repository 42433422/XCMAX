"""Excel ETL OCR 桥接：扫描件/图片/PDF → 网格表 → 临时 xlsx → 走通用解析。

不硬编码送货单版式；OCR 只负责还原单元格，列映射仍由 universal/KB/LLM 完成。
"""

from __future__ import annotations

import logging
import os
import re
import tempfile
from pathlib import Path
from typing import Any

from app.utils.operational_errors import RECOVERABLE_ERRORS

logger = logging.getLogger(__name__)

_IMAGE_SUFFIXES = {".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"}
_PDF_SUFFIXES = {".pdf"}


def is_ocr_source(path: str | Path) -> bool:
    suffix = Path(path).suffix.lower()
    return suffix in _IMAGE_SUFFIXES or suffix in _PDF_SUFFIXES


def _cluster_axis(values: list[float], *, threshold: float) -> list[float]:
    if not values:
        return []
    ordered = sorted(values)
    clusters: list[list[float]] = [[ordered[0]]]
    for value in ordered[1:]:
        if abs(value - clusters[-1][-1]) <= threshold:
            clusters[-1].append(value)
        else:
            clusters.append([value])
    return [sum(c) / len(c) for c in clusters]


def _text_blocks_to_grid_result(
    blocks: list[dict[str, Any]],
    *,
    row_threshold: float | None = None,
    col_threshold: float | None = None,
) -> tuple[list[list[str]], list[dict[str, Any]]]:
    """Return the clustered grid plus auditable block-to-cell coordinates."""
    cleaned: list[dict[str, Any]] = []
    for raw in blocks or []:
        text = str(raw.get("text") or "").strip()
        if not text:
            continue
        if "center" in raw and isinstance(raw.get("center"), (tuple, list)):
            cx, cy = float(raw["center"][0]), float(raw["center"][1])
        else:
            left = float(raw.get("left") or 0)
            top = float(raw.get("top") or 0)
            width = float(raw.get("width") or 0)
            height = float(raw.get("height") or 0)
            cx = left + width / 2.0
            cy = float(raw.get("y_center") or (top + height / 2.0))
        cleaned.append({"text": text, "cx": cx, "cy": cy, "raw": raw})
    if not cleaned:
        return [], []

    ys = [b["cy"] for b in cleaned]
    xs = [b["cx"] for b in cleaned]
    y_span = max(ys) - min(ys) if ys else 1.0
    x_span = max(xs) - min(xs) if xs else 1.0
    row_th = float(row_threshold if row_threshold is not None else max(18.0, y_span * 0.03))
    col_th = float(col_threshold if col_threshold is not None else max(24.0, x_span * 0.04))

    row_centers = _cluster_axis(ys, threshold=row_th)
    col_centers = _cluster_axis(xs, threshold=col_th)
    if not row_centers or not col_centers:
        return [], []

    grid: list[list[str]] = [["" for _ in col_centers] for _ in row_centers]
    slots: dict[tuple[int, int], list[dict[str, Any]]] = {}
    for block in cleaned:
        r = min(range(len(row_centers)), key=lambda i: abs(row_centers[i] - block["cy"]))
        c = min(range(len(col_centers)), key=lambda i: abs(col_centers[i] - block["cx"]))
        slots.setdefault((r, c), []).append(block)

    cells: list[dict[str, Any]] = []
    for (row_index, column_index), parts in slots.items():
        ordered = sorted(parts, key=lambda item: item["cx"])
        text = " ".join(part["text"] for part in ordered).strip()
        grid[row_index][column_index] = text
        raw_parts = [part["raw"] for part in ordered]
        confidences = [
            float(raw.get("confidence", raw.get("score")))
            for raw in raw_parts
            if isinstance(raw.get("confidence", raw.get("score")), (int, float))
        ]
        lefts = [float(raw.get("left") or 0) for raw in raw_parts]
        tops = [float(raw.get("top") or 0) for raw in raw_parts]
        rights = [float(raw.get("left") or 0) + float(raw.get("width") or 0) for raw in raw_parts]
        bottoms = [float(raw.get("top") or 0) + float(raw.get("height") or 0) for raw in raw_parts]
        left = min(lefts, default=0.0)
        top = min(tops, default=0.0)
        right = max(rights, default=left)
        bottom = max(bottoms, default=top)
        cells.append(
            {
                "grid_row": row_index + 1,
                "grid_column": column_index + 1,
                "text": text,
                "confidence": min(confidences) if confidences else None,
                "left": left,
                "top": top,
                "width": max(0.0, right - left),
                "height": max(0.0, bottom - top),
                "center": [(left + right) / 2.0, (top + bottom) / 2.0],
            }
        )
    return grid, cells


def text_blocks_to_grid(
    blocks: list[dict[str, Any]],
    *,
    row_threshold: float | None = None,
    col_threshold: float | None = None,
) -> list[list[str]]:
    """将带坐标文本块聚类为二维表格（行优先、列从左到右）。"""
    grid, _cells = _text_blocks_to_grid_result(
        blocks,
        row_threshold=row_threshold,
        col_threshold=col_threshold,
    )
    return grid


def grid_to_workbook_path(
    grid: list[list[str]],
    *,
    output_path: str | Path | None = None,
    sheet_name: str = "OCR",
    meta_lines: list[str] | None = None,
) -> Path:
    from openpyxl import Workbook

    path = (
        Path(output_path)
        if output_path
        else Path(tempfile.mkstemp(prefix="etl_ocr_", suffix=".xlsx")[1])
    )
    path = path.expanduser().resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = (sheet_name or "OCR")[:31]
    row_idx = 1
    for line in meta_lines or []:
        text = str(line or "").strip()
        if not text:
            continue
        ws.cell(row_idx, 1, text)
        row_idx += 1
    for row in grid:
        for col, value in enumerate(row, start=1):
            ws.cell(row_idx, col, value)
        row_idx += 1
    wb.save(path)
    wb.close()
    return path


def page_grids_to_workbook_path(
    pages: list[dict[str, Any]],
    *,
    base_sheet_name: str = "OCR",
) -> Path:
    """把 OCR 多页结果写成逐页工作表，避免不同页面按相同坐标错误叠加。"""
    from openpyxl import Workbook

    # Always allocate the workbook ourselves.  The public OCR route may receive
    # an output_path for legacy compatibility, but user input must never select
    # the file opened by this writer.
    path = Path(tempfile.mkstemp(prefix="etl_ocr_", suffix=".xlsx")[1]).resolve()
    workbook = Workbook()
    workbook.remove(workbook.active)
    for index, page in enumerate(pages, start=1):
        raw_name = f"{base_sheet_name[:24]}_P{index}"
        worksheet = workbook.create_sheet(raw_name[:31])
        page["sheet_name"] = worksheet.title
        row_index = 1
        for line in page.get("meta_lines") or []:
            text = str(line or "").strip()
            if text:
                worksheet.cell(row_index, 1, text)
                row_index += 1
        page["data_start_row"] = row_index
        for cell in page.get("grid_cells") or []:
            cell["workbook_row"] = row_index + int(cell.get("grid_row") or 1) - 1
            cell["workbook_column"] = int(cell.get("grid_column") or 1)
        for row in page.get("grid") or []:
            for column, value in enumerate(row, start=1):
                worksheet.cell(row_index, column, value)
            row_index += 1
    if not workbook.worksheets:
        workbook.create_sheet("OCR")
    workbook.save(path)
    workbook.close()
    return path


def _safe_ocr_block(raw: dict[str, Any]) -> dict[str, Any]:
    """仅保留审计所需 OCR 证据，避免把后端对象或提示文本带入执行上下文。"""
    result: dict[str, Any] = {"text": str(raw.get("text") or "").strip()}
    for key in ("left", "top", "width", "height", "confidence", "score"):
        value = raw.get(key)
        if isinstance(value, (int, float)):
            result[key] = float(value)
    center = raw.get("center")
    if isinstance(center, (tuple, list)) and len(center) >= 2:
        result["center"] = [float(center[0]), float(center[1])]
    return result


def _load_image_arrays(path: Path) -> list[Any]:
    """返回可 OCR 的图像数组列表（PDF 多页）。"""
    suffix = path.suffix.lower()
    if suffix in _IMAGE_SUFFIXES:
        import numpy as np
        from PIL import Image

        image = Image.open(path).convert("RGB")
        return [np.array(image)]

    if suffix in _PDF_SUFFIXES:
        # 优先 pypdfium2（轻）；否则 pdf2image
        try:
            import numpy as np
            import pypdfium2 as pdfium

            pdf = pdfium.PdfDocument(str(path))
            try:
                pages = []
                max_pages = int(os.environ.get("FHD_EXCEL_ETL_OCR_MAX_PAGES") or 5)
                for i in range(min(len(pdf), max(1, max_pages))):
                    page = pdf[i]
                    try:
                        bitmap = page.render(scale=2).to_pil().convert("RGB")
                        pages.append(np.array(bitmap))
                    finally:
                        page.close()
                return pages
            finally:
                pdf.close()
        except RECOVERABLE_ERRORS:
            logger.debug("pypdfium2 pdf render unavailable", exc_info=True)
        try:
            import numpy as np
            from pdf2image import convert_from_path

            max_pages = int(os.environ.get("FHD_EXCEL_ETL_OCR_MAX_PAGES") or 5)
            images = convert_from_path(str(path), first_page=1, last_page=max_pages)
            return [np.array(img.convert("RGB")) for img in images]
        except RECOVERABLE_ERRORS as exc:
            raise RuntimeError(f"PDF OCR 依赖不可用（需 pypdfium2 或 pdf2image）: {exc}") from exc

    raise RuntimeError(f"不支持的 OCR 源类型: {suffix}")


def _guess_meta_lines(grid: list[list[str]]) -> list[str]:
    """把疑似抬头行抽成单列文本，便于 buyer regex。"""
    meta: list[str] = []
    for row in grid[:4]:
        cells = [str(cell).strip() for cell in row if str(cell).strip()]
        joined = " ".join(cells)
        if not joined:
            continue
        header_tokens = sum(
            bool(
                re.fullmatch(
                    r"(客户(?:名称)?|购货单位|产品(?:名称)?|品名|型号|规格|数量|单价|金额|"
                    r"联系人|电话|地址|日期|单号|订单号|备注)",
                    cell,
                    re.I,
                )
            )
            for cell in cells
        )
        if len(cells) >= 2 and header_tokens >= 2 and not re.search(r"[:：]", joined):
            # A real column-header row is table structure, not document metadata.
            continue
        if (len(cells) == 1 or bool(re.search(r"[:：]", joined))) and re.search(
            r"客户|购货|收货|联系人|日期|单号|订单|buyer|customer|date|order",
            joined,
            re.I,
        ):
            meta.append(joined)
    return meta[:3]


def ocr_source_to_workbook(
    file_path: str | Path,
    *,
    output_path: str | Path | None = None,
) -> dict[str, Any]:
    """扫描件/图片/PDF → 临时 xlsx。"""
    path = Path(file_path).expanduser().resolve()
    if not path.is_file():
        return {"success": False, "message": f"文件不存在: {path}", "error_code": "missing_file"}
    if not is_ocr_source(path):
        return {
            "success": False,
            "message": f"不是 OCR 源（支持图片/PDF）: {path.suffix}",
            "error_code": "not_ocr_source",
        }

    try:
        from app.services.ocr_service import get_ocr_service

        ocr = get_ocr_service()
        images = _load_image_arrays(path)
        all_blocks: list[dict[str, Any]] = []
        pages: list[dict[str, Any]] = []
        for page_number, img in enumerate(images, start=1):
            blocks = ocr.recognize_text_blocks(img) or []
            all_blocks.extend(blocks)
            grid, grid_cells = _text_blocks_to_grid_result(blocks)
            if not grid:
                continue
            meta_lines = _guess_meta_lines(grid)
            body = grid
            body_row_indexes = list(range(1, len(grid) + 1))
            if meta_lines:
                selected = [
                    (index, row)
                    for index, row in enumerate(grid, start=1)
                    if " ".join(str(c).strip() for c in row if str(c).strip()) not in meta_lines
                ]
                if selected:
                    body_row_indexes = [index for index, _row in selected]
                    body = [row for _index, row in selected]
            row_remap = {
                original_index: body_index
                for body_index, original_index in enumerate(body_row_indexes, start=1)
            }
            body_cells = [
                {
                    **cell,
                    "grid_row": row_remap[int(cell["grid_row"])],
                }
                for cell in grid_cells
                if int(cell["grid_row"]) in row_remap
            ]
            pages.append(
                {
                    "page_number": page_number,
                    "grid": body,
                    "grid_cells": body_cells,
                    "meta_lines": meta_lines,
                    "blocks": [_safe_ocr_block(block) for block in blocks],
                    "row_count": len(body),
                    "col_count": max((len(row) for row in body), default=0),
                }
            )
        if not all_blocks:
            return {
                "success": False,
                "message": "OCR 未识别到文本块",
                "error_code": "ocr_empty",
                "backend": getattr(ocr, "get_active_ocr_backend", lambda: "unknown")(),
            }
        if not pages:
            return {
                "success": False,
                "message": "OCR 文本无法聚类成表格",
                "error_code": "ocr_grid_empty",
            }
        xlsx = page_grids_to_workbook_path(
            pages,
            base_sheet_name=path.stem[:24] or "OCR",
        )
        meta_lines = [line for page in pages for line in page.get("meta_lines") or []]
        return {
            "success": True,
            "file_path": str(xlsx),
            "source_path": str(path),
            "block_count": len(all_blocks),
            "row_count": sum(int(page["row_count"]) for page in pages),
            "col_count": max((int(page["col_count"]) for page in pages), default=0),
            "meta_lines": meta_lines,
            "pages": pages,
            "message": f"OCR 已生成表格 {xlsx.name}",
        }
    except RECOVERABLE_ERRORS as exc:
        logger.info("ocr_source_to_workbook failed: %s", exc, exc_info=True)
        return {"success": False, "message": f"OCR 失败: {exc}", "error_code": "ocr_failed"}


def parse_ocr_document(
    file_path: str | Path,
    *,
    include_ledger: bool | str = "auto",
    unit_name_hint: str | None = None,
    profile_id: str | None = None,
    workspace_root: str | Path | None = None,
) -> dict[str, Any]:
    """OCR 源文件 → 解析 notes（内部先转 xlsx）。"""
    from app.application.shipment_excel_etl_app_service import parse_delivery_notes
    from app.application.shipment_excel_etl_security import (
        ShipmentEtlPathError,
        resolve_etl_output_path,
        resolve_etl_path,
    )

    try:
        src = resolve_etl_path(file_path, workspace_root=workspace_root, must_exist=True)
    except ShipmentEtlPathError as exc:
        return {
            "success": False,
            "message": f"非法文件路径: {exc}",
            "error_code": "unsafe_path",
            "notes": [],
        }

    out_dir = Path(tempfile.mkdtemp(prefix="etl_ocr_out_"))
    try:
        out_xlsx = resolve_etl_output_path(
            out_dir / f"{src.stem}_ocr.xlsx",
            workspace_root=workspace_root or out_dir,
        )
    except ShipmentEtlPathError:
        out_xlsx = out_dir / f"{src.stem}_ocr.xlsx"

    ocr_result = ocr_source_to_workbook(src, output_path=out_xlsx)
    if not ocr_result.get("success"):
        return {**ocr_result, "notes": [], "ocr": ocr_result}

    parsed = parse_delivery_notes(
        ocr_result["file_path"],
        include_ledger=include_ledger,
        unit_name_hint=unit_name_hint or src.stem,
        profile_id=profile_id,
        allow_ocr=False,
    )
    return {
        **parsed,
        "ocr": ocr_result,
        "source_kind_input": "ocr",
        "source_path": str(src),
        "message": f"OCR→解析：{parsed.get('message')}",
    }


__all__ = [
    "grid_to_workbook_path",
    "is_ocr_source",
    "ocr_source_to_workbook",
    "parse_ocr_document",
    "text_blocks_to_grid",
]
