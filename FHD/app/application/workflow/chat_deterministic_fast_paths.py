"""Planner 对话确定性快路径：常见计数/读表问题直接查库或读 Excel，避免 LLM 幻觉。"""

from __future__ import annotations

import re
from typing import Any

from app.domain.context.session_context import (
    enrich_excel_tool_arguments,
)
from app.utils.operational_errors import RECOVERABLE_ERRORS

_PRODUCT_COUNT_RE = re.compile(
    r"(?:"
    r"(?:产品表|产品库|products?\s*表?).{0,24}?(?:多少|几条|总数|记录数|count|条记录)"
    r"|"
    r"(?:多少|几条|总数|记录数).{0,24}?(?:产品表|产品库|products?\s*表?)"
    r")",
    re.IGNORECASE,
)

_EXCEL_ROW_COUNT_RE = re.compile(
    r"(?:多少行|几行|行数|总行数|row\s*count|有多少行|第一个\s*sheet|第一个工作表)",
    re.IGNORECASE,
)

_NUMERIC_ONLY_RE = re.compile(
    r"(?:只|仅|就)?(?:回答|回复|说|输出)?(?:一个)?数字",
    re.IGNORECASE,
)


def _wants_numeric_only(message: str) -> bool:
    return bool(_NUMERIC_ONLY_RE.search(str(message or "")))


def _query_product_count() -> int | None:
    try:
        from app.db.models import Product
        from app.db.session import get_db

        with get_db() as db:
            return int(db.query(Product).count())
    except RECOVERABLE_ERRORS:
        return None


def _count_excel_rows_openpyxl(file_path: str, sheet_name: str | None) -> int | None:
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        try:
            names = wb.sheetnames
            if not names:
                return None
            target = sheet_name if sheet_name and sheet_name in names else names[0]
            ws = wb[target]
            return sum(1 for _ in ws.iter_rows())
        finally:
            wb.close()
    except RECOVERABLE_ERRORS:
        return None


def _query_excel_row_count(
    message: str,
    runtime_context: dict[str, Any] | None,
    *,
    workspace_root: str | None,
) -> int | None:
    ea = runtime_context.get("excel_analysis") if isinstance(runtime_context, dict) else None
    if not ea:
        return None
    if not _EXCEL_ROW_COUNT_RE.search(str(message or "")):
        return None
    file_path = str(ea.get("file_path") or "").strip()
    sheet_name = str(ea.get("sheet_name") or "").strip() or None
    if not file_path:
        return None
    try:
        from app.application.tools.workflow import handle_excel_analysis

        args = enrich_excel_tool_arguments(
            "excel_analysis",
            {"action": "read", "sheet_name": sheet_name} if sheet_name else {"action": "read"},
            runtime_context or {},
        )
        file_path = str(args.get("file_path") or file_path).strip()
        result = handle_excel_analysis(args, workspace_root=workspace_root)
        if isinstance(result, dict) and result.get("success"):
            return int(result.get("row_count") or 0)
    except RECOVERABLE_ERRORS:
        pass
    return _count_excel_rows_openpyxl(file_path, sheet_name)


def try_deterministic_chat_reply(
    message: str,
    *,
    runtime_context: dict[str, Any] | None = None,
    workspace_root: str | None = None,
) -> dict[str, str] | None:
    """命中常见计数问题时返回 {response, text, thinking_steps}，否则 None。"""
    text = str(message or "").strip()
    if not text:
        return None

    numeric_only = _wants_numeric_only(text)

    if _PRODUCT_COUNT_RE.search(text):
        count = _query_product_count()
        if count is not None:
            answer = str(count) if numeric_only else f"产品表共有 {count} 条记录。"
            return {
                "response": answer,
                "text": answer,
                "thinking_steps": "[调用工具: db_query/product_count]",
            }

    row_count = _query_excel_row_count(text, runtime_context, workspace_root=workspace_root)
    if row_count is not None:
        answer = str(row_count) if numeric_only else f"第一个工作表共有 {row_count} 行（含表头）。"
        return {
            "response": answer,
            "text": answer,
            "thinking_steps": "[调用工具: excel_analysis/read]",
        }

    return None


__all__ = ["try_deterministic_chat_reply"]
