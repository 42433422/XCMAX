# ruff: noqa
# mypy: ignore-errors
"""Implementation extracted from the public facade module."""
from __future__ import annotations
import importlib

def _facade():
    return importlib.import_module('app.application.workflow.planner')

def _execute_shipment_records_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    try:
        from app.bootstrap import get_shipment_app_service
        unit = params.get('unit_name') or params.get('keyword') or params.get('customer_name')
        limit = int(params.get('limit', 50))
        svc = get_shipment_app_service()
        rows = svc.get_shipment_records(unit_name=str(unit).strip() if unit else None, limit=limit)
        return {'success': True, 'data': rows, 'message': f'共 {len(rows)} 条出货记录'}
    except ImportError as e:
        _facade().logger.error('出货记录服务导入失败: %s', e)
        return {'success': False, 'message': '出货记录服务不可用', 'error_code': 'service_unavailable'}
    except (ValueError, TypeError) as e:
        _facade().logger.warning('出货记录查询参数错误: %s', e)
        return {'success': False, 'message': '查询参数错误，请检查单位名称', 'error_code': 'invalid_parameters'}
    except RuntimeError as e:
        _facade().logger.error('出货记录查询运行时错误: %s', e)
        return {'success': False, 'message': '查询失败，请稍后重试', 'error_code': 'query_failed'}

def _execute_materials_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    try:
        from app.bootstrap import get_materials_service
        search = str(params.get('keyword') or params.get('search') or '').strip() or None
        category = str(params.get('category') or '').strip() or None
        page = int(params.get('page', 1))
        per_page = int(params.get('per_page', 20))
        return get_materials_service().get_all_materials(search=search, category=category, page=page, per_page=per_page)
    except ImportError as e:
        _facade().logger.error('原材料服务导入失败: %s', e)
        return {'success': False, 'message': '原材料服务不可用', 'error_code': 'service_unavailable'}
    except (ValueError, TypeError) as e:
        _facade().logger.warning('原材料查询参数错误: %s', e)
        return {'success': False, 'message': '查询参数错误，请检查输入', 'error_code': 'invalid_parameters'}
    except RuntimeError as e:
        _facade().logger.error('原材料查询运行时错误: %s', e)
        return {'success': False, 'message': '查询失败，请稍后重试', 'error_code': 'query_failed'}

def _execute_print_label_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    try:
        import os
        from app.infrastructure.documents.shipment_document_generator_impl import SimpleLabelGenerator
        from app.utils.path_io.path_utils import get_resource_path
        products = params.get('products')
        if not isinstance(products, list) or not products:
            return {'success': False, 'message': '缺少 products 数组', 'error_code': 'missing_products'}
        labels_dir = get_resource_path('ai_assistant', '商标导出')
        os.makedirs(labels_dir, exist_ok=True)
        order_number = str(params.get('order_number') or params.get('doc_name') or 'LABEL').strip()
        gen = SimpleLabelGenerator(labels_dir)
        labels = gen.generate_labels_for_order(order_number=order_number, products=products)
        return {'success': True, 'data': labels, 'message': f'已生成 {len(labels)} 张标签'}
    except ImportError as e:
        _facade().logger.error('标签生成服务导入失败: %s', e)
        return {'success': False, 'message': '标签生成服务不可用', 'error_code': 'service_unavailable'}
    except (ValueError, TypeError) as e:
        _facade().logger.warning('标签生成参数错误: %s', e)
        return {'success': False, 'message': '标签参数错误，请检查产品数据', 'error_code': 'invalid_parameters'}
    except OSError as e:
        _facade().logger.error('标签文件生成失败: %s', e)
        return {'success': False, 'message': '标签导出失败，请检查磁盘空间', 'error_code': 'file_io_error'}
    except RuntimeError as e:
        _facade().logger.error('标签生成运行时错误: %s', e)
        return {'success': False, 'message': '生成失败，请稍后重试', 'error_code': 'generation_failed'}

def _execute_excel_decompose_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    try:
        from app.bootstrap import get_template_app_service
        file_path = str(params.get('file_path') or '').strip()
        if not file_path:
            return {'success': False, 'message': '缺少 file_path', 'error_code': 'missing_file_path'}
        template_type = params.get('template_type') or params.get('scope')
        return get_template_app_service().decompose_template(file_path, str(template_type).strip() if template_type else None)
    except ImportError as e:
        _facade().logger.error('模板服务导入失败: %s', e)
        return {'success': False, 'message': '模板服务不可用', 'error_code': 'service_unavailable'}
    except (ValueError, TypeError) as e:
        _facade().logger.warning('模板分解参数错误: %s', e)
        return {'success': False, 'message': '模板参数错误，请检查文件', 'error_code': 'invalid_parameters'}
    except OSError as e:
        _facade().logger.error('模板文件读取失败: %s', e)
        return {'success': False, 'message': '文件读取失败，请检查文件是否存在', 'error_code': 'file_not_found'}
    except RuntimeError as e:
        _facade().logger.error('模板分解运行时错误: %s', e)
        return {'success': False, 'message': '分解失败，请稍后重试', 'error_code': 'decomposition_failed'}

def _execute_template_extract_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    """与 excel_decompose 共用模板分解能力。"""
    return _facade()._execute_excel_decompose_tool(params)

def _execute_excel_schema_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    """分析 Excel 文件的表结构。"""
    file_path = str(params.get('file_path') or '').strip()
    if not file_path:
        return {'success': False, 'message': '缺少 file_path 参数', 'error_code': 'missing_file_path'}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get('sheet_name') or wb.sheetnames[0]
        ws = wb[sheet_name]
        fields = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value is not None:
                fields.append({'name': str(cell.column_letter), 'label': str(cell.value).strip(), 'column_index': cell.column})
        row_count = ws.max_row or 0
        wb.close()
        return {'success': True, 'file_path': file_path, 'sheet_name': sheet_name, 'fields': fields, 'row_count': max(0, row_count - 1), 'message': f'Excel 结构分析完成：{len(fields)} 列，{max(0, row_count - 1)} 行数据'}
    except ImportError as e:
        _facade().logger.error('Excel 分析库导入失败: %s', e)
        return {'success': False, 'message': 'Excel 处理库不可用', 'error_code': 'library_unavailable'}
    except (ValueError, TypeError) as e:
        _facade().logger.warning('Excel 结构参数错误: %s', e)
        return {'success': False, 'message': '文件参数错误，请检查 Excel 文件', 'error_code': 'invalid_parameters'}
    except OSError as e:
        _facade().logger.error('Excel 文件读取失败: %s', e)
        return {'success': False, 'message': '文件读取失败，请检查文件是否存在', 'error_code': 'file_not_found'}
    except RuntimeError as e:
        _facade().logger.error('Excel 结构分析运行时错误: %s', e)
        return {'success': False, 'message': '分析失败，请稍后重试', 'error_code': 'analysis_failed'}

def _execute_excel_analysis_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    """读取/查询/聚合 Excel 数据。"""
    file_path = str(params.get('file_path') or '').strip()
    if not file_path:
        return {'success': False, 'message': '缺少 file_path 参数', 'error_code': 'missing_file_path'}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get('sheet_name') or wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else '')
        target_columns = params.get('columns')
        col_indices = list(range(len(headers)))
        if target_columns:
            col_indices = [i for (i, h) in enumerate(headers) if h in target_columns]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 2, 52)):
            row_data = {}
            for i in col_indices:
                if i < len(row):
                    row_data[headers[i]] = row[i].value
            if any((v is not None for v in row_data.values())):
                rows.append(row_data)
        wb.close()
        return {'success': True, 'file_path': file_path, 'sheet_name': sheet_name, 'headers': headers, 'rows': rows[:50], 'total_rows': len(rows), 'message': f'Excel 数据读取完成：{len(headers)} 列，{len(rows)} 行'}
    except ImportError as e:
        _facade().logger.error('Excel 分析库导入失败: %s', e)
        return {'success': False, 'message': 'Excel 处理库不可用', 'error_code': 'library_unavailable'}
    except (ValueError, TypeError) as e:
        _facade().logger.warning('Excel 数据参数错误: %s', e)
        return {'success': False, 'message': '文件参数错误，请检查 Excel 文件', 'error_code': 'invalid_parameters'}
    except OSError as e:
        _facade().logger.error('Excel 文件读取失败: %s', e)
        return {'success': False, 'message': '文件读取失败，请检查文件是否存在', 'error_code': 'file_not_found'}
    except RuntimeError as e:
        _facade().logger.error('Excel 数据分析运行时错误: %s', e)
        return {'success': False, 'message': '分析失败，请稍后重试', 'error_code': 'analysis_failed'}

def _execute_import_excel_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    """将 Excel 数据导入数据库。"""
    file_path = str(params.get('file_path') or '').strip()
    if not file_path:
        return {'success': False, 'message': '缺少 file_path 参数', 'error_code': 'missing_file_path'}
    unit_name = str(params.get('unit_name') or '').strip()
    price_column = str(params.get('price_column') or '').strip()
    create_customer = params.get('create_customer_if_missing', True)
    skip_duplicates = params.get('skip_duplicates', True)
    try:
        from app.bootstrap import get_products_service
        products_service = get_products_service()
    except ImportError as e:
        _facade().logger.error('产品服务导入失败: %s', e)
        return {'success': False, 'message': '产品服务不可用', 'error_code': 'service_unavailable'}
    except RuntimeError as e:
        _facade().logger.error('产品服务初始化失败: %s', e)
        return {'success': False, 'message': '产品服务初始化失败', 'error_code': 'service_init_failed'}
    customer_service = None
    try:
        from app.bootstrap import get_customer_app_service
        customer_service = get_customer_app_service()
    except ImportError:
        _facade().logger.warning('客户服务不可用，降级为仅产品入库')
    except RuntimeError as e:
        _facade().logger.warning('客户服务初始化失败，降级为仅产品入库: %s', e)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get('sheet_name') or wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else '')
        name_col = None
        model_col = None
        price_col = None
        unit_col = None
        for (i, h) in enumerate(headers):
            if not name_col and any((k in h for k in ('产品名称', '名称', '品名'))):
                name_col = i
            if not model_col and any((k in h for k in ('编号', '型号', '产品编号', '规格型号'))):
                model_col = i
            if not unit_col and any((k in h for k in ('单位', '客户', '购买单位'))):
                unit_col = i
        resolved_price_col_name = ''
        if not price_column:
            try:
                from app.application.ai_chat_app_service import AIChatApplicationService
                merged_intent = AIChatApplicationService._merge_user_intent_for_price_resolution(str(params.get('_user_message') or ''), params.get('_request_context'))
                overrides = params.get('excel_import_column_overrides')
                (resolved_price_col_name, price_err) = AIChatApplicationService._resolve_unit_price_column(keys=headers, current='', user_message=merged_intent, overrides=overrides if isinstance(overrides, dict) else {})
                if price_err == 'ambiguous_price_columns':
                    wb.close()
                    return {'success': False, 'message': "检测到「调价前」和「调价后」两列价格，请明确指定使用哪一列（如传入 price_column='调价前含税单价'）", 'error_code': 'ambiguous_price_columns'}
                if resolved_price_col_name:
                    price_column = resolved_price_col_name
                    _facade().logger.info("智能价格列消歧: 选中列 '%s'", price_column)
            except ImportError:
                _facade().logger.debug('AI 服务不可用，回退简单匹配')
            except (ValueError, TypeError) as e:
                _facade().logger.debug('智能价格列消歧参数错误，回退简单匹配: %s', e)
            except RuntimeError as e:
                _facade().logger.warning('智能价格列消歧运行时错误，回退简单匹配: %s', e)
        for (i, h) in enumerate(headers):
            if not price_col:
                if price_column and price_column in h or (not price_column and any((k in h for k in ('单价', '价格', '价')))):
                    price_col = i
        if price_column and price_col is None:
            for (i, h) in enumerate(headers):
                if price_column in h:
                    price_col = i
                    break
        created_units = 0
        created_products = 0
        skipped_products = 0
        touched_units: set = set()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_values = [c.value for c in row]
            product_name = str(row_values[name_col] or '').strip() if name_col is not None and name_col < len(row_values) else ''
            model_number = str(row_values[model_col] or '').strip().upper() if model_col is not None and model_col < len(row_values) else ''
            unit_price = 0.0
            if price_col is not None and price_col < len(row_values):
                try:
                    unit_price = float(str(row_values[price_col] or 0))
                except (ValueError, TypeError):
                    unit_price = 0.0
            row_unit = str(row_values[unit_col] or '').strip() if unit_col is not None and unit_col < len(row_values) else ''
            effective_unit = unit_name or row_unit
            if not effective_unit and (not product_name) and (not model_number):
                continue
            touched_units.add(effective_unit)
            if effective_unit and customer_service is not None and create_customer:
                matched = customer_service.match_purchase_unit(effective_unit)
                if not matched:
                    create_result = customer_service.create({'customer_name': effective_unit})
                    if create_result.get('success'):
                        created_units += 1
            if (product_name or model_number) and products_service is not None:
                exists_result = products_service.get_products(unit_name=effective_unit or None, model_number=model_number or None, keyword=product_name or model_number or None, page=1, per_page=5)
                existed = False
                if exists_result.get('success'):
                    for item in exists_result.get('data') or []:
                        item_model = str(item.get('model_number') or '').strip().upper()
                        item_name = str(item.get('name') or item.get('product_name') or '').strip()
                        if model_number and item_model == model_number:
                            existed = True
                            break
                        if product_name and item_name == product_name:
                            existed = True
                            break
                if existed and skip_duplicates:
                    skipped_products += 1
                    continue
                create_product = products_service.create_product({'name': product_name or model_number, 'product_name': product_name or model_number, 'product_code': model_number or None, 'model_number': model_number or None, 'unit_price': unit_price, 'price': unit_price, 'unit': effective_unit})
                if create_product.get('success'):
                    created_products += 1
        wb.close()
        return {'success': True, 'records': len(touched_units) + created_products + skipped_products, 'touched_units': len(touched_units), 'created_units': created_units, 'created_products': created_products, 'skipped_products': skipped_products, 'price_column_used': headers[price_col] if price_col is not None else '未指定', 'message': f'导入完成：新增客户 {created_units}，新增产品 {created_products}，跳过重复 {skipped_products}'}
    except ImportError as e:
        _facade().logger.error('Excel 处理库导入失败: %s', e)
        return {'success': False, 'message': 'Excel 处理库不可用', 'error_code': 'library_unavailable'}
    except (ValueError, TypeError) as e:
        _facade().logger.warning('Excel 导入参数错误: %s', e)
        return {'success': False, 'message': '导入参数错误，请检查文件格式', 'error_code': 'invalid_parameters'}
    except OSError as e:
        _facade().logger.error('Excel 文件读取失败: %s', e)
        return {'success': False, 'message': '文件读取失败，请检查文件是否存在', 'error_code': 'file_not_found'}
    except RuntimeError as e:
        _facade().logger.error('Excel 导入运行时错误: %s', e)
        return {'success': False, 'message': '导入失败，请检查数据格式后重试', 'error_code': 'import_failed'}

def _execute_employee_list_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    from app.application.facades.tools_facade import execute_registered_workflow_tool
    return execute_registered_workflow_tool('employee', 'list', params)

def _execute_employee_execute_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    from app.application.facades.tools_facade import execute_registered_workflow_tool
    return execute_registered_workflow_tool('employee', 'execute', params)

def _execute_business_db_read_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    from app.application.facades.tools_facade import execute_registered_workflow_tool
    return execute_registered_workflow_tool('business_db', 'read', params)

def _execute_business_db_write_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    from app.application.facades.tools_facade import execute_registered_workflow_tool
    return execute_registered_workflow_tool('business_db', 'write', params)
