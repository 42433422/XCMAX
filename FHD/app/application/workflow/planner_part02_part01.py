# mypy: disable-error-code="no-any-return, valid-type"
"""Implementation extracted from the public facade module."""

from __future__ import annotations

import importlib


def _facade():
    return importlib.import_module("app.application.workflow.planner")


def _execute_shipment_records_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    try:
        from app.bootstrap import get_shipment_app_service

        unit = params.get("unit_name") or params.get("keyword") or params.get("customer_name")
        limit = int(params.get("limit", 50))
        svc = get_shipment_app_service()
        rows = svc.get_shipment_records(unit_name=str(unit).strip() if unit else None, limit=limit)
        return {"success": True, "data": rows, "message": f"共 {len(rows)} 条出货记录"}
    except ImportError as e:
        _facade().logger.error("出货记录服务导入失败: %s", e)
        return {
            "success": False,
            "message": "出货记录服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        _facade().logger.warning("出货记录查询参数错误: %s", e)
        return {
            "success": False,
            "message": "查询参数错误，请检查单位名称",
            "error_code": "invalid_parameters",
        }
    except RuntimeError as e:
        _facade().logger.error("出货记录查询运行时错误: %s", e)
        return {"success": False, "message": "查询失败，请稍后重试", "error_code": "query_failed"}


def _execute_materials_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    try:
        from app.bootstrap import get_materials_service

        search = str(params.get("keyword") or params.get("search") or "").strip() or None
        category = str(params.get("category") or "").strip() or None
        page = int(params.get("page", 1))
        per_page = int(params.get("per_page", 20))
        return get_materials_service().get_all_materials(
            search=search, category=category, page=page, per_page=per_page
        )
    except ImportError as e:
        _facade().logger.error("原材料服务导入失败: %s", e)
        return {
            "success": False,
            "message": "原材料服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        _facade().logger.warning("原材料查询参数错误: %s", e)
        return {
            "success": False,
            "message": "查询参数错误，请检查输入",
            "error_code": "invalid_parameters",
        }
    except RuntimeError as e:
        _facade().logger.error("原材料查询运行时错误: %s", e)
        return {"success": False, "message": "查询失败，请稍后重试", "error_code": "query_failed"}


def _execute_print_label_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    try:
        import os

        from app.infrastructure.documents.shipment_document_generator_impl import (
            SimpleLabelGenerator,
        )
        from app.utils.path_io.path_utils import get_resource_path

        products = params.get("products")
        if not isinstance(products, list) or not products:
            return {
                "success": False,
                "message": "缺少 products 数组",
                "error_code": "missing_products",
            }
        labels_dir = get_resource_path("ai_assistant", "商标导出")
        os.makedirs(labels_dir, exist_ok=True)
        order_number = str(params.get("order_number") or params.get("doc_name") or "LABEL").strip()
        gen = SimpleLabelGenerator(labels_dir)
        labels = gen.generate_labels_for_order(order_number=order_number, products=products)
        return {"success": True, "data": labels, "message": f"已生成 {len(labels)} 张标签"}
    except ImportError as e:
        _facade().logger.error("标签生成服务导入失败: %s", e)
        return {
            "success": False,
            "message": "标签生成服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        _facade().logger.warning("标签生成参数错误: %s", e)
        return {
            "success": False,
            "message": "标签参数错误，请检查产品数据",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        _facade().logger.error("标签文件生成失败: %s", e)
        return {
            "success": False,
            "message": "标签导出失败，请检查磁盘空间",
            "error_code": "file_io_error",
        }
    except RuntimeError as e:
        _facade().logger.error("标签生成运行时错误: %s", e)
        return {
            "success": False,
            "message": "生成失败，请稍后重试",
            "error_code": "generation_failed",
        }


def _execute_excel_decompose_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    try:
        from app.bootstrap import get_template_app_service

        file_path = str(params.get("file_path") or "").strip()
        if not file_path:
            return {
                "success": False,
                "message": "缺少 file_path",
                "error_code": "missing_file_path",
            }
        template_type = params.get("template_type") or params.get("scope")
        return get_template_app_service().decompose_template(
            file_path, str(template_type).strip() if template_type else None
        )
    except ImportError as e:
        _facade().logger.error("模板服务导入失败: %s", e)
        return {"success": False, "message": "模板服务不可用", "error_code": "service_unavailable"}
    except (ValueError, TypeError) as e:
        _facade().logger.warning("模板分解参数错误: %s", e)
        return {
            "success": False,
            "message": "模板参数错误，请检查文件",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        _facade().logger.error("模板文件读取失败: %s", e)
        return {
            "success": False,
            "message": "文件读取失败，请检查文件是否存在",
            "error_code": "file_not_found",
        }
    except RuntimeError as e:
        _facade().logger.error("模板分解运行时错误: %s", e)
        return {
            "success": False,
            "message": "分解失败，请稍后重试",
            "error_code": "decomposition_failed",
        }


def _execute_template_extract_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    """与 excel_decompose 共用模板分解能力。"""
    return _facade()._execute_excel_decompose_tool(params)


def _execute_excel_schema_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    """分析 Excel 文件的表结构。"""
    file_path = str(params.get("file_path") or "").strip()
    if not file_path:
        return {
            "success": False,
            "message": "缺少 file_path 参数",
            "error_code": "missing_file_path",
        }
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get("sheet_name") or wb.sheetnames[0]
        ws = wb[sheet_name]
        fields = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value is not None:
                fields.append(
                    {
                        "name": str(cell.column_letter),
                        "label": str(cell.value).strip(),
                        "column_index": cell.column,
                    }
                )
        row_count = ws.max_row or 0
        wb.close()
        return {
            "success": True,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "fields": fields,
            "row_count": max(0, row_count - 1),
            "message": f"Excel 结构分析完成：{len(fields)} 列，{max(0, row_count - 1)} 行数据",
        }
    except ImportError as e:
        _facade().logger.error("Excel 分析库导入失败: %s", e)
        return {
            "success": False,
            "message": "Excel 处理库不可用",
            "error_code": "library_unavailable",
        }
    except (ValueError, TypeError) as e:
        _facade().logger.warning("Excel 结构参数错误: %s", e)
        return {
            "success": False,
            "message": "文件参数错误，请检查 Excel 文件",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        _facade().logger.error("Excel 文件读取失败: %s", e)
        return {
            "success": False,
            "message": "文件读取失败，请检查文件是否存在",
            "error_code": "file_not_found",
        }
    except RuntimeError as e:
        _facade().logger.error("Excel 结构分析运行时错误: %s", e)
        return {
            "success": False,
            "message": "分析失败，请稍后重试",
            "error_code": "analysis_failed",
        }


def _execute_excel_analysis_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    """读取/查询/聚合 Excel 数据。"""
    file_path = str(params.get("file_path") or "").strip()
    if not file_path:
        return {
            "success": False,
            "message": "缺少 file_path 参数",
            "error_code": "missing_file_path",
        }
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get("sheet_name") or wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else "")
        target_columns = params.get("columns")
        col_indices = list(range(len(headers)))
        if target_columns:
            col_indices = [i for i, h in enumerate(headers) if h in target_columns]
        rows = []
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 2, 52)):
            row_data = {}
            for i in col_indices:
                if i < len(row):
                    row_data[headers[i]] = row[i].value
            if any(v is not None for v in row_data.values()):
                rows.append(row_data)
        wb.close()
        return {
            "success": True,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": rows[:50],
            "total_rows": len(rows),
            "message": f"Excel 数据读取完成：{len(headers)} 列，{len(rows)} 行",
        }
    except ImportError as e:
        _facade().logger.error("Excel 分析库导入失败: %s", e)
        return {
            "success": False,
            "message": "Excel 处理库不可用",
            "error_code": "library_unavailable",
        }
    except (ValueError, TypeError) as e:
        _facade().logger.warning("Excel 数据参数错误: %s", e)
        return {
            "success": False,
            "message": "文件参数错误，请检查 Excel 文件",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        _facade().logger.error("Excel 文件读取失败: %s", e)
        return {
            "success": False,
            "message": "文件读取失败，请检查文件是否存在",
            "error_code": "file_not_found",
        }
    except RuntimeError as e:
        _facade().logger.error("Excel 数据分析运行时错误: %s", e)
        return {
            "success": False,
            "message": "分析失败，请稍后重试",
            "error_code": "analysis_failed",
        }
