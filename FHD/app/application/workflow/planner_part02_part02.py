# mypy: disable-error-code="valid-type"
"""Implementation extracted from the public facade module."""

from __future__ import annotations

import importlib


def _facade():
    return importlib.import_module("app.application.workflow.planner")


def _execute_import_excel_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    """将 Excel 数据导入数据库。"""
    file_path = str(params.get("file_path") or "").strip()
    if not file_path:
        return {
            "success": False,
            "message": "缺少 file_path 参数",
            "error_code": "missing_file_path",
        }
    unit_name = str(params.get("unit_name") or "").strip()
    price_column = str(params.get("price_column") or "").strip()
    create_customer = params.get("create_customer_if_missing", True)
    skip_duplicates = params.get("skip_duplicates", True)
    try:
        from app.bootstrap import get_products_service

        products_service = get_products_service()
    except ImportError as e:
        _facade().logger.error("产品服务导入失败: %s", e)
        return {"success": False, "message": "产品服务不可用", "error_code": "service_unavailable"}
    except RuntimeError as e:
        _facade().logger.error("产品服务初始化失败: %s", e)
        return {
            "success": False,
            "message": "产品服务初始化失败",
            "error_code": "service_init_failed",
        }
    customer_service = None
    try:
        from app.bootstrap import get_customer_app_service

        customer_service = get_customer_app_service()
    except ImportError:
        _facade().logger.warning("客户服务不可用，降级为仅产品入库")
    except RuntimeError as e:
        _facade().logger.warning("客户服务初始化失败，降级为仅产品入库: %s", e)
    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get("sheet_name") or wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else "")
        name_col = None
        model_col = None
        price_col = None
        unit_col = None
        for i, h in enumerate(headers):
            if not name_col and any(k in h for k in ("产品名称", "名称", "品名")):
                name_col = i
            if not model_col and any(k in h for k in ("编号", "型号", "产品编号", "规格型号")):
                model_col = i
            if not unit_col and any(k in h for k in ("单位", "客户", "购买单位")):
                unit_col = i
        resolved_price_col_name = ""
        if not price_column:
            try:
                from app.application.ai_chat_app_service import AIChatApplicationService

                merged_intent = AIChatApplicationService._merge_user_intent_for_price_resolution(
                    str(params.get("_user_message") or ""), params.get("_request_context")
                )
                overrides = params.get("excel_import_column_overrides")
                resolved_price_col_name, price_err = (
                    AIChatApplicationService._resolve_unit_price_column(
                        keys=headers,
                        current="",
                        user_message=merged_intent,
                        overrides=overrides if isinstance(overrides, dict) else {},
                    )
                )
                if price_err == "ambiguous_price_columns":
                    wb.close()
                    return {
                        "success": False,
                        "message": "检测到「调价前」和「调价后」两列价格，请明确指定使用哪一列（如传入 price_column='调价前含税单价'）",
                        "error_code": "ambiguous_price_columns",
                    }
                if resolved_price_col_name:
                    price_column = resolved_price_col_name
                    _facade().logger.info("智能价格列消歧: 选中列 '%s'", price_column)
            except ImportError:
                _facade().logger.debug("AI 服务不可用，回退简单匹配")
            except (ValueError, TypeError) as e:
                _facade().logger.debug("智能价格列消歧参数错误，回退简单匹配: %s", e)
            except RuntimeError as e:
                _facade().logger.warning("智能价格列消歧运行时错误，回退简单匹配: %s", e)
        for i, h in enumerate(headers):
            if not price_col:
                if (
                    price_column
                    and price_column in h
                    or (not price_column and any(k in h for k in ("单价", "价格", "价")))
                ):
                    price_col = i
        if price_column and price_col is None:
            for i, h in enumerate(headers):
                if price_column in h:
                    price_col = i
                    break
        created_units = 0
        created_products = 0
        skipped_products = 0
        touched_units: set = set()
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_values = [c.value for c in row]
            product_name = (
                str(row_values[name_col] or "").strip()
                if name_col is not None and name_col < len(row_values)
                else ""
            )
            model_number = (
                str(row_values[model_col] or "").strip().upper()
                if model_col is not None and model_col < len(row_values)
                else ""
            )
            unit_price = 0.0
            if price_col is not None and price_col < len(row_values):
                try:
                    unit_price = float(str(row_values[price_col] or 0))
                except (ValueError, TypeError):
                    unit_price = 0.0
            row_unit = (
                str(row_values[unit_col] or "").strip()
                if unit_col is not None and unit_col < len(row_values)
                else ""
            )
            effective_unit = unit_name or row_unit
            if not effective_unit and (not product_name) and (not model_number):
                continue
            touched_units.add(effective_unit)
            if effective_unit and customer_service is not None and create_customer:
                matched = customer_service.match_purchase_unit(effective_unit)
                if not matched:
                    create_result = customer_service.create({"customer_name": effective_unit})
                    if create_result.get("success"):
                        created_units += 1
            if (product_name or model_number) and products_service is not None:
                exists_result = products_service.get_products(
                    unit_name=effective_unit or None,
                    model_number=model_number or None,
                    keyword=product_name or model_number or None,
                    page=1,
                    per_page=5,
                )
                existed = False
                if exists_result.get("success"):
                    for item in exists_result.get("data") or []:
                        item_model = str(item.get("model_number") or "").strip().upper()
                        item_name = str(item.get("name") or item.get("product_name") or "").strip()
                        if model_number and item_model == model_number:
                            existed = True
                            break
                        if product_name and item_name == product_name:
                            existed = True
                            break
                if existed and skip_duplicates:
                    skipped_products += 1
                    continue
                create_product = products_service.create_product(
                    {
                        "name": product_name or model_number,
                        "product_name": product_name or model_number,
                        "product_code": model_number or None,
                        "model_number": model_number or None,
                        "unit_price": unit_price,
                        "price": unit_price,
                        "unit": effective_unit,
                    }
                )
                if create_product.get("success"):
                    created_products += 1
        wb.close()
        return {
            "success": True,
            "records": len(touched_units) + created_products + skipped_products,
            "touched_units": len(touched_units),
            "created_units": created_units,
            "created_products": created_products,
            "skipped_products": skipped_products,
            "price_column_used": headers[price_col] if price_col is not None else "未指定",
            "message": f"导入完成：新增客户 {created_units}，新增产品 {created_products}，跳过重复 {skipped_products}",
        }
    except ImportError as e:
        _facade().logger.error("Excel 处理库导入失败: %s", e)
        return {
            "success": False,
            "message": "Excel 处理库不可用",
            "error_code": "library_unavailable",
        }
    except (ValueError, TypeError) as e:
        _facade().logger.warning("Excel 导入参数错误: %s", e)
        return {
            "success": False,
            "message": "导入参数错误，请检查文件格式",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        _facade().logger.error("Excel 文件读取失败: %s", e)
        return {
            "success": False,
            "message": "文件读取失败，请检查文件是否存在",
            "error_code": "file_not_found",
        }
    except RuntimeError as e:
        _facade().logger.error("Excel 导入运行时错误: %s", e)
        return {
            "success": False,
            "message": "导入失败，请检查数据格式后重试",
            "error_code": "import_failed",
        }


def _execute_employee_list_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    from app.application.facades.tools_facade import execute_registered_workflow_tool

    return execute_registered_workflow_tool("employee", "list", params)


def _execute_employee_execute_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    from app.application.facades.tools_facade import execute_registered_workflow_tool

    return execute_registered_workflow_tool("employee", "execute", params)


def _execute_business_db_read_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    from app.application.facades.tools_facade import execute_registered_workflow_tool

    return execute_registered_workflow_tool("business_db", "read", params)


def _execute_business_db_write_tool(params: dict[str, _facade().Any]) -> dict[str, _facade().Any]:
    from app.application.facades.tools_facade import execute_registered_workflow_tool

    return execute_registered_workflow_tool("business_db", "write", params)
