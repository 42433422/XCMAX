"""Migrated from legacy_static.py (v10)."""

from __future__ import annotations

import base64
import logging
import os
import shutil

from fastapi import APIRouter, Body, File, Form, Query, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response, StreamingResponse

from app.traditional_mode_fs import (
    ROOT_DIR,
    copy_response,
    list_files_response,
    move_response,
    read_file_response,
    resolve_safe_path,
    root_info_response,
    sse_watch_events,
    stat_response,
    write_base64_response,
    write_text_response,
)
from app.utils.operational_errors import RECOVERABLE_ERRORS
from app.utils.path_utils import get_base_dir
from app.utils.secure_filename import secure_filename

logger = logging.getLogger(__name__)

router = APIRouter(tags=["legacy-static"], deprecated=True)


@router.get("/")
def gap_batch1_index():
    base_dir = get_base_dir()
    vue_index = os.path.join(base_dir, "templates", "vue-dist", "index.html")
    if os.path.exists(vue_index):
        return FileResponse(vue_index, media_type="text/html")
    legacy = os.path.join(base_dir, "templates", "ai_assistant_console.html")
    if os.path.exists(legacy):
        return FileResponse(legacy, media_type="text/html")
    return JSONResponse({"success": False, "message": "前端模板未找到"}, status_code=404)


def _vue_dist_dir() -> str:
    return os.path.join(get_base_dir(), "templates", "vue-dist")


@router.get("/static/{path:path}")
def gap_batch2_serve_static(path: str):
    vue_dist_dir = _vue_dist_dir()
    static_dir = os.path.join(vue_dist_dir, "static")
    static_path = os.path.join(static_dir, path)
    if os.path.exists(static_path) and not os.path.isdir(static_path):
        return FileResponse(static_path)
    return JSONResponse({"success": False, "message": f"静态资源不存在：{path}"}, status_code=404)


@router.get("/vite.svg")
def gap_batch2_vite_svg():
    p = os.path.join(_vue_dist_dir(), "vite.svg")
    if os.path.exists(p):
        return FileResponse(p, media_type="image/svg+xml")
    return JSONResponse({"success": False, "message": "vite.svg 不存在"}, status_code=404)


@router.get("/brand-xc-logo.jpg")
def gap_batch2_brand_xc_logo():
    p = os.path.join(_vue_dist_dir(), "brand-xc-logo.jpg")
    if os.path.exists(p):
        return FileResponse(p, media_type="image/jpeg")
    return JSONResponse({"success": False, "message": "brand-xc-logo.jpg 不存在"}, status_code=404)


@router.get("/brand-xc-logo.png")
def gap_batch2_brand_xc_logo_png():
    p = os.path.join(_vue_dist_dir(), "brand-xc-logo.png")
    if os.path.exists(p):
        return FileResponse(p, media_type="image/png")
    return JSONResponse({"success": False, "message": "brand-xc-logo.png 不存在"}, status_code=404)


@router.get("/workflow-employee-docs.json")
def gap_batch2_workflow_employee_docs_json():
    p = os.path.join(_vue_dist_dir(), "workflow-employee-docs.json")
    if os.path.exists(p):
        return FileResponse(p, media_type="application/json")
    return JSONResponse(
        {"success": False, "message": "workflow-employee-docs.json 不存在"}, status_code=404
    )


@router.get("/sw.js")
def gap_batch2_service_worker_js():
    """须在 SPA fallback 之前注册；否则 /sw.js 会被当成 index.html（MIME text/html）。"""
    p = os.path.join(_vue_dist_dir(), "sw.js")
    if os.path.exists(p):
        return FileResponse(p, media_type="application/javascript")
    return JSONResponse({"success": False, "message": "sw.js 不存在"}, status_code=404)


def _workflow_employees_json_paths() -> list[str]:
    base = get_base_dir()
    return [
        os.path.join(_vue_dist_dir(), "workflow-employees.json"),
        os.path.join(base, "frontend", "public", "workflow-employees.json"),
        os.path.join(base, "frontend", "src", "data", "workflow-employees.json"),
    ]


@router.get("/workflow-employees.json")
def gap_batch2_workflow_employees_json():
    for p in _workflow_employees_json_paths():
        if os.path.isfile(p):
            return FileResponse(p, media_type="application/json")
    return JSONResponse(
        {"success": False, "message": "workflow-employees.json 不存在"}, status_code=404
    )


@router.get("/favicon.ico")
def gap_batch2_favicon():
    gif = base64.b64decode("R0lGODlhAQABAIAAAAAAAP///yH5BAEAAAAALAAAAAABAAEAAAIBRAA7")
    return Response(content=gif, media_type="image/gif")


@router.get("/outputs/{filename:path}")
def gap_batch2_outputs(filename: str):
    try:
        from app.utils.path_utils import get_app_data_dir, get_resource_path

        shipment_outputs_dir = os.path.join(get_app_data_dir(), "shipment_outputs")
        if os.path.isdir(shipment_outputs_dir):
            outputs_dir = shipment_outputs_dir
        else:
            outputs_dir = get_resource_path("ai_assistant", "outputs")
            if not os.path.isdir(outputs_dir):
                outputs_dir = os.path.join(get_base_dir(), "AI助手", "outputs")
        if not os.path.isdir(outputs_dir):
            return JSONResponse(
                {"success": False, "message": f"输出目录不存在: {outputs_dir}"}, status_code=404
            )
        file_path = os.path.join(outputs_dir, filename)
        if not os.path.exists(file_path):
            return JSONResponse(
                {"success": False, "message": f"文件不存在：{filename}"}, status_code=404
            )
        return FileResponse(file_path, filename=os.path.basename(filename))
    except RECOVERABLE_ERRORS as e:
        return JSONResponse({"success": False, "message": f"下载失败：{str(e)}"}, status_code=500)


@router.get("/test-buttons")
def gap_batch2_test_buttons():
    p = os.path.join(get_base_dir(), "templates", "test-buttons.html")
    if os.path.exists(p):
        return FileResponse(p, media_type="text/html")
    return JSONResponse({"success": False, "message": "test-buttons.html 未找到"}, status_code=404)


@router.get("/products-test")
def gap_batch2_products_test():
    p = os.path.join(get_base_dir(), "templates", "products_test.html")
    if os.path.exists(p):
        return FileResponse(p, media_type="text/html")
    return JSONResponse({"success": False, "message": "products_test.html 未找到"}, status_code=404)


@router.get("/console")
def gap_batch2_console():
    vue_index = os.path.join(_vue_dist_dir(), "index.html")
    if os.path.exists(vue_index):
        return FileResponse(vue_index, media_type="text/html")
    legacy = os.path.join(get_base_dir(), "templates", "ai_assistant_console.html")
    if os.path.exists(legacy):
        return FileResponse(legacy, media_type="text/html")
    return JSONResponse({"success": False, "message": "前端模板未找到"}, status_code=404)


@router.get("/api/traditional-mode/list")
def traditional_list(path: str = Query(default="")):
    payload, code = list_files_response(path)
    return JSONResponse(payload, status_code=code)


@router.get("/api/traditional-mode/root")
def traditional_root():
    return root_info_response()


@router.get("/api/traditional-mode/read")
def traditional_read(file: str = Query(default="")):
    payload, code = read_file_response(file)
    return JSONResponse(payload, status_code=code)


@router.get("/api/traditional-mode/watch")
def traditional_watch(path: str = Query(default="")):
    return StreamingResponse(sse_watch_events(path), media_type="text/event-stream")


@router.get("/api/traditional-mode/agent/stat")
def traditional_agent_stat(path: str = Query(default="")):
    payload, code = stat_response(path)
    return JSONResponse(payload, status_code=code)


@router.post("/api/traditional-mode/agent/write-text")
def traditional_agent_write_text(body: dict = Body(default_factory=dict)):
    data = body or {}
    payload, code = write_text_response(
        str(data.get("file") or data.get("path") or ""),
        str(data.get("content") or ""),
        append=bool(data.get("append", False)),
    )
    return JSONResponse(payload, status_code=code)


@router.post("/api/traditional-mode/agent/write-base64")
def traditional_agent_write_base64(body: dict = Body(default_factory=dict)):
    data = body or {}
    payload, code = write_base64_response(
        str(data.get("file") or data.get("path") or ""),
        str(data.get("content_base64") or data.get("content") or ""),
    )
    return JSONResponse(payload, status_code=code)


@router.post("/api/traditional-mode/agent/move")
def traditional_agent_move(body: dict = Body(default_factory=dict)):
    data = body or {}
    payload, code = move_response(
        str(data.get("src") or data.get("source") or ""),
        str(data.get("dst") or data.get("target") or ""),
        overwrite=bool(data.get("overwrite", False)),
    )
    return JSONResponse(payload, status_code=code)


@router.post("/api/traditional-mode/agent/copy")
def traditional_agent_copy(body: dict = Body(default_factory=dict)):
    data = body or {}
    payload, code = copy_response(
        str(data.get("src") or data.get("source") or ""),
        str(data.get("dst") or data.get("target") or ""),
        overwrite=bool(data.get("overwrite", False)),
    )
    return JSONResponse(payload, status_code=code)


@router.post("/api/traditional-mode/write")
def traditional_mode_write(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "error": "请求体为空或格式错误"}, status_code=400)
    rel_file = data.get("file", "")
    file_data = data.get("data", {})
    file_type = data.get("type", "")
    full_path = resolve_safe_path(rel_file)
    if full_path is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if file_type != "excel":
        return JSONResponse(
            {"success": False, "error": f"不支持的写入类型: {file_type}"}, status_code=400
        )
    try:
        import openpyxl
    except ImportError:
        return JSONResponse(
            {"success": False, "error": "openpyxl 未安装，无法写入 Excel 文件"}, status_code=500
        )
    parent_dir = os.path.dirname(full_path)
    if parent_dir and not os.path.exists(parent_dir):
        os.makedirs(parent_dir, exist_ok=True)
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    default_sheet.title = file_data.get("active_sheet", "Sheet")
    sheets_content = file_data.get("content", {})
    if isinstance(sheets_content, dict):
        for sheet_name, sheet_data_item in sheets_content.items():
            if sheet_name == default_sheet.title:
                ws = default_sheet
            else:
                ws = wb.create_sheet(title=sheet_name)
            rows = sheet_data_item.get("rows", []) if isinstance(sheet_data_item, dict) else []
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, cell_value in enumerate(row, start=1):
                    if cell_value is not None:
                        ws.cell(row=r_idx, column=c_idx, value=cell_value)
    if len(wb.sheetnames) > 1 and default_sheet.title in wb.sheetnames:
        if not sheets_content or default_sheet.title not in sheets_content:
            wb.remove(default_sheet)
    wb.save(full_path)
    wb.close()
    return {"success": True}


@router.post("/api/traditional-mode/mkdir")
def traditional_mode_mkdir(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "error": "请求体为空或格式错误"}, status_code=400)
    rel_path = data.get("path", "")
    folder_name = (data.get("name") or "").strip()
    if not folder_name:
        return JSONResponse({"success": False, "error": "文件夹名称不能为空"}, status_code=400)
    if "/" in folder_name or "\\" in folder_name or ".." in folder_name:
        return JSONResponse({"success": False, "error": "文件夹名称包含非法字符"}, status_code=400)
    full_parent = resolve_safe_path(rel_path)
    if full_parent is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    full_new_path = os.path.join(full_parent, folder_name)
    if not os.path.abspath(full_new_path).startswith(os.path.abspath(ROOT_DIR)):
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if os.path.exists(full_new_path):
        return JSONResponse({"success": False, "error": "文件夹已存在"}, status_code=409)
    os.makedirs(full_new_path, exist_ok=False)
    return {"success": True}


@router.post("/api/traditional-mode/rename")
def traditional_mode_rename(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "error": "请求体为空或格式错误"}, status_code=400)
    rel_path = data.get("path", "")
    old_name = (data.get("old_name") or "").strip()
    new_name = (data.get("new_name") or "").strip()
    if not old_name or not new_name:
        return JSONResponse({"success": False, "error": "旧名称和新名称不能为空"}, status_code=400)
    if "/" in new_name or "\\" in new_name or ".." in new_name:
        return JSONResponse({"success": False, "error": "新名称包含非法字符"}, status_code=400)
    full_parent = resolve_safe_path(rel_path)
    if full_parent is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    full_old_path = os.path.join(full_parent, old_name)
    full_new_path = os.path.join(full_parent, new_name)
    if not os.path.abspath(full_old_path).startswith(os.path.abspath(ROOT_DIR)):
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if not os.path.abspath(full_new_path).startswith(os.path.abspath(ROOT_DIR)):
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if not os.path.exists(full_old_path):
        return JSONResponse({"success": False, "error": "源文件或文件夹不存在"}, status_code=404)
    if os.path.exists(full_new_path):
        return JSONResponse({"success": False, "error": "目标名称已存在"}, status_code=409)
    os.rename(full_old_path, full_new_path)
    return {"success": True}


@router.post("/api/traditional-mode/delete")
def traditional_mode_delete(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "error": "请求体为空或格式错误"}, status_code=400)
    rel_path = data.get("path", "")
    name = (data.get("name") or "").strip()
    if not name:
        return JSONResponse({"success": False, "error": "名称不能为空"}, status_code=400)
    full_parent = resolve_safe_path(rel_path)
    if full_parent is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    full_target = os.path.join(full_parent, name)
    if not os.path.abspath(full_target).startswith(os.path.abspath(ROOT_DIR)):
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if not os.path.exists(full_target):
        return JSONResponse({"success": False, "error": "文件或文件夹不存在"}, status_code=404)
    if os.path.isdir(full_target):
        shutil.rmtree(full_target)
    else:
        os.remove(full_target)
    return {"success": True}


@router.post("/api/traditional-mode/upload")
async def traditional_mode_upload(
    file: UploadFile = File(...),
    path: str = Form(default=""),
):
    if not file.filename:
        return JSONResponse({"success": False, "error": "文件名为空"}, status_code=400)
    full_target_dir = resolve_safe_path(path)
    if full_target_dir is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    filename = secure_filename(file.filename)
    if not filename:
        filename = "uploaded_file"
    if not os.path.exists(full_target_dir):
        os.makedirs(full_target_dir, exist_ok=True)
    save_path = os.path.join(full_target_dir, filename)
    if os.path.exists(save_path):
        base, ext = os.path.splitext(filename)
        counter = 1
        while os.path.exists(save_path):
            filename = f"{base}_{counter}{ext}"
            save_path = os.path.join(full_target_dir, filename)
            counter += 1
    content = await file.read()
    with open(save_path, "wb") as f:
        f.write(content)
    return {"success": True, "filename": filename, "path": save_path}


@router.get("/api/customers/import")
def customers_import_stub():
    return {"success": True, "message": "购买单位导入接口，请使用 POST 上传 .xlsx 文件"}


@router.delete("/api/customers/batch-delete")
def customers_batch_delete_delete(
    ids: str | None = Query(default=None),
    force: str = Query(default="false"),
    body: dict | None = Body(default=None),
):
    from app.application import get_customer_app_service

    try:
        if isinstance(body, dict) and body.get("ids"):
            id_list = body.get("ids") or []
            force_b = bool(body.get("force", False))
        elif ids:
            id_list = [int(x.strip()) for x in ids.split(",") if x.strip()]
            force_b = force.lower() in ("true", "1", "yes")
        else:
            return JSONResponse({"success": False, "message": "ID 列表不能为空"}, status_code=400)
        if not id_list:
            return JSONResponse({"success": False, "message": "ID 列表不能为空"}, status_code=400)
        result = get_customer_app_service().batch_delete(id_list, force=force_b)
        code = 200 if result["success"] else (409 if result.get("has_associations") else 400)
        return JSONResponse(result, status_code=code)
    except ValueError as e:
        return JSONResponse(
            {"success": False, "message": f"ID 格式错误：{str(e)}"}, status_code=400
        )
    except RECOVERABLE_ERRORS as e:
        return JSONResponse({"success": False, "message": f"删除失败：{str(e)}"}, status_code=500)


@router.delete("/api/preferences/{key}")
def preferences_delete_key(key: str, user_id: str = Query(default="default")):
    try:
        from app.application.facades.conversation_facade import get_user_preference_service

        success = get_user_preference_service().delete_preference(user_id, key)
        return {"success": success, "message": "偏好已删除" if success else "删除失败"}
    except RECOVERABLE_ERRORS as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)
