"""
XCAGI 前端兼容 API — 模板 / Excel 网格提取路由。
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import uuid
from pathlib import Path

from fastapi import APIRouter, File, Form, HTTPException, UploadFile

from app.utils.operational_errors import RECOVERABLE_ERRORS

router = APIRouter(tags=["xcagi-compat"])
logger = logging.getLogger(__name__)


def _excel_cell_to_text(v) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _form_bool(v: str | None) -> bool:
    if v is None:
        return False
    return str(v).strip().lower() in ("1", "true", "yes", "on")


def _pick_header_row(values_rows: list[list[str]]) -> tuple[int, list[str]]:
    best_idx = 0
    best_non_empty = -1
    for i, row in enumerate(values_rows[:10]):
        non_empty = sum(1 for c in row if str(c or "").strip())
        if non_empty > best_non_empty:
            best_non_empty = non_empty
            best_idx = i
    header = values_rows[best_idx] if values_rows else []
    if sum(1 for c in header if str(c or "").strip()) < 2 and values_rows:
        best_idx = 0
        header = values_rows[0]
    return best_idx, header


def _detect_effective_col_count(values_rows: list[list[str]], fallback_cols: int) -> int:
    max_used = 0
    for row in values_rows:
        last_non_empty = 0
        for idx, cell in enumerate(row, start=1):
            if str(cell or "").strip():
                last_non_empty = idx
        if last_non_empty > max_used:
            max_used = last_non_empty
    if max_used > 0:
        return max_used
    return max(1, int(fallback_cols or 1))


def _detect_effective_row_count(values_rows: list[list[str]], fallback_rows: int) -> int:
    max_used = 0
    for row_idx, row in enumerate(values_rows, start=1):
        if any(str(c or "").strip() for c in row):
            max_used = row_idx
    if max_used > 0:
        return max_used
    return max(1, int(fallback_rows or 1))


def _excel_col_width_to_px(width: float | int | None) -> int:
    w = float(width or 8.43)
    return max(40, int(w * 7 + 5))


def _excel_row_height_to_px(height: float | int | None) -> int:
    h = float(height or 15.0)
    return max(20, int(h * 96.0 / 72.0))


def _merge_anchor_and_skip(
    ws, row_count: int, col_count: int
) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    merge_anchor: dict[tuple[int, int], tuple[int, int]] = {}
    merge_skip: set[tuple[int, int]] = set()
    for rg in ws.merged_cells.ranges:
        min_r, min_c, max_r, max_c = rg.min_row, rg.min_col, rg.max_row, rg.max_col
        if min_r > row_count or min_c > col_count:
            continue
        rowspan = max(1, min(max_r, row_count) - min_r + 1)
        colspan = max(1, min(max_c, col_count) - min_c + 1)
        merge_anchor[(min_r, min_c)] = (rowspan, colspan)
        for rr in range(min_r, min(max_r, row_count) + 1):
            for cc in range(min_c, min(max_c, col_count) + 1):
                if rr == min_r and cc == min_c:
                    continue
                merge_skip.add((rr, cc))
    return merge_anchor, merge_skip


def _serialize_cell_style(cell) -> dict:
    d: dict = {}
    try:
        font = cell.font
        if font:
            fd: dict = {}
            nm = getattr(font, "name", None)
            if nm:
                fd["name"] = str(nm)
            if getattr(font, "sz", None) is not None:
                try:
                    fd["size"] = float(font.sz)
                except (TypeError, ValueError):
                    pass
            if getattr(font, "b", None) is not None:
                fd["bold"] = bool(font.b)
            if getattr(font, "i", None) is not None:
                fd["italic"] = bool(font.i)
            col = getattr(font, "color", None)
            rgb = getattr(col, "rgb", None) if col is not None else None
            if rgb:
                fd["color"] = str(rgb)
            if fd:
                d["font"] = fd
        fill = cell.fill
        if fill is not None:
            fg = getattr(fill, "fgColor", None)
            rgb2 = getattr(fg, "rgb", None) if fg is not None else None
            if rgb2:
                d["fill"] = {"fgColor": str(rgb2)}
        al = cell.alignment
        if al is not None:
            ad: dict = {}
            if getattr(al, "horizontal", None):
                ad["horizontal"] = str(al.horizontal)
            if getattr(al, "vertical", None):
                ad["vertical"] = str(al.vertical)
            if getattr(al, "wrapText", None) is not None:
                ad["wrapText"] = bool(al.wrapText)
            if getattr(al, "textRotation", None) not in (None, 0):
                try:
                    ad["textRotation"] = int(al.textRotation)
                except (TypeError, ValueError):
                    pass
            if ad:
                d["alignment"] = ad
        border = cell.border
        if border is not None:
            sides = ("left", "right", "top", "bottom")
            bd: dict = {}
            for side in sides:
                b = getattr(border, side, None)
                if not b or not getattr(b, "style", None):
                    continue
                st = str(b.style)
                rgb3 = None
                colo = getattr(b, "color", None)
                if colo is not None:
                    rgb3 = getattr(colo, "rgb", None)
                bd[side] = {"style": st, "color": str(rgb3) if rgb3 else None}
            if bd:
                d["border"] = bd
    except RECOVERABLE_ERRORS:
        return {}
    return d


def _build_grid_style_cache(
    ws, row_count: int, col_count: int, merge_skip: set[tuple[int, int]]
) -> dict:
    from openpyxl.utils import get_column_letter

    styles: dict[str, dict] = {}
    cell_style_refs: dict[str, str] = {}
    max_styles = 256
    for r in range(1, row_count + 1):
        for c in range(1, col_count + 1):
            if (r, c) in merge_skip:
                continue
            cell = ws.cell(row=r, column=c)
            sd = _serialize_cell_style(cell)
            if not sd:
                continue
            key = hashlib.sha256(
                json.dumps(sd, sort_keys=True, ensure_ascii=False).encode("utf-8")
            ).hexdigest()[:16]
            if key not in styles:
                if len(styles) >= max_styles:
                    continue
                styles[key] = sd
            coord = f"{get_column_letter(c)}{r}"
            cell_style_refs[coord] = key
    return {"styles": styles, "cell_style_refs": cell_style_refs}


def _matrix_to_real_grid_rows(
    ws, values_rows: list[list[str]], row_count: int, col_count: int
) -> list[list[dict]]:
    from openpyxl.utils import get_column_letter

    merge_anchor, merge_skip = _merge_anchor_and_skip(ws, row_count, col_count)

    col_width_px: dict[int, int] = {}
    for c in range(1, col_count + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        col_width_px[c] = _excel_col_width_to_px(getattr(dim, "width", None))

    row_height_px: dict[int, int] = {}
    for r in range(1, row_count + 1):
        dim = ws.row_dimensions.get(r)
        row_height_px[r] = _excel_row_height_to_px(getattr(dim, "height", None))

    out: list[list[dict]] = []
    for r_idx in range(1, row_count + 1):
        row = values_rows[r_idx - 1] if r_idx - 1 < len(values_rows) else []
        line: list[dict] = []
        for col_idx in range(1, col_count + 1):
            if (r_idx, col_idx) in merge_skip:
                continue
            text = row[col_idx - 1] if col_idx - 1 < len(row) else ""
            rowspan, colspan = merge_anchor.get((r_idx, col_idx), (1, 1))
            width_px = sum(
                col_width_px.get(col_idx + k, col_width_px.get(col_idx, 92)) for k in range(colspan)
            )
            height_px = sum(
                row_height_px.get(r_idx + k, row_height_px.get(r_idx, 20)) for k in range(rowspan)
            )
            line.append(
                {
                    "col": col_idx,
                    "text": str(text or ""),
                    "rowspan": rowspan,
                    "colspan": colspan,
                    "width_px": width_px,
                    "height_px": height_px,
                }
            )
        out.append(line)
    return out


_EXTRACT_GRID_MAX_SHEETS = 30


def _extract_single_sheet_bundle(
    wb,
    sn: str,
    _sheet_names_full: list[str],
    _persisted_rel: str,
    _original_filename: str,
) -> dict:
    ws = wb[sn]
    max_row = int(ws.max_row or 0)
    max_col = int(ws.max_column or 0)
    read_rows = min(max_row, 60)
    read_cols = min(max_col, 30)

    matrix: list[list[str]] = []
    for r in range(1, read_rows + 1):
        row_vals: list[str] = []
        for c in range(1, read_cols + 1):
            row_vals.append(_excel_cell_to_text(ws.cell(row=r, column=c).value))
        matrix.append(row_vals)

    effective_cols = _detect_effective_col_count(matrix, read_cols)
    effective_rows = _detect_effective_row_count(matrix, read_rows)
    matrix = matrix[:effective_rows]
    matrix = [row[:effective_cols] for row in matrix]

    empty_preview = {
        "rows": [],
        "max_row": max_row,
        "max_col": 0,
        "header_row_index": 1,
    }
    empty_cache: object = {"styles": {}, "cell_style_refs": {}}

    if not matrix:
        return {
            "sheet_name": sn,
            "fields": [],
            "sample_rows": [],
            "grid_preview": empty_preview,
            "grid_style_cache": empty_cache,
            "tables": [],
        }

    header_idx, header_row = _pick_header_row(matrix)
    headers: list[str] = []
    for i, h in enumerate(header_row, start=1):
        txt = str(h or "").strip()
        headers.append(txt or f"列{i}")

    fields = [{"label": h, "name": h, "type": "dynamic"} for h in headers if str(h).strip()]

    sample_rows: list[dict] = []
    for row in matrix[header_idx + 1 : header_idx + 11]:
        if not any(str(x or "").strip() for x in row):
            continue
        item = {}
        for i, key in enumerate(headers):
            item[key] = row[i] if i < len(row) else ""
        sample_rows.append(item)

    preview_row_count = len(matrix)
    grid_rows = _matrix_to_real_grid_rows(
        ws, matrix[:preview_row_count], preview_row_count, effective_cols
    )
    _, merge_skip = _merge_anchor_and_skip(ws, preview_row_count, effective_cols)
    grid_style_cache = _build_grid_style_cache(ws, preview_row_count, effective_cols, merge_skip)

    tables = [
        {
            "table_index": 0,
            "header_row": header_idx + 1,
            "fields": list(fields),
            "sample_rows": list(sample_rows[:12]),
        }
    ]

    grid_preview = {
        "rows": grid_rows,
        "max_row": max_row,
        "max_col": effective_cols,
        "header_row_index": header_idx + 1,
    }

    return {
        "sheet_name": sn,
        "fields": fields,
        "sample_rows": sample_rows,
        "grid_preview": grid_preview,
        "grid_style_cache": grid_style_cache,
        "tables": tables,
    }


def _bundle_to_sheet_entry(bundle: dict, sheet_index: int) -> dict:
    st = bundle["grid_style_cache"]
    return {
        "sheet_index": sheet_index,
        "sheet_name": bundle["sheet_name"],
        "fields": bundle["fields"],
        "sample_rows": bundle["sample_rows"],
        "grid_preview": bundle["grid_preview"],
        "style_cache": st,
        "tables": bundle["tables"],
    }


@router.post("/templates/extract-grid")
@router.post("/templates/extract-grid/", include_in_schema=False)
async def templates_extract_grid(
    file: UploadFile = File(...),
    sheet_name: str | None = Form(default=None),
    analyze_all_sheets: str | None = Form(default=None),
) -> dict:
    name = (file.filename or "").strip()
    logger.info(
        "extract-grid: request filename=%r analyze_all_sheets=%r sheet_name=%r",
        name,
        analyze_all_sheets,
        sheet_name,
    )
    suffix = Path(name).suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise HTTPException(status_code=400, detail="仅支持 .xlsx/.xlsm 文件")

    try:
        from io import BytesIO

        from openpyxl import load_workbook

        raw = await file.read()
        logger.info("extract-grid: read upload %d bytes, persisting + load_workbook", len(raw))
        workspace_root = Path(os.environ.get("WORKSPACE_ROOT", os.getcwd())).resolve()
        upload_dir = workspace_root / "uploads"
        upload_dir.mkdir(parents=True, exist_ok=True)
        persisted_path = upload_dir / f"{uuid.uuid4().hex}{suffix}"
        persisted_path.write_bytes(raw)
        try:
            persisted_rel = persisted_path.relative_to(workspace_root).as_posix()
        except ValueError:
            persisted_rel = str(persisted_path)

        wb = load_workbook(filename=BytesIO(raw), data_only=True)
    except RECOVERABLE_ERRORS as e:
        raise HTTPException(status_code=400, detail=f"Excel 读取失败: {e}") from e

    sheet_names = list(wb.sheetnames or [])
    logger.info("extract-grid: workbook loaded filename=%r sheet_count=%d", name, len(sheet_names))
    if not sheet_names:
        return {
            "success": True,
            "template_name": name,
            "file_path": persisted_rel,
            "fields": [],
            "sheets": [],
            "preview_data": {
                "sheet_names": [],
                "selected_sheet_name": "",
                "sheet_name": "",
                "file_path": persisted_rel,
                "sample_rows": [],
                "grid_preview": {"rows": [], "max_row": 0, "max_col": 0},
                "grid_style_cache": {"styles": {}, "cell_style_refs": {}},
                "tables": [],
                "all_sheets": [],
            },
        }

    analyze_all = _form_bool(analyze_all_sheets)
    sn_arg = (sheet_name or "").strip()
    multi_mode = analyze_all and (not sn_arg or sn_arg not in sheet_names)

    if multi_mode:
        targets = sheet_names[:_EXTRACT_GRID_MAX_SHEETS]
        bundles = [
            _extract_single_sheet_bundle(wb, sn, sheet_names, persisted_rel, name) for sn in targets
        ]
        first = bundles[0]
        sheets_top = [_bundle_to_sheet_entry(b, i + 1) for i, b in enumerate(bundles)]
        all_sheets = [_bundle_to_sheet_entry(b, i + 1) for i, b in enumerate(bundles)]
        return {
            "success": True,
            "template_name": name,
            "file_path": persisted_rel,
            "fields": first["fields"],
            "sheets": sheets_top,
            "preview_data": {
                "sheet_names": sheet_names,
                "selected_sheet_name": first["sheet_name"],
                "sheet_name": first["sheet_name"],
                "file_path": persisted_rel,
                "sample_rows": first["sample_rows"],
                "grid_preview": first["grid_preview"],
                "grid_style_cache": first["grid_style_cache"],
                "tables": first["tables"],
                "all_sheets": all_sheets,
            },
        }

    chosen = sn_arg if sn_arg in sheet_names else sheet_names[0]
    bundle = _extract_single_sheet_bundle(wb, chosen, sheet_names, persisted_rel, name)
    one = _bundle_to_sheet_entry(bundle, 1)
    return {
        "success": True,
        "template_name": name,
        "file_path": persisted_rel,
        "fields": bundle["fields"],
        "sheets": [one],
        "preview_data": {
            "sheet_names": sheet_names,
            "selected_sheet_name": bundle["sheet_name"],
            "sheet_name": bundle["sheet_name"],
            "file_path": persisted_rel,
            "sample_rows": bundle["sample_rows"],
            "grid_preview": bundle["grid_preview"],
            "grid_style_cache": bundle["grid_style_cache"],
            "tables": bundle["tables"],
            "all_sheets": [one],
        },
    }
