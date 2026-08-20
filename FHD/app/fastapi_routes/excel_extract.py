"""Excel 提取与生成 API（自归档 excel_extract 蓝图迁移）。"""

from __future__ import annotations

import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Any

from fastapi import APIRouter, Body, File, Form, UploadFile
from fastapi.responses import FileResponse, JSONResponse

from app.application.shipment_excel_etl_security import (
    ShipmentEtlPathError,
    resolve_etl_output_path,
    resolve_etl_path,
)
from app.utils.operational_errors import RECOVERABLE_ERRORS
from app.utils.path_io.path_utils import get_app_data_dir
from app.utils.security.secure_filename import secure_filename

logger = logging.getLogger(__name__)
_EXCEL_UNAVAILABLE = "Excel 服务暂时不可用，请稍后重试"

router: APIRouter = APIRouter(prefix="/api/excel/data", tags=["excel-data"])

TEMP_EXCEL_DIR = os.path.join(get_app_data_dir(), "temp_excel")
os.makedirs(TEMP_EXCEL_DIR, exist_ok=True)


def _form_truthy(value: str, default: bool = True) -> bool:
    text = str(value if value is not None else "").strip().lower()
    if not text:
        return default
    return text in {"1", "true", "yes", "on"}


def _form_include_ledger(value: str, default: str = "auto") -> bool | str:
    text = str(value if value is not None else "").strip().lower()
    if not text:
        text = str(default or "auto").strip().lower()
    if text in {"1", "true", "yes", "on"}:
        return True
    if text in {"0", "false", "no", "off"}:
        return False
    return "auto"


def _get_ai_product_parser():
    from app.application.facades.excel_facade import get_ai_product_parser

    return get_ai_product_parser()


def _extract_from_excel(file_path, sheet_name=None, header_row=1) -> tuple[dict, int]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter

        safe_file_path = resolve_etl_path(file_path)
        if not safe_file_path.is_file():
            return {"success": False, "message": "文件不存在"}, 404

        wb = load_workbook(safe_file_path, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]

        headers: list[dict[str, Any]] = []
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v is not None:
                headers.append(
                    {
                        "column": get_column_letter(c),
                        "column_index": c,
                        "value": str(v).strip() if v else "",
                    }
                )

        rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            row_data: dict[str, Any] = {}
            for h in headers:
                v = ws.cell(r, h["column_index"]).value
                row_data[h["value"]] = v
            if any(v is not None and v != "" for v in row_data.values()):
                rows.append(row_data)

        return {
            "success": True,
            "file": safe_file_path.name,
            "sheet": ws.title,
            "header_row": header_row,
            "headers": headers,
            "rows": rows,
            "total_rows": len(rows),
        }, 200

    except ShipmentEtlPathError:
        return {"success": False, "message": "文件不存在"}, 404
    except RECOVERABLE_ERRORS:
        logger.exception("提取 Excel 数据失败")
        return {"success": False, "message": _EXCEL_UNAVAILABLE}, 500


def _generate_excel(data, filename=None, sheet_name="Sheet1") -> tuple[dict, int]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"export_{timestamp}.xlsx"

        safe_name = secure_filename(str(filename))
        if not safe_name:
            return {"success": False, "message": "文件名无效"}, 400
        file_path = resolve_etl_output_path(Path(TEMP_EXCEL_DIR) / safe_name)

        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.title = sheet_name

        if not data or not isinstance(data, list):
            return {"success": False, "message": "数据格式错误，需要数组类型"}, 400

        headers = list(data[0].keys()) if data else []

        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col, header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row_idx, col_idx, row_data.get(header))

        wb.save(file_path)

        return {
            "success": True,
            "file_path": str(file_path),
            "filename": safe_name,
            "sheet": sheet_name,
            "rows": len(data),
        }, 200

    except ShipmentEtlPathError:
        return {"success": False, "message": "输出路径无效"}, 400
    except RECOVERABLE_ERRORS:
        logger.exception("生成 Excel 文件失败")
        return {"success": False, "message": _EXCEL_UNAVAILABLE}, 500


def _extract_attendance_detail_roster(
    file_path: str | Path, sheet_name: str | None = None
) -> tuple[dict, int]:
    """解析太阳鸟/钉钉考勤统计表「明细」：自第 4 行起每 6 行一块，A 部门、B 性质、C 姓名。"""
    try:
        from openpyxl import load_workbook

        safe_file_path = resolve_etl_path(file_path)
        if not safe_file_path.is_file():
            return {"success": False, "message": "文件不存在"}, 404

        header_rows = 3
        block_rows = 6

        wb = load_workbook(safe_file_path, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif "明细" in wb.sheetnames:
                ws = wb["明细"]
            else:
                ws = wb[wb.sheetnames[0]]

            rows_out: list[dict[str, Any]] = []
            for base_row in range(header_rows + 1, ws.max_row + 1, block_rows):
                raw_name = ws.cell(base_row, 3).value
                if raw_name in (None, ""):
                    continue
                name = str(raw_name).strip()
                if not name:
                    continue
                dept = str(ws.cell(base_row, 1).value or "").strip()
                nature = str(ws.cell(base_row, 2).value or "").strip()
                rows_out.append(
                    {
                        "product_code": "",
                        "product_name": name,
                        "specification": nature,
                        "unit": dept or "个",
                        "unit_price": 0,
                        "remark": "__from_attendance_detail__",
                    }
                )

            headers = [
                {"column": "A", "column_index": 1, "value": "部门"},
                {"column": "B", "column_index": 2, "value": "性质"},
                {"column": "C", "column_index": 3, "value": "姓名"},
            ]
            return {
                "success": True,
                "file": safe_file_path.name,
                "sheet": ws.title,
                "header_row": header_rows,
                "headers": headers,
                "rows": rows_out,
                "total_rows": len(rows_out),
                "parse_mode": "attendance_detail",
            }, 200
        finally:
            wb.close()
    except ShipmentEtlPathError:
        return {"success": False, "message": "文件不存在"}, 404
    except RECOVERABLE_ERRORS:
        logger.exception("考勤明细人员提取失败")
        return {"success": False, "message": _EXCEL_UNAVAILABLE}, 500


@router.post("/extract")
def extract_from_excel(data: dict[str, Any] = Body(default_factory=dict)):
    try:
        file_path = data.get("file_path")
        sheet_name = data.get("sheet_name")
        header_row = data.get("header_row", 1)
        if not file_path:
            return JSONResponse(
                {"success": False, "message": "请提供 file_path 参数"}, status_code=400
            )
        result, status = _extract_from_excel(file_path, sheet_name, header_row)
        return JSONResponse(result, status_code=status)
    except RECOVERABLE_ERRORS:
        logger.exception("提取 Excel 数据失败")
        return JSONResponse({"success": False, "message": _EXCEL_UNAVAILABLE}, status_code=500)


@router.post("/extract/upload")
async def extract_upload(
    excel_file: UploadFile | None = File(default=None),
    sheet_name: str | None = Form(default=None),
    header_row: int = Form(default=1),
    parse_mode: str = Form(default=""),
):
    try:
        if excel_file is None or not excel_file.filename:
            return JSONResponse({"success": False, "message": "请上传 Excel 文件"}, status_code=400)
        if not excel_file.filename.lower().endswith((".xlsx", ".xls")):
            return JSONResponse(
                {"success": False, "message": "只支持 .xlsx 和 .xls 格式"}, status_code=400
            )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        upload_name = secure_filename(excel_file.filename)
        if not upload_name:
            return JSONResponse({"success": False, "message": "文件名无效"}, status_code=400)
        filename = f"extract_{timestamp}_{upload_name}"
        file_path = resolve_etl_output_path(Path(TEMP_EXCEL_DIR) / filename)
        body = await excel_file.read()
        with open(file_path, "wb") as f:
            f.write(body)
        mode = (parse_mode or "").strip().lower()
        if mode in ("attendance_detail", "attendance", "考勤明细"):
            result, status = _extract_attendance_detail_roster(file_path, sheet_name)
        else:
            result, status = _extract_from_excel(file_path, sheet_name, header_row)
        if os.path.exists(file_path):
            os.remove(file_path)
        return JSONResponse(result, status_code=status)
    except RECOVERABLE_ERRORS:
        logger.exception("上传并提取 Excel 数据失败")
        return JSONResponse({"success": False, "message": _EXCEL_UNAVAILABLE}, status_code=500)


@router.post("/generate")
def generate_excel(data: dict[str, Any] = Body(default_factory=dict)):
    try:
        excel_data = data.get("data", [])
        filename = data.get("filename")
        sheet_name = data.get("sheet_name", "Sheet1")
        if not excel_data:
            return JSONResponse(
                {"success": False, "message": "请提供数据 data 参数"}, status_code=400
            )
        result, status = _generate_excel(excel_data, filename, sheet_name)
        return JSONResponse(result, status_code=status)
    except RECOVERABLE_ERRORS:
        logger.exception("生成 Excel 文件失败")
        return JSONResponse({"success": False, "message": _EXCEL_UNAVAILABLE}, status_code=500)


@router.post("/generate/download")
def download_generated_excel(data: dict[str, Any] = Body(default_factory=dict)):
    try:
        excel_data = data.get("data", [])
        filename = data.get("filename")
        sheet_name = data.get("sheet_name", "Sheet1")
        if not excel_data:
            return JSONResponse(
                {"success": False, "message": "请提供数据 data 参数"}, status_code=400
            )
        result, status = _generate_excel(excel_data, filename, sheet_name)
        if status != 200:
            return JSONResponse(result, status_code=status)
        file_path = result.get("file_path")
        download_filename = result.get("filename")
        if not file_path or not Path(file_path).is_file():
            return JSONResponse({"success": False, "message": "生成的文件不存在"}, status_code=500)
        return FileResponse(
            file_path,
            filename=download_filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except RECOVERABLE_ERRORS:
        logger.exception("生成并下载 Excel 文件失败")
        return JSONResponse({"success": False, "message": _EXCEL_UNAVAILABLE}, status_code=500)


@router.get("/extract/test")
def extract_test():
    return JSONResponse(
        {
            "success": True,
            "message": "Excel 提取服务运行正常",
            "timestamp": datetime.now().isoformat(),
        }
    )


@router.post("/import/products")
def import_products(data: dict[str, Any] = Body(default_factory=dict)):
    try:
        from app.application.facades.excel_facade import get_product_import_service
        from app.bootstrap import get_extract_log_service

        extracted_data = data.get("data", [])
        options = data.get("options", {})
        if not extracted_data:
            return JSONResponse({"success": False, "message": "缺少数据"}, status_code=400)

        log_service = get_extract_log_service()
        log_id = log_service.create_log(
            file_name=data.get("file_name", "products_import"),
            data_type="products",
            total_rows=len(extracted_data),
            field_mapping=data.get("field_mapping"),
        )

        use_ai_parse = bool(options.get("use_ai_parse", False))
        ai_source_field = options.get("ai_source_field") or ""
        if use_ai_parse and ai_source_field:
            parser = _get_ai_product_parser()
            normalized_rows = []
            for row in extracted_data:
                raw_text = str(row.get(ai_source_field, "") or "")
                if not raw_text.strip():
                    normalized_rows.append(row)
                    continue
                parsed = parser.parse_single(raw_text, use_ai=True, fallback_to_rule=True)
                if parsed.get("success"):
                    row = dict(row)
                    if parsed.get("product_code"):
                        row["product_code"] = parsed["product_code"]
                    if parsed.get("product_name"):
                        row["product_name"] = parsed["product_name"]
                    if parsed.get("specification"):
                        row["specification"] = parsed["specification"]
                    if parsed.get("unit") and "unit" not in row:
                        row["unit"] = parsed["unit"]
                    if parsed.get("unit_price") is not None and "unit_price" not in row:
                        row["unit_price"] = parsed.get("unit_price")
                normalized_rows.append(row)
            extracted_data = normalized_rows

        service = get_product_import_service()
        result = service.import_data(
            data=extracted_data,
            skip_duplicates=options.get("skip_duplicates", True),
            validate_before_import=options.get("validate_before_import", True),
            clean_data=options.get("clean_data", True),
            replace_attendance_detail_tagged=bool(options.get("replace_attendance_detail_tagged")),
        )

        log_service.update_log(
            log_id=log_id,
            status="completed" if result["failed"] == 0 else "partial",
            imported_rows=result["imported"],
            skipped_rows=result["skipped"],
            failed_rows=result["failed"],
        )

        return JSONResponse(
            {
                "success": True,
                "log_id": log_id,
                "imported": result["imported"],
                "skipped": result["skipped"],
                "failed": result["failed"],
                "details": result["details"],
            }
        )
    except RECOVERABLE_ERRORS:
        logger.exception("导入产品数据失败")
        return JSONResponse(
            {"success": False, "message": "产品导入服务暂时不可用，请稍后重试"}, status_code=500
        )


@router.post("/import/customers")
def import_customers(data: dict[str, Any] = Body(default_factory=dict)):
    try:
        from app.application import get_customer_app_service
        from app.bootstrap import get_extract_log_service

        extracted_data = data.get("data", [])
        options = data.get("options", {})
        if not extracted_data:
            return JSONResponse({"success": False, "message": "缺少数据"}, status_code=400)

        log_service = get_extract_log_service()
        log_id = log_service.create_log(
            file_name=data.get("file_name", "customers_import"),
            data_type="customers",
            total_rows=len(extracted_data),
            field_mapping=data.get("field_mapping"),
        )

        service = get_customer_app_service()
        result = service.import_data(
            data=extracted_data,
            skip_duplicates=options.get("skip_duplicates", True),
            validate_before_import=options.get("validate_before_import", True),
            clean_data=options.get("clean_data", True),
        )

        log_service.update_log(
            log_id=log_id,
            status="completed" if result["failed"] == 0 else "partial",
            imported_rows=result["imported"],
            skipped_rows=result["skipped"],
            failed_rows=result["failed"],
        )

        return JSONResponse(
            {
                "success": True,
                "log_id": log_id,
                "imported": result["imported"],
                "skipped": result["skipped"],
                "failed": result["failed"],
                "details": result["details"],
            }
        )
    except RECOVERABLE_ERRORS:
        logger.exception("导入客户数据失败")
        return JSONResponse(
            {"success": False, "message": "客户导入服务暂时不可用，请稍后重试"}, status_code=500
        )


from app.fastapi_routes.excel_extract_shipment import (
    get_extract_log as get_extract_log,
)
from app.fastapi_routes.excel_extract_shipment import (
    get_extract_logs as get_extract_logs,
)
from app.fastapi_routes.excel_extract_shipment import (
    get_preview as get_preview,
)
from app.fastapi_routes.excel_extract_shipment import (
    shipment_etl_batch_execute as shipment_etl_batch_execute,
)
from app.fastapi_routes.excel_extract_shipment import (
    shipment_etl_batch_preview as shipment_etl_batch_preview,
)
from app.fastapi_routes.excel_extract_shipment import (
    shipment_etl_execute as shipment_etl_execute,
)
from app.fastapi_routes.excel_extract_shipment import (
    shipment_etl_generate_template as shipment_etl_generate_template,
)
from app.fastapi_routes.excel_extract_shipment import (
    shipment_etl_ocr_preview as shipment_etl_ocr_preview,
)
from app.fastapi_routes.excel_extract_shipment import (
    shipment_etl_preview as shipment_etl_preview,
)
from app.fastapi_routes.excel_extract_shipment import (
    shipment_etl_regenerate as shipment_etl_regenerate,
)
