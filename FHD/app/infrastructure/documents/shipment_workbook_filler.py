"""Fill detected delivery-note layouts without assuming fixed row numbers."""

from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from app.application.etl.parser_structure import clean_cell_text, semantic_key


def _header_field(value: Any) -> str:
    key = semantic_key(value)
    if key in {"型号", "编号", "产品型号", "货品型号", "商品型号"}:
        return "model_number"
    if key in {"品名", "名称", "产品", "产品名称", "货品名称", "商品名称"}:
        return "name"
    if any(token in key for token in ("数量件", "数量桶", "件数", "桶数")):
        return "quantity_tins"
    if key == "数量":
        return "quantity_tins"
    if key.startswith("规格") or key in {"净重", "每桶kg"}:
        return "tin_spec"
    if "数量kg" in key or "总重量" in key or "重量kg" in key:
        return "quantity_kg"
    if any(token in key for token in ("单价", "价格")):
        return "unit_price"
    if any(token in key for token in ("金额", "合计金额")):
        return "amount"
    if key in {"备注", "说明"}:
        return "description"
    return ""


def _layout(worksheet: Any) -> tuple[int, int, dict[str, int]]:
    for row in range(1, min(worksheet.max_row, 80) + 1):
        mapping: dict[str, int] = {}
        for column in range(1, min(worksheet.max_column, 30) + 1):
            field = _header_field(worksheet.cell(row=row, column=column).value)
            if field and field not in mapping:
                mapping[field] = column
        if ("name" in mapping or "model_number" in mapping) and len(mapping) >= 4:
            total_row = 0
            for candidate in range(row + 1, worksheet.max_row + 1):
                first_cells = " ".join(
                    clean_cell_text(worksheet.cell(candidate, column).value)
                    for column in range(1, min(4, worksheet.max_column) + 1)
                )
                if semantic_key(first_cells) in {"合计", "总计", "小计", "汇总"}:
                    total_row = candidate
                    break
            if total_row:
                return row, total_row, mapping
    raise ValueError("发货单版式缺少可识别的产品表头或合计行")


def _replace_meta(
    worksheet: Any,
    *,
    header_row: int,
    unit_name: str,
    contact_person: str,
    order_number: str,
    date: str | None,
) -> None:
    target_row = max(1, header_row - 1)
    for row in range(header_row - 1, max(0, header_row - 6), -1):
        text = " ".join(
            clean_cell_text(worksheet.cell(row, column).value)
            for column in range(1, min(worksheet.max_column, 12) + 1)
        )
        if "购货单位" in text or "购买单位" in text or "订单" in text:
            target_row = row
            break
    display_date = str(date or "").strip()
    if not display_date:
        now = datetime.now()
        display_date = f"{now.year}年{now.month:02d}月{now.day:02d}日"
    worksheet.cell(target_row, 1).value = (
        f"购货单位：{unit_name}       联系人：{contact_person}       "
        f"{display_date}      订单编号：{order_number}"
    )


def _number(value: Any) -> float:
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _line_kg(item: dict[str, Any]) -> float:
    return _number(item.get("quantity_kg")) or (
        _number(item.get("quantity_tins")) * _number(item.get("tin_spec"))
    )


def _line_amount(item: dict[str, Any]) -> float:
    return _number(item.get("amount")) or (
        _line_kg(item) * _number(item.get("unit_price") or item.get("price"))
    )


def fill_shipment_workbook(
    template_path: str | Path,
    *,
    output_path: str | Path,
    unit_name: str,
    contact_person: str,
    products: list[dict[str, Any]],
    order_number: str,
    date: str | None = None,
) -> dict[str, Any]:
    source = Path(template_path)
    keep_vba = source.suffix.lower() == ".xlsm"
    workbook = load_workbook(source, data_only=False, keep_vba=keep_vba, keep_links=False)
    try:
        worksheet = workbook.active
        header_row, total_row, mapping = _layout(worksheet)
        capacity = total_row - header_row - 1
        if len(products) > capacity:
            raise ValueError(f"当前发货单版式最多容纳 {capacity} 个产品，请拆分开单")
        _replace_meta(
            worksheet,
            header_row=header_row,
            unit_name=unit_name,
            contact_person=contact_person,
            order_number=order_number,
            date=date,
        )
        for row in range(header_row + 1, total_row):
            for column in set(mapping.values()):
                worksheet.cell(row, column).value = None
        for index, product in enumerate(products, start=header_row + 1):
            quantity_tins = _number(product.get("quantity_tins"))
            tin_spec = _number(product.get("tin_spec"))
            quantity_kg = _number(product.get("quantity_kg")) or quantity_tins * tin_spec
            unit_price = _number(product.get("unit_price") or product.get("price"))
            amount = _number(product.get("amount")) or quantity_kg * unit_price
            values = {
                "model_number": product.get("model_number") or "",
                "name": product.get("name") or product.get("product_name") or "",
                "quantity_tins": int(quantity_tins)
                if quantity_tins.is_integer()
                else quantity_tins,
                "tin_spec": int(tin_spec) if tin_spec.is_integer() else tin_spec,
                "quantity_kg": quantity_kg,
                "unit_price": unit_price,
                "amount": round(amount, 2),
                "description": product.get("description") or "",
            }
            for field, column in mapping.items():
                worksheet.cell(index, column).value = values.get(field, "")
        first_data_row = header_row + 1
        last_data_row = total_row - 1
        for field in ("quantity_tins", "quantity_kg", "amount"):
            column = mapping.get(field)
            if column:
                letter = get_column_letter(column)
                worksheet.cell(
                    total_row, column
                ).value = f"=SUM({letter}{first_data_row}:{letter}{last_data_row})"
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True
        destination = Path(output_path)
        destination.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(destination)
    finally:
        workbook.close()
    return {
        "file_path": str(output_path),
        "filename": Path(output_path).name,
        "order_number": order_number,
        "total_amount": round(sum(_line_amount(item) for item in products), 2),
        "total_quantity": sum(_line_kg(item) for item in products),
    }


def safe_shipment_filename(order_number: str) -> str:
    token = re.sub(r"[^0-9A-Za-z._-]+", "-", str(order_number or "")).strip("-")
    return f"发货单_{token or datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"


__all__ = ["fill_shipment_workbook", "safe_shipment_filename"]
