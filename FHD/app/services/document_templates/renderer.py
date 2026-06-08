from __future__ import annotations

import json
import logging

logger = logging.getLogger(__name__)


def _is_unreadable_workbook_error(error_message: str) -> bool:
    text = str(error_message or "").lower()
    markers = [
        "unable to read workbook",
        "could not read worksheets",
        "invalid xml",
        "badzipfile",
    ]
    return any(m in text for m in markers)


def _extract_structured_excel_preview(
    file_path: str, sheet_name: str = None, sample_limit: int = 6
):
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return {"fields": [], "sample_rows": [], "sheet_name": ""}

            header_row_idx = None
            header_entries = []
            max_scan_rows = min(ws.max_row or 0, 30)
            max_scan_cols = min(ws.max_column or 0, 25)
            best_candidate = None

            def _norm(text):
                return str(text or "").strip().replace(" ", "").lower()

            def _is_number_like(text: str) -> bool:
                t = str(text or "").strip().replace(",", "")
                if not t:
                    return False
                try:
                    float(t)
                    return True
                except Exception:
                    return False

            def _is_date_like(text: str) -> bool:
                t = str(text or "").strip()
                if not t:
                    return False
                if any(ch in t for ch in ("年", "月", "日", "-")) and any(ch.isdigit() for ch in t):
                    return True
                if len(t) >= 8 and t.isdigit():
                    return True
                return False

            for r in range(1, max_scan_rows + 1):
                row_headers = []
                norm_texts = []
                str_like = 0
                num_like = 0
                date_like = 0
                for c in range(1, max_scan_cols + 1):
                    value = ws.cell(r, c).value
                    if value is None:
                        continue
                    text = str(value).strip()
                    if not text:
                        continue
                    row_headers.append({"name": text, "column_index": c})
                    nt = _norm(text)
                    norm_texts.append(nt)
                    if _is_number_like(text):
                        num_like += 1
                    elif _is_date_like(text):
                        date_like += 1
                    else:
                        str_like += 1

                if len(row_headers) < 3:
                    continue

                unique_ratio = (len(set(norm_texts)) / len(norm_texts)) if norm_texts else 0.0
                text_ratio = (str_like / len(row_headers)) if row_headers else 0.0
                number_ratio = (num_like / len(row_headers)) if row_headers else 0.0
                date_ratio = (date_like / len(row_headers)) if row_headers else 0.0

                score = 0.0
                score += text_ratio * 12.0
                score += unique_ratio * 8.0
                score -= number_ratio * 10.0
                score -= date_ratio * 4.0
                score -= 1.2 if r <= 2 else 0.0

                next_row_non_empty = 0
                if r + 1 <= (ws.max_row or 0):
                    for h in row_headers[: min(len(row_headers), 16)]:
                        nv = ws.cell(r + 1, h["column_index"]).value
                        if nv is not None and str(nv).strip() != "":
                            next_row_non_empty += 1
                score += min(next_row_non_empty, 8)

                if best_candidate is None or score > best_candidate["score"]:
                    best_candidate = {
                        "row": r,
                        "headers": row_headers,
                        "score": score,
                    }

            if best_candidate is not None:
                header_row_idx = best_candidate["row"]
                header_entries = best_candidate["headers"]

            if header_row_idx is None:
                return {"fields": [], "sample_rows": [], "sheet_name": ws.title}

            fields = [{"label": h["name"], "value": "", "type": "dynamic"} for h in header_entries]

            sample_rows = []
            for r in range(
                header_row_idx + 1, min((ws.max_row or 0), header_row_idx + sample_limit + 15) + 1
            ):
                row_data = {}
                has_non_empty = False
                for h in header_entries:
                    value = ws.cell(r, h["column_index"]).value
                    if value is not None and str(value).strip() != "":
                        has_non_empty = True
                    row_data[h["name"]] = value
                if has_non_empty:
                    sample_rows.append(row_data)
                if len(sample_rows) >= sample_limit:
                    break

            return {"fields": fields, "sample_rows": sample_rows, "sheet_name": ws.title}
        finally:
            wb.close()
    except Exception:
        return {"fields": [], "sample_rows": [], "sheet_name": sheet_name or ""}


def _extract_excel_grid_preview(
    file_path: str, sheet_name: str = None, max_rows: int = 18, max_cols: int = 12
):
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.cell import range_boundaries

        wb = load_workbook(file_path, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return {"sheet_name": "", "rows": []}

            row_limit = min(ws.max_row or 1, max_rows)
            col_limit = min(ws.max_column or 1, max_cols)

            merged_top_left = {}
            merged_covered = set()
            for merged in ws.merged_cells.ranges:
                min_col, min_row, max_col, max_row = range_boundaries(str(merged))
                if min_row > row_limit or min_col > col_limit:
                    continue
                span_row = max(1, min(max_row, row_limit) - min_row + 1)
                span_col = max(1, min(max_col, col_limit) - min_col + 1)
                merged_top_left[(min_row, min_col)] = (span_row, span_col)
                for r in range(min_row, min(max_row, row_limit) + 1):
                    for c in range(min_col, min(max_col, col_limit) + 1):
                        if (r, c) != (min_row, min_col):
                            merged_covered.add((r, c))

            rows = []
            for r in range(1, row_limit + 1):
                row_cells = []
                for c in range(1, col_limit + 1):
                    if (r, c) in merged_covered:
                        continue
                    rowspan, colspan = merged_top_left.get((r, c), (1, 1))
                    value = ws.cell(r, c).value
                    row_cells.append(
                        {
                            "row": r,
                            "col": c,
                            "text": "" if value is None else str(value),
                            "rowspan": rowspan,
                            "colspan": colspan,
                        }
                    )
                rows.append(row_cells)

            return {"sheet_name": ws.title, "rows": rows}
        finally:
            wb.close()
    except Exception:
        return {"sheet_name": sheet_name or "", "rows": []}


def _extract_excel_grid_style_cache(
    file_path: str, sheet_name: str = None, max_rows: int = 24, max_cols: int = 14
):
    try:
        from copy import copy

        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return {"sheet_name": "", "styles": {}, "cell_style_refs": {}}

            row_limit = min(ws.max_row or 1, max_rows)
            col_limit = min(ws.max_column or 1, max_cols)

            styles = {}
            cell_style_refs = {}

            def _style_snapshot(cell):
                f = copy(cell.font)
                fill = copy(cell.fill)
                border = copy(cell.border)
                align = copy(cell.alignment)
                return {
                    "font": {
                        "name": f.name,
                        "size": f.size,
                        "bold": bool(f.bold),
                        "italic": bool(f.italic),
                        "color": str(getattr(f.color, "rgb", "") or ""),
                    },
                    "fill": {
                        "fill_type": fill.fill_type,
                        "fg_color": str(getattr(fill.fgColor, "rgb", "") or ""),
                    },
                    "border": {
                        "left": str(getattr(border.left, "style", "") or ""),
                        "right": str(getattr(border.right, "style", "") or ""),
                        "top": str(getattr(border.top, "style", "") or ""),
                        "bottom": str(getattr(border.bottom, "style", "") or ""),
                    },
                    "alignment": {
                        "horizontal": align.horizontal,
                        "vertical": align.vertical,
                        "wrap_text": bool(align.wrap_text),
                    },
                    "number_format": cell.number_format or "",
                }

            def _style_key(snapshot):
                import json

                return str(abs(hash(json.dumps(snapshot, ensure_ascii=False, sort_keys=True))))

            for r in range(1, row_limit + 1):
                for c in range(1, col_limit + 1):
                    cell = ws.cell(r, c)
                    snapshot = _style_snapshot(cell)
                    key = _style_key(snapshot)
                    if key not in styles:
                        styles[key] = snapshot
                    cell_style_refs[f"{r},{c}"] = key

            return {
                "sheet_name": ws.title,
                "styles": styles,
                "cell_style_refs": cell_style_refs,
            }
        finally:
            wb.close()
    except Exception:
        return {"sheet_name": sheet_name or "", "styles": {}, "cell_style_refs": {}}


def _extract_excel_all_sheets_preview(
    file_path: str,
    sample_limit: int = 8,
    max_rows: int = 24,
    max_cols: int = 14,
):
    sheets = []
    for idx, name in enumerate(_list_excel_sheet_names(file_path), start=1):
        structured = _extract_structured_excel_preview(
            file_path, sheet_name=name, sample_limit=sample_limit
        )
        grid_preview = _extract_excel_grid_preview(
            file_path, sheet_name=name, max_rows=max_rows, max_cols=max_cols
        )
        style_cache = _extract_excel_grid_style_cache(
            file_path, sheet_name=name, max_rows=max_rows, max_cols=max_cols
        )
        logical_tables = _extract_logical_tables_from_sheet(file_path, sheet_name=name)
        sheets.append(
            {
                "sheet_index": idx,
                "sheet_name": name,
                "fields": structured.get("fields") or [],
                "sample_rows": structured.get("sample_rows") or [],
                "grid_preview": grid_preview,
                "style_cache": style_cache,
                "tables": logical_tables,
            }
        )
    return sheets


def _extract_logical_tables_from_sheet(
    file_path: str, sheet_name: str, max_scan_rows: int = 400, max_scan_cols: int = 25
):
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            if sheet_name not in wb.sheetnames:
                return []
            ws = wb[sheet_name]
            row_limit = min(ws.max_row or 0, max_scan_rows)
            col_limit = min(ws.max_column or 0, max_scan_cols)
            tables = []
            idx = 0
            r = 1
            while r <= row_limit:
                non_empty_cells = []
                for c in range(1, col_limit + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    t = str(v).strip()
                    if t:
                        non_empty_cells.append({"name": t, "column_index": c})
                if len(non_empty_cells) >= 3:
                    header_entries = non_empty_cells
                    sample_rows = []
                    rr = r + 1
                    while rr <= row_limit:
                        row_data = {}
                        has_non_empty = False
                        for h in header_entries:
                            v = ws.cell(rr, h["column_index"]).value
                            if v is not None and str(v).strip() != "":
                                has_non_empty = True
                            row_data[h["name"]] = v
                        if not has_non_empty:
                            break
                        sample_rows.append(row_data)
                        if len(sample_rows) >= 8:
                            break
                        rr += 1
                    fields = [
                        {"label": h["name"], "value": "", "type": "dynamic"} for h in header_entries
                    ]
                    idx += 1
                    tables.append(
                        {
                            "table_index": idx,
                            "header_row": r,
                            "fields": fields,
                            "sample_rows": sample_rows,
                        }
                    )
                    r = max(rr + 1, r + 1)
                    continue
                r += 1
            return tables
        finally:
            wb.close()
    except Exception:
        return []


def _list_excel_sheet_names(file_path: str):
    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=False)
        try:
            names = [str(n).strip() for n in (wb.sheetnames or []) if str(n).strip()]
            if names:
                return names
        finally:
            wb.close()
    except Exception:
        logger.debug("suppressed exception", exc_info=True)

    try:
        from openpyxl import load_workbook

        wb = load_workbook(file_path, data_only=True, read_only=True)
        try:
            names = [str(n).strip() for n in (wb.sheetnames or []) if str(n).strip()]
            if names:
                return names
        finally:
            wb.close()
    except Exception:
        logger.debug("suppressed exception", exc_info=True)

    try:
        import pandas as pd

        excel_file = pd.ExcelFile(file_path)
        names = [str(n).strip() for n in (excel_file.sheet_names or []) if str(n).strip()]
        return names
    except Exception:
        return []


def _parse_json_dict(raw_value):
    if isinstance(raw_value, dict):
        return raw_value
    if not raw_value:
        return {}
    try:
        data = json.loads(raw_value)
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _parse_json_list(raw_value):
    if isinstance(raw_value, list):
        return raw_value
    if not raw_value:
        return []
    try:
        data = json.loads(raw_value)
        return data if isinstance(data, list) else []
    except Exception:
        return []
