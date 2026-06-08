"""传统模式文件浏览与 AI 托管目录能力。

该模块只允许访问一个开放托管根目录，默认仍是仓库下的 ``bang``，可通过
``TRADITIONAL_MODE_ROOT`` 指向任意本机文件夹。前端契约保持不变。
"""

from __future__ import annotations

import base64
import json
import logging
import os
import queue
import shutil
import threading
from collections.abc import Iterator
from datetime import UTC, datetime
from typing import Any

from app.utils.path_utils import get_base_dir

logger = logging.getLogger(__name__)


def _resolve_root_dir() -> str:
    raw = (os.environ.get("TRADITIONAL_MODE_ROOT") or "").strip()
    root = raw if raw else os.path.join(get_base_dir(), "bang")
    root = os.path.abspath(os.path.expanduser(os.path.expandvars(root)))
    os.makedirs(root, exist_ok=True)
    return root


ROOT_DIR = _resolve_root_dir()

IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico", ".tiff", ".svg"}
EXCEL_EXTENSIONS = {".xlsx", ".xls"}

_watch_clients: list[queue.Queue[str]] = []
_watch_clients_lock = threading.Lock()
_last_snapshot: dict[str, str] = {}
_snapshot_lock = threading.Lock()
_watchdog_running = False
_watchdog_thread: threading.Thread | None = None
_watchdog_prev: dict[str, float] = {}


def resolve_safe_path(relative_path: str = "") -> str | None:
    root_abs = os.path.abspath(ROOT_DIR)
    safe = os.path.abspath(os.path.normpath(os.path.join(root_abs, relative_path or "")))
    try:
        common = os.path.commonpath([root_abs, safe])
    except ValueError:
        return None
    if common != root_abs:
        return None
    return safe


def root_info_response() -> dict[str, Any]:
    return {
        "success": True,
        "data": {
            "root": ROOT_DIR,
            "logical_root": "bang",
            "capabilities": [
                "list",
                "read",
                "download",
                "upload",
                "write_excel",
                "write_text",
                "write_base64",
                "append_text",
                "mkdir",
                "rename",
                "delete",
                "move",
                "copy",
            ],
        },
    }


def _rel_from_root(path: str) -> str:
    return os.path.relpath(path, ROOT_DIR).replace(os.sep, "/")


def stat_response(rel_path: str = "") -> tuple[dict[str, Any], int]:
    target = resolve_safe_path(rel_path)
    if target is None:
        return {"success": False, "error": "路径越权访问被拒绝"}, 403
    if not os.path.exists(target):
        return {"success": False, "error": "路径不存在"}, 404
    st = os.stat(target)
    is_dir = os.path.isdir(target)
    return {
        "success": True,
        "data": {
            "path": _rel_from_root(target) if target != ROOT_DIR else "",
            "name": os.path.basename(target) or "bang",
            "is_dir": is_dir,
            "size": 0 if is_dir else st.st_size,
            "modified_time": _format_time(st.st_mtime),
            "type": "文件夹" if is_dir else _get_file_type(os.path.basename(target)),
        },
    }, 200


def write_text_response(
    rel_file: str, content: str, append: bool = False
) -> tuple[dict[str, Any], int]:
    target = resolve_safe_path(rel_file)
    if target is None:
        return {"success": False, "error": "路径越权访问被拒绝"}, 403
    if os.path.isdir(target):
        return {"success": False, "error": "目标路径是目录"}, 400
    os.makedirs(os.path.dirname(target), exist_ok=True)
    mode = "a" if append else "w"
    with open(target, mode, encoding="utf-8", newline="") as f:
        f.write(content)
    return {
        "success": True,
        "data": {"path": _rel_from_root(target), "bytes": os.path.getsize(target)},
    }, 200


def write_base64_response(rel_file: str, content_base64: str) -> tuple[dict[str, Any], int]:
    target = resolve_safe_path(rel_file)
    if target is None:
        return {"success": False, "error": "路径越权访问被拒绝"}, 403
    if os.path.isdir(target):
        return {"success": False, "error": "目标路径是目录"}, 400
    try:
        data = base64.b64decode(content_base64, validate=True)
    except Exception:
        return {"success": False, "error": "base64 内容无效"}, 400
    os.makedirs(os.path.dirname(target), exist_ok=True)
    with open(target, "wb") as f:
        f.write(data)
    return {"success": True, "data": {"path": _rel_from_root(target), "bytes": len(data)}}, 200


def move_response(
    src_rel: str, dst_rel: str, overwrite: bool = False
) -> tuple[dict[str, Any], int]:
    src = resolve_safe_path(src_rel)
    dst = resolve_safe_path(dst_rel)
    if src is None or dst is None:
        return {"success": False, "error": "路径越权访问被拒绝"}, 403
    if not os.path.exists(src):
        return {"success": False, "error": "源路径不存在"}, 404
    if os.path.exists(dst):
        if not overwrite:
            return {"success": False, "error": "目标路径已存在"}, 409
        if os.path.isdir(dst):
            shutil.rmtree(dst)
        else:
            os.remove(dst)
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    shutil.move(src, dst)
    return {"success": True, "data": {"path": _rel_from_root(dst)}}, 200


def copy_response(
    src_rel: str, dst_rel: str, overwrite: bool = False
) -> tuple[dict[str, Any], int]:
    src = resolve_safe_path(src_rel)
    dst = resolve_safe_path(dst_rel)
    if src is None or dst is None:
        return {"success": False, "error": "路径越权访问被拒绝"}, 403
    if not os.path.exists(src):
        return {"success": False, "error": "源路径不存在"}, 404
    if os.path.exists(dst):
        if not overwrite:
            return {"success": False, "error": "目标路径已存在"}, 409
        if os.path.isdir(dst):
            shutil.rmtree(dst)
        else:
            os.remove(dst)
    os.makedirs(os.path.dirname(dst), exist_ok=True)
    if os.path.isdir(src):
        shutil.copytree(src, dst)
    else:
        shutil.copy2(src, dst)
    return {"success": True, "data": {"path": _rel_from_root(dst)}}, 200


def _get_file_type(filename: str) -> str:
    name_lower = filename.lower()
    ext = os.path.splitext(name_lower)[1]
    if ext in EXCEL_EXTENSIONS:
        return ext[1:]
    if ext in IMAGE_EXTENSIONS:
        return ext[1:]
    return ext[1:] if "." in name_lower else "文件"


def _format_time(ts: float) -> str:
    return datetime.fromtimestamp(ts, tz=UTC).isoformat()


def list_files_response(rel_path: str = "") -> tuple[dict[str, Any], int]:
    try:
        full_path = resolve_safe_path(rel_path)
        if full_path is None:
            return {"success": False, "error": "路径越权访问被拒绝"}, 403
        if not os.path.exists(full_path):
            return {"success": True, "data": {"path": rel_path, "files": []}}, 200
        if not os.path.isdir(full_path):
            return {"success": False, "error": "指定路径不是目录"}, 400

        dirs: list[dict[str, Any]] = []
        files: list[dict[str, Any]] = []
        for name in sorted(os.listdir(full_path)):
            entry_path = os.path.join(full_path, name)
            try:
                stat_info = os.stat(entry_path)
                is_dir = os.path.isdir(entry_path)
                entry = {
                    "name": name,
                    "is_dir": is_dir,
                    "size": 0 if is_dir else stat_info.st_size,
                    "modified_time": _format_time(stat_info.st_mtime),
                    "type": "文件夹" if is_dir else _get_file_type(name),
                }
                if is_dir:
                    dirs.append(entry)
                else:
                    files.append(entry)
            except OSError as e:
                logger.warning("无法获取文件信息: %s 错误: %s", entry_path, e)
                continue

        entries = dirs + files
        return {"success": True, "data": {"path": rel_path, "files": entries}}, 200
    except Exception as e:
        logger.error("列出目录失败: %s", e, exc_info=True)
        return {"success": False, "error": str(e)}, 500


def read_file_response(rel_file: str) -> tuple[dict[str, Any], int]:
    try:
        full_path = resolve_safe_path(rel_file)
        if full_path is None:
            return {"success": False, "error": "路径越权访问被拒绝"}, 403
        if not os.path.exists(full_path):
            return {"success": False, "error": "文件不存在"}, 404
        if os.path.isdir(full_path):
            return {"success": False, "error": "指定路径是目录而非文件"}, 400

        ext = os.path.splitext(full_path)[1].lower()

        if ext in EXCEL_EXTENSIONS:
            try:
                import openpyxl

                wb = openpyxl.load_workbook(full_path, data_only=True)
                sheet_names = list(wb.sheetnames)
                sheets_data: dict[str, Any] = {}
                for sheet_name in sheet_names:
                    sheet = wb[sheet_name]
                    rows_data = []
                    for row in sheet.iter_rows(values_only=True):
                        row_list = []
                        for cell in row:
                            if cell is None:
                                row_list.append(None)
                            elif isinstance(cell, datetime):
                                row_list.append(cell.isoformat())
                            else:
                                row_list.append(cell)
                        rows_data.append(row_list)
                    sheets_data[sheet_name] = {
                        "rows": rows_data,
                        "max_row": sheet.max_row,
                        "max_column": sheet.max_column,
                    }
                wb.close()
                return {
                    "success": True,
                    "data": {"type": "excel", "sheets": sheet_names, "content": sheets_data},
                }, 200
            except ImportError:
                return {"success": False, "error": "openpyxl 未安装，无法读取 Excel 文件"}, 500
            except Exception as e:
                logger.error("读取 Excel 文件失败: %s", e, exc_info=True)
                return {"success": False, "error": f"读取 Excel 失败: {str(e)}"}, 500

        if ext in IMAGE_EXTENSIONS:
            with open(full_path, "rb") as f:
                img_base64 = base64.b64encode(f.read()).decode("utf-8")
            mime_map = {
                ".png": "image/png",
                ".jpg": "image/jpeg",
                ".jpeg": "image/jpeg",
                ".gif": "image/gif",
                ".webp": "image/webp",
                ".bmp": "image/bmp",
                ".ico": "image/x-icon",
                ".tiff": "image/tiff",
                ".svg": "image/svg+xml",
            }
            return {
                "success": True,
                "data": {
                    "type": "image",
                    "mime": mime_map.get(ext, "application/octet-stream"),
                    "content": img_base64,
                },
            }, 200

        try:
            with open(full_path, encoding="utf-8") as f:
                content = f.read()
            return {"success": True, "data": {"type": "text", "content": content}}, 200
        except UnicodeDecodeError:
            with open(full_path, "rb") as f:
                content = base64.b64encode(f.read()).decode("utf-8")
            return {"success": True, "data": {"type": "binary", "content": content}}, 200

    except Exception as e:
        logger.error("读取文件失败: %s", e, exc_info=True)
        return {"success": False, "error": str(e)}, 500


def _build_snapshot() -> dict[str, float]:
    snap: dict[str, float] = {}
    if os.path.exists(ROOT_DIR) and os.path.isdir(ROOT_DIR):
        try:
            for dirpath, _dirnames, filenames in os.walk(ROOT_DIR):
                try:
                    rel = os.path.relpath(dirpath, ROOT_DIR)
                    prefix = "" if rel == "." else rel.replace(os.sep, "/") + "/"
                    for fname in filenames:
                        fpath = os.path.join(dirpath, fname)
                        try:
                            snap[prefix + fname] = os.stat(fpath).st_mtime
                        except OSError:
                            continue
                except OSError:
                    continue
        except OSError:
            pass
    return snap


def _format_snapshot(mtime_map: dict[str, float]) -> dict[str, str]:
    return {name: _format_time(ts) for name, ts in mtime_map.items()}


def _watchdog_loop() -> None:
    global _watchdog_running, _watchdog_prev, _last_snapshot
    _watchdog_running = True
    try:
        _watchdog_prev = _build_snapshot()
        while _watchdog_running:
            threading.Event().wait(3.0)
            if not _watchdog_running:
                break
            curr = _build_snapshot()
            changed: list[str] = []
            for fname, mtime in curr.items():
                if _watchdog_prev.get(fname) != mtime:
                    changed.append(fname)
            for fname in list(_watchdog_prev.keys()):
                if fname not in curr:
                    changed.append(f"__deleted__:{fname}")
            _watchdog_prev = curr
            if changed:
                snap_str = _format_snapshot(curr)
                with _snapshot_lock:
                    _last_snapshot = snap_str
                payload = json.dumps({"changed": changed}, ensure_ascii=False)
                dead: list[queue.Queue[str]] = []
                with _watch_clients_lock:
                    for q in _watch_clients:
                        try:
                            q.put_nowait(payload)
                        except queue.Full:
                            dead.append(q)
                    for q in dead:
                        if q in _watch_clients:
                            _watch_clients.remove(q)
    except Exception as e:
        logger.error("Watchdog 线程异常: %s", e, exc_info=True)
    finally:
        _watchdog_running = False


def _ensure_watchdog() -> None:
    global _watchdog_thread, _watchdog_running
    if not _watchdog_running or (_watchdog_thread and not _watchdog_thread.is_alive()):
        _watchdog_thread = threading.Thread(target=_watchdog_loop, daemon=True, name="fs-watchdog")
        _watchdog_thread.start()


def sse_watch_events(rel_path: str) -> Iterator[str]:
    full_path = resolve_safe_path(rel_path)
    if full_path is None:
        yield f"data: {json.dumps({'success': False, 'error': '路径越权访问被拒绝'}, ensure_ascii=False)}\n\n"
        return

    _ensure_watchdog()

    client_q: queue.Queue[str] = queue.Queue(maxsize=20)
    with _watch_clients_lock:
        _watch_clients.append(client_q)

    try:
        with _snapshot_lock:
            initial = json.dumps(
                {"changed": [], "snapshot": dict(_last_snapshot)}, ensure_ascii=False
            )
        yield f"data: {initial}\n\n"
        while True:
            try:
                msg = client_q.get(timeout=55)
                yield f"data: {msg}\n\n"
            except queue.Empty:
                yield ": heartbeat\n\n"
    finally:
        with _watch_clients_lock:
            if client_q in _watch_clients:
                _watch_clients.remove(client_q)
