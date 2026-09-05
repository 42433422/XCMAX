"""Owner-isolated rules, template and conversion endpoints."""

from __future__ import annotations

from pathlib import Path
from uuid import uuid4
from zipfile import BadZipFile

from fastapi import File, Form, HTTPException, Request, UploadFile
from fastapi.responses import Response

from app.mod_sdk.attendance_roster import read_attendance_roster
from app.mod_sdk.customer_features import require_attendance_conversion
from app.mod_sdk.owner_workspace import owner_workspace
from app.mod_sdk.runtime_frontend import authorized_runtime

from .convert import convert_attendance_file
from .owner_config import MOD_ID, TEMPLATE_NAME, read_policy, save_policy

MAX_UPLOAD_BYTES = 32 * 1024 * 1024


def require_access(request: Request):
    require_attendance_conversion(request)
    authorized_runtime(request, MOD_ID)


async def _upload_bytes(file: UploadFile) -> bytes:
    content = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "文件不能超过 32 MB")
    if not content:
        raise HTTPException(400, "文件内容为空")
    return content


def register(router):
    @router.get("/attendance/rules")
    async def rules():
        workspace = owner_workspace(MOD_ID)
        return {
            "success": True,
            "data": {
                "template_ready": workspace.file_path(TEMPLATE_NAME).is_file(),
                "attendance_policy": read_policy(),
                "roster_count": len(read_attendance_roster()),
                "description": "太阳鸟账号专属模板、班制规则与考勤表转换",
            },
        }

    @router.get("/attendance/policy")
    async def policy_get():
        return {"success": True, "attendance_policy": read_policy()}

    @router.post("/attendance/policy")
    async def policy_post(body: dict):
        value = body.get("attendance_policy")
        if not isinstance(value, dict):
            raise HTTPException(400, "请提供 attendance_policy 对象")
        return {"success": True, "attendance_policy": save_policy(value)}

    @router.post("/attendance/template")
    async def template_upload(file: UploadFile = File(...), replace_existing: bool = Form(False)):
        if Path(file.filename or "").suffix.lower() != ".xlsx":
            raise HTTPException(400, "模板需为 .xlsx 文件")
        content = await _upload_bytes(file)
        # Validate the workbook without executing macros or touching existing data.
        from io import BytesIO

        from openpyxl import load_workbook

        try:
            workbook = load_workbook(BytesIO(content), read_only=True)
            valid = "明细" in workbook.sheetnames
            workbook.close()
        except (OSError, ValueError, KeyError, BadZipFile):
            raise HTTPException(400, "模板不是有效的 Excel 工作簿") from None
        if not valid:
            raise HTTPException(400, "模板必须包含明细工作表")
        workspace = owner_workspace(MOD_ID)
        workspace.root.mkdir(parents=True, exist_ok=True)
        target = workspace.file_path(TEMPLATE_NAME)
        if target.exists() and not replace_existing:
            raise HTTPException(409, "当前账号已有模板，请明确确认替换")
        temporary = workspace.file_path(f".template-{uuid4().hex}.xlsx")
        temporary.write_bytes(content)
        try:
            if replace_existing:
                temporary.replace(target)
            else:
                import os

                try:
                    os.link(temporary, target)
                except FileExistsError:
                    raise HTTPException(409, "当前账号已有模板，请明确确认替换") from None
        finally:
            temporary.unlink(missing_ok=True)
        return {"success": True, "message": "模板已保存"}

    @router.post("/attendance/convert-upload")
    async def convert_upload(
        file: UploadFile = File(...),
        month: str = Form(""),
        header_row: int = Form(0),
    ):
        suffix = Path(file.filename or "").suffix.lower()
        if suffix not in {".xlsx", ".xlsm", ".xls"}:
            raise HTTPException(400, "不支持此考勤文件类型")
        workspace = owner_workspace(MOD_ID)
        try:
            template = workspace.existing_file(TEMPLATE_NAME)
        except (OSError, ValueError):
            raise HTTPException(409, "请先安装或上传当前账号的考勤模板") from None
        roster = read_attendance_roster()
        if not roster:
            raise HTTPException(409, "请先在共享考勤工作区维护当前账号的人员名单")
        content = await _upload_bytes(file)
        workspace.root.mkdir(parents=True, exist_ok=True)
        source = workspace.file_path(f"upload-{uuid4().hex}{suffix}")
        output = workspace.file_path(f"output-{uuid4().hex}.xlsx")
        source.write_bytes(content)
        result = convert_attendance_file(
            str(source),
            str(output),
            template_path=str(template),
            month=month or None,
            header_row=header_row,
            use_llm=False,
            personnel_roster=roster,
        )
        if not result.get("success"):
            raise HTTPException(422, str(result.get("error") or "考勤转换失败"))
        result.pop("input", None)
        result.pop("output", None)
        result["download_path"] = f"/api/mod/{MOD_ID}/attendance/download?file={output.name}"
        return {"success": True, "data": result}

    @router.get("/attendance/download")
    async def download(file: str):
        if not file.startswith("output-") or not file.endswith(".xlsx"):
            raise HTTPException(404, "文件不存在")
        try:
            path = owner_workspace(MOD_ID).existing_file(file)
            content = path.read_bytes()
        except (OSError, ValueError):
            raise HTTPException(404, "当前账号无此输出文件") from None
        return Response(
            content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{path.name}"',
                "Cache-Control": "no-store",
            },
        )
