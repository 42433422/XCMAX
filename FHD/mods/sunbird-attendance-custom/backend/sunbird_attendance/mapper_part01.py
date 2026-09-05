"""Implementation extracted from the public facade module."""

from __future__ import annotations

import importlib


def _facade():
    return importlib.import_module("sunbird_attendance.mapper")


def _chinese_block_label_to_int(text: str) -> int | None:
    """块内左侧标签常见「一…三十一」纯中文数字串 → 整数；含其它汉字则不改。"""
    if not text:
        return None
    text = _facade().unicodedata.normalize("NFKC", text).strip()
    if not text:
        return None
    allowed = set(_facade()._CN_BLOCK_DIGITS) | {"十"}
    if any(c not in allowed for c in text):
        return None
    d = _facade()._CN_BLOCK_DIGITS
    if len(text) == 1:
        if text in d:
            return d[text]
        if text == "十":
            return 10
        return None
    if len(text) == 2:
        if text[0] == "十" and text[1] in d:
            return 10 + d[text[1]]
        if text[1] == "十" and text[0] in d:
            return d[text[0]] * 10
    if len(text) == 3 and text[1] == "十" and (text[0] in d) and (text[2] in d):
        return d[text[0]] * 10 + d[text[2]]
    return None


def _plain_cell_text(value: object) -> str:
    """与 ``str()`` 相比，显式支持 openpyxl 的富文本单元格（拼接各段，便于匹配 ``18:30记加班``）。"""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    try:
        from openpyxl.cell.rich_text import CellRichText

        if isinstance(value, CellRichText):
            return "".join(str(part) for part in value)
    except ImportError:
        pass
    return str(value)


def _replace_line_all_cn_numerals(ls: str) -> tuple[str, bool]:
    """把一行里每一段「仅由数字用字组成」的串按最长可解析规则换成阿拉伯数字（含无换行的「二一」）。"""
    if not ls:
        return (ls, False)
    if ls.strip() == "〇":
        return (ls, False)
    if ls[0] == "〇" and len(ls.strip()) > 1:
        return (ls, False)
    allowed = set(_facade()._CN_BLOCK_DIGITS) | {"十"}
    out: list[str] = []
    i = 0
    changed = False
    while i < len(ls):
        if ls[i] not in allowed:
            out.append(ls[i])
            i += 1
            continue
        j = i
        while j < len(ls) and ls[j] in allowed:
            j += 1
        run = ls[i:j]
        tup: tuple[int, int] | None = None
        for take in range(len(run), 0, -1):
            prefix = run[:take]
            n = _facade()._chinese_block_label_to_int(prefix)
            if n is None:
                continue
            if n == 0 and prefix == "〇":
                break
            tup = (n, take)
            break
        if tup is None:
            out.append(ls[i])
            i += 1
            continue
        n, take = tup
        out.append(str(n))
        i += take
        changed = True
    return ("".join(out), changed)


def _cell_text_replace_chinese_numerals(text: str) -> str | None:
    """多行、富文本拼成的单行、或「中文数+符号」混排：替换行内所有可解析的中文数字段。"""
    text = _facade().unicodedata.normalize("NFKC", text)
    lines = text.splitlines()
    out: list[str] = []
    changed = False
    for line in lines:
        ls = line.lstrip(" \t")
        indent_len = len(line) - len(ls)
        indent = line[:indent_len]
        if not ls:
            out.append(line)
            continue
        nl, _ = _facade()._replace_line_all_cn_numerals(ls)
        new_line = indent + nl
        if new_line != line:
            changed = True
        out.append(new_line)
    if not changed:
        return None
    return "\n".join(out)


def _merge_anchor_for_cell(
    row: int, col: int, overlaps: list[tuple[int, int, int, int]]
) -> tuple[int, int]:
    """在预筛选的合并矩形列表中解析 (row,col) 所属合并的左上角。"""
    for mr1, mc1, mr2, mc2 in overlaps:
        if mr1 <= row <= mr2 and mc1 <= col <= mc2:
            return (mr1, mc1)
    return (row, col)


def _normalize_chinese_numerals_in_rect(ws, r_lo: int, r_hi: int, col_lo: int, col_hi: int) -> None:
    """矩形 [r_lo..r_hi]×[col_lo..col_hi] 内：中文数字改为阿拉伯数字（合并格写左上角）。"""
    overlaps: list[tuple[int, int, int, int]] = []
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = _facade().range_boundaries(str(rng))
        if max_row < r_lo or min_row > r_hi or max_col < col_lo or (min_col > col_hi):
            continue
        overlaps.append((min_row, min_col, max_row, max_col))
    processed_merge_tops: set[tuple[int, int]] = set()
    for row in range(r_lo, r_hi + 1):
        for col in range(col_lo, col_hi + 1):
            wr, wc = _facade()._merge_anchor_for_cell(row, col, overlaps)
            key = (wr, wc)
            if key in processed_merge_tops:
                continue
            processed_merge_tops.add(key)
            v = ws.cell(wr, wc).value
            if v is None:
                continue
            if isinstance(v, (int, float)) and type(v) is not bool:
                continue
            text = _facade()._plain_cell_text(v).strip().replace("\u3000", "").strip()
            if not text:
                continue
            text = _facade().unicodedata.normalize("NFKC", text).strip()
            if not text:
                continue
            if _facade().re.fullmatch("-?\\d+\\.", text):
                try:
                    coerced = float(text)
                except ValueError:
                    coerced = None
                if coerced is not None:
                    c = ws.cell(wr, wc)
                    c.value = coerced
                    if (
                        _facade().DETAIL_SUM_COL_START
                        <= wc
                        < _facade().DETAIL_TEMPLATE_SUMMARY_BEGIN_COL
                    ):
                        _facade()._force_arabic_number_format(c)
                    continue
            if text.startswith("="):
                continue
            if text in _facade()._ATT_MARK:
                continue
            n = _facade()._chinese_block_label_to_int(text)
            if n is not None:
                ws.cell(wr, wc).value = n
                continue
            mixed = _facade()._cell_text_replace_chinese_numerals(text)
            if mixed is not None:
                ws.cell(wr, wc).value = mixed


def _normalize_block_chinese_numerals(
    ws, block_top: int, col_lo: int = 1, col_hi: int = _facade().DETAIL_MAX_COL
) -> None:
    """每人 6 行块内 [col_lo..col_hi]：纯中文数字改为阿拉伯数字（合并格写左上角）。

    整格为考勤符号（如请假「〇」）的不改。合并区只预筛与本块矩形相交的范围，避免全表 merged 扫描卡死。
    """
    r_hi = block_top + _facade().DETAIL_PERSON_BLOCK_ROWS - 1
    _facade()._normalize_chinese_numerals_in_rect(ws, block_top, r_hi, col_lo, col_hi)


@_facade().dataclass
class DayBandEntry:
    symbol: str
    value: float


@_facade().dataclass
class EmployeeDayTemplateData:
    work_date: _facade().date
    morning: list[_facade().DayBandEntry] = _facade().field(default_factory=list)
    afternoon: list[_facade().DayBandEntry] = _facade().field(default_factory=list)
    night: list[_facade().DayBandEntry] = _facade().field(default_factory=list)
    notes: list[str] = _facade().field(default_factory=list)


@_facade().dataclass
class EmployeeMonthTemplateData:
    employee_name: str
    attendance_group: str
    department: str
    employee_no: str
    days: dict[int, _facade().EmployeeDayTemplateData] = _facade().field(default_factory=dict)
    normal_hours: float = 0.0
    weekday_overtime_hours: float = 0.0
    sunday_overtime_hours: float = 0.0
    leave_hours: float = 0.0
    absent_hours: float = 0.0
    late_count: float = 0.0
    early_count: float = 0.0
    warnings: list[str] = _facade().field(default_factory=list)


@_facade().dataclass
class TemplateWriteResult:
    matched_employee_count: int
    unmatched_employee_names: list[str]


@_facade().dataclass(frozen=True)
class TemplateEmployeeProfile:
    employee_name: str
    base_row: int
    department: str
    nature_text: str
    block_values: tuple[float, float, float, float]
    overtime_start: _facade().time
    morning_work_start: _facade().time | None = None
