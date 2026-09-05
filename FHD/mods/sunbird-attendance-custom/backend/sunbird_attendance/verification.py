"""Exercise the installed converter without writing any customer business data."""

import sqlite3
from pathlib import Path
from tempfile import TemporaryDirectory

from fastapi import Request
from openpyxl import Workbook, load_workbook

from app.mod_sdk.attendance_roster import read_attendance_roster
from app.mod_sdk.owner_workspace import (
    attendance_database_path,
    authenticated_owner,
    owner_context,
)

from .convert import convert_attendance_file

CASE_ID = "sunbird-owner-conversion-v1"
SAMPLE_NAME = "交付验证样例"


def write_conversion_sample(directory: Path) -> tuple[Path, Path]:
    """Public synthetic data only; never copies a customer template or roster."""
    source = directory / "sample.xlsx"
    template = directory / "template.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "每日统计"
    sheet.append(
        [
            "姓名",
            "日期",
            "考勤组",
            "部门",
            "上班1打卡时间",
            "下班1打卡时间",
            "上班2打卡时间",
            "下班2打卡时间",
        ]
    )
    sheet.append(
        [
            SAMPLE_NAME,
            "2026-09-01",
            "验证组",
            "验证部门",
            "08:00",
            "12:00",
            "13:30",
            "17:30",
        ]
    )
    workbook.save(source)
    workbook.close()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "明细"
    sheet.cell(1, 1, "交付验证合成模板")
    sheet.cell(3, 70, "正班")
    sheet.cell(4, 3, SAMPLE_NAME)
    for row in range(4, 10):
        sheet.cell(row, 4, "验证")
    workbook.save(template)
    workbook.close()
    return source, template


def verify_conversion(request: Request) -> dict:
    owner = authenticated_owner(request)
    with owner_context(owner):
        database = attendance_database_path()
        schema_ready = False
        if database.is_file():
            with sqlite3.connect(f"{database.as_uri()}?mode=ro", uri=True) as connection:
                columns = {
                    row[1] for row in connection.execute("PRAGMA table_info(attendance_employees)")
                }
                schema_ready = {
                    "id",
                    "employee_name",
                    "department",
                    "position",
                } <= columns
        if not schema_ready:
            return {"pending": True, "reason": "workspace_not_ready", "case_id": CASE_ID}
        # Reads the same owner-scoped contract used by the actual conversion API.
        roster = read_attendance_roster()
        with TemporaryDirectory(prefix="sunbird-delivery-probe-") as temporary:
            directory = Path(temporary)
            source, template = write_conversion_sample(directory)
            output = directory / "converted.xlsx"
            result = convert_attendance_file(
                str(source),
                str(output),
                template_path=str(template),
                month="2026-09",
                use_llm=False,
                personnel_roster=[("验证部门", "验证岗位", SAMPLE_NAME)],
            )
            if not result.get("success"):
                return {
                    "passed": False,
                    "case_id": CASE_ID,
                    "observations": {
                        "conversion_error": str(result.get("error") or "转换未成功"),
                        "owner_schema_ready": schema_ready,
                    },
                }
            workbook = load_workbook(output, read_only=True, data_only=False)
            try:
                sheets = workbook.sheetnames
                name_matches = workbook["明细"].cell(4, 3).value == SAMPLE_NAME
                # A real generated monthly formula must refer back to detail totals.
                linked_totals = any(
                    isinstance(cell.value, str)
                    and cell.value.startswith("=")
                    and "明细" in cell.value
                    for row in workbook["月度统计"]
                    for cell in row
                )
            finally:
                workbook.close()
    observations = {
        "owner_schema_ready": schema_ready,
        "owner_roster_count": len(roster),
        "rows_parsed": result.get("rows_in"),
        "rows_matched": result.get("rows_used_for_template"),
        "employees_matched": result.get("employees_matched"),
        "output_sheets": sheets,
        "sample_name_matches": name_matches,
        "monthly_formulas_link_detail": linked_totals,
        "external_model_used": result.get("used_llm"),
        "customer_data_written": False,
    }
    passed = bool(
        schema_ready
        and name_matches
        and linked_totals
        and result.get("rows_used_for_template") == 1
        and result.get("employees_matched") == 1
        and sheets == ["明细", "月度统计"]
        and result.get("used_llm") is False
    )
    return {"passed": passed, "case_id": CASE_ID, "observations": observations}
