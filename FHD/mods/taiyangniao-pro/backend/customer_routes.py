"""太阳鸟 PRO — 部门/客户管理 FastAPI 路由（自 blueprints.py 拆出，行为零变更）。"""

from __future__ import annotations

from pathlib import Path

from fastapi import File, UploadFile
from fastapi.responses import JSONResponse


def register(router, *, logger, get_database_path, _load_products_personnel_roster_from_host) -> None:
    """在给定 router 上注册 /customers*、/purchase_units、shipment 单位兼容路由。"""

    @router.get("/customers", response_model=None)
    @router.get("/customers/", response_model=None, include_in_schema=False)
    async def customers_all(page: int = 1, per_page: int = 20, keyword: str = ""):
        return await customers_list(page=page, per_page=per_page, keyword=keyword)

    @router.get("/customers/list", response_model=None)
    async def customers_list(
        page: int = 1,
        per_page: int = 20,
        keyword: str = "",
        purchase_unit: str = "",
    ):
        import sqlite3

        db_path = get_database_path()
        if not db_path.exists():
            return {"success": True, "data": [], "total": 0}
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cond = []
        args = []
        if keyword:
            cond.append("(customer_name LIKE ? OR contact_person LIKE ? OR contact_phone LIKE ?)")
            args.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
        if purchase_unit:
            cond.append("purchase_unit = ?")
            args.append(purchase_unit)
        where = " AND ".join(cond) if cond else "1=1"
        cur.execute(f"SELECT COUNT(*) FROM customers WHERE {where}", args)
        total = cur.fetchone()[0]
        offset = (page - 1) * per_page
        cur.execute(
            f"SELECT id, customer_name, contact_person, contact_phone, address, purchase_unit "
            f"FROM customers WHERE {where} ORDER BY id LIMIT ? OFFSET ?",
            [*args, per_page, offset],
        )
        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return {"success": True, "data": items, "total": total}

    def _distinct_customer_purchase_units() -> list[str]:
        import sqlite3

        db_path = get_database_path()
        if not db_path.exists():
            return []
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        cur.execute(
            "SELECT DISTINCT TRIM(purchase_unit) FROM customers "
            "WHERE purchase_unit IS NOT NULL AND TRIM(purchase_unit) != '' "
            "ORDER BY purchase_unit COLLATE NOCASE"
        )
        units = [str(row[0]).strip() for row in cur.fetchall() if row and str(row[0]).strip()]
        conn.close()
        return units

    def _distinct_attendance_departments() -> list[str]:
        """Return real departments from the personnel and attendance stores."""
        import sqlite3

        values: set[str] = {
            str(department).strip()
            for department, _specification, _name in _load_products_personnel_roster_from_host()
            if str(department).strip()
        }
        db_path = get_database_path()
        if db_path.exists():
            conn = sqlite3.connect(str(db_path))
            cur = conn.cursor()
            try:
                table_names = {
                    str(row[0])
                    for row in cur.execute(
                        "SELECT name FROM sqlite_master WHERE type = 'table'"
                    ).fetchall()
                }
                if "attendance_departments" in table_names:
                    for row in cur.execute(
                        "SELECT department, main_department FROM attendance_departments"
                    ).fetchall():
                        values.update(
                            str(value).strip() for value in row if str(value or "").strip()
                        )
                if "attendance_employees" in table_names:
                    for row in cur.execute(
                        "SELECT department, main_department FROM attendance_employees"
                    ).fetchall():
                        values.update(
                            str(value).strip() for value in row if str(value or "").strip()
                        )
            except sqlite3.Error:
                logger.exception("读取太阳鸟考勤部门失败")
            finally:
                conn.close()
        if not values:
            values.update(_distinct_customer_purchase_units())
        return sorted(values, key=str.casefold)

    @router.get("/purchase_units")
    @router.get("/purchase_units/", include_in_schema=False)
    async def purchase_units_list():
        units = _distinct_customer_purchase_units()
        return {"success": True, "data": units}

    @router.get("/shipment/shipment-records/units")
    @router.get("/shipment/shipment-records/units/", include_in_schema=False)
    async def shipment_record_units_compat():
        units = _distinct_attendance_departments()
        return {"success": True, "data": units, "units": units}

    @router.get("/customers/{customer_id}", response_model=None)
    async def customers_get(customer_id: int):
        import sqlite3

        db_path = get_database_path()
        if not db_path.exists():
            return JSONResponse({"success": False, "error": "not found"}, status_code=404)
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE id = ?", (customer_id,))
        row = cur.fetchone()
        conn.close()
        if not row:
            return JSONResponse({"success": False, "error": "not found"}, status_code=404)
        return {"success": True, "data": dict(row)}

    @router.post("/customers", response_model=None)
    async def customers_add(data: dict):
        import datetime
        import sqlite3

        db_path = get_database_path()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        now = datetime.datetime.now().isoformat()
        cur.execute(
            "INSERT INTO customers (source_file, customer_name, contact_person, contact_phone, address, purchase_unit, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                data.get("source_file", ""),
                data.get("customer_name", ""),
                data.get("contact_person", ""),
                data.get("contact_phone", ""),
                data.get("address", ""),
                data.get("purchase_unit", ""),
                now,
                now,
            ),
        )
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        return {"success": True, "data": {"id": new_id}}

    @router.put("/customers/{customer_id}", response_model=None)
    async def customers_update(customer_id: int, data: dict):
        import datetime
        import sqlite3

        db_path = get_database_path()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        now = datetime.datetime.now().isoformat()
        cur.execute(
            "UPDATE customers SET customer_name=?, contact_person=?, contact_phone=?, address=?, purchase_unit=?, updated_at=? WHERE id=?",
            (
                data.get("customer_name", ""),
                data.get("contact_person", ""),
                data.get("contact_phone", ""),
                data.get("address", ""),
                data.get("purchase_unit", ""),
                now,
                customer_id,
            ),
        )
        conn.commit()
        conn.close()
        return {"success": True}

    @router.delete("/customers/{customer_id}", response_model=None)
    async def customers_delete(customer_id: int):
        import sqlite3

        db_path = get_database_path()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        cur.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
        conn.commit()
        conn.close()
        return {"success": True}

    @router.post("/customers/batch-delete", response_model=None)
    async def customers_batch_delete(data: dict):
        import sqlite3

        ids = data.get("ids") or []
        if not ids:
            return {"success": True}
        db_path = get_database_path()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        placeholders = ",".join("?" * len(ids))
        cur.execute(f"DELETE FROM customers WHERE id IN ({placeholders})", ids)
        conn.commit()
        conn.close()
        return {"success": True}

    @router.post("/customers/import", response_model=None)
    async def customers_import(file: UploadFile = File(...)):
        import datetime
        import shutil
        import sqlite3
        import tempfile

        import openpyxl

        db_path = get_database_path()
        suffix = Path(file.filename or "import.xlsx").suffix.lower()
        if suffix not in {".xlsx", ".xlsm", ".xls"}:
            return JSONResponse(
                {"success": False, "error": "unsupported file type"}, status_code=400
            )
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in ws[1]]
            name_idx = headers.index("客户名称") if "客户名称" in headers else 0
            contact_idx = headers.index("联系人") if "联系人" in headers else -1
            phone_idx = headers.index("电话") if "电话" in headers else -1
            addr_idx = headers.index("地址") if "地址" in headers else -1
            now = datetime.datetime.now().isoformat()
            conn = sqlite3.connect(str(db_path))
            cur = conn.cursor()
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
                if not name:
                    continue
                contact = (
                    str(row[contact_idx]) if contact_idx >= 0 and contact_idx < len(row) else ""
                )
                phone = str(row[phone_idx]) if phone_idx >= 0 and phone_idx < len(row) else ""
                addr = str(row[addr_idx]) if addr_idx >= 0 and addr_idx < len(row) else ""
                cur.execute(
                    "INSERT INTO customers (source_file, customer_name, contact_person, contact_phone, address, purchase_unit, created_at, updated_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (file.filename or "", name, contact, phone, addr, "", now, now),
                )
                count += 1
            conn.commit()
            conn.close()
            return {"success": True, "imported": count}
        finally:
            Path(tmp_path).unlink(missing_ok=True)
