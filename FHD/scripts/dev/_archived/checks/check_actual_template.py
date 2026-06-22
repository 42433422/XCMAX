"""检查实际使用的模板文件"""

import os
from pathlib import Path
from openpyxl import load_workbook

# 设置 WORKSPACE_ROOT
os.environ["WORKSPACE_ROOT"] = r"e:\FHD"

from app.shell.taiyangniao_attendance.paths import resolve_workspace_excel

# 测试模板路径
template_relpath = "424/考勤 -2026-3 月份考勤统计表.xlsx"

try:
    template_path = resolve_workspace_excel(template_relpath)
    print(f"✓ 模板路径解析成功：{template_path}")
    print(f"  文件存在：{template_path.is_file()}")

    if template_path.is_file():
        wb = load_workbook(str(template_path))
        print(f"  工作表：{wb.sheetnames}")

        for i, ws_name in enumerate(wb.sheetnames):
            ws = wb[ws_name]
            merged_count = len(list(ws.merged_cells.ranges)) if ws.merged_cells.ranges else 0
            print(f"    [{i+1}] {ws_name}: {merged_count} 个合并单元格")
            if merged_count > 0:
                for m in ws.merged_cells.ranges:
                    print(f"        {m}")

        wb.close()
    else:
        print(f"✗ 模板文件不存在")

except Exception as e:
    print(f"✗ 错误：{e}")

# 列出 424 目录下所有 Excel 文件
print("\n\n424 目录下的 Excel 文件:")
dir_424 = Path(r"e:\FHD\424")
for f in sorted(dir_424.glob("*.xlsx")):
    print(f"  {f.name} ({f.stat().st_size} 字节)")
