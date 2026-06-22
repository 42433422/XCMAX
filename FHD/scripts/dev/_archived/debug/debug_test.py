"""调试测试"""

import sys

sys.path.insert(0, r"e:\FHD\backend\shell")

from pathlib import Path
from taiyangniao_attendance.convert import _write_workbook_to_path
import pandas as pd

# 创建简单的测试数据
stats_df = pd.DataFrame(
    {
        "考勤组": ["公司 - 考勤"],
        "工号": ["001"],
        "姓名": ["张三"],
        "部门": ["技术部"],
        "日期": ["2026-03-01"],
        "上班打卡": [pd.Timestamp("2026-03-01 09:00")],
        "下班打卡": [pd.Timestamp("2026-03-01 17:30")],
        "工作时长": [8.0],
        "考勤结果": ["正常"],
        "周类型": ["大周"],
        "是否周六": [False],
        "周六工作有效": [False],
        "规则违规": [""],
        "备注": [""],
    }
)

dingtalk_df = pd.DataFrame(
    {
        "姓名": ["张三"],
        "工号": ["001"],
        "部门": ["技术部"],
        "日期": ["2026-03-01"],
        "上班打卡": [pd.Timestamp("2026-03-01 09:00")],
        "下班打卡": [pd.Timestamp("2026-03-01 17:30")],
        "工作时长": [8.0],
    }
)

template_path = Path(r"e:\FHD\424\考勤 -2026-3 月份考勤统计表.xlsx")  # 无空格
output_path = Path(r"e:\FHD\424\调试输出 2.xlsx")

print("=== 调试 _write_workbook_to_path ===")
print(f"模板路径：{template_path}")
print(f"模板路径.is_file(): {template_path.is_file()}")

# 手动执行函数中的逻辑
import os

template_file_path = None
if template_path is not None:
    if template_path.is_file():
        template_file_path = str(template_path)
        print(f"直接访问成功：{template_file_path}")
    else:
        print("直接访问失败，尝试目录列表...")
        try:
            parent_dir = str(template_path.parent)
            file_name = template_path.name
            dir_contents = os.listdir(parent_dir)
            print(f"目录内容：{len(dir_contents)} 个文件")
            if file_name in dir_contents:
                template_file_path = os.path.join(parent_dir, file_name)
                print(f"目录列表成功：{template_file_path}")
            else:
                print(f"文件名 {file_name} 不在目录中")
        except Exception as e:
            print(f"目录列表失败：{e}")

print(f"\n最终 template_file_path: {template_file_path}")

if template_file_path:
    try:
        from openpyxl import load_workbook

        wb = load_workbook(template_file_path)
        print(f"✓ 加载成功！工作表：{wb.sheetnames}")
    except Exception as e:
        print(f"✗ 加载失败：{e}")
else:
    print("模板文件路径为 None，将使用无模板模式")

# 调用函数
print("\n调用函数...")
_write_workbook_to_path(
    stats_df, output_path, template_path=template_path, dingtalk_detail=dingtalk_df
)
print(f"完成！输出文件：{output_path}")

# 检查输出
if output_path.exists():
    wb_out = load_workbook(str(output_path))
    print(f"输出文件工作表：{wb_out.sheetnames}")
