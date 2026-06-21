import os
from openpyxl import load_workbook

folder = r"e:\FHD\424"

# 方法：直接使用 os.listdir 找到的文件名，不重新构建路径
dir_contents = os.listdir(folder)
target_file = None
for f in dir_contents:
    if f == "考勤 -2026-3 月份考勤统计表.xlsx":
        target_file = f
        break

if target_file:
    print(f"找到文件：{target_file}")
    print(f"repr: {repr(target_file)}")

    # 方法 1: 切换到目录，然后用相对路径
    old_cwd = os.getcwd()
    try:
        os.chdir(folder)
        print(f"\n当前目录：{os.getcwd()}")

        wb = load_workbook(target_file)
        print(f"✓ 加载成功！工作表：{wb.sheetnames}")

        # 保存
        wb.save("测试直接加载.xlsx")
        print(f"✓ 保存成功！")

        # 验证
        wb2 = load_workbook("测试直接加载.xlsx")
        print(f"验证 - 工作表：{wb2.sheetnames}")
    except Exception as e:
        print(f"✗ 失败：{e}")
        import traceback

        traceback.print_exc()
    finally:
        os.chdir(old_cwd)
else:
    print("未找到文件")
