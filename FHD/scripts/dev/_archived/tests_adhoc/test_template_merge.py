"""测试模板合并单元格保留功能"""

import sys
from pathlib import Path
from openpyxl import Workbook, load_workbook
import pandas as pd

# 创建一个带合并单元格的模板
template_path = Path(r"e:\FHD\424\test_template_with_merge.xlsx")
output_path = Path(r"e:\FHD\424\test_output.xlsx")

print("1. 创建带合并单元格的模板...")
wb = Workbook()
ws = wb.active
ws.title = "封面"

# 添加合并单元格
ws.merge_cells("A1:D1")
ws.cell(row=1, column=1, value="考勤统计表封面")
ws.cell(row=1, column=1).alignment = ws.cell(row=1, column=1).alignment.copy(horizontal="center")

ws.merge_cells("A2:B2")
ws.cell(row=2, column=1, value="公司名称")
ws.cell(row=2, column=3, value="XXX 公司")

ws.cell(row=3, column=1, value="统计月份")
ws.cell(row=3, column=3, value="2026-03")

# 添加第二个工作表（作为参考表）
ws2 = wb.create_sheet(title="参考表")
ws2.cell(row=1, column=1, value="参考数据")
ws2.merge_cells("A2:C2")
ws2.cell(row=2, column=1, value="这是合并的单元格")

wb.save(str(template_path))
wb.close()

print(f"   模板已保存：{template_path}")
print(f"   文件大小：{template_path.stat().st_size} 字节")

# 检查模板
wb_check = load_workbook(str(template_path))
print(f"   模板工作表：{wb_check.sheetnames}")
for ws_name in wb_check.sheetnames:
    ws_temp = wb_check[ws_name]
    merged_count = len(list(ws_temp.merged_cells.ranges)) if ws_temp.merged_cells.ranges else 0
    print(f"     - {ws_name}: {merged_count} 个合并单元格")
wb_check.close()

# 模拟转换过程
print("\n2. 模拟转换过程...")

# 创建测试数据
stats_data = {
    "考勤组": ["考勤组 A", "考勤组 B"],
    "工号": ["001", "002"],
    "姓名": ["张三", "李四"],
    "部门": ["销售部", "生产部"],
    "日期": ["2026-03-01", "2026-03-01"],
    "上班打卡": ["09:00", "08:30"],
    "下班打卡": ["18:00", "17:30"],
}
stats_df = pd.DataFrame(stats_data)

detail_data = {
    "姓名": ["张三", "李四"],
    "日期": ["2026-03-01", "2026-03-01"],
    "上班打卡": ["09:00", "08:30"],
    "下班打卡": ["18:00", "17:30"],
}
detail_df = pd.DataFrame(detail_data)

# 使用修复后的逻辑
from openpyxl import load_workbook

wb = load_workbook(str(template_path))

print(f"   加载模板后工作表：{wb.sheetnames}")
for ws_name in wb.sheetnames:
    ws_temp = wb[ws_name]
    merged_count = len(list(ws_temp.merged_cells.ranges)) if ws_temp.merged_cells.ranges else 0
    print(f"     - {ws_name}: {merged_count} 个合并单元格")

# 删除已存在的工作表
for sheet_name in ["月度统计", "钉钉解析"]:
    if sheet_name in wb.sheetnames:
        print(f"   删除已有工作表：{sheet_name}")
        del wb[sheet_name]

# 创建新的工作表
print("   创建新工作表...")
ws_stats = wb.create_sheet(title="月度统计")
for col_idx, col_name in enumerate(stats_df.columns, 1):
    cell = ws_stats.cell(row=1, column=col_idx, value=col_name)
    cell.font = cell.font.copy(bold=True)
for row_idx in range(len(stats_df)):
    for col_idx in range(len(stats_df.columns)):
        ws_stats.cell(row=row_idx + 2, column=col_idx + 1, value=stats_df.iloc[row_idx, col_idx])

ws_detail = wb.create_sheet(title="钉钉解析")
for col_idx, col_name in enumerate(detail_df.columns, 1):
    cell = ws_detail.cell(row=1, column=col_idx, value=col_name)
    cell.font = cell.font.copy(bold=True)
for row_idx in range(len(detail_df)):
    for col_idx in range(len(detail_df.columns)):
        ws_detail.cell(row=row_idx + 2, column=col_idx + 1, value=detail_df.iloc[row_idx, col_idx])

# 保存
wb.save(str(output_path))
wb.close()

print(f"   输出已保存：{output_path}")
print(f"   文件大小：{output_path.stat().st_size} 字节")

# 检查输出
print("\n3. 检查输出文件...")
wb_out = load_workbook(str(output_path))
print(f"   输出工作表：{wb_out.sheetnames}")
for ws_name in wb_out.sheetnames:
    ws_temp = wb_out[ws_name]
    merged_count = len(list(ws_temp.merged_cells.ranges)) if ws_temp.merged_cells.ranges else 0
    print(f"     - {ws_name}: {merged_count} 个合并单元格")
    if merged_count > 0:
        for m in ws_temp.merged_cells.ranges:
            print(f"       {m}")

wb_out.close()

print("\n✓ 测试完成！")
print(f"\n模板路径：{template_path}")
print(f"输出路径：{output_path}")
print("\n请手动打开这两个文件对比，确认合并单元格是否保留！")
