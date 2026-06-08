import openpyxl
from openpyxl import load_workbook
import os

folder = r"e:\FHD\424"

# 查看钉钉数据文件
print("=== 钉钉数据文件：钉钉导出来的考勤数据.xlsx ===")
dingtalk_file = os.path.join(folder, "钉钉导出来的考勤数据.xlsx")
wb_dingtalk = load_workbook(dingtalk_file)
print(f"工作表：{wb_dingtalk.sheetnames}")
for sheet_name in wb_dingtalk.sheetnames:
    ws = wb_dingtalk[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"最大行：{ws.max_row}, 最大列：{ws.max_column}")
    print("前 5 行内容:")
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row)), 1):
        values = [cell.value for cell in row if cell.value is not None]
        if values:
            print(f"  Row {row_idx}: {values[:10]}...")  # 只显示前 10 列

# 查看模板文件
print("\n\n=== 模板文件：考勤 -2026-3 月份考勤统计表.xlsx ===")
template_file = os.path.join(folder, "考勤 -2026-3 月份考勤统计表.xlsx")
wb_template = load_workbook(template_file)
print(f"工作表：{wb_template.sheetnames}")
for sheet_name in wb_template.sheetnames:
    ws = wb_template[sheet_name]
    print(f"\n=== Sheet: {sheet_name} ===")
    print(f"最大行：{ws.max_row}, 最大列：{ws.max_column}")
    print("表头 (前 3 行):")
    for row_idx in range(1, min(4, ws.max_row + 1)):
        values = []
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                values.append(f"{cell.coordinate}:{cell.value}")
        if values:
            print(f"  Row {row_idx}: {values}")

    # 查找姓名列的位置
    print("\n查找员工姓名区域:")
    for row_idx in range(4, min(40, ws.max_row + 1)):
        for col_idx in range(1, min(20, ws.max_column + 1)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if (
                cell.value
                and isinstance(cell.value, str)
                and len(cell.value) > 1
                and "勇" in cell.value
                or "陶" in cell.value
            ):
                print(f"  找到姓名：{cell.coordinate}={cell.value}")
                break
