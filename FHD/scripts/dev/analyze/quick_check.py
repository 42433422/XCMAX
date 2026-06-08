"""快速检查输出文件"""

from openpyxl import load_workbook
import os
import tempfile
import shutil

output_file = r"e:\FHD\424\考勤转换输出.xlsx"

if os.path.exists(output_file):
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, "temp.xlsx")
    shutil.copy2(output_file, temp_file)
    try:
        wb = load_workbook(temp_file)
        print(f"工作表：{wb.sheetnames}")
    finally:
        os.remove(temp_file)
        os.rmdir(temp_dir)
else:
    print("文件不存在")
