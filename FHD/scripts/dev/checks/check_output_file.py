"""检查输出文件"""

from openpyxl import load_workbook
import os

output_file = r"e:\FHD\424\测试上传输出.xlsx"

print(f"检查输出文件：{output_file}")
print(f"文件存在：{os.path.exists(output_file)}")

if os.path.exists(output_file):
    wb = load_workbook(output_file)
    print(f"工作表：{wb.sheetnames}")

    if "明细" in wb.sheetnames:
        ws = wb["明细"]
        print(f'\n✓ 成功！包含"明细"工作表')
        print(f"最大行：{ws.max_row}, 最大列：{ws.max_column}")

        # 检查有多少行有数据
        data_rows = 0
        for row_idx in range(4, ws.max_row + 1):
            # 检查 C 列（姓名）是否有值
            name_cell = ws.cell(row=row_idx, column=3).value
            if name_cell:
                data_rows += 1

        print(f"有数据的行数：{data_rows}")

        # 显示前几个员工
        print(f"\n前 5 个员工:")
        shown_count = 0
        for row_idx in range(4, ws.max_row + 1):
            name_cell = ws.cell(row=row_idx, column=3).value
            if name_cell and shown_count < 5:
                dept = ws.cell(row=row_idx, column=1).value
                print(f"  Row {row_idx}: {dept} - {name_cell}")
                shown_count += 1
    else:
        print(f'\n✗ 错误：不包含"明细"工作表')
else:
    print("文件不存在")
