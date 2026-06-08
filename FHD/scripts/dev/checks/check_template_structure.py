"""查看模板文件结构"""

import os
from openpyxl import load_workbook

dir_path = r"e:\FHD\424"
file_name = "考勤 -2026-3 月份考勤统计表.xlsx"
file_path = os.path.join(dir_path, file_name)

# 确认文件存在
if not os.path.isfile(file_path):
    print(f"文件不存在：{file_path}")
    print("\n目录内容:")
    for f in os.listdir(dir_path):
        if "考勤" in f:
            print(f"  {f}")
    exit(1)

print(f"文件存在：{file_path}")
print(f"文件大小：{os.path.getsize(file_path)} 字节\n")

wb = load_workbook(file_path, data_only=True)

print(f"工作表：{wb.sheetnames}\n")

for ws_name in wb.sheetnames:
    ws = wb[ws_name]
    print(f"工作表：{ws_name}")
    print(f"  最大行：{ws.max_row}, 最大列：{ws.max_column}")
    merged = list(ws.merged_cells.ranges) if ws.merged_cells.ranges else []
    print(f"  合并单元格：{len(merged)} 个")

    # 显示前 10 行
    print(f"  前 10 行:")
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, min(15, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            row_data.append(str(val)[:15] if val else "")
        print(f"    行{row_idx}: {' | '.join(row_data)}")
    print()

wb.close()
