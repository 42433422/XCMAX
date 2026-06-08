"""调试模板加载过程"""

import sys

sys.path.insert(0, r"e:\FHD\backend\shell")

from pathlib import Path
import os

# 模板路径
template_path = Path(r"e:\FHD\424\考勤 -2026-3 月份考勤统计表.xlsx")

print("=== 调试模板加载 ===")
print(f"模板路径：{template_path}")
print(f"路径字符串：{str(template_path)}")
print(f"路径绝对：{template_path.absolute()}")
print(f"路径存在：{template_path.exists()}")
print(f"路径是文件：{template_path.is_file()}")

# 检查父目录
parent_dir = template_path.parent
file_name = template_path.name
print(f"\n父目录：{parent_dir}")
print(f"文件名：{file_name}")

# 列出目录内容
try:
    dir_contents = os.listdir(str(parent_dir))
    print(f"\n目录内容 ({len(dir_contents)} 个文件):")
    for f in dir_contents[:10]:
        print(f"  {f}")

    # 检查文件名是否在目录中
    if file_name in dir_contents:
        print(f"\n✓ 文件名在目录中")
    else:
        print(f"\n✗ 文件名不在目录中")

    # 尝试找到匹配的文件
    for f in dir_contents:
        if "考勤统计表" in f:
            print(f"找到匹配文件：{f}")
except Exception as e:
    print(f"列出目录失败：{e}")

# 尝试直接加载
print("\n=== 尝试加载模板 ===")
try:
    from openpyxl import load_workbook

    wb = load_workbook(str(template_path))
    print(f"✓ 成功加载！工作表：{wb.sheetnames}")
except Exception as e:
    print(f"✗ 加载失败：{e}")
    import traceback

    traceback.print_exc()
