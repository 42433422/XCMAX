#!/usr/bin/env python3
"""生成快速上手教程用 Excel / Word 样本（frontend/public/tutorial/）。"""
from __future__ import annotations

import sys
import zipfile
from pathlib import Path
from xml.sax.saxutils import escape

ROOT = Path(__file__).resolve().parents[2]
OUT = ROOT / "frontend" / "public" / "tutorial"
PREFIX = "教程示例-"


def _write_minimal_xlsx(path: Path, sheet_name: str, headers: list[str], rows: list[list[str]]) -> None:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise SystemExit("需要 openpyxl: pip install openpyxl") from exc
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    ws.append(headers)
    for row in rows:
        ws.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _write_minimal_docx(path: Path, title: str, paragraphs: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    body = "".join(f"<w:p><w:r><w:t>{escape(p)}</w:t></w:r></w:p>" for p in paragraphs)
    document_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>{body}<w:sectPr/></w:body>
</w:document>"""
    content_types = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""
    rels = """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="word/document.xml"/>
</Relationships>"""
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", content_types)
        zf.writestr("_rels/.rels", rels)
        zf.writestr("word/document.xml", document_xml)


def main() -> int:
    OUT.mkdir(parents=True, exist_ok=True)
    _write_minimal_xlsx(
        OUT / "xcagi-quickstart-sample-a.xlsx",
        f"{PREFIX}产品",
        ["产品名称", "单位", "数量"],
        [
            [f"{PREFIX}签字笔", "支", "12"],
            [f"{PREFIX}笔记本", "本", "6"],
        ],
    )
    _write_minimal_xlsx(
        OUT / "xcagi-quickstart-sample-b.xlsx",
        f"{PREFIX}联系人",
        ["姓名", "部门", "电话"],
        [
            [f"{PREFIX}张三", "市场部", "13800000001"],
            [f"{PREFIX}李四", "研发部", "13800000002"],
        ],
    )
    _write_minimal_docx(
        OUT / "xcagi-quickstart-sample.docx",
        f"{PREFIX}说明",
        [
            f"{PREFIX}办公包读 Word 测试",
            "这是快速上手教程用的短文，读完会在对话里显示摘要。",
            f"行前缀「{PREFIX}」便于识别与清理。",
        ],
    )
    # 进阶教程沿用
    _write_minimal_xlsx(
        OUT / "xcagi-tutorial-dept-employee.xlsx",
        f"{PREFIX}部门",
        ["部门名称", "负责人"],
        [
            [f"{PREFIX}市场部", "王五"],
            [f"{PREFIX}研发部", "赵六"],
        ],
    )
    wb_path = OUT / "xcagi-tutorial-dept-employee.xlsx"
    try:
        from openpyxl import load_workbook

        wb = load_workbook(wb_path)
        ws = wb.create_sheet(f"{PREFIX}人员")
        ws.append(["产品单位", "产品名称", "规格"])
        ws.append([f"{PREFIX}市场部", f"{PREFIX}样品A", "标准"])
        ws.append([f"{PREFIX}市场部", f"{PREFIX}样品B", "大号"])
        ws.append([f"{PREFIX}研发部", f"{PREFIX}样品C", "定制"])
        wb.save(wb_path)
    except Exception as exc:
        print("warn: dept-employee second sheet skipped:", exc, file=sys.stderr)

    print("wrote samples to", OUT)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
