"""验证输出文件"""

from openpyxl import load_workbook
import os

output_file = r"e:\FHD\424\测试上传输出.xlsx"

print(f"检查输出文件：{output_file}")
print(f"文件存在：{os.path.exists(output_file)}")

if os.path.exists(output_file):
    wb = load_workbook(output_file)
    print(f"工作表：{wb.sheetnames}")

    if "明细" in wb.sheetnames:
        ws = wb["明细"]
        print(f'\n✓ 成功！包含"明细"工作表')
        print(f"最大行：{ws.max_row}, 最大列：{ws.max_column}")

        # 检查前几行
        print(f"\n前 10 行:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    row_values.append(f"{chr(64+col_idx)}{row_idx}={cell.value}")
            if row_values:
                print(f"  Row {row_idx}: {row_values}")
    else:
        print(f'\n✗ 错误：不包含"明细"工作表')
else:
    print("文件不存在")
