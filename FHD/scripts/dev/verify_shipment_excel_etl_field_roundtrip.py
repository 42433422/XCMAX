#!/usr/bin/env python3
"""送货单 ETL 字段级验收：写模板单元格 → 解析字段 → 反推指纹 → 流水读写。

用法：
  cd FHD && FHD_SHIPMENT_ETL_LLM=0 .venv/bin/python scripts/dev/verify_shipment_excel_etl_field_roundtrip.py
"""

from __future__ import annotations

import json
import sys
import tempfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def main() -> int:
    import os

    # 字段复验走通用知识库；execute 需 shipment target
    os.environ.setdefault("FHD_SHIPMENT_ETL_LLM", "0")
    os.environ.setdefault("FHD_EXCEL_ETL_DEFAULT_TARGET", "shipment")
    os.environ.pop("FHD_EXCEL_ETL_ALLOW_BUILTIN", None)

    from openpyxl import load_workbook

    from app.application.excel_etl_kb import reset_excel_etl_kb_for_tests
    from app.application.shipment_etl_profile import clear_profile_cache, get_shipment_etl_profile
    from app.application.shipment_excel_etl_app_service import (
        note_fingerprint,
        parse_delivery_notes,
        regenerate_delivery_notes_from_file,
        write_delivery_note_workbook,
        write_ledger_workbook,
    )

    errors: list[str] = []
    oks: list[str] = []

    def fail(msg: str) -> None:
        errors.append(msg)

    def ok(msg: str) -> None:
        oks.append(msg)

    td_kb = Path(tempfile.mkdtemp(prefix="etl_kb_"))
    os.environ["FHD_EXCEL_ETL_KB_PATH"] = str(td_kb / "kb.json")
    reset_excel_etl_kb_for_tests(td_kb / "kb.json")
    clear_profile_cache()

    prof = get_shipment_etl_profile()
    write_cfg = prof.write
    if not write_cfg.get("header_row"):
        fail("profile write.header_row missing")
    else:
        ok(f"profile loaded id={prof.id}")

    td = Path(tempfile.mkdtemp(prefix="etl_field_verify_"))
    golden = [
        {
            "unit_name": "验收客户甲",
            "contact_person": "张三",
            "order_date": "2026年07月25日",
            "order_number": "V-1001",
            "sheet": "甲单",
            "items": [
                {
                    "model_number": "RX-V1",
                    "product_name": "PU哑光清漆",
                    "quantity_tins": 2,
                    "tin_spec": 25.0,
                    "quantity_kg": 50.0,
                    "unit_price": 18.0,
                    "amount": 900.0,
                },
                {
                    "model_number": "GS621",
                    "product_name": "PE白底漆",
                    "quantity_tins": 1,
                    "tin_spec": 28.0,
                    "quantity_kg": 28.0,
                    "unit_price": 8.5,
                    "amount": 238.0,
                },
            ],
        },
        {
            "unit_name": "验收客户乙",
            "contact_person": "李四",
            "order_date": "2026年07月25日",
            "order_number": "V-1002",
            "sheet": "乙单",
            "items": [
                {
                    "model_number": "6821A",
                    "product_name": "白底",
                    "quantity_tins": 3,
                    "tin_spec": 20.0,
                    "quantity_kg": 60.0,
                    "unit_price": 7.0,
                    "amount": 420.0,
                }
            ],
        },
    ]

    delivery_path = td / "verify_delivery.xlsx"
    written = write_delivery_note_workbook(golden, delivery_path)
    if not written.get("success"):
        fail(f"write delivery failed: {written}")
    else:
        ok(f"write delivery sheets={written.get('sheet_count')}")

    wb = load_workbook(delivery_path, data_only=True)
    try:
        if len(wb.sheetnames) != 2:
            fail(f"expected 2 sheets, got {wb.sheetnames}")
        item_cols = write_cfg["item_columns"]
        headers = write_cfg["header_row"]
        for idx, note in enumerate(golden):
            ws = wb[wb.sheetnames[idx]]
            if str(ws["A1"].value or "") != str(write_cfg.get("seller_title") or ""):
                fail(f"sheet{idx} A1 title mismatch: {ws['A1'].value!r}")
            a2 = str(ws["A2"].value or "")
            for token in (note["unit_name"], note["contact_person"], note["order_number"]):
                if token not in a2:
                    fail(f"sheet{idx} A2 missing {token!r}")
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(3, col).value
                if str(cell or "") != str(h):
                    fail(f"sheet{idx} header col{col}: {cell!r} != {h!r}")
            for r, item in enumerate(note["items"], start=4):
                for field, col in item_cols.items():
                    got = ws.cell(r, int(col)).value
                    exp = item[field]
                    if field in {"model_number", "product_name"}:
                        if str(got or "") != str(exp):
                            fail(f"sheet{idx} r{r} {field}: {got!r} != {exp!r}")
                    elif abs(float(got or 0) - float(exp)) > 1e-6:
                        fail(f"sheet{idx} r{r} {field}: {got} != {exp}")
            ok(f"xlsx cells ok sheet={ws.title}")
    finally:
        wb.close()

    parsed = parse_delivery_notes(delivery_path, include_ledger=False)
    if not parsed.get("success") or int(parsed.get("note_count") or 0) != 2:
        fail(f"parse delivery failed note_count={parsed.get('note_count')}")
    by_order = {n.get("order_number"): n for n in (parsed.get("notes") or [])}
    for note in golden:
        got = by_order.get(note["order_number"])
        if not got:
            fail(f"missing parsed order {note['order_number']}")
            continue
        for k in ("unit_name", "contact_person", "order_date", "order_number"):
            if str(got.get(k) or "") != str(note[k]):
                fail(f"{note['order_number']} {k}: {got.get(k)!r} != {note[k]!r}")
        gmap = {str(i.get("model_number")): i for i in (got.get("items") or [])}
        for item in note["items"]:
            gi = gmap.get(item["model_number"])
            if not gi:
                fail(f"missing item {item['model_number']}")
                continue
            if str(gi.get("product_name") or "") != item["product_name"]:
                fail(f"{item['model_number']} name mismatch")
            for num_k in ("quantity_tins", "tin_spec", "quantity_kg", "unit_price", "amount"):
                if abs(float(gi.get(num_k) or 0) - float(item[num_k])) > 1e-6:
                    fail(f"{item['model_number']} {num_k}: {gi.get(num_k)} != {item[num_k]}")
        ok(f"parsed fields ok order={note['order_number']}")

    regen_path = td / "verify_regen.xlsx"
    regen = regenerate_delivery_notes_from_file(delivery_path, regen_path, include_ledger=False)
    if not regen.get("success") or not regen.get("fingerprint_match"):
        fail(f"regenerate fingerprint_match failed: {regen.get('fingerprint_match')}")
    else:
        ok("regenerate fingerprint_match=true")
    reparsed = parse_delivery_notes(regen_path, include_ledger=False)
    re_by = {n.get("order_number"): n for n in (reparsed.get("notes") or [])}
    for note in golden:
        got = re_by.get(note["order_number"])
        if not got:
            fail(f"regen missing {note['order_number']}")
        elif note_fingerprint(got) != note_fingerprint(by_order[note["order_number"]]):
            fail(f"regen fingerprint drift {note['order_number']}")
        else:
            ok(f"regen stable {note['order_number']}")

    ledger_rows = [
        {
            "order_date": "2026-07-01",
            "order_number": "LV-001",
            "model_number": "GS621",
            "product_name": "PE白底漆",
            "quantity_tins": 2,
            "tin_spec": 25,
            "quantity_kg": 50,
            "unit_price": 8.5,
            "amount": 425,
        },
        {
            "order_date": "2026-07-02",
            "order_number": "LV-002",
            "model_number": "RX001",
            "product_name": "PU哑光漆",
            "quantity_tins": 1,
            "tin_spec": 20,
            "quantity_kg": 20,
            "unit_price": 17,
            "amount": 340,
        },
    ]
    ledger_path = td / "verify_ledger.xlsx"
    lw = write_ledger_workbook(ledger_rows, ledger_path, unit_name="流水验收客户")
    if not lw.get("success"):
        fail(f"write ledger failed: {lw}")
    else:
        ok(f"write ledger rows={lw.get('row_count')}")

    wb2 = load_workbook(ledger_path, data_only=True)
    try:
        ws = wb2.active
        lcols = write_cfg["ledger_item_columns"]
        for r, row in enumerate(ledger_rows, start=2):
            for field, col in lcols.items():
                got = ws.cell(r, int(col)).value
                exp = row[field]
                if field in {"order_date", "order_number", "model_number", "product_name"}:
                    if str(got or "") != str(exp):
                        fail(f"ledger r{r} {field}: {got!r} != {exp!r}")
                elif abs(float(got or 0) - float(exp)) > 1e-6:
                    fail(f"ledger r{r} {field}: {got} != {exp}")
        ok("ledger xlsx cells match")
    finally:
        wb2.close()

    lparsed = parse_delivery_notes(
        ledger_path, include_ledger=True, unit_name_hint="流水验收客户"
    )
    notes = lparsed.get("notes") or []
    if len(notes) < 2:
        fail(f"ledger parse expected>=2 got={len(notes)} msg={lparsed.get('message')}")
    else:
        lby = {n.get("order_number"): n for n in notes}
        for row in ledger_rows:
            n = lby.get(row["order_number"])
            if not n or n.get("source_kind") != "shipment_ledger":
                fail(f"ledger missing/wrong kind {row['order_number']}")
                continue
            it = (n.get("items") or [None])[0]
            if not it:
                fail(f"ledger empty items {row['order_number']}")
                continue
            for field in ("model_number", "product_name"):
                if str(it.get(field) or "") != str(row[field]):
                    fail(f"ledger {row['order_number']} {field}")
            for field in ("quantity_tins", "tin_spec", "quantity_kg", "unit_price", "amount"):
                if abs(float(it.get(field) or 0) - float(row[field])) > 1e-6:
                    fail(f"ledger {row['order_number']} {field}")
            ok(f"ledger parsed {row['order_number']}")

    # execute 入库字段（内存仓储 + 独立指纹目录，避免污染本机幂等库）
    from app.application.shipment_app_service import ShipmentApplicationService
    from app.domain.shipment.aggregates import Shipment

    class _Mem:
        def __init__(self) -> None:
            self.items: dict[int, Shipment] = {}
            self._n = 0

        def save(self, shipment: Shipment) -> Shipment:
            if shipment.id is None:
                self._n += 1
                shipment.id = self._n
            self.items[int(shipment.id)] = shipment
            return shipment

        def find_by_id(self, shipment_id: int):
            return self.items.get(int(shipment_id))

        def delete(self, shipment_id: int) -> bool:
            return self.items.pop(int(shipment_id), None) is not None

        def find_all(self, page: int = 1, per_page: int = 20):
            return list(self.items.values())

        def find_by_unit(self, unit_name: str):
            return [s for s in self.items.values() if s.purchase_unit_name == unit_name]

        def count(self) -> int:
            return len(self.items)

    import os

    from app.application.shipment_excel_etl_app_service import execute_shipment_excel_etl

    fp_root = td / "data"
    fp_root.mkdir(parents=True, exist_ok=True)
    os.environ["WORKSPACE_ROOT"] = str(td)
    os.environ["FHD_SHIPMENT_ETL_ALLOW_TMP"] = "1"
    os.environ["FHD_SHIPMENT_ETL_FINGERPRINT_BACKEND"] = "legacy"

    import app.utils.path_io.path_utils as path_utils

    old_get_data = path_utils.get_data_dir
    old_get_app = getattr(path_utils, "get_app_data_dir", None)
    path_utils.get_data_dir = lambda: str(fp_root)  # type: ignore[assignment]
    if old_get_app is not None:
        path_utils.get_app_data_dir = lambda: str(fp_root)  # type: ignore[assignment]

    repo = _Mem()
    svc = ShipmentApplicationService(repository=repo)

    import app.bootstrap as bootstrap

    old = getattr(bootstrap, "get_shipment_app_service", None)
    bootstrap.get_shipment_app_service = lambda: svc  # type: ignore[assignment]
    try:
        import app.services.tools_workflow_registered as tw

        old_imp = getattr(tw, "_execute_excel_import_records", None)
        tw._execute_excel_import_records = lambda records: {  # type: ignore[assignment]
            "success": True,
            "imported": len(records),
        }
        try:
            exe = execute_shipment_excel_etl(
                delivery_path,
                workspace_root=td,
                include_ledger=False,
                idempotent=True,
            )
        finally:
            if old_imp is not None:
                tw._execute_excel_import_records = old_imp  # type: ignore[assignment]
    finally:
        if old is not None:
            bootstrap.get_shipment_app_service = old  # type: ignore[assignment]
        path_utils.get_data_dir = old_get_data  # type: ignore[assignment]
        if old_get_app is not None:
            path_utils.get_app_data_dir = old_get_app  # type: ignore[assignment]

    if not exe.get("success") or int(exe.get("shipment_created") or 0) != 2:
        fail(f"execute failed: {exe}")
    else:
        ok(f"execute created={exe.get('shipment_created')}")
        saved_units = {s.purchase_unit_name for s in repo.items.values()}
        if saved_units != {"验收客户甲", "验收客户乙"}:
            fail(f"execute units mismatch: {sorted(saved_units)}")
        else:
            ok("execute units match")
        for s in repo.items.values():
            raw = str(getattr(s, "raw_text", "") or "")
            if "external_order_number=" not in raw or "fingerprint=" not in raw:
                fail(f"execute meta missing in raw_text: {raw}")
        if all("external_order_number=" in str(getattr(s, "raw_text", "") or "") for s in repo.items.values()):
            ok("execute order meta written")

    report = {
        "success": not errors,
        "ok_count": len(oks),
        "error_count": len(errors),
        "errors": errors,
        "oks": oks,
        "workdir": str(td),
    }
    out = ROOT / "tests" / "fixtures" / "shipment_etl" / "field_verify_report.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
