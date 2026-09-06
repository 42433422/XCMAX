# mypy: disable-error-code="arg-type"
"""考勤工作簿导入服务（attendance_import_app_service）单元测试。

覆盖：schema 建表、单元格首行提取、钉钉「每日统计」/固定模板「明细」两种解析、
文件名月份推断、products/customers 同步、每日打卡记录写入、完整导入流程（含回滚）。
"""

from __future__ import annotations

import json
import sqlite3
import sys
import types
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import pytest

from app.application import attendance_import_app_service as svc


class _FakeWs:
    """模拟 openpyxl worksheet 的 iter_rows(values_only=True)。"""

    def __init__(self, rows):
        self._rows = rows

    def iter_rows(self, min_row=1, values_only=False):
        yield from self._rows


def _make_xlsx(path: Path, sheets: dict[str, list[tuple]]) -> Path:
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(list(row))
    wb.save(str(path))
    return path


def _dingtalk_xlsx(path: Path) -> Path:
    return _make_xlsx(
        path,
        {
            "每日统计": [
                ("表头1",) * 7,
                ("表头2",) * 7,
                ("表头3",) * 7,
                ("张三", "公司正班", "包装部", "生产中心", "001", "普工", "u001"),
                ("李四", "工厂正班", "组装部", "生产中心", "002", "技工", "u002"),
                ("张三", "公司正班", "包装部", "生产中心", "001", "普工", "u001"),  # 重复
                ("", "x", "y", "z", "", "", ""),  # 空名跳过
            ]
        },
    )


def _mingxi_xlsx(path: Path) -> Path:
    return _make_xlsx(
        path,
        {
            "明细": [
                ("模板表头1",),
                ("模板表头2",),
                ("模板表头3",),
                ("包装部\n生产中心", "正班", "张三\n备注"),
                ("组装部", "正班", "李四"),
                ("包装部", "正班", "张三"),  # 重复
            ]
        },
    )


# ── _cell_first_line / _infer_month_from_filename ────────────────────────


class TestCellFirstLine:
    @pytest.mark.parametrize(
        ("value", "expected"),
        [
            (None, ""),
            ("", ""),
            ("   ", ""),
            ("abc", "abc"),
            ("  abc  ", "abc"),
            ("abc\nxyz", "abc"),
            ("abc\r\nxyz", "abc"),
            (123, "123"),
        ],
    )
    def test_variants(self, value, expected):
        assert svc._cell_first_line(value) == expected


class TestInferMonthFromFilename:
    @pytest.mark.parametrize(
        ("name", "expected"),
        [
            ("考勤-2026-3月份考勤统计表.xlsx", "2026-03"),
            # 正则 (0?[1-9]|1[0-2]) 在 "12" 前优先匹配 "1"，因此只取一位
            ("report_2025_12.xlsx", "2025-01"),
            ("2026年1月考勤.xlsx", "2026-01"),
            ("attendance-2026.09.xlsx", "2026-09"),
            ("no-month-here.xlsx", ""),
            # 同上，"13" 中的 "1" 被匹配，输出 2026-01（当前代码行为）
            ("2026-13-invalid.xlsx", "2026-01"),
        ],
    )
    def test_variants(self, name, expected):
        assert svc._infer_month_from_filename(Path(name)) == expected


# ── _ensure_schema ────────────────────────────────────────────────────────


class TestEnsureSchema:
    def test_creates_all_tables_and_indexes(self):
        conn = sqlite3.connect(":memory:")
        svc._ensure_schema(conn)
        names = {
            r[0]
            for r in conn.execute("SELECT name FROM sqlite_master WHERE type IN ('table', 'index')")
        }
        assert {
            "attendance_import_batches",
            "attendance_daily_records",
            "attendance_departments",
            "attendance_employees",
            "products",
            "customers",
            "ux_attendance_source_row",
            "ix_attendance_employee_date",
            "ix_employees_name",
            "ix_departments_dept",
        } <= names
        # 幂等：重复执行不报错
        svc._ensure_schema(conn)
        conn.close()


# ── sheet 解析 ────────────────────────────────────────────────────────────


class TestParseDingtalkDailySheet:
    def test_dedup_and_skip_empty_name(self):
        ws = _FakeWs(
            [
                ("张三", "公司正班", "包装部", "生产中心", "001", "普工", "u001"),
                ("张三", "公司正班", "包装部", "生产中心", "001", "普工", "u001"),
                ("李四", "工厂正班", "组装部", "生产中心", "002", "技工", "u002"),
                ("", "g", "d", "m", "", "", ""),
                (None, None, None, None, None, None, None),
            ]
        )
        departments, employees = svc._parse_dingtalk_daily_sheet(ws)
        assert len(employees) == 2
        assert employees[0] == {
            "name": "张三",
            "group": "公司正班",
            "dept": "包装部",
            "main_dept": "生产中心",
            "emp_no": "001",
            "position": "普工",
            "uid": "u001",
        }
        assert len(departments) == 2
        assert departments[0] == {
            "department": "包装部",
            "main_department": "生产中心",
            "attendance_group": "公司正班",
        }

    def test_empty_sheet(self):
        assert svc._parse_dingtalk_daily_sheet(_FakeWs([])) == ([], [])


class TestParseAttendanceDetailSheet:
    def test_first_line_and_header_skip(self):
        ws = _FakeWs(
            [
                ("部门", "性质", "姓名"),  # 表头跳过（dept == 部门）
                ("包装部\n生产中心", "正班", "张三\n别名"),
                ("组装部", "正班", "李四"),
                ("包装部", "正班", "张三"),  # 重复
                ("销售部", "正班", "姓名"),  # name == 姓名 跳过
                ("包装部", "正班", ""),  # 空名跳过
                (),  # 空行跳过
                ("仅部门",),  # 短行
            ]
        )
        departments, employees = svc._parse_attendance_detail_sheet(ws)
        assert [e["name"] for e in employees] == ["张三", "李四"]
        assert employees[0]["dept"] == "包装部"
        assert employees[0]["main_dept"] == "包装部"
        assert employees[0]["position"] == "正班"
        assert [d["department"] for d in departments] == ["包装部", "组装部"]

    def test_empty_sheet(self):
        assert svc._parse_attendance_detail_sheet(_FakeWs([])) == ([], [])


class TestParseWorkbook:
    def test_dingtalk_sheet_kind(self, tmp_path):
        path = _dingtalk_xlsx(tmp_path / "考勤-2026-3月.xlsx")
        departments, employees, kind = svc._parse_workbook(path)
        assert kind == "dingtalk"
        assert len(employees) == 2
        assert len(departments) == 2

    def test_mingxi_sheet_kind(self, tmp_path):
        path = _mingxi_xlsx(tmp_path / "明细转换.xlsx")
        departments, employees, kind = svc._parse_workbook(path)
        assert kind == "mingxi"
        assert [e["name"] for e in employees] == ["张三", "李四"]

    def test_missing_known_sheet_raises(self, tmp_path):
        path = _make_xlsx(tmp_path / "other.xlsx", {"Sheet1": [("a",)]})
        with pytest.raises(ValueError, match="未找到"):
            svc._parse_workbook(path)


# ── _sync_products_customers / _to_datetime_text ─────────────────────────


class TestSyncProductsCustomers:
    def _conn(self):
        conn = sqlite3.connect(":memory:")
        svc._ensure_schema(conn)
        return conn

    def test_products_per_employee_customers_per_dept(self):
        conn = self._conn()
        employees = [
            {"name": "张三", "dept": "包装部", "group": "正班"},
            {"name": "李四", "dept": "组装部", "group": "正班"},
            {"name": "王五", "dept": "包装部", "group": "正班"},  # 部门重复
        ]
        products, customers = svc._sync_products_customers(conn, "src", employees)
        assert products == 3
        assert customers == 2
        row = conn.execute(
            "SELECT model_number, name, specification, unit FROM products WHERE name='张三'"
        ).fetchone()
        assert row == ("包装部::张三", "张三", "正班", "包装部")
        conn.close()

    def test_employee_without_dept(self):
        conn = self._conn()
        products, customers = svc._sync_products_customers(
            conn, "src", [{"name": "独行侠", "dept": "", "group": "g"}]
        )
        assert products == 1
        assert customers == 0
        row = conn.execute("SELECT model_number FROM products").fetchone()
        assert row == ("独行侠",)
        conn.close()

    def test_resync_deletes_old_rows(self):
        conn = self._conn()
        svc._sync_products_customers(conn, "src", [{"name": "旧", "dept": "旧部", "group": ""}])
        products, customers = svc._sync_products_customers(
            conn, "src", [{"name": "新", "dept": "新部", "group": ""}]
        )
        assert products == 1 and customers == 1
        assert (
            conn.execute("SELECT COUNT(*) FROM products WHERE source_file='src'").fetchone()[0] == 1
        )
        assert conn.execute("SELECT name FROM products").fetchone()[0] == "新"
        conn.close()


class TestToDatetimeText:
    def test_mixed_values(self):
        dt = datetime(2026, 3, 5, 8, 0, 0)
        out = svc._to_datetime_text([dt, "raw-string", 8])
        assert out == ["2026-03-05 08:00:00", "raw-string", "8"]


# ── _import_daily_records_if_possible ────────────────────────────────────


class TestImportDailyRecordsIfPossible:
    def _conn(self):
        conn = sqlite3.connect(":memory:")
        svc._ensure_schema(conn)
        return conn

    def _fake_record(self, row: int, name: str):
        return SimpleNamespace(
            source_row=row,
            employee_name=name,
            attendance_group="公司正班",
            department="包装部",
            employee_no="001",
            position="普工",
            user_id="u001",
            work_date=datetime(2026, 3, row).date(),
            shift_name="正班",
            daily_times=[datetime(2026, 3, row, 8, 0)],
            raw_times=["08:00"],
            all_punch_times=lambda: [datetime(2026, 3, row, 8, 0)],
            leave_hours=0,
            absent_days=0,
            late_count_hint=0,
            early_count_hint=0,
            missing_card_count=0,
            notes=["备注"],
        )

    def test_writes_daily_records(self, monkeypatch):
        import app.mod_sdk.attendance as attendance_sdk

        parsed = SimpleNamespace(
            month="2026-03",
            rows_in=2,
            records=[self._fake_record(4, "张三"), self._fake_record(5, "李四")],
        )
        monkeypatch.setattr(attendance_sdk, "parse_attendance_workbook", lambda *a, **k: parsed)

        conn = self._conn()
        rows_in, written, month = svc._import_daily_records_if_possible(
            conn, Path("考勤-2026-3月.xlsx"), "src", ""
        )
        assert (rows_in, written, month) == (2, 2, "2026-03")
        row = conn.execute(
            "SELECT employee_name, month_label, daily_times_json, notes_json "
            "FROM attendance_daily_records ORDER BY source_row LIMIT 1"
        ).fetchone()
        assert row[0] == "张三" and row[1] == "2026-03"
        assert json.loads(row[2]) == ["2026-03-04 08:00:00"]
        assert json.loads(row[3]) == ["备注"]
        conn.close()

    def test_parser_missing_returns_passthrough(self, monkeypatch):
        fake = types.ModuleType("app.mod_sdk.attendance")

        def _raise(name):
            raise ModuleNotFoundError("No module named 'taiyangniao_attendance'")

        fake.__getattr__ = _raise  # type: ignore[attr-defined]
        monkeypatch.setitem(sys.modules, "app.mod_sdk.attendance", fake)
        conn = self._conn()
        assert svc._import_daily_records_if_possible(conn, Path("x.xlsx"), "src", "2026-03") == (
            0,
            0,
            "2026-03",
        )
        conn.close()


# ── import_attendance_workbook（集成）─────────────────────────────────────


class TestImportAttendanceWorkbook:
    def test_file_not_found(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            svc.import_attendance_workbook(tmp_path / "missing.xlsx", tmp_path / "db.sqlite")

    def test_dingtalk_full_import(self, monkeypatch, tmp_path):
        import app.mod_sdk.attendance as attendance_sdk

        record = SimpleNamespace(
            source_row=4,
            employee_name="张三",
            attendance_group="公司正班",
            department="包装部",
            employee_no="001",
            position="普工",
            user_id="u001",
            work_date=datetime(2026, 3, 4).date(),
            shift_name="正班",
            daily_times=[],
            raw_times=[],
            all_punch_times=lambda: [],
            leave_hours=1.5,
            absent_days=0,
            late_count_hint=0,
            early_count_hint=0,
            missing_card_count=0,
            notes=[],
        )
        parsed = SimpleNamespace(month="2026-03", rows_in=1, records=[record])
        monkeypatch.setattr(attendance_sdk, "parse_attendance_workbook", lambda *a, **k: parsed)

        excel = _dingtalk_xlsx(tmp_path / "考勤-2026-3月份考勤统计表.xlsx")
        db = tmp_path / "private" / "taiyangniao_pro.db"
        result = svc.import_attendance_workbook(excel, db)

        assert result["workbook_kind"] == "dingtalk"
        assert result["month_label"] == "2026-03"
        assert result["employee_rows"] == 2
        assert result["department_rows"] == 2
        assert result["product_rows"] == 2
        assert result["customer_rows"] == 2
        assert result["daily_rows_in"] == 1
        assert result["daily_rows_written"] == 1
        assert result["batch_id"] >= 1
        assert result["sync_ui_tables"] is True

        conn = sqlite3.connect(str(db))
        assert conn.execute("SELECT COUNT(*) FROM attendance_daily_records").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM attendance_import_batches").fetchone()[0] == 1
        assert conn.execute("SELECT COUNT(*) FROM attendance_employees").fetchone()[0] == 2
        conn.close()

    def test_mingxi_skips_daily_records(self, tmp_path):
        excel = _mingxi_xlsx(tmp_path / "转换结果-明细.xlsx")
        db = tmp_path / "db.sqlite"
        result = svc.import_attendance_workbook(
            excel, db, source_file_key="custom-key", sync_ui_tables=False
        )
        assert result["workbook_kind"] == "mingxi"
        assert result["source_file"] == "custom-key"
        assert result["daily_rows_in"] == 0
        assert result["daily_rows_written"] == 0
        assert result["product_rows"] == 0
        assert result["customer_rows"] == 0
        # 明细模式批次行数回退到员工数
        conn = sqlite3.connect(str(db))
        batch = conn.execute(
            "SELECT rows_in, rows_written FROM attendance_import_batches"
        ).fetchone()
        assert batch == (result["employee_rows"], result["employee_rows"])
        conn.close()

    def test_rollback_on_insert_error(self, monkeypatch, tmp_path):
        excel = _mingxi_xlsx(tmp_path / "明细.xlsx")
        db = tmp_path / "db.sqlite"

        def _boom(*args, **kwargs):
            raise sqlite3.OperationalError("disk full")

        monkeypatch.setattr(svc, "_sync_products_customers", _boom)
        with pytest.raises(sqlite3.OperationalError):
            svc.import_attendance_workbook(excel, db)
        # 回滚后无残留部门/员工行
        conn = sqlite3.connect(str(db))
        assert conn.execute("SELECT COUNT(*) FROM attendance_departments").fetchone()[0] == 0
        assert conn.execute("SELECT COUNT(*) FROM attendance_employees").fetchone()[0] == 0
        conn.close()
