"""Tests for app.application.customer_app_service — extended coverage.

Focus: import_data edge cases, import_from_excel, export_to_excel,
get_purchase_unit_by_name, match_purchase_unit, _check_shipment_associations,
engine/SessionLocal properties.
"""

from __future__ import annotations

from datetime import datetime
from unittest.mock import MagicMock, Mock, patch

import pytest

from app.application.customer_app_service import (
    CustomerApplicationService,
    get_customer_app_service,
    get_customers_session,
    reset_customers_engine,
)

# ---------------------------------------------------------------------------
# CustomerApplicationService — properties
# ---------------------------------------------------------------------------


class TestCustomerAppServiceProperties:
    def test_engine_property(self):
        svc = CustomerApplicationService()
        with patch("app.db.engine", "fake_engine"):
            result = svc._engine
        assert result == "fake_engine"

    def test_session_local_property(self):
        svc = CustomerApplicationService()
        with patch("app.db.SessionLocal", "fake_sl"):
            result = svc._SessionLocal
        assert result == "fake_sl"


# ---------------------------------------------------------------------------
# get_all — error path
# ---------------------------------------------------------------------------


class TestGetAllErrorPath:
    def test_recoverable_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.get_all()
        assert result["success"] is False
        assert "db error" in result["message"]
        assert result["data"] == []
        assert result["total"] == 0


# ---------------------------------------------------------------------------
# get_by_id — error path
# ---------------------------------------------------------------------------


class TestGetByIdErrorPath:
    def test_recoverable_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.get_by_id(1)
        assert result["success"] is False
        assert "db error" in result["message"]
        assert result["data"] is None


# ---------------------------------------------------------------------------
# create — error path
# ---------------------------------------------------------------------------


class TestCreateErrorPath:
    def test_recoverable_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.create({"customer_name": "test"})
        assert result["success"] is False
        assert "db error" in result["message"]


# ---------------------------------------------------------------------------
# update — error path
# ---------------------------------------------------------------------------


class TestUpdateErrorPath:
    def test_recoverable_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.update(1, {"customer_name": "test"})
        assert result["success"] is False
        assert "db error" in result["message"]


# ---------------------------------------------------------------------------
# delete — error path
# ---------------------------------------------------------------------------


class TestDeleteErrorPath:
    def test_recoverable_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.delete(1)
        assert result["success"] is False
        assert "db error" in result["message"]
        assert result["deleted_count"] == 0


# ---------------------------------------------------------------------------
# batch_delete — error path
# ---------------------------------------------------------------------------


class TestBatchDeleteErrorPath:
    def test_recoverable_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.batch_delete([1, 2])
        assert result["success"] is False
        assert "db error" in result["message"]
        assert result["deleted_count"] == 0


# ---------------------------------------------------------------------------
# import_data — extended
# ---------------------------------------------------------------------------


class TestImportDataExtended:
    def test_empty_data(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data([])
        assert result["success"] is True
        assert result["imported"] == 0
        assert result["skipped"] == 0
        assert result["failed"] == 0

    def test_import_with_empty_name_skipped(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data([{"customer_name": ""}])
        assert result["success"] is True
        assert result["skipped"] == 1

    def test_import_with_name_from_unit_name(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None  # no existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data([{"unit_name": "TestCo"}])
        assert result["success"] is True
        assert result["imported"] == 1

    def test_import_with_name_from_name_field(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data([{"name": "TestCo"}])
        assert result["success"] is True
        assert result["imported"] == 1

    def test_import_duplicate_skip(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_existing = Mock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = mock_existing  # existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data([{"customer_name": "Existing"}], skip_duplicates=True)
        assert result["success"] is True
        assert result["skipped"] == 1
        assert result["imported"] == 0  # skipped, not imported

    def test_import_duplicate_update(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_existing = Mock()
        mock_existing.contact_person = ""
        mock_existing.contact_phone = ""
        mock_existing.address = ""
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = mock_existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data(
                [{"customer_name": "Existing", "contact_person": "New"}],
                skip_duplicates=False,
            )
        assert result["success"] is True
        assert result["imported"] == 1

    def test_import_with_clean_data(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data([{"customer_name": "  TestCo  "}], clean_data=True)
        assert result["success"] is True
        assert result["imported"] == 1

    def test_import_item_error(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        # Force an error in the query
        mock_session.query.side_effect = RuntimeError("query error")
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data([{"customer_name": "TestCo"}])
        assert result["success"] is True
        assert result["failed"] == 1

    def test_import_outer_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("outer error")):
            result = svc.import_data([{"customer_name": "TestCo"}])
        assert result["success"] is False
        assert result["imported"] == 0


# ---------------------------------------------------------------------------
# import_from_excel — extended
# ---------------------------------------------------------------------------


class TestImportFromExcelExtended:
    def test_outer_error(self):
        svc = CustomerApplicationService()
        with patch(
            "app.db.sqlite_write_guard.sqlite_write_guard",
            side_effect=RuntimeError("guard error"),
        ):
            result = svc.import_from_excel(None)
        assert result["success"] is False
        assert result["imported"] == 0

    def test_inner_error(self):
        svc = CustomerApplicationService()
        mock_guard = MagicMock()
        mock_guard.__enter__ = MagicMock(return_value=None)
        mock_guard.__exit__ = MagicMock(return_value=False)

        with (
            patch("app.db.sqlite_write_guard.sqlite_write_guard", return_value=mock_guard),
            patch.object(svc, "_import_from_excel_locked", side_effect=RuntimeError("inner error")),
        ):
            result = svc.import_from_excel(None)
        assert result["success"] is False


# ---------------------------------------------------------------------------
# _import_from_excel_locked — extended
# ---------------------------------------------------------------------------


class TestImportFromExcelLockedExtended:
    def test_happy_path(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["客户名称", "联系人", "电话", "地址"])
        ws.append(["TestCo", "Zhang", "138", "Addr"])
        ws.append(["", "", "", ""])  # empty row should be skipped

        file_path = tmp_path / "test.xlsx"
        wb.save(str(file_path))

        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_session.query.return_value = mock_query

        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc._import_from_excel_locked(str(file_path))

        assert result["success"] is True
        assert result["inserted"] == 1

    def test_update_existing(self, tmp_path):
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["客户名称", "联系人", "电话", "地址"])
        ws.append(["Existing", "New", "139", "NewAddr"])

        file_path = tmp_path / "test.xlsx"
        wb.save(str(file_path))

        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_existing = Mock()
        mock_existing.contact_person = ""
        mock_existing.contact_phone = ""
        mock_existing.address = ""
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = mock_existing
        mock_session.query.return_value = mock_query

        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc._import_from_excel_locked(str(file_path))

        assert result["success"] is True
        assert result["updated"] == 1

    def test_load_workbook_error(self):
        svc = CustomerApplicationService()
        with patch("openpyxl.load_workbook", side_effect=RuntimeError("load error")):
            result = svc._import_from_excel_locked("/nonexistent.xlsx")
        assert result["success"] is False


# ---------------------------------------------------------------------------
# export_to_excel — extended
# ---------------------------------------------------------------------------


class TestExportToExcelExtended:
    def test_happy_path_no_template(self, tmp_path):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_unit = Mock()
        mock_unit.id = 1
        mock_unit.unit_name = "TestCo"
        mock_unit.contact_person = "Zhang"
        mock_unit.contact_phone = "138"
        mock_unit.address = "Addr"
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.all.return_value = [mock_unit]
        mock_session.query.return_value = mock_query

        with (
            patch.object(svc, "_get_session", return_value=mock_session),
            patch("app.utils.path_utils.get_data_dir", return_value=str(tmp_path)),
        ):
            result = svc.export_to_excel()
        assert result["success"] is True
        assert "file_path" in result
        assert "filename" in result

    def test_with_keyword(self, tmp_path):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.all.return_value = []
        mock_session.query.return_value = mock_query

        with (
            patch.object(svc, "_get_session", return_value=mock_session),
            patch("app.utils.path_utils.get_data_dir", return_value=str(tmp_path)),
        ):
            result = svc.export_to_excel(keyword="test")
        assert result["success"] is True

    def test_with_template_id(self, tmp_path):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.all.return_value = []
        mock_session.query.return_value = mock_query

        with (
            patch.object(svc, "_get_session", return_value=mock_session),
            patch("app.utils.path_utils.get_data_dir", return_value=str(tmp_path)),
            patch("app.application.get_template_app_service") as mock_get_tpl,
        ):
            mock_tpl_svc = Mock()
            mock_tpl_svc.get_templates.return_value = {"templates": []}
            mock_get_tpl.return_value = mock_tpl_svc
            result = svc.export_to_excel(template_id="1")
        assert result["success"] is True

    def test_template_lookup_error(self, tmp_path):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.all.return_value = []
        mock_session.query.return_value = mock_query

        with (
            patch.object(svc, "_get_session", return_value=mock_session),
            patch("app.utils.path_utils.get_data_dir", return_value=str(tmp_path)),
            patch(
                "app.application.get_template_app_service", side_effect=RuntimeError("tpl error")
            ),
        ):
            result = svc.export_to_excel(template_id="1")
        assert result["success"] is True  # Should fall back to no template

    def test_outer_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.export_to_excel()
        assert result["success"] is False


# ---------------------------------------------------------------------------
# get_purchase_unit_by_name — extended
# ---------------------------------------------------------------------------


class TestGetPurchaseUnitByNameExtended:
    def test_found(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_unit = Mock()
        mock_unit.id = 1
        mock_unit.unit_name = "TestCo"
        mock_unit.contact_person = "Zhang"
        mock_unit.contact_phone = "138"
        mock_unit.address = "Addr"
        mock_unit.discount_rate = 0.9
        mock_unit.is_active = True
        mock_unit.created_at = datetime(2024, 1, 1)
        mock_unit.updated_at = None
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = mock_unit
        mock_session.query.return_value = mock_query

        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_purchase_unit_by_name("TestCo")
        assert result is not None
        assert result.unit_name == "TestCo"

    def test_not_found(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_session.query.return_value = mock_query

        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_purchase_unit_by_name("Nonexistent")
        assert result is None

    def test_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.get_purchase_unit_by_name("TestCo")
        assert result is None


# ---------------------------------------------------------------------------
# match_purchase_unit — extended
# ---------------------------------------------------------------------------


class TestMatchPurchaseUnitExtended:
    def test_empty_name_returns_none(self):
        svc = CustomerApplicationService()
        result = svc.match_purchase_unit("")
        assert result is None

    def test_none_name_returns_none(self):
        svc = CustomerApplicationService()
        result = svc.match_purchase_unit(None)
        assert result is None

    def test_exact_match(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_unit = Mock()
        mock_unit.id = 1
        mock_unit.unit_name = "TestCo"
        mock_unit.contact_person = "Zhang"
        mock_unit.contact_phone = "138"
        mock_unit.address = "Addr"
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = mock_unit
        mock_session.query.return_value = mock_query

        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.match_purchase_unit("TestCo")
        assert result is not None
        assert result.unit_name == "TestCo"

    def test_substring_match(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()

        # First query (exact) returns None
        exact_query = MagicMock()
        exact_query.filter.return_value = exact_query
        exact_query.first.return_value = None

        # Second query (all) returns list with one unit
        all_query = MagicMock()
        all_query.filter.return_value = all_query
        all_query.all.return_value = []
        # Set up query to return exact_query first, then all_query
        mock_session.query.side_effect = [exact_query, all_query]

        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.match_purchase_unit("Test")
        assert result is None

    def test_error(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.match_purchase_unit("TestCo")
        assert result is None

    def test_short_name_no_substring(self):
        """Names shorter than 2 chars should not do substring matching."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        exact_query = MagicMock()
        exact_query.filter.return_value = exact_query
        exact_query.first.return_value = None
        all_query = MagicMock()
        all_query.filter.return_value = all_query
        all_query.all.return_value = []
        mock_session.query.side_effect = [exact_query, all_query]

        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.match_purchase_unit("X")
        assert result is None


# ---------------------------------------------------------------------------
# _check_shipment_associations — extended
# ---------------------------------------------------------------------------


class TestCheckShipmentAssociationsExtended:
    def test_no_associations(self):
        svc = CustomerApplicationService()
        mock_db = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.limit.return_value.all.return_value = []
        mock_query.count.return_value = 0
        mock_db.query.return_value = mock_query

        with patch("app.db.session.get_db") as mock_get_db:
            mock_ctx = MagicMock()
            mock_ctx.__enter__ = MagicMock(return_value=mock_db)
            mock_ctx.__exit__ = MagicMock(return_value=False)
            mock_get_db.return_value = mock_ctx
            result = svc._check_shipment_associations("TestCo")

        assert result["has_associations"] is False
        assert result["shipment_count"] == 0

    def test_with_associations(self):
        svc = CustomerApplicationService()
        mock_db = MagicMock()
        mock_record = Mock()
        mock_record.id = 1
        mock_record.product_name = "Widget"
        mock_record.quantity_kg = 10.5
        mock_record.created_at = datetime(2024, 1, 1)
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.limit.return_value.all.return_value = [mock_record]
        mock_query.count.return_value = 5
        mock_db.query.return_value = mock_query

        with patch("app.db.session.get_db") as mock_get_db:
            mock_ctx = MagicMock()
            mock_ctx.__enter__ = MagicMock(return_value=mock_db)
            mock_ctx.__exit__ = MagicMock(return_value=False)
            mock_get_db.return_value = mock_ctx
            result = svc._check_shipment_associations("TestCo")

        assert result["has_associations"] is True
        assert result["shipment_count"] == 5
        assert len(result["sample_records"]) == 1

    def test_error(self):
        svc = CustomerApplicationService()
        with patch("app.db.session.get_db", side_effect=RuntimeError("db error")):
            result = svc._check_shipment_associations("TestCo")
        assert result["has_associations"] is False
        assert result["shipment_count"] == 0
        assert "db error" in result["message"]


# ---------------------------------------------------------------------------
# get_customer_app_service — extended
# ---------------------------------------------------------------------------


class TestGetCustomerAppServiceExtended:
    def test_returns_singleton(self):
        mock_registry = Mock()
        mock_svc = Mock()
        mock_registry.customer_application_service = mock_svc
        with patch(
            "app.application.customer_app_service.get_service_registry", return_value=mock_registry
        ):
            result = get_customer_app_service()
        assert result is mock_svc


# ---------------------------------------------------------------------------
# get_customers_session — extended
# ---------------------------------------------------------------------------


class TestGetCustomersSessionExtended:
    def test_resolve_succeeds(self):
        mock_session = Mock()
        with patch(
            "app.mod_sdk.erp_repository_registry.resolve_customers_session",
            return_value=mock_session,
        ):
            result = get_customers_session()
        assert result is mock_session

    def test_resolve_fails_value_error(self):
        with (
            patch(
                "app.mod_sdk.erp_repository_registry.resolve_customers_session",
                side_effect=ValueError("no mod"),
            ),
            patch("app.db.SessionLocal") as mock_sl,
        ):
            get_customers_session()
            mock_sl.assert_called_once()


# ---------------------------------------------------------------------------
# reset_customers_engine — extended
# ---------------------------------------------------------------------------


class TestResetCustomersEngineExtended:
    def test_calls_invalidate(self):
        mock_registry = Mock()
        with patch(
            "app.application.customer_app_service.get_service_registry", return_value=mock_registry
        ):
            reset_customers_engine()
        mock_registry.invalidate_customer_application_service.assert_called_once()
