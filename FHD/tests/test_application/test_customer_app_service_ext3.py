"""Tests for app.application.customer_app_service — additional coverage (ext3).

Focus on REMAINING uncovered lines:
- get_customers_session with resolve_customers_session success path
- reset_customers_engine
- get_all with keyword filter, pagination, None values
- get_by_id with None values (contact_person, contact_phone, address, created_at, updated_at)
- create with various data combinations
- update with all field combinations and name conflict
- _check_shipment_associations with records and error path
- delete with force=True and associations
- batch_delete with force=True and mixed associations
- import_data with address/contact_address fallback, validate_before_import
- _import_from_excel_locked with update existing (row[1], row[2], row[3] truthy/falsy)
- export_to_excel with keyword, template_id, template lookup
- get_purchase_unit_by_name with discount_rate and is_active
- match_purchase_unit with exact match, substring match, short name, empty name
- get_customer_app_service singleton
"""

from __future__ import annotations

from datetime import datetime
from unittest.mock import MagicMock, Mock, patch

import pytest

from app.application.customer_app_service import (
    CustomerApplicationService,
    get_customer_app_service,
    get_customers_session,
    reset_customers_engine,
)


# ---------------------------------------------------------------------------
# get_customers_session — additional
# ---------------------------------------------------------------------------


class TestGetCustomersSessionAdditional:
    def test_resolve_customers_session_success(self):
        """Test that resolve_customers_session success path is used."""
        mock_session = MagicMock()
        with patch(
            "app.mod_sdk.erp_repository_registry.resolve_customers_session",
            return_value=mock_session,
        ):
            result = get_customers_session()
        assert result is mock_session

    def test_resolve_customers_session_import_error_fallback(self):
        """Test that ImportError falls back to SessionLocal."""
        mock_session_local = MagicMock()
        with patch(
            "app.mod_sdk.erp_repository_registry.resolve_customers_session",
            side_effect=ImportError("module not found"),
        ), patch("app.db.SessionLocal", return_value=mock_session_local):
            result = get_customers_session()
        assert result is mock_session_local

    def test_resolve_customers_session_runtime_error_fallback(self):
        """Test that RuntimeError falls back to SessionLocal."""
        mock_session_local = MagicMock()
        with patch(
            "app.mod_sdk.erp_repository_registry.resolve_customers_session",
            side_effect=RuntimeError("runtime error"),
        ), patch("app.db.SessionLocal", return_value=mock_session_local):
            result = get_customers_session()
        assert result is mock_session_local


# ---------------------------------------------------------------------------
# reset_customers_engine
# ---------------------------------------------------------------------------


class TestResetCustomersEngine:
    def test_reset_calls_invalidate(self):
        """Test that reset_customers_engine calls invalidate on registry."""
        mock_registry = MagicMock()
        with patch(
            "app.application.customer_app_service.get_service_registry", return_value=mock_registry
        ):
            reset_customers_engine()
        mock_registry.invalidate_customer_application_service.assert_called_once()


# ---------------------------------------------------------------------------
# get_all — additional branches
# ---------------------------------------------------------------------------


class TestGetAllAdditional:
    def _make_unit(self, **kwargs):
        unit = Mock()
        unit.id = kwargs.get("id", 1)
        unit.unit_name = kwargs.get("unit_name", "TestCo")
        unit.contact_person = kwargs.get("contact_person", "Zhang")
        unit.contact_phone = kwargs.get("contact_phone", "13800000000")
        unit.address = kwargs.get("address", "Addr")
        unit.created_at = kwargs.get("created_at", datetime(2026, 1, 1))
        unit.updated_at = kwargs.get("updated_at", datetime(2026, 1, 2))
        return unit

    def test_with_keyword_filter(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.count.return_value = 1
        unit = self._make_unit()
        mock_query.order_by.return_value.offset.return_value.limit.return_value.all.return_value = [
            unit
        ]
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_all(keyword="Test", page=1, per_page=20)
        assert result["success"] is True
        assert result["total"] == 1
        assert len(result["data"]) == 1
        assert result["data"][0]["customer_name"] == "TestCo"

    def test_with_none_values(self):
        """Test that None values are converted to empty strings."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.count.return_value = 1
        unit = self._make_unit(
            contact_person=None,
            contact_phone=None,
            address=None,
            created_at=None,
            updated_at=None,
        )
        mock_query.order_by.return_value.offset.return_value.limit.return_value.all.return_value = [
            unit
        ]
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_all()
        assert result["success"] is True
        assert result["data"][0]["contact_person"] == ""
        assert result["data"][0]["contact_phone"] == ""
        assert result["data"][0]["contact_address"] == ""
        assert result["data"][0]["created_at"] is None
        assert result["data"][0]["updated_at"] is None

    def test_with_pagination_page_2(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.count.return_value = 50
        mock_query.order_by.return_value.offset.return_value.limit.return_value.all.return_value = []
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_all(page=2, per_page=10)
        assert result["success"] is True
        assert result["page"] == 2
        assert result["per_page"] == 10
        # Verify offset is (2-1)*10 = 10
        mock_query.order_by.return_value.offset.assert_called_once_with(10)

    def test_empty_result(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.count.return_value = 0
        mock_query.order_by.return_value.offset.return_value.limit.return_value.all.return_value = []
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_all()
        assert result["success"] is True
        assert result["data"] == []
        assert result["total"] == 0


# ---------------------------------------------------------------------------
# get_by_id — additional branches
# ---------------------------------------------------------------------------


class TestGetByIdAdditional:
    def test_not_found(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_by_id(999)
        assert result["success"] is False
        assert result["message"] == "客户不存在"
        assert result["data"] is None

    def test_with_none_values(self):
        """Test get_by_id with None values converted to empty strings."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCo"
        unit.contact_person = None
        unit.contact_phone = None
        unit.address = None
        unit.created_at = None
        unit.updated_at = None
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_by_id(1)
        assert result["success"] is True
        assert result["data"]["contact_person"] == ""
        assert result["data"]["contact_phone"] == ""
        assert result["data"]["contact_address"] == ""
        assert result["data"]["created_at"] is None
        assert result["data"]["updated_at"] is None

    def test_with_valid_values(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCo"
        unit.contact_person = "Zhang"
        unit.contact_phone = "138"
        unit.address = "Addr"
        unit.created_at = datetime(2026, 1, 1)
        unit.updated_at = datetime(2026, 1, 2)
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_by_id(1)
        assert result["success"] is True
        assert result["data"]["customer_name"] == "TestCo"
        assert result["data"]["contact_person"] == "Zhang"
        assert "2026-01-01" in result["data"]["created_at"]


# ---------------------------------------------------------------------------
# create — additional branches
# ---------------------------------------------------------------------------


class TestCreateAdditional:
    def test_empty_customer_name(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.create({"customer_name": ""})
        assert result["success"] is False
        assert result["message"] == "客户名称不能为空"

    def test_missing_customer_name(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.create({})
        assert result["success"] is False
        assert result["message"] == "客户名称不能为空"

    def test_none_customer_name(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.create({"customer_name": None})
        assert result["success"] is False
        assert result["message"] == "客户名称不能为空"

    def test_existing_customer_name(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = Mock()  # existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.create({"customer_name": "Existing"})
        assert result["success"] is False
        assert result["message"] == "客户名称已存在"

    def test_create_success_with_all_fields(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None  # no existing

        # Mock the unit that gets created
        def add_side_effect(unit):
            unit.id = 1
            unit.created_at = datetime(2026, 1, 1)
            unit.updated_at = datetime(2026, 1, 1)

        mock_session.add.side_effect = add_side_effect
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.create(
                {
                    "customer_name": "NewCo",
                    "contact_person": "Zhang",
                    "contact_phone": "138",
                    "contact_address": "Addr",
                }
            )
        assert result["success"] is True
        assert result["message"] == "客户创建成功"
        assert result["data"]["customer_name"] == "NewCo"

    def test_create_success_with_minimal_data(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None

        def add_side_effect(unit):
            unit.id = 2
            unit.created_at = None
            unit.updated_at = None

        mock_session.add.side_effect = add_side_effect
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.create({"customer_name": "MinimalCo"})
        assert result["success"] is True
        assert result["data"]["contact_person"] == ""
        assert result["data"]["contact_phone"] == ""
        assert result["data"]["contact_address"] == ""


# ---------------------------------------------------------------------------
# update — additional branches
# ---------------------------------------------------------------------------


class TestUpdateAdditional:
    def test_not_found(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.update(999, {"customer_name": "Test"})
        assert result["success"] is False
        assert result["message"] == "客户不存在"

    def test_update_name_conflict(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        # First call returns the unit, second call (name conflict check) returns existing
        mock_query.first.side_effect = [Mock(), Mock()]
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.update(1, {"customer_name": "ConflictName"})
        assert result["success"] is False
        assert result["message"] == "客户名称已存在"

    def test_update_all_fields(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        # First query: find unit to update (returns unit)
        # Second query: check name conflict (returns None - no conflict)
        mock_query1 = MagicMock()
        mock_query1.filter.return_value = mock_query1
        mock_query1.first.return_value = None  # no name conflict

        unit = Mock()
        unit.id = 1
        unit.unit_name = "OldName"
        unit.contact_person = "Old"
        unit.contact_phone = "111"
        unit.address = "OldAddr"
        unit.created_at = datetime(2026, 1, 1)
        unit.updated_at = datetime(2026, 1, 2)

        mock_query0 = MagicMock()
        mock_query0.filter.return_value = mock_query0
        mock_query0.first.return_value = unit

        mock_session.query.return_value = mock_query0
        # After first call, return mock_query1 for conflict check
        mock_session.query.side_effect = [mock_query0, mock_query1]
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.update(
                1,
                {
                    "customer_name": "NewName",
                    "contact_person": "New",
                    "contact_phone": "222",
                    "contact_address": "NewAddr",
                },
            )
        assert result["success"] is True
        assert result["message"] == "客户更新成功"
        # Verify unit attributes were updated
        assert unit.unit_name == "NewName"
        assert unit.contact_person == "New"
        assert unit.contact_phone == "222"
        assert unit.address == "NewAddr"

    def test_update_partial_fields(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "KeepName"
        unit.contact_person = "Old"
        unit.contact_phone = "111"
        unit.address = "OldAddr"
        unit.created_at = datetime(2026, 1, 1)
        unit.updated_at = datetime(2026, 1, 2)
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.update(1, {"contact_phone": "999"})
        assert result["success"] is True
        # Only contact_phone should be updated
        assert unit.contact_phone == "999"
        # Other fields unchanged
        assert unit.unit_name == "KeepName"
        assert unit.contact_person == "Old"

    def test_update_no_name_change(self):
        """Test update without changing customer_name (no conflict check)."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "KeepName"
        unit.contact_person = "Old"
        unit.contact_phone = "111"
        unit.address = "OldAddr"
        unit.created_at = datetime(2026, 1, 1)
        unit.updated_at = datetime(2026, 1, 2)
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.update(1, {"contact_person": "New"})
        assert result["success"] is True
        # query.first should only be called once (no name conflict check)
        assert mock_query.first.call_count == 1


# ---------------------------------------------------------------------------
# _check_shipment_associations — additional
# ---------------------------------------------------------------------------


class TestCheckShipmentAssociationsAdditional:
    def test_with_records(self):
        svc = CustomerApplicationService()
        mock_db = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.limit.return_value.all.return_value = []
        mock_query.count.return_value = 5
        mock_db.query.return_value = mock_query
        mock_db.__enter__ = MagicMock(return_value=mock_db)
        mock_db.__exit__ = MagicMock(return_value=False)
        with patch("app.db.session.get_db", return_value=mock_db):
            result = svc._check_shipment_associations("TestCo")
        assert result["has_associations"] is True
        assert result["shipment_count"] == 5

    def test_no_records(self):
        svc = CustomerApplicationService()
        mock_db = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.limit.return_value.all.return_value = []
        mock_query.count.return_value = 0
        mock_db.query.return_value = mock_query
        mock_db.__enter__ = MagicMock(return_value=mock_db)
        mock_db.__exit__ = MagicMock(return_value=False)
        with patch("app.db.session.get_db", return_value=mock_db):
            result = svc._check_shipment_associations("TestCo")
        assert result["has_associations"] is False
        assert result["shipment_count"] == 0

    def test_with_sample_records(self):
        svc = CustomerApplicationService()
        mock_db = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query

        record = Mock()
        record.id = 1
        record.product_name = "Product A"
        record.quantity_kg = 100.5
        record.created_at = datetime(2026, 1, 1)

        mock_query.order_by.return_value.limit.return_value.all.return_value = [record]
        mock_query.count.return_value = 1
        mock_db.query.return_value = mock_query
        mock_db.__enter__ = MagicMock(return_value=mock_db)
        mock_db.__exit__ = MagicMock(return_value=False)
        with patch("app.db.session.get_db", return_value=mock_db):
            result = svc._check_shipment_associations("TestCo")
        assert result["has_associations"] is True
        assert len(result["sample_records"]) == 1
        assert result["sample_records"][0]["product_name"] == "Product A"
        assert result["sample_records"][0]["quantity_kg"] == 100.5

    def test_with_none_created_at(self):
        svc = CustomerApplicationService()
        mock_db = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query

        record = Mock()
        record.id = 1
        record.product_name = "Product A"
        record.quantity_kg = 100.5
        record.created_at = None

        mock_query.order_by.return_value.limit.return_value.all.return_value = [record]
        mock_query.count.return_value = 1
        mock_db.query.return_value = mock_query
        mock_db.__enter__ = MagicMock(return_value=mock_db)
        mock_db.__exit__ = MagicMock(return_value=False)
        with patch("app.db.session.get_db", return_value=mock_db):
            result = svc._check_shipment_associations("TestCo")
        assert result["sample_records"][0]["created_at"] is None

    def test_error_path(self):
        svc = CustomerApplicationService()
        with patch("app.db.session.get_db", side_effect=RuntimeError("db error")):
            result = svc._check_shipment_associations("TestCo")
        assert result["has_associations"] is False
        assert result["shipment_count"] == 0
        assert result["sample_records"] == []
        assert "db error" in result["message"]


# ---------------------------------------------------------------------------
# delete — additional branches
# ---------------------------------------------------------------------------


class TestDeleteAdditional:
    def test_not_found(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.delete(999)
        assert result["success"] is False
        assert result["message"] == "客户不存在"
        assert result["deleted_count"] == 0

    def test_delete_with_associations_no_force(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCo"
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session), patch.object(
            svc,
            "_check_shipment_associations",
            return_value={
                "has_associations": True,
                "shipment_count": 3,
                "sample_records": [{"id": 1}],
            },
        ):
            result = svc.delete(1, force=False)
        assert result["success"] is False
        assert result["has_associations"] is True
        assert result["association_details"]["shipment_count"] == 3
        assert "force=True" in result["suggestion"]

    def test_delete_with_associations_force(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCo"
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session), patch.object(
            svc,
            "_check_shipment_associations",
            return_value={
                "has_associations": True,
                "shipment_count": 3,
                "sample_records": [],
            },
        ):
            result = svc.delete(1, force=True)
        assert result["success"] is True
        assert result["deleted_count"] == 1
        mock_session.delete.assert_called_once_with(unit)
        mock_session.commit.assert_called_once()

    def test_delete_no_associations(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCo"
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session), patch.object(
            svc,
            "_check_shipment_associations",
            return_value={"has_associations": False, "shipment_count": 0, "sample_records": []},
        ):
            result = svc.delete(1)
        assert result["success"] is True
        assert result["deleted_count"] == 1
        assert result["has_associations"] is False


# ---------------------------------------------------------------------------
# batch_delete — additional branches
# ---------------------------------------------------------------------------


class TestBatchDeleteAdditional:
    def test_not_found(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.all.return_value = []
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.batch_delete([999])
        assert result["success"] is False
        assert result["message"] == "未找到要删除的客户"
        assert result["deleted_count"] == 0

    def test_with_associations_no_force(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit1 = Mock()
        unit1.id = 1
        unit1.unit_name = "Co1"
        unit2 = Mock()
        unit2.id = 2
        unit2.unit_name = "Co2"
        mock_query.all.return_value = [unit1, unit2]
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session), patch.object(
            svc,
            "_check_shipment_associations",
            side_effect=[
                {"has_associations": True, "shipment_count": 2, "sample_records": []},
                {"has_associations": False, "shipment_count": 0, "sample_records": []},
            ],
        ):
            result = svc.batch_delete([1, 2], force=False)
        assert result["success"] is False
        assert result["has_associations"] is True
        assert len(result["affected_units"]) == 1
        assert result["affected_units"][0]["id"] == 1

    def test_with_associations_force(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit1 = Mock()
        unit1.id = 1
        unit1.unit_name = "Co1"
        unit2 = Mock()
        unit2.id = 2
        unit2.unit_name = "Co2"
        mock_query.all.return_value = [unit1, unit2]
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.batch_delete([1, 2], force=True)
        assert result["success"] is True
        assert result["deleted_count"] == 2
        assert mock_session.delete.call_count == 2

    def test_no_associations(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "Co1"
        mock_query.all.return_value = [unit]
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session), patch.object(
            svc,
            "_check_shipment_associations",
            return_value={"has_associations": False, "shipment_count": 0, "sample_records": []},
        ):
            result = svc.batch_delete([1], force=False)
        assert result["success"] is True
        assert result["deleted_count"] == 1


# ---------------------------------------------------------------------------
# import_data — additional branches
# ---------------------------------------------------------------------------


class TestImportDataAdditional:
    def test_import_with_address_fallback(self):
        """Test that address falls back to contact_address."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None

        def add_side_effect(unit):
            unit.id = 1

        mock_session.add.side_effect = add_side_effect
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data(
                [{"customer_name": "TestCo", "contact_address": "Addr"}]
            )
        assert result["success"] is True
        assert result["imported"] == 1

    def test_import_with_both_address_fields(self):
        """Test that address takes priority over contact_address."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None

        def add_side_effect(unit):
            unit.id = 1

        mock_session.add.side_effect = add_side_effect
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data(
                [
                    {
                        "customer_name": "TestCo",
                        "address": "PrimaryAddr",
                        "contact_address": "FallbackAddr",
                    }
                ]
            )
        assert result["success"] is True
        assert result["imported"] == 1

    def test_import_update_existing_with_address(self):
        """Test updating existing customer with address fields."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_existing = Mock()
        mock_existing.contact_person = "Old"
        mock_existing.contact_phone = "111"
        mock_existing.address = "OldAddr"
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = mock_existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data(
                [
                    {
                        "customer_name": "Existing",
                        "contact_person": "New",
                        "contact_phone": "222",
                        "address": "NewAddr",
                    }
                ],
                skip_duplicates=False,
            )
        assert result["success"] is True
        assert result["imported"] == 1
        assert mock_existing.contact_person == "New"
        assert mock_existing.contact_phone == "222"
        assert mock_existing.address == "NewAddr"

    def test_import_update_existing_with_contact_address(self):
        """Test updating existing customer with contact_address fallback."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_existing = Mock()
        mock_existing.contact_person = "Old"
        mock_existing.contact_phone = "111"
        mock_existing.address = "OldAddr"
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = mock_existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data(
                [
                    {
                        "customer_name": "Existing",
                        "contact_address": "NewAddr",
                    }
                ],
                skip_duplicates=False,
            )
        assert result["success"] is True
        assert mock_existing.address == "NewAddr"

    def test_import_update_existing_no_new_values(self):
        """Test updating existing customer with no new values (keeps old)."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_existing = Mock()
        mock_existing.contact_person = "Old"
        mock_existing.contact_phone = "111"
        mock_existing.address = "OldAddr"
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = mock_existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data(
                [{"customer_name": "Existing"}],
                skip_duplicates=False,
            )
        assert result["success"] is True
        # Values should remain unchanged
        assert mock_existing.contact_person == "Old"

    def test_import_multiple_items(self):
        """Test importing multiple items with mixed results."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        # First item: no existing, Second item: existing (skip), Third: empty name
        mock_query.first.side_effect = [None, Mock(), None]

        def add_side_effect(unit):
            unit.id = 1

        mock_session.add.side_effect = add_side_effect
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.import_data(
                [
                    {"customer_name": "NewCo"},
                    {"customer_name": "Existing"},
                    {"customer_name": ""},
                ]
            )
        assert result["success"] is True
        assert result["imported"] == 1
        assert result["skipped"] == 2  # existing + empty name


# ---------------------------------------------------------------------------
# _import_from_excel_locked — additional branches
# ---------------------------------------------------------------------------


class TestImportFromExcelLockedAdditional:
    def test_update_existing_with_all_fields(self, tmp_path):
        """Test updating existing customer with all fields from Excel."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["客户名称", "联系人", "电话", "地址"])
        ws.append(["ExistingCo", "NewPerson", "NewPhone", "NewAddr"])

        file_path = tmp_path / "test.xlsx"
        wb.save(str(file_path))

        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_existing = Mock()
        mock_existing.contact_person = "Old"
        mock_existing.contact_phone = "Old"
        mock_existing.address = "Old"
        mock_query.first.return_value = mock_existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc._import_from_excel_locked(str(file_path))
        assert result["success"] is True
        assert result["updated"] == 1
        assert result["inserted"] == 0
        assert mock_existing.contact_person == "NewPerson"
        assert mock_existing.contact_phone == "NewPhone"
        assert mock_existing.address == "NewAddr"

    def test_update_existing_with_partial_fields(self, tmp_path):
        """Test updating existing customer with partial fields (falsy values)."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["客户名称", "联系人", "电话", "地址"])
        # row[1], row[2], row[3] are None/empty → existing values not updated
        ws.append(["ExistingCo", None, None, None])

        file_path = tmp_path / "test.xlsx"
        wb.save(str(file_path))

        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_existing = Mock()
        mock_existing.contact_person = "KeepPerson"
        mock_existing.contact_phone = "KeepPhone"
        mock_existing.address = "KeepAddr"
        mock_query.first.return_value = mock_existing
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc._import_from_excel_locked(str(file_path))
        assert result["success"] is True
        assert result["updated"] == 1
        # Values should remain unchanged
        assert mock_existing.contact_person == "KeepPerson"

    def test_insert_new_customer(self, tmp_path):
        """Test inserting new customer from Excel."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["客户名称", "联系人", "电话", "地址"])
        ws.append(["NewCo", "Person", "Phone", "Addr"])

        file_path = tmp_path / "test.xlsx"
        wb.save(str(file_path))

        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None  # no existing

        def add_side_effect(unit):
            unit.id = 1

        mock_session.add.side_effect = add_side_effect
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc._import_from_excel_locked(str(file_path))
        assert result["success"] is True
        assert result["inserted"] == 1
        assert result["updated"] == 0

    def test_skip_empty_row(self, tmp_path):
        """Test that empty rows are skipped (not counted in skipped, just continue)."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.append(["客户名称", "联系人", "电话", "地址"])
        ws.append([None, None, None, None])  # empty row - skipped via `continue`
        ws.append(["", "", "", ""])  # row with empty strings - row[0] is "" which is falsy

        file_path = tmp_path / "test.xlsx"
        wb.save(str(file_path))

        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc._import_from_excel_locked(str(file_path))
        assert result["success"] is True
        # Empty rows are skipped via `continue` before reaching the skipped counter
        # So skipped count is 0, but inserted/updated should also be 0
        assert result["inserted"] == 0
        assert result["updated"] == 0

    def test_error_path(self, tmp_path):
        """Test error path in _import_from_excel_locked."""
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc._import_from_excel_locked("nonexistent.xlsx")
        assert result["success"] is False
        assert "db error" in result["message"]


# ---------------------------------------------------------------------------
# export_to_excel — additional branches
# ---------------------------------------------------------------------------


class TestExportToExcelAdditional:
    def test_export_with_keyword(self, tmp_path):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.all.return_value = []
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session), patch(
            "app.utils.path_utils.get_data_dir", return_value=str(tmp_path)
        ):
            result = svc.export_to_excel(keyword="Test")
        assert result["success"] is True
        assert "file_path" in result

    def test_export_error_path(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.export_to_excel()
        assert result["success"] is False
        assert "db error" in result["message"]

    def test_export_with_template_id_not_found(self, tmp_path):
        """Test export with template_id that doesn't match any template."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.all.return_value = []
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session), patch(
            "app.utils.path_utils.get_data_dir", return_value=str(tmp_path)
        ), patch(
            "app.application.get_template_app_service"
        ) as mock_get_templates:
            mock_template_svc = MagicMock()
            mock_template_svc.get_templates.return_value = {"templates": []}
            mock_get_templates.return_value = mock_template_svc
            result = svc.export_to_excel(template_id="nonexistent")
        assert result["success"] is True

    def test_export_with_template_id_found(self, tmp_path):
        """Test export with template_id that matches a template."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.all.return_value = []
        mock_session.query.return_value = mock_query
        # Create a template file
        template_file = tmp_path / "template.xlsx"
        from openpyxl import Workbook

        wb = Workbook()
        wb.save(str(template_file))
        with patch.object(svc, "_get_session", return_value=mock_session), patch(
            "app.utils.path_utils.get_data_dir", return_value=str(tmp_path)
        ), patch(
            "app.application.get_template_app_service"
        ) as mock_get_templates:
            mock_template_svc = MagicMock()
            mock_template_svc.get_templates.return_value = {
                "templates": [{"id": "1", "path": str(template_file)}]
            }
            mock_get_templates.return_value = mock_template_svc
            result = svc.export_to_excel(template_id="1")
        assert result["success"] is True

    def test_export_with_template_id_error(self, tmp_path):
        """Test export with template_id where template lookup fails."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.order_by.return_value.all.return_value = []
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session), patch(
            "app.utils.path_utils.get_data_dir", return_value=str(tmp_path)
        ), patch(
            "app.application.get_template_app_service",
            side_effect=RuntimeError("template error"),
        ):
            result = svc.export_to_excel(template_id="1")
        # Should still succeed (template error is caught)
        assert result["success"] is True


# ---------------------------------------------------------------------------
# get_purchase_unit_by_name — additional
# ---------------------------------------------------------------------------


class TestGetPurchaseUnitByNameAdditional:
    def test_found_with_all_fields(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCo"
        unit.contact_person = "Zhang"
        unit.contact_phone = "138"
        unit.address = "Addr"
        unit.discount_rate = 0.9
        unit.is_active = True
        unit.created_at = datetime(2026, 1, 1)
        unit.updated_at = datetime(2026, 1, 2)
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_purchase_unit_by_name("TestCo")
        assert result is not None
        assert result.id == 1
        assert result.unit_name == "TestCo"
        assert result.discount_rate == 0.9
        assert result.is_active is True

    def test_not_found(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_purchase_unit_by_name("Nonexistent")
        assert result is None

    def test_with_none_values(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCo"
        unit.contact_person = None
        unit.contact_phone = None
        unit.address = None
        unit.discount_rate = None
        unit.is_active = True
        unit.created_at = None
        unit.updated_at = None
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.get_purchase_unit_by_name("TestCo")
        assert result is not None
        assert result.contact_person == ""
        assert result.contact_phone == ""
        assert result.address == ""
        assert result.discount_rate == 1.0  # default

    def test_error_path(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.get_purchase_unit_by_name("TestCo")
        assert result is None


# ---------------------------------------------------------------------------
# match_purchase_unit — additional
# ---------------------------------------------------------------------------


class TestMatchPurchaseUnitAdditional:
    def test_empty_name(self):
        svc = CustomerApplicationService()
        result = svc.match_purchase_unit("")
        assert result is None

    def test_none_name(self):
        svc = CustomerApplicationService()
        result = svc.match_purchase_unit(None)
        assert result is None

    def test_whitespace_name(self):
        svc = CustomerApplicationService()
        result = svc.match_purchase_unit("   ")
        assert result is None

    def test_exact_match(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCo"
        unit.contact_person = "Zhang"
        unit.contact_phone = "138"
        unit.address = "Addr"
        mock_query.first.return_value = unit
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.match_purchase_unit("TestCo")
        assert result is not None
        assert result.id == 1
        assert result.unit_name == "TestCo"

    def test_substring_match_name_in_unit(self):
        """Test substring match where input name is substring of unit name."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        # First query (exact) returns None
        mock_query.first.return_value = None
        # Second query (all) returns list with matching unit
        unit = Mock()
        unit.id = 1
        unit.unit_name = "TestCompany"
        unit.contact_person = "Zhang"
        unit.contact_phone = "138"
        unit.address = "Addr"
        mock_query.all.return_value = [unit]
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.match_purchase_unit("Test")
        assert result is not None
        assert result.unit_name == "TestCompany"

    def test_substring_match_unit_in_name(self):
        """Test substring match where unit name is substring of input name."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        unit = Mock()
        unit.id = 1
        unit.unit_name = "Test"
        unit.contact_person = "Zhang"
        unit.contact_phone = "138"
        unit.address = "Addr"
        mock_query.all.return_value = [unit]
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.match_purchase_unit("TestCompany")
        assert result is not None
        assert result.unit_name == "Test"

    def test_short_name_no_substring_match(self):
        """Test that short names (< 2 chars) don't do substring matching."""
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_query.all.return_value = []
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.match_purchase_unit("A")
        assert result is None

    def test_no_match(self):
        svc = CustomerApplicationService()
        mock_session = MagicMock()
        mock_query = MagicMock()
        mock_query.filter.return_value = mock_query
        mock_query.first.return_value = None
        mock_query.all.return_value = []
        mock_session.query.return_value = mock_query
        with patch.object(svc, "_get_session", return_value=mock_session):
            result = svc.match_purchase_unit("Nonexistent")
        assert result is None

    def test_error_path(self):
        svc = CustomerApplicationService()
        with patch.object(svc, "_get_session", side_effect=RuntimeError("db error")):
            result = svc.match_purchase_unit("TestCo")
        assert result is None


# ---------------------------------------------------------------------------
# get_customer_app_service — singleton
# ---------------------------------------------------------------------------


class TestGetCustomerAppService:
    def test_returns_singleton(self):
        mock_registry = MagicMock()
        mock_service = MagicMock()
        mock_registry.customer_application_service = mock_service
        with patch(
            "app.application.customer_app_service.get_service_registry", return_value=mock_registry
        ):
            result = get_customer_app_service()
        assert result is mock_service
