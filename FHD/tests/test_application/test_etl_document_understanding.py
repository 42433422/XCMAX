from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook

from app.application.etl.document_routing import (
    build_document_routes,
    build_sheet_inventory,
)
from app.application.etl.document_understanding import understand_workbook
from app.application.etl.llm_assist import LlmAssistResult, advise_document_understanding
from app.application.etl.parser_document_plan import parse_workbook_with_document_plan
from app.application.etl.parsers import parse_file
from app.application.etl.target_detection import detect_etl_target
from app.application.etl.workbook_evidence import build_workbook_evidence


def _save(rows: list[list[object]], path: Path) -> Path:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "业务单"
    for row in rows:
        worksheet.append(row)
    workbook.save(path)
    return path


def test_document_understanding_rejects_invented_cells_and_ranges(tmp_path, monkeypatch):
    path = _save(
        [
            ["采购订单"],
            ["订单号", "PO-1"],
            ["品名", "数量", "单价", "金额"],
            ["底漆", 2, 10, 20],
        ],
        tmp_path / "order.xlsx",
    )
    evidence = build_workbook_evidence(path)
    monkeypatch.setattr(
        "app.application.etl.llm_assist._complete",
        lambda *_args, **_kwargs: LlmAssistResult(
            used_llm=True,
            data={
                "file_structure": "single_document",
                "summary": "采购订单",
                "documents": [
                    {
                        "document_id": "valid",
                        "document_type": "purchase_order",
                        "sheet": "业务单",
                        "title_cell_ids": ["s1:r1:c1", "invented"],
                        "header_fields": [
                            {
                                "role": "document_number",
                                "label_cell_id": "s1:r2:c1",
                                "value_cell_id": "s1:r2:c2",
                                "reason": "相邻键值",
                            },
                            {
                                "role": "supplier",
                                "label_cell_id": "invented",
                                "value_cell_id": "s1:r2:c2",
                                "reason": "伪造",
                            },
                        ],
                        "tables": [
                            {
                                "header_start_row": 3,
                                "header_end_row": 3,
                                "data_start_row": 4,
                                "data_end_row": 4,
                                "first_column": 1,
                                "last_column": 4,
                                "columns": [
                                    {
                                        "column": 1,
                                        "role": "product_name",
                                        "header_cell_id": "s1:r3:c1",
                                        "reason": "品名",
                                    },
                                    {
                                        "column": 2,
                                        "role": "quantity",
                                        "header_cell_id": "invented",
                                        "reason": "伪造",
                                    },
                                ],
                            },
                            {
                                "header_start_row": 9,
                                "header_end_row": 9,
                                "data_start_row": 10,
                                "data_end_row": 12,
                                "first_column": 1,
                                "last_column": 4,
                                "columns": [],
                            },
                        ],
                        "total_amount_cell_id": "invented",
                        "confidence": 0.98,
                        "requires_review": False,
                        "issues": [],
                    },
                    {
                        "document_id": "invented-sheet",
                        "document_type": "invoice",
                        "sheet": "不存在",
                        "title_cell_ids": [],
                        "header_fields": [],
                        "tables": [],
                        "total_amount_cell_id": "",
                        "confidence": 1,
                        "requires_review": False,
                        "issues": [],
                    },
                ],
            },
        ),
    )

    result = advise_document_understanding(evidence)

    assert [item["document_id"] for item in result.data["documents"]] == ["valid"]
    document = result.data["documents"][0]
    assert [item["cell_id"] for item in document["title_cells"]] == ["s1:r1:c1"]
    assert [item["role"] for item in document["header_fields"]] == ["document_number"]
    assert len(document["tables"]) == 1
    assert [item["role"] for item in document["tables"][0]["columns"]] == ["product_name"]
    assert document["total_amount_cell_id"] == ""


def test_fallback_understanding_projects_purchase_order_header_into_line_rows(monkeypatch):
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    path = Path(__file__).parents[1] / "fixtures" / "network_forms" / "form_采购订单_七彩乐园.xlsx"

    plan = understand_workbook(
        path,
        hinted_target_type="customer_products",
        hint_confidence=0.35,
    )
    dataset = parse_workbook_with_document_plan(
        path,
        target_type="purchase_orders",
        document_plan=plan,
        max_rows=1000,
    )

    assert plan["source"] == "deterministic_fallback"
    assert plan["recommended_target_type"] == "purchase_orders"
    assert plan["documents"][0]["document_type"] == "purchase_order"
    assert plan["documents"][0]["title_cells"][0]["coordinate"] == "A1"
    assert dataset is not None
    assert len(dataset.rows) == 2
    assert dataset.rows[0].values["订单号"] == "PO-2026-7788"
    assert dataset.rows[0].values["供应商"] == "星光涂料厂"
    assert dataset.rows[0].values["日期"] == "2026-07-20"
    assert dataset.rows[1].values["品名"] == "PU哑光清漆"


def test_understanding_reuses_prebuilt_workbook_evidence(tmp_path, monkeypatch):
    path = tmp_path / "multi-sheet.xlsx"
    workbook = Workbook()
    workbook.active.title = "发货单"
    workbook.active.append(["品名", "数量"])
    workbook.create_sheet("产品表").append(["产品名称", "型号"])
    workbook.save(path)
    workbook.close()
    evidence = build_workbook_evidence(path)

    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.document_understanding.build_workbook_evidence",
        lambda *_args, **_kwargs: (_ for _ in ()).throw(
            AssertionError("prebuilt evidence must be reused")
        ),
    )

    plan = understand_workbook(
        path,
        hinted_target_type="export_xlsx",
        evidence=evidence,
    )

    assert [item["sheet"] for item in plan["sheet_inventory"]] == ["发货单", "产品表"]
    assert plan["evidence"]["sheet_count"] == 2


def test_low_confidence_fallback_uses_safe_export_instead_of_guessing_business_target(
    monkeypatch,
):
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    path = Path(__file__).parents[1] / "fixtures" / "network_forms" / "net_invoice_sample2.xlsx"

    plan = understand_workbook(
        path,
        hinted_target_type="customer_products",
        hint_confidence=0.35,
    )

    assert plan["source"] == "deterministic_fallback"
    assert plan["recommended_target_type"] == "export_xlsx"
    assert {item["document_type"] for item in plan["documents"]} <= {"generic_table", "invoice"}
    assert plan["requires_confirmation"] is True


def test_document_plan_blocks_line_and_document_total_mismatch(tmp_path, monkeypatch):
    path = _save(
        [
            ["采购订单"],
            ["订单号", "PO-2"],
            ["品名", "数量", "单价", "金额"],
            ["底漆", 2, 10, 19],
            ["面漆", 1, 5, 5],
            ["总计", 30],
        ],
        tmp_path / "mismatch.xlsx",
    )
    plan = {
        "documents": [
            {
                "document_id": "po-2",
                "document_type": "purchase_order",
                "sheet": "业务单",
                "header_fields": [
                    {
                        "role": "document_number",
                        "label": "订单号",
                        "value": "PO-2",
                        "label_cell_id": "s1:r2:c1",
                        "value_cell_id": "s1:r2:c2",
                    }
                ],
                "tables": [
                    {
                        "header_start_row": 3,
                        "header_end_row": 3,
                        "data_start_row": 4,
                        "data_end_row": 5,
                        "first_column": 1,
                        "last_column": 4,
                        "columns": [
                            {"column": 1, "role": "product_name"},
                            {"column": 2, "role": "quantity"},
                            {"column": 3, "role": "unit_price"},
                            {"column": 4, "role": "amount"},
                        ],
                    }
                ],
                "total_amount": 30,
            }
        ]
    }
    from openpyxl.worksheet._read_only import ReadOnlyWorksheet

    def reject_random_cell_access(*_args, **_kwargs):
        raise AssertionError("read-only document parsing must stream ranges")

    monkeypatch.setattr(ReadOnlyWorksheet, "cell", reject_random_cell_access)

    dataset = parse_workbook_with_document_plan(
        path,
        target_type="purchase_orders",
        document_plan=plan,
        max_rows=1000,
    )

    assert dataset is not None
    first_codes = {issue["code"] for issue in dataset.rows[0].provenance["validation_issues"]}
    second_codes = {issue["code"] for issue in dataset.rows[1].provenance["validation_issues"]}
    assert "ETL_LINE_AMOUNT_MISMATCH" in first_codes
    assert "ETL_DOCUMENT_TOTAL_MISMATCH" in first_codes
    assert "ETL_DOCUMENT_TOTAL_MISMATCH" in second_codes


def test_document_plan_parses_two_documents_on_one_sheet(tmp_path):
    path = _save(
        [
            ["采购订单 A"],
            ["订单号", "PO-A"],
            ["品名", "数量"],
            ["底漆", 1],
            [],
            ["采购订单 B"],
            ["订单号", "PO-B"],
            ["品名", "数量"],
            ["面漆", 2],
        ],
        tmp_path / "two-orders.xlsx",
    )
    documents = []
    for document_id, value, header_row, data_row in (
        ("po-a", "PO-A", 3, 4),
        ("po-b", "PO-B", 8, 9),
    ):
        documents.append(
            {
                "document_id": document_id,
                "document_type": "purchase_order",
                "sheet": "业务单",
                "header_fields": [
                    {
                        "role": "document_number",
                        "label": "订单号",
                        "value": value,
                    }
                ],
                "tables": [
                    {
                        "header_start_row": header_row,
                        "header_end_row": header_row,
                        "data_start_row": data_row,
                        "data_end_row": data_row,
                        "first_column": 1,
                        "last_column": 2,
                        "columns": [
                            {"column": 1, "role": "product_name"},
                            {"column": 2, "role": "quantity"},
                        ],
                    }
                ],
            }
        )

    dataset = parse_workbook_with_document_plan(
        path,
        target_type="purchase_orders",
        document_plan={"documents": documents},
        max_rows=1000,
    )

    assert dataset is not None
    assert [(row.values["订单号"], row.values["品名"]) for row in dataset.rows] == [
        ("PO-A", "底漆"),
        ("PO-B", "面漆"),
    ]


def test_sheet_inventory_precedes_document_routing_and_preserves_order(tmp_path):
    path = tmp_path / "inventory.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "两张报价"
    first.append(["报价单 A"])
    first.append(["品名", "数量"])
    first.append(["底漆", 1])
    second = workbook.create_sheet("空表")
    third = workbook.create_sheet("考勤")
    third.append(["姓名", "日期"])
    third.append(["张三", "2026-07-30"])
    workbook.save(path)
    workbook.close()

    evidence = build_workbook_evidence(path)
    documents = [
        {
            "document_id": "quote-a",
            "document_type": "quotation",
            "sheet": "两张报价",
            "tables": [],
        },
        {
            "document_id": "quote-b",
            "document_type": "quotation",
            "sheet": "两张报价",
            "tables": [],
        },
        {
            "document_id": "attendance",
            "document_type": "attendance",
            "sheet": "考勤",
            "tables": [],
        },
    ]

    inventory = build_sheet_inventory(evidence, documents)
    understanding = {
        "documents": documents,
        "document_count": len(documents),
        "sheet_inventory": inventory,
    }
    routes = build_document_routes(understanding, hinted_target_type="export_xlsx")

    assert evidence["workbook_sheet_count"] == 3
    assert [item["sheet"] for item in inventory] == ["两张报价", "空表", "考勤"]
    assert inventory[0]["structure"] == "multi_document"
    assert inventory[0]["document_count"] == 2
    assert inventory[1]["structure"] == "empty"
    assert [route["target_type"] for route in routes] == [
        "export_xlsx",
        "export_xlsx",
        "attendance",
    ]


def test_export_route_keeps_arithmetic_mismatch_as_non_blocking_warning(tmp_path):
    path = _save(
        [
            ["报价单"],
            ["品名", "数量", "单价", "金额"],
            ["底漆", 2, 10, 19],
        ],
        tmp_path / "quotation-warning.xlsx",
    )
    plan = {
        "source": "llm",
        "routing_scope": {"route_id": "quotation"},
        "documents": [
            {
                "document_id": "quotation",
                "document_type": "quotation",
                "sheet": "业务单",
                "header_fields": [],
                "tables": [
                    {
                        "header_start_row": 2,
                        "header_end_row": 2,
                        "data_start_row": 3,
                        "data_end_row": 3,
                        "first_column": 1,
                        "last_column": 4,
                        "columns": [
                            {"column": 1, "role": "product_name"},
                            {"column": 2, "role": "quantity"},
                            {"column": 3, "role": "unit_price"},
                            {"column": 4, "role": "amount"},
                        ],
                    }
                ],
            }
        ],
    }

    dataset = parse_workbook_with_document_plan(
        path,
        target_type="export_xlsx",
        document_plan=plan,
        max_rows=1000,
    )

    assert dataset is not None
    issues = dataset.rows[0].provenance["validation_issues"]
    assert issues == [
        {
            "code": "ETL_LINE_AMOUNT_MISMATCH",
            "severity": "warning",
            "field": "金额",
            "message": "数量 × 单价为 20，与明细金额 19 不一致",
        }
    ]


def test_scoped_empty_document_plan_does_not_fall_back_to_whole_workbook(tmp_path):
    path = _save(
        [
            ["客户", "品名", "数量"],
            ["不应跨路由读取", "底漆", 10],
        ],
        tmp_path / "empty-scope.xlsx",
    )

    dataset = parse_file(
        path,
        target_type="shipment_records",
        document_plan={
            "source": "llm",
            "routing_scope": {"route_id": "empty-document", "sheet": "业务单"},
            "documents": [
                {
                    "document_id": "empty-document",
                    "document_type": "delivery_note",
                    "sheet": "业务单",
                    "header_fields": [],
                    "tables": [],
                }
            ],
        },
    )

    assert dataset.rows == []
    assert dataset.source_features["structure_detection"] == "evidence_document_plan_v1"


def test_overlapping_document_tables_are_deduplicated_by_physical_row(tmp_path):
    path = _save(
        [
            ["报价单"],
            ["品名", "数量"],
            ["底漆", 2],
            ["面漆", 3],
        ],
        tmp_path / "overlap.xlsx",
    )
    table = {
        "header_start_row": 2,
        "header_end_row": 2,
        "data_start_row": 3,
        "data_end_row": 4,
        "first_column": 1,
        "last_column": 2,
        "columns": [
            {"column": 1, "role": "product_name"},
            {"column": 2, "role": "quantity"},
        ],
    }

    dataset = parse_workbook_with_document_plan(
        path,
        target_type="export_xlsx",
        document_plan={
            "source": "llm",
            "routing_scope": {"route_id": "overlap"},
            "documents": [
                {
                    "document_id": "overlap",
                    "document_type": "quotation",
                    "sheet": "业务单",
                    "header_fields": [],
                    "tables": [table, dict(table)],
                }
            ],
        },
        max_rows=100,
    )

    assert dataset is not None
    assert [(row.row_number, row.values["品名"]) for row in dataset.rows] == [
        (3, "底漆"),
        (4, "面漆"),
    ]
    assert dataset.warnings[0]["code"] == "ETL_OVERLAPPING_DOCUMENT_TABLE_ROWS_SKIPPED"


def test_inline_header_fields_are_split_into_canonical_values(tmp_path):
    path = _save(
        [
            ["送货单"],
            ["购货单位：金汉武家私 联系人：张总 2026年01月21日 订单编号：26-010057A"],
            ["品名", "数量"],
            ["底漆", 2],
        ],
        tmp_path / "inline-header.xlsx",
    )
    combined = "购货单位：金汉武家私 联系人：张总 2026年01月21日 订单编号：26-010057A"
    fields = [
        {
            "role": "customer",
            "label": combined,
            "value": combined.split("：", 1)[1],
        }
    ]

    dataset = parse_workbook_with_document_plan(
        path,
        target_type="shipment_records",
        document_plan={
            "source": "llm",
            "routing_scope": {"route_id": "inline"},
            "documents": [
                {
                    "document_id": "inline",
                    "document_type": "delivery_note",
                    "sheet": "业务单",
                    "header_fields": fields,
                    "tables": [
                        {
                            "header_start_row": 3,
                            "header_end_row": 3,
                            "data_start_row": 4,
                            "data_end_row": 4,
                            "first_column": 1,
                            "last_column": 2,
                            "columns": [
                                {"column": 1, "role": "product_name"},
                                {"column": 2, "role": "quantity"},
                            ],
                        }
                    ],
                }
            ],
        },
        max_rows=100,
    )

    assert dataset is not None
    assert dataset.rows[0].values == {
        "客户": "金汉武家私",
        "联系人": "张总",
        "日期": "2026-01-21",
        "订单号": "26-010057A",
        "品名": "底漆",
        "数量": 2,
    }


def test_unlabeled_title_does_not_infer_business_header_roles(tmp_path):
    path = _save(
        [
            ["19/9统计：金汉武色漆样品表"],
            ["编号", "名称"],
            ["S16A", "亮光米灰"],
        ],
        tmp_path / "unlabeled-title.xlsx",
    )
    dataset = parse_workbook_with_document_plan(
        path,
        target_type="export_xlsx",
        document_plan={
            "source": "llm",
            "routing_scope": {"route_id": "catalog"},
            "documents": [
                {
                    "document_id": "catalog",
                    "document_type": "product_catalog",
                    "sheet": "业务单",
                    "header_fields": [
                        {
                            "role": "other",
                            "label": "19/9统计：金汉武色漆样品表",
                            "value": "金汉武色漆样品表",
                        }
                    ],
                    "tables": [
                        {
                            "header_start_row": 2,
                            "header_end_row": 2,
                            "data_start_row": 3,
                            "data_end_row": 3,
                            "first_column": 1,
                            "last_column": 2,
                            "columns": [
                                {"column": 1, "role": "product_code"},
                                {"column": 2, "role": "product_name"},
                            ],
                        }
                    ],
                }
            ],
        },
        max_rows=100,
    )

    assert dataset is not None
    assert dataset.rows[0].values == {"编号": "S16A", "名称": "亮光米灰"}


def test_amount_validation_uses_weight_when_piece_count_is_not_price_basis(tmp_path):
    path = _save(
        [
            ["送货单"],
            ["品名", "数量/件", "数量/KG", "单价/元", "金额/元"],
            ["底漆", 2, 50, 17, 850],
            ["合   计", 2, 50, None, 850],
        ],
        tmp_path / "weight-price.xlsx",
    )
    dataset = parse_workbook_with_document_plan(
        path,
        target_type="shipment_records",
        document_plan={
            "source": "llm",
            "routing_scope": {"route_id": "weight-price"},
            "documents": [
                {
                    "document_id": "weight-price",
                    "document_type": "delivery_note",
                    "sheet": "业务单",
                    "header_fields": [
                        {"role": "customer", "label": "客户", "value": "甲公司"}
                    ],
                    "tables": [
                        {
                            "header_start_row": 2,
                            "header_end_row": 2,
                            "data_start_row": 3,
                            "data_end_row": 4,
                            "first_column": 1,
                            "last_column": 5,
                            "columns": [
                                {"column": 1, "role": "product_name"},
                                {"column": 2, "role": "quantity"},
                                {"column": 3, "role": "other"},
                                {"column": 4, "role": "unit_price"},
                                {"column": 5, "role": "amount"},
                            ],
                        }
                    ],
                }
            ],
        },
        max_rows=100,
    )

    assert dataset is not None
    assert [row.row_number for row in dataset.rows] == [3]
    assert dataset.rows[0].provenance["validation_issues"] == []


def test_document_routes_keep_unsafe_large_tables_out_of_write_targets():
    understanding = {
        "sheet_inventory": [
            {"sheet_index": 1, "sheet": "历史出货"},
            {"sheet_index": 2, "sheet": "产品目录"},
        ],
        "documents": [
            {
                "document_id": "ledger",
                "document_type": "delivery_note",
                "sheet": "历史出货",
                "header_fields": [],
                "tables": [
                    {
                        "data_start_row": 3,
                        "data_end_row": 311,
                    }
                ],
            },
            {
                "document_id": "catalog",
                "document_type": "product_catalog",
                "sheet": "产品目录",
                "header_fields": [],
                "tables": [
                    {
                        "data_start_row": 2,
                        "data_end_row": 30,
                    }
                ],
            },
        ],
    }

    routes = build_document_routes(understanding, hinted_target_type="shipment_records")

    assert [route["target_type"] for route in routes] == [
        "export_xlsx",
        "export_xlsx",
    ]
    assert [route["route_reason"] for route in routes] == [
        "large_table_without_delivery_header",
        "product_owner_context_missing",
    ]
    assert all(route["requires_review"] for route in routes)


def test_safety_routed_export_does_not_report_target_mismatch(tmp_path):
    path = _save(
        [
            ["历史出货"],
            ["品名", "数量"],
            ["底漆", 2],
        ],
        tmp_path / "safety-export.xlsx",
    )
    dataset = parse_workbook_with_document_plan(
        path,
        target_type="export_xlsx",
        document_plan={
            "source": "llm",
            "routing_scope": {
                "route_id": "safe-export",
                "target_type": "export_xlsx",
                "recommended_target_type": "shipment_records",
                "route_reason": "large_table_without_delivery_header",
            },
            "documents": [
                {
                    "document_id": "ledger",
                    "document_type": "delivery_note",
                    "sheet": "业务单",
                    "header_fields": [],
                    "tables": [
                        {
                            "header_start_row": 2,
                            "header_end_row": 2,
                            "data_start_row": 3,
                            "data_end_row": 3,
                            "first_column": 1,
                            "last_column": 2,
                            "columns": [
                                {"column": 1, "role": "product_name"},
                                {"column": 2, "role": "quantity"},
                            ],
                        }
                    ],
                }
            ],
        },
        max_rows=100,
    )

    assert dataset is not None
    codes = {
        issue["code"]
        for row in dataset.rows
        for issue in row.provenance["validation_issues"]
    }
    assert "ETL_DOCUMENT_TARGET_MISMATCH" not in codes


def test_successful_llm_result_still_falls_back_for_missing_sheets(tmp_path, monkeypatch):
    path = tmp_path / "hybrid.xlsx"
    workbook = Workbook()
    first = workbook.active
    first.title = "已识别"
    first.append(["品名", "数量"])
    first.append(["底漆", 1])
    second = workbook.create_sheet("待降级")
    second.append(["客户", "金额"])
    second.append(["甲公司", 100])
    workbook.save(path)
    workbook.close()
    monkeypatch.setattr(
        "app.application.etl.document_understanding.advise_document_understanding",
        lambda *_args, **_kwargs: LlmAssistResult(
            used_llm=True,
            data={
                "file_structure": "one_per_sheet",
                "summary": "第一张表已识别",
                "documents": [
                    {
                        "document_id": "recognized",
                        "document_type": "quotation",
                        "sheet": "已识别",
                        "header_fields": [],
                        "tables": [],
                        "confidence": 0.9,
                        "requires_review": False,
                        "issues": [],
                    }
                ],
            },
        ),
    )

    plan = understand_workbook(
        path,
        hinted_target_type="export_xlsx",
        hint_confidence=0.8,
    )

    assert plan["source"] == "hybrid"
    assert [document["sheet"] for document in plan["documents"]] == [
        "已识别",
        "待降级",
    ]
    assert plan["documents"][0]["document_type"] == "quotation"
    assert plan["documents"][1]["requires_review"] is True


def test_document_eval_manifest_covers_public_and_sanitized_workbooks():
    fixture_dir = Path(__file__).parents[1] / "fixtures" / "network_forms"
    manifest = json.loads((fixture_dir / "document_eval_manifest.json").read_text(encoding="utf-8"))
    cases = manifest["cases"]

    assert len(cases) >= 12
    assert {case["source_kind"] for case in cases} == {"public", "sanitized"}
    for case in cases:
        path = fixture_dir / case["file"]
        assert path.is_file()
        evidence = build_workbook_evidence(path)
        assert evidence["sheets"]
        assert evidence["cell_index"]
        assert evidence["evidence_hash"]


def test_unknown_structured_table_defaults_to_reviewable_export(tmp_path):
    path = tmp_path / "unknown.csv"
    path.write_text("alpha,beta\n1,2\n", encoding="utf-8")

    detection = detect_etl_target(path)

    assert detection["target_type"] == "export_xlsx"
    assert detection["confidence"] < 0.5
