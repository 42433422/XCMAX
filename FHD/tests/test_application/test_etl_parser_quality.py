from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook

from app.application.etl.parsers import parse_file
from app.application.etl.service import EtlService
from app.application.etl.shipment_template_extractor import extract_shipment_template
from app.application.etl.target_detection import detect_etl_target
from app.application.etl.targets import get_adapter
from app.application.etl.transforms import apply_mapping
from app.application.shipment_excel_etl_ocr import _guess_meta_lines
from app.infrastructure.documents.shipment_workbook_filler import fill_shipment_workbook


def _save_workbook(path: Path, build) -> Path:
    workbook = Workbook()
    build(workbook)
    workbook.save(path)
    workbook.close()
    return path


def test_complex_header_skips_preamble_repeated_header_and_footer(tmp_path, monkeypatch):
    path = tmp_path / "complex-customer-products.xlsx"

    def build(workbook):
        sheet = workbook.active
        sheet.title = "业务明细"
        sheet.append(["2026 年客户产品资料导入"])
        sheet.append(["客户信息", None, None, "产品信息", None, None])
        sheet.merge_cells("A2:C2")
        sheet.merge_cells("D2:F2")
        sheet.append(["名称", "电话", "地址", "名称", "型号", "单价"])
        sheet.append([" 甲公司 ", " 13800000000 ", " 上海 ", " 底漆 ", " P-1 ", "￥ 1,200 元"])
        sheet.append(["名称", "电话", "地址", "名称", "型号", "单价"])
        sheet.append(["合计", None, None, None, None, 1200])
        notes = workbook.create_sheet("导入说明")
        notes.append(["请使用标准模板"])
        notes.append(["红色字段为必填"])

    _save_workbook(path, build)
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    dataset = parse_file(path, target_type="customer_products")

    assert len(dataset.rows) == 1
    assert dataset.rows[0].row_number == 4
    assert "客户信息/名称" in dataset.headers
    assert "产品信息/名称" in dataset.headers
    assert dataset.source_features["sheets"][0]["header_depth"] == 2
    assert dataset.source_features["skipped_sheets"] == [
        {"name": "导入说明", "reason": "no_tabular_header"}
    ]
    warning_codes = {warning["code"] for warning in dataset.warnings}
    assert "ETL_REPEATED_HEADERS_SKIPPED" in warning_codes
    assert "ETL_FOOTER_ROWS_SKIPPED" in warning_codes

    mappings = EtlService()._suggest_mappings(
        dataset,
        get_adapter("customer_products"),
    )
    by_target = {mapping["target"]: mapping for mapping in mappings}
    assert by_target["customer_name"]["source"] == "客户信息/名称"
    assert by_target["name"]["source"] == "产品信息/名称"
    assert by_target["price"]["source"] == "单价"
    assert by_target["customer_name"]["confidence"] >= 0.95


def test_multi_sheet_keeps_business_sheets_and_skips_instructions(tmp_path, monkeypatch):
    path = tmp_path / "regions.xlsx"

    def build(workbook):
        east = workbook.active
        east.title = "华东"
        east.append(["客户名称", "产品名称", "型号"])
        east.append(["甲公司", "底漆", "P-1"])
        south = workbook.create_sheet("华南")
        south.append(["客户名称", "产品名称", "型号"])
        south.append(["乙公司", "面漆", "P-2"])
        notes = workbook.create_sheet("README")
        notes.append(["文件填写说明"])

    _save_workbook(path, build)
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    dataset = parse_file(path, target_type="customer_products")

    assert [(row.sheet, row.values["客户名称"]) for row in dataset.rows] == [
        ("华东", "甲公司"),
        ("华南", "乙公司"),
    ]
    assert [sheet["name"] for sheet in dataset.source_features["sheets"]] == ["华东", "华南"]
    assert dataset.source_features["skipped_sheets"][0]["name"] == "README"


def test_mixed_workbook_splits_two_delivery_regions_and_excludes_other_domains(
    tmp_path, monkeypatch
):
    path = tmp_path / "salesperson-workbook.xlsx"

    def build(workbook):
        sheet = workbook.active
        sheet.title = "业务员甲"
        sheet.append(["某公司送货单"])
        sheet.append(["购货单位：甲家具  联系人：张总  2026年01月21日  订单编号：A-1"])
        sheet.append(
            ["产品型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"]
        )
        sheet.append(["P-1", "底漆", 1, 20, 20, 10, 200])
        sheet.append([None, "固化剂", 1, 20, 20, 12, 240])
        sheet.append(["合 计", None, 2, None, None, None, 440])
        sheet.append([])
        sheet.append(["某公司送货单"])
        sheet.append(["购货单位（乙方）：乙家具  联系人：王总  日期2025年04月14日  订单编号：B-1"])
        sheet.append(
            ["产品型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"]
        )
        sheet.append([None, "面漆", 2, 25, 50, 17, 850])
        sheet.append([None, "运费", None, None, None, None, 60])
        sheet.append(["合计", None, 2, None, None, None, 910])
        finance = workbook.create_sheet("回款")
        finance.append(["客户名", "回款金额", "余额"])
        finance.append(["甲家具", 100, 200])
        catalog = workbook.create_sheet("报价")
        catalog.append(["甲家具报价"])
        catalog.append(["名称", "规格", "现金价", "备注"])
        catalog.append(["清漆", 20, 17, ""])

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    dataset = parse_file(path, target_type="customer_products")

    assert len(dataset.rows) == 4
    assert {row.values["customer_name"] for row in dataset.rows} == {"甲家具", "乙家具"}
    assert {row.values["name"] for row in dataset.rows} == {"底漆", "固化剂", "面漆", "清漆"}
    assert all(row.values["customer_name"] != "业务员甲" for row in dataset.rows)
    summary = dataset.source_features["region_summary"]
    assert summary["selected"] == 2
    assert summary["business_rows"] == 4
    assert any(region["role"] == "product_catalog" for region in dataset.source_features["regions"])
    assert any(warning["code"] == "ETL_NON_PRODUCT_CHARGES_SKIPPED" for warning in dataset.warnings)


def test_delivery_workbook_plans_companion_history_without_misreading_finance(
    tmp_path, monkeypatch
):
    """A delivery-note upload exposes companion product data, but never writes it as shipments."""
    path = tmp_path / "multi-sheet-delivery.xlsx"

    def build(workbook):
        delivery = workbook.active
        delivery.title = "送货单"
        delivery.append(["成都国圣送货单"])
        delivery.append(["购货单位：金汉武家私  订单编号：A-1"])
        delivery.append(["型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"])
        delivery.append(["9803", "测试面漆", 1, 20, 20, 17, 340])
        delivery.append(["合计", None, 1, None, 20, None, 340])

        history = workbook.create_sheet("出货历史")
        history.append(
            [
                "金汉武（宾驰）",
                "45659",
                "2",
                "方和",
                None,
                None,
                "黑棕面用修色精",
                3,
                4,
                12,
                48,
                576,
            ]
        )

        quote = workbook.create_sheet("报价")
        quote.append(["金汉武报价"])
        quote.append(["名称", "规格", "单位", "现金价"])
        quote.append(["PU实色漆", 25, "kg/桶", 17])

        reconciliation = workbook.create_sheet("Sheet1")
        reconciliation.append(["名品对账单"])
        reconciliation.append(["欠款年月", "欠款金额/元", "收款金额/元", "期末欠款金额/元"])
        reconciliation.append(["2026年1月", 1000, 0, 1000])

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    shipment = parse_file(path, target_type="shipment_records")
    assert len(shipment.rows) == 1
    assert shipment.rows[0].values["purchase_unit"] == "金汉武家私"
    assert shipment.source_features["shipment_history_product_candidates"] == 2
    shipment_plan = {item["sheet"]: item for item in shipment.source_features["sheet_plan"]}
    assert shipment_plan["出货历史"]["role"] == "supporting_customer_product_data"
    assert shipment_plan["Sheet1"]["role"] == "finance_or_reconciliation"
    assert any(
        warning["code"] == "ETL_COMPANION_CUSTOMER_PRODUCT_DATA_FOUND"
        for warning in shipment.warnings
    )

    customer_products = parse_file(path, target_type="customer_products")
    black = next(
        row for row in customer_products.rows if row.values.get("name") == "黑棕面用修色精"
    )
    assert black.values == {
        "customer_name": "金汉武家私",
        "name": "黑棕面用修色精",
        "specification": 4.0,
        "price": 48.0,
        "model_number": "方和",
    }
    assert all(row.sheet != "Sheet1" for row in customer_products.rows)


def test_companion_history_uses_source_date_not_workbook_row_order(tmp_path, monkeypatch):
    path = tmp_path / "latest-history.xlsx"

    def build(workbook):
        delivery = workbook.active
        delivery.title = "送货单"
        delivery.append(["成都国圣送货单"])
        delivery.append(["购货单位：金汉武家私  订单编号：A-1"])
        delivery.append(["型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"])
        delivery.append(["9803", "测试面漆", 1, 20, 20, 17, 340])
        delivery.append(["合计", None, 1, None, 20, None, 340])
        history = workbook.create_sheet("出货历史")
        # The latest source date intentionally appears first. A simple
        # last-row-wins import would regress the 48 price to 40.
        history.append(
            ["金汉武", 46000, "2", "方和", None, None, "黑棕面用修色精", 1, 4, 4, 48, 192]
        )
        history.append(
            ["金汉武", 45900, "2", "方和", None, None, "黑棕面用修色精", 1, 4, 4, 40, 160]
        )

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    dataset = parse_file(path, target_type="customer_products")
    product = next(row for row in dataset.rows if row.values.get("name") == "黑棕面用修色精")
    assert product.values["price"] == 48.0
    assert product.provenance["source_date"]
    assert dataset.source_features["latest_record_selection"] == {
        "basis": "source_date_then_same_sheet_row",
        "unique_candidates": 1,
        "stale_records_skipped": 1,
        "future_dated_records_skipped": 0,
        "same_date_conflicts": 0,
        "model_identity_ambiguity_groups": 0,
    }
    assert any(
        warning["code"] == "ETL_LATEST_PRODUCT_DATA_SELECTED" for warning in dataset.warnings
    )


def test_shipment_preview_keeps_companion_notice_when_history_has_stale_rows(tmp_path, monkeypatch):
    """Shipment preview must explain that appendices were read despite stale-row selection."""
    path = tmp_path / "shipment-companion-stale-history.xlsx"

    def build(workbook):
        delivery = workbook.active
        delivery.title = "送货单"
        delivery.append(["成都国圣送货单"])
        delivery.append(["购货单位：金汉武家私  订单编号：A-1"])
        delivery.append(["型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"])
        delivery.append(["9803", "测试面漆", 1, 20, 20, 17, 340])
        delivery.append(["合计", None, 1, None, 20, None, 340])

        history = workbook.create_sheet("出货历史")
        history.append(
            ["金汉武", 46000, "2", "方和", None, None, "黑棕面用修色精", 1, 4, 4, 48, 192]
        )
        history.append(
            ["金汉武", 45900, "2", "方和", None, None, "黑棕面用修色精", 1, 4, 4, 40, 160]
        )

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    shipment = parse_file(path, target_type="shipment_records")
    warning_codes = {warning["code"] for warning in shipment.warnings}

    assert shipment.source_features["shipment_history_product_candidates"] == 1
    assert shipment.source_features["latest_record_selection"]["stale_records_skipped"] == 1
    assert "ETL_LATEST_PRODUCT_DATA_SELECTED" in warning_codes
    assert "ETL_COMPANION_CUSTOMER_PRODUCT_DATA_FOUND" in warning_codes


def test_companion_history_excludes_future_dated_fact_from_latest_selection(tmp_path, monkeypatch):
    path = tmp_path / "future-history.xlsx"
    future_date = f"{date.today().year + 1}-01-17"

    def build(workbook):
        delivery = workbook.active
        delivery.title = "送货单"
        delivery.append(["成都国圣送货单"])
        delivery.append(["购货单位：金汉武家私  订单编号：A-1"])
        delivery.append(["型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"])
        delivery.append(["9803", "测试面漆", 1, 20, 20, 17, 340])
        delivery.append(["合计", None, 1, None, 20, None, 340])
        history = workbook.create_sheet("出货历史")
        history.append(
            ["金汉武", "2026-01-17", "2", "方和", None, None, "黑棕面用修色精", 1, 4, 4, 48, 192]
        )
        history.append(
            ["金汉武", future_date, "2", "方和", None, None, "黑棕面用修色精", 1, 4, 4, 99, 396]
        )

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    dataset = parse_file(path, target_type="customer_products")
    product = next(row for row in dataset.rows if row.values.get("name") == "黑棕面用修色精")

    assert product.values["price"] == 48.0
    assert dataset.source_features["latest_record_selection"]["future_dated_records_skipped"] == 1
    warning = next(
        warning for warning in dataset.warnings if warning["code"] == "ETL_FUTURE_DATED_SOURCE_ROW"
    )
    assert warning["rows"] == [{"sheet": "出货历史", "row": 2, "source_date": future_date}]


def test_companion_history_keeps_evidenced_numeric_models_separate(tmp_path, monkeypatch):
    """A numeric SKU after an order marker must not collapse product variants."""
    path = tmp_path / "numeric-history-models.xlsx"

    def build(workbook):
        delivery = workbook.active
        delivery.title = "送货单"
        delivery.append(["成都国圣送货单"])
        delivery.append(["购货单位：金汉武家私  订单编号：A-1"])
        delivery.append(["型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"])
        delivery.append(["9803", "测试面漆", 1, 20, 20, 17, 340])
        delivery.append(["合计", None, 1, None, 20, None, 340])

        history = workbook.create_sheet("出货历史")
        # A positive 9804 variant, its return, and a 6832 variant mirror a
        # real ledger shape: order marker | numeric SKU | blank columns | product.
        history.append(
            ["金汉武", 45702, "1号", 9804, None, None, "PE 白底漆", 5, 28, 140, 9.2, 1288]
        )
        history.append(
            ["金汉武", 45702, "1号", "退货9804", None, None, "PE 白底漆", -3, 28, -84, 9.2, -772.8]
        )
        history.append(["金汉武", 45702, "1号", 6832, None, None, "PE 白底漆", 2, 30, 60, 8, 480])

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    dataset = parse_file(path, target_type="customer_products")
    products = [
        row
        for row in dataset.rows
        if row.values.get("customer_name") == "金汉武家私" and row.values.get("name") == "PE 白底漆"
    ]

    assert {
        (
            row.values.get("model_number"),
            row.values.get("specification"),
            row.values.get("price"),
        )
        for row in products
    } == {("9804", 28.0, 9.2), ("6832", 30.0, 8.0)}
    assert len(products) == 2


def test_modeled_delivery_and_model_less_newer_quote_are_blocked_for_review(tmp_path, monkeypatch):
    """Regression for the 侯雪梅 workbook pattern: never guess a duplicate product."""
    path = tmp_path / "侯雪梅-产品歧义.xlsx"

    def build(workbook):
        delivery = workbook.active
        delivery.title = "送货单"
        delivery.append(["成都国圣送货单"])
        delivery.append(["购货单位：金汉武家私  订单编号：A-1"])
        delivery.append(["型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"])
        delivery.append(["6824A", "PE封固底漆", 1, 20, 20, 12.5, 250])
        delivery.append(["合计", None, 1, None, 20, None, 250])

        quote = workbook.create_sheet("报价")
        quote.append(["金汉武家私报价 2025年05月25日"])
        quote.append(["名称", "规格", "现金价"])
        # The later quote deliberately has no model, just like the real workbook.
        quote.append(["PE封固底漆", 20, 11.3])

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    dataset = parse_file(path, target_type="customer_products")
    conflict_rows = [
        row
        for row in dataset.rows
        if row.values.get("customer_name") == "金汉武家私"
        and row.values.get("name") == "PE封固底漆"
    ]

    assert len(conflict_rows) == 2
    assert {str(row.values.get("model_number") or "") for row in conflict_rows} == {"", "6824A"}
    assert all(
        issue["code"] == "ETL_PRODUCT_MODEL_AMBIGUITY"
        for row in conflict_rows
        for issue in row.provenance["validation_issues"]
    )
    assert (
        dataset.source_features["latest_record_selection"]["model_identity_ambiguity_groups"] == 1
    )
    assert any(warning["code"] == "ETL_PRODUCT_MODEL_AMBIGUITY" for warning in dataset.warnings)


def test_shipment_history_order_note_is_not_treated_as_product_model(tmp_path, monkeypatch):
    """Status annotations must not manufacture a SKU conflict in a ledger."""
    path = tmp_path / "shipment-history-order-note.xlsx"

    def build(workbook):
        delivery = workbook.active
        delivery.title = "送货单"
        delivery.append(["成都国圣送货单"])
        delivery.append(["购货单位：名品（晶美鑫）  订单编号：A-1"])
        delivery.append(["型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"])
        delivery.append(["9803", "测试面漆", 1, 20, 20, 17, 340])
        delivery.append(["合计", None, 1, None, 20, None, 340])

        history = workbook.create_sheet("出货历史")
        # "未签单" is an order-state note in the model column, not a SKU.
        history.append(
            ["名品（晶美鑫）", 46000, "2", "未签单", None, None, "全哑黑面漆", 1, 20, 20, 18, 360]
        )
        history.append(
            ["名品（晶美鑫）", 45900, "2", None, None, None, "全哑黑面漆", 1, 20, 20, 17, 340]
        )

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    dataset = parse_file(path, target_type="customer_products")
    rows = [
        row
        for row in dataset.rows
        if row.values.get("customer_name") == "名品（晶美鑫）"
        and row.values.get("name") == "全哑黑面漆"
    ]

    assert len(rows) == 1
    assert rows[0].values.get("model_number") is None
    assert rows[0].values["price"] == 18.0
    assert rows[0].provenance["source_date"] == "2025-12-09"
    assert not any(
        issue["code"] == "ETL_PRODUCT_MODEL_AMBIGUITY"
        for issue in rows[0].provenance.get("validation_issues", [])
    )


def test_mixed_workbook_projects_delivery_regions_and_reuses_extracted_layout(
    tmp_path, monkeypatch
):
    path = tmp_path / "侯雪梅.xlsx"

    def build(workbook):
        sheet = workbook.active
        sheet.title = "侯雪梅"
        sheet.append(["某公司送货单"])
        sheet.merge_cells("A1:I1")
        sheet.append(["购货单位：甲家具  联系人：张总  2026年01月21日  订单编号：A-1"])
        sheet.merge_cells("A2:I2")
        sheet.append(
            [
                "产品型号",
                None,
                None,
                "产品名称",
                "数量/件",
                "规格/KG",
                "数量/KG",
                "单价/元",
                "金额/元",
            ]
        )
        sheet.merge_cells("A3:C3")
        for row in range(4, 8):
            sheet.merge_cells(start_row=row, end_row=row, start_column=1, end_column=3)
        sheet.append(["P-1", None, None, "底漆", 1, 20, 20, 10, 200])
        sheet.append([None, None, None, "固化剂", 1, 20, 20, 12, 240])
        sheet.append([None])
        sheet.append(
            ["合 计", None, None, None, "=SUM(E4:E7)", None, "=SUM(G4:G7)", None, "=SUM(I4:I7)"]
        )
        sheet.merge_cells("A8:C8")
        sheet.append(["销售协议", "测试协议"])
        sheet.append(["销售单位：某公司"])
        finance = workbook.create_sheet("回款")
        finance.append(["客户名", "回款金额", "余额"])
        finance.append(["甲家具", 100, 200])

    _save_workbook(path, build)
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    monkeypatch.setattr(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile",
        lambda *_args, **_kwargs: None,
    )

    detection = detect_etl_target(path)
    assert detection["target_type"] == "shipment_records"

    dataset = parse_file(path, target_type="shipment_records")
    assert len(dataset.rows) == 2
    assert dataset.source_features["business_document_type"] == "delivery_note"
    assert {
        key: value
        for key, value in dataset.rows[0].values.items()
        if key not in {"source_fingerprint", "legacy_note_fingerprint"}
    } == {
        "purchase_unit": "甲家具",
        "model_number": "P-1",
        "product_name": "底漆",
        "quantity_tins": 1,
        "tin_spec": 20,
        "quantity_kg": 20,
        "unit_price": 10,
        "amount": 200,
        "external_order_no": "A-1",
    }
    assert len(dataset.rows[0].values["source_fingerprint"]) == 64
    assert len(dataset.rows[0].values["legacy_note_fingerprint"]) == 28
    assert all(row.sheet != "回款" for row in dataset.rows)

    template = tmp_path / "侯雪梅-发货单版式.xlsx"
    extracted = extract_shipment_template(
        path,
        source_features=dataset.source_features,
        destination=template,
    )
    assert extracted["source_region_id"] == "侯雪梅!R3C1:9"

    generated = tmp_path / "发货单_TEST-001.xlsx"
    fill_shipment_workbook(
        template,
        output_path=generated,
        unit_name="侯雪梅",
        contact_person="侯女士",
        products=[
            {
                "model_number": "9803",
                "name": "测试面漆",
                "quantity_tins": 3,
                "tin_spec": 28,
                "unit_price": 17,
            }
        ],
        order_number="TEST-001",
    )
    from openpyxl import load_workbook

    workbook = load_workbook(generated, data_only=False)
    try:
        sheet = workbook.active
        assert "购货单位：侯雪梅" in sheet["A2"].value
        assert "订单编号：TEST-001" in sheet["A2"].value
        assert sheet["A4"].value == "9803"
        assert sheet["D4"].value == "测试面漆"
        assert sheet["E4"].value == 3
        assert sheet["F4"].value == 28
        assert sheet["G4"].value == 84
        assert sheet["I4"].value == 1428
        assert any(sheet.cell(row, 1).value == "销售协议" for row in range(1, sheet.max_row + 1))
    finally:
        workbook.close()


def test_shipment_template_stops_before_next_document_route_in_same_sheet(tmp_path):
    path = tmp_path / "two-delivery-notes.xlsx"

    def build(workbook):
        sheet = workbook.active
        sheet.title = "送货单合集"
        sheet.append(["某公司送货单"])
        sheet.append(["购货单位：甲家具  联系人：张总  订单编号：A-1"])
        sheet.append(
            [
                "产品型号",
                None,
                None,
                "产品名称",
                "数量/件",
                "规格/KG",
                "数量/KG",
                "单价/元",
                "金额/元",
                "备注",
            ]
        )
        sheet.append(["P-1", None, None, "底漆", 1, 20, 20, 10, 200])
        sheet.append(["合计", None, None, None, None, None, None, None, 200])
        sheet.append(["销售协议", "测试协议"])
        sheet.append(["销售单位：某公司"])
        sheet.append([])
        sheet.append([])
        sheet.append(["某公司送货单"])
        sheet.append(["购货单位：乙家具  联系人：李总  订单编号：B-1"])
        sheet.append(
            [
                "产品型号",
                None,
                None,
                "产品名称",
                "数量/件",
                "规格/KG",
                "数量/KG",
                "单价/元",
                "金额/元",
                "备注",
            ]
        )
        sheet.append(["P-2", None, None, "面漆", 2, 20, 40, 12, 480])
        sheet.append(["合计", None, None, None, None, None, None, None, 480])

    _save_workbook(path, build)
    source_features = {
        "regions": [
            {
                "id": "route-a",
                "sheet": "送货单合集",
                "status": "selected",
                "header_row": 3,
                "last_column": 10,
            }
        ],
        "document_understanding": {
            "document_routes": [
                {
                    "route_id": "route-a",
                    "sheet": "送货单合集",
                    "data_ranges": [
                        {"header_start_row": 3, "data_start_row": 4, "data_end_row": 4}
                    ],
                },
                {
                    "route_id": "route-b",
                    "sheet": "送货单合集",
                    "data_ranges": [
                        {"header_start_row": 12, "data_start_row": 13, "data_end_row": 13}
                    ],
                },
            ]
        },
    }
    template = tmp_path / "甲家具-发货单版式.xlsx"

    extracted = extract_shipment_template(
        path,
        source_features=source_features,
        destination=template,
        source_region_id="route-a",
    )

    assert extracted["rows"] == 7
    from openpyxl import load_workbook

    workbook = load_workbook(template, data_only=False)
    try:
        sheet = workbook.active
        values = [
            str(sheet.cell(row=row, column=column).value or "")
            for row in range(1, sheet.max_row + 1)
            for column in range(1, sheet.max_column + 1)
        ]
        assert sheet.max_row == 7
        assert str(sheet.print_area) == "'送货单'!$A$1:$J$7"
        assert not any("乙家具" in value or "B-1" in value for value in values)
    finally:
        workbook.close()


def test_csv_preamble_and_dirty_values_are_normalized_deterministically(tmp_path):
    path = tmp_path / "dirty.csv"
    path.write_text(
        "客户产品导出\n"
        "导出时间,2026-07-27\n"
        "客户名称,单价,日期\n"
        '\ufeff\u200b 甲\u3000公司 , "￥ 1,234.50 元" ,20260727\n',
        encoding="utf-8",
    )

    dataset = parse_file(path, target_type="customer_products")

    assert dataset.source_features["header_row"] == 3
    assert dataset.rows[0].row_number == 4
    normalized = apply_mapping(
        dataset.rows[0].values,
        [
            {"source": "客户名称", "target": "customer", "transforms": [{"op": "trim"}]},
            {"source": "单价", "target": "price", "transforms": [{"op": "number"}]},
            {"source": "日期", "target": "date", "transforms": [{"op": "date"}]},
        ],
    )
    assert normalized == {
        "customer": "甲 公司",
        "price": "1234.50",
        "date": "2026-07-27",
    }
    assert dataset.rows[0].provenance["original_fragment"]["客户名称"].startswith("\ufeff")


def test_ocr_provenance_uses_cell_coordinates_for_repeated_text(tmp_path, monkeypatch):
    source = tmp_path / "scan.png"
    source.write_bytes(b"placeholder")
    derived = tmp_path / "ocr-derived.xlsx"

    def build(workbook):
        sheet = workbook.active
        sheet.title = "scan_P1"
        sheet.append(["客户名称", "电话"])
        sheet.append(["同名", "同名"])

    _save_workbook(derived, build)
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_ocr.ocr_source_to_workbook",
        lambda *_args, **_kwargs: {
            "success": True,
            "file_path": str(derived),
            "block_count": 4,
            "meta_lines": [],
            "pages": [
                {
                    "sheet_name": "scan_P1",
                    "page_number": 1,
                    "data_start_row": 1,
                    "blocks": [
                        {"text": "同名", "confidence": 0.99, "left": 999},
                    ],
                    "grid_cells": [
                        {
                            "workbook_row": 2,
                            "workbook_column": 1,
                            "text": "同名",
                            "confidence": 0.95,
                            "left": 10,
                        },
                        {
                            "workbook_row": 2,
                            "workbook_column": 2,
                            "text": "同名",
                            "confidence": 0.6,
                            "left": 100,
                        },
                    ],
                }
            ],
        },
    )

    dataset = parse_file(source, target_type="customers")
    provenance = dataset.rows[0].provenance

    assert provenance["cells"]["客户名称"]["confidence"] == 0.95
    assert provenance["cells"]["客户名称"]["position"]["left"] == 10
    assert provenance["cells"]["电话"]["confidence"] == 0.6
    assert provenance["cells"]["电话"]["position"]["left"] == 100
    assert provenance["low_confidence_fields"] == ["电话"]
    assert provenance["requires_confirmation"] is True


def test_ocr_metadata_detection_does_not_remove_business_header():
    grid = [
        ["客户：甲公司", "日期：2026-07-27", ""],
        ["客户名称", "产品名称", "单价"],
        ["甲公司", "底漆", "100"],
    ]

    assert _guess_meta_lines(grid) == ["客户：甲公司 日期：2026-07-27"]
