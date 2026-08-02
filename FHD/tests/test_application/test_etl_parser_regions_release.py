"""Release regression coverage for the deterministic multi-sheet ETL parser.

These cases use only temporary workbooks.  They exercise the same mixed
delivery/history/quote layouts accepted by the desktop data-import centre.
"""

from __future__ import annotations

from datetime import date, datetime, timedelta

import pytest
from openpyxl import Workbook

from app.application.etl import parser_regions as regions
from app.application.etl import parser_shipment_history as history
from app.application.etl.errors import EtlError
from app.application.etl.parser_types import ParsedRow


def _row(*values, sheet: str = "送货单", number: int = 1, source_date: str = "") -> ParsedRow:
    provenance = {"source_date": source_date} if source_date else {}
    return ParsedRow(sheet=sheet, row_number=number, values=dict(values), provenance=provenance)


def _mixed_workbook(tmp_path):
    workbook = Workbook()
    delivery = workbook.active
    delivery.title = "送货单"
    delivery.append(["送货单"])
    delivery.append(["购货单位：甲方家具有限公司"])
    delivery.append(["联系人：张三 电话：13800138000"])
    delivery.append(["订单编号：DD-001 2025年08月01日"])
    delivery.append(["型号", "产品名称", "数量桶", "每桶kg", "数量kg", "单价", "金额", "备注"])
    delivery.append(["P-01", "白底漆", 2, 5, 10, 20, 200, "首单"])
    delivery.append(["", "运费", 1, 1, 1, 10, 10, "应排除"])
    delivery.append(["合计", "", "", "", "", "", 210, ""])

    historical = workbook.create_sheet("出货历史")
    historical.append(
        ["甲方家具有限公司", "2025-09-10", "1号", "9804", "PE白底漆", 2, 5, 10, 25, 250]
    )

    structured = workbook.create_sheet("销售明细")
    structured.append(["客户名称", "型号", "产品名称", "规格", "数量桶", "单价", "出货日期"])
    structured.append(["甲方家具", "S-01", "清面漆", 5, 2, 28, "2025-10-10"])

    quote = workbook.create_sheet("报价-甲方家具")
    quote.append(["甲方家具有限公司 2025-11-11"])
    quote.append(["产品名称", "规格", "单价"])
    quote.append(["透明底漆", 5, 30])

    finance = workbook.create_sheet("回款对账")
    finance.append(["客户", "回款金额"])
    finance.append(["甲方家具有限公司", 100])

    path = tmp_path / "mixed-delivery.xlsx"
    workbook.save(path)
    return path


def test_header_meta_role_and_total_row_helpers_cover_document_shapes():
    assert regions._field_for_header("产品型号") == "model_number"
    assert regions._field_for_header("货品名称") == "name"
    assert regions._field_for_header("数量kg") == "quantity_kg"
    assert regions._field_for_header("金额") == "amount"
    assert regions._field_for_header("无关列") == ""

    header = regions._header_candidate(["型号", "产品名称", "数量桶", "单价", "金额"])
    assert header is not None
    assert header["by_field"]["model_number"] == 1
    assert regions._header_candidate(["备注", "金额"]) is None

    meta = regions._extract_meta(
        [
            (1, ("购货单位：甲方家具有限公司",)),
            (2, ("联系人：张三 电话：13800138000",)),
            (3, ("订单编号：DD-1 2025年08月01日",)),
        ],
        max_col=8,
    )
    assert meta["customer_name"] == "甲方家具有限公司"
    assert meta["contact_person"] == "张三"
    assert meta["contact_phone"] == "13800138000"
    assert meta["order_number"] == "DD-1"
    assert meta["order_date"] == "2025年08月01日"
    assert (
        regions._region_role(sheet_name="送货单", context_rows=[], meta=meta, header=header)
        == "delivery_note"
    )
    assert (
        regions._region_role(sheet_name="对账", context_rows=[], meta={}, header=header)
        == "finance"
    )
    assert (
        regions._region_role(sheet_name="价目表", context_rows=[], meta={}, header=header)
        == "product_catalog"
    )
    assert (
        regions._region_role(sheet_name="出货记录", context_rows=[], meta={}, header=header)
        == "shipment_ledger"
    )
    assert (
        regions._region_role(sheet_name="Sheet1", context_rows=[], meta={}, header=header)
        == "ignore"
    )
    assert regions._is_total_row(("合计", 100), max_col=2) is True
    assert regions._is_total_row(("白底漆", 100), max_col=2) is False


def test_region_helpers_keep_only_evidenced_latest_facts_and_fingerprints():
    current = _row(
        ("customer_name", "甲方"),
        ("model_number", "P-1"),
        ("specification", 5),
        ("price", 20),
        sheet="送货单",
        number=2,
        source_date="2025-01-01",
    )
    newer = _row(
        ("customer_name", "甲方"),
        ("model_number", "P-1"),
        ("specification", 5),
        ("price", 22),
        sheet="出货历史",
        number=3,
        source_date="2025-02-01",
    )
    same_day_conflict = _row(
        ("customer_name", "甲方"),
        ("model_number", "P-1"),
        ("specification", 6),
        ("price", 22),
        sheet="报价",
        number=4,
        source_date="2025-02-01",
    )
    assert regions._prefer_newer_companion(newer, current) is True
    assert regions._same_date_conflict(same_day_conflict, newer) is True
    regions._mark_same_date_conflict(newer, same_day_conflict)
    assert newer.provenance["validation_issues"][0]["code"] == "ETL_LATEST_SOURCE_CONFLICT"
    assert regions._normalized_order_date("2025年2月3日") == "2025-02-03"
    assert regions._normalized_order_date("日期未知") == ""
    assert regions._is_future_companion(
        _row(("customer_name", "甲方"), source_date=(date.today() + timedelta(days=1)).isoformat())
    )
    assert not regions._is_future_companion(_row(("customer_name", "甲方"), source_date="bad"))

    shipment_one = _row(
        ("purchase_unit", "甲方"),
        ("external_order_no", "DD-1"),
        ("model_number", "P-1"),
        ("product_name", "白底漆"),
        ("quantity_tins", 2),
        ("quantity_kg", 10),
        ("unit_price", 20),
        sheet="送货单",
        number=8,
    )
    shipment_one.provenance["region_id"] = "送货单!R5C1:7"
    shipment_two = _row(
        ("purchase_unit", "甲方"),
        ("external_order_no", "DD-1"),
        ("model_number", "P-2"),
        ("product_name", "清面漆"),
        ("quantity_tins", 1),
        ("quantity_kg", 5),
        ("unit_price", 30),
        sheet="送货单",
        number=9,
    )
    shipment_two.provenance["region_id"] = "送货单!R5C1:7"
    regions._attach_delivery_fingerprints([shipment_one, shipment_two])
    assert (
        shipment_one.values["legacy_note_fingerprint"]
        == shipment_two.values["legacy_note_fingerprint"]
    )
    assert shipment_one.values["source_fingerprint"] != shipment_two.values["source_fingerprint"]


def test_region_sheet_plan_and_shape_helpers_cover_non_target_appendices(tmp_path):
    workbook = Workbook()
    delivery = workbook.active
    delivery.title = "交付"
    delivery.append(["送货单"])
    finance = workbook.create_sheet("Sheet1")
    finance.append(["对账单", "收款金额"])
    catalog = workbook.create_sheet("价格目录")
    catalog.append(["报价", "产品名称"])
    assert regions._sheet_domain_hint(finance) == "finance_or_reconciliation"
    assert regions._sheet_domain_hint(catalog) == "reference_catalog"
    plan = regions._build_sheet_plan(
        workbook_sheet_names=["交付", "Sheet1", "价格目录", "附录"],
        regions=[{"sheet": "交付", "status": "selected", "row_count": 2}],
        companion_sheet_counts={"价格目录": 1},
        sheet_domain_hints={
            "Sheet1": "finance_or_reconciliation",
            "价格目录": "reference_catalog",
        },
    )
    assert [item["status"] for item in plan] == ["included", "excluded", "included", "excluded"]
    assert regions._value_at(("a", "b"), 2) == "b"
    assert regions._value_at(("a",), 2) is None
    assert regions._has_measure(("P-1", 2), {"quantity_tins": 2}) is True
    assert regions._has_measure(("P-1", ""), {"quantity_tins": 2}) is False
    assert regions._unique_source_headers({1: "数量", 2: "数量", 3: "金额"}) == {
        1: "数量",
        2: "数量_2",
        3: "金额",
    }


def test_shipment_history_helpers_parse_headerless_structured_and_quote_sheets(tmp_path):
    assert history.customer_alias_key("甲方家具有限公司（东莞）") == "甲方"
    assert history._number(True) is None
    assert history._number("2.5") == 2.5
    assert history._source_date(datetime(2025, 2, 3, 12, 0)) == "2025-02-03"
    assert history._source_date("2025年2月3日") == "2025-02-03"
    assert history._source_date("9804") == ""
    assert history._looks_like_model("未签单") is False
    assert history._looks_like_model("P-9804") is True
    assert history._nearby_model_number(("甲方", "1号", "9804", "PE白底漆"), 3) == "9804"
    assert history._line_candidate(("甲方", "PE白底漆", 2, 5, 10, 25, 250)) is not None
    assert history._line_candidate(("甲方", "PE白底漆", 2, 5, 8, 25, 250)) is None

    workbook = Workbook()
    ledger = workbook.active
    ledger.title = "出货历史"
    ledger.append(["甲方家具有限公司", "2025-02-03", "1号", "9804", "PE白底漆", 2, 5, 10, 25, 250])
    structured = workbook.create_sheet("销售明细")
    structured.append(["客户名称", "型号", "产品名称", "规格", "数量桶", "单价", "出货日期"])
    structured.append(["甲方家具", "S-1", "清面漆", 5, 2, 28, "2025-03-01"])
    quote = workbook.create_sheet("甲方报价")
    quote.append(["甲方家具有限公司 2025-04-01"])
    quote.append(["产品名称", "规格", "单价"])
    quote.append(["透明底漆", 5, 30])
    finance = workbook.create_sheet("回款明细")
    finance.append(["甲方家具有限公司", "PE白底漆", 2, 5, 10, 25, 250])

    aliases = {"甲方": "甲方家具有限公司"}
    ledger_rows = history.parse_shipment_history_rows(
        ledger, canonical_by_alias=aliases, max_rows=10
    )
    assert ledger_rows[0].values["model_number"] == "9804"
    assert ledger_rows[0].provenance["source_date"] == "2025-02-03"
    structured_rows = history.parse_structured_shipment_history_rows(
        structured, canonical_by_alias=aliases, max_rows=10
    )
    assert structured_rows[0].values["customer_name"] == "甲方家具有限公司"
    quote_rows = history.parse_quote_rows(quote, canonical_by_alias=aliases, max_rows=10)
    assert quote_rows[0].values["name"] == "透明底漆"
    assert (
        history.parse_shipment_history_rows(finance, canonical_by_alias=aliases, max_rows=10) == []
    )
    assert history.product_match_key({"customer_name": "甲方", "model_number": "P-1"}) == (
        "甲方",
        "model",
        "P-1",
    )


def test_mixed_workbook_keeps_delivery_and_latest_companion_product_data(monkeypatch, tmp_path):
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    dataset = regions.parse_customer_product_regions(_mixed_workbook(tmp_path), max_rows=50)
    assert dataset is not None
    assert {row.values.get("name") for row in dataset.rows} >= {
        "白底漆",
        "PE白底漆",
        "清面漆",
        "透明底漆",
    }
    assert all(row.values.get("name") != "运费" for row in dataset.rows)
    warning_codes = {warning["code"] for warning in dataset.warnings}
    assert "ETL_NON_PRODUCT_CHARGES_SKIPPED" in warning_codes
    assert "ETL_SHIPMENT_HISTORY_PRODUCTS_INCLUDED" in warning_codes
    assert dataset.source_features["sheet_plan"][-1]["status"] == "excluded"


def test_shipment_preview_stays_with_delivery_lines_and_has_stable_fingerprints(
    monkeypatch, tmp_path
):
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    dataset = regions.parse_customer_product_regions(
        _mixed_workbook(tmp_path), max_rows=50, target_type="shipment_records"
    )
    assert dataset is not None
    assert len(dataset.rows) == 1
    row = dataset.rows[0]
    assert row.values["purchase_unit"] == "甲方家具有限公司"
    assert row.values["product_name"] == "白底漆"
    assert len(row.values["legacy_note_fingerprint"]) == 28
    assert len(row.values["source_fingerprint"]) == 64
    assert "ETL_COMPANION_CUSTOMER_PRODUCT_DATA_FOUND" in {
        warning["code"] for warning in dataset.warnings
    }


def test_parser_returns_none_for_finance_only_workbook_and_enforces_row_limit(
    monkeypatch, tmp_path
):
    monkeypatch.setenv("FHD_ETL_LLM", "off")
    finance_only = Workbook()
    sheet = finance_only.active
    sheet.title = "回款对账"
    sheet.append(["客户", "回款金额"])
    sheet.append(["甲方", 100])
    finance_path = tmp_path / "finance.xlsx"
    finance_only.save(finance_path)
    assert regions.parse_customer_product_regions(finance_path, max_rows=10) is None

    with pytest.raises(EtlError) as exc_info:
        regions.parse_customer_product_regions(_mixed_workbook(tmp_path), max_rows=0)
    assert exc_info.value.code == "ETL_ROW_LIMIT_EXCEEDED"
