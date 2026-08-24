# mypy: disable-error-code="index, union-attr"
"""Branch/behaviour coverage for the deterministic ETL parsing modules.

Covers:
  * parser_structure.py   (layout detection, header scoring, row classifiers)
  * parser_target_match.py (header hints + required-field coverage)
  * parser_ocr_provenance.py (OCR evidence enrichment)
  * parser_region_projection.py (delivery-region projection / features)
  * parser_types.py        (ParsedDataset / ParsedRow value objects)
  * parsers.py             (parse_file entry-point + workbook/csv/document paths)
  * parser_regions.py      (multi-region customer/product parsing)

Heavy external dependencies (LLM advice, legacy compat parser, OCR extraction)
are stubbed with unittest.mock so the deterministic branches are exercised with
real temporary xlsx/csv files.
"""

from __future__ import annotations

import csv
from pathlib import Path
from unittest import mock
from unittest.mock import patch

import pytest

from app.application.etl.errors import EtlError
from app.application.etl.parser_ocr_provenance import enrich_ocr_provenance
from app.application.etl.parser_region_projection import (
    project_delivery_region,
    region_source_features,
)
from app.application.etl.parser_structure import (
    TableLayout,
    clean_cell_text,
    detect_table_layout,
    header_match_score,
    header_semantic_keys,
    is_auxiliary_sheet_name,
    is_footer_or_note_row,
    is_repeated_header,
    semantic_key,
)
from app.application.etl.parser_target_match import (
    covers_required_target_fields,
    target_header_hints,
)
from app.application.etl.parser_types import ParsedDataset, ParsedRow

# ---------------------------------------------------------------------------
# parser_types.py
# ---------------------------------------------------------------------------


def test_parsed_row_defaults_and_fields():
    row = ParsedRow(sheet="S", row_number=3, values={"a": 1})
    assert row.provenance == {}
    assert row.sheet == "S"
    assert row.row_number == 3
    # dataclass slots + frozen-less mutation
    row.provenance["x"] = 1
    assert row.provenance == {"x": 1}


def test_parsed_dataset_default_warnings():
    ds = ParsedDataset(headers=["a"], rows=[], source_features={"kind": "csv"})
    assert ds.warnings == []
    ds.warnings.append({"code": "X"})
    assert ds.headers == ["a"]
    assert ds.source_features["kind"] == "csv"


# ---------------------------------------------------------------------------
# parser_structure.py
# ---------------------------------------------------------------------------


def test_clean_cell_text_normalizes():
    assert clean_cell_text(None) == ""
    assert clean_cell_text("\ufeffabc\u200b\u00a0def") == "abc def"
    assert clean_cell_text("  a   b  ") == "a b"


def test_semantic_key_drops_punctuation_casefolds():
    assert semantic_key("产品-名称/型号") == "产品名称型号"
    assert semantic_key(123) == "123"
    assert semantic_key("") == ""


def test_header_semantic_keys_composed_parenthesis():
    keys = header_semantic_keys("购买单位/客户")
    assert "购买单位客户" in keys
    assert "客户" in keys
    assert keys[0] == "购买单位客户"


def test_header_semantic_keys_simple():
    keys = header_semantic_keys("产品型号")
    assert "产品型号" in keys
    assert keys[0] == "产品型号"


def test_header_match_score_exact_and_partial():
    assert header_match_score("购货单位", ("购货单位",)) >= 0.95
    assert header_match_score("购买单位", ("购买单位", "客户")) >= 0.98
    # partial: leaf key matches
    assert header_match_score("产品名称/规格", ("规格",)) >= 0.84
    # no overlap -> 0
    assert header_match_score("完全无关", ("型号",)) == 0.0
    # empty header -> 0
    assert header_match_score("", ("型号",)) == 0.0


def test_is_auxiliary_sheet_name():
    assert is_auxiliary_sheet_name("说明文档")
    assert is_auxiliary_sheet_name("封面")
    assert is_auxiliary_sheet_name("read me")
    assert is_auxiliary_sheet_name("Sheet1") is False


def test_is_footer_or_note_row():
    assert is_footer_or_note_row(["合计", "100"]) is True
    assert is_footer_or_note_row(["备注：测试"]) is True
    assert is_footer_or_note_row(["签字"]) is True
    assert is_footer_or_note_row(["A", "B", "C"]) is False
    assert is_footer_or_note_row([]) is False
    assert is_footer_or_note_row(["M01", 10]) is False


def test_is_repeated_header():
    headers = ["产品型号", "产品名称", "数量"]
    assert is_repeated_header(["产品型号", "产品名称", "数量"], headers) is True
    assert is_repeated_header(["M01", "面漆", 5], headers) is False
    assert is_repeated_header(["M01"], headers) is False  # <2 populated


def test_detect_table_layout_empty_and_noise():
    assert detect_table_layout([]) is None
    # single noise cell with no hints and non-data -> low score, still not -50
    noise = [["随便一段文字"]]
    layout = detect_table_layout(noise)
    assert layout is not None
    assert layout.matched_hint_count == 0


def test_detect_table_layout_with_target_header():
    rows = [
        ["某公司送货单"],
        ["产品型号", "产品名称", "数量/桶", "金额/元"],
        ["M01", "面漆", 2, 200],
        ["M02", "底漆", 1, 100],
    ]
    layout = detect_table_layout(rows, header_hints=("产品名称", "产品型号", "数量", "金额"))
    assert layout is not None
    assert layout.matched_hint_count >= 1
    assert layout.header_start == 1  # leading preamble skipped
    assert layout.header_end == 1
    assert "多行表头" not in layout.reasons  # single row header
    assert layout.confidence >= 0.35


def test_detect_table_layout_multi_row_header():
    rows = [
        ["产品信息", "类别"],
        ["型号", "名称"],
        ["M01", "面漆"],
    ]
    layout = detect_table_layout(rows, header_hints=("产品名称", "型号"))
    assert layout is not None
    assert layout.header_start == 0
    assert layout.header_end == 1
    assert "multi_row_header" in layout.reasons
    composed = "".join(layout.headers)
    assert "型号" in composed and "名称" in composed


def test_detect_table_layout_all_data_no_valid_header():
    rows = [
        ["2026-07-01", "M01", "20"],
        ["2026-07-02", "M02", "30"],
    ]
    layout = detect_table_layout(rows, header_hints=("产品名称",))
    # numeric/dates rows score very low -> None
    assert layout is None or layout.matched_hint_count == 0


# ---------------------------------------------------------------------------
# parser_target_match.py
# ---------------------------------------------------------------------------


def test_target_header_hints_known_target():
    hints = target_header_hints("products")
    assert "产品名称" in hints
    assert "型号" in hints


def test_target_header_hints_unknown_target_downgrades_to_empty():
    assert target_header_hints("no_such_target") == []


def test_covers_required_fields_true():
    ds = ParsedDataset(
        headers=["购买单位", "产品名称"],
        rows=[],
        source_features={"kind": "csv"},
    )
    assert covers_required_target_fields(ds, "customer_products") is True


def test_covers_required_fields_false():
    ds = ParsedDataset(headers=["型号", "数量"], rows=[], source_features={"kind": "csv"})
    assert covers_required_target_fields(ds, "customer_products") is False


def test_covers_required_fields_no_required():
    # knowledge has no required fields
    ds = ParsedDataset(headers=[], rows=[], source_features={"kind": "doc"})
    assert covers_required_target_fields(ds, "knowledge") is True


# ---------------------------------------------------------------------------
# parser_region_projection.py
# ---------------------------------------------------------------------------


def test_project_delivery_region_non_shipment_returns_copy():
    values = {"name": "面漆", "price": 10, "extra": "x"}
    out = project_delivery_region(values, target_type="customer_products", meta={})
    assert out == values
    assert out is not values


def test_project_delivery_region_shipment_maps_fields():
    values = {
        "customer_name": "客户A",
        "name": "面漆",
        "specification": "20",
        "price": 10,
        "quantity_tins": 2,
        "quantity_kg": 40,
        "amount": 200,
        "junk": "drop",
    }
    out = project_delivery_region(
        values,
        target_type="shipment_records",
        meta={"order_number": "M-1"},
    )
    assert out["purchase_unit"] == "客户A"
    assert out["product_name"] == "面漆"
    assert out["tin_spec"] == "20"
    assert out["unit_price"] == 10
    assert out["external_order_no"] == "M-1"
    assert "junk" not in out


def test_project_delivery_region_shipment_whitelist_drops_extra():
    out = project_delivery_region(
        {"name": "x", "description": "y"},
        target_type="shipment_records",
        meta={},
    )
    assert "description" not in out
    assert "product_name" in out


def test_region_source_features():
    regions = [
        {"status": "selected", "customer_name": "B客户"},
        {"status": "selected", "customer_name": "A客户"},
        {"status": "excluded"},
    ]
    feats = region_source_features(target_type="shipment_records", regions=regions, rows=5)
    assert feats["kind"] == "workbook_delivery_regions"
    assert feats["suggested_target_type"] == "shipment_records"
    assert feats["region_summary"]["selected"] == 2
    assert feats["region_summary"]["excluded"] == 1
    assert feats["region_summary"]["customers"] == ["A客户", "B客户"]

    feats2 = region_source_features(target_type="customer_products", regions=[], rows=0)
    assert feats2["kind"] == "workbook_regions"
    assert feats2["suggested_target_type"] == "customer_products"
    assert feats2["region_summary"]["customers"] == []


# ---------------------------------------------------------------------------
# parser_ocr_provenance.py
# ---------------------------------------------------------------------------


def _ocr_dataset():
    row = ParsedRow(
        sheet="OCR1",
        row_number=2,
        values={"产品名称": "面漆", "数量": "20", "空字段": ""},
        provenance={"columns": {"产品名称": 1, "数量": 2}},
    )
    return ParsedDataset(headers=["产品名称", "数量"], rows=[row], source_features={})


def test_enrich_ocr_provenance_cells_and_confidence():
    result = {
        "success": True,
        "block_count": 3,
        "meta_lines": ["line1"],
        "pages": [
            {
                "sheet_name": "OCR1",
                "page_number": 1,
                "data_start_row": 2,
                "grid_cells": [
                    {"workbook_row": 2, "workbook_column": 1, "text": "面漆", "confidence": 0.9},
                    {"workbook_row": 2, "workbook_column": 2, "text": "20", "score": 150},
                ],
                "blocks": [],
            }
        ],
    }
    ds = enrich_ocr_provenance(_ocr_dataset(), result, source_suffix=".pdf")
    assert ds.source_features["kind"] == "ocr"
    assert ds.source_features["ocr_page_count"] == 1
    assert ds.source_features["ocr_block_count"] == 3
    row = ds.rows[0]
    assert row.provenance["ocr"] is True
    assert row.provenance["requires_confirmation"] is True
    assert row.provenance["low_confidence_fields"] == []
    assert row.provenance["confidence"] == 0.9
    assert row.provenance["cells"]["产品名称"]["confidence"] == 0.9
    assert row.provenance["cells"]["数量"]["confidence"] == 1.5  # >1 -> /100? no: 150/100=1.5
    assert any(w["code"] == "ETL_OCR_REVIEW_REQUIRED" for w in ds.warnings)


def test_enrich_ocr_provenance_no_evidence_low_confidence():
    result = {
        "success": True,
        "pages": [{"sheet_name": "OCR1", "page_number": 2, "grid_cells": [], "blocks": []}],
    }
    ds = enrich_ocr_provenance(_ocr_dataset(), result, source_suffix=".png")
    row = ds.rows[0]
    assert set(row.provenance["low_confidence_fields"]) == {"产品名称", "数量"}
    assert row.provenance["confidence"] is None
    assert row.provenance["cells"]["产品名称"]["confidence"] is None
    assert row.provenance["page"] == 2
    assert row.provenance["table_position"]["data_start_row"] is None


def test_enrich_ocr_provenance_confidence_under_threshold():
    result = {
        "success": True,
        "pages": [
            {
                "sheet_name": "OCR1",
                "page_number": 1,
                "grid_cells": [
                    {"workbook_row": 2, "workbook_column": 1, "text": "面漆", "confidence": 0.5}
                ],
                "blocks": [],
            }
        ],
    }
    ds = enrich_ocr_provenance(_ocr_dataset(), result, source_suffix=".jpg")
    row = ds.rows[0]
    assert "产品名称" in row.provenance["low_confidence_fields"]
    assert row.provenance["cells"]["产品名称"]["confidence"] == 0.5


def test_enrich_ocr_provenance_block_fallback():
    result = {
        "success": True,
        "pages": [
            {
                "sheet_name": "OCR1",
                "page_number": 1,
                "grid_cells": [],
                "blocks": [
                    {"text": "面漆", "confidence": 0.95, "left": 1, "top": 2, "width": 3},
                    {"text": "999", "confidence": 0.95},
                ],
            }
        ],
    }
    ds = enrich_ocr_provenance(_ocr_dataset(), result, source_suffix=".pdf")
    row = ds.rows[0]
    assert row.provenance["cells"]["产品名称"]["confidence"] == 0.95
    assert row.provenance["cells"]["产品名称"]["position"]["left"] == 1
    # 数量 "20" not in blocks -> low confidence
    assert "数量" in row.provenance["low_confidence_fields"]


# ---------------------------------------------------------------------------
# helpers: build real files
# ---------------------------------------------------------------------------


def _write_csv(path: Path, rows: list[list[str]]) -> Path:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.writer(handle)
        for row in rows:
            writer.writerow(row)
    return path


def _delivery_workbook(path: Path, *, sheet_name: str = "送货") -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws["A1"] = "测试工厂送货单"
    ws["A2"] = "购货单位：混合客户     联系人：张总       2026年07月24日         订单编号：M-1"
    ws["A3"] = "产品型号"
    ws["D3"] = "产品名称"
    ws["E3"] = "数量/件"
    ws["F3"] = "规格/KG"
    ws["G3"] = "数量/KG"
    ws["H3"] = "单价/元"
    ws["I3"] = "金额/元"
    ws["A4"] = "M01"
    ws["D4"] = "面漆"
    ws["E4"] = 1
    ws["F4"] = 20
    ws["G4"] = 20
    ws["H4"] = 10
    ws["I4"] = 200
    ws["A5"] = "M02"
    ws["D5"] = "底漆"
    ws["E5"] = 2
    ws["F5"] = 25
    ws["G5"] = 50
    ws["H5"] = 8
    ws["I5"] = 400
    wb.save(path)
    return path


def _simple_products_workbook(path: Path) -> Path:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "产品"
    ws.append(["购买单位", "产品名称", "规格", "价格"])
    ws.append(["客户甲", "面漆", "20", 10])
    ws.append(["客户乙", "底漆", "25", 8])
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# parsers.py — parse_file
# ---------------------------------------------------------------------------


def test_parse_file_missing_path(tmp_path):
    with pytest.raises(EtlError) as exc:
        from app.application.etl.parsers import parse_file

        parse_file(tmp_path / "nope.xlsx", target_type="products")
    assert exc.value.code == "ETL_UPLOAD_MISSING"


def test_parse_file_unsupported_suffix(tmp_path):
    from app.application.etl.parsers import parse_file

    p = tmp_path / "file.txt"
    p.write_text("hello")
    with pytest.raises(EtlError) as exc:
        parse_file(p, target_type="products")
    assert exc.value.code == "ETL_FILE_TYPE_UNSUPPORTED"


def test_parse_file_docx_wrong_target(tmp_path):
    from app.application.etl.parsers import parse_file

    p = tmp_path / "notes.docx"
    p.write_bytes(b"%PDF-ish")
    with pytest.raises(EtlError) as exc:
        parse_file(p, target_type="products")
    assert exc.value.code == "ETL_KNOWLEDGE_ONLY_FILE"


def test_parse_file_docx_knowledge(tmp_path):
    from app.application.etl.parsers import parse_file

    p = tmp_path / "notes.docx"
    p.write_bytes(b"x")
    ds = parse_file(p, target_type="knowledge")
    assert ds.source_features["kind"] == "document"
    assert ds.source_features["knowledge_only"] is True
    assert ds.rows[0].values["document_path"] == str(p.resolve())


def test_parse_file_xlsx_knowledge_preserves_original_document_and_inventory(tmp_path):
    from openpyxl import Workbook

    from app.application.etl.parsers import parse_file

    p = tmp_path / "mixed-business.xlsx"
    workbook = Workbook()
    workbook.active.title = "发货单"
    workbook.active.append(["购货单位", "产品名称", "数量"])
    workbook.active.append(["国圣化工", "面漆", 10])
    hidden = workbook.create_sheet("对账")
    hidden.sheet_state = "hidden"
    hidden.append(["回款", 100])
    workbook.save(p)

    ds = parse_file(p, target_type="knowledge")

    assert ds.headers == ["document_path", "source_key"]
    assert len(ds.rows) == 1
    assert ds.rows[0].values == {
        "document_path": str(p.resolve()),
        "source_key": p.name,
    }
    assert ds.source_features["kind"] == "document"
    assert ds.source_features["structured_source"] is True
    assert ds.source_features["preserves_original_layout"] is True
    assert ds.source_features["workbook_inventory"] == [
        {"name": "发货单", "state": "visible", "max_row": 2, "max_column": 3},
        {"name": "对账", "state": "hidden", "max_row": 1, "max_column": 2},
    ]


def test_parse_file_csv_with_preset_rejected(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _write_csv(tmp_path / "a.csv", [["产品名称"], ["面漆"]])
    with pytest.raises(EtlError) as exc:
        parse_file(p, target_type="products", compatibility_preset_id="legacy")
    assert exc.value.code == "ETL_COMPATIBILITY_PRESET_FILE_UNSUPPORTED"


def test_parse_file_csv_happy_path(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _write_csv(
        tmp_path / "a.csv",
        [
            ["购买单位", "产品名称", "规格", "价格"],
            ["客户甲", "面漆", "20", "10"],
            ["客户乙", "底漆", "25", "8"],
        ],
    )
    ds = parse_file(p, target_type="products")
    assert ds.source_features["kind"] == "csv"
    assert len(ds.rows) == 2
    assert ds.rows[0].values["产品名称"] == "面漆"
    assert ds.headers and "产品名称" in ds.headers


def test_parse_file_csv_row_limit(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _write_csv(
        tmp_path / "a.csv",
        [
            ["购买单位", "产品名称", "规格", "价格"],
            ["客户甲", "面漆", "20", "10"],
            ["客户乙", "底漆", "25", "8"],
        ],
    )
    with pytest.raises(EtlError) as exc:
        parse_file(p, target_type="products", max_rows=1)
    assert exc.value.code == "ETL_ROW_LIMIT_EXCEEDED"


def test_parse_file_xlsx_compat_preset_wrong_target(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _simple_products_workbook(tmp_path / "a.xlsx")
    with pytest.raises(EtlError) as exc:
        parse_file(p, target_type="knowledge", compatibility_preset_id="legacy")
    assert exc.value.code == "ETL_COMPATIBILITY_PRESET_TARGET_MISMATCH"


def test_parse_file_xlsx_customer_products_regional_preferred(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _delivery_workbook(tmp_path / "delivery.xlsx")
    # patch legacy compat parser so the regional (preferred) path is the only one
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = None
        ds = parse_file(p, target_type="customer_products")
    assert ds.source_features["structure_detection"] == "deterministic_regions_v1"
    assert len(ds.rows) >= 1
    assert ds.rows[0].values.get("customer_name") == "混合客户"


def test_parse_file_xlsx_shipment_records_regional(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _delivery_workbook(tmp_path / "delivery.xlsx")
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = None
        ds = parse_file(p, target_type="shipment_records")
    assert ds.source_features["structure_detection"] == "deterministic_regions_v1"
    assert len(ds.rows) >= 1
    assert "source_fingerprint" in ds.rows[0].values


def test_parse_file_xlsx_generic_workbook(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _simple_products_workbook(tmp_path / "products.xlsx")
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = None
        ds = parse_file(p, target_type="products")
    assert ds.source_features["structure_detection"] == "deterministic_v2"
    assert len(ds.rows) == 2
    assert ds.rows[0].values["产品名称"] == "面漆"


def test_parse_file_xlsx_generic_preferred_over_compat(tmp_path):
    from app.application.etl.parsers import ParsedDataset, ParsedRow, parse_file

    p = _simple_products_workbook(tmp_path / "products.xlsx")
    compat_ds = ParsedDataset(
        headers=["产品名称"],
        rows=[ParsedRow(sheet="CSV", row_number=1, values={"产品名称": "底漆"})],
        source_features={"kind": "compat"},
    )
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = compat_ds
        ds = parse_file(p, target_type="products")
    # generic rows present and cover required fields -> generic preferred
    assert ds.source_features["structure_detection"] == "deterministic_v2"
    assert any(w["code"] == "ETL_GENERIC_STRUCTURE_PREFERRED" for w in ds.warnings)


def test_parse_file_xlsx_compat_fallback_on_generic_error(tmp_path):
    from openpyxl import Workbook

    from app.application.etl.parsers import ParsedDataset, ParsedRow, parse_file

    # workbook whose generic parse fails
    p = tmp_path / "weird.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "数据"
    ws.append(["随便", "文本", "没有", "表头"])
    ws.append(["a", "b", "c", "d"])
    wb.save(p)

    compat_ds = ParsedDataset(
        headers=["产品名称"],
        rows=[ParsedRow(sheet="CSV", row_number=1, values={"产品名称": "底漆"})],
        source_features={"kind": "compat"},
    )
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = compat_ds
        with patch("app.application.etl.parsers._parse_workbook", side_effect=RuntimeError("boom")):
            ds = parse_file(p, target_type="products")
    assert ds.source_features["kind"] == "compat"


def test_parse_file_xlsx_compat_preset_no_match(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _simple_products_workbook(tmp_path / "products.xlsx")
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = None
        with pytest.raises(EtlError) as exc:
            parse_file(p, target_type="products", compatibility_preset_id="legacy")
    assert exc.value.code == "ETL_COMPATIBILITY_PRESET_NO_MATCH"


def test_parse_file_xlsx_compat_preset_returned(tmp_path):
    from app.application.etl.parsers import ParsedDataset, ParsedRow, parse_file

    p = _simple_products_workbook(tmp_path / "products.xlsx")
    compat_ds = ParsedDataset(
        headers=["产品名称"],
        rows=[ParsedRow(sheet="CSV", row_number=1, values={"产品名称": "底漆"})],
        source_features={"kind": "compat"},
    )
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = compat_ds
        ds = parse_file(p, target_type="products", compatibility_preset_id="legacy")
    assert ds.source_features["kind"] == "compat"


def test_parse_file_xlsx_shipment_records_uses_compat_directly(tmp_path):
    from app.application.etl.parsers import ParsedDataset, ParsedRow, parse_file

    p = _simple_products_workbook(tmp_path / "products.xlsx")
    compat_ds = ParsedDataset(
        headers=["purchase_unit"],
        rows=[ParsedRow(sheet="CSV", row_number=1, values={"purchase_unit": "客户"})],
        source_features={"kind": "compat"},
    )
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = compat_ds
        ds = parse_file(p, target_type="shipment_records")
    assert ds.source_features["kind"] == "compat"


def test_parse_file_ocr_success(tmp_path):
    from app.application.etl.parsers import parse_file

    # OCR module returns an xlsx path; parse_file then runs _parse_workbook on it
    ocr_out_path = _simple_products_workbook(tmp_path / "ocr_out.xlsx")
    src_pdf = tmp_path / "scan.pdf"
    src_pdf.write_bytes(b"%PDF-1.4")
    with patch("app.application.shipment_excel_etl_ocr.ocr_source_to_workbook") as ocr_mock:
        ocr_mock.return_value = {
            "success": True,
            "file_path": str(ocr_out_path),
            "block_count": 2,
            "meta_lines": ["meta"],
            "pages": [{"sheet_name": "产品", "page_number": 1}],
        }
        ds = parse_file(src_pdf, target_type="products")
    assert ds.source_features["kind"] == "ocr"
    assert ds.source_features["ocr_block_count"] == 2
    assert len(ds.rows) == 2
    assert any(w["code"] == "ETL_OCR_REVIEW_REQUIRED" for w in ds.warnings)


def test_parse_file_ocr_failure(tmp_path):
    from app.application.etl.parsers import parse_file

    src = tmp_path / "img.pdf"
    src.write_bytes(b"pdf")
    with patch("app.application.shipment_excel_etl_ocr.ocr_source_to_workbook") as ocr_mock:
        ocr_mock.return_value = {"success": False, "error_code": "low_confidence"}
        with pytest.raises(EtlError) as exc:
            parse_file(src, target_type="products")
    assert exc.value.code == "LOW_CONFIDENCE"


def test_parse_file_ocr_with_compat_preset_rejected(tmp_path):
    from app.application.etl.parsers import parse_file

    src = tmp_path / "img.pdf"
    src.write_bytes(b"pdf")
    with pytest.raises(EtlError) as exc:
        parse_file(src, target_type="products", compatibility_preset_id="legacy")
    assert exc.value.code == "ETL_COMPATIBILITY_PRESET_FILE_UNSUPPORTED"


# ---------------------------------------------------------------------------
# parsers.py — _parse_workbook auxiliary/repeated/footer handling
# ---------------------------------------------------------------------------


def test_parse_workbook_auxiliary_sheet_and_repeated_footer(tmp_path):
    from openpyxl import Workbook

    from app.application.etl.parsers import parse_file

    path = tmp_path / "mixed.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "产品"
    ws.append(["购买单位", "产品名称", "规格", "价格"])
    ws.append(["客户甲", "面漆", "20", 10])
    ws.append(["客户甲", "面漆", "20", 10])  # repeated header-like row not matched
    main = wb.create_sheet("说明")
    main.append(["这是说明工作表", "无业务字段", "x", "y"])
    wb.save(path)

    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = None
        ds = parse_file(path, target_type="products")
    assert ds.source_features["kind"] == "workbook"
    # auxiliary sheet skipped warning present
    assert any(w["code"] == "ETL_AUXILIARY_SHEETS_SKIPPED" for w in ds.warnings)


def test_parse_workbook_footer_and_repeated_headers(tmp_path):
    from openpyxl import Workbook

    from app.application.etl.parsers import parse_file

    path = tmp_path / "foot.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "产品"
    ws.append(["购买单位", "产品名称", "规格", "价格"])
    ws.append(["客户甲", "面漆", "20", 10])
    ws.append(["购买单位", "产品名称", "规格", "价格"])  # repeated header
    ws.append(["合计", "", "", ""])  # footer
    wb.save(path)

    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = None
        ds = parse_file(path, target_type="products")
    codes = {w["code"] for w in ds.warnings}
    assert "ETL_REPEATED_HEADERS_SKIPPED" in codes
    assert "ETL_FOOTER_ROWS_SKIPPED" in codes
    assert len(ds.rows) == 1


def test_parse_workbook_row_limit(tmp_path):
    from openpyxl import Workbook

    from app.application.etl.parsers import parse_file

    path = tmp_path / "many.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "产品"
    ws.append(["购买单位", "产品名称", "规格", "价格"])
    ws.append(["客户甲", "面漆", "20", 10])
    ws.append(["客户乙", "底漆", "25", 8])
    wb.save(path)

    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = None
        with pytest.raises(EtlError) as exc:
            parse_file(path, target_type="products", max_rows=1)
    assert exc.value.code == "ETL_ROW_LIMIT_EXCEEDED"


def test_parse_workbook_knowledge_keeps_original_source(tmp_path):
    from openpyxl import Workbook

    from app.application.etl.parsers import parse_file

    # Knowledge ingestion keeps the original workbook instead of flattening it
    # into generic rows; the database preview performs tabular parsing separately.
    path = tmp_path / "kb.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "知识"
    ws.append(["标题", "内容"])
    ws.append(["第一条", "正文内容"])
    wb.save(path)

    ds = parse_file(path, target_type="knowledge")
    assert ds.source_features["kind"] == "document"
    assert ds.source_features["workbook_inventory"][0]["name"] == "知识"
    assert len(ds.rows) == 1
    assert ds.rows[0].values["document_path"] == str(path.resolve())


# ---------------------------------------------------------------------------
# parser_regions.py — deterministic multi-region parser
# ---------------------------------------------------------------------------


@pytest.fixture
def _fake_llm_advice():
    from app.application.etl.llm_assist import LlmAssistResult

    with patch(
        "app.application.etl.parser_regions.advise_workbook_regions",
        return_value=LlmAssistResult(),
    ):
        yield


def test_parse_regions_customer_products(tmp_path, _fake_llm_advice):
    from app.application.etl.parser_regions import parse_customer_product_regions

    p = _delivery_workbook(tmp_path / "delivery.xlsx")
    ds = parse_customer_product_regions(p, max_rows=1000, target_type="customer_products")
    assert ds is not None
    assert ds.source_features["structure_detection"] == "deterministic_regions_v1"
    assert len(ds.rows) >= 1
    assert ds.rows[0].values.get("customer_name") == "混合客户"
    assert any(w["code"] == "ETL_MULTI_REGION_WORKBOOK_PLANNED" for w in ds.warnings)


def test_parse_regions_shipment_records(tmp_path, _fake_llm_advice):
    from app.application.etl.parser_regions import parse_customer_product_regions

    p = _delivery_workbook(tmp_path / "delivery.xlsx")
    ds = parse_customer_product_regions(p, max_rows=1000, target_type="shipment_records")
    assert ds is not None
    assert len(ds.rows) >= 1
    assert "source_fingerprint" in ds.rows[0].values
    assert ds.rows[0].provenance["source_kind"] == "delivery_note_region"


def test_parse_regions_no_target_returns_none(tmp_path, _fake_llm_advice):
    from app.application.etl.parser_regions import parse_customer_product_regions

    p = tmp_path / "plain.xlsx"
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["a", "b", "c"])
    ws.append([1, 2, 3])
    wb.save(p)
    ds = parse_customer_product_regions(p, max_rows=1000, target_type="customer_products")
    assert ds is None


def test_parse_regions_row_limit(tmp_path, _fake_llm_advice):
    from app.application.etl.parser_regions import parse_customer_product_regions

    p = _delivery_workbook(tmp_path / "delivery.xlsx")
    with pytest.raises(EtlError) as exc:
        parse_customer_product_regions(p, max_rows=1, target_type="customer_products")
    assert exc.value.code == "ETL_ROW_LIMIT_EXCEEDED"


def test_parse_regions_finance_sheet_excluded(tmp_path, _fake_llm_advice):
    from openpyxl import Workbook

    from app.application.etl.parser_regions import parse_customer_product_regions

    p = tmp_path / "finance.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "回款对账"
    ws.append(["日期", "客户", "金额"])
    ws.append(["2026-07-01", "客户A", 100])
    wb.save(p)

    ds = parse_customer_product_regions(p, max_rows=1000, target_type="customer_products")
    # finance-only sheet -> no parsed rows -> None
    assert ds is None


def test_parse_regions_region_role_helpers():
    # exercise private helpers directly for branch coverage
    from app.application.etl import parser_regions as pr

    # _field_for_header mapping branches
    assert pr._field_for_header("产品型号") == "model_number"
    assert pr._field_for_header("货号") == "model_number"
    assert pr._field_for_header("产品名称") == "name"
    assert pr._field_for_header("品名") == "name"
    assert pr._field_for_header("数量/KG") == "quantity_kg"
    assert pr._field_for_header("数量/件") == "quantity_tins"
    assert pr._field_for_header("数量") == "quantity_tins"
    assert pr._field_for_header("规格") == "specification"
    assert pr._field_for_header("单价") == "price"
    assert pr._field_for_header("金额") == "amount"
    assert pr._field_for_header("备注") == "description"
    assert pr._field_for_header("这个不是字段") == ""

    # _header_candidate
    cand = pr._header_candidate(["产品型号", "产品名称", "数量/件", "金额/元"])
    assert cand is not None
    assert cand["identity_count"] == 2
    assert cand["commerce_count"] >= 1
    # identities only branch
    cand2 = pr._header_candidate(["产品型号", "产品名称", "x", "y"])
    assert cand2 is not None and cand2["identity_count"] == 2
    # no identity -> None
    assert pr._header_candidate(["单价", "金额"]) is None

    # _is_total_row
    assert pr._is_total_row(["合计", 100], max_col=10) is True
    assert pr._is_total_row(["", ""], max_col=10) is False
    assert pr._is_total_row(["M01", 10], max_col=10) is False

    # _joined_row truncation
    long = pr._joined_row(["a" * 3000], max_col=10)
    assert len(long) <= 2000

    # _value_at bounds
    assert pr._value_at(["a", "b"], 0) is None
    assert pr._value_at(["a", "b"], 3) is None
    assert pr._value_at(["a", "b"], 1) == "a"
    assert pr._value_at(["a", "b"], None) is None

    # _has_measure
    assert pr._has_measure(["x", 5], {"price": 2}) is True
    assert pr._has_measure(["x", ""], {"price": 2}) is False

    # _unique_source_headers
    assert pr._unique_source_headers({1: "型号", 2: "型号"}) == {1: "型号", 2: "型号_2"}

    # _normalized_order_date
    assert pr._normalized_order_date("2026年07月24日") == "2026-07-24"
    assert pr._normalized_order_date("没有日期") == ""
    assert pr._normalized_order_date("2026-13-99") == ""

    # _companion_source_date / _is_future_companion
    past = ParsedRow(sheet="S", row_number=1, values={}, provenance={"source_date": "2000-01-01"})
    bad = ParsedRow(sheet="S", row_number=2, values={}, provenance={"source_date": "not-a-date"})
    from datetime import date

    future = ParsedRow(
        sheet="S",
        row_number=3,
        values={},
        provenance={"source_date": (date.today().replace(year=date.today().year + 1)).isoformat()},
    )
    assert pr._is_future_companion(past) is False
    assert pr._is_future_companion(bad) is False
    assert pr._is_future_companion(future) is True

    # _same_date_conflict
    row_a = ParsedRow(
        sheet="A",
        row_number=1,
        values={"price": 10},
        provenance={"source_date": "2026-07-24"},
    )
    row_b = ParsedRow(
        sheet="B",
        row_number=1,
        values={"price": 20},
        provenance={"source_date": "2026-07-24"},
    )
    assert pr._same_date_conflict(row_a, row_b) is True
    assert pr._same_date_conflict(row_a, row_a) is False
    row_b2 = ParsedRow(
        sheet="B",
        row_number=1,
        values={"price": 20},
        provenance={"source_date": "2026-07-25"},
    )
    assert pr._same_date_conflict(row_a, row_b2) is False

    # _mark_same_date_conflict de-dup
    pr._mark_same_date_conflict(row_a, row_a)
    issues = row_a.provenance["validation_issues"]
    assert len(issues) == 1
    pr._mark_same_date_conflict(row_a, row_a)
    assert len(row_a.provenance["validation_issues"]) == 1

    # _prefer_newer_companion
    newer = ParsedRow(
        sheet="B",
        row_number=2,
        values={},
        provenance={"source_date": "2026-07-25"},
    )
    assert pr._prefer_newer_companion(newer, row_a) is True
    assert pr._prefer_newer_companion(row_a, newer) is False
    # candidate has date, current none
    assert (
        pr._prefer_newer_companion(
            newer, ParsedRow(sheet="C", row_number=1, values={}, provenance={})
        )
        is True
    )
    # same date, different sheet -> False
    same = ParsedRow(
        sheet="C",
        row_number=1,
        values={},
        provenance={"source_date": "2026-07-24"},
    )
    assert pr._prefer_newer_companion(same, row_a) is False
    # same sheet higher row -> True
    same_master = ParsedRow(
        sheet="A",
        row_number=5,
        values={},
        provenance={"source_date": "2026-07-24"},
    )
    assert pr._prefer_newer_companion(same_master, row_a) is True


def test_parse_regions_private_helpers():
    from app.application.etl import parser_regions as pr

    # _build_sheet_plan: every branch
    plan = pr._build_sheet_plan(
        workbook_sheet_names=["送货", "出货", "回款", "对账明细", "价目表", "杂项"],
        regions=[{"status": "selected", "sheet": "送货", "row_count": 3}],
        companion_sheet_counts={"出货": 5},
        sheet_domain_hints={"回款": "finance_or_reconciliation", "价目表": "reference_catalog"},
    )
    roles = {p["sheet"]: p["role"] for p in plan}
    assert roles["送货"] == "delivery_note_template_and_records"
    assert roles["出货"] == "supporting_customer_product_data"
    assert roles["回款"] == "finance_or_reconciliation"
    assert roles["对账明细"] == "finance_or_reconciliation"  # via sheet-name regex
    assert roles["价目表"] == "reference_catalog"
    assert roles["杂项"] == "non_target_appendix"

    # _sheet_domain_hint
    class FakeWS:
        def __init__(self, title, content):
            self.title = title
            self._content = content

        def iter_rows(self, max_row=20, max_col=18, values_only=True):
            yield from self._content

    assert pr._sheet_domain_hint(FakeWS("Sheet1", [["对账单", "收款金额"], ["", ""]])) == (
        "finance_or_reconciliation"
    )
    assert pr._sheet_domain_hint(FakeWS("Sheet2", [["价目表", "红漆"], ["", ""]])) == (
        "reference_catalog"
    )
    assert pr._sheet_domain_hint(FakeWS("Sheet3", [["普通", "数据"]])) == ""

    # _extract_meta
    meta = pr._extract_meta(
        [
            (1, ("购货单位：测试公司  联系人：张三  电话：13800138000",)),
            (2, ("订单编号：M-OP-1  2026年07月24日",)),
            (3, ("",)),
        ],
        max_col=10,
    )
    assert meta["customer_name"] == "测试公司"
    assert meta["contact_person"] == "张三"
    assert meta["contact_phone"] == "13800138000"
    assert meta["order_number"] == "M-OP-1"
    assert meta["order_date"] == "2026年07月24日"
    assert meta["evidence_rows"]

    # _region_role branches
    assert (
        pr._region_role(sheet_name="对账", context_rows=[], meta={}, header={"last_col": 5})
        == "finance"
    )
    assert (
        pr._region_role(
            sheet_name="Sheet1",
            context_rows=[(1, ("送货",))],
            meta={"customer_name": "测试公司"},
            header={"last_col": 5, "commerce_count": 2},
        )
        == "delivery_note"
    )
    assert (
        pr._region_role(sheet_name="价目表", context_rows=[], meta={}, header={"last_col": 5})
        == "product_catalog"
    )
    assert (
        pr._region_role(
            sheet_name="Sheet1",
            context_rows=[(1, ("出货",))],
            meta={},
            header={"last_col": 5},
        )
        == "shipment_ledger"
    )
    assert (
        pr._region_role(
            sheet_name="Sheet1",
            context_rows=[(1, ("普通内容",))],
            meta={},
            header={"last_col": 5},
        )
        == "ignore"
    )

    # _attach_delivery_fingerprints: grouping + non-numeric number branch
    r1 = ParsedRow(
        sheet="A",
        row_number=1,
        values={
            "model_number": "M1",
            "product_name": "面漆",
            "quantity_tins": 2,
            "quantity_kg": 40,
            "unit_price": 10,
            "purchase_unit": "客户",
            "external_order_no": "X-1",
        },
        provenance={"region_id": "R1", "order_date": "2026-07-24"},
    )
    r2 = ParsedRow(
        sheet="A",
        row_number=2,
        values={
            "model_number": "M2",
            "product_name": "底漆",
            "quantity_tins": "abc",
            "quantity_kg": 50,
            "unit_price": 8,
        },
        provenance={"region_id": "R1"},
    )
    r3 = ParsedRow(sheet="A", row_number=3, values={}, provenance={})
    pr._attach_delivery_fingerprints([r1, r2, r3])
    assert "legacy_note_fingerprint" in r1.values
    assert "source_fingerprint" in r1.values
    assert "legacy_note_fingerprint" in r2.values  # "abc" -> 0.0 via except branch
    assert "legacy_note_fingerprint" not in r3.values  # no region_id


def test_parse_regions_companion_and_conflict_warnings(tmp_path, _fake_llm_advice):
    from openpyxl import Workbook

    from app.application.etl.parser_regions import parse_customer_product_regions
    from app.application.etl.parser_types import ParsedRow

    p = _delivery_workbook(tmp_path / "delivery.xlsx")
    from openpyxl import load_workbook

    wb = load_workbook(p)
    # add a companion sheet (not finance/catalog/delivery) to the delivery workbook
    ledger = wb.create_sheet("出货流水")
    ledger.append(["日期", "型号", "品名", "数量", "价格"])
    ledger.append(["2026-07-01", "C01", "面漆C", 1, 10])
    wb.save(p)

    today = None
    from datetime import date

    future_iso = date.today().replace(year=date.today().year + 1).isoformat()

    def _row(rn, model, name, price, sdate, kind="shipment_history_ledger"):
        return ParsedRow(
            sheet="出货流水",
            row_number=rn,
            values={
                "customer_name": "混合客户",
                "model_number": model,
                "name": name,
                "price": price,
            },
            provenance={"source_date": sdate, "source_kind": kind},
        )

    companion_rows = [
        # Same product + same day but *different sheet* than the delivery note ->
        # a same-day cross-sheet conflict (delivery M01 is 10 on "送货").
        _row(1, "M01", "面漆", 20, "2026-07-24"),
        # Newer business date for the same product -> replaces delivery M01.
        _row(2, "M01", "面漆", 12, "2026-07-25"),
        # Future-dated record -> isolated, never treated as latest fact.
        _row(3, "C03", "面漆C", 11, future_iso),
    ]
    with (
        patch(
            "app.application.etl.parser_regions.parse_shipment_history_rows",
            return_value=companion_rows,
        ),
        patch(
            "app.application.etl.parser_regions.parse_structured_shipment_history_rows",
            return_value=[],
        ),
        patch(
            "app.application.etl.parser_regions.parse_quote_rows",
            return_value=[],
        ),
    ):
        ds = parse_customer_product_regions(p, max_rows=1000, target_type="customer_products")
    assert ds is not None
    codes = {w["code"] for w in ds.warnings}
    assert "ETL_LATEST_SOURCE_CONFLICT" in codes
    assert "ETL_FUTURE_DATED_SOURCE_ROW" in codes
    assert "ETL_LATEST_PRODUCT_DATA_SELECTED" in codes
    assert "ETL_SHIPMENT_HISTORY_PRODUCTS_INCLUDED" in codes


def test_parse_regions_model_ambiguity_warning(tmp_path, _fake_llm_advice):
    from openpyxl import load_workbook

    from app.application.etl.parser_regions import parse_customer_product_regions
    from app.application.etl.parser_types import ParsedRow

    p = _delivery_workbook(tmp_path / "delivery.xlsx")
    wb = load_workbook(p)
    ledger = wb.create_sheet("出货流水")
    ledger.append(["日期", "型号", "品名", "数量", "价格"])
    ledger.append(["2026-07-01", "C01", "面漆C", 1, 10])
    wb.save(p)

    modeled = ParsedRow(
        sheet="出货流水",
        row_number=1,
        values={"customer_name": "混合客户", "model_number": "C01", "name": "面漆C", "price": 6},
        provenance={"source_date": "2026-07-21", "source_kind": "shipment_history_ledger"},
    )
    nomodel = ParsedRow(
        sheet="出货流水",
        row_number=2,
        values={"customer_name": "混合客户", "name": "面漆C", "price": 5},
        provenance={"source_date": "2026-07-20", "source_kind": "shipment_history_ledger"},
    )
    with (
        patch(
            "app.application.etl.parser_regions.parse_shipment_history_rows",
            return_value=[modeled, nomodel],
        ),
        patch(
            "app.application.etl.parser_regions.parse_structured_shipment_history_rows",
            return_value=[],
        ),
        patch(
            "app.application.etl.parser_regions.parse_quote_rows",
            return_value=[],
        ),
    ):
        ds = parse_customer_product_regions(p, max_rows=1000, target_type="customer_products")
    assert ds is not None
    assert any(w["code"] == "ETL_PRODUCT_MODEL_AMBIGUITY" for w in ds.warnings)


def test_parse_regions_shipment_fingerprint_non_numeric(tmp_path, _fake_llm_advice):
    from openpyxl import Workbook

    from app.application.etl.parser_regions import parse_customer_product_regions

    p = tmp_path / "shipment.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "送货"
    ws["A1"] = "购货单位：客户甲  订单编号：S-1  2026年07月24日"
    ws["A2"] = "产品型号"
    ws["D2"] = "产品名称"
    ws["E2"] = "数量/件"
    ws["G2"] = "数量/KG"
    ws["H2"] = "单价/元"
    ws["A3"] = "S01"
    ws["D3"] = "面漆"
    ws["E3"] = "N/A"  # non-numeric quantity_tins
    ws["G3"] = 20
    ws["H3"] = 10
    wb.save(p)

    ds = parse_customer_product_regions(p, max_rows=1000, target_type="shipment_records")
    assert ds is not None
    assert "source_fingerprint" in ds.rows[0].values


# ---------------------------------------------------------------------------
# parsers.py — additional branch coverage
# ---------------------------------------------------------------------------


def test_parse_file_csv_no_tabular_header(tmp_path):
    from app.application.etl.parsers import parse_file

    # Empty CSV: no detectable table layout -> empty dataset (edge case).
    p = _write_csv(tmp_path / "empty.csv", [])
    ds = parse_file(p, target_type="products")
    assert ds.source_features["kind"] == "csv"
    assert ds.rows == []


def test_parse_file_csv_repeated_header_and_footer(tmp_path):
    from app.application.etl.parsers import parse_file

    p = _write_csv(
        tmp_path / "a.csv",
        [
            ["购买单位", "产品名称", "规格", "价格"],
            ["客户甲", "面漆", "20", "10"],
            ["购买单位", "产品名称", "规格", "价格"],  # repeated header
            ["合计", "", "", ""],  # footer
        ],
    )
    ds = parse_file(p, target_type="products")
    codes = {w["code"] for w in ds.warnings}
    assert "ETL_REPEATED_HEADERS_SKIPPED" in codes
    assert "ETL_FOOTER_ROWS_SKIPPED" in codes
    assert len(ds.rows) == 1


def test_parse_file_xlsx_regional_raises_non_etl_falls_back(tmp_path):
    from app.application.etl.parsers import ParsedDataset, ParsedRow, parse_file

    p = _delivery_workbook(tmp_path / "delivery.xlsx")
    compat_ds = ParsedDataset(
        headers=["产品名称"],
        rows=[ParsedRow(sheet="CSV", row_number=1, values={"产品名称": "底漆"})],
        source_features={"kind": "compat"},
    )
    with patch(
        "app.application.etl.shipment_compat_parser.parse_delivery_note_with_compat_profile"
    ) as compat:
        compat.return_value = compat_ds
        with patch(
            "app.application.etl.parser_regions.parse_customer_product_regions",
            side_effect=RuntimeError("boom"),
        ):
            ds = parse_file(p, target_type="customer_products")
    assert ds.source_features["kind"] == "compat"
