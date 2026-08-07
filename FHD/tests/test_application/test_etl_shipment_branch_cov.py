"""ETL 发货/转换模块行为测试（分支覆盖补测）。

覆盖 app/application/etl/ 下 8 个模块：
  shipment_compat_parser / shipment_preview_fallback / shipment_template_extractor
  parser_shipment_history / target_detection / transforms / mapping_assist / product_identity
"""

from __future__ import annotations

from contextlib import contextmanager
from datetime import date, datetime
from pathlib import Path
from typing import Any
from unittest.mock import MagicMock, patch

import pytest

from app.application.etl import (
    mapping_assist,
    parser_shipment_history,
    product_identity,
    shipment_compat_parser,
    shipment_preview_fallback,
    shipment_template_extractor,
    target_detection,
    transforms,
)
from app.application.etl.errors import EtlError
from app.application.etl.parser_types import ParsedDataset, ParsedRow
from app.application.etl.targets import TargetAdapter, TargetField


# ---------------------------------------------------------------------------
# transforms.py
# ---------------------------------------------------------------------------
class TestTransforms:
    def test_neutralize_spreadsheet_formula(self):
        assert transforms.neutralize_spreadsheet_formula("=SUM(A1)") == "'=SUM(A1)"
        assert transforms.neutralize_spreadsheet_formula("+1") == "'+1"
        assert transforms.neutralize_spreadsheet_formula("-1") == "'-1"
        assert transforms.neutralize_spreadsheet_formula("@cmd") == "'@cmd"
        assert transforms.neutralize_spreadsheet_formula("  =bad") == "'  =bad"
        assert transforms.neutralize_spreadsheet_formula(123) == 123
        assert transforms.neutralize_spreadsheet_formula("plain") == "plain"

    def test_apply_transform_trim(self):
        assert transforms.apply_transform("  a\tb  ", {"op": "trim"}, {}) == "a b"
        assert transforms.apply_transform("\ufeff\u200ba\nb\u3000", {"op": "trim"}, {}) == "a b"
        assert transforms.apply_transform(123, {"op": "trim"}, {}) == 123

    def test_apply_transform_cast(self):
        assert transforms.apply_transform(None, {"op": "cast", "type": "string"}, {}) == ""
        assert transforms.apply_transform(5, {"op": "cast", "type": "string"}, {}) == "5"
        assert transforms.apply_transform("1,234.5", {"op": "cast", "type": "float"}, {}) == 1234.5
        assert transforms.apply_transform("12", {"op": "cast", "type": "number"}, {}) == "12"
        assert transforms.apply_transform("12", {"op": "cast", "type": "integer"}, {}) == 12
        assert transforms.apply_transform("", {"op": "cast", "type": "number"}, {}) == ""
        assert transforms.apply_transform("", {"op": "cast", "type": "integer"}, {}) == ""
        assert transforms.apply_transform(True, {"op": "cast", "type": "boolean"}, {}) is True
        assert transforms.apply_transform("是", {"op": "cast", "type": "boolean"}, {}) is True
        assert transforms.apply_transform("否", {"op": "cast", "type": "boolean"}, {}) is False
        assert transforms.apply_transform("no", {"op": "cast", "type": "boolean"}, {}) is False
        assert transforms.apply_transform("2026-08-01", {"op": "cast", "type": "date"}, {}) == "2026-08-01"
        with pytest.raises(EtlError):
            transforms.apply_transform("maybe", {"op": "cast", "type": "boolean"}, {})
        with pytest.raises(EtlError):
            transforms.apply_transform("x", {"op": "cast", "type": "point"}, {})

    def test_apply_transform_date(self):
        assert transforms.apply_transform(None, {"op": "date"}, {}) == ""
        assert transforms.apply_transform(datetime(2026, 8, 1, 12, 0), {"op": "date"}, {}) == "2026-08-01"
        assert transforms.apply_transform(date(2026, 8, 2), {"op": "date"}, {}) == "2026-08-02"
        assert transforms.apply_transform("20260801", {"op": "date"}, {}) == "2026-08-01"
        assert transforms.apply_transform("2026年08月03日", {"op": "date"}, {}) == "2026-08-03"
        # %d/%m/%Y is tried before %m/%d/%Y in the default format list
        assert transforms.apply_transform("01/02/2026", {"op": "date"}, {"f": "x"}) == "2026-02-01"
        assert (
            transforms.apply_transform(
                "2026-08-04T10:00:00", {"op": "date"}, {}
            )
            == "2026-08-04"
        )
        with pytest.raises(EtlError):
            transforms.apply_transform("not-adate", {"op": "date"}, {})

    def test_apply_transform_number_and_default(self):
        assert transforms.apply_transform("", {"op": "number"}, {}) == ""
        assert transforms.apply_transform("(1,200)", {"op": "number"}, {}) == "-1200"
        assert transforms.apply_transform("￥50", {"op": "number"}, {}) == "50"
        assert transforms.apply_transform("30元", {"op": "number"}, {}) == "30"
        with pytest.raises(EtlError):
            transforms.apply_transform("abc", {"op": "number"}, {})
        assert transforms.apply_transform("", {"op": "default", "value": "D"}, {}) == "D"
        assert transforms.apply_transform("keep", {"op": "default", "value": "D"}, {}) == "keep"

    def test_apply_transform_map_split_concat(self):
        assert transforms.apply_transform("a", {"op": "map", "values": {"a": "甲"}}, {}) == "甲"
        assert transforms.apply_transform("z", {"op": "map", "values": {"a": "甲"}, "fallback": "F"}, {}) == "F"
        assert transforms.apply_transform("z", {"op": "map", "values": {"a": "甲"}}, {}) == "z"
        with pytest.raises(EtlError):
            transforms.apply_transform("a", {"op": "map", "values": []}, {})
        assert transforms.apply_transform("x,y,z", {"op": "split", "delimiter": ",", "index": 1}, {}) == "y"
        assert transforms.apply_transform("x,y", {"op": "split", "delimiter": ",", "index": 9}, {}) == ""
        assert transforms.apply_transform(None, {"op": "split", "delimiter": ","}, {}) == ""
        assert (
            transforms.apply_transform("IGN", {"op": "concat", "fields": ["a", "b"], "separator": "-"}, {"a": "甲", "b": "乙"})
            == "甲-乙"
        )
        with pytest.raises(EtlError):
            transforms.apply_transform("x", {"op": "concat", "fields": "bad"}, {})

    def test_apply_transform_formula(self):
        row = {"a": 10, "b": 5}
        assert transforms.apply_transform(None, {"op": "formula", "operator": "add", "operands": [{"field": "a"}, {"literal": 2}]}, row) == "12"
        assert transforms.apply_transform(None, {"op": "formula", "operator": "sub", "operands": [{"field": "a"}, {"field": "b"}]}, row) == "5"
        assert transforms.apply_transform(None, {"op": "formula", "operator": "mul", "operands": [{"field": "a"}, {"field": "b"}]}, row) == "50"
        assert transforms.apply_transform(None, {"op": "formula", "operator": "div", "operands": [{"field": "a"}, {"field": "b"}]}, row) == "2"
        assert transforms.apply_transform(None, {"op": "formula", "operator": "coalesce", "operands": [{"field": "missing"}, {"literal": "fallback"}]}, row) == "fallback"
        with pytest.raises(EtlError):
            transforms.apply_transform(None, {"op": "formula", "operator": "div", "operands": [{"field": "a"}, {"literal": 0}]}, row)
        with pytest.raises(EtlError):
            transforms.apply_transform(None, {"op": "formula", "operator": "pow", "operands": [{"field": "a"}]}, row)
        with pytest.raises(EtlError):
            transforms.apply_transform(None, {"op": "formula", "operator": "add", "operands": []}, row)
        with pytest.raises(EtlError):
            transforms.apply_transform(None, {"op": "formula", "operator": "add", "operands": [{"both": 1}]}, row)
        with pytest.raises(EtlError):
            transforms.apply_transform(None, {"op": "evil"}, row)

    def test_apply_mapping(self):
        out = transforms.apply_mapping(
            {"src": " 漆  "},
            [
                {"target": "name", "source": "src", "transforms": [{"op": "trim"}]},
                {"target": "price", "transforms": [{"op": "number"}], "source": "price"},
            ],
        )
        assert out["name"] == "漆"
        assert out["price"] == ""
        # empty target skipped
        assert transforms.apply_mapping({"a": 1}, [{"target": "", "source": "a"}]) == {}
        # transforms must be list
        with pytest.raises(EtlError):
            transforms.apply_mapping({"a": 1}, [{"target": "t", "source": "a", "transforms": "x"}])
        # rule must be dict
        with pytest.raises(EtlError):
            transforms.apply_mapping({"a": 1}, [{"target": "t", "source": "a", "transforms": ["x"]}])


# ---------------------------------------------------------------------------
# product_identity.py
# ---------------------------------------------------------------------------
class TestProductIdentity:
    def test_product_name_key_and_token(self):
        assert product_identity.product_name_key({"unit": " 甲 ", "name": " 漆 "}) == ("甲", "漆")
        assert product_identity.product_name_key({"customer_name": "甲", "name": "漆"}, unit_field="customer_name") == ("甲", "漆")
        assert product_identity.model_token("  ABC  ") == "abc"
        assert product_identity.model_token(None) == ""

    def test_validation_issue(self):
        issue = product_identity.validation_issue()
        assert issue["code"] == "ETL_PRODUCT_MODEL_AMBIGUITY"
        assert issue["severity"] == "error"

    def test_provenance_validation_issues(self):
        assert product_identity.provenance_validation_issues(None) == []
        assert product_identity.provenance_validation_issues({}) == []
        pv = {
            "validation_issues": [
                {"code": "C1", "message": "m1", "field": "f", "severity": "error"},
                {"code": "", "message": "no code"},
                {"code": "C2", "message": ""},
                "not-a-dict",
                {"code": "C3", "message": "m3", "field": "f3", "severity": ""},
            ]
        }
        issues = product_identity.provenance_validation_issues(pv)
        assert len(issues) == 2
        assert issues[0]["code"] == "C1"
        assert issues[1]["severity"] == "error"

    def test_source_model_ambiguity_issues(self):
        rows = [
            {"unit": "甲", "name": "漆", "model_number": "M1"},
            {"unit": "甲", "name": "漆", "model_number": ""},
            {"unit": "甲", "name": "其他", "model_number": "M2"},
            {"unit": "", "name": "漆", "model_number": "M3"},
            {"unit": "乙", "name": "漆", "model_number": "M9"},
            {"unit": "乙", "name": "漆", "model_number": "M10"},
        ]
        result = product_identity.source_model_ambiguity_issues(rows, unit_field="unit")
        assert set(result) == {0, 1}
        assert product_identity.MODEL_AMBIGUITY_CODE in {i["code"] for i in result[0]}

    def test_candidate_model_token(self):
        assert product_identity.candidate_model_token({"model_number": "  M1  "}) == "m1"
        assert product_identity.candidate_model_token({"after": {"model_number": "M2"}}) == "m2"
        assert product_identity.candidate_model_token({"name": "x"}) == ""

        class _Obj:
            model_number = "M3"

        assert product_identity.candidate_model_token(_Obj()) == "m3"

    def test_database_model_ambiguity_issue(self):
        assert product_identity.database_model_ambiguity_issue({"model_number": "M1"}, [], exact_match=True) is None
        # incoming missing model but known candidate -> issue
        issue = product_identity.database_model_ambiguity_issue(
            {"model_number": ""}, [{"model_number": "M1"}], exact_match=False
        )
        assert issue is not None
        # incoming model, non-exact, but a candidate lacks model -> issue
        issue2 = product_identity.database_model_ambiguity_issue(
            {"model_number": "M2"}, [{"model_number": ""}], exact_match=False
        )
        assert issue2 is not None
        # incoming model and all candidates have models, inexact -> safe
        assert product_identity.database_model_ambiguity_issue(
            {"model_number": "M2"}, [{"model_number": "M9"}], exact_match=False
        ) is None


# ---------------------------------------------------------------------------
# target_detection.py
# ---------------------------------------------------------------------------
class TestTargetDetection:
    def test_knowledge_only_suffix(self, tmp_path):
        path = tmp_path / "a.docx"
        path.write_text("x")
        out = target_detection.detect_etl_target(path)
        assert out["target_type"] == "knowledge"
        assert out["confidence"] == 1.0

    def test_csv_classify_branches(self, tmp_path):
        cases = [
            ("送货单 产品名称 数量", "shipment_records"),
            ("考勤 上班时间 打卡", "attendance"),
            ("采购订单 供应商 产品名称", "purchase_orders"),
            ("客户名称 产品名称", "customer_products"),
            ("客户 购货单位", "customers"),
            ("商品名称 品名", "products"),
            ("随便 的数据", "customer_products"),
        ]
        for content, expected in cases:
            path = tmp_path / "f.csv"
            path.write_text(content, encoding="utf-8")
            out = target_detection.detect_etl_target(path)
            assert out["target_type"] == expected, content

    def test_xlsx_detection(self, tmp_path):
        from openpyxl import Workbook

        path = tmp_path / "w.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "送货单"
        ws.append(["产品名称", "数量"])
        wb.save(path)
        wb.close()
        out = target_detection.detect_etl_target(path)
        assert out["target_type"] == "shipment_records"

    def test_filename_delivery(self, tmp_path):
        path = tmp_path / "发货单.pdf"
        path.write_text("x")
        out = target_detection.detect_etl_target(path, suffix=".pdf")
        assert out["target_type"] == "shipment_records"
        assert out["reason"] == "delivery_filename"

    def test_unknown_suffix(self, tmp_path):
        path = tmp_path / "other.pdf"
        path.write_text("x")
        out = target_detection.detect_etl_target(path)
        assert out["target_type"] == "customer_products"
        assert out["reason"] == "manual_review_required"


# ---------------------------------------------------------------------------
# parser_shipment_history.py
# ---------------------------------------------------------------------------
def _make_sheet(title: str, rows: list[list]) -> Any:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    return ws


class TestParserShipmentHistory:
    def test_customer_alias_key(self):
        # 有限公司/家具 suffix markers and parenthetical markers are stripped
        assert parser_shipment_history.customer_alias_key("星光家具（华南）有限公司") == "星光"
        assert parser_shipment_history.customer_alias_key(None) == ""

    def test_product_match_key(self):
        assert parser_shipment_history.product_match_key({"customer_name": "甲", "model_number": "M1"}) == ("甲", "model", "M1")
        assert parser_shipment_history.product_match_key({"customer_name": "甲", "name": "漆"}) == ("甲", "name", "漆")

    def test_parse_shipment_history_rows(self):
        ws = _make_sheet(
            "出货单",
            [
                ["客户甲", "PE白底漆", 2, 25, 50, 10, 500],
                ["客户甲", "面漆", 3, 20, 60, 12, 720],
                [],  # empty row -> candidate None
                ["客户乙", "1号", "9804", "PE白底漆", 2, 25, 50, 10, 500],
            ],
        )
        rows = parser_shipment_history.parse_shipment_history_rows(
            ws, canonical_by_alias={"客户甲": "客户甲"}, max_rows=10
        )
        assert len(rows) == 3
        assert rows[0].values["customer_name"] == "客户甲"
        assert rows[0].values["name"] == "PE白底漆"
        assert rows[0].values["specification"] == 25
        assert rows[2].values["model_number"] == "9804"
        assert rows[0].provenance["source_kind"] == "shipment_history_ledger"

    def test_parse_shipment_history_finance_sheet(self):
        ws = _make_sheet("对账表", [["客户甲", "漆", 1, 10, 10, 5, 50]])
        assert parser_shipment_history.parse_shipment_history_rows(ws, canonical_by_alias={}, max_rows=5) == []

    def test_parse_shipment_history_finance_row(self):
        # sheet ok but a row explicitly names a finance concept
        ws = _make_sheet("出货单", [["回款", "漆", 1, 10, 10, 5, 50]])
        assert parser_shipment_history.parse_shipment_history_rows(ws, canonical_by_alias={}, max_rows=5) == []

    def test_parse_shipment_history_unnamed_low_evidence(self):
        ws = _make_sheet("明细", [["客户甲", "漆", 1, 10, 10, 5, 50]])
        rows = parser_shipment_history.parse_shipment_history_rows(ws, canonical_by_alias={}, max_rows=5)
        assert rows == []

    def test_parse_shipment_history_max_rows(self):
        ws = _make_sheet(
            "出货单",
            [
                ["客户甲", "漆A", 1, 10, 10, 5, 50],
                ["客户甲", "漆B", 1, 10, 10, 5, 50],
                ["客户甲", "漆C", 1, 10, 10, 5, 50],
            ],
        )
        rows = parser_shipment_history.parse_shipment_history_rows(ws, canonical_by_alias={}, max_rows=2)
        assert len(rows) == 2

    def test_parse_structured_shipment_history_rows(self):
        ws = _make_sheet(
            "出货明细",
            [
                ["客户", "型号", "品名", "规格", "数量桶", "单价"],
                ["客户甲", "M1", "面漆", 20, 2, 15],
                ["客户乙", "", "清漆", 10, 1, 99],
            ],
        )
        rows = parser_shipment_history.parse_structured_shipment_history_rows(
            ws, canonical_by_alias={"客户甲": "客户甲"}, max_rows=10
        )
        assert len(rows) == 2
        assert rows[0].values["model_number"] == "M1"
        assert "model_number" not in rows[1].values
        assert rows[0].values["customer_name"] == "客户甲"

    def test_parse_structured_finance_sheet(self):
        ws = _make_sheet("对账", [["客户", "品名", "规格", "数量桶", "单价"]])
        assert (
            parser_shipment_history.parse_structured_shipment_history_rows(ws, canonical_by_alias={}, max_rows=5)
            == []
        )

    def test_parse_structured_no_header(self):
        ws = _make_sheet("出货明细", [["a", "b", "c", "d", "e", "f"]])
        assert (
            parser_shipment_history.parse_structured_shipment_history_rows(ws, canonical_by_alias={}, max_rows=5)
            == []
        )

    def test_parse_quote_rows(self):
        ws = _make_sheet(
            "报价A",
            [
                ["客户甲 报价单"],
                ["品名", "规格", "单价"],
                ["面漆", 20, 15],
                ["清漆", "", 10],
            ],
        )
        rows = parser_shipment_history.parse_quote_rows(ws, canonical_by_alias={"客户甲": "客户甲"}, max_rows=5)
        assert len(rows) == 1
        assert rows[0].values["customer_name"] == "客户甲"
        assert rows[0].values["name"] == "面漆"
        assert rows[0].values["price"] == 15

    def test_parse_quote_not_quote(self):
        ws = _make_sheet("普通", [["客户甲"], ["品名", "规格", "单价"], ["面漆", 20, 15]])
        assert parser_shipment_history.parse_quote_rows(ws, canonical_by_alias={"客户甲": "客户甲"}, max_rows=5) == []

    def test_parse_quote_no_customer(self):
        ws = _make_sheet("报价B", [["品名", "规格", "单价"], ["面漆", 20, 15]])
        assert parser_shipment_history.parse_quote_rows(ws, canonical_by_alias={}, max_rows=5) == []


# ---------------------------------------------------------------------------
# shipment_compat_parser.py
# ---------------------------------------------------------------------------
_NOTE = {
    "unit_name": "客户甲",
    "order_number": "ORD-1",
    "sheet": "送货单",
    "contact_person": "王",
    "contact_phone": "123",
    "contact_address": "addr",
    "fingerprint": "fp",
    "profile_id": "p1",
    "items": [
        {
            "model_number": "M1",
            "product_name": "漆",
            "quantity_kg": 20,
            "quantity_tins": 2,
            "tin_spec": 10,
            "unit_price": 5,
            "amount": 100,
        },
        {"model_number": "合计", "product_name": "合计"},
        {"model_number": "大写人民币壹佰", "product_name": "漆"},
    ],
}


class TestShipmentCompatParser:
    def _patched_preview(self, monkeypatch, result):
        monkeypatch.setattr(
            "app.application.shipment_excel_etl_app_service.preview_shipment_excel_etl",
            lambda *a, **k: result,
        )

    def test_non_xlsx_returns_none(self, tmp_path):
        path = tmp_path / "a.csv"
        path.write_text("x")
        assert (
            shipment_compat_parser.parse_delivery_note_with_compat_profile(
                path, target_type="customers", max_rows=10
            )
            is None
        )

    def test_unsupported_target_type(self, tmp_path):
        path = tmp_path / "a.xlsx"
        path.write_text("x")
        assert (
            shipment_compat_parser.parse_delivery_note_with_compat_profile(
                path, target_type="attendance", max_rows=10
            )
            is None
        )

    def test_preview_exception_returns_none(self, tmp_path, monkeypatch):
        def _boom(*a, **k):
            raise RuntimeError("preview down")

        monkeypatch.setattr(
            "app.application.shipment_excel_etl_app_service.preview_shipment_excel_etl", _boom
        )
        path = tmp_path / "a.xlsx"
        path.write_text("x")
        assert (
            shipment_compat_parser.parse_delivery_note_with_compat_profile(
                path, target_type="customers", max_rows=10
            )
            is None
        )

    def test_not_success_or_bad_shape(self, tmp_path, monkeypatch):
        path = tmp_path / "a.xlsx"
        path.write_text("x")
        for result in ({"success": False, "notes": []}, {"success": True, "notes": "x"}, {"success": True, "notes": []}):
            self._patched_preview(monkeypatch, result)
            assert (
                shipment_compat_parser.parse_delivery_note_with_compat_profile(
                    path, target_type="customers", max_rows=10
                )
                is None
            )

    def test_customers_target(self, tmp_path, monkeypatch):
        self._patched_preview(monkeypatch, {"success": True, "notes": [_NOTE]})
        path = tmp_path / "a.xlsx"
        path.write_text("x")
        out = shipment_compat_parser.parse_delivery_note_with_compat_profile(
            path, target_type="customers", max_rows=10
        )
        assert out is not None
        assert out.rows[0].values["customer_name"] == "客户甲"
        assert out.rows[0].values["contact_person"] == "王"
        assert out.rows[0].values["contact_phone"] == "123"
        assert out.rows[0].values["contact_address"] == "addr"
        assert out.source_features["compatibility_preset"] is True

    def test_customer_products_target(self, tmp_path, monkeypatch):
        self._patched_preview(monkeypatch, {"success": True, "notes": [_NOTE]})
        path = tmp_path / "a.xlsx"
        path.write_text("x")
        out = shipment_compat_parser.parse_delivery_note_with_compat_profile(
            path, target_type="customer_products", max_rows=10, compatibility_preset_id="p1"
        )
        assert out is not None
        row = out.rows[0]
        assert row.values["name"] == "漆"
        assert row.values["customer_name"] == "客户甲"
        assert row.values["model_number"] == "M1"
        assert out.source_features["compatibility_preset_id"] == "p1"
        # total row skipped
        assert len(out.rows) == 1

    def test_products_target(self, tmp_path, monkeypatch):
        self._patched_preview(monkeypatch, {"success": True, "notes": [_NOTE]})
        path = tmp_path / "a.xlsx"
        path.write_text("x")
        out = shipment_compat_parser.parse_delivery_note_with_compat_profile(
            path, target_type="products", max_rows=10
        )
        assert out is not None
        assert out.rows[0].values["unit"] == "客户甲"

    def test_shipment_records_target(self, tmp_path, monkeypatch):
        self._patched_preview(monkeypatch, {"success": True, "notes": [_NOTE]})
        path = tmp_path / "a.xlsx"
        path.write_text("x")
        out = shipment_compat_parser.parse_delivery_note_with_compat_profile(
            path, target_type="shipment_records", max_rows=10
        )
        assert out is not None
        row = out.rows[0]
        assert row.values["purchase_unit"] == "客户甲"
        assert row.values["external_order_no"] == "ORD-1"
        assert row.values["quantity_tins"] == 2
        assert row.values["quantity_kg"] == 20
        assert bool(row.values["source_fingerprint"])

    def test_inherited_unit(self, tmp_path, monkeypatch):
        reliable = {"unit_name": "客户甲", "order_number": "ORD-1", "sheet": "送货单", "items": [{"product_name": "漆", "model_number": "M1"}]}
        unreliable = {"unit_name": "客户甲", "sheet": "明细", "items": [{"product_name": "底漆", "model_number": "M2"}]}
        self._patched_preview(monkeypatch, {"success": True, "notes": [reliable, unreliable]})
        path = tmp_path / "客户甲.xlsx"
        path.write_text("x")
        out = shipment_compat_parser.parse_delivery_note_with_compat_profile(
            path, target_type="products", max_rows=10
        )
        assert out is not None
        assert len(out.rows) == 2
        assert any(row.values.get("unit") == "客户甲" for row in out.rows)
        codes = {w["code"] for w in out.warnings}
        assert "ETL_COMPATIBILITY_UNIT_INHERITED" in codes

    def test_max_rows_exceeded(self, tmp_path, monkeypatch):
        note = {**_NOTE, "items": [{"product_name": "漆", "model_number": "M1"}]}
        self._patched_preview(monkeypatch, {"success": True, "notes": [note]})
        path = tmp_path / "a.xlsx"
        path.write_text("x")
        with pytest.raises(EtlError):
            shipment_compat_parser.parse_delivery_note_with_compat_profile(
                path, target_type="products", max_rows=0
            )

    def test_no_rows_returns_none(self, tmp_path, monkeypatch):
        # unreliabable-only note with no matching primary unit -> skipped
        unreliable = {"unit_name": "客户甲", "sheet": "明细", "items": []}
        self._patched_preview(monkeypatch, {"success": True, "notes": [unreliable, "not-a-dict"]})
        path = tmp_path / "客户甲.xlsx"
        path.write_text("x")
        out = shipment_compat_parser.parse_delivery_note_with_compat_profile(
            path, target_type="products", max_rows=10
        )
        assert out is None

    def test_contact_person_date_and_skipped_sheet(self, tmp_path, monkeypatch):
        note = {
            **_NOTE,
            "items": [{"product_name": "漆", "model_number": "M1"}],
            "contact_person": "日期：2026年08月01日",
        }
        unreliable = {"unit_name": "客户甲", "sheet": "回款", "items": []}
        self._patched_preview(monkeypatch, {"success": True, "notes": [note, unreliable]})
        path = tmp_path / "客户甲.xlsx"
        path.write_text("x")
        out = shipment_compat_parser.parse_delivery_note_with_compat_profile(
            path, target_type="customer_products", max_rows=10
        )
        assert out is not None
        # date-like contact person is dropped entirely from the row values
        assert "contact_person" not in out.rows[0].values
        codes = {w["code"] for w in out.warnings}
        assert "ETL_COMPATIBILITY_LOW_CONFIDENCE_SHEETS_SKIPPED" in codes


# ---------------------------------------------------------------------------
# shipment_template_extractor.py
# ---------------------------------------------------------------------------
class TestShipmentTemplateExtractor:
    def _make_source(self, tmp_path, *, with_total=True):
        from openpyxl import Workbook

        path = tmp_path / "src.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "送货单"
        ws["A1"] = "某客户 送货单"
        ws["A3"] = "型号"
        ws["B3"] = "名称"
        ws["C3"] = "数量桶"
        ws["D3"] = "单价"
        ws["E3"] = "金额"
        ws["A4"] = "M1"
        ws["B4"] = "漆"
        ws["C4"] = 2
        ws["D4"] = 10
        ws["E4"] = 20
        if with_total:
            ws["A5"] = "合计"
            ws["E5"] = 20
        ws.merge_cells("A1:B1")
        wb.save(path)
        wb.close()
        return path

    def _features(self, *, source_region_id="r1", status="selected"):
        return {
            "regions": [
                {
                    "id": "r1",
                    "sheet": "送货单",
                    "header_row": 3,
                    "last_column": 5,
                    "status": status,
                    "evidence_rows": [{"row": 3}],
                }
            ]
        }

    def test_extract_happy_path(self, tmp_path):
        src = self._make_source(tmp_path)
        dest = tmp_path / "out" / "t.xlsx"
        out = shipment_template_extractor.extract_shipment_template(
            src, source_features=self._features(), destination=dest, source_region_id="r1"
        )
        assert dest.is_file()
        assert out["sheet"] == "送货单"
        assert out["header_row"] >= 1
        assert out["columns"] == 5
        assert out["source_region_id"] == "r1"

    def test_extract_no_selected_region(self, tmp_path):
        src = self._make_source(tmp_path)
        dest = tmp_path / "o.xlsx"
        with pytest.raises(EtlError) as exc:
            shipment_template_extractor.extract_shipment_template(
                src, source_features={"regions": [{"id": "r1", "status": "rejected", "header_row": 3}]}, destination=dest
            )
        assert exc.value.code == "ETL_SHIPMENT_TEMPLATE_REGION_MISSING"

    def test_extract_requested_region_not_found(self, tmp_path):
        src = self._make_source(tmp_path)
        dest = tmp_path / "o.xlsx"
        with pytest.raises(EtlError) as exc:
            shipment_template_extractor.extract_shipment_template(
                src, source_features=self._features(), destination=dest, source_region_id="missing"
            )
        assert exc.value.code == "ETL_SHIPMENT_TEMPLATE_REGION_NOT_FOUND"
        assert exc.value.status_code == 409

    def test_extract_sheet_missing(self, tmp_path):
        src = self._make_source(tmp_path)
        dest = tmp_path / "o.xlsx"
        features = {
            "regions": [
                {
                    "id": "r1",
                    "sheet": "不存在",
                    "header_row": 3,
                    "status": "selected",
                    "evidence_rows": [{"row": 3}],
                }
            ]
        }
        with pytest.raises(EtlError) as exc:
            shipment_template_extractor.extract_shipment_template(src, source_features=features, destination=dest)
        assert exc.value.code == "ETL_SHIPMENT_TEMPLATE_SHEET_MISSING"

    def test_extract_total_missing(self, tmp_path):
        src = self._make_source(tmp_path, with_total=False)
        dest = tmp_path / "o.xlsx"
        with pytest.raises(EtlError) as exc:
            shipment_template_extractor.extract_shipment_template(
                src, source_features=self._features(), destination=dest
            )
        assert exc.value.code == "ETL_SHIPMENT_TEMPLATE_TOTAL_MISSING"

    def test_template_bounds_no_requested_sorts(self, tmp_path):
        src = self._make_source(tmp_path)
        dest = tmp_path / "o.xlsx"
        features = {
            "regions": [
                {"id": "b", "sheet": "送货单", "header_row": 9, "status": "selected"},
                {"id": "a", "sheet": "送货单", "header_row": 3, "status": "selected"},
            ]
        }
        out = shipment_template_extractor.extract_shipment_template(
            src, source_features=features, destination=dest
        )
        assert out is not None


# ---------------------------------------------------------------------------
# mapping_assist.py
# ---------------------------------------------------------------------------
class _Adapter(TargetAdapter):
    type = "test"
    label = "test"
    fields = (
        TargetField("name", "名称", required=True),
        TargetField("price", "单价", type="number"),
    )


class TestMappingAssist:
    def _dataset(self, headers=None, rows=None):
        headers = headers or ["hdr1"]
        rows = rows or [ParsedRow(sheet="s", row_number=1, values={"hdr1": "漆"})]
        return ParsedDataset(headers=headers, rows=rows, source_features={})

    def test_dynamic_adapter_skips(self):
        class _Dyn(_Adapter):
            allow_dynamic_fields = True

        ds = self._dataset()
        mappings, meta = mapping_assist.enhance_mappings_with_llm(ds, _Dyn(), [])
        assert meta["used_llm"] is False
        assert meta["reason"] == "dynamic_or_empty_dataset"

    def test_empty_dataset_skips(self):
        ds = ParsedDataset(headers=[], rows=[], source_features={})
        mappings, meta = mapping_assist.enhance_mappings_with_llm(ds, _Adapter(), [])
        assert meta["used_llm"] is False

    def test_applies_llm_suggestions(self, monkeypatch):
        from app.application.etl.llm_assist import LlmAssistResult

        def _fake(**kwargs):
            return LlmAssistResult(
                used_llm=True,
                data={
                    "mappings": [
                        {"target": "name", "source": "hdr1", "confidence": 0.9, "transform": "trim", "reason": "r"},
                        # current confidence too high -> skip
                        {"target": "price", "source": "p2", "confidence": 0.9},
                        # source already used -> skip
                        {"target": "name", "source": "hdr1", "confidence": 0.9},
                        # unknown target -> skip
                        {"target": "ghost", "source": "x", "confidence": 0.9},
                        # low llm confidence -> skip
                        {"target": "price", "source": "p3", "confidence": 0.8},
                    ]
                },
            )

        monkeypatch.setattr(mapping_assist, "advise_field_mappings", _fake)
        ds = self._dataset()
        deterministic = [
            {"target": "name", "source": "old", "confidence": 0.5},
            {"target": "price", "source": "p", "confidence": 0.95},
        ]
        mappings, meta = mapping_assist.enhance_mappings_with_llm(ds, _Adapter(), deterministic)
        assert meta["used_llm"] is True
        assert meta["suggestion_count"] == 5
        assert meta["applied_count"] == 1
        assert mappings[0]["source"] == "hdr1"
        assert mappings[0]["suggested_by"] == "llm"
        assert mappings[0]["transforms"] == [{"op": "trim"}]
        # price mapping unchanged
        assert any(m["target"] == "price" and m["source"] == "p" for m in mappings)


# ---------------------------------------------------------------------------
# shipment_preview_fallback.py
# ---------------------------------------------------------------------------
class TestShipmentPreviewFallback:
    def test_normalize_customer_name(self):
        assert shipment_preview_fallback._normalize_customer_name(None) == ""
        assert shipment_preview_fallback._normalize_customer_name("星光家具（华南）有限公司") == "星光"
        assert shipment_preview_fallback._normalize_customer_name("  甲  ") == "甲"

    def test_normalize_product_name(self):
        assert shipment_preview_fallback._normalize_product_name(None) == ""
        assert shipment_preview_fallback._normalize_product_name(" 面漆-1 ") == "面漆1"

    def test_valid_owner_and_tenant(self, monkeypatch):
        monkeypatch.setattr(
            "app.application.etl.shipment_preview_fallback.current_tenant_id", lambda: "123"
        )
        assert shipment_preview_fallback._valid_owner_and_tenant(7) == (123, 7)
        assert shipment_preview_fallback._valid_owner_and_tenant(0) is None
        monkeypatch.setattr(
            "app.application.etl.shipment_preview_fallback.current_tenant_id", lambda: None
        )
        assert shipment_preview_fallback._valid_owner_and_tenant(7) is None
        assert shipment_preview_fallback._valid_owner_and_tenant("abc") is None

    def test_row_is_valid_candidate(self):
        class _Row:
            final_action = "update"
            validation_json = "[]"
            normalized_json = '{"name": "漆"}'

        assert shipment_preview_fallback._row_is_valid_candidate(_Row()) == {"name": "漆"}

        class _RowBadAction:
            final_action = "skip"
            validation_json = "[]"
            normalized_json = "x"

        assert shipment_preview_fallback._row_is_valid_candidate(_RowBadAction()) is None

        class _RowIssues:
            final_action = "new"
            validation_json = '[{"code": "x"}]'
            normalized_json = "x"

        assert shipment_preview_fallback._row_is_valid_candidate(_RowIssues()) is None

        class _RowNotDict:
            final_action = "new"
            validation_json = "[]"
            normalized_json = "not-json"

        assert shipment_preview_fallback._row_is_valid_candidate(_RowNotDict()) is None

    def test_candidate_names_and_price(self):
        assert shipment_preview_fallback._candidate_customer_name({"customer_name": "甲"}) == "甲"
        assert shipment_preview_fallback._candidate_customer_name({"purchase_unit": "乙"}) == "乙"
        assert shipment_preview_fallback._candidate_customer_name({}) == ""
        assert shipment_preview_fallback._candidate_product_name({"name": "漆"}) == "漆"
        assert shipment_preview_fallback._candidate_product_name({"product_name": "清漆"}) == "清漆"
        assert shipment_preview_fallback._candidate_product_name({}) == ""
        assert shipment_preview_fallback._candidate_price({"price": "1,234.5"}) == 1234.5
        assert shipment_preview_fallback._candidate_price({"unit_price": "￥50"}) == 50
        assert shipment_preview_fallback._candidate_price({}) is None
        assert shipment_preview_fallback._candidate_price({"price": "bad"}) is None
        assert shipment_preview_fallback._candidate_price({"price": -1}) is None
        assert shipment_preview_fallback._candidate_specification({"specification": "20KG"}) == 20
        assert shipment_preview_fallback._candidate_specification({"tin_spec": 10}) == 10
        assert shipment_preview_fallback._candidate_specification({}) is None
        assert shipment_preview_fallback._candidate_specification({"specification": "x"}) is None
        assert shipment_preview_fallback._candidate_specification({"specification": 0}) is None

    def test_candidate_source_date(self):
        class _Row:
            provenance_json = '{"source_date": "2026-08-01"}'

        assert shipment_preview_fallback._candidate_source_date(_Row()) == "2026-08-01"

        class _RowBad:
            provenance_json = '{"source_date": "bad"}'

        assert shipment_preview_fallback._candidate_source_date(_RowBad()) == ""

        class _RowNotDict:
            provenance_json = '"x"'

        assert shipment_preview_fallback._candidate_source_date(_RowNotDict()) == ""

    def test_cleanup_ephemeral(self, tmp_path):
        bad = tmp_path / "not_ours.xlsx"
        bad.write_text("x")
        shipment_preview_fallback.cleanup_ephemeral_preview_layout(bad)
        assert bad.exists()
        shipment_preview_fallback.cleanup_ephemeral_preview_layout(None)
        shipment_preview_fallback.cleanup_ephemeral_preview_layout(123)

    def _db_cm(self, db):
        @contextmanager
        def _cm():
            yield db

        return _cm

    def test_resolve_preview_product_candidate_outcome(self, monkeypatch):
        mod = shipment_preview_fallback

        class _Row:
            final_action = "new"
            validation_json = "[]"
            normalized_json = '{"name": "漆", "model_number": "M1", "price": 10, "specification": 20, "customer_name": "甲"}'
            provenance_json = '{"source_date": "2026-08-01"}'
            source_sheet = "送货单"
            source_row = 2
            id = 1

        class _Run:
            id = 42
            updated_at = None
            created_at = None

        db = MagicMock()
        db.query.return_value.filter.return_value.order_by.return_value.all.return_value = [_Row()]
        monkeypatch.setattr(mod, "get_db", self._db_cm(db))
        monkeypatch.setattr(mod, "current_tenant_id", lambda: "123")
        monkeypatch.setattr(mod, "_preview_runs", lambda *a, **k: [_Run()])

        out = mod.resolve_preview_product_candidate_outcome(
            owner_user_id=7, unit_name="甲", product_name="漆"
        )
        assert out["status"] == "resolved"
        assert out["candidate"]["model_number"] == "M1"
        assert out["candidate"]["price"] == 10

    def test_resolve_preview_not_found(self, monkeypatch):
        mod = shipment_preview_fallback

        class _Run:
            id = 1

        db = MagicMock()
        db.query.return_value.filter.return_value.order_by.return_value.all.return_value = []
        monkeypatch.setattr(mod, "get_db", self._db_cm(db))
        monkeypatch.setattr(mod, "current_tenant_id", lambda: "123")
        monkeypatch.setattr(mod, "_preview_runs", lambda *a, **k: [_Run()])
        out = mod.resolve_preview_product_candidate_outcome(
            owner_user_id=7, unit_name="甲", product_name="不存在"
        )
        assert out["status"] == "not_found"

    def test_resolve_preview_conflict(self, monkeypatch):
        mod = shipment_preview_fallback

        class _Run:
            id = 1

        rows = []
        for model, price in [("M1", 10), ("M1", 11)]:
            rows.append(
                MagicMock(
                    final_action="new",
                    validation_json="[]",
                    normalized_json=f'{{"name": "漆", "model_number": "{model}", "price": {price}, "customer_name": "甲"}}',
                    provenance_json='{"source_date": "2026-08-01"}',
                    source_sheet="送货单",
                    source_row=1,
                    id=1,
                )
            )
        db = MagicMock()
        db.query.return_value.filter.return_value.order_by.return_value.all.return_value = rows
        monkeypatch.setattr(mod, "get_db", self._db_cm(db))
        monkeypatch.setattr(mod, "current_tenant_id", lambda: "123")
        monkeypatch.setattr(mod, "_preview_runs", lambda *a, **k: [_Run()])
        out = mod.resolve_preview_product_candidate_outcome(owner_user_id=7, unit_name="甲", product_name="漆")
        assert out["status"] == "conflict"

    def test_resolve_preview_unavailable_scope(self, monkeypatch):
        mod = shipment_preview_fallback
        monkeypatch.setattr(mod, "current_tenant_id", lambda: None)
        out = mod.resolve_preview_product_candidate_outcome(owner_user_id=0, unit_name="甲", product_name="漆")
        assert out["status"] == "unavailable"

    def test_resolve_preview_recoverable_error(self, monkeypatch):
        mod = shipment_preview_fallback

        def _boom(*a, **k):
            raise RuntimeError("db down")

        monkeypatch.setattr(mod, "get_db", lambda: (_ for _ in ()).throw(RuntimeError("x")))
        monkeypatch.setattr(mod, "current_tenant_id", lambda: "123")
        monkeypatch.setattr(mod, "_preview_runs", _boom)
        out = mod.resolve_preview_product_candidate_outcome(owner_user_id=7, unit_name="甲", product_name="漆")
        assert out["status"] == "unavailable"

    def test_resolve_preview_product_candidate_wrapper(self, monkeypatch):
        mod = shipment_preview_fallback
        monkeypatch.setattr(
            mod,
            "resolve_preview_product_candidate_outcome",
            lambda **k: {"status": "resolved", "candidate": {"name": "漆"}},
        )
        assert mod.resolve_preview_product_candidate(owner_user_id=1, unit_name="甲", product_name="漆") == {"name": "漆"}
        monkeypatch.setattr(
            mod, "resolve_preview_product_candidate_outcome", lambda **k: {"status": "not_found", "candidate": None}
        )
        assert mod.resolve_preview_product_candidate(owner_user_id=1, unit_name="甲", product_name="漆") is None

    def test_safe_owned_upload_path(self, tmp_path, monkeypatch):
        mod = shipment_preview_fallback
        root = Path(tmp_path) / "etl" / "uploads" / "1" / "7"
        root.mkdir(parents=True)
        f = root / "del.xlsx"
        f.write_text("x")
        monkeypatch.setattr(mod, "get_app_data_dir", lambda: str(tmp_path))
        path = mod._safe_owned_upload_path(str(f), ".xlsx", None, tenant_id=1, owner_user_id=7)
        assert path == f.resolve()
        # outside sandbox
        outside = Path(tmp_path) / "out.xlsx"
        outside.write_text("x")
        assert mod._safe_owned_upload_path(str(outside), ".xlsx", None, tenant_id=1, owner_user_id=7) is None
        # bad suffix
        txt = root / "a.txt"
        txt.write_text("x")
        assert mod._safe_owned_upload_path(str(txt), ".xlsx", None, tenant_id=1, owner_user_id=7) is None
        # expired
        from datetime import UTC, datetime, timedelta

        expired = datetime.now(UTC) - timedelta(hours=1)
        assert mod._safe_owned_upload_path(str(f), ".xlsx", expired, tenant_id=1, owner_user_id=7) is None
        # naive expiry in future ok
        future_naive = datetime.now() + timedelta(hours=1)
        assert mod._safe_owned_upload_path(str(f), ".xlsx", future_naive, tenant_id=1, owner_user_id=7) == f.resolve()
        # path resolution raises ValueError -> None
        assert mod._safe_owned_upload_path("\x00", ".xlsx", None, tenant_id=1, owner_user_id=7) is None

    def test_selected_region(self):
        mod = shipment_preview_fallback
        assert mod._selected_region({"regions": [{"id": "r1", "status": "selected"}]}, "r1") == {
            "id": "r1",
            "status": "selected",
        }
        assert mod._selected_region({"regions": [{"id": "r1", "status": "rejected"}]}, "r1") is None
        assert mod._selected_region({"regions": [{"id": "r2", "status": "selected"}]}, "r1") is None
        assert mod._selected_region({"regions": ["x"]}, "r1") is None

    def test_valid_owner_tenant_nonint(self, monkeypatch):
        mod = shipment_preview_fallback
        monkeypatch.setattr(mod, "current_tenant_id", lambda: "abc")
        assert mod._valid_owner_and_tenant(7) is None

    def test_preview_runs(self):
        mod = shipment_preview_fallback
        db = MagicMock()
        db.query.return_value.filter.return_value.order_by.return_value.limit.return_value.all.return_value = ["run"]
        assert mod._preview_runs(db, tenant_id=1, owner_user_id=2, target_type="t") == ["run"]

    def test_public_layout_candidate(self):
        mod = shipment_preview_fallback
        record = {
            "run_id": "r1",
            "template_id": "t1",
            "name": "版式",
            "customer_name": "甲",
            "source_region_id": "reg1",
            "sheet": "送货单",
            "header_row": 3,
            "file_name": "f.xlsx",
        }
        out = mod._public_layout_candidate(record)
        assert out["run_id"] == "r1"
        assert out["warning"] == mod.LAYOUT_PREVIEW_WARNING
        assert out["provenance"]["source_region_id"] == "reg1"

    def _layout_run(self):
        import json as _json

        class _Run:
            source_features_json = _json.dumps(
                {
                    "shipment_template_candidates": [
                        {
                            "status": "detected",
                            "customer_name": "甲",
                            "source_region_id": "r1",
                            "name": "版式",
                            "sheet": "送货单",
                            "header_row": 3,
                        }
                    ],
                    "regions": [{"id": "r1", "status": "selected", "customer_name": "甲", "sheet": "送货单", "header_row": 3}],
                }
            )
            id = 1
            upload_id = 5
            file_sha256 = "abc"

        return _Run()

    def test_layout_candidate_for_run(self):
        mod = shipment_preview_fallback
        run = self._layout_run()
        upload = MagicMock()
        upload.file_name = "f.xlsx"
        upload.storage_path = "/x"
        upload.suffix = ".xlsx"
        upload.expires_at = None
        db = MagicMock()
        db.query.return_value.filter.return_value.first.return_value = upload
        out = mod._layout_candidate_for_run(db, run=run, tenant_id=1, owner_user_id=2, unit_name="甲")
        assert out is not None
        assert out["run_id"] == "1"
        assert out["sheet"] == "送货单"
        assert out["header_row"] == 3

    def test_layout_candidate_for_run_branches(self, monkeypatch):
        import json as _json

        mod = shipment_preview_fallback
        # source_features not dict
        class _RunBad:
            source_features_json = "not-json"
            upload_id = 5
            id = 1
            file_sha256 = "abc"

        assert mod._layout_candidate_for_run(MagicMock(), run=_RunBad(), tenant_id=1, owner_user_id=2, unit_name="甲") is None

        # candidate not detected -> None
        class _RunRejected(_RunBad):
            source_features_json = _json.dumps(
                {
                    "shipment_template_candidates": [
                        {"status": "rejected", "customer_name": "甲", "source_region_id": "r1"}
                    ]
                }
            )

        assert (
            mod._layout_candidate_for_run(MagicMock(), run=_RunRejected(), tenant_id=1, owner_user_id=2, unit_name="甲")
            is None
        )
        # region not selected -> None
        class _RunNoRegion(_RunBad):
            source_features_json = _json.dumps(
                {
                    "shipment_template_candidates": [
                        {"status": "detected", "customer_name": "甲", "source_region_id": "r1"}
                    ]
                }
            )

        assert (
            mod._layout_candidate_for_run(MagicMock(), run=_RunNoRegion(), tenant_id=1, owner_user_id=2, unit_name="甲")
            is None
        )
        # upload missing -> None
        class _RunOk(_RunBad):
            source_features_json = _json.dumps(
                {
                    "shipment_template_candidates": [
                        {"status": "detected", "customer_name": "甲", "source_region_id": "r1"}
                    ],
                    "regions": [{"id": "r1", "status": "selected", "customer_name": "甲", "sheet": "s", "header_row": 2}],
                }
            )

        db = MagicMock()
        db.query.return_value.filter.return_value.first.return_value = None
        assert mod._layout_candidate_for_run(db, run=_RunOk(), tenant_id=1, owner_user_id=2, unit_name="甲") is None

        # single candidate (shipment_template_candidate) + matching region -> resolved
        class _RunSingle(_RunBad):
            source_features_json = _json.dumps(
                {
                    "shipment_template_candidate": {
                        "status": "detected",
                        "customer_name": "甲",
                        "source_region_id": "r1",
                    },
                    "regions": [{"id": "r1", "status": "selected", "customer_name": "甲", "sheet": "s", "header_row": 2}],
                }
            )

        db_single = MagicMock()
        upload = MagicMock()
        upload.file_name = "f"
        upload.storage_path = "/x"
        upload.suffix = ".xlsx"
        upload.expires_at = None
        db_single.query.return_value.filter.return_value.first.return_value = upload
        assert (
            mod._layout_candidate_for_run(db_single, run=_RunSingle(), tenant_id=1, owner_user_id=2, unit_name="甲")
            is not None
        )

        # no embedded candidates -> fall back to shipment_template_candidates
        class _RunFallback(_RunBad):
            source_features_json = _json.dumps(
                {"regions": [{"id": "r1", "status": "selected", "customer_name": "甲", "sheet": "s", "header_row": 2}]}
            )

        monkeypatch.setattr(
            "app.application.etl.service_shipment_templates.shipment_template_candidates",
            lambda *a: [{"status": "detected", "customer_name": "甲", "source_region_id": "r1"}],
        )
        db_fb = MagicMock()
        upload_fb = MagicMock()
        upload_fb.file_name = "f"
        upload_fb.storage_path = "/x"
        upload_fb.suffix = ".xlsx"
        upload_fb.expires_at = None
        db_fb.query.return_value.filter.return_value.first.return_value = upload_fb
        assert (
            mod._layout_candidate_for_run(db_fb, run=_RunFallback(), tenant_id=1, owner_user_id=2, unit_name="甲")
            is not None
        )

        # source_features is a non-dict list -> None
        class _RunList(_RunBad):
            source_features_json = "[]"

        assert (
            mod._layout_candidate_for_run(MagicMock(), run=_RunList(), tenant_id=1, owner_user_id=2, unit_name="甲")
            is None
        )

    def test_find_preview_layout_record(self, monkeypatch):
        mod = shipment_preview_fallback
        # scope None
        monkeypatch.setattr(mod, "current_tenant_id", lambda: None)
        assert mod._find_preview_layout_record(owner_user_id=0, unit_name="甲") is None
        # empty unit
        monkeypatch.setattr(mod, "current_tenant_id", lambda: "1")
        assert mod._find_preview_layout_record(owner_user_id=1, unit_name="") is None

        # run_id path, run found
        run = MagicMock()
        db = MagicMock()
        db.query.return_value.filter.return_value.first.return_value = run
        monkeypatch.setattr(mod, "get_db", partial_context(db))
        monkeypatch.setattr(mod, "_layout_candidate_for_run", lambda *a, **k: {"rec": True})
        assert mod._find_preview_layout_record(owner_user_id=1, unit_name="甲", run_id="r") == {"rec": True}

        # run_id path, run missing
        db2 = MagicMock()
        db2.query.return_value.filter.return_value.first.return_value = None
        monkeypatch.setattr(mod, "get_db", partial_context(db2))
        assert mod._find_preview_layout_record(owner_user_id=1, unit_name="甲", run_id="r") is None

        # no run_id loop
        monkeypatch.setattr(mod, "_preview_runs", lambda *a, **k: ["r1", "r2"])
        calls = {"n": 0}

        def _layout(*a, **k):
            calls["n"] += 1
            return {"rec": True} if calls["n"] == 2 else None

        monkeypatch.setattr(mod, "_layout_candidate_for_run", _layout)
        assert mod._find_preview_layout_record(owner_user_id=1, unit_name="甲") == {"rec": True}

        # no run_id loop nothing found
        monkeypatch.setattr(mod, "_layout_candidate_for_run", lambda *a, **k: None)
        assert mod._find_preview_layout_record(owner_user_id=1, unit_name="甲") is None

        # recoverable error
        def _boom(*a, **k):
            raise RuntimeError("db down")

        monkeypatch.setattr(mod, "get_db", lambda: (_ for _ in ()).throw(RuntimeError("x")))
        monkeypatch.setattr(mod, "_preview_runs", _boom)
        assert mod._find_preview_layout_record(owner_user_id=1, unit_name="甲") is None

    def test_find_latest_preview_layout_candidate(self, monkeypatch):
        mod = shipment_preview_fallback
        monkeypatch.setattr(
            mod,
            "_find_preview_layout_record",
            lambda **k: {
                "run_id": "r1",
                "template_id": "t1",
                "name": "n",
                "customer_name": "甲",
                "source_region_id": "reg",
                "sheet": "s",
                "header_row": 3,
                "file_name": "f",
            },
        )
        out = mod.find_latest_preview_layout_candidate(owner_user_id=1, unit_name="甲")
        assert out["run_id"] == "r1"
        assert out["warning"] == mod.LAYOUT_PREVIEW_WARNING
        monkeypatch.setattr(mod, "_find_preview_layout_record", lambda **k: None)
        assert mod.find_latest_preview_layout_candidate(owner_user_id=1, unit_name="甲") is None

    def test_materialize_preview_layout_candidate(self, tmp_path, monkeypatch):
        mod = shipment_preview_fallback
        record = {
            "run_id": "1",
            "template_id": "t1",
            "name": "n",
            "customer_name": "甲",
            "source_region_id": "reg",
            "sheet": "s",
            "header_row": 3,
            "file_name": "f",
            "upload_storage_path": "/x.xlsx",
            "upload_suffix": ".xlsx",
            "upload_expires_at": None,
            "source_features": {},
        }
        # record None -> None
        monkeypatch.setattr(mod, "_find_preview_layout_record", lambda **k: None)
        assert mod.materialize_preview_layout_candidate(owner_user_id=1, unit_name="甲") is None

        # upload_path None -> None
        monkeypatch.setattr(mod, "_find_preview_layout_record", lambda **k: record)
        monkeypatch.setattr(mod, "_safe_owned_upload_path", lambda *a, **k: None)
        assert mod.materialize_preview_layout_candidate(owner_user_id=1, unit_name="甲") is None

        # success
        src = tmp_path / "src.xlsx"
        src.write_text("x")

        def _extract(source_path, *, source_features, destination, source_region_id):
            Path(destination).write_text("x")

        monkeypatch.setattr(mod, "_safe_owned_upload_path", lambda *a, **k: src)
        monkeypatch.setattr(mod, "extract_shipment_template", _extract)
        out = mod.materialize_preview_layout_candidate(owner_user_id=1, unit_name="甲")
        assert out is not None
        assert out["path"]
        assert out["source"] == "etl_preview_candidate"

        # recoverable error -> None
        def _boom(*a, **k):
            raise OSError("fail")

        monkeypatch.setattr(mod, "extract_shipment_template", _boom)
        assert mod.materialize_preview_layout_candidate(owner_user_id=1, unit_name="甲") is None

        # extract produces no file -> OSError -> None
        monkeypatch.setattr(mod, "extract_shipment_template", lambda *a, **k: None)
        assert mod.materialize_preview_layout_candidate(owner_user_id=1, unit_name="甲") is None

    def test_resolve_preview_mismatch_row(self, monkeypatch):
        # row whose customer does not match -> continue (line 235)
        mod = shipment_preview_fallback

        class _Run:
            id = 1

        class _Row:
            final_action = "new"
            validation_json = "[]"
            normalized_json = '{"name": "漆", "model_number": "M1", "price": 10, "customer_name": "其他"}'
            provenance_json = '{"source_date": "2026-08-01"}'
            source_sheet = "s"
            source_row = 1
            id = 1

        db = MagicMock()
        db.query.return_value.filter.return_value.order_by.return_value.all.return_value = [_Row()]
        monkeypatch.setattr(mod, "get_db", partial_context(db))
        monkeypatch.setattr(mod, "current_tenant_id", lambda: "123")
        monkeypatch.setattr(mod, "_preview_runs", lambda *a, **k: [_Run()])
        out = mod.resolve_preview_product_candidate_outcome(owner_user_id=7, unit_name="甲", product_name="漆")
        assert out["status"] == "not_found"

    def test_resolve_preview_legacy_path(self, monkeypatch):
        # match without provenance source_date -> legacy run-rank ordering
        mod = shipment_preview_fallback

        class _Run:
            id = 1

        class _Row:
            final_action = "new"
            validation_json = "[]"
            normalized_json = '{"name": "漆", "model_number": "M1", "price": 10, "customer_name": "甲"}'
            provenance_json = "{}"
            source_sheet = "s"
            source_row = 1
            id = 1

        db = MagicMock()
        db.query.return_value.filter.return_value.order_by.return_value.all.return_value = [_Row()]
        monkeypatch.setattr(mod, "get_db", partial_context(db))
        monkeypatch.setattr(mod, "current_tenant_id", lambda: "123")
        monkeypatch.setattr(mod, "_preview_runs", lambda *a, **k: [_Run()])
        out = mod.resolve_preview_product_candidate_outcome(owner_user_id=7, unit_name="甲", product_name="漆")
        assert out["status"] == "resolved"


def partial_context(db):
    @contextmanager
    def _cm():
        yield db

    return _cm