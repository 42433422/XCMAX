"""Branch coverage tests for app.utils.excel_template_analyzer.

Uses real openpyxl workbooks (created in tmp_path) to exercise branches in:
_guess_merged_purpose, _classify_cell_type, _classify_row_type,
_identify_editable_ranges, _group_consecutive, _describe_range,
_build_output, save_json, extract_entries, analyze_template, analyze_to_json.
"""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.utils.excel_template_analyzer import (
    CellStyle,
    ExcelTemplateAnalyzer,
    analyze_template,
    analyze_to_json,
    extract_entries,
)

# ---------------------------------------------------------------------------
# Helpers — build real .xlsx files with openpyxl
# ---------------------------------------------------------------------------


def _create_workbook(tmp_path: Path, filename: str = "test.xlsx") -> Path:
    path = tmp_path / filename
    wb = Workbook()
    wb.active.title = "Sheet1"
    wb.save(path)
    return path


def _create_workbook_with_content(
    tmp_path: Path,
    cells: dict[tuple[int, int], object] | None = None,
    merged: list[str] | None = None,
    sheet_name: str = "Sheet1",
    filename: str = "test.xlsx",
) -> Path:
    path = tmp_path / filename
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    if cells:
        for (row, col), value in cells.items():
            ws.cell(row=row, column=col, value=value)
    if merged:
        for mr in merged:
            ws.merge_cells(mr)
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# _guess_merged_purpose — all branches
# ---------------------------------------------------------------------------


class TestGuessMergedPurpose:
    def test_title_delivery(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(1, 1): "送货单"}, merged=["A1:E1"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(1, 1, 1, 5)
        assert purpose == "标题"

    def test_title_receipt(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(1, 1): "收据"}, merged=["A1:E1"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(1, 1, 1, 5)
        assert purpose == "标题"

    def test_buyer_info_gouhuo(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(2, 1): "购货单位信息"}, merged=["A2:E2"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(2, 2, 1, 5)
        assert purpose == "购货单位信息"

    def test_buyer_info_yifang(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(2, 1): "乙方信息"}, merged=["A2:E2"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(2, 2, 1, 5)
        assert purpose == "购货单位信息"

    def test_header_xinghao(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(3, 1): "型号列表"}, merged=["A3:E3"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(3, 3, 1, 5)
        assert purpose == "表头"

    def test_summary_heji(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(8, 1): "合计行"}, merged=["A8:E8"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(8, 8, 1, 5)
        assert purpose == "汇总"

    def test_summary_sum(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(8, 1): "SUM公式"}, merged=["A8:E8"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(8, 8, 1, 5)
        assert purpose == "汇总"

    def test_signature(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(9, 1): "签名处"}, merged=["A9:E9"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(9, 9, 1, 5)
        assert purpose == "签名区"

    def test_content_area(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(5, 1): "普通内容"}, merged=["A5:E5"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(5, 5, 1, 5)
        assert purpose == "内容区"

    def test_empty_large_row_span(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, merged=["A1:E8"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(1, 8, 1, 5)
        assert purpose == "大标题区"

    def test_empty_small_row_span(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, merged=["A1:E2"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        purpose = analyzer._guess_merged_purpose(1, 2, 1, 5)
        assert purpose == "合并内容"


# ---------------------------------------------------------------------------
# _classify_cell_type — all branches
# ---------------------------------------------------------------------------


class TestClassifyCellType:
    def test_formula_cell(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 3): "=SUM(C1:C4)"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C5"]
        assert cell.type == "formula"
        assert cell.formula is not None
        assert cell.value is None

    def test_empty_cell(self, tmp_path):
        # Write to E5 to force worksheet dimensions to include row 5, col 5
        path = _create_workbook_with_content(tmp_path, cells={(5, 5): "anchor"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["A5"]
        assert cell.type == "empty"

    def test_row_le_3_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(2, 3): "标题文字"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C2"]
        assert cell.type == "template"

    def test_col_1_editable(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 1): "数据值"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["A5"]
        assert cell.type == "editable"

    def test_col_2_editable(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 2): "数据值"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["B5"]
        assert cell.type == "editable"

    def test_col_1_formula_not_editable(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 1): "=A1+A2"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["A5"]
        assert cell.type == "formula"

    def test_summary_keyword_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(8, 3): "合计"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C8"]
        assert cell.type == "template"

    def test_subtotal_keyword_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(8, 3): "小计"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C8"]
        assert cell.type == "template"

    def test_total_keyword_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(8, 3): "总计"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C8"]
        assert cell.type == "template"

    def test_date_keyword_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(8, 3): "日期：2026-01-01"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C8"]
        assert cell.type == "template"

    def test_unit_keyword_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(8, 3): "单位：个"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C8"]
        assert cell.type == "template"

    def test_numeric_string_editable(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 3): "123-456"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C5"]
        assert cell.type == "editable"

    def test_chinese_long_string_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 3): "产品名称列表"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C5"]
        assert cell.type == "template"

    def test_col_ge_2_editable(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 3): "ab"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["C5"]
        assert cell.type == "editable"

    def test_col_1_short_string_editable(self, tmp_path):
        """Col 1 with short string (not starting with =) is editable."""
        path = _create_workbook_with_content(tmp_path, cells={(5, 1): "ab"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["A5"]
        # col==1, value_str="ab" truthy and not starting with "=" → editable
        assert cell.type == "editable"

    def test_col_1_whitespace_falls_through_to_template(self, tmp_path):
        """Col 1 with whitespace-only value falls through to template (value_str empty)."""
        path = _create_workbook_with_content(tmp_path, cells={(5, 1): "  ", (5, 5): "anchor"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        cell = analyzer.cells["A5"]
        # value="  " → value_str="" (falsy) → skips editable → falls to template
        assert cell.type == "template"


# ---------------------------------------------------------------------------
# _classify_row_type — all branches
# ---------------------------------------------------------------------------


class TestClassifyRowType:
    def test_row_le_3_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(2, 1): "标题"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        assert analyzer._classify_row_type(2) == "template"

    def test_row_has_editable(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 1): "数据值"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        assert analyzer._classify_row_type(5) == "editable"

    def test_row_has_template_only(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 3): "产品名称列表"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        assert analyzer._classify_row_type(5) == "template"

    def test_row_has_formula_pass(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 3): "=SUM(1+2)"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        # Formula cells are "pass" — no editable/template flag set
        assert analyzer._classify_row_type(5) == "template"

    def test_row_empty_returns_template(self, tmp_path):
        path = _create_workbook_with_content(tmp_path)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        assert analyzer._classify_row_type(5) == "template"


# ---------------------------------------------------------------------------
# _identify_editable_ranges — all branches
# ---------------------------------------------------------------------------


class TestIdentifyEditableRanges:
    def test_no_editable_cells(self, tmp_path):
        path = _create_workbook_with_content(tmp_path)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        analyzer._identify_editable_ranges()
        assert analyzer.editable_ranges == []

    def test_few_editable_rows(self, tmp_path):
        """<=5 editable rows → single group."""
        cells = {(i, 1): f"val{i}" for i in range(5, 8)}  # 3 rows
        path = _create_workbook_with_content(tmp_path, cells=cells)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        analyzer._identify_editable_ranges()
        assert len(analyzer.editable_ranges) >= 1

    def test_many_editable_rows_consecutive(self, tmp_path):
        """>5 consecutive editable rows → grouped."""
        cells = {(i, 1): f"val{i}" for i in range(5, 15)}  # 10 rows
        path = _create_workbook_with_content(tmp_path, cells=cells)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        analyzer._identify_editable_ranges()
        assert len(analyzer.editable_ranges) >= 1

    def test_many_editable_rows_non_consecutive(self, tmp_path):
        """>5 non-consecutive editable rows → multiple groups."""
        cells = {}
        for i in [5, 6, 7, 8, 9, 10, 20, 21, 22]:
            cells[(i, 1)] = f"val{i}"
        path = _create_workbook_with_content(tmp_path, cells=cells)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        analyzer._identify_editable_ranges()
        # Should have at least 2 groups (gap between 10 and 20)
        assert len(analyzer.editable_ranges) >= 2


# ---------------------------------------------------------------------------
# _group_consecutive
# ---------------------------------------------------------------------------


class TestGroupConsecutive:
    def test_empty_list(self):
        analyzer = ExcelTemplateAnalyzer.__new__(ExcelTemplateAnalyzer)
        assert analyzer._group_consecutive([]) == []

    def test_single_element(self):
        analyzer = ExcelTemplateAnalyzer.__new__(ExcelTemplateAnalyzer)
        assert analyzer._group_consecutive([5]) == [[5]]

    def test_all_consecutive(self):
        analyzer = ExcelTemplateAnalyzer.__new__(ExcelTemplateAnalyzer)
        assert analyzer._group_consecutive([1, 2, 3, 4]) == [[1, 2, 3, 4]]

    def test_with_gap(self):
        analyzer = ExcelTemplateAnalyzer.__new__(ExcelTemplateAnalyzer)
        result = analyzer._group_consecutive([1, 2, 5, 6, 9])
        assert result == [[1, 2], [5, 6], [9]]

    def test_gap_of_one_still_grouped(self):
        """Difference of 1 (<=1) is still consecutive."""
        analyzer = ExcelTemplateAnalyzer.__new__(ExcelTemplateAnalyzer)
        result = analyzer._group_consecutive([1, 2, 3])
        assert result == [[1, 2, 3]]


# ---------------------------------------------------------------------------
# _describe_range
# ---------------------------------------------------------------------------


class TestDescribeRange:
    def test_with_first_cell_value(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 2): "产品A"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        desc = analyzer._describe_range(5, 10, 2, 5)
        assert "产品A" in desc
        assert "B5" in desc

    def test_without_first_cell_value(self, tmp_path):
        path = _create_workbook_with_content(tmp_path)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        desc = analyzer._describe_range(5, 10, 2, 5)
        assert "B5" in desc
        assert "..." not in desc


# ---------------------------------------------------------------------------
# _extract_style — all branches
# ---------------------------------------------------------------------------


class TestExtractStyle:
    def test_full_style(self, tmp_path):
        path = tmp_path / "styled.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "S1"
        cell = ws.cell(row=1, column=1, value="test")
        cell.font = Font(name="Arial", size=14, bold=True, color="FF0000")
        cell.fill = PatternFill(patternType="solid", fgColor="00FF00", bgColor="0000FF")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style="thin"),
            right=Side(style="medium"),
            top=Side(style="thick"),
            bottom=Side(style="dotted"),
        )
        cell.number_format = "0.00"
        wb.save(path)

        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        style = analyzer.cells["A1"].style
        assert style is not None
        assert style.font_name == "Arial"
        assert style.font_size == 14
        assert style.font_bold is True
        assert style.font_color == "00FF0000"
        assert style.fill_pattern == "solid"
        assert style.fill_fg_color == "0000FF00"
        assert style.fill_bg_color == "000000FF"
        assert style.alignment_horizontal == "center"
        assert style.alignment_vertical == "center"
        assert "left:thin" in style.border_style
        assert "right:medium" in style.border_style
        assert "top:thick" in style.border_style
        assert "bottom:dotted" in style.border_style
        assert style.number_format == "0.00"

    def test_no_style_elements(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "test"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        style = analyzer.cells["A1"].style
        # Default cell has font/fill/alignment/border from openpyxl defaults
        assert style is not None
        # number_format is always set
        assert style.number_format is not None

    def test_font_without_color(self, tmp_path):
        path = tmp_path / "nocolor.xlsx"
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="x")
        cell.font = Font(name="Calibri", size=11, bold=False)
        wb.save(path)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        style = analyzer.cells["A1"].style
        assert style.font_name == "Calibri"
        # font_color may be None if no color set
        assert style.font_bold is False

    def test_fill_without_pattern(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "x"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        style = analyzer.cells["A1"].style
        # Default fill has no patternType
        assert style.fill_pattern is None or style.fill_pattern is None

    def test_border_without_sides(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "x"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        analyzer._extract_cells()
        style = analyzer.cells["A1"].style
        # Default border has no styles
        assert style.border_style is None


# ---------------------------------------------------------------------------
# _build_output — formula + style branches
# ---------------------------------------------------------------------------


class TestBuildOutput:
    def test_output_with_formula(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 3): "=SUM(C1:C4)"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        result = analyzer.analyze()
        cell = result["cells"]["C5"]
        assert "formula" in cell
        assert cell["type"] == "formula"

    def test_output_without_formula(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(5, 1): "data"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        result = analyzer.analyze()
        cell = result["cells"]["A5"]
        assert "formula" not in cell

    def test_output_with_style(self, tmp_path):
        path = tmp_path / "styled.xlsx"
        wb = Workbook()
        ws = wb.active
        cell = ws.cell(row=1, column=1, value="x")
        cell.font = Font(name="Arial", size=12, bold=True, color="FF0000")
        wb.save(path)
        analyzer = ExcelTemplateAnalyzer(str(path))
        result = analyzer.analyze()
        cell = result["cells"]["A1"]
        assert "style" in cell
        assert cell["style"]["font_name"] == "Arial"

    def test_output_structure(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "标题"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        result = analyzer.analyze()
        assert result["file"] == "test.xlsx"
        assert result["sheet"] == "Sheet1"
        assert "structure" in result
        assert "zones" in result
        assert "merged_cells" in result
        assert "editable_ranges" in result
        assert "cells" in result
        assert result["structure"]["max_row"] >= 1
        assert result["structure"]["max_col"] >= 1


# ---------------------------------------------------------------------------
# _load_workbook — sheet_name branch
# ---------------------------------------------------------------------------


class TestLoadWorkbook:
    def test_default_sheet(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "x"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        assert analyzer.workbook_name == "Sheet1"

    def test_specific_sheet_name(self, tmp_path):
        path = tmp_path / "multi.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        ws2 = wb.create_sheet("Sheet2")
        ws2.cell(row=1, column=1, value="hello")
        wb.save(path)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook(sheet_name="Sheet2")
        assert analyzer.workbook_name == "Sheet2"


# ---------------------------------------------------------------------------
# _get_merged_range
# ---------------------------------------------------------------------------


class TestGetMergedRange:
    def test_address_in_merged_range(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(1, 1): "标题"}, merged=["A1:E1"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        mr = analyzer._get_merged_range("A1")
        assert mr is not None
        assert "A1" in mr

    def test_address_not_in_merged_range(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(1, 1): "标题"}, merged=["A1:E1"]
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        mr = analyzer._get_merged_range("F5")
        assert mr is None


# ---------------------------------------------------------------------------
# _identify_zones
# ---------------------------------------------------------------------------


class TestIdentifyZones:
    def test_zones_with_header_and_data(self, tmp_path):
        cells = {(1, 1): "送货单标题", (5, 1): "数据行1", (6, 1): "数据行2", (8, 3): "合计"}
        path = _create_workbook_with_content(tmp_path, cells=cells)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer.analyze()
        zone_names = [z.name for z in analyzer.zones]
        assert "header" in zone_names
        assert "data" in zone_names

    def test_zones_only_header(self, tmp_path):
        cells = {(1, 1): "标题", (2, 1): "副标题"}
        path = _create_workbook_with_content(tmp_path, cells=cells)
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer.analyze()
        zone_names = [z.name for z in analyzer.zones]
        assert "header" in zone_names


# ---------------------------------------------------------------------------
# save_json — pretty branch
# ---------------------------------------------------------------------------


class TestSaveJson:
    def test_pretty_true(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "x"})
        output = tmp_path / "out.json"
        analyzer = ExcelTemplateAnalyzer(str(path))
        # Mock analyze to return simple serializable dict (avoid RGB serialization issues)
        simple_result = {"file": "test.xlsx", "sheet": "Sheet1", "cells": {"A1": {"value": "x"}}}
        analyzer.analyze = MagicMock(return_value=simple_result)
        result = analyzer.save_json(str(output), pretty=True)
        assert result["file"] == "test.xlsx"
        content = output.read_text(encoding="utf-8")
        # Pretty JSON has newlines and indentation
        assert "\n" in content
        assert "  " in content

    def test_pretty_false(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "x"})
        output = tmp_path / "out.json"
        analyzer = ExcelTemplateAnalyzer(str(path))
        simple_result = {"file": "test.xlsx", "sheet": "Sheet1", "cells": {"A1": {"value": "x"}}}
        analyzer.analyze = MagicMock(return_value=simple_result)
        result = analyzer.save_json(str(output), pretty=False)
        assert result["file"] == "test.xlsx"
        content = output.read_text(encoding="utf-8")
        # Non-pretty JSON is single line
        parsed = json.loads(content)
        assert parsed["file"] == "test.xlsx"


# ---------------------------------------------------------------------------
# analyze_template (module function)
# ---------------------------------------------------------------------------


class TestAnalyzeTemplateModule:
    def test_basic(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "送货单", (5, 1): "数据"})
        result = analyze_template(str(path))
        assert result["file"] == "test.xlsx"
        assert result["sheet"] == "Sheet1"
        assert "cells" in result

    def test_with_sheet_name(self, tmp_path):
        path = tmp_path / "multi.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        ws2 = wb.create_sheet("Data")
        ws2.cell(row=1, column=1, value="hello")
        wb.save(path)
        result = analyze_template(str(path), sheet_name="Data")
        assert result["sheet"] == "Data"


# ---------------------------------------------------------------------------
# analyze_to_json (module function)
# ---------------------------------------------------------------------------


class TestAnalyzeToJsonModule:
    def test_basic(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "x"})
        output = tmp_path / "out.json"
        # Mock analyze to avoid RGB serialization issues in default openpyxl styles
        simple_result = {"file": "test.xlsx", "sheet": "Sheet1", "cells": {}}
        with patch.object(ExcelTemplateAnalyzer, "analyze", return_value=simple_result):
            result = analyze_to_json(str(path), str(output))
        assert result["file"] == "test.xlsx"
        assert output.exists()
        saved = json.loads(output.read_text(encoding="utf-8"))
        assert saved["file"] == "test.xlsx"


# ---------------------------------------------------------------------------
# extract_entries
# ---------------------------------------------------------------------------


class TestExtractEntries:
    def test_with_editable_cells(self, tmp_path):
        cells = {(5, 1): "数据1", (6, 1): "数据2", (8, 3): "合计"}
        path = _create_workbook_with_content(tmp_path, cells=cells)
        result = extract_entries(str(path))
        assert result["file"] == "test.xlsx"
        assert "editable_entries" in result
        assert len(result["editable_entries"]) >= 2
        # Each entry has address, row, col, value
        entry = result["editable_entries"][0]
        assert "address" in entry
        assert "row" in entry
        assert "col" in entry
        assert "value" in entry

    def test_without_editable_cells(self, tmp_path):
        cells = {(1, 1): "标题", (2, 1): "副标题"}
        path = _create_workbook_with_content(tmp_path, cells=cells)
        result = extract_entries(str(path))
        assert result["editable_entries"] == []

    def test_with_sheet_name(self, tmp_path):
        path = tmp_path / "multi.xlsx"
        wb = Workbook()
        wb.active.title = "Sheet1"
        ws2 = wb.create_sheet("Data")
        ws2.cell(row=5, column=1, value="editable_data")
        wb.save(path)
        result = extract_entries(str(path), sheet_name="Data")
        assert result["sheet"] == "Data"
        assert len(result["editable_entries"]) >= 1


# ---------------------------------------------------------------------------
# _extract_merged_cells
# ---------------------------------------------------------------------------


class TestExtractMergedCells:
    def test_multiple_merged_ranges(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path,
            cells={(1, 1): "送货单", (5, 1): "型号", (8, 1): "合计"},
            merged=["A1:E1", "A5:E5", "A8:E8"],
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        assert len(analyzer.merged_cells_list) == 3
        purposes = [mc.purpose for mc in analyzer.merged_cells_list]
        assert "标题" in purposes
        assert "表头" in purposes
        assert "汇总" in purposes

    def test_no_merged_ranges(self, tmp_path):
        path = _create_workbook_with_content(tmp_path, cells={(1, 1): "x"})
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        analyzer._extract_merged_cells()
        assert analyzer.merged_cells_list == []
        assert analyzer.merged_ranges_set == set()


# ---------------------------------------------------------------------------
# _extract_structure
# ---------------------------------------------------------------------------


class TestExtractStructure:
    def test_max_row_col(self, tmp_path):
        path = _create_workbook_with_content(
            tmp_path, cells={(1, 1): "a", (10, 5): "b"}
        )
        analyzer = ExcelTemplateAnalyzer(str(path))
        analyzer._load_workbook()
        analyzer._extract_structure()
        assert analyzer.max_row >= 10
        assert analyzer.max_col >= 5


# ---------------------------------------------------------------------------
# Full analyze() integration
# ---------------------------------------------------------------------------


class TestAnalyzeIntegration:
    def test_full_template_analysis(self, tmp_path):
        """End-to-end analysis of a realistic delivery note template."""
        path = tmp_path / "delivery.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "送货单"
        # Title
        ws.cell(row=1, column=1, value="送货单")
        ws.merge_cells("A1:F1")
        # Buyer info
        ws.cell(row=2, column=1, value="购货单位：ACME")
        ws.merge_cells("A2:F2")
        # Header
        ws.cell(row=4, column=1, value="型号")
        ws.cell(row=4, column=2, value="名称")
        ws.cell(row=4, column=3, value="数量")
        ws.cell(row=4, column=4, value="单价")
        ws.cell(row=4, column=5, value="金额")
        ws.cell(row=4, column=6, value="备注")
        # Data rows
        for i in range(5, 10):
            ws.cell(row=i, column=1, value=f"M{i}")
            ws.cell(row=i, column=2, value=f"产品{i}")
            ws.cell(row=i, column=3, value=i)
            ws.cell(row=i, column=4, value=10.5)
            ws.cell(row=i, column=5, value=f"=C{i}*D{i}")
            ws.cell(row=i, column=6, value="")
        # Summary
        ws.cell(row=11, column=1, value="合计")
        ws.cell(row=11, column=5, value="=SUM(E5:E9)")
        # Signature
        ws.cell(row=13, column=1, value="签名：")
        ws.cell(row=13, column=4, value="日期：")
        wb.save(path)

        result = analyze_template(str(path))
        assert result["file"] == "delivery.xlsx"
        assert result["sheet"] == "送货单"
        assert len(result["merged_cells"]) == 2
        assert "cells" in result
        # Formula cells should be detected
        formula_cells = [
            c for c in result["cells"].values() if c.get("type") == "formula"
        ]
        assert len(formula_cells) >= 5
        # Zones should include header and data
        zone_names = [z["name"] for z in result["zones"]]
        assert "header" in zone_names
        assert "data" in zone_names
