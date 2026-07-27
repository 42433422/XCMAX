from __future__ import annotations

import json
import sqlite3
from datetime import UTC, datetime, timedelta
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pytest
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

from app.application.dataset_rag_app_service import DatasetRagApplicationService
from app.application.etl.adviser import EtlRowAdviser
from app.application.etl.errors import EtlConflict, EtlError, EtlNotFound
from app.application.etl.parsers import parse_file
from app.application.etl.service import EtlService, mark_interrupted_runs_on_startup
from app.application.etl.targets import (
    AttendanceAdapter,
    CustomerAdapter,
    CustomerProductsAdapter,
    ExportCsvAdapter,
    ExportXlsxAdapter,
    KnowledgeAdapter,
    ProductAdapter,
    PurchaseOrderAdapter,
    ShipmentAdapter,
    TargetAdapter,
    WebhookAdapter,
    get_adapter,
    target_capabilities,
)
from app.application.etl.transforms import (
    apply_mapping,
    apply_transform,
    neutralize_spreadsheet_formula,
)
from app.application.excel_etl_kb import ExcelEtlKnowledgeBase, TemplateMemory
from app.application.shipment_excel_etl_app_service import write_delivery_note_workbook
from app.db.base import Base
from app.db.models.etl import (
    EtlRun,
    EtlRunRow,
    EtlTargetConfig,
    EtlTemplate,
    EtlTemplateVersion,
    EtlUpload,
)
from app.db.models.inventory import Warehouse
from app.db.models.product import Product
from app.db.models.purchase import PurchaseOrder, PurchaseOrderItem, Supplier
from app.db.models.purchase_unit import PurchaseUnit
from app.db.models.shipment import ShipmentRecord
from app.db.models.shipment_etl_fingerprint import ShipmentEtlImportFingerprint
from app.infrastructure.tenant_scope import tenant_scope
from app.services.ocr_service import OCRService


@pytest.fixture()
def etl_db(tmp_path, monkeypatch):
    db_path = tmp_path / "etl.sqlite"
    engine = create_engine(f"sqlite:///{db_path}")
    Base.metadata.create_all(
        engine,
        tables=[
            EtlUpload.__table__,
            EtlTemplate.__table__,
            EtlTemplateVersion.__table__,
            EtlRun.__table__,
            EtlRunRow.__table__,
            EtlTargetConfig.__table__,
            PurchaseUnit.__table__,
            Product.__table__,
            Supplier.__table__,
            Warehouse.__table__,
            PurchaseOrder.__table__,
            PurchaseOrderItem.__table__,
            ShipmentRecord.__table__,
            ShipmentEtlImportFingerprint.__table__,
        ],
    )
    maker = sessionmaker(bind=engine, expire_on_commit=False)
    monkeypatch.setenv("XCAGI_DATA_DIR", str(tmp_path / "app-data"))
    monkeypatch.setattr("app.application.etl.service.SessionLocal", maker)
    return maker


def test_safe_transform_dsl_rejects_dynamic_execution():
    with pytest.raises(EtlError, match="不允许的转换操作"):
        apply_transform("x", {"op": "python", "code": "__import__('os')"}, {})

    with pytest.raises(EtlError, match="不允许的公式操作符"):
        apply_transform("1", {"op": "formula", "operator": "eval", "operands": ["1"]}, {})

    result = apply_mapping(
        {"单价": "1,200.50", "数量": "2"},
        [
            {"source": "单价", "target": "price", "transforms": [{"op": "number"}]},
            {
                "source": "",
                "target": "amount",
                "transforms": [
                    {
                        "op": "formula",
                        "operator": "mul",
                        "operands": [{"field": "price"}, {"field": "数量"}],
                    }
                ],
            },
        ],
    )
    assert result == {"price": "1200.50", "amount": "2401.00"}
    assert neutralize_spreadsheet_formula("=HYPERLINK('x')").startswith("'")


def test_global_excel_etl_knowledge_base_is_read_only_in_runtime(tmp_path):
    path = tmp_path / "excel_etl_kb.json"
    path.write_text(
        json.dumps(
            {
                "templates": {
                    "existing": TemplateMemory(
                        fingerprint="existing",
                        columns={"product_name": 1},
                    ).to_dict()
                }
            },
            ensure_ascii=False,
        ),
        encoding="utf-8",
    )
    original = path.read_bytes()
    kb = ExcelEtlKnowledgeBase(path)
    assert kb.touch("existing") is not None
    kb.remember(TemplateMemory(fingerprint="new", columns={"product_name": 2}))
    assert kb.get_template("new") is None
    assert path.read_bytes() == original


def test_csv_parser_preserves_source_row_and_rejects_word_business_target(tmp_path):
    csv_path = tmp_path / "customers.csv"
    csv_path.write_text("客户名称,电话\n甲公司,13800000000\n", encoding="utf-8")
    dataset = parse_file(csv_path, target_type="customers")
    assert dataset.headers == ["客户名称", "电话"]
    assert dataset.rows[0].row_number == 2
    assert dataset.rows[0].provenance["original_fragment"]["客户名称"] == "甲公司"

    doc_path = tmp_path / "memo.docx"
    doc_path.write_bytes(b"placeholder")
    with pytest.raises(EtlError, match="仅可导入知识库"):
        parse_file(doc_path, target_type="customers")


def test_delivery_profile_converts_to_general_etl_rows(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    path = tmp_path / "delivery.xlsx"
    result = write_delivery_note_workbook(
        [
            {
                "unit_name": "兼容客户",
                "order_number": "COMPAT-1",
                "items": [
                    {
                        "model_number": "A1",
                        "product_name": "底漆",
                        "quantity_tins": 1,
                        "tin_spec": 20,
                        "quantity_kg": 20,
                        "unit_price": 10,
                        "amount": 200,
                    },
                    {
                        "model_number": "B2",
                        "product_name": "面漆",
                        "quantity_tins": 2,
                        "tin_spec": 20,
                        "quantity_kg": 40,
                        "unit_price": 12,
                        "amount": 480,
                    },
                ],
            }
        ],
        path,
    )
    assert result["success"] is True

    with tenant_scope(1):
        dataset = parse_file(path, target_type="shipment_records")

    assert dataset.source_features["kind"] == "shipment_profile"
    assert dataset.source_features["compatibility_preset"] is True
    assert len(dataset.rows) == 2
    values = [row.values for row in dataset.rows]
    assert {row["product_name"] for row in values} == {"底漆", "面漆"}
    assert len({row["source_fingerprint"] for row in values}) == 2
    assert len({row["legacy_note_fingerprint"] for row in values}) == 1


def test_explicit_compatibility_preset_is_used_for_parsing(tmp_path, monkeypatch):
    path = tmp_path / "explicit.xlsx"
    path.write_bytes(b"placeholder")
    captured: dict[str, object] = {}

    def preview(*_args, **kwargs):
        captured.update(kwargs)
        return {
            "success": True,
            "notes": [
                {
                    "sheet": "送货单",
                    "unit_name": "甲公司",
                    "order_number": "A-1",
                    "assist": {"ok": True},
                    "profile_id": "custom-profile",
                    "fingerprint": "fp-1",
                    "items": [{"product_name": "底漆", "model_number": "P-1"}],
                }
            ],
        }

    monkeypatch.setattr(
        "app.application.shipment_excel_etl_app_service.preview_shipment_excel_etl",
        preview,
    )

    dataset = parse_file(
        path,
        target_type="customer_products",
        compatibility_preset_id="custom-profile",
    )

    assert captured["profile_id"] == "custom-profile"
    assert dataset.source_features["compatibility_preset_id"] == "custom-profile"
    assert dataset.rows[0].values["customer_name"] == "甲公司"


def test_delivery_profile_skips_filename_fallback_sheets_and_total_rows(tmp_path, monkeypatch):
    path = tmp_path / "855237eb-59a4-419e-b22c-360ea04a8a56.xlsx"
    path.write_bytes(b"placeholder")
    notes = [
        {
            "sheet": "送货单",
            "unit_name": "甲公司",
            "order_number": "A-1",
            "score": 50,
            "assist": {"ok": True},
            "profile_id": "universal",
            "fingerprint": "good",
            "contact_person": "日期：2026年01月19日",
            "items": [
                {"product_name": "底漆", "model_number": "P-1", "unit_price": 10},
                {
                    "product_name": "100",
                    "model_number": "大 写 人 民 币",
                    "unit_price": 0,
                },
            ],
        },
        {
            "sheet": "回款明细",
            "unit_name": path.stem,
            "order_number": "",
            "score": 12,
            "assist": {"ok": False},
            "profile_id": "universal",
            "fingerprint": "bad",
            "items": [{"product_name": "伪产品", "model_number": ""}],
        },
    ]
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_app_service.preview_shipment_excel_etl",
        lambda *_args, **_kwargs: {"success": True, "notes": notes},
    )

    customers = parse_file(path, target_type="customers")
    products = parse_file(path, target_type="products")
    linked = parse_file(path, target_type="customer_products")

    assert [row.values["customer_name"] for row in customers.rows] == ["甲公司"]
    assert "contact_person" not in customers.rows[0].values
    assert [row.values["name"] for row in products.rows] == ["底漆"]
    assert linked.rows[0].values == {
        "customer_name": "甲公司",
        "model_number": "P-1",
        "name": "底漆",
        "price": 10,
    }
    assert customers.source_features["skipped_note_count"] == 1
    assert any(
        warning["code"] == "ETL_COMPATIBILITY_LOW_CONFIDENCE_SHEETS_SKIPPED"
        for warning in customers.warnings
    )


def test_macos_vision_fallback_returns_auditable_positioned_blocks(monkeypatch):
    service = OCRService.__new__(OCRService)
    service.macos_vision_available = True
    monkeypatch.setattr(
        "app.services.ocr_service.subprocess.run",
        lambda *_args, **_kwargs: SimpleNamespace(
            returncode=0,
            stdout=json.dumps(
                [
                    {
                        "text": "客户名称",
                        "confidence": 0.92,
                        "x": 0.1,
                        "y": 0.7,
                        "width": 0.2,
                        "height": 0.1,
                    }
                ]
            ),
            stderr="",
        ),
    )
    blocks = service._recognize_macos_vision_blocks(np.zeros((100, 200, 3), dtype=np.uint8))
    assert blocks == [
        {
            "text": "客户名称",
            "left": 20.0,
            "top": pytest.approx(20.0),
            "width": 40.0,
            "height": 10.0,
            "confidence": 0.92,
            "center": (40.0, pytest.approx(25.0)),
            "y_center": pytest.approx(25.0),
        }
    ]


def test_llm_adviser_failure_never_changes_deterministic_action(etl_db, monkeypatch):
    def unavailable(_payload):
        raise RuntimeError("model unavailable")

    service = EtlService(adviser=EtlRowAdviser(unavailable))
    monkeypatch.setattr(
        service,
        "_submit_preview",
        lambda run_id, _tenant_id, owner_user_id: service._preview_worker(run_id, owner_user_id),
    )
    with tenant_scope(6):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=10,
            file_name="llm-fallback.csv",
            content_type="text/csv",
            stream=BytesIO("客户名称\n确定性客户\n".encode()),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=10,
            upload_id=upload["upload_id"],
            target_type="customers",
        )
        rows = service.get_rows(
            db,
            run_id=run["id"],
            owner_user_id=10,
            page=1,
            page_size=10,
        )
        assert rows["items"][0]["suggested_action"] == "new"
        assert rows["items"][0]["final_action"] == "new"
        assert rows["items"][0]["llm_suggestion"]["degradation_code"] == "ETL_LLM_UNAVAILABLE"
        persisted = service.get_run(db, run_id=run["id"], owner_user_id=10)
        assert persisted["details"]["llm_degraded"] is True
        db.close()


def test_batch_llm_advice_never_overrides_adapter_action(etl_db, monkeypatch):
    def batch_advice(_payloads):
        return {
            "items": [{"index": 0, "action": "skip", "reason": "模型建议跳过"}],
            "metadata": {
                "used_llm": True,
                "advisory_only": True,
                "degraded": False,
                "model": "software-model",
            },
        }

    service = EtlService(adviser=EtlRowAdviser(batch_provider=batch_advice))
    monkeypatch.setattr(
        service,
        "_submit_preview",
        lambda run_id, _tenant_id, owner_user_id: service._preview_worker(run_id, owner_user_id),
    )
    with tenant_scope(6):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=10,
            file_name="llm-advice.csv",
            content_type="text/csv",
            stream=BytesIO("客户名称\n确定性客户\n".encode()),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=10,
            upload_id=upload["upload_id"],
            target_type="customers",
        )
        rows = service.get_rows(
            db,
            run_id=run["id"],
            owner_user_id=10,
            page=1,
            page_size=10,
        )
        assert rows["items"][0]["suggested_action"] == "new"
        assert rows["items"][0]["final_action"] == "new"
        assert rows["items"][0]["llm_suggestion"]["action"] == "skip"
        assert rows["items"][0]["llm_suggestion"]["advisory_only"] is True
        db.close()


def test_preview_blocks_invalid_rows_execute_valid_rows_and_rollback(etl_db, monkeypatch):
    service = EtlService()

    def run_inline(run_id: str, _tenant_id: int, owner_user_id: int) -> None:
        service._preview_worker(run_id, owner_user_id)

    monkeypatch.setattr(service, "_submit_preview", run_inline)
    monkeypatch.setattr(
        service,
        "_submit_execution",
        lambda run_id, _tenant_id, owner_user_id, valid_rows_only: service._execute_worker(
            run_id, owner_user_id, valid_rows_only
        ),
    )
    with tenant_scope(7):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=11,
            file_name="customers.csv",
            content_type="text/csv",
            stream=BytesIO("客户名称,联系人,电话\n甲公司,张三,138\n,李四,139\n".encode()),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=11,
            upload_id=upload["upload_id"],
            target_type="customers",
        )
        assert run["status"] == "preview_ready"
        assert run["summary"] == {
            "new": 1,
            "update": 0,
            "skip": 0,
            "error": 1,
            "executed": 0,
        }

        with pytest.raises(EtlConflict) as blocked:
            service.execute(
                db,
                run_id=run["id"],
                owner_user_id=11,
                confirmed=True,
                valid_rows_only=False,
            )
        assert blocked.value.code == "ETL_INVALID_ROWS_BLOCKED"

        completed = service.execute(
            db,
            run_id=run["id"],
            owner_user_id=11,
            confirmed=True,
            valid_rows_only=True,
        )
        assert completed["status"] == "completed"
        assert completed["receipt"]["partial"] is True
        assert db.query(PurchaseUnit).filter(PurchaseUnit.unit_name == "甲公司").count() == 1

        rolled_back = service.rollback(db, run_id=run["id"], owner_user_id=11)
        assert rolled_back["rollback_status"] == "completed"
        assert db.query(PurchaseUnit).filter(PurchaseUnit.unit_name == "甲公司").count() == 0
        db.close()


def test_internal_partial_failure_retries_only_unfinished_rows(etl_db, monkeypatch):
    service = EtlService()

    def run_inline(run_id: str, _tenant_id: int, owner_user_id: int) -> None:
        service._preview_worker(run_id, owner_user_id)

    monkeypatch.setattr(service, "_submit_preview", run_inline)
    monkeypatch.setattr(
        service,
        "_submit_execution",
        lambda run_id, _tenant_id, owner_user_id, valid_rows_only: service._execute_worker(
            run_id, owner_user_id, valid_rows_only
        ),
    )
    monkeypatch.setattr(
        service,
        "_submit_revalidation",
        lambda run_id, _tenant_id, owner_user_id, _overrides=None: (
            service._revalidate_existing_rows(db, run_id, owner_user_id)
        ),
    )
    adapter = get_adapter("customers")
    original_execute = adapter.execute_row
    failed_once = False

    def fail_second_row(db, data, **kwargs):
        nonlocal failed_once
        if data.get("customer_name") == "乙公司" and not failed_once:
            failed_once = True
            raise EtlError("ETL_TEST_TRANSIENT", "模拟第二行瞬时失败")
        return original_execute(db, data, **kwargs)

    monkeypatch.setattr(adapter, "execute_row", fail_second_row)
    with tenant_scope(8):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=12,
            file_name="retry.csv",
            content_type="text/csv",
            stream=BytesIO("客户名称\n甲公司\n乙公司\n".encode()),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=12,
            upload_id=upload["upload_id"],
            target_type="customers",
        )
        failed = service.execute(
            db,
            run_id=run["id"],
            owner_user_id=12,
            confirmed=True,
            valid_rows_only=False,
        )
        assert failed["status"] == "failed"
        assert failed["error"]["code"] == "ETL_TEST_TRANSIENT"
        assert db.query(PurchaseUnit).filter(PurchaseUnit.unit_name == "甲公司").count() == 1
        assert db.query(PurchaseUnit).filter(PurchaseUnit.unit_name == "乙公司").count() == 0

        retried = service.retry(db, run_id=run["id"], owner_user_id=12)
        assert retried["status"] == "preview_ready"
        completed = service.execute(
            db,
            run_id=run["id"],
            owner_user_id=12,
            confirmed=True,
            valid_rows_only=False,
        )
        assert completed["summary"]["executed"] == 2
        assert db.query(PurchaseUnit).count() == 2

        service.rollback(db, run_id=run["id"], owner_user_id=12)
        assert db.query(PurchaseUnit).count() == 0
        db.close()


def test_templates_are_private_and_versions_are_immutable(etl_db):
    service = EtlService()
    draft = {
        "field_mappings": [
            {
                "source": "客户",
                "target": "customer_name",
                "transforms": [{"op": "trim"}],
                "confidence": 1,
                "required": True,
            }
        ],
        "allowed_update_fields": [],
        "match_keys": ["customer_name"],
        "action_rules": {"duplicate": "skip"},
    }
    with tenant_scope(9):
        db = etl_db()
        template = service.create_template(
            db,
            owner_user_id=1,
            name="客户导入",
            target_type="customers",
            draft=draft,
        )
        db.commit()
        service.update_template(
            db,
            template_id=template["id"],
            owner_user_id=1,
            draft={**draft, "allowed_update_fields": ["contact_phone"]},
        )
        db.commit()
        versions = service.template_versions(db, template_id=template["id"], owner_user_id=1)
        assert [item["version"]["number"] for item in versions] == [2, 1]
        assert versions[1]["version"]["allowed_update_fields"] == []
        with pytest.raises(EtlNotFound):
            service.template_versions(db, template_id=template["id"], owner_user_id=2)
        db.close()
    with tenant_scope(10):
        db = etl_db()
        with pytest.raises(EtlNotFound):
            service.get_template(db, template_id=template["id"], owner_user_id=1)
        db.close()


def test_preview_validates_and_audits_selected_compatibility_preset(etl_db, monkeypatch):
    service = EtlService()
    monkeypatch.setattr(service, "_submit_preview", lambda *_args: None)
    monkeypatch.setattr(
        "app.application.shipment_etl_profile.list_profiles",
        lambda: [
            {
                "id": "legacy-profile",
                "label": "旧送货单",
                "source": "yaml",
                "target": "shipment",
            }
        ],
    )
    with tenant_scope(11):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=3,
            file_name="delivery.xlsx",
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            stream=BytesIO(b"placeholder"),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=3,
            upload_id=upload["upload_id"],
            target_type="customer_products",
            compatibility_preset_id="legacy-profile",
        )
        assert run["draft"]["compatibility_preset_id"] == "legacy-profile"

        template = service.create_template(
            db,
            owner_user_id=3,
            name="我的旧送货单模板",
            target_type="customer_products",
            draft=run["draft"],
            source_features={"compatibility_preset_id": "legacy-profile"},
        )
        db.commit()
        inherited = service.create_preview(
            db,
            owner_user_id=3,
            upload_id=upload["upload_id"],
            target_type="customer_products",
            template_id=template["id"],
        )
        assert inherited["draft"]["compatibility_preset_id"] == "legacy-profile"

        with pytest.raises(EtlError) as exc:
            service.create_preview(
                db,
                owner_user_id=3,
                upload_id=upload["upload_id"],
                target_type="customer_products",
                compatibility_preset_id="missing-profile",
            )
        assert exc.value.code == "ETL_COMPATIBILITY_PRESET_NOT_FOUND"
        db.close()


def test_product_adapter_requires_confirmed_update_fields_and_rolls_back(etl_db):
    adapter = ProductAdapter()
    with tenant_scope(12):
        db = etl_db()
        data = {"unit": "甲公司", "model_number": "A1", "name": "底漆", "price": "10"}
        preview = adapter.preview(db, data, allowed_update_fields=set(), context={})
        assert preview.action == "new"
        created = adapter.execute_row(
            db,
            data,
            action="new",
            match_ref="",
            allowed_update_fields=set(),
            context={},
        )
        db.commit()

        unchanged = adapter.preview(
            db,
            {**data, "price": "12"},
            allowed_update_fields=set(),
            context={},
        )
        assert unchanged.action == "skip"
        update = adapter.preview(
            db,
            {**data, "price": "12"},
            allowed_update_fields={"price"},
            context={},
        )
        assert update.action == "update"
        adapter.execute_row(
            db,
            {**data, "price": "12"},
            action="update",
            match_ref=update.match_ref,
            allowed_update_fields={"price"},
            context={},
        )
        db.commit()
        assert str(db.get(Product, int(created["match_ref"])).price) == "12.00"
        adapter.rollback_row(
            db,
            match_ref=update.match_ref,
            before=update.before or {},
            after=update.after or {},
            context={},
        )
        db.commit()
        assert str(db.get(Product, int(created["match_ref"])).price) == "10.00"
        db.close()


def test_customer_products_adapter_links_parent_and_children_and_rolls_back(etl_db):
    adapter = CustomerProductsAdapter()
    first = {
        "customer_name": "甲公司",
        "contact_person": "张总",
        "model_number": "A1",
        "name": "底漆",
        "price": "10",
    }
    second = {
        "customer_name": "甲公司",
        "contact_person": "张总",
        "model_number": "A2",
        "name": "面漆",
        "price": "20",
    }
    with tenant_scope(121):
        db = etl_db()
        preview_context = {"_preview_cache": {}}
        first_preview = adapter.preview(
            db,
            first,
            allowed_update_fields=set(),
            context=preview_context,
        )
        second_preview = adapter.preview(
            db,
            second,
            allowed_update_fields=set(),
            context=preview_context,
        )
        repeated_preview = adapter.preview(
            db,
            first,
            allowed_update_fields=set(),
            context=preview_context,
        )
        conflicting_parent = adapter.preview(
            db,
            {
                **first,
                "contact_person": "李总",
                "model_number": "A3",
                "name": "清漆",
            },
            allowed_update_fields=set(),
            context=preview_context,
        )
        assert first_preview.action == "new"
        assert second_preview.action == "new"
        assert repeated_preview.action == "skip"
        assert conflicting_parent.action == "error"
        assert conflicting_parent.issues[0]["code"] == "ETL_PARENT_FIELDS_CONFLICT"

        execution_context = {"_execution_cache": {}}
        first_result = adapter.execute_row(
            db,
            first,
            action="new",
            match_ref=first_preview.match_ref,
            allowed_update_fields=set(),
            context=execution_context,
        )
        second_result = adapter.execute_row(
            db,
            second,
            action="new",
            match_ref=second_preview.match_ref,
            allowed_update_fields=set(),
            context=execution_context,
        )
        db.commit()

        assert db.query(PurchaseUnit).filter(PurchaseUnit.unit_name == "甲公司").count() == 1
        products = (
            db.query(Product).filter(Product.unit == "甲公司").order_by(Product.model_number).all()
        )
        assert [item.model_number for item in products] == ["A1", "A2"]
        assert first_result["after"]["_etl"]["customer_created"] is True
        assert second_result["after"]["_etl"]["customer_created"] is False

        duplicate = adapter.preview(
            db,
            first,
            allowed_update_fields=set(),
            context={"_preview_cache": {}},
        )
        assert duplicate.action == "skip"

        adapter.rollback_row(
            db,
            match_ref=second_result["match_ref"],
            before=second_preview.before or {},
            after=second_result["after"],
            context={},
        )
        db.commit()
        adapter.rollback_row(
            db,
            match_ref=first_result["match_ref"],
            before=first_preview.before or {},
            after=first_result["after"],
            context={},
        )
        db.commit()
        assert db.query(Product).count() == 0
        assert db.query(PurchaseUnit).count() == 0
        db.close()


def test_customer_products_adapter_updates_both_images_and_restores_them(etl_db):
    adapter = CustomerProductsAdapter()
    with tenant_scope(123):
        db = etl_db()
        customer = PurchaseUnit(
            tenant_id=123,
            unit_name="甲公司",
            contact_person="旧联系人",
            is_active=True,
        )
        product = Product(
            tenant_id=123,
            unit="甲公司",
            model_number="A1",
            name="底漆",
            price="10",
        )
        db.add_all([customer, product])
        db.commit()
        data = {
            "customer_name": "甲公司",
            "contact_person": "新联系人",
            "model_number": "A1",
            "name": "底漆",
            "price": "12",
        }
        preview = adapter.preview(
            db,
            data,
            allowed_update_fields={"contact_person", "price"},
            context={"_preview_cache": {}},
        )
        assert preview.action == "update"
        result = adapter.execute_row(
            db,
            data,
            action="update",
            match_ref=preview.match_ref,
            allowed_update_fields={"contact_person", "price"},
            context={"_execution_cache": {}},
        )
        db.commit()
        assert customer.contact_person == "新联系人"
        assert float(product.price) == 12

        adapter.rollback_row(
            db,
            match_ref=result["match_ref"],
            before=preview.before or {},
            after=result["after"],
            context={},
        )
        db.commit()
        assert customer.contact_person == "旧联系人"
        assert float(product.price) == 10
        db.close()


def test_customer_products_service_executes_one_confirmed_linked_run(etl_db, monkeypatch):
    service = EtlService()
    monkeypatch.setattr(
        service,
        "_submit_preview",
        lambda run_id, _tenant_id, owner_user_id: service._preview_worker(run_id, owner_user_id),
    )
    monkeypatch.setattr(
        service,
        "_submit_execution",
        lambda run_id, _tenant_id, owner_user_id, valid_rows_only: service._execute_worker(
            run_id, owner_user_id, valid_rows_only
        ),
    )
    with tenant_scope(122):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=21,
            file_name="linked.csv",
            content_type="text/csv",
            stream=BytesIO(
                (
                    "客户名称,联系人,产品型号,产品名称,价格\n"
                    "甲公司,张总,A1,底漆,10\n"
                    "甲公司,张总,A2,面漆,20\n"
                ).encode()
            ),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=21,
            upload_id=upload["upload_id"],
            target_type="customer_products",
        )
        assert run["status"] == "preview_ready"
        assert run["file_name"] == "linked.csv"
        assert run["file_sha256"] == upload["sha256"]
        persisted_run = db.get(EtlRun, run["id"])
        persisted_run.summary_json = "{}"
        db.commit()
        assert service.get_run(db, run_id=run["id"], owner_user_id=21)["file_name"] == "linked.csv"
        assert run["summary"]["new"] == 2
        completed = service.execute(
            db,
            run_id=run["id"],
            owner_user_id=21,
            confirmed=True,
            valid_rows_only=False,
        )
        assert completed["status"] == "completed"
        assert completed["summary"]["executed"] == 2
        assert db.query(PurchaseUnit).filter(PurchaseUnit.unit_name == "甲公司").count() == 1
        assert db.query(Product).filter(Product.unit == "甲公司").count() == 2

        rolled_back = service.rollback(db, run_id=run["id"], owner_user_id=21)
        assert rolled_back["rollback_status"] == "completed"
        assert db.query(PurchaseUnit).count() == 0
        assert db.query(Product).count() == 0
        db.close()


def test_customer_products_retry_replays_parent_without_orphaning(etl_db, monkeypatch):
    service = EtlService()
    monkeypatch.setattr(
        service,
        "_submit_preview",
        lambda run_id, _tenant_id, owner_user_id: service._preview_worker(run_id, owner_user_id),
    )
    monkeypatch.setattr(
        service,
        "_submit_execution",
        lambda run_id, _tenant_id, owner_user_id, valid_rows_only: service._execute_worker(
            run_id, owner_user_id, valid_rows_only
        ),
    )
    monkeypatch.setattr(
        service,
        "_submit_revalidation",
        lambda run_id, _tenant_id, owner_user_id, _overrides=None: (
            service._revalidate_existing_rows(db, run_id, owner_user_id)
        ),
    )
    adapter = get_adapter("customer_products")
    original_execute = adapter.execute_row
    failed_once = False

    def fail_second_product(db, data, **kwargs):
        nonlocal failed_once
        if data.get("model_number") == "A2" and not failed_once:
            failed_once = True
            raise EtlError("ETL_TEST_TRANSIENT", "模拟关联产品瞬时失败")
        return original_execute(db, data, **kwargs)

    monkeypatch.setattr(adapter, "execute_row", fail_second_product)
    with tenant_scope(124):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=22,
            file_name="linked-retry.csv",
            content_type="text/csv",
            stream=BytesIO(
                ("客户名称,产品型号,产品名称\n甲公司,A1,底漆\n甲公司,A2,面漆\n").encode()
            ),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=22,
            upload_id=upload["upload_id"],
            target_type="customer_products",
        )
        failed = service.execute(
            db,
            run_id=run["id"],
            owner_user_id=22,
            confirmed=True,
            valid_rows_only=False,
        )
        assert failed["status"] == "failed"
        assert db.query(PurchaseUnit).count() == 1
        assert db.query(Product).filter(Product.unit == "甲公司").count() == 1

        retried = service.retry(db, run_id=run["id"], owner_user_id=22)
        assert retried["status"] == "preview_ready"
        completed = service.execute(
            db,
            run_id=run["id"],
            owner_user_id=22,
            confirmed=True,
            valid_rows_only=False,
        )
        assert completed["status"] == "completed"
        assert db.query(PurchaseUnit).count() == 1
        assert db.query(Product).filter(Product.unit == "甲公司").count() == 2

        service.rollback(db, run_id=run["id"], owner_user_id=22)
        assert db.query(PurchaseUnit).count() == 0
        assert db.query(Product).count() == 0
        db.close()


def test_customer_adapter_uses_visible_purchase_units_and_tenant_scope(etl_db):
    adapter = CustomerAdapter()
    with tenant_scope(13):
        db = etl_db()
        original = PurchaseUnit(
            tenant_id=13,
            unit_name="同名客户",
            contact_phone="100",
            is_active=True,
        )
        db.add(original)
        db.commit()
        preview = adapter.preview(
            db,
            {"customer_name": "同名客户", "contact_phone": "200"},
            allowed_update_fields={"contact_phone"},
            context={},
        )
        assert preview.action == "update"
        updated = adapter.execute_row(
            db,
            {"customer_name": "同名客户", "contact_phone": "200"},
            action="update",
            match_ref=preview.match_ref,
            allowed_update_fields={"contact_phone"},
            context={},
        )
        db.commit()
        assert original.contact_phone == "200"
        adapter.rollback_row(
            db,
            match_ref=preview.match_ref,
            before=preview.before or {},
            after=updated["after"],
            context={},
        )
        db.commit()
        assert original.contact_phone == "100"
        db.close()

    with tenant_scope(14):
        db = etl_db()
        isolated = adapter.preview(
            db,
            {"customer_name": "同名客户"},
            allowed_update_fields=set(),
            context={},
        )
        assert isolated.action == "new"
        db.close()


def test_purchase_order_v1_add_skip_and_rollback(etl_db):
    adapter = PurchaseOrderAdapter()
    with tenant_scope(14):
        db = etl_db()
        supplier = Supplier(tenant_id=14, code="SUP-1", name="甲供应商")
        product = Product(
            tenant_id=14,
            unit="内部采购",
            model_number="M1",
            name="树脂",
        )
        db.add_all([supplier, product])
        db.commit()
        data = {
            "external_order_no": "PO-100",
            "supplier_name": "甲供应商",
            "order_date": "2026-07-26",
            "product_model": "M1",
            "product_name": "树脂",
            "quantity": "2",
            "unit": "桶",
            "unit_price": "88",
        }
        preview = adapter.preview(db, data, allowed_update_fields=set(), context={})
        assert preview.action == "new"
        created = adapter.execute_row(
            db,
            data,
            action="new",
            match_ref="",
            allowed_update_fields=set(),
            context={},
        )
        db.commit()
        duplicate = adapter.preview(db, data, allowed_update_fields=set(), context={})
        assert duplicate.action == "skip"
        assert duplicate.reason == "existing_order_v1_no_update"

        adapter.rollback_row(
            db,
            match_ref=created["match_ref"],
            before={},
            after=created["after"],
            context={},
        )
        db.commit()
        assert db.query(PurchaseOrderItem).count() == 0
        assert db.query(PurchaseOrder).count() == 0
        db.close()


def test_purchase_order_number_is_unique_per_tenant_for_etl(etl_db):
    adapter = PurchaseOrderAdapter()
    shared_order_no = "PO-SHARED-100"

    with tenant_scope(141):
        db = etl_db()
        supplier = Supplier(tenant_id=141, code="SUP-141", name="甲供应商")
        product = Product(
            tenant_id=141,
            unit="内部采购",
            model_number="M-SHARED",
            name="共享型号树脂",
        )
        db.add_all([supplier, product])
        db.commit()
        adapter.execute_row(
            db,
            {
                "external_order_no": shared_order_no,
                "supplier_name": supplier.name,
                "order_date": "2026-07-27",
                "product_model": product.model_number,
                "product_name": product.name,
                "quantity": "1",
            },
            action="new",
            match_ref="",
            allowed_update_fields=set(),
            context={},
        )
        db.commit()
        assert (
            db.query(PurchaseOrder).filter(PurchaseOrder.order_no == shared_order_no).count() == 1
        )
        db.close()

    with tenant_scope(142):
        db = etl_db()
        supplier = Supplier(tenant_id=142, code="SUP-142", name="乙供应商")
        product = Product(
            tenant_id=142,
            unit="内部采购",
            model_number="M-SHARED",
            name="共享型号树脂",
        )
        db.add_all([supplier, product])
        db.commit()
        data = {
            "external_order_no": shared_order_no,
            "supplier_name": supplier.name,
            "order_date": "2026-07-27",
            "product_model": product.model_number,
            "product_name": product.name,
            "quantity": "2",
        }
        assert adapter.preview(db, data, allowed_update_fields=set(), context={}).action == "new"
        adapter.execute_row(
            db,
            data,
            action="new",
            match_ref="",
            allowed_update_fields=set(),
            context={},
        )
        db.commit()
        assert (
            db.query(PurchaseOrder).filter(PurchaseOrder.order_no == shared_order_no).count() == 1
        )
        db.close()


def test_shipment_adapter_consults_legacy_and_general_fingerprints(etl_db):
    adapter = ShipmentAdapter()
    with tenant_scope(15):
        db = etl_db()
        legacy = ShipmentRecord(
            tenant_id=15,
            purchase_unit="甲公司",
            product_name="底漆",
            quantity_kg=20,
            quantity_tins=1,
            status="pending",
            raw_text="source=shipment_excel_etl|external_order_number=SO-1",
        )
        db.add(legacy)
        db.flush()
        db.add(
            ShipmentEtlImportFingerprint(
                tenant_key="tenant:15",
                fingerprint="legacy-fingerprint",
                shipment_id=legacy.id,
                unit_name="甲公司",
                order_number="SO-1",
                file_name="legacy.xlsx",
                source_kind="shipment",
            )
        )
        db.commit()
        data = {
            "purchase_unit": "甲公司",
            "external_order_no": "SO-1",
            "product_name": "底漆",
            "quantity_tins": "1",
        }
        legacy_duplicate = adapter.preview(
            db,
            data,
            allowed_update_fields=set(),
            context={
                "file_sha256": "different-hash",
                "file_name": "renamed.xlsx",
                "source_row": 2,
            },
        )
        assert legacy_duplicate.action == "skip"
        assert legacy_duplicate.reason == "legacy_source_duplicate"

        fresh = {**data, "external_order_no": "SO-2"}
        context = {
            "file_sha256": "new-hash",
            "file_name": "new.xlsx",
            "source_row": 2,
            "run_id": "shipment-run",
        }
        preview = adapter.preview(db, fresh, allowed_update_fields=set(), context=context)
        assert preview.action == "new"
        created = adapter.execute_row(
            db,
            fresh,
            action="new",
            match_ref="",
            allowed_update_fields=set(),
            context=context,
        )
        db.commit()
        duplicate = adapter.preview(db, fresh, allowed_update_fields=set(), context=context)
        assert duplicate.action == "skip"
        assert duplicate.reason == "legacy_fingerprint_duplicate"

        adapter.rollback_row(
            db,
            match_ref=created["match_ref"],
            before={},
            after=created["after"],
            context=context,
        )
        db.commit()
        assert db.get(ShipmentRecord, int(created["match_ref"])) is None
        assert (
            db.query(ShipmentEtlImportFingerprint)
            .filter(ShipmentEtlImportFingerprint.source_kind == "general_etl")
            .count()
            == 0
        )
        db.close()


def test_export_neutralizes_formulas_and_webhook_chunks_have_idempotency(tmp_path, monkeypatch):
    monkeypatch.setenv("XCAGI_DATA_DIR", str(tmp_path / "app-data"))
    export = ExportCsvAdapter().execute_batch(
        [{"name": "=HYPERLINK('https://bad')", "value": 1}],
        {"run_id": "export-run"},
    )
    assert export["executed"] == 1
    exported_path = tmp_path / "app-data" / "etl" / "exports" / "etl-export-run.csv"
    assert "'=HYPERLINK" in exported_path.read_text(encoding="utf-8-sig")

    calls: list[dict] = []

    class Response:
        status_code = 200

    class Client:
        def __init__(self, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def post(self, endpoint, *, json, headers):
            calls.append({"endpoint": endpoint, "json": json, "headers": headers})
            return Response()

    monkeypatch.setattr("httpx.Client", Client)
    monkeypatch.setattr("app.application.etl.targets._assert_safe_webhook_url", lambda _url: None)
    result = WebhookAdapter().execute_batch(
        [{"index": index} for index in range(501)],
        {
            "run_id": "webhook-run",
            "target_config": {
                "endpoint_url": "https://example.com/hook",
                "headers": {},
                "secret_ref": None,
            },
        },
    )
    assert result["executed"] == 501
    assert len(calls) == 2
    assert calls[0]["headers"]["Idempotency-Key"] == "webhook-run:0"
    assert calls[1]["headers"]["Idempotency-Key"] == "webhook-run:1"
    assert len(calls[0]["json"]["rows"]) == 500
    assert len(calls[1]["json"]["rows"]) == 1


def test_target_contract_catalog_covers_v1_targets_without_secret_fields():
    capabilities = {item["type"]: item for item in target_capabilities()}
    assert set(capabilities) == {
        "knowledge",
        "customer_products",
        "customers",
        "products",
        "purchase_orders",
        "shipment_records",
        "attendance",
        "export_xlsx",
        "export_csv",
        "webhook",
    }
    assert capabilities["customer_products"]["reversible"] is True
    assert capabilities["customers"]["reversible"] is True
    assert capabilities["attendance"]["reversible"] is True
    assert capabilities["webhook"]["reversible"] is False
    for target_type, capability in capabilities.items():
        adapter = get_adapter(target_type)
        fields = {field["key"] for field in capability["fields"]}
        assert capability["type"] == target_type
        assert capability["label"]
        assert set(capability["required_fields"]).issubset(fields)
        assert set(capability["supported_actions"]).issubset({"new", "update", "skip"})
        assert capability["supported_actions"]
        if target_type not in {"knowledge", "attendance"}:
            assert set(capability["default_match_keys"]).issubset(fields)
        has_batch_execute = callable(getattr(adapter, "execute_batch", None))
        assert has_batch_execute or type(adapter).execute_row is not TargetAdapter.execute_row
        if capability["reversible"]:
            has_batch_rollback = callable(getattr(adapter, "rollback_batch", None))
            assert (
                has_batch_rollback or type(adapter).rollback_row is not TargetAdapter.rollback_row
            )
        assert "secret_ref" not in json.dumps(capability, ensure_ascii=False)


def test_attendance_adapter_is_idempotent_and_rolls_back_only_its_source(tmp_path, monkeypatch):
    data_root = tmp_path / "app-data"
    monkeypatch.setenv("XCAGI_DATA_DIR", str(data_root))
    source_path = tmp_path / "attendance.xlsx"
    source_path.write_bytes(b"workbook-placeholder")
    adapter = AttendanceAdapter()
    source_key = f"attendance-hash:{source_path.name}"

    preview_context = {
        "upload_path": str(source_path),
        "file_sha256": "attendance-hash",
        "_preview_cache": {},
    }
    preview = adapter.preview(
        None,
        {},
        allowed_update_fields=set(),
        context=preview_context,
    )
    assert preview.action == "new"
    assert preview.after["source_file"] == source_key

    def fake_import(_source_path, db_path, *, source_file_key, sync_ui_tables):
        assert source_file_key == source_key
        assert sync_ui_tables is True
        db_path.parent.mkdir(parents=True, exist_ok=True)
        with sqlite3.connect(str(db_path)) as conn:
            conn.executescript(
                """
                CREATE TABLE attendance_import_batches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    source_file TEXT NOT NULL,
                    rows_written INTEGER NOT NULL,
                    imported_at TEXT NOT NULL
                );
                CREATE TABLE attendance_daily_records (source_file TEXT NOT NULL);
                CREATE TABLE attendance_employees (source_file TEXT NOT NULL);
                CREATE TABLE attendance_departments (source_file TEXT NOT NULL);
                CREATE TABLE products (source_file TEXT NOT NULL);
                CREATE TABLE customers (source_file TEXT NOT NULL);
                """
            )
            cursor = conn.execute(
                """
                INSERT INTO attendance_import_batches
                    (source_file, rows_written, imported_at)
                VALUES (?, 2, '2026-07-27T10:00:00')
                """,
                (source_file_key,),
            )
            for table in (
                "attendance_daily_records",
                "attendance_employees",
                "attendance_departments",
                "products",
                "customers",
            ):
                conn.execute(f"INSERT INTO {table} (source_file) VALUES (?)", (source_file_key,))
                conn.execute(f"INSERT INTO {table} (source_file) VALUES ('unrelated-source')")
            conn.commit()
            batch_id = int(cursor.lastrowid)
        return {
            "source_file": source_file_key,
            "db_path": str(db_path),
            "rows_written": 2,
            "batch_id": batch_id,
        }

    monkeypatch.setattr(
        "app.application.attendance_import_app_service.import_attendance_workbook",
        fake_import,
    )
    execution_context = {
        "upload_path": str(source_path),
        "file_sha256": "attendance-hash",
        "row_count": 2,
    }
    result = adapter.execute_batch([{}, {}], execution_context)
    assert result["executed"] == 2

    duplicate = adapter.preview(
        None,
        {},
        allowed_update_fields=set(),
        context={
            "upload_path": str(source_path),
            "file_sha256": "attendance-hash",
            "_preview_cache": {},
        },
    )
    assert duplicate.action == "skip"
    assert duplicate.reason == "duplicate_attendance_source"

    deleted = adapter.rollback_batch({}, result["receipt"])
    assert deleted == 5
    with sqlite3.connect(result["receipt"]["db_path"]) as conn:
        for table in (
            "attendance_daily_records",
            "attendance_employees",
            "attendance_departments",
            "products",
            "customers",
        ):
            assert conn.execute(f"SELECT source_file FROM {table}").fetchall() == [
                ("unrelated-source",)
            ]
        assert conn.execute("SELECT COUNT(*) FROM attendance_import_batches").fetchone()[0] == 0

    invalid = adapter.preview(
        None,
        {},
        allowed_update_fields=set(),
        context={"upload_path": str(tmp_path / "attendance.csv"), "file_sha256": "bad"},
    )
    assert invalid.action == "error"
    assert invalid.issues[0]["code"] == "ETL_ATTENDANCE_FILE_INVALID"


def test_export_xlsx_neutralizes_spreadsheet_formulas(tmp_path, monkeypatch):
    monkeypatch.setenv("XCAGI_DATA_DIR", str(tmp_path / "app-data"))
    result = ExportXlsxAdapter().execute_batch(
        [{"name": "=2+2", "value": 4}],
        {
            "run_id": "xlsx-formula-run",
            "row_count": 1,
            "output_headers": ["name", "value"],
        },
    )
    path = tmp_path / "app-data" / "etl" / "exports" / result["receipt"]["file_name"]
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        rows = list(workbook.active.iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows == [("name", "value"), ("'=2+2", 4)]


def test_webhook_retries_server_errors_and_fails_with_stable_code(monkeypatch):
    statuses = iter((500, 200, 503, 503, 503))
    waits: list[int] = []

    class Response:
        def __init__(self, status_code):
            self.status_code = status_code

    class Client:
        def __init__(self, **_kwargs):
            pass

        def __enter__(self):
            return self

        def __exit__(self, *_args):
            return False

        def post(self, *_args, **_kwargs):
            return Response(next(statuses))

    monkeypatch.setattr("httpx.Client", Client)
    monkeypatch.setattr("app.application.etl.targets._assert_safe_webhook_url", lambda _url: None)
    monkeypatch.setattr(
        "app.application.etl.targets.batch.Event.wait",
        lambda _self, seconds: waits.append(seconds),
    )
    adapter = WebhookAdapter()
    context = {
        "run_id": "webhook-retry",
        "row_count": 1,
        "target_config": {
            "endpoint_url": "https://example.com/hook",
            "headers": {},
            "secret_ref": None,
        },
    }
    assert adapter.execute_batch([{"id": 1}], context)["executed"] == 1
    assert waits == [1]

    with pytest.raises(EtlError) as exc_info:
        adapter.execute_batch([{"id": 2}], {**context, "run_id": "webhook-fail"})
    assert exc_info.value.code == "ETL_WEBHOOK_DELIVERY_FAILED"
    assert exc_info.value.status_code == 502
    assert waits == [1, 1, 2]


def test_knowledge_hash_dedup_source_replacement_and_rollback(tmp_path, monkeypatch):
    rag = DatasetRagApplicationService(
        storage_path=tmp_path / "knowledge.json",
        rebuild_workers_enabled=False,
    )
    monkeypatch.setattr(
        "app.application.dataset_rag_app_service.get_dataset_rag_app_service",
        lambda: rag,
    )
    adapter = KnowledgeAdapter()

    def context() -> dict[str, object]:
        return {
            "owner_user_id": 8,
            "file_name": "制度.md",
            "file_sha256": "upload-hash",
            "run_id": "knowledge-run",
            "_preview_cache": {},
        }

    with tenant_scope(22):
        original = {"content": "第一版", "source_key": "员工制度"}
        preview = adapter.preview(None, original, allowed_update_fields=set(), context=context())
        assert preview.action == "new"
        first = adapter.execute_row(
            None,
            original,
            action="new",
            match_ref="",
            allowed_update_fields=set(),
            context=context(),
        )

        duplicate = adapter.preview(None, original, allowed_update_fields=set(), context=context())
        assert duplicate.action == "skip"
        assert duplicate.reason == "duplicate_content_hash"

        changed = {"content": "第二版", "source_key": "员工制度"}
        unconfirmed = adapter.preview(None, changed, allowed_update_fields=set(), context=context())
        assert unconfirmed.action == "skip"
        confirmed = adapter.preview(
            None,
            changed,
            allowed_update_fields={"content"},
            context=context(),
        )
        assert confirmed.action == "update"
        second = adapter.execute_row(
            None,
            changed,
            action="update",
            match_ref=confirmed.match_ref,
            allowed_update_fields={"content"},
            context=context(),
        )
        assert second["match_ref"] != first["match_ref"]

        adapter.rollback_row(
            None,
            match_ref=second["match_ref"],
            before=confirmed.before or {},
            after=second["after"],
            context=context(),
        )
        status = rag.status(
            "office-docking",
            tenant_id="22",
            access_context={
                "actor_id": "8",
                "tenant_id": "22",
                "permissions": ["dataset.read"],
            },
        )
        assert status["document_count"] == 1
        assert status["documents"][0]["document_id"] == first["match_ref"]


def test_webhook_headers_cannot_bypass_credential_store(etl_db):
    service = EtlService()
    with tenant_scope(30):
        db = etl_db()
        with pytest.raises(EtlError) as forbidden:
            service.create_target_config(
                db,
                owner_user_id=3,
                name="危险配置",
                endpoint_url="https://example.com/hook",
                headers={"Authorization": "Bearer plaintext"},
                secret=None,
            )
        assert forbidden.value.code == "ETL_WEBHOOK_SECRET_HEADER_FORBIDDEN"
        db.close()


def test_retention_removes_payloads_but_keeps_run_summary(etl_db):
    service = EtlService()
    with tenant_scope(31):
        db = etl_db()
        upload_data = service.save_upload(
            db,
            owner_user_id=4,
            file_name="old.csv",
            content_type="text/csv",
            stream=BytesIO(b"name\nold\n"),
        )
        upload = db.get(EtlUpload, upload_data["upload_id"])
        assert upload is not None
        upload.expires_at = datetime.now(UTC) - timedelta(days=1)
        run = EtlRun(
            id="old-run",
            tenant_id=31,
            owner_user_id=4,
            upload_id=upload.id,
            target_type="customers",
            status="completed",
            stage="completed",
            file_sha256=upload.sha256,
            reversible=True,
            created_at=datetime.now(UTC) - timedelta(days=91),
        )
        db.add(run)
        db.add(
            EtlRunRow(
                tenant_id=31,
                owner_user_id=4,
                run_id=run.id,
                source_sheet="Sheet1",
                source_row=2,
            )
        )
        db.commit()
        stored_path = upload.storage_path

        result = service.cleanup_retention(db, owner_user_id=4)

        assert result == {"removed_upload_files": 1, "removed_run_rows": 1}
        assert not Path(stored_path).exists()
        assert db.get(EtlUpload, upload.id).storage_path == ""
        retained = db.get(EtlRun, run.id)
        assert retained is not None
        assert retained.reversible is False
        assert retained.rollback_status == "expired"
        db.close()


def test_dynamic_export_preserves_source_columns_and_values(etl_db, monkeypatch, tmp_path):
    service = EtlService()
    monkeypatch.setattr(
        service,
        "_submit_preview",
        lambda run_id, _tenant_id, owner_user_id: service._preview_worker(run_id, owner_user_id),
    )
    monkeypatch.setattr(
        service,
        "_submit_execution",
        lambda run_id, _tenant_id, owner_user_id, valid_rows_only: service._execute_worker(
            run_id, owner_user_id, valid_rows_only
        ),
    )
    with tenant_scope(40):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=5,
            file_name="export.csv",
            content_type="text/csv",
            stream=BytesIO("姓名,金额\n甲,10\n乙,20\n".encode()),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=5,
            upload_id=upload["upload_id"],
            target_type="export_csv",
        )
        assert [item["target"] for item in run["draft"]["field_mappings"]] == ["姓名", "金额"]
        rows = service.get_rows(
            db,
            run_id=run["id"],
            owner_user_id=5,
            page=1,
            page_size=10,
        )
        assert rows["items"][0]["normalized"] == {"姓名": "甲", "金额": "10"}
        completed = service.execute(
            db,
            run_id=run["id"],
            owner_user_id=5,
            confirmed=True,
            valid_rows_only=False,
        )
        assert completed["status"] == "completed"
        exported = service.download_path(db, run_id=run["id"], owner_user_id=5)
        text = exported.read_text(encoding="utf-8-sig")
        assert "姓名,金额" in text
        assert "甲,10" in text
        db.close()


def test_duplicate_customers_inside_one_file_are_skipped(etl_db, monkeypatch):
    service = EtlService()
    monkeypatch.setattr(
        service,
        "_submit_preview",
        lambda run_id, _tenant_id, owner_user_id: service._preview_worker(run_id, owner_user_id),
    )
    monkeypatch.setattr(
        service,
        "_submit_execution",
        lambda run_id, _tenant_id, owner_user_id, valid_rows_only: service._execute_worker(
            run_id, owner_user_id, valid_rows_only
        ),
    )
    with tenant_scope(41):
        db = etl_db()
        upload = service.save_upload(
            db,
            owner_user_id=6,
            file_name="duplicates.csv",
            content_type="text/csv",
            stream=BytesIO("客户名称,电话\n甲公司,138\n甲公司,138\n".encode()),
        )
        db.commit()
        run = service.create_preview(
            db,
            owner_user_id=6,
            upload_id=upload["upload_id"],
            target_type="customers",
        )
        assert run["summary"]["new"] == 1
        assert run["summary"]["skip"] == 1
        completed = service.execute(
            db,
            run_id=run["id"],
            owner_user_id=6,
            confirmed=True,
            valid_rows_only=False,
        )
        assert completed["summary"]["executed"] == 1
        assert db.query(PurchaseUnit).filter(PurchaseUnit.unit_name == "甲公司").count() == 1
        db.close()


def test_knowledge_adapter_rejects_arbitrary_local_document_path(tmp_path):
    adapter = KnowledgeAdapter()
    upload = tmp_path / "safe.docx"
    upload.write_bytes(b"safe")
    with tenant_scope(42):
        decision = adapter.preview(
            None,
            {"document_path": "/etc/passwd"},
            allowed_update_fields=set(),
            context={
                "upload_path": str(upload),
                "file_sha256": "hash",
                "file_name": upload.name,
                "_preview_cache": {},
            },
        )
    assert decision.action == "error"
    assert decision.issues[0]["code"] == "ETL_DOCUMENT_PATH_FORBIDDEN"


def test_rollback_refuses_to_overwrite_post_import_changes(etl_db):
    adapter = ProductAdapter()
    with tenant_scope(43):
        db = etl_db()
        created = adapter.execute_row(
            db,
            {"unit": "甲公司", "model_number": "A1", "name": "底漆", "price": "10"},
            action="new",
            match_ref="",
            allowed_update_fields=set(),
            context={},
        )
        db.commit()
        product = db.get(Product, int(created["match_ref"]))
        product.description = "导入后人工补充"
        db.commit()
        with pytest.raises(EtlError) as conflict:
            adapter.rollback_row(
                db,
                match_ref=created["match_ref"],
                before={},
                after=created["after"],
                context={},
            )
        assert conflict.value.code == "ETL_ROLLBACK_CONCURRENT_CHANGE"
        assert db.get(Product, int(created["match_ref"])) is not None
        db.close()


def test_shipment_rollback_refuses_post_import_status_change(etl_db):
    adapter = ShipmentAdapter()
    with tenant_scope(45):
        db = etl_db()
        data = {
            "purchase_unit": "甲公司",
            "external_order_no": "ROLLBACK-SHIPMENT-1",
            "product_name": "底漆",
            "quantity_tins": "1",
        }
        context = {
            "file_sha256": "rollback-shipment",
            "file_name": "shipment.csv",
            "source_row": 2,
            "run_id": "rollback-shipment-run",
        }
        created = adapter.execute_row(
            db,
            data,
            action="new",
            match_ref="",
            allowed_update_fields=set(),
            context=context,
        )
        db.commit()
        shipment = db.get(ShipmentRecord, int(created["match_ref"]))
        shipment.status = "shipped"
        db.commit()
        with pytest.raises(EtlError) as conflict:
            adapter.rollback_row(
                db,
                match_ref=created["match_ref"],
                before={},
                after=created["after"],
                context=context,
            )
        assert conflict.value.code == "ETL_ROLLBACK_CONCURRENT_CHANGE"
        assert db.get(ShipmentRecord, int(created["match_ref"])) is not None
        db.close()


def test_startup_recovery_marks_inflight_runs_interrupted(etl_db):
    with tenant_scope(44):
        db = etl_db()
        upload = EtlUpload(
            id="startup-upload",
            tenant_id=44,
            owner_user_id=7,
            file_name="run.csv",
            suffix=".csv",
            size_bytes=1,
            sha256="a" * 64,
            storage_path="/tmp/missing",
        )
        run = EtlRun(
            id="startup-run",
            tenant_id=44,
            owner_user_id=7,
            upload_id=upload.id,
            target_type="customers",
            status="executing",
            stage="executing",
            file_sha256=upload.sha256,
        )
        db.add(upload)
        db.commit()
        db.add(run)
        db.commit()
        assert mark_interrupted_runs_on_startup(db.get_bind()) == 1
        db.expire_all()
        recovered = db.get(EtlRun, run.id)
        assert recovered.status == "interrupted"
        assert recovered.error_code == "ETL_EXECUTION_INTERRUPTED"
        db.close()
