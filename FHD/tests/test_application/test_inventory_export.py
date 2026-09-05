"""Inventory export contracts through HTTP and real isolated SQL data."""

from __future__ import annotations

from contextlib import contextmanager
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, event
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.db.base import Base
from app.db.models import InventoryLedger, Product, Warehouse
from app.fastapi_routes import inventory
from app.infrastructure.tenant_scope import tenant_scope


@pytest.fixture
def inventory_http(monkeypatch):
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    Base.metadata.create_all(engine)
    factory = sessionmaker(bind=engine)

    @contextmanager
    def isolated_db():
        with factory() as db:
            yield db

    monkeypatch.setattr("app.services.inventory_service.get_db", isolated_db)
    with factory() as db:
        for tenant_id in (1, 2):
            with tenant_scope(tenant_id):
                warehouse = Warehouse(
                    tenant_id=tenant_id, code=f"WH-{tenant_id}", name=f"仓库{tenant_id}"
                )
                second = Warehouse(
                    tenant_id=tenant_id, code=f"WH-{tenant_id}-SECOND", name=f"分仓{tenant_id}"
                )
                db.add_all([warehouse, second])
                db.flush()
                for index in range(60 if tenant_id == 1 else 1):
                    product = Product(
                        tenant_id=tenant_id,
                        unit="测试客户",
                        name=f"涂料{index}" if tenant_id == 1 else "其他租户保密产品",
                        model_number=f"SKU-{index:03}",
                    )
                    db.add(product)
                    db.flush()
                    db.add(
                        InventoryLedger(
                            tenant_id=tenant_id,
                            product_id=product.id,
                            warehouse_id=warehouse.id if index < 55 else second.id,
                            quantity=Decimal("1.2345"),
                            available_quantity=Decimal("0.7500"),
                            batch_no=f"B-{index}",
                            unit="千克",
                            in_date=date(2026, 9, 5),
                            created_at=datetime(2026, 9, 5, 12),
                        )
                    )
                db.commit()
    app = FastAPI()
    app.include_router(inventory.router)
    with TestClient(app) as client:
        yield client, factory, engine
    engine.dispose()


def workbook_rows(response):
    assert response.status_code == 200, response.text[:300] if response.status_code != 200 else ""
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(BytesIO(response.content), data_only=False)
    return workbook, list(workbook.active.values)


def test_export_fetches_all_pages_in_one_snapshot_and_preserves_quantities(inventory_http):
    client, _factory, engine = inventory_http
    statements = []

    def record_statement(_conn, _cursor, statement, params, _context, _executemany):
        if statement.lstrip().upper().startswith("SELECT"):
            statements.append((statement, params))

    event.listen(engine, "before_cursor_execute", record_statement)
    response = client.get("/api/inventory/export.xlsx")
    event.remove(engine, "before_cursor_execute", record_statement)
    workbook, rows = workbook_rows(response)

    assert len(statements) == 1
    assert "LIMIT" in statements[0][0].upper()
    assert statements[0][1][-2:] == (50_001, 0)
    assert len(rows) == 61
    assert rows[0] == (
        "产品名称",
        "型号",
        "仓库",
        "批次",
        "库存数量",
        "可用数量",
        "单位",
        "入库日期",
    )
    assert rows[1][:7] == ("涂料59", "SKU-059", "分仓1", "B-59", 1.2345, 0.75, "千克")
    assert rows[1][7].date() == date(2026, 9, 5)
    assert workbook.active["E2"].data_type == "n"
    assert response.headers["x-inventory-row-count"] == "60"
    assert "attachment" in response.headers["content-disposition"]
    assert response.headers["cache-control"] == "no-store"
    page1 = client.get("/api/inventory?page=1&per_page=50").json()
    page2 = client.get("/api/inventory?page=2&per_page=50").json()
    assert [row[1] for row in rows[1:]] == [
        row["product_code"] for row in page1["data"] + page2["data"]
    ]


@pytest.mark.parametrize("keyword, count", [("涂料3", 11), ("sku-003", 1), ("  SKU-003  ", 1)])
def test_keyword_matches_name_or_model_for_both_list_and_export(inventory_http, keyword, count):
    client, _factory, _engine = inventory_http
    params = {"keyword": keyword}
    listing = client.get("/api/inventory", params=params).json()
    _workbook, rows = workbook_rows(client.get("/api/inventory/export.xlsx", params=params))
    assert listing["total"] == count
    assert len(rows) - 1 == count
    assert [row[1] for row in rows[1:]] == [item["product_code"] for item in listing["data"]]


def test_export_combines_warehouse_keyword_product_and_batch_filters(inventory_http):
    client, factory, _engine = inventory_http
    with factory() as db:
        warehouse = db.query(Warehouse).filter(Warehouse.code == "WH-1-SECOND").one()
        product = db.query(Product).filter(Product.model_number == "SKU-057").one()
        params = {"warehouse_id": warehouse.id, "keyword": "SKU-0"}
        response = client.get("/api/inventory/export.xlsx", params=params)
        _book, rows = workbook_rows(response)
        assert len(rows) == 6
        params.update(product_id=product.id, batch_no="B-57")
        _book, rows = workbook_rows(client.get("/api/inventory/export.xlsx", params=params))
        assert len(rows) == 2
        assert rows[1][1] == "SKU-057"


def test_keyword_treats_sql_wildcards_literally(inventory_http):
    client, factory, _engine = inventory_http
    with factory() as db:
        product = db.query(Product).filter(Product.model_number == "SKU-000").one()
        product.name = "100%_配方"
        db.commit()
    for keyword in ("%", "_", "%_"):
        listing = client.get("/api/inventory", params={"keyword": keyword}).json()
        assert listing["total"] == 1
        _book, rows = workbook_rows(
            client.get("/api/inventory/export.xlsx", params={"keyword": keyword})
        )
        assert len(rows) == 2
        assert rows[1][0] == "100%_配方"


def test_export_is_tenant_scoped_including_warehouse_filter(inventory_http):
    client, factory, _engine = inventory_http
    with tenant_scope(2):
        with factory() as db:
            other_warehouse_id = db.query(Warehouse).filter(Warehouse.code == "WH-2").one().id
        _book, rows = workbook_rows(client.get("/api/inventory/export.xlsx"))
        assert len(rows) == 2
        assert rows[1][0] == "其他租户保密产品"
    response = client.get("/api/inventory/export.xlsx", params={"warehouse_id": other_warehouse_id})
    assert response.status_code == 404
    assert response.json()["error_code"] == "INVENTORY_EXPORT_EMPTY"


def test_export_denies_missing_tenant_and_cross_tenant_relationships(inventory_http):
    client, factory, _engine = inventory_http
    with tenant_scope(None):
        response = client.get("/api/inventory/export.xlsx")
        assert response.status_code == 404
        assert client.get("/api/inventory").json()["total"] == 0
    with tenant_scope(2), factory() as db:
        other_product_id = db.query(Product).one().id
        other_warehouse_id = db.query(Warehouse).filter(Warehouse.code == "WH-2").one().id
    with factory() as db:
        ledgers = db.query(InventoryLedger).order_by(InventoryLedger.id).limit(2).all()
        ledgers[0].product_id = other_product_id
        ledgers[1].warehouse_id = other_warehouse_id
        db.commit()
    # Exercise both the explicit product JOIN and the eager warehouse relationship.
    _book, rows = workbook_rows(client.get("/api/inventory/export.xlsx"))
    assert len(rows) == 60
    assert all("其他租户保密产品" not in row and "仓库2" not in row for row in rows)
    assert next(row for row in rows if row[1] == "SKU-001")[2] is None


def test_export_neutralizes_formulas_in_all_text_columns(inventory_http):
    client, factory, _engine = inventory_http
    with factory() as db:
        product = db.query(Product).filter(Product.model_number == "SKU-000").one()
        product.name = '=HYPERLINK("https://invalid.example")'
        product.model_number = "+1+2"
        ledger = db.query(InventoryLedger).filter(InventoryLedger.product_id == product.id).one()
        ledger.warehouse.name = "@SUM(1)"
        ledger.batch_no = "\t-1+1"
        ledger.unit = "=1+1"
        product_id = product.id
        db.commit()
    workbook, rows = workbook_rows(
        client.get("/api/inventory/export.xlsx", params={"product_id": product_id})
    )
    for index in (0, 1, 2, 3, 6):
        assert rows[1][index].startswith("'")
        assert workbook.active.cell(2, index + 1).data_type == "s"
    assert all(cell.data_type != "f" for row in workbook.active for cell in row)


@pytest.mark.parametrize("params", [{"keyword": "不存在"}, {"batch_no": "不存在"}])
def test_empty_export_returns_error_without_download(inventory_http, params):
    client, _factory, _engine = inventory_http
    response = client.get("/api/inventory/export.xlsx", params=params)
    assert response.status_code == 404
    assert response.json()["error_code"] == "INVENTORY_EXPORT_EMPTY"
    assert "content-disposition" not in response.headers


def test_export_limit_rejects_instead_of_silently_truncating(inventory_http, monkeypatch):
    client, _factory, _engine = inventory_http
    monkeypatch.setattr("app.services.inventory_lookup.INVENTORY_EXPORT_ROW_LIMIT", 2)
    response = client.get("/api/inventory/export.xlsx")
    assert response.status_code == 413
    assert response.json()["error_code"] == "INVENTORY_EXPORT_LIMIT"
    assert "2" in response.json()["message"] and "筛选" in response.json()["message"]
    assert "content-disposition" not in response.headers


def test_export_failure_returns_no_workbook(inventory_http, monkeypatch):
    client, _factory, _engine = inventory_http
    service = MagicMock()
    service.export_inventory.side_effect = RuntimeError("internal database unavailable")
    monkeypatch.setattr(inventory, "_svc", lambda: service)
    response = client.get("/api/inventory/export.xlsx")
    assert response.status_code == 500
    assert response.json()["error_code"] == "INVENTORY_EXPORT_FAILED"
    assert "internal database" not in response.text
    assert "content-disposition" not in response.headers
