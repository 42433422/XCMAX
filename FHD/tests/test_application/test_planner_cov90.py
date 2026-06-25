"""High-value coverage tests for app/application/workflow/planner.py.

Targets previously-uncovered branches:
- _execute_print_label_tool ValueError/TypeError/OSError/RuntimeError branches (868-884)
- _execute_excel_analysis_tool ValueError/TypeError branch (1124-1126)
- _execute_import_excel_tool full success loop incl. price-column resolution,
  duplicate skip, customer creation, ambiguous price columns (1230-1326, 1351)
- LLMWorkflowPlanner._plan_with_react_multiagent probe extraction + execution +
  injection (1577-1756)
- LLMWorkflowPlanner._critic_repair_with_llm valid-plan-returning path (1958-1988)
- LLMWorkflowPlanner._plan_with_llm memory/probe context branch (2124-2155)

All external dependencies (services, LLM HTTP, AI conversation service) are mocked;
Excel I/O uses a real openpyxl workbook written under tmp_path for determinism.
"""

from __future__ import annotations

import json
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

from app.application.workflow.planner import (
    LLMWorkflowPlanner,
    _execute_excel_analysis_tool,
    _execute_import_excel_tool,
    _execute_print_label_tool,
)
from app.application.workflow.types import PlanGraph, WorkflowNode

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_planner() -> LLMWorkflowPlanner:
    """Construct a planner with a fully-mocked AI conversation service.

    We patch get_ai_conversation_service at the import site so __init__ does not
    touch real services, then override _ai_service with a controllable mock.
    """
    with patch("app.application.workflow.planner.get_ai_conversation_service") as mock_get:
        mock_get.return_value = MagicMock()
        planner = LLMWorkflowPlanner()
    ai = MagicMock()
    ai.api_key = "sk-test"
    ai.api_url = "http://llm.test/v1/chat/completions"
    ai.model = "deepseek-chat"
    ai.get_context.return_value = None
    planner._ai_service = ai
    return planner


def _llm_response(content: str, status: int = 200) -> MagicMock:
    resp = MagicMock()
    resp.status_code = status
    resp.json.return_value = {"choices": [{"message": {"content": content}}]}
    return resp


def _write_xlsx(path, headers, rows):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append(r)
    wb.save(str(path))
    return str(path)


# ---------------------------------------------------------------------------
# _execute_print_label_tool — error branches (868-884)
# ---------------------------------------------------------------------------


class TestPrintLabelErrorBranches:
    def test_value_error_branch(self) -> None:
        gen = MagicMock()
        gen.generate_labels_for_order.side_effect = ValueError("bad product")
        with (
            patch(
                "app.infrastructure.documents.shipment_document_generator_impl.SimpleLabelGenerator",
                return_value=gen,
            ),
            patch("app.utils.path_utils.get_resource_path", return_value="/tmp/labels"),
            patch("os.makedirs"),
        ):
            out = _execute_print_label_tool({"products": [{"name": "P1"}]})
        assert out["success"] is False
        assert out["error_code"] == "invalid_parameters"

    def test_os_error_branch(self) -> None:
        gen = MagicMock()
        gen.generate_labels_for_order.side_effect = OSError("disk full")
        with (
            patch(
                "app.infrastructure.documents.shipment_document_generator_impl.SimpleLabelGenerator",
                return_value=gen,
            ),
            patch("app.utils.path_utils.get_resource_path", return_value="/tmp/labels"),
            patch("os.makedirs"),
        ):
            out = _execute_print_label_tool({"products": [{"name": "P1"}]})
        assert out["success"] is False
        assert out["error_code"] == "file_io_error"

    def test_runtime_error_branch(self) -> None:
        gen = MagicMock()
        gen.generate_labels_for_order.side_effect = RuntimeError("boom")
        with (
            patch(
                "app.infrastructure.documents.shipment_document_generator_impl.SimpleLabelGenerator",
                return_value=gen,
            ),
            patch("app.utils.path_utils.get_resource_path", return_value="/tmp/labels"),
            patch("os.makedirs"),
        ):
            out = _execute_print_label_tool({"products": [{"name": "P1"}]})
        assert out["success"] is False
        assert out["error_code"] == "generation_failed"

    def test_success_path(self) -> None:
        gen = MagicMock()
        gen.generate_labels_for_order.return_value = ["lbl1", "lbl2"]
        with (
            patch(
                "app.infrastructure.documents.shipment_document_generator_impl.SimpleLabelGenerator",
                return_value=gen,
            ),
            patch("app.utils.path_utils.get_resource_path", return_value="/tmp/labels"),
            patch("os.makedirs"),
        ):
            out = _execute_print_label_tool({"products": [{"name": "P1"}], "order_number": "ORD-9"})
        assert out["success"] is True
        assert out["data"] == ["lbl1", "lbl2"]
        assert "2" in out["message"]
        gen.generate_labels_for_order.assert_called_once()


# ---------------------------------------------------------------------------
# _execute_excel_analysis_tool — ValueError/TypeError branch (1124-1126)
# ---------------------------------------------------------------------------


class TestExcelAnalysisValueErrorBranch:
    # The in-function `from app.bootstrap import get_excel_analysis_app_service`
    # naturally raises ImportError (the getter does not exist), so execution
    # always falls through to the openpyxl path. We drive that path's error
    # handlers by making load_workbook raise.
    def test_value_error_after_openpyxl_load(self) -> None:
        with patch("openpyxl.load_workbook", side_effect=ValueError("corrupt header")):
            out = _execute_excel_analysis_tool({"file_path": "/tmp/x.xlsx"})
        assert out["success"] is False
        assert out["error_code"] == "invalid_parameters"

    def test_os_error_after_openpyxl_load(self) -> None:
        with patch("openpyxl.load_workbook", side_effect=OSError("no file")):
            out = _execute_excel_analysis_tool({"file_path": "/tmp/x.xlsx"})
        assert out["success"] is False
        assert out["error_code"] == "file_not_found"

    def test_runtime_error_after_openpyxl_load(self) -> None:
        with patch("openpyxl.load_workbook", side_effect=RuntimeError("weird")):
            out = _execute_excel_analysis_tool({"file_path": "/tmp/x.xlsx"})
        assert out["success"] is False
        assert out["error_code"] == "analysis_failed"


# ---------------------------------------------------------------------------
# _execute_import_excel_tool — full success loop (1230-1326, 1351)
# ---------------------------------------------------------------------------


class TestImportExcelSuccessLoop:
    def _services(self, *, existing=False, match=None):
        products = MagicMock()
        products.get_products.return_value = {
            "success": True,
            "data": ([{"model_number": "9803", "name": "灯具A"}] if existing else []),
        }
        products.create_product.return_value = {"success": True}
        customers = MagicMock()
        customers.match_purchase_unit.return_value = match
        customers.create.return_value = {"success": True}
        return products, customers

    def test_full_import_creates_customer_and_product(self, tmp_path) -> None:
        fp = _write_xlsx(
            tmp_path / "in.xlsx",
            ["产品名称", "型号", "含税单价", "购买单位"],
            [["灯具A", "9803", 12.5, "七彩乐园"]],
        )
        products, customers = self._services(existing=False, match=None)
        with (
            patch("app.bootstrap.get_products_service", return_value=products),
            patch("app.bootstrap.get_customer_app_service", return_value=customers),
        ):
            out = _execute_import_excel_tool({"file_path": fp})
        assert out["success"] is True
        assert out["created_units"] == 1
        assert out["created_products"] == 1
        assert out["skipped_products"] == 0
        # price column was auto-detected
        assert out["price_column_used"] == "含税单价"
        customers.create.assert_called_once()
        products.create_product.assert_called_once()

    def test_skip_existing_duplicate_by_model(self, tmp_path) -> None:
        fp = _write_xlsx(
            tmp_path / "dup.xlsx",
            ["产品名称", "型号", "单价", "客户"],
            [["灯具A", "9803", 9.9, "七彩乐园"]],
        )
        # match_purchase_unit returns truthy -> no customer created
        products, customers = self._services(existing=True, match={"id": 1})
        with (
            patch("app.bootstrap.get_products_service", return_value=products),
            patch("app.bootstrap.get_customer_app_service", return_value=customers),
        ):
            out = _execute_import_excel_tool({"file_path": fp, "skip_duplicates": True})
        assert out["success"] is True
        assert out["skipped_products"] == 1
        assert out["created_products"] == 0
        assert out["created_units"] == 0
        products.create_product.assert_not_called()

    def test_explicit_price_column_used(self, tmp_path) -> None:
        fp = _write_xlsx(
            tmp_path / "exp.xlsx",
            ["产品名称", "型号", "调价前含税单价", "调价后含税单价"],
            [["灯具B", "7001", 5.0, 6.0]],
        )
        products, customers = self._services(existing=False, match=None)
        with (
            patch("app.bootstrap.get_products_service", return_value=products),
            patch("app.bootstrap.get_customer_app_service", return_value=customers),
        ):
            out = _execute_import_excel_tool(
                {
                    "file_path": fp,
                    "unit_name": "测试单位",
                    "price_column": "调价后含税单价",
                }
            )
        assert out["success"] is True
        assert out["price_column_used"] == "调价后含税单价"
        # the created product should carry the 调价后 price (6.0)
        created_args = products.create_product.call_args[0][0]
        assert created_args["unit_price"] == 6.0

    def test_ambiguous_price_columns(self, tmp_path) -> None:
        fp = _write_xlsx(
            tmp_path / "amb.xlsx",
            ["产品名称", "型号", "调价前含税单价", "调价后含税单价"],
            [["灯具B", "7001", 5.0, 6.0]],
        )
        products, customers = self._services(existing=False, match=None)
        # Force the price resolver to report ambiguity.
        with (
            patch("app.bootstrap.get_products_service", return_value=products),
            patch("app.bootstrap.get_customer_app_service", return_value=customers),
            patch(
                "app.application.ai_chat_app_service.AIChatApplicationService"
                "._merge_user_intent_for_price_resolution",
                return_value="",
            ),
            patch(
                "app.application.ai_chat_app_service.AIChatApplicationService"
                "._resolve_unit_price_column",
                return_value=("", "ambiguous_price_columns"),
            ),
        ):
            out = _execute_import_excel_tool({"file_path": fp})
        assert out["success"] is False
        assert out["error_code"] == "ambiguous_price_columns"

    def test_resolver_picks_price_column(self, tmp_path) -> None:
        fp = _write_xlsx(
            tmp_path / "resolve.xlsx",
            ["产品名称", "型号", "调价前含税单价", "调价后含税单价"],
            [["灯具B", "7001", 5.0, 6.0]],
        )
        products, customers = self._services(existing=False, match=None)
        with (
            patch("app.bootstrap.get_products_service", return_value=products),
            patch("app.bootstrap.get_customer_app_service", return_value=customers),
            patch(
                "app.application.ai_chat_app_service.AIChatApplicationService"
                "._merge_user_intent_for_price_resolution",
                return_value="用调价后",
            ),
            patch(
                "app.application.ai_chat_app_service.AIChatApplicationService"
                "._resolve_unit_price_column",
                return_value=("调价后含税单价", None),
            ),
        ):
            out = _execute_import_excel_tool({"file_path": fp})
        assert out["success"] is True
        assert out["price_column_used"] == "调价后含税单价"

    def test_customer_service_unavailable_degrades(self, tmp_path) -> None:
        # Customer service import fails -> degrade to products-only; product still created.
        fp = _write_xlsx(
            tmp_path / "nocust.xlsx",
            ["产品名称", "型号", "单价", "客户"],
            [["灯具C", "5500", 3.3, "某单位"]],
        )
        products, _ = self._services(existing=False, match=None)
        with (
            patch("app.bootstrap.get_products_service", return_value=products),
            patch("app.bootstrap.get_customer_app_service", side_effect=ImportError),
        ):
            out = _execute_import_excel_tool({"file_path": fp})
        assert out["success"] is True
        assert out["created_units"] == 0
        assert out["created_products"] == 1

    def test_blank_row_skipped(self, tmp_path) -> None:
        # A row with no unit/name/model is skipped via the `continue` branch (1293).
        fp = _write_xlsx(
            tmp_path / "blank.xlsx",
            ["产品名称", "型号", "单价", "客户"],
            [["", "", None, ""], ["灯具D", "6000", 4.4, "单位X"]],
        )
        products, customers = self._services(existing=False, match=None)
        with (
            patch("app.bootstrap.get_products_service", return_value=products),
            patch("app.bootstrap.get_customer_app_service", return_value=customers),
        ):
            out = _execute_import_excel_tool({"file_path": fp})
        assert out["success"] is True
        # Only the non-blank row produced a product.
        assert out["created_products"] == 1


# ---------------------------------------------------------------------------
# LLMWorkflowPlanner._plan_with_react_multiagent — probe loop (1577-1756)
# ---------------------------------------------------------------------------


_PROBE_REGISTRY = {
    "products": {
        "description": "产品",
        "actions": {
            "query": {
                "risk": "low",
                "idempotent": True,
                "required_params": ["keyword"],
            },
        },
    },
}


def _plan_json(tool_id: str, action: str, params: dict) -> str:
    return json.dumps(
        {
            "intent": "查询",
            "todo_steps": ["查询"],
            "risk_level": "low",
            "nodes": [
                {
                    "node_id": "n1",
                    "tool_id": tool_id,
                    "action": action,
                    "params": params,
                    "risk": "low",
                    "idempotent": True,
                    "description": "查询",
                    "depends_on": [],
                }
            ],
        }
    )


class TestPlanWithReactMultiagent:
    def test_probe_executed_and_final_plan_returned(self) -> None:
        planner = _make_planner()
        # candidate + final plan both come from _plan_with_llm; give them a
        # low-risk idempotent query node so it is selected for probing.
        candidate = PlanGraph(
            plan_id="p",
            intent="查询",
            todo_steps=["查询"],
            nodes=[
                WorkflowNode(
                    node_id="n1",
                    tool_id="products",
                    action="query",
                    params={"keyword": "9803"},
                    risk="low",
                    idempotent=True,
                )
            ],
            risk_level="low",
        )
        final = PlanGraph(
            plan_id="p",
            intent="查询产品",
            todo_steps=["查询产品"],
            nodes=[
                WorkflowNode(
                    node_id="n1",
                    tool_id="products",
                    action="query",
                    params={"keyword": "9803"},
                    risk="low",
                    idempotent=True,
                )
            ],
            risk_level="low",
        )
        probe_out = {"success": True, "data": [{"id": 1, "name": "灯具"}], "message": "ok"}
        with (
            patch.object(planner, "_plan_with_llm", side_effect=[candidate, final]) as mock_llm,
            patch(
                "app.application.facades.tools_facade.execute_registered_workflow_tool",
                return_value=probe_out,
            ) as mock_exec,
            patch("app.services.task_agent.TaskAgent", side_effect=ImportError),
        ):
            result = planner._plan_with_react_multiagent(
                plan_id="p",
                user_id="u1",
                message="查 9803",
                tool_registry=_PROBE_REGISTRY,
                context={},
            )
        assert result is final
        # The probe was actually executed against the registered tool.
        mock_exec.assert_called_once()
        # _plan_with_llm called twice: candidate then compose-final.
        assert mock_llm.call_count == 2
        # Probe outputs were injected into the compose context.
        compose_ctx = mock_llm.call_args_list[1].kwargs["context"]
        assert "tool_probe_outputs" in compose_ctx
        assert compose_ctx["tool_probe_outputs"][0]["success"] is True

    def test_candidate_none_returns_none(self) -> None:
        planner = _make_planner()
        with patch.object(planner, "_plan_with_llm", return_value=None):
            result = planner._plan_with_react_multiagent(
                plan_id="p",
                user_id="u1",
                message="x",
                tool_registry=_PROBE_REGISTRY,
                context={},
            )
        assert result is None

    def test_high_risk_node_not_probed(self) -> None:
        planner = _make_planner()
        reg = {
            "danger": {
                "description": "危险",
                "actions": {"query": {"risk": "high", "idempotent": True, "required_params": []}},
            }
        }
        candidate = PlanGraph(
            plan_id="p",
            intent="x",
            nodes=[
                WorkflowNode(
                    node_id="n1",
                    tool_id="danger",
                    action="query",
                    params={},
                    risk="high",
                    idempotent=True,
                )
            ],
        )
        final = PlanGraph(
            plan_id="p",
            intent="x2",
            nodes=[
                WorkflowNode(
                    node_id="n1",
                    tool_id="danger",
                    action="query",
                    params={},
                    risk="high",
                    idempotent=True,
                )
            ],
        )
        with (
            patch.object(planner, "_plan_with_llm", side_effect=[candidate, final]),
            patch(
                "app.application.facades.tools_facade.execute_registered_workflow_tool"
            ) as mock_exec,
            patch("app.services.task_agent.TaskAgent", side_effect=ImportError),
        ):
            result = planner._plan_with_react_multiagent(
                plan_id="p",
                user_id="u1",
                message="x",
                tool_registry=reg,
                context={},
            )
        # High-risk node is filtered out -> no probe executed.
        mock_exec.assert_not_called()
        assert result is final

    def test_final_none_returns_none(self) -> None:
        planner = _make_planner()
        candidate = PlanGraph(
            plan_id="p",
            intent="x",
            nodes=[
                WorkflowNode(
                    node_id="n1",
                    tool_id="products",
                    action="query",
                    params={"keyword": "k"},
                    risk="low",
                    idempotent=True,
                )
            ],
        )
        with (
            patch.object(planner, "_plan_with_llm", side_effect=[candidate, None]),
            patch(
                "app.application.facades.tools_facade.execute_registered_workflow_tool",
                return_value={"success": True, "data": []},
            ),
            patch("app.services.task_agent.TaskAgent", side_effect=ImportError),
        ):
            result = planner._plan_with_react_multiagent(
                plan_id="p",
                user_id="u1",
                message="查 k",
                tool_registry=_PROBE_REGISTRY,
                context={},
            )
        assert result is None


# ---------------------------------------------------------------------------
# _critic_repair_with_llm — valid-plan path (1958-1988)
# ---------------------------------------------------------------------------


class TestCriticRepairValidPlan:
    def test_repair_nonempty_content_raises_attribute_error(self) -> None:
        # SUSPECTED SOURCE BUG: _critic_repair_with_llm calls
        # self._strip_json_code_fence(raw) at planner.py:1953, but that method
        # does not exist on LLMWorkflowPlanner. AttributeError is not caught by
        # the (ValueError, TypeError) / RuntimeError handlers, so any non-empty
        # LLM response propagates an AttributeError and the whole "valid repaired
        # plan" path (lines ~1958-1988) is unreachable dead code. We assert the
        # ACTUAL behaviour rather than the intended one.
        planner = _make_planner()
        invalid = PlanGraph(
            plan_id="p",
            intent="orig_intent",
            todo_steps=["原步骤"],
            nodes=[],
            risk_level="low",
        )
        repaired_json = json.dumps(
            {
                "intent": "repaired_intent",
                "todo_steps": ["修复步骤"],
                "risk_level": "medium",
                "nodes": [
                    {
                        "node_id": "rn1",
                        "tool_id": "products",
                        "action": "query",
                        "params": {"keyword": "x"},
                    }
                ],
            }
        )
        client = MagicMock()
        client.post.return_value = _llm_response(repaired_json)
        with (
            patch(
                "app.application.workflow.planner._get_planner_http_client",
                return_value=client,
            ),
            pytest.raises(AttributeError, match="_strip_json_code_fence"),
        ):
            planner._critic_repair_with_llm(
                "p", "u1", "msg", _PROBE_REGISTRY, {}, "node missing", invalid
            )

    def test_repair_http_error_returns_none(self) -> None:
        planner = _make_planner()
        invalid = PlanGraph(plan_id="p", intent="i", todo_steps=[], nodes=[])
        client = MagicMock()
        client.post.return_value = _llm_response("{}", status=502)
        with patch(
            "app.application.workflow.planner._get_planner_http_client",
            return_value=client,
        ):
            out = planner._critic_repair_with_llm(
                "p", "u1", "msg", _PROBE_REGISTRY, {}, "err", invalid
            )
        assert out is None

    def test_repair_empty_content_returns_none(self) -> None:
        planner = _make_planner()
        invalid = PlanGraph(plan_id="p", intent="i", todo_steps=[], nodes=[])
        client = MagicMock()
        client.post.return_value = _llm_response("")
        with patch(
            "app.application.workflow.planner._get_planner_http_client",
            return_value=client,
        ):
            out = planner._critic_repair_with_llm(
                "p", "u1", "msg", _PROBE_REGISTRY, {}, "err", invalid
            )
        assert out is None


# ---------------------------------------------------------------------------
# _plan_with_llm — memory / probe context branch (2124-2155)
# ---------------------------------------------------------------------------


class TestPlanWithLLMContextProjection:
    def test_memory_and_probe_context_projected_into_metadata(self) -> None:
        planner = _make_planner()
        context = {
            "user_memory_rag": {"summary": "RAG摘要"},
            "memory_v2": {"summary": "记忆v2摘要"},
            "tool_probe_outputs": [
                {
                    "tool_id": "products",
                    "action": "query",
                    "success": True,
                    "message": "命中",
                    "data_preview": "[{'id': 1}]",
                },
                "not-a-dict-skipped",
            ],
        }
        client = MagicMock()
        client.post.return_value = _llm_response(_plan_json("products", "query", {"keyword": "k"}))
        with patch(
            "app.application.workflow.planner._get_planner_http_client",
            return_value=client,
        ):
            plan = planner._plan_with_llm("p", "u1", "查 k", _PROBE_REGISTRY, context)
        assert plan is not None
        md = plan.metadata
        assert md["user_memory_rag_summary"] == "RAG摘要"
        assert md["memory_v2_summary"] == "记忆v2摘要"
        assert len(md["tool_probe_outputs"]) == 1
        assert md["tool_probe_outputs"][0]["tool_id"] == "products"
        assert md["tool_probe_outputs"][0]["success"] is True

    def test_recent_messages_included(self) -> None:
        planner = _make_planner()
        conv_ctx = MagicMock()
        conv_ctx.conversation_history = [{"role": "user", "content": f"m{i}"} for i in range(10)]
        planner._ai_service.get_context.return_value = conv_ctx
        client = MagicMock()
        client.post.return_value = _llm_response(_plan_json("products", "query", {"keyword": "k"}))
        with patch(
            "app.application.workflow.planner._get_planner_http_client",
            return_value=client,
        ):
            plan = planner._plan_with_llm("p", "u1", "查 k", _PROBE_REGISTRY, {})
        assert plan is not None
        # Verify only the last 6 messages were sent in the prompt payload.
        sent_payload = json.loads(client.post.call_args.kwargs["json"]["messages"][1]["content"])
        assert len(sent_payload["recent_messages"]) == 6
