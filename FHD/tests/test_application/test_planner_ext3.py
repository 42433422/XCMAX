"""Tests for app.application.workflow.planner — uncovered branches (ext3).

Focus: _execute_excel_schema_tool openpyxl fallback, _execute_excel_analysis_tool
openpyxl fallback, _execute_import_excel_tool branches, _execute_wechat_preview_tool,
LLMWorkflowPlanner._validate_required_params, _fallback_plan, _get_planner_http_client,
and _WORKFLOW_TOOL_HANDLERS mapping.
"""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, Mock, patch

import pytest

from app.application.workflow.planner import (
    LLMWorkflowPlanner,
    _WORKFLOW_TOOL_HANDLERS,
    _filter_tool_registry_for_profile,
    execute_tool,
    get_tool_registry,
)
from app.application.workflow.types import PlanGraph, WorkflowNode


# ========================= _execute_excel_schema_tool - openpyxl fallback ===


class TestExecuteExcelSchemaToolOpenpyxlFallback:
    def test_service_import_falls_back_to_openpyxl(self, tmp_path):
        from app.application.workflow.planner import _execute_excel_schema_tool

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["产品名称", "单价", "数量"])
        ws.append(["涂料A", 100, 10])
        xlsx_path = str(tmp_path / "schema_test.xlsx")
        wb.save(xlsx_path)

        # get_excel_analysis_app_service does not exist in app.bootstrap,
        # so ImportError is raised naturally, falling back to openpyxl
        result = _execute_excel_schema_tool({"file_path": xlsx_path})
        assert result["success"] is True
        assert len(result["fields"]) == 3
        assert result["row_count"] == 1

    def test_service_runtime_error_falls_back(self, tmp_path):
        from app.application.workflow.planner import _execute_excel_schema_tool

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["名称", "价格"])
        xlsx_path = str(tmp_path / "schema_rt.xlsx")
        wb.save(xlsx_path)

        # Inject a mock that raises RuntimeError to test the RECOVERABLE_ERRORS fallback
        with patch(
            "app.bootstrap.get_excel_analysis_app_service",
            side_effect=RuntimeError("svc fail"),
            create=True,
        ):
            result = _execute_excel_schema_tool({"file_path": xlsx_path})
        assert result["success"] is True

    def test_openpyxl_import_error(self):
        from app.application.workflow.planner import _execute_excel_schema_tool

        # ImportError from app.bootstrap happens naturally; then block openpyxl too
        with patch.dict("sys.modules", {"openpyxl": None}):
            result = _execute_excel_schema_tool({"file_path": "/fake.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "library_unavailable"

    def test_file_not_found_os_error(self):
        from app.application.workflow.planner import _execute_excel_schema_tool

        # ImportError from app.bootstrap happens naturally; then openpyxl can't find file
        result = _execute_excel_schema_tool({"file_path": "/nonexistent_file.xlsx"})
        assert result["success"] is False
        assert result["error_code"] in ("file_not_found", "invalid_parameters")

    def test_value_error(self):
        from app.application.workflow.planner import _execute_excel_schema_tool

        with patch("openpyxl.load_workbook", side_effect=ValueError("bad params")):
            result = _execute_excel_schema_tool({"file_path": "/test.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "invalid_parameters"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_excel_schema_tool

        with patch("openpyxl.load_workbook", side_effect=RuntimeError("runtime")):
            result = _execute_excel_schema_tool({"file_path": "/test.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "analysis_failed"

    def test_missing_file_path(self):
        from app.application.workflow.planner import _execute_excel_schema_tool

        result = _execute_excel_schema_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_file_path"


# ========================= _execute_excel_analysis_tool - openpyxl fallback


class TestExecuteExcelAnalysisToolOpenpyxlFallback:
    def test_service_import_falls_back_to_openpyxl(self, tmp_path):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["名称", "价格"])
        ws.append(["涂料A", 100])
        ws.append(["涂料B", 200])
        xlsx_path = str(tmp_path / "analysis_test.xlsx")
        wb.save(xlsx_path)

        # ImportError from app.bootstrap happens naturally, falling back to openpyxl
        result = _execute_excel_analysis_tool({"file_path": xlsx_path})
        assert result["success"] is True
        assert len(result["headers"]) == 2
        assert result["total_rows"] == 2

    def test_with_columns_filter(self, tmp_path):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["名称", "价格", "备注"])
        ws.append(["涂料A", 100, "ok"])
        xlsx_path = str(tmp_path / "cols_test.xlsx")
        wb.save(xlsx_path)

        result = _execute_excel_analysis_tool(
            {
                "file_path": xlsx_path,
                "columns": ["名称", "价格"],
            }
        )
        assert result["success"] is True

    def test_service_runtime_error_falls_back(self, tmp_path):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["名称"])
        xlsx_path = str(tmp_path / "rt_test.xlsx")
        wb.save(xlsx_path)

        with patch(
            "app.bootstrap.get_excel_analysis_app_service",
            side_effect=RuntimeError("svc fail"),
            create=True,
        ):
            result = _execute_excel_analysis_tool({"file_path": xlsx_path})
        assert result["success"] is True

    def test_missing_file_path(self):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        result = _execute_excel_analysis_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_file_path"

    def test_openpyxl_import_error(self):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        with patch.dict("sys.modules", {"openpyxl": None}):
            result = _execute_excel_analysis_tool({"file_path": "/fake.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "library_unavailable"

    def test_os_error(self):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        with patch("openpyxl.load_workbook", side_effect=OSError("file not found")):
            result = _execute_excel_analysis_tool({"file_path": "/test.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "file_not_found"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        with patch("openpyxl.load_workbook", side_effect=RuntimeError("runtime")):
            result = _execute_excel_analysis_tool({"file_path": "/test.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "analysis_failed"


# ========================= _execute_import_excel_tool - branches ============


class TestExecuteImportExcelToolBranches:
    def test_missing_file_path(self):
        from app.application.workflow.planner import _execute_import_excel_tool

        result = _execute_import_excel_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_file_path"

    def test_products_service_import_error(self):
        from app.application.workflow.planner import _execute_import_excel_tool

        with patch(
            "app.bootstrap.get_products_service",
            side_effect=ImportError("no svc"),
        ):
            result = _execute_import_excel_tool({"file_path": "/test.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"

    def test_products_service_runtime_error(self):
        from app.application.workflow.planner import _execute_import_excel_tool

        with patch(
            "app.bootstrap.get_products_service",
            side_effect=RuntimeError("init fail"),
        ):
            result = _execute_import_excel_tool({"file_path": "/test.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "service_init_failed"

    def test_customer_service_import_error_degrades(self, tmp_path):
        from app.application.workflow.planner import _execute_import_excel_tool

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["产品名称", "型号", "单价"])
        ws.append(["涂料A", "5003A", 100])
        xlsx_path = str(tmp_path / "import_test.xlsx")
        wb.save(xlsx_path)

        mock_products_svc = Mock()
        mock_products_svc.get_products.return_value = {"success": True, "data": []}
        mock_products_svc.create_product.return_value = {"success": True}

        with (
            patch(
                "app.bootstrap.get_products_service",
                return_value=mock_products_svc,
            ),
            patch(
                "app.bootstrap.get_customer_app_service",
                side_effect=ImportError("no svc"),
            ),
        ):
            result = _execute_import_excel_tool({"file_path": xlsx_path})
        assert result["success"] is True

    def test_customer_service_runtime_error_degrades(self, tmp_path):
        from app.application.workflow.planner import _execute_import_excel_tool

        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["产品名称"])
        ws.append(["涂料A"])
        xlsx_path = str(tmp_path / "import_rt.xlsx")
        wb.save(xlsx_path)

        mock_products_svc = Mock()
        mock_products_svc.get_products.return_value = {"success": True, "data": []}
        mock_products_svc.create_product.return_value = {"success": True}

        with (
            patch(
                "app.bootstrap.get_products_service",
                return_value=mock_products_svc,
            ),
            patch(
                "app.bootstrap.get_customer_app_service",
                side_effect=RuntimeError("init fail"),
            ),
        ):
            result = _execute_import_excel_tool({"file_path": xlsx_path})
        assert result["success"] is True


# ========================= _execute_wechat_preview_tool - extended ==========


class TestExecuteWechatPreviewToolExtended:
    def test_with_keyword(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        mock_svc = Mock()
        mock_svc.get_contacts.return_value = [{"name": "张三"}]
        with patch(
            "app.bootstrap.get_wechat_contact_app_service",
            return_value=mock_svc,
        ):
            result = _execute_wechat_preview_tool({"keyword": "张"})
        assert result["success"] is True
        assert len(result["data"]) == 1

    def test_no_contacts_found(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        mock_svc = Mock()
        mock_svc.get_contacts.return_value = []
        with patch(
            "app.bootstrap.get_wechat_contact_app_service",
            return_value=mock_svc,
        ):
            result = _execute_wechat_preview_tool({"keyword": "不存在"})
        assert result["success"] is True
        assert "未找到" in result["message"]

    def test_import_error(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        with patch(
            "app.bootstrap.get_wechat_contact_app_service",
            side_effect=ImportError("no svc"),
        ):
            result = _execute_wechat_preview_tool({"keyword": "test"})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"

    def test_value_error(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        with patch(
            "app.bootstrap.get_wechat_contact_app_service",
            side_effect=ValueError("bad params"),
        ):
            result = _execute_wechat_preview_tool({"keyword": "test"})
        assert result["success"] is False
        assert result["error_code"] == "invalid_parameters"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        with patch(
            "app.bootstrap.get_wechat_contact_app_service",
            side_effect=RuntimeError("fail"),
        ):
            result = _execute_wechat_preview_tool({"keyword": "test"})
        assert result["success"] is False
        assert result["error_code"] == "query_failed"


# ========================= LLMWorkflowPlanner._validate_required_params ====


class TestLLMWorkflowPlannerValidateRequiredParams:
    def test_missing_required_params(self):
        with patch("app.application.workflow.planner.get_ai_conversation_service"):
            planner = LLMWorkflowPlanner()
        # _validate_required_params returns str|None: error string or None if valid
        plan = PlanGraph(
            plan_id="p1",
            intent="test",
            nodes=[
                WorkflowNode(
                    node_id="n1",
                    tool_id="customers",
                    action="ensure_exists",
                    params={"customer_name": "公司A"},
                ),
            ],
        )
        tool_registry = {
            "customers": {
                "actions": {
                    "ensure_exists": {"required_params": ["customer_name", "unit_name"]},
                },
            },
        }
        result = planner._validate_required_params(plan, tool_registry)
        assert result is not None
        assert "unit_name" in result

    def test_all_params_present(self):
        with patch("app.application.workflow.planner.get_ai_conversation_service"):
            planner = LLMWorkflowPlanner()
        plan = PlanGraph(
            plan_id="p1",
            intent="test",
            nodes=[
                WorkflowNode(
                    node_id="n1",
                    tool_id="customers",
                    action="query",
                    params={"customer_name": "公司A"},
                ),
            ],
        )
        tool_registry = {
            "customers": {
                "actions": {
                    "query": {"required_params": ["customer_name"]},
                },
            },
        }
        result = planner._validate_required_params(plan, tool_registry)
        assert result is None

    def test_no_required_params(self):
        with patch("app.application.workflow.planner.get_ai_conversation_service"):
            planner = LLMWorkflowPlanner()
        plan = PlanGraph(plan_id="p1", intent="test", nodes=[])
        tool_registry = {}
        result = planner._validate_required_params(plan, tool_registry)
        assert result is None


# ========================= LLMWorkflowPlanner._fallback_plan ===============


class TestLLMWorkflowPlannerFallbackPlan:
    def test_fallback_plan_with_known_intent(self):
        with patch("app.application.workflow.planner.get_ai_conversation_service"):
            planner = LLMWorkflowPlanner()
        # _fallback_plan(plan_id, message, tool_registry)
        result = planner._fallback_plan("p1", "添加产品", {"customers": {}})
        assert isinstance(result, PlanGraph)

    def test_fallback_plan_with_unknown_intent(self):
        with patch("app.application.workflow.planner.get_ai_conversation_service"):
            planner = LLMWorkflowPlanner()
        result = planner._fallback_plan("p1", "随便聊聊", {})
        assert isinstance(result, PlanGraph)


# ========================= _WORKFLOW_TOOL_HANDLERS mapping ==================


class TestWorkflowToolHandlersMapping:
    def test_all_handlers_present(self):
        expected_keys = {
            ("price_list", "export"),
            ("products", "query"),
            ("customers", "query"),
            ("customers", "ensure_exists"),
            ("shipment_generate", "generate"),
            ("shipment_records", "query"),
            ("shipments", "query"),
            ("materials", "query"),
            ("print_label", "generate"),
            ("excel_decompose", "decompose"),
            ("template_extract", "extract"),
            ("wechat_send", "preview"),
            ("excel_schema", "analyze"),
            ("excel_analysis", "analyze"),
            ("import_excel", "import"),
        }
        for key in expected_keys:
            assert key in _WORKFLOW_TOOL_HANDLERS, f"Missing handler: {key}"

    def test_handlers_are_callable(self):
        for key, handler in _WORKFLOW_TOOL_HANDLERS.items():
            assert callable(handler), f"Handler for {key} is not callable"


# ========================= get_tool_registry - structure ====================


class TestGetToolRegistryStructure:
    def test_registry_has_all_tools(self):
        reg = get_tool_registry()
        expected_tools = [
            "price_list",
            "products",
            "customers",
            "shipment_generate",
            "shipment_records",
            "materials",
            "print_label",
            "excel_decompose",
            "template_extract",
            "wechat_send",
            "excel_schema",
            "excel_analysis",
            "import_excel",
        ]
        for tool in expected_tools:
            assert tool in reg, f"Missing tool: {tool}"

    def test_each_tool_has_description(self):
        reg = get_tool_registry()
        for tool_id, tool_def in reg.items():
            assert "description" in tool_def, f"Tool {tool_id} missing description"

    def test_each_tool_has_actions(self):
        reg = get_tool_registry()
        for tool_id, tool_def in reg.items():
            assert "actions" in tool_def, f"Tool {tool_id} missing actions"


# ========================= _get_planner_http_client ========================


class TestGetPlannerHttpClient:
    def test_returns_client(self):
        from app.application.workflow.planner import _get_planner_http_client

        # Reset singleton to ensure httpx.Client is actually invoked.
        import app.application.workflow.planner as planner_mod

        old_client = planner_mod._planner_http_client
        planner_mod._planner_http_client = None
        try:
            with patch("httpx.Client", return_value=Mock()) as mock_cls:
                client = _get_planner_http_client()
            mock_cls.assert_called_once()
        finally:
            planner_mod._planner_http_client = old_client

    def test_singleton_reuse(self):
        from app.application.workflow.planner import _get_planner_http_client, _planner_http_client

        # Reset singleton
        import app.application.workflow.planner as planner_mod

        old_client = planner_mod._planner_http_client
        try:
            planner_mod._planner_http_client = None
            with patch("httpx.Client", return_value=Mock()) as mock_cls:
                client1 = _get_planner_http_client()
                client2 = _get_planner_http_client()
            assert mock_cls.call_count == 1
        finally:
            planner_mod._planner_http_client = old_client


# ========================= _filter_tool_registry_for_profile - edge cases ===


class TestFilterToolRegistryForProfileEdgeCases:
    def test_tool_without_availability_field(self):
        reg = {"tool_a": {"actions": {"query": {"risk": "low"}}}}
        result = _filter_tool_registry_for_profile(reg, "normal")
        # Should include tool without availability (defaults to shared)
        assert "tool_a" in result

    def test_action_without_availability(self):
        reg = {
            "tool_a": {
                "availability": "shared",
                "actions": {"query": {"risk": "low"}},
            }
        }
        result = _filter_tool_registry_for_profile(reg, "normal")
        assert "tool_a" in result
        assert "query" in result["tool_a"]["actions"]

    def test_pro_default_profile(self):
        reg = {
            "tool_a": {
                "availability": "pro_only",
                "actions": {"query": {"availability": "pro_only", "risk": "low"}},
            },
        }
        result = _filter_tool_registry_for_profile(reg, "pro_default")
        assert "tool_a" in result
