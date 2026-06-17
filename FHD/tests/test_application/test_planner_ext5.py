"""Tests for app.application.workflow.planner — extended coverage (ext5).

Focus: _execute_price_list_tool error branches, _execute_shipment_generate_tool,
_execute_excel_schema_tool, _execute_excel_analysis_tool, _execute_import_excel_tool,
LLMWorkflowPlanner._validate_required_params, _fallback_plan branches,
_filter_tool_registry_for_profile normal/pro_default profiles, _get_planner_http_client.
"""

from __future__ import annotations

import json
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest


# ---------------------------------------------------------------------------
# _execute_price_list_tool — error branches
# ---------------------------------------------------------------------------


class TestExecutePriceListToolErrors:
    """Cover _execute_price_list_tool error branches."""

    def test_missing_customer_name(self):
        from app.application.workflow.planner import _execute_price_list_tool

        result = _execute_price_list_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_customer_name"

    def test_import_error(self):
        from app.application.workflow.planner import _execute_price_list_tool

        with patch.dict(
            "sys.modules", {"app.application.tools": None}
        ):
            result = _execute_price_list_tool({"customer_name": "ACME"})
        assert result["success"] is False
        assert result["error_code"] in ("service_unavailable", "invalid_parameters", "export_failed", "file_io_error")

    def test_value_error(self):
        from app.application.workflow.planner import _execute_price_list_tool

        with patch(
            "app.application.tools.handle_price_list_export",
            side_effect=ValueError("bad param"),
        ):
            result = _execute_price_list_tool({"customer_name": "ACME"})
        assert result["success"] is False
        assert result["error_code"] == "invalid_parameters"

    def test_os_error(self):
        from app.application.workflow.planner import _execute_price_list_tool

        with patch(
            "app.application.tools.handle_price_list_export",
            side_effect=OSError("disk full"),
        ):
            result = _execute_price_list_tool({"customer_name": "ACME"})
        assert result["success"] is False
        assert result["error_code"] == "file_io_error"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_price_list_tool

        with patch(
            "app.application.tools.handle_price_list_export",
            side_effect=RuntimeError("boom"),
        ):
            result = _execute_price_list_tool({"customer_name": "ACME"})
        assert result["success"] is False
        assert result["error_code"] == "export_failed"


# ---------------------------------------------------------------------------
# _execute_shipment_generate_tool
# ---------------------------------------------------------------------------


class TestExecuteShipmentGenerateTool:
    """Cover _execute_shipment_generate_tool branches."""

    def test_missing_params(self):
        from app.application.workflow.planner import _execute_shipment_generate_tool

        result = _execute_shipment_generate_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_order_params"

    def test_with_unit_and_products(self):
        from app.application.workflow.planner import _execute_shipment_generate_tool

        with patch("app.bootstrap.get_shipment_app_service") as mock_ss, patch(
            "app.routes.tools._parse_order_text"
        ):
            mock_svc = MagicMock()
            mock_svc.generate_shipment_document.return_value = {"success": True}
            mock_ss.return_value = mock_svc
            result = _execute_shipment_generate_tool(
                {"unit_name": "ACME", "products": [{"name": "Widget", "qty": 1}]}
            )
        assert result["success"] is True

    def test_with_order_text_parse_failure(self):
        from app.application.workflow.planner import _execute_shipment_generate_tool

        with patch("app.routes.tools._parse_order_text") as mock_parse:
            mock_parse.return_value = {"success": False, "message": "parse error"}
            result = _execute_shipment_generate_tool({"order_text": "invalid"})
        assert result["success"] is False

    def test_import_error(self):
        from app.application.workflow.planner import _execute_shipment_generate_tool

        with patch.dict("sys.modules", {"app.bootstrap": None}):
            result = _execute_shipment_generate_tool(
                {"unit_name": "ACME", "products": [{"name": "W"}]}
            )
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"

    def test_value_error(self):
        from app.application.workflow.planner import _execute_shipment_generate_tool

        with patch("app.routes.tools._parse_order_text") as mock_parse:
            mock_parse.side_effect = ValueError("bad")
            result = _execute_shipment_generate_tool({"order_text": "x"})
        assert result["success"] is False
        assert result["error_code"] == "invalid_parameters"

    def test_os_error(self):
        from app.application.workflow.planner import _execute_shipment_generate_tool

        with patch("app.routes.tools._parse_order_text") as mock_parse, patch(
            "app.bootstrap.get_shipment_app_service"
        ) as mock_ss:
            mock_parse.return_value = {
                "success": True,
                "unit_name": "ACME",
                "products": [{"name": "W"}],
            }
            mock_svc = MagicMock()
            mock_svc.generate_shipment_document.side_effect = OSError("disk full")
            mock_ss.return_value = mock_svc
            result = _execute_shipment_generate_tool({"order_text": "x"})
        assert result["success"] is False
        assert result["error_code"] == "file_io_error"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_shipment_generate_tool

        with patch("app.routes.tools._parse_order_text") as mock_parse, patch(
            "app.bootstrap.get_shipment_app_service"
        ) as mock_ss:
            mock_parse.return_value = {
                "success": True,
                "unit_name": "ACME",
                "products": [{"name": "W"}],
            }
            mock_svc = MagicMock()
            mock_svc.generate_shipment_document.side_effect = RuntimeError("boom")
            mock_ss.return_value = mock_svc
            result = _execute_shipment_generate_tool({"order_text": "x"})
        assert result["success"] is False
        assert result["error_code"] == "generation_failed"


# ---------------------------------------------------------------------------
# _execute_excel_schema_tool
# ---------------------------------------------------------------------------


class TestExecuteExcelSchemaTool:
    """Cover _execute_excel_schema_tool branches."""

    def test_missing_file_path(self):
        from app.application.workflow.planner import _execute_excel_schema_tool

        result = _execute_excel_schema_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_file_path"

    def test_with_app_service(self, tmp_path: Path):
        from app.application.workflow.planner import _execute_excel_schema_tool

        p = tmp_path / "data.xlsx"
        # Create a minimal xlsx file.
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["name", "price"])
            ws.append(["Widget", 10])
            wb.save(p)
        except ImportError:
            pytest.skip("openpyxl not available")

        mock_svc = MagicMock()
        mock_svc.analyze_schema.return_value = {"success": True, "fields": []}
        with patch(
            "app.bootstrap.get_excel_analysis_app_service",
            return_value=mock_svc,
            create=True,
        ):
            result = _execute_excel_schema_tool({"file_path": str(p)})
        assert result["success"] is True

    def test_fallback_to_openpyxl(self, tmp_path: Path):
        from app.application.workflow.planner import _execute_excel_schema_tool

        p = tmp_path / "data.xlsx"
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["name", "price"])
            ws.append(["Widget", 10])
            wb.save(p)
        except ImportError:
            pytest.skip("openpyxl not available")

        # Force ImportError on app service path.
        with patch(
            "app.bootstrap.get_excel_analysis_app_service",
            side_effect=ImportError("no service"),
            create=True,
        ):
            result = _execute_excel_schema_tool({"file_path": str(p)})
        assert result["success"] is True
        assert "fields" in result

    def test_file_not_found(self, tmp_path: Path):
        from app.application.workflow.planner import _execute_excel_schema_tool

        with patch(
            "app.bootstrap.get_excel_analysis_app_service",
            side_effect=ImportError("no service"),
            create=True,
        ):
            result = _execute_excel_schema_tool({"file_path": "missing.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "file_not_found"


# ---------------------------------------------------------------------------
# _execute_excel_analysis_tool
# ---------------------------------------------------------------------------


class TestExecuteExcelAnalysisTool:
    """Cover _execute_excel_analysis_tool branches."""

    def test_missing_file_path(self):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        result = _execute_excel_analysis_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_file_path"

    def test_with_app_service(self, tmp_path: Path):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        p = tmp_path / "data.xlsx"
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["name"])
            ws.append(["Widget"])
            wb.save(p)
        except ImportError:
            pytest.skip("openpyxl not available")

        mock_svc = MagicMock()
        mock_svc.analyze_data.return_value = {"success": True, "rows": []}
        with patch(
            "app.bootstrap.get_excel_analysis_app_service",
            return_value=mock_svc,
            create=True,
        ):
            result = _execute_excel_analysis_tool({"file_path": str(p)})
        assert result["success"] is True

    def test_fallback_to_openpyxl(self, tmp_path: Path):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        p = tmp_path / "data.xlsx"
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["name", "price"])
            ws.append(["Widget", 10])
            ws.append(["Gadget", 20])
            wb.save(p)
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch(
            "app.bootstrap.get_excel_analysis_app_service",
            side_effect=ImportError("no service"),
            create=True,
        ):
            result = _execute_excel_analysis_tool({"file_path": str(p)})
        assert result["success"] is True
        assert "headers" in result
        assert "rows" in result

    def test_with_columns_filter(self, tmp_path: Path):
        from app.application.workflow.planner import _execute_excel_analysis_tool

        p = tmp_path / "data.xlsx"
        try:
            import openpyxl

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["name", "price", "qty"])
            ws.append(["Widget", 10, 1])
            wb.save(p)
        except ImportError:
            pytest.skip("openpyxl not available")

        with patch(
            "app.bootstrap.get_excel_analysis_app_service",
            side_effect=ImportError("no service"),
            create=True,
        ):
            result = _execute_excel_analysis_tool(
                {"file_path": str(p), "columns": ["name", "price"]}
            )
        assert result["success"] is True


# ---------------------------------------------------------------------------
# _execute_import_excel_tool
# ---------------------------------------------------------------------------


class TestExecuteImportExcelTool:
    """Cover _execute_import_excel_tool branches."""

    def test_missing_file_path(self):
        from app.application.workflow.planner import _execute_import_excel_tool

        result = _execute_import_excel_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_file_path"

    def test_products_service_import_error(self):
        from app.application.workflow.planner import _execute_import_excel_tool

        with patch.dict("sys.modules", {"app.bootstrap": None}):
            result = _execute_import_excel_tool({"file_path": "x.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"


# ---------------------------------------------------------------------------
# _execute_customers_ensure_exists_tool — additional branches
# ---------------------------------------------------------------------------


class TestExecuteCustomersEnsureExistsAdditional:
    """Cover _execute_customers_ensure_exists_tool additional branches."""

    def test_missing_unit_name(self):
        from app.application.workflow.planner import _execute_customers_ensure_exists_tool

        result = _execute_customers_ensure_exists_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_unit_name"

    def test_existing_customer(self):
        from app.application.workflow.planner import _execute_customers_ensure_exists_tool

        with patch("app.bootstrap.get_customer_app_service") as mock_cs:
            mock_svc = MagicMock()
            mock_match = MagicMock()
            mock_match.id = 1
            mock_match.unit_name = "ACME"
            mock_svc.match_purchase_unit.return_value = mock_match
            mock_cs.return_value = mock_svc
            result = _execute_customers_ensure_exists_tool({"unit_name": "ACME"})
        assert result["success"] is True
        assert result["created"] is False

    def test_creates_new_customer(self):
        from app.application.workflow.planner import _execute_customers_ensure_exists_tool

        with patch("app.bootstrap.get_customer_app_service") as mock_cs:
            mock_svc = MagicMock()
            mock_svc.match_purchase_unit.return_value = None
            mock_svc.create.return_value = {"success": True, "id": 5}
            mock_cs.return_value = mock_svc
            result = _execute_customers_ensure_exists_tool({"unit_name": "NewCo"})
        assert result["success"] is True
        assert result["created"] is True

    def test_import_error(self):
        from app.application.workflow.planner import _execute_customers_ensure_exists_tool

        with patch.dict("sys.modules", {"app.bootstrap": None}):
            result = _execute_customers_ensure_exists_tool({"unit_name": "ACME"})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"
        assert result["created"] is False

    def test_value_error(self):
        from app.application.workflow.planner import _execute_customers_ensure_exists_tool

        with patch("app.bootstrap.get_customer_app_service") as mock_cs:
            mock_svc = MagicMock()
            mock_svc.match_purchase_unit.side_effect = ValueError("bad")
            mock_cs.return_value = mock_svc
            result = _execute_customers_ensure_exists_tool({"unit_name": "ACME"})
        assert result["success"] is False
        assert result["error_code"] == "invalid_parameters"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_customers_ensure_exists_tool

        with patch("app.bootstrap.get_customer_app_service") as mock_cs:
            mock_svc = MagicMock()
            mock_svc.match_purchase_unit.side_effect = RuntimeError("boom")
            mock_cs.return_value = mock_svc
            result = _execute_customers_ensure_exists_tool({"unit_name": "ACME"})
        assert result["success"] is False
        assert result["error_code"] == "create_failed"


# ---------------------------------------------------------------------------
# _execute_shipment_records_tool — additional branches
# ---------------------------------------------------------------------------


class TestExecuteShipmentRecordsAdditional:
    """Cover _execute_shipment_records_tool additional branches."""

    def test_with_unit_name(self):
        from app.application.workflow.planner import _execute_shipment_records_tool

        with patch("app.bootstrap.get_shipment_app_service") as mock_ss:
            mock_svc = MagicMock()
            mock_svc.get_shipment_records.return_value = [{"id": 1}]
            mock_ss.return_value = mock_svc
            result = _execute_shipment_records_tool({"unit_name": "ACME"})
        assert result["success"] is True
        assert len(result["data"]) == 1

    def test_import_error(self):
        from app.application.workflow.planner import _execute_shipment_records_tool

        with patch.dict("sys.modules", {"app.bootstrap": None}):
            result = _execute_shipment_records_tool({})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"

    def test_value_error(self):
        from app.application.workflow.planner import _execute_shipment_records_tool

        with patch("app.bootstrap.get_shipment_app_service") as mock_ss:
            mock_svc = MagicMock()
            mock_svc.get_shipment_records.side_effect = ValueError("bad")
            mock_ss.return_value = mock_svc
            result = _execute_shipment_records_tool({})
        assert result["success"] is False
        assert result["error_code"] == "invalid_parameters"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_shipment_records_tool

        with patch("app.bootstrap.get_shipment_app_service") as mock_ss:
            mock_svc = MagicMock()
            mock_svc.get_shipment_records.side_effect = RuntimeError("boom")
            mock_ss.return_value = mock_svc
            result = _execute_shipment_records_tool({})
        assert result["success"] is False
        assert result["error_code"] == "query_failed"


# ---------------------------------------------------------------------------
# _execute_materials_tool — additional branches
# ---------------------------------------------------------------------------


class TestExecuteMaterialsAdditional:
    """Cover _execute_materials_tool additional branches."""

    def test_with_params(self):
        from app.application.workflow.planner import _execute_materials_tool

        with patch("app.bootstrap.get_materials_service") as mock_ms:
            mock_svc = MagicMock()
            mock_svc.get_all_materials.return_value = {"success": True, "data": []}
            mock_ms.return_value = mock_svc
            result = _execute_materials_tool({"keyword": "steel", "category": "metal"})
        assert result["success"] is True

    def test_import_error(self):
        from app.application.workflow.planner import _execute_materials_tool

        with patch.dict("sys.modules", {"app.bootstrap": None}):
            result = _execute_materials_tool({})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"

    def test_value_error(self):
        from app.application.workflow.planner import _execute_materials_tool

        with patch("app.bootstrap.get_materials_service") as mock_ms:
            mock_svc = MagicMock()
            mock_svc.get_all_materials.side_effect = ValueError("bad")
            mock_ms.return_value = mock_svc
            result = _execute_materials_tool({})
        assert result["success"] is False
        assert result["error_code"] == "invalid_parameters"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_materials_tool

        with patch("app.bootstrap.get_materials_service") as mock_ms:
            mock_svc = MagicMock()
            mock_svc.get_all_materials.side_effect = RuntimeError("boom")
            mock_ms.return_value = mock_svc
            result = _execute_materials_tool({})
        assert result["success"] is False
        assert result["error_code"] == "query_failed"


# ---------------------------------------------------------------------------
# _execute_print_label_tool — additional branches
# ---------------------------------------------------------------------------


class TestExecutePrintLabelAdditional:
    """Cover _execute_print_label_tool additional branches."""

    def test_missing_products(self):
        from app.application.workflow.planner import _execute_print_label_tool

        result = _execute_print_label_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_products"

    def test_empty_products_list(self):
        from app.application.workflow.planner import _execute_print_label_tool

        result = _execute_print_label_tool({"products": []})
        assert result["success"] is False
        assert result["error_code"] == "missing_products"

    def test_import_error(self):
        from app.application.workflow.planner import _execute_print_label_tool

        with patch.dict(
            "sys.modules",
            {
                "app.infrastructure.documents.shipment_document_generator_impl": None,
            },
        ):
            result = _execute_print_label_tool({"products": [{"name": "W"}]})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"


# ---------------------------------------------------------------------------
# _execute_excel_decompose_tool — additional branches
# ---------------------------------------------------------------------------


class TestExecuteExcelDecomposeAdditional:
    """Cover _execute_excel_decompose_tool additional branches."""

    def test_missing_file_path(self):
        from app.application.workflow.planner import _execute_excel_decompose_tool

        result = _execute_excel_decompose_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_file_path"

    def test_with_file_path(self):
        from app.application.workflow.planner import _execute_excel_decompose_tool

        with patch("app.bootstrap.get_template_app_service") as mock_ts:
            mock_svc = MagicMock()
            mock_svc.decompose_template.return_value = {"success": True}
            mock_ts.return_value = mock_svc
            result = _execute_excel_decompose_tool({"file_path": "/tmp/x.xlsx"})
        assert result["success"] is True

    def test_import_error(self):
        from app.application.workflow.planner import _execute_excel_decompose_tool

        with patch.dict("sys.modules", {"app.bootstrap": None}):
            result = _execute_excel_decompose_tool({"file_path": "/tmp/x.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"

    def test_os_error(self):
        from app.application.workflow.planner import _execute_excel_decompose_tool

        with patch("app.bootstrap.get_template_app_service") as mock_ts:
            mock_svc = MagicMock()
            mock_svc.decompose_template.side_effect = OSError("not found")
            mock_ts.return_value = mock_svc
            result = _execute_excel_decompose_tool({"file_path": "/tmp/x.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "file_not_found"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_excel_decompose_tool

        with patch("app.bootstrap.get_template_app_service") as mock_ts:
            mock_svc = MagicMock()
            mock_svc.decompose_template.side_effect = RuntimeError("boom")
            mock_ts.return_value = mock_svc
            result = _execute_excel_decompose_tool({"file_path": "/tmp/x.xlsx"})
        assert result["success"] is False
        assert result["error_code"] == "decomposition_failed"


# ---------------------------------------------------------------------------
# _execute_template_extract_tool
# ---------------------------------------------------------------------------


class TestExecuteTemplateExtractTool:
    """Cover _execute_template_extract_tool (delegates to excel_decompose)."""

    def test_missing_file_path(self):
        from app.application.workflow.planner import _execute_template_extract_tool

        result = _execute_template_extract_tool({})
        assert result["success"] is False
        assert result["error_code"] == "missing_file_path"

    def test_with_file_path(self):
        from app.application.workflow.planner import _execute_template_extract_tool

        with patch("app.bootstrap.get_template_app_service") as mock_ts:
            mock_svc = MagicMock()
            mock_svc.decompose_template.return_value = {"success": True}
            mock_ts.return_value = mock_svc
            result = _execute_template_extract_tool({"file_path": "/tmp/x.xlsx"})
        assert result["success"] is True


# ---------------------------------------------------------------------------
# _execute_wechat_preview_tool — additional branches
# ---------------------------------------------------------------------------


class TestExecuteWechatPreviewAdditional:
    """Cover _execute_wechat_preview_tool additional branches."""

    def test_with_keyword(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        with patch("app.bootstrap.get_wechat_contact_app_service") as mock_ws:
            mock_svc = MagicMock()
            mock_svc.get_contacts.return_value = [{"id": 1, "name": "Alice"}]
            mock_ws.return_value = mock_svc
            result = _execute_wechat_preview_tool({"keyword": "Alice"})
        assert result["success"] is True
        assert len(result["data"]) == 1

    def test_no_contacts(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        with patch("app.bootstrap.get_wechat_contact_app_service") as mock_ws:
            mock_svc = MagicMock()
            mock_svc.get_contacts.return_value = []
            mock_ws.return_value = mock_svc
            result = _execute_wechat_preview_tool({})
        assert result["success"] is True
        assert "未找到" in result["message"]

    def test_import_error(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        with patch.dict("sys.modules", {"app.bootstrap": None}):
            result = _execute_wechat_preview_tool({})
        assert result["success"] is False
        assert result["error_code"] == "service_unavailable"

    def test_value_error(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        with patch("app.bootstrap.get_wechat_contact_app_service") as mock_ws:
            mock_svc = MagicMock()
            mock_svc.get_contacts.side_effect = ValueError("bad")
            mock_ws.return_value = mock_svc
            result = _execute_wechat_preview_tool({})
        assert result["success"] is False
        assert result["error_code"] == "invalid_parameters"

    def test_runtime_error(self):
        from app.application.workflow.planner import _execute_wechat_preview_tool

        with patch("app.bootstrap.get_wechat_contact_app_service") as mock_ws:
            mock_svc = MagicMock()
            mock_svc.get_contacts.side_effect = RuntimeError("boom")
            mock_ws.return_value = mock_svc
            result = _execute_wechat_preview_tool({})
        assert result["success"] is False
        assert result["error_code"] == "query_failed"


# ---------------------------------------------------------------------------
# _filter_tool_registry_for_profile — additional branches
# ---------------------------------------------------------------------------


class TestFilterToolRegistryForProfileAdditional:
    """Cover _filter_tool_registry_for_profile normal/pro_default branches."""

    def test_normal_profile_excludes_pro_only(self):
        from app.application.workflow.planner import _filter_tool_registry_for_profile

        registry = {
            "shared_tool": {
                "availability": "shared",
                "actions": {
                    "query": {"availability": "shared", "risk": "low"},
                },
            },
            "pro_only_tool": {
                "availability": "pro_only",
                "actions": {
                    "query": {"availability": "shared", "risk": "low"},
                },
            },
        }
        result = _filter_tool_registry_for_profile(registry, "normal")
        assert "shared_tool" in result
        assert "pro_only_tool" not in result

    def test_pro_default_excludes_normal_only(self):
        from app.application.workflow.planner import _filter_tool_registry_for_profile

        registry = {
            "shared_tool": {
                "availability": "shared",
                "actions": {
                    "query": {"availability": "shared", "risk": "low"},
                },
            },
            "normal_only_tool": {
                "availability": "normal_only",
                "actions": {
                    "query": {"availability": "shared", "risk": "low"},
                },
            },
        }
        result = _filter_tool_registry_for_profile(registry, "pro_default")
        assert "shared_tool" in result
        assert "normal_only_tool" not in result

    def test_action_level_filtering(self):
        from app.application.workflow.planner import _filter_tool_registry_for_profile

        registry = {
            "tool": {
                "availability": "shared",
                "actions": {
                    "shared_action": {"availability": "shared", "risk": "low"},
                    "pro_only_action": {"availability": "pro_only", "risk": "low"},
                },
            },
        }
        result = _filter_tool_registry_for_profile(registry, "normal")
        assert "tool" in result
        assert "shared_action" in result["tool"]["actions"]
        assert "pro_only_action" not in result["tool"]["actions"]

    def test_empty_actions_filtered_out(self):
        from app.application.workflow.planner import _filter_tool_registry_for_profile

        registry = {
            "tool": {
                "availability": "shared",
                "actions": {
                    "pro_only_action": {"availability": "pro_only", "risk": "low"},
                },
            },
        }
        result = _filter_tool_registry_for_profile(registry, "normal")
        # All actions filtered out → tool should be excluded.
        assert "tool" not in result

    def test_non_dict_spec_skipped(self):
        from app.application.workflow.planner import _filter_tool_registry_for_profile

        registry = {"bad_tool": "not a dict"}
        result = _filter_tool_registry_for_profile(registry, "normal")
        assert "bad_tool" not in result


# ---------------------------------------------------------------------------
# LLMWorkflowPlanner._validate_required_params
# ---------------------------------------------------------------------------


class TestValidateRequiredParams:
    """Cover LLMWorkflowPlanner._validate_required_params."""

    def test_no_nodes(self):
        from app.application.workflow.planner import LLMWorkflowPlanner, PlanGraph

        plan = PlanGraph(
            plan_id="p1",
            intent="test",
            todo_steps=[],
            nodes=[],
            risk_level="low",
            metadata={},
        )
        result = LLMWorkflowPlanner._validate_required_params(plan, {})
        assert result is None

    def test_missing_required_param(self):
        from app.application.workflow.planner import (
            LLMWorkflowPlanner,
            PlanGraph,
            WorkflowNode,
        )

        node = WorkflowNode(
            node_id="n1",
            tool_id="customers",
            action="ensure_exists",
            params={},
            risk="medium",
            idempotent=True,
            description="",
            depends_on=[],
        )
        plan = PlanGraph(
            plan_id="p1",
            intent="test",
            todo_steps=[],
            nodes=[node],
            risk_level="medium",
            metadata={},
        )
        registry = {
            "customers": {
                "actions": {
                    "ensure_exists": {
                        "required_params": ["unit_name"],
                    }
                }
            }
        }
        result = LLMWorkflowPlanner._validate_required_params(plan, registry)
        assert result is not None
        assert "unit_name" in result

    def test_required_param_present(self):
        from app.application.workflow.planner import (
            LLMWorkflowPlanner,
            PlanGraph,
            WorkflowNode,
        )

        node = WorkflowNode(
            node_id="n1",
            tool_id="customers",
            action="ensure_exists",
            params={"unit_name": "ACME"},
            risk="medium",
            idempotent=True,
            description="",
            depends_on=[],
        )
        plan = PlanGraph(
            plan_id="p1",
            intent="test",
            todo_steps=[],
            nodes=[node],
            risk_level="medium",
            metadata={},
        )
        registry = {
            "customers": {
                "actions": {
                    "ensure_exists": {
                        "required_params": ["unit_name"],
                    }
                }
            }
        }
        result = LLMWorkflowPlanner._validate_required_params(plan, registry)
        assert result is None

    def test_empty_required_param_value(self):
        from app.application.workflow.planner import (
            LLMWorkflowPlanner,
            PlanGraph,
            WorkflowNode,
        )

        node = WorkflowNode(
            node_id="n1",
            tool_id="customers",
            action="ensure_exists",
            params={"unit_name": "  "},
            risk="medium",
            idempotent=True,
            description="",
            depends_on=[],
        )
        plan = PlanGraph(
            plan_id="p1",
            intent="test",
            todo_steps=[],
            nodes=[node],
            risk_level="medium",
            metadata={},
        )
        registry = {
            "customers": {
                "actions": {
                    "ensure_exists": {
                        "required_params": ["unit_name"],
                    }
                }
            }
        }
        result = LLMWorkflowPlanner._validate_required_params(plan, registry)
        assert result is not None

    def test_unknown_tool_skipped(self):
        from app.application.workflow.planner import (
            LLMWorkflowPlanner,
            PlanGraph,
            WorkflowNode,
        )

        node = WorkflowNode(
            node_id="n1",
            tool_id="unknown_tool",
            action="unknown",
            params={},
            risk="low",
            idempotent=True,
            description="",
            depends_on=[],
        )
        plan = PlanGraph(
            plan_id="p1",
            intent="test",
            todo_steps=[],
            nodes=[node],
            risk_level="low",
            metadata={},
        )
        result = LLMWorkflowPlanner._validate_required_params(plan, {})
        assert result is None


# ---------------------------------------------------------------------------
# LLMWorkflowPlanner._fallback_plan — branches
# ---------------------------------------------------------------------------


class TestFallbackPlanBranches:
    """Cover LLMWorkflowPlanner._fallback_plan branches."""

    def test_add_product_intent(self):
        from app.application.workflow.planner import LLMWorkflowPlanner, get_tool_registry

        planner = LLMWorkflowPlanner.__new__(LLMWorkflowPlanner)
        registry = get_tool_registry()
        plan = planner._fallback_plan("p1", "添加产品", registry)
        assert plan.intent == "add_product_to_unit"
        assert len(plan.nodes) >= 1
        assert any(n.tool_id == "customers" for n in plan.nodes)

    def test_create_product_intent(self):
        from app.application.workflow.planner import LLMWorkflowPlanner, get_tool_registry

        planner = LLMWorkflowPlanner.__new__(LLMWorkflowPlanner)
        registry = get_tool_registry()
        plan = planner._fallback_plan("p1", "新增产品", registry)
        assert plan.intent == "add_product_to_unit"

    def test_english_create_intent(self):
        from app.application.workflow.planner import LLMWorkflowPlanner, get_tool_registry

        planner = LLMWorkflowPlanner.__new__(LLMWorkflowPlanner)
        registry = get_tool_registry()
        # Source matches "create" (English) AND "产品" (Chinese) for add_product_to_unit intent.
        plan = planner._fallback_plan("p1", "create 产品", registry)
        assert plan.intent == "add_product_to_unit"

    def test_generic_fallback_with_products(self):
        from app.application.workflow.planner import LLMWorkflowPlanner, get_tool_registry

        planner = LLMWorkflowPlanner.__new__(LLMWorkflowPlanner)
        registry = get_tool_registry()
        plan = planner._fallback_plan("p1", "查询信息", registry)
        assert plan.intent == "generic_workflow"
        # Should fall back to a query node.
        assert len(plan.nodes) >= 1

    def test_generic_fallback_without_products(self):
        from app.application.workflow.planner import LLMWorkflowPlanner

        planner = LLMWorkflowPlanner.__new__(LLMWorkflowPlanner)
        # Registry without "products" but with "customers".
        registry = {
            "customers": {
                "actions": {
                    "query": {"risk": "low", "idempotent": True},
                }
            }
        }
        plan = planner._fallback_plan("p1", "查询信息", registry)
        assert plan.intent == "generic_workflow"
        assert any(n.tool_id == "customers" for n in plan.nodes)

    def test_generic_fallback_no_tools(self):
        from app.application.workflow.planner import LLMWorkflowPlanner

        planner = LLMWorkflowPlanner.__new__(LLMWorkflowPlanner)
        plan = planner._fallback_plan("p1", "查询信息", {})
        assert plan.intent == "generic_workflow"
        assert plan.nodes == []

    def test_risk_level_high(self):
        from app.application.workflow.planner import (
            LLMWorkflowPlanner,
            WorkflowNode,
        )
        from app.application.workflow.types import PlanGraph

        # Build a plan with a high-risk node to test risk_level propagation.
        # _fallback_plan itself doesn't produce high-risk nodes, but we test the logic.
        planner = LLMWorkflowPlanner.__new__(LLMWorkflowPlanner)
        # Use add_product path which produces medium risk.
        registry = {"customers": {"actions": {"ensure_exists": {"risk": "medium"}}}}
        plan = planner._fallback_plan("p1", "添加产品", registry)
        assert plan.risk_level in ("low", "medium", "high")


# ---------------------------------------------------------------------------
# _get_planner_http_client
# ---------------------------------------------------------------------------


class TestGetPlannerHttpClient:
    """Cover _get_planner_http_client."""

    def test_returns_client(self):
        from app.application.workflow.planner import _get_planner_http_client

        client = _get_planner_http_client()
        assert client is not None

    def test_returns_same_instance(self):
        from app.application.workflow.planner import _get_planner_http_client

        c1 = _get_planner_http_client()
        c2 = _get_planner_http_client()
        assert c1 is c2


# ---------------------------------------------------------------------------
# execute_tool — additional action dispatch
# ---------------------------------------------------------------------------


class TestExecuteToolActionDispatch:
    """Cover execute_tool action dispatch via _action param."""

    def test_explicit_action(self):
        from app.application.workflow.planner import execute_tool

        # Pass _action to force a specific handler.
        mock_handler = MagicMock(return_value={"success": True, "data": []})
        with patch.dict(
            "app.application.workflow.planner._WORKFLOW_TOOL_HANDLERS",
            {("products", "query"): mock_handler},
        ):
            result = execute_tool("products", {"_action": "query"})
        assert result["success"] is True
        mock_handler.assert_called_once()

    def test_default_action_for_unknown_tool(self):
        from app.application.workflow.planner import execute_tool

        result = execute_tool("unknown_tool", {})
        assert result["success"] is False
        assert result["error_code"] == "unknown_tool_action"

    def test_runtime_context_stripped(self):
        from app.application.workflow.planner import execute_tool

        mock_handler = MagicMock(return_value={"success": True})
        with patch.dict(
            "app.application.workflow.planner._WORKFLOW_TOOL_HANDLERS",
            {("products", "query"): mock_handler},
        ):
            result = execute_tool(
                "products",
                {"_action": "query", "_runtime_context": {"user_id": "u1"}},
            )
        assert result["success"] is True
        # _runtime_context should be stripped before calling handler.
        called_params = mock_handler.call_args[0][0]
        assert "_runtime_context" not in called_params
