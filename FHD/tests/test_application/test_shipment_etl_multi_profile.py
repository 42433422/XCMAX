"""通用知识库 + 自定义 YAML；默认无内置送货单模板。"""

from __future__ import annotations

from pathlib import Path

import yaml

from app.application.excel_etl_kb import (
    reset_excel_etl_kb_for_tests,
    sheet_layout_fingerprint,
)
from app.application.shipment_etl_profile import clear_profile_cache, list_profiles
from app.application.shipment_excel_etl_app_service import (
    parse_delivery_notes,
    write_delivery_note_workbook,
)


def test_list_profiles_default_is_universal(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    monkeypatch.delenv("FHD_EXCEL_ETL_ALLOW_BUILTIN", raising=False)
    monkeypatch.delenv("FHD_EXCEL_ETL_PROFILE_DIR", raising=False)
    monkeypatch.delenv("FHD_SHIPMENT_ETL_PROFILE_DIR", raising=False)
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    ids = {p["id"] for p in list_profiles()}
    assert "universal" in ids
    assert "default" not in ids
    assert "generic_table" not in ids


def test_universal_recognizes_generic_table_without_delivery_title(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    from openpyxl import Workbook

    path = tmp_path / "quote.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "报价"
    ws["A1"] = "产品报价单"
    ws["A2"] = "客户：竞分客户甲     联系人：赵     日期：2026年07月25日         单号：Q-9"
    ws["A3"] = "型号"
    ws["B3"] = "名称"
    ws["C3"] = "数量"
    ws["D3"] = "规格"
    ws["E3"] = "单价"
    ws["F3"] = "金额"
    ws["A4"] = "GT-01"
    ws["B4"] = "面漆"
    ws["C4"] = 2
    ws["D4"] = 20
    ws["E4"] = 15
    ws["F4"] = 30
    wb.save(path)
    wb.close()

    out = parse_delivery_notes(path, include_ledger=False)
    assert out["success"] is True
    assert out["note_count"] == 1
    note = out["notes"][0]
    assert note.get("profile_id") == "universal"
    assert note["unit_name"] == "竞分客户甲"
    assert note["items"][0]["model_number"] == "GT-01"
    assert note["items"][0]["quantity_tins"] == 2


def test_knowledge_base_remembers_and_hits(tmp_path, monkeypatch):
    kb_path = tmp_path / "kb.json"
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(kb_path))
    kb = reset_excel_etl_kb_for_tests(kb_path)
    clear_profile_cache()
    from openpyxl import Workbook

    path = tmp_path / "learn.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "S1"
    ws["A1"] = "Document"
    ws["A2"] = "客户：学习客户     联系人：钱     日期：2026年07月25日         单号：L-1"
    ws["A3"] = "型号"
    ws["B3"] = "名称"
    ws["C3"] = "数量"
    ws["D3"] = "单价"
    ws["E3"] = "金额"
    ws["A4"] = "KB-01"
    ws["B4"] = "样品"
    ws["C4"] = 3
    ws["D4"] = 9
    ws["E4"] = 27
    wb.save(path)
    wb.close()

    first = parse_delivery_notes(path, include_ledger=False)
    assert first["note_count"] == 1
    fp = sheet_layout_fingerprint(
        sheet_title="S1",
        header_cells=["型号", "名称", "数量", "单价", "金额"],
    )
    mem = kb.get_template(fp)
    assert mem is not None
    assert mem.columns.get("model_number") == 1
    assert mem.columns.get("product_name") == 2

    # 第二次应走 KB 命中
    second = parse_delivery_notes(path, include_ledger=False)
    assert second["note_count"] == 1
    assist = (second["notes"][0].get("assist") or {})
    assert assist.get("reason") == "knowledge_base_hit"
    assert assist.get("cache_hit") is True


def test_custom_profile_dir_overrides(tmp_path, monkeypatch):
    """把自定义 YAML 丢进 PROFILE_DIR 即可识别新模板。"""
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    data = {
        "id": "custom_pack_slip",
        "kind": "pack_slip",
        "label": "装箱单",
        "target": "preview_only",
        "detect": {
            "primary": {
                "title_patterns": ["装箱单"],
                "title_weight": 60,
                "buyer_token": "收货方",
                "buyer_weight": 20,
                "header_hit_tokens": ["货号", "品名", "箱数", "单价"],
                "header_hit_weight": 8,
                "header_hit_cap": 5,
                "min_score": 50,
                "probe_rows": 8,
            }
        },
        "meta": {
            "buyer_label": "收货方",
            "buyer_pattern": r"收货方[：:\s]*([^\s联系人日期]+)",
            "buyer_split_pattern": r"收货方[：:]",
            "buyer_stop_pattern": r"联系人|日期|单号",
            "contact_pattern": r"联系人[：:\s]*([^\s日期]*)",
            "date_pattern": r"(\d{4}[-/]\d{1,2}[-/]\d{1,2})",
            "order_no_pattern": r"单号[：:\s]*([A-Za-z0-9\-]+)",
            "stop_row_pattern": r"合计",
        },
        "header_detect": {
            "primary": {
                "max_scan_rows": 12,
                "require_groups": [["货号"], ["品名"], ["箱数", "数量"]],
            }
        },
        "columns": {
            "model_number": [{"contains_any": ["货号"]}],
            "product_name": [{"contains_any": ["品名"]}],
            "quantity_tins": [{"contains_any": ["箱数", "数量"]}],
            "unit_price": [{"contains_any": ["单价"]}],
            "amount": [{"contains_any": ["金额"]}],
        },
        "write": {
            "seller_title": "装箱单",
            "header_row": ["货号", "品名", "箱数", "单价", "金额"],
            "item_columns": {
                "model_number": 1,
                "product_name": 2,
                "quantity_tins": 3,
                "unit_price": 4,
                "amount": 5,
            },
            "meta_line_template": "收货方：{unit} 联系人：{contact} 日期：{order_date} 单号：{order_no}",
            "demo_meta_line": "收货方：示例 联系人：测 日期：2026-07-25 单号：P-1",
            "default_sheet_name": "装箱",
            "sheet_name_prefix": "箱",
            "date_format": "%Y-%m-%d",
            "footer_label": "",
            "demo_item": {
                "model_number": "P1",
                "product_name": "样",
                "quantity_tins": 1,
                "unit_price": 1,
                "amount": 1,
            },
        },
    }
    (tmp_path / "custom_pack_slip.yaml").write_text(
        yaml.safe_dump(data, allow_unicode=True), encoding="utf-8"
    )
    monkeypatch.setenv("FHD_EXCEL_ETL_PROFILE_DIR", str(tmp_path))
    clear_profile_cache()
    assert any(p["id"] == "custom_pack_slip" for p in list_profiles())

    from openpyxl import Workbook

    xlsx = tmp_path / "pack.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "装箱单"
    ws["A2"] = "收货方：仓库A     联系人：周     日期：2026-07-25         单号：PK-1"
    ws["A3"] = "货号"
    ws["B3"] = "品名"
    ws["C3"] = "箱数"
    ws["D3"] = "单价"
    ws["E3"] = "金额"
    ws["A4"] = "PK01"
    ws["B4"] = "零件"
    ws["C4"] = 4
    ws["D4"] = 3
    ws["E4"] = 12
    wb.save(xlsx)
    wb.close()

    out = parse_delivery_notes(xlsx, include_ledger=False)
    assert out["note_count"] == 1
    note = out["notes"][0]
    assert note["profile_id"] == "custom_pack_slip"
    assert note["unit_name"] == "仓库A"
    assert note["items"][0]["model_number"] == "PK01"
    assert note["items"][0]["quantity_tins"] == 4


def test_universal_write_roundtrip(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    path = tmp_path / "delivery.xlsx"
    write_delivery_note_workbook(
        [
            {
                "unit_name": "送货优先客户",
                "contact_person": "吴",
                "order_date": "2026年07月25日",
                "order_number": "D-1",
                "items": [
                    {
                        "model_number": "D01",
                        "product_name": "漆",
                        "quantity_tins": 1,
                        "tin_spec": 20,
                        "quantity_kg": 20,
                        "unit_price": 10,
                        "amount": 200,
                    }
                ],
            }
        ],
        path,
    )
    out = parse_delivery_notes(path, include_ledger=False)
    assert out["note_count"] == 1
    assert out["notes"][0].get("profile_id") == "universal"
    assert out["notes"][0]["unit_name"] == "送货优先客户"
