"""Excel ETL Profile：通用知识库 + 可选自定义 YAML（无仓库内置送货单）。"""

from __future__ import annotations

from pathlib import Path

import yaml

from app.application.excel_etl_kb import reset_excel_etl_kb_for_tests
from app.application.shipment_etl_profile import (
    ShipmentEtlProfileError,
    clear_profile_cache,
    get_shipment_etl_profile,
    load_profile_from_path,
)
from app.application.shipment_excel_etl_app_service import parse_delivery_notes


def test_universal_profile_loads_from_kb(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    monkeypatch.delenv("FHD_EXCEL_ETL_ALLOW_BUILTIN", raising=False)
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    prof = get_shipment_etl_profile("universal")
    assert prof.id == "universal"
    assert "model_number" in prof.columns
    assert prof.kind == "universal_document"
    # default 别名也指向通用 KB，不再绑送货单 YAML
    assert get_shipment_etl_profile("default").id == "universal"


def test_custom_profile_dir_switches_column_aliases(tmp_path, monkeypatch):
    """换 YAML 列别名即可解析非默认表头，无需改引擎。"""
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    universal = get_shipment_etl_profile("universal")
    data = {
        "id": "alt_headers",
        "kind": "document",
        "label": "英文表头",
        "target": "preview_only",
        "detect": {
            "primary": {
                "title_patterns": [],
                "title_weight": 0,
                "buyer_token": "购货单位",
                "buyer_weight": 20,
                "header_hit_tokens": ["SKU", "ItemName", "QtyPcs", "Price"],
                "header_hit_weight": 8,
                "header_hit_cap": 5,
                "min_score": 32,
                "probe_rows": 8,
            }
        },
        "meta": {
            "buyer_label": "购货单位",
            "buyer_pattern": r"购货单位[（(]?[^)）]*[)）]?[：:\s]*([^\s]+)",
            "buyer_split_pattern": r"购货单位[（(]?[^)）]*[)）]?[：:]",
            "buyer_stop_pattern": r"联系人|日期|订单编号",
            "contact_pattern": r"联系人[：:\s]*([^\s]+)",
            "date_pattern": r"((?:20)?\d{2}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日|\d{4}[-/]\d{1,2}[-/]\d{1,2})",
            "order_no_pattern": r"订单编号[：:\s]*([A-Za-z0-9\-]+)",
            "stop_row_pattern": r"合计",
        },
        "header_detect": {
            "primary": {
                "max_scan_rows": 12,
                "require_groups": [
                    ["sku", "model"],
                    ["itemname", "name"],
                    ["qtypcs", "qty", "数量"],
                ],
            }
        },
        "columns": {
            "model_number": [{"contains_any": ["sku", "model"]}],
            "product_name": [{"contains_any": ["itemname", "name"]}],
            "quantity_tins": [{"contains_any": ["qtypcs", "qty"]}],
            "tin_spec": [{"contains_any": ["speckg", "spec"]}],
            "quantity_kg": [{"contains_any": ["qtykg"]}],
            "unit_price": [{"contains_any": ["price", "unitprice"]}],
            "amount": [{"contains_any": ["amount", "total"]}],
            "order_number": [{"contains_any": ["orderno", "order"]}],
            "order_date": [{"contains_any": ["date"]}],
        },
        "write": {
            "seller_title": "ALT Factory Delivery",
            "header_row": ["SKU", "ItemName", "QtyPcs", "SpecKg", "QtyKg", "Price", "Amount"],
            "item_columns": {
                "model_number": 1,
                "product_name": 2,
                "quantity_tins": 3,
                "tin_spec": 4,
                "quantity_kg": 5,
                "unit_price": 6,
                "amount": 7,
            },
            "meta_line_template": (
                "购货单位（乙方）：{unit}     联系人：{contact}        "
                "日期：{order_date}         订单编号：{order_no}"
            ),
            "demo_meta_line": "购货单位（乙方）：示例 联系人：测 日期：2026年07月25日 订单编号：A-1",
            "default_sheet_name": "N1",
            "sheet_name_prefix": "N",
            "date_format": "%Y年%m月%d日",
            "footer_label": "",
            "demo_item": {
                "model_number": "A1",
                "product_name": "样",
                "quantity_tins": 1,
                "tin_spec": 1,
                "quantity_kg": 1,
                "unit_price": 1,
                "amount": 1,
            },
        },
    }

    profile_path = tmp_path / "alt_headers.yaml"
    profile_path.write_text(yaml.safe_dump(data, allow_unicode=True), encoding="utf-8")
    monkeypatch.setenv("FHD_SHIPMENT_ETL_PROFILE_DIR", str(tmp_path))
    clear_profile_cache()

    alt = get_shipment_etl_profile("alt_headers")
    assert alt.id == "alt_headers"
    assert alt.write["seller_title"] == "ALT Factory Delivery"

    from openpyxl import Workbook

    xlsx = tmp_path / "alt.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "N1"
    ws["A1"] = "ALT Factory Delivery"
    ws["A2"] = (
        "购货单位（乙方）：切换客户     联系人：李     日期：2026年07月25日         订单编号：ALT-1"
    )
    ws["A3"] = "SKU"
    ws["B3"] = "ItemName"
    ws["C3"] = "QtyPcs"
    ws["D3"] = "SpecKg"
    ws["E3"] = "QtyKg"
    ws["F3"] = "Price"
    ws["G3"] = "Amount"
    ws["A4"] = "ALT-01"
    ws["B4"] = "测试漆"
    ws["C4"] = 2
    ws["D4"] = 25
    ws["E4"] = 50
    ws["F4"] = 10
    ws["G4"] = 500
    wb.save(xlsx)
    wb.close()

    default_parsed = parse_delivery_notes(xlsx, include_ledger=False, profile=universal)
    alt_parsed = parse_delivery_notes(xlsx, include_ledger=False, profile=alt)
    assert alt_parsed["success"] is True
    assert alt_parsed["note_count"] == 1
    note = alt_parsed["notes"][0]
    assert note["unit_name"] == "切换客户"
    assert note["items"][0]["model_number"] == "ALT-01"
    assert note["items"][0]["product_name"] == "测试漆"
    assert note["items"][0]["quantity_tins"] == 2
    assert alt_parsed["profile_id"] == "alt_headers"
    # 通用 profile 对纯英文表头通常解析失败或字段不全；自定义 YAML 必须成功
    assert alt_parsed["note_count"] >= (default_parsed.get("note_count") or 0)


def test_unknown_profile_raises(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    try:
        get_shipment_etl_profile("no_such_profile_xyz")
        raise AssertionError("expected ShipmentEtlProfileError")
    except ShipmentEtlProfileError as exc:
        assert "unknown" in str(exc).lower() or "no_such" in str(exc)


def test_load_example_yaml_when_allow_builtin(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    monkeypatch.setenv("FHD_EXCEL_ETL_ALLOW_BUILTIN", "1")
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    path = Path("resources/config/shipment_etl/examples/default_delivery.yaml")
    prof = load_profile_from_path(path)
    assert prof.id == "default"
    assert prof.delivery_min_score == 60
    from app.application.shipment_etl_profile import load_all_profiles

    ids = {p.id for p in load_all_profiles()}
    assert "universal" in ids
    assert "default" in ids
    assert "generic_table" in ids
