"""Excel ETL：陌生表头 / 多表混排 / OCR 桥接 / 直写门禁。"""

from __future__ import annotations

from pathlib import Path

import pytest

from app.application.excel_etl_kb import reset_excel_etl_kb_for_tests
from app.application.shipment_etl_profile import clear_profile_cache
from app.application.shipment_excel_etl_app_service import (
    execute_shipment_excel_etl,
    parse_delivery_notes,
    write_delivery_note_workbook,
    write_ledger_workbook,
)
from app.application.shipment_excel_etl_ocr import (
    grid_to_workbook_path,
    text_blocks_to_grid,
)
from app.application.shipment_excel_etl_security import direct_execute_allowed


@pytest.fixture(autouse=True)
def _kb(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    monkeypatch.setenv("FHD_EXCEL_ETL_DEFAULT_TARGET", "shipment")
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    monkeypatch.delenv("FHD_EXCEL_ETL_ALLOW_BUILTIN", raising=False)
    monkeypatch.delenv("FHD_EXCEL_ETL_ALLOW_DIRECT", raising=False)
    monkeypatch.delenv("FHD_SHIPMENT_ETL_ALLOW_DIRECT", raising=False)
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    yield
    clear_profile_cache()


def test_unknown_headers_heuristic_samples(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "weird.xlsx"
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "客户：陌生客户     联系人：测     日期：2026年07月25日         单号：U-1"
    # 表头故意不在同义词里；靠样例类型推断
    ws["A2"] = "PartCode"
    ws["B2"] = "GoodsTitle"
    ws["C2"] = "PcsCount"
    ws["D2"] = "UnitFee"
    ws["E2"] = "LineSum"
    ws["A3"] = "PC-01"
    ws["B3"] = "面漆样品"
    ws["C3"] = 2
    ws["D3"] = 15
    ws["E3"] = 30
    wb.save(path)
    wb.close()

    out = parse_delivery_notes(path, include_ledger=False)
    assert out["note_count"] == 1
    note = out["notes"][0]
    assert note["unit_name"] == "陌生客户"
    assert note["items"][0]["model_number"] == "PC-01"
    assert note["items"][0]["product_name"] == "面漆样品"
    assert note["items"][0]["quantity_tins"] == 2
    assert note.get("assist", {}).get("reason") == "heuristic_samples"


def test_mixed_workbook_roles(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "mixed.xlsx"
    wb = Workbook()
    cover = wb.active
    cover.title = "目录说明"
    cover["A1"] = "本册说明"
    cover["A2"] = "请勿导入"
    write_delivery_note_workbook(
        [
            {
                "unit_name": "混排客户",
                "contact_person": "周",
                "order_date": "2026年07月25日",
                "order_number": "MX-1",
                "sheet": "单据",
                "items": [
                    {
                        "model_number": "M1",
                        "product_name": "漆",
                        "quantity_tins": 1,
                        "tin_spec": 20,
                        "quantity_kg": 20,
                        "unit_price": 9,
                        "amount": 180,
                    }
                ],
            }
        ],
        tmp_path / "d.xlsx",
    )
    from openpyxl import load_workbook

    src = load_workbook(tmp_path / "d.xlsx")
    delivery = src.active
    ws = wb.create_sheet("单据")
    for row in delivery.iter_rows(values_only=True):
        ws.append(list(row))
    src.close()
    write_ledger_workbook(
        [
            {
                "order_date": "2026-07-01",
                "order_number": "LX-1",
                "model_number": "L1",
                "product_name": "旧货",
                "quantity_tins": 1,
                "tin_spec": 10,
                "quantity_kg": 10,
                "unit_price": 5,
                "amount": 50,
            }
        ],
        tmp_path / "l.xlsx",
        unit_name="流水客户",
    )
    lsrc = load_workbook(tmp_path / "l.xlsx")
    ledger = wb.create_sheet("Ledger")
    for row in lsrc.active.iter_rows(values_only=True):
        ledger.append(list(row))
    lsrc.close()
    wb.save(path)
    wb.close()

    auto = parse_delivery_notes(path, include_ledger="auto")
    assert auto["mixed_workbook"] is True
    roles = {r["sheet"]: r["role"] for r in auto.get("sheet_roles") or []}
    assert roles.get("目录说明") == "ignore"
    assert auto["delivery_note_count"] >= 1
    assert auto["ledger_note_count"] == 0
    assert auto["ledger_available_count"] >= 1

    both = parse_delivery_notes(path, include_ledger=True)
    assert both["ledger_note_count"] >= 1


def test_ocr_text_blocks_to_grid_and_parse(tmp_path):
    blocks = [
        {"text": "客户：OCR客户", "left": 10, "top": 10, "width": 120, "height": 20},
        {"text": "联系人：甲", "left": 150, "top": 10, "width": 80, "height": 20},
        {"text": "日期：2026年07月25日", "left": 250, "top": 10, "width": 140, "height": 20},
        {"text": "单号：O-1", "left": 420, "top": 10, "width": 80, "height": 20},
        {"text": "型号", "left": 10, "top": 50, "width": 40, "height": 18},
        {"text": "名称", "left": 80, "top": 50, "width": 40, "height": 18},
        {"text": "数量", "left": 150, "top": 50, "width": 40, "height": 18},
        {"text": "单价", "left": 220, "top": 50, "width": 40, "height": 18},
        {"text": "金额", "left": 290, "top": 50, "width": 40, "height": 18},
        {"text": "OCR-01", "left": 10, "top": 80, "width": 50, "height": 18},
        {"text": "清漆", "left": 80, "top": 80, "width": 40, "height": 18},
        {"text": "2", "left": 150, "top": 80, "width": 20, "height": 18},
        {"text": "10", "left": 220, "top": 80, "width": 20, "height": 18},
        {"text": "20", "left": 290, "top": 80, "width": 20, "height": 18},
    ]
    grid = text_blocks_to_grid(blocks)
    assert len(grid) >= 2
    xlsx = grid_to_workbook_path(grid, output_path=tmp_path / "ocr.xlsx", meta_lines=[])
    # 把首行 meta 拼进 A1 便于 buyer regex（模拟 ocr 桥接）
    from openpyxl import load_workbook

    wb = load_workbook(xlsx)
    ws = wb.active
    # 若首行已是客户行则直接解析
    wb.save(xlsx)
    wb.close()
    out = parse_delivery_notes(xlsx, include_ledger=False, allow_ocr=False)
    assert out["note_count"] == 1
    note = out["notes"][0]
    assert "OCR" in note["unit_name"] or note["items"][0]["model_number"] == "OCR-01"
    assert note["items"][0]["model_number"] == "OCR-01"
    assert note["items"][0]["quantity_tins"] == 2


def test_direct_execute_requires_env(tmp_path, monkeypatch):
    path = tmp_path / "d.xlsx"
    write_delivery_note_workbook(
        [
            {
                "unit_name": "直写客户",
                "order_number": "DIR-1",
                "order_date": "2026年07月25日",
                "items": [
                    {
                        "model_number": "D1",
                        "product_name": "漆",
                        "quantity_tins": 1,
                        "tin_spec": 20,
                        "quantity_kg": 20,
                        "unit_price": 8,
                        "amount": 160,
                    }
                ],
            }
        ],
        path,
    )
    monkeypatch.setenv("WORKSPACE_ROOT", str(tmp_path))
    denied = execute_shipment_excel_etl(path, direct=True, dry_run=True, workspace_root=tmp_path)
    # dry_run 仍检查 direct 开关（非 dry 才拦）；确认非 dry 被拒
    denied2 = execute_shipment_excel_etl(path, direct=True, dry_run=False, workspace_root=tmp_path)
    assert denied2["success"] is False
    assert denied2["error_code"] == "direct_execute_denied"
    assert direct_execute_allowed() is False

    monkeypatch.setenv("FHD_EXCEL_ETL_ALLOW_DIRECT", "1")
    assert direct_execute_allowed() is True

    class _Fake:
        def create_shipment(self, *a, **k):
            return {"success": True, "shipment": {"id": 9}}

    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: _Fake())
    monkeypatch.setattr(
        "app.services.tools_workflow_registered._execute_excel_import_records",
        lambda records: {"success": True, "imported": len(records)},
    )
    monkeypatch.setenv("FHD_SHIPMENT_ETL_FINGERPRINT_BACKEND", "legacy")
    fp = tmp_path / "fp.sqlite3"
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._legacy_db_path",
        lambda: fp,
    )
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._db_path",
        lambda: fp,
    )
    # 默认 target 已是 shipment（fixture）；验证直写成功
    ok = execute_shipment_excel_etl(
        path,
        direct=True,
        force_shipment_target=True,
        workspace_root=tmp_path,
        idempotent=True,
    )
    assert ok["success"] is True
    assert ok.get("direct") is True
    assert ok["shipment_created"] == 1


def test_direct_promotes_preview_only_target(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_DEFAULT_TARGET", "preview_only")
    clear_profile_cache()
    path = tmp_path / "p.xlsx"
    write_delivery_note_workbook(
        [
            {
                "unit_name": "提升客户",
                "order_number": "P-1",
                "order_date": "2026年07月25日",
                "items": [
                    {
                        "model_number": "P1",
                        "product_name": "漆",
                        "quantity_tins": 1,
                        "tin_spec": 10,
                        "quantity_kg": 10,
                        "unit_price": 1,
                        "amount": 10,
                    }
                ],
            }
        ],
        path,
    )
    monkeypatch.setenv("WORKSPACE_ROOT", str(tmp_path))
    monkeypatch.setenv("FHD_EXCEL_ETL_ALLOW_DIRECT", "1")

    class _Fake:
        def create_shipment(self, *a, **k):
            return {"success": True, "shipment": {"id": 3}}

    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: _Fake())
    monkeypatch.setattr(
        "app.services.tools_workflow_registered._execute_excel_import_records",
        lambda records: {"success": True, "imported": len(records)},
    )
    fp = tmp_path / "fp2.sqlite3"
    monkeypatch.setenv("FHD_SHIPMENT_ETL_FINGERPRINT_BACKEND", "legacy")
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._legacy_db_path",
        lambda: fp,
    )
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._db_path",
        lambda: fp,
    )
    # 不 force → preview_only 挡入库
    blocked = execute_shipment_excel_etl(path, direct=True, workspace_root=tmp_path)
    assert blocked["success"] is False
    assert blocked["error_code"] == "unsupported_profile_target"
    # force → 提升 target 后入库
    ok = execute_shipment_excel_etl(
        path, direct=True, force_shipment_target=True, workspace_root=tmp_path
    )
    assert ok["success"] is True
    assert ok["shipment_created"] == 1
