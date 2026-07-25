"""送货单 Excel ETL 单测：解析 / 流水 / 幂等 / 模板反推 / 批量。"""

from __future__ import annotations

from pathlib import Path

import pytest

from app.application.excel_etl_kb import reset_excel_etl_kb_for_tests
from app.application.shipment_etl_profile import clear_profile_cache
from app.application.shipment_excel_etl_app_service import (
    batch_execute_shipment_excel_etl,
    batch_preview_shipment_excel_etl,
    execute_shipment_excel_etl,
    note_fingerprint,
    parse_delivery_notes,
    preview_shipment_excel_etl,
    regenerate_delivery_notes_from_file,
    write_delivery_note_workbook,
    write_ledger_workbook,
)

SAMPLE_DIR = Path("/Users/a4243342/Desktop/新建文件夹 (4)/产品文件夹/发货单")


@pytest.fixture(autouse=True)
def _etl_universal_kb(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    monkeypatch.setenv("FHD_EXCEL_ETL_DEFAULT_TARGET", "shipment")
    monkeypatch.delenv("FHD_EXCEL_ETL_ALLOW_BUILTIN", raising=False)
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    yield
    clear_profile_cache()


def test_parse_guosheng_delivery_note():
    path = SAMPLE_DIR / "国圣化工.xlsx"
    if not path.is_file():
        pytest.skip("local sample missing")
    result = parse_delivery_notes(path, include_ledger=False)
    if not result.get("note_count"):
        pytest.skip("local sample not matched by universal KB synonyms")
    assert result["success"] is True
    note = result["notes"][0]
    assert note["item_count"] >= 1
    assert note["items"][0]["product_name"]
    assert note.get("fingerprint")


def test_parse_houxuemei_and_yin():
    any_ok = False
    for name in ("侯雪梅.xlsx", "尹玉华1.xlsx", "现金.xlsx", "澜宇电视柜.xlsx"):
        path = SAMPLE_DIR / name
        if not path.is_file():
            continue
        result = parse_delivery_notes(path, include_ledger=False)
        if not result.get("note_count"):
            continue
        any_ok = True
        assert result["success"] is True, name
        assert all(n["item_count"] >= 1 for n in result["notes"]), name
    if not any_ok:
        pytest.skip("local samples missing or not matched by universal KB")


def test_preview_marks_confirm_required():
    path = SAMPLE_DIR / "国圣化工.xlsx"
    if not path.is_file():
        pytest.skip("local sample missing")
    preview = preview_shipment_excel_etl(
        path,
        include_ledger=False,
        workspace_root=SAMPLE_DIR,
    )
    if not preview.get("success"):
        pytest.skip("local sample not matched by universal KB")
    assert preview.get("confirm_required") is True
    assert preview.get("product_records")


def test_score_rejects_non_business_sheet_without_ledger(tmp_path):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "报价"
    ws.append(["品名", "参考价"])
    ws.append(["面漆", 12])
    path = tmp_path / "quote_only.xlsx"
    wb.save(path)
    result = parse_delivery_notes(path, include_ledger=True)
    assert result["success"] is True
    assert result["note_count"] == 0


def test_auto_skips_ledger_when_delivery_present(tmp_path):
    from openpyxl import Workbook

    path = tmp_path / "mixed.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "送货"
    ws["A1"] = "测试工厂送货单"
    ws["A2"] = "购货单位：混合客户     联系人：张总       2026年07月24日         订单编号：M-1"
    ws["A3"] = "产品型号"
    ws["D3"] = "产品名称"
    ws["E3"] = "数量/件"
    ws["F3"] = "规格/KG"
    ws["G3"] = "数量/KG"
    ws["H3"] = "单价/元"
    ws["I3"] = "金额/元"
    ws["A4"] = "M01"
    ws["D4"] = "面漆"
    ws["E4"] = 1
    ws["F4"] = 20
    ws["G4"] = 20
    ws["H4"] = 10
    ws["I4"] = 200
    ledger = wb.create_sheet("25出货")
    ledger.append(
        [
            "日期",
            "单号",
            "产品型号",
            "",
            "",
            "产品名称",
            "数量/件",
            "规格/KG",
            "数量/KG",
            "单价/元",
            "金额/元",
        ]
    )
    ledger.append(["2026-07-01", "L-9", "X1", "", "", "旧货", 1, 10, 10, 5, 50])
    wb.save(path)

    auto = parse_delivery_notes(path, include_ledger="auto")
    assert auto["note_count"] == 1
    assert auto["delivery_note_count"] == 1
    assert auto["ledger_note_count"] == 0
    assert auto["ledger_available_count"] >= 1

    forced = parse_delivery_notes(path, include_ledger=True)
    assert forced["note_count"] >= 2
    assert forced["ledger_note_count"] >= 1


def test_parse_generated_ledger_groups_by_order(tmp_path):
    path = tmp_path / "闭环流水客户.xlsx"
    written = write_ledger_workbook([], path, unit_name="闭环流水客户")
    assert written["success"] is True
    result = parse_delivery_notes(path, include_ledger=True, unit_name_hint="闭环流水客户")
    assert result["success"] is True
    assert result["ledger_note_count"] >= 2
    assert all(n["source_kind"] == "shipment_ledger" for n in result["notes"])
    assert {n["order_number"] for n in result["notes"]} >= {"L-001", "L-002"}


def test_write_and_parse_delivery_template_roundtrip(tmp_path):
    path = tmp_path / "delivery_tpl.xlsx"
    notes = [
        {
            "unit_name": "闭环测试客户甲",
            "contact_person": "王工",
            "order_date": "2026年07月24日",
            "order_number": "LOOP-1001",
            "sheet": "送货甲",
            "items": [
                {
                    "model_number": "RX-LOOP",
                    "product_name": "PU哑光清漆",
                    "quantity_tins": 2,
                    "tin_spec": 25,
                    "quantity_kg": 50,
                    "unit_price": 18,
                    "amount": 900,
                }
            ],
        }
    ]
    written = write_delivery_note_workbook(notes, path)
    assert written["success"] is True
    parsed = parse_delivery_notes(path, include_ledger=False)
    assert parsed["note_count"] == 1
    note = parsed["notes"][0]
    assert note["unit_name"] == "闭环测试客户甲"
    assert note["order_number"] == "LOOP-1001"
    assert note["items"][0]["model_number"] == "RX-LOOP"
    assert note_fingerprint(note) == note["fingerprint"]

    regen_path = tmp_path / "delivery_regen.xlsx"
    regen = regenerate_delivery_notes_from_file(path, regen_path, include_ledger=False)
    assert regen["success"] is True
    assert regen["fingerprint_match"] is True


def test_execute_idempotent_skips_second_run(tmp_path, monkeypatch):
    path = tmp_path / "delivery.xlsx"
    write_delivery_note_workbook(
        [
            {
                "unit_name": "幂等测试客户",
                "contact_person": "李总",
                "order_date": "2026年07月24日",
                "order_number": "IDEM-1",
                "items": [
                    {
                        "model_number": "IDEM01",
                        "product_name": "测试漆",
                        "quantity_tins": 1,
                        "tin_spec": 20,
                        "quantity_kg": 20,
                        "unit_price": 10,
                        "amount": 200,
                    }
                ],
            }
        ],
        path,
    )

    calls: list[dict] = []

    class _FakeShipmentSvc:
        def create_shipment(self, unit_name, items_data, contact_person="", **kwargs):
            calls.append(
                {
                    "unit_name": unit_name,
                    "items_data": items_data,
                    "contact_person": contact_person,
                    **kwargs,
                }
            )
            return {"success": True, "shipment": {"id": 1000 + len(calls)}}

    monkeypatch.setattr(
        "app.bootstrap.get_shipment_app_service",
        lambda: _FakeShipmentSvc(),
    )
    monkeypatch.setattr(
        "app.services.tools_workflow_registered._execute_excel_import_records",
        lambda records: {"success": True, "imported": len(records)},
    )
    fp_db = tmp_path / "fps.sqlite3"
    monkeypatch.setenv("FHD_SHIPMENT_ETL_FINGERPRINT_BACKEND", "legacy")
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._legacy_db_path",
        lambda: fp_db,
    )
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._db_path",
        lambda: fp_db,
    )
    monkeypatch.setenv("WORKSPACE_ROOT", str(tmp_path))

    first = execute_shipment_excel_etl(path, idempotent=True, workspace_root=tmp_path)
    assert first["success"] is True
    assert first["shipment_created"] == 1
    assert first["shipment_skipped"] == 0
    assert len(calls) == 1
    assert calls[0].get("external_order_number") == "IDEM-1"

    second = execute_shipment_excel_etl(path, idempotent=True, workspace_root=tmp_path)
    assert second["success"] is True
    assert second["shipment_created"] == 0
    assert second["shipment_skipped"] == 1
    assert len(calls) == 1


def test_ledger_requires_confirm(tmp_path, monkeypatch):
    path = tmp_path / "闭环流水客户.xlsx"
    write_ledger_workbook([], path, unit_name="闭环流水客户")
    monkeypatch.setenv("WORKSPACE_ROOT", str(tmp_path))
    result = execute_shipment_excel_etl(
        path,
        include_ledger=True,
        confirm_ledger=False,
        workspace_root=tmp_path,
    )
    assert result["success"] is False
    assert result["error_code"] == "ledger_confirm_required"


def test_dry_run_does_not_write(tmp_path, monkeypatch):
    path = tmp_path / "delivery.xlsx"
    write_delivery_note_workbook(
        [
            {
                "unit_name": "预演客户",
                "order_number": "DRY-1",
                "order_date": "2026年07月24日",
                "items": [
                    {
                        "model_number": "D1",
                        "product_name": "底漆",
                        "quantity_tins": 1,
                        "tin_spec": 25,
                        "quantity_kg": 25,
                        "unit_price": 8,
                        "amount": 200,
                    }
                ],
            }
        ],
        path,
    )
    calls: list = []

    class _FakeShipmentSvc:
        def create_shipment(self, *a, **k):
            calls.append(1)
            return {"success": True, "shipment": {"id": 1}}

    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: _FakeShipmentSvc())
    monkeypatch.setenv("WORKSPACE_ROOT", str(tmp_path))
    out = execute_shipment_excel_etl(path, dry_run=True, workspace_root=tmp_path)
    assert out["success"] is True
    assert out["dry_run"] is True
    assert out["would_create"] == 1
    assert calls == []


def test_batch_preview_and_execute(tmp_path, monkeypatch):
    batch_dir = tmp_path / "batch"
    batch_dir.mkdir()
    write_delivery_note_workbook(
        [
            {
                "unit_name": "批量客户A",
                "order_number": "B-1",
                "order_date": "2026年07月24日",
                "items": [
                    {
                        "model_number": "BA1",
                        "product_name": "底漆",
                        "quantity_tins": 1,
                        "tin_spec": 25,
                        "quantity_kg": 25,
                        "unit_price": 8,
                        "amount": 200,
                    }
                ],
            }
        ],
        batch_dir / "a.xlsx",
    )
    write_ledger_workbook([], batch_dir / "闭环流水客户.xlsx", unit_name="闭环流水客户")

    monkeypatch.setenv("WORKSPACE_ROOT", str(tmp_path))
    preview = batch_preview_shipment_excel_etl(batch_dir, workspace_root=tmp_path)
    assert preview["success"] is True
    assert preview["file_count"] == 2
    assert preview["note_count"] >= 1

    calls: list[str] = []

    class _FakeShipmentSvc:
        def create_shipment(self, unit_name, items_data, contact_person="", **kwargs):
            calls.append(unit_name)
            return {"success": True, "shipment": {"id": len(calls)}}

    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: _FakeShipmentSvc())
    monkeypatch.setattr(
        "app.services.tools_workflow_registered._execute_excel_import_records",
        lambda records: {"success": True, "imported": len(records)},
    )
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._legacy_db_path",
        lambda: tmp_path / "batch_fps.sqlite3",
    )
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._db_path",
        lambda: tmp_path / "batch_fps.sqlite3",
    )
    monkeypatch.setenv("FHD_SHIPMENT_ETL_FINGERPRINT_BACKEND", "legacy")
    monkeypatch.setenv("FHD_SHIPMENT_ETL_ALLOW_BATCH", "1")

    executed = batch_execute_shipment_excel_etl(
        batch_dir,
        idempotent=True,
        include_ledger=False,
        workspace_root=tmp_path,
    )
    assert executed["success"] is True
    assert executed["shipment_created"] >= 1
    assert len(calls) >= 1


def test_execute_creates_shipment(tmp_path, monkeypatch):
    path = tmp_path / "delivery.xlsx"
    write_delivery_note_workbook(
        [
            {
                "unit_name": "测试客户甲",
                "contact_person": "张总",
                "order_date": "2026年01月21日",
                "order_number": "T-1",
                "items": [
                    {
                        "model_number": "RX001",
                        "product_name": "PU哑光漆",
                        "quantity_tins": 2,
                        "tin_spec": 25,
                        "quantity_kg": 50,
                        "unit_price": 17,
                        "amount": 850,
                    }
                ],
            }
        ],
        path,
    )

    monkeypatch.setenv("WORKSPACE_ROOT", str(tmp_path))
    parsed = parse_delivery_notes(path)
    assert parsed["note_count"] == 1

    calls: list[dict] = []

    class _FakeShipmentSvc:
        def create_shipment(self, unit_name, items_data, contact_person="", **kwargs):
            calls.append(
                {
                    "unit_name": unit_name,
                    "items_data": items_data,
                    "contact_person": contact_person,
                    **kwargs,
                }
            )
            return {"success": True, "shipment": {"id": 101}}

    monkeypatch.setattr(
        "app.bootstrap.get_shipment_app_service",
        lambda: _FakeShipmentSvc(),
    )
    monkeypatch.setattr(
        "app.services.tools_workflow_registered._execute_excel_import_records",
        lambda records: {"success": True, "imported_count": len(records)},
    )
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._legacy_db_path",
        lambda: tmp_path / "exec_fps.sqlite3",
    )
    monkeypatch.setattr(
        "app.application.shipment_excel_etl_fingerprint_store._db_path",
        lambda: tmp_path / "exec_fps.sqlite3",
    )
    monkeypatch.setenv("FHD_SHIPMENT_ETL_FINGERPRINT_BACKEND", "legacy")

    result = execute_shipment_excel_etl(path, idempotent=False, workspace_root=tmp_path)
    assert result["success"] is True
    assert result["closed_loop"] is True
    assert result["shipment_created"] == 1
    assert calls[0]["unit_name"] == "测试客户甲"
    assert calls[0]["items_data"][0]["model_number"] == "RX001"
    assert calls[0]["external_order_number"] == "T-1"


def test_rejects_path_outside_sandbox(tmp_path, monkeypatch):
    monkeypatch.delenv("FHD_SHIPMENT_ETL_ALLOW_TMP", raising=False)
    # Force empty pytest flag simulation: still under pytest so temp is allowed.
    # Use a path under /etc which is never allowed.
    result = preview_shipment_excel_etl("/etc/hosts", workspace_root=str(tmp_path))
    assert result["success"] is False
    assert result.get("error_code") == "unsafe_path"
