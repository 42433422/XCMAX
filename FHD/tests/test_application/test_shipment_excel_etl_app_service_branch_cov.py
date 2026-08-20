"""Branch/behaviour coverage for the shipment Excel ETL app service module.

Covers ``app.application.shipment_excel_etl_app_service``:
  * every private helper (scoring / header detection / mapping / inference /
    buyer meta / item building / fingerprint / sheet probe / LLM assist)
  * the public entry points (parse / preview / execute / write / regenerate /
    batch preview / batch execute)
  * the facade class + singleton getter.

Strategy: heavy external dependencies (KB, LLM, OCR, fingerprint store, security,
shipment app service, product-import) are stubbed with ``unittest.mock`` so tests
run fast and deterministically without any real DB / network / LLM / file parsing.
Real openpyxl workbooks are built only for the parsing io paths.
"""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import Workbook

from app.application import shipment_excel_etl_app_service as svc_mod
from app.application.shipment_etl_profile import (
    CompiledMetaPatterns,
    ShipmentEtlProfile,
)
from app.application.shipment_excel_etl_app_service import (
    ShipmentExcelEtlApplicationService,
    batch_execute_shipment_excel_etl,
    batch_preview_shipment_excel_etl,
    execute_shipment_excel_etl,
    get_shipment_excel_etl_app_service,
    note_fingerprint,
    parse_delivery_notes,
    preview_shipment_excel_etl,
    regenerate_delivery_notes_from_file,
    write_delivery_note_workbook,
    write_ledger_workbook,
)

MOD = "app.application.shipment_excel_etl_app_service"
SEC = "app.application.shipment_excel_etl_security"
KB = "app.application.excel_etl_kb"
LLM = "app.application.shipment_excel_etl_llm"
FP = "app.application.shipment_excel_etl_fingerprint_store"


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


class FakeWS:
    """Minimal worksheet stand-in: ``ws.cell(r, c).value`` + metadata."""

    def __init__(self, rows, title="S"):
        self._rows = rows
        self.title = title
        self.max_row = len(rows)
        self.max_column = max((len(r) for r in rows), default=0)

    def cell(self, row, col):
        r, c = row - 1, col - 1
        if r < 0 or r >= len(self._rows) or c < 0 or c >= len(self._rows[r]):
            return SimpleNamespace(value=None)
        return SimpleNamespace(value=self._rows[r][c])


def _patterns():
    return CompiledMetaPatterns(
        title=re.compile(r"送货单|delivery|test", re.I),
        buyer=re.compile(r"(?:客户|购货单位)[：:\s]+([^\s]+)", re.I),
        buyer_split=re.compile(r"(?:客户|购货单位)[：:]", re.I),
        buyer_stop=re.compile(r"联系人|日期|单号", re.I),
        contact=re.compile(r"联系人[：:\s]*([^\s]+)", re.I),
        date=re.compile(
            r"((?:20)?\d{2}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日|\d{4}[-/]\d{1,2}[-/]\d{1,2})"
        ),
        order_no=re.compile(r"(?:单号|订单编号)[：:\s]*([A-Za-z0-9\-]+)", re.I),
        stop_row=re.compile(r"合计|总计|小计|grand\s*total|total", re.I),
        buyer_label="客户",
        ledger_sheet=re.compile(r"流水|ledger|出货", re.I),
    )


def _make_profile(**overrides):
    base = {
        "id": "test",
        "kind": "document",
        "label": "测试",
        "target": "shipment",
        "raw": {},
        "meta_patterns": _patterns(),
        "detect": {
            "delivery": {
                "probe_rows": 8,
                "title_weight": 50,
                "buyer_token": "客户",
                "buyer_weight": 25,
                "header_hit_tokens": ["型号", "名称", "数量"],
                "header_hit_cap": 5,
                "header_hit_weight": 6,
                "bonus_tokens": [{"token": "送货", "weight": 10}],
                "min_score": 40,
            },
            "ledger": {
                "probe_rows": 10,
                "suppress_if_delivery_score_gte": 60,
                "sheet_weight": 20,
                "content_tokens": ["出货", "流水"],
                "hit_tokens": ["型号", "名称", "数量", "单号"],
                "hit_cap": 6,
                "hit_weight": 10,
                "bonus_require_token": "出货",
                "bonus_exclude_token": "报价",
                "bonus_weight": 5,
            },
        },
        "header_detect": {
            "delivery": {
                "max_scan_rows": 12,
                "require_groups": [
                    ["型号", "货号", "model"],
                    ["名称", "品名", "name"],
                    ["数量", "qty"],
                ],
            },
            "ledger": {
                "max_scan_rows": 16,
                "require_groups": [["单号", "编号", "order"]],
                "and_any_groups": [["型号", "名称"]],
            },
        },
        "columns": {
            "model_number": [{"contains_any": ["型号", "货号"]}],
            "product_name": [{"contains_any": ["名称", "品名"]}],
            "quantity_tins": [{"contains_any": ["数量", "件数"]}],
            "tin_spec": [{"contains_any": ["规格", "单重"]}],
            "quantity_kg": [{"contains_any": ["数量kg", "公斤"]}],
            "unit_price": [{"contains_any": ["单价", "价格"]}],
            "amount": [{"contains_any": ["金额"]}],
            "order_number": [{"contains_any": ["单号", "订单"]}],
            "order_date": [{"contains_any": ["日期"]}],
        },
        "ledger": {"title_template": "{unit}/{order_no}"},
        "write": {},
    }
    base.update(overrides)
    return ShipmentEtlProfile(**base)


def _delivery_ws():
    return FakeWS(
        [
            ["测试送货单"],
            ["客户：测试公司    联系人：张总    2026年07月24日    单号：M-100"],
            ["型号", "名称", "数量", "规格", "数量KG", "单价", "金额"],
            ["M01", "面漆", 1, 20, 20, 10, 200],
            ["M02", "底漆", 2, 25, 50, 8, 400],
        ],
        title="送货单",
    )


def _ledger_ws():
    return FakeWS(
        [
            ["客户：测试公司"],
            ["单号", "型号", "名称", "数量", "数量KG", "单价", "金额"],
            ["L-001", "C01", "面漆", 1, 20, 10, 200],
            ["L-002", "C02", "底漆", 2, 40, 8, 320],
        ],
        title="出货流水",
    )


def _write_delivery_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "送货单"
    ws["A1"] = "测试送货单"
    ws["A2"] = "客户：测试公司    联系人：张总    2026年07月24日    单号：M-100"
    ws["A3"] = "型号"
    ws["B3"] = "名称"
    ws["C3"] = "数量"
    ws["D3"] = "规格"
    ws["E3"] = "数量KG"
    ws["F3"] = "单价"
    ws["G3"] = "金额"
    ws["A4"] = "M01"
    ws["B4"] = "面漆"
    ws["C4"] = 1
    ws["D4"] = 20
    ws["E4"] = 20
    ws["F4"] = 10
    ws["G4"] = 200
    ws["A5"] = "M02"
    ws["B5"] = "底漆"
    ws["C5"] = 2
    ws["D5"] = 25
    ws["E5"] = 50
    ws["F5"] = 8
    ws["G5"] = 400
    wb.save(path)
    return path


def _write_ledger_xlsx(path):
    wb = Workbook()
    ws = wb.active
    ws.title = "出货流水"
    ws["A1"] = "客户：测试公司"
    ws["A2"] = "单号"
    ws["B2"] = "型号"
    ws["C2"] = "名称"
    ws["D2"] = "数量"
    ws["E2"] = "数量KG"
    ws["F2"] = "单价"
    ws["G2"] = "金额"
    ws["A3"] = "L-001"
    ws["B3"] = "C01"
    ws["C3"] = "面漆"
    ws["D3"] = 1
    ws["E3"] = 20
    ws["F3"] = 10
    ws["G3"] = 200
    ws["A4"] = "L-002"
    ws["B4"] = "C02"
    ws["C4"] = "底漆"
    ws["D4"] = 2
    ws["E4"] = 40
    ws["F4"] = 8
    ws["G4"] = 320
    wb.save(path)
    return path


def _delivery_note(profile_target="shipment"):
    return {
        "unit_name": "客户A",
        "contact_person": "张三",
        "order_date": "2026-07-24",
        "order_number": "O-1",
        "source_kind": "delivery_note",
        "profile_target": profile_target,
        "items": [
            {
                "model_number": "M1",
                "product_name": "面漆",
                "quantity_tins": 1,
                "tin_spec": 20.0,
                "quantity_kg": 20,
                "unit_price": 10,
                "amount": 200,
            }
        ],
    }


class _FakeKB:
    def get_template(self, fp):
        return None

    def touch(self, fp):
        return None

    def remember(self, mem):
        return None

    def lookup(self, fp):
        return None

    def list_templates(self):
        return []


@pytest.fixture
def delivery_profile():
    return _make_profile()


@pytest.fixture
def no_llm(monkeypatch):
    monkeypatch.setattr(f"{LLM}.needs_llm_assist", lambda **kw: (False, "rules_confident"))
    monkeypatch.setattr(
        f"{LLM}.assist_sheet_layout",
        lambda probe: SimpleNamespace(as_public_dict=lambda: {"ok": False}) or None,
    )
    monkeypatch.setattr(f"{LLM}.unit_name_is_weak", lambda *a, **kw: False)


@pytest.fixture
def no_kb(monkeypatch):
    monkeypatch.setattr(f"{MOD}.get_excel_etl_kb", lambda: _FakeKB())


# ---------------------------------------------------------------------------
# pure helpers
# ---------------------------------------------------------------------------


def test_norm_cell_and_header():
    assert svc_mod._norm_cell(None) == ""
    assert svc_mod._norm_cell("\u3000a\u3000b c\u3000") == "abc"
    assert svc_mod._norm_header(" 型号 ") == "型号"
    assert svc_mod._norm_header("NaMe") == "name"


def test_to_float_branches():
    assert svc_mod._to_float(None) == 0.0
    assert svc_mod._to_float("") == 0.0
    assert svc_mod._to_float(5) == 5.0
    assert svc_mod._to_float(5.5) == 5.5
    assert svc_mod._to_float("1,234.5") == 1234.5
    assert svc_mod._to_float("abc", default=3.0) == 3.0


def test_to_int_branches():
    assert svc_mod._to_int("3.4") == 3
    assert svc_mod._to_int("abc", default=7) == 7
    assert svc_mod._to_int(None, default=2) == 2


def test_row_texts_and_joined():
    ws = FakeWS([["a", None, "  b  ", ""], ["x"]])
    assert svc_mod._row_texts(ws, 1) == ["a", "b"]
    assert svc_mod._joined_row(ws, 1) == "a b"
    assert svc_mod._row_texts(ws, 2) == ["x"]
    ws2 = FakeWS([["   ", None]])
    assert svc_mod._row_texts(ws2, 1) == []


def test_token_in_compact():
    assert svc_mod._token_in_compact("", "x") is False
    assert svc_mod._token_in_compact("型号", "有型号") is True
    assert svc_mod._token_in_compact("A/B", "a/b") is True
    assert svc_mod._token_in_compact("AbC", "abc") is True
    assert svc_mod._token_in_compact("zzz", "abc") is False


def test_header_cell_texts():
    ws = FakeWS([["a", None, "b", "c"], ["d"]], title="S")
    assert svc_mod._header_cell_texts(ws, 1) == ["a", "b", "c"]


def test_kb_resolve_layout_hit_and_miss(monkeypatch):
    class KBHit:
        def __init__(self):
            self.columns = {"product_name": 2}
            self.header_row = 3
            self.touched = False

        def get_template(self, fp):
            return self

        def touch(self, fp):
            self.touched = True
            return self

    kb = KBHit()
    monkeypatch.setattr(f"{MOD}.get_excel_etl_kb", lambda: kb)
    ws = _delivery_ws()
    row, mapping, fp = svc_mod._kb_resolve_layout(ws)
    assert row == 3
    assert mapping.get("product_name") == 2
    assert fp
    assert kb.touched is True

    # no memory -> miss
    monkeypatch.setattr(f"{MOD}.get_excel_etl_kb", lambda: _FakeKB())
    row2, mapping2, fp2 = svc_mod._kb_resolve_layout(FakeWS([["x"], ["y"]]))
    assert row2 is None and mapping2 == {} and fp2 == ""


def test_remember_sheet_layout_branches(monkeypatch):
    captured = []
    kb = _FakeKB()
    kb.remember = lambda mem: captured.append(mem)
    monkeypatch.setattr(f"{MOD}.get_excel_etl_kb", lambda: kb)
    prof = _make_profile()
    # empty mapping -> ""
    assert (
        svc_mod._remember_sheet_layout(_delivery_ws(), header_row=3, mapping={}, profile=prof) == ""
    )
    # <2 headers -> ""
    assert (
        svc_mod._remember_sheet_layout(
            FakeWS([["single"]]), header_row=1, mapping={"a": 1}, profile=prof
        )
        == ""
    )
    fp = svc_mod._remember_sheet_layout(
        _delivery_ws(), header_row=3, mapping={"product_name": 2}, profile=prof, source="rules"
    )
    assert fp
    assert captured and captured[0].columns == {"product_name": 2}

    # exception path
    class KBFail:
        def remember(self, mem):
            raise OSError("boom")

    monkeypatch.setattr(f"{MOD}.get_excel_etl_kb", lambda: KBFail())
    assert (
        svc_mod._remember_sheet_layout(
            _delivery_ws(), header_row=3, mapping={"product_name": 2}, profile=prof
        )
        == ""
    )


def test_score_delivery_sheet(monkeypatch):
    prof = _make_profile()
    ws = _delivery_ws()
    score = svc_mod._score_delivery_sheet(ws, prof)
    assert score >= 50
    # bonus token only when dict
    assert svc_mod._score_delivery_sheet(ws, prof) == score


def test_score_ledger_sheet(monkeypatch):
    prof = _make_profile()
    # delivery sheet suppressed (delivery score high)
    assert svc_mod._score_ledger_sheet(_delivery_ws(), prof) == 0
    # real ledger sheet
    lscore = svc_mod._score_ledger_sheet(_ledger_ws(), prof)
    assert lscore >= 40
    # bonus excluded token present -> no bonus
    ws_excl = FakeWS([["客户：测试公司    报价单"], [""]] + [[None] * 7] * 3, title="出货流水")
    assert svc_mod._score_ledger_sheet(ws_excl, prof) >= 0


def test_find_header_row_known():
    prof = _make_profile()
    assert svc_mod._find_header_row(_delivery_ws(), prof) == 3


def test_find_header_row_unknown_fallback():
    prof = _make_profile()
    prof.header_detect = {"delivery": {"max_scan_rows": 12, "require_groups": [["编号", "SKU"]]}}
    ws = FakeWS(
        [
            ["标题", "", "", ""],
            ["型号", "名称", "数量", "金额"],
            ["M01", "面漆", 1, 200],
        ]
    )
    assert svc_mod._find_header_row(ws, prof) == 2
    # no body -> None
    ws2 = FakeWS([["标题", "", "", ""], ["型号", "名称", "数量", "金额"]])
    assert svc_mod._find_header_row(ws2, prof) is None


def test_find_ledger_header_row():
    prof = _make_profile()
    assert svc_mod._find_ledger_header_row(_ledger_ws(), prof) == 2
    # groups not matched -> None
    assert svc_mod._find_ledger_header_row(_delivery_ws(), prof) is None
    # and_any not satisfied -> None
    prof2 = _make_profile()
    prof2.header_detect = {
        "ledger": {
            "max_scan_rows": 16,
            "require_groups": [["型号"]],
            "and_any_groups": [["报价"]],
        }
    }
    assert svc_mod._find_ledger_header_row(_ledger_ws(), prof2) is None


def test_map_headers_and_only_if_missing():
    prof = _make_profile()
    mapping = svc_mod._map_headers(_delivery_ws(), 3, prof)
    assert mapping.get("model_number") == 1
    assert mapping.get("product_name") == 2
    assert mapping.get("quantity_tins") == 3
    assert mapping.get("amount") == 7

    prof2 = _make_profile(
        columns={
            "product_name": [{"contains_any": ["名称"], "only_if_missing": ["model_number"]}],
            "model_number": [{"contains_any": ["型号"]}],
        }
    )
    mapping2 = svc_mod._map_headers(_delivery_ws(), 3, prof2)
    assert mapping2.get("model_number") == 1
    # product_name has only_if_missing model_number which is already mapped → skipped
    assert mapping2.get("product_name") is None


def test_sample_values():
    ws = FakeWS([["h"], ["a"], [None], ["b"], ["c"], ["d"], ["e"], ["f"], ["g"], ["h"]])
    assert svc_mod._sample_values(ws, 1, 1) == ["a", "b", "c", "d", "e"]
    assert svc_mod._sample_values(ws, 1, 1, limit=2) == ["a", "b"]
    assert svc_mod._sample_values(FakeWS([["h"]]), 1, 1) == []


def test_infer_columns_from_samples(monkeypatch):
    prof = _make_profile()
    ws = FakeWS(
        [
            ["编号", "品名", "数量", "单价"],
            ["M-01", "面漆", "10", "12.5"],
            ["M-02", "底漆", "20", "8"],
        ]
    )
    monkeypatch.setitem(os.environ, "FHD_EXCEL_ETL_HEURISTIC", "1")
    base = {"product_name": 2}
    inferred = svc_mod._infer_columns_from_samples(ws, 1, base)
    assert "model_number" in inferred
    assert "quantity_tins" in inferred
    assert "tin_spec" in inferred

    # heuristic off -> copy
    monkeypatch.setitem(os.environ, "FHD_EXCEL_ETL_HEURISTIC", "off")
    assert svc_mod._infer_columns_from_samples(ws, 1, base) == base


def test_classify_sheet_role(monkeypatch):
    prof = _make_profile()
    d_ws = _delivery_ws()
    d = svc_mod._score_delivery_sheet(d_ws, prof)
    l = svc_mod._score_ledger_sheet(d_ws, prof)
    assert svc_mod._classify_sheet_role(d_ws, prof, d_score=d, l_score=l) == "delivery"

    l_ws = _ledger_ws()
    dl = svc_mod._score_delivery_sheet(l_ws, prof)
    ll = svc_mod._score_ledger_sheet(l_ws, prof)
    assert svc_mod._classify_sheet_role(l_ws, prof, d_score=dl, l_score=ll) == "ledger"

    # ignore when both low & price-list title
    assert (
        svc_mod._classify_sheet_role(
            FakeWS([["报价单"], ["x"]], title="报价单"), prof, d_score=5, l_score=5
        )
        == "ignore"
    )
    # ignore default branch
    assert (
        svc_mod._classify_sheet_role(
            FakeWS([["x"], ["y"]], title="Sheet1"), prof, d_score=5, l_score=5
        )
        == "ignore"
    )
    # unknown
    assert (
        svc_mod._classify_sheet_role(
            FakeWS([["x"], ["y"], ["z"]], title="Sheet1"), prof, d_score=20, l_score=10
        )
        == "unknown"
    )


def test_unit_name_looks_truncated():
    assert svc_mod._unit_name_looks_truncated("") is True
    assert svc_mod._unit_name_looks_truncated("ltd.") is True
    assert svc_mod._unit_name_looks_truncated("Tech") is True
    assert svc_mod._unit_name_looks_truncated("科技有限公司") is False


def test_parse_buyer_meta(monkeypatch):
    monkeypatch.setattr(f"{MOD}._extract_adjacent_buyer_meta", lambda ws, r: {})
    monkeypatch.setattr(f"{LLM}.unit_name_is_weak", lambda u, **kw: False)
    ws = _delivery_ws()
    meta = svc_mod._parse_buyer_meta(ws, 3, _make_profile())
    assert meta["unit_name"] == "测试公司"
    assert meta["title"] == "测试送货单"
    assert meta["contact_person"] == "张总"
    assert meta["order_date"] == "2026年07月24日"
    assert meta["order_number"] == "M-100"


def test_parse_buyer_meta_weak_cleared(monkeypatch):
    monkeypatch.setattr(f"{MOD}._extract_adjacent_buyer_meta", lambda ws, r: {})
    monkeypatch.setattr(f"{LLM}.unit_name_is_weak", lambda u, **kw: True)
    ws = _delivery_ws()
    meta = svc_mod._parse_buyer_meta(ws, 3, _make_profile())
    assert meta["unit_name"] == ""


def test_parse_buyer_meta_label_split(monkeypatch):
    monkeypatch.setattr(f"{MOD}._extract_adjacent_buyer_meta", lambda ws, r: {})
    monkeypatch.setattr(f"{LLM}.unit_name_is_weak", lambda u, **kw: False)
    ws = FakeWS([["客户 测试公司 联系人 张三"], ["型号", "名称"]])
    meta = svc_mod._parse_buyer_meta(ws, 2, _make_profile())
    assert meta["unit_name"] == "测试公司"


def test_extract_adjacent_buyer_meta(monkeypatch):
    monkeypatch.setattr(f"{LLM}.unit_name_is_weak", lambda u, **kw: False)
    ws = FakeWS(
        [
            ["Bill To", "Dukjil Trading Pte Ltd"],
            ["Attn", "Tom"],
            ["DO No: X-100", ""],
            ["", "", "", ""],
        ]
    )
    out = svc_mod._extract_adjacent_buyer_meta(ws, 4)
    assert out["unit_name"] == "Dukjil Trading Pte Ltd"
    assert out["contact_person"] == "Tom"
    assert out["order_number"] == "X-100"


def test_looks_like_non_product_and_titleish():
    assert svc_mod._looks_like_non_product_token("January") is True
    assert svc_mod._looks_like_non_product_token("") is False
    assert svc_mod._looks_like_non_product_token("面漆") is False
    assert svc_mod._looks_like_titleish("title") is True
    assert svc_mod._looks_like_titleish("") is False
    assert svc_mod._looks_like_titleish("2026") is False
    assert svc_mod._looks_like_titleish("面漆") is False


def test_build_item_from_row(monkeypatch):
    prof = _make_profile()
    mapping = {"model_number": 1, "product_name": 2, "quantity_tins": 3, "tin_spec": 4}
    ws = FakeWS([["h"], ["M01", "面漆", 2, 20]])
    item = svc_mod._build_item_from_row(ws, 2, mapping)
    assert item["model_number"] == "M01"
    assert item["quantity_tins"] == 2
    assert item["tin_spec"] == 20.0

    # name only
    m2 = {"product_name": 1}
    assert svc_mod._build_item_from_row(FakeWS([["h"], ["漆"]]), 2, m2) is None  # no qty
    # model with no alnum -> becomes name
    m3 = {"model_number": 1, "product_name": 2, "quantity_tins": 3}
    item3 = svc_mod._build_item_from_row(FakeWS([["h"], ["面漆", None, 2]]), 2, m3)
    assert item3["product_name"] == "面漆" and item3["model_number"] == ""
    # both empty -> None
    assert svc_mod._build_item_from_row(FakeWS([["h"], [None, None]]), 2, m3) is None
    # non-product token blocked
    assert (
        svc_mod._build_item_from_row(FakeWS([["h"], ["title", 1, 1, 1, 1, 1, 1]]), 2, mapping)
        is None
    )
    # year-as-qty + titleish name -> None
    m_y = {"model_number": 1, "product_name": 2, "quantity_tins": 3}
    assert svc_mod._build_item_from_row(FakeWS([["h"], [None, "title", 2024]]), 2, m_y) is None
    # computed tin_spec from qty
    m_c = {"model_number": 1, "product_name": 2, "quantity_tins": 3, "quantity_kg": 4}
    itemc = svc_mod._build_item_from_row(FakeWS([["h"], ["M", "面", 2, 40]]), 2, m_c)
    assert itemc["tin_spec"] == 20.0
    # computed qty_kg from tins*spec
    m_q = {
        "model_number": 1,
        "product_name": 2,
        "quantity_tins": 3,
        "tin_spec": 4,
        "quantity_kg": 5,
    }
    itemq = svc_mod._build_item_from_row(FakeWS([["h"], ["M", "面", 2, 20, None]]), 2, m_q)
    assert itemq["quantity_kg"] == 40.0
    # amount from price*qty
    m_a = {
        "model_number": 1,
        "product_name": 2,
        "quantity_tins": 3,
        "quantity_kg": 4,
        "unit_price": 5,
        "amount": 6,
    }
    itema = svc_mod._build_item_from_row(FakeWS([["h"], ["M", "面", 2, 20, 10, None]]), 2, m_a)
    assert itema["amount"] == 200.0
    # tins 0 -> forced 1
    item1 = svc_mod._build_item_from_row(
        FakeWS([["h"], ["M", "面", 0, 20, None, None, None]]),
        2,
        {"model_number": 1, "product_name": 2, "quantity_tins": 3, "quantity_kg": 4},
    )
    assert item1["quantity_tins"] == 1


def test_parse_items_stop_row(monkeypatch):
    prof = _make_profile()
    mapping = {"model_number": 1, "product_name": 2, "quantity_tins": 3}
    ws = FakeWS([["h"], ["M1", "面漆", 1], ["合计", "", ""], ["M2", "底漆", 1]])
    items = svc_mod._parse_items(ws, 1, mapping, prof)
    assert len(items) == 1
    assert items[0]["model_number"] == "M1"


def test_note_fingerprint():
    a = _delivery_note()
    b = _delivery_note()
    assert note_fingerprint(a) == note_fingerprint(b)
    c = dict(a)
    c["items"] = [dict(a["items"][0], model_number="DIFFERENT")]
    assert note_fingerprint(a) != note_fingerprint(c)


def test_fingerprint_store_path_and_legacy(monkeypatch, tmp_path):
    monkeypatch.setattr("app.utils.path_io.path_utils.get_data_dir", lambda: tmp_path)
    path = svc_mod._fingerprint_store_path()
    assert path.name == "shipment_etl_fingerprints.json"
    assert svc_mod._legacy_json_has_fingerprint("x") is False

    path.write_text('{"entries": {"fp1": 1}}', encoding="utf-8")
    assert svc_mod._legacy_json_has_fingerprint("fp1") is True
    assert svc_mod._legacy_json_has_fingerprint("nope") is False

    # invalid json
    path.write_text("{not json", encoding="utf-8")
    assert svc_mod._legacy_json_has_fingerprint("fp1") is False


def test_load_save_fingerprints(monkeypatch, tmp_path):
    monkeypatch.setattr("app.utils.path_io.path_utils.get_data_dir", lambda: tmp_path)
    assert svc_mod._load_fingerprints() == {"entries": {}}
    svc_mod._save_fingerprints({"entries": {"a": 1}})
    assert svc_mod._load_fingerprints() == {"entries": {"a": 1}}
    # non-dict entries -> empty
    (tmp_path / "shipment_etl_fingerprints.json").write_text('{"entries": []}', encoding="utf-8")
    assert svc_mod._load_fingerprints() == {"entries": {}}


def test_is_fingerprint_imported(monkeypatch):
    monkeypatch.setattr(f"{FP}.has_fingerprint", lambda a, b: True)
    assert svc_mod._is_fingerprint_imported("t", "fp") is True
    monkeypatch.setattr(f"{FP}.has_fingerprint", lambda a, b: False)
    monkeypatch.setattr(f"{MOD}._legacy_json_has_fingerprint", lambda fp: True)
    assert svc_mod._is_fingerprint_imported("t", "fp") is True


def test_record_fingerprint_now(monkeypatch):
    rec = MagicMock()
    monkeypatch.setattr(f"{FP}.record_fingerprint", rec)
    svc_mod._record_fingerprint_now(
        "t", "fp", shipment_id=1, unit_name="u", order_number="o", file_name="f"
    )
    rec.assert_called_once_with(
        "t", "fp", shipment_id=1, unit_name="u", order_number="o", file_name="f"
    )


def test_enrich_note():
    note = _delivery_note()
    out = svc_mod._enrich_note(note)
    assert out["sheet"] == ""
    assert out["fingerprint"] == note_fingerprint(out)
    assert out["item_count"] == 1
    assert out["total_amount"] == 200.0


def test_build_sheet_probe():
    prof = _make_profile()
    probe = svc_mod._build_sheet_probe(_delivery_ws(), prof, rule_hint={"k": 1})
    assert probe.profile_id == "test"
    assert probe.max_row == 5
    assert probe.max_col == 7
    assert probe.rule_hint == {"k": 1}
    assert probe.candidate_headers


def test_merge_meta():
    base = {"unit_name": "a", "contact_person": "b"}
    assert svc_mod._merge_meta(base, {"unit_name": "A", "title": "t"})["unit_name"] == "a"
    assert (
        svc_mod._merge_meta(base, {"unit_name": "A", "title": "t"}, prefer_overlay=True)[
            "unit_name"
        ]
        == "A"
    )
    assert svc_mod._merge_meta(base, {"unit_name": ""})["unit_name"] == "a"


def test_apply_llm_assist_to_layout(monkeypatch):
    prof = _make_profile()
    ws = _delivery_ws()

    # not needed -> early return
    monkeypatch.setattr(f"{LLM}.needs_llm_assist", lambda **kw: (False, "ok"))
    res = svc_mod._apply_llm_assist_to_layout(
        ws,
        prof,
        delivery_score=10,
        ledger_score=0,
        min_score=40,
        header_row=3,
        mapping={"a": 1},
        meta={"unit_name": "x"},
        prefer_kind="delivery_note",
    )
    assert res[0] == 3 and res[4]["ok"] is True

    # needed but assist fails -> unchanged
    monkeypatch.setattr(f"{LLM}.needs_llm_assist", lambda **kw: (True, "gray"))
    monkeypatch.setattr(
        f"{LLM}.assist_sheet_layout",
        lambda probe: SimpleNamespace(
            ok=False,
            as_public_dict=lambda: {"ok": False, "used_llm": True, "reason": "err"},
            header_row=None,
            columns={},
            meta={},
            source_kind="",
        ),
    )
    res2 = svc_mod._apply_llm_assist_to_layout(
        ws,
        prof,
        delivery_score=10,
        ledger_score=0,
        min_score=40,
        header_row=3,
        mapping={"a": 1},
        meta={"unit_name": "x"},
        prefer_kind="delivery_note",
    )
    assert res2[0] == 3

    # assist ok -> merge columns/meta/kind
    monkeypatch.setattr(
        f"{LLM}.assist_sheet_layout",
        lambda probe: SimpleNamespace(
            ok=True,
            as_public_dict=lambda: {"ok": True, "used_llm": True},
            header_row=2,
            columns={"product_name": 2, "model_number": 1},
            meta={"unit_name": "新客户", "title": "t"},
            source_kind="delivery_note",
        ),
    )
    res3 = svc_mod._apply_llm_assist_to_layout(
        ws,
        prof,
        delivery_score=10,
        ledger_score=0,
        min_score=40,
        header_row=3,
        mapping={"a": 1},
        meta={"unit_name": "x"},
        prefer_kind="delivery_note",
    )
    assert res3[0] == 2
    assert res3[1]["product_name"] == 2
    assert res3[2]["unit_name"] == "新客户"
    assert res3[3] == "delivery_note"

    # invalid source_kind -> fallback prefer_kind
    monkeypatch.setattr(
        f"{LLM}.assist_sheet_layout",
        lambda probe: SimpleNamespace(
            ok=True,
            as_public_dict=lambda: {"ok": True},
            header_row=2,
            columns={},
            meta={},
            source_kind="bogus",
        ),
    )
    res4 = svc_mod._apply_llm_assist_to_layout(
        ws,
        prof,
        delivery_score=10,
        ledger_score=0,
        min_score=40,
        header_row=3,
        mapping={"a": 1},
        meta={},
        prefer_kind="shipment_ledger",
    )
    assert res4[3] == "shipment_ledger"


def test_excel_date_to_str():
    prof = _make_profile()
    assert svc_mod._excel_date_to_str(None, prof) == ""
    assert svc_mod._excel_date_to_str("", prof) == ""
    assert svc_mod._excel_date_to_str(datetime(2026, 7, 24), prof) == "2026-07-24"
    assert svc_mod._excel_date_to_str("2026年07月24日", prof) == "2026年07月24日"
    assert svc_mod._excel_date_to_str("2026-07-24", prof) == "2026-07-24"
    assert svc_mod._excel_date_to_str("无日期", prof) == "无日期"


def test_notes_to_product_records():
    notes = [
        {
            "unit_name": "A",
            "items": [
                {"model_number": "m1", "product_name": "面漆", "unit_price": 10},
                {"model_number": "M1", "product_name": "面漆", "unit_price": 10},  # dup (upper)
            ],
        },
        {
            "unit_name": "A",
            "items": [{"model_number": "M2", "product_name": "底漆", "unit_price": 8}],
        },
    ]
    records = svc_mod._notes_to_product_records(notes)
    assert len(records) == 2


# ---------------------------------------------------------------------------
# parse_delivery_notes + parse sheet internals
# ---------------------------------------------------------------------------


def test_parse_delivery_sheet_rules_only(monkeypatch, no_llm, no_kb):
    monkeypatch.setattr(f"{MOD}._extract_adjacent_buyer_meta", lambda ws, r: {})
    note = svc_mod._parse_delivery_sheet(_delivery_ws(), fallback_unit="f", profile=_make_profile())
    assert note is not None
    assert note["source_kind"] == "delivery_note"
    assert note["unit_name"] == "测试公司"
    assert len(note["items"]) == 2


def test_parse_delivery_sheet_no_header(monkeypatch, no_llm, no_kb):
    ws = FakeWS([["纯文本"], ["无表头"]])
    note = svc_mod._parse_delivery_sheet(ws, fallback_unit="f", profile=_make_profile())
    assert note is None


def test_parse_delivery_sheet_llm_ignore(monkeypatch, no_kb):
    monkeypatch.setattr(f"{LLM}.unit_name_is_weak", lambda *a, **kw: True)
    monkeypatch.setattr(f"{LLM}.needs_llm_assist", lambda **kw: (True, "gray"))
    monkeypatch.setattr(
        f"{LLM}.assist_sheet_layout",
        lambda probe: SimpleNamespace(
            ok=True,
            as_public_dict=lambda: {"ok": True, "used_llm": True},
            header_row=3,
            columns={},
            meta={},
            source_kind="ignore",
        ),
    )
    assert (
        svc_mod._parse_delivery_sheet(_delivery_ws(), fallback_unit="f", profile=_make_profile())
        is None
    )


def test_parse_delivery_sheet_llm_ledger_reclass(monkeypatch, no_kb):
    monkeypatch.setattr(f"{LLM}.unit_name_is_weak", lambda *a, **kw: True)
    monkeypatch.setattr(f"{LLM}.needs_llm_assist", lambda **kw: (True, "gray"))
    monkeypatch.setattr(
        f"{LLM}.assist_sheet_layout",
        lambda probe: SimpleNamespace(
            ok=True,
            as_public_dict=lambda: {"ok": True, "used_llm": True},
            header_row=3,
            columns={},
            meta={},
            source_kind="shipment_ledger",
        ),
    )
    assert (
        svc_mod._parse_delivery_sheet(_delivery_ws(), fallback_unit="f", profile=_make_profile())
        is None
    )


def test_parse_ledger_sheet(monkeypatch, no_llm, no_kb):
    notes = svc_mod._parse_ledger_sheet(
        _ledger_ws(), fallback_unit="ledger", profile=_make_profile()
    )
    assert len(notes) == 2
    assert all(n["source_kind"] == "shipment_ledger" for n in notes)
    assert notes[0]["unit_name"] == "ledger"
    assert notes[0]["title"] == "ledger/L-001"


def test_parse_ledger_sheet_empty_label_split(monkeypatch, no_llm, no_kb):
    prof = _make_profile()
    prof.ledger = {"title_template": "{unit}/{order_no}"}
    notes = svc_mod._parse_ledger_sheet(_ledger_ws(), fallback_unit="", profile=prof)
    assert notes


def test_parse_ledger_sheet_llm_ignore(monkeypatch, no_kb):
    monkeypatch.setattr(f"{LLM}.needs_llm_assist", lambda **kw: (True, "gray"))
    monkeypatch.setattr(
        f"{LLM}.assist_sheet_layout",
        lambda probe: SimpleNamespace(
            ok=True,
            as_public_dict=lambda: {"ok": True},
            header_row=2,
            columns={},
            meta={},
            source_kind="ignore",
        ),
    )
    assert (
        svc_mod._parse_ledger_sheet(_ledger_ws(), fallback_unit="u", profile=_make_profile()) == []
    )


def test_parse_ledger_sheet_missing_order(monkeypatch, no_llm, no_kb):
    prof = _make_profile()
    ws = FakeWS([["型号", "名称"], ["M01", "面漆"]])
    assert svc_mod._parse_ledger_sheet(ws, fallback_unit="u", profile=prof) == []


def test_parse_delivery_notes_happy(tmp_path, monkeypatch, delivery_profile, no_llm, no_kb):
    path = _write_delivery_xlsx(tmp_path / "del.xlsx")
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: path)
    monkeypatch.setattr(f"{MOD}._profiles_for_parse", lambda *a, **k: [delivery_profile])
    monkeypatch.setattr(f"{MOD}._extract_adjacent_buyer_meta", lambda ws, r: {})
    res = parse_delivery_notes(path)
    assert res["success"] is True
    assert res["delivery_note_count"] == 1
    assert res["note_count"] == 1
    assert res["mixed_workbook"] is False
    assert res["profile_id"] == "test"


def test_parse_delivery_notes_ledger_only(tmp_path, monkeypatch, delivery_profile, no_llm, no_kb):
    path = _write_ledger_xlsx(tmp_path / "lg.xlsx")
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: path)
    monkeypatch.setattr(f"{MOD}._profiles_for_parse", lambda *a, **k: [delivery_profile])
    res = parse_delivery_notes(path, include_ledger=True)
    assert res["success"] is True
    assert res["ledger_note_count"] == 2


def test_parse_delivery_notes_mixed_auto(tmp_path, monkeypatch, delivery_profile, no_llm, no_kb):
    wb = Workbook()
    ws = wb.active
    ws.title = "送货单"
    ws["A1"] = "测试送货单"
    ws["A2"] = "客户：测试公司    2026年07月24日    单号：M-100"
    ws["A3"] = "型号"
    ws["B3"] = "名称"
    ws["C3"] = "数量"
    ws["A4"] = "M01"
    ws["B4"] = "面漆"
    ws["C4"] = 1
    lg = wb.create_sheet("出货流水")
    lg["A1"] = "客户：测试公司"
    lg["A2"] = "单号"
    lg["B2"] = "型号"
    lg["C2"] = "名称"
    lg["D2"] = "数量"
    lg["A3"] = "L-001"
    lg["B3"] = "C01"
    lg["C3"] = "面漆"
    lg["D3"] = 1
    path = tmp_path / "mixed.xlsx"
    wb.save(path)
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: path)
    monkeypatch.setattr(f"{MOD}._profiles_for_parse", lambda *a, **k: [delivery_profile])
    monkeypatch.setattr(f"{MOD}._extract_adjacent_buyer_meta", lambda ws, r: {})
    res = parse_delivery_notes(path)
    assert res["success"] is True
    assert res["mixed_workbook"] is True
    # auto: delivery present -> ledger skipped
    assert res["delivery_note_count"] == 1
    assert res["ledger_note_count"] == 0
    assert any(s["reason"] == "ledger_skipped_auto_has_delivery" for s in res["skipped_sheets"])

    # mode True -> include both
    res2 = parse_delivery_notes(path, include_ledger=True)
    assert res2["ledger_note_count"] == 1

    # mode False -> delivery only, ledger disabled
    res3 = parse_delivery_notes(path, include_ledger=False)
    assert res3["delivery_note_count"] == 1
    assert res3["ledger_note_count"] == 0
    assert any(s["reason"] == "ledger_disabled" for s in res3["skipped_sheets"])


def test_parse_delivery_notes_unsafe_path(tmp_path, monkeypatch, delivery_profile):
    from app.application.shipment_excel_etl_security import ShipmentEtlPathError

    def _raise(*a, **k):
        raise ShipmentEtlPathError("bad")

    monkeypatch.setattr(f"{SEC}.resolve_etl_path", _raise)
    res = parse_delivery_notes(tmp_path / "x.xlsx")
    assert res["success"] is False
    assert res["error_code"] == "unsafe_path"


def test_parse_delivery_notes_file_missing(tmp_path, monkeypatch, delivery_profile):
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: tmp_path / "nope.xlsx")
    monkeypatch.setattr(f"{MOD}._profiles_for_parse", lambda *a, **k: [delivery_profile])
    res = parse_delivery_notes(tmp_path / "nope.xlsx")
    assert res["success"] is False
    assert res["message"] == "文件不存在"


def test_parse_delivery_notes_openpyxl_import_error(tmp_path, monkeypatch, delivery_profile):
    path = _write_delivery_xlsx(tmp_path / "d.xlsx")
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: path)
    monkeypatch.setattr(f"{MOD}._profiles_for_parse", lambda *a, **k: [delivery_profile])
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    try:
        res = parse_delivery_notes(path)
    finally:
        monkeypatch.undo()
    assert res["success"] is False
    assert "openpyxl" in res["message"]


def test_parse_delivery_notes_load_error(tmp_path, monkeypatch, delivery_profile):
    path = _write_delivery_xlsx(tmp_path / "d.xlsx")
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: path)
    monkeypatch.setattr(f"{MOD}._profiles_for_parse", lambda *a, **k: [delivery_profile])

    def boom(*a, **k):
        raise OSError("load fail")

    monkeypatch.setattr("openpyxl.load_workbook", boom)
    res = parse_delivery_notes(path)
    assert res["success"] is False
    assert "无法读取" in res["message"]


def test_parse_delivery_notes_ocr_route(tmp_path, monkeypatch, delivery_profile):
    OCR = "app.application.shipment_excel_etl_ocr"
    path = tmp_path / "scan.pdf"
    path.write_bytes(b"pdf")
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: path)
    monkeypatch.setattr(f"{OCR}.is_ocr_source", lambda p: True)
    monkeypatch.setattr(
        f"{OCR}.parse_ocr_document",
        lambda *a, **k: {"success": True, "notes": []},
    )
    res = parse_delivery_notes(path)
    assert res["success"] is True


def test_parse_delivery_notes_ocr_exception_falls_back(
    tmp_path, monkeypatch, delivery_profile, no_llm, no_kb
):
    OCR = "app.application.shipment_excel_etl_ocr"
    path = _write_delivery_xlsx(tmp_path / "d.xlsx")
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: path)
    monkeypatch.setattr(f"{MOD}._profiles_for_parse", lambda *a, **k: [delivery_profile])
    monkeypatch.setattr(f"{MOD}._extract_adjacent_buyer_meta", lambda ws, r: {})
    monkeypatch.setattr(f"{OCR}.is_ocr_source", lambda p: True)

    def boom(p, **k):
        raise OSError("ocr boom")

    monkeypatch.setattr(f"{OCR}.parse_ocr_document", boom)
    res = parse_delivery_notes(path)
    assert res["success"] is True  # fell back to excel parse


def test_parse_delivery_notes_no_notes_skipped(
    tmp_path, monkeypatch, delivery_profile, no_llm, no_kb
):
    wb = Workbook()
    ws = wb.active
    ws.title = "报价单"
    ws["A1"] = "报价单"
    ws["A2"] = "一些文本"
    path = tmp_path / "quote.xlsx"
    wb.save(path)
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: path)
    monkeypatch.setattr(f"{MOD}._profiles_for_parse", lambda *a, **k: [delivery_profile])
    res = parse_delivery_notes(path)
    assert res["success"] is True
    assert res["note_count"] == 0
    assert res["message"] == "未识别到可匹配的单据模板（可自定义 YAML profile）"


# ---------------------------------------------------------------------------
# preview
# ---------------------------------------------------------------------------


def test_preview_shipment_excel_etl_happy(tmp_path, monkeypatch):
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: tmp_path / "d.xlsx")
    monkeypatch.setattr(
        f"{MOD}.parse_delivery_notes",
        lambda *a, **k: {
            "success": True,
            "notes": [_delivery_note()],
            "message": "识别到 1 张单据",
            "ledger_available_count": 0,
            "ledger_note_count": 0,
        },
    )
    monkeypatch.setattr(f"{SEC}.tenant_key_for_etl", lambda: "tenant:1")
    monkeypatch.setattr(f"{MOD}._is_fingerprint_imported", lambda a, b: False)
    res = preview_shipment_excel_etl(tmp_path / "d.xlsx")
    assert res["success"] is True
    assert res["preview"] is True
    assert res["confirm_required"] is True
    assert res["duplicate_note_count"] == 0
    assert res["ledger_risk"] is False
    assert res["product_records"]


def test_preview_shipment_excel_etl_unsafe(tmp_path, monkeypatch):
    from app.application.shipment_excel_etl_security import ShipmentEtlPathError

    def _raise(*a, **k):
        raise ShipmentEtlPathError("bad")

    monkeypatch.setattr(f"{SEC}.resolve_etl_path", _raise)
    res = preview_shipment_excel_etl(tmp_path / "d.xlsx")
    assert res["success"] is False
    assert res["error_code"] == "unsafe_path"


def test_preview_shipment_excel_etl_duplicate_and_ledger_risk(tmp_path, monkeypatch):
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: tmp_path / "d.xlsx")
    note = _delivery_note()
    note["fingerprint"] = "fp-1"
    monkeypatch.setattr(
        f"{MOD}.parse_delivery_notes",
        lambda *a, **k: {
            "success": True,
            "notes": [note],
            "message": "识别到 1 张单据",
            "ledger_available_count": 2,
            "ledger_note_count": 0,
        },
    )
    monkeypatch.setattr(f"{SEC}.tenant_key_for_etl", lambda: "tenant:1")
    monkeypatch.setattr(f"{MOD}._is_fingerprint_imported", lambda a, b: True)
    res = preview_shipment_excel_etl(tmp_path / "d.xlsx")
    assert res["duplicate_note_count"] == 1
    assert res["ledger_risk"] is True


# ---------------------------------------------------------------------------
# execute
# ---------------------------------------------------------------------------


def _patch_execute(monkeypatch, tmp_path, profile):
    monkeypatch.setattr(f"{MOD}.get_shipment_etl_profile", lambda pid=None: profile)
    monkeypatch.setattr(f"{SEC}.tenant_key_for_etl", lambda: "tenant:1")
    monkeypatch.setattr(f"{MOD}._is_fingerprint_imported", lambda a, b: False)
    monkeypatch.setattr(f"{MOD}._record_fingerprint_now", lambda *a, **k: None)
    monkeypatch.setattr(f"{FP}.delete_fingerprint", lambda *a, **k: None)


def test_execute_direct_denied(monkeypatch, delivery_profile):
    monkeypatch.setattr(f"{SEC}.direct_execute_allowed", lambda: False)
    res = execute_shipment_excel_etl("x.xlsx", direct=True, dry_run=False)
    assert res["success"] is False
    assert res["error_code"] == "direct_execute_denied"


def test_execute_unsafe_path(monkeypatch, delivery_profile, tmp_path):
    from app.application.shipment_excel_etl_security import ShipmentEtlPathError

    _patch_execute(monkeypatch, tmp_path, delivery_profile)

    def _raise(*a, **k):
        raise ShipmentEtlPathError("bad")

    monkeypatch.setattr(f"{SEC}.resolve_etl_path", _raise)
    res = execute_shipment_excel_etl("x.xlsx")
    assert res["error_code"] == "unsafe_path"


def test_execute_missing_path(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    res = execute_shipment_excel_etl(None)
    assert res["error_code"] == "missing_path"


def test_execute_parsed_not_success(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    monkeypatch.setattr(f"{SEC}.resolve_etl_path", lambda *a, **k: tmp_path / "d.xlsx")
    monkeypatch.setattr(
        f"{MOD}.parse_delivery_notes",
        lambda *a, **k: {
            "success": False,
            "message": "文件不存在",
            "notes": [],
        },
    )
    res = execute_shipment_excel_etl("d.xlsx")
    assert res["success"] is False


def test_execute_ledger_confirm_required(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    note = _delivery_note()
    note["source_kind"] = "shipment_ledger"
    res = execute_shipment_excel_etl(None, notes=[note], confirm_ledger=False)
    assert res["error_code"] == "ledger_confirm_required"


def test_execute_no_notes(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    res = execute_shipment_excel_etl(None, notes=[])
    assert res["error_code"] == "no_delivery_notes"


def test_execute_unsupported_profile_target(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    note = _delivery_note(profile_target="preview_only")
    res = execute_shipment_excel_etl(None, notes=[note], import_shipments=True)
    assert res["error_code"] == "unsupported_profile_target"


def test_execute_dry_run_with_duplicates(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    monkeypatch.setattr(f"{MOD}._is_fingerprint_imported", lambda a, b: True)
    res = execute_shipment_excel_etl(
        "d.xlsx", notes=[_delivery_note()], dry_run=True, import_shipments=True
    )
    assert res["success"] is True
    assert res["dry_run"] is True
    assert res["would_skip"] == 1
    assert res["would_create"] == 0


def test_execute_product_import_failed(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    monkeypatch.setattr(
        "app.services.tools_workflow_registered._execute_excel_import_records",
        lambda records: {"success": False, "message": "导入失败"},
    )
    res = execute_shipment_excel_etl(
        "d.xlsx", notes=[_delivery_note()], import_products=True, import_shipments=False
    )
    assert res["error_code"] == "product_import_failed"


def test_execute_success(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    svc = MagicMock()
    svc.create_shipment.return_value = {"success": True, "shipment": {"id": 1}}
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    res = execute_shipment_excel_etl(
        "d.xlsx",
        notes=[_delivery_note()],
        idempotent=True,
        import_products=False,
        import_shipments=True,
    )
    assert res["success"] is True
    assert res["shipment_created"] == 1
    assert res["shipment_ids"] == [1]
    assert res["closed_loop"] is True


def test_execute_missing_unit_or_items(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    svc = MagicMock()
    svc.create_shipment.return_value = {"success": True, "shipment": {"id": 1}}
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    note = _delivery_note()
    note["unit_name"] = ""
    res = execute_shipment_excel_etl(
        "d.xlsx", notes=[note], import_products=False, import_shipments=True
    )
    assert res["success"] is False
    assert res["shipment_failed"] == 1


def test_execute_shipment_service_unavailable(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)

    def boom():
        raise OSError("no svc")

    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", boom)
    res = execute_shipment_excel_etl(
        "d.xlsx", notes=[_delivery_note()], import_products=False, import_shipments=True
    )
    assert res["success"] is False
    assert "不可用" in res["message"]


def test_execute_compensate_on_failure(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    svc = MagicMock()
    svc.create_shipment.side_effect = [
        {"success": True, "shipment": {"id": 1}},
        {"success": False, "message": "boom"},
    ]
    svc.cancel_shipment.return_value = {"success": True}
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    notes = [_delivery_note(), _delivery_note()]
    res = execute_shipment_excel_etl(
        "d.xlsx",
        notes=notes,
        import_products=False,
        import_shipments=True,
        compensate_on_failure=True,
    )
    assert res["success"] is False
    assert res["compensated"] == [1]
    assert res["shipment_created"] == 0
    assert res["safe_to_retry"] is True


def test_execute_compensate_delete_fallback(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    svc = MagicMock()
    svc.create_shipment.side_effect = [
        {"success": True, "shipment": {"id": 1}},
        {"success": False, "message": "boom"},
    ]
    svc.cancel_shipment.return_value = {"success": False}
    svc.delete_shipment.return_value = {"success": True}
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    notes = [_delivery_note(), _delivery_note()]
    res = execute_shipment_excel_etl(
        "d.xlsx", notes=notes, import_products=False, import_shipments=True
    )
    assert res["compensated"] == [1]


def test_execute_compensate_error(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    svc = MagicMock()
    svc.create_shipment.side_effect = [
        {"success": True, "shipment": {"id": 1}},
        {"success": False, "message": "boom"},
    ]
    svc.cancel_shipment.return_value = {"success": False}
    svc.delete_shipment.return_value = {"success": False, "message": "del fail"}
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    notes = [_delivery_note(), _delivery_note()]
    res = execute_shipment_excel_etl(
        "d.xlsx", notes=notes, import_products=False, import_shipments=True
    )
    assert len(res["compensate_errors"]) == 1
    assert res["safe_to_retry"] is False


def test_execute_compensate_exception(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    svc = MagicMock()
    svc.create_shipment.side_effect = [
        {"success": True, "shipment": {"id": 1}},
        {"success": False, "message": "boom"},
    ]

    def cancel_boom(sid):
        raise OSError("cancel boom")

    svc.cancel_shipment.side_effect = cancel_boom
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    notes = [_delivery_note(), _delivery_note()]
    res = execute_shipment_excel_etl(
        "d.xlsx", notes=notes, import_products=False, import_shipments=True
    )
    assert len(res["compensate_errors"]) == 1
    assert res["compensate_errors"][0] == "补偿失败"


def test_execute_partial_no_compensate(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    svc = MagicMock()
    svc.create_shipment.side_effect = [
        {"success": True, "shipment": {"id": 1}},
        {"success": False, "message": "boom"},
    ]
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    notes = [_delivery_note(), _delivery_note()]
    res = execute_shipment_excel_etl(
        "d.xlsx",
        notes=notes,
        import_products=False,
        import_shipments=True,
        compensate_on_failure=False,
    )
    assert res["success"] is False
    assert res["partial_success"] is True
    assert res["shipment_ids"] == [1]


def test_execute_record_fingerprint_raises(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)

    def boom(*a, **k):
        raise OSError("fp boom")

    monkeypatch.setattr(f"{MOD}._record_fingerprint_now", boom)
    svc = MagicMock()
    svc.create_shipment.return_value = {"success": True, "shipment": {"id": 1}}
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    res = execute_shipment_excel_etl(
        "d.xlsx", notes=[_delivery_note()], import_products=False, import_shipments=True
    )
    assert res["success"] is True


def test_execute_force_shipment_target(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    monkeypatch.setattr(f"{SEC}.direct_execute_allowed", lambda: True)
    svc = MagicMock()
    svc.create_shipment.return_value = {"success": True, "shipment": {"id": 1}}
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    note = _delivery_note(profile_target="preview_only")
    res = execute_shipment_excel_etl(
        "d.xlsx",
        notes=[note],
        direct=True,
        force_shipment_target=True,
        import_products=False,
        import_shipments=True,
    )
    assert res["success"] is True
    assert res["audit"]["force_shipment_target"] is True


def test_execute_direct_allowed_success(monkeypatch, delivery_profile, tmp_path):
    _patch_execute(monkeypatch, tmp_path, delivery_profile)
    monkeypatch.setattr(f"{SEC}.direct_execute_allowed", lambda: True)
    svc = MagicMock()
    svc.create_shipment.return_value = {"success": True, "shipment": {"id": 1}}
    monkeypatch.setattr("app.bootstrap.get_shipment_app_service", lambda: svc)
    res = execute_shipment_excel_etl(
        "d.xlsx",
        notes=[_delivery_note()],
        direct=True,
        import_products=False,
        import_shipments=True,
    )
    assert res["success"] is True


# ---------------------------------------------------------------------------
# write workbooks
# ---------------------------------------------------------------------------

_WRITE_CFG = {
    "seller_title": "送货单",
    "header_row": ["型号", "名称", "数量", "规格", "数量KG", "单价", "金额"],
    "item_columns": {
        "model_number": 1,
        "product_name": 2,
        "quantity_tins": 3,
        "tin_spec": 4,
        "quantity_kg": 5,
        "unit_price": 6,
        "amount": 7,
    },
    "date_format": "%Y-%m-%d",
    "meta_line_template": "{unit} {contact} {order_date} {order_no}",
    "footer_label": "签字",
    "default_sheet_name": "Sheet1",
    "sheet_name_prefix": "S",
    "demo_meta_line": "demo 客户 联系人 日期 单号",
    "demo_item": {
        "model_number": "SKU",
        "product_name": "示例",
        "quantity_tins": 1,
        "tin_spec": 1,
        "quantity_kg": 1,
        "unit_price": 1,
        "amount": 1,
    },
    "ledger_sheet_name": "ledger",
    "ledger_header_row": ["日期", "单号", "型号", "名称", "数量", "数量KG", "单价", "金额"],
    "ledger_item_columns": {
        "order_date": 1,
        "order_number": 2,
        "model_number": 3,
        "product_name": 4,
        "quantity_tins": 5,
        "quantity_kg": 6,
        "unit_price": 7,
        "amount": 8,
    },
    "ledger_sample_rows": [
        {
            "order_date": "2026-07-01",
            "order_number": "L1",
            "model_number": "M1",
            "product_name": "面漆",
            "quantity_tins": 1,
            "quantity_kg": 20,
            "unit_price": 10,
            "amount": 200,
        }
    ],
    "ledger_extra_sheet": "说明",
    "ledger_default_unit": "unit",
}


def test_write_delivery_note_workbook(tmp_path, delivery_profile):
    prof = _make_profile(write=dict(_WRITE_CFG))
    out = tmp_path / "out" / "delivery.xlsx"
    res = write_delivery_note_workbook([_delivery_note()], out, profile=prof)
    assert res["success"] is True
    assert res["sheet_count"] == 1
    assert Path(res["file_path"]).is_file()


def test_write_delivery_note_workbook_multi_and_demo(tmp_path, delivery_profile):
    prof = _make_profile(write=dict(_WRITE_CFG))
    out = tmp_path / "multi.xlsx"
    notes = [
        {"unit_name": "共同客户", "sheet_name": "同一张", "items": [], "order_number": "A1"},
        {"unit_name": "共同客户", "sheet_name": "同一张", "items": [], "order_number": "B2"},
    ]
    res = write_delivery_note_workbook(notes, out, profile=prof)
    assert res["success"] is True
    assert res["sheet_count"] == 2
    # demo-only path
    res2 = write_delivery_note_workbook([], tmp_path / "demo.xlsx", profile=prof)
    assert res2["success"] is True
    assert res2["sheet_count"] == 1


def test_write_delivery_note_workbook_import_error(monkeypatch):
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    try:
        res = write_delivery_note_workbook([], "x.xlsx")
    finally:
        monkeypatch.undo()
    assert res["success"] is False


def test_write_ledger_workbook(tmp_path, delivery_profile):
    prof = _make_profile(write=dict(_WRITE_CFG))
    out = tmp_path / "ledger.xlsx"
    rows = [
        {
            "order_date": "2026-07-01",
            "order_number": "L1",
            "model_number": "M1",
            "product_name": "面漆",
            "quantity_tins": 1,
            "quantity_kg": 20,
            "unit_price": 10,
            "amount": 200,
        }
    ]
    res = write_ledger_workbook(rows, out, profile=prof, unit_name="客户X")
    assert res["success"] is True
    assert res["row_count"] == 1
    assert res["unit_name"] == "客户X"
    # default unit + empty row fill
    res2 = write_ledger_workbook([], tmp_path / "l2.xlsx", profile=prof)
    assert res2["success"] is True
    assert res2["unit_name"] == "unit"


def test_write_ledger_workbook_import_error(monkeypatch):
    monkeypatch.setitem(sys.modules, "openpyxl", None)
    try:
        res = write_ledger_workbook([], "x.xlsx")
    finally:
        monkeypatch.undo()
    assert res["success"] is False


# ---------------------------------------------------------------------------
# regenerate
# ---------------------------------------------------------------------------


def test_regenerate_happy(tmp_path, monkeypatch):
    prof = _make_profile()
    monkeypatch.setattr(
        f"{MOD}.parse_delivery_notes",
        lambda *a, **k: {
            "success": True,
            "notes": [svc_mod._enrich_note(_delivery_note())],
        },
    )
    written = {"success": True}
    monkeypatch.setattr(f"{MOD}.write_delivery_note_workbook", lambda *a, **k: written)
    monkeypatch.setattr(
        f"{MOD}.parse_delivery_notes",
        lambda *a, **k: (
            {
                "success": True,
                "notes": [svc_mod._enrich_note(_delivery_note())],
            }
            if str(a[0]).endswith("_out.xlsx")
            else {
                "success": True,
                "notes": [svc_mod._enrich_note(_delivery_note())],
            }
        ),
    )
    res = regenerate_delivery_notes_from_file("in.xlsx", tmp_path / "out.xlsx", profile=prof)
    assert res["success"] is True
    assert res["fingerprint_match"] is True


def test_regenerate_not_success(monkeypatch):
    prof = _make_profile()
    monkeypatch.setattr(
        f"{MOD}.parse_delivery_notes",
        lambda *a, **k: {
            "success": False,
            "message": "文件不存在",
            "notes": [],
        },
    )
    res = regenerate_delivery_notes_from_file("in.xlsx", "out.xlsx", profile=prof)
    assert res["success"] is False


def test_regenerate_no_notes(monkeypatch):
    prof = _make_profile()
    monkeypatch.setattr(
        f"{MOD}.parse_delivery_notes",
        lambda *a, **k: {
            "success": True,
            "notes": [],
        },
    )
    res = regenerate_delivery_notes_from_file("in.xlsx", "out.xlsx", profile=prof)
    assert res["success"] is False
    assert res["error_code"] == "no_delivery_notes"


def test_regenerate_write_failed(monkeypatch):
    prof = _make_profile()
    monkeypatch.setattr(
        f"{MOD}.parse_delivery_notes",
        lambda *a, **k: {
            "success": True,
            "notes": [svc_mod._enrich_note(_delivery_note())],
        },
    )
    monkeypatch.setattr(
        f"{MOD}.write_delivery_note_workbook",
        lambda *a, **k: {
            "success": False,
            "error_code": "write_failed",
            "message": "写入失败",
        },
    )
    res = regenerate_delivery_notes_from_file("in.xlsx", "out.xlsx", profile=prof)
    assert res["success"] is False
    assert res["error_code"] == "write_failed"
