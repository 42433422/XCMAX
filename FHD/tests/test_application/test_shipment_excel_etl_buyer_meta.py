"""相邻单元格买家抽取 + 弱客户名规则。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from app.application.shipment_etl_profile import clear_profile_cache
from app.application.shipment_excel_etl_app_service import parse_delivery_notes


def test_english_to_adjacent_buyer(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    clear_profile_cache()
    from app.application.excel_etl_kb import reset_excel_etl_kb_for_tests

    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")

    path = tmp_path / "do.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "DO"
    ws["A5"] = "To:"
    ws["C5"] = "Semiconductor Technologies Pte Ltd"
    ws["A8"] = "SKU"
    ws["B8"] = "Description"
    ws["C8"] = "Qty"
    ws["A9"] = "A1"
    ws["B9"] = "Cable"
    ws["C9"] = 2
    wb.save(path)

    out = parse_delivery_notes(path, include_ledger=False)
    assert out["success"] is True
    note = out["notes"][0]
    assert "Semiconductor" in str(note.get("unit_name") or "")


def test_pi_bill_to_not_buyer_po(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    clear_profile_cache()
    from app.application.excel_etl_kb import reset_excel_etl_kb_for_tests

    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")

    fixture = (
        Path(__file__).resolve().parents[1]
        / "fixtures"
        / "network_forms"
        / "net_PI_sample.xlsx"
    )
    if not fixture.is_file():
        return
    out = parse_delivery_notes(fixture, include_ledger=False)
    assert out["success"] is True
    unit = str(out["notes"][0].get("unit_name") or "")
    assert "Dukjil" in unit
    assert "PO:" not in unit


def test_purchase_unit_label(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    clear_profile_cache()
    from app.application.excel_etl_kb import reset_excel_etl_kb_for_tests

    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")

    fixture = (
        Path(__file__).resolve().parents[1]
        / "fixtures"
        / "network_forms"
        / "form_采购订单_七彩乐园.xlsx"
    )
    if not fixture.is_file():
        return
    out = parse_delivery_notes(fixture, include_ledger=False)
    assert out["success"] is True
    unit = str(out["notes"][0].get("unit_name") or "")
    assert "七彩乐园" in unit
