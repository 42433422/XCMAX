"""送货单 ETL LLM 辅助：开关 / 校验 / 缓存 / 低置信补列。"""

from __future__ import annotations

import pytest

from app.application.shipment_excel_etl_app_service import parse_delivery_notes
from app.application.shipment_excel_etl_llm import (
    SheetProbe,
    assist_sheet_layout,
    clear_assist_cache,
    llm_assist_enabled,
    needs_llm_assist,
)
from app.infrastructure.llm.structured_output import StructuredResult


@pytest.fixture(autouse=True)
def _reset_llm_env(monkeypatch, tmp_path):
    clear_assist_cache()
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    monkeypatch.setenv("FHD_EXCEL_ETL_HEURISTIC", "0")
    monkeypatch.setenv("FHD_EXCEL_ETL_KB_PATH", str(tmp_path / "kb.json"))
    monkeypatch.delenv("FHD_EXCEL_ETL_ALLOW_BUILTIN", raising=False)
    from app.application.excel_etl_kb import reset_excel_etl_kb_for_tests
    from app.application.shipment_etl_profile import clear_profile_cache

    reset_excel_etl_kb_for_tests(tmp_path / "kb.json")
    clear_profile_cache()
    yield
    clear_assist_cache()
    clear_profile_cache()


def test_llm_disabled_by_env(monkeypatch):
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    assert llm_assist_enabled() is False
    probe = SheetProbe(
        profile_id="universal",
        sheet_title="S1",
        probe_rows=[],
        candidate_headers=[],
        max_row=5,
        max_col=5,
    )
    out = assist_sheet_layout(probe)
    assert out.used_llm is False
    assert out.reason == "llm_disabled"


def test_needs_llm_assist_gray_and_missing_columns():
    need, reason = needs_llm_assist(
        delivery_score=50,
        ledger_score=0,
        min_score=60,
        header_row=3,
        mapping={"product_name": 4},
        meta={"unit_name": "A"},
        prefer_kind="delivery_note",
    )
    assert need is True
    assert reason == "delivery_score_gray"

    need2, reason2 = needs_llm_assist(
        delivery_score=80,
        ledger_score=0,
        min_score=60,
        header_row=3,
        mapping={},
        meta={"unit_name": "A"},
        prefer_kind="delivery_note",
    )
    assert need2 is True
    assert reason2 == "delivery_columns_incomplete"

    need3, reason3 = needs_llm_assist(
        delivery_score=80,
        ledger_score=0,
        min_score=60,
        header_row=3,
        mapping={"model_number": 1, "product_name": 4},
        meta={"unit_name": "客户"},
        prefer_kind="delivery_note",
    )
    assert need3 is False
    assert reason3 == "rules_confident"


def test_invalid_llm_columns_are_rejected(monkeypatch):
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "1")
    calls: list[int] = []

    def _fake_complete(messages, **kwargs):
        calls.append(1)
        return StructuredResult(
            data={
                "source_kind": "delivery_note",
                "header_row": 3,
                "columns": {"model_number": 99, "product_name": 4},
                "meta": {"unit_name": "X"},
                "confidence": 0.9,
                "reason": "bad_col",
            },
            attempts=1,
            repaired=False,
        )

    monkeypatch.setattr(
        "app.infrastructure.llm.structured_output.complete_structured_sync",
        _fake_complete,
    )
    probe = SheetProbe(
        profile_id="universal",
        sheet_title="S1",
        probe_rows=[{"row": 1, "cells": [{"col": 1, "text": "送货单"}]}],
        candidate_headers=[
            {
                "row": 3,
                "cells": [
                    {"col": 1, "header": "SKU", "samples": ["A"]},
                    {"col": 4, "header": "Name", "samples": ["漆"]},
                ],
            }
        ],
        max_row=10,
        max_col=9,
    )
    # model_number=99 not in allowed cols → validate fails because only product_name kept
    # wait: product_name=4 is allowed, so delivery has product_name → ok=True with only product_name
    out = assist_sheet_layout(probe)
    assert out.used_llm is True
    assert out.ok is True
    assert "model_number" not in out.columns
    assert out.columns.get("product_name") == 4


def test_llm_assist_fills_alt_headers(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "1")
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM_TIMEOUT", "5")
    calls: list[int] = []

    def _fake_complete(messages, **kwargs):
        calls.append(1)
        return StructuredResult(
            data={
                "source_kind": "delivery_note",
                "header_row": 3,
                "columns": {
                    "model_number": 1,
                    "product_name": 4,
                    "quantity_tins": 5,
                    "tin_spec": 6,
                    "quantity_kg": 7,
                    "unit_price": 8,
                    "amount": 9,
                },
                "meta": {
                    "unit_name": "LLM客户",
                    "contact_person": "王",
                    "order_date": "2026年07月25日",
                    "order_number": "LLM-1",
                    "title": "ALT 送货单",
                },
                "confidence": 0.92,
                "reason": "mapped_alt_headers",
            },
            attempts=1,
            repaired=False,
        )

    monkeypatch.setattr(
        "app.infrastructure.llm.structured_output.complete_structured_sync",
        _fake_complete,
    )

    from openpyxl import Workbook

    path = tmp_path / "alt_llm.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "N1"
    # 故意使用知识库同义词未覆盖的表头，逼出 LLM 补列
    ws["A1"] = "Factory Delivery Note"
    ws["A2"] = "Buyer：LLM客户     Contact：王        2026年07月25日         No：LLM-1"
    ws["A3"] = "PartCode"
    ws["D3"] = "GoodsTitle"
    ws["E3"] = "PcsCount"
    ws["F3"] = "NetKgEach"
    ws["G3"] = "TotalKg"
    ws["H3"] = "UnitFee"
    ws["I3"] = "LineSum"
    ws["A4"] = "SKU-1"
    ws["D4"] = "清漆"
    ws["E4"] = 2
    ws["F4"] = 25
    ws["G4"] = 50
    ws["H4"] = 10
    ws["I4"] = 500
    wb.save(path)
    wb.close()

    # 无 LLM 时通用同义词不应识别这些表头
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "0")
    off = parse_delivery_notes(path, include_ledger=False)
    assert off["note_count"] == 0

    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "1")
    clear_assist_cache()
    first = parse_delivery_notes(path, include_ledger=False)
    assert first["success"] is True
    assert first["note_count"] == 1
    note = first["notes"][0]
    assert note["unit_name"] == "LLM客户"
    assert note["items"][0]["model_number"] == "SKU-1"
    assert note["items"][0]["quantity_tins"] == 2
    assert note.get("assist", {}).get("used_llm") is True
    assert first.get("assist", {}).get("used_llm") is True
    assert len(calls) == 1

    # 二次解析走缓存
    second = parse_delivery_notes(path, include_ledger=False)
    assert second["note_count"] == 1
    assert second["notes"][0].get("assist", {}).get("cache_hit") is True
    assert len(calls) == 1


def test_llm_failure_degrades_quietly(tmp_path, monkeypatch):
    monkeypatch.setenv("FHD_SHIPMENT_ETL_LLM", "1")

    def _boom(*a, **k):
        raise RuntimeError("llm down")

    monkeypatch.setattr(
        "app.infrastructure.llm.structured_output.complete_structured_sync",
        _boom,
    )
    from app.application.shipment_excel_etl_app_service import write_delivery_note_workbook

    path = tmp_path / "ok.xlsx"
    write_delivery_note_workbook(
        [
            {
                "unit_name": "稳定客户",
                "contact_person": "测",
                "order_date": "2026年07月25日",
                "order_number": "S-1",
                "items": [
                    {
                        "model_number": "M1",
                        "product_name": "漆",
                        "quantity_tins": 1,
                        "tin_spec": 20,
                        "quantity_kg": 20,
                        "unit_price": 9,
                        "amount": 180,
                    }
                ],
            }
        ],
        path,
    )
    # 默认版式规则可解析；LLM 故障不应阻断
    out = parse_delivery_notes(path, include_ledger=False)
    assert out["success"] is True
    assert out["note_count"] == 1
    assert out["notes"][0]["unit_name"] == "稳定客户"
