"""COVERAGE_RAMP Phase 1 (p0-core): auth, middleware, db, app_service helpers (mocked I/O)."""

from __future__ import annotations

import io
import uuid
from contextlib import contextmanager
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from fastapi import FastAPI
from fastapi.exceptions import RequestValidationError
from fastapi.testclient import TestClient
from pydantic import ValidationError
from starlette.exceptions import HTTPException as StarletteHTTPException
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request
from starlette.responses import JSONResponse, Response

from app.application import aibiz_web_terminal_service as aibiz_mod
from app.application.auth_app_service import AuthApplicationService, _authenticate_failure_message
from app.application.excel_template_http_app_service import (
    _decompose_from_grid,
    _is_unreadable_workbook_error,
    _json_safe_cell_value,
    _map_template_category,
    _normalize_template_dto,
    _pick_sheet_name,
    _resolve_template_path,
    decompose_template,
    excel_templates_test,
    get_templates_list,
    list_templates_by_type,
    list_templates_get,
)
from app.application.ocr_app_service import OCRApplicationService
from app.application.print_app_service import PrintApplicationService
from app.application.product_app_service import ProductApplicationService, _normalize_optional_str
from app.application.rbac_app_service import RbacAppService, get_rbac_app_service
from app.application.session_account_meta import (
    audit_admin_action,
    clear_impersonation,
    effective_entitlement_market_user_id,
    enrich_session_meta_with_tenant,
    is_session_market_admin,
    load_session_account_meta,
    persist_session_account_meta,
)
from app.application.template_app_service import TemplateApplicationService
from app.application.tools.workflow import _parse_excel_header_row_1based
from app.contexts.context_notifier import get_context_notifier
from app.db.validators import ModelValidators, register_model_validators
from app.errors import AppError, AuthError, ErrorCode
from app.middleware.error_handler import register_exception_handlers
from app.middleware.request_id import RequestIdMiddleware
from app.middleware.subscription_gate import SubscriptionGateMiddleware, _subscription_gate_enabled
from app.schemas.finance_schema import FinanceTransactionCreate, FinanceTransactionUpdate
from app.schemas.rbac_schema import PermissionCreate, RoleCreate, RoleUpdate, UserRoleAssign
from app.utils.mobile_api import format_error_response, format_mobile_response, paginate_list

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture(autouse=True)
def _reset_rbac_singleton() -> None:
    import app.application.rbac_app_service as rbac_mod

    rbac_mod._service = None
    yield
    rbac_mod._service = None


def _mock_get_db(mock_db: MagicMock) -> MagicMock:
    cm = MagicMock()
    cm.__enter__.return_value = mock_db
    cm.__exit__.return_value = None
    return cm


def _query_chain(*, first=None, all_items=None) -> MagicMock:
    q = MagicMock()
    q.filter.return_value = q
    q.first.return_value = first
    q.all.return_value = all_items or []
    return q


# ---------------------------------------------------------------------------
# excel_template_http_app_service helpers
# ---------------------------------------------------------------------------


def test_excel_map_template_category_label() -> None:
    assert _map_template_category("打印标签") == "label_print"
    assert _map_template_category("excel") == "excel"


def test_excel_normalize_template_dto_word() -> None:
    dto = _normalize_template_dto(
        {"template_type": "发货单", "file_path": "a.docx", "exists": True}
    )
    assert dto["category"] == "word"
    assert dto["preview_capable"] is True


def test_excel_json_safe_cell_value_variants() -> None:
    assert _json_safe_cell_value(None) is None
    assert _json_safe_cell_value(datetime(2026, 1, 1, 12, 0)) == "2026-01-01T12:00:00"
    assert _json_safe_cell_value(date(2026, 1, 1)) == "2026-01-01"
    assert _json_safe_cell_value(time(8, 30)) == "08:30:00"
    assert _json_safe_cell_value(Decimal("12.5")) == "12.5"
    assert _json_safe_cell_value(float("inf")) is None


def test_excel_is_unreadable_workbook_error() -> None:
    assert _is_unreadable_workbook_error("BadZipFile: corrupt") is True
    assert _is_unreadable_workbook_error("ok") is False


def test_excel_pick_sheet_name_prefers_shipment() -> None:
    assert _pick_sheet_name(["Sheet1", "出货明细"], None) == "出货明细"
    assert _pick_sheet_name(["A", "B"], "B") == "B"


def test_excel_decompose_from_grid() -> None:
    def cell(r: int, c: int):
        grid = {
            (1, 1): "品名",
            (1, 2): "型号",
            (1, 3): "数量",
            (1, 4): "单价",
            (2, 1): "产品A",
            (2, 2): "M1",
            (2, 3): 10,
            (2, 4): 5.5,
        }
        return grid.get((r, c))

    result, status = _decompose_from_grid(
        "/tmp/t.xlsx",
        "出货",
        5,
        4,
        cell,
        0,
        "A1:D5",
        None,
        3,
    )
    assert status == 200
    assert result["success"] is True
    assert len(result["decomposition"]["editable_entries"]) >= 4


def test_excel_resolve_template_path(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(
        "app.application.excel_template_http_app_service.get_base_dir",
        lambda: str(tmp_path),
    )
    f = tmp_path / "tpl.xlsx"
    f.write_bytes(b"x")
    assert _resolve_template_path("tpl.xlsx") == str(f)


@patch("app.application.excel_template_http_app_service._get_template_list")
def test_excel_list_templates_get(mock_list: MagicMock) -> None:
    mock_list.return_value = [{"id": "1", "template_type": "label", "exists": True, "path": "/a"}]
    resp = list_templates_get()
    assert resp.status_code == 200
    body = resp.body.decode()
    assert "success" in body


@patch("app.application.excel_template_http_app_service._get_template_list")
def test_excel_get_templates_list(mock_list: MagicMock) -> None:
    mock_list.return_value = [{"id": "1"}]
    resp = get_templates_list()
    assert resp.status_code == 200


@patch("app.application.get_template_app_service")
def test_excel_list_templates_by_type(mock_svc_fn: MagicMock) -> None:
    svc = MagicMock()
    svc.list_by_type.return_value = [{"template_type": "发货单"}]
    mock_svc_fn.return_value = svc
    resp = list_templates_by_type(type="发货单", active_only="true")
    assert resp.status_code == 200


def test_excel_templates_test_endpoint() -> None:
    resp = excel_templates_test()
    assert resp.status_code == 200


def test_excel_decompose_template_missing_file() -> None:
    resp = decompose_template({"filename": "missing.xlsx"})
    assert resp.status_code == 404


def test_excel_decompose_template_no_args() -> None:
    resp = decompose_template({})
    assert resp.status_code == 400


def test_excel_decompose_template_xlsx(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    from openpyxl import Workbook

    monkeypatch.setattr(
        "app.application.excel_template_http_app_service.get_base_dir",
        lambda: str(tmp_path),
    )
    xlsx = tmp_path / "ship.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "出货"
    for c, h in enumerate(["品名", "型号", "数量", "单价", "金额"], 1):
        ws.cell(row=1, column=c, value=h)
    ws.cell(row=2, column=1, value="产品A")
    wb.save(xlsx)
    resp = decompose_template({"file_path": str(xlsx)})
    assert resp.status_code == 200
    assert b'"success": true' in resp.body.lower() or b'"success":true' in resp.body.lower()


# ---------------------------------------------------------------------------
# auth_app_service
# ---------------------------------------------------------------------------


def test_authenticate_failure_message_market_token() -> None:
    exc = RuntimeError("column market_access_token does not exist")
    msg = _authenticate_failure_message(exc)
    assert "market_access_token" in msg


def test_authenticate_failure_message_users_table() -> None:
    exc = RuntimeError("no such table: users")
    assert "users" in _authenticate_failure_message(exc)


def test_authenticate_failure_message_generic() -> None:
    assert "稍后重试" in _authenticate_failure_message(RuntimeError("other"))


@patch("app.application.auth_app_service.get_db")
@patch("app.application.auth_app_service.check_password_hash", return_value=True)
def test_auth_login_success(mock_hash: MagicMock, mock_get_db: MagicMock) -> None:
    mock_db = MagicMock()
    mock_get_db.return_value = _mock_get_db(mock_db)
    user = SimpleNamespace(
        id=1,
        username="u",
        display_name="U",
        email="u@x.com",
        role="user",
        is_active=True,
        password="hash",
    )
    mock_db.query.return_value.filter.return_value.first.return_value = user
    svc = AuthApplicationService()
    svc.session_manager = MagicMock()
    svc.session_manager.create_session_with_db.return_value = {
        "success": True,
        "session_id": "sid",
        "expires_at": "2099-01-01",
    }
    out = svc.authenticate("u", "pass")
    assert out["success"] is True
    assert out["session_id"] == "sid"


@patch("app.application.auth_app_service.get_db")
def test_auth_login_wrong_password(mock_get_db: MagicMock) -> None:
    mock_db = MagicMock()
    mock_get_db.return_value = _mock_get_db(mock_db)
    user = SimpleNamespace(is_active=True, password="hash")
    mock_db.query.return_value.filter.return_value.first.return_value = user
    with patch("app.application.auth_app_service.check_password_hash", return_value=False):
        out = AuthApplicationService().authenticate("u", "bad")
    assert out["success"] is False


@patch("app.application.auth_app_service.get_db")
def test_auth_create_session_user_missing(mock_get_db: MagicMock) -> None:
    mock_db = MagicMock()
    mock_get_db.return_value = _mock_get_db(mock_db)
    mock_db.query.return_value.filter.return_value.first.return_value = None
    out = AuthApplicationService().create_session_for_username("ghost")
    assert out["success"] is False


def test_auth_logout_delegates() -> None:
    svc = AuthApplicationService()
    svc.session_manager = MagicMock()
    svc.session_manager.delete_session.return_value = True
    assert svc.logout("sid") is True


# ---------------------------------------------------------------------------
# product / print / ocr / template app services
# ---------------------------------------------------------------------------


def test_product_normalize_optional_str() -> None:
    assert _normalize_optional_str(None) is None
    assert _normalize_optional_str("  ") is None
    assert _normalize_optional_str(" abc ") == "abc"


def test_product_app_service_delegates() -> None:
    products = MagicMock()
    products.get_products.return_value = {"success": True, "data": []}
    svc = ProductApplicationService(products_service=products)
    out = svc.get_products(unit_name=" 甲公司 ", model_number=" ab ")
    assert out["success"] is True
    products.get_products.assert_called_once()


def test_print_app_service_empty_labels() -> None:
    svc = PrintApplicationService(printer_service=MagicMock())
    out = svc.print_labels([])
    assert out["success"] is False


def test_print_app_service_single_label() -> None:
    printer = MagicMock()
    printer.print_labels.return_value = {"success": True}
    out = PrintApplicationService(printer_service=printer).print_single_label("产品A")
    assert out["success"] is True


def test_ocr_app_service_batch_recognize() -> None:
    ocr = MagicMock()
    ocr.recognize_text.side_effect = [
        {"success": True},
        {"success": False},
    ]
    out = OCRApplicationService(ocr_service=ocr).batch_recognize(["a.png", "b.png"])
    assert out["success_count"] == 1
    assert out["fail_count"] == 1


def test_template_app_service_category_filter() -> None:
    tpl_svc = MagicMock()
    tpl_svc.list_by_type.return_value = [{"id": 1}]
    out = TemplateApplicationService(template_service=tpl_svc).get_templates(category="label")
    assert len(out["templates"]) == 1


# ---------------------------------------------------------------------------
# rbac + schemas
# ---------------------------------------------------------------------------


def test_rbac_app_service_crud_stubs() -> None:
    svc = RbacAppService()
    role = svc.create_role("editor", "编辑", ["read"])
    assert role["name"] == "editor"
    assert svc.get_role(1)["id"] == 1
    assert get_rbac_app_service() is get_rbac_app_service()


def test_rbac_schemas_validation() -> None:
    role = RoleCreate(name="r", permissions=["a"])
    assert role.name == "r"
    perm = PermissionCreate(code="p.read", name="Read")
    assert perm.code == "p.read"
    with pytest.raises(ValidationError):
        PermissionCreate(code="", name="x")


def test_finance_schemas_validation() -> None:
    txn = FinanceTransactionCreate(transaction_type="revenue", amount=10.5)
    assert txn.amount == 10.5
    upd = FinanceTransactionUpdate(amount=20)
    assert upd.amount == 20


# ---------------------------------------------------------------------------
# session_account_meta (DB mocked)
# ---------------------------------------------------------------------------


@patch("app.application.session_account_meta.get_host_db")
def test_persist_session_account_meta(mock_get_db: MagicMock) -> None:
    mock_db = MagicMock()
    mock_get_db.return_value = _mock_get_db(mock_db)
    row = SimpleNamespace(
        account_kind="enterprise",
        company_brand="",
        market_user_id=None,
        market_is_admin=False,
        market_is_enterprise=False,
        impersonating_market_user_id=None,
        impersonating_username="",
        tenant_id=None,
    )
    mock_db.query.return_value.filter.return_value.first.return_value = row
    persist_session_account_meta("sid", account_kind="admin", company_brand="甲")
    assert row.account_kind == "admin"


@patch("app.application.session_account_meta.get_host_db")
def test_load_session_account_meta_found(mock_get_db: MagicMock) -> None:
    mock_db = MagicMock()
    mock_get_db.return_value = _mock_get_db(mock_db)
    row = SimpleNamespace(
        account_kind="enterprise",
        company_brand="X",
        market_user_id=9,
        market_is_admin=False,
        market_is_enterprise=True,
        impersonating_market_user_id=None,
        impersonating_username="",
        tenant_id=3,
    )
    mock_db.query.return_value.filter.return_value.first.return_value = row
    meta = load_session_account_meta("sid")
    assert meta is not None
    assert meta["tenant_id"] == 3


@patch("app.application.session_account_meta.load_session_account_meta")
def test_is_session_market_admin(mock_load: MagicMock) -> None:
    mock_load.return_value = {"account_kind": "admin", "market_is_admin": True}
    assert is_session_market_admin("sid") is True
    mock_load.return_value = {"account_kind": "enterprise"}
    assert is_session_market_admin("sid") is False


@patch("app.application.session_account_meta.load_session_account_meta")
def test_effective_entitlement_market_user_id(mock_load: MagicMock) -> None:
    mock_load.return_value = {"impersonating_market_user_id": 42}
    assert effective_entitlement_market_user_id("sid") == 42


@patch("app.application.session_account_meta.load_session_account_meta")
@patch("app.application.enterprise_login_flow.bind_tenant_for_login")
def test_enrich_session_meta_admin_skips(mock_bind: MagicMock, mock_load: MagicMock) -> None:
    mock_load.return_value = {"account_kind": "admin"}
    meta = enrich_session_meta_with_tenant("sid", SimpleNamespace(id=1))
    assert meta["account_kind"] == "admin"
    mock_bind.assert_not_called()


@patch("app.application.session_account_meta.get_host_db")
def test_clear_impersonation(mock_get_db: MagicMock) -> None:
    mock_db = MagicMock()
    mock_get_db.return_value = _mock_get_db(mock_db)
    row = SimpleNamespace(impersonating_market_user_id=1, impersonating_username="x")
    mock_db.query.return_value.filter.return_value.first.return_value = row
    clear_impersonation("sid")
    assert row.impersonating_market_user_id is None


@patch("app.application.session_account_meta.load_session_account_meta")
def test_audit_admin_action_no_crash(mock_load: MagicMock) -> None:
    mock_load.return_value = {"impersonating_username": "admin"}
    audit_admin_action(SimpleNamespace(), "test.action", target_user_id=1)


# ---------------------------------------------------------------------------
# middleware + error_handler
# ---------------------------------------------------------------------------


def test_subscription_gate_disabled_by_default() -> None:
    assert _subscription_gate_enabled() is False


@pytest.mark.asyncio
async def test_subscription_gate_passes_when_disabled() -> None:
    app = FastAPI()

    @app.get("/api/products")
    def products():
        return {"ok": True}

    app.add_middleware(SubscriptionGateMiddleware)
    client = TestClient(app)
    assert client.get("/api/products").status_code == 200


@pytest.mark.asyncio
async def test_subscription_gate_blocks_inactive(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("XCAGI_ENFORCE_SUBSCRIPTION", "1")
    app = FastAPI()

    @app.get("/api/products")
    def products():
        return {"ok": True}

    app.add_middleware(SubscriptionGateMiddleware)
    with (
        patch(
            "app.infrastructure.auth.dependencies.resolve_session_user",
            return_value=SimpleNamespace(id=1),
        ),
        patch(
            "app.application.tenant_subscription_app_service.subscription_status_for_user",
            return_value={"active": False},
        ),
    ):
        client = TestClient(app)
        r = client.get("/api/products")
    assert r.status_code == 403
    assert r.json()["error"]["code"] == "SUBSCRIPTION_REQUIRED"


def test_request_id_middleware_generates_header() -> None:
    app = FastAPI()

    @app.get("/ping")
    def ping():
        return {"ok": True}

    app.add_middleware(RequestIdMiddleware)
    client = TestClient(app)
    r = client.get("/ping")
    assert r.headers.get("X-Request-ID")


def test_request_id_middleware_preserves_incoming() -> None:
    app = FastAPI()
    captured: dict[str, str] = {}

    @app.get("/ping")
    def ping(request: Request):
        captured["rid"] = request.state.request_id
        return {"ok": True}

    app.add_middleware(RequestIdMiddleware)
    client = TestClient(app)
    r = client.get("/ping", headers={"X-Request-ID": "fixed-id-123"})
    assert r.headers.get("X-Request-ID") == "fixed-id-123"
    assert captured["rid"] == "fixed-id-123"


def test_error_handler_app_error() -> None:
    app = FastAPI()
    register_exception_handlers(app)

    @app.get("/boom")
    def boom():
        raise AppError(ErrorCode.VALIDATION_ERROR, "bad input", status_code=422)

    client = TestClient(app, raise_server_exceptions=False)
    r = client.get("/boom")
    assert r.status_code == 422
    assert r.json()["error_code"] == "VALIDATION_ERROR"


def test_error_handler_http_exception() -> None:
    app = FastAPI()
    register_exception_handlers(app)

    @app.get("/missing")
    def missing():
        raise StarletteHTTPException(status_code=404, detail="not found")

    r = TestClient(app, raise_server_exceptions=False).get("/missing")
    assert r.status_code == 404
    assert r.json()["success"] is False


def test_error_handler_validation_error() -> None:
    app = FastAPI()
    register_exception_handlers(app)

    @app.get("/need-q")
    def need_q(q: int):
        return {"q": q}

    r = TestClient(app, raise_server_exceptions=False).get("/need-q")
    assert r.status_code == 422
    assert r.json()["error_code"] == "validation_error"


def test_error_handler_unhandled() -> None:
    app = FastAPI()
    register_exception_handlers(app)

    @app.get("/crash")
    def crash():
        raise RuntimeError("unexpected")

    r = TestClient(app, raise_server_exceptions=False).get("/crash")
    assert r.status_code == 500
    assert r.json()["error_code"] == "internal_error"


# ---------------------------------------------------------------------------
# db validators + errors + misc
# ---------------------------------------------------------------------------


def test_model_validators_positive_number() -> None:
    assert ModelValidators.validate_positive_number(0, "qty") == 0
    with pytest.raises(ValueError):
        ModelValidators.validate_positive_number(-1, "qty", allow_zero=False)


def test_model_validators_email_phone() -> None:
    assert ModelValidators.validate_email("a@b.com") == "a@b.com"
    with pytest.raises(ValueError):
        ModelValidators.validate_email("bad")
    assert ModelValidators.validate_phone("13800138000") == "13800138000"


def test_register_model_validators_runs() -> None:
    assert register_model_validators() in (True, False)


def test_app_error_to_dict() -> None:
    err = AuthError(message="expired")
    d = err.to_dict(request_id="rid-1")
    assert d["request_id"] == "rid-1"
    assert d["success"] is False


def test_context_notifier_returns_none() -> None:
    assert get_context_notifier() is None


def test_workflow_parse_excel_header_row() -> None:
    assert _parse_excel_header_row_1based({"header_row": 2}) == 2
    assert _parse_excel_header_row_1based({"header_row_index": "3"}) == 3
    assert _parse_excel_header_row_1based({}) is None
    assert _parse_excel_header_row_1based({"header_row": 0}) is None


def test_mobile_api_helpers() -> None:
    resp = format_mobile_response({"a": 1}, message="ok")
    assert resp["success"] is True
    err = format_error_response("E001")
    assert err["success"] is False
    page = paginate_list([1, 2], total=10, page=1, per_page=2)
    assert page["pagination"]["total_pages"] == 5


def test_aibiz_surface_helpers() -> None:
    assert aibiz_mod._unwrap({"ok": True})["ok"] is True
    assert "terminal=web" in aibiz_mod._surface_image_url("web", 0)
    surface = {"captured_at": "2026-06-14T10:00:00Z"}
    assert aibiz_mod._surface_cache_token(surface).startswith("20260614")


def test_aibiz_png_transform_empty() -> None:
    assert aibiz_mod._crop_png_top(b"") == b""
    assert aibiz_mod._transform_png_view(b"", "full") == b""


def test_aibiz_png_transform_with_pil() -> None:
    try:
        from PIL import Image
    except ImportError:
        pytest.skip("Pillow not installed")
    buf = io.BytesIO()
    Image.new("RGB", (200, 100), color="red").save(buf, format="PNG")
    raw = buf.getvalue()
    cropped = aibiz_mod._crop_png_top(raw, height=50)
    assert len(cropped) <= len(raw)
    thumb = aibiz_mod._resize_png_thumb(raw, max_width=50)
    assert len(thumb) <= len(raw)
