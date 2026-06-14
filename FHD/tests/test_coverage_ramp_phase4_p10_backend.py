"""COVERAGE_RAMP Phase 4 round 10: traditional_mode_fs (12%→) on a temp ROOT_DIR."""

from __future__ import annotations

import base64
from pathlib import Path

import openpyxl
import pytest

import app.traditional_mode_fs as tmf


@pytest.fixture
def root(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> Path:
    monkeypatch.setattr(tmf, "ROOT_DIR", str(tmp_path))
    return tmp_path


# ---------------------------------------------------------------------------
# path safety
# ---------------------------------------------------------------------------


def test_resolve_safe_path_ok(root: Path) -> None:
    p = tmf.resolve_safe_path("a/b.txt")
    assert p is not None and p.startswith(str(root))


def test_resolve_safe_path_traversal(root: Path) -> None:
    assert tmf.resolve_safe_path("../../etc/passwd") is None


def test_root_info_response(root: Path) -> None:
    out = tmf.root_info_response()
    assert out["success"] is True
    assert "list" in out["data"]["capabilities"]


# ---------------------------------------------------------------------------
# file type / time helpers
# ---------------------------------------------------------------------------


def test_get_file_type() -> None:
    assert tmf._get_file_type("a.xlsx") == "xlsx"
    assert tmf._get_file_type("p.PNG") == "png"
    assert tmf._get_file_type("doc.pdf") == "pdf"
    assert tmf._get_file_type("README") == "文件"


def test_format_time() -> None:
    out = tmf._format_time(0.0)
    assert out.startswith("1970-01-01")


# ---------------------------------------------------------------------------
# stat
# ---------------------------------------------------------------------------


def test_stat_response_missing(root: Path) -> None:
    body, code = tmf.stat_response("nope.txt")
    assert code == 404


def test_stat_response_traversal(root: Path) -> None:
    body, code = tmf.stat_response("../x")
    assert code == 403


def test_stat_response_file(root: Path) -> None:
    (root / "f.txt").write_text("hi", encoding="utf-8")
    body, code = tmf.stat_response("f.txt")
    assert code == 200
    assert body["data"]["name"] == "f.txt"
    assert body["data"]["is_dir"] is False


# ---------------------------------------------------------------------------
# write text / base64
# ---------------------------------------------------------------------------


def test_write_text_and_append(root: Path) -> None:
    body, code = tmf.write_text_response("sub/note.txt", "hello")
    assert code == 200
    assert (root / "sub" / "note.txt").read_text(encoding="utf-8") == "hello"
    body2, code2 = tmf.write_text_response("sub/note.txt", "+more", append=True)
    assert code2 == 200
    assert (root / "sub" / "note.txt").read_text(encoding="utf-8") == "hello+more"


def test_write_text_dir_target(root: Path) -> None:
    (root / "d").mkdir()
    body, code = tmf.write_text_response("d", "x")
    assert code == 400


def test_write_text_traversal(root: Path) -> None:
    body, code = tmf.write_text_response("../escape.txt", "x")
    assert code == 403


def test_write_base64_ok(root: Path) -> None:
    encoded = base64.b64encode(b"binary-data").decode()
    body, code = tmf.write_base64_response("blob.bin", encoded)
    assert code == 200
    assert (root / "blob.bin").read_bytes() == b"binary-data"


def test_write_base64_invalid(root: Path) -> None:
    body, code = tmf.write_base64_response("blob.bin", "!!!not-base64!!!")
    assert code == 400


# ---------------------------------------------------------------------------
# move / copy
# ---------------------------------------------------------------------------


def test_move_success(root: Path) -> None:
    (root / "src.txt").write_text("a", encoding="utf-8")
    body, code = tmf.move_response("src.txt", "dst.txt")
    assert code == 200
    assert not (root / "src.txt").exists()
    assert (root / "dst.txt").exists()


def test_move_src_missing(root: Path) -> None:
    body, code = tmf.move_response("ghost.txt", "dst.txt")
    assert code == 404


def test_move_dst_exists_no_overwrite(root: Path) -> None:
    (root / "a.txt").write_text("a", encoding="utf-8")
    (root / "b.txt").write_text("b", encoding="utf-8")
    body, code = tmf.move_response("a.txt", "b.txt")
    assert code == 409


def test_move_overwrite(root: Path) -> None:
    (root / "a.txt").write_text("a", encoding="utf-8")
    (root / "b.txt").write_text("b", encoding="utf-8")
    body, code = tmf.move_response("a.txt", "b.txt", overwrite=True)
    assert code == 200
    assert (root / "b.txt").read_text(encoding="utf-8") == "a"


def test_copy_file(root: Path) -> None:
    (root / "src.txt").write_text("c", encoding="utf-8")
    body, code = tmf.copy_response("src.txt", "copy.txt")
    assert code == 200
    assert (root / "copy.txt").read_text(encoding="utf-8") == "c"


def test_copy_dir_tree(root: Path) -> None:
    d = root / "folder"
    d.mkdir()
    (d / "inner.txt").write_text("x", encoding="utf-8")
    body, code = tmf.copy_response("folder", "folder_copy")
    assert code == 200
    assert (root / "folder_copy" / "inner.txt").is_file()


# ---------------------------------------------------------------------------
# list files
# ---------------------------------------------------------------------------


def test_list_files(root: Path) -> None:
    (root / "f1.txt").write_text("1", encoding="utf-8")
    (root / "sub").mkdir()
    body, code = tmf.list_files_response("")
    assert code == 200
    names = {e["name"] for e in body["data"]["files"]}
    assert "f1.txt" in names and "sub" in names


def test_list_files_nonexistent(root: Path) -> None:
    body, code = tmf.list_files_response("missing-dir")
    assert code == 200
    assert body["data"]["files"] == []


def test_list_files_not_a_dir(root: Path) -> None:
    (root / "file.txt").write_text("x", encoding="utf-8")
    body, code = tmf.list_files_response("file.txt")
    assert code == 400


def test_list_files_traversal(root: Path) -> None:
    body, code = tmf.list_files_response("../..")
    assert code == 403


# ---------------------------------------------------------------------------
# read file (text / excel / image / binary / errors)
# ---------------------------------------------------------------------------


def test_read_text_file(root: Path) -> None:
    (root / "t.txt").write_text("文本内容", encoding="utf-8")
    body, code = tmf.read_file_response("t.txt")
    assert code == 200
    assert body["data"]["type"] == "text"
    assert body["data"]["content"] == "文本内容"


def test_read_excel_file(root: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "x"
    ws["B1"] = 2
    wb.save(str(root / "data.xlsx"))
    body, code = tmf.read_file_response("data.xlsx")
    assert code == 200
    assert body["data"]["type"] == "excel"
    assert "sheets" in body["data"]


def test_read_image_file(root: Path) -> None:
    (root / "pic.png").write_bytes(b"\x89PNG\r\n\x1a\n fake")
    body, code = tmf.read_file_response("pic.png")
    assert code == 200
    assert body["data"]["type"] == "image"
    assert body["data"]["mime"] == "image/png"


def test_read_binary_file(root: Path) -> None:
    (root / "raw.dat").write_bytes(b"\xff\xfe\x00\x01bad-utf8")
    body, code = tmf.read_file_response("raw.dat")
    assert code == 200
    assert body["data"]["type"] == "binary"


def test_read_missing(root: Path) -> None:
    body, code = tmf.read_file_response("nope.txt")
    assert code == 404


def test_read_dir(root: Path) -> None:
    (root / "adir").mkdir()
    body, code = tmf.read_file_response("adir")
    assert code == 400


# ---------------------------------------------------------------------------
# snapshot helpers
# ---------------------------------------------------------------------------


def test_build_and_format_snapshot(root: Path) -> None:
    (root / "x.txt").write_text("a", encoding="utf-8")
    (root / "nested").mkdir()
    (root / "nested" / "y.txt").write_text("b", encoding="utf-8")
    snap = tmf._build_snapshot()
    assert "x.txt" in snap
    assert "nested/y.txt" in snap
    formatted = tmf._format_snapshot(snap)
    assert all(isinstance(v, str) for v in formatted.values())
