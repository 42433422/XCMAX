"""COVERAGE_RAMP Phase 4 round 8: excel_template_analyzer (0%→) via a real xlsx.

Builds a representative 送货单-style template and drives the full analyze()
pipeline so structure/merged/zone/editable-range branches are exercised.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.utils.excel_template_analyzer import (
    analyze_template,
    extract_entries,
)


@pytest.fixture
def template_xlsx(tmp_path: Path) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "送货单模板"

    # Row 1: title (merged) -> 标题
    ws["A1"] = "送货单"
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.font = Font(name="宋体", size=18, bold=True, color="FF0000")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.fill = PatternFill(fill_type="solid", fgColor="FFFFFF00")
    title_cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Row 2: buyer info (merged) -> 购货单位信息
    ws["A2"] = "购货单位：甲公司"
    ws.merge_cells("A2:C2")
    ws["D2"] = "乙方：本公司"
    ws.merge_cells("D2:F2")

    # Row 3: header (merged contains 型号) -> 表头
    ws["A3"] = "产品型号名称"
    ws.merge_cells("A3:B3")
    ws["C3"] = "数量"
    ws["D3"] = "规格"
    ws["E3"] = "单价"
    ws["F3"] = "金额"

    # Rows 4-10: data rows (editable col1/col2 text + numeric cols)
    for i, row in enumerate(range(4, 11), start=1):
        ws.cell(row, 1, f"型号{i}")
        ws.cell(row, 2, f"名称{i}")
        ws.cell(row, 3, i * 2)  # numeric -> editable
        ws.cell(row, 4, "规格说明")  # pure chinese len>2 in col>2 -> template
        ws.cell(row, 5, 10.5)
        ws.cell(row, 6, i * 21.0)

    # Row 11: 合计 (merged) -> 汇总 ; formula cell
    ws["A11"] = "合计"
    ws.merge_cells("A11:E11")
    ws["F11"] = "=SUM(F4:F10)"

    # Row 12: 签名 (merged) -> 签名区
    ws["A12"] = "签名："
    ws.merge_cells("A12:F12")

    # Row 13: other text merged -> 内容区
    ws["A13"] = "其它说明事项"
    ws.merge_cells("A13:F13")

    # Rows 15-20: large empty merged -> 大标题区 (max_row-min_row>3)
    ws.merge_cells("A15:A20")

    # Row 22: small empty merged -> 合并内容
    ws.merge_cells("A22:B22")

    # An editable cell far away to force a non-consecutive group split
    ws.cell(25, 1, "尾项X")

    out = tmp_path / "delivery_template.xlsx"
    wb.save(str(out))
    return str(out)


def test_analyze_template_full_pipeline(template_xlsx: str) -> None:
    result = analyze_template(template_xlsx)
    assert result["file"] == "delivery_template.xlsx"
    assert result["sheet"] == "送货单模板"
    assert result["structure"]["max_col"] == 6
    assert result["structure"]["max_col_letter"] == "F"
    assert isinstance(result["zones"], list) and result["zones"]
    assert isinstance(result["merged_cells"], list) and result["merged_cells"]
    assert isinstance(result["cells"], dict) and result["cells"]


def test_merged_purposes_cover_branches(template_xlsx: str) -> None:
    result = analyze_template(template_xlsx)
    purposes = {mc["purpose"] for mc in result["merged_cells"]}
    # at least the title + several semantic categories should be recognised
    assert "标题" in purposes
    assert "购货单位信息" in purposes
    assert "大标题区" in purposes
    assert "合并内容" in purposes


def test_zones_present(template_xlsx: str) -> None:
    result = analyze_template(template_xlsx)
    zone_names = {z["name"] for z in result["zones"]}
    assert "header" in zone_names
    # data zone exists because editable rows were produced
    assert "data" in zone_names


def test_editable_ranges_and_group_split(template_xlsx: str) -> None:
    result = analyze_template(template_xlsx)
    ranges = result["editable_ranges"]
    assert isinstance(ranges, list) and ranges
    # the far-away editable cell (row 25) forces more than one editable group
    assert len(ranges) >= 2


def test_formula_and_style_in_cells(template_xlsx: str) -> None:
    result = analyze_template(template_xlsx)
    cells = result["cells"]
    assert "F11" in cells
    assert cells["F11"].get("formula") == "=SUM(F4:F10)"
    # title cell carries style metadata
    assert "style" in cells["A1"]
    assert cells["A1"]["style"]


def test_cell_type_classification(template_xlsx: str) -> None:
    result = analyze_template(template_xlsx)
    cells = result["cells"]
    # row<=3 -> template
    assert cells["A1"]["type"] == "template"
    # data row col1 text -> editable
    assert cells["A4"]["type"] == "editable"
    # numeric data cell col>1 -> editable
    assert cells["C4"]["type"] == "editable"
    # pure-chinese in col>2 -> template
    assert cells["D4"]["type"] == "template"


def test_extract_entries(template_xlsx: str) -> None:
    entries = extract_entries(template_xlsx)
    assert entries["file"] == "delivery_template.xlsx"
    assert isinstance(entries["editable_entries"], list)
    assert entries["editable_entries"]
    addresses = {e["address"] for e in entries["editable_entries"]}
    assert "A4" in addresses


def test_analyze_named_sheet(template_xlsx: str) -> None:
    # explicit sheet_name path in _load_workbook
    result = analyze_template(template_xlsx, sheet_name="送货单模板")
    assert result["sheet"] == "送货单模板"
