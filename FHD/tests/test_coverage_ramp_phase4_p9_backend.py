"""COVERAGE_RAMP Phase 4 round 9: template_export_utils (0%) + upload_helpers (0%)."""

from __future__ import annotations

import io
from datetime import date, datetime
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest
from fastapi import UploadFile

from app.utils.template_export_utils import (
    _format_cell_value,
    _normalize_header,
    _to_header_lookup,
    fill_workbook_from_template,
)
from app.utils.upload_helpers import save_upload_bytes, save_upload_file


# ---------------------------------------------------------------------------
# small helpers
# ---------------------------------------------------------------------------


def test_normalize_header() -> None:
    assert _normalize_header("  型 号 ") == "型号"
    assert _normalize_header(None) == ""
    assert _normalize_header("ABC ") == "abc"


def test_to_header_lookup_variants() -> None:
    lookup = _to_header_lookup(
        {
            "model": ["型号", "产品型号"],
            "name": "名称",
            "qty": [],  # falls back to record_key
            "": "ignored",  # empty key skipped
            "other": 123,  # non-str/list -> falls back to key
        }
    )
    keys = {rk for _, _, rk in lookup}
    assert "model" in keys and "name" in keys and "qty" in keys and "other" in keys
    assert "" not in keys
    normalized = {n for n, _, _ in lookup}
    assert "型号" in normalized
    assert "产品型号" in normalized


def test_format_cell_value() -> None:
    assert _format_cell_value(datetime(2024, 1, 2, 3, 4, 5)) == "2024-01-02 03:04:05"
    assert _format_cell_value(date(2024, 1, 2)) == "2024-01-02"
    assert _format_cell_value("plain") == "plain"
    assert _format_cell_value(42) == 42


# ---------------------------------------------------------------------------
# fill_workbook_from_template
# ---------------------------------------------------------------------------


def _make_template(tmp_path: Path, *, with_preamble: bool = False) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "模板"
    start = 1
    if with_preamble:
        ws["A1"] = "样例说明：请勿删除"
        ws.merge_cells("A1:C1")
        start = 3
    ws.cell(start, 1, "型号")
    ws.cell(start, 2, "名称")
    ws.cell(start, 3, "数量")
    # one stale sample data row
    ws.cell(start + 1, 1, "旧型号")
    ws.cell(start + 1, 2, "旧名称")
    ws.cell(start + 1, 3, 99)
    out = tmp_path / "tpl.xlsx"
    wb.save(str(out))
    return str(out)


_ALIAS = {"model": ["型号"], "name": ["名称"], "qty": ["数量"], "spec": ["规格"]}
_RECORDS = [
    {"model": "A1", "name": "甲", "qty": 3, "spec": "10kg"},
    {"model": "B2", "name": "乙", "qty": 5, "spec": None},
]


def test_fill_natural_header_found(tmp_path: Path) -> None:
    tpl = _make_template(tmp_path)
    wb = fill_workbook_from_template(tpl, _RECORDS, _ALIAS)
    ws = wb.active
    assert ws.cell(2, 1).value == "A1"
    assert ws.cell(2, 2).value == "甲"
    assert ws.cell(3, 1).value == "B2"


def test_fill_synthetic_header_fallback(tmp_path: Path) -> None:
    # empty workbook -> no natural header -> synthetic header written at row 1
    wb0 = openpyxl.Workbook()
    out = tmp_path / "empty.xlsx"
    wb0.save(str(out))
    wb = fill_workbook_from_template(str(out), _RECORDS, _ALIAS)
    ws = wb.active
    headers = {ws.cell(1, c).value for c in range(1, ws.max_column + 1)}
    assert "型号" in headers


def test_fill_append_missing_columns(tmp_path: Path) -> None:
    # template only has 型号/名称/数量; 规格 should be appended
    tpl = _make_template(tmp_path)
    wb = fill_workbook_from_template(
        tpl, _RECORDS, _ALIAS, append_missing_field_columns=True
    )
    ws = wb.active
    header_vals = {ws.cell(1, c).value for c in range(1, ws.max_column + 1)}
    assert "规格" in header_vals


def test_fill_clear_rows_above_header(tmp_path: Path) -> None:
    tpl = _make_template(tmp_path, with_preamble=True)
    wb = fill_workbook_from_template(
        tpl, _RECORDS, _ALIAS, clear_rows_above_header=True
    )
    ws = wb.active
    # header moved up to row 1 after preamble removal
    assert ws.cell(1, 1).value == "型号"


def test_fill_clear_existing_and_truncate(tmp_path: Path) -> None:
    tpl = _make_template(tmp_path)
    wb = fill_workbook_from_template(
        tpl,
        _RECORDS,
        _ALIAS,
        clear_existing_data_rows_all_columns=True,
        truncate_rows_after_data_area=True,
    )
    ws = wb.active
    # only header + 2 data rows remain
    assert ws.max_row == 3


def test_fill_with_sheet_name(tmp_path: Path) -> None:
    tpl = _make_template(tmp_path)
    wb = fill_workbook_from_template(tpl, _RECORDS, _ALIAS, sheet_name="模板")
    assert wb["模板"].cell(2, 1).value == "A1"


# ---------------------------------------------------------------------------
# upload_helpers
# ---------------------------------------------------------------------------


def test_save_upload_bytes(tmp_path: Path) -> None:
    with patch("app.utils.upload_helpers.get_upload_dir", return_value=str(tmp_path)):
        path = save_upload_bytes(b"hello", subdir="imports", filename="data.xlsx")
    assert Path(path).is_file()
    assert Path(path).read_bytes() == b"hello"
    assert "imports" in path


@pytest.mark.asyncio
async def test_save_upload_file(tmp_path: Path) -> None:
    upload = UploadFile(filename="up.txt", file=io.BytesIO(b"content-bytes"))
    with patch("app.utils.upload_helpers.get_upload_dir", return_value=str(tmp_path)):
        path = await save_upload_file(upload, subdir="docs")
    assert Path(path).is_file()
    assert Path(path).read_bytes() == b"content-bytes"


@pytest.mark.asyncio
async def test_save_upload_file_missing_filename() -> None:
    upload = UploadFile(filename="", file=io.BytesIO(b"x"))
    with pytest.raises(ValueError):
        await save_upload_file(upload, subdir="docs")
