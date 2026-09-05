"""Signed private module, real auth/session, real Excel conversion, isolated data."""

import importlib.util
import sys
from pathlib import Path

import pytest
from fastapi import FastAPI, Request
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.mod_sdk.attendance_roster import initialize_roster_once
from app.mod_sdk.owner_workspace import owner_context, owner_workspace

MOD_ID = "sunbird-attendance-custom"
BASE = f"/api/mod/{MOD_ID}/attendance"


@pytest.fixture
def sunbird_client(signed_runtime_mod):
    source = Path(__file__).resolve().parents[2] / "mods" / MOD_ID
    installed = signed_runtime_mod.install(mod_id=MOD_ID, source=source)
    for key in list(sys.modules):
        if key == "sunbird_attendance" or key.startswith("sunbird_attendance."):
            del sys.modules[key]
    spec = importlib.util.spec_from_file_location(
        "fixture_sunbird_backend", installed / "backend/blueprints.py"
    )
    backend = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(backend)
    app = FastAPI()
    backend.register_fastapi_routes(app, MOD_ID)
    from app.infrastructure.mods.install_receipts import mark_runtime_loaded

    mark_runtime_loaded(MOD_ID, mods_root=str(signed_runtime_mod.root), api_registered=True)

    @app.get("/probe")
    async def probe(request: Request):
        return backend.verify_delivery(request)

    with TestClient(app) as client:
        client.cookies.set("session_id", "mod-session-1")
        yield client


def test_real_template_conversion_and_download_are_owner_scoped(sunbird_client, tmp_path):
    from sunbird_attendance.verification import write_conversion_sample

    source, template = write_conversion_sample(tmp_path)
    with owner_context("tenant:1"):
        assert initialize_roster_once(
            [{"name": "交付验证样例", "dept": "验证部门", "group": "验证岗位"}]
        )
    with template.open("rb") as stream:
        response = sunbird_client.post(
            BASE + "/template", files={"file": ("template.xlsx", stream)}
        )
    assert response.status_code == 200, response.text
    with template.open("rb") as stream:
        assert (
            sunbird_client.post(
                BASE + "/template", files={"file": ("template.xlsx", stream)}
            ).status_code
            == 409
        )
    with source.open("rb") as stream:
        response = sunbird_client.post(
            BASE + "/convert-upload",
            files={"file": ("attendance.xlsx", stream)},
            data={"month": "2026-09"},
        )
    assert response.status_code == 200, response.text
    result = response.json()["data"]
    assert result["rows_used_for_template"] == result["employees_matched"] == 1
    assert result["used_llm"] is False
    assert "input" not in result and "output" not in result
    download = sunbird_client.get(result["download_path"])
    assert download.status_code == 200
    output = tmp_path / "output.xlsx"
    output.write_bytes(download.content)
    workbook = load_workbook(output)
    assert workbook.sheetnames == ["明细", "月度统计"]
    assert workbook["明细"].cell(4, 3).value == "交付验证样例"
    workbook.close()
    sunbird_client.cookies.set("session_id", "mod-session-2")
    assert sunbird_client.get(result["download_path"]).status_code == 403
    assert sunbird_client.get(BASE + "/rules").status_code == 403


def test_probe_performs_conversion_without_writing_customer_data(sunbird_client):
    with owner_context("tenant:1"):
        initialize_roster_once([{"name": "Private Name", "dept": "Private Dept"}])
        workspace = owner_workspace("attendance-industry").root
    before = {p.name: p.read_bytes() for p in workspace.iterdir()}
    response = sunbird_client.get("/probe")
    assert response.status_code == 200, response.text
    result = response.json()
    assert result["passed"] is True, result
    assert result["case_id"] == "sunbird-owner-conversion-v1"
    observations = result["observations"]
    assert observations["owner_roster_count"] == 1
    assert observations["monthly_formulas_link_detail"] is True
    assert observations["customer_data_written"] is False
    assert "Private Name" not in response.text
    assert before == {p.name: p.read_bytes() for p in workspace.iterdir()}


def test_probe_does_not_claim_uninitialized_owner_schema_is_ready(sunbird_client):
    response = sunbird_client.get("/probe")
    assert response.status_code == 200
    assert response.json()["pending"] is True
    assert response.json()["reason"] == "workspace_not_ready"
    assert "passed" not in response.json()


def test_invalid_template_is_a_friendly_error_and_cannot_write(sunbird_client):
    response = sunbird_client.post(
        BASE + "/template", files={"file": ("broken.xlsx", b"not-a-zip")}
    )
    assert response.status_code == 400
    with owner_context("tenant:1"):
        assert not owner_workspace(MOD_ID).root.exists()


@pytest.mark.parametrize("session", ["forged", "mod-session-4"])
def test_private_route_and_probe_require_a_valid_current_session(sunbird_client, session):
    sunbird_client.cookies.set("session_id", session)
    assert sunbird_client.get(BASE + "/rules").status_code == 401
    assert sunbird_client.get("/probe").status_code == 401


def test_policy_is_private_validated_and_isolated_from_host(sunbird_client, tmp_path, monkeypatch):
    def fail_host_config(*args, **kwargs):
        raise AssertionError("private policy must not read or write host approval config")

    monkeypatch.setattr("resources.config.approval_config.get_approval_config", fail_host_config)
    response = sunbird_client.post(
        BASE + "/policy",
        json={
            "attendance_policy": {
                "weekday_segments": ["09:00-12:00"],
                "sunday_empty_schedule": False,
            }
        },
    )
    assert response.status_code == 200, response.text
    assert sunbird_client.get(BASE + "/policy").json()["attendance_policy"]["weekday_segments"] == [
        "09:00-12:00"
    ]
    invalid = sunbird_client.post(
        BASE + "/policy",
        json={"attendance_policy": {"weekday_segments": ["29:00-12:00"]}},
    )
    assert invalid.status_code == 400
    assert sunbird_client.get(BASE + "/policy").json()["attendance_policy"]["weekday_segments"] == [
        "09:00-12:00"
    ]
    with owner_context("tenant:2"):
        from sunbird_attendance.owner_config import read_policy

        assert read_policy()["weekday_segments"] == ["08:00-12:00", "13:30-17:30"]


def test_parallel_owner_policies_do_not_share_process_global_state(sunbird_client):
    from concurrent.futures import ThreadPoolExecutor
    from threading import Barrier

    from sunbird_attendance.owner_config import read_policy, save_policy
    from sunbird_attendance.rules import ACTIVE_POLICY, set_attendance_policy

    barrier = Barrier(2)

    def read_one(owner, segment):
        with owner_context(owner):
            save_policy({"weekday_segments": [segment]})
            set_attendance_policy(read_policy())
            barrier.wait(timeout=5)
            return ACTIVE_POLICY.get("weekday_segments")

    with ThreadPoolExecutor(max_workers=2) as pool:
        one = pool.submit(read_one, "tenant:1", "08:00-12:00")
        two = pool.submit(read_one, "tenant:2", "09:00-11:00")
        assert one.result() == ["08:00-12:00"]
        assert two.result() == ["09:00-11:00"]
