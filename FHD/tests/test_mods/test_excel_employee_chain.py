# mypy: disable-error-code="union-attr"
from __future__ import annotations

import asyncio
import importlib.util
import json
from pathlib import Path
from types import ModuleType

from openpyxl import Workbook, load_workbook

FHD_ROOT = Path(__file__).resolve().parents[2]


def _load_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _sample_attendance_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "考勤明细"
    ws.append(["太阳鸟考勤报表"])
    ws.append(["统计周期", "2025-10-01 至 2025-10-31"])
    ws.append(["姓名", "部门", "考勤日期", "工时"])
    ws.append(["张三", "生产部", "2025-10-01", 8])
    ws.append(["李四", "仓储部", "2025-10-01", 7.5])
    wb.save(path)
    wb.close()


def test_excel_reader_outputs_flat_rows_with_detected_headers(tmp_path: Path) -> None:
    reader = _load_module(
        "excel_full_read_convert_test",
        FHD_ROOT
        / "mods/_employees/excel-full-read-employee/backend/vendor/excel_full_read/convert.py",
    )
    src = tmp_path / "attendance.xlsx"
    out = tmp_path / "workbook.json"
    _sample_attendance_workbook(src)

    result = reader.convert_file(
        src,
        out,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "workbook.json"},
    )

    data = json.loads(out.read_text(encoding="utf-8"))
    sheet = data["sheets"][0]
    assert result["output_path"] == str(out)
    assert sheet["header_row"] == 3
    assert sheet["data_start_row"] == 4
    assert sheet["columns"][:4] == ["姓名", "部门", "考勤日期", "工时"]
    assert sheet["rows"][0]["姓名"] == "张三"
    assert sheet["rows"][1]["工时"] == 7.5
    assert sheet["row_records"][0]["cells"]["部门"] == "生产部"


def test_excel_generator_accepts_reader_workbook_json(tmp_path: Path) -> None:
    reader = _load_module(
        "excel_full_read_convert_chain_test",
        FHD_ROOT
        / "mods/_employees/excel-full-read-employee/backend/vendor/excel_full_read/convert.py",
    )
    generator = _load_module(
        "excel_generate_convert_chain_test",
        FHD_ROOT
        / "mods/_employees/excel-generate-employee/backend/vendor/excel_generate/convert.py",
    )
    src = tmp_path / "attendance.xlsx"
    workbook_json = tmp_path / "workbook.json"
    generated = tmp_path / "generated.xlsx"
    _sample_attendance_workbook(src)

    reader.convert_file(
        src,
        workbook_json,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "workbook.json"},
    )
    result = asyncio.run(
        generator.convert_file(
            workbook_json,
            generated,
            payload={},
            ctx={},
            rule_spec={"default_output_relpath": "output.xlsx"},
        )
    )

    assert result["output_path"] == str(generated)
    assert result["sheet_count"] == 1
    assert result["row_count"] == 2
    wb = load_workbook(generated, data_only=True)
    try:
        ws = wb["考勤明细"]
        assert [ws.cell(1, col).value for col in range(1, 5)] == [
            "姓名",
            "部门",
            "考勤日期",
            "工时",
        ]
        assert [ws.cell(2, col).value for col in range(1, 5)] == ["张三", "生产部", "2025-10-01", 8]
        assert [ws.cell(3, col).value for col in range(1, 5)] == [
            "李四",
            "仓储部",
            "2025-10-01",
            7.5,
        ]
    finally:
        wb.close()
