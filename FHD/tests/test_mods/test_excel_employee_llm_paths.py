# mypy: disable-error-code="import-not-found, no-any-return, union-attr"
"""规则映射员 / 质检员的 LLM 协作路径测试（mock call_llm）。

分工契约：LLM 提议、确定性验证——非法提议必须被拒绝且留痕；
LLM 不可用/关闭时行为与纯启发式完全一致；LLM 不可推翻确定性检查结论。
"""

from __future__ import annotations

import asyncio
import importlib.util
import json
from pathlib import Path
from types import ModuleType

FHD_ROOT = Path(__file__).resolve().parents[2]

READER_CONVERT = (
    FHD_ROOT / "mods/_employees/excel-full-read-employee/backend/vendor/excel_full_read/convert.py"
)
WRITER_CONVERT = (
    FHD_ROOT
    / "mods/_employees/excel-template-write-employee/backend/vendor/excel_template_write/convert.py"
)
MAPPER_VENDOR = FHD_ROOT / "mods/_employees/excel-rules-map-employee/backend/vendor"
QC_VENDOR = FHD_ROOT / "mods/_employees/excel-qc-employee/backend/vendor"


def _load_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _load_pkgs():
    import sys

    for p in (str(MAPPER_VENDOR), str(QC_VENDOR)):
        if p not in sys.path:
            sys.path.insert(0, p)
    import excel_qc.checks as qc_checks
    import excel_qc.convert as qc_convert
    import excel_rules_map.compile_plan as compile_mod
    import excel_rules_map.convert as mapper_convert
    import excel_rules_map.infer as infer_mod
    import excel_rules_map.llm_refine as llm_refine

    return infer_mod, compile_mod, mapper_convert, llm_refine, qc_checks, qc_convert


def _synthetic_template(path: Path) -> None:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.cell(1, 2, 2026)
    ws.cell(1, 4, 7)
    for day in range(1, 16):
        ws.cell(2, 4 + (day - 1) * 2, day)
    for i, key in enumerate(["甲", "乙", "丙", "丁"]):
        top = 3 + i * 3
        ws.cell(top, 1, "组")
        ws.merge_cells(start_row=top, start_column=1, end_row=top + 2, end_column=1)
        ws.cell(top, 2, key)
        ws.merge_cells(start_row=top, start_column=2, end_row=top + 2, end_column=2)
        ws.cell(top, 35, f"=SUM(D{top}:AF{top + 2})")
    wb.save(path)
    wb.close()


def _workbook_json(tmp_path: Path) -> dict:
    reader = _load_module(f"reader_llm_{tmp_path.name}", READER_CONVERT)
    xlsx = tmp_path / "template.xlsx"
    _synthetic_template(xlsx)
    out = tmp_path / "wb.json"
    reader.convert_file(
        xlsx, out, payload={}, ctx={}, rule_spec={"default_output_relpath": "workbook.json"}
    )
    return json.loads(out.read_text(encoding="utf-8"))


def _fake_llm(response: dict | str, *, ok: bool = True):
    calls: list[dict] = []

    async def call_llm(messages, **kwargs):
        calls.append({"messages": messages, "kwargs": kwargs})
        content = (
            response if isinstance(response, str) else json.dumps(response, ensure_ascii=False)
        )
        return {"ok": ok, "content": content, "error": "" if ok else "mock down"}

    call_llm.calls = calls  # type: ignore[attr-defined]
    return call_llm


# ---------------------------------------------------------------------------
# 映射员：LLM 提议 → 机器验证
# ---------------------------------------------------------------------------


def test_mapper_llm_adopts_valid_bands_and_clear_zone(tmp_path: Path) -> None:
    infer_mod, _, _, llm_refine, _, _ = _load_pkgs()
    workbook = _workbook_json(tmp_path)
    rules = infer_mod.infer_rules(workbook, source_name="synthetic")
    sheet = workbook["sheets"][0]

    call_llm = _fake_llm(
        {
            "bands": {
                "am": {"row_offset": 0, "max_entries": 1},
                "pm": {"row_offset": 1, "max_entries": 2},
            },
            "key_col": None,
            "clear_zone": {"col_start": 4, "col_end": 33},
            "notes": ["块内 3 行推断为上午/下午两带"],
        }
    )
    rules = asyncio.run(llm_refine.llm_refine_rules(rules, sheet, call_llm))

    assert rules["bands"] == {
        "am": {"row_offset": 0, "max_entries": 1},
        "pm": {"row_offset": 1, "max_entries": 2},
    }
    assert rules["template_map"]["clear_zone"] == {"col_start": 4, "col_end": 33}
    llm_ev = rules["evidence"]["llm"]
    assert llm_ev["used"] is True
    assert {a["field"] for a in llm_ev["adopted"]} == {"bands", "clear_zone"}
    assert llm_ev["rejected"] == []
    assert any("LLM 已提议并通过机器验证" in q for q in rules["evidence"]["open_questions"])
    assert len(call_llm.calls) == 1  # type: ignore[attr-defined]


def test_mapper_llm_rejects_invalid_proposals(tmp_path: Path) -> None:
    """越块 bands、与公式区重叠的 clear_zone → 全部拒绝并留痕，规则保持原样。"""
    infer_mod, _, _, llm_refine, _, _ = _load_pkgs()
    workbook = _workbook_json(tmp_path)
    rules = infer_mod.infer_rules(workbook, source_name="synthetic")
    sheet = workbook["sheets"][0]
    bands_before = json.loads(json.dumps(rules["bands"]))
    zone_before = json.loads(json.dumps(rules["template_map"]["clear_zone"]))

    call_llm = _fake_llm(
        {
            "bands": {"夜班": {"row_offset": 9, "max_entries": 1}},  # 越块（块高 3）
            "clear_zone": {"col_start": 30, "col_end": 35},  # 与公式区 35..35 重叠
            "notes": [],
        }
    )
    rules = asyncio.run(llm_refine.llm_refine_rules(rules, sheet, call_llm))

    assert rules["bands"] == bands_before
    assert rules["template_map"]["clear_zone"] == zone_before
    reasons = "；".join(r["reason"] for r in rules["evidence"]["llm"]["rejected"])
    assert "越块" in reasons and "重叠" in reasons


def test_mapper_llm_garbage_and_outage_are_recorded(tmp_path: Path) -> None:
    infer_mod, _, _, llm_refine, _, _ = _load_pkgs()
    workbook = _workbook_json(tmp_path)
    sheet = workbook["sheets"][0]

    rules1 = infer_mod.infer_rules(workbook, source_name="synthetic")
    rules1 = asyncio.run(llm_refine.llm_refine_rules(rules1, sheet, _fake_llm("这不是 JSON")))
    assert any("无法解析" in r["reason"] for r in rules1["evidence"]["llm"]["rejected"])

    rules2 = infer_mod.infer_rules(workbook, source_name="synthetic")
    rules2 = asyncio.run(llm_refine.llm_refine_rules(rules2, sheet, _fake_llm({}, ok=False)))
    assert any("不可用" in r["reason"] for r in rules2["evidence"]["llm"]["rejected"])


def test_mapper_convert_no_llm_matches_heuristic_baseline(tmp_path: Path) -> None:
    """无 call_llm 时 convert_file 行为与纯启发式一致（回归保障）。"""
    infer_mod, _, mapper_convert, _, _, _ = _load_pkgs()
    workbook = _workbook_json(tmp_path)
    wb_path = tmp_path / "wb.json"

    result = asyncio.run(
        mapper_convert.convert_file(
            wb_path,
            tmp_path / "outputs" / "rules.json",
            payload={},
            ctx={},
            rule_spec={"default_output_relpath": "outputs/rules.json"},
        )
    )
    assert result["action"] == "infer"
    assert result["llm"] == {"used": False}
    produced = json.loads(Path(result["output_path"]).read_text(encoding="utf-8"))
    baseline = infer_mod.infer_rules(workbook, source_name=produced["evidence"]["source"])
    assert produced["template_map"] == baseline["template_map"]
    assert produced["bands"] == baseline["bands"]


def test_mapper_convert_with_llm_ctx(tmp_path: Path) -> None:
    _, _, mapper_convert, _, _, _ = _load_pkgs()
    _workbook_json(tmp_path)
    call_llm = _fake_llm(
        {
            "bands": {"am": {"row_offset": 0, "max_entries": 3}},
            "clear_zone": None,
            "key_col": None,
            "notes": [],
        }
    )
    result = asyncio.run(
        mapper_convert.convert_file(
            tmp_path / "wb.json",
            tmp_path / "outputs" / "rules.json",
            payload={},
            ctx={"call_llm": call_llm},
            rule_spec={"default_output_relpath": "outputs/rules.json"},
        )
    )
    assert result["llm"]["used"] is True
    assert {a["field"] for a in result["llm"]["adopted"]} == {"bands"}
    # use_llm=false 显式关闭
    result2 = asyncio.run(
        mapper_convert.convert_file(
            tmp_path / "wb.json",
            tmp_path / "outputs2" / "rules.json",
            payload={"use_llm": False},
            ctx={"call_llm": call_llm},
            rule_spec={"default_output_relpath": "outputs/rules.json"},
        )
    )
    assert result2["llm"] == {"used": False}
    assert len(call_llm.calls) == 1  # type: ignore[attr-defined]


# ---------------------------------------------------------------------------
# 质检员：semantic 节
# ---------------------------------------------------------------------------


def _qc_chain(tmp_path: Path):
    infer_mod, compile_mod, _, _, qc_checks, qc_convert = _load_pkgs()
    writer = _load_module(f"writer_llm_{tmp_path.name}", WRITER_CONVERT)
    workbook = _workbook_json(tmp_path)
    rules = infer_mod.infer_rules(workbook, source_name="synthetic")
    rules["bands"] = {"am": {"row_offset": 0, "max_entries": 1}}
    plan = compile_mod.compile_plan(
        rules,
        [{"key": "甲", "day": 1, "band": "am", "entries": [{"symbol": "√", "value": 2.0}]}],
        month_label="2026-07",
    )
    plan_path = tmp_path / "plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")
    filled = tmp_path / "filled.xlsx"
    writer.convert_file(
        plan_path,
        filled,
        template_path=tmp_path / "template.xlsx",
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    return rules, plan, filled, qc_convert


def test_qc_semantic_findings_merge_into_verdict(tmp_path: Path) -> None:
    rules, plan, filled, qc_convert = _qc_chain(tmp_path)
    call_llm = _fake_llm(
        {
            "findings": [
                {
                    "severity": "fail",
                    "detail": "甲的单日数值 2.0 与班次常理不符（示例）",
                    "evidence": "per_key_numeric_sum",
                },
                {"severity": "info", "detail": "其余键无记录属正常", "evidence": ""},
            ],
            "summary_zh": "本次回填结构正确，但甲的数值需人工复核。",
        }
    )
    result = asyncio.run(
        qc_convert.convert_file(
            filled,
            tmp_path / "outputs" / "qc_report.json",
            payload={"plan": plan, "rules": rules},
            ctx={"call_llm": call_llm},
            rule_spec={"default_output_relpath": "outputs/qc_report.json"},
        )
    )
    assert result["verdict"] == "FAIL"
    assert "semantic_llm" in result["blame"]
    assert result["human_summary"].startswith("本次回填结构正确")
    report = json.loads(Path(result["output_path"]).read_text(encoding="utf-8"))
    semantic = report["sections"]["semantic"]
    assert semantic["status"] == "fail"
    assert all(i.get("source") == "llm" for i in semantic["issues"])
    # 确定性节不被 LLM 推翻
    assert report["sections"]["conformance"]["status"] == "pass"


def test_qc_semantic_strict_false_downgrades_llm_fail(tmp_path: Path) -> None:
    rules, plan, filled, qc_convert = _qc_chain(tmp_path)
    call_llm = _fake_llm(
        {"findings": [{"severity": "fail", "detail": "可疑", "evidence": ""}], "summary_zh": "略"}
    )
    result = asyncio.run(
        qc_convert.convert_file(
            filled,
            tmp_path / "outputs" / "qc_report.json",
            payload={"plan": plan, "rules": rules, "llm_strict": False},
            ctx={"call_llm": call_llm},
            rule_spec={"default_output_relpath": "outputs/qc_report.json"},
        )
    )
    assert result["verdict"] == "WARN"
    assert result["blame"] == []


def test_qc_semantic_skipped_without_llm_and_on_garbage(tmp_path: Path) -> None:
    rules, plan, filled, qc_convert = _qc_chain(tmp_path)

    r1 = asyncio.run(
        qc_convert.convert_file(
            filled,
            tmp_path / "o1" / "qc_report.json",
            payload={"plan": plan, "rules": rules},
            ctx={},
            rule_spec={"default_output_relpath": "outputs/qc_report.json"},
        )
    )
    rep1 = json.loads(Path(r1["output_path"]).read_text(encoding="utf-8"))
    assert rep1["sections"]["semantic"]["status"] == "skipped"

    r2 = asyncio.run(
        qc_convert.convert_file(
            filled,
            tmp_path / "o2" / "qc_report.json",
            payload={"plan": plan, "rules": rules},
            ctx={"call_llm": _fake_llm("乱码非JSON")},
            rule_spec={"default_output_relpath": "outputs/qc_report.json"},
        )
    )
    rep2 = json.loads(Path(r2["output_path"]).read_text(encoding="utf-8"))
    assert rep2["sections"]["semantic"]["status"] == "warn"
    assert any("无法解析" in i["detail"] for i in rep2["sections"]["semantic"]["issues"])
