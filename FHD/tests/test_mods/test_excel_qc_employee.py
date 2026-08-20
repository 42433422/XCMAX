# mypy: disable-error-code="import-not-found, union-attr"
from __future__ import annotations

import asyncio
import importlib.util
import json
from pathlib import Path
from types import ModuleType

import pytest
from openpyxl import load_workbook

FHD_ROOT = Path(__file__).resolve().parents[2]

READER_CONVERT = (
    FHD_ROOT / "mods/_employees/excel-full-read-employee/backend/vendor/excel_full_read/convert.py"
)
WRITER_CONVERT = (
    FHD_ROOT
    / "mods/_employees/excel-template-write-employee/backend/vendor/excel_template_write/convert.py"
)
MAPPER_VENDOR = FHD_ROOT / "mods/_employees/excel-rules-map-employee/backend/vendor"
QC_VENDOR = FHD_ROOT / "mods/_employees/excel-qc-employee/backend/vendor"
QC_EMPLOYEE = FHD_ROOT / "mods/_employees/excel-qc-employee/backend/employees/excel_qc_employee.py"

SUNBIRD_TEMPLATE = Path(
    "/workspace/成都修茈科技有限公司/MODstore_deploy/var/employee_draft_assets/9/real-files-smoke/00_考勤-2026-3月份考勤统计表.xlsx"
)


def _load_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _load_vendor_pkgs():
    import sys

    for p in (str(MAPPER_VENDOR), str(QC_VENDOR)):
        if p not in sys.path:
            sys.path.insert(0, p)
    import excel_qc.checks as qc_checks
    import excel_rules_map.compile_plan as compile_mod
    import excel_rules_map.infer as infer_mod

    return infer_mod, compile_mod, qc_checks


def _synthetic_template(path: Path) -> None:
    """与规则映射员测试同构的合成模板：4 块×3 行、日历 1..15×2、公式区 AI..AJ。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.cell(1, 2, 2026)
    ws.cell(1, 4, 7)
    for day in range(1, 16):
        ws.cell(2, 4 + (day - 1) * 2, day)
    for i, key in enumerate(["甲", "乙", "丙", "丁"]):
        top = 3 + i * 3
        ws.cell(top, 1, "一组" if i < 2 else "二组")
        ws.merge_cells(start_row=top, start_column=1, end_row=top + 2, end_column=1)
        ws.cell(top, 2, key)
        ws.merge_cells(start_row=top, start_column=2, end_row=top + 2, end_column=2)
        ws.cell(top, 33, "残留数据")  # 位于 clear_zone（col4..34），必须被清掉
        ws.cell(top, 35, f"=SUM(D{top}:AF{top + 2})")
        ws.cell(top, 36, "=COUNT($D$2:$AF$2)")
    wb.save(path)
    wb.close()


def _build_chain(tmp_path: Path):
    """跑完 读取→infer→(人确认 bands)→compile→写入，返回 (rules, plan, filled, template)。"""
    infer_mod, compile_mod, _ = _load_vendor_pkgs()
    reader = _load_module(f"reader_qc_{tmp_path.name}", READER_CONVERT)
    writer = _load_module(f"writer_qc_{tmp_path.name}", WRITER_CONVERT)

    template = tmp_path / "template.xlsx"
    _synthetic_template(template)
    wb_json = tmp_path / "wb.json"
    reader.convert_file(
        template, wb_json, payload={}, ctx={}, rule_spec={"default_output_relpath": "workbook.json"}
    )
    rules = infer_mod.infer_rules(
        json.loads(wb_json.read_text(encoding="utf-8")), source_name="synthetic"
    )
    rules["bands"] = {
        "am": {"row_offset": 0, "max_entries": 1},
        "pm": {"row_offset": 1, "max_entries": 1},
    }
    records = [
        {"key": "甲", "day": 1, "band": "am", "entries": [{"symbol": "√", "value": 2.0}]},
        {"key": "乙", "day": 3, "band": "pm", "entries": [{"symbol": "☆", "value": 1.5}]},
        {"key": "路人", "day": 2, "band": "am", "entries": [{"symbol": "√", "value": 1}]},
    ]
    plan = compile_mod.compile_plan(rules, records, month_label="2026-07")
    plan_path = tmp_path / "plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")

    filled = tmp_path / "filled.xlsx"
    writer.convert_file(
        plan_path,
        filled,
        template_path=template,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    return rules, plan, filled, template


# ---------------------------------------------------------------------------
# 正向：全链产物 QC 通过（dropped 记录 → WARN 呈现给人）
# ---------------------------------------------------------------------------


def test_qc_passes_on_honest_chain(tmp_path: Path) -> None:
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, template = _build_chain(tmp_path)

    report = qc.run_qc(filled, plan, template_path=template, rules=rules)
    sections = report["sections"]
    assert sections["conformance"]["status"] == "pass", sections["conformance"]
    # 合成链路公式列全被模板覆盖 → 计划无 protected_ranges → 该节如实 skipped
    assert sections["protection"]["status"] == "skipped"
    # 映射员丢弃了「路人」→ expected 节 warn（呈现给人确认），不是 fail
    assert sections["expected"]["status"] == "warn"
    assert sections["formulas"]["status"] == "pass"
    assert sections["traceability"]["status"] == "pass"
    assert sections["structure"]["status"] == "pass"
    assert report["verdict"] == "WARN"
    assert report["blame"] == []
    # 残留数据被 clear 且未误报
    assert sections["conformance"]["stats"]["clear_residues"] == 0


def test_qc_pass_without_optional_inputs(tmp_path: Path) -> None:
    """只给 plan（无 rules/模板）：可跳过的节标 skipped，不假装通过也不误报。"""
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, _ = _build_chain(tmp_path)
    plan2 = json.loads(json.dumps(plan))
    plan2["expected"] = {k: v for k, v in plan["expected"].items() if k != "records_dropped"}

    report = qc.run_qc(filled, plan2)
    sections = report["sections"]
    assert sections["protection"]["status"] == "skipped"
    assert sections["structure"]["status"] == "skipped"
    # expected 有 per_key_numeric_sum 但无 rules → warn 提示补充
    assert sections["expected"]["status"] == "warn"
    assert sections["traceability"]["status"] == "warn"
    assert report["verdict"] == "WARN"


# ---------------------------------------------------------------------------
# 篡改检出：每条问责路由
# ---------------------------------------------------------------------------


def test_qc_detects_writer_tampering(tmp_path: Path) -> None:
    """改动输出文件的计划格值 → conformance FAIL，blame writer_or_plan。"""
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, template = _build_chain(tmp_path)

    wb = load_workbook(filled)
    wb["明细"].cell(3, 5).value = 999  # 计划写 2.0 的数值格
    wb.save(filled)
    wb.close()

    report = qc.run_qc(filled, plan, template_path=template, rules=rules)
    assert report["verdict"] == "FAIL"
    assert "writer_or_plan" in report["blame"]
    conf = report["sections"]["conformance"]
    assert any("格值与计划不符" in i["detail"] for i in conf["issues"])
    # expected 三方对账也应发现 file ≠ plan
    assert report["sections"]["expected"]["status"] == "fail"
    assert "mapper" in report["blame"] or any(
        "输出文件重算" in i["detail"] for i in report["sections"]["expected"]["issues"]
    )


def test_qc_detects_clear_residue(tmp_path: Path) -> None:
    """clear 范围内塞回残值 → conformance FAIL。"""
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, template = _build_chain(tmp_path)
    wb = load_workbook(filled)
    wb["明细"].cell(4, 20).value = "偷偷残留"  # clear_zone 内、非计划格
    wb.save(filled)
    wb.close()

    report = qc.run_qc(filled, plan, template_path=template, rules=rules)
    assert report["verdict"] == "FAIL"
    assert any("残值" in i["detail"] for i in report["sections"]["conformance"]["issues"])


def test_qc_detects_mapper_false_claim(tmp_path: Path) -> None:
    """伪造 expected 自述 → expected FAIL，blame mapper。"""
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, template = _build_chain(tmp_path)
    plan2 = json.loads(json.dumps(plan))
    plan2["expected"]["cells_planned"] = 999
    plan2["expected"]["per_key_numeric_sum"]["甲"] = 123.0

    report = qc.run_qc(filled, plan2, template_path=template, rules=rules)
    assert report["verdict"] == "FAIL"
    assert report["blame"] == ["mapper"]
    issues = "；".join(i["detail"] for i in report["sections"]["expected"]["issues"])
    assert "cells_planned" in issues and "统计失真" in issues


def test_qc_detects_protection_breach(tmp_path: Path) -> None:
    """保护区（未拟合公式列）被改 → protection FAIL，blame writer。"""
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, template = _build_chain(tmp_path)
    plan2 = json.loads(json.dumps(plan))
    plan2["protected_ranges"] = ["明细!AJ:AJ"]  # 人为声明 AJ 为保护列
    # 同时从计划里去掉 AJ 的 formula_writes（模拟「AJ 不归本计划管」）
    for ph in plan2["phases"]:
        if ph["phase"] == "formula_writes":
            ph["writes"] = [w for w in ph["writes"] if not w["ref"].startswith("AJ")]

    wb = load_workbook(filled)
    wb["明细"]["AJ3"] = "公式被覆盖"
    wb.save(filled)
    wb.close()

    report = qc.run_qc(filled, plan2, template_path=template, rules=rules)
    assert report["verdict"] == "FAIL"
    assert "writer" in report["blame"]
    assert any(
        "保护区格与原模板不一致" in i["detail"] for i in report["sections"]["protection"]["issues"]
    )


def test_qc_detects_ref_error_and_dangling_sheet(tmp_path: Path) -> None:
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, template = _build_chain(tmp_path)
    wb = load_workbook(filled)
    wb["明细"]["AK1"] = "=SUM(#REF!)"
    wb["明细"]["AK2"] = "=不存在的表!A1"
    wb.save(filled)
    wb.close()

    report = qc.run_qc(filled, plan, template_path=template, rules=rules)
    assert report["verdict"] == "FAIL"
    issues = "；".join(i["detail"] for i in report["sections"]["formulas"]["issues"])
    assert "#REF!" in issues and "不存在的 sheet" in issues


def test_qc_detects_rules_ref_mismatch(tmp_path: Path) -> None:
    """规则文件被改但计划未重编 → traceability FAIL，blame pipeline。"""
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, template = _build_chain(tmp_path)
    rules2 = json.loads(json.dumps(rules))
    rules2["policy"]["value_number_format"] = "0.00"

    report = qc.run_qc(filled, plan, template_path=template, rules=rules2)
    assert report["verdict"] == "FAIL"
    assert "pipeline" in report["blame"]
    assert any(
        "规则版本不一致" in i["detail"] for i in report["sections"]["traceability"]["issues"]
    )


def test_qc_detects_structure_drift(tmp_path: Path) -> None:
    """输出键列与 rules.blocks 不符（模板重排）→ structure FAIL，blame rules_stale。"""
    _, _, qc = _load_vendor_pkgs()
    rules, plan, filled, template = _build_chain(tmp_path)
    wb = load_workbook(filled)
    wb["明细"].cell(3, 2).value = "戊"  # 甲块键被改
    wb.save(filled)
    wb.close()

    report = qc.run_qc(filled, plan, rules=rules)
    assert report["verdict"] == "FAIL"
    assert "rules_stale" in report["blame"]
    assert any("块键与规则不符" in i["detail"] for i in report["sections"]["structure"]["issues"])


# ---------------------------------------------------------------------------
# 员工入口
# ---------------------------------------------------------------------------


def test_qc_employee_run_end_to_end(tmp_path: Path) -> None:
    employee = _load_module("excel_qc_employee_run_test", QC_EMPLOYEE)
    rules, plan, filled, template = _build_chain(tmp_path)
    rules_path = tmp_path / "rules.json"
    rules_path.write_text(json.dumps(rules, ensure_ascii=False), encoding="utf-8")

    result = asyncio.run(
        employee.run(
            {
                "file_path": str(filled),
                "plan_path": str(tmp_path / "plan.json"),
                "rules_path": str(rules_path),
                "template_path": str(template),
                "workspace_root": str(tmp_path),
            },
            {"workspace_root": str(tmp_path)},
        )
    )
    assert result["ok"], result["error"]
    item = result["items"][0]
    assert item["verdict"] == "WARN"  # dropped 记录呈现
    assert Path(item["output_path"]).is_file()
    report = json.loads(Path(item["output_path"]).read_text(encoding="utf-8"))
    assert set(report["sections"]) == {
        "conformance",
        "protection",
        "expected",
        "formulas",
        "traceability",
        "structure",
        "semantic",
    }
    # 无 call_llm 的测试环境：semantic 节如实 skipped
    assert report["sections"]["semantic"]["status"] == "skipped"
    assert result["meta"]["verdict"] == "WARN"


def test_qc_employee_missing_plan_fails(tmp_path: Path) -> None:
    employee = _load_module("excel_qc_employee_noplan_test", QC_EMPLOYEE)
    _, _, filled, _ = _build_chain(tmp_path)
    result = asyncio.run(
        employee.run(
            {"file_path": str(filled), "workspace_root": str(tmp_path)},
            {"workspace_root": str(tmp_path)},
        )
    )
    assert not result["ok"]
    assert "plan" in result["error"]


# ---------------------------------------------------------------------------
# 真实太阳鸟模板全链 QC
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not SUNBIRD_TEMPLATE.is_file(), reason="真实太阳鸟模板不在本机")
def test_qc_on_real_sunbird_chain(tmp_path: Path) -> None:
    infer_mod, compile_mod, qc = _load_vendor_pkgs()
    reader = _load_module("reader_qc_sunbird", READER_CONVERT)
    writer = _load_module("writer_qc_sunbird", WRITER_CONVERT)

    wb_json = tmp_path / "wb.json"
    reader.convert_file(
        SUNBIRD_TEMPLATE,
        wb_json,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "workbook.json"},
    )
    rules = infer_mod.infer_rules(
        json.loads(wb_json.read_text(encoding="utf-8")), source_name="sunbird"
    )
    rules["bands"] = {
        "morning": {"row_offset": 0, "max_entries": 2},
        "afternoon": {"row_offset": 2, "max_entries": 2},
        "night": {"row_offset": 4, "max_entries": 2},
    }
    records = [
        {"key": "胡超", "day": 2, "band": "morning", "entries": [{"symbol": "√", "value": 4.0}]},
        {"key": "樊琪麒", "day": 8, "band": "night", "entries": [{"symbol": "☆", "value": 1.5}]},
    ]
    plan = compile_mod.compile_plan(rules, records, month_label="2026-03")
    plan_path = tmp_path / "plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")
    filled = tmp_path / "filled.xlsx"
    write_result = writer.convert_file(
        plan_path,
        filled,
        template_path=SUNBIRD_TEMPLATE,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    write_report = json.loads((filled.parent / "write_report.json").read_text(encoding="utf-8"))

    report = qc.run_qc(
        filled, plan, template_path=SUNBIRD_TEMPLATE, rules=rules, write_report=write_report
    )
    sections = report["sections"]
    assert sections["conformance"]["status"] == "pass", sections["conformance"]["issues"][:3]
    assert sections["expected"]["status"] == "pass", sections["expected"]["issues"][:3]
    assert sections["traceability"]["status"] == "pass"
    assert sections["structure"]["status"] == "pass"
    assert sections["formulas"]["stats"]["formulas_total"] >= 1280
    assert report["verdict"] in ("PASS", "WARN")
    assert write_result["violations"] == []
    # 数值对账命中
    assert sections["expected"]["stats"]["file_sum"]["胡超"] == 4.0
