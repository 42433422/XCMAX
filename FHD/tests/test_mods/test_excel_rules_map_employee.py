# mypy: disable-error-code="import-not-found, no-any-return, union-attr"
from __future__ import annotations

import asyncio
import importlib.util
import json
from pathlib import Path
from types import ModuleType

import pytest
from openpyxl import Workbook, load_workbook

FHD_ROOT = Path(__file__).resolve().parents[2]

READER_CONVERT = (
    FHD_ROOT / "mods/_employees/excel-full-read-employee/backend/vendor/excel_full_read/convert.py"
)
WRITER_CONVERT = (
    FHD_ROOT
    / "mods/_employees/excel-template-write-employee/backend/vendor/excel_template_write/convert.py"
)
MAPPER_VENDOR = FHD_ROOT / "mods/_employees/excel-rules-map-employee/backend/vendor"
MAPPER_EMPLOYEE = (
    FHD_ROOT
    / "mods/_employees/excel-rules-map-employee/backend/employees/excel_rules_map_employee.py"
)

SUNBIRD_TEMPLATE = Path(
    "/workspace/成都修茈科技有限公司/MODstore_deploy/var/employee_draft_assets/9/real-files-smoke/00_考勤-2026-3月份考勤统计表.xlsx"
)


def _load_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _load_mapper():
    """vendor 包内相对导入：以包形式加载 excel_rules_map。"""
    import sys

    if str(MAPPER_VENDOR) not in sys.path:
        sys.path.insert(0, str(MAPPER_VENDOR))
    import excel_rules_map.compile_plan as compile_mod
    import excel_rules_map.convert as convert_mod
    import excel_rules_map.infer as infer_mod

    return infer_mod, compile_mod, convert_mod


def _synthetic_template(path: Path) -> None:
    """合成通用模板：2 行表头（年/月 + 日历 1..15×2）、4 块×3 行、2 公式列。

    布局：A 列块标签（竖向合并 3 行）、B 列键、C 列备注、日历 D..（col 4 起，
    day15 占 col 32/33）、公式区 AI..AJ（col 35 等差公式 / col 36 常数引用公式）。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.cell(1, 2, 2026)
    ws.cell(1, 4, 7)
    for day in range(1, 16):
        ws.cell(2, 4 + (day - 1) * 2, day)
    keys = ["甲", "乙", "丙", "丁"]
    for i, key in enumerate(keys):
        top = 3 + i * 3
        ws.cell(top, 1, "一组" if i < 2 else "二组")
        ws.merge_cells(start_row=top, start_column=1, end_row=top + 2, end_column=1)
        ws.cell(top, 2, key)
        ws.merge_cells(start_row=top, start_column=2, end_row=top + 2, end_column=2)
        ws.cell(top, 3, "备注")
        # 等差公式列 AI：数据区行号随块平移；常数公式列 AJ
        ws.cell(top, 35, f"=SUM(D{top}:AF{top + 2})")
        ws.cell(top, 36, "=COUNT($D$2:$AF$2)")
    wb.save(path)
    wb.close()


def _reader_json(reader, xlsx: Path, out: Path) -> dict:
    reader.convert_file(
        xlsx, out, payload={}, ctx={}, rule_spec={"default_output_relpath": "workbook.json"}
    )
    return json.loads(out.read_text(encoding="utf-8"))


# ---------------------------------------------------------------------------
# infer：合成模板全字段
# ---------------------------------------------------------------------------


def test_infer_on_synthetic_template(tmp_path: Path) -> None:
    infer_mod, _, _ = _load_mapper()
    reader = _load_module("reader_for_infer_synth", READER_CONVERT)
    xlsx = tmp_path / "template.xlsx"
    _synthetic_template(xlsx)
    workbook = _reader_json(reader, xlsx, tmp_path / "wb.json")

    rules = infer_mod.infer_rules(workbook, source_name="synthetic")
    tm = rules["template_map"]

    assert tm["sheet"] == "明细"
    assert tm["block"] == {"rows": 3, "first_top": 3, "count": 4}
    assert tm["header_rows"] == 2
    assert tm["key_col"] == 2
    assert [b["key"] for b in tm["blocks"]] == ["甲", "乙", "丙", "丁"]

    cal = tm["calendar"]
    assert cal["anchor_col"] == 4
    assert cal["slots_per_day"] == 2
    assert cal["day_count"] == 15
    assert cal["layout"] == "symbol_value"

    assert tm["month_cells"] == [
        {"ref": "B1", "part": "year"},
        {"ref": "D1", "part": "month"},
    ]

    assert tm["formula_zones"] == [{"col_start": 35, "col_end": 36}]
    assert tm["clear_zone"] == {"col_start": 4, "col_end": 34}

    templates = {t["col"]: t for t in rules["formula_templates"]}
    assert set(templates) == {"AI", "AJ"}
    ai_params = templates["AI"]["params"]
    linear = [p for p in ai_params if p["kind"] == "linear"]
    assert linear and all(p["step"] == 3 for p in linear)
    assert all(p["kind"] == "const" for p in templates["AJ"]["params"])
    # 实例化第 2 块（top=9）应还原原公式
    assert infer_mod.instantiate_formula(templates["AI"], 2) == "=SUM(D9:AF11)"
    assert infer_mod.instantiate_formula(templates["AJ"], 3) == "=COUNT($D$2:$AF$2)"

    assert rules["evidence"]["confidences"]["block"] > 0.7
    assert isinstance(rules["evidence"]["open_questions"], list)


# ---------------------------------------------------------------------------
# compile：坐标/公式/expected/防线
# ---------------------------------------------------------------------------


def _rules_for_compile(infer_mod, reader, tmp_path: Path) -> dict:
    xlsx = tmp_path / "template.xlsx"
    if not xlsx.exists():
        _synthetic_template(xlsx)
    workbook = _reader_json(reader, xlsx, tmp_path / "wb_compile.json")
    rules = infer_mod.infer_rules(workbook, source_name="synthetic")
    # 人工确认环节：定义双带布局（上午第 1 行 / 下午第 2 行 / 合计第 3 行留白）
    rules["bands"] = {
        "am": {"row_offset": 0, "max_entries": 1},
        "pm": {"row_offset": 1, "max_entries": 1},
    }
    return rules


def test_compile_plan_coordinates_and_expected(tmp_path: Path) -> None:
    infer_mod, compile_mod, _ = _load_mapper()
    reader = _load_module("reader_for_compile_synth", READER_CONVERT)
    rules = _rules_for_compile(infer_mod, reader, tmp_path)

    records = [
        {"key": "甲", "day": 1, "band": "am", "entries": [{"symbol": "√", "value": 2.0}]},
        {"key": "乙", "day": 3, "band": "pm", "entries": [{"symbol": "☆", "value": 1.5}]},
        {"key": "甲", "cells": [{"col": "C", "row_offset": 2, "value": "手工备注"}]},
        {"key": "路人", "day": 2, "band": "am", "entries": [{"symbol": "√", "value": 1}]},
        {"key": "丙", "day": 99, "band": "am", "entries": [{"symbol": "√", "value": 1}]},
        {"key": "丁", "day": 2, "band": "夜班", "entries": [{"symbol": "★", "value": 1}]},
    ]
    plan = compile_mod.compile_plan(rules, records, month_label="2026-07")

    phases = {p["phase"]: p for p in plan["phases"]}
    assert list(phases) == ["clear_ranges", "cell_writes", "formula_writes"]
    assert phases["clear_ranges"]["ranges"] == ["明细!D3:AH14"]

    writes = phases["cell_writes"]["writes"]
    by_pos = {(w["sheet"], w["row"], w["col"]): w for w in writes}
    # month_cells
    assert by_pos[("明细", 1, 2)]["value"] == 2026
    assert by_pos[("明细", 1, 4)]["value"] == 7
    # 甲 day1 am：块 top=3，符号列=4，数值列=5
    assert by_pos[("明细", 3, 4)]["value"] == "√"
    assert by_pos[("明细", 3, 5)]["value"] == 2.0
    assert by_pos[("明细", 3, 5)]["number_format"] == "0.0"
    # 乙 day3 pm：top=6 offset1=7，符号列=4+(3-1)*2=8
    assert by_pos[("明细", 7, 8)]["value"] == "☆"
    assert by_pos[("明细", 7, 9)]["value"] == 1.5
    # 直写：甲块 top=3 offset2=5，C 列
    assert by_pos[("明细", 5, 3)]["value"] == "手工备注"

    formulas = {w["ref"]: w["formula"] for w in phases["formula_writes"]["writes"]}
    assert formulas["AI3"] == "=SUM(D3:AF5)"
    assert formulas["AI9"] == "=SUM(D9:AF11)"
    assert formulas["AJ12"] == "=COUNT($D$2:$AF$2)"
    assert len(formulas) == 8  # 2 列 × 4 块

    expected = plan["expected"]
    assert expected["records_in"] == 6
    reasons = "；".join(d["reason"] for d in expected["records_dropped"])
    assert "键不在模板块清单中" in reasons
    assert "day 越界" in reasons
    assert "band 未定义" in reasons
    assert expected["keys_unmatched_source"] == ["路人"]
    assert expected["per_key_numeric_sum"]["甲"] == 2.0
    assert "乙" in expected["per_key_numeric_sum"]
    # 公式列全部被模板覆盖 → 无剩余保护区
    assert plan["protected_ranges"] == []
    assert plan["meta"]["rules_ref"]["sha256"]

    # rules_ref 随规则内容变化
    rules2 = json.loads(json.dumps(rules))
    rules2["policy"]["value_number_format"] = "0.00"
    plan2 = compile_mod.compile_plan(rules2, records[:1])
    assert plan2["meta"]["rules_ref"]["sha256"] != plan["meta"]["rules_ref"]["sha256"]


def test_compile_validation_errors(tmp_path: Path) -> None:
    infer_mod, compile_mod, _ = _load_mapper()
    reader = _load_module("reader_for_compile_err", READER_CONVERT)
    rules = _rules_for_compile(infer_mod, reader, tmp_path)

    with pytest.raises(ValueError, match="rules_version"):
        compile_mod.compile_plan({"rules_version": 9}, [])
    broken = json.loads(json.dumps(rules))
    broken["template_map"]["calendar"] = None
    with pytest.raises(ValueError, match="calendar"):
        compile_mod.compile_plan(
            broken, [{"key": "甲", "day": 1, "band": "am", "entries": [{"value": 1}]}]
        )
    with pytest.raises(ValueError, match="month_label"):
        compile_mod.compile_plan(rules, [], month_label="2026年7月")
    with pytest.raises(ValueError, match="空计划"):
        no_zone = json.loads(json.dumps(rules))
        no_zone["template_map"]["clear_zone"] = None
        no_zone["formula_templates"] = []
        compile_mod.compile_plan(no_zone, [])


# ---------------------------------------------------------------------------
# e2e：读取员 → infer → compile → 写入员 → 重读
# ---------------------------------------------------------------------------


def test_full_chain_read_infer_compile_write_reread(tmp_path: Path) -> None:
    infer_mod, compile_mod, _ = _load_mapper()
    reader = _load_module("reader_e2e_rules_map", READER_CONVERT)
    writer = _load_module("writer_e2e_rules_map", WRITER_CONVERT)

    rules = _rules_for_compile(infer_mod, reader, tmp_path)
    records = [
        {"key": "甲", "day": 1, "band": "am", "entries": [{"symbol": "√", "value": 2.0}]},
        {"key": "丁", "day": 15, "band": "pm", "entries": [{"symbol": "★", "value": 3.0}]},
    ]
    plan = compile_mod.compile_plan(rules, records, month_label="2026-07")
    plan_path = tmp_path / "plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")

    filled = tmp_path / "filled.xlsx"
    result = writer.convert_file(
        plan_path,
        filled,
        template_path=tmp_path / "template.xlsx",
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    assert result["violations"] == []
    assert result["expected"]["per_key_numeric_sum"] == {"甲": 2.0, "丁": 3.0}

    wb = load_workbook(filled, data_only=False)
    ws = wb["明细"]
    assert ws.cell(1, 2).value == 2026 and ws.cell(1, 4).value == 7
    assert ws.cell(3, 4).value == "√" and ws.cell(3, 5).value == 2.0
    # 丁块 top=12，pm 行=13；day15 符号列=4+14*2=32，数值列=33
    assert ws.cell(13, 32).value == "★" and ws.cell(13, 33).value == 3.0
    assert ws.cell(3, 35).value == "=SUM(D3:AF5)"
    assert ws.cell(12, 35).value == "=SUM(D12:AF14)"
    # 模板合并保留
    assert "A3:A5" in {str(r) for r in ws.merged_cells.ranges}
    wb.close()


def test_calendar_slot_cannot_enter_formula_zone(tmp_path: Path) -> None:
    infer_mod, compile_mod, _ = _load_mapper()
    reader = _load_module("reader_guard_rules_map", READER_CONVERT)
    rules = _rules_for_compile(infer_mod, reader, tmp_path)
    # 人为把公式区前沿改到 col33：day15 值列=33 落入 → 必须被丢弃，且剩余公式列进保护区
    rules["template_map"]["formula_zones"] = [{"col_start": 33, "col_end": 36}]
    plan = compile_mod.compile_plan(
        rules,
        [{"key": "丁", "day": 15, "band": "pm", "entries": [{"symbol": "★", "value": 3.0}]}],
    )
    expected = plan["expected"]
    assert expected["cells_planned"] == 0
    assert any("公式区" in d["reason"] for d in expected["records_dropped"])
    # 模板未覆盖的公式列 AG..AH（33..34）成为保护区
    assert plan["protected_ranges"] == ["明细!AG:AH"]


# ---------------------------------------------------------------------------
# 员工入口（async run）
# ---------------------------------------------------------------------------


def test_employee_run_infer_and_compile(tmp_path: Path) -> None:
    employee = _load_module("excel_rules_map_employee_run_test", MAPPER_EMPLOYEE)
    reader = _load_module("reader_employee_rules_map", READER_CONVERT)
    xlsx = tmp_path / "template.xlsx"
    _synthetic_template(xlsx)
    wb_json = tmp_path / "wb.json"
    _reader_json(reader, xlsx, wb_json)

    r1 = asyncio.run(
        employee.run(
            {"file_path": str(wb_json), "workspace_root": str(tmp_path)},
            {"workspace_root": str(tmp_path)},
        )
    )
    assert r1["ok"], r1["error"]
    item1 = r1["items"][0]
    assert item1["action"] == "infer"
    rules_path = Path(item1["output_path"])
    assert rules_path.is_file()

    rules = json.loads(rules_path.read_text(encoding="utf-8"))
    rules["bands"] = {"am": {"row_offset": 0, "max_entries": 1}}
    r2 = asyncio.run(
        employee.run(
            {
                "rules": rules,
                "records": [
                    {"key": "甲", "day": 1, "band": "am", "entries": [{"symbol": "√", "value": 2}]}
                ],
                "workspace_root": str(tmp_path),
                "output_relpath": "outputs/plan.json",
            },
            {"workspace_root": str(tmp_path)},
        )
    )
    assert r2["ok"], r2["error"]
    item2 = r2["items"][0]
    assert item2["action"] == "compile"
    assert Path(item2["output_path"]).is_file()
    assert item2["rules_ref"]["sha256"]


# ---------------------------------------------------------------------------
# 真实太阳鸟模板（金样级：公式实例化与模板原文一致）
# ---------------------------------------------------------------------------


@pytest.mark.skipif(not SUNBIRD_TEMPLATE.is_file(), reason="真实太阳鸟模板不在本机")
def test_infer_on_real_sunbird_template(tmp_path: Path) -> None:
    infer_mod, compile_mod, _ = _load_mapper()
    reader = _load_module("reader_real_sunbird", READER_CONVERT)
    workbook = _reader_json(reader, SUNBIRD_TEMPLATE, tmp_path / "sunbird_wb.json")

    rules = infer_mod.infer_rules(workbook, source_name="sunbird")
    tm = rules["template_map"]

    assert tm["block"] == {"rows": 6, "first_top": 4, "count": 151}
    assert tm["header_rows"] == 3
    assert tm["key_col"] == 3  # C 列姓名
    assert tm["blocks"][0]["key"] == "胡超"

    cal = tm["calendar"]
    assert cal == {
        "anchor_col": 5,
        "slots_per_day": 2,
        "day_count": 31,
        "header_row": 3,
        "layout": "symbol_value",
    }
    assert tm["month_cells"] == [
        {"ref": "M1", "part": "year"},
        {"ref": "S1", "part": "month"},
    ]
    # 公式区 BR..CG（70..85）；clear_zone 与太阳鸟本尊 DETAIL_SUM_COL_START..69 一致
    assert tm["formula_zones"] == [{"col_start": 70, "col_end": 85}]
    assert tm["clear_zone"] == {"col_start": 5, "col_end": 69}

    template_cols = {t["col"] for t in rules["formula_templates"]}
    assert {"BR", "CC", "CD"}.issubset(template_cols)

    # 金样对照：实例化公式必须与模板原文逐字一致
    src_wb = load_workbook(SUNBIRD_TEMPLATE)
    ws = src_wb["明细"]
    by_col = {t["col"]: t for t in rules["formula_templates"]}
    for col_letter, block_index, row in (("BR", 0, 4), ("BR", 1, 10), ("CD", 2, 16), ("CC", 3, 22)):
        got = infer_mod.instantiate_formula(by_col[col_letter], block_index)
        assert got == ws[f"{col_letter}{row}"].value, f"{col_letter}{row} 公式实例化不一致"
    src_wb.close()

    # compile 冒烟：真实 151 块 × 拟合列全部实例化
    rules["bands"] = {
        "morning": {"row_offset": 0, "max_entries": 2},
        "afternoon": {"row_offset": 2, "max_entries": 2},
        "night": {"row_offset": 4, "max_entries": 2},
    }
    plan = compile_mod.compile_plan(
        rules,
        [
            {"key": "胡超", "day": 1, "band": "morning", "entries": [{"symbol": "√", "value": 2}]},
            {"key": "胡超", "day": 1, "band": "night", "entries": [{"symbol": "☆", "value": 1.5}]},
        ],
        month_label="2026-03",
    )
    phases = {p["phase"]: p for p in plan["phases"]}
    writes = {(w["row"], w["col"]): w for w in phases["cell_writes"]["writes"]}
    assert writes[(4, 5)]["value"] == "√"  # 胡超块 top=4，day1 符号列=5
    assert writes[(4, 6)]["value"] == 2
    assert writes[(8, 5)]["value"] == "☆"  # night row_offset=4
    formulas = {w["ref"]: w["formula"] for w in phases["formula_writes"]["writes"]}
    assert formulas["BR4"].startswith("=SUMIF(OFFSET($E$4:$BM$9,ROWS($1:1)*6-6,)")
    assert formulas["BR10"].startswith("=SUMIF(OFFSET($E$4:$BM$9,ROWS($1:7)*6-6,)")
    assert plan["expected"]["keys_matched"] == 1
    # 151 块中 80 块有姓名，其中胡超有记录 → 79 个在住块无记录
    assert len(plan["expected"]["blocks_without_records"]) == 79
