# mypy: disable-error-code="arg-type, import-not-found, no-any-return, union-attr"
"""固化循环测试：LLM 写 records 生产者脚本 → 金样对账 → 固化。

金样判据全确定性（反读器 + diff）；LLM 用 mock 演练协作契约：
一轮通过 / 反馈迭代 / 黑名单拒绝 / 屡试不过失败留证。
太阳鸟端到端：金样由太阳鸟单体（参考实现）生成，验证固化管道能逐槽复现其转化。
"""

from __future__ import annotations

import asyncio
import importlib.util
import json
from pathlib import Path
from types import ModuleType

import pytest
from openpyxl import Workbook

FHD_ROOT = Path(__file__).resolve().parents[2]

READER_CONVERT = (
    FHD_ROOT / "mods/_employees/excel-full-read-employee/backend/vendor/excel_full_read/convert.py"
)
WRITER_CONVERT = (
    FHD_ROOT
    / "mods/_employees/excel-template-write-employee/backend/vendor/excel_template_write/convert.py"
)
MAPPER_VENDOR = FHD_ROOT / "mods/_employees/excel-rules-map-employee/backend/vendor"

SUNBIRD_DIR = Path(
    "/workspace/成都修茈科技有限公司/MODstore_deploy/var/employee_draft_assets/9/real-files-smoke"
)
SUNBIRD_TEMPLATE = SUNBIRD_DIR / "00_考勤-2026-3月份考勤统计表.xlsx"
SUNBIRD_SOURCE = SUNBIRD_DIR / "01_钉钉导出来的考勤数据.xlsx"
SUNBIRD_VENDOR = Path("/workspace/成都修茈科技有限公司/taiyangniao-pro/backend")


def _load_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _load_pkg():
    import sys

    if str(MAPPER_VENDOR) not in sys.path:
        sys.path.insert(0, str(MAPPER_VENDOR))
    import excel_rules_map.compile_plan as compile_mod
    import excel_rules_map.convert as convert_mod
    import excel_rules_map.golden as golden_mod
    import excel_rules_map.infer as infer_mod
    import excel_rules_map.solidify as solidify_mod

    return infer_mod, compile_mod, golden_mod, solidify_mod, convert_mod


def _synthetic_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "明细"
    ws.cell(1, 2, 2026)
    ws.cell(1, 4, 7)
    for day in range(1, 16):
        ws.cell(2, 4 + (day - 1) * 2, day)
    for i, key in enumerate(["甲", "乙", "丙", "丁"]):
        top = 3 + i * 3
        ws.cell(top, 1, "组")
        ws.merge_cells(start_row=top, start_column=1, end_row=top + 2, end_column=1)
        ws.cell(top, 2, key)
        ws.merge_cells(start_row=top, start_column=2, end_row=top + 2, end_column=2)
        ws.cell(top, 35, f"=SUM(D{top}:AF{top + 2})")
    wb.save(path)
    wb.close()


def _read_json(reader, xlsx: Path, out: Path) -> dict:
    reader.convert_file(
        xlsx, out, payload={}, ctx={}, rule_spec={"default_output_relpath": "workbook.json"}
    )
    return json.loads(out.read_text(encoding="utf-8"))


def _rules_with_bands(infer_mod, reader, tmp_path: Path) -> dict:
    xlsx = tmp_path / "template.xlsx"
    if not xlsx.exists():
        _synthetic_template(xlsx)
    workbook = _read_json(reader, xlsx, tmp_path / "tpl_wb.json")
    rules = infer_mod.infer_rules(workbook, source_name="synthetic")
    rules["bands"] = {
        "am": {"row_offset": 0, "max_entries": 1},
        "pm": {"row_offset": 1, "max_entries": 1},
    }
    return rules


def _fake_llm_seq(responses: list):
    """按调用顺序返回预设响应（str 直接作为 content）。"""
    calls: list = []

    async def call_llm(messages, **kwargs):
        idx = min(len(calls), len(responses) - 1)
        calls.append(messages)
        content = responses[idx]
        return {"ok": True, "content": content, "error": ""}

    call_llm.calls = calls  # type: ignore[attr-defined]
    return call_llm


# ---------------------------------------------------------------------------
# 反读器与 diff
# ---------------------------------------------------------------------------


def test_golden_reverse_read_roundtrip(tmp_path: Path) -> None:
    """compile → 写入员 → 读取员重读 → 反读 records 与原 records 逐槽一致。"""
    infer_mod, compile_mod, golden_mod, _, _ = _load_pkg()
    reader = _load_module(f"reader_rt_{tmp_path.name}", READER_CONVERT)
    writer = _load_module(f"writer_rt_{tmp_path.name}", WRITER_CONVERT)
    rules = _rules_with_bands(infer_mod, reader, tmp_path)

    records = [
        {"key": "甲", "day": 1, "band": "am", "entries": [{"symbol": "√", "value": 2.0}]},
        {"key": "甲", "day": 3, "band": "pm", "entries": [{"symbol": "☆", "value": 1.5}]},
        {"key": "丁", "day": 15, "band": "am", "entries": [{"symbol": "★", "value": 4.0}]},
    ]
    plan = compile_mod.compile_plan(rules, records, month_label="2026-07")
    plan_path = tmp_path / "plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")
    filled = tmp_path / "filled.xlsx"
    writer.convert_file(
        plan_path,
        filled,
        template_path=tmp_path / "template.xlsx",
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    filled_wb = _read_json(reader, filled, tmp_path / "filled_wb.json")

    extracted = golden_mod.extract_records_from_workbook(filled_wb, rules)
    diff = golden_mod.diff_records(extracted, records)
    assert diff["ok"], diff
    assert diff["stats"]["matched"] == 3


def test_diff_records_reports_mismatch_missing_extra() -> None:
    _, _, golden_mod, _, _ = _load_pkg()
    expected = [
        {"key": "甲", "day": 1, "band": "am", "entries": [{"symbol": "√", "value": 2.0}]},
        {"key": "乙", "day": 2, "band": "am", "entries": [{"symbol": "√", "value": 1.0}]},
    ]
    produced = [
        {
            "key": "甲",
            "day": 1,
            "band": "am",
            "entries": [{"symbol": "√", "value": 9.9}],
        },  # mismatch
        {"key": "丙", "day": 3, "band": "am", "entries": [{"symbol": "√", "value": 1.0}]},  # extra
    ]
    diff = golden_mod.diff_records(produced, expected)
    assert not diff["ok"]
    assert diff["stats"] == {
        "expected_slots": 2,
        "produced_slots": 2,
        "matched": 0,
        "mismatched": 1,
        "missing": 1,
        "extra": 1,
    }


# ---------------------------------------------------------------------------
# 固化循环（合成场景）
# ---------------------------------------------------------------------------

_GOOD_SCRIPT = """```python
def produce_records(source_workbook: dict, rules: dict) -> list[dict]:
    records = []
    for sheet in source_workbook.get("sheets") or []:
        for row in sheet.get("rows") or []:
            name = str(row.get("姓名") or "").strip()
            if not name:
                continue
            day = int(row.get("日期"))
            band = "am" if str(row.get("时段")).strip() == "上午" else "pm"
            hours = float(row.get("工时") or 0)
            symbol = "√" if hours >= 4 else "☆"
            records.append({"key": name, "day": day, "band": band,
                            "entries": [{"symbol": symbol, "value": hours}]})
    return records
```"""

_BAD_BAND_SCRIPT = """```python
def produce_records(source_workbook: dict, rules: dict) -> list[dict]:
    records = []
    for sheet in source_workbook.get("sheets") or []:
        for row in sheet.get("rows") or []:
            name = str(row.get("姓名") or "").strip()
            if not name:
                continue
            records.append({"key": name, "day": int(row.get("日期")), "band": "夜班",
                            "entries": [{"symbol": "√", "value": 1.0}]})
    return records
```"""

_FORBIDDEN_SCRIPT = """```python
import subprocess

def produce_records(source_workbook: dict, rules: dict) -> list[dict]:
    subprocess.run(["ls"])
    return []
```"""


def _source_workbook() -> dict:
    return {
        "source": "punch.xlsx",
        "sheets": [
            {
                "name": "打卡",
                "header_row": 1,
                "columns": ["姓名", "日期", "时段", "工时"],
                "rows": [
                    {"姓名": "甲", "日期": 1, "时段": "上午", "工时": 4.0},
                    {"姓名": "甲", "日期": 1, "时段": "下午", "工时": 2.0},
                    {"姓名": "乙", "日期": 2, "时段": "上午", "工时": 4.0},
                ],
                "row_count": 3,
            }
        ],
    }


def _expected_records() -> list:
    return [
        {"key": "甲", "day": 1, "band": "am", "entries": [{"symbol": "√", "value": 4.0}]},
        {"key": "甲", "day": 1, "band": "pm", "entries": [{"symbol": "☆", "value": 2.0}]},
        {"key": "乙", "day": 2, "band": "am", "entries": [{"symbol": "√", "value": 4.0}]},
    ]


def test_solidify_succeeds_first_round(tmp_path: Path) -> None:
    infer_mod, _, _, solidify_mod, _ = _load_pkg()
    reader = _load_module(f"reader_s1_{tmp_path.name}", READER_CONVERT)
    rules = _rules_with_bands(infer_mod, reader, tmp_path)
    call_llm = _fake_llm_seq([_GOOD_SCRIPT])

    result = asyncio.run(
        solidify_mod.solidify_transform(_source_workbook(), rules, _expected_records(), call_llm)
    )
    assert result["ok"], result["iterations"]
    assert len(result["iterations"]) == 1
    assert result["diff"]["stats"]["matched"] == 3
    assert "def produce_records" in result["script"]
    assert result["script_sha256"]


def test_solidify_iterates_on_feedback(tmp_path: Path) -> None:
    """第一轮 band 越界 → 契约校验失败反馈；第二轮修正 → 通过，iterations=2。"""
    infer_mod, _, _, solidify_mod, _ = _load_pkg()
    reader = _load_module(f"reader_s2_{tmp_path.name}", READER_CONVERT)
    rules = _rules_with_bands(infer_mod, reader, tmp_path)
    call_llm = _fake_llm_seq([_BAD_BAND_SCRIPT, _GOOD_SCRIPT])

    result = asyncio.run(
        solidify_mod.solidify_transform(_source_workbook(), rules, _expected_records(), call_llm)
    )
    assert result["ok"]
    assert len(result["iterations"]) == 2
    assert "band 未定义" in result["iterations"][0]["error"]
    # 第二轮 prompt 里带了失败反馈
    second_prompt = call_llm.calls[1][1]["content"]  # type: ignore[attr-defined]
    assert "上一轮失败反馈" in second_prompt


def test_solidify_rejects_forbidden_and_fails_out(tmp_path: Path) -> None:
    infer_mod, _, _, solidify_mod, _ = _load_pkg()
    reader = _load_module(f"reader_s3_{tmp_path.name}", READER_CONVERT)
    rules = _rules_with_bands(infer_mod, reader, tmp_path)
    call_llm = _fake_llm_seq([_FORBIDDEN_SCRIPT])

    result = asyncio.run(
        solidify_mod.solidify_transform(
            _source_workbook(), rules, _expected_records(), call_llm, max_iterations=2
        )
    )
    assert not result["ok"]
    assert len(result["iterations"]) == 2
    assert all("禁止的操作" in it["error"] for it in result["iterations"])


def test_solidify_convert_action_end_to_end(tmp_path: Path) -> None:
    """convert 层 solidify：组合 JSON 上传 → transform.py + solidify_report.json 落盘。"""
    infer_mod, _, _, _, convert_mod = _load_pkg()
    reader = _load_module(f"reader_s4_{tmp_path.name}", READER_CONVERT)
    rules = _rules_with_bands(infer_mod, reader, tmp_path)

    bundle = {
        "source_workbook": _source_workbook(),
        "expected_records": _expected_records(),
        "rules": rules,
    }
    bundle_path = tmp_path / "solidify_input.json"
    bundle_path.write_text(json.dumps(bundle, ensure_ascii=False), encoding="utf-8")

    result = asyncio.run(
        convert_mod.convert_file(
            bundle_path,
            tmp_path / "outputs" / "rules.json",
            payload={},
            ctx={"call_llm": _fake_llm_seq([_GOOD_SCRIPT])},
            rule_spec={"default_output_relpath": "outputs/rules.json"},
        )
    )
    assert result["action"] == "solidify"
    assert result["iterations"] == 1
    assert result["diff_stats"]["matched"] == 3
    script = Path(result["output_path"]).read_text(encoding="utf-8")
    assert "def produce_records" in script
    report = json.loads(Path(result["report_path"]).read_text(encoding="utf-8"))
    assert report["ok"] and report["script_sha256"] == result["script_sha256"]
    assert report["rules_ref"]["sha256"]

    # 固化产物可直接复跑（零 LLM）：加载脚本重算 records
    namespace: dict = {}
    exec(compile(script, "<t>", "exec"), namespace)  # noqa: S102
    records = namespace["produce_records"](_source_workbook(), rules)
    assert len(records) == 3

    # 无 LLM 时 solidify 明确报错
    with pytest.raises(ValueError, match="LLM"):
        asyncio.run(
            convert_mod.convert_file(
                bundle_path,
                tmp_path / "outputs2" / "rules.json",
                payload={},
                ctx={},
                rule_spec={"default_output_relpath": "outputs/rules.json"},
            )
        )


# ---------------------------------------------------------------------------
# 太阳鸟金样端到端：固化脚本复现太阳鸟单体转化
# ---------------------------------------------------------------------------


SOLIDIFIED_EXAMPLE = (
    FHD_ROOT / "mods/_employees/excel-rules-map-employee/examples/sunbird-solidified/transform.py"
)


@pytest.mark.skipif(
    not (SUNBIRD_TEMPLATE.is_file() and SUNBIRD_SOURCE.is_file() and SUNBIRD_VENDOR.is_dir()),
    reason="真实太阳鸟文件/vendor 不在本机",
)
def test_solidified_example_passes_golden_gate(tmp_path: Path) -> None:
    """LLM 亲笔固化脚本（examples/sunbird-solidified）常驻回归：金样对账必须全绿。

    金样 = 太阳鸟单体输出；脚本零 taiyangniao 依赖，纯从源 workbook.json 计算。
    该测试同时守护固化脚本与金样反读器/records 契约的兼容性。
    """
    import sys

    infer_mod, _, golden_mod, _, _ = _load_pkg()
    reader = _load_module("reader_solidified_example", READER_CONVERT)

    if str(SUNBIRD_VENDOR) not in sys.path:
        sys.path.insert(0, str(SUNBIRD_VENDOR))
    from taiyangniao_attendance.convert import convert_attendance_file

    golden_xlsx = tmp_path / "golden.xlsx"
    assert convert_attendance_file(
        str(SUNBIRD_SOURCE), str(golden_xlsx), template_path=str(SUNBIRD_TEMPLATE)
    )["success"]

    rules = infer_mod.infer_rules(
        _read_json(reader, SUNBIRD_TEMPLATE, tmp_path / "tpl_wb.json"), source_name="sunbird"
    )
    rules["bands"] = {
        "morning": {"row_offset": 0, "max_entries": 2},
        "afternoon": {"row_offset": 2, "max_entries": 2},
        "night": {"row_offset": 4, "max_entries": 2},
    }
    rules["template_map"]["clear_zone"]["col_end"] = 67
    expected = golden_mod.extract_records_from_workbook(
        _read_json(reader, golden_xlsx, tmp_path / "golden_wb.json"), rules
    )

    code = SOLIDIFIED_EXAMPLE.read_text(encoding="utf-8")
    assert "taiyangniao" not in code.replace("taiyangniao 依赖", "")
    namespace: dict = {}
    exec(compile(code, str(SOLIDIFIED_EXAMPLE), "exec"), namespace)  # noqa: S102
    source_wb = _read_json(reader, SUNBIRD_SOURCE, tmp_path / "src_wb.json")
    records = namespace["produce_records"](source_wb, rules)

    diff = golden_mod.diff_records(records, expected)
    assert diff["ok"], diff["stats"]
    assert diff["stats"]["matched"] == len(expected) > 3000


@pytest.mark.skipif(
    not (SUNBIRD_TEMPLATE.is_file() and SUNBIRD_SOURCE.is_file() and SUNBIRD_VENDOR.is_dir()),
    reason="真实太阳鸟文件/vendor 不在本机",
)
def test_solidify_reproduces_sunbird_conversion(tmp_path: Path) -> None:
    import sys

    infer_mod, compile_mod, golden_mod, solidify_mod, _ = _load_pkg()
    reader = _load_module("reader_sunbird_solidify", READER_CONVERT)
    writer = _load_module("writer_sunbird_solidify", WRITER_CONVERT)

    # 1) 金样：太阳鸟单体（参考实现）跑真实钉钉表 + 模板
    if str(SUNBIRD_VENDOR) not in sys.path:
        sys.path.insert(0, str(SUNBIRD_VENDOR))
    from taiyangniao_attendance.convert import convert_attendance_file

    golden_xlsx = tmp_path / "golden.xlsx"
    golden_result = convert_attendance_file(
        str(SUNBIRD_SOURCE), str(golden_xlsx), template_path=str(SUNBIRD_TEMPLATE)
    )
    assert golden_result["success"], golden_result
    golden_wb = _read_json(reader, golden_xlsx, tmp_path / "golden_wb.json")

    # 2) 结构规则：真实模板 infer + 人确认环节——补三带、并按 open_question 提示
    #    把 clear_zone 右界修到 67（68/69 是太阳鸟侧栏序号/姓名列，非日历数据区）
    tpl_wb = _read_json(reader, SUNBIRD_TEMPLATE, tmp_path / "tpl_wb.json")
    rules = infer_mod.infer_rules(tpl_wb, source_name="sunbird")
    assert any("clear_zone" in q for q in rules["evidence"]["open_questions"])
    rules["bands"] = {
        "morning": {"row_offset": 0, "max_entries": 2},
        "afternoon": {"row_offset": 2, "max_entries": 2},
        "night": {"row_offset": 4, "max_entries": 2},
    }
    rules["template_map"]["clear_zone"]["col_end"] = 67

    # 3) 金样反读出期望 records（确定性判据）
    expected_records = golden_mod.extract_records_from_workbook(golden_wb, rules)
    assert expected_records, "金样反读不能为空"

    # 4) mock LLM「生成」桥接脚本：调太阳鸟真实算子转 records（验证窄接口契约
    #    足以无损承载太阳鸟转化；真 LLM 场景由宿主环境执行同一循环）
    bridge_script = f"""```python
import sys

def produce_records(source_workbook: dict, rules: dict) -> list[dict]:
    sys.path.insert(0, {str(SUNBIRD_VENDOR)!r})
    from taiyangniao_attendance.convert import _aggregate_employee_records, _filter_records_to_template_roster
    from taiyangniao_attendance.mapper import build_template_profiles
    from taiyangniao_attendance.parser import parse_attendance_workbook
    from openpyxl import load_workbook

    parsed = parse_attendance_workbook({str(SUNBIRD_SOURCE)!r})
    wb = load_workbook({str(SUNBIRD_TEMPLATE)!r})
    ws = wb["明细"]
    profiles = build_template_profiles(ws)
    wb.close()
    filtered = _filter_records_to_template_roster(parsed.records, profiles)
    employees, _, _ = _aggregate_employee_records(filtered, template_profiles=profiles)

    records = []
    for name, month in employees.items():
        for day, dp in month.days.items():
            for band, entries in (("morning", dp.morning), ("afternoon", dp.afternoon), ("night", dp.night)):
                trimmed = [
                    {{"symbol": e.symbol, "value": round(float(e.value), 1)}}
                    for e in entries[:2]
                ]
                if trimmed:
                    records.append({{"key": name, "day": int(day), "band": band, "entries": trimmed}})
    return records
```"""

    result = asyncio.run(
        solidify_mod.solidify_transform(
            {"source": SUNBIRD_SOURCE.name, "sheets": [], "_note": "桥接脚本自读源表"},
            rules,
            expected_records,
            _fake_llm_seq([bridge_script]),
            max_iterations=2,
        )
    )
    assert result["ok"], (result["iterations"], (result.get("diff") or {}).get("samples"))
    assert result["diff"]["stats"]["mismatched"] == 0
    assert result["diff"]["stats"]["missing"] == 0
    assert result["diff"]["stats"]["extra"] == 0

    # 5) 固化脚本产 records → 白盒管道 → 与金样数据区逐格一致（端到端验收）
    plan = compile_mod.compile_plan(
        rules, result["records"], month_label=str(golden_result.get("month") or "2026-03")
    )
    plan_path = tmp_path / "plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")
    ours_xlsx = tmp_path / "ours.xlsx"
    writer.convert_file(
        plan_path,
        ours_xlsx,
        template_path=SUNBIRD_TEMPLATE,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )

    from openpyxl import load_workbook

    wb_g = load_workbook(golden_xlsx)
    wb_o = load_workbook(ours_xlsx)
    ws_g, ws_o = wb_g["明细"], wb_o["明细"]
    tm = rules["template_map"]
    cz = tm["clear_zone"]
    mismatches = []
    for block in tm["blocks"]:
        if not str(block.get("key") or "").strip():
            continue
        top = int(block["top"])
        for r in range(top, top + tm["block"]["rows"]):
            for c in range(int(cz["col_start"]), int(cz["col_end"]) + 1):
                vg, vo = ws_g.cell(r, c).value, ws_o.cell(r, c).value
                if isinstance(vg, (int, float)) and isinstance(vo, (int, float)):
                    if abs(float(vg) - float(vo)) > 1e-6:
                        mismatches.append((r, c, vg, vo))
                elif str(vg or "").strip() != str(vo or "").strip():
                    mismatches.append((r, c, vg, vo))
    wb_g.close()
    wb_o.close()
    assert not mismatches, f"数据区与金样不一致（前5处）：{mismatches[:5]}"
