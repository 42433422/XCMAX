# mypy: disable-error-code="index, union-attr"
from __future__ import annotations

import asyncio
import importlib.util
import json
from pathlib import Path
from types import ModuleType

import pytest
from openpyxl import Workbook, load_workbook

FHD_ROOT = Path(__file__).resolve().parents[2]

READER_CONVERT = (
    FHD_ROOT / "mods/_employees/excel-full-read-employee/backend/vendor/excel_full_read/convert.py"
)
WRITER_CONVERT = (
    FHD_ROOT
    / "mods/_employees/excel-template-write-employee/backend/vendor/excel_template_write/convert.py"
)
WRITER_EMPLOYEE = (
    FHD_ROOT
    / "mods/_employees/excel-template-write-employee/backend/employees/excel_template_write_employee.py"
)


def _load_module(name: str, path: Path) -> ModuleType:
    spec = importlib.util.spec_from_file_location(name, path)
    assert spec is not None
    assert spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def _sample_template_workbook(path: Path) -> None:
    """通用模板样例：标题合并、[DBNum1] 格式格、右侧 SUMIF 保护列、多 sheet。"""
    wb = Workbook()
    detail = wb.active
    detail.title = "明细"
    detail["A1"] = "考勤明细模板"
    detail.merge_cells("A1:D1")
    detail.append(["姓名", "部门", "出勤", "工时"])  # row 2 表头
    detail["A3"] = "张三"
    detail["B3"] = "生产部"
    detail["A4"] = "李四"
    detail["B4"] = "仓储部"
    detail["D3"] = 9
    detail["D3"].number_format = "[DBNum1][$-804]General"
    # 右侧 F 列为模板预置 SUMIF 小计（保护区）
    detail["F1"] = "小计"
    detail["F3"] = "=SUMIF(A:A,A3,D:D)"
    detail["F4"] = "=SUMIF(A:A,A4,D:D)"

    monthly = wb.create_sheet("月度统计")
    monthly["A1"] = "姓名"
    monthly["B1"] = "出勤合计"
    monthly["A2"] = "张三"
    monthly["A3"] = "李四"

    wb.create_sheet("草稿")
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# 读取员：模板级元数据（merged_ranges / number_format）
# ---------------------------------------------------------------------------


def test_reader_exports_merged_ranges_and_number_format(tmp_path: Path) -> None:
    reader = _load_module("excel_full_read_convert_meta_test", READER_CONVERT)
    src = tmp_path / "template.xlsx"
    out = tmp_path / "workbook.json"
    _sample_template_workbook(src)

    reader.convert_file(
        src, out, payload={}, ctx={}, rule_spec={"default_output_relpath": "workbook.json"}
    )
    data = json.loads(out.read_text(encoding="utf-8"))
    detail = next(s for s in data["sheets"] if s["name"] == "明细")

    assert "A1:D1" in detail["merged_ranges"]
    d3 = next(c for c in detail["cells"] if c["row"] == 3 and c["col"] == 4)
    assert d3["number_format"].startswith("[DBNum1]")
    # General 格不输出 number_format 字段（省空间、向后兼容）
    a3 = next(c for c in detail["cells"] if c["row"] == 3 and c["col"] == 1)
    assert "number_format" not in a3
    # 公式仍照旧导出
    f3 = next(c for c in detail["cells"] if c["row"] == 3 and c["col"] == 6)
    assert f3["formula"].startswith("=SUMIF")


# ---------------------------------------------------------------------------
# 写入员：核心回填
# ---------------------------------------------------------------------------


def _base_plan() -> dict:
    return {
        "plan_version": 1,
        "template": {"sheet_names": ["明细", "月度统计"]},
        "protected_ranges": ["明细!F:F"],
        "phases": [
            {"phase": "clear_ranges", "ranges": ["明细!C3:D4"]},
            {
                "phase": "cell_writes",
                "writes": [
                    {"sheet": "明细", "row": 3, "col": 3, "value": "√"},
                    {
                        "sheet": "明细",
                        "ref": "D3",
                        "value": "8",
                        "value_type": "number",
                        "number_format": "0.0",
                    },
                    {"sheet": "明细", "ref": "C4", "value": "√"},
                    {"sheet": "明细", "ref": "B6", "value": "2026-03-05", "value_type": "date"},
                ],
            },
            {
                "phase": "formula_writes",
                "writes": [
                    {"sheet": "月度统计", "ref": "B2", "formula": "=SUMIF(明细!A:A,A2,明细!D:D)"},
                ],
            },
            {"phase": "retain_sheets", "names": ["明细", "月度统计"]},
        ],
        "expected": {"cells": 4, "note": "for QC"},
    }


def test_writer_executes_plan_and_preserves_template(tmp_path: Path) -> None:
    writer = _load_module("excel_template_write_convert_test", WRITER_CONVERT)
    template = tmp_path / "template.xlsx"
    plan_path = tmp_path / "plan.json"
    out = tmp_path / "filled.xlsx"
    _sample_template_workbook(template)
    plan_path.write_text(json.dumps(_base_plan(), ensure_ascii=False), encoding="utf-8")

    result = writer.convert_file(
        plan_path,
        out,
        template_path=template,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )

    assert result["output_path"] == str(out)
    assert result["cells_written"] == 4
    assert result["formulas_written"] == 1
    assert result["cells_cleared"] == 1  # C3:D4 中仅 D3 原有值
    assert result["violations"] == []
    assert result["sheets_removed"] == ["草稿"]
    assert result["expected"] == {"cells": 4, "note": "for QC"}

    wb = load_workbook(out, data_only=False)
    detail = wb["明细"]
    assert wb.sheetnames == ["明细", "月度统计"]
    assert detail["C3"].value == "√"
    assert detail["D3"].value == 8
    assert detail["D3"].number_format == "0.0"
    from datetime import datetime

    assert detail["B6"].value == datetime(2026, 3, 5)  # openpyxl 落盘后 date → datetime
    # 模板保真：标题合并、既有公式原样保留
    assert "A1:D1" in {str(r) for r in detail.merged_cells.ranges}
    assert detail["F3"].value == "=SUMIF(A:A,A3,D:D)"
    assert wb["月度统计"]["B2"].value == "=SUMIF(明细!A:A,A2,明细!D:D)"
    wb.close()

    report = json.loads((tmp_path / "write_report.json").read_text(encoding="utf-8"))
    assert report["cells_written"] == 4
    assert report["output_sheet_names"] == ["明细", "月度统计"]
    assert report["expected"] == {"cells": 4, "note": "for QC"}


def test_writer_skips_protected_ranges_and_records_violations(tmp_path: Path) -> None:
    writer = _load_module("excel_template_write_protected_test", WRITER_CONVERT)
    template = tmp_path / "template.xlsx"
    out = tmp_path / "filled.xlsx"
    _sample_template_workbook(template)

    plan = {
        "plan_version": 1,
        "protected_ranges": ["明细!F:F"],
        "phases": [
            {
                "phase": "cell_writes",
                "writes": [
                    {"sheet": "明细", "ref": "C3", "value": "√"},
                    {"sheet": "明细", "ref": "F3", "value": "覆盖公式"},
                ],
            }
        ],
    }
    result = writer.convert_file(
        None,
        out,
        template_path=template,
        payload={"plan": plan},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    assert result["cells_written"] == 1
    assert len(result["violations"]) == 1
    assert result["violations"][0]["sheet"] == "明细"
    assert any("保护区" in w for w in result["warnings"])

    wb = load_workbook(out)
    assert wb["明细"]["F3"].value == "=SUMIF(A:A,A3,D:D)"  # 公式未被覆盖
    wb.close()

    with pytest.raises(ValueError, match="strict_protected"):
        writer.convert_file(
            None,
            out,
            template_path=template,
            payload={"plan": plan, "strict_protected": True},
            ctx={},
            rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
        )


def test_writer_clear_ranges_respects_protection_and_counts(tmp_path: Path) -> None:
    writer = _load_module("excel_template_write_clear_test", WRITER_CONVERT)
    template = tmp_path / "template.xlsx"
    out = tmp_path / "filled.xlsx"
    _sample_template_workbook(template)

    plan = {
        "plan_version": 1,
        "protected_ranges": ["明细!F:F"],
        "phases": [{"phase": "clear_ranges", "ranges": ["明细!A3:F4"]}],
    }
    result = writer.convert_file(
        None,
        out,
        template_path=template,
        payload={"plan": plan},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    # A3/B3/D3/A4/B4 五个有值格被清；F3/F4 命中保护区跳过
    assert result["cells_cleared"] == 5
    assert len(result["violations"]) == 2

    wb = load_workbook(out)
    detail = wb["明细"]
    assert detail["A3"].value is None
    assert detail["F3"].value == "=SUMIF(A:A,A3,D:D)"
    wb.close()


def test_writer_plan_validation_errors(tmp_path: Path) -> None:
    writer = _load_module("excel_template_write_validate_test", WRITER_CONVERT)
    template = tmp_path / "template.xlsx"
    out = tmp_path / "filled.xlsx"
    _sample_template_workbook(template)

    def _run(plan: dict):
        return writer.convert_file(
            None,
            out,
            template_path=template,
            payload={"plan": plan},
            ctx={},
            rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
        )

    with pytest.raises(ValueError, match="phases"):
        _run({"plan_version": 1})
    with pytest.raises(ValueError, match="plan_version"):
        _run({"plan_version": 99, "phases": [{"phase": "cell_writes", "writes": [{}]}]})
    with pytest.raises(ValueError, match="未知阶段"):
        _run({"plan_version": 1, "phases": [{"phase": "nope"}]})
    with pytest.raises(ValueError, match="sheet 不存在"):
        _run(
            {
                "plan_version": 1,
                "phases": [
                    {
                        "phase": "cell_writes",
                        "writes": [{"sheet": "不存在", "ref": "A1", "value": 1}],
                    }
                ],
            }
        )
    with pytest.raises(ValueError, match="模板缺少计划声明的 sheet"):
        _run(
            {
                "plan_version": 1,
                "template": {"sheet_names": ["不存在的表"]},
                "phases": [
                    {"phase": "cell_writes", "writes": [{"sheet": "明细", "ref": "A9", "value": 1}]}
                ],
            }
        )
    with pytest.raises(ValueError, match="公式必须以 = 开头"):
        _run(
            {
                "plan_version": 1,
                "phases": [
                    {
                        "phase": "formula_writes",
                        "writes": [{"sheet": "明细", "ref": "G1", "formula": "SUM(A:A)"}],
                    }
                ],
            }
        )
    with pytest.raises(ValueError, match="模板缺失"):
        writer.convert_file(
            None,
            out,
            template_path=None,
            payload={
                "plan": {
                    "plan_version": 1,
                    "phases": [
                        {
                            "phase": "cell_writes",
                            "writes": [{"sheet": "明细", "ref": "A1", "value": 1}],
                        }
                    ],
                }
            },
            ctx={},
            rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
        )


def test_writer_top_level_writes_shorthand(tmp_path: Path) -> None:
    writer = _load_module("excel_template_write_shorthand_test", WRITER_CONVERT)
    template = tmp_path / "template.xlsx"
    out = tmp_path / "filled.xlsx"
    _sample_template_workbook(template)

    result = writer.convert_file(
        None,
        out,
        template_path=template,
        payload={"plan": {"writes": [{"sheet": "明细", "ref": "C3", "value": "√"}]}},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    assert result["cells_written"] == 1


# ---------------------------------------------------------------------------
# 员工入口（async run）
# ---------------------------------------------------------------------------


def test_writer_employee_run_with_inline_plan(tmp_path: Path) -> None:
    employee = _load_module("excel_template_write_employee_run_test", WRITER_EMPLOYEE)
    template = tmp_path / "template.xlsx"
    _sample_template_workbook(template)

    payload = {
        "plan": _base_plan(),
        "template_path": str(template),
        "workspace_root": str(tmp_path),
    }
    result = asyncio.run(employee.run(payload, {"workspace_root": str(tmp_path)}))
    assert result["ok"], result["error"]
    item = result["items"][0]
    assert item["cells_written"] == 4
    assert Path(item["output_path"]).is_file()
    assert Path(item["report_path"]).is_file()
    assert (tmp_path / "outputs" / "filled.xlsx").is_file()


def test_writer_employee_run_missing_plan_fails(tmp_path: Path) -> None:
    employee = _load_module("excel_template_write_employee_noplan_test", WRITER_EMPLOYEE)
    result = asyncio.run(employee.run({"workspace_root": str(tmp_path)}, {}))
    assert not result["ok"]
    assert "plan" in result["error"] or "file_path" in result["error"]


# ---------------------------------------------------------------------------
# 闭环：读取员读模板 → 构造 plan → 写入员回填 → 读取员重读输出
# ---------------------------------------------------------------------------


def test_read_plan_write_reread_roundtrip(tmp_path: Path) -> None:
    reader = _load_module("excel_full_read_roundtrip_test", READER_CONVERT)
    writer = _load_module("excel_template_write_roundtrip_test", WRITER_CONVERT)

    template = tmp_path / "template.xlsx"
    template_json = tmp_path / "template_workbook.json"
    filled = tmp_path / "filled.xlsx"
    filled_json = tmp_path / "filled_workbook.json"
    _sample_template_workbook(template)

    # 1) 读取员读模板，拿到结构元数据
    reader.convert_file(
        template,
        template_json,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "workbook.json"},
    )
    tpl = json.loads(template_json.read_text(encoding="utf-8"))
    detail = next(s for s in tpl["sheets"] if s["name"] == "明细")
    assert "A1:D1" in detail["merged_ranges"]

    # 2) 模拟规则映射员：基于读取结果构造写入计划（公式列 F 设为保护区）
    formula_cols = sorted({c["letter"][:1] for c in detail["cells"] if c.get("formula")})
    assert formula_cols == ["F"]
    name_rows = [c["row"] for c in detail["cells"] if c["col"] == 1 and c["row"] >= 3]
    plan = {
        "plan_version": 1,
        "template": {"sheet_names": ["明细"]},
        "protected_ranges": [f"明细!{formula_cols[0]}:{formula_cols[0]}"],
        "phases": [
            {
                "phase": "cell_writes",
                "writes": [{"sheet": "明细", "row": r, "col": 3, "value": "√"} for r in name_rows],
            }
        ],
    }
    plan_path = tmp_path / "plan.json"
    plan_path.write_text(json.dumps(plan, ensure_ascii=False), encoding="utf-8")

    # 3) 写入员回填
    result = writer.convert_file(
        plan_path,
        filled,
        template_path=template,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/filled.xlsx"},
    )
    assert result["cells_written"] == len(name_rows) == 2

    # 4) 读取员重读输出，验证闭环
    reader.convert_file(
        filled,
        filled_json,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "workbook.json"},
    )
    out = json.loads(filled_json.read_text(encoding="utf-8"))
    detail_out = next(s for s in out["sheets"] if s["name"] == "明细")
    marks = [c for c in detail_out["cells"] if c["col"] == 3 and c["value"] == "√"]
    assert len(marks) == 2
    # 模板公式与合并在输出中保持
    assert "A1:D1" in detail_out["merged_ranges"]
    assert any((c.get("formula") or "").startswith("=SUMIF") for c in detail_out["cells"])
