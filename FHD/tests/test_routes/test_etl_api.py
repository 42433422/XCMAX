from __future__ import annotations

import json
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker

import app.fastapi_routes.etl as etl_routes
from app.application.etl.service import EtlService
from app.db.base import Base
from app.db.models.etl import (
    EtlRun,
    EtlRunRow,
    EtlTargetConfig,
    EtlTemplate,
    EtlTemplateVersion,
    EtlUpload,
)
from app.db.models.product import Product
from app.db.models.purchase_unit import PurchaseUnit
from app.db.models.shipment import ShipmentRecord
from app.db.models.shipment_etl_fingerprint import ShipmentEtlImportFingerprint
from app.fastapi_routes.shipment_etl_compat import shipment_etl_preview
from app.infrastructure.tenant_scope import tenant_scope


def _test_app(tmp_path, monkeypatch):
    engine = create_engine(
        f"sqlite:///{tmp_path / 'api.sqlite'}", connect_args={"check_same_thread": False}
    )
    Base.metadata.create_all(
        engine,
        tables=[
            EtlUpload.__table__,
            EtlTemplate.__table__,
            EtlTemplateVersion.__table__,
            EtlRun.__table__,
            EtlRunRow.__table__,
            EtlTargetConfig.__table__,
            PurchaseUnit.__table__,
            Product.__table__,
            ShipmentRecord.__table__,
            ShipmentEtlImportFingerprint.__table__,
        ],
    )
    maker = sessionmaker(bind=engine, expire_on_commit=False)
    service = EtlService()
    monkeypatch.setattr("app.application.etl.service.SessionLocal", maker)
    monkeypatch.setattr(etl_routes, "get_etl_service", lambda: service)
    monkeypatch.setenv("XCAGI_DATA_DIR", str(tmp_path / "app-data"))
    monkeypatch.setenv("XCAGI_PRODUCT_SKU", "enterprise")
    monkeypatch.setenv("FHD_ETL_CENTER_ENABLED", "1")
    monkeypatch.setenv("FHD_ETL_LLM", "off")

    def inline_preview(run_id: str, _tenant_id: int, owner_user_id: int) -> None:
        service._preview_worker(run_id, owner_user_id)

    monkeypatch.setattr(service, "_submit_preview", inline_preview)
    monkeypatch.setattr(
        service,
        "_submit_execution",
        lambda run_id, _tenant_id, owner_user_id, valid_rows_only: service._execute_worker(
            run_id, owner_user_id, valid_rows_only
        ),
    )

    current_user = {"id": 101}

    def user():
        return SimpleNamespace(id=current_user["id"])

    def db_dependency():
        with tenant_scope(4):
            db = maker()
            try:
                yield db
                db.commit()
            finally:
                db.close()

    app = FastAPI()
    app.include_router(etl_routes.router)
    app.dependency_overrides[etl_routes._read] = user
    app.dependency_overrides[etl_routes._execute] = user
    app.dependency_overrides[etl_routes._rollback] = user
    app.dependency_overrides[etl_routes._template_manage] = user
    app.dependency_overrides[etl_routes._target_manage] = user
    app.dependency_overrides[etl_routes.get_db_dependency] = db_dependency
    return TestClient(app), current_user


def _mixed_delivery_workbook_bytes(tmp_path: Path) -> bytes:
    """A compact mixed workbook: delivery + product history + finance appendix."""

    path = tmp_path / "mixed-delivery.xlsx"
    workbook = Workbook()
    delivery = workbook.active
    delivery.title = "侯雪梅"
    delivery.append(["某公司送货单"])
    delivery.append(["购货单位：金汉武家私  联系人：张总  2026年01月21日  订单编号：A-9803"])
    delivery.append(["型号", "产品名称", "数量/件", "规格/KG", "数量/KG", "单价/元", "金额/元"])
    delivery.append(["9803", "黑棕面用修色精", 3, 28, 84, 48, 4032])
    delivery.append(["合计", None, 3, None, 84, None, 4032])

    history = workbook.create_sheet("出货历史")
    history.append(
        ["金汉武（宾驰）", "45659", "2", "方和", None, None, "黑棕面用修色精", 3, 4, 12, 48, 576]
    )

    finance = workbook.create_sheet("25年回款")
    finance.append(["客户名", "回款金额", "余额"])
    finance.append(["金汉武家私", 1000, 200])
    workbook.save(path)
    workbook.close()
    return path.read_bytes()


def test_etl_api_upload_preview_execute_and_owner_isolation(tmp_path, monkeypatch):
    client, current_user = _test_app(tmp_path, monkeypatch)
    capabilities = client.get("/api/etl/capabilities")
    assert capabilities.status_code == 200
    assert {row["type"] for row in capabilities.json()["data"]["targets"]} >= {
        "customer_products",
        "customers",
        "products",
        "shipment_records",
        "webhook",
    }
    assert capabilities.json()["data"]["inputs"]["folder_upload"] is True
    assert capabilities.json()["data"]["limits"]["max_file_bytes"] == 100 * 1024 * 1024

    uploaded = client.post(
        "/api/etl/uploads",
        data={
            "batch_id": "11111111-1111-4111-8111-111111111111",
            "relative_path": "客户资料/customers.csv",
        },
        files={"file": ("customers.csv", "客户名称,电话\n甲公司,138\n".encode(), "text/csv")},
    )
    assert uploaded.status_code == 201
    assert uploaded.json()["data"]["batch_id"] == "11111111-1111-4111-8111-111111111111"
    assert uploaded.json()["data"]["relative_path"] == "客户资料/customers.csv"
    upload_id = uploaded.json()["data"]["upload_id"]

    preview = client.post(
        "/api/etl/runs/preview",
        json={"upload_id": upload_id, "target_type": "customers"},
    )
    assert preview.status_code == 202
    run = preview.json()["data"]
    assert run["status"] == "preview_ready"
    assert run["summary"]["new"] == 1
    assert run["batch_id"] == "11111111-1111-4111-8111-111111111111"
    assert run["relative_path"] == "客户资料/customers.csv"

    batch_runs = client.get(
        "/api/etl/runs",
        params={"batch_id": "11111111-1111-4111-8111-111111111111", "limit": 500},
    )
    assert batch_runs.status_code == 200
    assert [item["id"] for item in batch_runs.json()["data"]] == [run["id"]]

    rows = client.get(f"/api/etl/runs/{run['id']}/rows")
    assert rows.status_code == 200
    assert rows.json()["data"]["items"][0]["final_action"] == "new"

    executed = client.post(
        f"/api/etl/runs/{run['id']}/execute",
        json={"confirmed": True, "valid_rows_only": False},
    )
    assert executed.status_code == 202
    assert executed.json()["data"]["status"] == "completed"

    current_user["id"] = 202
    hidden = client.get(f"/api/etl/runs/{run['id']}")
    assert hidden.status_code == 404
    assert hidden.json()["detail"]["code"] == "ETL_NOT_FOUND"


def test_etl_api_rejects_non_enterprise_build(tmp_path, monkeypatch):
    client, _ = _test_app(tmp_path, monkeypatch)
    monkeypatch.setenv("XCAGI_PRODUCT_SKU", "personal")
    response = client.get("/api/etl/capabilities")
    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "ETL_ENTERPRISE_REQUIRED"


def test_etl_api_auto_target_resolves_before_preview(tmp_path, monkeypatch):
    client, _ = _test_app(tmp_path, monkeypatch)
    uploaded = client.post(
        "/api/etl/uploads",
        files={
            "file": (
                "customers.csv",
                "客户名称,电话\n自动识别客户,13800000000\n".encode(),
                "text/csv",
            )
        },
    )
    upload_id = uploaded.json()["data"]["upload_id"]

    preview = client.post(
        "/api/etl/runs/preview",
        json={"upload_id": upload_id, "target_type": "auto"},
    )

    assert preview.status_code == 202
    run = preview.json()["data"]
    assert run["target_type"] == "customers"
    assert run["status"] == "preview_ready"
    assert run["details"]["requested_target_type"] == "auto"
    assert run["source_features"]["target_detection"] == {
        "target_type": "customers",
        "document_type": "customer_table",
        "confidence": 0.74,
        "reason": "customer_headers",
    }


def test_etl_api_mixed_delivery_preview_exposes_sheet_routes_and_private_layout(
    tmp_path,
    monkeypatch,
):
    """The docking loop stays preview-only until the caller explicitly executes it."""

    client, _ = _test_app(tmp_path, monkeypatch)
    uploaded = client.post(
        "/api/etl/uploads",
        files={
            "file": (
                "侯雪梅.xlsx",
                _mixed_delivery_workbook_bytes(tmp_path),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert uploaded.status_code == 201
    upload_id = uploaded.json()["data"]["upload_id"]

    shipment_preview = client.post(
        "/api/etl/runs/preview",
        json={"upload_id": upload_id, "target_type": "auto"},
    )
    assert shipment_preview.status_code == 202
    shipment_run = shipment_preview.json()["data"]
    assert shipment_run["target_type"] == "shipment_records"
    assert shipment_run["status"] == "preview_ready"
    assert shipment_run["details"]["requested_target_type"] == "auto"
    assert shipment_run["source_features"]["target_detection"]["document_type"] == (
        "delivery_note_workbook"
    )
    inventory = shipment_run["details"]["sheet_inventory"]
    assert [sheet["sheet"] for sheet in inventory] == ["侯雪梅", "出货历史", "25年回款"]
    assert inventory[0]["structure"] == "single_document"
    assert inventory[1]["status"] == "review_required"
    routes = shipment_run["details"]["document_routes"]
    assert routes[0]["sheet"] == "侯雪梅"
    assert routes[0]["target_type"] == "shipment_records"
    assert routes[0]["status"] == "preview_ready"

    saved_layout = client.post(
        f"/api/etl/runs/{shipment_run['id']}/shipment-template",
        json={"name": ""},
    )
    assert saved_layout.status_code == 200
    layout = saved_layout.json()["data"]
    assert layout["template_id"].startswith("etl:")
    assert layout["name"] == "金汉武家私-发货单版式"
    assert Path(layout["file_path"]).is_file()
    assert "/document_templates/101/" in layout["file_path"]

    # A printing layout is deliberately hidden from import-template selection
    # and fails closed if a caller tries to use its raw ID as an ETL mapping.
    templates = client.get("/api/etl/templates")
    assert templates.status_code == 200
    assert layout["template_id"].removeprefix("etl:") not in {
        row["id"] for row in templates.json()["data"]
    }
    invalid_mapping = client.post(
        "/api/etl/runs/preview",
        json={
            "upload_id": upload_id,
            "target_type": "shipment_records",
            "template_id": layout["template_id"].removeprefix("etl:"),
        },
    )
    assert invalid_mapping.status_code == 409
    assert invalid_mapping.json()["detail"]["code"] == ("ETL_SHIPMENT_TEMPLATE_NOT_IMPORT_TEMPLATE")

    # Preview routing creates no customer/product records. Only /execute can
    # perform a state change, and this test intentionally never calls it.
    from app.application.etl import service as etl_service_module

    with tenant_scope(4):
        db = etl_service_module.SessionLocal()
        try:
            assert db.query(PurchaseUnit).count() == 0
            assert db.query(Product).count() == 0
        finally:
            db.close()


def test_etl_api_is_fail_closed_when_feature_flag_is_missing(tmp_path, monkeypatch):
    client, _ = _test_app(tmp_path, monkeypatch)
    monkeypatch.delenv("FHD_ETL_CENTER_ENABLED")
    response = client.get("/api/etl/capabilities")
    assert response.status_code == 403
    assert response.json()["detail"]["code"] == "ETL_CENTER_DISABLED"


def test_etl_api_rejects_invalid_folder_batch_id(tmp_path, monkeypatch):
    client, _ = _test_app(tmp_path, monkeypatch)
    response = client.post(
        "/api/etl/uploads",
        data={"batch_id": "../../not-a-batch", "relative_path": "../../customers.csv"},
        files={"file": ("customers.csv", b"name\nAcme\n", "text/csv")},
    )

    assert response.status_code == 400
    assert response.json()["detail"]["code"] == "ETL_BATCH_ID_INVALID"


@pytest.mark.asyncio
async def test_legacy_shipment_preview_rejects_caller_selected_path():
    response = await shipment_etl_preview(
        file=None,
        file_path="/etc/passwd",
        workspace_root="/",
        include_ledger="auto",
        save_as_template="1",
        template_name="unsafe",
        template_scope="global",
        db=object(),
        user=SimpleNamespace(id=101),
    )
    payload = json.loads(response.body)
    assert response.status_code == 409
    assert payload["error_code"] == "ETL_UPLOAD_REQUIRED"
    assert payload["file_path_ignored"] is True
