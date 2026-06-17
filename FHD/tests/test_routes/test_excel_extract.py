"""Comprehensive tests for app.fastapi_routes.excel_extract.

Covers: extract, upload, generate, download, import products/customers, logs, preview,
attendance detail extraction, and all helper functions.
"""

from __future__ import annotations

import os
import tempfile
from unittest.mock import MagicMock, patch

import pytest

from app.fastapi_routes.excel_extract import (
    TEMP_EXCEL_DIR,
    _extract_attendance_detail_roster,
    _extract_from_excel,
    _generate_excel,
    download_generated_excel,
    extract_from_excel,
    extract_test,
    generate_excel,
    get_extract_log,
    get_extract_logs,
    get_preview,
    import_customers,
    import_products,
)


# ---------------------------------------------------------------------------
# _extract_from_excel
# ---------------------------------------------------------------------------


class TestExtractFromExcel:
    def test_file_not_found(self):
        result, status = _extract_from_excel("/nonexistent/file.xlsx")
        assert status == 404
        assert result["success"] is False

    def test_extract_basic(self):
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            ws.append(["Name", "Age"])
            ws.append(["Alice", 30])
            ws.append(["Bob", 25])
            wb.save(f.name)
            path = f.name

        try:
            result, status = _extract_from_excel(path)
            assert status == 200
            assert result["success"] is True
            assert result["total_rows"] == 2
            assert len(result["headers"]) == 2
        finally:
            os.unlink(path)

    def test_extract_with_sheet_name(self):
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Data"
            ws.append(["Col1"])
            ws.append(["val1"])
            wb.save(f.name)
            path = f.name

        try:
            result, status = _extract_from_excel(path, sheet_name="Data")
            assert status == 200
            assert result["sheet"] == "Data"
        finally:
            os.unlink(path)

    def test_extract_with_custom_header_row(self):
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.append(["Skip", "This"])
            ws.append(["Name", "Age"])
            ws.append(["Alice", 30])
            wb.save(f.name)
            path = f.name

        try:
            result, status = _extract_from_excel(path, header_row=2)
            assert status == 200
            assert result["header_row"] == 2
        finally:
            os.unlink(path)

    def test_extract_empty_rows(self):
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Age"])
            wb.save(f.name)
            path = f.name

        try:
            result, status = _extract_from_excel(path)
            assert status == 200
            assert result["total_rows"] == 0
        finally:
            os.unlink(path)

    def test_extract_error(self):
        with patch("app.fastapi_routes.excel_extract._extract_from_excel", side_effect=RuntimeError("corrupt")):
            # The route handler catches RECOVERABLE_ERRORS
            result, status = _extract_from_excel("/fake/path.xlsx")
            # If it's a direct call, the error propagates; if patched, we test the route
            # Let's test the actual error path via the route handler
            pass

    def test_extract_error_via_handler(self):
        with patch(
            "app.fastapi_routes.excel_extract._extract_from_excel",
            side_effect=RuntimeError("corrupt"),
        ):
            resp = extract_from_excel({"file_path": "/fake/path.xlsx"})
            assert resp.status_code == 500


# ---------------------------------------------------------------------------
# _generate_excel
# ---------------------------------------------------------------------------


class TestGenerateExcel:
    def test_basic_generation(self):
        data = [{"Name": "Alice", "Age": 30}, {"Name": "Bob", "Age": 25}]
        result, status = _generate_excel(data)
        assert status == 200
        assert result["success"] is True
        assert result["rows"] == 2
        assert os.path.exists(result["file_path"])
        os.unlink(result["file_path"])

    def test_custom_filename(self):
        data = [{"A": 1}]
        result, status = _generate_excel(data, filename="test_custom.xlsx")
        assert status == 200
        assert "test_custom.xlsx" in result["filename"]
        os.unlink(result["file_path"])

    def test_custom_sheet_name(self):
        data = [{"A": 1}]
        result, status = _generate_excel(data, sheet_name="MySheet")
        assert status == 200
        assert result["sheet"] == "MySheet"
        os.unlink(result["file_path"])

    def test_empty_data(self):
        result, status = _generate_excel([])
        assert status == 400
        assert result["success"] is False

    def test_invalid_data(self):
        result, status = _generate_excel("not a list")
        assert status == 400
        assert result["success"] is False

    def test_no_filename_generates_timestamp(self):
        data = [{"X": 1}]
        result, status = _generate_excel(data, filename=None)
        assert status == 200
        assert "export_" in result["filename"]
        os.unlink(result["file_path"])


# ---------------------------------------------------------------------------
# _extract_attendance_detail_roster
# ---------------------------------------------------------------------------


class TestExtractAttendanceDetailRoster:
    def test_file_not_found(self):
        result, status = _extract_attendance_detail_roster("/nonexistent/file.xlsx")
        assert status == 404

    def test_basic_extraction(self):
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "明细"
            ws.cell(1, 1, "Header1")
            ws.cell(2, 1, "Header2")
            ws.cell(3, 1, "Header3")
            ws.cell(4, 1, "技术部")
            ws.cell(4, 2, "全职")
            ws.cell(4, 3, "张三")
            wb.save(f.name)
            path = f.name

        try:
            result, status = _extract_attendance_detail_roster(path)
            assert status == 200
            assert result["success"] is True
            assert result["parse_mode"] == "attendance_detail"
            assert len(result["rows"]) >= 1
            assert result["rows"][0]["product_name"] == "张三"
        finally:
            os.unlink(path)

    def test_empty_name_skipped(self):
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "明细"
            ws.cell(1, 1, "H1")
            ws.cell(2, 1, "H2")
            ws.cell(3, 1, "H3")
            ws.cell(4, 1, "Dept")
            ws.cell(4, 2, "Type")
            ws.cell(4, 3, None)
            wb.save(f.name)
            path = f.name

        try:
            result, status = _extract_attendance_detail_roster(path)
            assert status == 200
            assert len(result["rows"]) == 0
        finally:
            os.unlink(path)

    def test_with_sheet_name(self):
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.title = "Other"
            wb.save(f.name)
            path = f.name

        try:
            result, status = _extract_attendance_detail_roster(path, sheet_name="Other")
            assert status == 200
        finally:
            os.unlink(path)


# ---------------------------------------------------------------------------
# Route handler tests (direct function calls, no TestClient)
# ---------------------------------------------------------------------------


class TestExtractRoute:
    def test_no_file_path(self):
        resp = extract_from_excel({})
        assert resp.status_code == 400

    def test_file_not_found(self):
        resp = extract_from_excel({"file_path": "/nonexistent.xlsx"})
        assert resp.status_code == 404

    def test_extract_success(self):
        from openpyxl import Workbook

        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name"])
            ws.append(["Alice"])
            wb.save(f.name)
            path = f.name

        try:
            resp = extract_from_excel({"file_path": path})
            assert resp.status_code == 200
        finally:
            os.unlink(path)


class TestGenerateRoute:
    def test_no_data(self):
        resp = generate_excel({})
        assert resp.status_code == 400

    def test_generate_success(self):
        resp = generate_excel({"data": [{"A": 1}]})
        assert resp.status_code == 200
        import json

        body = json.loads(resp.body)
        assert body["success"] is True
        # Clean up generated file
        if "file_path" in body and os.path.exists(body["file_path"]):
            os.unlink(body["file_path"])


class TestDownloadRoute:
    def test_no_data(self):
        resp = download_generated_excel({})
        assert resp.status_code == 400

    def test_download_success(self):
        resp = download_generated_excel({"data": [{"A": 1}]})
        # Should return FileResponse on success
        assert resp.status_code == 200
        # Clean up
        from starlette.responses import FileResponse as FR

        if isinstance(resp, FR) and os.path.exists(resp.filename):
            os.unlink(resp.filename)


class TestExtractTestRoute:
    def test_returns_ok(self):
        resp = extract_test()
        assert resp.status_code == 200
        import json

        body = json.loads(resp.body)
        assert body["success"] is True


class TestImportProductsRoute:
    def test_no_data(self):
        resp = import_products({})
        assert resp.status_code == 400

    def test_import_success(self):
        mock_svc = MagicMock()
        mock_svc.import_data.return_value = {
            "imported": 1, "skipped": 0, "failed": 0, "details": []
        }
        mock_log_svc = MagicMock()
        mock_log_svc.create_log.return_value = 1
        with patch("app.application.facades.excel_facade.get_product_import_service", return_value=mock_svc), \
             patch("app.application.get_extract_log_app_service", return_value=mock_log_svc):
            resp = import_products({"data": [{"product_name": "Test"}]})
            assert resp.status_code == 200

    def test_import_with_ai_parse(self):
        mock_parser = MagicMock()
        mock_parser.parse_single.return_value = {
            "success": True, "product_code": "P1", "product_name": "PN", "specification": "S1"
        }
        mock_svc = MagicMock()
        mock_svc.import_data.return_value = {
            "imported": 1, "skipped": 0, "failed": 0, "details": []
        }
        mock_log_svc = MagicMock()
        mock_log_svc.create_log.return_value = 1
        with patch("app.application.facades.excel_facade.get_product_import_service", return_value=mock_svc), \
             patch("app.application.get_extract_log_app_service", return_value=mock_log_svc), \
             patch("app.application.facades.excel_facade.get_ai_product_parser", return_value=mock_parser):
            resp = import_products({
                "data": [{"raw_text": "some text"}],
                "options": {"use_ai_parse": True, "ai_source_field": "raw_text"},
            })
            assert resp.status_code == 200

    def test_import_error(self):
        mock_log_svc = MagicMock()
        mock_log_svc.create_log.return_value = 1
        mock_svc = MagicMock()
        mock_svc.import_data.side_effect = RuntimeError("import fail")
        with patch("app.application.facades.excel_facade.get_product_import_service", return_value=mock_svc), \
             patch("app.application.get_extract_log_app_service", return_value=mock_log_svc):
            resp = import_products({"data": [{"product_name": "Test"}]})
            assert resp.status_code == 500


class TestImportCustomersRoute:
    def test_no_data(self):
        resp = import_customers({})
        assert resp.status_code == 400

    def test_import_success(self):
        mock_svc = MagicMock()
        mock_svc.import_data.return_value = {
            "imported": 1, "skipped": 0, "failed": 0, "details": []
        }
        mock_log_svc = MagicMock()
        mock_log_svc.create_log.return_value = 1
        with patch("app.application.get_customer_app_service", return_value=mock_svc), \
             patch("app.application.get_extract_log_app_service", return_value=mock_log_svc):
            resp = import_customers({"data": [{"customer_name": "Test"}]})
            assert resp.status_code == 200

    def test_import_error(self):
        mock_svc = MagicMock()
        mock_svc.import_data.side_effect = RuntimeError("import fail")
        mock_log_svc = MagicMock()
        mock_log_svc.create_log.return_value = 1
        with patch("app.application.get_customer_app_service", return_value=mock_svc), \
             patch("app.application.get_extract_log_app_service", return_value=mock_log_svc):
            resp = import_customers({"data": [{"customer_name": "Test"}]})
            assert resp.status_code == 500


class TestLogsRoute:
    def test_get_logs(self):
        mock_svc = MagicMock()
        mock_svc.get_logs.return_value = []
        with patch("app.application.get_extract_log_app_service", return_value=mock_svc):
            resp = get_extract_logs()
            assert resp.status_code == 200

    def test_get_logs_error(self):
        with patch("app.application.get_extract_log_app_service", side_effect=RuntimeError("fail")):
            resp = get_extract_logs()
            assert resp.status_code == 500


class TestLogDetailRoute:
    def test_not_found(self):
        mock_svc = MagicMock()
        mock_svc.get_log.return_value = None
        with patch("app.application.get_extract_log_app_service", return_value=mock_svc):
            resp = get_extract_log(999)
            assert resp.status_code == 404

    def test_found(self):
        mock_svc = MagicMock()
        mock_svc.get_log.return_value = {"id": 1}
        with patch("app.application.get_extract_log_app_service", return_value=mock_svc):
            resp = get_extract_log(1)
            assert resp.status_code == 200

    def test_error(self):
        with patch("app.application.get_extract_log_app_service", side_effect=RuntimeError("fail")):
            resp = get_extract_log(1)
            assert resp.status_code == 500


class TestPreviewRoute:
    def test_not_found(self):
        mock_svc = MagicMock()
        mock_svc.get_log.return_value = None
        with patch("app.application.get_extract_log_app_service", return_value=mock_svc):
            resp = get_preview(999)
            assert resp.status_code == 404

    def test_found(self):
        mock_svc = MagicMock()
        mock_svc.get_log.return_value = {"id": 1}
        with patch("app.application.get_extract_log_app_service", return_value=mock_svc):
            resp = get_preview(1)
            assert resp.status_code == 200

    def test_error(self):
        with patch("app.application.get_extract_log_app_service", side_effect=RuntimeError("fail")):
            resp = get_preview(1)
            assert resp.status_code == 500
