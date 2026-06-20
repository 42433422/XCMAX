"""template/routes 路由测试 — 覆盖 Excel 网格提取、纯函数、多 Sheet 分析等。"""

from __future__ import annotations

import io
import os
import tempfile
from types import SimpleNamespace
from unittest.mock import MagicMock, patch

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient

from app.fastapi_routes.domains.template import routes as template_routes


@pytest.fixture
def client() -> TestClient:
    app = FastAPI()
    app.include_router(template_routes.router)
    return TestClient(app, raise_server_exceptions=False)


# ---------------------------------------------------------------------------
# 纯函数
# ---------------------------------------------------------------------------


class TestExcelCellToText:
    def test_none(self):
        assert template_routes._excel_cell_to_text(None) == ""

    def test_number(self):
        assert template_routes._excel_cell_to_text(42) == "42"

    def test_string(self):
        assert template_routes._excel_cell_to_text(" hello ") == "hello"

    def test_float(self):
        assert template_routes._excel_cell_to_text(3.14) == "3.14"


class TestFormBool:
    def test_none(self):
        assert template_routes._form_bool(None) is False

    def test_true_values(self):
        for v in ("1", "true", "yes", "on", "True", "YES", "ON"):
            assert template_routes._form_bool(v) is True

    def test_false_values(self):
        for v in ("0", "false", "no", "off", ""):
            assert template_routes._form_bool(v) is False


class TestPickHeaderRow:
    def test_picks_row_with_most_non_empty(self):
        rows = [["", ""], ["名称", "数量", "单位"], ["A", "1", "kg"]]
        idx, header = template_routes._pick_header_row(rows)
        assert idx == 1
        assert "名称" in header

    def test_empty_rows(self):
        idx, header = template_routes._pick_header_row([])
        assert idx == 0
        assert header == []

    def test_all_empty_rows(self):
        rows = [["", ""], ["", ""]]
        idx, header = template_routes._pick_header_row(rows)
        assert idx == 0

    def test_fallback_when_header_too_sparse(self):
        rows = [["a"], ["", ""], ["", ""]]
        idx, header = template_routes._pick_header_row(rows)
        assert idx == 0


class TestDetectEffectiveColCount:
    def test_detects_max_used(self):
        rows = [["a", "", "c"], ["", "b", ""]]
        assert template_routes._detect_effective_col_count(rows, 1) == 3

    def test_empty_rows_fallback(self):
        rows = [["", ""], ["", ""]]
        assert template_routes._detect_effective_col_count(rows, 5) == 5

    def test_fallback_minimum(self):
        rows = [["", ""], ["", ""]]
        assert template_routes._detect_effective_col_count(rows, 0) == 1


class TestDetectEffectiveRowCount:
    def test_detects_max_used(self):
        rows = [["a"], [""], ["b"]]
        assert template_routes._detect_effective_row_count(rows, 1) == 3

    def test_empty_rows_fallback(self):
        rows = [["", ""], ["", ""]]
        assert template_routes._detect_effective_row_count(rows, 5) == 5


class TestExcelColWidthToPx:
    def test_default_width(self):
        assert template_routes._excel_col_width_to_px(8.43) >= 40

    def test_none(self):
        assert template_routes._excel_col_width_to_px(None) >= 40

    def test_small_width_clamped(self):
        assert template_routes._excel_col_width_to_px(0) >= 40


class TestExcelRowHeightToPx:
    def test_default_height(self):
        assert template_routes._excel_row_height_to_px(15.0) >= 20

    def test_none(self):
        assert template_routes._excel_row_height_to_px(None) >= 20

    def test_small_height_clamped(self):
        assert template_routes._excel_row_height_to_px(0) >= 20


class TestMergeAnchorAndSkip:
    def test_basic_merge(self):
        ws = MagicMock()
        rg = MagicMock()
        rg.min_row, rg.min_col, rg.max_row, rg.max_col = 1, 1, 2, 2
        ws.merged_cells.ranges = [rg]
        anchor, skip = template_routes._merge_anchor_and_skip(ws, 3, 3)
        assert (1, 1) in anchor
        assert anchor[(1, 1)] == (2, 2)
        assert (2, 2) in skip
        assert (1, 1) not in skip

    def test_merge_out_of_bounds(self):
        ws = MagicMock()
        rg = MagicMock()
        rg.min_row, rg.min_col, rg.max_row, rg.max_col = 10, 10, 12, 12
        ws.merged_cells.ranges = [rg]
        anchor, skip = template_routes._merge_anchor_and_skip(ws, 3, 3)
        assert len(anchor) == 0
        assert len(skip) == 0

    def test_no_merges(self):
        ws = MagicMock()
        ws.merged_cells.ranges = []
        anchor, skip = template_routes._merge_anchor_and_skip(ws, 5, 5)
        assert len(anchor) == 0
        assert len(skip) == 0


class TestSerializeCellStyle:
    def test_empty_cell(self):
        cell = MagicMock()
        cell.font = None
        cell.fill = None
        cell.alignment = None
        cell.border = None
        result = template_routes._serialize_cell_style(cell)
        assert result == {}

    def test_font_with_name_and_size(self):
        cell = MagicMock()
        font = MagicMock()
        font.name = "Arial"
        font.sz = 12.0
        font.b = None
        font.i = None
        font.color = None
        cell.font = font
        cell.fill = None
        cell.alignment = None
        cell.border = None
        result = template_routes._serialize_cell_style(cell)
        assert "font" in result
        assert result["font"]["name"] == "Arial"
        assert result["font"]["size"] == 12.0

    def test_font_bold_italic(self):
        cell = MagicMock()
        font = MagicMock()
        font.name = None
        font.sz = None
        font.b = True
        font.i = False
        font.color = None
        cell.font = font
        cell.fill = None
        cell.alignment = None
        cell.border = None
        result = template_routes._serialize_cell_style(cell)
        assert result["font"]["bold"] is True
        assert result["font"]["italic"] is False

    def test_fill_fgcolor(self):
        cell = MagicMock()
        cell.font = None
        fill = MagicMock()
        fg = MagicMock()
        fg.rgb = "FFFF0000"
        fill.fgColor = fg
        cell.fill = fill
        cell.alignment = None
        cell.border = None
        result = template_routes._serialize_cell_style(cell)
        assert "fill" in result
        assert result["fill"]["fgColor"] == "FFFF0000"

    def test_alignment(self):
        cell = MagicMock()
        cell.font = None
        cell.fill = None
        al = MagicMock()
        al.horizontal = "center"
        al.vertical = "top"
        al.wrapText = True
        al.textRotation = 0
        cell.alignment = al
        cell.border = None
        result = template_routes._serialize_cell_style(cell)
        assert "alignment" in result
        assert result["alignment"]["horizontal"] == "center"

    def test_border(self):
        cell = MagicMock()
        cell.font = None
        cell.fill = None
        cell.alignment = None
        border = MagicMock()
        side = MagicMock()
        side.style = "thin"
        side.color = None
        border.left = side
        border.right = None
        border.top = None
        border.bottom = None
        cell.border = border
        result = template_routes._serialize_cell_style(cell)
        assert "border" in result
        assert "left" in result["border"]


class TestBuildGridStyleCache:
    def test_empty_ws(self):
        ws = MagicMock()
        ws.cell.return_value = MagicMock()
        ws.cell.return_value.font = None
        ws.cell.return_value.fill = None
        ws.cell.return_value.alignment = None
        ws.cell.return_value.border = None
        result = template_routes._build_grid_style_cache(ws, 2, 2, set())
        assert "styles" in result
        assert "cell_style_refs" in result


class TestBundleToSheetEntry:
    def test_converts(self):
        bundle = {
            "sheet_name": "Sheet1",
            "fields": [{"label": "A", "name": "A", "type": "dynamic"}],
            "sample_rows": [],
            "grid_preview": {"rows": [], "max_row": 0, "max_col": 0, "header_row_index": 1},
            "grid_style_cache": {"styles": {}, "cell_style_refs": {}},
            "tables": [],
        }
        entry = template_routes._bundle_to_sheet_entry(bundle, 1)
        assert entry["sheet_index"] == 1
        assert entry["sheet_name"] == "Sheet1"
        assert "style_cache" in entry


# ---------------------------------------------------------------------------
# 路由测试
# ---------------------------------------------------------------------------


class TestTemplatesExtractGrid:
    def test_rejects_non_excel(self, client: TestClient):
        r = client.post(
            "/templates/extract-grid",
            files={"file": ("test.txt", b"hello", "text/plain")},
        )
        assert r.status_code == 400

    def test_rejects_non_excel_with_data(self, client: TestClient):
        r = client.post(
            "/templates/extract-grid",
            files={"file": ("data.csv", b"a,b\n1,2", "text/csv")},
        )
        assert r.status_code == 400

    def test_accepts_xlsx(self, client: TestClient, tmp_path):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["名称", "数量", "单位"])
        ws.append(["苹果", 10, "kg"])
        ws.append(["香蕉", 5, "kg"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        r = client.post(
            "/templates/extract-grid",
            files={
                "file": (
                    "test.xlsx",
                    buf.getvalue(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert r.status_code == 200
        data = r.json()
        assert data["success"] is True
        assert len(data["fields"]) > 0
        assert data["preview_data"]["sheet_names"] == ["Sheet1"]

    def test_analyze_all_sheets(self, client: TestClient):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A", "B"])
        ws1.append(["1", "2"])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["C", "D"])
        ws2.append(["3", "4"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        r = client.post(
            "/templates/extract-grid",
            files={"file": ("multi.xlsx", buf.getvalue(), "application/octet-stream")},
            data={"analyze_all_sheets": "true"},
        )
        assert r.status_code == 200
        data = r.json()
        assert len(data["sheets"]) >= 2

    def test_empty_workbook(self, client: TestClient):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empty"
        # No data at all

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        r = client.post(
            "/templates/extract-grid",
            files={"file": ("empty.xlsx", buf.getvalue(), "application/octet-stream")},
        )
        assert r.status_code == 200
        assert r.json()["success"] is True

    def test_specific_sheet_name(self, client: TestClient):
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "First"
        ws1.append(["X", "Y"])
        ws2 = wb.create_sheet("Second")
        ws2.append(["A", "B"])
        ws2.append(["1", "2"])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        r = client.post(
            "/templates/extract-grid",
            files={"file": ("sheets.xlsx", buf.getvalue(), "application/octet-stream")},
            data={"sheet_name": "Second"},
        )
        assert r.status_code == 200
        assert r.json()["preview_data"]["selected_sheet_name"] == "Second"
