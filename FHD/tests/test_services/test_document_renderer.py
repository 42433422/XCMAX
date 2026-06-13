"""app/services/document_templates/renderer 单测：Excel 结构/网格/样式/逻辑表抽取 + JSON 辅助。

用 openpyxl 造真实临时 xlsx（仅文件系统边界，铁律4），覆盖正常解析、合并单元格、
多表与坏文件/缺文件的 RECOVERABLE_ERRORS 回退分支（铁律3）。
"""

from __future__ import annotations

import openpyxl
import pytest

from app.services.document_templates.renderer import (
    _extract_excel_all_sheets_preview,
    _extract_excel_grid_preview,
    _extract_excel_grid_style_cache,
    _extract_logical_tables_from_sheet,
    _extract_structured_excel_preview,
    _is_unreadable_workbook_error,
    _list_excel_sheet_names,
    _parse_json_dict,
    _parse_json_list,
)


@pytest.fixture()
def sample_xlsx(tmp_path):
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["产品名称", "数量", "单价"])
    ws.append(["油漆A", 10, 99.5])
    ws.append(["油漆B", 5, 50.0])
    ws.append(["油漆C", 3, 30.0])
    ws.merge_cells("A6:C6")
    ws["A6"] = "合计说明"
    wb.save(path)
    return str(path)


class TestUnreadableWorkbookError:
    def test_marker_present(self):
        assert _is_unreadable_workbook_error("Unable to read workbook") is True
        assert _is_unreadable_workbook_error("BadZipFile: not a zip") is True

    def test_no_marker(self):
        assert _is_unreadable_workbook_error("some other error") is False

    def test_none(self):
        assert _is_unreadable_workbook_error(None) is False


class TestParseJsonHelpers:
    def test_parse_dict_passthrough(self):
        assert _parse_json_dict({"a": 1}) == {"a": 1}

    def test_parse_dict_empty(self):
        assert _parse_json_dict(None) == {}
        assert _parse_json_dict("") == {}

    def test_parse_dict_from_json(self):
        assert _parse_json_dict('{"k": 2}') == {"k": 2}

    def test_parse_dict_json_list_returns_empty(self):
        assert _parse_json_dict("[1,2]") == {}

    def test_parse_dict_invalid(self):
        assert _parse_json_dict("{bad") == {}

    def test_parse_list_passthrough(self):
        assert _parse_json_list([1, 2]) == [1, 2]

    def test_parse_list_empty(self):
        assert _parse_json_list(None) == []

    def test_parse_list_from_json(self):
        assert _parse_json_list("[1,2,3]") == [1, 2, 3]

    def test_parse_list_json_dict_returns_empty(self):
        assert _parse_json_list('{"a":1}') == []

    def test_parse_list_invalid(self):
        assert _parse_json_list("[bad") == []


class TestListSheetNames:
    def test_lists_sheets(self, sample_xlsx):
        assert _list_excel_sheet_names(sample_xlsx) == ["Data"]

    def test_missing_file_returns_empty(self):
        assert _list_excel_sheet_names("/no/such/file.xlsx") == []


class TestStructuredPreview:
    def test_extracts_fields_and_rows(self, sample_xlsx):
        out = _extract_structured_excel_preview(sample_xlsx)
        labels = [f["label"] for f in out["fields"]]
        assert "产品名称" in labels
        assert out["sheet_name"] == "Data"
        assert len(out["sample_rows"]) >= 1

    def test_specific_sheet(self, sample_xlsx):
        out = _extract_structured_excel_preview(sample_xlsx, sheet_name="Data", sample_limit=2)
        assert out["sheet_name"] == "Data"

    def test_missing_file_returns_default(self):
        out = _extract_structured_excel_preview("/no/file.xlsx", sheet_name="X")
        assert out == {"fields": [], "sample_rows": [], "sheet_name": "X"}


class TestGridPreview:
    def test_grid_rows_with_merge(self, sample_xlsx):
        out = _extract_excel_grid_preview(sample_xlsx)
        assert out["sheet_name"] == "Data"
        assert isinstance(out["rows"], list) and out["rows"]
        first_cell = out["rows"][0][0]
        assert {"row", "col", "text", "rowspan", "colspan"} <= set(first_cell.keys())

    def test_missing_file_returns_default(self):
        assert _extract_excel_grid_preview("/no/file.xlsx") == {"sheet_name": "", "rows": []}


class TestStyleCache:
    def test_styles_and_refs(self, sample_xlsx):
        out = _extract_excel_grid_style_cache(sample_xlsx)
        assert out["sheet_name"] == "Data"
        assert out["styles"]
        assert out["cell_style_refs"]

    def test_missing_file_returns_default(self):
        out = _extract_excel_grid_style_cache("/no/file.xlsx")
        assert out == {"sheet_name": "", "styles": {}, "cell_style_refs": {}}


class TestLogicalTables:
    def test_detects_table(self, sample_xlsx):
        tables = _extract_logical_tables_from_sheet(sample_xlsx, sheet_name="Data")
        assert tables
        assert tables[0]["header_row"] == 1
        assert any(f["label"] == "产品名称" for f in tables[0]["fields"])

    def test_unknown_sheet_returns_empty(self, sample_xlsx):
        assert _extract_logical_tables_from_sheet(sample_xlsx, sheet_name="NoSuch") == []


class TestAllSheetsPreview:
    def test_collects_sheets(self, sample_xlsx):
        sheets = _extract_excel_all_sheets_preview(sample_xlsx)
        assert len(sheets) == 1
        assert sheets[0]["sheet_name"] == "Data"
        assert "grid_preview" in sheets[0]
        assert "style_cache" in sheets[0]
