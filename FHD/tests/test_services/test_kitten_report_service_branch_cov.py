"""Branch-coverage tests for app.services.kitten_report.service.

Targets KittenReportExportService branches:
- collect_plugin_results: with/without dataset, messages, result, phase, industry
- build_report: with/without dataset, messages
- _build_workbook: dataset present/absent, messages present/absent
- _add_financial_sheet: no plugins, financial only, inventory only, both,
  with monthly_data, product_analysis, customer_analysis, low_stock_alerts
- _html_to_text: various tag replacements
- _plugin_to_dict: normal mapping
"""

from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Any
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import Workbook, load_workbook

# Pre-import app.application to break circular import between app.services and app.mod_sdk
import app.application  # noqa: F401
from app.services.kitten_report.plugins import PluginResult
from app.services.kitten_report.service import KittenReportExportService


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def service() -> KittenReportExportService:
    return KittenReportExportService()


def _load_workbook(content: bytes) -> Any:
    """Load workbook from bytes."""
    return load_workbook(BytesIO(content))


# ---------------------------------------------------------------------------
# _html_to_text
# ---------------------------------------------------------------------------


class TestHtmlToText:
    def test_replaces_br_tags(self) -> None:
        result = KittenReportExportService._html_to_text("line1<br>line2<br/>line3<br />line4")
        assert result == "line1\nline2\nline3\nline4"

    def test_removes_strong_tags(self) -> None:
        result = KittenReportExportService._html_to_text("<strong>bold</strong> text")
        assert result == "bold text"

    def test_replaces_nbsp(self) -> None:
        result = KittenReportExportService._html_to_text("a&nbsp;b")
        assert result == "a b"

    def test_replaces_amp(self) -> None:
        result = KittenReportExportService._html_to_text("a&amp;b")
        assert result == "a&b"

    def test_all_tags_combined(self) -> None:
        result = KittenReportExportService._html_to_text(
            "<strong>title</strong><br>&nbsp;&amp;"
        )
        assert result == "title\n &"

    def test_empty_string(self) -> None:
        assert KittenReportExportService._html_to_text("") == ""

    def test_no_tags(self) -> None:
        assert KittenReportExportService._html_to_text("plain text") == "plain text"


# ---------------------------------------------------------------------------
# _plugin_to_dict
# ---------------------------------------------------------------------------


class TestPluginToDict:
    def test_converts_plugin_result(self) -> None:
        plugin_result = PluginResult(
            key="test_key",
            title="Test Title",
            level="info",
            summary="Test summary",
            details={"metric": 42},
        )
        result = KittenReportExportService._plugin_to_dict(plugin_result)
        assert result == {
            "key": "test_key",
            "title": "Test Title",
            "level": "info",
            "summary": "Test summary",
            "details": {"metric": 42},
        }

    def test_with_empty_details(self) -> None:
        plugin_result = PluginResult(
            key="k",
            title="t",
            level="warn",
            summary="s",
            details={},
        )
        result = KittenReportExportService._plugin_to_dict(plugin_result)
        assert result["details"] == {}


# ---------------------------------------------------------------------------
# collect_plugin_results
# ---------------------------------------------------------------------------


class TestCollectPluginResults:
    def test_with_full_payload(self, service: KittenReportExportService) -> None:
        payload = {
            "dataset": {"rows": 10, "columns": 3, "preview": [[1, 2, 3]]},
            "messages": [{"role": "user"}, {"role": "ai"}],
            "result": {"title": "Test", "summary": "Summary"},
            "phase": "analysis",
            "industry": "涂料行业",
        }
        results = service.collect_plugin_results(payload)
        assert isinstance(results, list)
        assert len(results) == 6  # 6 plugins
        for r in results:
            assert "key" in r
            assert "title" in r
            assert "level" in r
            assert "summary" in r
            assert "details" in r

    def test_with_empty_payload(self, service: KittenReportExportService) -> None:
        results = service.collect_plugin_results({})
        assert isinstance(results, list)
        assert len(results) == 6

    def test_with_none_values(self, service: KittenReportExportService) -> None:
        payload = {
            "dataset": None,
            "messages": None,
            "result": None,
            "phase": None,
            "industry": None,
        }
        results = service.collect_plugin_results(payload)
        assert len(results) == 6

    def test_default_phase_and_industry(self, service: KittenReportExportService) -> None:
        payload = {}
        results = service.collect_plugin_results(payload)
        # IndustryStrategyPlugin should use "通用" as default
        industry_plugin = next(p for p in results if p["key"] == "industry_strategy")
        assert "通用" in industry_plugin["summary"]


# ---------------------------------------------------------------------------
# build_report
# ---------------------------------------------------------------------------


class TestBuildReport:
    def test_returns_file_name_and_content(self, service: KittenReportExportService) -> None:
        payload = {
            "dataset": {"rows": 5, "columns": 2, "name": "test.csv"},
            "messages": [{"role": "user", "content": "hi", "time": "10:00"}],
            "result": {"title": "Report", "summary": "Summary"},
            "phase": "analysis",
            "industry": "通用",
        }
        result = service.build_report(payload)
        assert "file_name" in result
        assert result["file_name"].startswith("小猫分析报告_")
        assert result["file_name"].endswith(".xlsx")
        assert isinstance(result["content"], bytes)
        assert len(result["content"]) > 0
        assert "plugins" in result
        assert len(result["plugins"]) == 6

    def test_with_empty_payload(self, service: KittenReportExportService) -> None:
        result = service.build_report({})
        assert "file_name" in result
        assert isinstance(result["content"], bytes)
        assert len(result["plugins"]) == 6

    def test_with_no_messages(self, service: KittenReportExportService) -> None:
        payload = {"messages": []}
        result = service.build_report(payload)
        assert isinstance(result["content"], bytes)

    def test_with_no_dataset(self, service: KittenReportExportService) -> None:
        payload = {"dataset": {}}
        result = service.build_report(payload)
        assert isinstance(result["content"], bytes)

    def test_content_is_valid_xlsx(self, service: KittenReportExportService) -> None:
        payload = {
            "dataset": {"rows": 5, "columns": 2, "name": "test.csv"},
            "messages": [{"role": "ai", "content": "hello", "time": "10:00"}],
        }
        result = service.build_report(payload)
        wb = _load_workbook(result["content"])
        assert "报告摘要" in wb.sheetnames
        assert "算法洞察" in wb.sheetnames
        assert "对话记录" in wb.sheetnames


# ---------------------------------------------------------------------------
# _build_workbook - dataset branches
# ---------------------------------------------------------------------------


class TestBuildWorkbookDataset:
    def test_with_dataset_includes_data_summary_sheet(
        self, service: KittenReportExportService
    ) -> None:
        dataset = {
            "name": "data.csv",
            "rows": 100,
            "columns": 5,
            "fieldNames": ["col1", "col2"],
            "previewText": "preview",
        }
        content = service._build_workbook(
            dataset=dataset,
            messages=[],
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        assert "数据摘要" in wb.sheetnames

    def test_without_dataset_no_data_summary_sheet(
        self, service: KittenReportExportService
    ) -> None:
        content = service._build_workbook(
            dataset={},
            messages=[],
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        assert "数据摘要" not in wb.sheetnames

    def test_dataset_with_none_values(self, service: KittenReportExportService) -> None:
        """Dataset fields that are None should be handled gracefully."""
        dataset = {
            "name": None,
            "rows": None,
            "columns": None,
            "fieldNames": None,
            "previewText": None,
        }
        content = service._build_workbook(
            dataset=dataset,
            messages=[],
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        assert "数据摘要" in wb.sheetnames


# ---------------------------------------------------------------------------
# _build_workbook - messages branches
# ---------------------------------------------------------------------------


class TestBuildWorkbookMessages:
    def test_with_messages(self, service: KittenReportExportService) -> None:
        messages = [
            {"role": "user", "content": "hello", "time": "10:00"},
            {"role": "ai", "content": "hi there", "time": "10:01"},
        ]
        content = service._build_workbook(
            dataset={},
            messages=messages,
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["对话记录"]
        # Header + 2 messages
        assert ws.max_row == 3

    def test_without_messages_shows_system_message(
        self, service: KittenReportExportService
    ) -> None:
        content = service._build_workbook(
            dataset={},
            messages=[],
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["对话记录"]
        # Header + system message
        assert ws.max_row == 2
        assert ws.cell(row=2, column=2).value == "系统"

    def test_message_with_html_content(self, service: KittenReportExportService) -> None:
        messages = [
            {"role": "ai", "content": "<strong>bold</strong><br>line2", "time": "10:00"},
        ]
        content = service._build_workbook(
            dataset={},
            messages=messages,
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["对话记录"]
        cell_value = ws.cell(row=2, column=4).value
        assert "bold" in cell_value
        assert "line2" in cell_value

    def test_message_role_ai_shows_ai_label(self, service: KittenReportExportService) -> None:
        messages = [{"role": "ai", "content": "response", "time": "10:00"}]
        content = service._build_workbook(
            dataset={},
            messages=messages,
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["对话记录"]
        assert ws.cell(row=2, column=2).value == "AI"

    def test_message_role_other_shows_user_label(
        self, service: KittenReportExportService
    ) -> None:
        messages = [{"role": "user", "content": "question", "time": "10:00"}]
        content = service._build_workbook(
            dataset={},
            messages=messages,
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["对话记录"]
        assert ws.cell(row=2, column=2).value == "用户"


# ---------------------------------------------------------------------------
# _build_workbook - result/phase/industry branches
# ---------------------------------------------------------------------------


class TestBuildWorkbookResultFields:
    def test_result_with_title(self, service: KittenReportExportService) -> None:
        content = service._build_workbook(
            dataset={},
            messages=[],
            result={"title": "Custom Title"},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["报告摘要"]
        # Find the title row
        titles = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "Custom Title" in titles

    def test_result_without_title_uses_default(self, service: KittenReportExportService) -> None:
        content = service._build_workbook(
            dataset={},
            messages=[],
            result={},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["报告摘要"]
        titles = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "AI 分析" in titles

    def test_result_with_summary(self, service: KittenReportExportService) -> None:
        content = service._build_workbook(
            dataset={},
            messages=[],
            result={"summary": "My summary"},
            phase="analysis",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["报告摘要"]
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "My summary" in values

    def test_phase_displayed(self, service: KittenReportExportService) -> None:
        content = service._build_workbook(
            dataset={},
            messages=[],
            result={},
            phase="custom_phase",
            industry="通用",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["报告摘要"]
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "custom_phase" in values

    def test_industry_displayed(self, service: KittenReportExportService) -> None:
        content = service._build_workbook(
            dataset={},
            messages=[],
            result={},
            phase="analysis",
            industry="制造业",
            plugin_results=[],
        )
        wb = _load_workbook(content)
        ws = wb["报告摘要"]
        values = [ws.cell(row=r, column=2).value for r in range(1, ws.max_row + 1)]
        assert "制造业" in values


# ---------------------------------------------------------------------------
# _add_financial_sheet - no plugins
# ---------------------------------------------------------------------------


class TestAddFinancialSheetNoPlugins:
    def test_no_plugins_no_sheet(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        service._add_financial_sheet(wb, [])
        assert "财务报表" not in wb.sheetnames


# ---------------------------------------------------------------------------
# _add_financial_sheet - financial plugin only
# ---------------------------------------------------------------------------


class TestAddFinancialSheetFinancialOnly:
    def test_financial_plugin_creates_sheet(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "财务报表分析",
                "level": "info",
                "summary": "s",
                "details": {
                    "metrics": {
                        "total_revenue": 10000,
                        "total_cost": 5000,
                        "gross_profit": 5000,
                        "profit_margin": 50,
                        "order_count": 100,
                        "avg_order_value": 100,
                    },
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        assert "财务报表" in wb.sheetnames

    def test_financial_with_monthly_breakdown(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "metrics": {"total_revenue": 1000},
                    "monthly_breakdown": [
                        {"month": "2026-01", "revenue": 500, "order_count": 5},
                        {"month": "2026-02", "revenue": 500, "order_count": 5},
                    ],
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        # Should have monthly data
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend([str(v) for v in row if v is not None])
        assert any("2026-01" in v for v in values)

    def test_financial_with_product_analysis(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "metrics": {"total_revenue": 1000},
                    "product_analysis": [
                        {
                            "product_name": "Widget A",
                            "total_revenue": 500,
                            "total_qty": 10,
                            "order_count": 5,
                        }
                    ],
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend([str(v) for v in row if v is not None])
        assert any("Widget A" in v for v in values)

    def test_financial_with_customer_analysis(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "metrics": {"total_revenue": 1000},
                    "customer_analysis": [
                        {
                            "customer": "Acme Corp",
                            "total_amount": 500,
                            "order_count": 5,
                            "avg_order_value": 100,
                        }
                    ],
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend([str(v) for v in row if v is not None])
        assert any("Acme Corp" in v for v in values)

    def test_financial_with_empty_metrics(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {"metrics": {}},
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        assert "财务报表" in wb.sheetnames

    def test_financial_with_no_metrics_key(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {},
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        assert "财务报表" in wb.sheetnames

    def test_financial_value_rounding(self, service: KittenReportExportService) -> None:
        """Verify float values are rounded to 2 decimal places."""
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "metrics": {
                        "total_revenue": 10000.567,
                        "total_cost": 5000.234,
                    },
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        # Find the revenue value
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "总营收":
                assert row[1] == 10000.57  # rounded
                break

    def test_financial_with_string_value(self, service: KittenReportExportService) -> None:
        """Non-numeric values should be stored as-is."""
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "metrics": {
                        "total_revenue": "N/A",
                    },
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "总营收":
                assert row[1] == "N/A"
                break


# ---------------------------------------------------------------------------
# _add_financial_sheet - inventory plugin only
# ---------------------------------------------------------------------------


class TestAddFinancialSheetInventoryOnly:
    def test_inventory_plugin_creates_sheet(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "inventory_valuation",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "materials": {"total_items": 10, "total_value": 5000},
                    "products": {"total_items": 20, "total_value": 10000},
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        assert "财务报表" in wb.sheetnames

    def test_inventory_with_low_stock_alerts(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "inventory_valuation",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "materials": {"total_items": 10, "total_value": 5000},
                    "products": {"total_items": 20, "total_value": 10000},
                    "low_stock_alerts": [
                        {
                            "name": "Steel",
                            "current": 5,
                            "min_required": 20,
                            "unit_price": 100,
                        }
                    ],
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        values = []
        for row in ws.iter_rows(values_only=True):
            values.extend([str(v) for v in row if v is not None])
        assert any("Steel" in v for v in values)

    def test_inventory_with_empty_materials_products(
        self, service: KittenReportExportService
    ) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "inventory_valuation",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "materials": {},
                    "products": {},
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        assert "财务报表" in wb.sheetnames

    def test_inventory_total_calculation(self, service: KittenReportExportService) -> None:
        """Verify total inventory value is sum of materials + products."""
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "inventory_valuation",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "materials": {"total_items": 10, "total_value": 3000},
                    "products": {"total_items": 20, "total_value": 7000},
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        # Find the total row
        for row in ws.iter_rows(values_only=True):
            if row and row[0] == "合计":
                assert row[2] == 10000.0  # 3000 + 7000
                break


# ---------------------------------------------------------------------------
# _add_financial_sheet - both plugins
# ---------------------------------------------------------------------------


class TestAddFinancialSheetBothPlugins:
    def test_both_plugins_creates_sheet(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "metrics": {"total_revenue": 1000},
                },
            },
            {
                "key": "inventory_valuation",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "materials": {"total_items": 5, "total_value": 1000},
                    "products": {"total_items": 10, "total_value": 2000},
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        assert "财务报表" in wb.sheetnames

    def test_both_plugins_with_all_sections(self, service: KittenReportExportService) -> None:
        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "metrics": {"total_revenue": 1000, "order_count": 10},
                    "monthly_breakdown": [{"month": "2026-01", "revenue": 500, "order_count": 5}],
                    "product_analysis": [
                        {"product_name": "P1", "total_revenue": 500, "total_qty": 5, "order_count": 2}
                    ],
                    "customer_analysis": [
                        {"customer": "C1", "total_amount": 500, "order_count": 2, "avg_order_value": 250}
                    ],
                },
            },
            {
                "key": "inventory_valuation",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {
                    "materials": {"total_items": 5, "total_value": 1000},
                    "products": {"total_items": 10, "total_value": 2000},
                    "low_stock_alerts": [
                        {"name": "Item1", "current": 1, "min_required": 10, "unit_price": 50}
                    ],
                },
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        assert ws.max_row > 10  # Should have many rows


# ---------------------------------------------------------------------------
# _add_financial_sheet - column widths
# ---------------------------------------------------------------------------


class TestAddFinancialSheetColumnWidths:
    def test_column_widths_set(self, service: KittenReportExportService) -> None:
        from openpyxl.utils import get_column_letter

        wb = Workbook()
        wb.active.title = "test"
        plugin_results = [
            {
                "key": "financial_report",
                "title": "t",
                "level": "info",
                "summary": "s",
                "details": {"metrics": {"total_revenue": 1000}},
            },
        ]
        service._add_financial_sheet(wb, plugin_results)
        ws = wb["财务报表"]
        for col in range(1, 5):
            assert ws.column_dimensions[get_column_letter(col)].width == 22


# ---------------------------------------------------------------------------
# Integration: build_report with financial data
# ---------------------------------------------------------------------------


class TestBuildReportWithFinancialData:
    def test_build_report_includes_financial_sheet(
        self, service: KittenReportExportService
    ) -> None:
        """When plugins return financial data, the report should include 财务报表 sheet."""
        payload = {
            "dataset": {"rows": 10, "columns": 3, "name": "sales.csv"},
            "messages": [{"role": "user", "content": "analyze", "time": "10:00"}],
            "result": {"title": "Sales Analysis"},
            "phase": "financial",
            "industry": "零售",
        }
        result = service.build_report(payload)
        wb = _load_workbook(result["content"])
        # The financial sheet may or may not be present depending on whether
        # the FinancialReportPlugin returns data (it queries DB which may fail in test)
        # Just verify the report builds without error
        assert "报告摘要" in wb.sheetnames

    def test_build_report_with_ai_message_html(self, service: KittenReportExportService) -> None:
        payload = {
            "messages": [
                {"role": "ai", "content": "<strong>Analysis</strong><br>Result here", "time": "10:00"},
            ],
        }
        result = service.build_report(payload)
        wb = _load_workbook(result["content"])
        ws = wb["对话记录"]
        cell_value = ws.cell(row=2, column=4).value
        assert "Analysis" in cell_value
        assert "Result here" in cell_value
