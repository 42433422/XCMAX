"""Behavioural tests for app.services.kitten_report.service.KittenReportExportService.

Focus: the ``_add_financial_sheet`` worksheet builder (lines 156-400) which holds
most of the previously-uncovered branches — financial metrics table, monthly
breakdown, product/customer Top-10 rankings, inventory valuation and the
low-stock warning section — plus the early-return and helper paths.

These tests call ``_add_financial_sheet`` directly with hand-built
``plugin_results`` (the exact dict shape produced by ``_plugin_to_dict``), so no
DB / plugin / network dependency is touched. openpyxl is a pure in-memory lib,
so the assertions inspect real cell values written into the workbook.
"""

from __future__ import annotations

from openpyxl import Workbook

from app.services.kitten_report.service import KittenReportExportService


# ---------------------------------------------------------------------------
# Helpers / fixtures
# ---------------------------------------------------------------------------
def _service() -> KittenReportExportService:
    return KittenReportExportService()


def _financial_result(details: dict) -> dict:
    return {
        "key": "financial_report",
        "title": "财务报表分析",
        "level": "info",
        "summary": "s",
        "details": details,
    }


def _inventory_result(details: dict) -> dict:
    return {
        "key": "inventory_valuation",
        "title": "库存价值评估",
        "level": "warn",
        "summary": "s",
        "details": details,
    }


def _sheet_cells(ws) -> set[str]:
    """All non-empty cell string values of a worksheet."""
    out: set[str] = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                out.add(str(cell.value))
    return out


def _column_values(ws, col_idx: int) -> list:
    return [ws.cell(row=r, column=col_idx).value for r in range(1, ws.max_row + 1)]


# ---------------------------------------------------------------------------
# _add_financial_sheet — early return (line 164-165)
# ---------------------------------------------------------------------------
class TestEarlyReturn:
    def test_no_financial_no_inventory_does_not_create_sheet(self):
        svc = _service()
        wb = Workbook()
        plugin_results = [
            {"key": "rule_stats", "title": "x", "level": "info", "summary": "", "details": {}},
            {"key": "forecast", "title": "y", "level": "info", "summary": "", "details": {}},
        ]
        svc._add_financial_sheet(wb, plugin_results)
        # Early return: no 财务报表 sheet must be created.
        assert "财务报表" not in wb.sheetnames

    def test_empty_plugin_results_returns_early(self):
        svc = _service()
        wb = Workbook()
        svc._add_financial_sheet(wb, [])
        assert "财务报表" not in wb.sheetnames


# ---------------------------------------------------------------------------
# _add_financial_sheet — financial core metrics table (lines 177-222)
# ---------------------------------------------------------------------------
class TestFinancialMetricsTable:
    def test_core_metrics_written_and_rounded(self):
        svc = _service()
        wb = Workbook()
        details = {
            "metrics": {
                "total_revenue": 10000.555,
                "total_cost": 6000.0,
                "gross_profit": 4000.555,
                "profit_margin": 40.123,
                "order_count": 50,
                "avg_order_value": 200.0,
            }
        }
        svc._add_financial_sheet(wb, [_financial_result(details)])

        assert "财务报表" in wb.sheetnames
        ws = wb["财务报表"]
        cells = _sheet_cells(ws)
        # Title + section header present.
        assert "📊 小猫财务分析报告" in cells
        assert "核心财务指标" in cells
        # Metric labels rendered.
        for label in ("总营收", "成本估算", "毛利润", "毛利率", "订单数量", "平均订单额"):
            assert label in cells
        # Float values are rounded to 2 dp before writing.
        col2 = _column_values(ws, 2)
        assert 10000.56 in col2  # 10000.555 -> 10000.56
        assert 40.12 in col2  # 40.123 -> 40.12

    def test_non_numeric_metric_value_passed_through(self):
        # round() branch is guarded by isinstance(value, (int, float)); a string
        # value must be written verbatim (covers the else of line 217).
        svc = _service()
        wb = Workbook()
        details = {"metrics": {"total_revenue": "N/A"}}
        svc._add_financial_sheet(wb, [_financial_result(details)])
        ws = wb["财务报表"]
        assert "N/A" in _column_values(ws, 2)

    def test_metrics_missing_keys_default_to_zero(self):
        svc = _service()
        wb = Workbook()
        # Empty metrics dict -> every metric defaults to 0 via .get(..., 0).
        svc._add_financial_sheet(wb, [_financial_result({"metrics": {}})])
        ws = wb["财务报表"]
        col2 = _column_values(ws, 2)
        # All six metric rows default to 0.
        assert col2.count(0) >= 6


# ---------------------------------------------------------------------------
# monthly breakdown section (lines 224-251)
# ---------------------------------------------------------------------------
class TestMonthlyBreakdown:
    def test_monthly_rows_written(self):
        svc = _service()
        wb = Workbook()
        details = {
            "metrics": {"total_revenue": 1.0},
            "monthly_breakdown": [
                {"month": "2026-01", "revenue": 5000.126, "order_count": 25},
                {"month": "2026-02", "revenue": 4999.874, "order_count": 30},
            ],
        }
        svc._add_financial_sheet(wb, [_financial_result(details)])
        ws = wb["财务报表"]
        cells = _sheet_cells(ws)
        assert "月度营收趋势" in cells
        assert "2026-01" in cells
        assert "2026-02" in cells
        # Revenue is rounded to 2 dp.
        col2 = _column_values(ws, 2)
        assert 5000.13 in col2
        assert 4999.87 in col2
        # Order counts present in col 3.
        col3 = _column_values(ws, 3)
        assert 25 in col3 and 30 in col3

    def test_empty_monthly_breakdown_skips_section(self):
        svc = _service()
        wb = Workbook()
        details = {"metrics": {"total_revenue": 1.0}, "monthly_breakdown": []}
        svc._add_financial_sheet(wb, [_financial_result(details)])
        ws = wb["财务报表"]
        assert "月度营收趋势" not in _sheet_cells(ws)


# ---------------------------------------------------------------------------
# product analysis Top-10 (lines 253-283)
# ---------------------------------------------------------------------------
class TestProductAnalysis:
    def test_product_rows_and_top10_cap(self):
        svc = _service()
        wb = Workbook()
        products = [
            {
                "product_name": f"P{i}",
                "total_revenue": float(i) + 0.005,
                "total_qty": float(i),
                "order_count": i,
            }
            for i in range(1, 13)  # 12 products -> only first 10 written
        ]
        details = {"metrics": {"total_revenue": 1.0}, "product_analysis": products}
        svc._add_financial_sheet(wb, [_financial_result(details)])
        ws = wb["财务报表"]
        col1 = _column_values(ws, 1)
        assert "产品销售排行 (Top 10)" in col1
        # P1..P10 present, P11/P12 excluded by [:10] slice.
        assert "P1" in col1 and "P10" in col1
        assert "P11" not in col1
        assert "P12" not in col1

    def test_empty_product_analysis_skips_section(self):
        svc = _service()
        wb = Workbook()
        details = {"metrics": {"total_revenue": 1.0}, "product_analysis": []}
        svc._add_financial_sheet(wb, [_financial_result(details)])
        assert "产品销售排行 (Top 10)" not in _column_values(wb["财务报表"], 1)


# ---------------------------------------------------------------------------
# customer analysis Top-10 (lines 285-317)
# ---------------------------------------------------------------------------
class TestCustomerAnalysis:
    def test_customer_rows_and_rounding(self):
        svc = _service()
        wb = Workbook()
        customers = [
            {
                "customer": "客户甲",
                "total_amount": 3000.005,
                "order_count": 15,
                "avg_order_value": 200.004,
            },
            {
                "customer": "客户乙",
                "total_amount": 1500.0,
                "order_count": 8,
                "avg_order_value": 187.5,
            },
        ]
        details = {"metrics": {"total_revenue": 1.0}, "customer_analysis": customers}
        svc._add_financial_sheet(wb, [_financial_result(details)])
        ws = wb["财务报表"]
        cells = _sheet_cells(ws)
        assert "客户销售排行 (Top 10)" in cells
        assert "客户甲" in cells and "客户乙" in cells
        col2 = _column_values(ws, 2)
        assert 3000.01 in col2  # rounded total_amount
        # avg_order_value rounded -> 200.0 in col4
        assert 200.0 in _column_values(ws, 4)

    def test_customer_top10_cap(self):
        svc = _service()
        wb = Workbook()
        customers = [
            {
                "customer": f"C{i}",
                "total_amount": float(i),
                "order_count": i,
                "avg_order_value": float(i),
            }
            for i in range(1, 15)
        ]
        details = {"metrics": {"total_revenue": 1.0}, "customer_analysis": customers}
        svc._add_financial_sheet(wb, [_financial_result(details)])
        col1 = _column_values(wb["财务报表"], 1)
        assert "C10" in col1
        assert "C11" not in col1


# ---------------------------------------------------------------------------
# inventory valuation section (lines 319-363)
# ---------------------------------------------------------------------------
class TestInventoryValuation:
    def test_inventory_totals_and_grand_total(self):
        svc = _service()
        wb = Workbook()
        details = {
            "materials": {"total_items": 5, "total_value": 1000.126},
            "products": {"total_items": 3, "total_value": 2000.0},
        }
        svc._add_financial_sheet(wb, [_inventory_result(details)])
        ws = wb["财务报表"]
        cells = _sheet_cells(ws)
        assert "📦 库存价值评估" in cells
        assert "原材料" in cells
        assert "成品/产品" in cells
        assert "合计" in cells
        col3 = _column_values(ws, 3)
        # material value rounded
        assert 1000.13 in col3
        assert 2000.0 in col3
        # grand total = 1000.126 + 2000.0 -> round(3000.126, 2) = 3000.13
        assert 3000.13 in col3

    def test_inventory_missing_sections_default_zero(self):
        svc = _service()
        wb = Workbook()
        # materials / products absent -> .get("materials", {}) -> defaults to 0.
        svc._add_financial_sheet(wb, [_inventory_result({})])
        ws = wb["财务报表"]
        col2 = _column_values(ws, 2)  # item counts
        assert col2.count(0) >= 2


# ---------------------------------------------------------------------------
# low-stock alerts section (lines 365-397)
# ---------------------------------------------------------------------------
class TestLowStockAlerts:
    def test_low_stock_rows_written(self):
        svc = _service()
        wb = Workbook()
        details = {
            "materials": {"total_items": 1, "total_value": 100.0},
            "products": {"total_items": 0, "total_value": 0.0},
            "low_stock_alerts": [
                {"name": "螺丝", "current": 5.0, "min_required": 10.0, "unit_price": 0.5},
                {"name": "钢板", "current": 2.0, "min_required": 20.0, "unit_price": 80.0},
            ],
        }
        svc._add_financial_sheet(wb, [_inventory_result(details)])
        ws = wb["财务报表"]
        cells = _sheet_cells(ws)
        assert "⚠️ 低库存预警" in cells
        assert "螺丝" in cells and "钢板" in cells
        # alert header row written.
        assert "当前库存" in cells and "安全库存" in cells
        col2 = _column_values(ws, 2)
        assert 5.0 in col2 and 2.0 in col2

    def test_low_stock_top10_cap(self):
        svc = _service()
        wb = Workbook()
        alerts = [
            {"name": f"M{i}", "current": float(i), "min_required": 100.0, "unit_price": 1.0}
            for i in range(1, 15)
        ]
        details = {
            "materials": {"total_items": 1, "total_value": 1.0},
            "products": {"total_items": 0, "total_value": 0.0},
            "low_stock_alerts": alerts,
        }
        svc._add_financial_sheet(wb, [_inventory_result(details)])
        col1 = _column_values(wb["财务报表"], 1)
        assert "M10" in col1
        assert "M11" not in col1

    def test_no_low_stock_skips_alert_section(self):
        svc = _service()
        wb = Workbook()
        details = {
            "materials": {"total_items": 1, "total_value": 1.0},
            "products": {"total_items": 0, "total_value": 0.0},
            "low_stock_alerts": [],
        }
        svc._add_financial_sheet(wb, [_inventory_result(details)])
        assert "⚠️ 低库存预警" not in _sheet_cells(wb["财务报表"])


# ---------------------------------------------------------------------------
# combined financial + inventory in one sheet
# ---------------------------------------------------------------------------
class TestCombinedSheet:
    def test_financial_and_inventory_both_render(self):
        svc = _service()
        wb = Workbook()
        fin = _financial_result(
            {
                "metrics": {"total_revenue": 100.0},
                "monthly_breakdown": [{"month": "2026-03", "revenue": 100.0, "order_count": 1}],
                "product_analysis": [
                    {"product_name": "Z", "total_revenue": 1.0, "total_qty": 1.0, "order_count": 1}
                ],
                "customer_analysis": [
                    {"customer": "K", "total_amount": 1.0, "order_count": 1, "avg_order_value": 1.0}
                ],
            }
        )
        inv = _inventory_result(
            {
                "materials": {"total_items": 1, "total_value": 10.0},
                "products": {"total_items": 1, "total_value": 20.0},
                "low_stock_alerts": [
                    {"name": "W", "current": 1.0, "min_required": 5.0, "unit_price": 2.0}
                ],
            }
        )
        svc._add_financial_sheet(wb, [fin, inv])
        cells = _sheet_cells(wb["财务报表"])
        # Only one 财务报表 sheet is created and holds both sections.
        assert wb.sheetnames.count("财务报表") == 1
        assert "核心财务指标" in cells
        assert "📦 库存价值评估" in cells
        assert "月度营收趋势" in cells
        assert "⚠️ 低库存预警" in cells

    def test_column_widths_set(self):
        svc = _service()
        wb = Workbook()
        svc._add_financial_sheet(wb, [_financial_result({"metrics": {"total_revenue": 1.0}})])
        ws = wb["财务报表"]
        # Final loop sets columns A..D width to 22.
        for col in ("A", "B", "C", "D"):
            assert ws.column_dimensions[col].width == 22


# ---------------------------------------------------------------------------
# End-to-end through build_report with patched plugins to drive financial sheet
# ---------------------------------------------------------------------------
class TestBuildReportIntegration:
    def test_build_report_returns_valid_xlsx_with_financial_section(self):
        from openpyxl import load_workbook

        svc = _service()
        # Replace plugin instances with stubs so no DB is touched and we control
        # the exact financial/inventory data flowing into the financial sheet.
        from app.services.kitten_report.plugins import PluginResult

        class _StubFin:
            def run(self, payload):
                return PluginResult(
                    key="financial_report",
                    title="财务报表分析",
                    level="info",
                    summary="ok",
                    details={
                        "metrics": {"total_revenue": 500.0, "order_count": 3},
                        "monthly_breakdown": [
                            {"month": "2026-04", "revenue": 500.0, "order_count": 3}
                        ],
                        "product_analysis": [],
                        "customer_analysis": [],
                    },
                )

        class _StubInv:
            def run(self, payload):
                return PluginResult(
                    key="inventory_valuation",
                    title="库存价值评估",
                    level="info",
                    summary="ok",
                    details={
                        "materials": {"total_items": 0, "total_value": 0.0},
                        "products": {"total_items": 0, "total_value": 0.0},
                        "low_stock_alerts": [],
                    },
                )

        svc.plugins = [_StubFin(), _StubInv()]

        out = svc.build_report({"phase": "p", "industry": "制造业"})
        assert out["file_name"].endswith(".xlsx")
        assert out["file_name"].startswith("小猫分析报告_")
        content = out["content"]
        assert isinstance(content, bytes) and len(content) > 0

        # The bytes are a real, openable xlsx with the financial sheet present.
        from io import BytesIO

        wb = load_workbook(BytesIO(content))
        assert "财务报表" in wb.sheetnames
        assert "📊 小猫财务分析报告" in _sheet_cells(wb["财务报表"])
        # plugins list returned in the result.
        keys = {p["key"] for p in out["plugins"]}
        assert "financial_report" in keys and "inventory_valuation" in keys
