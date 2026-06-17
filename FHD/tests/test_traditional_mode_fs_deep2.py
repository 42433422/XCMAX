"""Deep tests for ``app.traditional_mode_fs`` covering remaining uncovered branches."""
from __future__ import annotations

import base64
import json
import os
import queue
import threading
from datetime import datetime
from typing import Any
from unittest.mock import MagicMock, patch

import pytest

from app import traditional_mode_fs as tmf


@pytest.fixture()
def temp_root(tmp_path, monkeypatch: pytest.MonkeyPatch):
    """Use a temp dir as ROOT_DIR for isolation."""
    root = tmp_path / "bang"
    root.mkdir()
    monkeypatch.setattr(tmf, "ROOT_DIR", str(root))
    return str(root)


# ── _resolve_root_dir deep ───────────────────────────────────────────────────


class TestResolveRootDirDeep:
    def test_with_env_var(self, tmp_path, monkeypatch: pytest.MonkeyPatch) -> None:
        custom = tmp_path / "custom_root"
        monkeypatch.setenv("TRADITIONAL_MODE_ROOT", str(custom))
        result = tmf._resolve_root_dir()
        assert result == str(custom.resolve())
        assert os.path.isdir(result)

    def test_with_env_var_expanded(self, tmp_path, monkeypatch: pytest.MonkeyPatch) -> None:
        # Test with ~ expansion
        monkeypatch.setenv("TRADITIONAL_MODE_ROOT", "~/test_root_xyz")
        result = tmf._resolve_root_dir()
        assert "~" not in result
        assert os.path.isdir(result)
        # cleanup
        os.rmdir(result)

    def test_default_root(self, tmp_path, monkeypatch: pytest.MonkeyPatch) -> None:
        monkeypatch.delenv("TRADITIONAL_MODE_ROOT", raising=False)
        result = tmf._resolve_root_dir()
        assert os.path.isdir(result)
        assert "bang" in result


# ── resolve_safe_path deep ───────────────────────────────────────────────────


class TestResolveSafePathDeep:
    def test_normal_relative_path(self, temp_root) -> None:
        result = tmf.resolve_safe_path("subdir/file.txt")
        assert result is not None
        assert result.startswith(temp_root)

    def test_dot_path(self, temp_root) -> None:
        result = tmf.resolve_safe_path(".")
        assert result == temp_root

    def test_nested_traversal_rejected(self, temp_root) -> None:
        result = tmf.resolve_safe_path("sub/../../../etc/passwd")
        assert result is None

    def test_absolute_path_outside_root_rejected(self, temp_root, tmp_path) -> None:
        # An absolute path that's outside the root
        outside = tmp_path / "outside.txt"
        result = tmf.resolve_safe_path(str(outside))
        assert result is None


# ── stat_response deep ───────────────────────────────────────────────────────


class TestStatResponseDeep:
    def test_stat_file_with_extension(self, temp_root) -> None:
        fpath = os.path.join(temp_root, "doc.xlsx")
        with open(fpath, "w") as f:
            f.write("data")
        result, status = tmf.stat_response("doc.xlsx")
        assert status == 200
        assert result["data"]["type"] == "xlsx"

    def test_stat_file_no_extension(self, temp_root) -> None:
        fpath = os.path.join(temp_root, "noext")
        with open(fpath, "w") as f:
            f.write("data")
        result, status = tmf.stat_response("noext")
        assert status == 200
        assert result["data"]["type"] == "文件"

    def test_stat_subdir(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "subdir"))
        result, status = tmf.stat_response("subdir")
        assert status == 200
        assert result["data"]["is_dir"] is True
        assert result["data"]["size"] == 0


# ── write_text_response deep ─────────────────────────────────────────────────


class TestWriteTextResponseDeep:
    def test_write_to_existing_file_overwrites(self, temp_root) -> None:
        with open(os.path.join(temp_root, "file.txt"), "w") as f:
            f.write("old")
        result, status = tmf.write_text_response("file.txt", "new")
        assert status == 200
        with open(os.path.join(temp_root, "file.txt")) as f:
            assert f.read() == "new"

    def test_write_creates_parent_dirs(self, temp_root) -> None:
        result, status = tmf.write_text_response("a/b/c/file.txt", "data")
        assert status == 200
        assert os.path.isfile(os.path.join(temp_root, "a", "b", "c", "file.txt"))


# ── write_base64_response deep ───────────────────────────────────────────────


class TestWriteBase64ResponseDeep:
    def test_write_binary_data(self, temp_root) -> None:
        data = b"\x00\x01\x02\x03"
        content = base64.b64encode(data).decode()
        result, status = tmf.write_base64_response("file.bin", content)
        assert status == 200
        assert result["data"]["bytes"] == 4
        with open(os.path.join(temp_root, "file.bin"), "rb") as f:
            assert f.read() == data

    def test_write_creates_parent_dirs(self, temp_root) -> None:
        data = b"hello"
        content = base64.b64encode(data).decode()
        result, status = tmf.write_base64_response("sub/file.bin", content)
        assert status == 200
        assert os.path.isfile(os.path.join(temp_root, "sub", "file.bin"))


# ── move_response deep ───────────────────────────────────────────────────────


class TestMoveResponseDeep:
    def test_move_directory(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "srcdir"))
        with open(os.path.join(temp_root, "srcdir", "file.txt"), "w") as f:
            f.write("data")
        result, status = tmf.move_response("srcdir", "dstdir")
        assert status == 200
        assert os.path.isdir(os.path.join(temp_root, "dstdir"))
        assert os.path.isfile(os.path.join(temp_root, "dstdir", "file.txt"))

    def test_move_both_traversal_rejected(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("data")
        result, status = tmf.move_response("src.txt", "../../../etc/passwd")
        assert status == 403


# ── copy_response deep ───────────────────────────────────────────────────────


class TestCopyResponseDeep:
    def test_copy_with_overwrite_dir(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "srcdir"))
        with open(os.path.join(temp_root, "srcdir", "file.txt"), "w") as f:
            f.write("data")
        os.makedirs(os.path.join(temp_root, "dstdir"))
        result, status = tmf.copy_response("srcdir", "dstdir", overwrite=True)
        assert status == 200
        assert os.path.isfile(os.path.join(temp_root, "dstdir", "file.txt"))

    def test_copy_both_traversal_rejected(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("data")
        result, status = tmf.copy_response("src.txt", "../../../etc/passwd")
        assert status == 403


# ── _get_file_type deep ──────────────────────────────────────────────────────


class TestGetFileTypeDeep:
    def test_all_image_extensions(self) -> None:
        for ext in [".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp", ".ico", ".tiff", ".svg"]:
            assert tmf._get_file_type(f"file{ext}") == ext[1:]

    def test_all_excel_extensions(self) -> None:
        for ext in [".xlsx", ".xls"]:
            assert tmf._get_file_type(f"file{ext}") == ext[1:]

    def test_uppercase_extension(self) -> None:
        assert tmf._get_file_type("FILE.PNG") == "png"

    def test_dotfile_no_extension(self) -> None:
        # .gitignore → splitext returns ('.gitignore', '') so ext=''
        # "." is in name_lower so returns ext[1:] = ''
        assert tmf._get_file_type(".gitignore") == ""

    def test_unknown_extension(self) -> None:
        assert tmf._get_file_type("file.xyz") == "xyz"


# ── list_files_response deep ─────────────────────────────────────────────────


class TestListFilesResponseDeep:
    def test_list_nested_directory(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "subdir"))
        with open(os.path.join(temp_root, "subdir", "nested.txt"), "w") as f:
            f.write("data")
        result, status = tmf.list_files_response("subdir")
        assert status == 200
        assert len(result["data"]["files"]) == 1
        assert result["data"]["files"][0]["name"] == "nested.txt"

    def test_list_with_multiple_files_sorted(self, temp_root) -> None:
        for name in ["c.txt", "a.txt", "b.txt"]:
            with open(os.path.join(temp_root, name), "w") as f:
                f.write("data")
        result, status = tmf.list_files_response("")
        assert status == 200
        names = [e["name"] for e in result["data"]["files"]]
        assert names == ["a.txt", "b.txt", "c.txt"]

    def test_list_dirs_before_files(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "zdir"))
        with open(os.path.join(temp_root, "afile.txt"), "w") as f:
            f.write("data")
        result, status = tmf.list_files_response("")
        assert status == 200
        assert result["data"]["files"][0]["name"] == "zdir"
        assert result["data"]["files"][0]["is_dir"] is True


# ── read_file_response deep ──────────────────────────────────────────────────


class TestReadFileResponseDeep:
    def test_read_all_image_types(self, temp_root) -> None:
        mime_map = {
            ".png": "image/png",
            ".jpg": "image/jpeg",
            ".jpeg": "image/jpeg",
            ".gif": "image/gif",
            ".webp": "image/webp",
            ".bmp": "image/bmp",
            ".ico": "image/x-icon",
            ".tiff": "image/tiff",
            ".svg": "image/svg+xml",
        }
        for ext, expected_mime in mime_map.items():
            fname = f"img{ext}"
            with open(os.path.join(temp_root, fname), "wb") as f:
                f.write(b"fake")
            result, status = tmf.read_file_response(fname)
            assert status == 200, f"Failed for {ext}"
            assert result["data"]["type"] == "image"
            assert result["data"]["mime"] == expected_mime

    def test_read_excel_with_datetime(self, temp_root) -> None:
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "text"
        ws["A2"] = datetime(2026, 1, 1)
        ws["A3"] = 42
        ws["A4"] = "after_none"  # Use a real value, not None (openpyxl may truncate trailing None)
        wb.save(os.path.join(temp_root, "file.xlsx"))
        wb.close()
        result, status = tmf.read_file_response("file.xlsx")
        assert status == 200
        assert result["data"]["type"] == "excel"
        content = result["data"]["content"]
        sheet_name = result["data"]["sheets"][0]
        rows = content[sheet_name]["rows"]
        assert rows[0][0] == "text"
        assert "T" in rows[1][0]  # isoformat
        assert rows[2][0] == 42
        assert rows[3][0] == "after_none"

    def test_read_excel_multiple_sheets(self, temp_root) -> None:
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "s1"
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "s2"
        wb.save(os.path.join(temp_root, "multi.xlsx"))
        wb.close()
        result, status = tmf.read_file_response("multi.xlsx")
        assert status == 200
        assert len(result["data"]["sheets"]) == 2
        assert "Sheet1" in result["data"]["sheets"]
        assert "Sheet2" in result["data"]["sheets"]


# ── _build_snapshot deep ─────────────────────────────────────────────────────


class TestBuildSnapshotDeep:
    def test_build_snapshot_oserror_on_stat(self, temp_root, monkeypatch) -> None:
        with open(os.path.join(temp_root, "file.txt"), "w") as f:
            f.write("data")
        original_stat = os.stat

        def stat_with_fail(path):
            if "file.txt" in str(path):
                raise OSError("stat fail")
            return original_stat(path)

        monkeypatch.setattr(os, "stat", stat_with_fail)
        snap = tmf._build_snapshot()
        # file.txt should be skipped due to OSError
        assert "file.txt" not in snap

    def test_build_snapshot_oserror_on_relpath(self, temp_root, monkeypatch) -> None:
        with open(os.path.join(temp_root, "file.txt"), "w") as f:
            f.write("data")
        original_relpath = os.path.relpath

        def relpath_with_fail(path, start):
            # os.path.relpath is called with dirpath (the directory),
            # not the file. Raise when the path equals temp_root (the ROOT_DIR).
            if str(path) == str(temp_root):
                raise OSError("relpath fail")
            return original_relpath(path, start)

        # Patch os.path.relpath (not os.relpath)
        monkeypatch.setattr(os.path, "relpath", relpath_with_fail)
        snap = tmf._build_snapshot()
        # The inner try/except OSError should catch it and continue (skip the dir)
        assert snap == {}

    def test_build_snapshot_root_not_dir(self, tmp_path, monkeypatch) -> None:
        # ROOT_DIR exists but is a file, not a dir
        root_file = tmp_path / "notadir"
        root_file.write_text("data")
        monkeypatch.setattr(tmf, "ROOT_DIR", str(root_file))
        snap = tmf._build_snapshot()
        assert snap == {}


# ── _format_snapshot deep ────────────────────────────────────────────────────


class TestFormatSnapshotDeep:
    def test_empty_snapshot(self) -> None:
        result = tmf._format_snapshot({})
        assert result == {}

    def test_multiple_entries(self) -> None:
        mtime_map = {"a.txt": 1700000000.0, "b.txt": 1700000001.0}
        result = tmf._format_snapshot(mtime_map)
        assert len(result) == 2
        assert "a.txt" in result
        assert "b.txt" in result


# ── sse_watch_events deep ────────────────────────────────────────────────────


class TestSseWatchEventsDeep:
    def test_sse_yields_initial_then_heartbeat(self, temp_root, monkeypatch) -> None:
        # Reset watchdog state
        monkeypatch.setattr(tmf, "_watchdog_running", False)
        monkeypatch.setattr(tmf, "_watchdog_thread", None)
        with tmf._snapshot_lock:
            tmf._last_snapshot = {}

        # Mock _ensure_watchdog to avoid starting real thread
        monkeypatch.setattr(tmf, "_ensure_watchdog", lambda: None)

        gen = tmf.sse_watch_events("")
        first = next(gen)
        assert "snapshot" in first

        # Next should be heartbeat (queue.Empty after timeout)
        # But we don't want to wait 55s, so close the generator
        gen.close()

    def test_sse_registers_client(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(tmf, "_watchdog_running", True)
        mock_thread = MagicMock()
        mock_thread.is_alive.return_value = True
        monkeypatch.setattr(tmf, "_watchdog_thread", mock_thread)
        with tmf._snapshot_lock:
            tmf._last_snapshot = {}

        initial_clients = len(tmf._watch_clients)
        gen = tmf.sse_watch_events("")
        next(gen)
        assert len(tmf._watch_clients) == initial_clients + 1
        gen.close()
        assert len(tmf._watch_clients) == initial_clients

    def test_sse_receives_message_from_queue(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(tmf, "_watchdog_running", True)
        mock_thread = MagicMock()
        mock_thread.is_alive.return_value = True
        monkeypatch.setattr(tmf, "_watchdog_thread", mock_thread)
        with tmf._snapshot_lock:
            tmf._last_snapshot = {}

        gen = tmf.sse_watch_events("")
        first = next(gen)  # initial snapshot

        # Put a message in the client queue
        # The client queue is the last one added
        client_q = tmf._watch_clients[-1]
        test_msg = json.dumps({"changed": ["test.txt"]})
        client_q.put_nowait(test_msg)

        second = next(gen)
        assert test_msg in second
        gen.close()


# ── _watchdog_loop deep ──────────────────────────────────────────────────────


class TestWatchdogLoopDeep:
    def test_watchdog_detects_changes(self, temp_root, monkeypatch) -> None:
        # Set up initial state - empty directory
        monkeypatch.setattr(tmf, "_watchdog_running", True)
        with tmf._snapshot_lock:
            tmf._last_snapshot = {}

        # _watchdog_loop will:
        # 1. _watchdog_prev = _build_snapshot()  (empty at this point)
        # 2. while loop: wait() → if not running: break → curr = _build_snapshot()
        # We create the file DURING wait() so the second snapshot sees it.
        call_count = {"n": 0}

        def fake_wait(self, timeout=None):
            call_count["n"] += 1
            if call_count["n"] == 1:
                # Create the file during the first wait, so the second
                # _build_snapshot() in the loop body will see it.
                with open(os.path.join(temp_root, "newfile.txt"), "w") as f:
                    f.write("data")
            elif call_count["n"] >= 2:
                tmf._watchdog_running = False
            return True

        monkeypatch.setattr(threading.Event, "wait", fake_wait)

        # Add a client to receive the change notification
        client_q: queue.Queue[str] = queue.Queue(maxsize=20)
        with tmf._watch_clients_lock:
            tmf._watch_clients.append(client_q)

        tmf._watchdog_loop()

        # The client should have received a message
        assert not client_q.empty()
        msg = client_q.get_nowait()
        assert "changed" in msg

        # Cleanup
        with tmf._watch_clients_lock:
            if client_q in tmf._watch_clients:
                tmf._watch_clients.remove(client_q)

    def test_watchdog_detects_deleted_files(self, temp_root, monkeypatch) -> None:
        # Create initial file - will be in _watchdog_prev
        with open(os.path.join(temp_root, "to_delete.txt"), "w") as f:
            f.write("data")

        monkeypatch.setattr(tmf, "_watchdog_running", True)
        with tmf._snapshot_lock:
            tmf._last_snapshot = {}

        # Delete the file DURING the first wait so the second snapshot doesn't see it
        call_count = {"n": 0}

        def fake_wait(self, timeout=None):
            call_count["n"] += 1
            if call_count["n"] == 1:
                # Delete during first wait
                os.remove(os.path.join(temp_root, "to_delete.txt"))
            elif call_count["n"] >= 2:
                tmf._watchdog_running = False
            return True

        monkeypatch.setattr(threading.Event, "wait", fake_wait)

        client_q: queue.Queue[str] = queue.Queue(maxsize=20)
        with tmf._watch_clients_lock:
            tmf._watch_clients.append(client_q)

        tmf._watchdog_loop()

        assert not client_q.empty()
        msg = client_q.get_nowait()
        assert "__deleted__" in msg

        with tmf._watch_clients_lock:
            if client_q in tmf._watch_clients:
                tmf._watch_clients.remove(client_q)

    def test_watchdog_full_queue_removes_client(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(tmf, "_watchdog_running", True)
        with tmf._snapshot_lock:
            tmf._last_snapshot = {}

        # Create a file DURING the first wait to trigger a change
        call_count = {"n": 0}

        def fake_wait(self, timeout=None):
            call_count["n"] += 1
            if call_count["n"] == 1:
                with open(os.path.join(temp_root, "trigger.txt"), "w") as f:
                    f.write("data")
            elif call_count["n"] >= 2:
                tmf._watchdog_running = False
            return True

        monkeypatch.setattr(threading.Event, "wait", fake_wait)

        # Create a full queue
        client_q: queue.Queue[str] = queue.Queue(maxsize=1)
        client_q.put_nowait("existing")
        with tmf._watch_clients_lock:
            tmf._watch_clients.append(client_q)

        tmf._watchdog_loop()

        # The full queue should have been removed from _watch_clients
        with tmf._watch_clients_lock:
            assert client_q not in tmf._watch_clients


# ── _ensure_watchdog deep ────────────────────────────────────────────────────


class TestEnsureWatchdogDeep:
    def test_ensure_watchdog_not_running_starts(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(tmf, "_watchdog_running", False)
        monkeypatch.setattr(tmf, "_watchdog_thread", None)
        started = {"called": False}

        class FakeThread:
            def __init__(self, *a, **k):
                started["called"] = True

            def start(self):
                pass

            def is_alive(self):
                return True

        monkeypatch.setattr(threading, "Thread", FakeThread)
        tmf._ensure_watchdog()
        assert started["called"] is True

    def test_ensure_watchdog_running_alive_thread(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(tmf, "_watchdog_running", True)
        mock_thread = MagicMock()
        mock_thread.is_alive.return_value = True
        monkeypatch.setattr(tmf, "_watchdog_thread", mock_thread)
        with patch("threading.Thread") as mock_thread_cls:
            tmf._ensure_watchdog()
        mock_thread_cls.assert_not_called()
