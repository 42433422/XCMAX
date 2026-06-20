"""Extended tests for ``app.traditional_mode_fs`` covering low-coverage branches."""

from __future__ import annotations

import base64
import json
import os
import queue
import threading
from datetime import datetime
from typing import Any
from unittest.mock import MagicMock, patch

import pytest

from app import traditional_mode_fs as tmf


@pytest.fixture()
def temp_root(tmp_path, monkeypatch: pytest.MonkeyPatch):
    """Use a temp dir as ROOT_DIR for isolation."""
    root = tmp_path / "bang"
    root.mkdir()
    monkeypatch.setattr(tmf, "ROOT_DIR", str(root))
    return str(root)


class TestResolveSafePath:
    def test_resolve_safe_path_within_root(self, temp_root) -> None:
        result = tmf.resolve_safe_path("subdir/file.txt")
        assert result is not None
        assert result.startswith(temp_root)

    def test_resolve_safe_path_empty(self, temp_root) -> None:
        result = tmf.resolve_safe_path("")
        assert result == temp_root

    def test_resolve_safe_path_none_input(self, temp_root) -> None:
        result = tmf.resolve_safe_path(None)  # type: ignore[arg-type]
        assert result == temp_root

    def test_resolve_safe_path_traversal_rejected(self, temp_root) -> None:
        result = tmf.resolve_safe_path("../../../etc/passwd")
        assert result is None

    def test_resolve_safe_path_commonpath_value_error(self, temp_root, monkeypatch) -> None:
        # Force commonpath to raise ValueError
        monkeypatch.setattr(
            os.path, "commonpath", lambda paths: (_ for _ in ()).throw(ValueError("diff drives"))
        )
        result = tmf.resolve_safe_path("file.txt")
        assert result is None


class TestRootInfoResponse:
    def test_root_info_response_shape(self, temp_root) -> None:
        result = tmf.root_info_response()
        assert result["success"] is True
        assert result["data"]["root"] == temp_root
        assert result["data"]["logical_root"] == "bang"
        assert "list" in result["data"]["capabilities"]
        assert "delete" in result["data"]["capabilities"]


class TestStatResponse:
    def test_stat_root(self, temp_root) -> None:
        result, status = tmf.stat_response("")
        assert status == 200
        assert result["success"] is True
        assert result["data"]["name"] == "bang"
        assert result["data"]["is_dir"] is True

    def test_stat_existing_file(self, temp_root) -> None:
        fpath = os.path.join(temp_root, "file.txt")
        with open(fpath, "w") as f:
            f.write("hello")
        result, status = tmf.stat_response("file.txt")
        assert status == 200
        assert result["data"]["name"] == "file.txt"
        assert result["data"]["is_dir"] is False
        assert result["data"]["size"] == 5

    def test_stat_path_traversal(self, temp_root) -> None:
        result, status = tmf.stat_response("../../../etc")
        assert status == 403
        assert result["success"] is False

    def test_stat_not_found(self, temp_root) -> None:
        result, status = tmf.stat_response("nonexistent.txt")
        assert status == 404
        assert result["success"] is False


class TestWriteTextResponse:
    def test_write_text_new_file(self, temp_root) -> None:
        result, status = tmf.write_text_response("sub/file.txt", "hello")
        assert status == 200
        assert result["success"] is True
        assert result["data"]["bytes"] == 5
        with open(os.path.join(temp_root, "sub", "file.txt")) as f:
            assert f.read() == "hello"

    def test_write_text_append(self, temp_root) -> None:
        tmf.write_text_response("file.txt", "hello")
        result, status = tmf.write_text_response("file.txt", " world", append=True)
        assert status == 200
        with open(os.path.join(temp_root, "file.txt")) as f:
            assert f.read() == "hello world"

    def test_write_text_to_dir_fails(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "adir"))
        result, status = tmf.write_text_response("adir", "content")
        assert status == 400
        assert result["success"] is False

    def test_write_text_traversal_rejected(self, temp_root) -> None:
        result, status = tmf.write_text_response("../../../etc/foo", "x")
        assert status == 403


class TestWriteBase64Response:
    def test_write_base64_success(self, temp_root) -> None:
        content = base64.b64encode(b"hello").decode()
        result, status = tmf.write_base64_response("file.bin", content)
        assert status == 200
        assert result["data"]["bytes"] == 5

    def test_write_base64_invalid(self, temp_root) -> None:
        result, status = tmf.write_base64_response("file.bin", "not valid base64!!!")
        assert status == 400
        assert result["success"] is False

    def test_write_base64_to_dir_fails(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "adir"))
        content = base64.b64encode(b"hello").decode()
        result, status = tmf.write_base64_response("adir", content)
        assert status == 400

    def test_write_base64_traversal_rejected(self, temp_root) -> None:
        content = base64.b64encode(b"hello").decode()
        result, status = tmf.write_base64_response("../../../etc/foo", content)
        assert status == 403


class TestMoveResponse:
    def test_move_success(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("data")
        result, status = tmf.move_response("src.txt", "dst.txt")
        assert status == 200
        assert not os.path.exists(os.path.join(temp_root, "src.txt"))
        assert os.path.exists(os.path.join(temp_root, "dst.txt"))

    def test_move_src_not_found(self, temp_root) -> None:
        result, status = tmf.move_response("nope.txt", "dst.txt")
        assert status == 404

    def test_move_dst_exists_no_overwrite(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("src")
        with open(os.path.join(temp_root, "dst.txt"), "w") as f:
            f.write("dst")
        result, status = tmf.move_response("src.txt", "dst.txt")
        assert status == 409

    def test_move_dst_exists_with_overwrite_file(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("src")
        with open(os.path.join(temp_root, "dst.txt"), "w") as f:
            f.write("dst")
        result, status = tmf.move_response("src.txt", "dst.txt", overwrite=True)
        assert status == 200
        with open(os.path.join(temp_root, "dst.txt")) as f:
            assert f.read() == "src"

    def test_move_dst_exists_with_overwrite_dir(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("src")
        os.makedirs(os.path.join(temp_root, "dstdir"))
        result, status = tmf.move_response("src.txt", "dstdir", overwrite=True)
        assert status == 200

    def test_move_traversal_rejected(self, temp_root) -> None:
        result, status = tmf.move_response("../../../etc", "dst.txt")
        assert status == 403


class TestCopyResponse:
    def test_copy_file_success(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("data")
        result, status = tmf.copy_response("src.txt", "dst.txt")
        assert status == 200
        assert os.path.exists(os.path.join(temp_root, "src.txt"))
        assert os.path.exists(os.path.join(temp_root, "dst.txt"))

    def test_copy_dir_success(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "srcdir"))
        with open(os.path.join(temp_root, "srcdir", "file.txt"), "w") as f:
            f.write("data")
        result, status = tmf.copy_response("srcdir", "dstdir")
        assert status == 200
        assert os.path.exists(os.path.join(temp_root, "dstdir", "file.txt"))

    def test_copy_src_not_found(self, temp_root) -> None:
        result, status = tmf.copy_response("nope.txt", "dst.txt")
        assert status == 404

    def test_copy_dst_exists_no_overwrite(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("src")
        with open(os.path.join(temp_root, "dst.txt"), "w") as f:
            f.write("dst")
        result, status = tmf.copy_response("src.txt", "dst.txt")
        assert status == 409

    def test_copy_dst_exists_with_overwrite_file(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("src")
        with open(os.path.join(temp_root, "dst.txt"), "w") as f:
            f.write("dst")
        result, status = tmf.copy_response("src.txt", "dst.txt", overwrite=True)
        assert status == 200

    def test_copy_dst_exists_with_overwrite_dir(self, temp_root) -> None:
        with open(os.path.join(temp_root, "src.txt"), "w") as f:
            f.write("src")
        os.makedirs(os.path.join(temp_root, "dstdir"))
        result, status = tmf.copy_response("src.txt", "dstdir", overwrite=True)
        assert status == 200

    def test_copy_traversal_rejected(self, temp_root) -> None:
        result, status = tmf.copy_response("../../../etc", "dst.txt")
        assert status == 403


class TestGetFileType:
    def test_excel_extension(self) -> None:
        assert tmf._get_file_type("file.xlsx") == "xlsx"
        assert tmf._get_file_type("file.xls") == "xls"

    def test_image_extension(self) -> None:
        assert tmf._get_file_type("file.png") == "png"
        assert tmf._get_file_type("file.jpg") == "jpg"

    def test_other_extension(self) -> None:
        assert tmf._get_file_type("file.txt") == "txt"

    def test_no_extension(self) -> None:
        assert tmf._get_file_type("file") == "文件"


class TestListFilesResponse:
    def test_list_empty_dir(self, temp_root) -> None:
        result, status = tmf.list_files_response("")
        assert status == 200
        assert result["data"]["files"] == []

    def test_list_with_files_and_dirs(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "subdir"))
        with open(os.path.join(temp_root, "file.txt"), "w") as f:
            f.write("data")
        result, status = tmf.list_files_response("")
        assert status == 200
        names = [e["name"] for e in result["data"]["files"]]
        assert "subdir" in names
        assert "file.txt" in names
        # dirs first
        assert result["data"]["files"][0]["name"] == "subdir"

    def test_list_not_found_returns_empty(self, temp_root) -> None:
        result, status = tmf.list_files_response("nonexistent")
        assert status == 200
        assert result["data"]["files"] == []

    def test_list_path_is_file(self, temp_root) -> None:
        with open(os.path.join(temp_root, "file.txt"), "w") as f:
            f.write("data")
        result, status = tmf.list_files_response("file.txt")
        assert status == 400

    def test_list_traversal_rejected(self, temp_root) -> None:
        result, status = tmf.list_files_response("../../../etc")
        assert status == 403

    def test_list_recoverable_error(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(
            tmf, "resolve_safe_path", lambda p: (_ for _ in ()).throw(RuntimeError("boom"))
        )
        result, status = tmf.list_files_response("")
        assert status == 500

    def test_list_oserror_on_entry_skipped(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "subdir"))
        with open(os.path.join(temp_root, "file.txt"), "w") as f:
            f.write("data")
        # Force os.stat to fail on one entry
        original_stat = os.stat

        def stat_with_fail(path):
            if "subdir" in str(path):
                raise OSError("stat fail")
            return original_stat(path)

        with patch("os.stat", side_effect=stat_with_fail):
            result, status = tmf.list_files_response("")
        assert status == 200
        names = [e["name"] for e in result["data"]["files"]]
        assert "subdir" not in names
        assert "file.txt" in names


class TestReadFileResponse:
    def test_read_text_file(self, temp_root) -> None:
        with open(os.path.join(temp_root, "file.txt"), "w") as f:
            f.write("hello")
        result, status = tmf.read_file_response("file.txt")
        assert status == 200
        assert result["data"]["type"] == "text"
        assert result["data"]["content"] == "hello"

    def test_read_not_found(self, temp_root) -> None:
        result, status = tmf.read_file_response("nope.txt")
        assert status == 404

    def test_read_directory(self, temp_root) -> None:
        os.makedirs(os.path.join(temp_root, "adir"))
        result, status = tmf.read_file_response("adir")
        assert status == 400

    def test_read_traversal_rejected(self, temp_root) -> None:
        result, status = tmf.read_file_response("../../../etc/passwd")
        assert status == 403

    def test_read_image(self, temp_root) -> None:
        with open(os.path.join(temp_root, "img.png"), "wb") as f:
            f.write(b"fake png")
        result, status = tmf.read_file_response("img.png")
        assert status == 200
        assert result["data"]["type"] == "image"
        assert result["data"]["mime"] == "image/png"
        assert base64.b64decode(result["data"]["content"]) == b"fake png"

    def test_read_binary_fallback(self, temp_root) -> None:
        # Write non-utf8 bytes
        with open(os.path.join(temp_root, "file.bin"), "wb") as f:
            f.write(b"\xff\xfe\x00")
        result, status = tmf.read_file_response("file.bin")
        assert status == 200
        assert result["data"]["type"] == "binary"
        assert base64.b64decode(result["data"]["content"]) == b"\xff\xfe\x00"

    def test_read_excel(self, temp_root) -> None:
        try:
            import openpyxl
        except ImportError:
            pytest.skip("openpyxl not installed")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "hello"
        ws["A2"] = datetime(2026, 1, 1)
        wb.save(os.path.join(temp_root, "file.xlsx"))
        wb.close()
        result, status = tmf.read_file_response("file.xlsx")
        assert status == 200
        assert result["data"]["type"] == "excel"
        assert "Sheet" in result["data"]["sheets"]

    def test_read_excel_import_error(self, temp_root, monkeypatch) -> None:
        # Create a fake xlsx file
        with open(os.path.join(temp_root, "file.xlsx"), "wb") as f:
            f.write(b"fake")
        import builtins

        real_import = builtins.__import__

        def fake_import(name, *args, **kwargs):
            if name == "openpyxl":
                raise ImportError("no openpyxl")
            return real_import(name, *args, **kwargs)

        monkeypatch.setattr(builtins, "__import__", fake_import)
        result, status = tmf.read_file_response("file.xlsx")
        assert status == 500
        assert "openpyxl" in result["error"]

    def test_read_excel_recoverable_error(self, temp_root, monkeypatch) -> None:
        with open(os.path.join(temp_root, "file.xlsx"), "wb") as f:
            f.write(b"fake")
        import openpyxl

        def fake_load(*args, **kwargs):
            raise RuntimeError("load fail")

        monkeypatch.setattr(openpyxl, "load_workbook", fake_load)
        result, status = tmf.read_file_response("file.xlsx")
        assert status == 500

    def test_read_recoverable_error(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(
            tmf, "resolve_safe_path", lambda p: (_ for _ in ()).throw(RuntimeError("boom"))
        )
        result, status = tmf.read_file_response("file.txt")
        assert status == 500


class TestBuildSnapshot:
    def test_build_snapshot_empty_root(self, temp_root) -> None:
        snap = tmf._build_snapshot()
        assert snap == {}

    def test_build_snapshot_with_files(self, temp_root) -> None:
        with open(os.path.join(temp_root, "file.txt"), "w") as f:
            f.write("data")
        os.makedirs(os.path.join(temp_root, "sub"))
        with open(os.path.join(temp_root, "sub", "nested.txt"), "w") as f:
            f.write("nested")
        snap = tmf._build_snapshot()
        assert "file.txt" in snap
        assert "sub/nested.txt" in snap

    def test_build_snapshot_nonexistent_root(self, tmp_path, monkeypatch) -> None:
        monkeypatch.setattr(tmf, "ROOT_DIR", str(tmp_path / "nonexistent"))
        snap = tmf._build_snapshot()
        assert snap == {}

    def test_build_snapshot_oserror_walk(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(os, "walk", lambda *a, **k: (_ for _ in ()).throw(OSError("walk fail")))
        snap = tmf._build_snapshot()
        assert snap == {}


class TestFormatSnapshot:
    def test_format_snapshot(self) -> None:
        result = tmf._format_snapshot({"file.txt": 1700000000.0})
        assert "file.txt" in result
        assert isinstance(result["file.txt"], str)


class TestFormatTime:
    def test_format_time(self) -> None:
        result = tmf._format_time(1700000000.0)
        assert isinstance(result, str)
        assert "T" in result


class TestSseWatchEvents:
    def test_sse_traversal_rejected(self, temp_root) -> None:
        gen = tmf.sse_watch_events("../../../etc")
        first = next(gen)
        assert "路径越权访问被拒绝" in first
        # Generator should stop after the rejection
        with pytest.raises(StopIteration):
            next(gen)

    def test_sse_initial_snapshot(self, temp_root) -> None:
        # Pre-populate last_snapshot
        with tmf._snapshot_lock:
            tmf._last_snapshot = {"file.txt": "2024-01-01T00:00:00+00:00"}
        gen = tmf.sse_watch_events("")
        first = next(gen)
        assert "snapshot" in first
        # Clean up - close the generator
        gen.close()


class TestEnsureWatchdog:
    def test_ensure_watchdog_starts_thread(self, temp_root, monkeypatch) -> None:
        # Reset state
        monkeypatch.setattr(tmf, "_watchdog_running", False)
        monkeypatch.setattr(tmf, "_watchdog_thread", None)
        started = {"called": False}

        class FakeThread:
            def __init__(self, *a, **k):
                started["called"] = True

            def start(self):
                pass

            def is_alive(self):
                return True

        monkeypatch.setattr(threading, "Thread", FakeThread)
        tmf._ensure_watchdog()
        assert started["called"] is True

    def test_ensure_watchdog_already_running(self, temp_root, monkeypatch) -> None:
        mock_thread = MagicMock()
        mock_thread.is_alive.return_value = True
        monkeypatch.setattr(tmf, "_watchdog_running", True)
        monkeypatch.setattr(tmf, "_watchdog_thread", mock_thread)
        # Should not start a new thread
        with patch("threading.Thread") as mock_thread_cls:
            tmf._ensure_watchdog()
        mock_thread_cls.assert_not_called()

    def test_ensure_watchdog_dead_thread_restarts(self, temp_root, monkeypatch) -> None:
        mock_thread = MagicMock()
        mock_thread.is_alive.return_value = False
        monkeypatch.setattr(tmf, "_watchdog_running", True)
        monkeypatch.setattr(tmf, "_watchdog_thread", mock_thread)
        with patch("threading.Thread") as mock_thread_cls:
            tmf._ensure_watchdog()
        mock_thread_cls.assert_called_once()


class TestWatchdogLoop:
    def test_watchdog_loop_recoverable_error(self, temp_root, monkeypatch) -> None:
        monkeypatch.setattr(tmf, "_watchdog_running", True)
        # Force _build_snapshot to raise
        monkeypatch.setattr(
            tmf, "_build_snapshot", lambda: (_ for _ in ()).throw(RuntimeError("snap fail"))
        )
        # Should not raise
        tmf._watchdog_loop()
        assert tmf._watchdog_running is False
