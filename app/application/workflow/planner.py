from __future__ import annotations

import json
import logging
import uuid
from collections.abc import Callable
from typing import Any

import httpx

from app.services import get_ai_conversation_service
from app.utils.path_utils import ensure_fhd_repo_on_syspath

from .types import PlanGraph, WorkflowNode, validate_plan_graph

logger = logging.getLogger(__name__)

# 同步规划 LLM 复用 Client，减轻短时多次 DeepSeek 连接失败
_planner_http_client: httpx.Client | None = None


def get_tool_registry() -> dict[str, Any]:
    """
    返回工作流工具注册表，供 ai_chat_app_service 使用。
    覆盖报价、主数据、出货、模板与微信辅助等能力，与意图层 tool_key 对齐。
    """
    return {
        "price_list": {
            "description": "生成客户价格表/报价单 Word 文档",
            "availability": "shared",
            "actions": {
                "export": {
                    "description": "导出客户价格表为 Word 文档",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": ["customer_name"],
                },
            },
        },
        "products": {
            "description": "产品查询与管理",
            "availability": "shared",
            "actions": {
                "query": {
                    "description": "按关键词/型号/客户查询产品",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": [],
                },
            },
        },
        "customers": {
            "description": "客户查询与管理",
            "availability": "shared",
            "actions": {
                "query": {
                    "description": "按关键词查询客户",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": [],
                },
                "ensure_exists": {
                    "description": "按单位名称创建客户（不存在时）",
                    "risk": "medium",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": ["unit_name"],
                },
            },
        },
        "shipment_generate": {
            "description": "解析自然语言订单并生成发货单，写入 shipment_records",
            "availability": "shared",
            "actions": {
                "generate": {
                    "description": "生成发货单（优先 order_text；否则 unit_name+products）",
                    "risk": "medium",
                    "idempotent": False,
                    "availability": "shared",
                    "required_params": [],
                },
            },
        },
        "shipment_records": {
            "description": "出货记录 shipment_records 查询",
            "availability": "shared",
            "actions": {
                "query": {
                    "description": "按客户/关键词筛选出货记录",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": [],
                },
            },
        },
        "shipments": {
            "description": "与 shipment_records 相同的出货列表查询（兼容旧 tool_key）",
            "availability": "shared",
            "actions": {
                "query": {
                    "description": "列出最近出货记录",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": [],
                },
            },
        },
        "materials": {
            "description": "原材料档案与库存查询",
            "availability": "shared",
            "actions": {
                "query": {
                    "description": "按关键词/分类查询原材料",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": [],
                },
            },
        },
        "print_label": {
            "description": "根据产品行生成标签图（商标导出目录）",
            "availability": "shared",
            "actions": {
                "generate": {
                    "description": "生成标签图片列表",
                    "risk": "medium",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": ["products"],
                },
            },
        },
        "excel_decompose": {
            "description": "从 Excel 模板文件分解字段网格",
            "availability": "shared",
            "actions": {
                "decompose": {
                    "description": "分解指定路径的 Excel 模板",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": ["file_path"],
                },
            },
        },
        "template_extract": {
            "description": "提取/分解模板结构（与 excel_decompose 共用实现）",
            "availability": "shared",
            "actions": {
                "extract": {
                    "description": "从 file_path 提取模板结构",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": ["file_path"],
                },
            },
        },
        "wechat_send": {
            "description": "微信联系人检索（发送动作由客户端完成）",
            "availability": "shared",
            "actions": {
                "preview": {
                    "description": "按关键词列出可发送的微信联系人",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": [],
                },
            },
        },
        "excel_schema": {
            "description": "分析 Excel 文件的表结构（列名、数据类型、行数等）",
            "availability": "shared",
            "actions": {
                "analyze": {
                    "description": "分析指定 Excel 文件的表结构",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": ["file_path"],
                },
            },
        },
        "excel_analysis": {
            "description": "对 Excel 文件进行数据读取、查询和聚合分析",
            "availability": "shared",
            "actions": {
                "analyze": {
                    "description": "读取/查询/聚合 Excel 数据",
                    "risk": "low",
                    "idempotent": True,
                    "availability": "shared",
                    "required_params": ["file_path"],
                },
            },
        },
        "import_excel": {
            "description": "将 Excel 数据导入数据库（产品库/客户库）",
            "availability": "shared",
            "actions": {
                "import": {
                    "description": "将 Excel 表格写入产品/客户库",
                    "risk": "medium",
                    "idempotent": False,
                    "availability": "shared",
                    "required_params": ["file_path"],
                },
            },
        },
    }


def execute_tool(tool_name: str, params: dict[str, Any]) -> dict[str, Any]:
    """
    执行指定工具（支持 execute_registered_workflow_tool 注入的 _action）。

    与 get_tool_registry 中的工具 id 一致。
    """
    logger.info("execute_tool called: tool_name=%s, params=%s", tool_name, params)

    merged = dict(params or {})
    merged.pop("_runtime_context", None)
    action = str(merged.pop("_action", "") or "").strip().lower()
    if not action:
        action_defaults: dict[str, str] = {
            "price_list": "export",
            "products": "query",
            "customers": "query",
            "shipment_generate": "generate",
            "shipment_records": "query",
            "shipments": "query",
            "materials": "query",
            "print_label": "generate",
            "excel_decompose": "decompose",
            "template_extract": "extract",
            "wechat_send": "preview",
            "excel_schema": "analyze",
            "excel_analysis": "analyze",
            "import_excel": "import",
        }
        action = action_defaults.get(tool_name, "query")

    handler = _WORKFLOW_TOOL_HANDLERS.get((tool_name, action))
    if handler is not None:
        return handler(merged)

    return {
        "success": False,
        "message": f"未知工具或动作: {tool_name}.{action}",
        "error_code": "unknown_tool_action",
    }


def _execute_price_list_tool(params: dict[str, Any]) -> dict[str, Any]:
    """执行价格表导出工具"""
    try:
        customer_name = params.get("customer_name") or params.get("unit")
        keyword = params.get("keyword")
        date = params.get("date")

        if not customer_name:
            return {
                "success": False,
                "message": "缺少 customer_name 参数",
                "error_code": "missing_customer_name",
            }

        fhd_root = ensure_fhd_repo_on_syspath()

        from app.application.tools import handle_price_list_export

        result = handle_price_list_export(
            {"customer_name": customer_name, "keyword": keyword, "export_date": date},
            workspace_root=str(fhd_root) if fhd_root else None,
        )
        return result
    except ImportError as e:
        logger.error("价格表导出服务导入失败: %s", e)
        return {
            "success": False,
            "message": "价格表导出服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("价格表导出参数错误: %s", e)
        return {
            "success": False,
            "message": "参数错误：请检查客户名称和价格参数",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        logger.error("价格表导出文件操作失败: %s", e)
        return {
            "success": False,
            "message": "文件导出失败，请检查磁盘空间",
            "error_code": "file_io_error",
        }
    except RuntimeError as e:
        logger.error("价格表导出运行时错误: %s", e)
        return {
            "success": False,
            "message": "导出处理失败，请稍后重试",
            "error_code": "export_failed",
        }


def _execute_products_tool(params: dict[str, Any]) -> dict[str, Any]:
    """执行产品查询工具"""
    try:
        from app.bootstrap import get_products_service

        keyword = str(params.get("keyword") or "").strip()
        unit_name = str(params.get("unit_name") or params.get("unit") or "").strip() or None
        model_number = (
            str(params.get("model_number") or params.get("product_code") or "").strip() or None
        )
        page = int(params.get("page", 1))
        per_page = int(params.get("per_page", 20))

        svc = get_products_service()
        if model_number and unit_name:
            result = svc.get_products(
                unit_name=unit_name,
                model_number=model_number,
                keyword=None,
                page=page,
                per_page=per_page,
            )
        elif model_number:
            result = svc.get_products(
                unit_name=None,
                model_number=model_number,
                keyword=None,
                page=page,
                per_page=per_page,
            )
        elif unit_name:
            result = svc.get_products(
                unit_name=unit_name,
                model_number=None,
                keyword=keyword or None,
                page=page,
                per_page=per_page,
            )
        else:
            result = svc.get_products(
                unit_name=None,
                model_number=None,
                keyword=keyword or None,
                page=page,
                per_page=per_page,
            )
        return result
    except ImportError as e:
        logger.error("产品服务导入失败: %s", e)
        return {"success": False, "message": "产品服务不可用", "error_code": "service_unavailable"}
    except (ValueError, TypeError) as e:
        logger.warning("产品查询参数错误: %s", e)
        return {
            "success": False,
            "message": "查询参数错误，请检查输入",
            "error_code": "invalid_parameters",
        }
    except RuntimeError as e:
        logger.error("产品查询运行时错误: %s", e)
        return {"success": False, "message": "查询失败，请稍后重试", "error_code": "query_failed"}


def _execute_customers_tool(params: dict[str, Any]) -> dict[str, Any]:
    """执行客户查询工具"""
    try:
        from app.bootstrap import get_customer_app_service

        keyword = params.get("keyword") or params.get("customer_name") or ""
        page = int(params.get("page", 1))
        per_page = int(params.get("per_page", 20))

        svc = get_customer_app_service()
        return svc.get_all(
            keyword=str(keyword).strip() or None,
            page=page,
            per_page=per_page,
        )
    except ImportError as e:
        logger.error("客户服务导入失败: %s", e)
        return {"success": False, "message": "客户服务不可用", "error_code": "service_unavailable"}
    except (ValueError, TypeError) as e:
        logger.warning("客户查询参数错误: %s", e)
        return {
            "success": False,
            "message": "查询参数错误，请检查输入",
            "error_code": "invalid_parameters",
        }
    except RuntimeError as e:
        logger.error("客户查询运行时错误: %s", e)
        return {"success": False, "message": "查询失败，请稍后重试", "error_code": "query_failed"}


def _execute_customers_ensure_exists_tool(params: dict[str, Any]) -> dict[str, Any]:
    """创建客户（单位）如不存在。"""
    try:
        from app.bootstrap import get_customer_app_service

        unit = str(params.get("unit_name") or params.get("customer_name") or "").strip()
        if not unit:
            return {
                "success": False,
                "message": "缺少 unit_name",
                "error_code": "missing_unit_name",
            }

        svc = get_customer_app_service()
        matched = svc.match_purchase_unit(unit)
        if matched:
            return {
                "success": True,
                "created": False,
                "message": f"单位已存在：{unit}",
                "data": {
                    "id": getattr(matched, "id", None),
                    "customer_name": getattr(matched, "unit_name", None) or unit,
                    "unit_name": getattr(matched, "unit_name", None) or unit,
                },
            }
        created = svc.create({"customer_name": unit})
        out = dict(created) if isinstance(created, dict) else {"success": False}
        out["created"] = bool(out.get("success"))
        return out
    except ImportError as e:
        logger.error("客户创建服务导入失败: %s", e)
        return {
            "success": False,
            "message": "客户创建服务不可用",
            "error_code": "service_unavailable",
            "created": False,
        }
    except (ValueError, TypeError) as e:
        logger.warning("客户创建参数错误: %s", e)
        return {
            "success": False,
            "message": "创建参数错误，请检查单位名称",
            "error_code": "invalid_parameters",
            "created": False,
        }
    except RuntimeError as e:
        logger.error("客户创建运行时错误: %s", e)
        return {
            "success": False,
            "message": "创建失败，请稍后重试",
            "error_code": "create_failed",
            "created": False,
        }


def _execute_shipment_generate_tool(params: dict[str, Any]) -> dict[str, Any]:
    try:
        from app.bootstrap import get_shipment_app_service
        from app.routes.tools import _parse_order_text

        order_text = str(params.get("order_text") or "").strip()
        unit_name = str(params.get("unit_name") or "").strip()
        products = params.get("products")

        if order_text:
            parsed = _parse_order_text(order_text)
        elif unit_name and isinstance(products, list) and products:
            parsed = {"success": True, "unit_name": unit_name, "products": products}
        else:
            return {
                "success": False,
                "message": "缺少 order_text，或 unit_name+products",
                "error_code": "missing_order_params",
            }

        if not parsed.get("success"):
            return {
                "success": False,
                "message": parsed.get("message") or parsed.get("error") or "订单解析失败",
            }

        svc = get_shipment_app_service()
        return svc.generate_shipment_document(
            unit_name=str(parsed.get("unit_name") or ""),
            products=list(parsed.get("products") or []),
            template_name=params.get("template_name"),
            date=params.get("date"),
            order_number=params.get("order_number"),
            raw_text=order_text or str(params.get("raw_text") or ""),
        )
    except ImportError as e:
        logger.error("发货单服务导入失败: %s", e)
        return {
            "success": False,
            "message": "发货单服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("发货单生成参数错误: %s", e)
        return {
            "success": False,
            "message": "订单参数错误，请检查输入",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        logger.error("发货单文件生成失败: %s", e)
        return {
            "success": False,
            "message": "文档生成失败，请检查磁盘空间",
            "error_code": "file_io_error",
        }
    except RuntimeError as e:
        logger.error("发货单生成运行时错误: %s", e)
        return {
            "success": False,
            "message": "生成失败，请稍后重试",
            "error_code": "generation_failed",
        }


def _execute_shipment_records_tool(params: dict[str, Any]) -> dict[str, Any]:
    try:
        from app.bootstrap import get_shipment_app_service

        unit = params.get("unit_name") or params.get("keyword") or params.get("customer_name")
        limit = int(params.get("limit", 50))
        svc = get_shipment_app_service()
        rows = svc.get_shipment_records(unit_name=str(unit).strip() if unit else None, limit=limit)
        return {"success": True, "data": rows, "message": f"共 {len(rows)} 条出货记录"}
    except ImportError as e:
        logger.error("出货记录服务导入失败: %s", e)
        return {
            "success": False,
            "message": "出货记录服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("出货记录查询参数错误: %s", e)
        return {
            "success": False,
            "message": "查询参数错误，请检查单位名称",
            "error_code": "invalid_parameters",
        }
    except RuntimeError as e:
        logger.error("出货记录查询运行时错误: %s", e)
        return {"success": False, "message": "查询失败，请稍后重试", "error_code": "query_failed"}


def _execute_materials_tool(params: dict[str, Any]) -> dict[str, Any]:
    try:
        from app.bootstrap import get_materials_service

        search = str(params.get("keyword") or params.get("search") or "").strip() or None
        category = str(params.get("category") or "").strip() or None
        page = int(params.get("page", 1))
        per_page = int(params.get("per_page", 20))
        return get_materials_service().get_all_materials(
            search=search,
            category=category,
            page=page,
            per_page=per_page,
        )
    except ImportError as e:
        logger.error("原材料服务导入失败: %s", e)
        return {
            "success": False,
            "message": "原材料服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("原材料查询参数错误: %s", e)
        return {
            "success": False,
            "message": "查询参数错误，请检查输入",
            "error_code": "invalid_parameters",
        }
    except RuntimeError as e:
        logger.error("原材料查询运行时错误: %s", e)
        return {"success": False, "message": "查询失败，请稍后重试", "error_code": "query_failed"}


def _execute_print_label_tool(params: dict[str, Any]) -> dict[str, Any]:
    try:
        import os

        from app.infrastructure.documents.shipment_document_generator_impl import (
            SimpleLabelGenerator,
        )
        from app.utils.path_utils import get_resource_path

        products = params.get("products")
        if not isinstance(products, list) or not products:
            return {
                "success": False,
                "message": "缺少 products 数组",
                "error_code": "missing_products",
            }

        labels_dir = get_resource_path("ai_assistant", "商标导出")
        os.makedirs(labels_dir, exist_ok=True)
        order_number = str(params.get("order_number") or params.get("doc_name") or "LABEL").strip()
        gen = SimpleLabelGenerator(labels_dir)
        labels = gen.generate_labels_for_order(order_number=order_number, products=products)
        return {"success": True, "data": labels, "message": f"已生成 {len(labels)} 张标签"}
    except ImportError as e:
        logger.error("标签生成服务导入失败: %s", e)
        return {
            "success": False,
            "message": "标签生成服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("标签生成参数错误: %s", e)
        return {
            "success": False,
            "message": "标签参数错误，请检查产品数据",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        logger.error("标签文件生成失败: %s", e)
        return {
            "success": False,
            "message": "标签导出失败，请检查磁盘空间",
            "error_code": "file_io_error",
        }
    except RuntimeError as e:
        logger.error("标签生成运行时错误: %s", e)
        return {
            "success": False,
            "message": "生成失败，请稍后重试",
            "error_code": "generation_failed",
        }


def _execute_excel_decompose_tool(params: dict[str, Any]) -> dict[str, Any]:
    try:
        from app.bootstrap import get_template_app_service

        file_path = str(params.get("file_path") or "").strip()
        if not file_path:
            return {
                "success": False,
                "message": "缺少 file_path",
                "error_code": "missing_file_path",
            }
        template_type = params.get("template_type") or params.get("scope")
        return get_template_app_service().decompose_template(
            file_path, str(template_type).strip() if template_type else None
        )
    except ImportError as e:
        logger.error("模板服务导入失败: %s", e)
        return {"success": False, "message": "模板服务不可用", "error_code": "service_unavailable"}
    except (ValueError, TypeError) as e:
        logger.warning("模板分解参数错误: %s", e)
        return {
            "success": False,
            "message": "模板参数错误，请检查文件",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        logger.error("模板文件读取失败: %s", e)
        return {
            "success": False,
            "message": "文件读取失败，请检查文件是否存在",
            "error_code": "file_not_found",
        }
    except RuntimeError as e:
        logger.error("模板分解运行时错误: %s", e)
        return {
            "success": False,
            "message": "分解失败，请稍后重试",
            "error_code": "decomposition_failed",
        }


def _execute_template_extract_tool(params: dict[str, Any]) -> dict[str, Any]:
    """与 excel_decompose 共用模板分解能力。"""
    return _execute_excel_decompose_tool(params)


def _execute_wechat_preview_tool(params: dict[str, Any]) -> dict[str, Any]:
    try:
        from app.bootstrap import get_wechat_contact_app_service

        keyword = str(params.get("keyword") or params.get("unit_name") or "").strip() or None
        limit = int(params.get("limit", 30))
        contacts = get_wechat_contact_app_service().get_contacts(keyword=keyword, limit=limit)
        return {
            "success": True,
            "data": contacts,
            "message": "请在客户端选择联系人完成发送" if contacts else "未找到匹配的微信联系人",
        }
    except ImportError as e:
        logger.error("微信联系人服务导入失败: %s", e)
        return {
            "success": False,
            "message": "微信联系人服务不可用",
            "error_code": "service_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("微信联系人查询参数错误: %s", e)
        return {
            "success": False,
            "message": "查询参数错误，请检查关键词",
            "error_code": "invalid_parameters",
        }
    except RuntimeError as e:
        logger.error("微信联系人查询运行时错误: %s", e)
        return {"success": False, "message": "查询失败，请稍后重试", "error_code": "query_failed"}


def _execute_excel_schema_tool(params: dict[str, Any]) -> dict[str, Any]:
    """分析 Excel 文件的表结构。"""
    file_path = str(params.get("file_path") or "").strip()
    if not file_path:
        return {
            "success": False,
            "message": "缺少 file_path 参数",
            "error_code": "missing_file_path",
        }

    try:
        from app.bootstrap import get_excel_analysis_app_service

        service = get_excel_analysis_app_service()
        return service.analyze_schema(
            file_path=file_path,
            sheet_name=params.get("sheet_name"),
        )
    except ImportError:
        pass
    except Exception as e:
        logger.warning("excel_analysis_app_service 不可用，降级 openpyxl: %s", e)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get("sheet_name") or wb.sheetnames[0]
        ws = wb[sheet_name]

        fields = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value is not None:
                fields.append(
                    {
                        "name": str(cell.column_letter),
                        "label": str(cell.value).strip(),
                        "column_index": cell.column,
                    }
                )

        row_count = ws.max_row or 0
        wb.close()

        return {
            "success": True,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "fields": fields,
            "row_count": max(0, row_count - 1),
            "message": f"Excel 结构分析完成：{len(fields)} 列，{max(0, row_count - 1)} 行数据",
        }
    except ImportError as e:
        logger.error("Excel 分析库导入失败: %s", e)
        return {
            "success": False,
            "message": "Excel 处理库不可用",
            "error_code": "library_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("Excel 结构参数错误: %s", e)
        return {
            "success": False,
            "message": "文件参数错误，请检查 Excel 文件",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        logger.error("Excel 文件读取失败: %s", e)
        return {
            "success": False,
            "message": "文件读取失败，请检查文件是否存在",
            "error_code": "file_not_found",
        }
    except RuntimeError as e:
        logger.error("Excel 结构分析运行时错误: %s", e)
        return {
            "success": False,
            "message": "分析失败，请稍后重试",
            "error_code": "analysis_failed",
        }


def _execute_excel_analysis_tool(params: dict[str, Any]) -> dict[str, Any]:
    """读取/查询/聚合 Excel 数据。"""
    file_path = str(params.get("file_path") or "").strip()
    if not file_path:
        return {
            "success": False,
            "message": "缺少 file_path 参数",
            "error_code": "missing_file_path",
        }

    try:
        from app.bootstrap import get_excel_analysis_app_service

        service = get_excel_analysis_app_service()
        return service.analyze_data(
            file_path=file_path,
            sheet_name=params.get("sheet_name"),
            query=params.get("query"),
            columns=params.get("columns"),
        )
    except ImportError:
        pass
    except Exception as e:
        logger.warning("excel_analysis_app_service 不可用，降级 openpyxl: %s", e)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get("sheet_name") or wb.sheetnames[0]
        ws = wb[sheet_name]

        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else "")

        target_columns = params.get("columns")
        col_indices = list(range(len(headers)))
        if target_columns:
            col_indices = [i for i, h in enumerate(headers) if h in target_columns]

        rows = []
        for row in ws.iter_rows(min_row=2, max_row=min(ws.max_row or 2, 52)):
            row_data = {}
            for i in col_indices:
                if i < len(row):
                    row_data[headers[i]] = row[i].value
            if any(v is not None for v in row_data.values()):
                rows.append(row_data)

        wb.close()

        return {
            "success": True,
            "file_path": file_path,
            "sheet_name": sheet_name,
            "headers": headers,
            "rows": rows[:50],
            "total_rows": len(rows),
            "message": f"Excel 数据读取完成：{len(headers)} 列，{len(rows)} 行",
        }
    except ImportError as e:
        logger.error("Excel 分析库导入失败: %s", e)
        return {
            "success": False,
            "message": "Excel 处理库不可用",
            "error_code": "library_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("Excel 数据参数错误: %s", e)
        return {
            "success": False,
            "message": "文件参数错误，请检查 Excel 文件",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        logger.error("Excel 文件读取失败: %s", e)
        return {
            "success": False,
            "message": "文件读取失败，请检查文件是否存在",
            "error_code": "file_not_found",
        }
    except RuntimeError as e:
        logger.error("Excel 数据分析运行时错误: %s", e)
        return {
            "success": False,
            "message": "分析失败，请稍后重试",
            "error_code": "analysis_failed",
        }


def _execute_import_excel_tool(params: dict[str, Any]) -> dict[str, Any]:
    """将 Excel 数据导入数据库。"""
    file_path = str(params.get("file_path") or "").strip()
    if not file_path:
        return {
            "success": False,
            "message": "缺少 file_path 参数",
            "error_code": "missing_file_path",
        }

    unit_name = str(params.get("unit_name") or "").strip()
    price_column = str(params.get("price_column") or "").strip()
    create_customer = params.get("create_customer_if_missing", True)
    skip_duplicates = params.get("skip_duplicates", True)

    try:
        from app.bootstrap import get_products_service

        products_service = get_products_service()
    except ImportError as e:
        logger.error("产品服务导入失败: %s", e)
        return {"success": False, "message": "产品服务不可用", "error_code": "service_unavailable"}
    except RuntimeError as e:
        logger.error("产品服务初始化失败: %s", e)
        return {
            "success": False,
            "message": "产品服务初始化失败",
            "error_code": "service_init_failed",
        }

    customer_service = None
    try:
        from app.bootstrap import get_customer_app_service

        customer_service = get_customer_app_service()
    except ImportError:
        logger.warning("客户服务不可用，降级为仅产品入库")
    except RuntimeError as e:
        logger.warning("客户服务初始化失败，降级为仅产品入库: %s", e)

    try:
        import openpyxl

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        sheet_name = params.get("sheet_name") or wb.sheetnames[0]
        ws = wb[sheet_name]

        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value).strip() if cell.value else "")

        name_col = None
        model_col = None
        price_col = None
        unit_col = None

        for i, h in enumerate(headers):
            if not name_col and any(k in h for k in ("产品名称", "名称", "品名")):
                name_col = i
            if not model_col and any(k in h for k in ("编号", "型号", "产品编号", "规格型号")):
                model_col = i
            if not unit_col and any(k in h for k in ("单位", "客户", "购买单位")):
                unit_col = i

        resolved_price_col_name = ""
        if not price_column:
            try:
                from app.application.ai_chat_app_service import AIChatApplicationService

                merged_intent = AIChatApplicationService._merge_user_intent_for_price_resolution(
                    str(params.get("_user_message") or ""),
                    params.get("_request_context"),
                )
                overrides = params.get("excel_import_column_overrides")
                resolved_price_col_name, price_err = (
                    AIChatApplicationService._resolve_unit_price_column(
                        keys=headers,
                        current="",
                        user_message=merged_intent,
                        overrides=overrides if isinstance(overrides, dict) else {},
                    )
                )
                if price_err == "ambiguous_price_columns":
                    wb.close()
                    return {
                        "success": False,
                        "message": "检测到「调价前」和「调价后」两列价格，请明确指定使用哪一列（如传入 price_column='调价前含税单价'）",
                        "error_code": "ambiguous_price_columns",
                    }
                if resolved_price_col_name:
                    price_column = resolved_price_col_name
                    logger.info("智能价格列消歧: 选中列 '%s'", price_column)
            except ImportError:
                logger.debug("AI 服务不可用，回退简单匹配")
            except (ValueError, TypeError) as e:
                logger.debug("智能价格列消歧参数错误，回退简单匹配: %s", e)
            except RuntimeError as e:
                logger.warning("智能价格列消歧运行时错误，回退简单匹配: %s", e)

        for i, h in enumerate(headers):
            if not price_col:
                if (
                    price_column
                    and price_column in h
                    or not price_column
                    and any(k in h for k in ("单价", "价格", "价"))
                ):
                    price_col = i

        if price_column and price_col is None:
            for i, h in enumerate(headers):
                if price_column in h:
                    price_col = i
                    break

        created_units = 0
        created_products = 0
        skipped_products = 0
        touched_units: set = set()

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            row_values = [c.value for c in row]
            product_name = (
                str(row_values[name_col] or "").strip()
                if name_col is not None and name_col < len(row_values)
                else ""
            )
            model_number = (
                str(row_values[model_col] or "").strip().upper()
                if model_col is not None and model_col < len(row_values)
                else ""
            )
            unit_price = 0.0
            if price_col is not None and price_col < len(row_values):
                try:
                    unit_price = float(row_values[price_col] or 0)
                except (ValueError, TypeError):
                    unit_price = 0.0
            row_unit = (
                str(row_values[unit_col] or "").strip()
                if unit_col is not None and unit_col < len(row_values)
                else ""
            )

            effective_unit = unit_name or row_unit
            if not effective_unit and not product_name and not model_number:
                continue

            touched_units.add(effective_unit)

            if effective_unit and customer_service is not None and create_customer:
                matched = customer_service.match_purchase_unit(effective_unit)
                if not matched:
                    create_result = customer_service.create({"customer_name": effective_unit})
                    if create_result.get("success"):
                        created_units += 1

            if (product_name or model_number) and products_service is not None:
                exists_result = products_service.get_products(
                    unit_name=effective_unit or None,
                    model_number=model_number or None,
                    keyword=(product_name or model_number or None),
                    page=1,
                    per_page=5,
                )
                existed = False
                if exists_result.get("success"):
                    for item in exists_result.get("data") or []:
                        item_model = str(item.get("model_number") or "").strip().upper()
                        item_name = str(item.get("name") or item.get("product_name") or "").strip()
                        if model_number and item_model == model_number:
                            existed = True
                            break
                        if product_name and item_name == product_name:
                            existed = True
                            break

                if existed and skip_duplicates:
                    skipped_products += 1
                    continue

                create_product = products_service.create_product(
                    {
                        "name": product_name or model_number,
                        "product_name": product_name or model_number,
                        "product_code": model_number or None,
                        "model_number": model_number or None,
                        "unit_price": unit_price,
                        "price": unit_price,
                        "unit": effective_unit,
                    }
                )
                if create_product.get("success"):
                    created_products += 1

        wb.close()

        return {
            "success": True,
            "records": len(touched_units) + created_products + skipped_products,
            "touched_units": len(touched_units),
            "created_units": created_units,
            "created_products": created_products,
            "skipped_products": skipped_products,
            "price_column_used": headers[price_col] if price_col is not None else "未指定",
            "message": f"导入完成：新增客户 {created_units}，新增产品 {created_products}，跳过重复 {skipped_products}",
        }
    except ImportError as e:
        logger.error("Excel 处理库导入失败: %s", e)
        return {
            "success": False,
            "message": "Excel 处理库不可用",
            "error_code": "library_unavailable",
        }
    except (ValueError, TypeError) as e:
        logger.warning("Excel 导入参数错误: %s", e)
        return {
            "success": False,
            "message": "导入参数错误，请检查文件格式",
            "error_code": "invalid_parameters",
        }
    except OSError as e:
        logger.error("Excel 文件读取失败: %s", e)
        return {
            "success": False,
            "message": "文件读取失败，请检查文件是否存在",
            "error_code": "file_not_found",
        }
    except RuntimeError as e:
        logger.error("Excel 导入运行时错误: %s", e)
        return {
            "success": False,
            "message": "导入失败，请检查数据格式后重试",
            "error_code": "import_failed",
        }


# 与 get_tool_registry / execute_tool 默认 action 对齐；(tool_id, action) -> 实现函数
_WORKFLOW_TOOL_HANDLERS: dict[tuple[str, str], Callable[[dict[str, Any]], dict[str, Any]]] = {
    ("price_list", "export"): _execute_price_list_tool,
    ("products", "query"): _execute_products_tool,
    ("customers", "query"): _execute_customers_tool,
    ("customers", "ensure_exists"): _execute_customers_ensure_exists_tool,
    ("shipment_generate", "generate"): _execute_shipment_generate_tool,
    ("shipment_records", "query"): _execute_shipment_records_tool,
    ("shipments", "query"): _execute_shipment_records_tool,
    ("materials", "query"): _execute_materials_tool,
    ("print_label", "generate"): _execute_print_label_tool,
    ("excel_decompose", "decompose"): _execute_excel_decompose_tool,
    ("template_extract", "extract"): _execute_template_extract_tool,
    ("wechat_send", "preview"): _execute_wechat_preview_tool,
    ("excel_schema", "analyze"): _execute_excel_schema_tool,
    ("excel_analysis", "analyze"): _execute_excel_analysis_tool,
    ("import_excel", "import"): _execute_import_excel_tool,
}


def _get_planner_http_client() -> httpx.Client:
    global _planner_http_client
    if _planner_http_client is None:
        _planner_http_client = httpx.Client(
            timeout=httpx.Timeout(20.0, connect=10.0),
            limits=httpx.Limits(max_keepalive_connections=10, max_connections=20),
        )
    return _planner_http_client


def _filter_tool_registry_for_profile(
    tool_registry: dict[str, Any],
    profile: str,
) -> dict[str, Any]:
    """
    - normal：剔除 pro_only 工具与动作（普通界面走槽位/共享工具）。
    - pro_default：剔除 normal_only 工具与动作（全专业链路不暴露纯普通槽位工具）。
    """
    filtered: dict[str, Any] = {}
    for tool_id, spec in tool_registry.items():
        if not isinstance(spec, dict):
            continue
        tool_av = str(spec.get("availability") or "shared").strip().lower()
        if profile == "normal" and tool_av == "pro_only":
            continue
        if profile == "pro_default" and tool_av == "normal_only":
            continue
        actions = spec.get("actions") or {}
        if not isinstance(actions, dict):
            continue
        kept_actions: dict[str, Any] = {}
        for aname, ameta in actions.items():
            if not isinstance(ameta, dict):
                continue
            av = str(ameta.get("availability") or "shared").strip().lower()
            if profile == "normal" and av == "pro_only":
                continue
            if profile == "pro_default" and av == "normal_only":
                continue
            kept_actions[aname] = ameta
        if not kept_actions:
            continue
        new_spec = dict(spec)
        new_spec["actions"] = kept_actions
        filtered[tool_id] = new_spec
    return filtered


class LLMWorkflowPlanner:
    def __init__(self) -> None:
        self._ai_service = get_ai_conversation_service()

    def plan(
        self,
        user_id: str,
        message: str,
        tool_registry: dict[str, Any],
        context: dict[str, Any] | None = None,
    ) -> PlanGraph:
        context = dict(context or {})
        plan_id = uuid.uuid4().hex

        from app.application.normal_chat_dispatch import resolve_tool_execution_profile

        profile = resolve_tool_execution_profile(context)
        registry_for_plan = _filter_tool_registry_for_profile(tool_registry, profile)

        # 专业模式 ReAct/CoT 前：拉取用户记忆 RAG 命中摘要，注入到 context（只放摘要文本）。
        # 目的：让规划器在“选择工具/补全关键 params”时更贴近用户习惯，但避免把全量历史注入 prompt。
        try:
            from app.application import get_user_memory_rag_app_service

            rag = get_user_memory_rag_app_service()
            rag_res = rag.query(user_id=user_id, query_text=message, top_k=3)
            hits = (rag_res or {}).get("hits") if isinstance(rag_res, dict) else None
            if isinstance(hits, list) and hits:
                summary = rag.format_for_prompt(
                    user_id=user_id, query_text=message, hits=hits, max_hits=4
                )
                context["user_memory_rag"] = {"summary": summary}
        except ImportError:
            logger.debug("用户记忆 RAG 服务不可用（不阻断主流程）")
        except (ValueError, TypeError) as e:
            logger.debug("用户记忆 RAG 参数错误（不阻断主流程）: %s", e)
        except RuntimeError as e:
            logger.warning("用户记忆 RAG 运行时错误（不阻断主流程）: %s", e)

        planned = self._plan_with_react_multiagent(
            plan_id=plan_id,
            user_id=user_id,
            message=message,
            tool_registry=registry_for_plan,
            context=context,
        )
        if planned is not None:
            err = validate_plan_graph(planned)
            if err is None:
                return planned
            logger.warning("ReAct/CoT 计划校验失败，回退规则规划: %s", err)

        return self._fallback_plan(plan_id, message, registry_for_plan)

    def _plan_with_react_multiagent(
        self,
        plan_id: str,
        user_id: str,
        message: str,
        tool_registry: dict[str, Any],
        context: dict[str, Any],
    ) -> PlanGraph | None:
        """
        多步 ReAct/CoT 风格规划（简化实现）：
        1) 先用 LLM 生成候选 PlanGraph（DecomposerAgent）。
        2) 基于候选 PlanGraph 抽取低风险只读节点做 ToolProbe（真实工具调用）。
        3) 将探测结果注入 prompt 再次规划得到最终 PlanGraph（PlanComposerAgent）。
        4) validate_plan_graph；失败则降级 fallback（CriticAgent）。
        """

        # 0) 用现有 Planner 生成候选计划
        candidate = self._plan_with_llm(
            plan_id=plan_id,
            user_id=user_id,
            message=message,
            tool_registry=tool_registry,
            context=context,
        )
        if candidate is None:
            return None

        # 1) 抽取 probe：只探测 low-risk + idempotent 的节点
        runtime_context_for_probe = dict(context or {})
        runtime_context_for_probe["message"] = str(message or "")
        probe_requests: list[dict[str, Any]] = []
        for node in candidate.nodes or []:
            tid = str(node.tool_id or "").strip()
            act = str(node.action or "").strip()
            if not tid or not act:
                continue
            tool_spec = tool_registry.get(tid)
            if not isinstance(tool_spec, dict):
                continue
            actions = tool_spec.get("actions") or {}
            if not isinstance(actions, dict):
                continue
            meta = actions.get(act)
            if not isinstance(meta, dict):
                continue

            risk = str(meta.get("risk") or "").strip().lower()
            idempotent = bool(meta.get("idempotent", False))
            if risk != "low" or not idempotent:
                continue

            # 只对“查询类/列举类”探测，避免意义不明的 view/info 探测
            if act not in (
                "query",
                "exists",
                "list",
                "view",
                "preview",
                "decompose",
                "extract",
                "refresh_contact_cache",
                "refresh_messages_cache",
            ):
                continue

            probe_requests.append(
                {
                    "tool_id": tid,
                    "action": act,
                    "params": node.params or {},
                }
            )

        # 最多 3 个 probe，避免 prompt 过长/探测过多
        probe_requests = probe_requests[:3]

        # 2) 执行 ToolProbe（并注入检索词：对 products/customers.query 补 keyword）
        probe_outputs: list[dict[str, Any]] = []

        task_agent = None
        try:
            from app.services.task_agent import TaskAgent

            task_agent = TaskAgent()
        except ImportError:
            logger.debug("TaskAgent 服务不可用")
            task_agent = None
        except RuntimeError as e:
            logger.warning("TaskAgent 初始化失败: %s", e)
            task_agent = None

        for pr in probe_requests:
            try:
                tool_id = str(pr.get("tool_id") or "").strip()
                action = str(pr.get("action") or "").strip()
                params = pr.get("params") if isinstance(pr.get("params"), dict) else {}

                # 安全约束校验：仅允许 low-risk & idempotent 的工具探测，并严格校验 required_params 非空
                tool_spec = tool_registry.get(tool_id) or {}
                actions = tool_spec.get("actions") or {}
                action_meta = actions.get(action) if isinstance(actions, dict) else None
                if not isinstance(action_meta, dict):
                    continue
                risk = str(action_meta.get("risk") or "").strip().lower()
                idempotent = bool(action_meta.get("idempotent", False))
                if risk != "low" or not idempotent:
                    continue

                required_params = action_meta.get("required_params") or []
                if not isinstance(required_params, list):
                    required_params = []
                missing_required = []
                for k in required_params:
                    if (
                        k not in (params or {})
                        or params.get(k) is None
                        or str(params.get(k)).strip() == ""
                    ):
                        missing_required.append(k)
                if missing_required:
                    continue

                # query 补检索词（避免探测空 keyword 导致无意义全量扫描）
                if tool_id == "products" and action == "query":
                    # node.params 优先；否则从 message 中尽量抽取
                    if not (
                        params.get("keyword")
                        or params.get("model_number")
                        or params.get("unit_name")
                    ):
                        try:
                            from app.application.normal_chat_dispatch import (
                                route_normal_mode_message,
                            )

                            rr = route_normal_mode_message(message)
                            if rr.get("intent") == "product_query":
                                slots = rr.get("slots") or {}
                                params.update(
                                    {
                                        "keyword": slots.get("keyword")
                                        or params.get("keyword")
                                        or "",
                                        "model_number": slots.get("model_number")
                                        or params.get("model_number")
                                        or "",
                                        "unit_name": slots.get("unit_name")
                                        or params.get("unit_name")
                                        or "",
                                    }
                                )
                        except (ImportError, RuntimeError):
                            # fallback: 使用原消息作为 keyword
                            if not params.get("keyword"):
                                params["keyword"] = str(message or "").strip()[:80]

                if tool_id == "customers" and action == "query":
                    if (
                        not (params.get("keyword") or params.get("customer_name"))
                        and task_agent is not None
                    ):
                        try:
                            cust_slots = task_agent._extract_customer_query_slots(
                                str(message or "")
                            )
                            if isinstance(cust_slots, dict):
                                params.update(cust_slots)
                        except (ImportError, RuntimeError):
                            if not params.get("keyword"):
                                params["keyword"] = str(message or "").strip()[:80]

                from app.routes.tools import execute_registered_workflow_tool

                merged_params = dict(params or {})
                merged_params["_runtime_context"] = dict(runtime_context_for_probe)

                out = execute_registered_workflow_tool(
                    tool_id=tool_id, action=action, params=merged_params
                )

                # 只保留概要信息，避免 prompt 爆长
                data_preview = ""
                if isinstance(out, dict):
                    if isinstance(out.get("data"), list):
                        data_preview = str(out.get("data")[:3])[:600]
                    elif out.get("data") is not None:
                        data_preview = str(out.get("data"))[:600]
                    elif out.get("raw") is not None:
                        data_preview = str(out.get("raw"))[:600]
                    else:
                        data_preview = str(out)[:600]

                if isinstance(out, dict) and out.get("success") is True:
                    probe_outputs.append(
                        {
                            "tool_id": tool_id,
                            "action": action,
                            "success": True,
                            "message": str(
                                (out or {}).get("message") or (out or {}).get("error") or ""
                            ),
                            "data_preview": data_preview,
                        }
                    )
            except (ValueError, TypeError) as e:
                logger.debug("ToolProbe 参数错误（将跳过注入）: %s", e)
                continue
            except RuntimeError as e:
                logger.warning("ToolProbe 运行时错误（将跳过注入）: %s", e)
                continue

        # 3) 把探测结果塞回 context，再规划最终计划
        context_for_compose = dict(context or {})
        if probe_outputs:
            context_for_compose["tool_probe_outputs"] = probe_outputs

        final_plan = self._plan_with_llm(
            plan_id=plan_id,
            user_id=user_id,
            message=message,
            tool_registry=tool_registry,
            context=context_for_compose,
        )
        if final_plan is None:
            return None

        # 4) CriticAgent：validate_plan_graph + required_params 检查；失败则尝试修复一次
        err = validate_plan_graph(final_plan)
        if err is None:
            err = self._validate_required_params(final_plan, tool_registry)

        if err is None:
            return final_plan

        logger.warning("CriticAgent 校验失败，尝试 LLM 修复（最多 1 次）: %s", err)
        repaired = self._critic_repair_with_llm(
            plan_id=plan_id,
            user_id=user_id,
            message=message,
            tool_registry=tool_registry,
            context=context_for_compose,
            error=err,
            invalid_plan=final_plan,
        )
        if repaired is not None:
            err2 = validate_plan_graph(repaired)
            if err2 is None:
                err2 = self._validate_required_params(repaired, tool_registry)
            if err2 is None:
                return repaired

        logger.warning("CriticAgent 修复失败（回退 fallback）: %s", err)
        return None

    @staticmethod
    def _validate_required_params(plan: PlanGraph, tool_registry: dict[str, Any]) -> str | None:
        """检查节点 params 是否满足 tool_registry 的 required_params。"""
        for node in plan.nodes or []:
            tool_spec = (tool_registry or {}).get(str(node.tool_id) or "")
            if not isinstance(tool_spec, dict):
                continue
            actions = tool_spec.get("actions") or {}
            if not isinstance(actions, dict):
                continue
            action_meta = actions.get(str(node.action) or "")
            if not isinstance(action_meta, dict):
                continue
            required_params = action_meta.get("required_params") or []
            if not isinstance(required_params, list):
                required_params = []
            params = node.params or {}
            for key in required_params:
                if (
                    key not in params
                    or params.get(key) is None
                    or str(params.get(key)).strip() == ""
                ):
                    return f"节点 {node.node_id} 缺少 required_params: {key}"
        return None

    def _critic_repair_with_llm(
        self,
        plan_id: str,
        user_id: str,
        message: str,
        tool_registry: dict[str, Any],
        context: dict[str, Any],
        error: str,
        invalid_plan: PlanGraph,
    ) -> PlanGraph | None:
        """CriticAgent：LLM 修复无效 PlanGraph（只重试一次）。"""
        api_key = getattr(self._ai_service, "api_key", "") or ""
        if not api_key:
            return None

        try:
            tool_specs = []
            for tool_id, spec in tool_registry.items():
                actions = spec.get("actions") or {}
                action_specs = []
                for action_name, action_meta in actions.items():
                    if not isinstance(action_meta, dict):
                        continue
                    action_specs.append(
                        {
                            "action": action_name,
                            "risk": action_meta.get("risk", "low"),
                            "idempotent": bool(action_meta.get("idempotent", False)),
                            "required_params": action_meta.get("required_params", []),
                        }
                    )
                tool_specs.append(
                    {
                        "tool_id": tool_id,
                        "description": spec.get("description", ""),
                        "actions": action_specs,
                    }
                )

            invalid_dict = {
                "plan_id": invalid_plan.plan_id,
                "intent": invalid_plan.intent,
                "todo_steps": invalid_plan.todo_steps,
                "risk_level": invalid_plan.risk_level,
                "nodes": [
                    {
                        "node_id": n.node_id,
                        "tool_id": n.tool_id,
                        "action": n.action,
                        "params": n.params,
                        "risk": n.risk,
                        "idempotent": n.idempotent,
                        "description": n.description,
                        "depends_on": n.depends_on,
                    }
                    for n in (invalid_plan.nodes or [])
                ],
            }

            prompt = {
                "task": "修复一个无效的工作流 PlanGraph JSON，使其满足 validate_plan_graph 规则且满足 required_params 约束。",
                "rules": [
                    "只输出 JSON，不要 markdown。",
                    "node_id 必须唯一且非空。",
                    "所有 nodes 项必须包含 tool_id/action/params/risk/idempotent/description/depends_on 结构字段。",
                    "对于 required_params：必须在 params 中提供非空值（若无法从 user_message 推断，仍需给出最合理的非空占位/默认值，保证结构字段不缺失）。",
                ],
                "validation_error": error,
                "invalid_plan": invalid_dict,
                "user_message": message,
                "context": context,
                "tool_registry": tool_specs,
                "output_schema": {
                    "intent": "string",
                    "todo_steps": ["string"],
                    "risk_level": "low|medium|high",
                    "nodes": [
                        {
                            "node_id": "string",
                            "tool_id": "string",
                            "action": "string",
                            "params": {},
                            "risk": "low|medium|high",
                            "idempotent": "bool",
                            "description": "string",
                            "depends_on": ["node_id"],
                        }
                    ],
                },
            }

            messages = [
                {"role": "system", "content": "你是工作流计划修复器，只输出可执行 JSON。"},
                {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
            ]

            api_url = (
                getattr(self._ai_service, "api_url", "")
                or "https://api.deepseek.com/v1/chat/completions"
            )
            response = _get_planner_http_client().post(
                api_url,
                headers={"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"},
                json={
                    "model": getattr(self._ai_service, "model", "") or "deepseek-chat",
                    "messages": messages,
                    "temperature": 0.2,
                    "max_tokens": 1000,
                },
            )
            if response.status_code >= 400:
                return None

            response_data = response.json()
            raw = (
                (response_data.get("choices") or [{}])[0]
                .get("message", {})
                .get("content", "")
                .strip()
            )
            if not raw:
                return None

            raw = raw.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            raw = self._strip_json_code_fence(raw)
            if not raw:
                return None
            parsed = json.loads(raw)

            nodes: list[WorkflowNode] = []
            for idx, node in enumerate(parsed.get("nodes", []), start=1):
                nodes.append(
                    WorkflowNode(
                        node_id=str(node.get("node_id") or f"node_{idx}"),
                        tool_id=str(node.get("tool_id") or ""),
                        action=str(node.get("action") or ""),
                        params=node.get("params") or {},
                        risk=str(node.get("risk") or "low"),
                        idempotent=bool(node.get("idempotent", False)),
                        description=str(node.get("description") or ""),
                        depends_on=[str(x) for x in (node.get("depends_on") or [])],
                    )
                )

            return PlanGraph(
                plan_id=plan_id,
                intent=str(parsed.get("intent") or invalid_plan.intent or "dynamic_workflow"),
                todo_steps=[
                    str(x) for x in (parsed.get("todo_steps") or invalid_plan.todo_steps or [])
                ],
                nodes=nodes,
                risk_level=str(parsed.get("risk_level") or invalid_plan.risk_level or "low"),
                metadata={"planner": "critic_repair", "message": message},
            )
        except (ValueError, TypeError) as e:
            logger.debug("CriticAgent 修复参数错误: %s", e)
            return None
        except RuntimeError as e:
            logger.warning("CriticAgent 修复运行时错误: %s", e)
            return None

    def _plan_with_llm(
        self,
        plan_id: str,
        user_id: str,
        message: str,
        tool_registry: dict[str, Any],
        context: dict[str, Any],
    ) -> PlanGraph | None:
        try:
            tool_specs = []
            for tool_id, spec in tool_registry.items():
                actions = spec.get("actions", {})
                action_specs = []
                for action_name, action_meta in actions.items():
                    action_specs.append(
                        {
                            "action": action_name,
                            "risk": action_meta.get("risk", "low"),
                            "idempotent": bool(action_meta.get("idempotent", False)),
                            "required_params": action_meta.get("required_params", []),
                        }
                    )
                tool_specs.append(
                    {
                        "tool_id": tool_id,
                        "description": spec.get("description", ""),
                        "actions": action_specs,
                    }
                )

            recent_messages = []
            conv_ctx = self._ai_service.get_context(user_id)
            if conv_ctx and conv_ctx.conversation_history:
                recent_messages = conv_ctx.conversation_history[-6:]

            prompt = {
                "task": "根据用户意图生成可执行工作流计划（JSON）。",
                "rules": [
                    "只输出 JSON，不要 markdown。",
                    "优先使用 tool_registry 中已有工具与 action。",
                    "如果步骤有依赖，写到 depends_on。",
                    "todo_steps 要贴合用户语义，不要模板化。",
                    "risk_level 按节点最高风险确定。",
                    "对 products.query / customers.query：必须在 params 填入 keyword 或 model_number 等检索词，"
                    "从用户话中提取（如「七彩乐园的9803」→ keyword 含单位+型号），禁止留空对象 {}。",
                    "如果 context 中包含 tool_probe_outputs 且其中 success=true，请优先使用其中 data_preview 的信息来补全 nodes.params。",
                    "若 context 中 tool_execution_profile 为 normal 或 ui_surface 为 normal 且 intent_channel 为 pro："
                    "仅可使用 availability 为 shared 或 normal_only 的工具；产品查询优先 normal_slot_dispatch.product_query 或 products.query。",
                    "若 context 为全专业链路（未带上述混合标记）：仅使用 shared 或 pro_only，勿选 normal_only。",
                ],
                "user_message": message,
                "recent_messages": recent_messages,
                "context": context,
                "tool_probe_outputs": (
                    (context or {}).get("tool_probe_outputs") if isinstance(context, dict) else []
                ),
                "tool_registry": tool_specs,
                "output_schema": {
                    "intent": "string",
                    "todo_steps": ["string"],
                    "risk_level": "low|medium|high",
                    "nodes": [
                        {
                            "node_id": "string",
                            "tool_id": "string",
                            "action": "string",
                            "params": {},
                            "risk": "low|medium|high",
                            "idempotent": "bool",
                            "description": "string",
                            "depends_on": ["node_id"],
                        }
                    ],
                },
            }

            messages = [
                {"role": "system", "content": "你是工作流规划器，只输出可执行 JSON。"},
                {"role": "user", "content": json.dumps(prompt, ensure_ascii=False)},
            ]
            api_key = getattr(self._ai_service, "api_key", "") or ""
            api_url = (
                getattr(self._ai_service, "api_url", "")
                or "https://api.deepseek.com/v1/chat/completions"
            )
            model = getattr(self._ai_service, "model", "") or "deepseek-chat"
            if not api_key:
                return None

            response = _get_planner_http_client().post(
                api_url,
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": model,
                    "messages": messages,
                    "temperature": 0.1,
                    "max_tokens": 1200,
                },
            )
            if response.status_code >= 400:
                return None
            response_data = response.json()
            raw = (
                (response_data.get("choices") or [{}])[0]
                .get("message", {})
                .get("content", "")
                .strip()
            )
            if not raw:
                return None
            raw = raw.removeprefix("```json").removeprefix("```").removesuffix("```").strip()
            parsed = json.loads(raw)

            nodes: list[WorkflowNode] = []
            for idx, node in enumerate(parsed.get("nodes", []), start=1):
                nodes.append(
                    WorkflowNode(
                        node_id=str(node.get("node_id") or f"node_{idx}"),
                        tool_id=str(node.get("tool_id") or ""),
                        action=str(node.get("action") or ""),
                        params=node.get("params") or {},
                        risk=str(node.get("risk") or "low"),
                        idempotent=bool(node.get("idempotent", False)),
                        description=str(node.get("description") or ""),
                        depends_on=[str(x) for x in (node.get("depends_on") or [])],
                    )
                )

            tool_probe_outputs = []
            user_memory_rag_summary = ""
            try:
                if isinstance(context, dict):
                    user_memory_rag = context.get("user_memory_rag")
                    if isinstance(user_memory_rag, dict):
                        user_memory_rag_summary = str(user_memory_rag.get("summary") or "").strip()
                    tpo = context.get("tool_probe_outputs")
                    if isinstance(tpo, list):
                        tool_probe_outputs = []
                        for item in tpo[:2]:
                            if not isinstance(item, dict):
                                continue
                            tool_probe_outputs.append(
                                {
                                    "tool_id": item.get("tool_id"),
                                    "action": item.get("action"),
                                    "success": bool(item.get("success")),
                                    "message": str(item.get("message") or "").strip()[:120],
                                    "data_preview": str(item.get("data_preview") or "").strip()[
                                        :160
                                    ],
                                }
                            )
            except (ImportError, RuntimeError):
                tool_probe_outputs = []
                user_memory_rag_summary = ""

            return PlanGraph(
                plan_id=plan_id,
                intent=str(parsed.get("intent") or "dynamic_workflow"),
                todo_steps=[str(x) for x in (parsed.get("todo_steps") or [])],
                nodes=nodes,
                risk_level=str(parsed.get("risk_level") or "low"),
                metadata={
                    "planner": "llm",
                    "message": message,
                    "user_memory_rag_summary": user_memory_rag_summary,
                    "tool_probe_outputs": tool_probe_outputs,
                },
            )
        except (ValueError, TypeError) as err:
            logger.debug("LLM 规划参数错误，回退规则规划: %s", err)
            return None
        except RuntimeError as err:
            logger.warning("LLM 规划运行时错误，回退规则规划: %s", err)
            return None

    def _fallback_plan(
        self,
        plan_id: str,
        message: str,
        tool_registry: dict[str, Any],
    ) -> PlanGraph:
        lower = (message or "").lower()
        nodes: list[WorkflowNode] = []
        todo = ["理解用户目标", "执行可用工具", "输出执行结果"]
        intent = "generic_workflow"

        if ("添加" in message or "新增" in message or "create" in lower) and ("产品" in message):
            intent = "add_product_to_unit"
            todo = [
                "意图分析：识别产品新增任务",
                "全局检查单位是否存在",
                "单位不存在则先创建",
                "新增产品并绑定单位",
                "返回执行明细",
            ]
            if "customers" in tool_registry:
                nodes.append(
                    WorkflowNode(
                        node_id="check_or_create_unit",
                        tool_id="customers",
                        action="ensure_exists",
                        params={},
                        risk="medium",
                        description="确保客户存在",
                    )
                )
            if "products" in tool_registry:
                nodes.append(
                    WorkflowNode(
                        node_id="create_product",
                        tool_id="products",
                        action="create",
                        params={},
                        risk="medium",
                        description="创建产品",
                        depends_on=["check_or_create_unit"] if nodes else [],
                    )
                )

        if not nodes:
            # 兜底：挑一个低风险查询，避免空图
            if "products" in tool_registry:
                nodes.append(
                    WorkflowNode(
                        node_id="query_products",
                        tool_id="products",
                        action="query",
                        params={"keyword": message},
                        risk="low",
                        description="查询相关产品",
                        idempotent=True,
                    )
                )
            elif "customers" in tool_registry:
                nodes.append(
                    WorkflowNode(
                        node_id="query_customers",
                        tool_id="customers",
                        action="query",
                        params={"keyword": message},
                        risk="low",
                        description="查询相关客户",
                        idempotent=True,
                    )
                )

        risk = "low"
        if any(node.risk == "high" for node in nodes):
            risk = "high"
        elif any(node.risk == "medium" for node in nodes):
            risk = "medium"

        return PlanGraph(
            plan_id=plan_id,
            intent=intent,
            todo_steps=todo,
            nodes=nodes,
            risk_level=risk,
            metadata={"planner": "fallback", "message": message},
        )
