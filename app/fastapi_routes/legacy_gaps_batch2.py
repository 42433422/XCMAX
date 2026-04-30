"""
Flask → FastAPI 迁移缺口补齐 batch2 (原 ``archive_gap_batch2``)。

与 ``scripts/output/only_in_flask_native_batch2.json`` 对齐的原生 FastAPI 实现。

Phase 2C 起 History fallback 已搬到 :mod:`app.fastapi_routes.spa_fallback`
(``register_spa_history_fallback``),必须在应用所有路由注册之后调用。

**后续计划**: 按业务域继续拆分到 ``app/fastapi_routes/<domain>.py``
(Phase 2 遗留工作,见 ``docs/reports/LEGACY_CLEANUP_TRACKING.md``)。
"""

from __future__ import annotations

import base64
import json
import logging
import os
import shutil
import sys
import urllib.error
import urllib.request
import uuid
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Body, File, Form, Header, Query, Request, UploadFile
from fastapi.responses import FileResponse, JSONResponse, Response
from starlette.testclient import TestClient

from app.utils.secure_filename import secure_filename

from app.traditional_mode_fs import ROOT_DIR, resolve_safe_path
from app.utils.path_utils import get_base_dir, get_upload_dir
from app.fastapi_routes.ai_qclaw import _QCLOW_RUNTIME_STATE
from app.fastapi_routes.legacy_gaps_batch1 import (
    _mp_json_response,
    _mp_jwt_user_id,
    _require_login_user,
    _require_permission,
    _session_id_from_request,
)

logger = logging.getLogger(__name__)

router = APIRouter(tags=["legacy-gaps-batch2"])


def _emit_legacy_gaps_load_log() -> None:
    """启动阶段打一次日志,方便观测 legacy_gaps_batch2 当前承载的路由数。"""
    try:
        count = len([r for r in router.routes if hasattr(r, "path")])
    except Exception:
        count = -1
    logger.info(
        "[legacy-cleanup] legacy_gaps_batch2 loaded: routes=%s (further domain split pending; see LEGACY_CLEANUP_TRACKING.md)",
        count,
    )

def _mp_wechat_json_response(code: int, message: str, data: Any = None, *, success: bool = True) -> JSONResponse:
    body: dict[str, Any] = {"code": code, "message": message, "success": success}
    if data is not None:
        body["data"] = data
    return JSONResponse(body, status_code=code)


# --- wechat_contacts（缺口：GET /api/wechat_contacts/{id}、POST 变体）---
@router.get("/api/wechat_contacts/{contact_id}")
def wechat_contacts_get_by_id(contact_id: int):
    try:
        from app.application import get_wechat_contact_app_service

        service = get_wechat_contact_app_service()
        contact = service.get_contact_by_id(contact_id)
        if contact:
            return {"success": True, "data": contact}
        return JSONResponse({"success": False, "message": "联系人不存在"}, status_code=404)
    except Exception as e:
        return JSONResponse({"success": False, "message": f"查询失败：{str(e)}"}, status_code=500)


@router.post("/api/wechat_contacts/ensure_contact_cache")
def wechat_contacts_ensure_contact_cache_post():
    from app.services.wechat_contact_cache_import import refresh_wechat_contacts_from_decrypt

    payload, code = refresh_wechat_contacts_from_decrypt()
    return JSONResponse(payload, status_code=code)


@router.post("/api/wechat_contacts/send_message")
def wechat_contacts_send_message(body: dict = Body(default_factory=dict)):
    try:
        contact_name = (body.get("contact_name") or "").strip()
        message = (body.get("message") or "").strip()
        if not contact_name:
            return JSONResponse({"success": False, "message": "联系人名称不能为空"}, status_code=400)
        if not message:
            return JSONResponse({"success": False, "message": "消息内容不能为空"}, status_code=400)

        from app.utils.path_utils import get_resource_path

        sys_path = get_resource_path("wechat-decrypt")
        if sys_path not in sys.path:
            sys.path.insert(0, sys_path)

        from resources.wechat_cv.wechat_cv_send import search_and_send_by_cv

        result = search_and_send_by_cv(contact_name, message, delay=1.0, use_ocr=True)
        if result.get("status") == "success":
            return {"success": True, "message": f"已发送给 {contact_name}", "result": result}
        return JSONResponse(
            {
                "success": False,
                "message": f"发送失败: {result.get('message', '未知错误')}",
                "result": result,
            },
            status_code=500,
        )
    except Exception as e:
        logger.exception("wechat_contacts send_message: %s", e)
        return JSONResponse({"success": False, "message": f"发送失败：{str(e)}"}, status_code=500)


@router.post("/api/wechat_contacts/{contact_id}/send_message")
def wechat_contacts_send_message_to_id(contact_id: int, body: dict = Body(default_factory=dict)):
    try:
        message = (body.get("message") or "").strip()
        if not message:
            return JSONResponse({"success": False, "message": "消息内容不能为空"}, status_code=400)

        from app.application import get_wechat_contact_app_service

        service = get_wechat_contact_app_service()
        contact = service.get_contact_by_id(contact_id)
        if not contact:
            return JSONResponse({"success": False, "message": f"联系人不存在: {contact_id}"}, status_code=404)

        contact_name = (
            contact.get("contact_name") or contact.get("remark") or contact.get("wechat_id") or f"ID {contact_id}"
        )

        from app.utils.path_utils import get_resource_path

        sys_path = get_resource_path("wechat-decrypt")
        if sys_path not in sys.path:
            sys.path.insert(0, sys_path)

        from resources.wechat_cv.wechat_cv_send import search_and_send_by_cv

        result = search_and_send_by_cv(contact_name, message, delay=1.0, use_ocr=True)
        if result.get("status") == "success":
            return {"success": True, "message": f"已发送给 {contact_name}", "result": result}
        return JSONResponse(
            {
                "success": False,
                "message": f"发送失败: {result.get('message', '未知错误')}",
                "result": result,
            },
            status_code=500,
        )
    except Exception as e:
        logger.exception("wechat_contacts send to id: %s", e)
        return JSONResponse({"success": False, "message": f"发送失败：{str(e)}"}, status_code=500)


# --- 前端静态（Vue dist）---
def _vue_dist_dir() -> str:
    return os.path.join(get_base_dir(), "templates", "vue-dist")


@router.get("/assets/{path:path}")
def gap_batch2_serve_assets(path: str):
    vue_dist_dir = _vue_dist_dir()
    assets_dir = os.path.join(vue_dist_dir, "assets")
    asset_path = os.path.join(assets_dir, path)
    if os.path.exists(asset_path) and not os.path.isdir(asset_path):
        return FileResponse(asset_path)
    direct_path = os.path.join(vue_dist_dir, path)
    if os.path.exists(direct_path) and not os.path.isdir(direct_path):
        return FileResponse(direct_path)
    return JSONResponse({"success": False, "message": f"资源不存在：{path}"}, status_code=404)


@router.get("/static/{path:path}")
def gap_batch2_serve_static(path: str):
    vue_dist_dir = _vue_dist_dir()
    static_dir = os.path.join(vue_dist_dir, "static")
    static_path = os.path.join(static_dir, path)
    if os.path.exists(static_path) and not os.path.isdir(static_path):
        return FileResponse(static_path)
    return JSONResponse({"success": False, "message": f"静态资源不存在：{path}"}, status_code=404)


@router.get("/vite.svg")
def gap_batch2_vite_svg():
    p = os.path.join(_vue_dist_dir(), "vite.svg")
    if os.path.exists(p):
        return FileResponse(p, media_type="image/svg+xml")
    return JSONResponse({"success": False, "message": "vite.svg 不存在"}, status_code=404)


@router.get("/brand-xc-logo.jpg")
def gap_batch2_brand_xc_logo():
    p = os.path.join(_vue_dist_dir(), "brand-xc-logo.jpg")
    if os.path.exists(p):
        return FileResponse(p, media_type="image/jpeg")
    return JSONResponse({"success": False, "message": "brand-xc-logo.jpg 不存在"}, status_code=404)


@router.get("/workflow-employee-docs.json")
def gap_batch2_workflow_employee_docs_json():
    p = os.path.join(_vue_dist_dir(), "workflow-employee-docs.json")
    if os.path.exists(p):
        return FileResponse(p, media_type="application/json")
    return JSONResponse({"success": False, "message": "workflow-employee-docs.json 不存在"}, status_code=404)


@router.get("/favicon.ico")
def gap_batch2_favicon():
    gif = base64.b64decode("R0lGODlhAQABAIAAAAAAAP///yH5BAEAAAAALAAAAAABAAEAAAIBRAA7")
    return Response(content=gif, media_type="image/gif")


@router.get("/outputs/{filename:path}")
def gap_batch2_outputs(filename: str):
    try:
        from app.utils.path_utils import get_app_data_dir, get_resource_path

        shipment_outputs_dir = os.path.join(get_app_data_dir(), "shipment_outputs")
        if os.path.isdir(shipment_outputs_dir):
            outputs_dir = shipment_outputs_dir
        else:
            outputs_dir = get_resource_path("ai_assistant", "outputs")
            if not os.path.isdir(outputs_dir):
                outputs_dir = os.path.join(get_base_dir(), "AI助手", "outputs")
        if not os.path.isdir(outputs_dir):
            return JSONResponse({"success": False, "message": f"输出目录不存在: {outputs_dir}"}, status_code=404)
        file_path = os.path.join(outputs_dir, filename)
        if not os.path.exists(file_path):
            return JSONResponse({"success": False, "message": f"文件不存在：{filename}"}, status_code=404)
        return FileResponse(file_path, filename=os.path.basename(filename))
    except Exception as e:
        return JSONResponse({"success": False, "message": f"下载失败：{str(e)}"}, status_code=500)


@router.get("/test-buttons")
def gap_batch2_test_buttons():
    p = os.path.join(get_base_dir(), "templates", "test-buttons.html")
    if os.path.exists(p):
        return FileResponse(p, media_type="text/html")
    return JSONResponse({"success": False, "message": "test-buttons.html 未找到"}, status_code=404)


@router.get("/products-test")
def gap_batch2_products_test():
    p = os.path.join(get_base_dir(), "templates", "products_test.html")
    if os.path.exists(p):
        return FileResponse(p, media_type="text/html")
    return JSONResponse({"success": False, "message": "products_test.html 未找到"}, status_code=404)


@router.get("/console")
def gap_batch2_console():
    vue_index = os.path.join(_vue_dist_dir(), "index.html")
    if os.path.exists(vue_index):
        return FileResponse(vue_index, media_type="text/html")
    legacy = os.path.join(get_base_dir(), "templates", "ai_assistant_console.html")
    if os.path.exists(legacy):
        return FileResponse(legacy, media_type="text/html")
    return JSONResponse({"success": False, "message": "前端模板未找到"}, status_code=404)


# --- /api/ai/* batch2 ---
@router.post("/api/ai/message/save")
def ai_message_save(body: dict = Body(default_factory=dict)):
    from app.services.conversation_service import get_conversation_service as get_conversation_app_service

    service = get_conversation_app_service()
    if not body:
        return JSONResponse({"success": False, "message": "请求数据不能为空"}, status_code=400)
    session_id = body.get("session_id")
    user_raw = body.get("user_id", "default")
    role = body.get("role")
    content = body.get("content")
    intent = body.get("intent", "")
    metadata = body.get("metadata", "")
    if not session_id:
        return JSONResponse({"success": False, "message": "会话 ID 不能为空"}, status_code=400)
    if not role:
        return JSONResponse({"success": False, "message": "角色不能为空"}, status_code=400)
    if role in ("ai", "bot"):
        role = "assistant"
    if role not in ("user", "assistant", "system"):
        return JSONResponse({"success": False, "message": f"无效的角色：{role}"}, status_code=400)
    if not content:
        return JSONResponse({"success": False, "message": "消息内容不能为空"}, status_code=400)

    user_id_str = str(user_raw) if user_raw is not None else "default"
    try:
        message_id = service.save_message(session_id, user_id_str, role, content, intent, metadata)
        return {"success": True, "message_id": message_id}
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


@router.post("/api/ai/sqlite/import_unit_products")
def ai_sqlite_import_unit_products(body: dict = Body(default_factory=dict)):
    try:
        from app.application import get_unit_products_import_app_service

        service = get_unit_products_import_app_service()
        result = service.import_unit_products(
            saved_name=body.get("saved_name") or "",
            unit_name=(body.get("unit_name") or body.get("unit_name_guess") or "").strip(),
            create_purchase_unit=bool(body.get("create_purchase_unit", True)),
            skip_duplicates=bool(body.get("skip_duplicates", True)),
        )
        return JSONResponse(result, status_code=200 if result.get("success") else 400)
    except Exception as e:
        return JSONResponse({"success": False, "message": f"导入失败：{str(e)}"}, status_code=500)


@router.post("/api/ai/config/approval")
def ai_config_approval_post(body: dict = Body(default_factory=dict)):
    try:
        payload = body or {}
        enabled = payload.get("enabled", True)
        rules = payload.get("rules", [])
        from resources.config.approval_config import get_approval_config, reload_approval_config

        config = get_approval_config()
        config.enabled = enabled
        config.rules = rules
        if "attendance_policy" in payload and isinstance(payload["attendance_policy"], dict):
            from resources.config.approval_config import normalize_attendance_policy

            config.attendance_policy = normalize_attendance_policy(payload["attendance_policy"])
        config.save()
        reload_approval_config()
        from app.application.workflow import reload_approval_service

        reload_approval_service()
        return {"success": True, "message": "保存成功"}
    except Exception as e:
        logger.exception("save approval config: %s", e)
        return JSONResponse({"success": False, "message": f"保存失败：{str(e)}"}, status_code=500)


@router.post("/api/ai/approval/request")
def ai_approval_request(body: dict = Body(default_factory=dict)):
    try:
        payload = body or {}
        plan_id = payload.get("plan_id") or ""
        node_id = payload.get("node_id") or ""
        tool_id = payload.get("tool_id") or ""
        action = payload.get("action") or ""
        params = payload.get("params") or {}
        if not plan_id or not node_id:
            return JSONResponse({"success": False, "message": "缺少 plan_id 或 node_id"}, status_code=400)
        from app.application.workflow import WorkflowNode, get_approval_service

        approval_service = get_approval_service()
        node = WorkflowNode(node_id=node_id, tool_id=tool_id, action=action, params=params)
        approval_req = approval_service.create_approval_request(plan_id=plan_id, node=node)
        return {
            "success": True,
            "message": "审批请求已创建",
            "data": {
                "request_id": approval_req.request_id,
                "plan_id": approval_req.plan_id,
                "node_id": approval_req.node_id,
                "tool_id": approval_req.tool_id,
                "action": approval_req.action,
                "status": approval_req.status.value,
                "created_at": approval_req.created_at.isoformat() if approval_req.created_at else None,
            },
        }
    except Exception as e:
        logger.exception("approval request: %s", e)
        return JSONResponse({"success": False, "message": f"创建审批请求失败：{str(e)}"}, status_code=500)


def _dispatch_tool_for_approval(*, tool_id: str, action: str, params: dict | None = None) -> dict[str, Any]:
    """与 ``WorkflowEngine`` 约定一致：返回带 ``success`` 的字典。"""
    from app.routes.tools import execute_registered_workflow_tool

    return execute_registered_workflow_tool(tool_id=tool_id, action=action, params=params)


@router.post("/api/ai/approval/approve")
def ai_approval_approve(body: dict = Body(default_factory=dict)):
    try:
        payload = body or {}
        request_id = payload.get("request_id") or ""
        plan_id = payload.get("plan_id") or ""
        comment = payload.get("comment") or ""
        from app.application.workflow import WorkflowEngine, get_approval_service

        approval_service = get_approval_service()
        actual_request_id = None
        if request_id:
            success = approval_service.approve(request_id, comment)
            actual_request_id = request_id
        elif plan_id:
            pending_req = approval_service.get_pending_request_by_plan(plan_id)
            if not pending_req:
                return JSONResponse({"success": False, "message": "没有待审批的请求"}, status_code=404)
            success = approval_service.approve(pending_req.request_id, comment)
            actual_request_id = pending_req.request_id
        else:
            return JSONResponse({"success": False, "message": "缺少 request_id 或 plan_id"}, status_code=400)

        if not success:
            return JSONResponse({"success": False, "message": "审批失败"}, status_code=400)

        workflow_data = approval_service.get_pending_workflow(actual_request_id) if actual_request_id else None
        run_result_data = None
        if workflow_data:
            plan_obj = workflow_data.get("plan")
            runtime_ctx = workflow_data.get("runtime_context", {})
            if plan_obj:
                engine = WorkflowEngine(tool_dispatcher=_dispatch_tool_for_approval)
                run_result = engine.run(plan=plan_obj, runtime_context=runtime_ctx, max_retries=1)
                run_result_data = {
                    "plan_id": plan_obj.plan_id,
                    "intent": plan_obj.intent,
                    "nodes_executed": len(run_result.node_results),
                    "nodes_total": len(plan_obj.nodes),
                    "has_errors": any(bool(r.error) for r in run_result.node_results),
                    "results_summary": [
                        {
                            "node_id": r.node_id,
                            "success": r.success,
                            "output": str(r.output)[:200] if r.output else None,
                        }
                        for r in run_result.node_results[:5]
                    ],
                }
                approval_service.remove_pending_workflow(actual_request_id)

        response_data: dict[str, Any] = {"status": "approved", "workflow_executed": workflow_data is not None}
        if run_result_data:
            response_data["workflow_result"] = run_result_data
        return {
            "success": True,
            "message": "审批已通过" + ("，工作流已执行" if workflow_data else ""),
            "data": response_data,
        }
    except Exception as e:
        logger.exception("approval approve: %s", e)
        return JSONResponse({"success": False, "message": f"审批失败：{str(e)}"}, status_code=500)


@router.post("/api/ai/approval/reject")
def ai_approval_reject(body: dict = Body(default_factory=dict)):
    try:
        payload = body or {}
        request_id = payload.get("request_id") or ""
        plan_id = payload.get("plan_id") or ""
        comment = payload.get("comment") or ""
        from app.application.workflow import get_approval_service

        approval_service = get_approval_service()
        if request_id:
            success = approval_service.reject(request_id, comment)
        elif plan_id:
            pending_req = approval_service.get_pending_request_by_plan(plan_id)
            if not pending_req:
                return JSONResponse({"success": False, "message": "没有待审批的请求"}, status_code=404)
            success = approval_service.reject(pending_req.request_id, comment)
        else:
            return JSONResponse({"success": False, "message": "缺少 request_id 或 plan_id"}, status_code=400)
        if success:
            return {"success": True, "message": "审批已拒绝", "data": {"status": "rejected"}}
        return JSONResponse({"success": False, "message": "审批拒绝失败"}, status_code=400)
    except Exception as e:
        logger.exception("approval reject: %s", e)
        return JSONResponse({"success": False, "message": f"审批拒绝失败：{str(e)}"}, status_code=500)


@router.post("/api/ai/parse-single")
def ai_parse_single(body: dict = Body(default_factory=dict)):
    from app.services import get_ai_product_parser

    data = body or {}
    text = data.get("text", "") or ""
    if not text.strip():
        return JSONResponse(
            {
                "success": False,
                "message": "text 不能为空",
                "missing_fields": ["unit", "quantity", "specification", "product"],
                "invalid_reason": "输入为空，无法解析",
            },
            status_code=400,
        )
    parser = get_ai_product_parser()
    result = parser.parse_single(text, use_ai=bool(data.get("use_ai", True)), fallback_to_rule=bool(data.get("fallback_to_rule", True)))
    return JSONResponse(result, status_code=200 if result.get("success") else 422)


@router.post("/api/ai/parse-products")
def ai_parse_products(body: dict = Body(default_factory=dict)):
    from app.services import get_ai_product_parser

    data = body or {}
    texts = data.get("texts") or []
    if not isinstance(texts, list) or not texts:
        return JSONResponse({"success": False, "message": "texts 必须为非空数组"}, status_code=400)
    parser = get_ai_product_parser()
    result = parser.parse_batch(
        texts,
        use_ai=bool(data.get("use_ai", True)),
        fallback_to_rule=bool(data.get("fallback_to_rule", True)),
    )
    return result


@router.post("/api/ai/analyze")
async def ai_analyze_post(
    query: str = Form(default=""),
    file: UploadFile | None = File(default=None),
):
    try:
        from app.services.data_analysis_service import get_data_analysis_service
        from app.utils.secure_filename import secure_filename as _sf

        service = get_data_analysis_service()
        if file is not None and file.filename:
            upload_dir = get_upload_dir()
            os.makedirs(upload_dir, exist_ok=True)
            filename = _sf(file.filename)
            file_path = os.path.join(upload_dir, f"{uuid.uuid4().hex[:8]}_{filename}")
            content = await file.read()
            with open(file_path, "wb") as f:
                f.write(content)
            try:
                result = service.analyze_file(file_path, query)
                return result
            finally:
                try:
                    os.unlink(file_path)
                except OSError:
                    pass
        if (query or "").strip():
            return {
                "success": True,
                "file_info": {"rows": 0, "columns": []},
                "statistics": {},
                "chart_data": {
                    "type": "line",
                    "labels": ["1月", "2月", "3月", "4月"],
                    "datasets": [{"label": "销量", "data": [1200, 1900, 1500, 2300], "borderColor": "#3b82f6"}],
                },
                "insights": ["已理解查询意图", "生成趋势分析"],
                "message": "文本查询分析完成",
            }
        return JSONResponse({"success": False, "message": "请提供文件或查询内容"}, status_code=400)
    except Exception as e:
        logger.exception("ai analyze: %s", e)
        return JSONResponse({"success": False, "message": f"服务器错误: {str(e)}"}, status_code=500)


@router.post("/api/ai/file/analyze")
async def ai_file_analyze(file: UploadFile | None = File(default=None), purpose: str = Form(default="general")):
    try:
        from app.application import get_file_analysis_app_service

        if file is None or not file.filename:
            return JSONResponse({"success": False, "message": "未选择文件"}, status_code=400)
        raw = await file.read()

        class _UploadShim:
            def __init__(self, name: str, data: bytes):
                self.filename = name
                self._data = data

            def save(self, path: str) -> None:
                with open(path, "wb") as f:
                    f.write(self._data)

        service = get_file_analysis_app_service()
        result = service.analyze_file(_UploadShim(file.filename, raw), purpose)
        return JSONResponse(result, status_code=200 if result.get("success") else 400)
    except Exception as e:
        return JSONResponse({"success": False, "message": f"文件分析失败：{str(e)}"}, status_code=500)


# --- Auth / Users ---
@router.post("/api/auth/login")
def auth_login(request: Request, body: dict = Body(default_factory=dict)):
    from app.application import get_auth_app_service

    username = (body.get("username") or "").strip()
    password = body.get("password", "")
    if not username or not password:
        return JSONResponse(
            {"success": False, "error": {"code": "INVALID_INPUT", "message": "用户名和密码不能为空"}},
            status_code=400,
        )
    auth_app_service = get_auth_app_service()
    result = auth_app_service.login(username, password)
    if not result["success"]:
        return JSONResponse(result, status_code=401)
    resp = JSONResponse(result)
    session_id = result.get("session_id")
    if session_id:
        cookie_name = os.environ.get("SESSION_COOKIE_NAME", "session_id")
        max_age = int(os.environ.get("SESSION_COOKIE_MAX_AGE", "86400"))
        resp.set_cookie(
            key=cookie_name,
            value=session_id,
            max_age=max_age,
            httponly=os.environ.get("SESSION_COOKIE_HTTPONLY", "1") not in ("0", "false", "False"),
            secure=os.environ.get("SESSION_COOKIE_SECURE", "").lower() in ("1", "true", "yes"),
            samesite=os.environ.get("SESSION_COOKIE_SAMESITE", "Lax"),
            path="/",
        )
    return resp


@router.post("/api/auth/logout")
def auth_logout(request: Request):
    from app.application import get_auth_app_service

    sid = _session_id_from_request(request)
    if not sid:
        return JSONResponse(
            {"success": False, "error": {"code": "NO_SESSION", "message": "无有效会话"}},
            status_code=400,
        )
    auth_app_service = get_auth_app_service()
    result = auth_app_service.logout(sid)
    resp = JSONResponse(result)
    cookie_name = os.environ.get("SESSION_COOKIE_NAME", "session_id")
    resp.delete_cookie(cookie_name, path="/")
    return resp


@router.post("/api/auth/password/change")
def auth_password_change(request: Request, body: dict = Body(default_factory=dict)):
    user, err = _require_login_user(request)
    if err:
        return err
    from app.application import get_auth_app_service

    old_password = body.get("old_password", "")
    new_password = body.get("new_password", "")
    if not old_password or not new_password:
        return JSONResponse(
            {"success": False, "error": {"code": "INVALID_INPUT", "message": "请填写完整信息"}},
            status_code=400,
        )
    if len(new_password) < 6:
        return JSONResponse(
            {"success": False, "error": {"code": "WEAK_PASSWORD", "message": "新密码至少 6 个字符"}},
            status_code=400,
        )
    auth_app_service = get_auth_app_service()
    result = auth_app_service.change_password(user.id, old_password, new_password)
    if not result["success"]:
        return JSONResponse(result, status_code=400)
    return result


@router.post("/api/users")
def users_create(request: Request, body: dict = Body(default_factory=dict)):
    _, err = _require_permission(request, "admin.manage_users")
    if err:
        return err
    from app.application import get_user_app_service

    username = (body.get("username") or "").strip()
    password = body.get("password", "")
    if not username or not password:
        return JSONResponse(
            {"success": False, "error": {"code": "INVALID_INPUT", "message": "用户名和密码不能为空"}},
            status_code=400,
        )
    if len(password) < 6:
        return JSONResponse(
            {"success": False, "error": {"code": "WEAK_PASSWORD", "message": "密码至少6个字符"}},
            status_code=400,
        )
    role = body.get("role", "viewer")
    if role not in ["viewer", "operator", "admin"]:
        return JSONResponse(
            {"success": False, "error": {"code": "INVALID_ROLE", "message": "无效的角色"}},
            status_code=400,
        )
    user_service = get_user_app_service()
    result = user_service.create_user(
        username=username,
        password=password,
        display_name=body.get("display_name", ""),
        email=body.get("email", ""),
        role=role,
    )
    if not result["success"]:
        return JSONResponse(
            {"success": False, "error": {"code": "CREATE_FAILED", "message": result["error"]}},
            status_code=400,
        )
    return JSONResponse({"success": True, "data": {"user": result["user"]}}, status_code=201)


@router.put("/api/users/{user_id}")
def users_update(request: Request, user_id: int, body: dict = Body(default_factory=dict)):
    _, err = _require_permission(request, "admin.manage_users")
    if err:
        return err
    from app.application import get_user_app_service

    role = body.get("role")
    if role and role not in ["viewer", "operator", "admin"]:
        return JSONResponse(
            {"success": False, "error": {"code": "INVALID_ROLE", "message": "无效的角色"}},
            status_code=400,
        )
    user_service = get_user_app_service()
    result = user_service.update_user(
        user_id=user_id,
        display_name=body.get("display_name"),
        email=body.get("email"),
        role=role,
        is_active=body.get("is_active"),
    )
    if not result["success"]:
        return JSONResponse(
            {"success": False, "error": {"code": "UPDATE_FAILED", "message": result["error"]}},
            status_code=400,
        )
    return {"success": True, "data": {"user": result["user"]}}


@router.post("/api/users/{user_id}/reset-password")
def users_reset_password(request: Request, user_id: int, body: dict = Body(default_factory=dict)):
    _, err = _require_permission(request, "admin.manage_users")
    if err:
        return err
    from app.application import get_auth_app_service

    new_password = body.get("new_password", "admin123")
    if len(new_password) < 6:
        return JSONResponse(
            {"success": False, "error": {"code": "WEAK_PASSWORD", "message": "密码至少6个字符"}},
            status_code=400,
        )
    auth_app_service = get_auth_app_service()
    result = auth_app_service.reset_password(user_id, new_password)
    if not result["success"]:
        return JSONResponse(result, status_code=400)
    return result


# --- Database ---
@router.post("/api/database/backup")
def database_backup():
    try:
        from app.services import get_database_service

        db_service = get_database_service()
        result = db_service.backup_database()
        return JSONResponse(result, status_code=200 if result.get("success") else 500)
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


@router.post("/api/database/restore")
def database_restore(body: dict = Body(default_factory=dict)):
    try:
        from app.services import get_database_service

        data = body or {}
        backup_file = data.get("backup_file")
        if not backup_file:
            return JSONResponse(
                {"success": False, "message": "缺少参数：backup_file"},
                status_code=400,
            )
        db_service = get_database_service()
        result = db_service.restore_database(backup_file)
        return JSONResponse(result, status_code=200 if result.get("success") else 400)
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


# --- Intent packages / intent ---
# --- Inventory ---
@router.post("/api/inventory/in")
def inventory_in(body: dict = Body(default_factory=dict)):
    from app.services.inventory_service import InventoryService

    service = InventoryService()
    data = body or {}
    result = service.inventory_in(
        product_id=data.get("product_id"),
        warehouse_id=data.get("warehouse_id"),
        quantity=float(data.get("quantity", 0)),
        batch_no=data.get("batch_no"),
        location_id=data.get("location_id"),
        unit_price=float(data["unit_price"]) if data.get("unit_price") is not None else None,
        reference_type=data.get("reference_type"),
        reference_id=data.get("reference_id"),
        operator=data.get("operator"),
        remark=data.get("remark"),
    )
    return result


@router.post("/api/inventory/out")
def inventory_out(body: dict = Body(default_factory=dict)):
    from app.services.inventory_service import InventoryService

    service = InventoryService()
    data = body or {}
    result = service.inventory_out(
        product_id=data.get("product_id"),
        warehouse_id=data.get("warehouse_id"),
        quantity=float(data.get("quantity", 0)),
        batch_no=data.get("batch_no"),
        location_id=data.get("location_id"),
        unit_price=float(data["unit_price"]) if data.get("unit_price") is not None else None,
        reference_type=data.get("reference_type"),
        reference_id=data.get("reference_id"),
        operator=data.get("operator"),
        remark=data.get("remark"),
    )
    return result


@router.post("/api/inventory/transfer")
def inventory_transfer(body: dict = Body(default_factory=dict)):
    from app.services.inventory_service import InventoryService

    service = InventoryService()
    data = body or {}
    result = service.inventory_transfer(
        product_id=data.get("product_id"),
        from_warehouse_id=data.get("from_warehouse_id"),
        to_warehouse_id=data.get("to_warehouse_id"),
        quantity=float(data.get("quantity", 0)),
        batch_no=data.get("batch_no"),
        from_location_id=data.get("from_location_id"),
        to_location_id=data.get("to_location_id"),
        operator=data.get("operator"),
        remark=data.get("remark"),
    )
    return result


@router.post("/api/inventory/locations")
def inventory_locations_post(body: dict = Body(default_factory=dict)):
    from app.services.inventory_service import InventoryService

    service = InventoryService()
    return service.create_storage_location(body or {})


@router.post("/api/inventory/warehouses")
def inventory_warehouses_post(body: dict = Body(default_factory=dict)):
    from app.services.inventory_service import InventoryService

    service = InventoryService()
    return service.create_warehouse(body or {})


@router.put("/api/inventory/warehouses/{warehouse_id}")
def inventory_warehouses_put(warehouse_id: int, body: dict = Body(default_factory=dict)):
    from app.services.inventory_service import InventoryService

    service = InventoryService()
    return service.update_warehouse(warehouse_id, body or {})


# --- Performance ---
@router.post("/api/performance/cache/clear")
def performance_cache_clear(pattern: str | None = Query(default=None)):
    try:
        from app.utils.performance_initializer import get_performance_optimizer

        optimizer = get_performance_optimizer()
        if not optimizer.redis_cache:
            return JSONResponse({"success": False, "message": "Redis 缓存未初始化"}, status_code=503)
        if pattern:
            cleared = optimizer.redis_cache.clear_pattern(pattern)
            message = f"已清除模式 '{pattern}' 的缓存 ({cleared} 个键)"
        else:
            optimizer.redis_cache.clear_local_cache()
            message = "已清除本地缓存"
        return {"success": True, "message": message}
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


@router.post("/api/performance/cache/invalidate")
def performance_cache_invalidate(body: dict = Body(default_factory=dict)):
    try:
        data = body or {}
        keys = data.get("keys", [])
        if not keys:
            return JSONResponse({"success": False, "message": "请提供要失效的键列表"}, status_code=400)
        from app.utils.performance_initializer import get_performance_optimizer

        optimizer = get_performance_optimizer()
        if not optimizer.redis_cache:
            return JSONResponse({"success": False, "message": "Redis 缓存未初始化"}, status_code=503)
        deleted = optimizer.redis_cache.delete(*keys)
        return {
            "success": True,
            "data": {"deleted_count": deleted, "requested_keys": len(keys)},
            "message": f"已删除 {deleted} 个缓存键",
        }
    except Exception as e:
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


@router.post("/api/performance/optimize/reinitialize")
def performance_optimize_reinitialize():
    try:
        from app import create_app
        from app.config import DevelopmentConfig
        from app.utils.performance_initializer import init_performance_optimization

        app = create_app(DevelopmentConfig)
        with app.app_context():
            optimizer = init_performance_optimization(app)
        return {"success": True, "message": "性能优化系统已重新初始化", "data": optimizer.get_status()}
    except Exception as e:
        logger.exception("performance reinit: %s", e)
        return JSONResponse({"success": False, "message": str(e)}, status_code=500)


# --- Products ---
@router.post("/api/products/batch")
def products_batch(body: dict = Body(default_factory=dict)):
    from app.bootstrap import get_products_service

    data = body or {}
    products = data.get("products") or []
    if not isinstance(products, list) or not products:
        return JSONResponse({"success": False, "message": "products 必须为非空数组"}, status_code=400)
    service = get_products_service()
    return service.batch_add_products(products)


@router.post("/api/products/{product_id}")
def products_update_post(product_id: int, body: dict = Body(default_factory=dict)):
    from app.bootstrap import get_products_service

    service = get_products_service()
    result = service.update_product(product_id, body or {})
    return JSONResponse(result, status_code=200 if result.get("success") else 400)


@router.put("/api/products/{product_id}")
def products_put(product_id: int, body: dict = Body(default_factory=dict)):
    from app.bootstrap import get_products_service

    service = get_products_service()
    result = service.update_product(product_id, body or {})
    return JSONResponse(result, status_code=200 if result.get("success") else 400)


@router.patch("/api/products/{product_id}")
def products_patch(product_id: int, body: dict = Body(default_factory=dict)):
    from app.bootstrap import get_products_service

    service = get_products_service()
    result = service.update_product(product_id, body or {})
    return JSONResponse(result, status_code=200 if result.get("success") else 400)


# --- Purchase ---
@router.post("/api/purchase/suppliers")
def purchase_suppliers_post(body: dict = Body(default_factory=dict)):
    from app.services.purchase_service import PurchaseService

    return PurchaseService().create_supplier(body or {})


@router.put("/api/purchase/suppliers/{supplier_id}")
def purchase_suppliers_put(supplier_id: int, body: dict = Body(default_factory=dict)):
    from app.services.purchase_service import PurchaseService

    return PurchaseService().update_supplier(supplier_id, body or {})


@router.post("/api/purchase/orders")
def purchase_orders_post(body: dict = Body(default_factory=dict)):
    from app.services.purchase_service import PurchaseService

    return PurchaseService().create_purchase_order(body or {})


@router.put("/api/purchase/orders/{order_id}")
def purchase_orders_put(order_id: int, body: dict = Body(default_factory=dict)):
    from app.services.purchase_service import PurchaseService

    return PurchaseService().update_purchase_order(order_id, body or {})


@router.post("/api/purchase/orders/{order_id}/approve")
def purchase_orders_approve(order_id: int, approver: str = Query(default="system")):
    from app.services.purchase_service import PurchaseService

    return PurchaseService().approve_purchase_order(order_id, approver)


@router.post("/api/purchase/orders/{order_id}/cancel")
def purchase_orders_cancel(order_id: int):
    from app.services.purchase_service import PurchaseService

    return PurchaseService().cancel_purchase_order(order_id)


@router.post("/api/purchase/inbounds")
def purchase_inbounds_post(body: dict = Body(default_factory=dict)):
    from app.services.purchase_service import PurchaseService

    return PurchaseService().create_purchase_inbound(body or {})


# --- Report ---
@router.post("/api/report/export")
def report_export(body: dict = Body(default_factory=dict)):
    from app.services.report_service import ReportService

    data = body or {}
    return ReportService().export_to_excel(
        report_type=data.get("report_type", "report"),
        data=data.get("data", []),
        filename=data.get("filename", "report"),
    )


# --- Skills ---
@router.post("/api/skills/execute")
def skills_execute(body: dict = Body(default_factory=dict)):
    from app.routes.tools import run_archive_tools_execute

    data, code = run_archive_tools_execute(body)
    return JSONResponse(data, status_code=code)


@router.post("/api/skills/analyze/excel")
def skills_analyze_excel(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "message": "请求数据不能为空"}, status_code=400)
    file_path = data.get("file_path")
    if not file_path:
        return JSONResponse({"success": False, "message": "缺少参数: file_path"}, status_code=400)
    from app.infrastructure.skills.excel_analyzer.excel_template_analyzer import get_excel_analyzer_skill

    skill = get_excel_analyzer_skill()
    return skill.execute(file_path=file_path, sheet_name=data.get("sheet_name"))


@router.post("/api/skills/view/excel")
def skills_view_excel(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "message": "请求数据不能为空"}, status_code=400)
    file_path = data.get("file_path")
    if not file_path:
        return JSONResponse({"success": False, "message": "缺少参数：file_path"}, status_code=400)
    from app.infrastructure.skills.excel_toolkit.excel_toolkit import get_excel_toolkit_skill

    skill = get_excel_toolkit_skill()
    return skill.execute(
        file_path=file_path,
        action=data.get("action", "view"),
        sheet_name=data.get("sheet_name"),
    )


@router.post("/api/skills/generate-label-template")
def skills_generate_label_template(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "message": "请求数据不能为空"}, status_code=400)
    image_path = data.get("image_path")
    if not image_path:
        return JSONResponse({"success": False, "message": "缺少参数：image_path"}, status_code=400)
    from app.infrastructure.skills.label_template_generator import get_label_template_generator_skill

    skill = get_label_template_generator_skill()
    return skill.execute(
        image_path=image_path,
        class_name=data.get("class_name", "LabelTemplateGenerator"),
        enable_ocr=data.get("enable_ocr", True),
        verbose=True,
    )


# --- System ---
@router.post("/api/system/printer")
def system_printer_post(body: dict = Body(default_factory=dict)):
    from app.services import get_system_service

    data = body or {}
    if not data:
        return JSONResponse({"success": False, "message": "请求数据不能为空"}, status_code=400)
    printer_name = data.get("printer_name")
    if not printer_name:
        return JSONResponse({"success": False, "message": "缺少参数：printer_name"}, status_code=400)
    result = get_system_service().set_default_printer(printer_name)
    return JSONResponse(result, status_code=200 if result.get("success") else 500)


@router.post("/api/system/startup")
def system_startup_post():
    from app.services import get_system_service

    result = get_system_service().enable_startup()
    return JSONResponse(result, status_code=200 if result.get("success") else 500)


# --- Templates (归档实现代理) ---
@router.post("/api/templates/create")
def templates_create(body: dict = Body(default_factory=dict)):
    from app.routes.document_templates_compat import run_archive_template_create

    data, code = run_archive_template_create(body)
    return JSONResponse(data, status_code=code)


@router.post("/api/templates/update")
def templates_update(body: dict = Body(default_factory=dict)):
    from app.routes.document_templates_compat import run_archive_template_update

    data, code = run_archive_template_update(body)
    return JSONResponse(data, status_code=code)


@router.post("/api/templates/delete")
def templates_delete_post(request: Request, body: dict = Body(default_factory=dict)):
    from app.fastapi_routes.legacy_gaps_batch1 import templates_delete as batch1_templates_delete

    return batch1_templates_delete(request, body)


@router.post("/api/templates/analyze")
async def templates_analyze(
    file: UploadFile = File(...),
    template_name: str = Form(default=""),
    template_scope: str = Form(default=""),
):
    from app.routes.document_templates_compat import run_archive_template_analyze

    raw = await file.read()
    data, code = run_archive_template_analyze(
        file_body=raw,
        filename=file.filename or "upload.bin",
        template_name=template_name,
        template_scope=template_scope,
    )
    return JSONResponse(data, status_code=code)


# --- Tools ---
@router.post("/api/tools/execute")
def tools_execute_route(body: dict = Body(default_factory=dict)):
    from app.routes.tools import run_archive_tools_execute

    data, code = run_archive_tools_execute(body)
    return JSONResponse(data, status_code=code)


# --- Traditional mode writes ---
@router.post("/api/traditional-mode/write")
def traditional_mode_write(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "error": "请求体为空或格式错误"}, status_code=400)
    rel_file = data.get("file", "")
    file_data = data.get("data", {})
    file_type = data.get("type", "")
    full_path = resolve_safe_path(rel_file)
    if full_path is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if file_type != "excel":
        return JSONResponse({"success": False, "error": f"不支持的写入类型: {file_type}"}, status_code=400)
    try:
        import openpyxl
    except ImportError:
        return JSONResponse({"success": False, "error": "openpyxl 未安装，无法写入 Excel 文件"}, status_code=500)
    parent_dir = os.path.dirname(full_path)
    if parent_dir and not os.path.exists(parent_dir):
        os.makedirs(parent_dir, exist_ok=True)
    wb = openpyxl.Workbook()
    default_sheet = wb.active
    default_sheet.title = file_data.get("active_sheet", "Sheet")
    sheets_content = file_data.get("content", {})
    if isinstance(sheets_content, dict):
        for sheet_name, sheet_data_item in sheets_content.items():
            if sheet_name == default_sheet.title:
                ws = default_sheet
            else:
                ws = wb.create_sheet(title=sheet_name)
            rows = sheet_data_item.get("rows", []) if isinstance(sheet_data_item, dict) else []
            for r_idx, row in enumerate(rows, start=1):
                for c_idx, cell_value in enumerate(row, start=1):
                    if cell_value is not None:
                        ws.cell(row=r_idx, column=c_idx, value=cell_value)
    if len(wb.sheetnames) > 1 and default_sheet.title in wb.sheetnames:
        if not sheets_content or default_sheet.title not in sheets_content:
            wb.remove(default_sheet)
    wb.save(full_path)
    wb.close()
    return {"success": True}


@router.post("/api/traditional-mode/mkdir")
def traditional_mode_mkdir(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "error": "请求体为空或格式错误"}, status_code=400)
    rel_path = data.get("path", "")
    folder_name = (data.get("name") or "").strip()
    if not folder_name:
        return JSONResponse({"success": False, "error": "文件夹名称不能为空"}, status_code=400)
    if "/" in folder_name or "\\" in folder_name or ".." in folder_name:
        return JSONResponse({"success": False, "error": "文件夹名称包含非法字符"}, status_code=400)
    full_parent = resolve_safe_path(rel_path)
    if full_parent is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    full_new_path = os.path.join(full_parent, folder_name)
    if not os.path.abspath(full_new_path).startswith(os.path.abspath(ROOT_DIR)):
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if os.path.exists(full_new_path):
        return JSONResponse({"success": False, "error": "文件夹已存在"}, status_code=409)
    os.makedirs(full_new_path, exist_ok=False)
    return {"success": True}


@router.post("/api/traditional-mode/rename")
def traditional_mode_rename(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "error": "请求体为空或格式错误"}, status_code=400)
    rel_path = data.get("path", "")
    old_name = (data.get("old_name") or "").strip()
    new_name = (data.get("new_name") or "").strip()
    if not old_name or not new_name:
        return JSONResponse({"success": False, "error": "旧名称和新名称不能为空"}, status_code=400)
    if "/" in new_name or "\\" in new_name or ".." in new_name:
        return JSONResponse({"success": False, "error": "新名称包含非法字符"}, status_code=400)
    full_parent = resolve_safe_path(rel_path)
    if full_parent is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    full_old_path = os.path.join(full_parent, old_name)
    full_new_path = os.path.join(full_parent, new_name)
    if not os.path.abspath(full_old_path).startswith(os.path.abspath(ROOT_DIR)):
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if not os.path.abspath(full_new_path).startswith(os.path.abspath(ROOT_DIR)):
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if not os.path.exists(full_old_path):
        return JSONResponse({"success": False, "error": "源文件或文件夹不存在"}, status_code=404)
    if os.path.exists(full_new_path):
        return JSONResponse({"success": False, "error": "目标名称已存在"}, status_code=409)
    os.rename(full_old_path, full_new_path)
    return {"success": True}


@router.post("/api/traditional-mode/delete")
def traditional_mode_delete(body: dict = Body(default_factory=dict)):
    data = body or {}
    if not data:
        return JSONResponse({"success": False, "error": "请求体为空或格式错误"}, status_code=400)
    rel_path = data.get("path", "")
    name = (data.get("name") or "").strip()
    if not name:
        return JSONResponse({"success": False, "error": "名称不能为空"}, status_code=400)
    full_parent = resolve_safe_path(rel_path)
    if full_parent is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    full_target = os.path.join(full_parent, name)
    if not os.path.abspath(full_target).startswith(os.path.abspath(ROOT_DIR)):
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    if not os.path.exists(full_target):
        return JSONResponse({"success": False, "error": "文件或文件夹不存在"}, status_code=404)
    if os.path.isdir(full_target):
        shutil.rmtree(full_target)
    else:
        os.remove(full_target)
    return {"success": True}


@router.post("/api/traditional-mode/upload")
async def traditional_mode_upload(
    file: UploadFile = File(...),
    path: str = Form(default=""),
):
    if not file.filename:
        return JSONResponse({"success": False, "error": "文件名为空"}, status_code=400)
    full_target_dir = resolve_safe_path(path)
    if full_target_dir is None:
        return JSONResponse({"success": False, "error": "路径越权访问被拒绝"}, status_code=403)
    filename = secure_filename(file.filename)
    if not filename:
        filename = "uploaded_file"
    if not os.path.exists(full_target_dir):
        os.makedirs(full_target_dir, exist_ok=True)
    save_path = os.path.join(full_target_dir, filename)
    if os.path.exists(save_path):
        base, ext = os.path.splitext(filename)
        counter = 1
        while os.path.exists(save_path):
            filename = f"{base}_{counter}{ext}"
            save_path = os.path.join(full_target_dir, filename)
            counter += 1
    content = await file.read()
    with open(save_path, "wb") as f:
        f.write(content)
    return {"success": True, "filename": filename, "path": save_path}


# --- WeChat 桌面 API ---
@router.post("/api/wechat/task/{task_id}/confirm")
def wechat_task_confirm(task_id: int):
    from app.application import get_wechat_task_app_service

    result = get_wechat_task_app_service().confirm_task(task_id)
    return JSONResponse(result, status_code=200 if result.get("success") else 400)


@router.post("/api/wechat/task/{task_id}/ignore")
def wechat_task_ignore(task_id: int):
    from app.application import get_wechat_task_app_service

    result = get_wechat_task_app_service().ignore_task(task_id)
    return JSONResponse(result, status_code=200 if result.get("success") else 400)


@router.post("/api/wechat/contacts")
def wechat_contacts_post(body: dict = Body(default_factory=dict)):
    from app.application import get_wechat_contact_app_service

    data = body or {}
    contact_name = (data.get("contact_name") or "").strip()
    if not contact_name:
        return JSONResponse({"success": False, "message": "联系人名称不能为空"}, status_code=400)
    service = get_wechat_contact_app_service()
    result = service.add_contact(
        contact_name=contact_name,
        remark=(data.get("remark") or "").strip(),
        wechat_id=(data.get("wechat_id") or "").strip(),
        contact_type=data.get("contact_type", "contact"),
        is_starred=data.get("is_starred", True),
    )
    return JSONResponse(result, status_code=200 if result.get("success") else 400)


@router.put("/api/wechat/contacts/{contact_id}")
def wechat_contacts_put(contact_id: int, body: dict = Body(default_factory=dict)):
    from app.application import get_wechat_contact_app_service

    data = body or {}
    service = get_wechat_contact_app_service()
    result = service.update_contact(
        contact_id=contact_id,
        contact_name=(data.get("contact_name") or "").strip() or None,
        remark=(data.get("remark") or "").strip() or None,
        wechat_id=(data.get("wechat_id") or "").strip() or None,
        contact_type=data.get("contact_type"),
        is_starred=data.get("is_starred"),
    )
    return JSONResponse(result, status_code=200 if result.get("success") else 400)


@router.post("/api/wechat/contacts/{contact_id}/star")
def wechat_contacts_star(contact_id: int, body: dict = Body(default_factory=dict)):
    from app.application import get_wechat_contact_app_service

    data = body or {}
    starred = data.get("starred", True)
    service = get_wechat_contact_app_service()
    star_fn = getattr(service, "star_contact", None)
    if callable(star_fn):
        result = star_fn(contact_id, starred)
    else:
        result = service.update_contact(contact_id=contact_id, is_starred=starred)
    return JSONResponse(result, status_code=200 if result.get("success") else 400)


@router.post("/api/wechat/contacts/unstar-all")
def wechat_contacts_unstar_all():
    from app.application import get_wechat_contact_app_service

    return get_wechat_contact_app_service().unstar_all()


@router.post("/api/wechat/scan")
def wechat_scan(body: dict = Body(default_factory=dict)):
    data = body or {}
    try:
        from app.tasks.wechat_tasks import scan_wechat_messages

        task = scan_wechat_messages.delay(contact_id=data.get("contact_id"), limit=data.get("limit", 20))
        return JSONResponse(
            {"success": True, "message": "扫描任务已触发", "task_id": task.id, "count": 0},
            status_code=202,
        )
    except Exception as e:
        return JSONResponse({"success": False, "message": f"扫描失败：{str(e)}"}, status_code=500)


@router.post("/api/wechat/login")
def wechat_miniprogram_style_login(body: dict = Body(default_factory=dict)):
    """兼容 ``/api/wechat/login``（与 ``app.routes.wechat_miniprogram`` 行为对齐）。"""
    from app.services.wechat_miniprogram_auth import WechatMiniProgramError, miniprogram_login_data_for_wx_username_binding

    data = body or {}
    code = (data.get("code") or "").strip()
    if not code:
        return _mp_wechat_json_response(400, "code 不能为空", {"error": "missing_code"}, success=False)
    try:
        payload = miniprogram_login_data_for_wx_username_binding(code)
        return _mp_wechat_json_response(200, "登录成功", payload)
    except WechatMiniProgramError as e:
        logger.error("wechat login: %s", e)
        return _mp_wechat_json_response(500, str(e), {"error": "wechat_api_error"}, success=False)
    except Exception as e:
        logger.exception("wechat login: %s", e)
        return _mp_wechat_json_response(500, f"登录失败：{str(e)}", {"error": "internal_error"}, success=False)


@router.put("/api/wechat/user/info")
def wechat_user_info_put(request: Request, body: dict = Body(default_factory=dict)):
    from app.decorators.mp_auth import verify_jwt_token

    auth = request.headers.get("Authorization") or ""
    if not auth.startswith("Bearer "):
        return _mp_wechat_json_response(401, "未授权", {"error": "缺少 token"}, success=False)
    payload = verify_jwt_token(auth[7:].strip())
    if not payload:
        return _mp_wechat_json_response(401, "token 无效或已过期", {"error": "invalid_token"}, success=False)
    uid = payload.get("user_id")
    from app.db.models import User
    from app.db.session import get_db

    data = body or {}
    with get_db() as db:
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            return _mp_wechat_json_response(404, "用户不存在", {"error": "user_not_found"}, success=False)
        if "display_name" in data:
            user.display_name = data["display_name"]
        db.commit()
        db.refresh(user)
        return _mp_wechat_json_response(
            200,
            "更新成功",
            {
                "id": user.id,
                "username": user.username,
                "display_name": user.display_name,
                "email": user.email,
                "role": user.role,
                "avatar": "",
                "created_at": user.created_at.isoformat() if user.created_at else None,
            },
        )


# --- Conversations ---
@router.put("/api/conversations/{session_id}/title")
def conversations_title_put(session_id: str, body: dict = Body(default_factory=dict)):
    from app.services.conversation_service import get_conversation_service as get_conversation_app_service

    service = get_conversation_app_service()
    data = body or {}
    title = data.get("title", "")
    success = service.update_session_title(session_id, title)
    return {"success": success}


# --- 小程序 MP v1（写操作）---
def _mp_uid_or_401(authorization: str | None):
    uid = _mp_jwt_user_id(authorization)
    if uid is None:
        return None, _mp_json_response(401, "未授权", {"message": "missing_token"}, success=False)
    return uid, None


@router.post("/api/mp/v1/auth/login")
def mp_v1_auth_login(body: dict = Body(default_factory=dict)):
    from datetime import datetime

    import requests

    from app.db.models import User
    from app.db.session import get_db
    from app.decorators.mp_auth import generate_jwt_token

    data = body or {}
    code = (data.get("code") or "").strip()
    if not code:
        return _mp_json_response(400, "code 不能为空", {"error": "missing_code"}, success=False)

    config = {
        "appid": os.environ.get("WECHAT_MINIPROGRAM_APPID", ""),
        "secret": os.environ.get("WECHAT_MINIPROGRAM_SECRET", ""),
    }
    if not config["appid"] or not config["secret"]:
        return _mp_json_response(500, "微信小程序配置缺失", {"error": "wechat_api_error"}, success=False)
    try:
        url = "https://api.weixin.qq.com/sns/jscode2session"
        params = {
            "appid": config["appid"],
            "secret": config["secret"],
            "js_code": code,
            "grant_type": "authorization_code",
        }
        response = requests.get(url, params=params, timeout=10)
        response.raise_for_status()
        result = response.json()
        if "errcode" in result:
            return _mp_json_response(
                500,
                f"微信登录失败：{result.get('errmsg', '未知错误')}",
                {"error": "wechat_api_error"},
                success=False,
            )
        openid = result.get("openid")
        if not openid:
            return _mp_json_response(500, "微信登录失败，未获取到 openid", {"error": "no_openid"}, success=False)

        with get_db() as db:
            user = db.query(User).filter(User.wx_openid == openid).first()
            if not user:
                user = User(
                    username=f"wx_{openid}",
                    password=uuid.uuid4().hex,
                    display_name="微信用户",
                    email="",
                    role="mp_user",
                    is_active=True,
                    wx_openid=openid,
                    wx_unionid=result.get("unionid"),
                    created_at=datetime.now(),
                )
                db.add(user)
                db.commit()
                db.refresh(user)
            user.last_login = datetime.now()
            db.commit()
            token = generate_jwt_token(user.id, openid)
        return _mp_json_response(
            200,
            "登录成功",
            {
                "token": token,
                "expires_in": 720 * 3600,
                "user": {
                    "id": user.id,
                    "username": user.username,
                    "display_name": user.display_name or "微信用户",
                    "avatar": user.wx_avatar_url or "",
                    "role": user.role,
                    "created_at": user.created_at.isoformat() if user.created_at else None,
                },
            },
        )
    except Exception as e:
        logger.exception("mp login: %s", e)
        return _mp_json_response(500, f"登录失败：{str(e)}", {"error": "internal_error"}, success=False)


@router.post("/api/mp/v1/auth/logout")
def mp_v1_auth_logout(authorization: str | None = Header(default=None)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    _ = uid
    return _mp_json_response(200, "登出成功", None)


def _mp_generate_order_no() -> str:
    from datetime import datetime

    import random
    import string

    prefix = datetime.now().strftime("%Y%m%d%H%M%S")
    suffix = "".join(random.choices(string.digits, k=6))
    return f"MP{prefix}{suffix}"


@router.post("/api/mp/v1/cart/add")
def mp_cart_add(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpCart, Product
    from app.db.session import get_db

    data = body or {}
    product_id = data.get("product_id")
    quantity = max(1, int(data.get("quantity", 1) or 1))
    if not product_id:
        return _mp_json_response(400, "商品ID不能为空", None, success=False)
    with get_db() as db:
        product = db.query(Product).filter(Product.id == product_id, Product.is_active == 1).first()
        if not product:
            return _mp_json_response(404, "商品不存在", None, success=False)
        existing = db.query(MpCart).filter(MpCart.user_id == uid, MpCart.product_id == product_id).first()
        if existing:
            existing.quantity += quantity
        else:
            existing = MpCart(user_id=uid, product_id=product_id, quantity=quantity, selected=True)
            db.add(existing)
        db.commit()
        return _mp_json_response(200, "添加成功", {"cart_id": existing.id})


@router.put("/api/mp/v1/cart/update")
def mp_cart_update(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpCart
    from app.db.session import get_db

    data = body or {}
    product_id = data.get("product_id")
    quantity = data.get("quantity")
    if not product_id:
        return _mp_json_response(400, "商品ID不能为空", None, success=False)
    if quantity is None or int(quantity) < 1:
        return _mp_json_response(400, "数量必须大于0", None, success=False)
    with get_db() as db:
        cart = db.query(MpCart).filter(MpCart.user_id == uid, MpCart.product_id == product_id).first()
        if not cart:
            return _mp_json_response(404, "购物车中不存在该商品", None, success=False)
        cart.quantity = int(quantity)
        db.commit()
        return _mp_json_response(200, "更新成功", None)


@router.put("/api/mp/v1/cart/select")
def mp_cart_select(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpCart
    from app.db.session import get_db

    data = body or {}
    product_id = data.get("product_id")
    selected = data.get("selected", True)
    if not product_id:
        return _mp_json_response(400, "商品ID不能为空", None, success=False)
    with get_db() as db:
        cart = db.query(MpCart).filter(MpCart.user_id == uid, MpCart.product_id == product_id).first()
        if not cart:
            return _mp_json_response(404, "购物车中不存在该商品", None, success=False)
        cart.selected = bool(selected)
        db.commit()
        return _mp_json_response(200, "操作成功", None)


@router.post("/api/mp/v1/address/create")
def mp_address_create(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpAddress
    from app.db.session import get_db

    data = body or {}
    required_fields = ["contact_name", "contact_phone", "province", "city", "district", "detail_address"]
    for field in required_fields:
        if not data.get(field):
            return _mp_json_response(400, f"{field} 不能为空", None, success=False)
    with get_db() as db:
        is_default = data.get("is_default", False)
        if is_default:
            db.query(MpAddress).filter(MpAddress.user_id == uid).update({"is_default": False})
        count = db.query(MpAddress).filter(MpAddress.user_id == uid).count()
        if count == 0:
            is_default = True
        address = MpAddress(
            user_id=uid,
            contact_name=data["contact_name"],
            contact_phone=data["contact_phone"],
            province=data["province"],
            city=data["city"],
            district=data["district"],
            detail_address=data["detail_address"],
            is_default=is_default,
        )
        db.add(address)
        db.commit()
        db.refresh(address)
        return _mp_json_response(200, "地址添加成功", {"id": address.id, "is_default": address.is_default})


@router.put("/api/mp/v1/address/update/{address_id}")
def mp_address_update(
    address_id: int,
    authorization: str | None = Header(default=None),
    body: dict = Body(default_factory=dict),
):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpAddress
    from app.db.session import get_db

    data = body or {}
    with get_db() as db:
        address = db.query(MpAddress).filter(MpAddress.id == address_id, MpAddress.user_id == uid).first()
        if not address:
            return _mp_json_response(404, "地址不存在", None, success=False)
        updatable = ["contact_name", "contact_phone", "province", "city", "district", "detail_address"]
        for field in updatable:
            if field in data and data[field]:
                setattr(address, field, data[field])
        if data.get("is_default") and not address.is_default:
            db.query(MpAddress).filter(MpAddress.user_id == uid, MpAddress.id != address_id).update(
                {"is_default": False}
            )
            address.is_default = True
        db.commit()
        return _mp_json_response(200, "地址更新成功", None)


@router.put("/api/mp/v1/address/default/{address_id}")
def mp_address_default(
    address_id: int,
    authorization: str | None = Header(default=None),
):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpAddress
    from app.db.session import get_db

    with get_db() as db:
        address = db.query(MpAddress).filter(MpAddress.id == address_id, MpAddress.user_id == uid).first()
        if not address:
            return _mp_json_response(404, "地址不存在", None, success=False)
        db.query(MpAddress).filter(MpAddress.user_id == uid).update({"is_default": False})
        address.is_default = True
        db.commit()
        return _mp_json_response(200, "默认地址设置成功", None)


@router.post("/api/mp/v1/order/create")
def mp_order_create(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpAddress, MpCart, MpOrder, MpOrderItem, Product
    from app.db.session import get_db

    data = body or {}
    address_id = data.get("address_id")
    remark = data.get("remark", "")
    cart_item_ids = data.get("cart_item_ids", [])
    if not address_id:
        return _mp_json_response(400, "请选择收货地址", None, success=False)
    with get_db() as db:
        address = db.query(MpAddress).filter(MpAddress.id == address_id, MpAddress.user_id == uid).first()
        if not address:
            return _mp_json_response(404, "收货地址不存在", None, success=False)
        if cart_item_ids:
            carts = (
                db.query(MpCart)
                .filter(MpCart.id.in_(cart_item_ids), MpCart.user_id == uid, MpCart.selected == True)
                .all()
            )
        else:
            carts = db.query(MpCart).filter(MpCart.user_id == uid, MpCart.selected == True).all()
        if not carts:
            return _mp_json_response(400, "请选择要结算的商品", None, success=False)
        order_items_data = []
        total_amount = 0.0
        for cart in carts:
            product = db.query(Product).filter(Product.id == cart.product_id).first()
            if not product or product.is_active != 1:
                continue
            unit_price = float(product.price) if product.price else 0
            subtotal = round(unit_price * cart.quantity, 2)
            total_amount += subtotal
            order_items_data.append(
                {
                    "product_id": product.id,
                    "product_name": product.name,
                    "product_sku": product.model_number or "",
                    "quantity": cart.quantity,
                    "unit_price": unit_price,
                    "subtotal": subtotal,
                }
            )
        if not order_items_data:
            return _mp_json_response(400, "没有有效的商品", None, success=False)
        order = MpOrder(
            order_no=_mp_generate_order_no(),
            user_id=uid,
            status="pending",
            total_amount=round(total_amount, 2),
            pay_status="unpaid",
            delivery_name=address.contact_name,
            delivery_phone=address.contact_phone,
            delivery_address=f"{address.province}{address.city}{address.district}{address.detail_address}",
            delivery_province=address.province,
            delivery_city=address.city,
            delivery_district=address.district,
            remark=remark,
        )
        db.add(order)
        db.flush()
        for item_data in order_items_data:
            item = MpOrderItem(order_id=order.id, **item_data)
            db.add(item)
        db.query(MpCart).filter(MpCart.id.in_([c.id for c in carts])).delete(synchronize_session=False)
        db.commit()
        db.refresh(order)
        return _mp_json_response(
            200,
            "订单创建成功",
            {
                "order_id": order.id,
                "order_no": order.order_no,
                "total_amount": float(order.total_amount),
                "status": order.status,
            },
        )


@router.put("/api/mp/v1/order/cancel/{order_id}")
def mp_order_cancel(order_id: int, authorization: str | None = Header(default=None)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpOrder
    from app.db.session import get_db

    with get_db() as db:
        order = db.query(MpOrder).filter(MpOrder.id == order_id, MpOrder.user_id == uid).first()
        if not order:
            return _mp_json_response(404, "订单不存在", None, success=False)
        if order.status not in ("pending", "paid"):
            return _mp_json_response(400, "当前状态不允许取消", None, success=False)
        order.status = "cancelled"
        db.commit()
        return _mp_json_response(200, "订单已取消", {"order_id": order.id, "status": order.status})


@router.put("/api/mp/v1/order/confirm/{order_id}")
def mp_order_confirm(order_id: int, authorization: str | None = Header(default=None)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpOrder
    from app.db.session import get_db

    with get_db() as db:
        order = db.query(MpOrder).filter(MpOrder.id == order_id, MpOrder.user_id == uid).first()
        if not order:
            return _mp_json_response(404, "订单不存在", None, success=False)
        if order.status != "shipped":
            return _mp_json_response(400, "当前状态无法确认收货", None, success=False)
        order.status = "completed"
        db.commit()
        return _mp_json_response(200, "确认收货成功", {"order_id": order.id, "status": order.status})


@router.post("/api/mp/v1/order/rebuy/{order_id}")
def mp_order_rebuy(order_id: int, authorization: str | None = Header(default=None)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpCart, MpOrder
    from app.db.session import get_db

    with get_db() as db:
        order = db.query(MpOrder).filter(MpOrder.id == order_id, MpOrder.user_id == uid).first()
        if not order:
            return _mp_json_response(404, "订单不存在", None, success=False)
        for item in order.items:
            existing = db.query(MpCart).filter(MpCart.user_id == uid, MpCart.product_id == item.product_id).first()
            if existing:
                existing.quantity += item.quantity
                existing.selected = True
            else:
                db.add(
                    MpCart(
                        user_id=uid,
                        product_id=item.product_id,
                        quantity=item.quantity,
                        selected=True,
                    )
                )
        db.commit()
        return _mp_json_response(200, "已加入购物车", None)


@router.post("/api/mp/v1/favorite/add")
def mp_favorite_add(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpFavorite, Product
    from app.db.session import get_db

    data = body or {}
    product_id = data.get("product_id")
    if not product_id:
        return _mp_json_response(400, "商品ID不能为空", None, success=False)
    with get_db() as db:
        product = db.query(Product).filter(Product.id == product_id, Product.is_active == 1).first()
        if not product:
            return _mp_json_response(404, "商品不存在", None, success=False)
        existing = db.query(MpFavorite).filter(MpFavorite.user_id == uid, MpFavorite.product_id == product_id).first()
        if existing:
            return _mp_json_response(200, "已收藏", None)
        nf = MpFavorite(user_id=uid, product_id=product_id)
        db.add(nf)
        db.commit()
        return _mp_json_response(200, "收藏成功", {"fav_id": nf.id})


@router.post("/api/mp/v1/feedback/submit")
def mp_feedback_submit(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpFeedback
    from app.db.session import get_db

    data = body or {}
    fb_type = (data.get("type") or "").strip()
    content = (data.get("content") or "").strip()
    images = data.get("images", [])
    valid_types = ["bug", "suggestion", "complaint", "other"]
    if fb_type not in valid_types:
        return _mp_json_response(400, "反馈类型无效", None, success=False)
    if not content:
        return _mp_json_response(400, "反馈内容不能为空", None, success=False)
    if len(content) > 1000:
        return _mp_json_response(400, "反馈内容不能超过1000字", None, success=False)
    with get_db() as db:
        feedback = MpFeedback(
            user_id=uid,
            type=fb_type,
            content=content,
            images=json.dumps(images) if images else None,
            status="pending",
        )
        db.add(feedback)
        db.commit()
        db.refresh(feedback)
        return _mp_json_response(200, "提交成功，感谢您的反馈！", {"feedback_id": feedback.id})


@router.put("/api/mp/v1/message/read/{msg_id}")
def mp_message_read(msg_id: int, authorization: str | None = Header(default=None)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpNotification
    from app.db.session import get_db

    with get_db() as db:
        msg = db.query(MpNotification).filter(MpNotification.id == msg_id, MpNotification.user_id == uid).first()
        if not msg:
            return _mp_json_response(404, "消息不存在", None, success=False)
        msg.is_read = True
        db.commit()
        return _mp_json_response(200, "已标记为已读", None)


@router.put("/api/mp/v1/message/read-all")
def mp_message_read_all(authorization: str | None = Header(default=None)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import MpNotification
    from app.db.session import get_db

    with get_db() as db:
        db.query(MpNotification).filter(MpNotification.user_id == uid, MpNotification.is_read == False).update(
            {"is_read": True}
        )
        db.commit()
        return _mp_json_response(200, "全部已读", None)


@router.put("/api/mp/v1/user/info")
def mp_user_info_put(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import User
    from app.db.session import get_db

    data = body or {}
    with get_db() as db:
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            return _mp_json_response(404, "用户不存在", {"error": "user_not_found"}, success=False)
        if "display_name" in data and data["display_name"]:
            user.display_name = data["display_name"]
        if "nickname" in data:
            user.mp_nickname = data["nickname"]
        if "avatar" in data:
            user.wx_avatar_url = data["avatar"]
        db.commit()
        db.refresh(user)
        return _mp_json_response(
            200,
            "更新成功",
            {
                "id": user.id,
                "display_name": user.display_name,
                "nickname": user.mp_nickname,
                "avatar": user.wx_avatar_url or "",
            },
        )


@router.post("/api/mp/v1/user/phone")
def mp_user_phone(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    from app.db.models import User
    from app.db.session import get_db

    data = body or {}
    phone = (data.get("phone") or "").strip()
    if not phone:
        return _mp_json_response(400, "手机号不能为空", None, success=False)
    with get_db() as db:
        user = db.query(User).filter(User.id == uid).first()
        if not user:
            return _mp_json_response(404, "用户不存在", None, success=False)
        user.mp_phone = phone
        db.commit()
        return _mp_json_response(200, "绑定成功", {"phone": phone})


@router.post("/api/mp/v1/ai/chat")
def mp_v1_ai_chat(authorization: str | None = Header(default=None), body: dict = Body(default_factory=dict)):
    uid, err = _mp_uid_or_401(authorization)
    if err:
        return err
    data = body or {}
    message = (data.get("message") or "").strip()
    session_id = data.get("session_id")
    if not message:
        return _mp_json_response(400, "消息内容不能为空", None, success=False)
    try:
        from app.services.ai_conversation_service import AIConversationService

        service = AIConversationService()
        response_text = service.chat(
            message=message,
            user_id=uid,
            session_id=session_id,
            source="miniprogram",
        )
        return _mp_json_response(200, "ok", {"reply": response_text, "message": message})
    except Exception as e:
        logger.exception("mp ai chat: %s", e)
        return _mp_json_response(500, f"AI 服务暂时不可用: {str(e)}", None, success=False)


_emit_legacy_gaps_load_log()
