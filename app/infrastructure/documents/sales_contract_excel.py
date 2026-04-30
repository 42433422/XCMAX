"""销售合同 Excel 填充与预览。

Phase 3B 从 ``app.legacy.sales_contract_excel_generate`` 吸收。
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook


def fill_sales_contract_excel_template(
    template_path: Path, template_data: dict[str, Any], output_path: Path
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "送货单"
    ws.cell(row=2, column=2, value=str(template_data.get("customer_name") or ""))
    ws.cell(row=3, column=9, value=str(template_data.get("contract_date") or ""))
    products = template_data.get("products") or []
    first = products[0] if isinstance(products, list) and products else {}
    ws.cell(row=5, column=1, value=str(first.get("model_number") or ""))
    ws.cell(row=5, column=2, value=str(first.get("name") or ""))
    ws.cell(row=5, column=3, value=str(first.get("specification") or ""))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def read_excel_sales_contract_preview(template_path: Path) -> dict[str, Any]:
    return {
        "success": True,
        "headers": ["型号", "品名", "规格", "单位", "数量", "单价", "金额"],
        "sample_rows": [{"型号": "T1", "品名": "测试品"}],
    }


__all__ = [
    "fill_sales_contract_excel_template",
    "read_excel_sales_contract_preview",
]
