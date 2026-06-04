"""人员管理 / 部门（客户）/ 远端 yuangon 同步 — 可挂到任意 Mod 的 ``/api/mod/{id}`` 前缀下。"""

from __future__ import annotations

import logging
import os
from pathlib import Path

from fastapi import APIRouter, File, UploadFile
from fastapi.responses import JSONResponse

from app.mod_sdk.personnel_db import get_personnel_database_path
from app.mod_sdk.remote_yuangon_sync import (
    DEFAULT_REMOTE_ROOT,
    DEFAULT_SSH_TARGET,
    sync_remote_yuangon_employees,
)

logger = logging.getLogger(__name__)


def attach_personnel_crud_routes(router: APIRouter, mod_id: str) -> None:
    """向已有 ``APIRouter`` 注册 employees / products / customers CRUD（不含考勤转换）。"""

    def _db():
        return get_personnel_database_path()

    @router.post("/employees/sync-remote-yuangon", response_model=None)
    async def sync_remote_yuangon_employees_endpoint(data: dict | None = None):
        try:
            payload = data or {}
            ssh_target = (
                str(payload.get("ssh_target") or "").strip()
                or os.environ.get("XCAGI_REMOTE_EMPLOYEE_SSH_TARGET", "").strip()
                or DEFAULT_SSH_TARGET
            )
            remote_root = (
                str(payload.get("remote_root") or "").strip()
                or os.environ.get("XCAGI_REMOTE_EMPLOYEE_ROOT", "").strip()
                or DEFAULT_REMOTE_ROOT
            )
            connect_timeout = int(payload.get("connect_timeout") or 15)
            result = sync_remote_yuangon_employees(
                _db(),
                ssh_target=ssh_target,
                remote_root=remote_root,
                connect_timeout=connect_timeout,
            )
            return {"success": True, "data": result}
        except Exception as e:
            logger.exception("同步远端 yuangon 员工失败 mod_id=%s", mod_id)
            return JSONResponse(
                {"success": False, "message": f"同步远端员工失败: {e}", "error": str(e)},
                status_code=500,
            )

    @router.get("/employees", response_model=None)
    async def list_employees(page: int = 1, page_size: int = 50, search: str = ""):
        import sqlite3

        db_path = _db()
        if not db_path.exists():
            return {
                "success": True,
                "data": {"items": [], "total": 0, "page": page, "page_size": page_size},
            }
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        like = f"%{search}%"
        cur.execute(
            "SELECT COUNT(*) FROM attendance_employees WHERE employee_name LIKE ? OR department LIKE ?",
            (like, like),
        )
        total = cur.fetchone()[0]
        offset = (page - 1) * page_size
        cur.execute(
            "SELECT id, employee_name, department, main_department, attendance_group, employee_no, position, user_id "
            "FROM attendance_employees WHERE employee_name LIKE ? OR department LIKE ? "
            "ORDER BY id LIMIT ? OFFSET ?",
            (like, like, page_size, offset),
        )
        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return {
            "success": True,
            "data": {"items": items, "total": total, "page": page, "page_size": page_size},
        }

    @router.get("/departments", response_model=None)
    async def list_departments(page: int = 1, page_size: int = 50, search: str = ""):
        import sqlite3

        db_path = _db()
        if not db_path.exists():
            return {
                "success": True,
                "data": {"items": [], "total": 0, "page": page, "page_size": page_size},
            }
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        like = f"%{search}%"
        cur.execute(
            "SELECT COUNT(*) FROM attendance_departments WHERE department LIKE ? OR main_department LIKE ?",
            (like, like),
        )
        total = cur.fetchone()[0]
        offset = (page - 1) * page_size
        cur.execute(
            "SELECT id, department, main_department, attendance_group "
            "FROM attendance_departments WHERE department LIKE ? OR main_department LIKE ? "
            "ORDER BY id LIMIT ? OFFSET ?",
            (like, like, page_size, offset),
        )
        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return {
            "success": True,
            "data": {"items": items, "total": total, "page": page, "page_size": page_size},
        }

    @router.get("/products/list", response_model=None)
    async def products_list(
        page: int = 1,
        per_page: int = 20,
        keyword: str = "",
        unit: str = "",
    ):
        import sqlite3

        db_path = _db()
        if not db_path.exists():
            return {"success": True, "data": [], "total": 0}
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cond = []
        args = []
        if keyword:
            cond.append("(model_number LIKE ? OR name LIKE ?)")
            args.extend([f"%{keyword}%", f"%{keyword}%"])
        if unit:
            cond.append("unit = ?")
            args.append(unit)
        where = " AND ".join(cond) if cond else "1=1"
        cur.execute(f"SELECT COUNT(*) FROM products WHERE {where}", args)
        total = cur.fetchone()[0]
        offset = (page - 1) * per_page
        cur.execute(
            f"SELECT id, model_number, name, specification, price, unit "
            f"FROM products WHERE {where} ORDER BY id LIMIT ? OFFSET ?",
            [*args, per_page, offset],
        )
        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return {"success": True, "data": items, "total": total}

    @router.get("/products/{product_id}", response_model=None)
    async def products_get(product_id: int):
        import sqlite3

        db_path = _db()
        if not db_path.exists():
            return JSONResponse({"success": False, "error": "not found"}, status_code=404)
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM products WHERE id = ?", (product_id,))
        row = cur.fetchone()
        conn.close()
        if not row:
            return JSONResponse({"success": False, "error": "not found"}, status_code=404)
        return {"success": True, "data": dict(row)}

    @router.post("/products/add", response_model=None)
    async def products_add(data: dict):
        import datetime
        import sqlite3

        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        now = datetime.datetime.now().isoformat()
        cur.execute(
            "INSERT INTO products (source_file, model_number, name, specification, price, unit, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                data.get("source_file", ""),
                data.get("model_number", ""),
                data.get("name", ""),
                data.get("specification", ""),
                float(data.get("price") or 0),
                data.get("unit", ""),
                now,
                now,
            ),
        )
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        return {"success": True, "data": {"id": new_id}}

    @router.post("/products/update", response_model=None)
    async def products_update(data: dict):
        import datetime
        import sqlite3

        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        now = datetime.datetime.now().isoformat()
        cur.execute(
            "UPDATE products SET model_number=?, name=?, specification=?, price=?, unit=?, updated_at=? WHERE id=?",
            (
                data.get("model_number", ""),
                data.get("name", ""),
                data.get("specification", ""),
                float(data.get("price") or 0),
                data.get("unit", ""),
                now,
                data.get("id"),
            ),
        )
        conn.commit()
        conn.close()
        return {"success": True}

    @router.post("/products/delete", response_model=None)
    async def products_delete(data: dict):
        import sqlite3

        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        cur.execute("DELETE FROM products WHERE id = ?", (data.get("id"),))
        conn.commit()
        conn.close()
        return {"success": True}

    @router.post("/products/batch-delete", response_model=None)
    async def products_batch_delete(data: dict):
        import sqlite3

        ids = data.get("ids") or []
        if not ids:
            return {"success": True}
        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        placeholders = ",".join("?" * len(ids))
        cur.execute(f"DELETE FROM products WHERE id IN ({placeholders})", ids)
        conn.commit()
        conn.close()
        return {"success": True}

    @router.get("/products/product_names", response_model=None)
    async def products_names():
        import sqlite3

        db_path = _db()
        if not db_path.exists():
            return {"success": True, "data": []}
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT id, model_number, name FROM products ORDER BY id")
        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return {"success": True, "data": items}

    @router.get("/products/product_names/search", response_model=None)
    async def products_names_search(keyword: str = ""):
        import sqlite3

        db_path = _db()
        if not db_path.exists():
            return {"success": True, "data": []}
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute(
            "SELECT id, model_number, name FROM products WHERE model_number LIKE ? OR name LIKE ? LIMIT 20",
            (f"%{keyword}%", f"%{keyword}%"),
        )
        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return {"success": True, "data": items}

    @router.post("/products/batch", response_model=None)
    async def products_batch_add(data: dict):
        import datetime
        import sqlite3

        products_list = data.get("products") or []
        if not products_list:
            return {"success": True, "data": []}
        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        now = datetime.datetime.now().isoformat()
        rows = []
        for p in products_list:
            rows.append(
                (
                    "",
                    p.get("model_number", ""),
                    p.get("name", ""),
                    p.get("specification", ""),
                    float(p.get("price") or 0),
                    p.get("unit", ""),
                    now,
                    now,
                )
            )
        cur.executemany(
            "INSERT INTO products (source_file, model_number, name, specification, price, unit, created_at, updated_at) VALUES (?,?,?,?,?,?,?,?)",
            rows,
        )
        conn.commit()
        conn.close()
        return {"success": True, "data": []}

    @router.get("/customers", response_model=None)
    @router.get("/customers/", response_model=None)
    async def customers_all(page: int = 1, per_page: int = 20, keyword: str = ""):
        return await customers_list(page=page, per_page=per_page, keyword=keyword)

    @router.get("/customers/list", response_model=None)
    async def customers_list(
        page: int = 1,
        per_page: int = 20,
        keyword: str = "",
        purchase_unit: str = "",
    ):
        import sqlite3

        db_path = _db()
        if not db_path.exists():
            return {"success": True, "data": [], "total": 0}
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cond = []
        args = []
        if keyword:
            cond.append("(customer_name LIKE ? OR contact_person LIKE ? OR contact_phone LIKE ?)")
            args.extend([f"%{keyword}%", f"%{keyword}%", f"%{keyword}%"])
        if purchase_unit:
            cond.append("purchase_unit = ?")
            args.append(purchase_unit)
        where = " AND ".join(cond) if cond else "1=1"
        cur.execute(f"SELECT COUNT(*) FROM customers WHERE {where}", args)
        total = cur.fetchone()[0]
        offset = (page - 1) * per_page
        cur.execute(
            f"SELECT id, customer_name, contact_person, contact_phone, address, purchase_unit "
            f"FROM customers WHERE {where} ORDER BY id LIMIT ? OFFSET ?",
            [*args, per_page, offset],
        )
        items = [dict(r) for r in cur.fetchall()]
        conn.close()
        return {"success": True, "data": items, "total": total}

    @router.get("/customers/{customer_id}", response_model=None)
    async def customers_get(customer_id: int):
        import sqlite3

        db_path = _db()
        if not db_path.exists():
            return JSONResponse({"success": False, "error": "not found"}, status_code=404)
        conn = sqlite3.connect(str(db_path))
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT * FROM customers WHERE id = ?", (customer_id,))
        row = cur.fetchone()
        conn.close()
        if not row:
            return JSONResponse({"success": False, "error": "not found"}, status_code=404)
        return {"success": True, "data": dict(row)}

    @router.post("/customers", response_model=None)
    async def customers_add(data: dict):
        import datetime
        import sqlite3

        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        now = datetime.datetime.now().isoformat()
        cur.execute(
            "INSERT INTO customers (source_file, customer_name, contact_person, contact_phone, address, purchase_unit, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (
                data.get("source_file", ""),
                data.get("customer_name", ""),
                data.get("contact_person", ""),
                data.get("contact_phone", ""),
                data.get("address", ""),
                data.get("purchase_unit", ""),
                now,
                now,
            ),
        )
        new_id = cur.lastrowid
        conn.commit()
        conn.close()
        return {"success": True, "data": {"id": new_id}}

    @router.put("/customers/{customer_id}", response_model=None)
    async def customers_update(customer_id: int, data: dict):
        import datetime
        import sqlite3

        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        now = datetime.datetime.now().isoformat()
        cur.execute(
            "UPDATE customers SET customer_name=?, contact_person=?, contact_phone=?, address=?, purchase_unit=?, updated_at=? WHERE id=?",
            (
                data.get("customer_name", ""),
                data.get("contact_person", ""),
                data.get("contact_phone", ""),
                data.get("address", ""),
                data.get("purchase_unit", ""),
                now,
                customer_id,
            ),
        )
        conn.commit()
        conn.close()
        return {"success": True}

    @router.delete("/customers/{customer_id}", response_model=None)
    async def customers_delete(customer_id: int):
        import sqlite3

        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        cur.execute("DELETE FROM customers WHERE id = ?", (customer_id,))
        conn.commit()
        conn.close()
        return {"success": True}

    @router.post("/customers/batch-delete", response_model=None)
    async def customers_batch_delete(data: dict):
        import sqlite3

        ids = data.get("ids") or []
        if not ids:
            return {"success": True}
        db_path = _db()
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        placeholders = ",".join("?" * len(ids))
        cur.execute(f"DELETE FROM customers WHERE id IN ({placeholders})", ids)
        conn.commit()
        conn.close()
        return {"success": True}

    @router.post("/customers/import", response_model=None)
    async def customers_import(file: UploadFile = File(...)):
        import datetime
        import shutil
        import sqlite3
        import tempfile

        import openpyxl

        db_path = _db()
        suffix = Path(file.filename or "import.xlsx").suffix.lower()
        if suffix not in {".xlsx", ".xlsm", ".xls"}:
            return JSONResponse(
                {"success": False, "error": "unsupported file type"}, status_code=400
            )
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb.active
            headers = [str(c.value or "").strip() for c in ws[1]]
            name_idx = headers.index("客户名称") if "客户名称" in headers else 0
            contact_idx = headers.index("联系人") if "联系人" in headers else -1
            phone_idx = headers.index("电话") if "电话" in headers else -1
            addr_idx = headers.index("地址") if "地址" in headers else -1
            now = datetime.datetime.now().isoformat()
            conn = sqlite3.connect(str(db_path))
            cur = conn.cursor()
            count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
                if not name:
                    continue
                contact = (
                    str(row[contact_idx]) if contact_idx >= 0 and contact_idx < len(row) else ""
                )
                phone = str(row[phone_idx]) if phone_idx >= 0 and phone_idx < len(row) else ""
                addr = str(row[addr_idx]) if addr_idx >= 0 and addr_idx < len(row) else ""
                cur.execute(
                    "INSERT INTO customers (source_file, customer_name, contact_person, contact_phone, address, purchase_unit, created_at, updated_at) "
                    "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                    (file.filename or "", name, contact, phone, addr, "", now, now),
                )
                count += 1
            conn.commit()
            conn.close()
            return {"success": True, "imported": count}
        finally:
            Path(tmp_path).unlink(missing_ok=True)


__all__ = ["attach_personnel_crud_routes"]
