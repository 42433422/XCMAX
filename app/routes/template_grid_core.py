# -*- coding: utf-8 -*-
"""
Excel 模板网格与客户抬头解析（供聊天导入、工具链与 legacy 归档逻辑引用）。

历史上此模块在迁移中遗漏，导致 ``_customer_hint_from_preview_grid`` 等逻辑静默失败。
"""
from __future__ import annotations

import re
from typing import Any, List

# XCAGI 测试 monkeypatch 兼容（可选）
_TEMPLATE_SCOPE_REQUIRED_TERMS_CACHE = None

_MEASURE_UNIT_TOKENS = frozenset(
    {
        "件",
        "个",
        "只",
        "箱",
        "盒",
        "包",
        "袋",
        "瓶",
        "桶",
        "罐",
        "套",
        "组",
        "台",
        "条",
        "张",
        "支",
        "kg",
        "g",
        "吨",
        "升",
    }
)

# 「客户名称：xxx 报价日期：…」等单行抬头
_CUSTOMER_LABEL_RE = re.compile(
    r"(?:^|[\s\n])"
    r"(?:客户名称|购货单位|购买单位|订货单位|需方|甲方|买受人|采购单位)"
    r"(?:[（(][^）)]*[）)])?"
    r"\s*[：:]\s*"
    r"(?P<name>.+?)"
    r"(?=\s*(?:联系人|电话|传真|手机|报价日期|日期|订单|地址|开户|税号|邮编|乙方|供方|卖方)"
    r"\s*[：:]|\s*$)",
    re.UNICODE,
)

# 购货单位（简称）：全称
_PURCHASE_UNIT_PAREN_RE = re.compile(
    r"[购货购买]单位\s*[（(][^）)]*[）)]\s*[：:]\s*(?P<name>.+?)(?=\s+(?:联系人|电话|日期|购货|购买)|$)",
    re.UNICODE,
)

# 购货单位：…（无括号）
_PURCHASE_UNIT_PLAIN_RE = re.compile(
    r"[购货购买]单位\s*[：:]\s*(?P<name>.+?)(?=\s+(?:联系人|电话|日期|订单)|$)",
    re.UNICODE,
)


def _clean_customer_candidate(value: str) -> str:
    s = re.sub(r"\s+", " ", str(value or "").strip())
    s = s.strip(" ；;，,")
    return s


def _is_trivial_customer_token(text: str) -> bool:
    t = str(text or "").strip()
    if len(t) < 2:
        return True
    low = t.lower()
    if low in _MEASURE_UNIT_TOKENS:
        return True
    if re.fullmatch(r"[\d\s.,:/\-年月日]+", t):
        return True
    try:
        float(t.replace(",", ""))
        return True
    except Exception:
        pass
    return False


def _extract_inline_customer_hits_from_cell(text: str) -> List[str]:
    """从合并单元格/抬头行文本中解析客户公司名，返回按出现顺序的去重列表。"""
    raw = str(text or "").strip()
    if len(raw) < 3:
        return []
    hits: List[str] = []
    seen: set[str] = set()

    def _push(name: str) -> None:
        c = _clean_customer_candidate(name)
        if not c or _is_trivial_customer_token(c):
            return
        key = c.casefold()
        if key in seen:
            return
        seen.add(key)
        hits.append(c)

    for m in _CUSTOMER_LABEL_RE.finditer(raw):
        _push(m.group("name"))
    if not hits:
        m = _PURCHASE_UNIT_PAREN_RE.search(raw)
        if m:
            _push(m.group("name"))
    if not hits:
        m = _PURCHASE_UNIT_PLAIN_RE.search(raw)
        if m:
            _push(m.group("name"))

    return hits


def _extract_customer_hint_from_excel(file_path: str, sheet_name: str | None = None) -> str:
    """扫描表头区单元格，提取客户公司全称。"""
    try:
        from openpyxl import load_workbook
    except Exception:
        return ""

    try:
        wb = load_workbook(file_path, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return ""

            max_r = min(ws.max_row or 0, 22)
            max_c = min(ws.max_column or 0, 16)
            if max_r < 1 or max_c < 1:
                return ""

            for r in range(1, max_r + 1):
                parts: List[str] = []
                for c in range(1, max_c + 1):
                    v = ws.cell(r, c).value
                    if v is None:
                        continue
                    t = str(v).strip()
                    if not t:
                        continue
                    hits = _extract_inline_customer_hits_from_cell(t)
                    if hits:
                        return hits[0]
                    parts.append(t)
                joined = " ".join(parts).strip()
                if joined:
                    hits = _extract_inline_customer_hits_from_cell(joined)
                    if hits:
                        return hits[0]
            return ""
        finally:
            wb.close()
    except Exception:
        return ""


def _extract_rectangular_excel_preview(
    file_path: str,
    sheet_name: str | None = None,
    max_rows: int = 800,
    max_cols: int = 48,
) -> dict[str, Any]:
    """按列字母键读取矩形区域（parse_mode=rectangular），供导入重读。"""
    try:
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return {"fields": [], "sample_rows": [], "sheet_name": sheet_name or ""}

    try:
        wb = load_workbook(file_path, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return {"fields": [], "sample_rows": [], "sheet_name": ""}

            row_end = min(ws.max_row or 0, max(1, max_rows))
            col_end = min(ws.max_column or 0, max(1, max_cols))
            fields = [
                {"label": get_column_letter(c), "value": "", "type": "dynamic"}
                for c in range(1, col_end + 1)
            ]
            sample_rows: List[dict[str, Any]] = []
            for r in range(1, row_end + 1):
                row_data: dict[str, Any] = {}
                for c in range(1, col_end + 1):
                    row_data[get_column_letter(c)] = ws.cell(r, c).value
                if any(v is not None and str(v).strip() != "" for v in row_data.values()):
                    sample_rows.append(row_data)
            return {"fields": fields, "sample_rows": sample_rows, "sheet_name": ws.title}
        finally:
            wb.close()
    except Exception:
        return {"fields": [], "sample_rows": [], "sheet_name": sheet_name or ""}


def _extract_structured_excel_preview(
    file_path: str,
    sheet_name: str | None = None,
    sample_limit: int = 6,
    force_header_row_1based: int | None = None,
) -> dict[str, Any]:
    """结构化预览：默认同 ``document_templates_service``；可强制指定表头行。"""
    if force_header_row_1based is None:
        from app.services.document_templates_service import _extract_structured_excel_preview as _legacy

        return _legacy(file_path, sheet_name=sheet_name, sample_limit=sample_limit)

    try:
        from openpyxl import load_workbook
    except Exception:
        return {"fields": [], "sample_rows": [], "sheet_name": sheet_name or ""}

    hdr = int(force_header_row_1based)
    if hdr < 1:
        from app.services.document_templates_service import _extract_structured_excel_preview as _legacy

        return _legacy(file_path, sheet_name=sheet_name, sample_limit=sample_limit)

    try:
        wb = load_workbook(file_path, data_only=True)
        try:
            if sheet_name and sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            elif wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
            else:
                return {"fields": [], "sample_rows": [], "sheet_name": ""}

            max_col = min(ws.max_column or 0, 80)
            if max_col < 1:
                return {"fields": [], "sample_rows": [], "sheet_name": ws.title}

            header_entries: List[dict[str, Any]] = []
            for c in range(1, max_col + 1):
                value = ws.cell(hdr, c).value
                text = str(value).strip() if value is not None else ""
                if text:
                    header_entries.append({"name": text, "column_index": c})

            if len(header_entries) < 1:
                return {"fields": [], "sample_rows": [], "sheet_name": ws.title}

            fields = [{"label": h["name"], "value": "", "type": "dynamic"} for h in header_entries]
            sample_rows: List[dict[str, Any]] = []
            max_row = ws.max_row or 0
            for r in range(hdr + 1, min(max_row, hdr + sample_limit + 40) + 1):
                row_data: dict[str, Any] = {}
                has_non_empty = False
                for h in header_entries:
                    value = ws.cell(r, h["column_index"]).value
                    if value is not None and str(value).strip() != "":
                        has_non_empty = True
                    row_data[h["name"]] = value
                if has_non_empty:
                    sample_rows.append(row_data)
                if len(sample_rows) >= sample_limit:
                    break

            return {"fields": fields, "sample_rows": sample_rows, "sheet_name": ws.title}
        finally:
            wb.close()
    except Exception:
        return {"fields": [], "sample_rows": [], "sheet_name": sheet_name or ""}
e_limit:
                    break

            return {"fields": fields, "sample_rows": sample_rows, "sheet_name": ws.title}
        finally:
            wb.close()
    except Exception:
        return {"fields": [], "sample_rows": [], "sheet_name": sheet_name or ""}
