import openpyxl
from openpyxl import load_workbook
import os

folder = r'e:\FHD\424'

# 查找正确的文件名
files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and '考勤统计表' in f and not f.startswith('~$')]
print(f'找到模板文件：{files}')

if files:
    template_file = os.path.join(folder, files[0])
    
    # 查看模板文件
    print('\n=== 模板文件 ===')
    wb_template = load_workbook(template_file)
    print(f'工作表：{wb_template.sheetnames}')
    for sheet_name in wb_template.sheetnames:
        ws = wb_template[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        print(f'最大行：{ws.max_row}, 最大列：{ws.max_column}')
        print('表头 (前 3 行):')
        for row_idx in range(1, min(4, ws.max_row + 1)):
            values = []
            for col_idx in range(1, min(20, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value is not None:
                    values.append(f'{cell.coordinate}:{cell.value}')
            if values:
                print(f'  Row {row_idx}: {values}')
        
        # 查找姓名列的位置
        print('\n查找员工姓名区域 (第 4-30 行):')
        for row_idx in range(4, min(31, ws.max_row + 1)):
            for col_idx in range(1, min(10, ws.max_column + 1)):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value and isinstance(cell.value, str) and len(cell.value) >= 2:
                    # 查找中文字符
                    if any('\u4e00' <= c <= '\u9fff' for c in cell.value):
                        print(f'  Row {row_idx}, Col {col_idx} ({cell.coordinate}): {cell.value}')
                        break
