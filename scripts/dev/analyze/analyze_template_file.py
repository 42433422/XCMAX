"""查看模板文件结构"""
import subprocess
import tempfile
from pathlib import Path
from openpyxl import load_workbook
import shutil

# 使用 PowerShell 获取文件的完整路径
result = subprocess.run(
    ['powershell', '-Command', "Get-ChildItem 'e:\\FHD\\424' -Filter '考勤 -2026-3 月份考勤统计表.xlsx' | Select-Object -ExpandProperty FullName -First 1"],
    capture_output=True,
    text=True,
    encoding='utf-8'
)

file_path = result.stdout.strip()
print(f"文件路径：{file_path}")

if not Path(file_path).exists():
    print("文件不存在")
    exit(1)

print(f"文件大小：{Path(file_path).stat().st_size} 字节\n")

# 复制到临时文件
temp_file = Path(tempfile.mktemp(suffix='.xlsx'))
shutil.copy2(file_path, temp_file)

try:
    wb = load_workbook(str(temp_file), data_only=True)
    
    print(f"工作表：{wb.sheetnames}\n")
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"=== 工作表：{ws_name} ===")
        print(f"最大行：{ws.max_row}, 最大列：{ws.max_column}")
        merged = list(ws.merged_cells.ranges) if ws.merged_cells.ranges else []
        print(f"合并单元格：{len(merged)} 个")
        
        # 显示前 10 行
        print(f"\n前 10 行:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(15, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(val)[:20] if val else '')
            print(f"  行{row_idx}: {' | '.join(row_data)}")
        print()
    
    wb.close()
    
finally:
    if temp_file.exists():
        temp_file.unlink()
