import sys
from pathlib import Path
from openpyxl import load_workbook

# 使用 Unicode 路径
file_path = Path(r'e:\FHD\424\考勤转换输出.xlsx')

if not file_path.exists():
    print(f"文件不存在：{file_path}")
    print(f"当前目录：{Path.cwd()}")
    dir_424 = Path(r'e:\FHD\424')
    xlsx_files = list(dir_424.glob('*.xlsx'))
    print(f"424 目录内容：{xlsx_files}")
    sys.exit(1)

print(f"文件存在：{file_path}")
print(f"文件大小：{file_path.stat().st_size} 字节")

try:
    wb = load_workbook(str(file_path))
    print(f"工作表：{wb.sheetnames}")
    
    for i, ws in enumerate(wb.worksheets):
        print(f"\n工作表 {i+1}: {ws.title}")
        print(f"  最大行：{ws.max_row}, 最大列：{ws.max_column}")
        merged = list(ws.merged_cells.ranges) if ws.merged_cells.ranges else []
        print(f"  合并单元格数量：{len(merged)}")
        if merged:
            for m in merged[:5]:  # 只显示前 5 个
                print(f"    {m}")
        
        # 显示前几行数据
        print(f"  前 3 行数据:")
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, min(6, ws.max_column + 1))]
            print(f"    行{row_idx}: {row_data}")
            
except Exception as e:
    print(f"读取失败：{e}")
    import traceback
    traceback.print_exc()
