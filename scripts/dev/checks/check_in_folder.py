import openpyxl
from openpyxl import load_workbook
import os

folder = r'e:\FHD\424'
files = [f for f in os.listdir(folder) if f.endswith('.xlsx') and not f.startswith('~$')]

print('424 目录中的 xlsx 文件:')
for f in files:
    print(f'  {f}')

# 找到输出文件
output_file = None
for f in files:
    if '测试输出' in f:
        output_file = f
        break

if output_file:
    print(f'\n找到输出文件：{output_file}')
    full_path = os.path.join(folder, output_file)
    wb = load_workbook(full_path)
    print(f'工作表：{wb.sheetnames}')
    
    if "明细" in wb.sheetnames:
        print('\n✓ 成功！包含"明细"工作表')
        ws = wb["明细"]
        print(f'最大行：{ws.max_row}, 最大列：{ws.max_column}')
        
        # 查看前 10 行
        print('\n前 10 行:')
        for row_idx in range(1, 11):
            values = []
            for col_idx in range(1, 6):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    values.append(f'{cell.coordinate}={cell.value}')
            if values:
                print(f'  Row {row_idx}: {values}')
    else:
        print('\n✗ 错误：不包含"明细"工作表')
        print('这说明模板没有被正确加载和复制')
