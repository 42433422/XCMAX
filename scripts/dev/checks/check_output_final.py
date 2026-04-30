"""检查输出文件"""
from openpyxl import load_workbook
import os
import tempfile
import shutil

output_file = r'e:\FHD\424\考勤转换输出.xlsx'

print(f'检查输出文件：{output_file}')
print(f'文件存在：{os.path.exists(output_file)}')

if os.path.exists(output_file):
    # 复制到临时文件以避免中文路径问题
    temp_dir = tempfile.mkdtemp()
    temp_file = os.path.join(temp_dir, 'temp.xlsx')
    shutil.copy2(output_file, temp_file)

    try:
        wb = load_workbook(temp_file)
        print(f'工作表：{wb.sheetnames}')

        if "明细" in wb.sheetnames:
            ws = wb["明细"]
            print(f'\n✓ 成功！包含"明细"工作表')
            print(f'最大行：{ws.max_row}, 最大列：{ws.max_column}')

            # 检查有多少行有数据
            data_rows = 0
            employees = []
            for row_idx in range(4, ws.max_row + 1):
                name_cell = ws.cell(row=row_idx, column=3).value
                if name_cell:
                    data_rows += 1
                    if name_cell not in employees:
                        employees.append(name_cell)

            print(f'有数据的行数：{data_rows}')
            print(f'员工数量：{len(employees)}')

            # 显示前几个员工
            print(f'\n前 5 个员工:')
            shown_count = 0
            for row_idx in range(4, ws.max_row + 1):
                name_cell = ws.cell(row=row_idx, column=3).value
                if name_cell and shown_count < 5:
                    dept = ws.cell(row=row_idx, column=1).value
                    nature = ws.cell(row=row_idx, column=2).value
                    print(f'  Row {row_idx}: {dept} - {nature} - {name_cell}')
                    shown_count += 1
        elif "月度统计" in wb.sheetnames:
            print(f'\n✗ 问题：输出文件不包含"明细"工作表，而是"月度统计"')
            print(f'这说明模板没有被正确使用！')
            print(f'可能的原因：')
            print(f'  1. template_relpath 参数为空')
            print(f'  2. 模板路径无效')
            print(f'  3. resolve_workspace_excel 抛出异常被捕获')
        else:
            print(f'\n✗ 错误：工作表未知')
    finally:
        try:
            os.remove(temp_file)
            os.rmdir(temp_dir)
        except:
            pass
else:
    print('文件不存在')
