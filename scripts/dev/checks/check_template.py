import openpyxl
from openpyxl import load_workbook
import os

# 获取目录中的所有 xlsx 文件
folder = r'e:\FHD\424'
files = [f for f in os.listdir(folder) if f.endswith('.xlsx')]
print('424 目录中的 xlsx 文件:')
for f in files:
    print(f'  {f}')

# 查找模板文件
template_file = None
for f in files:
    if '考勤统计表' in f and not f.startswith('~$'):
        template_file = f
        break

if template_file:
    print(f'\n找到模板文件：{template_file}')
    full_path = os.path.join(folder, template_file)
    
    wb = load_workbook(full_path)
    print(f'工作表：{wb.sheetnames}')
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f'\n=== Sheet: {sheet_name} ===')
        print(f'合并单元格数量：{len(ws.merged_cells.ranges)}')
        print(f'最大行：{ws.max_row}, 最大列：{ws.max_column}')
        print('前 30 行内容:')
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(30, ws.max_row)), 1):
            values = []
            for cell in row:
                if cell.value is not None:
                    values.append(f'{cell.coordinate}:{cell.value}')
            if values:
                print(f'  Row {row_idx}: {values}')
        print('合并单元格 (前 20 个):')
        merged_list = list(ws.merged_cells.ranges)
        for i, merge in enumerate(merged_list[:20]):
            print(f'  {merge}')
else:
    print('未找到模板文件')
