import sys
from pathlib import Path
from openpyxl import load_workbook

# 检查模板文件
template_path = Path(r'e:\FHD\424\考勤 -2026-3 月份考勤统计表.xlsx')

if not template_path.exists():
    print(f"模板文件不存在：{template_path}")
    sys.exit(1)

print(f"模板文件：{template_path}")
print(f"文件大小：{template_path.stat().st_size} 字节")

wb = load_workbook(str(template_path))
print(f"\n工作表：{wb.sheetnames}")

for i, ws in enumerate(wb.worksheets):
    print(f"\n工作表 {i+1}: {ws.title}")
    print(f"  最大行：{ws.max_row}, 最大列：{ws.max_column}")
    merged = list(ws.merged_cells.ranges) if ws.merged_cells.ranges else []
    print(f"  合并单元格数量：{len(merged)}")
    if merged:
        for m in merged:
            print(f"    {m}")
    
    # 显示前 5 行数据
    print(f"  前 5 行数据:")
    for row_idx in range(1, min(6, ws.max_row + 1)):
        row_data = [ws.cell(row=row_idx, column=col).value for col in range(1, min(10, ws.max_column + 1))]
        print(f"    行{row_idx}: {row_data}")
