"""查看模板文件结构 - 使用复制方式"""
import os
import shutil
from pathlib import Path
from openpyxl import load_workbook

dir_path = Path(r'e:\FHD\424')
file_name = '考勤 -2026-3 月份考勤统计表.xlsx'

# 找到文件
target_file = None
for f in dir_path.iterdir():
    if f.name == file_name:
        target_file = f
        break

if not target_file:
    print(f"找不到文件：{file_name}")
    exit(1)

print(f"找到文件：{target_file}")
print(f"文件大小：{target_file.stat().st_size} 字节\n")

# 复制到临时文件
import tempfile
temp_file = Path(tempfile.mktemp(suffix='.xlsx'))
shutil.copy2(str(target_file), str(temp_file))

try:
    wb = load_workbook(str(temp_file), data_only=True)
    
    print(f"工作表：{wb.sheetnames}\n")
    
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        print(f"工作表：{ws_name}")
        print(f"  最大行：{ws.max_row}, 最大列：{ws.max_column}")
        merged = list(ws.merged_cells.ranges) if ws.merged_cells.ranges else []
        print(f"  合并单元格：{len(merged)} 个")
        
        # 显示前 10 行
        print(f"  前 10 行:")
        for row_idx in range(1, min(11, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(15, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(val)[:15] if val else '')
            print(f"    行{row_idx}: {' | '.join(row_data)}")
        print()
    
    wb.close()
    
finally:
    # 清理临时文件
    if temp_file.exists():
        temp_file.unlink()
        print(f"已清理临时文件")
