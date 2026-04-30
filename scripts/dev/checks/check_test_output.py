"""检查测试输出"""
from openpyxl import load_workbook
import os

output_file = r'e:\FHD\424\测试_直接转换.xlsx'

if os.path.exists(output_file):
    wb = load_workbook(output_file)
    print(f'工作表：{wb.sheetnames}')

    if "明细" in wb.sheetnames:
        ws = wb["明细"]
        print(f'\n✓ 成功！包含"明细"工作表')
        print(f'最大行：{ws.max_row}, 最大列：{ws.max_column}')

        # 检查有多少行有数据
        data_rows = 0
        for row_idx in range(4, min(ws.max_row + 1, 50)):
            name_cell = ws.cell(row=row_idx, column=3).value
            if name_cell:
                data_rows += 1
                dept = ws.cell(row=row_idx, column=1).value
                nature = ws.cell(row=row_idx, column=2).value
                if data_rows <= 5:
                    print(f'  Row {row_idx}: {dept} - {nature} - {name_cell}')

        print(f'\n前 50 行中有数据的行数：{data_rows}')
    else:
        print(f'\n✗ 问题：没有"明细"工作表')
else:
    print('文件不存在')
