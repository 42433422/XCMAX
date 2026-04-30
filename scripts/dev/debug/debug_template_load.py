import openpyxl
from openpyxl import load_workbook
import os

# 直接测试加载模板
template_file = r'e:\FHD\424\考勤 -2026-3 月份考勤统计表.xlsx'

print(f'文件是否存在：{os.path.exists(template_file)}')
print(f'文件大小：{os.path.getsize(template_file)} bytes')

try:
    wb = load_workbook(template_file)
    print(f'成功加载！工作表：{wb.sheetnames}')
except Exception as e:
    print(f'加载失败：{e}')

# 检查生成的文件
output_file = r'e:\FHD\424\测试输出.xlsx'
print(f'\n生成文件是否存在：{os.path.exists(output_file)}')
if os.path.exists(output_file):
    wb_out = load_workbook(output_file)
    print(f'生成文件的工作表：{wb_out.sheetnames}')
