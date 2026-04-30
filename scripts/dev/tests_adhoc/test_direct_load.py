import os
from openpyxl import load_workbook

folder = r'e:\FHD\424'
file_name = '考勤 -2026-3 月份考勤统计表.xlsx'

# 使用 os.listdir 找到文件
dir_contents = os.listdir(folder)
matching_files = [f for f in dir_contents if f == file_name]

print(f'查找的文件名：{file_name}')
print(f'匹配的文件：{matching_files}')

if matching_files:
    found_file = matching_files[0]
    full_path = os.path.join(folder, found_file)
    print(f'\n完整路径：{full_path}')
    
    # 尝试加载
    try:
        wb = load_workbook(full_path)
        print(f'✓ 加载成功！工作表：{wb.sheetnames}')
    except Exception as e:
        print(f'✗ 加载失败：{e}')
else:
    print('未找到匹配的文件')
