"""测试恢复后的完整功能"""
import os
from pathlib import Path
from openpyxl import load_workbook

# 设置 WORKSPACE_ROOT
os.environ['WORKSPACE_ROOT'] = r'e:\FHD'

from app.shell.taiyangniao_attendance.convert import convert_dingtalk_file

# 测试文件
source_file = Path(r'e:\FHD\424\钉钉导出来的考勤数据.xlsx')
output_file = Path(r'e:\FHD\424\测试完整功能输出.xlsx')

print(f"源文件：{source_file}")
print(f"输出文件：{output_file}")

# 执行转换（不使用模板）
print("\n执行转换（不使用模板）...")
try:
    result = convert_dingtalk_file(
        source_file,
        output_file,
        month="2026-03",
        sheet="打卡时间",
        header_row=2,
        template_path=None,  # 不使用模板
    )
    print(f"✓ 转换成功！")
    print(f"  输入行数：{result['rows_in']}")
    print(f"  输出行数：{result['rows_stats']}")
    print(f"  输出路径：{result['output']}")
    
    # 检查输出文件
    print("\n输出文件信息:")
    wb_out = load_workbook(str(output_file))
    print(f"工作表：{wb_out.sheetnames}")
    
    for ws_name in wb_out.sheetnames:
        ws = wb_out[ws_name]
        print(f"\n工作表：{ws_name}")
        print(f"  最大行：{ws.max_row}, 最大列：{ws.max_column}")
        
        # 显示前 3 行
        print(f"  前 3 行:")
        for row_idx in range(1, min(4, ws.max_row + 1)):
            row_data = []
            for col_idx in range(1, min(10, ws.max_column + 1)):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_data.append(str(val)[:20] if val else '')
            print(f"    行{row_idx}: {' | '.join(row_data)}")
    
    wb_out.close()
    
    print(f"\n✓ 测试完成！")
    
except Exception as e:
    print(f"✗ 转换失败：{e}")
    import traceback
    traceback.print_exc()
