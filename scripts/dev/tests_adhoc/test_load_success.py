import os
from openpyxl import load_workbook

folder = r'e:\FHD\424'
# 直接使用 os.listdir 返回的文件名
file_name = '考勤 -2026-3 月份考勤统计表.xlsx'

full_path = os.path.join(folder, file_name)
print(f'文件路径：{full_path}')
print(f'文件存在：{os.path.exists(full_path)}')

# 尝试加载
try:
    wb = load_workbook(full_path)
    print(f'✓ 加载成功！工作表：{wb.sheetnames}')
    
    # 保存到新文件
    output_path = os.path.join(folder, '测试模板加载.xlsx')
    wb.save(output_path)
    print(f'✓ 保存成功！')
    
    # 验证
    wb2 = load_workbook(output_path)
    print(f'输出文件工作表：{wb2.sheetnames}')
except Exception as e:
    print(f'✗ 失败：{e}')
    import traceback
    traceback.print_exc()
