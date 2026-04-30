import openpyxl
from openpyxl import load_workbook
import os

output_file = r'e:\FHD\424\测试输出.xlsx'

print('=== 验证生成的文件 ===')
wb = load_workbook(output_file)
print(f'工作表：{wb.sheetnames}')

ws = wb["明细"]
print(f'\n=== Sheet: 明细 ===')
print(f'最大行：{ws.max_row}, 最大列：{ws.max_column}')

# 查看前 20 行
print('\n前 20 行数据:')
for row_idx in range(1, min(21, ws.max_row + 1)):
    values = []
    for col_idx in range(1, min(10, ws.max_column + 1)):  # 只显示前 10 列
        cell = ws.cell(row=row_idx, column=col_idx)
        if cell.value is not None:
            cell_type = '公式' if isinstance(cell.value, str) and cell.value.startswith('=') else '数据'
            values.append(f'{cell.coordinate}:{cell.value} [{cell_type}]')
    if values:
        print(f'  Row {row_idx}: {values}')

# 检查员工数据
print('\n=== 检查员工数据区域 ===')
for row_idx in range(4, min(16, ws.max_row + 1)):
    dept = ws.cell(row=row_idx, column=1).value
    name = ws.cell(row=row_idx, column=3).value
    time_period = ws.cell(row=row_idx, column=4).value
    if name:
        print(f'Row {row_idx}: 部门={dept}, 姓名={name}, 时段={time_period}')
    
    # 检查第 1 天的数据
    day1_mark = ws.cell(row=row_idx, column=7).value  # G 列
    day1_hours = ws.cell(row=row_idx, column=8).value  # H 列
    if day1_mark:
        print(f'         1 日：标记={day1_mark}, 工时={day1_hours}')
