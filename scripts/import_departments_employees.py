"""
从考勤 Excel 提取部门和人员，导入 taiyangniao_pro.db（考勤表 + products/customers，供前端列表）。

支持两种工作表：
  - 「每日统计」：钉钉导出（姓名在第 1 列，部门在第 3 列…）
  - 「明细」：固定模板统计表（部门 / 性质 / 姓名 在前 3 列，自第 4 行起）

用法:
    python scripts/import_departments_employees.py --input 424/考勤-2026-3月份考勤统计表.xlsx
    python scripts/import_departments_employees.py --input <xlsx> --db data/mod_dbs/taiyangniao_pro.db --no-sync-ui
"""

from __future__ import annotations

import argparse
import json
import sqlite3
from collections import OrderedDict
from datetime import datetime
from pathlib import Path
import sys

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))


def _ensure_schema(conn: sqlite3.Connection) -> None:
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance_import_batches (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            month_label TEXT NOT NULL,
            rows_in INTEGER NOT NULL,
            rows_written INTEGER NOT NULL,
            imported_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance_daily_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            month_label TEXT NOT NULL,
            source_row INTEGER NOT NULL,
            employee_name TEXT NOT NULL,
            attendance_group TEXT NOT NULL,
            department TEXT NOT NULL,
            employee_no TEXT NOT NULL,
            position TEXT NOT NULL,
            user_id TEXT NOT NULL,
            work_date TEXT NOT NULL,
            shift_name TEXT NOT NULL,
            daily_times_json TEXT NOT NULL,
            raw_times_json TEXT NOT NULL,
            all_times_json TEXT NOT NULL,
            leave_hours REAL NOT NULL,
            absent_days REAL NOT NULL,
            late_count_hint REAL NOT NULL,
            early_count_hint REAL NOT NULL,
            missing_card_count REAL NOT NULL,
            notes_json TEXT NOT NULL,
            imported_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE UNIQUE INDEX IF NOT EXISTS ux_attendance_source_row
        ON attendance_daily_records (source_file, source_row)
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS ix_attendance_employee_date
        ON attendance_daily_records (employee_name, work_date)
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance_departments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            department TEXT NOT NULL,
            main_department TEXT NOT NULL,
            attendance_group TEXT NOT NULL,
            UNIQUE(source_file, department, attendance_group)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS attendance_employees (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            employee_name TEXT NOT NULL,
            department TEXT NOT NULL,
            main_department TEXT NOT NULL,
            attendance_group TEXT NOT NULL,
            employee_no TEXT NOT NULL,
            position TEXT NOT NULL,
            user_id TEXT NOT NULL,
            UNIQUE(source_file, employee_name, department)
        )
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS ix_employees_name
        ON attendance_employees (employee_name)
    """)
    conn.execute("""
        CREATE INDEX IF NOT EXISTS ix_departments_dept
        ON attendance_departments (department)
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS products (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL DEFAULT '',
            model_number TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL DEFAULT '',
            specification TEXT NOT NULL DEFAULT '',
            price REAL NOT NULL DEFAULT 0,
            unit TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS customers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL DEFAULT '',
            customer_name TEXT NOT NULL DEFAULT '',
            contact_person TEXT NOT NULL DEFAULT '',
            contact_phone TEXT NOT NULL DEFAULT '',
            address TEXT NOT NULL DEFAULT '',
            purchase_unit TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
    """)


def _cell_first_line(val) -> str:
    s = str(val or "").strip()
    if not s:
        return ""
    return s.split("\n")[0].strip()


def _parse_dingtalk_daily_sheet(ws) -> tuple[list[dict], list[dict]]:
    """钉钉「每日统计」：A 姓名、B 考勤组、C 部门、D 一级部门…"""
    seen_employees: dict[str, dict] = OrderedDict()
    seen_departments: dict[str, dict] = OrderedDict()

    for row in ws.iter_rows(min_row=4, values_only=True):
        name = str(row[0] or "").strip()
        if not name:
            continue

        group = str(row[1] or "").strip()
        dept = str(row[2] or "").strip()
        main_dept = str(row[3] or "").strip()
        emp_no = str(row[4] or "").strip()
        position = str(row[5] or "").strip()
        uid = str(row[6] or "").strip()

        emp_key = (name, dept)
        if emp_key not in seen_employees:
            seen_employees[emp_key] = {
                "name": name,
                "group": group,
                "dept": dept,
                "main_dept": main_dept,
                "emp_no": emp_no,
                "position": position,
                "uid": uid,
            }

        dept_key = (dept, main_dept, group)
        if dept_key not in seen_departments:
            seen_departments[dept_key] = {
                "department": dept,
                "main_department": main_dept,
                "attendance_group": group,
            }

    return list(seen_departments.values()), list(seen_employees.values())


def _parse_attendance_detail_sheet(ws) -> tuple[list[dict], list[dict]]:
    """固定模板「明细」：A 部门、B 性质、C 姓名（第 4 行起为数据块首行）。"""
    seen_employees: dict[tuple[str, str], dict] = OrderedDict()
    seen_departments: dict[tuple[str, str, str], dict] = OrderedDict()

    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row:
            continue
        dept = _cell_first_line(row[0])
        nature = _cell_first_line(row[1])
        name = _cell_first_line(row[2])
        if not name:
            continue
        if name == "姓名" or dept == "部门":
            continue

        emp_key = (name, dept)
        if emp_key not in seen_employees:
            seen_employees[emp_key] = {
                "name": name,
                "group": nature,
                "dept": dept,
                "main_dept": dept,
                "emp_no": "",
                "position": nature,
                "uid": "",
            }

        dept_key = (dept, dept, nature)
        if dept_key not in seen_departments:
            seen_departments[dept_key] = {
                "department": dept,
                "main_department": dept,
                "attendance_group": nature,
            }

    return list(seen_departments.values()), list(seen_employees.values())


def _parse_workbook(excel_path: Path) -> tuple[list[dict], list[dict], str]:
    import openpyxl

    wb = openpyxl.load_workbook(str(excel_path), read_only=True, data_only=True)
    try:
        names = wb.sheetnames
        if "每日统计" in names:
            return (*_parse_dingtalk_daily_sheet(wb["每日统计"]), "dingtalk")
        if "明细" in names:
            return (*_parse_attendance_detail_sheet(wb["明细"]), "mingxi")
        raise ValueError(
            f"未找到「每日统计」或「明细」工作表。实际工作表: {names!r} — {excel_path}"
        )
    finally:
        wb.close()


def _sync_products_customers(
    conn: sqlite3.Connection,
    source_file: str,
    employees: list[dict],
    _departments: list[dict],
) -> tuple[int, int]:
    """与 mod 前端一致：products=人员行，customers=部门行（purchase_unit 置空）。"""
    now = datetime.now().isoformat(timespec="seconds")
    conn.execute("DELETE FROM products WHERE source_file = ?", (source_file,))
    conn.execute("DELETE FROM customers WHERE source_file = ?", (source_file,))

    prod_rows = 0
    for e in employees:
        dept = (e.get("dept") or "").strip()
        name = (e.get("name") or "").strip()
        grp = (e.get("group") or "").strip()
        model_number = f"{dept}::{name}" if dept else name
        # unit 必须与产品页「产品单位」下拉一致（该下拉来自部门/客户名），否则按部门筛选会查不到行
        spec = grp or ""
        conn.execute(
            """
            INSERT INTO products (source_file, model_number, name, specification, price, unit, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (source_file, model_number, name, spec, 0.0, dept, now, now),
        )
        prod_rows += 1

    cust_rows = 0
    seen_dept: set[str] = set()
    for e in employees:
        dn = (e.get("dept") or "").strip()
        if not dn or dn in seen_dept:
            continue
        seen_dept.add(dn)
        conn.execute(
            """
            INSERT INTO customers (source_file, customer_name, contact_person, contact_phone, address, purchase_unit, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (source_file, dn, "", "", "", "", now, now),
        )
        cust_rows += 1

    return prod_rows, cust_rows


def import_departments_and_employees(
    excel_path: Path,
    db_path: Path,
    *,
    sync_ui_tables: bool = True,
) -> tuple[int, int, int, int]:
    departments, employees, _kind = _parse_workbook(excel_path)
    source_file = str(excel_path.resolve())

    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    try:
        _ensure_schema(conn)
        conn.execute("BEGIN")

        conn.execute(
            "DELETE FROM attendance_departments WHERE source_file = ?",
            (source_file,),
        )
        conn.execute(
            "DELETE FROM attendance_employees WHERE source_file = ?",
            (source_file,),
        )

        dept_rows = 0
        for d in departments:
            conn.execute(
                """
                INSERT OR IGNORE INTO attendance_departments
                    (source_file, department, main_department, attendance_group)
                VALUES (?, ?, ?, ?)
                """,
                (
                    source_file,
                    d["department"],
                    d["main_department"],
                    d["attendance_group"],
                ),
            )
            dept_rows += 1

        emp_rows = 0
        for e in employees:
            conn.execute(
                """
                INSERT OR IGNORE INTO attendance_employees
                    (source_file, employee_name, department, main_department,
                     attendance_group, employee_no, position, user_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    source_file,
                    e["name"],
                    e["dept"],
                    e["main_dept"],
                    e["group"],
                    e["emp_no"],
                    e["position"],
                    e["uid"],
                ),
            )
            emp_rows += 1

        prod_rows = cust_rows = 0
        if sync_ui_tables:
            prod_rows, cust_rows = _sync_products_customers(conn, source_file, employees, departments)

        conn.commit()
        return dept_rows, emp_rows, prod_rows, cust_rows
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def main() -> None:
    parser = argparse.ArgumentParser(
        description="从钉钉考勤报表提取部门和人员，导入 taiyangniao_pro 数据库。"
    )
    parser.add_argument(
        "--input",
        default=str(ROOT / "424" / "考勤-2026-3月份考勤统计表.xlsx"),
        help="考勤报表路径（钉钉用「每日统计」，固定模板用「明细」）",
    )
    parser.add_argument(
        "--db",
        default=str(ROOT / "data" / "mod_dbs" / "taiyangniao_pro.db"),
        help="SQLite 数据库路径",
    )
    parser.add_argument(
        "--no-sync-ui",
        action="store_true",
        help="不同步写入 products/customers（仅 attendance_* 表）",
    )
    args = parser.parse_args()

    excel_path = Path(args.input).resolve()
    db_path = Path(args.db).resolve()

    if not excel_path.exists():
        raise FileNotFoundError(f"Excel 文件不存在: {excel_path}")

    dept_count, emp_count, prod_count, cust_count = import_departments_and_employees(
        excel_path,
        db_path,
        sync_ui_tables=not args.no_sync_ui,
    )
    print(f"attendance_departments: {dept_count} 条")
    print(f"attendance_employees: {emp_count} 条")
    if not args.no_sync_ui:
        print(f"products（人员列表）: {prod_count} 条")
        print(f"customers（部门列表）: {cust_count} 条")
    print(f"数据库: {db_path}")


if __name__ == "__main__":
    main()
