"""Asset-driven employee_pack generation helpers.

This module upgrades the workbench "make employee" path from a pure prompt
manifest generator into a file-aware direct_python pack builder.  Uploaded
templates and examples stay as real files; the LLM only receives structured
summaries and generates code against a stable runtime scaffold.
"""

from __future__ import annotations

import ast
import io
import json
import os
import py_compile
import re
import shutil
import tempfile
import zipfile
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

from modstore_server.csv_tabular_runtime import (
    build_csv_generate_rule_spec,
    build_csv_read_rule_spec,
    is_csv_full_read,
    is_csv_generate,
    render_csv_generate_convert_module,
    render_csv_read_convert_module,
)
from modstore_server.employee_ai_scaffold import parse_employee_pack_llm_json
from modstore_server.employee_pack_blueprints_template import render_employee_pack_blueprints_py
from modstore_server.excel_tabular_runtime import (
    build_excel_generate_rule_spec,
    build_excel_read_rule_spec,
    is_excel_full_read,
    is_excel_generate,
    render_excel_generate_convert_module,
    render_excel_read_convert_module,
)
from modstore_server.json_report_runtime import (
    build_json_quant_report_rule_spec,
    is_json_quant_report,
    render_json_report_convert_module,
)
from modstore_server.kitten_chart_runtime import (
    build_kitten_chart_rule_spec,
    is_kitten_chart_viz,
    render_kitten_chart_convert_module,
)
from modstore_server.llm_key_resolver import (
    OAI_COMPAT_OPENAI_STYLE_PROVIDERS,
    resolve_api_key,
    resolve_base_url,
)
from modstore_server.mod_ai_scaffold import normalize_mod_id
from modstore_server.mod_employee_impl_scaffold import sanitize_employee_stem
from modstore_server.mod_scaffold_runner import (
    chat_dispatch,
    import_zip,
    modstore_library_path,
    resolve_llm_provider_model_auto,
)
from modstore_server.models import CatalogItem, User
from modstore_server.pdf_extract_runtime import (
    build_pdf_generate_rule_spec,
    build_pdf_read_rule_spec,
    is_pdf_full_read,
    is_pdf_generate,
    render_pdf_generate_convert_module,
    render_pdf_read_convert_module,
)
from modstore_server.ppt_extract_runtime import (
    build_ppt_generate_rule_spec,
    build_ppt_read_rule_spec,
    is_ppt_full_read,
    is_ppt_generate,
    render_ppt_generate_convert_module,
    render_ppt_read_convert_module,
)
from modstore_server.txt_extract_runtime import (
    build_txt_generate_rule_spec,
    build_txt_read_rule_spec,
    is_txt_full_read,
    is_txt_generate,
    render_txt_generate_convert_module,
    render_txt_read_convert_module,
)
from modstore_server.word_extract_runtime import (
    build_word_extract_rule_spec,
    is_word_full_extract,
    render_word_fallback_convert_module,
)
from modstore_server.word_generate_runtime import (
    build_word_generate_rule_spec,
    is_word_generate,
    render_word_generate_convert_module,
)

EXCEL_SUFFIXES = {".xlsx", ".xlsm", ".xls"}
TEXT_SUFFIXES = {".txt", ".md", ".json", ".csv", ".py", ".yaml", ".yml"}

_LLM_CHAIN_MARKERS = re.compile(
    r"初始想法|澄清对话|<<<PLAN_|需要简短|不超过\d+个|字符计数|计算字符|"
    r"总结：\s*-|现在，构建|确保不泄露|从用户指令看|作为需求摘要|"
    r"输出格式必须严格|不能输出流程图|不能输出选项|不能输出执行清单|"
    r"当前制作类型|这可能意味着|为安全起见|或许更简洁|标准中文字符|"
    r"计算字符数|字符数.*不超过|以上.*字符",
)
_LLM_CHAIN_BLOCK_START = re.compile(r"【.*?(想法|对话|澄清|规划|分析)】")
_LLM_CHAIN_BLOCK_END = re.compile(r"【.*?(助手|用户|回答|结果|输出)】|<<<END")


def _clean_brief_for_description(brief: str, max_len: int = 200) -> str:
    if not brief:
        return ""
    first_sentence = re.split(r"[。！？\n]", brief)[0].strip()
    if (
        first_sentence
        and not _LLM_CHAIN_MARKERS.search(first_sentence)
        and not _LLM_CHAIN_BLOCK_START.search(first_sentence)
    ):
        return first_sentence[:max_len]
    lines = brief.splitlines()
    clean: list[str] = []
    in_chain_block = False
    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        if _LLM_CHAIN_BLOCK_START.search(stripped):
            in_chain_block = True
            continue
        if _LLM_CHAIN_BLOCK_END.search(stripped):
            in_chain_block = False
            continue
        if in_chain_block:
            continue
        if _LLM_CHAIN_MARKERS.search(stripped):
            continue
        if re.match(r"^[-*]\s", stripped) and len(stripped) < 15:
            continue
        if re.match(r"^\d+\.\s", stripped) and len(stripped) < 20:
            continue
        clean.append(stripped)
    if not clean:
        for seg in re.split(r"[。！？]", brief):
            s = seg.strip()
            if (
                s
                and not _LLM_CHAIN_MARKERS.search(s)
                and not _LLM_CHAIN_BLOCK_START.search(s)
                and len(s) > 5
            ):
                return s[:max_len]
        for seg in re.split(r"[\n,，;；]", brief):
            s = seg.strip()
            if (
                s
                and not _LLM_CHAIN_MARKERS.search(s)
                and not _LLM_CHAIN_BLOCK_START.search(s)
                and len(s) > 3
            ):
                return s[:max_len]
        return brief[:max_len].strip()
    return " ".join(clean)[:max_len]


def _safe_basename(name: str, fallback: str = "asset.bin") -> str:
    base = Path(name or "").name
    if not base or base in {".", ".."}:
        return fallback
    if ".." in base or "/" in base or "\\" in base:
        return fallback
    return base[:180]


def _classify_asset(filename: str) -> str:
    low = filename.lower()
    suffix = Path(filename).suffix.lower()
    if suffix in EXCEL_SUFFIXES:
        if any(k in filename for k in ("模板", "template", "样板")):
            return "template"
        if any(k in filename for k in ("输出", "结果", "expected", "answer")):
            return "expected_output"
        return "example_input"
    if suffix == ".py":
        return "reference_code"
    if suffix in TEXT_SUFFIXES:
        return "rules"
    return "asset"


def _runtime_module_name(pack_id: str) -> str:
    raw = re.sub(r"[^a-z0-9_]+", "_", (pack_id or "employee").lower()).strip("_")
    if not raw:
        raw = "employee"
    if raw[0].isdigit():
        raw = "e_" + raw
    return f"{raw}_runtime"


def _runtime_package_name(pack_id: str, employee_id: str = "") -> str:
    base = employee_id or pack_id
    raw = re.sub(r"[^a-z0-9_]+", "_", (base or "employee").lower()).strip("_")
    if not raw:
        raw = "employee"
    if raw.endswith("_employee"):
        raw = raw[: -len("_employee")] or raw
    if raw[0].isdigit():
        raw = "e_" + raw
    return raw


DOC_SUFFIXES = {".docx", ".doc", ".pdf", ".rtf"}


def _infer_accepted_extensions(asset_manifest: Optional[Dict[str, Any]] = None) -> List[str]:
    if not isinstance(asset_manifest, dict):
        return []
    suffixes: set = set()
    for item in asset_manifest.get("assets") or []:
        if not isinstance(item, dict):
            continue
        s = str(item.get("suffix") or "").lower()
        if s:
            suffixes.add(s)
    if not suffixes:
        return []
    if suffixes & EXCEL_SUFFIXES:
        return sorted(suffixes & EXCEL_SUFFIXES) + sorted(suffixes - EXCEL_SUFFIXES)
    if suffixes & DOC_SUFFIXES:
        return sorted(suffixes & DOC_SUFFIXES) + sorted(suffixes - DOC_SUFFIXES)
    return sorted(suffixes)


def _infer_asset_runtime_kind(brief: str, asset_manifest: Optional[Dict[str, Any]] = None) -> str:
    """Classify by asset shape, not by a customer-specific package name."""
    if is_csv_generate(brief):
        return "csv_generate"
    if is_csv_full_read(brief):
        return "csv_full_read"
    if is_excel_generate(brief):
        return "excel_generate"
    if is_excel_full_read(brief):
        return "excel_full_read"
    if is_pdf_generate(brief):
        return "pdf_generate"
    if is_pdf_full_read(brief):
        return "pdf_full_read"
    if is_json_quant_report(brief):
        return "json_quant_report"
    if is_kitten_chart_viz(brief):
        return "kitten_chart_viz"
    if is_ppt_generate(brief):
        return "ppt_generate"
    if is_ppt_full_read(brief):
        return "ppt_full_read"
    if is_txt_generate(brief):
        return "txt_generate"
    if is_txt_full_read(brief):
        return "txt_full_read"
    if is_word_generate(brief):
        return "word_generate"
    if is_word_full_extract(brief):
        return "word_full_extract"
    has_excel = False
    has_doc = False
    has_rules = False
    has_reference_code = False
    if isinstance(asset_manifest, dict):
        for item in asset_manifest.get("assets") or []:
            if not isinstance(item, dict):
                continue
            suffix = str(item.get("suffix") or "").lower()
            kind = str(item.get("kind") or "")
            has_excel = has_excel or suffix in EXCEL_SUFFIXES
            has_doc = has_doc or suffix in DOC_SUFFIXES
            has_rules = has_rules or kind == "rules"
            has_reference_code = has_reference_code or kind == "reference_code"
    text = brief or ""
    _contract_keywords = ("合同", "法务", "合规", "审核", "条款", "contract", "legal", "compliance")
    _doc_keywords = ("文档", "报告", "方案", "标书", "简历", "document", "report", "proposal")
    if has_reference_code:
        return "reference_python_transform"
    if has_doc or any(k in text for k in _contract_keywords):
        if any(k in text for k in _contract_keywords):
            return "contract_doc_review"
        return "doc_template_transform"
    if has_excel and (has_rules or any(k in text for k in ("规则", "模板", "转换", "考勤"))):
        return "excel_rules_transform"
    if has_excel:
        return "generic_excel_transform"
    return "generic_file_transform"


def _read_text_preview(content: bytes, limit: int = 4000) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(enc)[:limit]
        except UnicodeDecodeError:
            continue
    return ""


def _excel_summary(path: Path) -> Dict[str, Any]:
    summary: Dict[str, Any] = {"ok": False, "sheets": [], "error": ""}
    try:
        from openpyxl import load_workbook
    except Exception as exc:  # noqa: BLE001
        summary["error"] = f"openpyxl unavailable: {exc}"
        return summary
    try:
        wb = load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:  # noqa: BLE001
        summary["error"] = str(exc)[:500]
        return summary
    sheets: List[Dict[str, Any]] = []
    for ws in wb.worksheets[:12]:
        formulas = 0
        non_empty = 0
        header_candidates: List[List[str]] = []
        max_row = min(ws.max_row or 0, 30)
        max_col = min(ws.max_column or 0, 30)
        for r in range(1, max_row + 1):
            vals: List[str] = []
            filled = 0
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip():
                    filled += 1
                    vals.append(str(v).strip()[:40])
                    if isinstance(v, str) and v.startswith("="):
                        formulas += 1
            non_empty += filled
            if filled >= 2 and len(header_candidates) < 5:
                header_candidates.append(vals[:12])
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_ranges": len(list(ws.merged_cells.ranges)),
                "formula_cells_sampled": formulas,
                "non_empty_sampled": non_empty,
                "header_candidates": header_candidates,
            }
        )
    summary["ok"] = True
    summary["sheets"] = sheets
    return summary


def prepare_employee_assets(
    *,
    session_id: str,
    user_id: int,
    raw_files: List[Dict[str, Any]],
    repo_root: Optional[Path] = None,
) -> Dict[str, Any]:
    root = (
        (repo_root or Path(__file__).resolve().parents[1])
        / "var"
        / "employee_draft_assets"
        / str(user_id)
        / session_id
    )
    if root.exists():
        shutil.rmtree(root, ignore_errors=True)
    root.mkdir(parents=True, exist_ok=True)
    assets: List[Dict[str, Any]] = []
    for idx, item in enumerate(raw_files or []):
        filename = _safe_basename(str(item.get("filename") or f"asset-{idx}.bin"))
        content = item.get("content") or b""
        if not isinstance(content, (bytes, bytearray)):
            continue
        kind = _classify_asset(filename)
        dest = root / f"{idx:02d}_{filename}"
        dest.write_bytes(bytes(content))
        rec: Dict[str, Any] = {
            "id": f"asset_{idx}",
            "filename": filename,
            "kind": kind,
            "suffix": Path(filename).suffix.lower(),
            "size": len(content),
            "path": str(dest),
        }
        if rec["suffix"] in EXCEL_SUFFIXES:
            rec["excel"] = _excel_summary(dest)
        elif rec["suffix"] in TEXT_SUFFIXES:
            rec["preview"] = _read_text_preview(bytes(content))
        assets.append(rec)
    manifest = {
        "session_id": session_id,
        "user_id": user_id,
        "root": str(root),
        "assets": assets,
        "templates": [a for a in assets if a["kind"] == "template"],
        "example_inputs": [a for a in assets if a["kind"] == "example_input"],
        "expected_outputs": [a for a in assets if a["kind"] == "expected_output"],
        "rules": [a for a in assets if a["kind"] in ("rules", "reference_code")],
    }
    (root / "asset_manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return manifest


def _preflight_scaffold_write_access(
    *,
    session_id: str,
    user_id: int,
    repo_root: Optional[Path] = None,
) -> Tuple[bool, List[str], Dict[str, str]]:
    """资产脚手架写路径权限预检（MODstore 库 + 会话草稿目录）。"""
    paths_checked: Dict[str, str] = {}
    errors: List[str] = []
    lib = modstore_library_path()
    paths_checked["modstore_library"] = str(lib)
    try:
        lib.mkdir(parents=True, exist_ok=True)
        if not os.access(lib, os.W_OK):
            errors.append(f"modstore 库目录不可写：{lib}")
    except OSError as exc:
        errors.append(f"modstore 库目录不可创建：{lib} ({exc})")

    draft_root = (
        (repo_root or Path(__file__).resolve().parents[1])
        / "var"
        / "employee_draft_assets"
        / str(user_id)
        / session_id
    )
    paths_checked["draft_assets_root"] = str(draft_root)
    try:
        draft_root.parent.mkdir(parents=True, exist_ok=True)
        probe = draft_root.parent / ".write_probe"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
    except OSError as exc:
        errors.append(f"会话资产目录不可写：{draft_root.parent} ({exc})")

    return (not errors), errors, paths_checked


def build_rule_spec(
    brief: str,
    asset_manifest: Dict[str, Any],
    *,
    payload: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    from modstore_server.employee_pipeline_routing import confident_word_full_extract_routing

    if confident_word_full_extract_routing(brief):
        return build_word_extract_rule_spec(brief)
    if is_csv_generate(brief):
        return build_csv_generate_rule_spec(brief)
    if is_csv_full_read(brief):
        return build_csv_read_rule_spec(brief)
    if is_excel_generate(brief):
        return build_excel_generate_rule_spec(brief)
    if is_excel_full_read(brief):
        return build_excel_read_rule_spec(brief)
    if is_pdf_generate(brief):
        return build_pdf_generate_rule_spec(brief)
    if is_pdf_full_read(brief):
        return build_pdf_read_rule_spec(brief)
    if is_json_quant_report(brief):
        return build_json_quant_report_rule_spec(brief)
    if is_kitten_chart_viz(brief):
        return build_kitten_chart_rule_spec(brief)
    if is_ppt_generate(brief):
        return build_ppt_generate_rule_spec(brief)
    if is_ppt_full_read(brief):
        return build_ppt_read_rule_spec(brief)
    if is_txt_generate(brief):
        return build_txt_generate_rule_spec(brief)
    if is_txt_full_read(brief):
        return build_txt_read_rule_spec(brief)
    if is_word_full_extract(brief):
        return build_word_extract_rule_spec(brief)
    if is_word_generate(brief):
        return build_word_generate_rule_spec(brief)
    templates = asset_manifest.get("templates") or []
    if not templates:
        excels = [
            a for a in asset_manifest.get("assets") or [] if a.get("suffix") in EXCEL_SUFFIXES
        ]
        templates = excels[:1]
    template_asset = templates[0] if templates else None
    template_relpath = ""
    if template_asset:
        template_relpath = _template_storage_relpath(str(template_asset["filename"]), brief)
    runtime_kind = _infer_asset_runtime_kind(brief, asset_manifest)
    accepted_exts = _infer_accepted_extensions(asset_manifest)
    _is_doc_kind = runtime_kind in ("contract_doc_review", "doc_template_transform")
    if not accepted_exts:
        if _is_doc_kind:
            accepted_exts = [".docx", ".pdf"]
        else:
            accepted_exts = [".xlsx", ".xlsm", ".xls"]
    output_ext = accepted_exts[0] if accepted_exts else ".xlsx"
    output_relpath = f"outputs/employee_output{output_ext}"
    if "考勤" in brief:
        output_relpath = f"424/考勤转换输出{output_ext}"
    _mode = (
        "llm_doc_review"
        if runtime_kind == "contract_doc_review"
        else "direct_python_file_transform"
    )
    spec = {
        "brief": brief,
        "mode": _mode,
        "accepted_extensions": accepted_exts,
        "default_action": "review" if _is_doc_kind else "convert",
        "default_output_relpath": output_relpath,
        "default_template_relpath": (
            template_relpath.removeprefix("backend/templates/") if template_relpath else ""
        ),
        "template_relpath": template_relpath,
        "template_asset_id": template_asset.get("id") if template_asset else "",
        "runtime_kind": runtime_kind,
        "assets_summary": {
            "templates": [
                {
                    "filename": a.get("filename"),
                    "sheets": [
                        {
                            "name": s.get("name"),
                            "max_row": s.get("max_row"),
                            "max_column": s.get("max_column"),
                            "headers": s.get("header_candidates", [])[:2],
                        }
                        for s in ((a.get("excel") or {}).get("sheets") or [])[:6]
                    ],
                }
                for a in templates[:3]
            ],
            "example_inputs": [
                a.get("filename") for a in (asset_manifest.get("example_inputs") or [])[:5]
            ],
            "expected_outputs": [
                a.get("filename") for a in (asset_manifest.get("expected_outputs") or [])[:5]
            ],
            "rules": [
                {"filename": a.get("filename"), "preview": str(a.get("preview") or "")[:1000]}
                for a in (asset_manifest.get("rules") or [])[:5]
            ],
        },
        "requirements": [
            "Use direct_python only; do not add echo or llm_md.",
            "Preserve uploaded templates as binary files and copy them into backend/templates.",
            "Never claim success unless an output file is actually written.",
            "Return {ok, summary, items, warnings, error, meta}.",
        ],
    }
    if runtime_kind == "contract_doc_review":
        spec["requirements"] = [
            "This employee reviews contracts/documents using LLM reasoning via the agent handler.",
            "Read the uploaded document, identify missing clauses, ambiguous terms, and compliance issues.",
            "Output a structured review with specific suggestions for each issue found.",
            "Never fabricate legal advice; clearly state when professional legal review is recommended.",
            "Return {ok, summary, items, warnings, error, meta}.",
        ]
    elif runtime_kind in {"excel_rules_transform", "reference_python_transform"}:
        spec["requirements"].extend(
            [
                "Generate deterministic Python from the uploaded rules, template, and examples.",
                "If reference Python is provided, adapt or call it instead of replacing it with a placeholder.",
                "Fail explicitly when the generated transform cannot map the input workbook to the output workbook.",
            ]
        )
    return spec


def _slug_from_brief(brief: str) -> str:
    text = brief or ""
    explicit = re.search(
        r"(?:pack_id|员工包\s*ID|员工包ID|员工包 id|包ID|包 id)\s*[:：=]\s*([A-Za-z0-9][A-Za-z0-9_-]{2,80})",
        text,
        re.I,
    )
    if explicit:
        return normalize_mod_id(explicit.group(1)) or "asset-worker-employee"
    if re.search(r"\btaiyangniao[-_ ]attendance(?:[-_ ]employee)?\b", text, re.I):
        return "taiyangniao-attendance-employee"
    if is_csv_generate(brief):
        return "csv-generate-employee"
    if is_csv_full_read(brief):
        return "csv-full-read-employee"
    if is_excel_generate(brief):
        return "excel-generate-employee"
    if is_excel_full_read(brief):
        return "excel-full-read-employee"
    if is_pdf_generate(brief):
        return "pdf-generate-employee"
    if is_pdf_full_read(brief):
        return "pdf-full-read-employee"
    if is_json_quant_report(brief):
        return "json-report-employee"
    if is_ppt_generate(brief):
        return "ppt-generate-employee"
    if is_ppt_full_read(brief):
        return "ppt-full-read-employee"
    if is_txt_generate(brief):
        return "txt-generate-employee"
    if is_txt_full_read(brief):
        return "txt-full-read-employee"
    if is_word_full_extract(brief):
        return "word-full-read-employee"
    if is_word_generate(brief):
        return "word-generate-employee"
    if "考勤" in brief:
        return "attendance-transform-employee"
    return "asset-worker-employee"


def _employee_name_from_brief(brief: str, fallback: str = "文件处理员工") -> str:
    text = (brief or "").strip()
    explicit = re.search(r"(?:员工名称|员工名|name)\s*[:：=]\s*([^\n\r,，。；;]{2,40})", text, re.I)
    if explicit:
        return explicit.group(1).strip()
    if "考勤" in text:
        return "考勤处理员"
    if "报表" in text:
        return "报表处理员"
    if re.search(r"审核|审查|合规|风控|合同", text):
        return "合同审核员工"
    if re.search(r"翻译|本地化|多语言", text):
        return "文档翻译员工"
    if re.search(r"比对|对比|校对|校验", text):
        return "文档比对员工"
    if is_csv_generate(text):
        return "CSV 生成员"
    if is_csv_full_read(text):
        return "CSV 全量读取员"
    if is_excel_generate(text):
        return "Excel 生成员"
    if is_excel_full_read(text):
        return "Excel 全量读取员"
    if is_pdf_generate(text):
        return "PDF 生成员"
    if is_pdf_full_read(text):
        return "PDF 全量读取员"
    if is_ppt_generate(text):
        return "PPT 生成员"
    if is_ppt_full_read(text):
        return "PPT 全量读取员"
    if is_txt_generate(text):
        return "TXT 生成员"
    if is_txt_full_read(text):
        return "TXT 全量读取员"
    if is_word_generate(text):
        return "Word 生成员"
    if is_word_full_extract(text):
        return "Word 全量读取员"
    return fallback


def _template_storage_relpath(filename: str, brief: str = "") -> str:
    safe = _safe_basename(filename, "template.xlsx")
    return (
        f"backend/templates/424/{safe}" if "考勤" in (brief or "") else f"backend/templates/{safe}"
    )


def _employee_id_from_pack_id(pack_id: str) -> str:
    pid = (pack_id or "").strip()
    if pid.endswith("-employee"):
        return pid[: -len("-employee")] or pid
    return pid


def _fallback_manifest(brief: str, rule_spec: Dict[str, Any]) -> Dict[str, Any]:
    pid = _slug_from_brief(brief)
    name = _employee_name_from_brief(brief)
    is_attendance = "考勤" in brief
    runtime_kind = rule_spec.get("runtime_kind") or "generic_excel_transform"
    _is_doc_review = runtime_kind in ("contract_doc_review", "doc_template_transform")
    _is_word_extract = runtime_kind == "word_full_extract"
    _is_word_gen = runtime_kind == "word_generate"
    _is_csv_read = runtime_kind == "csv_full_read"
    _is_csv_gen = runtime_kind == "csv_generate"
    _is_excel_read = runtime_kind == "excel_full_read"
    _is_excel_gen = runtime_kind == "excel_generate"
    _is_txt_read = runtime_kind == "txt_full_read"
    _is_txt_gen = runtime_kind == "txt_generate"
    _is_pdf_read = runtime_kind == "pdf_full_read"
    _is_pdf_gen = runtime_kind == "pdf_generate"
    _is_ppt_read = runtime_kind == "ppt_full_read"
    _is_ppt_gen = runtime_kind == "ppt_generate"
    _is_json_report = runtime_kind == "json_quant_report"
    _is_kitten_chart = runtime_kind == "kitten_chart_viz"
    employee_id = str(rule_spec.get("pack_id") or pid).strip() or pid
    accepted = rule_spec.get("accepted_extensions") or [".xlsx"]
    has_doc = any(e in DOC_SUFFIXES for e in accepted)
    has_xls = any(e in EXCEL_SUFFIXES for e in accepted)
    if _is_csv_read:
        capabilities = ["data.csv_read", "data.json_export"]
    elif _is_csv_gen:
        capabilities = ["data.json_read", "data.csv_write"]
    elif _is_excel_read:
        capabilities = ["excel.full_read", "data.json_export"]
    elif _is_excel_gen:
        capabilities = ["data.json_read", "excel.write"]
    elif _is_txt_read:
        capabilities = ["text.full_read", "text.encoding_detect"]
    elif _is_txt_gen:
        capabilities = ["text.parse", "text.write", "text.polish_optional"]
    elif _is_pdf_read:
        capabilities = ["pdf.native_text", "pdf.image_extract", "vision.vlm"]
    elif _is_pdf_gen:
        capabilities = ["pdf.parse", "pdf.write", "pdf.polish_optional"]
    elif _is_ppt_read:
        capabilities = ["ppt.parse", "ppt.notes_generate", "vision.vlm"]
    elif _is_ppt_gen:
        capabilities = ["ppt.write", "ppt.ooxml", "data.json_read", "llm.plan"]
    elif _is_json_report:
        capabilities = ["data.json_read", "report.write"]
    elif _is_kitten_chart:
        capabilities = ["data.json_read", "chart.echarts", "viz.dashboard"]
    elif _is_word_gen:
        capabilities = ["doc.generate", "doc.template_merge", "doc.styles"]
    elif _is_word_extract:
        capabilities = [
            "doc.full_extract",
            "doc.tables",
            "doc.images",
            "doc.metadata",
            "doc.styles",
        ]
    elif _is_doc_review:
        capabilities = ["doc.review", "doc.compliance_check", "doc.suggestion"]
    elif is_attendance:
        capabilities = [
            "attendance.rules",
            "attendance.convert_upload",
            "attendance.template_fill",
            "attendance.download_hint",
        ]
    else:
        capabilities = ["file.transform", "doc.template_fill" if has_doc else "excel.template_fill"]
    if _is_csv_read:
        prompt = (
            f"你是{name}。你负责将用户上传的 .csv 解析为结构化 JSON（outputs/data.json）。"
            "必须真实执行 direct_python，禁止 LLM 编造行列数据。"
        )
    elif _is_csv_gen:
        prompt = (
            f"你是{name}。你负责根据 JSON（columns/rows）写出 outputs/output.csv。"
            "JSON 为中介；必须真实执行 direct_python，禁止编造表格内容。"
        )
    elif _is_excel_read:
        prompt = (
            f"你是{name}。你负责将用户上传的 xlsx 全量解析为 outputs/workbook.json（含 sheet、表头、单元格）。"
            "必须真实执行 direct_python，禁止 LLM 编造单元格数据。"
        )
    elif _is_excel_gen:
        prompt = (
            f"你是{name}。你负责根据 JSON（sheets/columns/rows）写出 outputs/output.xlsx。"
            "JSON 为中介；必须真实执行 direct_python，禁止编造表格内容。"
        )
    elif _is_txt_read:
        prompt = (
            f"你是{name}。你负责读取用户上传的 .txt 文件并原样交付全部纯文本。"
            "必须真实执行 direct_python，输出 document_full.txt 与 document_meta.json，禁止编造正文。"
        )
    elif _is_txt_gen:
        prompt = (
            f"你是{name}。你负责读取 .txt、输出结构化 JSON，并写入 generated_document.txt。"
            "direct_python 必须真实解析；润色/改写任务可走 agent，禁止无输入编造内容。"
        )
    elif _is_pdf_read:
        prompt = (
            f"你是{name}。你负责读取 PDF 原生文字并导出分类图片目录；"
            "正文禁止 LLM 编造；图片须走 VLM 描述（ctx.call_llm vision）。"
        )
    elif _is_pdf_gen:
        prompt = (
            f"你是{name}。你负责读取 PDF、输出 JSON 中介并生成 generated_document.pdf。"
            "direct_python 必须真实解析；润色可走 agent，禁止无输入编造。"
        )
    elif _is_ppt_read:
        prompt = (
            f"你是{name}。你负责全量解析 PPT：大纲、每页正文、导出图片并 VLM 描述；"
            "按「为这份PPT生成每页的演讲备注」生成 notes_generated。"
            "正文禁止 LLM 编造；必须真实执行 direct_python。"
        )
    elif _is_ppt_gen:
        prompt = (
            f"你是{name}。compose-first：无模板时从零合成多页 output.pptx；"
            "enhance：复制 template 后按 ppt_edit_plan 注入 OOXML 动画。"
            "必须执行 modstore_server.ppt_generate_pipeline，禁止仅输出纯文字幻灯片冒充带动效作业。"
        )
    elif _is_json_report:
        prompt = (
            f"你是{name}。你负责读取 document_full.json 或 execute_result 包装的 JSON，"
            "在 direct_python 内调用 LLM 生成 outputs/quantitative_report.html。"
            "仅基于 JSON 事实撰写量化报告，禁止编造未出现的指标或结论。"
        )
    elif _is_word_gen:
        prompt = (
            f"你是{name}。你负责读取 document_full.json（或与 Word 全量读取同 schema 的 JSON），"
            "可选 template.docx 模板，生成 generated_document.docx。"
            "direct_python 必须真实写 docx；禁止无 JSON 编造正文。"
        )
    elif _is_word_extract:
        prompt = (
            f"你是{name}。你负责全量提取 Word 文档的所有格式与信息：段落、表格、图片、样式、"
            "页眉页脚、元数据与批注。必须真实执行 direct_python 解析，输出 document_full.json 与 txt，"
            "禁止 LLM 编造文档内容。"
        )
    elif _is_doc_review:
        prompt = (
            f"你是{name}。你负责审核用户上传的合同/文档，识别缺失条款、模糊表述和合规风险，"
            "并给出具体修改建议。必须基于文档实际内容进行分析，禁止编造不存在的条款。"
            "当涉及专业法律建议时，应明确建议咨询专业律师。"
        )
    elif is_attendance:
        prompt = (
            f"你是{name}。你负责把用户上传的 Excel 按规则和模板生成结果。"
            "必须真实执行 direct_python 转换；输入文件、模板或转换模块缺失时返回明确错误，禁止编造成功。"
        )
    elif has_doc:
        prompt = (
            f"你是{name}。你负责处理用户上传的文档文件，按规则和模板生成结果。"
            "必须真实执行 direct_python 转换；输入文件缺失时返回明确错误，禁止编造成功。"
        )
    else:
        prompt = "你是文件处理员工。必须真实执行 direct_python，失败时返回真实错误，禁止编造结果。"
    if _is_csv_read:
        expertise = ["CSV 解析", "JSON 结构化", "表格数据"]
        persona = "严谨的 CSV 全量读取员工"
        skill_brief = "上传 csv，输出 data.json 中介。"
    elif _is_csv_gen:
        expertise = ["JSON 解析", "CSV 写出", "表格生成"]
        persona = "严谨的 CSV 生成员工"
        skill_brief = "JSON 中介 → 写出 output.csv。"
    elif _is_excel_read:
        expertise = ["Excel 解析", "JSON 结构化", "单元格全量"]
        persona = "严谨的 Excel 全量读取员工"
        skill_brief = "上传 xlsx，输出 workbook.json 中介。"
    elif _is_excel_gen:
        expertise = ["JSON 解析", "Excel 写出", "多 sheet"]
        persona = "严谨的 Excel 生成员工"
        skill_brief = "JSON 中介 → 写出 output.xlsx。"
    elif _is_txt_read:
        expertise = ["TXT 读取", "编码检测", "纯文本"]
        persona = "严谨的 TXT 全量读取员工"
        skill_brief = "上传 txt，原样读出全部文本并交付。"
    elif _is_txt_gen:
        expertise = ["TXT 解析", "JSON 结构化", "文档生成"]
        persona = "严谨的 TXT 生成员工"
        skill_brief = "上传 txt → JSON → 写 generated txt，可选润色。"
    elif _is_pdf_read:
        expertise = ["PDF 原生文字", "图片分类", "VLM 描述"]
        persona = "严谨的 PDF 全量读取员工"
        skill_brief = "上传 pdf，原生文字 + 图片分类目录 + VLM sidecar。"
    elif _is_pdf_gen:
        expertise = ["PDF 解析", "JSON 中介", "PDF 生成"]
        persona = "严谨的 PDF 生成员工"
        skill_brief = "上传 pdf → JSON → 写 generated pdf，可选润色。"
    elif _is_ppt_read:
        expertise = ["PPT 解析", "演讲备注", "VLM 识图"]
        persona = "严谨的 PPT 全量读取员工"
        skill_brief = "上传 pptx → JSON 中介 + 演讲备注 + 图片 VLM。"
    elif _is_ppt_gen:
        expertise = ["PPT 合成", "OOXML 动画", "LLM 编排"]
        persona = "严谨的 PPT 生成员工"
        skill_brief = "plan → compose/enhance → output.pptx。"
    elif _is_json_report:
        expertise = ["JSON 解析", "量化报告", "HTML 撰写"]
        persona = "严谨的 JSON 量化报告员"
        skill_brief = "document_full JSON → 美观 HTML 量化报告。"
    elif _is_word_gen:
        expertise = ["Word 生成", "JSON 中介", "模板合并"]
        persona = "严谨的 Word 生成员工"
        skill_brief = "JSON + 可选模板 → 生成 docx。"
    elif _is_word_extract:
        expertise = ["Word 解析", "OOXML", "文档结构化"]
        persona = "严谨的 Word 全量提取员工"
        skill_brief = "全量解析 docx，输出 JSON/txt/图片等结构化结果。"
    elif _is_doc_review:
        expertise = ["合同审核", "条款分析", "合规检查"]
        persona = "严谨的合同审核员工"
        skill_brief = "审核上传的合同/文档，识别风险条款并给出修改建议。"
    elif has_doc:
        expertise = ["文档处理", "模板回填"]
        persona = "严谨的文档处理员工"
        skill_brief = "读取上传文档并按模板生成输出。"
    elif has_xls:
        expertise = ["Excel", "模板回填"]
        persona = "严谨的数据处理员工"
        skill_brief = "读取上传文件并按模板生成输出。"
    else:
        expertise = ["文件处理", "模板回填"]
        persona = "严谨的数据处理员工"
        skill_brief = "读取上传文件并按模板生成输出。"
    if _is_csv_read:
        panel_summary = "上传 .csv，解析为 JSON 中介 outputs/data.json。"
    elif _is_csv_gen:
        panel_summary = "上传 JSON/纯文本描述，按 columns/rows 写出 outputs/output.csv。"
    elif _is_excel_read:
        panel_summary = "上传 .xlsx，全量读取 sheet/表头/单元格并输出 workbook.json。"
    elif _is_excel_gen:
        panel_summary = "上传 JSON 或纯文本描述，按 sheets 写出 outputs/output.xlsx。"
    elif _is_txt_read:
        panel_summary = "上传 .txt，全量读取纯文本并交付 document_full.txt。"
    elif _is_txt_gen:
        panel_summary = "上传 .txt，解析为 JSON 并生成 txt 文档，可选润色。"
    elif _is_pdf_read:
        panel_summary = "上传 .pdf，只读原生文字；图片分类存储并 VLM 描述。"
    elif _is_pdf_gen:
        panel_summary = "上传 JSON/纯文本描述，生成 PDF，可选润色。"
    elif _is_ppt_read:
        panel_summary = "上传 .pptx，全量解析并生成演讲备注；图片 VLM 识图。"
    elif _is_ppt_gen:
        panel_summary = "文字/JSON 从零生成或基于 template.pptx 增强，输出 output.pptx（含动画）。"
    elif _is_json_report:
        panel_summary = (
            "上传 document_full.json（或 execute_result），生成 outputs/quantitative_report.html。"
        )
    elif _is_word_gen:
        panel_summary = "上传 JSON（document_full.json），可选模板 docx，生成 Word 文档。"
    elif _is_word_extract:
        panel_summary = "上传 Word，全量提取正文/表格/图片/样式/元数据并交付 JSON。"
    elif _is_doc_review:
        panel_summary = "上传合同/文档，AI 审核风险条款并给出修改建议。"
    elif is_attendance:
        panel_summary = "上传考勤表，按规则和模板生成考勤结果。"
    else:
        panel_summary = "读取上传文件并按模板生成输出。"
    if _is_csv_read:
        behavior_rules = [
            "必须真实解析 csv，禁止编造行列。",
            "成功必须以写出 data.json 且 row_count 正确为准。",
        ]
        few_shot = []
    elif _is_csv_gen:
        behavior_rules = [
            "必须根据 JSON columns/rows 写出 csv。",
            "成功必须以写出 output.csv 为准。",
        ]
        few_shot = []
    elif _is_excel_read:
        behavior_rules = [
            "必须真实解析 xlsx，禁止编造单元格。",
            "成功必须以写出 workbook.json 为准。",
        ]
        few_shot = []
    elif _is_excel_gen:
        behavior_rules = [
            "必须根据 JSON sheets/rows 写出 xlsx。",
            "成功必须以写出 output.xlsx 为准。",
        ]
        few_shot = []
    elif _is_txt_read:
        behavior_rules = [
            "必须真实读取 txt 原文，禁止编造内容。",
            "成功必须以写出 document_full.txt 为准。",
        ]
        few_shot = []
    elif _is_txt_gen:
        behavior_rules = [
            "必须真实读取 txt 并写出 document_parsed.json。",
            "成功必须以写出 generated_document.txt 为准。",
            "润色时须基于 JSON 摘要，禁止无输入编造。",
        ]
        few_shot = []
    elif _is_pdf_read:
        behavior_rules = [
            "正文必须来自 PDF 原生文字层，禁止 LLM 编造。",
            "图片须写入 outputs/images/<category>/ 并尽量生成 VLM sidecar。",
            "成功必须以写出 document_full.txt 与 images_index.json 为准。",
        ]
        few_shot = []
    elif _is_pdf_gen:
        behavior_rules = [
            "必须真实读取 PDF 并写出 document_parsed.json。",
            "成功必须以写出 generated_document.pdf 为准。",
            "润色时须基于 JSON，禁止无输入编造。",
        ]
        few_shot = []
    elif _is_ppt_read:
        behavior_rules = [
            "幻灯片正文必须来自 pptx 真实解析，禁止 LLM 编造。",
            "必须写出 presentation_full.json 与 speaker_notes.md。",
            "图片尽量生成 VLM sidecar。",
        ]
        few_shot = []
    elif _is_ppt_gen:
        behavior_rules = [
            "必须写出 outputs/output.pptx（含 ppt_edit_plan.json）。",
            "compose 无模板时须多页骨架；enhance 须保留 template 媒体。",
            "禁止无输入编造；禁止纯文字冒充带动效/带图 PPT。",
        ]
        few_shot = []
    elif _is_json_report:
        behavior_rules = [
            "必须写出 outputs/quantitative_report.html。",
            "禁止编造 JSON 中不存在的章节或数据。",
            "统计数字优先使用确定性摘要。",
        ]
        few_shot = []
    elif _is_word_gen:
        behavior_rules = [
            "必须基于 JSON 真实生成 docx，禁止无输入编造正文。",
            "成功必须以写出 generated_document.docx 为准。",
            "可选模板仅用于样式，不得覆盖 JSON 正文语义。",
        ]
        few_shot = []
    elif _is_word_extract:
        behavior_rules = [
            "必须真实解析 docx，禁止编造段落或表格内容。",
            "成功必须以写出 document_full.json 为准。",
            "JSON 须含 paragraphs、tables、images、core_properties 等字段。",
        ]
        few_shot = []
    elif _is_doc_review:
        behavior_rules = [
            "必须基于文档实际内容分析，禁止编造不存在的条款。",
            "涉及专业法律建议时，必须建议咨询专业律师。",
            "输出必须包含具体条款位置和修改建议。",
        ]
        few_shot = [
            {
                "input": "上传一份AI技术服务合同",
                "output": "审核报告：1) 第3条服务范围表述模糊，建议明确具体服务项；2) 缺少数据安全条款，建议补充；3) 违约责任条款不完整，建议增加违约金比例。",
            },
        ]
    else:
        behavior_rules = ["没有真实文件时必须报错。", "成功必须以实际写出输出文件为准。"]
        few_shot = []
    if _is_doc_review:
        actions_cfg: Dict[str, Any] = {
            "handlers": ["agent"],
        }
    elif _is_txt_gen:
        handlers_list = (
            ["direct_python", "agent"]
            if rule_spec.get("optional_llm_polish")
            else ["direct_python", "agent"]
        )
        actions_cfg = {
            "handlers": handlers_list,
            "direct_python": {
                "module": sanitize_employee_stem(employee_id),
                "action": "convert",
                "default_output_relpath": rule_spec.get("default_output_relpath")
                or "outputs/document_parsed.json",
                "default_text_output_relpath": rule_spec.get("default_text_output_relpath")
                or "outputs/generated_document.txt",
            },
        }
    elif _is_pdf_gen:
        handlers_list = (
            ["direct_python", "agent"]
            if rule_spec.get("optional_llm_polish")
            else ["direct_python", "agent"]
        )
        actions_cfg = {
            "handlers": handlers_list,
            "direct_python": {
                "module": sanitize_employee_stem(employee_id),
                "action": "convert",
                "default_output_relpath": rule_spec.get("default_output_relpath")
                or "outputs/document_parsed.json",
                "default_pdf_output_relpath": rule_spec.get("default_pdf_output_relpath")
                or "outputs/generated_document.pdf",
            },
        }
    elif _is_word_gen:
        handlers_list = (
            ["direct_python", "agent"]
            if rule_spec.get("optional_llm_polish")
            else ["direct_python"]
        )
        actions_cfg = {
            "handlers": handlers_list,
            "direct_python": {
                "module": sanitize_employee_stem(employee_id),
                "action": "convert",
                "default_output_relpath": rule_spec.get("default_output_relpath")
                or "outputs/generated_document.docx",
                "default_template_relpath": str(
                    rule_spec.get("default_template_relpath") or "inputs/template.docx"
                ),
            },
        }
    else:
        default_out = rule_spec.get("default_output_relpath") or (
            "outputs/data.json"
            if _is_csv_read
            else (
                "outputs/output.csv"
                if _is_csv_gen
                else (
                    "outputs/workbook.json"
                    if _is_excel_read
                    else (
                        "outputs/output.xlsx"
                        if _is_excel_gen
                        else (
                            "outputs/document_full.txt"
                            if _is_txt_read
                            else (
                                "outputs/document_full.txt"
                                if _is_pdf_read
                                else (
                                    "outputs/quantitative_report.html"
                                    if _is_json_report
                                    else "outputs/employee_output.xlsx"
                                )
                            )
                        )
                    )
                )
            )
        )
        actions_cfg = {
            "handlers": ["direct_python"],
            "direct_python": {
                "module": sanitize_employee_stem(employee_id),
                "action": "convert",
                "default_output_relpath": default_out,
                "default_template_relpath": (
                    str(rule_spec.get("template_relpath") or "").removeprefix("backend/templates/")
                ),
                "default_use_personnel_roster": not (
                    _is_csv_read
                    or _is_csv_gen
                    or _is_excel_read
                    or _is_excel_gen
                    or _is_txt_read
                    or _is_txt_gen
                    or _is_pdf_read
                    or _is_pdf_gen
                ),
            },
        }
    if _is_csv_read:
        skill_name = "data.csv_read"
    elif _is_csv_gen:
        skill_name = "data.csv_write"
    elif _is_excel_read:
        skill_name = "excel.full_read"
    elif _is_excel_gen:
        skill_name = "excel.write"
    elif _is_txt_read:
        skill_name = "text.full_read"
    elif _is_txt_gen:
        skill_name = "text.generate"
    elif _is_pdf_read:
        skill_name = "pdf.full_read"
    elif _is_pdf_gen:
        skill_name = "pdf.generate"
    elif _is_ppt_read:
        skill_name = "ppt.full_read"
    elif _is_ppt_gen:
        skill_name = "ppt.generate"
    elif _is_json_report:
        skill_name = "report.quantitative"
    elif _is_word_gen:
        skill_name = "doc.generate"
    elif _is_word_extract:
        skill_name = "doc.full_extract"
    elif _is_doc_review:
        skill_name = "doc.review"
    else:
        skill_name = "file.transform"
    return {
        "id": pid,
        "name": name,
        "version": "1.0.0",
        "author": "XCAGI",
        "description": _clean_brief_for_description(brief, 400),
        "artifact": "employee_pack",
        "scope": "global",
        "dependencies": {"xcagi": ">=1.0.0"},
        "employee": {"id": employee_id, "label": name, "capabilities": capabilities},
        "workflow_employees": [
            {
                "id": employee_id,
                "label": name,
                "panel_title": name,
                "panel_summary": panel_summary,
                "capabilities": capabilities,
                "api_base_path": f"employees/{employee_id}",
                "entry_action": "run",
            }
        ],
        "backend": {"entry": "blueprints", "init": "mod_init"},
        "employee_config_v2": {
            "identity": {
                "id": pid,
                "version": "1.0.0",
                "artifact": "employee_pack",
                "name": name,
                "description": _clean_brief_for_description(brief, 500),
            },
            "perception": {
                "type": (
                    "csv"
                    if _is_csv_read
                    else (
                        "json"
                        if _is_csv_gen or _is_excel_gen
                        else "excel" if _is_excel_read else "file_or_text"
                    )
                ),
                "accepted_extensions": rule_spec.get("accepted_extensions")
                or (
                    [".json"]
                    if _is_csv_gen or _is_excel_gen
                    else (
                        [".csv"]
                        if _is_csv_read
                        else (
                            [".xlsx", ".xlsm"]
                            if _is_excel_read
                            else [".docx", ".pdf"] if _is_doc_review else [".xlsx"]
                        )
                    )
                ),
            },
            "memory": {"type": "session"},
            "cognition": {
                "agent": {
                    "system_prompt": prompt,
                    "role": {
                        "name": name,
                        "persona": persona,
                        "tone": "professional",
                        "expertise": expertise,
                    },
                    "behavior_rules": behavior_rules,
                    "few_shot_examples": few_shot,
                    "model": {
                        "provider": "auto",
                        "model_name": "auto",
                        "temperature": 0.1 if not _is_doc_review else 0.3,
                        "max_tokens": 4000 if _is_doc_review else 2000,
                        "top_p": 0.9,
                    },
                },
                "skills": [{"name": skill_name, "brief": skill_brief}],
            },
            "collaboration": {"workflow": {"workflow_id": 0, "name": name}},
            "actions": actions_cfg,
        },
        "metadata": {"framework_version": "2.0.0", "created_by": "asset_pipeline"},
    }


def _normalize_manifest(
    manifest: Dict[str, Any], brief: str, rule_spec: Dict[str, Any]
) -> Dict[str, Any]:
    fallback = _fallback_manifest(brief, rule_spec)
    out = dict(fallback)
    out.update({k: v for k, v in manifest.items() if k not in ("employee_config_v2",)})
    explicit_pid = _slug_from_brief(brief)
    pack_id_hint = str(rule_spec.get("pack_id") or "").strip()
    pid = (
        pack_id_hint
        or normalize_mod_id(str(explicit_pid or out.get("id") or fallback["id"]))
        or fallback["id"]
    )
    out["id"] = pid
    out["artifact"] = "employee_pack"
    out.setdefault("version", "1.0.0")
    out.setdefault("name", fallback["name"])
    emp = out.get("employee") if isinstance(out.get("employee"), dict) else {}
    fallback_emp = fallback["employee"] if isinstance(fallback.get("employee"), dict) else {}
    emp = {**fallback_emp, **emp, "id": pid}
    emp.setdefault("label", out.get("name") or pid)
    out["employee"] = emp
    rows = out.get("workflow_employees")
    if not isinstance(rows, list) or not rows:
        out["workflow_employees"] = fallback.get("workflow_employees") or []
    else:
        normalized_rows = []
        for row in rows:
            if not isinstance(row, dict):
                continue
            r = dict(row)
            r["id"] = pid
            r.setdefault("label", emp.get("label") or out.get("name") or r["id"])
            r.setdefault("api_base_path", f"employees/{r['id']}")
            r.setdefault("entry_action", "run")
            normalized_rows.append(r)
        out["workflow_employees"] = normalized_rows or fallback.get("workflow_employees") or []
    out["backend"] = {"entry": "blueprints", "init": "mod_init"}
    v2 = (
        manifest.get("employee_config_v2")
        if isinstance(manifest.get("employee_config_v2"), dict)
        else {}
    )
    merged_v2 = dict(fallback["employee_config_v2"])
    merged_v2.update(v2)
    actions = dict(merged_v2.get("actions") if isinstance(merged_v2.get("actions"), dict) else {})
    runtime_kind = rule_spec.get("runtime_kind") or "generic_excel_transform"
    _is_doc_review = runtime_kind in ("contract_doc_review", "doc_template_transform")
    if _is_doc_review:
        actions["handlers"] = ["agent"]
        actions.pop("direct_python", None)
    else:
        direct = dict(
            actions.get("direct_python") if isinstance(actions.get("direct_python"), dict) else {}
        )
        direct["module"] = sanitize_employee_stem(pack_id_hint or pid)
        direct.setdefault("action", "convert")
        if rule_spec.get("default_output_relpath"):
            direct.setdefault("default_output_relpath", rule_spec["default_output_relpath"])
        if rule_spec.get("template_relpath"):
            direct.setdefault("default_template_relpath", rule_spec["template_relpath"])
        actions["handlers"] = ["direct_python"]
        actions["direct_python"] = direct
    merged_v2["actions"] = actions
    if rule_spec.get("accepted_extensions"):
        perception = (
            merged_v2.get("perception") if isinstance(merged_v2.get("perception"), dict) else {}
        )
        perception["accepted_extensions"] = rule_spec["accepted_extensions"]
        merged_v2["perception"] = perception
    ident = dict(merged_v2.get("identity") if isinstance(merged_v2.get("identity"), dict) else {})
    ident.update(
        {
            "id": pid,
            "version": str(out.get("version") or "1.0.0"),
            "artifact": "employee_pack",
            "name": str(out.get("name") or pid),
        }
    )
    raw_desc = str(ident.get("description") or out.get("description") or brief).strip()
    if re.match(r"^(你是[一]?(名|位|个)|角色[：:])", raw_desc):
        emp_name = str(out.get("name") or pid)
        raw_desc = f"{emp_name}。{_clean_brief_for_description(re.sub(r'^(你是[一]?(名|位|个)|角色[：:])', '', raw_desc), 200)}"
    ident["description"] = _clean_brief_for_description(raw_desc, 500)
    merged_v2["identity"] = ident
    out["description"] = _clean_brief_for_description(str(out.get("description") or brief), 400)
    top_desc = str(out.get("description") or "")
    if re.match(r"^(你是[一]?(名|位|个)|角色[：:])", top_desc):
        out["description"] = (
            f"{out.get('name') or pid}。{_clean_brief_for_description(re.sub(r'^(你是[一]?(名|位|个)|角色[：:])', '', top_desc), 200)}"
        )
    cog = merged_v2.get("cognition") if isinstance(merged_v2.get("cognition"), dict) else {}
    agent_cfg = cog.get("agent") if isinstance(cog.get("agent"), dict) else {}
    sp = str(agent_cfg.get("system_prompt") or "")
    if "llm_md" in sp or "echo" in sp or "调用 LLM" in sp:
        agent_cfg["system_prompt"] = fallback["employee_config_v2"]["cognition"]["agent"][
            "system_prompt"
        ]
        cog["agent"] = agent_cfg
        merged_v2["cognition"] = cog
    fb_few = (
        fallback.get("employee_config_v2", {})
        .get("cognition", {})
        .get("agent", {})
        .get("few_shot_examples")
    ) or []
    cur_few = (agent_cfg.get("few_shot_examples")) or []
    if not cur_few and fb_few:
        agent_cfg["few_shot_examples"] = fb_few
        cog["agent"] = agent_cfg
        merged_v2["cognition"] = cog
    out["employee_config_v2"] = merged_v2
    out["actions"] = dict(actions)
    bundles = out.get("workflow_bundles")
    if isinstance(bundles, list):
        for b in bundles:
            if not isinstance(b, dict):
                continue
            raw_desc = str(b.get("description") or "")
            cleaned = _clean_brief_for_description(raw_desc, 500)
            if not cleaned:
                cleaned = str(b.get("name") or "")
            b["description"] = cleaned
    return out


def _sanitize_workflow_bundles(manifest: Dict[str, Any]) -> None:
    """Clean bundle names/descriptions polluted by NL graph or voice planning."""
    name = str(manifest.get("name") or manifest.get("id") or "工作流").strip()
    bundles = manifest.get("workflow_bundles")
    if not isinstance(bundles, list):
        return
    for b in bundles:
        if not isinstance(b, dict):
            continue
        raw_name = str(b.get("name") or "").strip()
        if (
            not raw_name
            or raw_name in ("（无回复）", "(无回复)")
            or _PLACEHOLDER_BRIEF.search(raw_name)
        ):
            b["name"] = name
        raw_desc = str(b.get("description") or "")
        cleaned = _clean_brief_for_description(raw_desc, 500)
        b["description"] = cleaned or name


_PLACEHOLDER_BRIEF = re.compile(r"（无回复）|相处报备|开始写吧", re.I)


def reconcile_employee_pack_manifest(pack_dir: Path, *, brief: str = "") -> Dict[str, Any]:
    """Re-apply rule_spec + _normalize_manifest after workflow/register edits."""
    mf_path = pack_dir / "manifest.json"
    if not mf_path.is_file():
        raise FileNotFoundError(f"manifest.json missing under {pack_dir}")
    raw = json.loads(mf_path.read_text(encoding="utf-8"))
    rule_spec: Dict[str, Any] = {}
    rs_path = pack_dir / "rule_spec.json"
    if rs_path.is_file():
        try:
            loaded = json.loads(rs_path.read_text(encoding="utf-8"))
            if isinstance(loaded, dict):
                rule_spec = loaded
        except (OSError, json.JSONDecodeError):
            pass
    if not rule_spec:
        from modstore_server.csv_tabular_runtime import (
            build_csv_generate_rule_spec,
            build_csv_read_rule_spec,
            is_csv_full_read,
            is_csv_generate,
        )
        from modstore_server.employee_brief_utils import extract_routing_brief
        from modstore_server.excel_tabular_runtime import (
            build_excel_generate_rule_spec,
            build_excel_read_rule_spec,
            is_excel_full_read,
            is_excel_generate,
        )
        from modstore_server.pdf_extract_runtime import (
            build_pdf_generate_rule_spec,
            build_pdf_read_rule_spec,
            is_pdf_full_read,
            is_pdf_generate,
        )
        from modstore_server.txt_extract_runtime import (
            build_txt_generate_rule_spec,
            build_txt_read_rule_spec,
            is_txt_full_read,
            is_txt_generate,
        )
        from modstore_server.word_extract_runtime import (
            build_word_extract_rule_spec,
            is_word_full_extract,
        )
        from modstore_server.word_generate_runtime import (
            build_word_generate_rule_spec,
            is_word_generate,
        )

        rb = extract_routing_brief({"brief": brief}, fallback=brief)
        if is_csv_generate(rb):
            rule_spec = build_csv_generate_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_csv_full_read(rb):
            rule_spec = build_csv_read_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_excel_generate(rb):
            rule_spec = build_excel_generate_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_excel_full_read(rb):
            rule_spec = build_excel_read_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_pdf_generate(rb):
            rule_spec = build_pdf_generate_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_pdf_full_read(rb):
            rule_spec = build_pdf_read_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_json_quant_report(rb):
            rule_spec = build_json_quant_report_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_ppt_generate(rb):
            from modstore_server.ppt_extract_runtime import build_ppt_generate_rule_spec

            rule_spec = build_ppt_generate_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_ppt_full_read(rb):
            from modstore_server.ppt_extract_runtime import build_ppt_read_rule_spec

            rule_spec = build_ppt_read_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_txt_generate(rb):
            rule_spec = build_txt_generate_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_txt_full_read(rb):
            rule_spec = build_txt_read_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_word_generate(rb):
            rule_spec = build_word_generate_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
        elif is_word_full_extract(rb):
            rule_spec = build_word_extract_rule_spec(rb)
            rs_path.write_text(
                json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
            )
    rb = brief
    if not rb:
        from modstore_server.employee_brief_utils import extract_routing_brief

        rb = extract_routing_brief(
            {"brief": str(raw.get("description") or "")}, fallback=str(raw.get("description") or "")
        )
    if rule_spec:
        aligned = _normalize_manifest(raw, rb, rule_spec)
    else:
        aligned = dict(raw)
    _sanitize_workflow_bundles(aligned)
    mf_path.write_text(json.dumps(aligned, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    return aligned


async def enrich_manifest_productivity_fields(
    db: Any,
    user: User,
    *,
    brief: str,
    rule_spec: Dict[str, Any],
    base_manifest: Dict[str, Any],
    provider: Optional[str],
    model: Optional[str],
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """轻量 LLM：仅补 description / panel_summary / behavior_rules，与模板落盘并行，不阻塞 convert。"""
    from modstore_server.employee_pipeline_routing import is_direct_python_template_runtime

    runtime_kind = str(rule_spec.get("runtime_kind") or "")
    if not is_direct_python_template_runtime(runtime_kind):
        return base_manifest, {"source": "skipped", "reason": "not_template_runtime"}
    prov, mdl, err = await resolve_llm_provider_model_auto(db, user, provider, model)
    if err:
        return base_manifest, {"source": "fallback", "warning": err}
    api_key, _ = resolve_api_key(db, user.id, prov)  # type: ignore[arg-type]
    if not api_key:
        return base_manifest, {"source": "fallback", "warning": "missing api key"}
    base = (
        resolve_base_url(db, user.id, prov) if prov in OAI_COMPAT_OPENAI_STYLE_PROVIDERS else None
    )
    system = (
        "你是员工包文案助手。只输出 JSON，不要 markdown。字段：description（1-3句）、"
        "panel_summary（工作台卡片一句话）、behavior_rules（字符串数组，3-6条操作边界）。"
        "不得改写 runtime_kind、handlers 或技术契约；禁止编造已执行结果。"
    )
    user_msg = json.dumps(
        {"brief": brief[:4000], "runtime_kind": runtime_kind, "name": base_manifest.get("name")},
        ensure_ascii=False,
    )
    try:
        result = await chat_dispatch(
            prov,
            api_key=api_key,
            base_url=base,
            model=mdl,
            messages=[{"role": "system", "content": system}, {"role": "user", "content": user_msg}],
            max_tokens=900,
        )
    except Exception:
        return base_manifest, {"source": "fallback", "warning": "enrich dispatch failed"}
    if not result.get("ok"):
        return base_manifest, {"source": "fallback", "warning": str(result.get("error") or "")}
    try:
        raw = str(result.get("content") or "").strip()
        raw = re.sub(r"^```(?:json)?\s*", "", raw, flags=re.I)
        raw = re.sub(r"\s*```\s*$", "", raw).strip()
        patch = json.loads(raw)
    except json.JSONDecodeError:
        return base_manifest, {"source": "fallback", "warning": "enrich parse failed"}
    if not isinstance(patch, dict):
        return base_manifest, {"source": "fallback", "warning": "enrich not object"}
    out = dict(base_manifest)
    desc = str(patch.get("description") or "").strip()
    if desc:
        out["description"] = desc[:2000]
    panel = str(patch.get("panel_summary") or "").strip()
    if panel:
        out["panel_summary"] = panel[:500]
    rules = patch.get("behavior_rules")
    if isinstance(rules, list):
        cleaned = [str(x).strip() for x in rules if str(x).strip()][:8]
        if cleaned:
            v2 = out.get("employee_config_v2")
            if not isinstance(v2, dict):
                v2 = {}
                out["employee_config_v2"] = v2
            v2["behavior_rules"] = cleaned
    return out, {"provider": prov, "model": mdl, "source": "productivity_enrich"}


async def design_asset_employee_manifest(
    db: Any,
    user: User,
    *,
    brief: str,
    rule_spec: Dict[str, Any],
    provider: Optional[str],
    model: Optional[str],
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    from modstore_server.employee_pipeline_routing import is_direct_python_template_runtime

    runtime_kind = str(rule_spec.get("runtime_kind") or "")
    if is_direct_python_template_runtime(runtime_kind):
        return _fallback_manifest(brief, rule_spec), {
            "provider": "",
            "model": "",
            "source": "template_manifest",
        }
    prov, mdl, err = await resolve_llm_provider_model_auto(db, user, provider, model)
    if err and runtime_kind not in (
        "word_full_extract",
        "word_generate",
        "txt_full_read",
        "txt_generate",
        "pdf_full_read",
        "pdf_generate",
        "csv_full_read",
        "csv_generate",
        "excel_full_read",
        "excel_generate",
    ):
        return _fallback_manifest(brief, rule_spec), {"provider": "", "model": "", "warning": err}
    if runtime_kind in (
        "word_full_extract",
        "word_generate",
        "txt_full_read",
        "txt_generate",
        "pdf_full_read",
        "pdf_generate",
        "csv_full_read",
        "csv_generate",
        "excel_full_read",
        "excel_generate",
    ) and (err or not resolve_api_key(db, user.id, prov or "")[0]):
        return _fallback_manifest(brief, rule_spec), {
            "provider": prov or "",
            "model": mdl or "",
            "warning": err or "missing api key",
        }
    api_key, _ = resolve_api_key(db, user.id, prov)  # type: ignore[arg-type]
    if not api_key:
        return _fallback_manifest(brief, rule_spec), {
            "provider": prov,
            "model": mdl,
            "warning": "missing api key",
        }
    base = (
        resolve_base_url(db, user.id, prov) if prov in OAI_COMPAT_OPENAI_STYLE_PROVIDERS else None
    )
    runtime_kind = rule_spec.get("runtime_kind") or "generic_excel_transform"
    _is_doc_review = runtime_kind in ("contract_doc_review", "doc_template_transform")
    if _is_doc_review:
        system = (
            "你是 employee_pack manifest 设计器。只输出 JSON，不输出 Markdown。"
            '这个员工是文档审核/处理员工，需要 LLM 推理能力，actions.handlers 必须只有 ["agent"]。'
            "不要声明 echo、llm_md、direct_python。不要编造已经执行。"
            "cognition.agent.system_prompt 应包含文档审核的专业指令。"
        )
    else:
        system = (
            "你是 employee_pack manifest 设计器。只输出 JSON，不输出 Markdown。"
            '这个员工必须是 direct_python 文件处理员工，actions.handlers 必须只有 ["direct_python"]。'
            "不要声明 echo、llm_md。不要编造已经执行。"
        )
        if runtime_kind == "word_full_extract":
            system += (
                " rule_spec.runtime_kind 为 word_full_extract：perception.accepted_extensions 必须含 .docx；"
                "默认输出为 outputs/document_full.json（读取/提取，不是生成 docx）；"
                "capabilities 须含 doc.full_extract；禁止写成 Word 生成或仅接受 .json 的员工。"
            )
    user_msg = json.dumps({"brief": brief, "rule_spec": rule_spec}, ensure_ascii=False)[:12000]
    result = await chat_dispatch(
        prov,
        api_key=api_key,
        base_url=base,
        model=mdl,
        messages=[{"role": "system", "content": system}, {"role": "user", "content": user_msg}],
        max_tokens=4000,
    )
    if not result.get("ok"):
        return _fallback_manifest(brief, rule_spec), {
            "provider": prov,
            "model": mdl,
            "warning": str(result.get("error") or ""),
        }
    parsed, perr = parse_employee_pack_llm_json(str(result.get("content") or ""))
    if perr or not parsed:
        return _fallback_manifest(brief, rule_spec), {
            "provider": prov,
            "model": mdl,
            "warning": perr or "parse failed",
        }
    return _normalize_manifest(parsed, brief, rule_spec), {"provider": prov, "model": mdl}


def _rule_spec_python_literal(rule_spec: Dict[str, Any]) -> str:
    """Embed rule_spec as valid Python dict literal (json.dumps uses true/false/null)."""
    raw = json.dumps(rule_spec, ensure_ascii=False, indent=2)
    return raw.replace(": true", ": True").replace(": false", ": False").replace(": null", ": None")


def render_direct_python_asset_worker(
    *, employee_id: str, label: str, runtime_module: str, rule_spec: Dict[str, Any]
) -> str:
    prompt = (
        f"你是{label}。你必须按 direct_python 方式处理真实文件，读取 payload 中的 file_path/path/excel_path，"
        "必要时使用打包模板，成功条件是实际写出输出文件。任何输入缺失、模板缺失、转换模块异常都要返回明确错误，禁止编造已完成。"
    )
    rule_spec_lit = _rule_spec_python_literal(rule_spec)
    return f'''"""Generated direct_python employee entrypoint."""
from __future__ import annotations

import asyncio
import json
from pathlib import Path
from typing import Any, Dict, List, Optional
import sys

EMPLOYEE_ID = {json.dumps(employee_id, ensure_ascii=False)}
EMPLOYEE_LABEL = {json.dumps(label, ensure_ascii=False)}
SYSTEM_PROMPT = {json.dumps(prompt, ensure_ascii=False)}
RULE_SPEC = {rule_spec_lit}


def _ok(data: Any, *, warnings: Optional[List[str]] = None, meta: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    return {{"ok": True, "summary": _summary(data), "items": data if isinstance(data, list) else [data], "warnings": list(warnings or []), "error": "", "meta": dict(meta or {{}})}}


def _err(msg: str, *, warnings: Optional[List[str]] = None, meta: Optional[Dict[str, Any]] = None) -> Dict[str, Any]:
    return {{"ok": False, "summary": msg[:400], "items": [], "warnings": list(warnings or []), "error": msg[:1000], "meta": dict(meta or {{}})}}


def _summary(data: Any) -> str:
    if isinstance(data, str):
        return data[:4000]
    try:
        return json.dumps(data, ensure_ascii=False)[:4000]
    except TypeError:
        return str(data)[:4000]


def _pack_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _workspace_root(ctx: Dict[str, Any], payload: Dict[str, Any]) -> Path:
    raw = payload.get("workspace_root") or ctx.get("workspace_root") or Path.cwd()
    return Path(str(raw)).expanduser()


def _resolve_input(payload: Dict[str, Any], ctx: Dict[str, Any]) -> Path:
    raw = str(payload.get("file_path") or payload.get("path") or payload.get("excel_path") or "").strip()
    if not raw:
        raise FileNotFoundError("缺少 file_path：请上传或指定要处理的文件。")
    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = _workspace_root(ctx, payload) / raw
    if not p.is_file():
        raise FileNotFoundError(f"文件不存在：{{p}}")
    return p


def _resolve_output(payload: Dict[str, Any], ctx: Dict[str, Any]) -> Path:
    rel = str(payload.get("output_relpath") or RULE_SPEC.get("default_output_relpath") or "outputs/employee_output.xlsx").strip()
    p = Path(rel).expanduser()
    if not p.is_absolute():
        p = _workspace_root(ctx, payload) / rel
    p.parent.mkdir(parents=True, exist_ok=True)
    return p


def _resolve_template(payload: Dict[str, Any], ctx: Dict[str, Any]) -> Optional[Path]:
    raw = str(
        payload.get("template_relpath")
        or RULE_SPEC.get("default_template_relpath")
        or RULE_SPEC.get("template_relpath")
        or ""
    ).strip()
    if not raw:
        return None
    candidates = []
    p = Path(raw).expanduser()
    if p.is_absolute():
        candidates.append(p)
    else:
        candidates.append(_workspace_root(ctx, payload) / raw)
        candidates.append(_pack_root() / raw)
        candidates.append(_pack_root() / "backend" / "templates" / raw)
        if raw.startswith("backend/"):
            candidates.append(_pack_root() / raw[len("backend/"):])
    for cand in candidates:
        if cand.is_file():
            return cand
    bundled_templates = sorted((_pack_root() / "templates").rglob("*.xls*")) if (_pack_root() / "templates").is_dir() else []
    if bundled_templates:
        return bundled_templates[0]
    return None


async def run(payload: Dict[str, Any], ctx: Dict[str, Any]) -> Dict[str, Any]:
    payload = dict(payload or {{}})
    ctx = dict(ctx or {{}})
    action = str(payload.get("action") or RULE_SPEC.get("default_action") or "convert").strip().lower()
    if action in ("help", "说明", "status"):
        return _ok({{"employee": EMPLOYEE_LABEL, "rule_spec": RULE_SPEC}}, meta={{"handler": "direct_python", "action": "help"}})
    if action not in ("convert", "upload", "转换", ""):
        return _err(f"不支持的 action：{{action}}", meta={{"handler": "direct_python", "action": action}})
    try:
        vendor_dir = _pack_root() / "vendor"
        if str(vendor_dir) not in sys.path:
            sys.path.insert(0, str(vendor_dir))
        from {runtime_module}.convert import convert_file
        src = _resolve_input(payload, ctx)
        out = _resolve_output(payload, ctx)
        template = _resolve_template(payload, ctx)
        result = convert_file(src, out, template_path=template, payload=payload, ctx=ctx, rule_spec=RULE_SPEC)
        if asyncio.iscoroutine(result):
            result = await result
        if isinstance(result, dict):
            result.setdefault("output_path", str(out))
            result.setdefault("template_path", str(template or ""))
        else:
            result = {{"output_path": str(out), "template_path": str(template or ""), "result": result}}
        if not out.is_file():
            return _err(f"转换未生成输出文件：{{out}}", meta={{"handler": "direct_python", "action": "convert"}})
        normalized = _ok(result, meta={{"handler": "direct_python", "action": "convert", "runtime": "generated_python"}})
        return {{
            "ok": normalized["ok"],
            "summary": normalized["summary"],
            "items": normalized["items"],
            "warnings": normalized["warnings"],
            "error": normalized["error"],
            "meta": normalized["meta"],
        }}
    except Exception as exc:  # noqa: BLE001
        return _err(str(exc), warnings=["请检查输入文件、模板文件和题目规则是否匹配。"], meta={{"handler": "direct_python", "action": "convert", "runtime": "generated_python"}})
'''


def _fallback_convert_module() -> str:
    return """from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, List, Optional


def convert_file(src_path: Path, output_path: Path, *, template_path: Optional[Path], payload: Dict[str, Any], ctx: Dict[str, Any], rule_spec: Dict[str, Any]) -> Dict[str, Any]:
    suffix = src_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls"}:
        raise ValueError(f"不支持的文件类型：{suffix or '(无后缀)'}")
    from openpyxl import load_workbook, Workbook
    src_wb = load_workbook(src_path, data_only=True)
    src_ws = src_wb.active
    src_rows = src_ws.max_row or 0
    src_cols = src_ws.max_column or 0
    headers: List[str] = []
    if src_rows > 0:
        for col in range(1, min(src_cols + 1, 51)):
            val = src_ws.cell(row=1, column=col).value
            headers.append(str(val) if val is not None else "")
    data_rows = max(0, src_rows - 1)
    if template_path and template_path.is_file():
        try:
            wb = load_workbook(template_path)
        except Exception as exc:  # noqa: BLE001
            raise RuntimeError(f"读取模板失败：{template_path}: {exc}") from exc
    else:
        wb = Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]
    if "转换结果" in wb.sheetnames:
        ws = wb["转换结果"]
    else:
        ws = wb.create_sheet("转换结果")
    ws["A1"] = "源文件"
    ws["B1"] = src_path.name
    ws["A2"] = "源行数"
    ws["B2"] = data_rows
    ws["A3"] = "源列数"
    ws["B3"] = src_cols
    ws["A4"] = "规则摘要"
    ws["B4"] = str(rule_spec.get("brief") or "")[:200]
    ws["A5"] = "状态"
    ws["B5"] = "已根据上传资产生成 direct_python 员工并写出结果"
    start_row = 7
    for idx, h in enumerate(headers):
        ws.cell(row=start_row, column=idx + 1, value=h)
    for row_idx in range(2, min(src_rows + 1, start_row + 1000)):
        for col_idx in range(1, min(src_cols + 1, 51)):
            val = src_ws.cell(row=row_idx, column=col_idx).value
            ws.cell(row=start_row + row_idx - 1, column=col_idx, value=val)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {
        "output_path": str(output_path),
        "output_relpath": str(rule_spec.get("default_output_relpath") or output_path.name),
        "template_path": str(template_path or ""),
        "source_rows": data_rows,
        "source_cols": src_cols,
        "stat_rows": data_rows,
        "headers": headers[:20],
    }
"""


def render_runtime_modules(
    rule_spec: Dict[str, Any], generated_convert_py: Optional[str] = None
) -> Dict[str, str]:
    runtime_kind = rule_spec.get("runtime_kind") or ""
    if runtime_kind == "csv_full_read" and not (generated_convert_py or "").strip():
        convert_py = render_csv_read_convert_module()
    elif runtime_kind == "csv_generate" and not (generated_convert_py or "").strip():
        convert_py = render_csv_generate_convert_module()
    elif runtime_kind == "excel_full_read" and not (generated_convert_py or "").strip():
        convert_py = render_excel_read_convert_module()
    elif runtime_kind == "excel_generate" and not (generated_convert_py or "").strip():
        convert_py = render_excel_generate_convert_module()
    elif runtime_kind == "txt_full_read" and not (generated_convert_py or "").strip():
        convert_py = render_txt_read_convert_module()
    elif runtime_kind == "txt_generate" and not (generated_convert_py or "").strip():
        convert_py = render_txt_generate_convert_module()
    elif runtime_kind == "pdf_full_read" and not (generated_convert_py or "").strip():
        convert_py = render_pdf_read_convert_module()
    elif runtime_kind == "pdf_generate" and not (generated_convert_py or "").strip():
        convert_py = render_pdf_generate_convert_module()
    elif runtime_kind == "ppt_full_read" and not (generated_convert_py or "").strip():
        convert_py = render_ppt_read_convert_module()
    elif runtime_kind == "ppt_generate" and not (generated_convert_py or "").strip():
        convert_py = render_ppt_generate_convert_module()
    elif runtime_kind == "json_quant_report" and not (generated_convert_py or "").strip():
        convert_py = render_json_report_convert_module()
    elif runtime_kind == "word_full_extract" and not (generated_convert_py or "").strip():
        convert_py = render_word_fallback_convert_module()
    elif runtime_kind == "word_generate" and not (generated_convert_py or "").strip():
        convert_py = render_word_generate_convert_module()
    else:
        convert_py = (generated_convert_py or "").strip() or _fallback_convert_module()
    modules: Dict[str, str] = {
        "__init__.py": '"""Generated runtime modules for asset-driven employee."""\n',
        "convert.py": convert_py,
        "parser.py": '"""Parser extension point generated by asset pipeline."""\n',
        "mapper.py": '"""Mapper extension point generated by asset pipeline."""\n',
        "rules.py": '"""Rules extension point generated by asset pipeline."""\n',
        "paths.py": '"""Path helpers generated by asset pipeline."""\n',
        "mapping.py": '"""Mapping helpers generated by asset pipeline."""\n',
        "header_resolver.py": '"""Header resolver generated by asset pipeline."""\n',
    }
    if runtime_kind == "word_full_extract":
        from modstore_server.legacy_doc_convert import render_legacy_doc_vendor_module

        modules["legacy_doc.py"] = render_legacy_doc_vendor_module()
    return modules


def render_build_xcemp_py(pack_id: str) -> str:
    return f'''"""Build {pack_id}.xcemp from this employee_pack directory."""
from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path


ROOT = Path(__file__).resolve().parent
PACK_ID = {json.dumps(pack_id, ensure_ascii=False)}


def main() -> None:
    manifest = json.loads((ROOT / "manifest.json").read_text(encoding="utf-8"))
    out = ROOT / f"{{PACK_ID}}.xcemp"
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(f"{{PACK_ID}}/manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2) + "\\n")
        for path in sorted((ROOT / "backend").rglob("*")):
            if not path.is_file():
                continue
            if path.suffix.lower() not in {{".py", ".xlsx", ".xlsm", ".xls"}}:
                continue
            rel = path.relative_to(ROOT).as_posix()
            zf.write(path, f"{{PACK_ID}}/{{rel}}")
        readme = ROOT / "README.md"
        if readme.is_file():
            zf.write(readme, f"{{PACK_ID}}/README.md")
    out.write_bytes(buf.getvalue())
    print(out)


if __name__ == "__main__":
    main()
'''


def _extract_python_code(text: str) -> str:
    raw = (text or "").strip()
    if not raw:
        return ""
    try:
        from modstore_server.script_agent.llm_client import extract_code_block

        extracted = extract_code_block(raw, lang="python").strip()
        if extracted:
            return extracted
    except Exception:
        pass
    match = re.search(r"```(?:python|py)?\s*(.*?)```", raw, re.S | re.I)
    if match:
        return match.group(1).strip()
    lines = raw.splitlines()
    for i, line in enumerate(lines):
        stripped = line.strip()
        if stripped.startswith(("from ", "import ", "def ", "class ", "@")) or stripped in {
            "# -*- coding: utf-8 -*-",
            "from __future__ import annotations",
        }:
            return "\n".join(lines[i:]).strip()
    return raw


def _validate_generated_convert_py(src: str) -> Tuple[bool, str]:
    code = (src or "").strip()
    if not code:
        return False, "empty generated convert.py"
    if re.search(r"\b(eval|exec|compile|__import__)\s*\(", code):
        return False, "generated convert.py uses forbidden dynamic execution"
    if re.search(r"\b(subprocess|os\.system|ctypes|multiprocessing)\b", code):
        return False, "generated convert.py uses forbidden process/system API"
    if re.search(r"\b(globals|locals|getattr|setattr|delattr|breakpoint)\s*\(", code):
        return False, "generated convert.py uses forbidden reflection/builtin"
    try:
        tree = ast.parse(code)
    except SyntaxError as exc:
        return False, f"generated convert.py syntax error: {exc}"
    has_convert = any(
        isinstance(node, ast.FunctionDef) and node.name == "convert_file" for node in tree.body
    )
    if not has_convert:
        return False, "generated convert.py must define convert_file(...)"
    for node in ast.walk(tree):
        if isinstance(node, ast.Import):
            for alias in node.names:
                if alias.name in ("subprocess", "ctypes", "multiprocessing"):
                    return False, f"generated convert.py imports forbidden module: {alias.name}"
        if isinstance(node, ast.ImportFrom):
            if node.module in ("subprocess", "ctypes", "multiprocessing"):
                return False, f"generated convert.py imports from forbidden module: {node.module}"
    return True, ""


def _auto_fix_generated_convert_py(src: str) -> Tuple[str, List[str]]:
    fixes: List[str] = []
    code = (src or "").strip()
    if not code:
        return code, fixes
    lines = code.splitlines()
    filtered: List[str] = []
    skip_patterns = [
        re.compile(r"\b(eval|exec|compile|__import__)\s*\("),
        re.compile(r"\bimport\s+(subprocess|ctypes|multiprocessing)\b"),
        re.compile(r"\bfrom\s+(subprocess|ctypes|multiprocessing)\s+import\b"),
        re.compile(r"\b(globals|locals|getattr|setattr|delattr|breakpoint)\s*\("),
    ]
    for line in lines:
        stripped = line.strip()
        if any(p.search(stripped) for p in skip_patterns):
            fixes.append(f"removed: {stripped[:80]}")
            continue
        filtered.append(line)
    return "\n".join(filtered), fixes


async def generate_runtime_convert_module(
    db: Any,
    user: User,
    *,
    brief: str,
    rule_spec: Dict[str, Any],
    asset_manifest: Dict[str, Any],
    provider: Optional[str],
    model: Optional[str],
    force_llm_codegen: bool = False,
    allow_builtin_codegen: bool = False,
    payload: Optional[Dict[str, Any]] = None,
) -> Tuple[Optional[str], Dict[str, Any]]:
    """Use the configured coding model to make the asset runtime real.

    This is the workbench's vibecoding step for asset employees: the platform
    supplies inspected assets and a strict runtime contract; the model writes
    only the transform module.
    """
    _force_llm = bool(force_llm_codegen)
    _allow_builtin = bool(allow_builtin_codegen)
    runtime_kind = rule_spec.get("runtime_kind") or ""
    if _allow_builtin and runtime_kind == "csv_full_read":
        return render_csv_read_convert_module(), {
            "provider": "",
            "model": "",
            "source": "csv_read_builtin",
        }
    if _allow_builtin and runtime_kind == "csv_generate":
        return render_csv_generate_convert_module(), {
            "provider": "",
            "model": "",
            "source": "csv_generate_builtin",
        }
    if _allow_builtin and runtime_kind == "excel_full_read":
        return render_excel_read_convert_module(), {
            "provider": "",
            "model": "",
            "source": "excel_read_builtin",
        }
    if _allow_builtin and runtime_kind == "excel_generate":
        return render_excel_generate_convert_module(), {
            "provider": "",
            "model": "",
            "source": "excel_generate_builtin",
        }
    if _allow_builtin and runtime_kind == "txt_full_read":
        return render_txt_read_convert_module(), {
            "provider": "",
            "model": "",
            "source": "txt_read_builtin",
        }
    if _allow_builtin and runtime_kind == "txt_generate":
        return render_txt_generate_convert_module(), {
            "provider": "",
            "model": "",
            "source": "txt_generate_builtin",
        }
    if _allow_builtin and runtime_kind == "pdf_full_read":
        return render_pdf_read_convert_module(), {
            "provider": "",
            "model": "",
            "source": "pdf_read_builtin",
        }
    if _allow_builtin and runtime_kind == "pdf_generate":
        return render_pdf_generate_convert_module(), {
            "provider": "",
            "model": "",
            "source": "pdf_generate_builtin",
        }
    if _allow_builtin and runtime_kind == "ppt_full_read":
        return render_ppt_read_convert_module(), {
            "provider": "",
            "model": "",
            "source": "ppt_read_builtin",
        }
    if _allow_builtin and runtime_kind == "ppt_generate":
        return render_ppt_generate_convert_module(), {
            "provider": "",
            "model": "",
            "source": "ppt_generate_builtin",
        }
    if _allow_builtin and runtime_kind == "json_quant_report":
        return render_json_report_convert_module(), {
            "provider": "",
            "model": "",
            "source": "json_quant_report_builtin",
        }
    if _allow_builtin and runtime_kind == "word_full_extract" and not _force_llm:
        return render_word_fallback_convert_module(), {
            "provider": "",
            "model": "",
            "source": "word_extract_builtin",
        }
    if _allow_builtin and runtime_kind == "word_generate" and not _force_llm:
        return render_word_generate_convert_module(), {
            "provider": "",
            "model": "",
            "source": "word_generate_builtin",
        }
    prov, mdl, err = await resolve_llm_provider_model_auto(db, user, provider, model)
    if err:
        return None, {"provider": "", "model": "", "warning": err}
    api_key, _ = resolve_api_key(db, user.id, prov)  # type: ignore[arg-type]
    if not api_key:
        return None, {"provider": prov, "model": mdl, "warning": "missing api key"}
    base = (
        resolve_base_url(db, user.id, prov) if prov in OAI_COMPAT_OPENAI_STYLE_PROVIDERS else None
    )
    if runtime_kind == "word_full_extract":
        from modstore_server.employee_ai_pipeline import _build_vibe_coding_prompt

        system = _build_vibe_coding_prompt(runtime_kind, rule_spec)
        contract = {
            "input": (
                "src_path may be .docx (OOXML) or legacy .doc (OLE). "
                "For .doc or misnamed binary, call legacy_doc.ensure_docx_for_extract(src, work_dir) first, "
                "then parse the returned .docx via OOXML (zipfile). Do not fabricate document text."
            ),
            "output": (
                "write outputs/document_full.json with paragraphs, tables, outline, blocks, sections, "
                "images metadata, styles, headers_footers, core_properties, comments, metadata, plain_text; "
                "also document_full.txt and export images under outputs/images/. "
                "Record metadata.legacy_doc when conversion happened."
            ),
            "template": "template_path is usually None for Word extract. vendor includes legacy_doc.py.",
        }
        max_tokens = 16000
    else:
        system = (
            "你是工作台 vibecoding 的 Python 实现器。只输出一个 Python 代码块，内容是 backend/vendor/<runtime>/convert.py。"
            "必须定义 convert_file(src_path: Path, output_path: Path, *, template_path: Optional[Path], payload: Dict[str, Any], ctx: Dict[str, Any], rule_spec: Dict[str, Any]) -> Dict[str, Any]。"
            "必须真实读取输入 Excel、按模板/规则写出 output_path。不能调用 LLM，不能写伪结果。"
            "绝对禁止使用以下任何一种写法，否则代码将被拒绝：\n"
            "  - eval(...) / exec(...) / compile(...) / __import__(...)\n"
            "  - import subprocess / import os 后调用 os.system / import ctypes / import multiprocessing\n"
            "  - globals() / locals() / getattr(...) / setattr(...) / delattr(...) / input(...) / breakpoint()\n"
            "  - 任何形式的动态代码执行或反射调用\n"
            "允许使用 pathlib、json、datetime、re、typing、openpyxl、pandas、copy、collections、io。异常要抛出清晰错误。\n"
            "正确示例：\n"
            "  from pathlib import Path\n"
            "  from openpyxl import load_workbook\n"
            "  def convert_file(src_path, output_path, *, template_path=None, payload=None, ctx=None, rule_spec=None):\n"
            "      wb = load_workbook(src_path)\n"
            "      # ... 处理逻辑 ...\n"
            "      wb.save(output_path)\n"
            "      return {'output_path': str(output_path), 'stat_rows': 1}\n"
            "错误示例（会被拒绝）：\n"
            "  exec('print(1)')  # 禁止\n"
            "  __import__('os')   # 禁止\n"
            "  getattr(obj, 'x') # 禁止\n"
        )
        contract = {
            "input": "src_path points to the uploaded Excel file.",
            "template": "template_path may point to a bundled template workbook, or may be None.",
            "output": "write the final workbook to output_path and return useful stats.",
        }
        max_tokens = 8000
    user_msg = json.dumps(
        {
            "brief": brief,
            "rule_spec": rule_spec,
            "asset_manifest": {
                **{k: v for k, v in asset_manifest.items() if k != "assets"},
                "assets": [
                    {
                        k: v
                        for k, v in dict(item).items()
                        if k in {"id", "filename", "kind", "suffix", "size", "excel", "preview"}
                    }
                    for item in (asset_manifest.get("assets") or [])
                    if isinstance(item, dict)
                ],
            },
            "contract": contract,
        },
        ensure_ascii=False,
    )[:20000]
    result = await chat_dispatch(
        prov,
        api_key=api_key,
        base_url=base,
        model=mdl,
        messages=[{"role": "system", "content": system}, {"role": "user", "content": user_msg}],
        max_tokens=max_tokens,
    )
    if not result.get("ok"):
        return None, {"provider": prov, "model": mdl, "warning": str(result.get("error") or "")}
    code = _extract_python_code(str(result.get("content") or ""))
    ok, validation_error = _validate_generated_convert_py(code)
    if not ok:
        if _allow_builtin:
            fixed_code, fixes = _auto_fix_generated_convert_py(code)
            fixed_ok, fixed_error = _validate_generated_convert_py(fixed_code)
            if fixed_ok:
                return fixed_code.rstrip() + "\n", {
                    "provider": prov,
                    "model": mdl,
                    "generated": True,
                    "auto_fixed": True,
                    "fixes": fixes,
                    "source": "llm_codegen",
                }
        return None, {"provider": prov, "model": mdl, "warning": validation_error}
    return code.rstrip() + "\n", {
        "provider": prov,
        "model": mdl,
        "generated": True,
        "source": "llm_codegen",
    }


async def repair_runtime_convert_module(
    db: Any,
    user: User,
    *,
    brief: str,
    rule_spec: Dict[str, Any],
    previous_convert_py: str,
    failure: Dict[str, Any],
    provider: Optional[str],
    model: Optional[str],
    round_no: int,
    allow_auto_fix: bool = False,
) -> Tuple[Optional[str], Dict[str, Any]]:
    prov, mdl, err = await resolve_llm_provider_model_auto(db, user, provider, model)
    if err:
        return None, {"provider": "", "model": "", "warning": err, "round": round_no}
    api_key, _ = resolve_api_key(db, user.id, prov)  # type: ignore[arg-type]
    if not api_key:
        return None, {
            "provider": prov,
            "model": mdl,
            "warning": "missing api key",
            "round": round_no,
        }
    base = (
        resolve_base_url(db, user.id, prov) if prov in OAI_COMPAT_OPENAI_STYLE_PROVIDERS else None
    )
    system = (
        "你是 Python 代码修复器。只输出修复后的 convert.py Python 代码块。"
        "必须保留 convert_file 签名，必须真实读取 src_path/template_path 并保存 output_path。"
        "如果业务映射复杂，最低要求也必须基于模板 workbook 写出一个有效 xlsx 到 output_path，并返回统计信息。"
        "绝对禁止使用以下任何一种写法，否则代码将被拒绝：\n"
        "  - eval(...) / exec(...) / compile(...) / __import__(...)\n"
        "  - import subprocess / import os 后调用 os.system / import ctypes / import multiprocessing\n"
        "  - globals() / locals() / getattr(...) / setattr(...) / delattr(...) / input(...) / breakpoint()\n"
        "  - 任何形式的动态代码执行或反射调用\n"
        "允许使用 pathlib、json、datetime、re、typing、openpyxl、pandas、copy、collections、io。\n"
    )
    user_msg = json.dumps(
        {
            "round": round_no,
            "failure": failure,
            "previous_convert_py": previous_convert_py,
            "brief": brief,
            "rule_spec": rule_spec,
        },
        ensure_ascii=False,
    )[:24000]
    result = await chat_dispatch(
        prov,
        api_key=api_key,
        base_url=base,
        model=mdl,
        messages=[{"role": "system", "content": system}, {"role": "user", "content": user_msg}],
        max_tokens=8000,
    )
    if not result.get("ok"):
        return None, {
            "provider": prov,
            "model": mdl,
            "warning": str(result.get("error") or ""),
            "round": round_no,
        }
    code = _extract_python_code(str(result.get("content") or ""))
    ok, validation_error = _validate_generated_convert_py(code)
    if not ok:
        if allow_auto_fix:
            fixed_code, fixes = _auto_fix_generated_convert_py(code)
            fixed_ok, _fixed_error = _validate_generated_convert_py(fixed_code)
            if fixed_ok:
                return fixed_code.rstrip() + "\n", {
                    "provider": prov,
                    "model": mdl,
                    "repaired": True,
                    "round": round_no,
                    "auto_fixed": True,
                    "fixes": fixes,
                }
        return None, {
            "provider": prov,
            "model": mdl,
            "warning": validation_error,
            "round": round_no,
        }
    return code.rstrip() + "\n", {
        "provider": prov,
        "model": mdl,
        "repaired": True,
        "round": round_no,
    }


def manifest_actions_handlers(mf: Dict[str, Any]) -> List[str]:
    """Read handlers from employee_config_v2.actions or canvas-root actions."""
    v2 = mf.get("employee_config_v2") if isinstance(mf.get("employee_config_v2"), dict) else {}
    actions = v2.get("actions") if isinstance(v2.get("actions"), dict) else {}
    if not actions and isinstance(mf.get("actions"), dict):
        actions = mf["actions"]
    raw = actions.get("handlers") if isinstance(actions.get("handlers"), list) else []
    return [str(h).strip() for h in raw if str(h).strip()]


def manifest_expects_word_runtime(mf: Dict[str, Any], *, brief: str = "") -> bool:
    """True when manifest/brief indicates Word 全量提取 direct_python delivery."""
    rs_path_inline = mf.get("rule_spec") if isinstance(mf.get("rule_spec"), dict) else {}
    if rs_path_inline.get("runtime_kind") == "word_full_extract":
        return True
    from modstore_server.employee_brief_utils import extract_routing_brief

    rb = (brief or "").strip() or extract_routing_brief(
        {"brief": str(mf.get("description") or "")},
        fallback=str(mf.get("description") or ""),
    )
    if is_word_full_extract(rb):
        return True
    perception = mf.get("perception")
    if not isinstance(perception, dict):
        v2 = mf.get("employee_config_v2") if isinstance(mf.get("employee_config_v2"), dict) else {}
        perception = v2.get("perception") if isinstance(v2.get("perception"), dict) else {}
    exts = (
        perception.get("accepted_extensions")
        if isinstance(perception.get("accepted_extensions"), list)
        else []
    )
    ext_l = {str(x).lower() for x in exts}
    if ext_l & {".docx", ".doc"} and "direct_python" in manifest_actions_handlers(mf):
        return True
    return False


def pack_has_direct_python_runtime(pack_dir: Path) -> bool:
    """Disk pack contains vendor convert and/or rule_spec for asset/direct_python employees."""
    rs = pack_dir / "rule_spec.json"
    if rs.is_file():
        try:
            data = json.loads(rs.read_text(encoding="utf-8"))
            if isinstance(data, dict) and data.get("runtime_kind") in (
                "word_full_extract",
                "txt_full_read",
                "txt_generate",
                "pdf_full_read",
                "pdf_generate",
                "csv_full_read",
                "csv_generate",
                "generic_excel_transform",
                "contract_doc_review",
                "doc_template_transform",
            ):
                return True
        except (OSError, json.JSONDecodeError):
            pass
    backend = pack_dir / "backend"
    if not backend.is_dir():
        return False
    for py_path in backend.rglob("*.py"):
        try:
            text = py_path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            continue
        if "def convert_file" in text and "vendor" in py_path.as_posix().lower():
            return True
        if "def convert" in text and "_import_runtime" in text:
            return True
    return False


DIRECT_PYTHON_RUNTIME_MISSING_MSG = (
    "manifest 声明了 Word/direct_python，但本地库中缺少 rule_spec 与 backend/vendor/convert。"
    "请在工作台「做员工」流水线完成 generate 步后再在浏览室保存；"
    "画布保存不能替代资产生成，否则会覆盖为仅含 LLM 脚手架的空包。"
)


def persist_manifest_to_pack_dir(
    pack_dir: Path, manifest: Dict[str, Any], *, brief: str = ""
) -> Dict[str, Any]:
    """Write manifest.json on disk; reconcile Word packs when rule_spec exists or brief matches."""
    pack_dir.mkdir(parents=True, exist_ok=True)
    mf_path = pack_dir / "manifest.json"
    mf_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    if (
        manifest_expects_word_runtime(manifest, brief=brief)
        or (pack_dir / "rule_spec.json").is_file()
    ):
        return reconcile_employee_pack_manifest(pack_dir, brief=brief)
    return manifest


def build_employee_pack_zip_for_library(
    pack_id: str,
    manifest: Dict[str, Any],
    *,
    pack_dir: Optional[Path] = None,
    brief: str = "",
) -> bytes:
    """Build .xcemp bytes: prefer on-disk vendor/runtime; never strip Word packs to template-only."""
    from modstore_server.employee_ai_scaffold import build_employee_pack_zip

    pd = pack_dir or (modstore_library_path() / pack_id)
    handlers = manifest_actions_handlers(manifest)
    wants_word = manifest_expects_word_runtime(manifest, brief=brief)
    if "direct_python" in handlers and wants_word:
        if not pd.is_dir() or not pack_has_direct_python_runtime(pd):
            raise ValueError(DIRECT_PYTHON_RUNTIME_MISSING_MSG)
    if pd.is_dir() and (pd / "manifest.json").is_file():
        persist_manifest_to_pack_dir(pd, manifest, brief=brief)
        if pack_has_direct_python_runtime(pd) or any(
            p.is_file()
            for p in pd.rglob("*")
            if p.suffix.lower() in {".py", ".json"} and "__pycache__" not in p.parts
        ):
            try:
                return build_employee_pack_zip_from_dir(pack_id, pd)
            except Exception:
                pass
    return build_employee_pack_zip(pack_id, manifest)


def build_employee_pack_zip_from_dir(pack_id: str, pack_dir: Path) -> bytes:
    mf_path = pack_dir / "manifest.json"
    if mf_path.is_file():
        try:
            _raw = json.loads(mf_path.read_text(encoding="utf-8"))
            _dirty = False
            if str(_raw.get("id") or "") != pack_id:
                _raw["id"] = pack_id
                _dirty = True
            if (
                isinstance(_raw.get("employee"), dict)
                and str(_raw["employee"].get("id") or "") != pack_id
            ):
                _raw["employee"]["id"] = pack_id
                _dirty = True
            for _r in _raw.get("workflow_employees") or []:
                if isinstance(_r, dict) and str(_r.get("id") or "") != pack_id:
                    _r["id"] = pack_id
                    _dirty = True
            if _dirty:
                mf_path.write_text(
                    json.dumps(_raw, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
                )
        except Exception:
            pass
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(pack_dir.rglob("*")):
            if not path.is_file():
                continue
            rel = path.relative_to(pack_dir).as_posix()
            if "__pycache__" in Path(rel).parts:
                continue
            if Path(rel).suffix.lower() not in {
                ".json",
                ".md",
                ".py",
                ".xlsx",
                ".xlsm",
                ".xls",
                ".txt",
                ".yaml",
                ".yml",
            }:
                continue
            zf.write(path, f"{pack_id}/{rel}")
    return buf.getvalue()


def mirror_catalog_file_to_market_files(stored_filename: str) -> None:
    """Keep the browsable market_files copy aligned with catalog_data/files."""
    name = _safe_basename(stored_filename, "")
    if not name:
        return
    from modstore_server.catalog_store import files_dir

    src = files_dir() / name
    if not src.is_file():
        return
    dest_dir = Path(__file__).resolve().parent / "market_files"
    dest_dir.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest_dir / name)


def _copy_template_assets(
    pack_dir: Path, asset_manifest: Dict[str, Any], rule_spec: Dict[str, Any]
) -> None:
    assets = asset_manifest.get("templates") or []
    if not assets:
        assets = [
            a for a in asset_manifest.get("assets") or [] if a.get("suffix") in EXCEL_SUFFIXES
        ][:1]
    for asset in assets:
        src = Path(str(asset.get("path") or ""))
        if src.is_file():
            filename = str(asset.get("filename") or src.name)
            rel = str(
                rule_spec.get("template_relpath")
                or _template_storage_relpath(filename, str(rule_spec.get("brief") or ""))
            )
            dest = pack_dir / rel
            dest.parent.mkdir(parents=True, exist_ok=True)
            shutil.copy2(src, dest)


def materialize_asset_employee_pack(
    *,
    manifest: Dict[str, Any],
    rule_spec: Dict[str, Any],
    asset_manifest: Dict[str, Any],
    generated_convert_py: Optional[str] = None,
) -> Tuple[Path, bytes]:
    pack_id = str(manifest.get("id") or "").strip()
    if not pack_id:
        raise ValueError("manifest.id 缺失")
    emp = manifest.get("employee") if isinstance(manifest.get("employee"), dict) else {}
    employee_id = pack_id
    label = str(emp.get("label") or manifest.get("name") or employee_id).strip() or employee_id
    stem = sanitize_employee_stem(employee_id)
    runtime_mod = _runtime_package_name(pack_id, employee_id)
    runtime_kind = rule_spec.get("runtime_kind") or "generic_excel_transform"
    _is_doc_review = runtime_kind in ("contract_doc_review", "doc_template_transform")
    tmp_dir = Path(tempfile.mkdtemp(prefix=f"asset_emp_{pack_id}_"))
    pack_dir = tmp_dir / pack_id
    (pack_dir / "backend" / "employees").mkdir(parents=True, exist_ok=True)
    (pack_dir / "manifest.json").write_text(
        json.dumps(manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    _readme_desc = "LLM 驱动文档审核" if _is_doc_review else "direct_python"
    (pack_dir / "README.md").write_text(
        "# " + label + f"\n\n由上传资产生成的 {_readme_desc} 员工包。\n", encoding="utf-8"
    )
    (pack_dir / "build_xcemp.py").write_text(render_build_xcemp_py(pack_id), encoding="utf-8")
    (pack_dir / "backend" / "blueprints.py").write_text(
        render_employee_pack_blueprints_py(
            pack_id=pack_id, employee_id=employee_id, stem=stem, label=label
        ),
        encoding="utf-8",
    )
    (pack_dir / "backend" / "employees" / "__init__.py").write_text(
        '"""Generated employees."""\n', encoding="utf-8"
    )
    if _is_doc_review:
        from modstore_server.employee_pack_blueprints_template import (
            render_employee_pack_employee_py,
        )

        (pack_dir / "backend" / "employees" / f"{stem}.py").write_text(
            render_employee_pack_employee_py(employee_id=employee_id, stem=stem, label=label),
            encoding="utf-8",
        )
    else:
        (pack_dir / "backend" / "vendor" / runtime_mod).mkdir(parents=True, exist_ok=True)
        (pack_dir / "backend" / "employees" / f"{stem}.py").write_text(
            render_direct_python_asset_worker(
                employee_id=employee_id,
                label=label,
                runtime_module=runtime_mod,
                rule_spec=rule_spec,
            ),
            encoding="utf-8",
        )
        for name, src in render_runtime_modules(
            rule_spec, generated_convert_py=generated_convert_py
        ).items():
            (pack_dir / "backend" / "vendor" / runtime_mod / name).write_text(src, encoding="utf-8")
    _copy_template_assets(pack_dir, asset_manifest, rule_spec)
    (pack_dir / "asset_manifest.json").write_text(
        json.dumps(asset_manifest, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    (pack_dir / "rule_spec.json").write_text(
        json.dumps(rule_spec, ensure_ascii=False, indent=2) + "\n", encoding="utf-8"
    )
    raw = build_employee_pack_zip_from_dir(pack_id, pack_dir)
    return pack_dir, raw


def validate_asset_employee_pack(pack_dir: Path, manifest: Dict[str, Any]) -> List[str]:
    warnings: List[str] = []
    v2 = (
        manifest.get("employee_config_v2")
        if isinstance(manifest.get("employee_config_v2"), dict)
        else {}
    )
    actions = v2.get("actions") if isinstance(v2.get("actions"), dict) else {}
    top_actions = manifest.get("actions") if isinstance(manifest.get("actions"), dict) else {}
    top_handlers = (
        top_actions.get("handlers") if isinstance(top_actions.get("handlers"), list) else []
    )
    v2_handlers = actions.get("handlers") if isinstance(actions.get("handlers"), list) else []
    _is_agent_handler = "agent" in v2_handlers
    if not _is_agent_handler and actions.get("handlers") != ["direct_python"]:
        warnings.append("actions.handlers 必须为 ['direct_python'] 或 ['agent']")
    if top_handlers and v2_handlers and top_handlers != v2_handlers:
        warnings.append(
            f"顶层 actions.handlers={top_handlers} 与 v2 actions.handlers={v2_handlers} 不一致"
        )
    emp_dir = pack_dir / "backend" / "employees"
    py_files = (
        [p for p in emp_dir.glob("*.py") if p.name != "__init__.py"] if emp_dir.is_dir() else []
    )
    if not py_files:
        warnings.append("缺少 backend/employees 入口脚本")
    direct = actions.get("direct_python") if isinstance(actions.get("direct_python"), dict) else {}
    module_name = str(direct.get("module") or "").strip()
    if module_name:
        expected_file = emp_dir / f"{module_name}.py"
        if not expected_file.is_file():
            warnings.append(
                f"direct_python.module={module_name} 但文件 {expected_file.name} 不存在"
            )
    for pf in py_files:
        code = pf.read_text(encoding="utf-8")
        has_dispatch = "_DISPATCH" in code
        if "direct_python" in (v2_handlers or []) and has_dispatch:
            warnings.append(
                f"{pf.name} 含 _DISPATCH 字典但 handlers 声明为 direct_python，应使用 render_direct_python_asset_worker 模板"
            )
        if "direct_python" not in (v2_handlers or []) and not has_dispatch:
            pass
    for p in sorted((pack_dir / "backend").rglob("*.py")):
        try:
            py_compile.compile(str(p), doraise=True)
            ast.parse(p.read_text(encoding="utf-8"))
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"{p.relative_to(pack_dir).as_posix()}: {exc}")
    if not list((pack_dir / "backend" / "vendor").rglob("*.py")):
        if not _is_agent_handler:
            warnings.append("缺少 backend/vendor 运行模块")
    tpl = str(direct.get("default_template_relpath") or "").strip()
    if tpl:
        template_dir = pack_dir / "backend" / "templates"
        has_any_template = template_dir.is_dir() and any(
            p.suffix.lower() in EXCEL_SUFFIXES for p in template_dir.rglob("*")
        )
        if not (
            (pack_dir / tpl).is_file()
            or (pack_dir / "backend" / tpl).is_file()
            or (pack_dir / "backend" / "templates" / tpl).is_file()
            or has_any_template
        ):
            warnings.append(f"默认模板未打包：{tpl}")
    return warnings


async def run_asset_employee_scaffold_async(
    db: Any,
    user: User,
    *,
    session_id: str,
    brief: str,
    raw_files: List[Dict[str, Any]],
    replace: bool = True,
    provider: Optional[str] = None,
    model: Optional[str] = None,
    publish_to_catalog: bool = False,
    force_llm_codegen: bool = False,
    payload: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    import asyncio

    from modstore_server.artifact_generator_blueprint import artifact_generator_preflight
    from modstore_server.craft_failure_signals import (
        _employee_trigger_limits,
        emit_craft_step_failure,
    )
    from modstore_server.employee_pipeline_routing import is_direct_python_template_runtime
    from modstore_server.vibecoding_convert_loop import (
        is_llm_codegen_source,
        run_vibecoding_codegen_loop,
    )

    _limits = _employee_trigger_limits("artifact-generator")
    _max_repair_rounds = max(1, int(_limits.get("max_patch_steps") or 4))

    bp_preflight = artifact_generator_preflight(payload=payload, brief=brief)
    if bp_preflight.get("status") == "error":
        emit_craft_step_failure(
            step_id="generate",
            error=str(bp_preflight.get("error") or "上游蓝图校验失败"),
            employee_id="artifact-generator",
            user_id=int(user.id),
            extra={
                "missing_fields": bp_preflight.get("missing_fields") or [],
                "validation_result": bp_preflight.get("validation_result"),
                "downstream_context": "pack-registrar 需完整 manifest 后再入库",
            },
        )
        return {
            "ok": False,
            "error": bp_preflight.get("error"),
            "status": "error",
            "generation_mode": bp_preflight.get("generation_mode"),
            "artifact_paths": [],
            "validation_result": bp_preflight.get("validation_result"),
            "warnings": [],
            "missing_fields": bp_preflight.get("missing_fields") or [],
        }

    _paths_ok, _path_errors, _paths_checked = _preflight_scaffold_write_access(
        session_id=session_id,
        user_id=int(user.id),
    )
    if not _paths_ok:
        _perm_msg = "；".join(_path_errors[:5])
        emit_craft_step_failure(
            step_id="generate",
            error=_perm_msg,
            employee_id="artifact-generator",
            user_id=int(user.id),
            extra={
                "paths_checked": _paths_checked,
                "escalate_to_human": True,
                "downstream_context": "pack-registrar：路径不可写时勿尝试注册包",
            },
        )
        return {
            "ok": False,
            "error": _perm_msg,
            "status": "error",
            "generation_mode": "asset",
            "artifact_paths": [],
            "validation_result": {"paths": _paths_checked, "permission_errors": _path_errors},
            "warnings": _path_errors,
            "paths_checked": _paths_checked,
        }

    _brief = str(bp_preflight.get("brief_from_plan") or brief).strip() or brief

    asset_manifest = prepare_employee_assets(
        session_id=session_id, user_id=int(user.id), raw_files=raw_files
    )
    rule_spec = build_rule_spec(_brief, asset_manifest, payload=payload)
    runtime_kind = str(rule_spec.get("runtime_kind") or "")
    repair_history: List[Dict[str, Any]] = []
    domain_smoke: Dict[str, Any] = {}
    golden_comparison: Dict[str, Any] = {}

    use_vibecoding_loop = is_direct_python_template_runtime(runtime_kind) or bool(force_llm_codegen)
    base_manifest = _normalize_manifest(_fallback_manifest(_brief, rule_spec), _brief, rule_spec)

    if use_vibecoding_loop:
        manifest, llm_meta = await enrich_manifest_productivity_fields(
            db,
            user,
            brief=_brief,
            rule_spec=rule_spec,
            base_manifest=base_manifest,
            provider=provider,
            model=model,
        )
        manifest = _normalize_manifest(manifest, _brief, rule_spec)
        generated_convert_py, runtime_meta, domain_smoke, golden_comparison = (
            await run_vibecoding_codegen_loop(
                db,
                user,
                session_id=session_id,
                brief=_brief,
                rule_spec=rule_spec,
                manifest=manifest,
                asset_manifest=asset_manifest,
                provider=provider,
                model=model,
                payload=payload,
            )
        )
        if isinstance(runtime_meta.get("repair_history"), list):
            repair_history = runtime_meta["repair_history"]
    else:
        manifest, llm_meta = await design_asset_employee_manifest(
            db,
            user,
            brief=_brief,
            rule_spec=rule_spec,
            provider=provider,
            model=model,
        )
        manifest = _normalize_manifest(manifest, _brief, rule_spec)
        generated_convert_py, runtime_meta = await generate_runtime_convert_module(
            db,
            user,
            brief=_brief,
            rule_spec=rule_spec,
            asset_manifest=asset_manifest,
            provider=provider,
            model=model,
            force_llm_codegen=bool(force_llm_codegen),
            allow_builtin_codegen=False,
            payload=payload,
        )

    pack_dir, raw_zip = materialize_asset_employee_pack(
        manifest=manifest,
        rule_spec=rule_spec,
        asset_manifest=asset_manifest,
        generated_convert_py=generated_convert_py,
    )
    warnings = validate_asset_employee_pack(pack_dir, manifest)
    if use_vibecoding_loop:
        if not generated_convert_py:
            warnings.append(str(runtime_meta.get("error") or "vibecoding 未产出合格 convert"))
        elif domain_smoke.get("ok") is False and not domain_smoke.get("skipped"):
            warnings.append(f"领域冒烟失败：{domain_smoke.get('error') or ''}"[:200])
        elif golden_comparison and not golden_comparison.get("passed"):
            warnings.append(
                f"黄金对比未达标：parity={golden_comparison.get('parity_score')} "
                f"diffs={len(golden_comparison.get('diff_items') or [])}"
            )
    elif runtime_meta.get("warning"):
        warnings.append(f"vibecoding runtime：{runtime_meta['warning']}")
    while (
        warnings
        and generated_convert_py
        and provider
        and len(repair_history) < _max_repair_rounds
        and not use_vibecoding_loop
    ):
        _fail = {"errors": warnings[:8], "stage": "validate_pack"}
        repaired, repair_meta = await repair_runtime_convert_module(
            db,
            user,
            brief=_brief,
            rule_spec=rule_spec,
            previous_convert_py=generated_convert_py,
            failure=_fail,
            provider=provider,
            model=model,
            round_no=len(repair_history) + 1,
        )
        repair_history.append(repair_meta)
        if not repaired:
            break
        pack_dir, raw_zip = materialize_asset_employee_pack(
            manifest=manifest,
            rule_spec=rule_spec,
            asset_manifest=asset_manifest,
            generated_convert_py=repaired,
        )
        warnings = validate_asset_employee_pack(pack_dir, manifest)
        generated_convert_py = repaired
        runtime_meta = {**runtime_meta, "repaired_after_validate": True}
    if warnings and len(repair_history) >= _max_repair_rounds and not use_vibecoding_loop:
        _budget_msg = (
            f"动态修复预算已用尽（max_patch_steps={_max_repair_rounds}，"
            f"max_patch_budget_tokens={_limits.get('max_patch_budget_tokens')}）"
        )
        emit_craft_step_failure(
            step_id="generate",
            error=_budget_msg + "：" + "；".join(str(w) for w in warnings[:3]),
            employee_id="artifact-generator",
            user_id=int(user.id),
            extra={
                "repair_history": repair_history,
                "downstream_context": "pack-registrar：请人工复核 manifest/convert 后再注册",
                "escalate_to_human": True,
            },
        )
    if generated_convert_py:
        runtime_meta["validation"] = (
            "generated_convert_py compiled; execution validation is performed by workbench smoke/tests"
        )
    pid = str(manifest.get("id") or pack_dir.name)
    lib = modstore_library_path()
    with tempfile.NamedTemporaryFile(suffix=".zip", delete=False) as tmp:
        tmp.write(raw_zip)
        tmp_path = Path(tmp.name)
    try:
        dest = import_zip(tmp_path, lib, replace=replace)
    finally:
        tmp_path.unlink(missing_ok=True)
    saved_package: Dict[str, Any] = {}
    if publish_to_catalog:
        with tempfile.NamedTemporaryFile(suffix=".xcemp", delete=False) as tmp:
            tmp.write(raw_zip)
            pkg_tmp_path = Path(tmp.name)
        try:
            from modstore_server.catalog_store import append_package

            rec = {
                "id": pid,
                "name": str(manifest.get("name") or pid),
                "version": str(manifest.get("version") or "1.0.0"),
                "description": str(manifest.get("description") or ""),
                "artifact": "employee_pack",
                "industry": str(manifest.get("industry") or "通用"),
                "release_channel": "stable",
                "commerce": {"mode": "free", "price": 0},
                "license": {"type": "personal", "verify_url": None},
            }
            saved_package = append_package(rec, pkg_tmp_path)
            row = db.query(CatalogItem).filter(CatalogItem.pkg_id == pid).first()
            if not row:
                row = CatalogItem(pkg_id=pid, author_id=user.id)
                db.add(row)
            row.version = saved_package.get("version") or rec["version"]
            row.name = saved_package.get("name") or rec["name"]
            row.description = saved_package.get("description") or rec["description"]
            row.price = 0.0
            row.artifact = "employee_pack"
            row.industry = saved_package.get("industry") or rec["industry"]
            row.stored_filename = saved_package.get("stored_filename") or ""
            row.sha256 = saved_package.get("sha256") or ""
            db.commit()
            mirror_catalog_file_to_market_files(row.stored_filename)
        finally:
            pkg_tmp_path.unlink(missing_ok=True)
    vibecoding_ok = True
    if use_vibecoding_loop:
        golden_ok = bool(golden_comparison.get("passed")) if golden_comparison else True
        vibecoding_ok = bool(
            generated_convert_py
            and domain_smoke.get("ok") is not False
            and golden_ok
            and is_llm_codegen_source(runtime_meta)
        )

    return {
        "ok": vibecoding_ok and not warnings,
        "id": dest.name,
        "path": str(dest),
        "manifest": manifest,
        "asset_manifest": asset_manifest,
        "rule_spec": rule_spec,
        "validate_warnings": warnings,
        "package": saved_package,
        "published": bool(publish_to_catalog and saved_package),
        "llm": llm_meta,
        "runtime_generation": runtime_meta,
        "runtime_repair_history": repair_history,
        "domain_smoke": domain_smoke,
        "golden_comparison": golden_comparison,
    }


async def run_word_extract_employee_scaffold_async(
    db: Any,
    user: User,
    *,
    session_id: str,
    brief: str,
    raw_files: Optional[List[Dict[str, Any]]] = None,
    replace: bool = True,
    provider: Optional[str] = None,
    model: Optional[str] = None,
    publish_to_catalog: bool = False,
    force_llm_codegen: bool = True,
    payload: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    """Word 全量提取员工：走 direct_python 资产脚手架，无上传文件也可生成。"""
    return await run_asset_employee_scaffold_async(
        db,
        user,
        session_id=session_id,
        brief=brief,
        raw_files=list(raw_files or []),
        replace=replace,
        provider=provider,
        model=model,
        publish_to_catalog=publish_to_catalog,
        force_llm_codegen=True,
        payload=payload,
    )
