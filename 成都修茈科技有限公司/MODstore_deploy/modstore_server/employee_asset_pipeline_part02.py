# ruff: noqa
"""Implementation extracted from the public facade module."""
from __future__ import annotations
import importlib


def _facade():
    return importlib.import_module("modstore_server.employee_asset_pipeline")


def _infer_accepted_extensions(
    asset_manifest: _facade().Optional[_facade().Dict[str, _facade().Any]] = None
) -> _facade().List[str]:
    if not isinstance(asset_manifest, dict):
        return []
    suffixes: set = set()
    for item in asset_manifest.get("assets") or []:
        if not isinstance(item, dict):
            continue
        s = str(item.get("suffix") or "").lower()
        if s:
            suffixes.add(s)
    if not suffixes:
        return []
    if suffixes & _facade().EXCEL_SUFFIXES:
        return sorted(suffixes & _facade().EXCEL_SUFFIXES) + sorted(
            suffixes - _facade().EXCEL_SUFFIXES
        )
    if suffixes & _facade().DOC_SUFFIXES:
        return sorted(suffixes & _facade().DOC_SUFFIXES) + sorted(suffixes - _facade().DOC_SUFFIXES)
    return sorted(suffixes)


def _infer_asset_runtime_kind(
    brief: str, asset_manifest: _facade().Optional[_facade().Dict[str, _facade().Any]] = None
) -> str:
    """Classify by asset shape, not by a customer-specific package name."""
    if _facade().is_csv_generate(brief):
        return "csv_generate"
    if _facade().is_csv_full_read(brief):
        return "csv_full_read"
    if _facade().is_excel_generate(brief):
        return "excel_generate"
    if _facade().is_excel_full_read(brief):
        return "excel_full_read"
    if _facade().is_pdf_generate(brief):
        return "pdf_generate"
    if _facade().is_pdf_full_read(brief):
        return "pdf_full_read"
    if _facade().is_json_quant_report(brief):
        return "json_quant_report"
    if _facade().is_kitten_chart_viz(brief):
        return "kitten_chart_viz"
    if _facade().is_ppt_generate(brief):
        return "ppt_generate"
    if _facade().is_ppt_full_read(brief):
        return "ppt_full_read"
    if _facade().is_txt_generate(brief):
        return "txt_generate"
    if _facade().is_txt_full_read(brief):
        return "txt_full_read"
    if _facade().is_word_generate(brief):
        return "word_generate"
    if _facade().is_word_full_extract(brief):
        return "word_full_extract"
    has_excel = False
    has_doc = False
    has_rules = False
    has_reference_code = False
    if isinstance(asset_manifest, dict):
        for item in asset_manifest.get("assets") or []:
            if not isinstance(item, dict):
                continue
            suffix = str(item.get("suffix") or "").lower()
            kind = str(item.get("kind") or "")
            has_excel = has_excel or suffix in _facade().EXCEL_SUFFIXES
            has_doc = has_doc or suffix in _facade().DOC_SUFFIXES
            has_rules = has_rules or kind == "rules"
            has_reference_code = has_reference_code or kind == "reference_code"
    text = brief or ""
    _contract_keywords = ("合同", "法务", "合规", "审核", "条款", "contract", "legal", "compliance")
    _doc_keywords = ("文档", "报告", "方案", "标书", "简历", "document", "report", "proposal")
    if has_reference_code:
        return "reference_python_transform"
    if has_doc or any((k in text for k in _contract_keywords)):
        if any((k in text for k in _contract_keywords)):
            return "contract_doc_review"
        return "doc_template_transform"
    if has_excel and (has_rules or any((k in text for k in ("规则", "模板", "转换", "考勤")))):
        return "excel_rules_transform"
    if has_excel:
        return "generic_excel_transform"
    return "generic_file_transform"


def _read_text_preview(content: bytes, limit: int = 4000) -> str:
    for enc in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            return content.decode(enc)[:limit]
        except UnicodeDecodeError:
            continue
    return ""


def _excel_summary(path: _facade().Path) -> _facade().Dict[str, _facade().Any]:
    summary: _facade().Dict[str, _facade().Any] = {"ok": False, "sheets": [], "error": ""}
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        summary["error"] = f"openpyxl unavailable: {exc}"
        return summary
    try:
        wb = load_workbook(path, read_only=False, data_only=False)
    except Exception as exc:
        summary["error"] = str(exc)[:500]
        return summary
    sheets: _facade().List[_facade().Dict[str, _facade().Any]] = []
    for ws in wb.worksheets[:12]:
        formulas = 0
        non_empty = 0
        header_candidates: _facade().List[_facade().List[str]] = []
        max_row = min(ws.max_row or 0, 30)
        max_col = min(ws.max_column or 0, 30)
        for r in range(1, max_row + 1):
            vals: _facade().List[str] = []
            filled = 0
            for c in range(1, max_col + 1):
                v = ws.cell(r, c).value
                if v is not None and str(v).strip():
                    filled += 1
                    vals.append(str(v).strip()[:40])
                    if isinstance(v, str) and v.startswith("="):
                        formulas += 1
            non_empty += filled
            if filled >= 2 and len(header_candidates) < 5:
                header_candidates.append(vals[:12])
        sheets.append(
            {
                "name": ws.title,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
                "merged_ranges": len(list(ws.merged_cells.ranges)),
                "formula_cells_sampled": formulas,
                "non_empty_sampled": non_empty,
                "header_candidates": header_candidates,
            }
        )
    summary["ok"] = True
    summary["sheets"] = sheets
    return summary


def prepare_employee_assets(
    *,
    session_id: str,
    user_id: int,
    raw_files: _facade().List[_facade().Dict[str, _facade().Any]],
    repo_root: _facade().Optional[_facade().Path] = None,
) -> _facade().Dict[str, _facade().Any]:
    root = (
        (repo_root or _facade().Path(__file__).resolve().parents[1])
        / "var"
        / "employee_draft_assets"
        / str(user_id)
        / session_id
    )
    if root.exists():
        _facade().shutil.rmtree(root, ignore_errors=True)
    root.mkdir(parents=True, exist_ok=True)
    assets: _facade().List[_facade().Dict[str, _facade().Any]] = []
    for idx, item in enumerate(raw_files or []):
        filename = _facade()._safe_basename(str(item.get("filename") or f"asset-{idx}.bin"))
        content = item.get("content") or b""
        if not isinstance(content, (bytes, bytearray)):
            continue
        kind = _facade()._classify_asset(filename)
        dest = root / f"{idx:02d}_{filename}"
        dest.write_bytes(bytes(content))
        rec: _facade().Dict[str, _facade().Any] = {
            "id": f"asset_{idx}",
            "filename": filename,
            "kind": kind,
            "suffix": _facade().Path(filename).suffix.lower(),
            "size": len(content),
            "path": str(dest),
        }
        if rec["suffix"] in _facade().EXCEL_SUFFIXES:
            rec["excel"] = _facade()._excel_summary(dest)
        elif rec["suffix"] in _facade().TEXT_SUFFIXES:
            rec["preview"] = _facade()._read_text_preview(bytes(content))
        assets.append(rec)
    manifest = {
        "session_id": session_id,
        "user_id": user_id,
        "root": str(root),
        "assets": assets,
        "templates": [a for a in assets if a["kind"] == "template"],
        "example_inputs": [a for a in assets if a["kind"] == "example_input"],
        "expected_outputs": [a for a in assets if a["kind"] == "expected_output"],
        "rules": [a for a in assets if a["kind"] in ("rules", "reference_code")],
    }
    (root / "asset_manifest.json").write_text(
        _facade().json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    return manifest


def _preflight_scaffold_write_access(
    *, session_id: str, user_id: int, repo_root: _facade().Optional[_facade().Path] = None
) -> _facade().Tuple[bool, _facade().List[str], _facade().Dict[str, str]]:
    """资产脚手架写路径权限预检（MODstore 库 + 会话草稿目录）。"""
    paths_checked: _facade().Dict[str, str] = {}
    errors: _facade().List[str] = []
    lib = _facade().modstore_library_path()
    paths_checked["modstore_library"] = str(lib)
    try:
        lib.mkdir(parents=True, exist_ok=True)
        if not _facade().os.access(lib, _facade().os.W_OK):
            errors.append(f"modstore 库目录不可写：{lib}")
    except OSError as exc:
        errors.append(f"modstore 库目录不可创建：{lib} ({exc})")
    draft_root = (
        (repo_root or _facade().Path(__file__).resolve().parents[1])
        / "var"
        / "employee_draft_assets"
        / str(user_id)
        / session_id
    )
    paths_checked["draft_assets_root"] = str(draft_root)
    try:
        draft_root.parent.mkdir(parents=True, exist_ok=True)
        probe = draft_root.parent / ".write_probe"
        probe.write_text("ok", encoding="utf-8")
        probe.unlink(missing_ok=True)
    except OSError as exc:
        errors.append(f"会话资产目录不可写：{draft_root.parent} ({exc})")
    return (not errors, errors, paths_checked)


def build_rule_spec(
    brief: str,
    asset_manifest: _facade().Dict[str, _facade().Any],
    *,
    payload: _facade().Optional[_facade().Dict[str, _facade().Any]] = None,
) -> _facade().Dict[str, _facade().Any]:
    from modstore_server.employee_pipeline_routing import confident_word_full_extract_routing

    if confident_word_full_extract_routing(brief):
        return _facade().build_word_extract_rule_spec(brief)
    if _facade().is_csv_generate(brief):
        return _facade().build_csv_generate_rule_spec(brief)
    if _facade().is_csv_full_read(brief):
        return _facade().build_csv_read_rule_spec(brief)
    if _facade().is_excel_generate(brief):
        return _facade().build_excel_generate_rule_spec(brief)
    if _facade().is_excel_full_read(brief):
        return _facade().build_excel_read_rule_spec(brief)
    if _facade().is_pdf_generate(brief):
        return _facade().build_pdf_generate_rule_spec(brief)
    if _facade().is_pdf_full_read(brief):
        return _facade().build_pdf_read_rule_spec(brief)
    if _facade().is_json_quant_report(brief):
        return _facade().build_json_quant_report_rule_spec(brief)
    if _facade().is_kitten_chart_viz(brief):
        return _facade().build_kitten_chart_rule_spec(brief)
    if _facade().is_ppt_generate(brief):
        return _facade().build_ppt_generate_rule_spec(brief)
    if _facade().is_ppt_full_read(brief):
        return _facade().build_ppt_read_rule_spec(brief)
    if _facade().is_txt_generate(brief):
        return _facade().build_txt_generate_rule_spec(brief)
    if _facade().is_txt_full_read(brief):
        return _facade().build_txt_read_rule_spec(brief)
    if _facade().is_word_full_extract(brief):
        return _facade().build_word_extract_rule_spec(brief)
    if _facade().is_word_generate(brief):
        return _facade().build_word_generate_rule_spec(brief)
    templates = asset_manifest.get("templates") or []
    if not templates:
        excels = [
            a
            for a in asset_manifest.get("assets") or []
            if a.get("suffix") in _facade().EXCEL_SUFFIXES
        ]
        templates = excels[:1]
    template_asset = templates[0] if templates else None
    template_relpath = ""
    if template_asset:
        template_relpath = _facade()._template_storage_relpath(
            str(template_asset["filename"]), brief
        )
    runtime_kind = _facade()._infer_asset_runtime_kind(brief, asset_manifest)
    accepted_exts = _facade()._infer_accepted_extensions(asset_manifest)
    _is_doc_kind = runtime_kind in ("contract_doc_review", "doc_template_transform")
    if not accepted_exts:
        if _is_doc_kind:
            accepted_exts = [".docx", ".pdf"]
        else:
            accepted_exts = [".xlsx", ".xlsm", ".xls"]
    output_ext = accepted_exts[0] if accepted_exts else ".xlsx"
    output_relpath = f"outputs/employee_output{output_ext}"
    if "考勤" in brief:
        output_relpath = f"424/考勤转换输出{output_ext}"
    _mode = (
        "llm_doc_review"
        if runtime_kind == "contract_doc_review"
        else "direct_python_file_transform"
    )
    spec = {
        "brief": brief,
        "mode": _mode,
        "accepted_extensions": accepted_exts,
        "default_action": "review" if _is_doc_kind else "convert",
        "default_output_relpath": output_relpath,
        "default_template_relpath": (
            template_relpath.removeprefix("backend/templates/") if template_relpath else ""
        ),
        "template_relpath": template_relpath,
        "template_asset_id": template_asset.get("id") if template_asset else "",
        "runtime_kind": runtime_kind,
        "assets_summary": {
            "templates": [
                {
                    "filename": a.get("filename"),
                    "sheets": [
                        {
                            "name": s.get("name"),
                            "max_row": s.get("max_row"),
                            "max_column": s.get("max_column"),
                            "headers": s.get("header_candidates", [])[:2],
                        }
                        for s in ((a.get("excel") or {}).get("sheets") or [])[:6]
                    ],
                }
                for a in templates[:3]
            ],
            "example_inputs": [
                a.get("filename") for a in (asset_manifest.get("example_inputs") or [])[:5]
            ],
            "expected_outputs": [
                a.get("filename") for a in (asset_manifest.get("expected_outputs") or [])[:5]
            ],
            "rules": [
                {"filename": a.get("filename"), "preview": str(a.get("preview") or "")[:1000]}
                for a in (asset_manifest.get("rules") or [])[:5]
            ],
        },
        "requirements": [
            "Use direct_python only; do not add echo or llm_md.",
            "Preserve uploaded templates as binary files and copy them into backend/templates.",
            "Never claim success unless an output file is actually written.",
            "Return {ok, summary, items, warnings, error, meta}.",
        ],
    }
    if runtime_kind == "contract_doc_review":
        spec["requirements"] = [
            "This employee reviews contracts/documents using LLM reasoning via the agent handler.",
            "Read the uploaded document, identify missing clauses, ambiguous terms, and compliance issues.",
            "Output a structured review with specific suggestions for each issue found.",
            "Never fabricate legal advice; clearly state when professional legal review is recommended.",
            "Return {ok, summary, items, warnings, error, meta}.",
        ]
    elif runtime_kind in {"excel_rules_transform", "reference_python_transform"}:
        spec["requirements"].extend(
            [
                "Generate deterministic Python from the uploaded rules, template, and examples.",
                "If reference Python is provided, adapt or call it instead of replacing it with a placeholder.",
                "Fail explicitly when the generated transform cannot map the input workbook to the output workbook.",
            ]
        )
    return spec


def _slug_from_brief(brief: str) -> str:
    text = brief or ""
    explicit = _facade().re.search(
        "(?:pack_id|员工包\\s*ID|员工包ID|员工包 id|包ID|包 id)\\s*[:：=]\\s*([A-Za-z0-9][A-Za-z0-9_-]{2,80})",
        text,
        _facade().re.I,
    )
    if explicit:
        return _facade().normalize_mod_id(explicit.group(1)) or "asset-worker-employee"
    if _facade().re.search(
        "\\btaiyangniao[-_ ]attendance(?:[-_ ]employee)?\\b", text, _facade().re.I
    ):
        return "taiyangniao-attendance-employee"
    if _facade().is_csv_generate(brief):
        return "csv-generate-employee"
    if _facade().is_csv_full_read(brief):
        return "csv-full-read-employee"
    if _facade().is_excel_generate(brief):
        return "excel-generate-employee"
    if _facade().is_excel_full_read(brief):
        return "excel-full-read-employee"
    if _facade().is_pdf_generate(brief):
        return "pdf-generate-employee"
    if _facade().is_pdf_full_read(brief):
        return "pdf-full-read-employee"
    if _facade().is_json_quant_report(brief):
        return "json-report-employee"
    if _facade().is_ppt_generate(brief):
        return "ppt-generate-employee"
    if _facade().is_ppt_full_read(brief):
        return "ppt-full-read-employee"
    if _facade().is_txt_generate(brief):
        return "txt-generate-employee"
    if _facade().is_txt_full_read(brief):
        return "txt-full-read-employee"
    if _facade().is_word_full_extract(brief):
        return "word-full-read-employee"
    if _facade().is_word_generate(brief):
        return "word-generate-employee"
    if "考勤" in brief:
        return "attendance-transform-employee"
    return "asset-worker-employee"


def _employee_name_from_brief(brief: str, fallback: str = "文件处理员工") -> str:
    text = (brief or "").strip()
    explicit = _facade().re.search(
        "(?:员工名称|员工名|name)\\s*[:：=]\\s*([^\\n\\r,，。；;]{2,40})", text, _facade().re.I
    )
    if explicit:
        return explicit.group(1).strip()
    if "考勤" in text:
        return "考勤处理员"
    if "报表" in text:
        return "报表处理员"
    if _facade().re.search("审核|审查|合规|风控|合同", text):
        return "合同审核员工"
    if _facade().re.search("翻译|本地化|多语言", text):
        return "文档翻译员工"
    if _facade().re.search("比对|对比|校对|校验", text):
        return "文档比对员工"
    if _facade().is_csv_generate(text):
        return "CSV 生成员"
    if _facade().is_csv_full_read(text):
        return "CSV 全量读取员"
    if _facade().is_excel_generate(text):
        return "Excel 生成员"
    if _facade().is_excel_full_read(text):
        return "Excel 全量读取员"
    if _facade().is_pdf_generate(text):
        return "PDF 生成员"
    if _facade().is_pdf_full_read(text):
        return "PDF 全量读取员"
    if _facade().is_ppt_generate(text):
        return "PPT 生成员"
    if _facade().is_ppt_full_read(text):
        return "PPT 全量读取员"
    if _facade().is_txt_generate(text):
        return "TXT 生成员"
    if _facade().is_txt_full_read(text):
        return "TXT 全量读取员"
    if _facade().is_word_generate(text):
        return "Word 生成员"
    if _facade().is_word_full_extract(text):
        return "Word 全量读取员"
    return fallback


def _template_storage_relpath(filename: str, brief: str = "") -> str:
    safe = _facade()._safe_basename(filename, "template.xlsx")
    return (
        f"backend/templates/424/{safe}" if "考勤" in (brief or "") else f"backend/templates/{safe}"
    )


def _employee_id_from_pack_id(pack_id: str) -> str:
    pid = (pack_id or "").strip()
    if pid.endswith("-employee"):
        return pid[: -len("-employee")] or pid
    return pid
