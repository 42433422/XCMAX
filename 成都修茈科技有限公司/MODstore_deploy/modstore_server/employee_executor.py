"""AI员工执行器：基于 employee_config_v2 的真实执行管道。"""

from __future__ import annotations

import asyncio
import csv
import importlib.util
import io
import json
import logging
import os
import shutil
import threading
import time
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

import httpx

from modstore_server.catalog_store import files_dir
from modstore_server.employee_runtime import (
    build_employee_context,
    load_employee_pack_resolved,
    parse_employee_config_v2,
)
from modstore_server.llm_failure_classifier import FAILURE_KIND_QUOTA, classify_failure_kind
from modstore_server.models import EmployeeExecutionMetric, User, get_session_factory
from modstore_server.runtime_async import run_coro_sync as _run_coro_sync
from modstore_server.services.llm import chat_dispatch_via_session

logger = logging.getLogger(__name__)

_METRIC_TASK_MAX_LEN = 128


def _emp_im_notify_boss(
    employee_id: str,
    manifest: Any,
    body: str,
    hook: str,
) -> None:
    """员工执行管道 hook 公共入口：从 manifest 抽 display_name + mod_id，调 notify_boss。

    best-effort：任何异常只 log debug，不抛错。body 为空也跳过。
    4 个 hook 点（perception/cognition/verification/handoff）共用此 helper。
    """
    body_text = (body or "").strip()
    if not body_text:
        return
    try:
        from modstore_server.employee_im_bridge import notify_boss as _notify_boss_im

        _emp_display = ""
        _emp_mod_id = ""
        try:
            if isinstance(manifest, dict):
                _ident = (
                    manifest.get("identity")
                    if isinstance(manifest.get("identity"), dict)
                    else {}
                )
                _emp_display = str(
                    _ident.get("name") or manifest.get("name") or ""
                ).strip()
                _emp_mod_id = str(
                    manifest.get("mod_id") or manifest.get("id") or ""
                ).strip()
        except Exception:
            logger.debug("emp_im_notify manifest parse skipped", exc_info=True)
        _notify_boss_im(
            employee_id,
            body=body_text[:600],
            hook=hook,
            mod_id=_emp_mod_id,
            display_name=_emp_display,
        )
    except Exception:
        logger.debug("emp_im_notify %s skipped", hook, exc_info=True)


# 员工大会待机：manifest system_prompt 常要求「输出 JSON」，与四段 Markdown 汇报冲突。
_ALL_HANDS_COGNITION_SYSTEM_APPEND = """\
【员工大会模式 — 覆盖日常 JSON 输出要求】
当前任务为数字管家召集的「员工大会」汇报（非流水线执行、非工作台单次任务）。
硬性要求：
- 只输出 **简体中文 Markdown**，严格按用户消息中的四段标题结构作答；**禁止**输出 JSON、禁止 ``warnings`` / ``status`` 字段。
- 用 manifest / depends_on / handlers 与 input 中的节选说明职责；缺上游产物在待机模式下**属于正常**，不得写「输入不足」类流水线报错。
- 不得编造 research_context / yuangon_pack_excerpt / recent_failures 中未出现的路径或版本号。"""

_ALL_HANDS_ROLE_CONTEXT_MODES = frozenset({"all_hands_meeting", "all_hands_standby"})


# 10 项成熟度第 1 项「主动沟通」协议：让 LLM 知道有 requires_human 通道可用，
# 遇到下面任一情况必须主动向老板提问（输出 requires_human=true + human_question）。
# 不破坏原 JSON 输出格式，只是把 requires_human / human_question 作为可选字段加入。
_PHASE_D_PROTOCOL_APPEND = """\
【主动沟通协议 — 强制遵守】
你是真员工，不是提示词机器人。完成任务时遇到以下任一情况，**必须**在 JSON 输出里追加两个字段：
- "requires_human": true
- "human_question": "<你具体想问老板的问题，简短一句中文>"

需要主动提问的情况：
1. 任务优先级不明确（多个任务冲突，不知道先做哪个）
2. 缺资源（需要其他员工配合、需要数据、需要权限）
3. 发现风险（代码改动可能影响线上、安全风险、合规风险）
4. 需要老板决策（业务方向、产品取舍、用户影响）
5. 任务超出你的职责范围（应该转交给别的员工）
6. 失败 3 次仍无法解决（需要换思路或换人）
7. 你不确定的事（宁可问，不要瞎做）

提问要具体，不要套话。例：
- ❌ "需要老板确认" — 太空
- ✅ "本周修复 A 还是 B 优先级更高？目前 A 影响 100 用户/天" — 具体

没有上述情况时，正常输出你的工作 JSON，不需要带 requires_human 字段。

【任务转交协议 — 同事协作】
如果你判断任务**不属于你的职责范围**（如本员工 manifest 没覆盖的代码路径 / 接口 / 服务），
应该在 JSON 输出里追加：
- "handoff_to": "<目标员工的 employee_id>"
- "handoff_reason": "<为什么转交给他，一句话>"
- "handoff_context": "<上下文摘要，让目标员工不用重新看一遍 task>"

只在你**确实无法处理**时才转交，不要把简单任务推给别人。
转交后你仍然要正常完成你能做的部分，不要因为转交就停手。"""


def _is_all_hands_cognition_context(inp: Any) -> bool:
    if not isinstance(inp, dict):
        return False
    if inp.get("all_hands_standby") is True:
        return True
    rc = inp.get("role_context")
    if isinstance(rc, dict):
        mode = str(rc.get("mode") or "").strip()
        if mode in _ALL_HANDS_ROLE_CONTEXT_MODES:
            return True
    return False


def _build_all_hands_cognition_user_message(
    task: str,
    normalized_input: Dict[str, Any],
    *,
    session_context_json: str = "",
) -> str:
    """与 ``all_hands_report._standby_manifest_report_via_bench`` 一致：任务模板 + JSON 上下文。"""
    payload_json = json.dumps(normalized_input, ensure_ascii=False)
    if len(payload_json) > 14000:
        payload_json = payload_json[:14000]
    task_part = str(task or "").strip()
    if task_part:
        user_input = (
            f"{task_part}\n\n---\n\n"
            f"以下为结构化输入（JSON），请据此撰写四段 Markdown 汇报：\n"
            f"{payload_json}"
        )
    else:
        user_input = payload_json
    if session_context_json:
        user_input = f"{user_input}\n\n[session_context]\n{session_context_json}"
    return user_input


def _metric_task_preview(task: object) -> str:
    """单行动略预览：``employee_execution_metrics.task`` 列为 VARCHAR(128)。"""
    t = str(task or "").replace("\r\n", "\n").replace("\r", "\n")
    t = " ".join(t.split())
    if len(t) <= _METRIC_TASK_MAX_LEN:
        return t
    return t[: _METRIC_TASK_MAX_LEN - 1] + "…"


def _resolve_metric_user_id(session: Any, user_id: object) -> int:
    """定时任务 / 员工大会常传 ``user_id=0``；指标表 ``user_id`` 须指向真实 ``users.id``。"""
    try:
        uid = int(user_id)
    except (TypeError, ValueError):
        uid = 0
    if uid > 0:
        row = session.query(User.id).filter(User.id == uid).first()
        if row is not None:
            return int(row[0])
    row = session.query(User.id).order_by(User.id.asc()).limit(1).first()
    if row is None:
        raise RuntimeError("employee_execution_metrics: 库中无 users 行，无法写入指标")
    return int(row[0])


_executor_sem: threading.Semaphore | None = None
_executor_sem_n: int = 0


def _executor_max_concurrent() -> int:
    raw = (os.environ.get("MODSTORE_EXECUTOR_MAX_CONCURRENT") or "").strip()
    if not raw:
        return 0
    try:
        n = int(raw)
    except ValueError:
        return 0
    return n if n > 0 else 0


def _get_executor_semaphore() -> threading.Semaphore | None:
    """Lazy singleton semaphore when MODSTORE_EXECUTOR_MAX_CONCURRENT > 0."""
    global _executor_sem, _executor_sem_n
    n = _executor_max_concurrent()
    if n <= 0:
        return None
    if _executor_sem is None or _executor_sem_n != n:
        _executor_sem = threading.Semaphore(n)
        _executor_sem_n = n
    return _executor_sem


def _executor_extra_cognition_retries() -> int:
    raw = (os.environ.get("MODSTORE_COGNITION_TRANSIENT_RETRIES") or "1").strip()
    try:
        x = int(raw)
    except ValueError:
        x = 1
    return min(max(x, 0), 2)


def _executor_detail_log_enabled() -> bool:
    return (os.environ.get("MODSTORE_EXECUTOR_LOG_DETAIL") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )


def _is_transient_llm_error(msg: str) -> bool:
    s = (msg or "").lower()
    if not s.strip():
        return False
    needles = (
        "timeout",
        "timed out",
        "connection reset",
        "connection aborted",
        "temporarily unavailable",
        "rate limit",
        "429",
        "503",
        "502",
        "bad gateway",
        "service unavailable",
        "eof occurred",
        "broken pipe",
        "connection refused",
        "connecterror",
        "readtimeout",
        "remotedisconnected",
        "try again",
        "overloaded",
    )
    return any(n in s for n in needles)


def _run_cognition_with_transient_retries(
    config: Dict[str, Any],
    perceived: Dict[str, Any],
    memory: Dict[str, Any],
    session,
    user_id: int,
    *,
    employee_id: str,
    task: str,
    bench_llm_override: Optional[Tuple[str, str]] = None,
) -> Tuple[Dict[str, Any], Dict[str, Any]]:
    """Return ``(reasoning, recovery_meta)``. ``recovery_meta`` is set when retries fixed cognition."""
    max_extra = _executor_extra_cognition_retries()
    first_err = ""
    reasoning: Dict[str, Any] = {}
    for attempt in range(max_extra + 1):
        reasoning = _cognition_sync(
            config,
            perceived,
            memory,
            session,
            user_id,
            employee_id=employee_id,
            task=task,
            bench_llm_override=bench_llm_override,
        )
        err = str(reasoning.get("error") or "").strip()
        if not err:
            if attempt > 0:
                return reasoning, {
                    "recovered": True,
                    "attempts": attempt + 1,
                    "original_error": first_err[:2000],
                    "recovery_action": "cognition_retry",
                }
            return reasoning, {}
        if attempt == 0:
            first_err = err
        if attempt >= max_extra or not _is_transient_llm_error(err):
            return reasoning, {}
        delay = 0.4 * (2**attempt)
        logger.warning(
            "employee_executor cognition transient failure employee_id=%s attempt=%s max_attempts=%s error=%s retry_delay_s=%.2f",
            employee_id,
            attempt + 1,
            max_extra + 1,
            err[:400],
            delay,
        )
        time.sleep(delay)
    return reasoning, {}


def _get_section(config: Dict[str, Any], section: str) -> Dict[str, Any]:
    if not isinstance(config, dict):
        return {}
    if section in config and isinstance(config.get(section), dict):
        return config.get(section) or {}
    return config


def _perception_excel(input_data: Any) -> Dict[str, Any]:
    """解析 .xlsx 内容（base64 或 data URL）。"""
    import base64
    import io

    try:
        import openpyxl
    except ImportError:
        return {
            "normalized_input": input_data,
            "type": "excel",
            "parse_error": "请安装 openpyxl: pip install openpyxl",
        }

    raw = input_data
    if isinstance(input_data, dict):
        raw = input_data.get("content", input_data.get("base64", ""))
    if isinstance(raw, str) and raw.startswith("data:"):
        raw = raw.split(",", 1)[-1]
    if not raw:
        return {"normalized_input": input_data, "type": "excel", "parse_error": "empty payload"}

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(raw)), read_only=True, data_only=True
        )
        sheets_data: Dict[str, Any] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: List[List[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            sheets_data[sheet_name] = {
                "rows": rows,
                "row_count": len(rows),
                "col_count": len(rows[0]) if rows else 0,
            }
        wb.close()
        return {"normalized_input": sheets_data, "type": "excel", "parse_ok": True}
    except Exception as e:  # noqa: PERF203
        return {"normalized_input": input_data, "type": "excel", "parse_error": str(e)}


def _extract_vision_data_urls(payload: Any) -> List[str]:
    """从 payload 抽取可送入 chat/completions 的 image_url（data: 或 https:）。"""
    out: List[str] = []
    if isinstance(payload, dict):
        for key in ("images", "image_urls", "urls"):
            val = payload.get(key)
            if isinstance(val, list):
                for u in val:
                    if isinstance(u, str) and u.strip():
                        s = u.strip()
                        out.append(
                            s if s.startswith(("data:", "http")) else f"data:image/png;base64,{s}"
                        )
        if not out:
            for key in ("image_url", "url", "base64", "content"):
                v = payload.get(key)
                if not isinstance(v, str) or not v.strip():
                    continue
                s = v.strip()
                if s.startswith("data:") or s.startswith("http"):
                    out.append(s)
                else:
                    out.append(f"data:image/png;base64,{s}")
                break
    elif isinstance(payload, str) and payload.strip():
        s = payload.strip()
        if s.startswith("data:") or s.startswith("http"):
            out.append(s)
        else:
            out.append(f"data:image/png;base64,{s}")
    return out[:8]


def _perception_image(input_data: Any, session, user_id: int) -> Dict[str, Any]:
    """优先使用多模态 LLM 描述图片（需 OpenAI 兼容 Key）。"""
    vision_urls = _extract_vision_data_urls(input_data)
    raw = input_data
    if isinstance(input_data, dict):
        raw = input_data.get("base64", input_data.get("url", input_data.get("content", "")))
    if isinstance(raw, str) and raw.startswith("data:"):
        raw = raw.split(",", 1)[-1]

    if not raw:
        return {
            "normalized_input": input_data,
            "type": "image",
            "note": "图片解析需配置 OpenAI API Key，并在 input 中提供 base64",
            "vision_data_urls": vision_urls,
        }

    image_content = (
        raw if isinstance(raw, str) and raw.startswith("data:") else f"data:image/png;base64,{raw}"
    )

    async def _call():
        return await chat_dispatch_via_session(
            session,
            user_id,
            "openai",
            "gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "请简要描述图片中的文字与关键信息。"},
                        {"type": "image_url", "image_url": {"url": image_content}},
                    ],
                }
            ],
            max_tokens=800,
        )

    try:
        result = _run_coro_sync(_call())
        if result.get("ok"):
            return {
                "normalized_input": {"description": result.get("content", ""), "type": "image"},
                "type": "image",
                "parse_ok": True,
                "method": "vision",
                "vision_data_urls": vision_urls or ([image_content] if image_content else []),
            }
    except Exception as e:  # noqa: PERF203
        return {
            "normalized_input": input_data,
            "type": "image",
            "parse_error": str(e),
            "vision_data_urls": vision_urls,
        }
    return {
        "normalized_input": input_data,
        "type": "image",
        "note": "vision 调用未返回内容",
        "vision_data_urls": vision_urls,
    }


def _memory_long_term_chroma(
    employee_id: str, input_data: Dict[str, Any], _cfg: Dict[str, Any]
) -> Dict[str, Any]:
    """员工长期记忆：复用 ``vector_engine`` 的 PersistentClient 单例（修文件句柄泄露），
    集合保留 Chroma 默认 embedding function 以兼容历史 ``query_texts`` 写法。
    """
    query = str(input_data.get("memory_query") or input_data.get("query") or "").strip()
    if not query:
        return {
            "enabled": True,
            "memories": [],
            "note": "请在 input_data 中提供 memory_query 以检索长期记忆",
        }
    try:
        from modstore_server import vector_engine
        from modstore_server.vector_engine import VectorEngineError
    except ImportError:
        return {"enabled": True, "memories": [], "note": "请安装 chromadb: pip install chromadb"}

    try:
        client = vector_engine.get_client()
    except VectorEngineError as e:
        return {"enabled": True, "memories": [], "error": str(e)}

    coll_name = vector_engine.employee_memory_collection_name(employee_id)
    try:
        collection = client.get_or_create_collection(
            name=coll_name,
            metadata={"hnsw:space": "cosine"},
        )
    except Exception as e:  # noqa: BLE001
        return {"enabled": True, "memories": [], "error": str(e)}

    try:
        results = collection.query(query_texts=[query], n_results=5)
    except Exception as e:  # noqa: BLE001
        return {"enabled": True, "memories": [], "error": str(e)}

    documents = (results.get("documents") or [[]])[0] or []
    distances = (results.get("distances") or [[]])[0] or []
    memories = []
    for i, doc in enumerate(documents):
        dist = float(distances[i]) if i < len(distances) else 1.0
        if dist < 0.85:
            memories.append({"content": doc, "distance": dist})
    return {"enabled": True, "memories": memories, "count": len(memories)}


def _perception_real(
    config: Dict[str, Any],
    input_data: Dict[str, Any],
    session=None,
    user_id: int = 0,
) -> Dict[str, Any]:
    p_cfg = _get_section(config, "perception")
    p_type = str(p_cfg.get("type") or "text").strip().lower()
    payload = input_data or {}
    if p_type == "text":
        return {"normalized_input": payload, "type": "text"}
    if p_type == "json":
        if isinstance(payload, dict):
            return {"normalized_input": payload, "type": "json"}
        try:
            parsed = json.loads(payload) if isinstance(payload, str) else payload
            return {"normalized_input": parsed, "type": "json"}
        except Exception as e:  # noqa: PERF203
            return {"normalized_input": payload, "type": "json", "parse_error": str(e)}
    if p_type == "csv":
        if (
            isinstance(payload, dict)
            and str(payload.get("file_path") or payload.get("path") or "").strip()
        ):
            return {"normalized_input": payload, "type": "csv"}
        raw = payload.get("content", "") if isinstance(payload, dict) else str(payload)
        try:
            reader = csv.DictReader(io.StringIO(raw))
            rows = list(reader)
            return {"normalized_input": {"rows": rows}, "type": "csv", "row_count": len(rows)}
        except Exception as e:  # noqa: PERF203
            return {"normalized_input": payload, "type": "csv", "parse_error": str(e)}
    if p_type == "excel":
        return _perception_excel(payload)
    if p_type == "image":
        return _perception_image(payload, session, user_id)
    if p_type == "document":
        return _perception_document(payload)
    if p_type in ("web_rankings", "ai_model_rankings"):
        return _perception_web_rankings(payload)
    return {"normalized_input": payload, "type": p_type}


def _perception_document(input_data: Any) -> Dict[str, Any]:
    """文档类输入：优先抽取文本字段供认知层处理。"""
    if isinstance(input_data, dict):
        text = (
            input_data.get("content")
            or input_data.get("text")
            or input_data.get("body")
            or input_data.get("markdown")
        )
        if isinstance(text, str) and text.strip():
            meta = {
                k: v
                for k, v in input_data.items()
                if k not in ("content", "text", "body", "markdown", "base64", "url")
            }
            return {
                "normalized_input": {"text": text, "meta": meta},
                "type": "document",
                "parse_ok": True,
            }
        if input_data.get("url"):
            return {
                "normalized_input": input_data,
                "type": "document",
                "note": "document.url 需宿主或后续链路拉取正文；已原样传入认知层",
            }
    return {"normalized_input": input_data, "type": "document"}


def _perception_web_rankings(input_data: Any) -> Dict[str, Any]:
    """排行榜 / 模型对比类感知：结构化包裹后由认知层推理（执行器内无实时爬虫）。"""
    payload = input_data if isinstance(input_data, dict) else {"raw": input_data}
    return {
        "normalized_input": {
            "ranking_task": True,
            "instructions": "请基于给定 payload 完成排序、对比或摘要；若信息不足请明确说明。",
            "payload": payload,
        },
        "type": "web_rankings",
    }


def _memory_real(
    config: Dict[str, Any], ctx: Dict[str, Any], session, user_id: int
) -> Dict[str, Any]:
    mem_cfg = _get_section(config, "memory")
    employee_id = ctx["employee_id"]
    result: Dict[str, Any] = {"session": {"employee_id": employee_id}, "long_term": None}
    short_term_cfg = mem_cfg.get("short_term") or {}
    if short_term_cfg.get("enabled", True):
        q = session.query(EmployeeExecutionMetric).filter(
            EmployeeExecutionMetric.employee_id == employee_id
        )
        if user_id > 0:
            q = q.filter(EmployeeExecutionMetric.user_id == user_id)
        recent = q.order_by(EmployeeExecutionMetric.id.desc()).limit(5).all()
        result["session"]["recent_tasks"] = [
            {
                "task": r.task,
                "status": r.status,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in recent
        ]
    long_term_cfg = mem_cfg.get("long_term") or {}
    if long_term_cfg.get("enabled", False):
        result["long_term"] = _memory_long_term_chroma(
            employee_id, ctx.get("input_data") or {}, long_term_cfg
        )
    try:
        from modstore_server.models_project_context import gather_for_employee

        result["project_context"] = gather_for_employee(
            employee_id=employee_id,
            input_data=ctx.get("input_data") or {},
        )
    except Exception:
        logger.debug("attach project_context to memory failed", exc_info=True)
    return result


async def _cognition_real(
    config: Dict[str, Any],
    perceived: Dict[str, Any],
    memory: Dict[str, Any],
    session,
    user_id: int,
    *,
    employee_id: str = "",
    task: str = "",
    bench_llm_override: Optional[Tuple[str, str]] = None,
) -> Dict[str, Any]:
    cog_cfg = _get_section(config, "cognition")
    agent = cog_cfg.get("agent") if isinstance(cog_cfg.get("agent"), dict) else cog_cfg
    system_prompt = agent.get("system_prompt", "你是智能员工助手")
    if employee_id:
        try:
            from modstore_server.prompt_evolution_ab import get_effective_system_prompt

            system_prompt = get_effective_system_prompt(str(employee_id), str(system_prompt))
        except Exception:
            pass
    # 10 项成熟度第 1 项「主动沟通」协议注入 — 让 LLM 知道有 requires_human 通道可用，
    # 在不确定优先级 / 缺资源 / 发现风险 / 需要老板决策时必须主动提问。
    # 不修改 55 个 system.md，在 cognition 层统一注入。
    if str(task or "").strip() and not _is_all_hands_cognition_context(
        perceived.get("normalized_input") if isinstance(perceived, dict) else {}
    ):
        system_prompt = f"{system_prompt.rstrip()}\n\n{_PHASE_D_PROTOCOL_APPEND}"
    model_cfg = agent.get("model") if isinstance(agent.get("model"), dict) else {}
    normalized_inp = perceived.get("normalized_input", {})
    if not isinstance(normalized_inp, dict):
        normalized_inp = {}
    all_hands_cognition = _is_all_hands_cognition_context(normalized_inp)
    if all_hands_cognition and str(task or "").strip():
        system_prompt = f"{system_prompt.rstrip()}\n\n{_ALL_HANDS_COGNITION_SYSTEM_APPEND}"

    use_platform_dispatch = bool(bench_llm_override)

    # 后台 loop 平台作用域：未显式指定 bench 时，优先用平台 bench 模型并走平台 only，
    # 跳过用户模型解析（否则 admin 名下无模型/配额会先报错或触发 403）。
    if not bench_llm_override:
        try:
            from modstore_server.platform_llm_scope import platform_llm_scope_active

            if platform_llm_scope_active():
                from modstore_server.services.llm import resolve_platform_bench_llm

                _pp, _pm = resolve_platform_bench_llm()
                if _pp and _pm:
                    bench_llm_override = (_pp, _pm)
                    use_platform_dispatch = True
        except Exception:
            pass

    if bench_llm_override:
        provider, model_name = bench_llm_override
    else:
        provider = str(model_cfg.get("provider") or "auto").strip()
        model_name = str(model_cfg.get("model_name") or "auto").strip()
        wants_auto = provider.lower() == "auto" or model_name.lower() == "auto"
        if wants_auto:
            from modstore_server.mod_scaffold_runner import resolve_llm_provider_model_auto

            uid = int(user_id or 0)
            if uid <= 0:
                from modstore_server.services.llm import resolve_platform_bench_llm

                rp, rm = resolve_platform_bench_llm()
                if not rp or not rm:
                    return {
                        "reasoning": "",
                        "error": (
                            "自动选择模型失败：平台未配置可用 LLM（请配置平台 API Key 或 "
                            "MODSTORE_EMPLOYEE_BENCH_PROVIDER / MODSTORE_EMPLOYEE_BENCH_MODEL）"
                        ),
                        "input": perceived.get("normalized_input", {}),
                        "memory": memory,
                        "knowledge": {"enabled": False, "items": [], "error": ""},
                        "provider": "auto",
                        "model": "auto",
                    }
                provider, model_name = rp, rm
                use_platform_dispatch = True
            else:
                urow = session.query(User).filter(User.id == uid).first()
                if not urow:
                    return {
                        "reasoning": "",
                        "error": "自动选择模型失败：找不到用户记录",
                        "input": perceived.get("normalized_input", {}),
                        "memory": memory,
                        "knowledge": {"enabled": False, "items": [], "error": ""},
                        "provider": "auto",
                        "model": "auto",
                    }
                rp, rm, perr = await resolve_llm_provider_model_auto(session, urow, None, None)
                if perr or not rp or not rm:
                    err_msg = perr or "无法解析可用 LLM"
                    return {
                        "reasoning": "",
                        "error": err_msg,
                        "input": perceived.get("normalized_input", {}),
                        "memory": memory,
                        "knowledge": {"enabled": False, "items": [], "error": ""},
                        "provider": "auto",
                        "model": "auto",
                    }
                provider, model_name = rp, rm
    max_tokens = int(model_cfg.get("max_tokens") or 4000)
    messages = [{"role": "system", "content": system_prompt}]
    p_cfg = _get_section(config, "perception")
    vis_cfg = p_cfg.get("vision") if isinstance(p_cfg.get("vision"), dict) else {}
    vision_enabled = bool(vis_cfg.get("enabled", True))

    mem_session = memory.get("session") if isinstance(memory, dict) else None
    session_context_json = json.dumps(mem_session, ensure_ascii=False) if mem_session else ""
    if all_hands_cognition and str(task or "").strip():
        user_input = _build_all_hands_cognition_user_message(
            task,
            normalized_inp,
            session_context_json=session_context_json,
        )
    else:
        # 修复：task 文本必须进入 user message，否则 LLM 看不到自己被要求做什么，
        # 也就无法判断是否需要触发 requires_human（Phase-D 主动沟通）。
        _task_text = str(task or "").strip()
        _inp_json = json.dumps(normalized_inp, ensure_ascii=False)
        if _task_text:
            user_input = f"任务：{_task_text}\n\n输入数据：{_inp_json}"
        else:
            user_input = _inp_json
        if session_context_json:
            user_input = f"{user_input}\n\n[session_context]\n{session_context_json}"

    v_urls: List[str] = []
    if isinstance(perceived, dict):
        vu = perceived.get("vision_data_urls")
        if isinstance(vu, list):
            v_urls = [str(u).strip() for u in vu if isinstance(u, str) and str(u).strip()]

    if vision_enabled and v_urls:
        parts: List[Dict[str, Any]] = [{"type": "text", "text": user_input}]
        for u in v_urls[:6]:
            parts.append({"type": "image_url", "image_url": {"url": u}})
        messages.append({"role": "user", "content": parts})
    else:
        messages.append({"role": "user", "content": user_input})

    knowledge_cfg = cog_cfg.get("knowledge") if isinstance(cog_cfg.get("knowledge"), dict) else {}
    rag_meta: Dict[str, Any] = {"enabled": False, "items": [], "error": ""}
    if knowledge_cfg.get("enabled"):
        try:
            from modstore_server import rag_service

            top_k = int(knowledge_cfg.get("top_k") or 6)
            min_score = float(knowledge_cfg.get("min_score") or 0.0)
            collection_ids = knowledge_cfg.get("collection_ids")
            query_text = str(task or user_input or "").strip()[:1500]
            # 集合可见性由 user_id + employee_id 解析（见 rag_service.visible_collection_ids）；定时岗位简报可设 MODSTORE_DAILY_BRIEF_USER_ID。
            chunks = await rag_service.retrieve(
                user_id=int(user_id or 0),
                query=query_text,
                employee_id=str(employee_id or "") or None,
                extra_collection_ids=collection_ids if isinstance(collection_ids, list) else None,
                top_k=top_k,
                min_score=min_score,
            )
            messages = rag_service.inject_rag_into_messages(messages, chunks)
            rag_meta = {
                "enabled": True,
                "items": [c.to_dict() for c in chunks],
                "count": len(chunks),
            }
        except Exception as e:  # noqa: BLE001 — RAG 失败不阻塞员工执行
            logger.warning("cognition.knowledge retrieve 失败: %s", e)
            rag_meta = {"enabled": True, "items": [], "error": str(e)}

    if use_platform_dispatch:
        from modstore_server.services.llm import chat_dispatch_via_platform_only

        result = await chat_dispatch_via_platform_only(
            provider, model_name, messages, max_tokens=max_tokens
        )
    else:
        result = await chat_dispatch_via_session(
            session,
            user_id,
            provider,
            model_name,
            messages,
            max_tokens=max_tokens,
        )
    if not result.get("ok"):
        err = str(result.get("error") or "llm call failed")
        if "missing api key" in err.lower():
            err = f"missing api key for provider: {provider}"
        return {
            "reasoning": "",
            "error": err,
            # 透传上游 HTTP 状态码，供失败分类区分配额/计费(402/403)与瞬时(429/5xx)。
            "status": result.get("status"),
            "input": perceived.get("normalized_input", {}),
            "memory": memory,
            "knowledge": rag_meta,
            "provider": provider,
            "model": model_name,
        }
    return {
        "reasoning": result.get("content", ""),
        "input": perceived.get("normalized_input", {}),
        "memory": memory,
        "knowledge": rag_meta,
        "provider": provider,
        "model": model_name,
        "llm_raw": result.get("raw"),
        "system_prompt": system_prompt,  # forwarded so agent runner can use it
        "_bench_platform_only": bool(use_platform_dispatch),
    }


def _cognition_sync(
    config: Dict[str, Any],
    perceived: Dict[str, Any],
    memory: Dict[str, Any],
    session,
    user_id: int,
    *,
    employee_id: str = "",
    task: str = "",
    bench_llm_override: Optional[Tuple[str, str]] = None,
) -> Dict[str, Any]:
    return _run_coro_sync(
        _cognition_real(
            config,
            perceived,
            memory,
            session,
            user_id,
            employee_id=employee_id,
            task=task,
            bench_llm_override=bench_llm_override,
        )
    )


def _action_wechat_notify(
    actions_cfg: Dict[str, Any], reasoning: Dict[str, Any], task: str
) -> Dict[str, Any]:
    """企业微信机器人 Webhook。"""
    wechat_cfg = actions_cfg.get("wechat_notify") or {}
    webhook_url = str(wechat_cfg.get("webhook_url") or "").strip()
    if not webhook_url:
        return {
            "handler": "wechat_notify",
            "status": "not_configured",
            "message": "未配置 actions.wechat_notify.webhook_url",
        }
    message_type = str(wechat_cfg.get("message_type") or "text").strip()
    content = str(reasoning.get("reasoning") or "")[:2048]
    payload: Dict[str, Any] = {"msgtype": message_type}
    if message_type == "markdown":
        payload["markdown"] = {"content": f"**AI 员工通知**\n任务: {task}\n\n{content}"}
    else:
        payload["text"] = {"content": f"【AI员工】任务:{task}\n{content}"}
    try:
        resp = httpx.post(webhook_url, json=payload, timeout=10.0)
        try:
            j = resp.json()
        except Exception:
            j = {}
        if resp.status_code == 200 and int(j.get("errcode", 0)) == 0:
            return {"handler": "wechat_notify", "status": "ok"}
        return {"handler": "wechat_notify", "status": "failed", "response": resp.text[:500]}
    except Exception as e:  # noqa: PERF203
        return {"handler": "wechat_notify", "status": "error", "error": str(e)}


def _action_openapi_tool(
    actions_cfg: Dict[str, Any],
    reasoning: Dict[str, Any],
    task: str,
    employee_id: str,
    user_id: int,
) -> Dict[str, Any]:
    """通过受控 OpenAPI 连接器调用第三方 API。

    配置示例（员工 actions.openapi_tool）::

        {
          "connector_id": 12,
          "operation_id": "createIssue",
          "params": {"project": "MOD"},
          "body": {"title": "{{task}}", "body": "{{reasoning}}"},
          "headers": {"X-Trace": "ai-employee"},
          "timeout": 20
        }
    """
    cfg = actions_cfg.get("openapi_tool") or {}
    connector_id = cfg.get("connector_id")
    operation_id = cfg.get("operation_id")
    if not connector_id or not operation_id:
        return {
            "handler": "openapi_tool",
            "error": "missing connector_id or operation_id",
        }
    try:
        connector_id_int = int(connector_id)
    except (TypeError, ValueError):
        return {"handler": "openapi_tool", "error": f"invalid connector_id: {connector_id!r}"}

    def _render(value: Any) -> Any:
        if isinstance(value, str):
            return (
                value.replace("{{reasoning}}", str(reasoning.get("reasoning") or ""))
                .replace("{{task}}", task)
                .replace("{{employee_id}}", employee_id)
            )
        if isinstance(value, dict):
            return {str(k): _render(v) for k, v in value.items()}
        if isinstance(value, list):
            return [_render(v) for v in value]
        return value

    try:
        from modstore_server.openapi_connector_runtime import call_generated_operation
    except Exception as exc:  # noqa: BLE001
        return {"handler": "openapi_tool", "error": f"runtime unavailable: {exc}"}

    timeout = float(cfg.get("timeout") or 30)
    result = call_generated_operation(
        connector_id=connector_id_int,
        user_id=int(user_id or 0),
        operation_id=str(operation_id),
        params=_render(cfg.get("params") or {}),
        body=_render(cfg.get("body")) if cfg.get("body") is not None else None,
        headers=_render(cfg.get("headers") or {}),
        timeout=timeout,
        source="employee",
    )
    return {
        "handler": "openapi_tool",
        "connector_id": connector_id_int,
        "operation_id": operation_id,
        "ok": bool(result.get("ok")),
        "status_code": result.get("status_code"),
        "body": result.get("body"),
        "error": result.get("error") or "",
        "duration_ms": result.get("duration_ms"),
    }


def _tpl_str(s: str, reasoning: Dict[str, Any], task: str) -> str:
    rtxt = str((reasoning or {}).get("reasoning") or "")
    return (s or "").replace("{{reasoning}}", rtxt).replace("{{task}}", task or "")


def _tpl_obj(obj: Any, reasoning: Dict[str, Any], task: str) -> Any:
    if isinstance(obj, str):
        return _tpl_str(obj, reasoning, task)
    if isinstance(obj, dict):
        return {str(k): _tpl_obj(v, reasoning, task) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_tpl_obj(x, reasoning, task) for x in obj]
    return obj


def _action_fhd_business(
    actions_cfg: Dict[str, Any], reasoning: Dict[str, Any], task: str
) -> Dict[str, Any]:
    biz = actions_cfg.get("fhd_business") or {}
    base = str(
        biz.get("fhd_base_url")
        or biz.get("base_url")
        or os.environ.get("FHD_BUSINESS_BASE_URL")
        or ""
    ).strip()
    path = str(biz.get("api_path") or biz.get("path") or "").strip().lstrip("/")
    method = str(biz.get("method") or "POST").strip().upper()
    if not base:
        return {"handler": "fhd_business", "error": "missing fhd_base_url"}
    if not path:
        return {"handler": "fhd_business", "error": "missing api_path"}
    raw_body = biz.get("body")
    body: Dict[str, Any] = {}
    if isinstance(raw_body, dict):
        tb = _tpl_obj(raw_body, reasoning, task)
        body = tb if isinstance(tb, dict) else {}
    headers_in = biz.get("headers") if isinstance(biz.get("headers"), dict) else {}
    hdrs = {str(k): _tpl_str(str(v), reasoning, task) for k, v in headers_in.items()}
    key = str(biz.get("business_key") or os.environ.get("FHD_BUSINESS_API_KEY") or "").strip()
    if key:
        hdrs.setdefault("X-FHD-Business-Key", key)
    url = f"{base.rstrip('/')}/api/business/{path}"
    try:
        timeout = float(biz.get("timeout") or 30.0)
        resp = httpx.request(method, url, json=body or None, headers=hdrs, timeout=timeout)
        return {
            "handler": "fhd_business",
            "url": url,
            "status_code": resp.status_code,
            "response": (resp.text or "")[:2000],
        }
    except Exception as e:  # noqa: PERF203
        return {"handler": "fhd_business", "error": str(e), "url": url}


def _action_agent_runner(
    actions_cfg: Dict[str, Any],
    reasoning: Dict[str, Any],
    task: str,
    employee_id: str,
    user_id: int,
) -> Dict[str, Any]:
    """Dispatch the ``agent`` handler by running an EmployeeAgentRunner ReAct loop.

    Reads ``actions.agent.workspace`` to determine the project root and whether
    write tools should be available.  Falls back to the reasoning text when the
    runner is unavailable.
    """
    try:
        from modstore_server.mod_employee_agent_runner import EmployeeAgentRunner
    except ImportError as exc:
        return {
            "handler": "agent",
            "ok": False,
            "error": f"EmployeeAgentRunner 未导入: {exc}",
        }

    agent_cfg = actions_cfg.get("agent") if isinstance(actions_cfg.get("agent"), dict) else {}
    ws_cfg = agent_cfg.get("workspace") if isinstance(agent_cfg.get("workspace"), dict) else {}
    read_only = bool(ws_cfg.get("read_only", True))
    requires_root = bool(ws_cfg.get("requires_project_root", False))

    # Try to get project_root from input payload first, then from the cognition result.
    cog_input = reasoning.get("input") or {}
    project_root_raw = (
        cog_input.get("project_root")
        or cog_input.get("workspace_root")
        or (reasoning.get("input") or {}).get("project_root")
    )
    workspace_root = "."
    # 读仓库树依赖调用方在 input_data 提供 project_root；Cron 简报等路径不传则无法走此分支。

    # 为员工分配持久化工作区（不覆盖调用方明确指定的 project_root）
    if not project_root_raw:
        try:
            from modstore_server.employee_workspace_manager import (
                enforce_workspace_limit,
                get_workspace_path,
            )

            _ws_path = get_workspace_path(str(employee_id or ""))
            workspace_root = str(_ws_path)
            enforce_workspace_limit(str(employee_id or ""))
        except Exception:
            pass  # 工作区不可用时静默降级到 "."

    if project_root_raw:
        try:
            from modstore_server.integrations.vibe_adapter import (
                VibePathError,
                ensure_within_workspace,
            )

            resolved = str(
                ensure_within_workspace(str(project_root_raw), user_id=int(user_id or 0))
            )
            workspace_root = resolved
        except Exception as exc:  # noqa: BLE001
            return {
                "handler": "agent",
                "ok": False,
                "error": f"project_root 路径无效: {exc}",
            }
    elif requires_root:
        return {
            "handler": "agent",
            "ok": False,
            "error": (
                "该员工需要项目根目录才能分析文件。"
                "请在 input_data 中提供 project_root 字段（例如：{'project_root': '/path/to/project'}）。"
            ),
        }

    # Build the ctx for the runner — wire up a synchronous-compatible call_llm.
    sf = get_session_factory()

    async def _agent_call_llm(messages: List[Dict[str, Any]], **kwargs) -> Dict[str, Any]:
        mt = int(kwargs.get("max_tokens") or 2048)
        temp = float(kwargs.get("temperature") or 0.2)
        # Re-use the same provider/model stored in reasoning if available.
        provider = str(reasoning.get("provider") or "auto")
        model = str(reasoning.get("model") or "auto")
        if not reasoning.get("_bench_platform_only") and (
            provider.lower() == "auto" or model.lower() == "auto"
        ):
            uid_ai = int(user_id or 0)
            if uid_ai > 0:
                from modstore_server.mod_scaffold_runner import resolve_llm_provider_model_auto

                with sf() as sess:
                    urow = sess.query(User).filter(User.id == uid_ai).first()
                    if urow:
                        rp, rm, perr = await resolve_llm_provider_model_auto(sess, urow, None, None)
                        if rp and rm and not perr:
                            provider, model = rp, rm
            else:
                from modstore_server.services.llm import resolve_platform_bench_llm

                rp, rm = resolve_platform_bench_llm()
                if rp and rm:
                    provider, model = rp, rm
        if reasoning.get("_bench_platform_only"):
            from modstore_server.services.llm import chat_dispatch_via_platform_only

            return await chat_dispatch_via_platform_only(provider, model, messages, max_tokens=mt)
        with sf() as sess:
            return await chat_dispatch_via_session(
                sess, user_id, provider, model, messages, max_tokens=mt
            )

    def _agent_http_allow_hosts() -> set[str]:
        raw = os.environ.get("MODSTORE_AGENT_HTTP_ALLOW_HOSTS", "").strip()
        if not raw:
            return set()
        return {x.strip().lower() for x in raw.split(",") if x.strip()}

    async def _allowlist_http_get(url: str, **kwargs) -> Dict[str, Any]:
        hosts = _agent_http_allow_hosts()
        if not hosts:
            return {"ok": False, "error": "未配置 MODSTORE_AGENT_HTTP_ALLOW_HOSTS"}
        try:
            parsed = urlparse(url or "")
            host = (parsed.hostname or "").lower()
        except Exception:
            return {"ok": False, "error": "无效 URL"}
        if not host or host not in hosts:
            return {"ok": False, "error": f"host 不在白名单: {host or '?'}"}
        timeout = float(kwargs.get("timeout") or 30)
        headers = kwargs.get("headers") if isinstance(kwargs.get("headers"), dict) else {}
        try:
            async with httpx.AsyncClient(timeout=timeout) as client:
                r = await client.get(url, headers=headers)
            text = (r.text or "")[:500_000]
            return {"ok": r.status_code < 400, "status": r.status_code, "text": text, "error": ""}
        except Exception as exc:  # noqa: BLE001
            return {"ok": False, "status": 0, "text": "", "error": str(exc)[:400]}

    async def _noop_http_get(url: str, **kwargs) -> Dict[str, Any]:
        return {"ok": False, "error": "agent 模式下 HTTP 工具未启用"}

    async def _noop_http_post(url: str, **kwargs) -> Dict[str, Any]:
        return {"ok": False, "error": "agent 模式下 HTTP 工具未启用"}

    _http_get_impl = _allowlist_http_get if _agent_http_allow_hosts() else _noop_http_get

    ctx: Dict[str, Any] = {
        "call_llm": _agent_call_llm,
        "http_get": _http_get_impl,
        "http_post": _noop_http_post,
        "workspace_root": workspace_root,
        "employee_id": employee_id,
        "read_only": read_only,
        "research_tools_enabled": os.environ.get("MODSTORE_AGENT_RESEARCH_TOOLS_ENABLED", "")
        .strip()
        .lower()
        in ("1", "true", "yes"),
    }
    try:
        from modstore_server.employee_scope_policy import workspace_policy_from_manifest

        with sf() as _sess:
            _pack = load_employee_pack_resolved(_sess, employee_id)
            _manifest = _pack.get("manifest") if isinstance(_pack.get("manifest"), dict) else {}
            _sg, _fg, _ag = workspace_policy_from_manifest(_manifest)
            ctx["scope_globs"] = _sg
            ctx["forbidden_globs"] = _fg
            ctx["approval_required_globs"] = _ag
    except Exception:
        ctx["scope_globs"] = []
        ctx["forbidden_globs"] = []
        ctx["approval_required_globs"] = []

    try:
        from modstore_server.integrations.ops_action_handlers import (
            OPS_EMPLOYEE_IDS,
        )
        from modstore_server.integrations.ops_action_handlers import repo_root as _ops_repo_root

        if employee_id in OPS_EMPLOYEE_IDS:
            ctx["ops_readonly_repo_root"] = str(_ops_repo_root())
    except Exception:
        pass

    # Extract system_prompt from reasoning config (populated by cognition layer).
    system_prompt = str(reasoning.get("system_prompt") or "").strip()
    if not system_prompt:
        # Fall back to the cognition section if available.
        cog_cfg = reasoning.get("cognition_cfg") or {}
        ag = cog_cfg.get("agent") if isinstance(cog_cfg.get("agent"), dict) else cog_cfg
        system_prompt = str(ag.get("system_prompt") or "").strip()

    runner = EmployeeAgentRunner(ctx, workspace_root=workspace_root)

    async def _run() -> Dict[str, Any]:
        return await runner.run(task, system_prompt=system_prompt)

    try:
        result = _run_coro_sync(_run())
    except Exception as exc:  # noqa: BLE001
        logger.exception("agent runner raised employee=%s", employee_id)
        return {"handler": "agent", "ok": False, "error": f"agent 执行异常: {exc}"}

    tool_calls = result.get("tool_calls") if isinstance(result.get("tool_calls"), list) else []
    cr_ids: set[int] = set()
    files_changed: List[Dict[str, Any]] = []
    for tc in tool_calls:
        if not isinstance(tc, dict):
            continue
        tr = tc.get("result") if isinstance(tc.get("result"), dict) else {}
        cid_raw = tr.get("change_request_id")
        try:
            cid = int(cid_raw or 0)
        except (TypeError, ValueError):
            cid = 0
        if cid > 0:
            cr_ids.add(cid)
        cids_raw = (
            tr.get("change_request_ids") if isinstance(tr.get("change_request_ids"), list) else []
        )
        for one in cids_raw:
            try:
                _cid = int(one or 0)
            except (TypeError, ValueError):
                _cid = 0
            if _cid > 0:
                cr_ids.add(_cid)
        p = str(tr.get("path") or "").strip()
        if p:
            item = {"path": p}
            if cid > 0:
                item["change_request_id"] = cid
            files_changed.append(item)

    return {
        "handler": "agent",
        "ok": result.get("ok", False),
        "summary": result.get("summary") or "",
        "rounds": result.get("rounds", 0),
        "tool_calls_count": len(result.get("tool_calls") or []),
        "change_request_ids": sorted(cr_ids),
        "files_changed": files_changed[:200],
        "workspace_root": workspace_root,
        "error": result.get("error") or "",
    }


def _employee_pack_extract_root(employee_id: str, manifest: Dict[str, Any]) -> Path:
    """Extract the employee pack to a runtime directory so package-local Python files can run."""
    runtime_root = Path(
        os.environ.get("MODSTORE_RUNTIME_DIR") or "/tmp/modstore_runtime"
    ).expanduser()
    pack_id = str(manifest.get("id") or employee_id).strip() or employee_id
    version = str(manifest.get("version") or "dev").strip() or "dev"
    target = runtime_root / "employee_packs" / pack_id / version
    module_file = target / "backend" / "employees" / "taiyangniao_attendance.py"
    sf = get_session_factory()
    with sf() as session:
        pack = load_employee_pack_resolved(session, employee_id)
    stored = str(pack.get("stored_filename") or "").strip()
    if not stored:
        raise RuntimeError("employee pack missing stored_filename")
    zpath = files_dir() / stored
    if not zpath.is_file():
        raise RuntimeError(f"employee pack file not found: {zpath}")

    marker = target / ".source_mtime"
    source_mtime = str(zpath.stat().st_mtime_ns)
    if module_file.is_file():
        try:
            if marker.read_text(encoding="utf-8").strip() == source_mtime:
                return target
        except OSError:
            pass

    target.mkdir(parents=True, exist_ok=True)
    tmp = target.with_name(target.name + ".tmp")
    if tmp.exists():
        shutil.rmtree(tmp, ignore_errors=True)
    tmp.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(zpath, "r") as zf:
        root_prefix = f"{pack_id}/"
        for info in zf.infolist():
            name = info.filename.replace("\\", "/")
            if not name.startswith(root_prefix):
                continue
            rel = name[len(root_prefix) :]
            if not rel or rel.startswith("/") or ".." in Path(rel).parts:
                continue
            dest = tmp / rel
            if info.is_dir():
                dest.mkdir(parents=True, exist_ok=True)
                continue
            dest.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(info, "r") as src, dest.open("wb") as out:
                shutil.copyfileobj(src, out)
    if target.exists():
        shutil.rmtree(target, ignore_errors=True)
    tmp.replace(target)
    marker.write_text(source_mtime + "\n", encoding="utf-8")
    return target


def _action_direct_python(
    actions_cfg: Dict[str, Any],
    reasoning: Dict[str, Any],
    task: str,
    employee_id: str,
    user_id: int = 0,
) -> Dict[str, Any]:
    """Run package-local backend/employees/*.py for script-style employee packs."""
    direct_cfg = (
        actions_cfg.get("direct_python")
        if isinstance(actions_cfg.get("direct_python"), dict)
        else {}
    )
    try:
        sf = get_session_factory()
        with sf() as session:
            pack = load_employee_pack_resolved(session, employee_id)
        manifest = pack.get("manifest") or {}
        root = _employee_pack_extract_root(employee_id, manifest)
        module_name = (
            str(direct_cfg.get("module") or "taiyangniao_attendance").strip()
            or "taiyangniao_attendance"
        )
        module_path = root / "backend" / "employees" / f"{module_name}.py"
        if not module_path.is_file():
            return {
                "handler": "direct_python",
                "ok": False,
                "error": f"module not found: {module_path}",
            }

        spec = importlib.util.spec_from_file_location(
            f"_modstore_employee_pack_{employee_id.replace('-', '_')}_{module_name}",
            str(module_path),
        )
        if spec is None or spec.loader is None:
            return {
                "handler": "direct_python",
                "ok": False,
                "error": f"cannot load module spec: {module_path}",
            }
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        run_fn = getattr(module, "run", None)
        if not callable(run_fn):
            return {
                "handler": "direct_python",
                "ok": False,
                "error": "module has no callable run(payload, ctx)",
            }

        payload = dict((reasoning or {}).get("input") or {})
        if isinstance(reasoning, dict):
            for key in ("file_path", "workspace_root", "original_filename", "action"):
                if key in reasoning and key not in payload:
                    payload[key] = reasoning[key]
        payload.setdefault("task", task)
        payload.setdefault("action", str(direct_cfg.get("action") or "convert"))
        for src, dst in (
            ("default_output_relpath", "output_relpath"),
            ("default_template_relpath", "template_relpath"),
            ("default_backend_path", "taiyangniao_backend_path"),
        ):
            val = direct_cfg.get(src)
            if val and not payload.get(dst):
                payload[dst] = val
        if (
            direct_cfg.get("default_use_personnel_roster") is not None
            and "use_personnel_roster" not in payload
        ):
            payload["use_personnel_roster"] = bool(direct_cfg.get("default_use_personnel_roster"))

        async def _direct_call_llm(messages: List[Dict[str, Any]], **kwargs) -> Dict[str, Any]:
            mt = int(kwargs.get("max_tokens") or 8000)
            provider = str(kwargs.get("provider") or "auto")
            model = str(kwargs.get("model") or "auto")
            uid = int(user_id or 0)
            if uid > 0 and (provider.lower() == "auto" or model.lower() == "auto"):
                from modstore_server.mod_scaffold_runner import resolve_llm_provider_model_auto

                with sf() as sess:
                    urow = sess.query(User).filter(User.id == uid).first()
                    if urow:
                        rp, rm, perr = await resolve_llm_provider_model_auto(sess, urow, None, None)
                        if rp and rm and not perr:
                            provider, model = rp, rm
            with sf() as sess:
                return await chat_dispatch_via_session(
                    sess, uid, provider, model, messages, max_tokens=mt
                )

        ctx = {
            "employee_id": employee_id,
            "user_id": user_id,
            "workspace_root": payload.get("workspace_root") or "",
            "logger": logging.getLogger(f"employee.direct_python.{employee_id}"),
            "call_llm": _direct_call_llm,
        }
        out = run_fn(payload, ctx)
        if asyncio.iscoroutine(out):
            out = _run_coro_sync(out)
        if isinstance(out, dict):
            return {
                "handler": "direct_python",
                "ok": bool(out.get("ok", out.get("success", True))),
                "output": out,
            }
        return {"handler": "direct_python", "ok": True, "output": out}
    except Exception as exc:  # noqa: BLE001
        logger.exception("direct_python handler failed employee_id=%s", employee_id)
        return {"handler": "direct_python", "ok": False, "error": str(exc)[:1000]}


def _filter_handlers_vibe_coding_maintainer(
    handlers: List[str],
    reasoning: Dict[str, Any],
    task: str,
) -> List[str]:
    """按 payload 路由 vibe-coding-maintainer，避免每次任务跑完全部 handler。"""
    inp = reasoning.get("input") if isinstance(reasoning.get("input"), dict) else {}
    try:
        from modstore_server.para_delegate_handler import para_delegate_enabled
    except Exception:
        para_delegate_enabled = lambda: False  # type: ignore[assignment]

    if para_delegate_enabled():
        return ["para_delegate"]

    requested = str(inp.get("handler") or "").strip()
    if requested and requested in handlers:
        out = [requested]
        if "llm_md" in handlers and requested != "llm_md":
            out.append("llm_md")
        return out

    priority = str(inp.get("priority") or "").upper()
    delegate = str(inp.get("delegate") or "").lower()
    multi_step = bool(inp.get("multi_step") or inp.get("handler_mode") == "agent")

    if delegate == "cursor" or priority == "P0" or inp.get("fallback_cursor"):
        selected: List[str] = []
        if "cursor_delegate" in handlers:
            selected.append("cursor_delegate")
        elif os.environ.get("MODSTORE_CURSOR_DELEGATE_ENABLED", "1").strip().lower() in (
            "1",
            "true",
            "yes",
            "on",
        ):
            selected.append("cursor_delegate")
        if "direct_python" in handlers:
            selected.append("direct_python")
        if "llm_md" in handlers:
            selected.append("llm_md")
        return selected or list(handlers)

    if multi_step and "agent" in handlers:
        selected = ["agent"]
        if "llm_md" in handlers:
            selected.append("llm_md")
        return selected

    selected = []
    if "vibe_edit" in handlers:
        selected.append("vibe_edit")
    if "llm_md" in handlers:
        selected.append("llm_md")
    return selected or list(handlers)


def _actions_real(
    config: Dict[str, Any],
    reasoning: Dict[str, Any],
    task: str,
    employee_id: str,
    user_id: int = 0,
) -> Dict[str, Any]:
    actions_cfg = _get_section(config, "actions")
    if employee_id == "taiyangniao-attendance-employee":
        actions_cfg = dict(actions_cfg)
        direct_cfg = dict(actions_cfg.get("direct_python") or {})
        # This pack is fully self-contained; never let stale manifests fall back to LLM echo
        # or inject an external taiyangniao-pro backend path.
        direct_cfg.pop("default_backend_path", None)
        actions_cfg["direct_python"] = direct_cfg
        actions_cfg["handlers"] = ["direct_python"]
    handlers = actions_cfg.get("handlers") or ["echo"]
    if employee_id in ("vibe-coding-maintainer", "change-request-auditor", "test-qa-runner"):
        try:
            from modstore_server.para_delegate_handler import para_delegate_enabled

            if para_delegate_enabled():
                handlers = ["para_delegate"]
        except Exception:
            pass
    if employee_id == "vibe-coding-maintainer":
        handlers = _filter_handlers_vibe_coding_maintainer(handlers, reasoning, task)
    outputs: List[Dict[str, Any]] = []
    for handler in handlers:
        if handler == "echo":
            outputs.append({"handler": "echo", "output": reasoning.get("reasoning", "")})
        elif handler == "http_request":
            http_cfg = actions_cfg.get("http_request") or {}
            url = str(http_cfg.get("url") or "").strip()
            method = str(http_cfg.get("method") or "POST").strip().upper()
            headers = http_cfg.get("headers") or {}
            body_tpl = str(http_cfg.get("body") or "")
            body = body_tpl.replace("{{reasoning}}", str(reasoning.get("reasoning") or ""))
            body = body.replace("{{task}}", task)
            if not url:
                outputs.append({"handler": "http_request", "error": "missing url"})
                continue
            try:
                resp = httpx.request(method, url, headers=headers, content=body, timeout=30.0)
                outputs.append(
                    {
                        "handler": "http_request",
                        "status_code": resp.status_code,
                        "response": resp.text[:2000],
                    }
                )
            except Exception as e:  # noqa: PERF203
                outputs.append({"handler": "http_request", "error": str(e)})
        elif handler == "webhook":
            webhook_cfg = actions_cfg.get("webhook") or {}
            url = str(webhook_cfg.get("url") or "").strip()
            if not url:
                outputs.append({"handler": "webhook", "error": "missing url"})
                continue
            payload = {
                "employee_id": employee_id,
                "task": task,
                "result": reasoning.get("reasoning", ""),
            }
            try:
                resp = httpx.post(url, json=payload, timeout=30.0)
                outputs.append({"handler": "webhook", "status_code": resp.status_code})
            except Exception as e:  # noqa: PERF203
                outputs.append({"handler": "webhook", "error": str(e)})
        elif handler == "data_sync":
            target = str((actions_cfg.get("data_sync") or {}).get("target") or "log")
            if target == "log":
                logger.info(
                    "[data_sync] employee=%s task=%s result=%s",
                    employee_id,
                    task,
                    str(reasoning.get("reasoning") or "")[:500],
                )
            outputs.append({"handler": "data_sync", "target": target, "status": "ok"})
        elif handler == "direct_python":
            outputs.append(
                _action_direct_python(actions_cfg, reasoning, task, employee_id, user_id)
            )
        elif handler == "wechat_notify":
            outputs.append(_action_wechat_notify(actions_cfg, reasoning, task))
        elif handler == "openapi_tool":
            outputs.append(_action_openapi_tool(actions_cfg, reasoning, task, employee_id, user_id))
        elif handler == "fhd_business":
            outputs.append(_action_fhd_business(actions_cfg, reasoning, task))
        elif handler == "voice_output":
            vo = (
                actions_cfg.get("voice_output")
                if isinstance(actions_cfg.get("voice_output"), dict)
                else {}
            )
            text = str(reasoning.get("reasoning") or "").strip()
            outputs.append(
                {
                    "handler": "voice_output",
                    "status": "pending_tts",
                    "note": "未配置 TTS 服务：返回待合成文本，可由宿主接入阿里云/讯飞/OpenAI TTS",
                    "text_preview": text[:800],
                    "provider": str(vo.get("provider") or "").strip(),
                    "voice_id": str(vo.get("voice_id") or "").strip(),
                }
            )
        elif handler == "agent":
            outputs.append(_action_agent_runner(actions_cfg, reasoning, task, employee_id, user_id))
        elif handler == "para_delegate":
            from modstore_server.para_delegate_handler import dispatch_para_delegate

            cog_in = reasoning.get("input") if isinstance(reasoning.get("input"), dict) else {}
            outputs.append(
                dispatch_para_delegate(
                    task=task,
                    input_data=cog_in,
                    employee_id=employee_id,
                )
            )
        elif handler == "cursor_delegate":
            from modstore_server.cursor_delegate_handler import dispatch_cursor_delegate

            cog_in = reasoning.get("input") if isinstance(reasoning.get("input"), dict) else {}
            outputs.append(
                dispatch_cursor_delegate(
                    task=task,
                    input_data=cog_in,
                    employee_id=employee_id,
                )
            )
        elif handler == "llm_md":
            # Alias: llm_md is single-shot LLM already done via cognition; return it.
            outputs.append({"handler": "llm_md", "output": reasoning.get("reasoning", "")})
        elif handler in ("vibe_edit", "vibe_heal", "vibe_code"):
            try:
                from modstore_server.integrations.vibe_action_handlers import dispatch_vibe_handler

                vibe_out = dispatch_vibe_handler(
                    str(handler), actions_cfg, reasoning, task, employee_id, user_id
                )
            except Exception as exc:  # noqa: BLE001
                logger.exception("vibe handler dispatch failed handler=%s", handler)
                vibe_out = {"handler": str(handler), "ok": False, "error": f"dispatch error: {exc}"}
            outputs.append(vibe_out or {"handler": str(handler), "ok": False, "error": "no output"})
        elif handler == "doc_sync":
            from modstore_server.integrations.doc_sync_handler import dispatch_doc_sync_handler

            outputs.append(
                dispatch_doc_sync_handler(actions_cfg, reasoning, task, employee_id, user_id)
            )
        elif handler in ("shell_exec", "ssh_exec"):
            from modstore_server.integrations.ops_action_handlers import dispatch_ops_handler

            outputs.append(
                dispatch_ops_handler(handler, actions_cfg, reasoning, task, employee_id, user_id)
            )
        else:
            outputs.append({"handler": str(handler), "error": "unknown handler"})
    return {
        "task": task,
        "handlers": handlers,
        "outputs": outputs,
        "summary": f"executed {len(outputs)} handlers",
    }


def _extract_token_count(reasoning: Dict[str, Any]) -> int:
    raw = reasoning.get("llm_raw") if isinstance(reasoning, dict) else {}
    usage = raw.get("usage") if isinstance(raw, dict) else {}
    total = usage.get("total_tokens")
    if isinstance(total, int):
        return total
    pt = usage.get("prompt_tokens")
    ct = usage.get("completion_tokens")
    return int(pt or 0) + int(ct or 0)


def _auto_wrap_execution_result_to_change_requests(
    employee_id: str,
    user_id: int,
    input_payload: Dict[str, Any],
    result: Dict[str, Any],
) -> Dict[str, Any]:
    """Glue layer: execution outputs -> EmployeeChangeRequest.

    Priority:
    1) Respect CR ids already returned by handlers (agent deferred writes).
    2) For outputs carrying ``files_changed`` with ``path + content``, auto-create CR.
    """
    outputs = result.get("outputs") if isinstance(result.get("outputs"), list) else []
    existing_ids: set[int] = set()
    file_candidates: List[Dict[str, str]] = []

    top_level_proposed = (
        result.get("proposed_changes") if isinstance(result.get("proposed_changes"), list) else []
    )
    if top_level_proposed:
        synthetic_out = {
            "files_changed": list(top_level_proposed),
            "workspace_root": str(
                input_payload.get("project_root") or input_payload.get("workspace_root") or ""
            ),
        }
        outputs = list(outputs) + [synthetic_out]

    for out in outputs:
        if not isinstance(out, dict):
            continue
        cid_raw = out.get("change_request_id")
        try:
            cid = int(cid_raw or 0)
        except (TypeError, ValueError):
            cid = 0
        if cid > 0:
            existing_ids.add(cid)
        cid_list = (
            out.get("change_request_ids") if isinstance(out.get("change_request_ids"), list) else []
        )
        for one in cid_list:
            try:
                _cid = int(one or 0)
            except (TypeError, ValueError):
                _cid = 0
            if _cid > 0:
                existing_ids.add(_cid)

        files_changed = (
            out.get("files_changed") if isinstance(out.get("files_changed"), list) else []
        )
        proposed = (
            out.get("proposed_changes") if isinstance(out.get("proposed_changes"), list) else []
        )
        if proposed:
            files_changed = list(files_changed) + list(proposed)
        for f in files_changed:
            if isinstance(f, dict):
                path = str(f.get("path") or "").strip()
                content = f.get("content")
                if not isinstance(content, str):
                    content = ""
                ws = str(
                    f.get("workspace_root")
                    or out.get("workspace_root")
                    or input_payload.get("project_root")
                    or input_payload.get("workspace_root")
                    or ""
                ).strip()
                if path:
                    file_candidates.append(
                        {
                            "path": path,
                            "content": content,
                            "workspace_root": ws,
                        }
                    )
            elif isinstance(f, str) and f.strip():
                ws = str(
                    out.get("workspace_root")
                    or input_payload.get("project_root")
                    or input_payload.get("workspace_root")
                    or ""
                ).strip()
                file_candidates.append(
                    {
                        "path": f.strip(),
                        "content": "",
                        "workspace_root": ws,
                    }
                )

    created_ids: List[int] = []
    skipped: List[Dict[str, str]] = []
    if file_candidates:
        try:
            from modstore_server.employee_change_request_service import (
                defer_write_as_change_request,
            )
            from modstore_server.employee_scope_policy import workspace_policy_from_manifest

            sf = get_session_factory()
            with sf() as session:
                try:
                    pack = load_employee_pack_resolved(session, employee_id)
                except Exception:
                    pack = {}
            manifest = pack.get("manifest") if isinstance(pack.get("manifest"), dict) else {}
            scope_globs, forbidden_globs, approval_required_globs = workspace_policy_from_manifest(
                manifest
            )
        except Exception as exc:
            return {
                "ok": False,
                "error": f"prepare CR bridge failed: {str(exc)[:300]}",
                "change_request_ids": sorted(existing_ids),
                "existing_change_request_ids": sorted(existing_ids),
                "created_change_request_ids": [],
                "skipped": [{"reason": "prepare_failed"}],
            }

        # Optional fallback: employee dedicated workspace.
        default_workspace = ""
        try:
            from modstore_server.employee_workspace_manager import get_workspace_path

            default_workspace = str(get_workspace_path(employee_id))
        except Exception:
            default_workspace = ""

        dedup_keys: set[str] = set()
        for item in file_candidates:
            path = str(item.get("path") or "").strip()
            content = str(item.get("content") or "")
            ws = str(item.get("workspace_root") or "").strip() or default_workspace
            if not path:
                skipped.append({"reason": "empty_path"})
                continue
            if not content:
                skipped.append({"path": path[:500], "reason": "missing_content"})
                continue
            if not ws:
                skipped.append({"path": path[:500], "reason": "missing_workspace_root"})
                continue
            key = f"{ws}::{path}::{hash(content)}"
            if key in dedup_keys:
                continue
            dedup_keys.add(key)
            try:
                cid = defer_write_as_change_request(
                    employee_id,
                    ws,
                    path,
                    content,
                    scope_globs=scope_globs,
                    forbidden_globs=forbidden_globs,
                    approval_required_globs=approval_required_globs,
                )
                created_ids.append(int(cid))
            except Exception as exc:
                skipped.append({"path": path[:500], "reason": str(exc)[:300]})

    all_ids = sorted(existing_ids.union(created_ids))
    return {
        "ok": True,
        "existing_change_request_ids": sorted(existing_ids),
        "created_change_request_ids": created_ids,
        "change_request_ids": all_ids,
        "processed_file_candidates": len(file_candidates),
        "skipped": skipped[:100],
    }


def _handlers_execution_ok(result: Dict[str, Any]) -> bool:
    """actions 层 outputs 中任一 handler 显式 ok=False 则视为失败。"""
    outputs = result.get("outputs") if isinstance(result.get("outputs"), list) else []
    if not outputs:
        return True
    for out in outputs:
        if isinstance(out, dict) and out.get("ok") is False:
            return False
    return True


def execute_employee_task(
    employee_id: str,
    task: str,
    input_data: Dict[str, Any] = None,
    user_id: int = 0,
    *,
    bench_llm_override: Optional[Tuple[str, str]] = None,
) -> Dict[str, Any]:
    t0 = time.perf_counter()
    payload = input_data or {}
    detail_log = _executor_detail_log_enabled()
    recovery_meta: Dict[str, Any] = {}
    logger.info(
        "employee_execute_start employee_id=%s user_id=%s task_len=%s",
        employee_id,
        user_id,
        len(task or ""),
    )
    sem = _get_executor_semaphore()
    if sem:
        sem.acquire()
    try:
        sf = get_session_factory()
        with sf() as session:
            try:
                pack = load_employee_pack_resolved(session, employee_id)
                manifest = pack.get("manifest") or {}
                config = parse_employee_config_v2(manifest)
                try:
                    from modstore_server.employee_runtime_policy import apply_policy_to_config

                    config, runtime_policy = apply_policy_to_config(employee_id, config)
                except Exception:
                    logger.debug(
                        "employee runtime policy apply failed employee_id=%s",
                        employee_id,
                        exc_info=True,
                    )
                    runtime_policy = {}
                actions_section = config.get("actions") or {}
                actions_inner = (
                    actions_section.get("actions")
                    if isinstance(actions_section.get("actions"), dict)
                    else actions_section
                )
                handler_list = list((actions_inner or {}).get("handlers") or [])

                try:
                    from modstore_server.employee_risk_middleware import gate_action_or_block

                    gate = gate_action_or_block(employee_id, manifest, handler_list, payload)
                except Exception:
                    logger.exception(
                        "risk middleware error; default to allow for backwards compatibility"
                    )
                    gate = {"ok": True, "risk_level": "unknown", "reason": "middleware error"}

                if not gate.get("ok"):
                    duration_ms = round((time.perf_counter() - t0) * 1000, 3)
                    session.add(
                        EmployeeExecutionMetric(
                            user_id=_resolve_metric_user_id(session, user_id),
                            employee_id=employee_id,
                            task=_metric_task_preview(task),
                            status="blocked_by_risk_gate",
                            duration_ms=duration_ms,
                            llm_tokens=0,
                        )
                    )
                    session.commit()
                    logger.info(
                        "employee_execute_finish employee_id=%s user_id=%s status=blocked_by_risk_gate duration_ms=%s",
                        employee_id,
                        user_id,
                        duration_ms,
                    )
                    return {
                        "employee_id": employee_id,
                        "pack": {"id": pack["pack_id"], "version": pack["version"]},
                        "duration_ms": duration_ms,
                        "result": {
                            "task": task,
                            "handlers": handler_list,
                            "outputs": [],
                            "summary": "blocked by risk middleware",
                            "risk_gate": gate,
                        },
                        "executed_at": datetime.now(timezone.utc).isoformat(),
                        "llm_tokens": 0,
                        "blocked_by_risk_gate": True,
                        "runtime_policy": runtime_policy or None,
                        "risk_level": gate.get("risk_level"),
                    }

                ctx = build_employee_context(employee_id, payload)
                perceived = _perception_real(
                    config.get("perception", {}), payload, session, user_id
                )
                # 10 项成熟度第 1 项「看得见」— 注入 workspace_signals / recent_runs /
                # recent_failures / scope_summary，让 LLM 能看到自己负责的代码和历史执行。
                try:
                    from modstore_server.employee_perception_enricher import enrich_perception
                    _project_root = (
                        str(payload.get("project_root") or "").strip()
                        if isinstance(payload, dict) else ""
                    ) or os.environ.get("MODSTORE_REPO_ROOT", "")
                    enrich_perception(
                        employee_id=employee_id,
                        perceived=perceived if isinstance(perceived, dict) else {},
                        config=config,
                        session=session,
                        project_root=Path(_project_root) if _project_root else None,
                        manifest=manifest if isinstance(manifest, dict) else None,
                    )
                except Exception as _pe_exc:
                    logger.debug("perception_enricher failed employee_id=%s err=%s", employee_id, _pe_exc)
                # 10 项成熟度第 3 项「会判断任务」— 用关键词规则给 task 分类
                # （bug/feature/ops/test/release/security/doc/handoff/unknown），
                # 注入到 perceived.normalized_input._task_classification，让 LLM 知道任务类型
                # + 给 human_report「发现什么」段反映。
                try:
                    from modstore_server.employee_task_classifier import (
                        enrich_perception_with_classification,
                    )

                    enrich_perception_with_classification(
                        employee_id=employee_id,
                        task=str(task or ""),
                        perceived=perceived if isinstance(perceived, dict) else {},
                    )
                except Exception as _tc_exc:
                    logger.debug("task_classifier failed employee_id=%s err=%s", employee_id, _tc_exc)
                # perception hook：员工像真人一样跟老板说"收到任务"——给老板
                # 一条「正在处理」的实时反馈，避免老板以为员工没动静。
                # 只在任务正文非空且短（<=200 字）时推，避免长 payload 干扰。
                try:
                    _task_preview = str(task or "").strip()
                    if _task_preview and len(_task_preview) <= 200:
                        _perceived_kind = ""
                        if isinstance(perceived, dict):
                            _perceived_kind = str(perceived.get("type") or "").strip()
                        _perception_body = f"📋 收到任务：{_task_preview}"
                        if _perceived_kind:
                            _perception_body = f"{_perception_body}\n类型：{_perceived_kind}"
                        _emp_im_notify_boss(
                            employee_id, manifest, _perception_body, "perception"
                        )
                except Exception:
                    logger.debug("perception im hook skipped", exc_info=True)
                file_path_fast = (
                    isinstance(payload, dict)
                    and str(payload.get("file_path") or payload.get("path") or "").strip()
                )
                direct_only = handler_list == ["direct_python"] and file_path_fast
                if direct_only:
                    memory: Dict[str, Any] = {}
                    reasoning = {
                        "input": dict(payload) if isinstance(payload, dict) else {},
                        "reasoning": "",
                        "skipped_cognition": True,
                    }
                    recovery_meta = {}
                else:
                    memory = _memory_real(config.get("memory", {}), ctx, session, user_id)
                    reasoning, recovery_meta = _run_cognition_with_transient_retries(
                        config.get("cognition", {}),
                        perceived,
                        memory,
                        session,
                        user_id,
                        employee_id=employee_id,
                        task=task,
                        bench_llm_override=bench_llm_override,
                    )
                # Phase-D：员工 cognition 输出 requires_human=true 时阻塞等老板回答。
                # LLM 实际输出在 reasoning["reasoning"]（字符串），需解析 JSON 才能取
                # requires_human/human_question 等字段。旧实现直接读 reasoning 顶层 dict
                # 永远拿不到 LLM 输出，导致 Phase-D 永不触发（commit 修复）。
                if isinstance(reasoning, dict):
                    _llm_out_raw = reasoning.get("reasoning") or ""
                    _parsed_llm: Dict[str, Any] = {}
                    if isinstance(_llm_out_raw, str) and _llm_out_raw.strip():
                        # 1. 直接 json.loads（LLM 直接输出纯 JSON 时）
                        try:
                            _parsed_llm = json.loads(_llm_out_raw) or {}
                        except (ValueError, TypeError):
                            _parsed_llm = {}
                        # 2. 从 ```json ... ``` 代码块抽取（贪婪 .*，匹配嵌套 JSON）
                        if not isinstance(_parsed_llm, dict):
                            import re as _re
                            _m = _re.search(
                                r"```(?:json)?\s*(\{.*\})\s*```",
                                _llm_out_raw,
                                _re.DOTALL,
                            )
                            if _m:
                                try:
                                    _parsed_llm = json.loads(_m.group(1)) or {}
                                except (ValueError, TypeError):
                                    _parsed_llm = {}
                        # 3. 剥掉首尾 ```json / ``` 后整体 json.loads
                        if not isinstance(_parsed_llm, dict):
                            _stripped = _llm_out_raw.strip()
                            if _stripped.startswith("```"):
                                _stripped = _re.sub(r"^```(?:json)?\s*", "", _stripped)
                                _stripped = _re.sub(r"\s*```\s*$", "", _stripped)
                                try:
                                    _parsed_llm = json.loads(_stripped) or {}
                                except (ValueError, TypeError):
                                    _parsed_llm = {}
                        # 4. 兜底：从首个含 requires_human 的 {...} 抽取（处理嵌套：
                        #    找 requires_human 出现的位置，向前找最近的 {，向后匹配同层 }）
                        if not isinstance(_parsed_llm, dict):
                            import re as _re
                            _m2 = _re.search(
                                r"\{[^{}]*\"requires_human\"[^{}]*\}",
                                _llm_out_raw,
                                _re.DOTALL,
                            )
                            if _m2:
                                try:
                                    _parsed_llm = json.loads(_m2.group(0)) or {}
                                except (ValueError, TypeError):
                                    _parsed_llm = {}
                    _ask_human = None
                    _human_question_text = ""
                    if isinstance(_parsed_llm, dict):
                        _ask_human = _parsed_llm.get("requires_human") or _parsed_llm.get("ask_human")
                        _human_question_text = str(_parsed_llm.get("human_question") or _parsed_llm.get("question") or "")
                    # 兼容：万一上层直接把字段塞到 reasoning 顶层
                    if _ask_human is None:
                        _ask_human = reasoning.get("requires_human") or reasoning.get("ask_human")
                    if not _human_question_text:
                        _human_question_text = str(reasoning.get("human_question") or reasoning.get("question") or "")
                    if _ask_human is True or (isinstance(_ask_human, str) and _ask_human.strip()):
                        _question_text = (
                            _ask_human
                            if isinstance(_ask_human, str) and _ask_human.strip()
                            else (_human_question_text or "需要老板决策")
                        )
                        try:
                            from modstore_server.human_uncertainty_queue import (
                                ask_human_blocking,
                            )

                            _resp = ask_human_blocking(
                                employee_id=employee_id,
                                user_id=_resolve_metric_user_id(session, user_id),
                                question=_question_text,
                                task=task,
                                context={
                                    "perceived": perceived,
                                    "reasoning_summary": str(
                                        _parsed_llm.get("summary") or reasoning.get("summary", "")
                                    )[:500],
                                    "llm_parsed": _parsed_llm if isinstance(_parsed_llm, dict) else {},
                                },
                            )
                            reasoning["_human_answer"] = _resp
                            reasoning["_phase_d_triggered"] = True
                            reasoning["_phase_d_question"] = _question_text[:300]
                            if _resp.get("status") == "answered":
                                reasoning["human_answer"] = _resp.get("answer", "")
                        except Exception as _exc:
                            reasoning["_human_answer_error"] = str(_exc)
                # cognition hook：让员工像真人一样在 IM 里主动汇报思考结果给老板。
                # 触发 Phase-D 时推"问老板：X"；否则推 cognition summary（如有）。
                try:
                    _im_body = ""
                    if reasoning.get("_phase_d_triggered"):
                        _im_q = str(reasoning.get("_phase_d_question") or "").strip()
                        if _im_q:
                            _im_body = (
                                f"🤔 我有个问题想问你：{_im_q}\n\n"
                                "（已通过任务中心发起，等你在那里回复）"
                            )
                    if not _im_body and isinstance(_parsed_llm, dict):
                        _im_body = str(
                            _parsed_llm.get("summary") or _parsed_llm.get("reasoning") or ""
                        ).strip()
                    if not _im_body and isinstance(reasoning, dict):
                        _im_body = str(reasoning.get("summary") or "").strip()
                    if _im_body:
                        _emp_im_notify_boss(employee_id, manifest, _im_body, "cognition")
                except Exception:
                    logger.debug("cognition im hook skipped", exc_info=True)
                # 10 项成熟度第 6 项 — 会协作：解析 handoff_to 并触发任务转交。
                # LLM 输出 handoff_to + handoff_reason + handoff_context 表示
                # 「这事不归我管，转给 @xxx」。本员工仍继续执行能做的部分，不阻断。
                try:
                    from modstore_server.employee_handoff import (
                        perform_handoff,
                        _resolve_target_employee_id,
                    )

                    _handoff_to_raw = None
                    _handoff_reason = ""
                    _handoff_context = ""
                    if isinstance(_parsed_llm, dict):
                        _handoff_to_raw = _parsed_llm.get("handoff_to") or _parsed_llm.get("delegate_to")
                        _handoff_reason = str(_parsed_llm.get("handoff_reason") or _parsed_llm.get("delegate_reason") or "")
                        _handoff_context = str(_parsed_llm.get("handoff_context") or _parsed_llm.get("delegate_context") or "")
                    # 兼容：LLM 把字段塞到 reasoning 顶层
                    if _handoff_to_raw is None and isinstance(reasoning, dict):
                        _handoff_to_raw = reasoning.get("handoff_to") or reasoning.get("delegate_to")
                        if not _handoff_reason:
                            _handoff_reason = str(reasoning.get("handoff_reason") or reasoning.get("delegate_reason") or "")
                        if not _handoff_context:
                            _handoff_context = str(reasoning.get("handoff_context") or reasoning.get("delegate_context") or "")

                    _handoff_target = _resolve_target_employee_id(_handoff_to_raw) if _handoff_to_raw else ""
                    if _handoff_target:
                        # 上下文摘要：LLM 解析摘要 + 感知输入类型
                        _ctx_parts = []
                        if _handoff_context:
                            _ctx_parts.append(_handoff_context[:500])
                        if isinstance(_parsed_llm, dict) and _parsed_llm.get("summary"):
                            _ctx_parts.append(f"LLM 摘要：{str(_parsed_llm.get('summary'))[:300]}")
                        _handoff_ctx_str = " | ".join(_ctx_parts)[:1500]

                        _handoff_out = perform_handoff(
                            source_employee_id=employee_id,
                            target_employee_id=_handoff_target,
                            reason=_handoff_reason,
                            context=_handoff_ctx_str,
                            original_task=str(task or ""),
                            extra_payload={
                                "source_employee_id": employee_id,
                                "parsed_llm_excerpt": str(_llm_out_raw)[:500] if isinstance(_llm_out_raw, str) else "",
                            },
                        )
                        if isinstance(reasoning, dict):
                            reasoning["_handoff"] = _handoff_out
                except Exception as _ho_exc:
                    logger.debug("handoff perform failed employee_id=%s err=%s", employee_id, _ho_exc)
                # handoff hook：员工像真人一样通知老板"我把这事转给 @xxx 了"
                try:
                    if _handoff_target:
                        _ho_msg = f"🔁 已转交给 {_handoff_target}"
                        if _handoff_reason:
                            _ho_msg = f"{_ho_msg}：{_handoff_reason[:200]}"
                        _ho_msg = f"{_ho_msg}\n（我仍继续做我能做的部分）"
                        _emp_im_notify_boss(employee_id, manifest, _ho_msg, "handoff")
                except Exception:
                    logger.debug("handoff im hook skipped", exc_info=True)
                result = _actions_real(
                    config.get("actions", {}), reasoning, task, employee_id, user_id
                )
                # 10 项成熟度第 2 项 — 知道自己管什么：硬约束 changed_files 必须在 scope_globs
                # 内、不命中 forbidden_globs。越权不阻断返回（保留证据），但标记 blocked_by_path_guard。
                try:
                    from modstore_server.employee_path_guard import check_path_guard

                    _path_guard = check_path_guard(
                        config=config,
                        result=result if isinstance(result, dict) else {},
                        employee_id=employee_id,
                    )
                    if isinstance(result, dict):
                        result["path_guard"] = _path_guard
                except Exception as _pg_exc:
                    logger.debug("path_guard check failed employee_id=%s err=%s", employee_id, _pg_exc)
                # 10 项成熟度第 5 项 — 会验证：程序化检查文件存在/测试报告/status/summary/handler output
                try:
                    from modstore_server.employee_verification import run_verification

                    _verif = run_verification(
                        employee_id=employee_id,
                        task=task,
                        reasoning=reasoning if isinstance(reasoning, dict) else {},
                        result=result if isinstance(result, dict) else {},
                        config=config,
                        project_root=Path(_project_root) if _project_root else None,
                    )
                    if isinstance(result, dict):
                        result["verification"] = _verif
                except Exception as _v_exc:
                    logger.debug("verification failed employee_id=%s err=%s", employee_id, _v_exc)
                # verification hook：员工像真人一样告诉老板"我做完验证了，通过/失败"
                try:
                    _verif_dict = _verif if isinstance(_verif, dict) else {}
                    _v_status = str(_verif_dict.get("status") or _verif_dict.get("ok") or "").strip()
                    _v_summary = str(_verif_dict.get("summary") or _verif_dict.get("message") or "").strip()
                    if _v_status or _v_summary:
                        _icon = "✅" if _v_status.lower() in ("ok", "passed", "pass", "success", "true") else (
                            "❌" if _v_status.lower() in ("fail", "failed", "error", "false") else "🔍"
                        )
                        _verif_body = f"{_icon} 验证：{_v_summary or _v_status}"[:300]
                        _emp_im_notify_boss(employee_id, manifest, _verif_body, "verification")
                except Exception:
                    logger.debug("verification im hook skipped", exc_info=True)
                # 10 项成熟度第 9 项 — 会学习：失败 >=3 次触发 prompt evolution 信号
                # 实际触发 prompt override 由 admin API / employee_autonomy_service 完成，
                # 这里只检测 + 写到 result["evolution_signal"] + 在 human_report 里反映。
                try:
                    from modstore_server.employee_self_evolution import check_evolution_signal

                    _evo = check_evolution_signal(
                        employee_id=employee_id,
                        session=session,
                    )
                    if isinstance(result, dict):
                        result["evolution_signal"] = _evo
                except Exception as _evo_exc:
                    logger.debug("evolution_signal check failed employee_id=%s err=%s", employee_id, _evo_exc)
                # 把 handoff 结果也写到 result 让 human_report 能读到
                _ho = reasoning.get("_handoff") if isinstance(reasoning, dict) else None
                if isinstance(_ho, dict) and isinstance(result, dict):
                    result["handoff"] = _ho
                duration_ms = round((time.perf_counter() - t0) * 1000, 3)
                llm_tokens = 0 if direct_only else _extract_token_count(reasoning)
                handler_ok = _handlers_execution_ok(result if isinstance(result, dict) else {})
                exec_status = "success" if handler_ok else "handler_failed"
                metric_error = "" if handler_ok else "one or more handlers returned ok=False"
                metric_failure_kind = ""
                # path_guard 越权 → 标记 blocked_by_path_guard 并附违规路径到 metric
                _pg = result.get("path_guard") if isinstance(result, dict) else None
                if isinstance(_pg, dict) and _pg.get("checked") and not _pg.get("ok"):
                    exec_status = "blocked_by_path_guard"
                    violations = _pg.get("violations") or []
                    vstr = "; ".join(
                        f"{v.get('path','')}({v.get('reason','')})"
                        for v in violations[:5]
                        if isinstance(v, dict)
                    )
                    metric_error = f"path_guard violations: {vstr}"[:500]
                    metric_failure_kind = "path_guard_violation"
                    handler_ok = False
                # 10 项成熟度第 10 项 — 会说人话：把执行结果转成 Markdown 四段式汇报
                try:
                    from modstore_server.employee_human_report import build_human_report

                    _human_report = build_human_report(
                        employee_id=employee_id,
                        task=task,
                        reasoning=reasoning if isinstance(reasoning, dict) else {},
                        result=result if isinstance(result, dict) else {},
                        duration_ms=duration_ms,
                        llm_tokens=llm_tokens,
                        exec_status=exec_status,
                        perceived=perceived if isinstance(perceived, dict) else None,
                        memory=memory if isinstance(memory, dict) else None,
                        cognition_error=str(reasoning.get("error") or "") if isinstance(reasoning, dict) else "",
                    )
                    if isinstance(result, dict):
                        result["human_report"] = _human_report
                except Exception as _hr_exc:
                    logger.debug("build_human_report failed employee_id=%s err=%s", employee_id, _hr_exc)
                if not handler_ok:
                    # 上游 LLM(认知层)失败常是 handler 失败的根因（如返回空正文）；据其
                    # 错误文本+状态码分类，让配额/计费(403)失败也带上 failure_kind，
                    # 避免自进化引擎据 handler_failed 误把额度耗尽当 prompt 问题去重写。
                    cog_err = (
                        str(reasoning.get("error") or "").strip()
                        if isinstance(reasoning, dict)
                        else ""
                    )
                    cog_status = reasoning.get("status") if isinstance(reasoning, dict) else None
                    metric_failure_kind = classify_failure_kind(cog_err or metric_error, cog_status)
                    if cog_err:
                        metric_error = f"{metric_error}; cognition_error={cog_err[:500]}"
                session.add(
                    EmployeeExecutionMetric(
                        user_id=_resolve_metric_user_id(session, user_id),
                        employee_id=employee_id,
                        task=_metric_task_preview(task),
                        status=exec_status,
                        duration_ms=duration_ms,
                        llm_tokens=llm_tokens,
                        error=metric_error,
                        failure_kind=metric_failure_kind,
                    )
                )
                session.commit()
                if not handler_ok:
                    suppress_lifecycle_events = isinstance(payload, dict) and str(
                        payload.get("suppress_lifecycle_events") or ""
                    ).strip().lower() in {"1", "true", "yes", "on"}
                    if not suppress_lifecycle_events:
                        try:
                            from modstore_server.notification_service import (
                                notify_employee_execution_done,
                            )

                            notify_employee_execution_done(user_id, employee_id, task, exec_status)
                        except Exception:
                            pass
                    return {
                        "employee_id": employee_id,
                        "pack": {"id": pack["pack_id"], "version": pack["version"]},
                        "duration_ms": duration_ms,
                        "result": result,
                        "executed_at": datetime.now(timezone.utc).isoformat(),
                        "llm_tokens": llm_tokens,
                        "handler_failed": True,
                        "runtime_policy": runtime_policy or None,
                    }
                if recovery_meta.get("recovered"):
                    try:
                        from modstore_server.services.change_signal import (
                            emit_execution_recovery_event,
                        )

                        emit_execution_recovery_event(
                            employee_id,
                            task,
                            recovery_action=str(
                                recovery_meta.get("recovery_action") or "cognition_retry"
                            ),
                            success=True,
                            original_error=str(recovery_meta.get("original_error") or ""),
                            attempts=int(recovery_meta.get("attempts") or 0),
                        )
                    except Exception:
                        logger.debug("emit_execution_recovery_event failed", exc_info=True)
                cr_bridge = _auto_wrap_execution_result_to_change_requests(
                    employee_id,
                    user_id,
                    payload if isinstance(payload, dict) else {},
                    result if isinstance(result, dict) else {},
                )
                if isinstance(result, dict):
                    result["change_request_bridge"] = cr_bridge
                    cids = (
                        cr_bridge.get("change_request_ids")
                        if isinstance(cr_bridge.get("change_request_ids"), list)
                        else []
                    )
                    if cids:
                        normalized_cids: List[int] = []
                        for x in cids:
                            try:
                                _cid = int(x or 0)
                            except (TypeError, ValueError):
                                _cid = 0
                            if _cid > 0:
                                normalized_cids.append(_cid)
                        result["change_request_ids"] = normalized_cids
                try:
                    from modstore_server.notification_service import notify_employee_execution_done

                    notify_employee_execution_done(user_id, employee_id, task, "success")
                except Exception:
                    pass
                try:
                    from modstore_server.models_project_context import (
                        record_execution_outcome,
                    )

                    record_execution_outcome(
                        employee_id=employee_id,
                        task=task,
                        input_data=payload if isinstance(payload, dict) else {},
                        outcome=result if isinstance(result, dict) else {},
                        status="success",
                    )
                except Exception:
                    pass
                suppress_lifecycle_events = isinstance(payload, dict) and str(
                    payload.get("suppress_lifecycle_events") or ""
                ).strip().lower() in {"1", "true", "yes", "on"}
                if not suppress_lifecycle_events:
                    try:
                        from modstore_server.services.change_signal import (
                            emit_signal_on_execution_complete,
                            emit_task_lifecycle_event,
                        )

                        emit_signal_on_execution_complete(
                            employee_id, task, {"status": "success", "result": result}
                        )
                        emit_task_lifecycle_event(
                            employee_id,
                            task,
                            status="success",
                            result={"result": result},
                        )
                    except Exception:
                        pass
                cog_err = ""
                if isinstance(reasoning, dict):
                    cog_err = str(reasoning.get("error") or "").strip()
                rex = ""
                if isinstance(reasoning, dict):
                    rex = str(reasoning.get("reasoning") or "").strip()[:4000]

                cog_attempts = (
                    int(recovery_meta.get("attempts") or 1) if recovery_meta.get("recovered") else 1
                )
                if detail_log:
                    logger.info(
                        "employee_execute_finish employee_id=%s user_id=%s status=success duration_ms=%s llm_tokens=%s cognition_attempts=%s handlers=%s",
                        employee_id,
                        user_id,
                        duration_ms,
                        llm_tokens,
                        cog_attempts,
                        ",".join(handler_list),
                    )
                else:
                    logger.info(
                        "employee_execute_finish employee_id=%s user_id=%s status=success duration_ms=%s llm_tokens=%s cognition_attempts=%s",
                        employee_id,
                        user_id,
                        duration_ms,
                        llm_tokens,
                        cog_attempts,
                    )

                return {
                    "employee_id": employee_id,
                    "pack": {"id": pack["pack_id"], "version": pack["version"]},
                    "duration_ms": duration_ms,
                    "result": result,
                    "executed_at": datetime.now(timezone.utc).isoformat(),
                    "llm_tokens": llm_tokens,
                    "runtime_policy": runtime_policy or None,
                    "cognition_error": cog_err or None,
                    "cognition_help": (
                        "LLM 未返回有效内容。请检查 API Key、模型名、网络与平台余额。"
                        if cog_err
                        else None
                    ),
                    "reasoning_excerpt": rex or None,
                    "change_request_ids": (
                        result.get("change_request_ids")
                        if isinstance(result, dict)
                        and isinstance(result.get("change_request_ids"), list)
                        else []
                    ),
                }
            except Exception as e:
                duration_ms = round((time.perf_counter() - t0) * 1000, 3)
                err_text = str(e)
                # 配额闸门 quota_middleware.require_llm_credit 在额度耗尽时抛
                # HTTPException(403, "配额不足: llm_calls")，会在此被捕获。分类保留
                # 「配额/计费 vs prompt」区分，否则压扁成通用 failed 会喂给自进化引擎重写。
                failure_kind = classify_failure_kind(err_text)
                session.add(
                    EmployeeExecutionMetric(
                        user_id=_resolve_metric_user_id(session, user_id),
                        employee_id=employee_id,
                        task=_metric_task_preview(task),
                        status="failed",
                        duration_ms=duration_ms,
                        llm_tokens=0,
                        error=err_text,
                        failure_kind=failure_kind,
                    )
                )
                session.commit()
                if failure_kind == FAILURE_KIND_QUOTA:
                    logger.warning(
                        "employee_execute_finish employee_id=%s user_id=%s status=failed "
                        "failure_kind=quota duration_ms=%s error=%s "
                        "(配额/计费失败，非 prompt 问题，不应触发自进化 prompt 重写)",
                        employee_id,
                        user_id,
                        duration_ms,
                        err_text[:400],
                    )
                else:
                    logger.info(
                        "employee_execute_finish employee_id=%s user_id=%s status=failed "
                        "failure_kind=%s duration_ms=%s error=%s",
                        employee_id,
                        user_id,
                        failure_kind or "unknown",
                        duration_ms,
                        err_text[:400],
                    )
                try:
                    from modstore_server.notification_service import notify_employee_execution_done

                    if user_id:
                        notify_employee_execution_done(user_id, employee_id, task, "failed")
                except Exception:
                    pass
                try:
                    from modstore_server.models_project_context import (
                        record_execution_outcome,
                    )

                    record_execution_outcome(
                        employee_id=employee_id,
                        task=task,
                        input_data=payload if isinstance(payload, dict) else {},
                        outcome={"error": str(e)},
                        status="failed",
                    )
                except Exception:
                    pass
                suppress_lifecycle_events = isinstance(payload, dict) and str(
                    payload.get("suppress_lifecycle_events") or ""
                ).strip().lower() in {"1", "true", "yes", "on"}
                if not suppress_lifecycle_events:
                    try:
                        from modstore_server.services.change_signal import emit_task_lifecycle_event

                        emit_task_lifecycle_event(employee_id, task, status="failed", error=str(e))
                    except Exception:
                        pass
                raise
    finally:
        if sem:
            sem.release()


def get_employee_status(employee_id: str) -> Dict[str, Any]:
    sf = get_session_factory()
    with sf() as session:
        rows = (
            session.query(EmployeeExecutionMetric)
            .filter(EmployeeExecutionMetric.employee_id == employee_id)
            .order_by(EmployeeExecutionMetric.id.desc())
            .limit(100)
            .all()
        )
        ok = len([r for r in rows if r.status == "success"])
        return {
            "status": "active",
            "employee_id": employee_id,
            "execution_stats": {
                "total_executions": len(rows),
                "success_count": ok,
                "failed_count": len(rows) - ok,
                "success_rate": (ok / len(rows) * 100.0) if rows else 0,
            },
            "last_execution": rows[0].created_at.isoformat() if rows else None,
        }


def list_employees() -> List[Dict[str, Any]]:
    """列出可展示的员工包：合并数据库 ``catalog_items`` 与仅存在于 ``packages.json`` 的登记。

    - ``source`` = ``catalog``：数据库中有 ``artifact=employee_pack`` 行（执行与权限以此为准）。
    - ``source`` = ``v1_catalog``：仅本地 XC catalog（``/v1/packages``）中存在，需管理员同步入库后方可稳定执行。

    同一逻辑 ``pkg_id`` 只保留一行；数据库优先于 JSON（按 :func:`~modstore_server.catalog_store.norm_pkg_id` 去重）。
    """
    from modstore_server import catalog_store
    from modstore_server.models import CatalogItem

    merged_by_norm: Dict[str, Dict[str, Any]] = {}

    sf = get_session_factory()
    with sf() as session:
        employees = session.query(CatalogItem).filter(CatalogItem.artifact == "employee_pack").all()
        for e in employees:
            nid = catalog_store.norm_pkg_id(e.pkg_id)
            if not nid:
                continue
            merged_by_norm[nid] = {
                "id": e.pkg_id,
                "name": e.name,
                "version": e.version,
                "description": e.description,
                "price": e.price,
                "industry": e.industry,
                "created_at": e.created_at.isoformat() if e.created_at else "",
                "source": "catalog",
            }

    try:
        pending_norm_raw: Dict[str, str] = {}
        for r in catalog_store.load_store().get("packages") or []:
            if not isinstance(r, dict):
                continue
            if str(r.get("artifact") or "").strip().lower() != "employee_pack":
                continue
            nid = catalog_store.norm_pkg_id(r.get("id"))
            if not nid or nid in merged_by_norm:
                continue
            rid = str(r.get("id")).strip()
            if rid:
                pending_norm_raw.setdefault(nid, rid)

        for _nid, raw_id in pending_norm_raw.items():
            versions = catalog_store.list_versions(raw_id)
            best = versions[0] if versions else None
            if not isinstance(best, dict):
                continue
            pid = str(best.get("id") or raw_id).strip()
            merged_by_norm[_nid] = {
                "id": pid,
                "name": str(best.get("name") or pid),
                "version": best.get("version"),
                "description": best.get("description"),
                "price": 0.0,
                "industry": str(best.get("industry") or "通用"),
                "created_at": str(best.get("created_at") or ""),
                "source": "v1_catalog",
            }
    except Exception as ex:  # noqa: BLE001
        logger.warning("list_employees: merge packages.json failed: %s", ex)

    out = list(merged_by_norm.values())
    out.sort(key=lambda x: str(x.get("name") or x.get("id") or "").lower())
    return out
