# ruff: noqa
"""Implementation extracted from the public facade module."""
from __future__ import annotations
import importlib


def _facade():
    return importlib.import_module("modstore_server.employee_executor")


def _executor_max_concurrent() -> int:
    raw = (_facade().os.environ.get("MODSTORE_EXECUTOR_MAX_CONCURRENT") or "").strip()
    if not raw:
        return 0
    try:
        n = int(raw)
    except ValueError:
        return 0
    return n if n > 0 else 0


def _get_executor_semaphore() -> _facade().threading.Semaphore | None:
    """Lazy singleton semaphore when MODSTORE_EXECUTOR_MAX_CONCURRENT > 0."""
    global _executor_sem, _executor_sem_n
    n = _facade()._executor_max_concurrent()
    if n <= 0:
        return None
    if _facade()._executor_sem is None or _facade()._executor_sem_n != n:
        _facade()._executor_sem = _facade().threading.Semaphore(n)
        _facade()._executor_sem_n = n
    return _facade()._executor_sem


def _executor_extra_cognition_retries() -> int:
    raw = (_facade().os.environ.get("MODSTORE_COGNITION_TRANSIENT_RETRIES") or "1").strip()
    try:
        x = int(raw)
    except ValueError:
        x = 1
    return min(max(x, 0), 2)


def _executor_detail_log_enabled() -> bool:
    return (_facade().os.environ.get("MODSTORE_EXECUTOR_LOG_DETAIL") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )


def _is_transient_llm_error(msg: str) -> bool:
    s = (msg or "").lower()
    if not s.strip():
        return False
    needles = (
        "timeout",
        "timed out",
        "connection reset",
        "connection aborted",
        "temporarily unavailable",
        "rate limit",
        "429",
        "503",
        "502",
        "bad gateway",
        "service unavailable",
        "eof occurred",
        "broken pipe",
        "connection refused",
        "connecterror",
        "readtimeout",
        "remotedisconnected",
        "try again",
        "overloaded",
    )
    return any((n in s for n in needles))


def _run_cognition_with_transient_retries(
    config: _facade().Dict[str, _facade().Any],
    perceived: _facade().Dict[str, _facade().Any],
    memory: _facade().Dict[str, _facade().Any],
    session,
    user_id: int,
    *,
    employee_id: str,
    task: str,
    bench_llm_override: _facade().Optional[_facade().Tuple[str, str]] = None,
) -> _facade().Tuple[_facade().Dict[str, _facade().Any], _facade().Dict[str, _facade().Any]]:
    """Return ``(reasoning, recovery_meta)``. ``recovery_meta`` is set when retries fixed cognition."""
    max_extra = _facade()._executor_extra_cognition_retries()
    first_err = ""
    reasoning: _facade().Dict[str, _facade().Any] = {}
    for attempt in range(max_extra + 1):
        reasoning = _facade()._cognition_sync(
            config,
            perceived,
            memory,
            session,
            user_id,
            employee_id=employee_id,
            task=task,
            bench_llm_override=bench_llm_override,
        )
        err = str(reasoning.get("error") or "").strip()
        if not err:
            if attempt > 0:
                return (
                    reasoning,
                    {
                        "recovered": True,
                        "attempts": attempt + 1,
                        "original_error": first_err[:2000],
                        "recovery_action": "cognition_retry",
                    },
                )
            return (reasoning, {})
        if attempt == 0:
            first_err = err
        if attempt >= max_extra or not _facade()._is_transient_llm_error(err):
            return (reasoning, {})
        delay = 0.4 * 2**attempt
        _facade().logger.warning(
            "employee_executor cognition transient failure employee_id=%s attempt=%s max_attempts=%s error=%s retry_delay_s=%.2f",
            employee_id,
            attempt + 1,
            max_extra + 1,
            err[:400],
            delay,
        )
        _facade().time.sleep(delay)
    return (reasoning, {})


def _get_section(
    config: _facade().Dict[str, _facade().Any], section: str
) -> _facade().Dict[str, _facade().Any]:
    if not isinstance(config, dict):
        return {}
    if section in config and isinstance(config.get(section), dict):
        return config.get(section) or {}
    return config


def _perception_excel(input_data: _facade().Any) -> _facade().Dict[str, _facade().Any]:
    """解析 .xlsx 内容（base64 或 data URL）。"""
    import base64
    import io

    try:
        import openpyxl
    except ImportError:
        return {
            "normalized_input": input_data,
            "type": "excel",
            "parse_error": "请安装 openpyxl: pip install openpyxl",
        }
    raw = input_data
    if isinstance(input_data, dict):
        raw = input_data.get("content", input_data.get("base64", ""))
    if isinstance(raw, str) and raw.startswith("data:"):
        raw = raw.split(",", 1)[-1]
    if not raw:
        return {"normalized_input": input_data, "type": "excel", "parse_error": "empty payload"}
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(raw)), read_only=True, data_only=True
        )
        sheets_data: _facade().Dict[str, _facade().Any] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: _facade().List[_facade().List[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            sheets_data[sheet_name] = {
                "rows": rows,
                "row_count": len(rows),
                "col_count": len(rows[0]) if rows else 0,
            }
        wb.close()
        return {"normalized_input": sheets_data, "type": "excel", "parse_ok": True}
    except Exception as e:
        return {"normalized_input": input_data, "type": "excel", "parse_error": str(e)}


def _extract_vision_data_urls(payload: _facade().Any) -> _facade().List[str]:
    """从 payload 抽取可送入 chat/completions 的 image_url（data: 或 https:）。"""
    out: _facade().List[str] = []
    if isinstance(payload, dict):
        for key in ("images", "image_urls", "urls"):
            val = payload.get(key)
            if isinstance(val, list):
                for u in val:
                    if isinstance(u, str) and u.strip():
                        s = u.strip()
                        out.append(
                            s if s.startswith(("data:", "http")) else f"data:image/png;base64,{s}"
                        )
        if not out:
            for key in ("image_url", "url", "base64", "content"):
                v = payload.get(key)
                if not isinstance(v, str) or not v.strip():
                    continue
                s = v.strip()
                if s.startswith("data:") or s.startswith("http"):
                    out.append(s)
                else:
                    out.append(f"data:image/png;base64,{s}")
                break
    elif isinstance(payload, str) and payload.strip():
        s = payload.strip()
        if s.startswith("data:") or s.startswith("http"):
            out.append(s)
        else:
            out.append(f"data:image/png;base64,{s}")
    return out[:8]


def _perception_image(
    input_data: _facade().Any, session, user_id: int
) -> _facade().Dict[str, _facade().Any]:
    """优先使用多模态 LLM 描述图片（需 OpenAI 兼容 Key）。"""
    vision_urls = _facade()._extract_vision_data_urls(input_data)
    raw = input_data
    if isinstance(input_data, dict):
        raw = input_data.get("base64", input_data.get("url", input_data.get("content", "")))
    if isinstance(raw, str) and raw.startswith("data:"):
        raw = raw.split(",", 1)[-1]
    if not raw:
        return {
            "normalized_input": input_data,
            "type": "image",
            "note": "图片解析需配置 OpenAI API Key，并在 input 中提供 base64",
            "vision_data_urls": vision_urls,
        }
    image_content = (
        raw if isinstance(raw, str) and raw.startswith("data:") else f"data:image/png;base64,{raw}"
    )

    async def _call():
        return await _facade().chat_dispatch_via_session(
            session,
            user_id,
            "openai",
            "gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "请简要描述图片中的文字与关键信息。"},
                        {"type": "image_url", "image_url": {"url": image_content}},
                    ],
                }
            ],
            max_tokens=800,
        )

    try:
        result = _facade()._run_coro_sync(_call())
        if result.get("ok"):
            return {
                "normalized_input": {"description": result.get("content", ""), "type": "image"},
                "type": "image",
                "parse_ok": True,
                "method": "vision",
                "vision_data_urls": vision_urls or ([image_content] if image_content else []),
            }
    except Exception as e:
        return {
            "normalized_input": input_data,
            "type": "image",
            "parse_error": str(e),
            "vision_data_urls": vision_urls,
        }
    return {
        "normalized_input": input_data,
        "type": "image",
        "note": "vision 调用未返回内容",
        "vision_data_urls": vision_urls,
    }


def _memory_long_term_chroma(
    employee_id: str,
    input_data: _facade().Dict[str, _facade().Any],
    _cfg: _facade().Dict[str, _facade().Any],
) -> _facade().Dict[str, _facade().Any]:
    """员工长期记忆：复用 ``vector_engine`` 的 PersistentClient 单例（修文件句柄泄露），
    集合保留 Chroma 默认 embedding function 以兼容历史 ``query_texts`` 写法。
    """
    query = str(input_data.get("memory_query") or input_data.get("query") or "").strip()
    if not query:
        return {
            "enabled": True,
            "memories": [],
            "note": "请在 input_data 中提供 memory_query 以检索长期记忆",
        }
    try:
        from modstore_server import vector_engine
        from modstore_server.vector_engine import VectorEngineError
    except ImportError:
        return {"enabled": True, "memories": [], "note": "请安装 chromadb: pip install chromadb"}
    try:
        client = vector_engine.get_client()
    except VectorEngineError as e:
        return {"enabled": True, "memories": [], "error": str(e)}
    coll_name = vector_engine.employee_memory_collection_name(employee_id)
    try:
        collection = client.get_or_create_collection(
            name=coll_name, metadata={"hnsw:space": "cosine"}
        )
    except Exception as e:
        return {"enabled": True, "memories": [], "error": str(e)}
    try:
        results = collection.query(query_texts=[query], n_results=5)
    except Exception as e:
        return {"enabled": True, "memories": [], "error": str(e)}
    documents = (results.get("documents") or [[]])[0] or []
    distances = (results.get("distances") or [[]])[0] or []
    memories = []
    for i, doc in enumerate(documents):
        dist = float(distances[i]) if i < len(distances) else 1.0
        if dist < 0.85:
            memories.append({"content": doc, "distance": dist})
    return {"enabled": True, "memories": memories, "count": len(memories)}


def _perception_real(
    config: _facade().Dict[str, _facade().Any],
    input_data: _facade().Dict[str, _facade().Any],
    session=None,
    user_id: int = 0,
) -> _facade().Dict[str, _facade().Any]:
    p_cfg = _facade()._get_section(config, "perception")
    p_type = str(p_cfg.get("type") or "text").strip().lower()
    payload = input_data or {}
    if p_type == "text":
        return {"normalized_input": payload, "type": "text"}
    if p_type == "json":
        if isinstance(payload, dict):
            return {"normalized_input": payload, "type": "json"}
        try:
            parsed = _facade().json.loads(payload) if isinstance(payload, str) else payload
            return {"normalized_input": parsed, "type": "json"}
        except Exception as e:
            return {"normalized_input": payload, "type": "json", "parse_error": str(e)}
    if p_type == "csv":
        if (
            isinstance(payload, dict)
            and str(payload.get("file_path") or payload.get("path") or "").strip()
        ):
            return {"normalized_input": payload, "type": "csv"}
        raw = payload.get("content", "") if isinstance(payload, dict) else str(payload)
        try:
            reader = _facade().csv.DictReader(_facade().io.StringIO(raw))
            rows = list(reader)
            return {"normalized_input": {"rows": rows}, "type": "csv", "row_count": len(rows)}
        except Exception as e:
            return {"normalized_input": payload, "type": "csv", "parse_error": str(e)}
    if p_type == "excel":
        return _facade()._perception_excel(payload)
    if p_type == "image":
        return _facade()._perception_image(payload, session, user_id)
    if p_type == "document":
        return _facade()._perception_document(payload)
    if p_type in ("web_rankings", "ai_model_rankings"):
        return _facade()._perception_web_rankings(payload)
    return {"normalized_input": payload, "type": p_type}


def _perception_document(input_data: _facade().Any) -> _facade().Dict[str, _facade().Any]:
    """文档类输入：优先抽取文本字段供认知层处理。"""
    if isinstance(input_data, dict):
        text = (
            input_data.get("content")
            or input_data.get("text")
            or input_data.get("body")
            or input_data.get("markdown")
        )
        if isinstance(text, str) and text.strip():
            meta = {
                k: v
                for (k, v) in input_data.items()
                if k not in ("content", "text", "body", "markdown", "base64", "url")
            }
            return {
                "normalized_input": {"text": text, "meta": meta},
                "type": "document",
                "parse_ok": True,
            }
        if input_data.get("url"):
            return {
                "normalized_input": input_data,
                "type": "document",
                "note": "document.url 需宿主或后续链路拉取正文；已原样传入认知层",
            }
    return {"normalized_input": input_data, "type": "document"}


def _perception_web_rankings(input_data: _facade().Any) -> _facade().Dict[str, _facade().Any]:
    """排行榜 / 模型对比类感知：结构化包裹后由认知层推理（执行器内无实时爬虫）。"""
    payload = input_data if isinstance(input_data, dict) else {"raw": input_data}
    return {
        "normalized_input": {
            "ranking_task": True,
            "instructions": "请基于给定 payload 完成排序、对比或摘要；若信息不足请明确说明。",
            "payload": payload,
        },
        "type": "web_rankings",
    }


def _memory_real(
    config: _facade().Dict[str, _facade().Any],
    ctx: _facade().Dict[str, _facade().Any],
    session,
    user_id: int,
) -> _facade().Dict[str, _facade().Any]:
    mem_cfg = _facade()._get_section(config, "memory")
    employee_id = ctx["employee_id"]
    result: _facade().Dict[str, _facade().Any] = {
        "session": {"employee_id": employee_id},
        "long_term": None,
    }
    short_term_cfg = mem_cfg.get("short_term") or {}
    if short_term_cfg.get("enabled", True):
        q = session.query(_facade().EmployeeExecutionMetric).filter(
            _facade().EmployeeExecutionMetric.employee_id == employee_id
        )
        if user_id > 0:
            q = q.filter(_facade().EmployeeExecutionMetric.user_id == user_id)
        recent = q.order_by(_facade().EmployeeExecutionMetric.id.desc()).limit(5).all()
        result["session"]["recent_tasks"] = [
            {
                "task": r.task,
                "status": r.status,
                "created_at": r.created_at.isoformat() if r.created_at else None,
            }
            for r in recent
        ]
    long_term_cfg = mem_cfg.get("long_term") or {}
    if long_term_cfg.get("enabled", False):
        result["long_term"] = _facade()._memory_long_term_chroma(
            employee_id, ctx.get("input_data") or {}, long_term_cfg
        )
    try:
        from modstore_server.models_project_context import gather_for_employee

        result["project_context"] = gather_for_employee(
            employee_id=employee_id, input_data=ctx.get("input_data") or {}
        )
    except Exception:
        _facade().logger.debug("attach project_context to memory failed", exc_info=True)
    return result


async def _cognition_real(
    config: _facade().Dict[str, _facade().Any],
    perceived: _facade().Dict[str, _facade().Any],
    memory: _facade().Dict[str, _facade().Any],
    session,
    user_id: int,
    *,
    employee_id: str = "",
    task: str = "",
    bench_llm_override: _facade().Optional[_facade().Tuple[str, str]] = None,
) -> _facade().Dict[str, _facade().Any]:
    cog_cfg = _facade()._get_section(config, "cognition")
    agent = cog_cfg.get("agent") if isinstance(cog_cfg.get("agent"), dict) else cog_cfg
    system_prompt = agent.get("system_prompt", "你是智能员工助手")
    if employee_id:
        try:
            from modstore_server.prompt_evolution_ab import get_effective_system_prompt

            system_prompt = get_effective_system_prompt(str(employee_id), str(system_prompt))
        except Exception:
            pass
    if str(task or "").strip() and (
        not _facade()._is_all_hands_cognition_context(
            perceived.get("normalized_input") if isinstance(perceived, dict) else {}
        )
    ):
        system_prompt = f"{system_prompt.rstrip()}\n\n{_facade()._PHASE_D_PROTOCOL_APPEND}"
    model_cfg = agent.get("model") if isinstance(agent.get("model"), dict) else {}
    normalized_inp = perceived.get("normalized_input", {})
    if not isinstance(normalized_inp, dict):
        normalized_inp = {}
    all_hands_cognition = _facade()._is_all_hands_cognition_context(normalized_inp)
    if all_hands_cognition and str(task or "").strip():
        system_prompt = (
            f"{system_prompt.rstrip()}\n\n{_facade()._ALL_HANDS_COGNITION_SYSTEM_APPEND}"
        )
    use_platform_dispatch = bool(bench_llm_override)
    if not bench_llm_override:
        try:
            from modstore_server.platform_llm_scope import platform_llm_scope_active

            if platform_llm_scope_active():
                from modstore_server.services.llm import resolve_platform_bench_llm

                (_pp, _pm) = resolve_platform_bench_llm()
                if _pp and _pm:
                    bench_llm_override = (_pp, _pm)
                    use_platform_dispatch = True
        except Exception:
            pass
    if not bench_llm_override and int(user_id or 0) <= 0:
        from modstore_server.services.llm import resolve_platform_bench_llm

        (_pp, _pm) = resolve_platform_bench_llm()
        if not _pp or not _pm:
            return {
                "reasoning": "",
                "error": "平台自动驾驶选择模型失败：未配置可用 LLM（请配置平台 API Key 或 MODSTORE_EMPLOYEE_BENCH_PROVIDER / MODSTORE_EMPLOYEE_BENCH_MODEL）",
                "input": perceived.get("normalized_input", {}),
                "memory": memory,
                "knowledge": {"enabled": False, "items": [], "error": ""},
                "provider": "auto",
                "model": "auto",
            }
        bench_llm_override = (_pp, _pm)
        use_platform_dispatch = True
    if bench_llm_override:
        (provider, model_name) = bench_llm_override
    else:
        provider = str(model_cfg.get("provider") or "auto").strip()
        model_name = str(model_cfg.get("model_name") or "auto").strip()
        wants_auto = provider.lower() == "auto" or model_name.lower() == "auto"
        if wants_auto:
            from modstore_server.mod_scaffold_runner import resolve_llm_provider_model_auto

            uid = int(user_id or 0)
            if uid <= 0:
                from modstore_server.services.llm import resolve_platform_bench_llm

                (rp, rm) = resolve_platform_bench_llm()
                if not rp or not rm:
                    return {
                        "reasoning": "",
                        "error": "自动选择模型失败：平台未配置可用 LLM（请配置平台 API Key 或 MODSTORE_EMPLOYEE_BENCH_PROVIDER / MODSTORE_EMPLOYEE_BENCH_MODEL）",
                        "input": perceived.get("normalized_input", {}),
                        "memory": memory,
                        "knowledge": {"enabled": False, "items": [], "error": ""},
                        "provider": "auto",
                        "model": "auto",
                    }
                (provider, model_name) = (rp, rm)
                use_platform_dispatch = True
            else:
                urow = session.query(_facade().User).filter(_facade().User.id == uid).first()
                if not urow:
                    return {
                        "reasoning": "",
                        "error": "自动选择模型失败：找不到用户记录",
                        "input": perceived.get("normalized_input", {}),
                        "memory": memory,
                        "knowledge": {"enabled": False, "items": [], "error": ""},
                        "provider": "auto",
                        "model": "auto",
                    }
                (rp, rm, perr) = await resolve_llm_provider_model_auto(session, urow, None, None)
                if perr or not rp or (not rm):
                    err_msg = perr or "无法解析可用 LLM"
                    return {
                        "reasoning": "",
                        "error": err_msg,
                        "input": perceived.get("normalized_input", {}),
                        "memory": memory,
                        "knowledge": {"enabled": False, "items": [], "error": ""},
                        "provider": "auto",
                        "model": "auto",
                    }
                (provider, model_name) = (rp, rm)
    max_tokens = int(model_cfg.get("max_tokens") or 4000)
    messages = [{"role": "system", "content": system_prompt}]
    p_cfg = _facade()._get_section(config, "perception")
    vis_cfg = p_cfg.get("vision") if isinstance(p_cfg.get("vision"), dict) else {}
    vision_enabled = bool(vis_cfg.get("enabled", True))
    mem_session = memory.get("session") if isinstance(memory, dict) else None
    session_context_json = (
        _facade().json.dumps(mem_session, ensure_ascii=False) if mem_session else ""
    )
    if all_hands_cognition and str(task or "").strip():
        user_input = _facade()._build_all_hands_cognition_user_message(
            task, normalized_inp, session_context_json=session_context_json
        )
    else:
        _task_text = str(task or "").strip()
        _inp_json = _facade().json.dumps(normalized_inp, ensure_ascii=False)
        if _task_text:
            user_input = f"任务：{_task_text}\n\n输入数据：{_inp_json}"
        else:
            user_input = _inp_json
        if session_context_json:
            user_input = f"{user_input}\n\n[session_context]\n{session_context_json}"
    v_urls: _facade().List[str] = []
    if isinstance(perceived, dict):
        vu = perceived.get("vision_data_urls")
        if isinstance(vu, list):
            v_urls = [str(u).strip() for u in vu if isinstance(u, str) and str(u).strip()]
    if vision_enabled and v_urls:
        parts: _facade().List[_facade().Dict[str, _facade().Any]] = [
            {"type": "text", "text": user_input}
        ]
        for u in v_urls[:6]:
            parts.append({"type": "image_url", "image_url": {"url": u}})
        messages.append({"role": "user", "content": parts})
    else:
        messages.append({"role": "user", "content": user_input})
    knowledge_cfg = cog_cfg.get("knowledge") if isinstance(cog_cfg.get("knowledge"), dict) else {}
    rag_meta: _facade().Dict[str, _facade().Any] = {"enabled": False, "items": [], "error": ""}
    if knowledge_cfg.get("enabled"):
        try:
            from modstore_server import rag_service

            top_k = int(knowledge_cfg.get("top_k") or 6)
            min_score = float(knowledge_cfg.get("min_score") or 0.0)
            collection_ids = knowledge_cfg.get("collection_ids")
            query_text = str(task or user_input or "").strip()[:1500]
            chunks = await rag_service.retrieve(
                user_id=int(user_id or 0),
                query=query_text,
                employee_id=str(employee_id or "") or None,
                extra_collection_ids=collection_ids if isinstance(collection_ids, list) else None,
                top_k=top_k,
                min_score=min_score,
            )
            messages = rag_service.inject_rag_into_messages(messages, chunks)
            rag_meta = {
                "enabled": True,
                "items": [c.to_dict() for c in chunks],
                "count": len(chunks),
            }
        except Exception as e:
            _facade().logger.warning("cognition.knowledge retrieve 失败: %s", e)
            rag_meta = {"enabled": True, "items": [], "error": str(e)}
    from modstore_server.mod_employee_agent_runner import _llm_timeout_seconds

    llm_timeout_s = _llm_timeout_seconds()
    try:
        if use_platform_dispatch:
            from modstore_server.services.llm import chat_dispatch_via_platform_only

            result = await _facade().asyncio.wait_for(
                chat_dispatch_via_platform_only(
                    provider, model_name, messages, max_tokens=max_tokens
                ),
                timeout=llm_timeout_s,
            )
        else:
            result = await _facade().asyncio.wait_for(
                _facade().chat_dispatch_via_session(
                    session, user_id, provider, model_name, messages, max_tokens=max_tokens
                ),
                timeout=llm_timeout_s,
            )
    except _facade().asyncio.TimeoutError:
        result = {"ok": False, "error": f"employee cognition LLM timeout ({int(llm_timeout_s)}s)"}
    if not result.get("ok"):
        err = str(result.get("error") or "llm call failed")
        if "missing api key" in err.lower():
            err = f"missing api key for provider: {provider}"
        return {
            "reasoning": "",
            "error": err,
            "status": result.get("status"),
            "input": perceived.get("normalized_input", {}),
            "memory": memory,
            "knowledge": rag_meta,
            "provider": provider,
            "model": model_name,
        }
    return {
        "reasoning": result.get("content", ""),
        "input": perceived.get("normalized_input", {}),
        "memory": memory,
        "knowledge": rag_meta,
        "provider": provider,
        "model": model_name,
        "llm_raw": result.get("raw"),
        "system_prompt": system_prompt,
        "_bench_platform_only": bool(use_platform_dispatch),
    }


def _cognition_sync(
    config: _facade().Dict[str, _facade().Any],
    perceived: _facade().Dict[str, _facade().Any],
    memory: _facade().Dict[str, _facade().Any],
    session,
    user_id: int,
    *,
    employee_id: str = "",
    task: str = "",
    bench_llm_override: _facade().Optional[_facade().Tuple[str, str]] = None,
) -> _facade().Dict[str, _facade().Any]:
    return _facade()._run_coro_sync(
        _facade()._cognition_real(
            config,
            perceived,
            memory,
            session,
            user_id,
            employee_id=employee_id,
            task=task,
            bench_llm_override=bench_llm_override,
        )
    )
