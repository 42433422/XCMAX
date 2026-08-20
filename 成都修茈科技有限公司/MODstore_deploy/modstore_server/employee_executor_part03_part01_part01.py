# mypy: disable-error-code="valid-type, attr-defined, no-any-return"
# isort: skip_file
"""Implementation extracted from the public facade module."""

from __future__ import annotations

from modstore_server.operational_errors import RECOVERABLE_ERRORS
import importlib


def _facade():
    return importlib.import_module("modstore_server.employee_executor")


def _executor_max_concurrent() -> int:
    raw = (_facade().os.environ.get("MODSTORE_EXECUTOR_MAX_CONCURRENT") or "").strip()
    if not raw:
        return 0
    try:
        n = int(raw)
    except ValueError:
        return 0
    return n if n > 0 else 0


def _get_executor_semaphore() -> _facade().threading.Semaphore | None:
    """Lazy singleton semaphore when MODSTORE_EXECUTOR_MAX_CONCURRENT > 0."""
    global _executor_sem, _executor_sem_n
    n = _facade()._executor_max_concurrent()
    if n <= 0:
        return None
    if _facade()._executor_sem is None or _facade()._executor_sem_n != n:
        _facade()._executor_sem = _facade().threading.Semaphore(n)
        _facade()._executor_sem_n = n
    return _facade()._executor_sem


def _executor_extra_cognition_retries() -> int:
    raw = (_facade().os.environ.get("MODSTORE_COGNITION_TRANSIENT_RETRIES") or "1").strip()
    try:
        x = int(raw)
    except ValueError:
        x = 1
    return min(max(x, 0), 2)


def _executor_detail_log_enabled() -> bool:
    return (_facade().os.environ.get("MODSTORE_EXECUTOR_LOG_DETAIL") or "").strip().lower() in (
        "1",
        "true",
        "yes",
        "on",
    )


def _is_transient_llm_error(msg: str) -> bool:
    s = (msg or "").lower()
    if not s.strip():
        return False
    needles = (
        "timeout",
        "timed out",
        "connection reset",
        "connection aborted",
        "temporarily unavailable",
        "rate limit",
        "429",
        "503",
        "502",
        "bad gateway",
        "service unavailable",
        "eof occurred",
        "broken pipe",
        "connection refused",
        "connecterror",
        "readtimeout",
        "remotedisconnected",
        "try again",
        "overloaded",
    )
    return any((n in s for n in needles))


def _run_cognition_with_transient_retries(
    config: _facade().Dict[str, _facade().Any],
    perceived: _facade().Dict[str, _facade().Any],
    memory: _facade().Dict[str, _facade().Any],
    session,
    user_id: int,
    *,
    employee_id: str,
    task: str,
    bench_llm_override: _facade().Optional[_facade().Tuple[str, str]] = None,
) -> _facade().Tuple[_facade().Dict[str, _facade().Any], _facade().Dict[str, _facade().Any]]:
    """Return ``(reasoning, recovery_meta)``. ``recovery_meta`` is set when retries fixed cognition."""
    max_extra = _facade()._executor_extra_cognition_retries()
    first_err = ""
    reasoning: _facade().Dict[str, _facade().Any] = {}
    for attempt in range(max_extra + 1):
        reasoning = _facade()._cognition_sync(
            config,
            perceived,
            memory,
            session,
            user_id,
            employee_id=employee_id,
            task=task,
            bench_llm_override=bench_llm_override,
        )
        err = str(reasoning.get("error") or "").strip()
        if not err:
            if attempt > 0:
                return (
                    reasoning,
                    {
                        "recovered": True,
                        "attempts": attempt + 1,
                        "original_error": first_err[:2000],
                        "recovery_action": "cognition_retry",
                    },
                )
            return (reasoning, {})
        if attempt == 0:
            first_err = err
        if attempt >= max_extra or not _facade()._is_transient_llm_error(err):
            return (reasoning, {})
        delay = 0.4 * 2**attempt
        _facade().logger.warning(
            "employee_executor cognition transient failure employee_id=%s attempt=%s max_attempts=%s error=%s retry_delay_s=%.2f",
            employee_id,
            attempt + 1,
            max_extra + 1,
            err[:400],
            delay,
        )
        _facade().time.sleep(delay)
    return (reasoning, {})


def _get_section(
    config: _facade().Dict[str, _facade().Any], section: str
) -> _facade().Dict[str, _facade().Any]:
    if not isinstance(config, dict):
        return {}
    if section in config and isinstance(config.get(section), dict):
        return config.get(section) or {}
    return config


def _perception_excel(input_data: _facade().Any) -> _facade().Dict[str, _facade().Any]:
    """解析 .xlsx 内容（base64 或 data URL）。"""
    import base64
    import io

    try:
        import openpyxl
    except ImportError:
        return {
            "normalized_input": input_data,
            "type": "excel",
            "parse_error": "请安装 openpyxl: pip install openpyxl",
        }
    raw = input_data
    if isinstance(input_data, dict):
        raw = input_data.get("content", input_data.get("base64", ""))
    if isinstance(raw, str) and raw.startswith("data:"):
        raw = raw.split(",", 1)[-1]
    if not raw:
        return {
            "normalized_input": input_data,
            "type": "excel",
            "parse_error": "empty payload",
        }
    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(base64.b64decode(raw)), read_only=True, data_only=True
        )
        sheets_data: _facade().Dict[str, _facade().Any] = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows: _facade().List[_facade().List[str]] = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else "" for c in row])
            sheets_data[sheet_name] = {
                "rows": rows,
                "row_count": len(rows),
                "col_count": len(rows[0]) if rows else 0,
            }
        wb.close()
        return {"normalized_input": sheets_data, "type": "excel", "parse_ok": True}
    except RECOVERABLE_ERRORS as e:
        return {"normalized_input": input_data, "type": "excel", "parse_error": str(e)}


def _extract_vision_data_urls(payload: _facade().Any) -> _facade().List[str]:
    """从 payload 抽取可送入 chat/completions 的 image_url（data: 或 https:）。"""
    out: _facade().List[str] = []
    if isinstance(payload, dict):
        for key in ("images", "image_urls", "urls"):
            val = payload.get(key)
            if isinstance(val, list):
                for u in val:
                    if isinstance(u, str) and u.strip():
                        s = u.strip()
                        out.append(
                            s if s.startswith(("data:", "http")) else f"data:image/png;base64,{s}"
                        )
        if not out:
            for key in ("image_url", "url", "base64", "content"):
                v = payload.get(key)
                if not isinstance(v, str) or not v.strip():
                    continue
                s = v.strip()
                if s.startswith("data:") or s.startswith("http"):
                    out.append(s)
                else:
                    out.append(f"data:image/png;base64,{s}")
                break
    elif isinstance(payload, str) and payload.strip():
        s = payload.strip()
        if s.startswith("data:") or s.startswith("http"):
            out.append(s)
        else:
            out.append(f"data:image/png;base64,{s}")
    return out[:8]


def _perception_image(
    input_data: _facade().Any, session, user_id: int
) -> _facade().Dict[str, _facade().Any]:
    """优先使用多模态 LLM 描述图片（需 OpenAI 兼容 Key）。"""
    vision_urls = _facade()._extract_vision_data_urls(input_data)
    raw = input_data
    if isinstance(input_data, dict):
        raw = input_data.get("base64", input_data.get("url", input_data.get("content", "")))
    if isinstance(raw, str) and raw.startswith("data:"):
        raw = raw.split(",", 1)[-1]
    if not raw:
        return {
            "normalized_input": input_data,
            "type": "image",
            "note": "图片解析需配置 OpenAI API Key，并在 input 中提供 base64",
            "vision_data_urls": vision_urls,
        }
    image_content = (
        raw if isinstance(raw, str) and raw.startswith("data:") else f"data:image/png;base64,{raw}"
    )

    async def _call():
        return await _facade().chat_dispatch_via_session(
            session,
            user_id,
            "openai",
            "gpt-4o-mini",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "请简要描述图片中的文字与关键信息。"},
                        {"type": "image_url", "image_url": {"url": image_content}},
                    ],
                }
            ],
            max_tokens=800,
        )

    try:
        result = _facade()._run_coro_sync(_call())
        if result.get("ok"):
            return {
                "normalized_input": {
                    "description": result.get("content", ""),
                    "type": "image",
                },
                "type": "image",
                "parse_ok": True,
                "method": "vision",
                "vision_data_urls": vision_urls or ([image_content] if image_content else []),
            }
    except RECOVERABLE_ERRORS as e:
        return {
            "normalized_input": input_data,
            "type": "image",
            "parse_error": str(e),
            "vision_data_urls": vision_urls,
        }
    return {
        "normalized_input": input_data,
        "type": "image",
        "note": "vision 调用未返回内容",
        "vision_data_urls": vision_urls,
    }


def _memory_long_term_chroma(
    employee_id: str,
    input_data: _facade().Dict[str, _facade().Any],
    _cfg: _facade().Dict[str, _facade().Any],
) -> _facade().Dict[str, _facade().Any]:
    """员工长期记忆：复用 ``vector_engine`` 的 PersistentClient 单例（修文件句柄泄露），
    集合保留 Chroma 默认 embedding function 以兼容历史 ``query_texts`` 写法。
    """
    query = str(input_data.get("memory_query") or input_data.get("query") or "").strip()
    if not query:
        return {
            "enabled": True,
            "memories": [],
            "note": "请在 input_data 中提供 memory_query 以检索长期记忆",
        }
    try:
        from modstore_server import vector_engine
        from modstore_server.vector_engine import VectorEngineError
    except ImportError:
        return {
            "enabled": True,
            "memories": [],
            "note": "请安装 chromadb: pip install chromadb",
        }
    try:
        client = vector_engine.get_client()
    except VectorEngineError as e:
        return {"enabled": True, "memories": [], "error": str(e)}
    coll_name = vector_engine.employee_memory_collection_name(employee_id)
    try:
        collection = client.get_or_create_collection(
            name=coll_name, metadata={"hnsw:space": "cosine"}
        )
    except RECOVERABLE_ERRORS as e:
        return {"enabled": True, "memories": [], "error": str(e)}
    try:
        results = collection.query(query_texts=[query], n_results=5)
    except RECOVERABLE_ERRORS as e:
        return {"enabled": True, "memories": [], "error": str(e)}
    documents = (results.get("documents") or [[]])[0] or []
    distances = (results.get("distances") or [[]])[0] or []
    memories = []
    for i, doc in enumerate(documents):
        dist = float(distances[i]) if i < len(distances) else 1.0
        if dist < 0.85:
            memories.append({"content": doc, "distance": dist})
    return {"enabled": True, "memories": memories, "count": len(memories)}


def _perception_real(
    config: _facade().Dict[str, _facade().Any],
    input_data: _facade().Dict[str, _facade().Any],
    session=None,
    user_id: int = 0,
) -> _facade().Dict[str, _facade().Any]:
    p_cfg = _facade()._get_section(config, "perception")
    p_type = str(p_cfg.get("type") or "text").strip().lower()
    payload = input_data or {}
    if p_type == "text":
        return {"normalized_input": payload, "type": "text"}
    if p_type == "json":
        if isinstance(payload, dict):
            return {"normalized_input": payload, "type": "json"}
        try:
            parsed = _facade().json.loads(payload) if isinstance(payload, str) else payload
            return {"normalized_input": parsed, "type": "json"}
        except RECOVERABLE_ERRORS as e:
            return {"normalized_input": payload, "type": "json", "parse_error": str(e)}
    if p_type == "csv":
        if (
            isinstance(payload, dict)
            and str(payload.get("file_path") or payload.get("path") or "").strip()
        ):
            return {"normalized_input": payload, "type": "csv"}
        raw = payload.get("content", "") if isinstance(payload, dict) else str(payload)
        try:
            reader = _facade().csv.DictReader(_facade().io.StringIO(raw))
            rows = list(reader)
            return {
                "normalized_input": {"rows": rows},
                "type": "csv",
                "row_count": len(rows),
            }
        except RECOVERABLE_ERRORS as e:
            return {"normalized_input": payload, "type": "csv", "parse_error": str(e)}
    if p_type == "excel":
        return _facade()._perception_excel(payload)
    if p_type == "image":
        return _facade()._perception_image(payload, session, user_id)
    if p_type == "document":
        return _facade()._perception_document(payload)
    if p_type in ("web_rankings", "ai_model_rankings"):
        return _facade()._perception_web_rankings(payload)
    return {"normalized_input": payload, "type": p_type}


def _perception_document(
    input_data: _facade().Any,
) -> _facade().Dict[str, _facade().Any]:
    """文档类输入：优先抽取文本字段供认知层处理。"""
    if isinstance(input_data, dict):
        text = (
            input_data.get("content")
            or input_data.get("text")
            or input_data.get("body")
            or input_data.get("markdown")
        )
        if isinstance(text, str) and text.strip():
            meta = {
                k: v
                for (k, v) in input_data.items()
                if k not in ("content", "text", "body", "markdown", "base64", "url")
            }
            return {
                "normalized_input": {"text": text, "meta": meta},
                "type": "document",
                "parse_ok": True,
            }
        if input_data.get("url"):
            return {
                "normalized_input": input_data,
                "type": "document",
                "note": "document.url 需宿主或后续链路拉取正文；已原样传入认知层",
            }
    return {"normalized_input": input_data, "type": "document"}


def _perception_web_rankings(
    input_data: _facade().Any,
) -> _facade().Dict[str, _facade().Any]:
    """排行榜 / 模型对比类感知：结构化包裹后由认知层推理（执行器内无实时爬虫）。"""
    payload = input_data if isinstance(input_data, dict) else {"raw": input_data}
    return {
        "normalized_input": {
            "ranking_task": True,
            "instructions": "请基于给定 payload 完成排序、对比或摘要；若信息不足请明确说明。",
            "payload": payload,
        },
        "type": "web_rankings",
    }
