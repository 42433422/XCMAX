"""Excel 全量读取与 Excel 生成员工：检测、规则、兜底 convert 与包体验证（JSON 为中介）。"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Tuple

from modstore_server import excel_tabular_specs as _specs

build_excel_generate_rule_spec = _specs.build_excel_generate_rule_spec
build_excel_read_rule_spec = _specs.build_excel_read_rule_spec
excel_generate_structured_spec = _specs.excel_generate_structured_spec
excel_read_structured_spec = _specs.excel_read_structured_spec
is_excel_full_read = _specs.is_excel_full_read
is_excel_generate = _specs.is_excel_generate


def render_excel_read_convert_module() -> str:
    return r"""from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

MAX_ROW_CAP = 500
MAX_COL_CAP = 100


def _cell_payload(ws, row: int, col: int, *, formula_ws=None) -> Dict[str, Any]:
    from openpyxl.utils import get_column_letter

    cell = ws.cell(row, col)
    letter = get_column_letter(col)
    raw = cell.value
    display = "" if raw is None else str(raw)
    formula = None
    if formula_ws is not None:
        fcell = formula_ws.cell(row, col)
        if isinstance(fcell.value, str) and fcell.value.startswith("="):
            formula = fcell.value
    data_type = getattr(cell, "data_type", None)
    return {
        "row": row,
        "col": col,
        "letter": letter,
        "value": raw,
        "display": display,
        "formula": formula,
        "data_type": str(data_type) if data_type is not None else None,
    }


def _sheet_to_dict(ws, formula_ws=None) -> Dict[str, Any]:
    max_row = int(ws.max_row or 0)
    max_col = int(ws.max_column or 0)
    cap_row = min(max_row, MAX_ROW_CAP)
    cap_col = min(max_col, MAX_COL_CAP)
    truncated = max_row > cap_row or max_col > cap_col

    cells: List[Dict[str, Any]] = []
    for r in range(1, cap_row + 1):
        for c in range(1, cap_col + 1):
            cell = ws.cell(r, c)
            if cell.value is None and (formula_ws is None or formula_ws.cell(r, c).value is None):
                continue
            cells.append(_cell_payload(ws, r, c, formula_ws=formula_ws))

    headers: List[Dict[str, Any]] = []
    if cap_row >= 1:
        for c in range(1, cap_col + 1):
            v = ws.cell(1, c).value
            if v is not None and str(v).strip():
                headers.append(_cell_payload(ws, 1, c, formula_ws=formula_ws))

    header_names = [str(h.get("display") or h.get("value") or "").strip() for h in headers]
    header_names = [n for n in header_names if n]

    rows_out: List[Dict[str, Any]] = []
    for r in range(2, cap_row + 1):
        row_cells: Dict[str, Any] = {}
        has_data = False
        for c in range(1, cap_col + 1):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip() != "":
                has_data = True
            key = header_names[c - 1] if c - 1 < len(header_names) else f"col_{c}"
            row_cells[key] = v
        if has_data:
            rows_out.append({"row_index": r, "cells": row_cells})

    return {
        "name": ws.title,
        "max_row": max_row,
        "max_column": max_col,
        "truncated": truncated,
        "headers": headers,
        "rows": rows_out,
        "cells": cells,
        "cell_count": len(cells),
        "row_count": len(rows_out),
    }


def convert_file(
    src_path: Path,
    output_path: Path,
    *,
    template_path: Optional[Path] = None,
    payload: Dict[str, Any],
    ctx: Dict[str, Any],
    rule_spec: Dict[str, Any],
) -> Dict[str, Any]:
    suffix = src_path.suffix.lower()
    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError(f"不支持的文件类型：{suffix or '(无后缀)'}，仅支持 .xlsx / .xlsm")

    from openpyxl import load_workbook

    output_dir = output_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    json_path = output_dir / "workbook.json"
    if output_path.suffix.lower() == ".json":
        json_path = output_path
    elif str(rule_spec.get("default_output_relpath") or "").endswith(".json"):
        json_path = output_dir / Path(str(rule_spec.get("default_output_relpath"))).name

    wb_val = load_workbook(src_path, read_only=False, data_only=True)
    wb_formula = load_workbook(src_path, read_only=False, data_only=False)
    sheets: List[Dict[str, Any]] = []
    try:
        for ws in wb_val.worksheets:
            fws = wb_formula[ws.title] if ws.title in wb_formula.sheetnames else None
            sheets.append(_sheet_to_dict(ws, fws))
    finally:
        wb_val.close()
        wb_formula.close()

    payload_data: Dict[str, Any] = {
        "source": src_path.name,
        "sheet_count": len(sheets),
        "sheets": sheets,
        "meta": {
            "source": src_path.name,
            "byte_size": src_path.stat().st_size,
            "max_row_cap": MAX_ROW_CAP,
            "max_col_cap": MAX_COL_CAP,
        },
    }
    json_path.write_text(json.dumps(payload_data, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    total_cells = sum(int(s.get("cell_count") or 0) for s in sheets)
    return {
        "output_path": str(json_path),
        "sheet_count": len(sheets),
        "cell_count": total_cells,
        "output_schema": list(rule_spec.get("output_schema") or []),
    }
"""


def render_excel_generate_convert_module() -> str:
    return r"""from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from modstore_server.office_plaintext_generate import resolve_table_spec


def _normalize_sheets(table: Dict[str, Any]) -> List[Dict[str, Any]]:
    sheets_in = table.get("sheets")
    if isinstance(sheets_in, list) and sheets_in:
        out: List[Dict[str, Any]] = []
        for idx, sh in enumerate(sheets_in):
            if not isinstance(sh, dict):
                continue
            name = str(sh.get("name") or f"Sheet{idx + 1}")
            columns = [str(c) for c in (sh.get("columns") or []) if str(c).strip()]
            rows_in = sh.get("rows")
            if not isinstance(rows_in, list):
                rows_in = []
            if not columns and rows_in and isinstance(rows_in[0], dict):
                columns = [str(k) for k in rows_in[0].keys()]
            out.append({"name": name, "columns": columns, "rows": rows_in})
        if out:
            return out
    columns = [str(c) for c in (table.get("columns") or []) if str(c).strip()]
    rows_in = table.get("rows")
    if not isinstance(rows_in, list):
        rows_in = []
    if not columns and rows_in and isinstance(rows_in[0], dict):
        columns = [str(k) for k in rows_in[0].keys()]
    return [{"name": str(table.get("sheet") or table.get("sheet_name") or "Sheet1"), "columns": columns, "rows": rows_in}]


async def convert_file(
    src_path: Path,
    output_path: Path,
    *,
    template_path: Optional[Path] = None,
    payload: Dict[str, Any],
    ctx: Dict[str, Any],
    rule_spec: Dict[str, Any],
) -> Dict[str, Any]:
    from openpyxl import Workbook

    table, _warnings = await resolve_table_spec(src_path, payload or {}, ctx or {}, rule_spec or {}, fmt="excel")
    sheets = _normalize_sheets(table)

    output_dir = output_path.parent
    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_dir / "output.xlsx"
    if output_path.suffix.lower() in {".xlsx", ".xlsm"}:
        xlsx_path = output_path
    elif str(rule_spec.get("default_output_relpath") or "").endswith(".xlsx"):
        xlsx_path = output_dir / Path(str(rule_spec.get("default_output_relpath"))).name

    wb = Workbook()
    first = True
    total_rows = 0
    for sh in sheets:
        name = str(sh.get("name") or "Sheet1")
        columns = list(sh.get("columns") or [])
        rows_in = sh.get("rows") if isinstance(sh.get("rows"), list) else []
        if first:
            ws = wb.active
            ws.title = name[:31]
            first = False
        else:
            ws = wb.create_sheet(title=name[:31])
        for col_idx, header in enumerate(columns, 1):
            ws.cell(1, col_idx, header)
        for row_idx, row in enumerate(rows_in, 2):
            if not isinstance(row, dict):
                continue
            for col_idx, header in enumerate(columns, 1):
                ws.cell(row_idx, col_idx, row.get(header, ""))
            total_rows += 1
    wb.save(xlsx_path)
    wb.close()

    return {
        "output_path": str(xlsx_path),
        "sheet_count": len(sheets),
        "row_count": total_rows,
        "output_schema": list(rule_spec.get("output_schema") or []),
    }
"""


def validate_excel_read_backend(pack_dir: Path) -> Tuple[List[str], List[str]]:
    return _validate_excel_backend(
        pack_dir,
        runtime_kind="excel_full_read",
        required_tokens=("workbook.json", "openpyxl", "sheets"),
    )


def validate_excel_generate_backend(pack_dir: Path) -> Tuple[List[str], List[str]]:
    return _validate_excel_backend(
        pack_dir,
        runtime_kind="excel_generate",
        required_tokens=("output.xlsx", "openpyxl", "workbook"),
    )


def _validate_excel_backend(
    pack_dir: Path,
    *,
    runtime_kind: str,
    required_tokens: tuple[str, ...],
) -> Tuple[List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []
    backend = pack_dir / "backend"
    if not backend.is_dir():
        errors.append("缺少 backend 目录")
        return errors, warnings

    py_blob = ""
    has_convert = False
    for py_path in backend.rglob("*.py"):
        try:
            text = py_path.read_text(encoding="utf-8", errors="ignore")
            py_blob += text.lower()
            if "def convert_file" in text and "vendor" in str(py_path).lower():
                has_convert = True
        except OSError:
            pass

    mf_path = pack_dir / "manifest.json"
    handlers: List[str] = []
    if mf_path.is_file():
        try:
            from modstore_server.employee_asset_pipeline import manifest_actions_handlers

            mf = json.loads(mf_path.read_text(encoding="utf-8"))
            handlers = manifest_actions_handlers(mf)
        except (json.JSONDecodeError, OSError):
            warnings.append("manifest.json 无法解析")

    if handlers and "direct_python" not in handlers:
        errors.append(f"{runtime_kind} 员工 handlers 必须包含 direct_python")
    if not has_convert:
        errors.append("backend/vendor 中缺少 convert_file 实现")
    if "openpyxl" not in py_blob:
        warnings.append("未发现 openpyxl 相关代码")
    for tok in required_tokens:
        if tok.lower() not in py_blob:
            warnings.append(f"convert 模块可能未覆盖：{tok}")

    rs_path = pack_dir / "rule_spec.json"
    if rs_path.is_file():
        try:
            rs = json.loads(rs_path.read_text(encoding="utf-8"))
            if isinstance(rs, dict) and rs.get("runtime_kind") != runtime_kind:
                warnings.append(f"rule_spec.runtime_kind 期望 {runtime_kind}")
        except (OSError, json.JSONDecodeError):
            warnings.append("rule_spec.json 无法解析")

    return errors, warnings


def minimal_xlsx_fixture_bytes() -> bytes:
    from io import BytesIO

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "name"
    ws["B1"] = "score"
    ws["A2"] = "alice"
    ws["B2"] = 90
    ws["A3"] = "bob"
    ws["B3"] = 85
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def minimal_excel_table_json() -> Dict[str, Any]:
    return {
        "sheets": [
            {
                "name": "Sheet1",
                "columns": ["name", "score"],
                "rows": [{"name": "alice", "score": 90}, {"name": "bob", "score": 85}],
            }
        ],
        "meta": {"source": "fixture"},
    }


def minimal_json_fixture_bytes() -> bytes:
    return json.dumps(minimal_excel_table_json(), ensure_ascii=False, indent=2).encode("utf-8")


def excel_read_orchestration_plan(brief: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    from modstore_server.employee_brief_utils import compact_routing_brief

    clean = compact_routing_brief(brief, max_len=400) or (brief or "").strip()
    short = "Excel 全量读取员"
    return {
        "employee_name": short,
        "employee_brief": (
            f"{clean}\n\n"
            "员工必须使用 direct_python 将 xlsx 解析为 outputs/workbook.json（含 sheet、表头、单元格全量），禁止编造数据。"
        ),
        "script_workflow_name": f"{short} 脚本工作流",
        "script_brief": f"{clean}\n\n读取 inputs/*.xlsx，写出 outputs/workbook.json。",
        "script_runtime_notes": "只能读 inputs/、写 outputs/；使用 openpyxl。",
        "workflow_name": str(payload.get("employee_workflow_name") or short).strip() or short,
        "workflow_brief": f"{clean}\n\nSkill：上传 xlsx → 全量单元格 JSON。",
        "acceptance": [
            "handlers 为 direct_python",
            "workbook.json 含 sheets、cells",
        ],
    }


def excel_generate_orchestration_plan(brief: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    from modstore_server.employee_brief_utils import compact_routing_brief

    clean = compact_routing_brief(brief, max_len=400) or (brief or "").strip()
    short = "Excel 生成员"
    return {
        "employee_name": short,
        "employee_brief": (
            f"{clean}\n\n"
            "员工必须使用 direct_python 从 JSON（sheets/columns/rows）写出 outputs/output.xlsx，禁止编造表格。"
        ),
        "script_workflow_name": f"{short} 脚本工作流",
        "script_brief": f"{clean}\n\n读取 inputs/*.json，写出 outputs/output.xlsx。",
        "script_runtime_notes": "JSON 为中介；使用 openpyxl.Workbook。",
        "workflow_name": str(payload.get("employee_workflow_name") or short).strip() or short,
        "workflow_brief": f"{clean}\n\nSkill：JSON → xlsx 落盘。",
        "acceptance": [
            "handlers 为 direct_python",
            "output.xlsx 与 JSON 一致",
        ],
    }


def resolve_excel_orchestration_plan(brief: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    if is_excel_generate(brief):
        return excel_generate_orchestration_plan(brief, payload)
    return excel_read_orchestration_plan(brief, payload)
