"""Parse uploaded knowledge files into text chunks.

Supports two chunking strategies:
- ``fixed`` (default): sliding window with configurable size and overlap.
- ``semantic``: sentence-boundary splitting with similarity-breakpoint detection.
  Falls back to ``fixed`` when embedding service is unavailable.
"""

from __future__ import annotations

import csv
import io
import json
import logging
import math
import os
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from fastapi import HTTPException

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {".txt", ".md", ".json", ".csv", ".pdf", ".docx", ".xlsx"}
MAX_UPLOAD_BYTES = int(os.environ.get("MODSTORE_KB_MAX_UPLOAD_BYTES", str(20 * 1024 * 1024)))

_CHUNK_STRATEGY = (os.environ.get("MODSTORE_KB_CHUNK_STRATEGY") or "fixed").strip().lower()


def _chunk_size() -> int:
    return max(400, min(int(os.environ.get("MODSTORE_KB_CHUNK_SIZE", "1000")), 4000))


def _chunk_overlap() -> int:
    return max(0, min(int(os.environ.get("MODSTORE_KB_CHUNK_OVERLAP", "120")), 800))


def _semantic_threshold() -> float:
    return max(0.1, min(float(os.environ.get("MODSTORE_KB_SEMANTIC_THRESHOLD", "0.5")), 0.95))


def _semantic_min_chunk() -> int:
    return max(100, min(int(os.environ.get("MODSTORE_KB_SEMANTIC_MIN_CHUNK", "200")), 2000))


def _semantic_max_chunk() -> int:
    return max(400, min(int(os.environ.get("MODSTORE_KB_SEMANTIC_MAX_CHUNK", "2000")), 8000))


def _decode_text(raw: bytes) -> str:
    for enc in ("utf-8", "utf-8-sig", "gb18030"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")


def _normalize_text(text: str) -> str:
    t = (text or "").replace("\r\n", "\n").replace("\r", "\n")
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


def _parse_pdf(raw: bytes) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise HTTPException(503, "服务器未安装 pypdf，暂不能解析 PDF") from e
    reader = PdfReader(io.BytesIO(raw))
    return "\n\n".join((page.extract_text() or "").strip() for page in reader.pages)


def _parse_pdf_pages(raw: bytes) -> List[Tuple[int, str]]:
    try:
        from pypdf import PdfReader
    except ImportError as e:
        raise HTTPException(503, "服务器未安装 pypdf，暂不能解析 PDF") from e
    reader = PdfReader(io.BytesIO(raw))
    pages: List[Tuple[int, str]] = []
    for idx, page in enumerate(reader.pages, start=1):
        text = _normalize_text((page.extract_text() or "").strip())
        if text:
            pages.append((idx, text))
    return pages


def _parse_docx(raw: bytes) -> str:
    try:
        import docx
    except ImportError as e:
        raise HTTPException(503, "服务器未安装 python-docx，暂不能解析 DOCX") from e
    doc = docx.Document(io.BytesIO(raw))
    return "\n".join(p.text for p in doc.paragraphs if p.text)


def _parse_xlsx(raw: bytes) -> str:
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError as e:
        raise HTTPException(503, "服务器未安装 openpyxl，暂不能解析 XLSX") from e
    max_rows = max(10, min(int(os.environ.get("MODSTORE_XLSX_PREVIEW_ROWS", "80")), 300))
    max_cols = max(8, min(int(os.environ.get("MODSTORE_XLSX_PREVIEW_COLS", "40")), 120))
    max_formulas = max(0, min(int(os.environ.get("MODSTORE_XLSX_MAX_FORMULAS", "80")), 500))
    max_scan_cells = max(
        1000, min(int(os.environ.get("MODSTORE_XLSX_SCAN_CELLS", "30000")), 200000)
    )

    def fmt(v: Any) -> str:
        if v is None:
            return ""
        if isinstance(v, float):
            return f"{v:g}"
        return str(v).replace("\n", " ").strip()

    def cell_text(formula_value: Any, cached_value: Any) -> str:
        f = fmt(formula_value)
        c = fmt(cached_value)
        if f.startswith("="):
            return f"公式 {f}" + (f" -> {c}" if c and c != f else "")
        return c or f

    def md_cell(text: Any) -> str:
        return fmt(text).replace("|", "\\|")

    wb_values = openpyxl.load_workbook(io.BytesIO(raw), read_only=False, data_only=True)
    wb_formulas = openpyxl.load_workbook(io.BytesIO(raw), read_only=False, data_only=False)
    lines: List[str] = []
    for ws in wb_formulas.worksheets:
        vws = wb_values[ws.title]
        used_range = ws.calculate_dimension()
        lines.append(f"## Sheet: {ws.title}")
        lines.append(f"- Used range: {used_range}")
        lines.append(f"- Size: {ws.max_row or 0} rows x {ws.max_column or 0} columns")
        merged = [str(rng) for rng in getattr(ws, "merged_cells", []).ranges]
        if merged:
            lines.append(
                f"- Merged cells: {', '.join(merged[:40])}" + (" ..." if len(merged) > 40 else "")
            )

        row_limit = min(ws.max_row or 0, max_rows)
        col_limit = min(ws.max_column or 0, max_cols)
        if row_limit and col_limit:
            lines.append("")
            lines.append("### Grid preview with coordinates")
            headers = ["Row"] + [get_column_letter(c) for c in range(1, col_limit + 1)]
            lines.append("| " + " | ".join(headers) + " |")
            lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
            for r in range(1, row_limit + 1):
                cells: List[str] = []
                has_value = False
                for c in range(1, col_limit + 1):
                    fcell = ws.cell(row=r, column=c)
                    vcell = vws.cell(row=r, column=c)
                    text = cell_text(fcell.value, vcell.value)
                    if text:
                        has_value = True
                        cells.append(md_cell(f"{fcell.coordinate}={text}"))
                    else:
                        cells.append("")
                if has_value:
                    lines.append("| " + " | ".join([str(r), *cells]) + " |")
            if (ws.max_row or 0) > row_limit or (ws.max_column or 0) > col_limit:
                lines.append(f"... preview limited to {row_limit} rows x {col_limit} columns")

        if max_formulas > 0:
            formulas: List[str] = []
            scanned = 0
            for row in ws.iter_rows():
                for fcell in row:
                    scanned += 1
                    if scanned > max_scan_cells or len(formulas) >= max_formulas:
                        break
                    value = fcell.value
                    if isinstance(value, str) and value.startswith("="):
                        cached = fmt(vws[fcell.coordinate].value)
                        formulas.append(
                            f"- {fcell.coordinate}: {value}"
                            + (f" -> cached {cached}" if cached else "")
                        )
                if scanned > max_scan_cells or len(formulas) >= max_formulas:
                    break
            if formulas:
                lines.append("")
                lines.append("### Formulas")
                lines.extend(formulas)
                if len(formulas) >= max_formulas:
                    lines.append(f"... formulas limited to {max_formulas}")
        lines.append("")
    wb_values.close()
    wb_formulas.close()
    return "\n".join(lines)


def parse_upload(filename: str, raw: bytes) -> str:
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, f"文件过大（>{MAX_UPLOAD_BYTES // 1024 // 1024}MB）")
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise HTTPException(400, "仅支持 .txt/.md/.json/.csv/.pdf/.docx/.xlsx")

    if suffix in {".txt", ".md"}:
        text = _decode_text(raw)
    elif suffix == ".json":
        try:
            text = json.dumps(json.loads(_decode_text(raw)), ensure_ascii=False, indent=2)
        except json.JSONDecodeError:
            text = _decode_text(raw)
    elif suffix == ".csv":
        rows = csv.reader(io.StringIO(_decode_text(raw)))
        text = "\n".join(" | ".join(cell.strip() for cell in row if cell.strip()) for row in rows)
    elif suffix == ".pdf":
        text = _parse_pdf(raw)
    elif suffix == ".docx":
        text = _parse_docx(raw)
    elif suffix == ".xlsx":
        text = _parse_xlsx(raw)
    else:
        text = ""

    normalized = _normalize_text(text)
    if len(normalized) < 10:
        raise HTTPException(400, "未能从文件中提取有效文本")
    return normalized


def _split_sentences(text: str) -> List[str]:
    """Split text into sentences using punctuation boundaries.

    Handles CJK and Latin punctuation.  Returns non-empty stripped sentences.
    """
    parts = re.split(r"(?<=[。！？；\n])|(?<=[.!?;]\s)", text)
    return [s.strip() for s in parts if s.strip()]


def _cosine_sim(a: Sequence[float], b: Sequence[float]) -> float:
    """Compute cosine similarity between two vectors."""
    dot = sum(x * y for x, y in zip(a, b))
    na = math.sqrt(sum(x * x for x in a))
    nb = math.sqrt(sum(x * x for x in b))
    if na < 1e-9 or nb < 1e-9:
        return 0.0
    return dot / (na * nb)


class SemanticChunker:
    """Sentence-boundary chunker with similarity-breakpoint detection.

    Algorithm:
    1. Split text into sentences.
    2. Embed each sentence via the configured embedding service.
    3. Walk sentence pairs; when cosine similarity drops below *threshold*,
       insert a chunk boundary.
    4. Merge consecutive sentences into chunks, respecting *min_chunk* and
       *max_chunk* character limits.
    5. Fall back to fixed-window chunking when embedding is unavailable.
    """

    def __init__(
        self,
        *,
        threshold: float | None = None,
        min_chunk: int | None = None,
        max_chunk: int | None = None,
    ):
        self.threshold = threshold or _semantic_threshold()
        self.min_chunk = min_chunk or _semantic_min_chunk()
        self.max_chunk = max_chunk or _semantic_max_chunk()

    async def chunk(self, text: str) -> List[str]:
        sentences = _split_sentences(text)
        if len(sentences) <= 1:
            return [text] if text.strip() else []

        embeddings = await self._embed_sentences(sentences)
        if embeddings is None:
            logger.info("SemanticChunker: embedding unavailable, falling back to fixed chunking")
            return chunk_text(text)

        boundaries = self._detect_boundaries(sentences, embeddings)
        chunks = self._merge_sentences(sentences, boundaries)
        return chunks

    def chunk_sync(self, text: str) -> List[str]:
        """Synchronous wrapper that falls back to fixed chunking.

        Semantic chunking requires async embedding calls; callers that cannot
        await should use this method which degrades gracefully.
        """
        logger.info(
            "SemanticChunker.chunk_sync: async embedding not available, using fixed chunking"
        )
        return chunk_text(text)

    async def _embed_sentences(self, sentences: List[str]) -> Optional[List[List[float]]]:
        try:
            from modstore_server.embedding_service import embed_texts

            vecs = await embed_texts(sentences)
            if vecs and len(vecs) == len(sentences):
                return [list(v) for v in vecs]
        except Exception as exc:
            logger.warning("SemanticChunker: embedding failed: %s", exc)
        return None

    def _detect_boundaries(self, sentences: List[str], embeddings: List[List[float]]) -> List[int]:
        """Return sentence indices where a chunk boundary should be inserted."""
        boundaries: List[int] = [0]
        for i in range(1, len(sentences)):
            sim = _cosine_sim(embeddings[i - 1], embeddings[i])
            if sim < self.threshold:
                boundaries.append(i)
        return boundaries

    def _merge_sentences(self, sentences: List[str], boundaries: List[int]) -> List[str]:
        """Merge sentences between boundaries, respecting size limits."""
        chunks: List[str] = []
        for start_idx in range(len(boundaries)):
            start = boundaries[start_idx]
            end = boundaries[start_idx + 1] if start_idx + 1 < len(boundaries) else len(sentences)
            current = ""
            for si in range(start, end):
                candidate = (current + " " + sentences[si]).strip() if current else sentences[si]
                if len(candidate) > self.max_chunk and current:
                    chunks.append(current)
                    current = sentences[si]
                else:
                    current = candidate
            if current:
                if len(current) < self.min_chunk and chunks:
                    chunks[-1] = (chunks[-1] + " " + current).strip()
                else:
                    chunks.append(current)
        return [c for c in chunks if c.strip()]


def _get_chunker(
    strategy: str | None = None,
    *,
    threshold: float | None = None,
    min_chunk: int | None = None,
    max_chunk: int | None = None,
) -> SemanticChunker | None:
    """Return a SemanticChunker if strategy is 'semantic', else None."""
    effective = (strategy or _CHUNK_STRATEGY).strip().lower()
    if effective == "semantic":
        return SemanticChunker(threshold=threshold, min_chunk=min_chunk, max_chunk=max_chunk)
    return None


def chunk_text(text: str) -> List[str]:
    size = _chunk_size()
    overlap = min(_chunk_overlap(), max(0, size - 1))
    t = _normalize_text(text)
    chunks: List[str] = []
    pos = 0
    while pos < len(t):
        end = min(len(t), pos + size)
        chunk = t[pos:end].strip()
        if chunk:
            chunks.append(chunk)
        if end >= len(t):
            break
        pos = max(pos + 1, end - overlap)
    return chunks


def chunk_text_with_metadata(
    text: str,
    *,
    page_no: Optional[int] = None,
    chunk_strategy: Optional[str] = None,
) -> Tuple[List[str], List[Dict[str, Any]]]:
    strategy = (chunk_strategy or _CHUNK_STRATEGY).strip().lower()
    if strategy == "semantic":
        chunker = _get_chunker(strategy)
        chunks = chunker.chunk_sync(text) if chunker else chunk_text(text)
    else:
        chunks = chunk_text(text)
    metas: List[Dict[str, Any]] = []
    char_pos = 0
    for i, chunk in enumerate(chunks):
        meta: Dict[str, Any] = {
            "chunk_in_source": i,
            "chunk_strategy": strategy,
            "char_range": [char_pos, char_pos + len(chunk)],
        }
        if page_no is not None:
            meta["page_no"] = int(page_no)
        metas.append(meta)
        char_pos += len(chunk)
    return chunks, metas


def parse_and_chunk_with_metadata(
    filename: str,
    raw: bytes,
    *,
    chunk_strategy: Optional[str] = None,
) -> Tuple[str, List[str], List[Dict[str, Any]]]:
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(400, f"文件过大（>{MAX_UPLOAD_BYTES // 1024 // 1024}MB）")
    suffix = Path(filename or "").suffix.lower()
    if suffix not in SUPPORTED_EXTENSIONS:
        raise HTTPException(400, "仅支持 .txt/.md/.json/.csv/.pdf/.docx/.xlsx")

    if suffix == ".pdf":
        pages = _parse_pdf_pages(raw)
        if not pages:
            raise HTTPException(400, "未能从 PDF 中提取有效文本")
        all_text_parts: List[str] = []
        all_chunks: List[str] = []
        all_metas: List[Dict[str, Any]] = []
        for page_no, page_text in pages:
            all_text_parts.append(f"[第 {page_no} 页]\n{page_text}")
            chunks, metas = chunk_text_with_metadata(
                page_text, page_no=page_no, chunk_strategy=chunk_strategy
            )
            all_chunks.extend(chunks)
            all_metas.extend(metas)
        return "\n\n".join(all_text_parts), all_chunks, all_metas

    text = parse_upload(filename, raw)
    chunks, metas = chunk_text_with_metadata(text, chunk_strategy=chunk_strategy)
    if not chunks:
        raise HTTPException(400, "文本分块为空")
    return text, chunks, metas


def parse_and_chunk(filename: str, raw: bytes) -> Tuple[str, List[str]]:
    text, chunks, _metas = parse_and_chunk_with_metadata(filename, raw)
    if not chunks:
        raise HTTPException(400, "文本分块为空")
    return text, chunks
