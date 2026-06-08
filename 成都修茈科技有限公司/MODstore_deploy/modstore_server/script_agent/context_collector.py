"""``context_collector`` —— 把 Brief + 知识库 + SDK 文档拼成生成上下文。

设计要点：

- 不依赖 LLM；纯组装。失败的子步（如知识库不可用）静默降级。
- ``sdk_doc`` 是固定字符串（修改 SDK 时一并更新这里）。
- ``kb_chunks_md`` 仅当 ``brief.references.kb_collection_ids`` 提供时才查。
"""

from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from modstore_server.script_agent.brief import Brief, ContextBundle
from modstore_server.script_agent.package_allowlist import allowed_packages

SDK_DOC = """\
from modstore_runtime import ai, kb_search, employee_run, http_get, log, inputs, outputs

# 1) ai(prompt, *, text="", schema=None, model=None, max_tokens=1024) -> str|dict
#    非确定性兜底：从一段文本里"提取/分类/总结"。schema 给 dict 时尝试解析为对象。
#
# 2) kb_search(query, *, top_k=6) -> list[{collection_id, collection_name, score, text, metadata}]
#    跨当前用户可见的知识库做向量检索。
#
# 3) employee_run(employee_id, task="", payload=None) -> dict
#    以当前用户身份调平台员工。
#
# 4) http_get(url, *, params=None, headers=None, timeout=30) -> {status, text, headers}
#    走父进程的受控 HTTP GET（域名白名单）。
#
# 5) log.info(msg, **fields)  log.warning(...)  log.error(...)
#    结构化日志写 stderr，observer 可解析。
#
# 6) inputs / outputs：
#    inputs.path("a.xlsx")            -> Path("inputs/a.xlsx")
#    inputs.list()                    -> ["a.xlsx", "b.csv"]
#    outputs.write_text(name, text)   -> Path
#    outputs.write_bytes(name, data)  -> Path
#    outputs.write_json(name, data)   -> Path
#
# 强制约束：
#   - 只能读 inputs/、写 outputs/；超出走 SDK
#   - 禁止 import subprocess/ctypes/multiprocessing
#   - 禁止 eval/exec/compile/__import__
#   - 第三方包必须在 allowlist 内
"""


def _xlsx_preview_lines(raw: bytes, *, max_line_chars: int) -> Tuple[List[str], bool]:
    """Return up to 4 tab-separated preview rows from first sheet.

    Second element: True if openpyxl successfully opened the workbook (even if the sheet is empty).
    """
    try:
        from openpyxl import load_workbook  # type: ignore[import-untyped]
    except ImportError:
        return [], False
    try:
        wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
    except Exception:  # noqa: BLE001
        return [], False
    try:
        ws = wb.active
        lines: List[str] = []
        for _row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=4, values_only=True), start=1
        ):
            cells: List[str] = []
            for v in row:
                if v is None:
                    cells.append("")
                else:
                    cells.append(str(v).replace("\n", " ").replace("\r", "").strip())
            while cells and cells[-1] == "":
                cells.pop()
            if not cells:
                continue
            line = "\t".join(cells)
            if len(line) > max_line_chars:
                line = line[: max_line_chars - 1] + "…"
            lines.append(line)
            if len(lines) >= 4:
                break
        return lines, True
    finally:
        wb.close()


def tabular_upload_preview(
    upload_items: Optional[Sequence[Dict[str, Any]]],
    *,
    max_files: int = 6,
    max_line_chars: int = 2000,
) -> str:
    """从上传的二进制内容中抽取 CSV/TSV/XLSX 表头与样例行，供 LLM 列名接地。"""
    if not upload_items:
        return ""
    blocks: List[str] = []
    used = 0
    xlsx_failed_hint = False
    for item in upload_items:
        if used >= max_files:
            break
        fn = str((item or {}).get("filename") or "").strip()
        raw = (item or {}).get("content")
        if not isinstance(raw, (bytes, bytearray)):
            continue
        low = fn.lower()
        lines: List[str] = []

        if low.endswith((".csv", ".tsv")):
            try:
                text = bytes(raw).decode("utf-8-sig", errors="replace")
            except Exception:  # noqa: BLE001
                continue
            lines = [ln.rstrip("\r\n") for ln in text.splitlines() if ln.strip()][:4]
        elif low.endswith((".xlsx", ".xlsm")):
            lines, wb_ok = _xlsx_preview_lines(bytes(raw), max_line_chars=max_line_chars)
            if not lines:
                if wb_ok:
                    xlsx_failed_hint = True
                continue
        else:
            continue

        if not lines:
            continue
        used += 1
        sep = (
            "Excel 首工作表（单元格以 TAB 拼接展示）"
            if low.endswith((".xlsx", ".xlsm"))
            else ("制表符(TAB)" if low.endswith(".tsv") else "逗号")
        )
        blocks.append(f"#### `{fn}`\n")
        blocks.append(f"- 分隔符: {sep}；**必须使用下列确切列名，禁止臆造别名**\n")
        blocks.append(f"- 第1行(表头): {lines[0][:max_line_chars]}\n")
        for idx, ln in enumerate(lines[1:3], start=2):
            blocks.append(f"- 第{idx}行(样例): {ln[:max_line_chars]}\n")
        blocks.append("\n")

    header = "## 表格类输入（CSV/TSV/XLSX）列名接地\n\n"
    if xlsx_failed_hint and not blocks:
        header += (
            "> 提示：检测到 Excel 上传但未能解析预览（文件损坏或格式异常）。"
            "脚本不得臆造列名；请改用 CSV 导出后再上传，或在需求中写明确切列名。\n\n"
        )
    if not blocks:
        return header.rstrip() if xlsx_failed_hint else ""
    return header + "".join(blocks)


def _summarize_inputs(
    brief: Brief,
    upload_items: Optional[Sequence[Dict[str, Any]]] = None,
) -> str:
    if brief.inputs:
        parts = []
        for f in brief.inputs:
            suffix = ""
            if f.description:
                suffix = f" — {f.description}"
            parts.append(f"- {f.filename}{suffix}")
        base = "\n".join(parts)
    else:
        base = "(Brief 未列出输入文件名)"
    tab = tabular_upload_preview(upload_items)
    if tab:
        return base + "\n\n" + tab
    return base


async def _collect_kb_chunks(
    *,
    user_id: int,
    queries: Sequence[str],
    collection_ids: Optional[Sequence[int]] = None,
    top_k_per_query: int = 3,
) -> str:
    """对每条 query 检索知识库，拼成 markdown 列表。失败返回空串。"""
    try:
        from modstore_server.rag_service import retrieve
    except Exception:  # noqa: BLE001
        return ""
    rows: List[str] = []
    for q in queries:
        if not q.strip():
            continue
        try:
            chunks = await retrieve(
                user_id=user_id,
                query=q,
                top_k=top_k_per_query,
                extra_collection_ids=collection_ids,
            )
        except Exception:  # noqa: BLE001
            continue
        for c in chunks:
            text = str(getattr(c, "text", "") or "").strip()
            if not text:
                continue
            name = str(getattr(c, "collection_name", "") or "")
            rows.append(f"- ({name}) {text[:600]}")
    return "\n".join(rows)


async def collect_context(
    brief: Brief,
    *,
    user_id: int,
    extra_kb_queries: Iterable[str] = (),
    upload_items: Optional[Sequence[Dict[str, Any]]] = None,
) -> ContextBundle:
    """同步拼装 + 异步检索知识库，返回喂给 planner / code_writer 的 bundle。"""
    kb_md = ""
    refs = brief.references or {}
    coll_ids = refs.get("kb_collection_ids") or refs.get("kb_collections")
    if coll_ids:
        try:
            ids: List[int] = []
            for x in coll_ids:
                try:
                    ids.append(int(x))
                except Exception:  # noqa: BLE001
                    continue
            queries = list(extra_kb_queries) or [brief.goal[:200]]
            kb_md = await _collect_kb_chunks(
                user_id=user_id,
                queries=queries,
                collection_ids=ids,
            )
        except Exception:  # noqa: BLE001
            kb_md = ""

    return ContextBundle(
        brief_md=brief.as_markdown(),
        inputs_summary=_summarize_inputs(brief, upload_items),
        kb_chunks_md=kb_md,
        sdk_doc=SDK_DOC,
        allowlist_packages=sorted(allowed_packages()),
    )
