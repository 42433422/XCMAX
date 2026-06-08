from __future__ import annotations

import io
import json
import zipfile
from pathlib import Path


def test_asset_employee_entrypoint_has_system_prompt_and_generated_runtime(tmp_path):
    from openpyxl import Workbook

    from modstore_server.employee_asset_pipeline import (
        build_rule_spec,
        materialize_asset_employee_pack,
        prepare_employee_assets,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "每日统计"
    ws.append(["姓名", "日期", "上班1", "下班1"])
    src = tmp_path / "考勤-2026-3月份考勤统计表.xlsx"
    wb.save(src)

    asset_manifest = prepare_employee_assets(
        session_id="asset-test",
        user_id=1,
        raw_files=[{"filename": src.name, "content": src.read_bytes()}],
        repo_root=tmp_path,
    )
    rule_spec = build_rule_spec("用两个 Excel 和规则生成考勤转换员工", asset_manifest)
    manifest = {
        "id": "attendance-transform-employee",
        "name": "考勤处理员",
        "version": "1.0.0",
        "artifact": "employee_pack",
        "employee": {"id": "attendance-transform-employee", "label": "考勤处理员"},
        "employee_config_v2": {"actions": {"handlers": ["direct_python"]}},
    }
    generated_convert = """from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, Optional

def convert_file(src_path: Path, output_path: Path, *, template_path: Optional[Path], payload: Dict[str, Any], ctx: Dict[str, Any], rule_spec: Dict[str, Any]) -> Dict[str, Any]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(src_path.read_bytes())
    return {"copied": True, "output_path": str(output_path)}
"""

    pack_dir, raw = materialize_asset_employee_pack(
        manifest=manifest,
        rule_spec=rule_spec,
        asset_manifest=asset_manifest,
        generated_convert_py=generated_convert,
    )
    employee_py = next((pack_dir / "backend" / "employees").glob("*.py")).read_text(
        encoding="utf-8"
    )

    assert "SYSTEM_PROMPT =" in employee_py
    assert "RULE_SPEC =" in employee_py
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        names = {n.replace("\\", "/") for n in zf.namelist()}
    assert (
        "attendance-transform-employee/backend/templates/424/考勤-2026-3月份考勤统计表.xlsx"
        in names
    )
    assert "attendance-transform-employee/backend/vendor/attendance_transform/convert.py" in names
    assert "attendance-transform-employee/manifest.json" in names


def test_generated_convert_validation_rejects_placeholder_fragments():
    from modstore_server.employee_asset_pipeline import _validate_generated_convert_py

    ok, error = _validate_generated_convert_py("print('hello')")
    assert not ok
    assert "convert_file" in error


def test_extract_python_code_strips_llm_preface():
    from modstore_server.employee_asset_pipeline import (
        _extract_python_code,
        _validate_generated_convert_py,
    )

    raw = """下面是修复后的 convert.py：

from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, Optional

def convert_file(src_path: Path, output_path: Path, *, template_path: Optional[Path], payload: Dict[str, Any], ctx: Dict[str, Any], rule_spec: Dict[str, Any]) -> Dict[str, Any]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_bytes(src_path.read_bytes())
    return {"output_path": str(output_path)}
"""
    code = _extract_python_code(raw)
    ok, error = _validate_generated_convert_py(code)

    assert code.startswith("from __future__ import annotations")
    assert ok, error


def test_employee_manifest_alias_candidates():
    from modstore_server.employee_api import _candidate_employee_pack_ids

    candidates = _candidate_employee_pack_ids("attendance_transform")
    assert candidates[:2] == ["attendance_transform", "attendance-transform"]
    assert "attendance-transform-employee" in candidates


def test_directory_zip_contains_full_asset_pack(tmp_path):
    from modstore_server.employee_asset_pipeline import build_employee_pack_zip_from_dir

    pack = tmp_path / "asset-pack"
    (pack / "backend" / "vendor" / "asset_pack_runtime").mkdir(parents=True)
    (pack / "backend" / "templates").mkdir(parents=True)
    (pack / "manifest.json").write_text(json.dumps({"id": "asset-pack"}), encoding="utf-8")
    (pack / "backend" / "vendor" / "asset_pack_runtime" / "convert.py").write_text(
        "x = 1\n", encoding="utf-8"
    )
    (pack / "backend" / "templates" / "template.xlsx").write_bytes(b"xlsx")

    raw = build_employee_pack_zip_from_dir("asset-pack", pack)
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        names = {n.replace("\\", "/") for n in zf.namelist()}

    assert "asset-pack/backend/vendor/asset_pack_runtime/convert.py" in names
    assert "asset-pack/backend/templates/template.xlsx" in names


def test_attendance_prompt_generates_requested_pack_shape(tmp_path):
    from openpyxl import Workbook

    from modstore_server.employee_asset_pipeline import (
        build_rule_spec,
        materialize_asset_employee_pack,
        prepare_employee_assets,
    )

    template = tmp_path / "考勤-2026-3月份考勤统计表.xlsx"
    source = tmp_path / "钉钉导出来的考勤数据.xlsx"
    for path, title in ((template, "明细"), (source, "每日统计")):
        wb = Workbook()
        ws = wb.active
        ws.title = title
        ws.append(["姓名", "部门", "日期", "打卡时间"])
        wb.save(path)

    brief = """员工包 ID：taiyangniao-attendance-employee
员工名称：太阳鸟考勤员
运行方式：direct_python
禁止使用 echo
默认输出路径：424/考勤转换输出.xlsx
默认模板路径：424/考勤-2026-3月份考勤统计表.xlsx
"""
    asset_manifest = prepare_employee_assets(
        session_id="shape-test",
        user_id=1,
        raw_files=[
            {"filename": template.name, "content": template.read_bytes()},
            {"filename": source.name, "content": source.read_bytes()},
        ],
        repo_root=tmp_path,
    )
    rule_spec = build_rule_spec(brief, asset_manifest)
    manifest = {
        "id": "taiyangniao-attendance-employee",
        "name": "太阳鸟考勤员",
        "version": "1.0.0",
        "artifact": "employee_pack",
        "employee": {"id": "taiyangniao-attendance", "label": "太阳鸟考勤员"},
        "employee_config_v2": {
            "perception": {
                "type": "file_or_text",
                "accepted_extensions": [".xlsx", ".xlsm", ".xls"],
            },
            "actions": {
                "handlers": ["direct_python"],
                "direct_python": {
                    "default_output_relpath": "424/考勤转换输出.xlsx",
                    "default_template_relpath": "424/考勤-2026-3月份考勤统计表.xlsx",
                    "default_use_personnel_roster": True,
                },
            },
        },
    }
    generated_convert = """from __future__ import annotations
from pathlib import Path
from typing import Any, Dict, Optional

def convert_file(src_path: Path, output_path: Path, *, template_path: Optional[Path], payload: Dict[str, Any], ctx: Dict[str, Any], rule_spec: Dict[str, Any]) -> Dict[str, Any]:
    from openpyxl import load_workbook
    if template_path is None:
        raise FileNotFoundError("模板不存在")
    wb = load_workbook(template_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return {"output_path": str(output_path), "source_rows": 1, "stat_rows": 1, "use_personnel_roster": True}
"""

    pack_dir, raw = materialize_asset_employee_pack(
        manifest=manifest,
        rule_spec=rule_spec,
        asset_manifest=asset_manifest,
        generated_convert_py=generated_convert,
    )

    assert (pack_dir / "build_xcemp.py").is_file()
    assert (pack_dir / "backend" / "employees" / "taiyangniao_attendance.py").is_file()
    for name in (
        "convert.py",
        "mapper.py",
        "parser.py",
        "rules.py",
        "paths.py",
        "mapping.py",
        "header_resolver.py",
    ):
        assert (pack_dir / "backend" / "vendor" / "taiyangniao_attendance" / name).is_file()
    assert (pack_dir / "backend" / "templates" / "424" / template.name).is_file()

    manifest_out = json.loads((pack_dir / "manifest.json").read_text(encoding="utf-8"))
    actions = manifest_out["employee_config_v2"]["actions"]
    assert actions["handlers"] == ["direct_python"]
    assert "echo" not in json.dumps(actions, ensure_ascii=False)
    assert (
        actions["direct_python"]["default_template_relpath"] == "424/考勤-2026-3月份考勤统计表.xlsx"
    )

    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        names = {n.replace("\\", "/") for n in zf.namelist()}
    assert (
        "taiyangniao-attendance-employee/backend/templates/424/考勤-2026-3月份考勤统计表.xlsx"
        in names
    )
    assert (
        "taiyangniao-attendance-employee/backend/vendor/taiyangniao_attendance/header_resolver.py"
        in names
    )


def test_explicit_employee_pack_id_overrides_llm_manifest_id():
    from modstore_server.employee_asset_pipeline import _normalize_manifest

    brief = "员工包 ID：taiyangniao-attendance-employee\n员工名称：太阳鸟考勤员\n用于考勤转换"
    rule_spec = {
        "accepted_extensions": [".xlsx"],
        "default_output_relpath": "424/考勤转换输出.xlsx",
    }
    manifest = {
        "id": "attendance-transform-employee",
        "name": "考勤处理员",
        "employee": {"id": "attendance-transform"},
        "employee_config_v2": {"actions": {"handlers": ["direct_python"]}},
    }

    out = _normalize_manifest(manifest, brief, rule_spec)

    assert out["id"] == "taiyangniao-attendance-employee"
    assert out["employee_config_v2"]["identity"]["id"] == "taiyangniao-attendance-employee"


def test_build_rule_spec_prefers_word_extract_over_generate():
    from modstore_server.employee_asset_pipeline import build_rule_spec

    brief = (
        "Word 全量提取：真实解析 docx，输出 document_full.json 与 document_full.txt，"
        "含 paragraphs、tables、outline；禁止编造正文"
    )
    rule_spec = build_rule_spec(brief, {"assets": []})
    assert rule_spec["runtime_kind"] == "word_full_extract"


def test_word_full_extract_pack_has_direct_python_vendor_convert():
    from modstore_server.employee_asset_pipeline import (
        _fallback_manifest,
        _normalize_manifest,
        build_rule_spec,
        materialize_asset_employee_pack,
        render_runtime_modules,
    )
    from modstore_server.word_extract_runtime import (
        is_word_full_extract,
        validate_word_extract_backend,
    )

    brief = "全量提取 Word 文档所有格式和信息，输出 JSON 与 txt"
    assert is_word_full_extract(brief)
    rule_spec = build_rule_spec(brief, {"assets": []})
    assert rule_spec["runtime_kind"] == "word_full_extract"
    manifest = _normalize_manifest(_fallback_manifest(brief, rule_spec), brief, rule_spec)
    convert_py = render_runtime_modules(rule_spec, None)["convert.py"]
    pack_dir, raw = materialize_asset_employee_pack(
        manifest=manifest,
        rule_spec=rule_spec,
        asset_manifest={"assets": []},
        generated_convert_py=convert_py,
    )
    handlers = manifest["employee_config_v2"]["actions"]["handlers"]
    assert handlers == ["direct_python"]
    errs, _ = validate_word_extract_backend(pack_dir)
    assert not errs
    with zipfile.ZipFile(io.BytesIO(raw), "r") as zf:
        names = {n.replace("\\", "/") for n in zf.namelist()}
    assert any("/backend/vendor/" in n and n.endswith("convert.py") for n in names)
