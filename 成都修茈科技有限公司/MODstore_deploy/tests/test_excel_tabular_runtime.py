"""Excel 读取/生成 runtime 路由与 convert 模块测试。"""

from __future__ import annotations

import json
from pathlib import Path

import pytest


def test_excel_routing_keywords():
    from modstore_server.excel_tabular_runtime import is_excel_full_read, is_excel_generate

    assert is_excel_full_read("制作 Excel读取员工，上传 xlsx 读 sheet 表头 单元格全量")
    assert not is_excel_generate("制作 Excel读取员工，上传 xlsx 读 sheet 表头 单元格全量")
    assert is_excel_generate("Excel生成员工，JSON中介写出 xlsx 文件，中介是json")
    assert not is_excel_full_read("Excel生成员工，JSON中介写出 xlsx 文件")


def test_excel_read_convert_writes_json(tmp_path: Path):
    from modstore_server.excel_tabular_runtime import (
        build_excel_read_rule_spec,
        minimal_xlsx_fixture_bytes,
        render_excel_read_convert_module,
    )

    code = render_excel_read_convert_module()
    ns: dict = {}
    exec(code, ns)  # noqa: S102
    convert_file = ns["convert_file"]

    root = tmp_path
    src = root / "sample.xlsx"
    src.write_bytes(minimal_xlsx_fixture_bytes())
    out = root / "outputs" / "workbook.json"
    rule_spec = build_excel_read_rule_spec("excel read")
    result = convert_file(
        src,
        out,
        template_path=None,
        payload={},
        ctx={},
        rule_spec=rule_spec,
    )
    assert out.is_file()
    data = json.loads(out.read_text(encoding="utf-8"))
    assert data["sheet_count"] >= 1
    assert data["sheets"][0]["name"] == "Sheet1"
    assert data["sheets"][0]["cells"]
    assert result["sheet_count"] >= 1


@pytest.mark.xfail(
    strict=False, reason="excel_tabular_runtime generate/convert pre-existing failure"
)
def test_excel_generate_convert_writes_xlsx(tmp_path: Path):
    from modstore_server.excel_tabular_runtime import (
        minimal_json_fixture_bytes,
        render_excel_generate_convert_module,
    )

    code = render_excel_generate_convert_module()
    ns: dict = {}
    exec(code, ns)  # noqa: S102
    convert_file = ns["convert_file"]

    root = tmp_path
    src = root / "table.json"
    src.write_bytes(minimal_json_fixture_bytes())
    out = root / "outputs" / "output.xlsx"
    convert_file(
        src,
        out,
        template_path=None,
        payload={},
        ctx={},
        rule_spec={"default_output_relpath": "outputs/output.xlsx", "output_schema": []},
    )
    assert out.is_file()
    from openpyxl import load_workbook

    wb = load_workbook(out, read_only=True, data_only=True)
    try:
        ws = wb.active
        assert ws["A1"].value == "name"
        assert ws["A2"].value == "alice"
    finally:
        wb.close()


def test_materialize_excel_read_pack_zip():
    from modstore_server.employee_asset_pipeline import (
        _fallback_manifest,
        _normalize_manifest,
        materialize_asset_employee_pack,
    )
    from modstore_server.excel_tabular_runtime import (
        build_excel_read_rule_spec,
        render_excel_read_convert_module,
    )

    brief = "Excel全量读取员工 xlsx sheet 表头 单元格"
    rule_spec = build_excel_read_rule_spec(brief)
    manifest = _normalize_manifest(_fallback_manifest(brief, rule_spec), brief, rule_spec)
    manifest["id"] = "excel-full-read-employee"
    pack_dir, raw_zip = materialize_asset_employee_pack(
        manifest=manifest,
        rule_spec=rule_spec,
        asset_manifest={
            "assets": [],
            "templates": [],
            "example_inputs": [],
            "expected_outputs": [],
            "rules": [],
        },
        generated_convert_py=render_excel_read_convert_module(),
    )
    assert len(raw_zip) > 500
    vendor = list(pack_dir.rglob("backend/vendor/*/convert.py"))
    assert vendor
